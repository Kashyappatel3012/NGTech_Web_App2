import os
import re
import zipfile
import tempfile
import calendar
from datetime import datetime, date
from collections import defaultdict

from flask import Blueprint, current_app, flash, redirect, request, send_from_directory, url_for, abort, jsonify, send_file
from flask_login import login_required, current_user
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from werkzeug.utils import secure_filename
import sys
import os
# Add parent directory to path to import security_utils
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from security_utils import sanitize_email_header, sanitize_email_content, sanitize_filename, sanitize_path

hr_dashboard_bp = Blueprint('hr_dashboard_bp', __name__)

MIN_TRACKER_DATE = date(2025, 10, 1)


@hr_dashboard_bp.route('/hr/download_daily_activity_tracker', methods=['POST'])
@login_required
def download_daily_activity_tracker():
    if current_user.department != "HR":
        abort(403)

    tracker_date_str = request.form.get('tracker_date')
    if not tracker_date_str:
        flash('Please select a date for the activity tracker.', 'error')
        return redirect(url_for('hr_dashboard'))

    try:
        tracker_date = datetime.strptime(tracker_date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date format. Please pick a valid date.', 'error')
        return redirect(url_for('hr_dashboard'))

    today = datetime.now().date()
    if tracker_date < MIN_TRACKER_DATE or tracker_date > today:
        flash('Please choose a date between 01 Oct 2025 and today.', 'error')
        return redirect(url_for('hr_dashboard'))

    filename = f"{tracker_date.day}_{tracker_date.strftime('%b')}_{tracker_date.year}.xlsx"
    tracker_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Everyday_Updated_Work')
    file_path = os.path.join(tracker_dir, filename)

    if not os.path.exists(file_path):
        formatted_date = tracker_date.strftime('%d %b %Y')
        flash(f'No activity tracker found for {formatted_date}.', 'error')
        return redirect(url_for('hr_dashboard'))

    return send_from_directory(tracker_dir, filename, as_attachment=True)


@hr_dashboard_bp.route('/hr/get_available_tracker_dates', methods=['GET'])
@login_required
def get_available_tracker_dates():
    """Get list of dates for which activity tracker files exist"""
    if current_user.department != "HR":
        abort(403)
    
    try:
        tracker_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Everyday_Updated_Work')
        
        if not os.path.exists(tracker_dir):
            return jsonify({'success': True, 'dates': []})
        
        available_dates = []
        # Pattern to match: DD_Mon_YYYY.xlsx (e.g., 28_Oct_2025.xlsx)
        date_pattern = re.compile(r'^(\d{1,2})_([A-Za-z]{3})_(\d{4})\.xlsx$')
        
        month_map = {
            'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
            'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
        }
        
        for filename in os.listdir(tracker_dir):
            if filename.endswith('.xlsx'):
                match = date_pattern.match(filename)
                if match:
                    try:
                        day = int(match.group(1))
                        month_str = match.group(2)
                        year = int(match.group(3))
                        
                        if month_str in month_map:
                            month = month_map[month_str]
                            file_date = date(year, month, day)
                            
                            # Only include dates from Oct 1, 2025 onwards and not in the future
                            today = datetime.now().date()
                            if file_date >= MIN_TRACKER_DATE and file_date <= today:
                                available_dates.append({
                                    'date': file_date.isoformat(),
                                    'display': file_date.strftime('%d %b %Y'),
                                    'filename': filename
                                })
                    except (ValueError, KeyError):
                        # Skip invalid date formats
                        continue
        
        # Sort dates in descending order (newest first)
        available_dates.sort(key=lambda x: x['date'], reverse=True)
        
        return jsonify({'success': True, 'dates': available_dates})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@hr_dashboard_bp.route('/hr/download_time_range_tracker', methods=['POST'])
@login_required
def download_time_range_tracker():
    """Download all activity tracker files between start_date and end_date as a zip file"""
    if current_user.department != "HR":
        abort(403)

    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')
    
    if not start_date_str or not end_date_str:
        flash('Please select both start date and end date.', 'error')
        return redirect(url_for('hr_dashboard'))

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date format. Please select valid dates.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Validate date range
    if start_date > end_date:
        flash('Start date must be before or equal to end date.', 'error')
        return redirect(url_for('hr_dashboard'))

    today = datetime.now().date()
    if start_date < MIN_TRACKER_DATE or end_date > today:
        flash(f'Please choose dates between {MIN_TRACKER_DATE.strftime("%d %b %Y")} and today.', 'error')
        return redirect(url_for('hr_dashboard'))

    tracker_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Everyday_Updated_Work')
    
    if not os.path.exists(tracker_dir):
        flash('Activity tracker directory not found.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Pattern to match: DD_Mon_YYYY.xlsx (e.g., 28_Oct_2025.xlsx)
    date_pattern = re.compile(r'^(\d{1,2})_([A-Za-z]{3})_(\d{4})\.xlsx$')
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
    }

    # Find all files within the date range
    files_to_zip = []
    for filename in os.listdir(tracker_dir):
        if filename.endswith('.xlsx'):
            match = date_pattern.match(filename)
            if match:
                try:
                    day = int(match.group(1))
                    month_str = match.group(2)
                    year = int(match.group(3))
                    
                    if month_str in month_map:
                        month = month_map[month_str]
                        file_date = date(year, month, day)
                        
                        # Check if file date is within the range
                        if start_date <= file_date <= end_date:
                            file_path = os.path.join(tracker_dir, filename)
                            if os.path.exists(file_path):
                                files_to_zip.append((file_path, filename))
                except (ValueError, KeyError):
                    # Skip invalid date formats
                    continue

    if not files_to_zip:
        flash(f'No activity tracker files found between {start_date.strftime("%d %b %Y")} and {end_date.strftime("%d %b %Y")}.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Create a temporary zip file
    try:
        temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        temp_zip_path = temp_zip.name
        temp_zip.close()

        # Create zip file with all matching files
        with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path, filename in files_to_zip:
                zipf.write(file_path, filename)

        # Generate zip filename
        zip_filename = f"Activity_Tracker_{start_date.strftime('%d_%b_%Y')}_to_{end_date.strftime('%d_%b_%Y')}.zip"

        # Send the zip file
        return send_file(
            temp_zip_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name=zip_filename
        )
    except Exception as e:
        flash(f'Error creating zip file: {str(e)}', 'error')
        return redirect(url_for('hr_dashboard'))


@hr_dashboard_bp.route('/hr/generate_attendance_record', methods=['POST'])
@login_required
def generate_attendance_record():
    """Generate attendance record Excel file for selected month and year"""
    if current_user.department != "HR":
        abort(403)

    month_str = request.form.get('month')
    year_str = request.form.get('year')

    if not month_str or not year_str:
        flash('Please select both month and year.', 'error')
        return redirect(url_for('hr_dashboard'))

    try:
        month = int(month_str)
        year = int(year_str)
        
        if month < 1 or month > 12:
            flash('Invalid month selected.', 'error')
            return redirect(url_for('hr_dashboard'))
            
        if year < 2020 or year > 2030:
            flash('Invalid year selected.', 'error')
            return redirect(url_for('hr_dashboard'))
    except ValueError:
        flash('Invalid month or year format.', 'error')
        return redirect(url_for('hr_dashboard'))

    tracker_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Everyday_Updated_Work')
    
    if not os.path.exists(tracker_dir):
        flash('Activity tracker directory not found.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Pattern to match: DD_Mon_YYYY.xlsx (e.g., 28_Oct_2025.xlsx)
    date_pattern = re.compile(r'^(\d{1,2})_([A-Za-z]{3})_(\d{4})\.xlsx$')
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
    }

    # Find all files matching the selected month and year
    matching_files = []
    for filename in os.listdir(tracker_dir):
        if filename.endswith('.xlsx'):
            match = date_pattern.match(filename)
            if match:
                try:
                    day = int(match.group(1))
                    month_str_file = match.group(2)
                    year_file = int(match.group(3))
                    
                    if month_str_file in month_map:
                        month_file = month_map[month_str_file]
                        
                        # Check if file matches selected month and year
                        if month_file == month and year_file == year:
                            file_path = os.path.join(tracker_dir, filename)
                            if os.path.exists(file_path):
                                matching_files.append(file_path)
                except (ValueError, KeyError):
                    continue

    if not matching_files:
        month_name = calendar.month_name[month]
        flash(f'No activity tracker files found for {month_name} {year}.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Process all matching files and calculate attendance
    employee_attendance = defaultdict(float)  # {employee_name: total_working_days}
    
    for file_path in matching_files:
        try:
            workbook = load_workbook(file_path, data_only=True)
            
            # Check if "Updated Work" worksheet exists
            if 'Updated Work' not in workbook.sheetnames:
                continue
            
            worksheet = workbook['Updated Work']
            
            # Find header row (assuming headers are in row 1)
            header_row = 1
            employee_name_col = None
            total_time_col = None
            
            # Find column indices for Employee Name (Column A) and Total Time (Column D)
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip().lower()
                    if 'employee name' in cell_str or col_idx == 1:
                        employee_name_col = col_idx
                    if 'total time' in cell_str or col_idx == 4:
                        total_time_col = col_idx
            
            # Default to Column A (1) and Column D (4) if not found
            if employee_name_col is None:
                employee_name_col = 1
            if total_time_col is None:
                total_time_col = 4
            
            # Process data rows (starting from row 2)
            for row_idx in range(header_row + 1, worksheet.max_row + 1):
                employee_name_cell = worksheet.cell(row=row_idx, column=employee_name_col)
                total_time_cell = worksheet.cell(row=row_idx, column=total_time_col)
                
                employee_name = employee_name_cell.value
                total_time_value = total_time_cell.value
                
                # Skip empty rows
                if not employee_name:
                    continue
                
                employee_name = str(employee_name).strip()
                
                # Convert total_time to float
                try:
                    if total_time_value is None:
                        total_time = 0.0
                    else:
                        total_time = float(total_time_value)
                except (ValueError, TypeError):
                    total_time = 0.0
                
                # Calculate working day based on Total Time
                # > 6 hours = 1 working day
                # 3-6 hours = 0.5 working day
                # < 3 hours = 0 working day
                if total_time > 6:
                    working_day = 1.0
                elif total_time >= 3:
                    working_day = 0.5
                else:
                    working_day = 0.0
                
                employee_attendance[employee_name] += working_day
                
        except Exception as e:
            print(f"Error processing file {file_path}: {str(e)}")
            continue

    if not employee_attendance:
        flash('No attendance data found in the files.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Get total days in the selected month
    total_days_in_month = calendar.monthrange(year, month)[1]

    # Create Excel workbook with attendance summary
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Record"
        
        # Header row
        headers = ['Sr. No', 'Employee Name', 'Working Day', 'Non-Working Day']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        row_num = 2
        for sr_no, (employee_name, working_days) in enumerate(sorted(employee_attendance.items()), start=1):
            non_working_days = total_days_in_month - working_days
            
            ws.cell(row=row_num, column=1, value=sr_no).border = border
            ws.cell(row=row_num, column=2, value=employee_name).border = border
            ws.cell(row=row_num, column=3, value=working_days).border = border
            ws.cell(row=row_num, column=4, value=non_working_days).border = border
            
            # Center align Sr. No and numeric columns
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')
            
            row_num += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        
        # Save to temporary file
        month_name = calendar.month_name[month]
        temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_excel_path = temp_excel.name
        temp_excel.close()
        
        wb.save(temp_excel_path)
        
        # Generate filename
        excel_filename = f"Attendance_Record_{month_name}_{year}.xlsx"
        
        # Send the Excel file
        return send_file(
            temp_excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=excel_filename
        )
        
    except Exception as e:
        flash(f'Error generating attendance record: {str(e)}', 'error')
        return redirect(url_for('hr_dashboard'))


@hr_dashboard_bp.route('/hr/generate_personal_activity_tracker', methods=['POST'])
@login_required
def generate_personal_activity_tracker():
    """Generate personal activity tracker Excel file for selected employee, month and year"""
    if current_user.department != "HR":
        abort(403)

    employee_name = request.form.get('employee_name', '').strip()
    month_str = request.form.get('month')
    year_str = request.form.get('year')

    if not employee_name:
        flash('Please enter employee name.', 'error')
        return redirect(url_for('hr_dashboard'))

    if not month_str or not year_str:
        flash('Please select both month and year.', 'error')
        return redirect(url_for('hr_dashboard'))

    try:
        month = int(month_str)
        year = int(year_str)
        
        if month < 1 or month > 12:
            flash('Invalid month selected.', 'error')
            return redirect(url_for('hr_dashboard'))
            
        if year < 2020 or year > 2030:
            flash('Invalid year selected.', 'error')
            return redirect(url_for('hr_dashboard'))
    except ValueError:
        flash('Invalid month or year format.', 'error')
        return redirect(url_for('hr_dashboard'))

    tracker_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Everyday_Updated_Work')
    
    if not os.path.exists(tracker_dir):
        flash('Activity tracker directory not found.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Pattern to match: DD_Mon_YYYY.xlsx (e.g., 28_Oct_2025.xlsx)
    date_pattern = re.compile(r'^(\d{1,2})_([A-Za-z]{3})_(\d{4})\.xlsx$')
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
    }

    # Find all files matching the selected month and year
    matching_files = []
    for filename in os.listdir(tracker_dir):
        if filename.endswith('.xlsx'):
            match = date_pattern.match(filename)
            if match:
                try:
                    day = int(match.group(1))
                    month_str_file = match.group(2)
                    year_file = int(match.group(3))
                    
                    if month_str_file in month_map:
                        month_file = month_map[month_str_file]
                        
                        # Check if file matches selected month and year
                        if month_file == month and year_file == year:
                            file_path = os.path.join(tracker_dir, filename)
                            if os.path.exists(file_path):
                                matching_files.append((file_path, filename))
                except (ValueError, KeyError):
                    continue

    if not matching_files:
        month_name = calendar.month_name[month]
        flash(f'No activity tracker files found for {month_name} {year}.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Process all matching files and extract data for the selected employee
    all_rows = []  # List of dictionaries with employee data
    
    for file_path, filename in matching_files:
        try:
            workbook = load_workbook(file_path, data_only=True)
            
            # Check if "Updated Work" worksheet exists
            if 'Updated Work' not in workbook.sheetnames:
                continue
            
            worksheet = workbook['Updated Work']
            
            # Find header row (assuming headers are in row 1)
            header_row = 1
            
            # Find column indices
            employee_name_col = None
            date_col = None
            total_time_col = None
            task_columns = {}  # {task_num: {'task': col, 'client': col, 'time': col}}
            
            # Scan header row to find all columns
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip()
                    cell_lower = cell_str.lower()
                    
                    # Employee Name (Column A or header contains "employee name")
                    if col_idx == 1 or 'employee name' in cell_lower:
                        employee_name_col = col_idx
                    
                    # Date (Column B or header contains "date")
                    if col_idx == 2 or ('date' in cell_lower and 'client' not in cell_lower):
                        date_col = col_idx
                    
                    # Total Time (Column D or header contains "total time")
                    if col_idx == 4 or 'total time' in cell_lower:
                        total_time_col = col_idx
                    
                    # Task columns - pattern: Task1, Task1 Client Name, Time For Task1, etc.
                    task_match = re.match(r'^task(\d+)$', cell_lower)
                    if task_match:
                        task_num = int(task_match.group(1))
                        if task_num not in task_columns:
                            task_columns[task_num] = {}
                        task_columns[task_num]['task'] = col_idx
                    
                    client_match = re.match(r'^task(\d+)\s+client\s+name$', cell_lower)
                    if client_match:
                        task_num = int(client_match.group(1))
                        if task_num not in task_columns:
                            task_columns[task_num] = {}
                        task_columns[task_num]['client'] = col_idx
                    
                    time_match = re.match(r'^time\s+for\s+task(\d+)$', cell_lower)
                    if time_match:
                        task_num = int(time_match.group(1))
                        if task_num not in task_columns:
                            task_columns[task_num] = {}
                        task_columns[task_num]['time'] = col_idx
            
            # Default to Column A, B, D if not found
            if employee_name_col is None:
                employee_name_col = 1
            if date_col is None:
                date_col = 2
            if total_time_col is None:
                total_time_col = 4
            
            # Process data rows (starting from row 2)
            for row_idx in range(header_row + 1, worksheet.max_row + 1):
                employee_name_cell = worksheet.cell(row=row_idx, column=employee_name_col)
                employee_name_value = employee_name_cell.value
                
                # Skip empty rows
                if not employee_name_value:
                    continue
                
                employee_name_value = str(employee_name_value).strip()
                
                # Filter by selected employee name (case-insensitive)
                if employee_name_value.lower() != employee_name.lower():
                    continue
                
                # Extract Date
                date_cell = worksheet.cell(row=row_idx, column=date_col)
                date_value = date_cell.value
                if date_value:
                    if isinstance(date_value, datetime):
                        date_str = date_value.strftime('%d-%m-%Y')
                    else:
                        date_str = str(date_value)
                else:
                    date_str = ''
                
                # Extract Total Time
                total_time_cell = worksheet.cell(row=row_idx, column=total_time_col)
                total_time_value = total_time_cell.value
                try:
                    if total_time_value is None:
                        total_time = ''
                    else:
                        total_time = str(total_time_value)
                except:
                    total_time = ''
                
                # Extract all tasks
                task_list = []
                for task_num in sorted(task_columns.keys()):
                    task_info = task_columns[task_num]
                    task_text = ''
                    client_text = ''
                    time_text = ''
                    
                    if 'task' in task_info:
                        task_cell = worksheet.cell(row=row_idx, column=task_info['task'])
                        if task_cell.value:
                            task_text = str(task_cell.value).strip()
                    
                    if 'client' in task_info:
                        client_cell = worksheet.cell(row=row_idx, column=task_info['client'])
                        if client_cell.value:
                            client_text = str(client_cell.value).strip()
                    
                    if 'time' in task_info:
                        time_cell = worksheet.cell(row=row_idx, column=task_info['time'])
                        if time_cell.value:
                            time_text = str(time_cell.value).strip()
                    
                    # Combine task info: "Task - Task Client Name - Time for Task"
                    if task_text or client_text or time_text:
                        task_combined = ' - '.join(filter(None, [task_text, client_text, time_text]))
                        if task_combined:
                            task_list.append(task_combined)
                
                # Combine all tasks with newlines
                task_combined_str = '\n'.join(task_list) if task_list else ''
                
                # Add row data
                all_rows.append({
                    'employee_name': employee_name_value,
                    'task': task_combined_str,
                    'date': date_str,
                    'total_time': total_time
                })
                
        except Exception as e:
            print(f"Error processing file {file_path}: {str(e)}")
            continue

    if not all_rows:
        flash(f'No data found for employee "{employee_name}" in {calendar.month_name[month]} {year}.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Create Excel workbook with personal activity tracker
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Personal Activity Tracker"
        
        # Header row
        headers = ['Sr. No', 'Employee Name', 'Task', 'Date', 'Total Time']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Center alignment with text wrap for all cells
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Data rows
        row_num = 2
        for sr_no, row_data in enumerate(all_rows, start=1):
            # Create all cells with border
            cell1 = ws.cell(row=row_num, column=1, value=sr_no)
            cell2 = ws.cell(row=row_num, column=2, value=row_data['employee_name'])
            cell3 = ws.cell(row=row_num, column=3, value=row_data['task'])
            cell4 = ws.cell(row=row_num, column=4, value=row_data['date'])
            cell5 = ws.cell(row=row_num, column=5, value=row_data['total_time'])
            
            # Apply border and alignment to all cells
            for cell in [cell1, cell2, cell3, cell4, cell5]:
                cell.border = border
                cell.alignment = center_alignment
            
            row_num += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # Set row heights for better visibility of wrapped text
        for row_idx in range(2, row_num):
            ws.row_dimensions[row_idx].height = 60
        
        # Save to temporary file
        month_name = calendar.month_name[month]
        safe_employee_name = re.sub(r'[^\w\s-]', '', employee_name).strip().replace(' ', '_')
        temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_excel_path = temp_excel.name
        temp_excel.close()
        
        wb.save(temp_excel_path)
        
        # Generate filename
        excel_filename = f"Personal_Activity_Tracker_{safe_employee_name}_{month_name}_{year}.xlsx"
        
        # Send the Excel file
        return send_file(
            temp_excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=excel_filename
        )
        
    except Exception as e:
        flash(f'Error generating personal activity tracker: {str(e)}', 'error')
        return redirect(url_for('hr_dashboard'))


@hr_dashboard_bp.route('/hr/generate_client_wise_activity_tracker', methods=['POST'])
@login_required
def generate_client_wise_activity_tracker():
    """Generate client-wise activity tracker Excel file for selected month and year"""
    if current_user.department != "HR":
        abort(403)

    month_str = request.form.get('month')
    year_str = request.form.get('year')

    if not month_str or not year_str:
        flash('Please select both month and year.', 'error')
        return redirect(url_for('hr_dashboard'))

    try:
        month = int(month_str)
        year = int(year_str)
        
        if month < 1 or month > 12:
            flash('Invalid month selected.', 'error')
            return redirect(url_for('hr_dashboard'))
            
        if year < 2020 or year > 2030:
            flash('Invalid year selected.', 'error')
            return redirect(url_for('hr_dashboard'))
    except ValueError:
        flash('Invalid month or year format.', 'error')
        return redirect(url_for('hr_dashboard'))

    tracker_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Everyday_Updated_Work')
    
    if not os.path.exists(tracker_dir):
        flash('Activity tracker directory not found.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Pattern to match: DD_Mon_YYYY.xlsx (e.g., 28_Oct_2025.xlsx)
    date_pattern = re.compile(r'^(\d{1,2})_([A-Za-z]{3})_(\d{4})\.xlsx$')
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
    }

    # Find all files matching the selected month and year
    matching_files = []
    for filename in os.listdir(tracker_dir):
        if filename.endswith('.xlsx'):
            match = date_pattern.match(filename)
            if match:
                try:
                    day = int(match.group(1))
                    month_str_file = match.group(2)
                    year_file = int(match.group(3))
                    
                    if month_str_file in month_map:
                        month_file = month_map[month_str_file]
                        
                        # Check if file matches selected month and year
                        if month_file == month and year_file == year:
                            file_path = os.path.join(tracker_dir, filename)
                            if os.path.exists(file_path):
                                matching_files.append((file_path, filename, day, month_str_file, year_file))
                except (ValueError, KeyError):
                    continue

    if not matching_files:
        month_name = calendar.month_name[month]
        flash(f'No activity tracker files found for {month_name} {year}.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Process all matching files and aggregate data by Client Name
    client_data = defaultdict(lambda: {'total_time': 0.0, 'details': [], 'dates': set()})
    
    for file_path, filename, day, month_str_file, year_file in matching_files:
        try:
            workbook = load_workbook(file_path, data_only=True)
            
            # Check if "Client_Time" worksheet exists
            if 'Client_Time' not in workbook.sheetnames:
                continue
            
            worksheet = workbook['Client_Time']
            
            # Format date from filename: 21_Nov_2025 -> 21 Nov 2025
            date_str = f"{day} {month_str_file} {year_file}"
            
            # Find header row (assuming headers are in row 1)
            header_row = 1
            
            # Find column indices
            client_name_col = None
            time_col = None
            details_col = None
            
            # Scan header row to find columns
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip()
                    cell_lower = cell_str.lower()
                    
                    # Client Name (Column B or header contains "client name")
                    if col_idx == 2 or 'client name' in cell_lower:
                        client_name_col = col_idx
                    
                    # Time (Column C or header contains "time")
                    if col_idx == 3 or ('time' in cell_lower and 'total' not in cell_lower):
                        time_col = col_idx
                    
                    # Details/Employees and Task (Column D or header contains "employees" or "task" or "details")
                    if col_idx == 4 or 'employees' in cell_lower or ('task' in cell_lower and 'client' not in cell_lower) or 'details' in cell_lower:
                        details_col = col_idx
            
            # Default to Column B, C, D if not found
            if client_name_col is None:
                client_name_col = 2
            if time_col is None:
                time_col = 3
            if details_col is None:
                details_col = 4
            
            # Process data rows (starting from row 2)
            for row_idx in range(header_row + 1, worksheet.max_row + 1):
                client_name_cell = worksheet.cell(row=row_idx, column=client_name_col)
                client_name_value = client_name_cell.value
                
                # Skip empty rows
                if not client_name_value:
                    continue
                
                client_name_value = str(client_name_value).strip()
                
                # Extract Time
                time_cell = worksheet.cell(row=row_idx, column=time_col)
                time_value = time_cell.value
                try:
                    if time_value is None:
                        time_val = 0.0
                    else:
                        time_val = float(time_value)
                except (ValueError, TypeError):
                    time_val = 0.0
                
                # Extract Details
                details_cell = worksheet.cell(row=row_idx, column=details_col)
                details_value = details_cell.value
                details_str = str(details_value).strip() if details_value else ''
                
                # Append date to details if details exist
                if details_str:
                    details_with_date = f"{details_str} - {date_str}"
                    client_data[client_name_value]['details'].append(details_with_date)
                
                # Aggregate data by client name
                client_data[client_name_value]['total_time'] += time_val
                client_data[client_name_value]['dates'].add(date_str)
                
        except Exception as e:
            print(f"Error processing file {file_path}: {str(e)}")
            continue

    if not client_data:
        flash('No client data found in the files for the selected month and year.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Create Excel workbook with client-wise activity tracker
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Client Wise Activity Tracker"
        
        # Header row
        headers = ['Sr. No', 'Client Name', 'Total Time', 'Date', 'Details']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Center alignment with text wrap for all cells
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Data rows - sort by client name
        row_num = 2
        for sr_no, (client_name, data) in enumerate(sorted(client_data.items()), start=1):
            # Combine all dates
            dates_str = ', '.join(sorted(data['dates']))
            
            # Combine all details with newlines
            details_str = '\n'.join(data['details']) if data['details'] else ''
            
            # Create all cells
            cell1 = ws.cell(row=row_num, column=1, value=sr_no)
            cell2 = ws.cell(row=row_num, column=2, value=client_name)
            cell3 = ws.cell(row=row_num, column=3, value=data['total_time'])
            cell4 = ws.cell(row=row_num, column=4, value=dates_str)
            cell5 = ws.cell(row=row_num, column=5, value=details_str)
            
            # Apply border and alignment to all cells
            for cell in [cell1, cell2, cell3, cell4, cell5]:
                cell.border = border
                cell.alignment = center_alignment
            
            row_num += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 50
        
        # Set row heights for better visibility of wrapped text
        for row_idx in range(2, row_num):
            ws.row_dimensions[row_idx].height = 60
        
        # Save to temporary file
        month_name = calendar.month_name[month]
        temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_excel_path = temp_excel.name
        temp_excel.close()
        
        wb.save(temp_excel_path)
        
        # Generate filename
        excel_filename = f"Client_Wise_Activity_Tracker_{month_name}_{year}.xlsx"
        
        # Send the Excel file
        return send_file(
            temp_excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=excel_filename
        )
        
    except Exception as e:
        flash(f'Error generating client-wise activity tracker: {str(e)}', 'error')
        return redirect(url_for('hr_dashboard'))


@hr_dashboard_bp.route('/hr/get_extra_work_entries', methods=['GET'])
@login_required
def get_extra_work_entries():
    """Get last 30 unapproved extra work entries"""
    if current_user.department != "HR":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    try:
        extra_work_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Extra_Work')
        approved_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Approved_Extra_work')
        
        if not os.path.exists(extra_work_dir):
            return jsonify({'success': True, 'entries': []})
        
        # Get current month and year
        current_date = datetime.now()
        current_month = current_date.month
        current_year = current_date.year
        
        month_map = {
            'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
            'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
        }
        
        month_names = {v: k for k, v in month_map.items()}
        
        # Pattern to match: Mon_YYYY.xlsx (e.g., Nov_2025.xlsx)
        file_pattern = re.compile(r'^([A-Za-z]{3})_(\d{4})\.xlsx$')
        
        # Collect all files with their month/year
        all_files = []
        for filename in os.listdir(extra_work_dir):
            if filename.endswith('.xlsx'):
                match = file_pattern.match(filename)
                if match:
                    month_str = match.group(1)
                    year = int(match.group(2))
                    if month_str in month_map:
                        month = month_map[month_str]
                        all_files.append((filename, month, year))
        
        # Sort files by year and month (newest first)
        all_files.sort(key=lambda x: (x[2], x[1]), reverse=True)
        
        # Load approved entries to filter them out
        approved_entries = set()
        if os.path.exists(approved_dir):
            for approved_file in os.listdir(approved_dir):
                if approved_file.endswith('.xlsx'):
                    try:
                        approved_path = os.path.join(approved_dir, approved_file)
                        wb = load_workbook(approved_path, data_only=True)
                        ws = wb.active
                        
                        # Read all rows and create a set of unique identifiers
                        # Approved file structure: Employee Name, Date, Task, Time, Status, Approved By
                        for row_idx in range(2, ws.max_row + 1):
                            employee_name = ws.cell(row=row_idx, column=1).value
                            date_val = ws.cell(row=row_idx, column=2).value
                            task_val = ws.cell(row=row_idx, column=3).value
                            time_val = ws.cell(row=row_idx, column=4).value
                            
                            if employee_name and date_val:
                                entry_key = f"{str(employee_name).strip()}|{str(date_val).strip()}|{str(task_val).strip() if task_val else ''}|{str(time_val).strip() if time_val else ''}"
                                approved_entries.add(entry_key)
                    except Exception as e:
                        print(f"Error reading approved file {approved_file}: {e}")
                        continue
        
        # Collect entries from files (starting from current month, going backwards)
        all_entries = []
        entries_needed = 30
        
        for filename, month, year in all_files:
            if entries_needed <= 0:
                break
            
            file_path = os.path.join(extra_work_dir, filename)
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                
                # Read rows from bottom to top (most recent first)
                # Source file structure: Employee Name, Employee Team, Date, Time, Task, Task Description, Client, Concerned Person, Timestamp
                for row_idx in range(ws.max_row, 1, -1):
                    if entries_needed <= 0:
                        break
                    
                    employee_name = ws.cell(row=row_idx, column=1).value
                    date_val = ws.cell(row=row_idx, column=3).value if ws.max_column >= 3 else None  # Column 3 is Date
                    time_val = ws.cell(row=row_idx, column=4).value if ws.max_column >= 4 else None  # Column 4 is Time
                    task_val = ws.cell(row=row_idx, column=5).value if ws.max_column >= 5 else None  # Column 5 is Task
                    
                    # Skip empty rows
                    if not employee_name:
                        continue
                    
                    # Get all column values for the entry
                    employee_team = ws.cell(row=row_idx, column=2).value if ws.max_column >= 2 else None
                    task_description = ws.cell(row=row_idx, column=6).value if ws.max_column >= 6 else None
                    client = ws.cell(row=row_idx, column=7).value if ws.max_column >= 7 else None
                    concerned_person = ws.cell(row=row_idx, column=8).value if ws.max_column >= 8 else None
                    
                    # Create entry key to check if already approved (using Employee Name, Date, Task, Time)
                    entry_key = f"{str(employee_name).strip()}|{str(date_val).strip() if date_val else ''}|{str(task_val).strip() if task_val else ''}|{str(time_val).strip() if time_val else ''}"
                    
                    # Skip if already approved
                    if entry_key in approved_entries:
                        continue
                    
                    all_entries.append({
                        'employee_name': str(employee_name).strip() if employee_name else '',
                        'date': str(date_val).strip() if date_val else '',
                        'task': str(task_val).strip() if task_val else '',
                        'time': str(time_val).strip() if time_val else '',
                        'employee_team': str(employee_team).strip() if employee_team else '',
                        'task_description': str(task_description).strip() if task_description else '',
                        'client': str(client).strip() if client else '',
                        'concerned_person': str(concerned_person).strip() if concerned_person else '',
                        'source_file': filename,
                        'row_index': row_idx
                    })
                    entries_needed -= 1
                    
            except Exception as e:
                print(f"Error reading file {filename}: {e}")
                continue
        
        # Return last 30 entries
        return jsonify({'success': True, 'entries': all_entries[:30]})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@hr_dashboard_bp.route('/hr/approve_decline_extra_work', methods=['POST'])
@login_required
def approve_decline_extra_work():
    """Approve or decline an extra work entry"""
    if current_user.department != "HR":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    try:
        data = request.get_json()
        source_file = data.get('source_file')
        row_index = data.get('row_index')
        status = data.get('status')  # 'approved' or 'declined'
        
        if not source_file or not row_index or not status:
            return jsonify({'success': False, 'message': 'Missing required parameters'}), 400
        
        extra_work_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Extra_Work')
        approved_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Approved_Extra_work')
        
        # Ensure approved directory exists
        os.makedirs(approved_dir, exist_ok=True)
        
        source_path = os.path.join(extra_work_dir, source_file)
        approved_path = os.path.join(approved_dir, source_file)
        
        if not os.path.exists(source_path):
            return jsonify({'success': False, 'message': 'Source file not found'}), 404
        
        # Read the entry from source file
        wb_source = load_workbook(source_path, data_only=True)
        ws_source = wb_source.active
        
        # Read columns according to source file structure:
        # Column 1: Employee Name
        # Column 2: Employee Team
        # Column 3: Date
        # Column 4: Time
        # Column 5: Task
        # Column 6: Task Description
        # Column 7: Client
        # Column 8: Concerned Person
        # Column 9: Timestamp
        
        employee_name = ws_source.cell(row=row_index, column=1).value
        employee_team = ws_source.cell(row=row_index, column=2).value if ws_source.max_column >= 2 else None
        date_val = ws_source.cell(row=row_index, column=3).value if ws_source.max_column >= 3 else None
        time_val = ws_source.cell(row=row_index, column=4).value if ws_source.max_column >= 4 else None
        task_val = ws_source.cell(row=row_index, column=5).value if ws_source.max_column >= 5 else None
        task_description = ws_source.cell(row=row_index, column=6).value if ws_source.max_column >= 6 else None
        client = ws_source.cell(row=row_index, column=7).value if ws_source.max_column >= 7 else None
        concerned_person = ws_source.cell(row=row_index, column=8).value if ws_source.max_column >= 8 else None
        
        # Load or create approved file
        if os.path.exists(approved_path):
            wb_approved = load_workbook(approved_path)
            ws_approved = wb_approved.active
        else:
            wb_approved = Workbook()
            ws_approved = wb_approved.active
            ws_approved.title = "Approved Extra Work"
            
            # Add headers
            headers = ['Employee Name', 'Date', 'Task', 'Time', 'Status', 'Approved By']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_approved.cell(row=1, column=col_idx, value=header)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add new row
        next_row = ws_approved.max_row + 1
        ws_approved.cell(row=next_row, column=1, value=employee_name)
        ws_approved.cell(row=next_row, column=2, value=date_val)
        ws_approved.cell(row=next_row, column=3, value=task_val)
        ws_approved.cell(row=next_row, column=4, value=time_val)
        ws_approved.cell(row=next_row, column=5, value=status.capitalize())
        ws_approved.cell(row=next_row, column=6, value=current_user.username)
        
        # Apply formatting
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx in range(1, 7):
            cell = ws_approved.cell(row=next_row, column=col_idx)
            cell.alignment = center_alignment
            cell.border = border
        
        # Auto-adjust column widths
        ws_approved.column_dimensions['A'].width = 25
        ws_approved.column_dimensions['B'].width = 15
        ws_approved.column_dimensions['C'].width = 40
        ws_approved.column_dimensions['D'].width = 15
        ws_approved.column_dimensions['E'].width = 15
        ws_approved.column_dimensions['F'].width = 20
        
        wb_approved.save(approved_path)
        
        return jsonify({'success': True, 'message': f'Entry {status} successfully'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@hr_dashboard_bp.route('/hr/download_approved_extra_work', methods=['POST'])
@login_required
def download_approved_extra_work():
    """Download approved extra work file for selected month and year"""
    if current_user.department != "HR":
        abort(403)

    month_str = request.form.get('month')
    year_str = request.form.get('year')

    if not month_str or not year_str:
        flash('Please select both month and year.', 'error')
        return redirect(url_for('hr_dashboard'))

    try:
        month = int(month_str)
        year = int(year_str)
        
        if month < 1 or month > 12:
            flash('Invalid month selected.', 'error')
            return redirect(url_for('hr_dashboard'))
            
        if year < 2020 or year > 2030:
            flash('Invalid year selected.', 'error')
            return redirect(url_for('hr_dashboard'))
    except ValueError:
        flash('Invalid month or year format.', 'error')
        return redirect(url_for('hr_dashboard'))

    approved_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Approved_Extra_work')
    
    if not os.path.exists(approved_dir):
        flash('Approved extra work directory not found.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Map month number to abbreviation
    month_map = {
        1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
        7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
    }
    
    month_abbr = month_map.get(month)
    if not month_abbr:
        flash('Invalid month selected.', 'error')
        return redirect(url_for('hr_dashboard'))
    
    # Construct filename: Nov_2025.xlsx
    filename = f"{month_abbr}_{year}.xlsx"
    file_path = os.path.join(approved_dir, filename)
    
    if not os.path.exists(file_path):
        month_name = calendar.month_name[month]
        flash(f'No approved extra work file found for {month_name} {year}.', 'error')
        return redirect(url_for('hr_dashboard'))
    
    return send_from_directory(approved_dir, filename, as_attachment=True)


@hr_dashboard_bp.route('/hr/generate_attendance_with_extra_work', methods=['POST'])
@login_required
def generate_attendance_with_extra_work():
    """Generate attendance record with extra work Excel file for selected month and year"""
    if current_user.department != "HR":
        abort(403)

    month_str = request.form.get('month')
    year_str = request.form.get('year')

    if not month_str or not year_str:
        flash('Please select both month and year.', 'error')
        return redirect(url_for('hr_dashboard'))

    try:
        month = int(month_str)
        year = int(year_str)
        
        if month < 1 or month > 12:
            flash('Invalid month selected.', 'error')
            return redirect(url_for('hr_dashboard'))
            
        if year < 2020 or year > 2030:
            flash('Invalid year selected.', 'error')
            return redirect(url_for('hr_dashboard'))
    except ValueError:
        flash('Invalid month or year format.', 'error')
        return redirect(url_for('hr_dashboard'))

    tracker_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Everyday_Updated_Work')
    approved_extra_work_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Approved_Extra_work')
    
    if not os.path.exists(tracker_dir):
        flash('Activity tracker directory not found.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Pattern to match: DD_Mon_YYYY.xlsx (e.g., 28_Oct_2025.xlsx)
    date_pattern = re.compile(r'^(\d{1,2})_([A-Za-z]{3})_(\d{4})\.xlsx$')
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
    }
    
    month_names = {v: k for k, v in month_map.items()}

    # Find all files matching the selected month and year
    matching_files = []
    for filename in os.listdir(tracker_dir):
        if filename.endswith('.xlsx'):
            match = date_pattern.match(filename)
            if match:
                try:
                    day = int(match.group(1))
                    month_str_file = match.group(2)
                    year_file = int(match.group(3))
                    
                    if month_str_file in month_map:
                        month_file = month_map[month_str_file]
                        
                        # Check if file matches selected month and year
                        if month_file == month and year_file == year:
                            file_path = os.path.join(tracker_dir, filename)
                            if os.path.exists(file_path):
                                matching_files.append(file_path)
                except (ValueError, KeyError):
                    continue

    if not matching_files:
        month_name = calendar.month_name[month]
        flash(f'No activity tracker files found for {month_name} {year}.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Process all matching files and calculate attendance (same logic as attendance record)
    employee_attendance = defaultdict(float)  # {employee_name: total_working_days}
    
    for file_path in matching_files:
        try:
            workbook = load_workbook(file_path, data_only=True)
            
            # Check if "Updated Work" worksheet exists
            if 'Updated Work' not in workbook.sheetnames:
                continue
            
            worksheet = workbook['Updated Work']
            
            # Find header row (assuming headers are in row 1)
            header_row = 1
            employee_name_col = None
            total_time_col = None
            
            # Find column indices for Employee Name (Column A) and Total Time (Column D)
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip().lower()
                    if 'employee name' in cell_str or col_idx == 1:
                        employee_name_col = col_idx
                    if 'total time' in cell_str or col_idx == 4:
                        total_time_col = col_idx
            
            # Default to Column A (1) and Column D (4) if not found
            if employee_name_col is None:
                employee_name_col = 1
            if total_time_col is None:
                total_time_col = 4
            
            # Process data rows (starting from row 2)
            for row_idx in range(header_row + 1, worksheet.max_row + 1):
                employee_name_cell = worksheet.cell(row=row_idx, column=employee_name_col)
                total_time_cell = worksheet.cell(row=row_idx, column=total_time_col)
                
                employee_name = employee_name_cell.value
                total_time_value = total_time_cell.value
                
                # Skip empty rows
                if not employee_name:
                    continue
                
                employee_name = str(employee_name).strip()
                
                # Convert total_time to float
                try:
                    if total_time_value is None:
                        total_time = 0.0
                    else:
                        total_time = float(total_time_value)
                except (ValueError, TypeError):
                    total_time = 0.0
                
                # Calculate working day based on Total Time
                # > 6 hours = 1 working day
                # 3-6 hours = 0.5 working day
                # < 3 hours = 0 working day
                if total_time > 6:
                    working_day = 1.0
                elif total_time >= 3:
                    working_day = 0.5
                else:
                    working_day = 0.0
                
                employee_attendance[employee_name] += working_day
                
        except Exception as e:
            print(f"Error processing file {file_path}: {str(e)}")
            continue

    # Process approved extra work files for the selected month
    employee_extra_work = defaultdict(float)  # {employee_name: extra_work_days}
    
    if os.path.exists(approved_extra_work_dir):
        month_abbr = month_names.get(month)
        if month_abbr:
            extra_work_filename = f"{month_abbr}_{year}.xlsx"
            extra_work_path = os.path.join(approved_extra_work_dir, extra_work_filename)
            
            if os.path.exists(extra_work_path):
                try:
                    wb_extra = load_workbook(extra_work_path, data_only=True)
                    ws_extra = wb_extra.active
                    
                    # Read approved extra work entries
                    # Structure: Employee Name (col 1), Date (col 2), Task (col 3), Time (col 4), Status (col 5), Approved By (col 6)
                    for row_idx in range(2, ws_extra.max_row + 1):
                        employee_name = ws_extra.cell(row=row_idx, column=1).value
                        time_val = ws_extra.cell(row=row_idx, column=4).value
                        status_val = ws_extra.cell(row=row_idx, column=5).value
                        
                        # Skip empty rows or declined entries
                        if not employee_name or not status_val:
                            continue
                        
                        # Only count approved entries
                        if str(status_val).strip().lower() != 'approved':
                            continue
                        
                        employee_name = str(employee_name).strip()
                        
                        # Convert time to float
                        try:
                            if time_val is None:
                                time_hours = 0.0
                            else:
                                time_hours = float(time_val)
                        except (ValueError, TypeError):
                            time_hours = 0.0
                        
                        # Convert time to working days:
                        # 0-4 hours = 0
                        # 4-8 hours (or exactly 4) = 0.5
                        # >8 hours (or exactly 8) = 1
                        if time_hours < 4:
                            extra_work_days = 0.0
                        elif time_hours >= 4 and time_hours < 8:
                            extra_work_days = 0.5
                        else:  # >= 8
                            extra_work_days = 1.0
                        
                        employee_extra_work[employee_name] += extra_work_days
                        
                except Exception as e:
                    print(f"Error processing extra work file {extra_work_path}: {str(e)}")

    if not employee_attendance:
        flash('No attendance data found in the files.', 'error')
        return redirect(url_for('hr_dashboard'))

    # Get total days in the selected month
    total_days_in_month = calendar.monthrange(year, month)[1]

    # Create Excel workbook with attendance summary including extra work
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance with Extra Work"
        
        # Header row
        headers = ['Sr. No', 'Employee Name', 'Working Day', 'Non-Working Day', 'Extra Work', 'Total Working Days']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Center alignment with text wrap for all cells
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Data rows - combine attendance and extra work data
        all_employees = set(employee_attendance.keys()) | set(employee_extra_work.keys())
        
        row_num = 2
        for sr_no, employee_name in enumerate(sorted(all_employees), start=1):
            working_days = employee_attendance.get(employee_name, 0.0)
            extra_work_days = employee_extra_work.get(employee_name, 0.0)
            non_working_days = total_days_in_month - working_days
            total_working_days = working_days + extra_work_days
            
            # Create all cells
            cell1 = ws.cell(row=row_num, column=1, value=sr_no)
            cell2 = ws.cell(row=row_num, column=2, value=employee_name)
            cell3 = ws.cell(row=row_num, column=3, value=working_days)
            cell4 = ws.cell(row=row_num, column=4, value=non_working_days)
            cell5 = ws.cell(row=row_num, column=5, value=extra_work_days)
            cell6 = ws.cell(row=row_num, column=6, value=total_working_days)
            
            # Apply border and alignment to all cells
            for cell in [cell1, cell2, cell3, cell4, cell5, cell6]:
                cell.border = border
                cell.alignment = center_alignment
            
            row_num += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        
        # Save to temporary file
        month_name = calendar.month_name[month]
        temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_excel_path = temp_excel.name
        temp_excel.close()
        
        wb.save(temp_excel_path)
        
        # Generate filename
        excel_filename = f"Attendance_Report_with_Extra_Work_{month_name}_{year}.xlsx"
        
        # Send the Excel file
        return send_file(
            temp_excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=excel_filename
        )
        
    except Exception as e:
        flash(f'Error generating attendance report with extra work: {str(e)}', 'error')
        return redirect(url_for('hr_dashboard'))


@hr_dashboard_bp.route('/hr/get_client_mail_list', methods=['GET'])
@login_required
def get_client_mail_list():
    """Get list of all client mail addresses"""
    if current_user.department != "HR":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Import inside function to avoid circular import
        # Use current_app to ensure we're using the db bound to the current app
        from app import ClientMail
        # Access db through current_app's extensions
        db = current_app.extensions['sqlalchemy']
        client_mails = db.session.query(ClientMail).order_by(ClientMail.email).all()
        emails_list = [{'id': mail.id, 'email': mail.email} for mail in client_mails]
        return jsonify({'success': True, 'emails': emails_list})
    except KeyError:
        # Fallback to direct import if extensions don't work
        from app import db, ClientMail
        client_mails = ClientMail.query.order_by(ClientMail.email).all()
        emails_list = [{'id': mail.id, 'email': mail.email} for mail in client_mails]
        return jsonify({'success': True, 'emails': emails_list})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@hr_dashboard_bp.route('/hr/add_client_mail', methods=['POST'])
@login_required
def add_client_mail():
    """Add a new client mail address"""
    if current_user.department != "HR":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Import inside function to avoid circular import
        from app import ClientMail
        # Access db through current_app's extensions
        try:
            db = current_app.extensions['sqlalchemy']
        except KeyError:
            from app import db
        
        data = request.get_json()
        email = data.get('email', '').strip()
        
        if not email:
            return jsonify({'success': False, 'message': 'Email address is required'}), 400
        
        # Check if email already exists
        try:
            existing = db.session.query(ClientMail).filter_by(email=email).first()
        except:
            existing = ClientMail.query.filter_by(email=email).first()
        if existing:
            return jsonify({'success': False, 'message': 'Email address already exists'}), 400
        
        # Add new client mail
        new_mail = ClientMail(email=email)
        db.session.add(new_mail)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Email address added successfully'})
    
    except Exception as e:
        try:
            db = current_app.extensions.get('sqlalchemy')
            if db:
                db.session.rollback()
        except:
            try:
                from app import db
                db.session.rollback()
            except:
                pass
        return jsonify({'success': False, 'message': str(e)}), 500


@hr_dashboard_bp.route('/hr/remove_client_mail', methods=['POST'])
@login_required
def remove_client_mail():
    """Remove a client mail address"""
    if current_user.department != "HR":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Import inside function to avoid circular import
        from app import ClientMail
        # Access db through current_app's extensions
        try:
            db = current_app.extensions['sqlalchemy']
        except KeyError:
            from app import db
        
        data = request.get_json()
        mail_id = data.get('mail_id')
        
        if not mail_id:
            return jsonify({'success': False, 'message': 'Mail ID is required'}), 400
        
        try:
            client_mail = db.session.query(ClientMail).get(mail_id)
        except:
            client_mail = ClientMail.query.get(mail_id)
        if not client_mail:
            return jsonify({'success': False, 'message': 'Email address not found'}), 404
        
        db.session.delete(client_mail)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Email address removed successfully'})
    
    except Exception as e:
        try:
            db = current_app.extensions.get('sqlalchemy')
            if db:
                db.session.rollback()
        except:
            try:
                from app import db
                db.session.rollback()
            except:
                pass
        return jsonify({'success': False, 'message': str(e)}), 500


@hr_dashboard_bp.route('/hr/send_client_mail', methods=['POST'])
@login_required
def send_client_mail():
    """Send email to all client mail addresses"""
    if current_user.department != "HR":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from app import ClientMail, mail
        from flask_mail import Message
        import tempfile
        import os
        
        # Get form data and sanitize to prevent email header injection
        subject = sanitize_email_header(request.form.get('subject', '').strip())
        content = sanitize_email_content(request.form.get('content', '').strip())
        attachments = request.files.getlist('attachments')
        
        if not subject:
            return jsonify({'success': False, 'message': 'Subject is required'}), 400
        
        if not content:
            return jsonify({'success': False, 'message': 'Content is required'}), 400
        
        # Get all client mail addresses - use current_app context
        try:
            db = current_app.extensions['sqlalchemy']
            client_mails = db.session.query(ClientMail).all()
        except (KeyError, AttributeError):
            from app import db
            client_mails = ClientMail.query.all()
        
        if not client_mails:
            return jsonify({'success': False, 'message': 'No client mail addresses found. Please add client emails first.'}), 400
        
        # Store temporary attachment files
        temp_files = []
        sent_count = 0
        failed_count = 0
        
        try:
            # Save attachments to temporary files if any (with path traversal protection)
            attachment_paths = []
            if attachments and any(f.filename for f in attachments):
                temp_dir = tempfile.mkdtemp()
                for attachment in attachments:
                    if attachment.filename:
                        # Sanitize filename to prevent path traversal attacks
                        safe_filename = sanitize_filename(attachment.filename)
                        temp_path = os.path.join(temp_dir, safe_filename)
                        attachment.save(temp_path)
                        attachment_paths.append(temp_path)
                        temp_files.append(temp_path)
                temp_files.append(temp_dir)  # Add directory to cleanup list
            
            # Send email to each client separately
            for client_mail in client_mails:
                try:
                    msg = Message(
                        subject=subject,
                        recipients=[client_mail.email],
                        html=content
                    )
                    
                    # Attach files if any (with sanitized filenames)
                    for att_path in attachment_paths:
                        with open(att_path, 'rb') as f:
                            # Sanitize attachment filename to prevent path traversal
                            safe_attach_name = sanitize_filename(os.path.basename(att_path))
                            msg.attach(
                                safe_attach_name,
                                'application/octet-stream',
                                f.read()
                            )
                    
                    mail.send(msg)
                    sent_count += 1
                except Exception as e:
                    print(f"Failed to send email to {client_mail.email}: {str(e)}")
                    failed_count += 1
                    continue
            
            # Cleanup temporary files
            for temp_file in temp_files:
                try:
                    if os.path.isdir(temp_file):
                        import shutil
                        shutil.rmtree(temp_file)
                    elif os.path.isfile(temp_file):
                        os.remove(temp_file)
                except:
                    pass
            
            if sent_count > 0:
                return jsonify({
                    'success': True,
                    'message': f'Emails sent successfully',
                    'sent_count': sent_count,
                    'failed_count': failed_count
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'Failed to send emails to all clients'
                }), 500
                
        except Exception as e:
            # Cleanup on error
            for temp_file in temp_files:
                try:
                    if os.path.isdir(temp_file):
                        import shutil
                        shutil.rmtree(temp_file)
                    elif os.path.isfile(temp_file):
                        os.remove(temp_file)
                except:
                    pass
            raise e
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@hr_dashboard_bp.route('/hr/send_employee_mail', methods=['POST'])
@login_required
def send_employee_mail():
    """Send email to all employee email addresses (used for employee ID and 2FA)"""
    if current_user.department != "HR":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from app import User, mail
        from flask_mail import Message
        import tempfile
        import os
        
        # Get form data and sanitize to prevent email header injection
        subject = sanitize_email_header(request.form.get('subject', '').strip())
        content = sanitize_email_content(request.form.get('content', '').strip())
        attachments = request.files.getlist('attachments')
        
        if not subject:
            return jsonify({'success': False, 'message': 'Subject is required'}), 400
        
        if not content:
            return jsonify({'success': False, 'message': 'Content is required'}), 400
        
        # Get all active employee email addresses (exclude deleted and inactive) - use current_app context
        try:
            db = current_app.extensions['sqlalchemy']
            from app import UserStatus
            # Get all users first, then filter for active ones
            all_users = db.session.query(User).all()
        except (KeyError, AttributeError):
            from app import db, UserStatus
            # Get all users first, then filter for active ones
            all_users = db.session.query(User).all()
        
        # Filter for active employees only:
        # 1. Not deleted (deleted_at is None)
        # 2. Active status (status doesn't exist OR status.is_active is True)
        active_employees = [
            emp for emp in all_users 
            if not emp.deleted_at and (not emp.status or emp.status.is_active)
        ]
        
        if not active_employees:
            return jsonify({'success': False, 'message': 'No active employee email addresses found.'}), 400
        
        # Store temporary attachment files
        temp_files = []
        sent_count = 0
        failed_count = 0
        
        try:
            # Save attachments to temporary files if any (with path traversal protection)
            attachment_paths = []
            if attachments and any(f.filename for f in attachments):
                temp_dir = tempfile.mkdtemp()
                for attachment in attachments:
                    if attachment.filename:
                        # Sanitize filename to prevent path traversal attacks
                        safe_filename = sanitize_filename(attachment.filename)
                        temp_path = os.path.join(temp_dir, safe_filename)
                        attachment.save(temp_path)
                        attachment_paths.append(temp_path)
                        temp_files.append(temp_path)
                temp_files.append(temp_dir)  # Add directory to cleanup list
            
            # Send email to each active employee separately
            for employee in active_employees:
                try:
                    msg = Message(
                        subject=subject,
                        recipients=[employee.email],
                        html=content
                    )
                    
                    # Attach files if any (with sanitized filenames)
                    for att_path in attachment_paths:
                        with open(att_path, 'rb') as f:
                            # Sanitize attachment filename to prevent path traversal
                            safe_attach_name = sanitize_filename(os.path.basename(att_path))
                            msg.attach(
                                safe_attach_name,
                                'application/octet-stream',
                                f.read()
                            )
                    
                    mail.send(msg)
                    sent_count += 1
                except Exception as e:
                    print(f"Failed to send email to {employee.email}: {str(e)}")
                    failed_count += 1
                    continue
            
            # Cleanup temporary files
            for temp_file in temp_files:
                try:
                    if os.path.isdir(temp_file):
                        import shutil
                        shutil.rmtree(temp_file)
                    elif os.path.isfile(temp_file):
                        os.remove(temp_file)
                except:
                    pass
            
            if sent_count > 0:
                return jsonify({
                    'success': True,
                    'message': f'Emails sent successfully',
                    'sent_count': sent_count,
                    'failed_count': failed_count
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'Failed to send emails to all employees'
                }), 500
                
        except Exception as e:
            # Cleanup on error
            for temp_file in temp_files:
                try:
                    if os.path.isdir(temp_file):
                        import shutil
                        shutil.rmtree(temp_file)
                    elif os.path.isfile(temp_file):
                        os.remove(temp_file)
                except:
                    pass
            raise e
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@hr_dashboard_bp.route('/hr/download_catalog/<catalog_type>', methods=['GET'])
@login_required
def download_catalog(catalog_type):
    """Download catalog file based on catalog type"""
    if current_user.department != "HR":
        abort(403)
    
    # Map catalog types to their file names
    catalog_files = {
        'infrastructure': 'Infrastructure VAPT Catalog.xlsx',
        'public_ip': 'Public IP VAPT Catalog.xlsx',
        'website': 'Website VAPT Catalog.xlsx'
    }
    
    if catalog_type not in catalog_files:
        abort(404)
    
    filename = catalog_files[catalog_type]
    catalog_path = os.path.join('static', 'Formats_and_Catalog', filename)
    
    # Check if file exists
    if not os.path.exists(catalog_path):
        flash(f'Catalog file not found: {filename}', 'error')
        return redirect(url_for('hr_dashboard'))
    
    try:
        return send_file(
            catalog_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'Error downloading catalog: {str(e)}', 'error')
        return redirect(url_for('hr_dashboard'))

@hr_dashboard_bp.route('/hr/get_all_employees', methods=['GET'])
@login_required
def get_all_employees():
    """Get all active (non-deleted) employees for performance rating"""
    if current_user.department != "HR":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from app import User, UserStatus
        try:
            db = current_app.extensions['sqlalchemy']
            users = db.session.query(User).all()
        except (KeyError, AttributeError):
            from app import db
            users = User.query.all()
        
        employees = []
        for user in users:
            # Check if user is deleted
            is_deleted = False
            if hasattr(user, 'deleted_at') and user.deleted_at:
                is_deleted = True
            
            # Check if user is inactive
            is_inactive = False
            if user.status and not user.status.is_active:
                is_inactive = True
            
            # Only include users that are not deleted and are active
            if not is_deleted and not is_inactive:
                employees.append({
                    'id': user.id,
                    'username': user.username,
                    'employee_name': user.employee_name,
                    'email': user.email
                })
        
        # Sort by employee name for better UX
        employees.sort(key=lambda x: x.get('employee_name', '') or x.get('username', ''))
        
        return jsonify({'success': True, 'employees': employees})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@hr_dashboard_bp.route('/hr/save_performance_data', methods=['POST'])
@login_required
def save_performance_data():
    """Save performance data for all employees"""
    if current_user.department != "HR":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from app import User, Performance
        from datetime import datetime
        import calendar
        
        # Get db instance using current_app context
        try:
            db = current_app.extensions['sqlalchemy']
        except (KeyError, AttributeError):
            from app import db
        
        metric_type = request.form.get('metric_type', '').strip()
        
        if not metric_type:
            return jsonify({'success': False, 'message': 'Metric type is required'}), 400
        
        # Get current month and year
        current_date = datetime.now()
        current_month = current_date.month
        current_year = current_date.year
        
        # Get all employee ratings from form
        employee_ratings = {}
        for key, value in request.form.items():
            if key.startswith('employee_'):
                try:
                    employee_id_str = key.replace('employee_', '')
                    # Validate employee_id with type safety
                    from app import validate_type_safe
                    id_valid, employee_id, id_error = validate_type_safe(employee_id_str, int, min_value=1)
                    if not id_valid:
                        return jsonify({'success': False, 'message': f'Invalid employee ID: {id_error}'}), 400
                    
                    # Validate rating with type safety
                    rating_valid, rating, rating_error = validate_type_safe(value, float, min_value=1.0, max_value=10.0)
                    if not rating_valid:
                        return jsonify({'success': False, 'message': f'Rating for employee {employee_id} must be between 1 and 10'}), 400
                    
                    employee_ratings[employee_id] = rating
                except Exception as e:
                    return jsonify({'success': False, 'message': f'Error processing rating: {str(e)}'}), 400
        
        if not employee_ratings:
            return jsonify({'success': False, 'message': 'No ratings provided'}), 400
        
        # Save performance data for each employee
        saved_count = 0
        for employee_id, rating in employee_ratings.items():
            # Get or create performance record
            performance = db.session.query(Performance).filter_by(
                user_id=employee_id,
                month=current_month,
                year=current_year
            ).first()
            
            if not performance:
                performance = Performance(
                    user_id=employee_id,
                    month=current_month,
                    year=current_year
                )
                db.session.add(performance)
            
            # Update the specific metric
            if metric_type == 'punctuality':
                performance.punctuality = rating
            elif metric_type == 'behaviour':
                performance.behaviour = rating
            elif metric_type == 'team_coordination':
                performance.team_coordination = rating
            elif metric_type == 'communication_skills':
                performance.communication_skills = rating
            else:
                return jsonify({'success': False, 'message': 'Invalid metric type'}), 400
            
            saved_count += 1
        
        db.session.commit()
        
        # Update Excel file
        try:
            update_performance_excel(current_month, current_year, metric_type)
        except Exception as e:
            print(f"Warning: Could not update Excel file: {e}")
            # Don't fail the request if Excel update fails
        
        metric_name = metric_type.replace('_', ' ').title()
        return jsonify({
            'success': True,
            'message': f'{metric_name} ratings saved successfully for {saved_count} employee(s)!'
        })
        
    except Exception as e:
        try:
            db.session.rollback()
        except:
            pass
        print(f"Error saving performance data: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error saving performance data: {str(e)}'}), 500

def update_performance_excel(month, year, metric_type):
    """Update Performance_data.xlsx with new performance data"""
    try:
        import os
        from openpyxl import load_workbook, Workbook
        from app import User, Performance
        
        # Get db instance using current_app context
        try:
            db = current_app.extensions['sqlalchemy']
        except (KeyError, AttributeError):
            from app import db
        
        excel_path = os.path.join('static', 'Performance', 'Performance_data.xlsx')
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        
        # Worksheet name format: "MM_YYYY" (e.g., "11_2025")
        worksheet_name = f"{month:02d}_{year}"
        
        # Load or create workbook
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # Get or create worksheet
        if worksheet_name in wb.sheetnames:
            ws = wb[worksheet_name]
        else:
            ws = wb.create_sheet(worksheet_name)
            # Add headers
            ws['A1'] = 'Employee Name'
            ws['B1'] = 'Punctuality'
            ws['C1'] = 'Behaviour'
            ws['D1'] = 'Team Coordination'
            ws['E1'] = 'Communication Skills'
            
            # Style headers
            from openpyxl.styles import Font, PatternFill
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
        
        # Get all users
        users = db.session.query(User).all()
        
        # Get performance data for current month
        performances = db.session.query(Performance).filter_by(month=month, year=year).all()
        perf_dict = {p.user_id: p for p in performances}
        
        # Clear existing data (keep headers)
        ws.delete_rows(2, ws.max_row)
        
        # Add employee data
        row = 2
        for user in users:
            ws.cell(row=row, column=1, value=user.employee_name or user.username)
            
            perf = perf_dict.get(user.id)
            if perf:
                ws.cell(row=row, column=2, value=perf.punctuality or 0)
                ws.cell(row=row, column=3, value=perf.behaviour or 0)
                ws.cell(row=row, column=4, value=perf.team_coordination or 0)
                ws.cell(row=row, column=5, value=perf.communication_skills or 0)
            else:
                ws.cell(row=row, column=2, value=0)
                ws.cell(row=row, column=3, value=0)
                ws.cell(row=row, column=4, value=0)
                ws.cell(row=row, column=5, value=0)
            
            row += 1
        
        # Save workbook
        wb.save(excel_path)
        print(f"✅ Excel file updated: {excel_path}, Worksheet: {worksheet_name}")
        
    except Exception as e:
        print(f"❌ Error updating Excel file: {e}")
        import traceback
        traceback.print_exc()
        raise e

@hr_dashboard_bp.route('/hr/upload_catalog', methods=['POST'])
@login_required
def upload_catalog():
    """Upload and replace catalog file"""
    if current_user.department != "HR":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Get form data
        catalog_type = request.form.get('catalog_type', '').strip()
        catalog_file = request.files.get('catalog_file')
        
        if not catalog_type:
            return jsonify({'success': False, 'message': 'Catalog type is required'}), 400
        
        if not catalog_file or catalog_file.filename == '':
            return jsonify({'success': False, 'message': 'Please select a file to upload'}), 400
        
        # Map catalog types to their file names
        catalog_files = {
            'infrastructure': 'Infrastructure VAPT Catalog.xlsx',
            'public_ip': 'Public IP VAPT Catalog.xlsx',
            'website': 'Website VAPT Catalog.xlsx'
        }
        
        if catalog_type not in catalog_files:
            return jsonify({'success': False, 'message': 'Invalid catalog type'}), 400
        
        # Get target filename
        target_filename = catalog_files[catalog_type]
        catalog_dir = os.path.join('static', 'Formats_and_Catalog')
        target_path = os.path.join(catalog_dir, target_filename)
        
        # Ensure directory exists
        os.makedirs(catalog_dir, exist_ok=True)
        
        # Validate file extension
        allowed_extensions = ['.xlsx', '.xls']
        file_extension = os.path.splitext(catalog_file.filename)[1].lower()
        
        if file_extension not in allowed_extensions:
            return jsonify({'success': False, 'message': 'Please upload an Excel file (.xlsx or .xls)'}), 400
        
        # Remove existing file if it exists
        if os.path.exists(target_path):
            try:
                os.remove(target_path)
            except Exception as e:
                print(f"Warning: Could not remove existing file: {e}")
        
        # Validate file content before saving
        from file_upload_utils import validate_file_content, validate_file_size
        
        # Validate file size (max 50MB for catalog files)
        size_valid, size_msg, file_size = validate_file_size(catalog_file, max_size_mb=50)
        if not size_valid:
            return jsonify({'success': False, 'message': size_msg}), 400
        
        # Validate file content
        content_valid, content_msg = validate_file_content(catalog_file, {'.xlsx', '.xls'})
        if not content_valid:
            return jsonify({'success': False, 'message': content_msg}), 400
        
        # Sanitize target path to prevent path traversal
        from security_utils import sanitize_path
        target_filename_safe = sanitize_path(target_filename)
        target_path = os.path.join(catalog_dir, target_filename_safe)
        
        # Ensure path is within catalog directory (prevent path traversal)
        catalog_dir_abs = os.path.abspath(catalog_dir)
        target_path_abs = os.path.abspath(target_path)
        if not target_path_abs.startswith(catalog_dir_abs):
            return jsonify({'success': False, 'message': 'Invalid file path detected'}), 400
        
        # Save the uploaded file with the correct name
        catalog_file.save(target_path)
        
        return jsonify({
            'success': True,
            'message': f'{target_filename} uploaded and replaced successfully!'
        })
        
    except Exception as e:
        print(f"Error uploading catalog: {str(e)}")
        return jsonify({'success': False, 'message': f'Error uploading catalog: {str(e)}'}), 500