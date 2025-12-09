from flask import Blueprint, request, jsonify, send_file
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from copy import copy

# Create blueprint for GRC IS Audit Compliance
grc_is_audit_compliance_bp = Blueprint('grc_is_audit_compliance', __name__)

def create_is_audit_compliance_excel(form_data, excel_file_path):
    """Create IS Audit Compliance Excel file"""
    try:
        print(f"🔍 DEBUG: Starting IS Audit Compliance Excel creation...")
        
        # Get form data
        org_name = form_data.get('organizationName', '')
        if org_name == 'Other':
            org_name = form_data.get('organizationNameOther', '')
        
        report_id = form_data.get('reportId', '')
        report_date = form_data.get('reportDate', '')
        compliance_date = form_data.get('complianceDate', '')
        
        # Format dates
        try:
            report_date_obj = datetime.strptime(report_date, '%Y-%m-%d')
            report_date_formatted = report_date_obj.strftime('%d/%m/%Y')
        except:
            report_date_formatted = report_date
        
        try:
            compliance_date_obj = datetime.strptime(compliance_date, '%Y-%m-%d')
            compliance_date_formatted = compliance_date_obj.strftime('%d/%m/%Y')
        except:
            compliance_date_formatted = compliance_date
        
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "IS Audit - Head Office"
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 60
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 30
        
        # Define styles
        header_font = Font(name='Times New Roman', size=12, bold=False)
        header_alignment_center = Alignment(horizontal='center', vertical='center')
        header_alignment_left = Alignment(horizontal='left', vertical='center')
        
        column_header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
        column_header_fill = PatternFill(start_color='FF1F497D', end_color='FF1F497D', fill_type='solid')
        column_header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='FF000000'),
            right=Side(style='thin', color='FF000000'),
            top=Side(style='thin', color='FF000000'),
            bottom=Side(style='thin', color='FF000000')
        )
        
        # Row 1
        ws['A1'] = "Submitted To:"
        ws['A1'].font = header_font
        ws['A1'].alignment = header_alignment_center
        ws['A1'].border = thin_border
        
        ws.merge_cells('B1:I1')
        ws['B1'] = f"The Management, {org_name}"
        ws['B1'].font = header_font
        ws['B1'].alignment = header_alignment_left
        # Apply border to all merged cells in B1:I1
        for col in range(2, 10):  # B to I (columns 2-9)
            ws.cell(row=1, column=col).border = thin_border
        
        # Row 2
        ws['A2'] = "Report Name:"
        ws['A2'].font = header_font
        ws['A2'].alignment = header_alignment_center
        ws['A2'].border = thin_border
        
        ws.merge_cells('B2:I2')
        ws['B2'] = f"Compliance Of Information System Audit Report (Report ID no. {report_id} dated {report_date_formatted})"
        ws['B2'].font = header_font
        ws['B2'].alignment = header_alignment_left
        # Apply border to all merged cells in B2:I2
        for col in range(2, 10):
            ws.cell(row=2, column=col).border = thin_border
        
        # Row 3
        ws['A3'] = "Confidentiality:"
        ws['A3'].font = header_font
        ws['A3'].alignment = header_alignment_center
        ws['A3'].border = thin_border
        
        ws.merge_cells('B3:I3')
        ws['B3'] = "Very High & Not for circulation"
        ws['B3'].font = header_font
        ws['B3'].alignment = header_alignment_left
        # Apply border to all merged cells in B3:I3
        for col in range(2, 10):
            ws.cell(row=3, column=col).border = thin_border
        
        # Row 4
        ws['A4'] = "Compliance Date:"
        ws['A4'].font = header_font
        ws['A4'].alignment = header_alignment_center
        ws['A4'].border = thin_border
        
        ws.merge_cells('B4:I4')
        ws['B4'] = compliance_date_formatted
        ws['B4'].font = header_font
        ws['B4'].alignment = header_alignment_left
        # Apply border to all merged cells in B4:I4
        for col in range(2, 10):
            ws.cell(row=4, column=col).border = thin_border
        
        # Row 5 - Column headers
        headers = [
            "Sr.No.",
            "Asset Reviewed",
            "Sub Serial Number",
            "Reference Page No.",
            "Observation",
            "Risk Factor",
            "Compliance By Bank",
            "Status",
            "Pl attach evidence"
        ]
        
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = column_header_font
            cell.fill = column_header_fill
            cell.alignment = column_header_alignment
            cell.border = thin_border
        
        # Read data from uploaded Excel file
        print(f"🔍 DEBUG: Reading Excel file: {excel_file_path}")
        source_wb = load_workbook(excel_file_path)
        
        current_row = 6
        worksheet_number = 1
        
        # Process each worksheet
        for sheet_name in source_wb.sheetnames:
            source_ws = source_wb[sheet_name]
            
            print(f"🔍 DEBUG: Processing worksheet: '{sheet_name}' (max_row={source_ws.max_row})")
            
            # Skip completely empty sheets
            if source_ws.max_row < 1:
                print(f"   ⚠️  Skipping completely empty worksheet")
                continue
            
            # Find data rows (assuming row 1 is header, check from row 2 onwards)
            data_rows = []
            if source_ws.max_row >= 2:
                for row_idx in range(2, source_ws.max_row + 1):
                    # Check if observation column (D) has data
                    observation_cell = source_ws.cell(row=row_idx, column=4)
                    observation_value = str(observation_cell.value).strip() if observation_cell.value else ''
                    
                    if observation_value:
                        risk_cell = source_ws.cell(row=row_idx, column=5)
                        risk_value = str(risk_cell.value).strip() if risk_cell.value else ''
                        
                        print(f"   📝 Row {row_idx}: Found observation='{observation_value[:50]}...', risk='{risk_value}'")
                        
                        data_rows.append({
                            'observation': observation_value,
                            'risk': risk_value
                        })
            
            print(f"   📊 Total data rows found: {len(data_rows)}")
            
            # If no data rows found, add a default compliance row
            if len(data_rows) == 0:
                print(f"   ⚠️  No data in worksheet '{sheet_name}', adding default compliance row")
                
                row_num = current_row
                ws.row_dimensions[row_num].height = 30
                
                # A column - Worksheet number (bold)
                ws.cell(row=row_num, column=1).value = worksheet_number
                ws.cell(row=row_num, column=1).font = Font(name='Times New Roman', size=12, bold=True)
                ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_num, column=1).border = thin_border
                
                # B column - Worksheet name (bold)
                ws.cell(row=row_num, column=2).value = sheet_name
                ws.cell(row=row_num, column=2).font = Font(name='Times New Roman', size=12, bold=True)
                ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row=row_num, column=2).border = thin_border
                
                # C column - Sub Serial Number
                ws.cell(row=row_num, column=3).value = f"{worksheet_number}.1"
                ws.cell(row=row_num, column=3).font = Font(name='Times New Roman', size=12)
                ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_num, column=3).border = thin_border
                
                # D column - Page number (will be filled by post-processing)
                ws.cell(row=row_num, column=4).value = ""  # Empty for now, post-processing will fill it
                ws.cell(row=row_num, column=4).font = Font(name='Times New Roman', size=12)
                ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_num, column=4).border = thin_border
                
                # E column - Compliance message
                ws.cell(row=row_num, column=5).value = "As per Auditor's Observation all Points are Compliance"
                ws.cell(row=row_num, column=5).font = Font(name='Times New Roman', size=12)
                ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws.cell(row=row_num, column=5).border = thin_border
                
                # F column - "Good Point" with blue background
                ws.cell(row=row_num, column=6).value = "Good Point"
                ws.cell(row=row_num, column=6).font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFFFF')
                ws.cell(row=row_num, column=6).fill = PatternFill(start_color='FF00B0F0', end_color='FF00B0F0', fill_type='solid')  # Blue
                ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_num, column=6).border = thin_border
                
                # G, H, I columns - "-" centered
                for col in [7, 8, 9]:
                    ws.cell(row=row_num, column=col).value = "-"
                    ws.cell(row=row_num, column=col).font = Font(name='Times New Roman', size=12)
                    ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=row_num, column=col).border = thin_border
                
                current_row += 1
                worksheet_number += 1
                continue
            
            # Calculate the merge range for Sr.No. and Asset Reviewed
            start_row = current_row
            end_row = current_row + len(data_rows) - 1
            
            # Merge A column (Sr.No.)
            if len(data_rows) > 1:
                ws.merge_cells(f'A{start_row}:A{end_row}')
                # Apply border to all merged cells
                for merge_row in range(start_row, end_row + 1):
                    ws.cell(row=merge_row, column=1).border = thin_border
            ws.cell(row=start_row, column=1).value = worksheet_number
            ws.cell(row=start_row, column=1).font = Font(name='Times New Roman', size=12, bold=True)
            ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=start_row, column=1).border = thin_border
            
            # Merge B column (Asset Reviewed)
            if len(data_rows) > 1:
                ws.merge_cells(f'B{start_row}:B{end_row}')
                # Apply border to all merged cells
                for merge_row in range(start_row, end_row + 1):
                    ws.cell(row=merge_row, column=2).border = thin_border
            ws.cell(row=start_row, column=2).value = sheet_name
            ws.cell(row=start_row, column=2).font = Font(name='Times New Roman', size=12, bold=True)
            ws.cell(row=start_row, column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=start_row, column=2).border = thin_border
            
            # Add data rows
            for idx, data_row in enumerate(data_rows):
                row_num = current_row + idx
                
                # Set row height to 30 pixels
                ws.row_dimensions[row_num].height = 30
                
                # C column - Sub Serial Number
                sub_serial = f"{worksheet_number}.{idx + 1}"
                ws.cell(row=row_num, column=3).value = sub_serial
                ws.cell(row=row_num, column=3).font = Font(name='Times New Roman', size=12)
                ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_num, column=3).border = thin_border
                
                # D column - Reference Page No. (left empty for manual entry)
                ws.cell(row=row_num, column=4).value = ""
                ws.cell(row=row_num, column=4).font = Font(name='Times New Roman', size=12)
                ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_num, column=4).border = thin_border
                
                # E column - Observation
                ws.cell(row=row_num, column=5).value = data_row['observation']
                ws.cell(row=row_num, column=5).font = Font(name='Times New Roman', size=12)
                ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws.cell(row=row_num, column=5).border = thin_border
                
                # F column - Risk Factor with conditional formatting
                risk_value = data_row['risk']
                risk_cell = ws.cell(row=row_num, column=6)
                risk_cell.value = risk_value
                risk_cell.alignment = Alignment(horizontal='center', vertical='center')
                risk_cell.border = thin_border
                
                # Apply conditional formatting based on risk level
                risk_lower = risk_value.lower().strip()
                if risk_lower == 'critical':
                    risk_cell.fill = PatternFill(start_color='FF8B0000', end_color='FF8B0000', fill_type='solid')  # Dark Red
                    risk_cell.font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFFFF')  # White
                elif risk_lower == 'high':
                    risk_cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')  # Red
                    risk_cell.font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFFFF')  # White
                elif risk_lower == 'medium':
                    risk_cell.fill = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')  # Orange
                    risk_cell.font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFFFF')  # White
                elif risk_lower == 'low':
                    risk_cell.fill = PatternFill(start_color='FF008000', end_color='FF008000', fill_type='solid')  # Green
                    risk_cell.font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFFFF')  # White
                
                # G column - Compliance By Bank (empty for user to fill)
                ws.cell(row=row_num, column=7).value = ""
                ws.cell(row=row_num, column=7).font = Font(name='Times New Roman', size=12)
                ws.cell(row=row_num, column=7).border = thin_border
                
                # H column - Status (empty for user to fill)
                ws.cell(row=row_num, column=8).value = ""
                ws.cell(row=row_num, column=8).font = Font(name='Times New Roman', size=12)
                ws.cell(row=row_num, column=8).border = thin_border
                
                # I column - Pl attach evidence (empty for user to fill)
                ws.cell(row=row_num, column=9).value = ""
                ws.cell(row=row_num, column=9).font = Font(name='Times New Roman', size=12)
                ws.cell(row=row_num, column=9).border = thin_border
            
            current_row = end_row + 1
            worksheet_number += 1
        
        # Post-process: Fix page numbers for compliance rows
        print(f"\n🔧 Post-processing: Fixing page numbers for compliance rows...")
        
        # Get the last row with data
        last_data_row = current_row - 1
        
        # Scan from row 6 to last data row
        for row_num in range(6, last_data_row + 1):
            observation_cell = ws.cell(row=row_num, column=5)
            
            # Check if this is a compliance row
            if observation_cell.value and "As per Auditor's Observation all Points are Compliance" in str(observation_cell.value):
                print(f"   🔍 Row {row_num}: Found compliance row")
                
                # Find the next non-compliance row with a page number
                next_page = None
                for search_row in range(row_num + 1, last_data_row + 1):
                    next_obs = ws.cell(row=search_row, column=5).value
                    next_page_cell = ws.cell(row=search_row, column=4).value
                    
                    # Check if this is NOT a compliance row and has a page number
                    if next_obs and "As per Auditor's Observation all Points are Compliance" not in str(next_obs):
                        if next_page_cell and str(next_page_cell).strip():
                            try:
                                next_page = int(next_page_cell)
                                print(f"   ✅ Found next non-compliance page: {next_page} at row {search_row}")
                                break
                            except:
                                pass
                
                # If found a next page, calculate page number
                if next_page:
                    # Now count backwards for consecutive compliance rows
                    compliance_rows = []
                    for check_row in range(row_num, last_data_row + 1):
                        check_obs = ws.cell(row=check_row, column=5).value
                        if check_obs and "As per Auditor's Observation all Points are Compliance" in str(check_obs):
                            compliance_rows.append(check_row)
                        else:
                            break
                    
                    # Assign page numbers backwards
                    for idx, comp_row in enumerate(reversed(compliance_rows)):
                        page_to_assign = next_page - idx - 1
                        ws.cell(row=comp_row, column=4).value = page_to_assign
                        print(f"   📄 Row {comp_row}: Assigned page {page_to_assign}")
        
        print(f"✅ Post-processing complete\n")
        
        # Save and return (will be saved after adding branch sheet)
        return wb
        
    except Exception as e:
        print(f"❌ Error creating IS Audit Compliance Excel: {e}")
        import traceback
        traceback.print_exc()
        raise

def create_branch_worksheet(wb, form_data, excel_file2_path):
    """Create IS Audit - Branch worksheet"""
    try:
        print(f"\n🔍 Creating IS Audit - Branch worksheet...")
        
        # Create new worksheet
        ws_branch = wb.create_sheet(title="IS Audit - Branch")
        
        # Get the first worksheet to copy header rows
        ws_head_office = wb["IS Audit - Head Office"]
        
        # Get form data for headers
        org_name = form_data.get('organizationName', '')
        if org_name == 'Other':
            org_name = form_data.get('organizationNameOther', '')
        
        report_id = form_data.get('reportId', '')
        report_date = form_data.get('reportDate', '')
        compliance_date = form_data.get('complianceDate', '')
        
        # Format dates
        try:
            report_date_obj = datetime.strptime(report_date, '%Y-%m-%d')
            report_date_formatted = report_date_obj.strftime('%d/%m/%Y')
        except:
            report_date_formatted = report_date
        
        try:
            compliance_date_obj = datetime.strptime(compliance_date, '%Y-%m-%d')
            compliance_date_formatted = compliance_date_obj.strftime('%d/%m/%Y')
        except:
            compliance_date_formatted = compliance_date
        
        # Set column widths
        ws_branch.column_dimensions['A'].width = 20
        ws_branch.column_dimensions['B'].width = 20
        ws_branch.column_dimensions['C'].width = 30
        ws_branch.column_dimensions['D'].width = 30
        ws_branch.column_dimensions['E'].width = 23
        ws_branch.column_dimensions['F'].width = 60
        ws_branch.column_dimensions['G'].width = 30
        ws_branch.column_dimensions['H'].width = 60
        ws_branch.column_dimensions['I'].width = 30
        ws_branch.column_dimensions['J'].width = 50
        
        # Define styles
        header_font = Font(name='Times New Roman', size=12, bold=False)
        header_alignment_center = Alignment(horizontal='center', vertical='center')
        header_alignment_left = Alignment(horizontal='left', vertical='center')
        
        column_header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
        column_header_fill = PatternFill(start_color='FF1F497D', end_color='FF1F497D', fill_type='solid')
        column_header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='FF000000'),
            right=Side(style='thin', color='FF000000'),
            top=Side(style='thin', color='FF000000'),
            bottom=Side(style='thin', color='FF000000')
        )
        
        # Copy first 4 rows from Head Office sheet
        # Row 1
        ws_branch['A1'] = "Submitted To:"
        ws_branch['A1'].font = header_font
        ws_branch['A1'].alignment = header_alignment_center
        ws_branch['A1'].border = thin_border
        
        ws_branch.merge_cells('B1:J1')
        ws_branch['B1'] = f"The Management, {org_name}"
        ws_branch['B1'].font = header_font
        ws_branch['B1'].alignment = header_alignment_left
        for col in range(2, 11):  # B to J
            ws_branch.cell(row=1, column=col).border = thin_border
        
        # Row 2
        ws_branch['A2'] = "Report Name:"
        ws_branch['A2'].font = header_font
        ws_branch['A2'].alignment = header_alignment_center
        ws_branch['A2'].border = thin_border
        
        ws_branch.merge_cells('B2:J2')
        ws_branch['B2'] = f"Compliance Of Information System Audit Report (Report ID no. {report_id} dated {report_date_formatted})"
        ws_branch['B2'].font = header_font
        ws_branch['B2'].alignment = header_alignment_left
        for col in range(2, 11):
            ws_branch.cell(row=2, column=col).border = thin_border
        
        # Row 3
        ws_branch['A3'] = "Confidentiality:"
        ws_branch['A3'].font = header_font
        ws_branch['A3'].alignment = header_alignment_center
        ws_branch['A3'].border = thin_border
        
        ws_branch.merge_cells('B3:J3')
        ws_branch['B3'] = "Very High & Not for circulation"
        ws_branch['B3'].font = header_font
        ws_branch['B3'].alignment = header_alignment_left
        for col in range(2, 11):
            ws_branch.cell(row=3, column=col).border = thin_border
        
        # Row 4
        ws_branch['A4'] = "Compliance Date:"
        ws_branch['A4'].font = header_font
        ws_branch['A4'].alignment = header_alignment_center
        ws_branch['A4'].border = thin_border
        
        ws_branch.merge_cells('B4:J4')
        ws_branch['B4'] = compliance_date_formatted
        ws_branch['B4'].font = header_font
        ws_branch['B4'].alignment = header_alignment_left
        for col in range(2, 11):
            ws_branch.cell(row=4, column=col).border = thin_border
        
        # Load the second Excel file (branch data)
        print(f"   📂 Loading branch Excel file: {excel_file2_path}")
        source_wb = load_workbook(excel_file2_path)
        source_ws = source_wb.active  # Assuming data is in the first sheet
        
        print(f"   📊 Source has {source_ws.max_row} rows and {source_ws.max_column} columns")
        
        # Copy data from source Excel (columns A to G) starting from row 5
        # Also copy all formatting
        row_offset = 4  # Data starts at row 5 (4 header rows)
        
        for row_idx in range(1, source_ws.max_row + 1):
            target_row = row_idx + row_offset
            
            # Copy columns A to G
            for col_idx in range(1, 8):  # 1-7 for A-G
                source_cell = source_ws.cell(row=row_idx, column=col_idx)
                target_cell = ws_branch.cell(row=target_row, column=col_idx)
                
                # Copy value
                target_cell.value = source_cell.value
                
                # Copy formatting
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = copy(source_cell.number_format)
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)
        
        # Copy merged cells from source (columns A to G only)
        print(f"   🔗 Copying merged cells...")
        merged_count = 0
        for merged_range in source_ws.merged_cells.ranges:
            # Get the boundaries of the merged range
            min_col = merged_range.min_col
            max_col = merged_range.max_col
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            
            # Only copy if it's within columns A to G (1-7)
            if min_col >= 1 and max_col <= 7:
                # Adjust row numbers for target sheet
                target_min_row = min_row + row_offset
                target_max_row = max_row + row_offset
                
                # Create merged cell range in target
                min_col_letter = get_column_letter(min_col)
                max_col_letter = get_column_letter(max_col)
                target_range = f"{min_col_letter}{target_min_row}:{max_col_letter}{target_max_row}"
                ws_branch.merge_cells(target_range)
                
                print(f"      🔗 Merged {merged_range} → {target_range}")
                merged_count += 1
        
        print(f"   ✅ Copied {merged_count} merged cell ranges")
        
        # Get G6 cell formatting for reference
        first_data_row = 1 + row_offset  # Row 5
        g6_cell = ws_branch.cell(row=6, column=7)  # G6
        
        # Copy G6 formatting
        g6_font = copy(g6_cell.font) if g6_cell.has_style else Font(name='Times New Roman', size=12)
        g6_fill = copy(g6_cell.fill) if g6_cell.has_style else PatternFill()
        g6_alignment = copy(g6_cell.alignment) if g6_cell.has_style else Alignment(horizontal='center', vertical='center')
        g6_border = copy(g6_cell.border) if g6_cell.has_style else thin_border
        
        # Add headers with G6 formatting
        # H6 = "COMPLIANCE BY BANK"
        ws_branch['H6'] = "Compliance by Bank"
        ws_branch['H6'].font = g6_font
        ws_branch['H6'].fill = g6_fill
        ws_branch['H6'].alignment = g6_alignment
        ws_branch['H6'].border = g6_border
        
        # I6 = "Status"
        ws_branch['I6'] = "Status"
        ws_branch['I6'].font = g6_font
        ws_branch['I6'].fill = g6_fill
        ws_branch['I6'].alignment = g6_alignment
        ws_branch['I6'].border = g6_border
        
        # J7 = "Pl attach evidence"
        ws_branch['J6'] = "Pl attach evidence"
        ws_branch['J6'].font = g6_font
        ws_branch['J6'].fill = g6_fill
        ws_branch['J6'].alignment = g6_alignment
        ws_branch['J6'].border = g6_border
        
        # Keywords that trigger merging (case-sensitive, exact match)
        merge_keywords = [
            "Head Office and",  # Will match "Head Office and <any number> Branches"
            "CBS Access Control",
            "PHYSICAL AND ENVIRONMENTAL SECURITY",
            "PATCH MANAGEMENT",
            "NETWORK SECURITY",
            "ENDPOINTS VULNERABILITY",
            "ATM MACHINE ROOM",
            "EMAIL-SECURITY",
            "POWER BACK UP",
            "USER AWARENESS",
            "MAINTENANCE AND BUSINESS CONTINUITY CONTROLS",
            "REMOTE ACCESS",
            "UNAUTHORIZED APPLICATIONS / PERSONAL DATA",
            "Important Note"
        ]
        
        # Check for keywords and merge cells from A to J
        print(f"   🔍 Checking for merge keywords...")
        keyword_merged_count = 0
        
        # Start from first data row + offset and go through all copied rows
        for row_idx in range(first_data_row, source_ws.max_row + row_offset + 1):
            cell_a = ws_branch.cell(row=row_idx, column=1)
            cell_value = str(cell_a.value).strip() if cell_a.value else ""
            
            should_merge = False
            
            # Check for exact keyword match
            for keyword in merge_keywords:
                if keyword == "Head Office and":
                    # Special case: match "Head Office and <number> Branches"
                    if cell_value.startswith(keyword):
                        should_merge = True
                        print(f"      🔗 Row {row_idx}: Found '{cell_value}' - merging A{row_idx}:J{row_idx}")
                        break
                else:
                    # Exact match (case-sensitive)
                    if cell_value == keyword:
                        should_merge = True
                        print(f"      🔗 Row {row_idx}: Found '{keyword}' - merging A{row_idx}:J{row_idx}")
                        break
            
            if should_merge:
                # Check if this cell is already part of a merged range
                is_already_merged = False
                for merged_range in ws_branch.merged_cells.ranges:
                    if row_idx >= merged_range.min_row and row_idx <= merged_range.max_row:
                        if 1 >= merged_range.min_col and 1 <= merged_range.max_col:
                            is_already_merged = True
                            print(f"      ⚠️  Row {row_idx} is already merged, skipping keyword merge")
                            break
                
                if not is_already_merged:
                    # Store the current formatting
                    original_font = copy(cell_a.font)
                    original_fill = copy(cell_a.fill)
                    original_alignment = copy(cell_a.alignment)
                    original_border = copy(cell_a.border)
                    
                    # Merge cells A to J
                    ws_branch.merge_cells(f'A{row_idx}:J{row_idx}')
                    
                    # Reapply formatting to merged cell
                    merged_cell = ws_branch.cell(row=row_idx, column=1)
                    merged_cell.font = original_font
                    merged_cell.fill = original_fill
                    merged_cell.alignment = original_alignment
                    merged_cell.border = original_border
                    
                    # Apply border to all cells in merged range
                    for col in range(1, 11):  # A to J
                        ws_branch.cell(row=row_idx, column=col).border = original_border
                    
                    keyword_merged_count += 1
        
        print(f"   ✅ Merged {keyword_merged_count} rows based on keywords")
        
        # Apply Times New Roman 12 to all cells and add borders to H, I, J columns
        print(f"   🎨 Applying formatting and borders...")
        
        times_new_roman_font = Font(name='Times New Roman', size=12)
        last_row_with_content = 0
        
        # Find the last row with content in column A
        for row_idx in range(1, ws_branch.max_row + 1):
            cell_a = ws_branch.cell(row=row_idx, column=1)
            if cell_a.value:
                last_row_with_content = row_idx
        
        print(f"   📏 Last row with content in column A: {last_row_with_content}")
        
        # Apply formatting to all cells in the worksheet
        for row_idx in range(1, ws_branch.max_row + 1):
            for col_idx in range(1, 11):  # A to J (columns 1-10)
                cell = ws_branch.cell(row=row_idx, column=col_idx)
                
                # Apply Times New Roman 12 to all cells
                # Preserve existing font properties (bold, color, etc.) but change name and size
                if cell.font:
                    cell.font = Font(
                        name='Times New Roman',
                        size=12,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color,
                        underline=cell.font.underline,
                        strike=cell.font.strike
                    )
                else:
                    cell.font = times_new_roman_font
                
                # Add borders to H, I, J columns for rows where A has content
                if col_idx in [8, 9, 10] and row_idx <= last_row_with_content:  # H, I, J columns
                    # Check if row A has content
                    cell_a = ws_branch.cell(row=row_idx, column=1)
                    if cell_a.value:
                        # Apply border if not already set
                        if not cell.border or cell.border.left.style is None:
                            cell.border = thin_border
        
        print(f"   ✅ Applied Times New Roman 12 to all cells")
        print(f"   ✅ Added borders to H, I, J columns for rows with content")
        print(f"   ✅ Branch worksheet created successfully")
        
        return wb
        
    except Exception as e:
        print(f"   ❌ Error creating branch worksheet: {e}")
        import traceback
        traceback.print_exc()
        raise

@grc_is_audit_compliance_bp.route('/grc_process_is_audit_compliance', methods=['POST'])
def process_is_audit_compliance():
    """Process IS Audit Compliance form submission"""
    try:
        print("\n" + "="*80)
        print("🚀 Processing IS Audit Compliance")
        print("="*80)
        
        # Get form data
        form_data = {
            'organizationName': request.form.get('organizationName'),
            'organizationNameOther': request.form.get('organizationNameOther'),
            'reportId': request.form.get('reportId'),
            'reportDate': request.form.get('reportDate'),
            'complianceDate': request.form.get('complianceDate')
        }
        
        # Handle file uploads
        excel_file1 = request.files.get('excelFile1')
        excel_file2 = request.files.get('excelFile2')
        
        if not excel_file1:
            return jsonify({'success': False, 'error': 'Excel file 1 is required'}), 400
        
        if not excel_file2:
            return jsonify({'success': False, 'error': 'Excel file 2 is required'}), 400
        
        # Save uploaded files temporarily
        upload_dir = os.path.join('static', 'uploads', 'temp')
        os.makedirs(upload_dir, exist_ok=True)
        
        excel_path1 = os.path.join(upload_dir, f'temp_excel1_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        excel_path2 = os.path.join(upload_dir, f'temp_excel2_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
        excel_file1.save(excel_path1)
        excel_file2.save(excel_path2)
        
        # Create the compliance Excel (Head Office sheet)
        wb = create_is_audit_compliance_excel(form_data, excel_path1)
        
        # Create the Branch worksheet
        wb = create_branch_worksheet(wb, form_data, excel_path2)
        
        # Save the final workbook
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'IS_Audit_Compliance_{timestamp}.xlsx'
        filepath = os.path.join('static', 'uploads', filename)
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        wb.save(filepath)
        
        # Clean up temporary files immediately
        temp_files_deleted = False
        try:
            if os.path.exists(excel_path1):
                os.remove(excel_path1)
                print(f"   🗑️  Deleted temp file: {excel_path1}")
            if os.path.exists(excel_path2):
                os.remove(excel_path2)
                print(f"   🗑️  Deleted temp file: {excel_path2}")
            temp_files_deleted = True
        except Exception as e:
            print(f"   ⚠️  Warning: Could not delete temp files: {e}")
        
        print(f"\n✅ IS Audit Compliance Excel created: {filename}")
        print("="*80)
        
        return jsonify({
            'success': True,
            'filename': filename,
            'download_url': f'/static/uploads/{filename}',
            'temp_cleaned': temp_files_deleted
        })
    
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Clean up temporary files even if there's an error
        try:
            if 'excel_path1' in locals() and os.path.exists(excel_path1):
                os.remove(excel_path1)
                print(f"   🗑️  Cleaned up temp file: {excel_path1}")
            if 'excel_path2' in locals() and os.path.exists(excel_path2):
                os.remove(excel_path2)
                print(f"   🗑️  Cleaned up temp file: {excel_path2}")
        except Exception as cleanup_error:
            print(f"   ⚠️  Could not clean up temp files: {cleanup_error}")
        
        return jsonify({'success': False, 'error': str(e)}), 500

@grc_is_audit_compliance_bp.route('/grc_cleanup_is_audit_compliance', methods=['POST'])
def cleanup_is_audit_compliance():
    """Clean up IS Audit Compliance Excel file after download"""
    try:
        data = request.get_json()
        filename = data.get('filename')
        
        files_deleted = []
        files_not_found = []
        
        if filename:
            # Delete the main file from static/uploads
            file_path = os.path.join('static', 'uploads', filename)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    files_deleted.append(file_path)
                    print(f"✅ Deleted: {file_path}")
                except Exception as e:
                    print(f"❌ Error deleting {file_path}: {e}")
            else:
                files_not_found.append(file_path)
                print(f"⚠️  File not found: {file_path}")
        
        # Also clean up any remaining temp files in static/uploads/temp
        temp_dir = os.path.join('static', 'uploads', 'temp')
        if os.path.exists(temp_dir):
            try:
                # Get all files in temp directory
                temp_files = [f for f in os.listdir(temp_dir) if os.path.isfile(os.path.join(temp_dir, f))]
                for temp_file in temp_files:
                    temp_file_path = os.path.join(temp_dir, temp_file)
                    try:
                        # Check if file is older than 5 minutes (to avoid deleting files being processed)
                        file_age = datetime.now().timestamp() - os.path.getmtime(temp_file_path)
                        if file_age > 300:  # 5 minutes
                            os.remove(temp_file_path)
                            files_deleted.append(temp_file_path)
                            print(f"✅ Deleted old temp file: {temp_file_path}")
                    except Exception as e:
                        print(f"❌ Error deleting temp file {temp_file_path}: {e}")
            except Exception as e:
                print(f"❌ Error cleaning temp directory: {e}")
        
        print(f"📊 Cleanup summary: {len(files_deleted)} files deleted, {len(files_not_found)} files not found")
        
        return jsonify({
            'success': True,
            'files_deleted': len(files_deleted),
            'files_not_found': len(files_not_found)
        })
    except Exception as e:
        print(f"❌ Error cleaning up: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

