from flask import Blueprint, request, flash, redirect, url_for, send_file, Response
from flask_login import login_required
from datetime import datetime
import os
import tempfile
import shutil
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO
import pandas as pd
import copy

# Create blueprint
branch_console_bp = Blueprint('branch_console_bp', __name__)

# Allowed file extensions
ALLOWED_EXCEL_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename, allowed_extensions):
    """Check if the uploaded file has an allowed extension"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in allowed_extensions

def is_non_compliance_value(value):
    """Check if the value is a Non-Compliance variation"""
    if pd.isna(value) or value is None:
        return False
    
    value_str = str(value).strip().lower()
    
    non_compliance_variations = [
        'non-compliance', 'non-compliance', 'non-compliance', 
        'noncompliance', 'noncompliance', 'noncompliance'
    ]
    
    return value_str in non_compliance_variations

def analyze_excel_file(file_path):
    """Analyze Excel file for Non-Compliance entries in Column E"""
    try:
        print(f"Analyzing Excel file: {file_path}")
        
        # Load the workbook (without data_only to preserve images)
        workbook = load_workbook(file_path, data_only=False)
        
        analysis_results = []
        
        # Iterate through all worksheets
        for sheet_name in workbook.sheetnames:
            print(f"Processing worksheet: {sheet_name}")
            worksheet = workbook[sheet_name]
            
            # Find the maximum row and column
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            if max_row <= 1:
                print(f"Worksheet '{sheet_name}' has no data rows")
                continue
            
            worksheet_results = {
                'worksheet_name': sheet_name,
                'non_compliance_entries': []
            }
            
            # Check each row for Non-Compliance in Column E (column index 5)
            for row_num in range(1, max_row + 1):
                cell_a = worksheet.cell(row=row_num, column=1)  # Column A
                cell_b = worksheet.cell(row=row_num, column=2)  # Column B
                cell_d = worksheet.cell(row=row_num, column=4)  # Column D
                cell_e = worksheet.cell(row=row_num, column=5)  # Column E
                
                if is_non_compliance_value(cell_e.value):
                    # Found Non-Compliance entry
                    a_value = cell_a.value if cell_a.value is not None else ""
                    b_value = cell_b.value if cell_b.value is not None else ""
                    d_value = cell_d.value if cell_d.value is not None else ""
                    e_value = cell_e.value if cell_e.value is not None else ""
                    
                    # Extract images from POC columns G, H, I (columns 7, 8, 9 = indices 6, 7, 8)
                    images = extract_images_from_row(worksheet, row_num, [6, 7, 8])
                    
                    entry = {
                        'row_number': row_num,
                        'column_a_data': str(a_value),
                        'column_b_data': str(b_value),
                        'column_d_data': str(d_value),
                        'column_e_data': str(e_value),
                        'cell_reference': f"E{row_num}",
                        'images': images  # Store extracted images
                    }
                    
                    worksheet_results['non_compliance_entries'].append(entry)
                    print(f"Found Non-Compliance in {sheet_name} at E{row_num}: A{row_num}='{a_value}', B{row_num}='{b_value}', D{row_num}='{d_value}', Images={len(images)}")
            
            if worksheet_results['non_compliance_entries']:
                analysis_results.append(worksheet_results)
        
        return analysis_results
        
    except Exception as e:
        print(f"Error analyzing Excel file: {e}")
        raise e

def extract_branch_name_from_worksheet(worksheet_name):
    """Extract branch name from worksheet name like '2_Ratanpur Ga Branch' -> 'Ratanpur Ga Branch'"""
    import re
    
    # Remove leading numbers and underscores
    # Pattern: number(s) + underscore + rest of the name
    match = re.match(r'^\d+_(.+)$', worksheet_name)
    if match:
        branch_name = match.group(1).strip()
    else:
        branch_name = worksheet_name.strip()
    
    # Also remove leading numbers and spaces without underscore (e.g., "1 Wadhwan Branch" -> "Wadhwan Branch")
    branch_name = re.sub(r'^\d+\s+', '', branch_name).strip()
    
    return branch_name

def extract_images_from_row(worksheet, row_number, poc_columns=[6, 7, 8]):
    """Extract images from specific row and POC columns (G, H, I are columns 7, 8, 9 in 1-based, or 6, 7, 8 in 0-based)"""
    extracted_images = []
    
    try:
        # Check all images in the worksheet
        for image in worksheet._images:
            try:
                # Get image position
                row_num = None
                col_num = None
                
                if hasattr(image, 'anchor'):
                    anchor = image.anchor
                    if hasattr(anchor, '_from'):
                        if hasattr(anchor._from, 'row') and hasattr(anchor._from, 'col'):
                            row_num = anchor._from.row + 1  # Convert to 1-based
                            col_num = anchor._from.col
                    elif hasattr(anchor, 'row') and hasattr(anchor, 'col'):
                        row_num = anchor.row + 1
                        col_num = anchor.col
                    elif hasattr(anchor, '_from') and hasattr(anchor._from, '_row') and hasattr(anchor._from, '_col'):
                        row_num = anchor._from._row + 1
                        col_num = anchor._from._col
                
                # Check if image is in the target row and POC columns
                if row_num == row_number and col_num in poc_columns:
                    # Copy the image
                    if hasattr(image, '_data'):
                        img_data = image._data()
                    elif hasattr(image, 'ref'):
                        img_data = image.ref
                    else:
                        img_data = None
                    
                    if img_data:
                        extracted_images.append({
                            'data': img_data,
                            'column': col_num,
                            'row': row_num
                        })
                        print(f"Extracted image from row {row_num}, column {col_num}")
            
            except Exception as e:
                print(f"Error extracting image: {e}")
                continue
        
    except Exception as e:
        print(f"Error in extract_images_from_row: {e}")
    
    return extracted_images

def add_images_to_cell(sheet, cell_ref, images_list, max_images=5):
    """Add up to max_images images to a specific cell in the worksheet"""
    try:
        if not images_list:
            return
        
        # Limit to maximum images
        images_to_add = images_list[:max_images]
        
        # Add each image to the cell
        for idx, img_data in enumerate(images_to_add):
            try:
                # Create image from data
                img = Image(BytesIO(img_data['data']))
                
                # Get original dimensions
                original_width = img.width
                original_height = img.height
                
                # Resize to 30px height with aspect ratio maintained
                target_height = 30
                aspect_ratio = original_width / original_height if original_height else 1
                img.height = target_height
                img.width = int(target_height * aspect_ratio)
                
                # Set image position (anchor to cell)
                # For multiple images, offset them vertically
                img.anchor = cell_ref
                
                # Add image to worksheet
                sheet.add_image(img)
                print(f"Added image {idx + 1} to cell {cell_ref} (size: {img.width}x{img.height}, original: {original_width}x{original_height})")
                
            except Exception as e:
                print(f"Error adding image {idx + 1} to cell {cell_ref}: {e}")
    
    except Exception as e:
        print(f"Error in add_images_to_cell for {cell_ref}: {e}")

def generate_excel_report(analysis_results, total_worksheets):
    """Generate Excel report from analysis results grouped by Column B data"""
    try:
        # Create a new workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Branch Console Report"
        
        # Set column widths
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 43
        sheet.column_dimensions['G'].width = 34
        sheet.column_dimensions['H'].width = 40  # Recommendation column
        sheet.column_dimensions['I'].width = 15  # POC column 1
        sheet.column_dimensions['J'].width = 15  # POC column 2
        sheet.column_dimensions['K'].width = 15  # POC column 3
        sheet.column_dimensions['L'].width = 15  # POC column 4
        sheet.column_dimensions['M'].width = 15  # POC column 5
        
        # Merge cells for header row (A1:M1)
        sheet.merge_cells('A1:M1')
        
        # Calculate number of branches (total worksheets - 1 for Head Office)
        num_branches = total_worksheets - 1 if total_worksheets > 0 else total_worksheets
        
        # Create border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set header text
        header_cell = sheet['A1']
        header_cell.value = f"Head Office and {num_branches} Branches"
        
        # Apply header formatting
        header_cell.font = Font(
            name='Calibri',
            size=12,
            bold=True,
            color='FFFFFF'  # White text
        )
        header_cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        header_cell.fill = PatternFill(
            start_color='000080',  # Blue background
            end_color='000080',
            fill_type='solid'
        )
        header_cell.border = thin_border
        
        # Apply border to all cells in merged range A1:M1
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            cell = sheet[f'{col}1']
            cell.border = thin_border
        
        # Add second row headers
        # Merge C2:D2 for Observations
        sheet.merge_cells('C2:D2')
        
        # Merge I2:M2 for POC's
        sheet.merge_cells('I2:M2')
        
        # Set header values for row 2
        sheet['A2'].value = 'Sr. No.'
        sheet['B2'].value = 'Reference No.'
        sheet['C2'].value = 'Observations'
        sheet['E2'].value = 'Risk Factor'
        sheet['F2'].value = 'Impact'
        sheet['G2'].value = 'Branches'
        sheet['H2'].value = 'Recommendation'
        sheet['I2'].value = "POC's"
        
        # Apply formatting to all cells in row 2
        row2_cells = ['A2', 'B2', 'C2', 'E2', 'F2', 'G2', 'H2', 'I2']
        
        for cell_ref in row2_cells:
            cell = sheet[cell_ref]
            cell.font = Font(
                name='Calibri',
                size=12,
                bold=True,
                color='FFFFFF'  # White text
            )
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            cell.fill = PatternFill(
                start_color='007FFF',  # Light blue background
                end_color='007FFF',
                fill_type='solid'
            )
            cell.border = thin_border
        
        # Apply border to D2 (part of merged C2:D2)
        sheet['D2'].border = thin_border
        
        # Apply border to J2, K2, L2, M2 (part of merged I2:M2)
        for col in ['J', 'K', 'L', 'M']:
            sheet[f'{col}2'].border = thin_border
        
        # Add third row - Merge A3:M3
        sheet.merge_cells('A3:M3')
        
        # Set third row text
        third_row_cell = sheet['A3']
        third_row_cell.value = 'CBS Access Control'
        
        # Apply same formatting as first row (Row 1)
        third_row_cell.font = Font(
            name='Calibri',
            size=12,
            bold=True,
            color='FFFFFF'  # White text
        )
        third_row_cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        third_row_cell.fill = PatternFill(
            start_color='000080',  # Navy Blue background (same as Row 1)
            end_color='000080',
            fill_type='solid'
        )
        third_row_cell.border = thin_border
        
        # Apply border to all cells in merged range A3:M3
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            cell = sheet[f'{col}3']
            cell.border = thin_border
        
        # Define CBS Access Control items
        cbs_access_control_items = {
            "Do employees are using strong passwords?",
            "Do you mandate periodical password changes?",
            "Do passwords are shared among employees?",
            "What happens after consecutive failed login attempts?",
            "Is session timeout enforced after a pre-defined period of inactivity?",
            "Whether Multiple Logins is enabled?",
            "Is the CBS System isolated from the internet?",
            "Is internet access restricted to only trusted and officially approved sites?",
            "Is there any policy for default password while unlocking the account?",
            "When a new user joins then how do you give the user ID and Password for the user by using mail or telephonic conversation?",
            "SSL is using or not in CBS Application?",
            "CBS is browser based or desktop based?",
            "Is the CBS compatible with the latest version of the browser?",
            "Whether Two Factor Authentication is implemented for CBS login?"
        }
        
        # Define Physical and Environmental Security items
        physical_environmental_security_items = {
            "CCTV Camera Present?",
            "Branch Locker CCTV Camera Present?",
            "CCTV camera covering complete branch area?",
            "Check DVR working properly?",
            "Is NTP configured in the DVR?",
            "DVR is available in bank network or not?",
            "Camera is working in night vision.",
            "CCTV camera history recording duration.",
            "Locker CCTV camera history recording duration.",
            "Biometric Devices or Proximity Card is available or not?",
            "ID card issues to the Employee?",
            "Whether guard is available at branch or not?",
            "Are secure areas controlled?",
            "Smoke Detector is installed in the branch?",
            "Whether Panic Switch installed?",
            "If a panic switch or smoke detector is installed at the branch or head office does the notification reach higher authorities when triggered?",
            "LAN ports are open or not?",
            "Visitor Register is maintained or not?",
            "Any record maintains for drill the environmental controls?",
            "Whether external modems, data cards etc., are being used in the Branch to access internet?",
            "Fire extinguisher is available or not?",
            "Enough fire-extinguisher is available for area?"
        }
        
        # Define Power Back Up items
        power_backup_items = {
            "Does the Bank have made enough Battery Backup which supports Computer Systems?",
            "Do the AMC is given for power backup systems?",
            "Generator is available or not?"
        }
        
        # Define User Awareness items
        user_awareness_items = {
            "Is training conducted for end user with respect to the Information Security?",
            "Is training conducted for end user with respect to the application usage?",
            "Employees are aware about using of fire extinguisher & Panic Switch?",
            "User was aware about checking CCTV recording history?"
        }
        
        # Define Maintenance and Business Continuity Controls items
        maintenance_business_continuity_items = {
            "Is AMC is given for computer and related systems?",
            "Is Complaint register maintained?"
        }
        
        # Define Patch Management items
        patch_management_items = {
            "Antivirus available in all the desktop?",
            "Antivirus updated in all the desktop?",
            "Windows patch updated in all the system?",
            "Are there any system on which outdated version of windows is installed?",
            "Expired License antivirus installed?"
        }
        
        # Define Network Security items
        network_security_items = {
            "Are controls in place to ensure users only have access to the network resources they have been specially authorized to use and are required for their duties?",
            "Is there network diagram available?",
            "Is structured cabling is observed?",
            "Is cable tagging is observed?",
            "Is provision made for backup network?",
            "Networking devices are under secure location and there is any cabinet facility is provided or not?",
            "Proper Cooling for Networking Devices available or not?",
            "Wi-Fi is available or not?",
            "Whether Firewall is available in Branch ?"
        }
        
        # Define Endpoints Vulnerability items
        endpoints_vulnerability_items = {
            "Does User login with Administrator Rights?",
            "Can Group Policy be modified?",
            "Can Security Configuration Policy be modified?",
            "Whether Proxy can be modified by branch users?",
            "Whether internet is allowed or not?",
            "Whether internet access is restricted or unrestricted?",
            "Authentication Policy for internet is available or not?",
            "System Password is weak or not?",
            "System tagging is available or not?",
            "Is NTP configured in the systems?",
            "Whether USB ports/CD drives enabled in the System?",
            "USB Approval process?",
            "If bank have trusted USB, then check the USB open outside the network?",
            "Whether Firewall is enable in systems?",
            "Whether all external storage media are checked for virus?",
            "Whether Remote Desktop Protocol is enabled or disabled?"
        }
        
        # Define ATM Machine Room items
        atm_machine_room_items = {
            "Check the CCTV camera  LAN cable and ATM machine power cable concealed?",
            "Guard is available or not?",
            "CCTV Camera and check the position of camera?",
            "ATM is working 24*7?",
            "Whether the Do's and Don't for the user awareness is available or not?",
            "Whether the ATM is network segmented or not?",
            "Whether ATM Machine is properly grouted (floor and wall) ?",
            "Whether access to ATM room for maintenance purpose is restricted to the authorized persons only?",
            "Whether ATM power back up is supported by UPS.",
            "System of ATM cash replenishment, adherence to dual control mechanism records is maintained properly?",
            "Check system of ATM reconciliation status maintained?",
            "Whether ATM machine preservation is done of journal print for future reference?"
        }
        
        # Define Email Security items
        email_security_items = {
            "Personal Mail used or not?",
            "Two factor authentications are in used or not?",
            "Single mail with multiple users?",
            "Whether official email ID gets open outside bank N/w?"
        }
        
        # Define Remote Access items
        remote_access_items = {
            "Which Department is using Remote Access?",
            "Remote Access approval process?"
        }
        
        # Define Unauthorized Applications / Personal Data items (partial match)
        unauthorized_apps_personal_data_prefixes = {
            "There are no unauthorized applications on any systems",
            "Unauthorized applications have been found in the below systems IP's",
            "There are no Personal Data on any systems",
            "Personal Data have been found in the below systems IP's"
        }
        
        # Define Important Note items in specific order
        important_note_items_order = [
            "Auditor Identity was not verified.",
            "Windows was not activated.",
            "Password was saved in browser.",
            "Password was written on wall, desk, notepad, diary.",
            "Asset Movement Register was not available.",
            "Dust was present on the network devices & systems.",
            "Auto switchable mode was not available for connectivity.",
            "Preventive Maintenance is not carried out for Systems/UPS/Network Devices."
        ]
        
        # Define content replacement mapping for CBS Access Control items
        cbs_content_replacement = {
            "Do employees are using strong passwords?": "Global password policy is not followed in CBS access.",
            "Do you mandate periodical password changes?": "Password does not expires after 30 days. It is mandatory to change the password after 30 days.",
            "Do passwords are shared among employees?": "Employees do share passwords with each other.",
            "What happens after consecutive failed login attempts?": "Account does not get locked after three consecutive failed login attempts.",
            "Is session timeout enforced after a pre-defined period of inactivity?": "Session timeout is less than 10 minutes.",
            "Whether Multiple Logins is enabled?": "Multiple logins was enabled.",
            "Is the CBS System isolated from the internet?": "CBS System is not isolated from the internet.",
            "Is internet access restricted to only trusted and officially approved sites?": "Unrestricted internet access is given to the user.",
            "Is there any policy for default password while unlocking the account?": "Default password is in use for unlocking the account.",
            "When a new user joins then how do you give the user ID and Password for the user by using mail or telephonic conversation?": "New user password creation process was not defined.",
            "SSL is using or not in CBS Application?": "SSL was not implemented in CBS application.",
            "CBS is browser based or desktop based?": "Extension is used to run CBS application.",
            "Is the CBS compatible with the latest version of the browser?": "CBS is not compatible with the latest version of the browser.",
            "Whether Two Factor Authentication is implemented for CBS login?": "Two factor authentication is not implemented for CBS login."
        }
        
        # Define content replacement mapping for all other sections
        all_content_replacement = {
            # Physical and Environmental Security
            "CCTV Camera Present?": "CCTV cameras are not present or not functional.",
            "Branch Locker CCTV Camera Present?": "Branch locker CCTV cameras are not present or not functional.",
            "CCTV camera covering complete branch area?": "CCTV cameras are not covering the complete branch area.",
            "Check DVR working properly?": "DVR is not working properly or not recording footage.",
            "Is NTP configured in the DVR?": "NTP is not configured in the DVR.",
            "DVR is available in bank network or not?": "DVR is not available in the bank network.",
            "Camera is working in night vision.": "Cameras are not working in night vision mode.",
            "CCTV camera history recording duration.": "Sufficient Time period recording was not available with the bank.",
            "Locker CCTV camera history recording duration.": "Sufficient Time period recording was not available with the bank for the Branch Locker.",
            "Biometric Devices or Proximity Card is available or not?": "Biometric devices or proximity cards are not available.",
            "ID card issues to the Employee?": "ID cards are not issued to employees.",
            "Whether guard is available at branch or not?": "Guard is not available at the branch.",
            "Are secure areas controlled?": "Secure areas are not properly controlled.",
            "Smoke Detector is installed in the branch?": "Smoke detector is not installed in the branch.",
            "Whether Panic Switch installed?": "Panic switch is not installed.",
            "If a panic switch or smoke detector is installed at the branch or head office does the notification reach higher authorities when triggered?": "Notification does not reach higher authorities when panic switch or smoke detector is triggered.",
            "LAN ports are open or not?": "LAN ports are open and not secured.",
            "Visitor Register is maintained or not?": "Visitor register is not maintained.",
            "Any record maintains for drill the environmental controls?": "Records are not maintained for environmental control drills.",
            "Whether external modems, data cards etc., are being used in the Branch to access internet?": "External modems and data cards are being used in the branch.",
            "Fire extinguisher is available or not?": "Fire extinguisher is not available.",
            "Enough fire-extinguisher is available for area?": "Enough fire extinguishers are not available for the area.",
            
            # Power Back Up
            "Does the Bank have made enough Battery Backup which supports Computer Systems?": "UPS not able to provide sufficient power supply.",
            "Do the AMC is given for power backup systems?": "An Annual Maintenance Contract (AMC) is not maintained for the power backup systems.",
            "Generator is available or not?": "Generator is not available.",
            
            # User Awareness
            "Is training conducted for end user with respect to the Information Security?": "Training was not conducted for end user with respect to information security.",
            "Is training conducted for end user with respect to the application usage?": "Training was not conducted for end user with respect to the application usage.",
            "Employees are aware about using of fire extinguisher & Panic Switch?": "Employees were not aware of using fire extinguishers and panic switch.",
            "User was aware about checking CCTV recording history?": "Employees were not aware about checking CCTV recording history.",
            
            # Maintenance and Business Continuity Controls
            "Is AMC is given for computer and related systems?": "AMC was not managed for the Computer and related System.",
            
            # Patch Management
            "Antivirus available in all the desktop?": "Anti-virus was not available in systems.",
            "Antivirus updated in all the desktop?": "Anti-virus was not updated in systems.",
            "Windows patch updated in all the system?": "Windows patch was not updated in systems.",
            "Are there any system on which outdated version of windows is installed?": "Outdated windows version was installed in systems.",
            "Expired License antivirus installed?": "Expired license of anti-virus was installed in systems.",
            
            # Network Security
            "Are controls in place to ensure users only have access to the network resources they have been specially authorized to use and are required for their duties?": "Network devices were not arranged in a secure location.",
            "Is there network diagram available?": "Network diagram was not available at the branch.",
            "Is structured cabling is observed?": "Structured cabling was not present in the branch.",
            "Is cable tagging is observed?": "Cable tagging was not available in the branch.",
            "Is provision made for backup network?": "Branch is currently working on single connectivity.",
            "Networking devices are under secure location and there is any cabinet facility is provided or not?": "Cabinet facility was there and networking devices were not placed inside the cabinet, however it was not locked.",
            "Proper Cooling for Networking Devices available or not?": "Proper cooling for networking devices was not available.",
            "Wi-Fi is available or not?": "Wi-Fi was available in the branch.",
            "Whether Firewall is available in Branch ?": "Firewall is not available in Branch.",
            
            # Endpoints Vulnerability
            "Does User login with Administrator Rights?": "Admin rights are not restricted in the systems.",
            "Can Group Policy be modified?": "Group Policy can be modified in the systems.",
            "Can Security Configuration Policy be modified?": "Security Configuration Policy can be modified in the systems.",
            "Whether Proxy can be modified by branch users?": "Proxy can be modified in the systems.",
            "Whether internet is allowed or not?": "Internet was allowed.",
            "Whether internet access is restricted or unrestricted?": "Unrestricted internet access is given to users. It is recommended to implement whitelisting concept.",
            "Authentication Policy for internet is available or not?": "Authentication policy was not available for accessing internet.",
            "System Password is weak or not?": "System password is weak and common. Users need to follow standard password policy for accessing systems.",
            "System tagging is available or not?": "System tagging was not present in the systems.",
            "Is NTP configured in the systems?": "NTP was not configured in systems.",
            "Whether USB ports/CD drives enabled in the System?": "USB ports/CD drives were enabled in the systems.",
            "USB Approval process?": "USB approval process was not available.",
            "If bank have trusted USB, then check the USB open outside the network?": "Bank does not have trusted USB.",
            "Whether Firewall is enable in systems?": "Firewall is not enabled in the systems.",
            "Whether all external storage media are checked for virus?": "External storage media are not scanned for virus.",
            "Whether Remote Desktop Protocol is enabled or disabled?": "Remote Desktop Protocol (RDP) is not disabled on the system.",
            
            # ATM Machine Room
            "Check the CCTV camera  LAN cable and ATM machine power cable concealed?": "CCTV camera LAN cable and ATM machine power cable was not concealed.",
            "Guard is available or not?": "Guard was not available.",
            "CCTV Camera and check the position of camera?": "CCTV Camera was not installed in proper position.",
            "ATM is working 24*7?": "ATM was not working 24*7.",
            "Whether the Do's and Don't for the user awareness is available or not?": "Do's and Don't Instructions are not available.",
            "Whether the ATM is network segmented or not?": "Network is not segmented for ATM.",
            "Whether ATM Machine is properly grouted (floor and wall) ?": "Machine is not properly grouted in floor and wall.",
            "Whether access to ATM room for maintenance purpose is restricted to the authorized persons only?": "Access to ATM room for maintenance purpose was not restricted.",
            "Whether ATM power back up is supported by UPS.": "ATM power back up was not available.",
            "System of ATM cash replenishment, adherence to dual control mechanism records is maintained properly?": "System of ATM cash replenishment, adherence to dual control mechanism records is not maintained properly.",
            "Check system of ATM reconciliation status maintained?": "System of ATM reconciliation status is not maintained.",
            "Whether ATM machine preservation is done of journal print for future reference?": "ATM machine does not preserve journal print for future reference.",
            
            # Email Security
            "Personal Mail used or not?": "Personal email id is in use.",
            "Two factor authentications are in used or not?": "Two factor authentication was not enabled for mail. It is recommended to enable Two factor authentication.",
            "Single mail with multiple users?": "Single mail is used by multiple users. Accountability cannot be established in case of any data theft.",
            "Whether official email ID gets open outside bank N/w?": "Bank official email id can be accessed outside of the bank network.",
            
            # Remote Access
            "Which Department is using Remote Access?": "Other Department or Person is using remote access.",
            "Remote Access approval process?": "Remote access approval process was not available.",
            
            # Unauthorized Applications / Personal Data
            "Unauthorized applications have been found in the below systems IP's": "Unauthorized applications have been found in the systems.",
            "Personal Data have been found in the below systems IP's": "Personal Data have been found in the systems.",
            
            # Important Note
            "Auditor Identity was not verified.": "Auditor Identity was not verified by organization's employee.",
            "Windows was not activated.": "Windows was not activated in systems.",
            "Password was saved in browser.": "Password was saved in browser at systems.",
            "Password was written on wall, desk, notepad, diary.": "Password was written on wall, desk, notepad, diary at premises.",
            "Asset Movement Register was not available.": "Asset Movement Register was not available in Branch.",
            "Dust was present on the network devices & systems.": "Dust was present on the network devices and systems.",
            "Auto switchable mode was not available for connectivity.": "Auto switchable mode was not available/enable for connectivity.",
            "Preventive Maintenance is not carried out for Systems/UPS/Network Devices.": "Preventive Maintenance is not carried out for the Systems/UPS/Network Devices."
        }
        
        # Define recommendation mapping for all observations (Column H)
        recommendation_mapping = {
            "Do employees are using strong passwords?": "It is recommended to enforce a strong password policy requiring a mix of uppercase, lowercase, numbers, and special characters, along with a minimum password length of 8–12 characters. Regular awareness training should also be conducted to promote password security.",
            "Do you mandate periodical password changes?": "It is recommended to implement a password expiry policy, mandating users to change their passwords every 45–90 days to minimize unauthorized access risks due to compromised credentials.",
            "Do passwords are shared among employees?": "It is recommended to strictly prohibit password sharing among employees and implement user-specific logins to ensure accountability and traceability in system activities.",
            "What happens after consecutive failed login attempts?": "It is recommended to configure account lockout mechanisms after a defined number of failed login attempts (e.g., 3–5) to protect against brute-force attacks and unauthorized access attempts.",
            "Is session timeout enforced after a pre-defined period of inactivity?": "It is recommended to enforce automatic session timeouts after a defined period of inactivity (e.g., 10–15 minutes) to prevent unauthorized access in case a session is left unattended.",
            "Whether Multiple Logins is enabled?": "It is recommended to restrict multiple concurrent logins for a single user ID to prevent misuse of credentials and to enhance access control and accountability.",
            "Is the CBS System isolated from the internet?": "It is recommended to keep the CBS system isolated from direct internet connectivity and access it only through secure, controlled internal networks to reduce exposure to cyber threats.",
            "Is internet access restricted to only trusted and officially approved sites?": "It is recommended to configure firewalls and web filters to allow internet access only to trusted and approved websites necessary for official use, minimizing the risk of malware or phishing attacks.",
            "Is there any policy for default password while unlocking the account?": "It is recommended to define and implement a policy where default or temporary passwords are system-generated, unique, and require immediate change upon first login.",
            "When a new user joins then how do you give the user ID and Password for the user by using mail or telephonic conversation?": "It is recommended to share new user credentials through secure channels such as encrypted emails or in-person delivery and avoid sharing via unsecured telephonic conversations.",
            "SSL is using or not in CBS Application?": "It is recommended to ensure SSL/TLS encryption is implemented in CBS communication to protect data in transit from eavesdropping or tampering.",
            "CBS is browser based or desktop based?": "It is recommended to ensure that the CBS application, whether browser-based or desktop-based, is regularly updated and secured through proper patch management and endpoint protection mechanisms.",
            "Is the CBS compatible with the latest version of the browser?": "It is recommended to ensure CBS application compatibility with the latest stable browser versions and security updates to prevent vulnerabilities due to outdated components.",
            "Whether Two Factor Authentication is implemented for CBS login?": "It is recommended to implement Two-Factor Authentication (2FA) for CBS logins to add an extra layer of security beyond passwords, reducing the risk of unauthorized access.",
            "CCTV Camera Present?": "It is recommended to install CCTV cameras in all critical areas of the branch, including entry/exit points, server room, and cash handling areas, to ensure continuous monitoring and incident recording.",
            "Branch Locker CCTV Camera Present?": "It is recommended to ensure CCTV cameras are installed inside and outside the locker room with proper coverage and functioning to monitor all locker-related activities.",
            "CCTV camera covering complete branch area?": "It is recommended to ensure complete branch coverage by CCTV, leaving no blind spots, and periodically review camera placement to enhance security surveillance.",
            "Check DVR working properly?": "It is recommended to verify that the DVR/NVR system is working properly, recording continuously, and backed up regularly to avoid data loss during security incidents.",
            "Is NTP configured in the DVR?": "It is recommended to configure NTP in the DVR to synchronize date and time accurately, ensuring recorded footage timestamps remain reliable for audit and investigation.",
            "DVR is available in bank network or not?": "It is recommended to connect DVR securely to the bank's internal network with restricted access to authorized users only, ensuring monitoring and management through secure channels.",
            "Camera is working in night vision.": "It is recommended to ensure all CCTV cameras have functioning infrared or night-vision capabilities to maintain surveillance during low-light or night hours.",
            "CCTV camera history recording duration.": "It is recommended to maintain CCTV footage for at least 90 days (or as per regulatory requirements) and ensure adequate storage capacity for continuous recording.",
            "Locker CCTV camera history recording duration.": "It is recommended to retain locker CCTV footage for a minimum of 180 days or as per the bank's policy and RBI guidelines to support investigation needs if required.",
            "Biometric Devices or Proximity Card is available or not?": "It is recommended to implement biometric or proximity card-based access control systems to restrict unauthorized access to sensitive areas like server or locker rooms.",
            "ID card issues to the Employee?": "It is recommended to issue official ID cards to all employees and enforce mandatory wearing within premises to ensure identification and prevent unauthorized entry.",
            "Whether guard is available at branch or not?": "It is recommended to appoint trained security guards during branch working hours and after-hours to safeguard assets and respond promptly to security incidents.",
            "Are secure areas controlled?": "It is recommended to restrict access to secure areas such as server rooms, lockers, and cash vaults through controlled mechanisms like access cards or biometric authentication.",
            "Smoke Detector is installed in the branch?": "It is recommended to install smoke detectors across the branch, including critical areas, to detect and alert early signs of fire or smoke hazards.",
            "Whether Panic Switch installed?": "It is recommended to install panic switches at strategic locations like teller counters and locker rooms to enable quick alerting in case of emergency or threat.",
            "If a panic switch or smoke detector is installed at the branch or head office does the notification reach higher authorities when triggered?": "It is recommended to integrate panic switches and smoke detectors with a central monitoring or alerting system to ensure immediate notification to higher authorities and security personnel.",
            "LAN ports are open or not?": "It is recommended to disable all unused LAN ports and secure active ones with MAC binding or port security to prevent unauthorized network access.",
            "Visitor Register is maintained or not?": "It is recommended to maintain a visitor logbook at the branch entrance to record visitor details, purpose, and time of entry/exit for accountability and tracking.",
            "Any record maintains for drill the environmental controls?": "It is recommended to conduct and document periodic environmental and safety drills, including fire and evacuation exercises, to ensure employee preparedness during emergencies.",
            "Whether external modems, data cards etc., are being used in the Branch to access internet?": "It is recommended to restrict the use of external modems, data cards, or Wi-Fi devices in the branch and allow internet access only through secured and approved bank networks.",
            "Fire extinguisher is available or not?": "It is recommended to ensure fire extinguishers are installed in all key areas and employees are trained in their proper use for effective fire control.",
            "Enough fire-extinguisher is available for area?": "It is recommended to ensure the number and type of fire extinguishers are adequate for the branch size and layout, and to conduct periodic inspections for refilling and maintenance.",
            "Does the Bank have made enough Battery Backup which supports Computer Systems?": "It is recommended to ensure sufficient UPS and battery backup are available to support all critical computer systems, network devices, and servers for a minimum of 30–60 minutes during power failures to maintain uninterrupted operations.",
            "Do the AMC is given for power backup systems?": "It is recommended to maintain an Annual Maintenance Contract (AMC) for all UPS, inverter, and power backup systems to ensure timely servicing, health checks, and replacement of faulty batteries or components.",
            "Generator is available or not?": "It is recommended to install and maintain a functional generator at the branch to provide alternate power during extended outages, ensuring that all critical banking services remain operational without disruption.",
            "Is training conducted for end user with respect to the Information Security?": "It is recommended to conduct regular information security awareness training for all employees to educate them on phishing, password management, data handling, and incident reporting to reduce human-related security risks.",
            "Is training conducted for end user with respect to the application usage?": "It is recommended to provide periodic hands-on training for end users on core banking and other critical applications to ensure proper usage, minimize operational errors, and enhance efficiency.",
            "Employees are aware about using of fire extinguisher & Panic Switch?": "It is recommended to conduct periodic safety and emergency response training, including demonstrations on the correct use of fire extinguishers and panic switches, to ensure quick and effective response during emergencies.",
            "User was aware about checking CCTV recording history?": "It is recommended to train designated branch staff on how to access and verify CCTV recording history to ensure footage is available, properly retained, and can be retrieved promptly during investigations.",
            "Is AMC is given for computer and related systems?": "It is recommended to maintain an Annual Maintenance Contract (AMC) for all computer systems, peripherals, and IT equipment to ensure regular preventive maintenance, timely repairs, and minimal downtime during operations.",
            "Is Complaint register maintained?": "It is recommended to maintain a complaint or issue register (manual or digital) to record all IT and infrastructure-related issues along with resolution details, helping in tracking recurring problems and improving service response.",
            "Antivirus available in all the desktop?": "It is recommended to install licensed antivirus software on all desktops and laptops to protect against malware, ransomware, and other cyber threats, ensuring complete endpoint protection across the network.",
            "Antivirus updated in all the desktop?": "It is recommended to enable automatic antivirus updates on all systems to ensure the latest virus definitions and security patches are applied, providing real-time protection against evolving threats.",
            "Windows patch updated in all the system?": "It is recommended to apply the latest Windows security patches and updates regularly through centralized patch management to fix vulnerabilities and enhance system stability and security.",
            "Are there any system on which outdated version of windows is installed?": "It is recommended to upgrade or replace systems running outdated or unsupported versions of Windows to ensure continued security updates, compatibility, and compliance with regulatory requirements.",
            "Expired License antivirus installed?": "It is recommended to renew or replace all expired antivirus licenses immediately to restore real-time protection and maintain effective security coverage across all endpoints.",
            "Are controls in place to ensure users only have access to the network resources they have been specially authorized to use and are required for their duties?": "It is recommended to implement role-based access controls (RBAC) and periodic user access reviews to ensure employees can only access the network resources necessary for their job responsibilities, reducing the risk of misuse or unauthorized access.",
            "Is there network diagram available?": "It is recommended to maintain an up-to-date network diagram showing all devices, connections, and data flows. This helps in troubleshooting, security audits, and ensuring proper documentation for compliance.",
            "Is structured cabling is observed?": "It is recommended to ensure all network cabling follows structured cabling standards for organized layout, ease of maintenance, and minimal interference, improving both performance and reliability.",
            "Is cable tagging is observed?": "It is recommended to label or tag all network cables clearly to identify their source and destination, enabling quicker troubleshooting, maintenance, and infrastructure management.",
            "Is provision made for backup network?": "It is recommended to maintain a redundant or backup network connection (such as secondary ISP or failover link) to ensure uninterrupted connectivity during primary link failures or outages.",
            "Networking devices are under secure location and there is any cabinet facility is provided or not?": "It is recommended to place all networking devices like switches, routers, and firewalls inside a locked cabinet or server room with restricted access to authorized personnel only.",
            "Proper Cooling for Networking Devices available or not?": "It is recommended to provide adequate ventilation or air conditioning in areas housing network equipment to prevent overheating and ensure stable device performance.",
            "Wi-Fi is available or not?": "It is recommended to secure all Wi-Fi networks using WPA3 or at least WPA2 encryption, disable SSID broadcasting for internal networks, and allow access only to authorized users through MAC binding or authentication.",
            "Whether Firewall is available in Branch ?": "It is recommended to install and configure a network firewall at the branch level to monitor and control incoming and outgoing traffic, ensuring protection from external threats and unauthorized access.",
            "Does User login with Administrator Rights?": "It is recommended to restrict administrative rights to authorized IT personnel only and ensure normal users log in with limited privileges to prevent unauthorized system changes and malware execution.",
            "Can Group Policy be modified?": "It is recommended to restrict Group Policy modification access to central IT administrators only and regularly review configurations to ensure branch users cannot alter security or system settings.",
            "Can Security Configuration Policy be modified?": "It is recommended to protect security configuration policies with administrative controls and prevent local modifications by enforcing centralized policy management and regular integrity checks.",
            "Whether Proxy can be modified by branch users?": "It is recommended to disable proxy modification for end users and manage proxy settings centrally through Group Policy to prevent bypassing of internet access restrictions and monitoring controls.",
            "Whether internet is allowed or not?": "It is recommended to allow internet access only on systems where it is operationally required, and to monitor and log all access to ensure compliance with security policies.",
            "Whether internet access is restricted or unrestricted?": "It is recommended to restrict internet access to only business-related websites using firewall and web filtering tools to minimize the risk of malware, phishing, or data leakage.",
            "Authentication Policy for internet is available or not?": "It is recommended to implement an authentication policy (such as AD or proxy-based login) for internet access, ensuring accountability and traceability of all user activities.",
            "System Password is weak or not?": "It is recommended to enforce a strong password policy requiring complex passwords, regular changes, and password history restrictions to prevent unauthorized access.",
            "System tagging is available or not?": "It is recommended to tag all systems with unique identification numbers or asset codes for inventory tracking, accountability, and easier maintenance or audit verification.",
            "Is NTP configured in the systems?": "It is recommended to configure NTP in all systems to maintain synchronized and accurate timestamps, which is essential for event correlation, logging, and forensic investigations.",
            "Whether USB ports/CD drives enabled in the System?": "It is recommended to disable unused USB ports and CD drives to prevent unauthorized data transfer and malware infection. If required, allow access only through trusted and approved devices.",
            "USB Approval process?": "It is recommended to define and implement a formal USB approval process, allowing only authorized users and trusted encrypted USB devices for specific official purposes.",
            "If bank have trusted USB, then check the USB open outside the network?": "It is recommended to restrict trusted USB devices from being used outside the bank network through encryption and device control solutions to prevent data leakage or misuse.",
            "Whether Firewall is enable in systems?": "It is recommended to enable and configure Windows Firewall or equivalent endpoint firewalls on all systems to block unauthorized inbound/outbound traffic and enhance endpoint security.",
            "Whether all external storage media are checked for virus?": "It is recommended to mandate antivirus scanning for all external storage media before use within the network to prevent malware introduction from external sources.",
            "Whether Remote Desktop Protocol is enabled or disabled?": "It is recommended to disable Remote Desktop Protocol (RDP) on end-user systems unless specifically required, and where enabled, secure it with strong authentication and limited IP access.",
            "Check the CCTV camera  LAN cable and ATM machine power cable concealed?": "It is recommended to properly conceal CCTV and ATM power/LAN cables within conduits or walls to prevent tampering, accidental damage, and unauthorized access to the network or power line.",
            "Guard is available or not?": "It is recommended to deploy a trained security guard at the ATM premises round-the-clock to ensure physical security, assist customers, and respond to any suspicious activities.",
            "CCTV Camera and check the position of camera?": "It is recommended to ensure CCTV cameras are correctly positioned to cover the ATM area, entry/exit points, and customer space while maintaining user privacy at the keypad.",
            "ATM is working 24*7?": "It is recommended to maintain continuous 24×7 operation of the ATM with proper power backup, network connectivity, and regular health monitoring to minimize downtime and customer inconvenience.",
            "Whether the Do's and Don't for the user awareness is available or not?": "It is recommended to display clear Do's and Don'ts signage near the ATM machine to educate users on safe transaction practices and discourage sharing of card/PIN information.",
            "Whether the ATM is network segmented or not?": "It is recommended to ensure the ATM network is properly segmented from the bank's internal LAN to minimize cybersecurity risks and prevent unauthorized lateral movement in case of a breach.",
            "Whether ATM Machine is properly grouted (floor and wall) ?": "It is recommended to ensure the ATM machine is securely grouted to the floor and/or wall to prevent physical theft or unauthorized movement of the machine.",
            "Whether access to ATM room for maintenance purpose is restricted to the authorized persons only?": "It is recommended to restrict maintenance access to authorized vendor personnel only and maintain entry/exit records to ensure accountability and prevent unauthorized access.",
            "Whether ATM power back up is supported by UPS.": "It is recommended to connect the ATM to a UPS or inverter backup system to maintain operations during power outages and prevent transaction failures.",
            "System of ATM cash replenishment, adherence to dual control mechanism records is maintained properly?": "It is recommended to strictly follow the dual control mechanism during cash replenishment and maintain proper records for every cash loading/unloading activity to ensure accountability and prevent fraud.",
            "Check system of ATM reconciliation status maintained?": "It is recommended to regularly reconcile ATM transactions and maintain reconciliation reports to detect and resolve discrepancies promptly between CBS and ATM transaction logs.",
            "Whether ATM machine preservation is done of journal print for future reference?": "It is recommended to preserve ATM journal printouts or electronic logs as per regulatory requirements to support dispute resolution and transaction verification during audits or investigations.",
            "Personal Mail used or not?": "It is recommended to restrict the use of personal email accounts for official communication. All employees should use only authorized official email IDs to prevent data leakage and ensure message traceability.",
            "Two factor authentications are in used or not?": "It is recommended to implement Two-Factor Authentication (2FA) for all official email accounts to enhance account security and protect against unauthorized access or credential compromise.",
            "Single mail with multiple users?": "It is recommended to assign unique email IDs to every employee instead of using shared accounts, ensuring individual accountability, audit tracking, and secure communication practices.",
            "Whether official email ID gets open outside bank network?": "It is recommended to restrict access to official email accounts from outside the bank network unless accessed via secure VPN or authorized remote access solution, to prevent unauthorized logins and data compromise.",
            "Which Department is using Remote Access?": "It is recommended that only the IT Department be allowed remote access to critical systems. Access should be restricted to authorized personnel for maintenance or monitoring purposes only. Any additional departmental access should be granted strictly based on business need and with proper approval.",
            "Remote Access approval process?": "It is recommended that a formal remote access approval process be established, where requests are reviewed and authorized by the IT Head or higher authority. The approval should include justification, duration, and user details, and all granted access must be logged and periodically reviewed.",
            "There are no unauthorized applications on any systems": "It is recommended that all systems be periodically scanned and monitored to detect unauthorized applications. Such applications should be uninstalled immediately, and only approved software from the IT Department's whitelist should be permitted. A strict application control policy should be enforced to prevent recurrence.",
            "Unauthorized applications have been found in the below systems IP's": "It is recommended that all systems be periodically scanned and monitored to detect unauthorized applications. Such applications should be uninstalled immediately, and only approved software from the IT Department's whitelist should be permitted. A strict application control policy should be enforced to prevent recurrence.",
            "There are no Personal Data on any systems": "It is recommended to identify and classify all personal data residing on systems. Such data should be stored only on secure, authorized locations in compliance with data protection guidelines. Personal data on general user systems should be restricted or encrypted to prevent misuse or unauthorized access.",
            "Personal Data have been found in the below systems IP's": "It is recommended to identify and classify all personal data residing on systems. Such data should be stored only on secure, authorized locations in compliance with data protection guidelines. Personal data on general user systems should be restricted or encrypted to prevent misuse or unauthorized access.",
            "Auditor Identity was not verified.": "It is recommended that the bank verify and record the identity of all visiting auditors before granting access. A proper visitor verification and authorization process should be implemented at the entry point to ensure accountability and security of premises.",
            "Windows was not activated.": "It is recommended to activate all Windows operating systems with valid licenses to ensure authenticity, security updates, and vendor support. Unlicensed or inactive systems may miss critical patches, increasing vulnerability risks.",
            "Password was saved in browser.": "It is recommended to disable the save password feature in browsers and enforce the use of secure password managers or manual entry. Saving passwords in browsers increases the risk of credential theft in case of system compromise.",
            "Password was written on wall, desk, notepad, diary.": "It is recommended to strictly prohibit writing or displaying passwords in visible areas. Employees should be trained on password confidentiality, and strong password policies must be implemented to prevent unauthorized access.",
            "Asset Movement Register was not available.": "It is recommended to maintain a proper asset movement register to track IT equipment and related assets. The register should record asset details, responsible persons, and movement approvals to ensure traceability and accountability.",
            "Dust was present on the network devices & systems.": "It is recommended to clean and maintain all IT and network devices regularly to avoid dust accumulation, which can affect performance and hardware life. A preventive maintenance schedule should be implemented and documented.",
            "Auto switchable mode was not available for connectivity.": "It is recommended to configure an auto-switching mechanism between primary and backup connectivity (e.g., dual ISP or redundant link). This ensures uninterrupted operations in case of network or link failure.",
            "Preventive Maintenance is not carried out for Systems/UPS/Network Devices.": "It is recommended to implement a documented preventive maintenance schedule for all systems, UPS units, and network devices. Regular maintenance enhances performance, reduces downtime, and helps detect early signs of failure."
        }
        
        # Define impact mapping for all observations (Column F)
        impact_mapping = {
            "Global password policy is not followed in CBS access.": "The absence of a global password policy leads to weak password practices and increases the likelihood of unauthorized access to critical CBS systems. This may result in data compromise and non-compliance with RBI/NABARD cybersecurity directives.",
            "Password does not expires after 30 days. It is mandatory to change the password after 30 days.": "Without periodic password expiry, users may continue using old passwords for extended periods, heightening the risk of password theft and unauthorized access. This weakens overall system security and violates standard password management practices.",
            "Employees do share passwords with each other.": "Password sharing reduces individual accountability and allows misuse of user privileges. This poses a high risk of insider threats, data manipulation, and difficulty in forensic tracking during security incidents.",
            "Account does not get locked after three consecutive failed login attempts.": "Lack of account lockout controls enables brute-force attacks, where attackers can guess credentials through repeated attempts. This exposes CBS systems to unauthorized access and compromises user authentication integrity.",
            "Session timeout is less than 10 minutes.": "Short or misconfigured session timeouts can allow unattended terminals to remain active, increasing the risk of misuse or unauthorized access to sensitive CBS data by nearby users or malicious insiders.",
            "Multiple logins was enabled.": "Allowing multiple simultaneous logins increases the chance of account misuse or sharing, making it difficult to trace activities to specific users. This compromises user accountability and may violate access control requirements.",
            "CBS System is not isolated from the internet.": "Connecting CBS systems directly to the internet exposes them to external attacks, including malware infiltration and data exfiltration. This configuration severely violates network segmentation best practices mandated by regulatory bodies.",
            "Unrestricted internet access is given to the user.": "Unrestricted access to the internet increases the risk of phishing, malware downloads, and exposure to malicious websites. It also violates RBI guidelines on controlled internet access in critical banking systems.",
            "Default password is in use for unlocking the account.": "Using default passwords significantly weakens system security and makes unauthorized access easier for attackers. It indicates poor password hygiene and lack of enforcement of security controls at user provisioning.",
            "New user password creation process was not defined.": "Absence of a defined password creation and distribution process may lead to insecure communication of credentials, exposing them to interception or misuse. This also increases the risk of unauthorized access by unverified personnel.",
            "SSL was not implemented in CBS application.": "Without SSL encryption, sensitive banking transactions and credentials may be transmitted in plain text, exposing them to interception, man-in-the-middle attacks, and data theft over the network.",
            "Extension is used to run CBS application.": "Reliance on third-party or outdated extensions introduces compatibility and security risks. Such components may not be regularly updated and could contain vulnerabilities that compromise CBS integrity.",
            "CBS is not compatible with the latest version of the browser.": "Running CBS on outdated or unsupported browsers may result in security vulnerabilities and performance instability. It increases exposure to exploits that could compromise user sessions and transaction integrity.",
            "Two factor authentication is not implemented for CBS login.": "Lack of multi-factor authentication weakens the user verification process and increases the risk of unauthorized system access, even if credentials are compromised. It fails to meet modern security and compliance standards.",
            "CCTV cameras are not present or not functional.": "Non-functional or absent CCTV cameras severely affect the branch's ability to monitor activities, detect suspicious behavior, or investigate incidents. This increases the risk of undetected theft, fraud, or physical security breaches.",
            "Branch locker CCTV cameras are not present or not functional.": "Lack of functional CCTV coverage at the locker area compromises critical security monitoring. It can result in unrecorded incidents, making it difficult to investigate unauthorized locker access or internal fraud.",
            "CCTV cameras are not covering the complete branch area.": "Incomplete camera coverage leaves blind spots within the branch, providing opportunities for unauthorized activity without evidence. This reduces the effectiveness of surveillance and audit trail capabilities.",
            "DVR is not working properly or not recording footage.": "A malfunctioning DVR prevents continuous recording and storage of surveillance footage, hindering post-incident investigation. It may also cause non-compliance with RBI's mandate for minimum CCTV retention.",
            "NTP is not configured in the DVR.": "Without NTP synchronization, CCTV timestamps become inaccurate, making it difficult to correlate events with other system logs during investigations. This leads to inconsistencies in evidence and incident timelines.",
            "DVR is not available in the bank network.": "When DVRs are not integrated into the bank network, centralized monitoring and remote audit access are not possible. This results in delayed incident detection and increased operational inefficiency.",
            "Cameras are not working in night vision mode.": "Non-functional night vision restricts surveillance visibility during low-light hours, exposing the branch to security threats such as burglary or unauthorized access during non-working hours.",
            "Sufficient Time period recording was not available with the bank.": "Insufficient video retention reduces the bank's ability to review past incidents or respond to delayed complaints. It may lead to loss of critical evidence in case of disputes or theft.",
            "Sufficient Time period recording was not available with the bank for the Branch Locker.": "Lack of long-duration footage for locker areas can result in missing records of critical locker operations. This can impact accountability and make forensic investigation impossible in case of locker-related disputes.",
            "Biometric devices or proximity cards are not available.": "Absence of biometric or proximity-based access controls increases reliance on manual verification, which can be bypassed. This allows unauthorized individuals to access restricted areas.",
            "ID cards are not issued to employees.": "Without employee ID cards, verifying personnel identity within the premises becomes difficult. This increases the risk of impersonation and unauthorized access to sensitive or restricted areas.",
            "Guard is not available at the branch.": "Lack of a physical security guard weakens the first line of defense against theft or intrusion. It also increases the time required to respond to emergencies or physical breaches.",
            "Secure areas are not properly controlled.": "Weak control over secure zones (like server or locker rooms) exposes critical assets to unauthorized access. This increases the risk of data compromise, cash theft, and internal misuse.",
            "Smoke detector is not installed in the branch.": "Absence of smoke detectors increases the risk of undetected fire outbreaks. This could lead to severe property damage, data loss, and safety hazards for staff and customers.",
            "Panic switch is not installed.": "Without a panic switch, staff cannot quickly alert authorities in case of a security emergency. This delay can escalate robbery or violent incidents, endangering employees and assets.",
            "Notification does not reach higher authorities when panic switch or smoke detector is triggered.": "Failure in the alert mechanism prevents timely escalation during emergencies, delaying critical response actions and increasing potential damage or loss.",
            "LAN ports are open and not secured.": "Unsecured LAN ports allow unauthorized devices to connect to the internal network, leading to potential malware introduction or data theft. It violates basic network access control measures.",
            "Visitor register is not maintained.": "Lack of visitor entry logs prevents tracking of external individuals entering the branch. This weakens physical access accountability and investigation capabilities during incidents.",
            "Records are not maintained for environmental control drills.": "Without documentation of safety drills, there is no proof of staff preparedness for fire or disaster events. This indicates poor compliance with safety standards and emergency readiness.",
            "External modems and data cards are being used in the branch.": "Use of external internet devices bypasses network monitoring and firewall controls, creating potential backdoors for cyberattacks and data exfiltration from the internal network.",
            "Fire extinguisher is not available.": "Absence of fire extinguishers exposes the branch to uncontrolled fire hazards. This could lead to severe loss of infrastructure, data, and life in the event of a fire emergency.",
            "Enough fire extinguishers are not available for the area.": "Insufficient fire-fighting equipment increases the chance of small fires spreading uncontrollably, resulting in preventable damage and operational disruption.",
            "UPS not able to provide sufficient power supply.": "Inadequate UPS backup leads to abrupt system shutdowns during power failures, risking data corruption, hardware damage, and interruptions in CBS and transaction processing.",
            "An Annual Maintenance Contract (AMC) is not maintained for the power backup systems.": "Lack of AMC for UPS or inverter systems results in delayed maintenance and higher chances of equipment failure, directly affecting business continuity and data protection.",
            "Generator is not available.": "Absence of a generator leaves the branch without alternative power support during outages, causing downtime of CBS systems, ATM operations, and surveillance equipment.",
            "Training was not conducted for end user with respect to information security.": "Lack of user awareness training increases the likelihood of social engineering, phishing, and accidental data breaches. Employees remain unaware of safe cyber hygiene practices, which can compromise system security.",
            "Training was not conducted for end user with respect to the application usage.": "When users are not trained in application usage, operational errors and data entry mistakes increase. This affects system integrity, data accuracy, and business continuity.",
            "Employees were not aware of using fire extinguishers and panic switch.": "Lack of awareness about fire and emergency equipment results in ineffective response during critical situations. This can escalate small incidents into major safety hazards.",
            "Employees were not aware about checking CCTV recording history.": "Unawareness regarding CCTV monitoring practices may result in unnoticed security incidents. It delays detection of tampering or malfunctioning equipment, weakening surveillance effectiveness.",
            "AMC was not managed for the Computer and related System.": "Without proper AMC, system maintenance becomes irregular, increasing downtime and risk of hardware or software failure. This impacts daily banking operations and data availability.",
            "Complaint register was not maintained.": "Absence of a complaint log prevents systematic tracking and resolution of IT or operational issues. It leads to repeated problems, lack of accountability, and customer dissatisfaction.",
            "Anti-virus was not available in systems.": "Systems without antivirus protection remain vulnerable to malware, ransomware, and trojan attacks. This exposes sensitive banking data and customer information to compromise.",
            "Anti-virus was not updated in systems.": "Outdated antivirus definitions fail to detect new or emerging threats. This increases the risk of successful cyberattacks and malware infiltration within the bank network.",
            "Windows patch was not updated in systems.": "Missing or outdated system patches leave exploitable vulnerabilities unaddressed. Attackers can use these flaws to gain unauthorized access or deploy ransomware in the internal network.",
            "Outdated windows version was installed in systems.": "Unsupported operating systems no longer receive security updates, increasing susceptibility to zero-day exploits and compliance violations. This poses a serious operational and regulatory risk.",
            "Expired license of anti-virus was installed in systems.": "An expired antivirus license stops updates and real-time protection, leaving systems open to malware infections. This can lead to data breaches and non-compliance with cyber hygiene standards.",
            "Network devices were not arranged in a secure location.": "Improper placement of network devices exposes them to unauthorized access or tampering. Physical compromise of switches or routers can lead to internal network infiltration.",
            "Network diagram was not available at the branch.": "Absence of an updated network diagram hampers network management and incident response. It becomes difficult to isolate infected devices or troubleshoot connectivity issues.",
            "Structured cabling was not present in the branch.": "Unstructured cabling creates network instability and troubleshooting difficulties. It also increases the risk of accidental disconnection and prolonged downtime during maintenance.",
            "Cable tagging was not available in the branch.": "Without proper labeling, identifying cables during maintenance or troubleshooting becomes difficult. This can cause accidental disruptions and delay restoration during outages.",
            "Branch is currently working on single connectivity.": "Operating without redundant connectivity creates a single point of failure. Network downtime can directly impact CBS access, ATM transactions, and branch operations.",
            "Cabinet facility was there and networking devices were not placed inside the cabinet, however it was not locked.": "Unlocked cabinets allow unauthorized physical access to network devices, risking disconnection, data interception, or malicious reconfiguration.",
            "Proper cooling for networking devices was not available.": "Overheating of networking devices can cause unexpected shutdowns and reduce equipment lifespan. This leads to network outages and service degradation during business hours.",
            "Wi-Fi was available in the branch.": "Unsecured or unauthorized Wi-Fi access can allow attackers to connect into the bank's internal network. This creates serious data leakage and compliance risks if not properly segregated.",
            "Firewall is not available in Branch.": "Without a firewall, the branch network remains directly exposed to the internet. This increases vulnerability to external attacks, malware infections, and data exfiltration.",
            "Admin rights are not restricted in the systems.": "Allowing users administrative access increases the risk of unauthorized configuration changes, malware execution, and accidental system damage. It weakens endpoint security enforcement.",
            "Group Policy can be modified in the systems.": "If group policies can be altered by users, it compromises centralized security control. Attackers can disable protective measures or install malicious software undetected.",
            "Security Configuration Policy can be modified in the systems.": "When users can edit security configurations, it undermines standardization and allows bypassing of hardening controls. This leads to inconsistent and insecure system setups.",
            "Proxy can be modified in the systems.": "Unrestricted proxy modification lets users bypass web filtering or monitoring tools. It increases the risk of visiting malicious sites or leaking sensitive data outside the secured network.",
            "Internet was allowed.": "Uncontrolled internet access increases exposure to phishing, malware, and malicious downloads. It also leads to reduced employee productivity and potential data leakage.",
            "Unrestricted internet access is given to users. It is recommended to implement whitelisting concept.": "Allowing unrestricted browsing access heightens cyber risks and may lead to accidental access of harmful websites. Whitelisting ensures that users can only visit business-approved domains.",
            "Authentication policy was not available for accessing internet.": "Lack of user authentication while accessing the internet prevents tracking of browsing activity. This reduces accountability and makes it difficult to trace security incidents.",
            "System password is weak and common. Users need to follow standard password policy for accessing systems.": "Weak or reused passwords increase the likelihood of brute-force and credential theft attacks. It compromises individual accountability and overall system protection.",
            "System tagging was not present in the systems.": "Without asset tagging, tracking hardware devices and ownership becomes difficult. This creates challenges in asset management, theft detection, and audit reconciliation.",
            "NTP was not configured in systems.": "Unsynchronized system clocks cause log discrepancies, making correlation of security events inaccurate. It affects forensic investigations and audit evidence integrity.",
            "USB ports/CD drives were enabled in the systems.": "Enabled USB or media drives create opportunities for malware infection and unauthorized data transfer. This weakens endpoint data protection and compliance posture.",
            "USB approval process was not available.": "Absence of a defined USB approval process allows employees to use personal devices without authorization. This increases the risk of data exfiltration and malware propagation.",
            "Bank does not have trusted USB.": "Using non-trusted USB devices increases the chance of introducing malicious software or losing sensitive data. Trusted USB solutions ensure encryption and controlled usage.",
            "Firewall is not enabled in the systems.": "Disabled system firewalls allow unrestricted inbound and outbound traffic, leaving systems exposed to local and remote attacks. It reduces endpoint resilience against threats.",
            "External storage media are not scanned for virus.": "Failure to scan external media leads to potential malware introduction into the internal network. This can disrupt critical systems or leak confidential data.",
            "Remote Desktop Protocol (RDP) is not disabled on the system.": "Enabled RDP increases the attack surface for remote exploitation. Unauthorized users can gain control of systems, install malware, or steal data if access is not secured.",
            "CCTV camera LAN cable and ATM machine power cable was not concealed.": "Exposed cables increase the risk of tampering or unauthorized disconnection. This can lead to ATM or CCTV downtime, compromising branch security and surveillance integrity.",
            "Guard was not available.": "Absence of security personnel increases the risk of unauthorized access, theft, or physical security incidents. It reduces overall branch safety and may delay response to emergencies.",
            "CCTV Camera was not installed in proper position.": "Incorrect placement of cameras creates blind spots in surveillance. This can result in unmonitored areas, reducing the effectiveness of security monitoring and incident detection.",
            "ATM was not working 24*7.": "ATM downtime impacts customer service and can lead to dissatisfaction. It also affects cash availability and transactional reliability, impacting branch operations.",
            "Do's and Don't Instructions are not available.": "Lack of guidance for users and employees increases the likelihood of operational errors and security violations. This reduces awareness of safe practices in daily banking operations.",
            "Network is not segmented for ATM.": "Without proper network segmentation, a compromised ATM system can affect other bank systems. This increases the risk of malware spreading and potential data breaches.",
            "Machine is not properly grouted in floor and wall.": "Improper installation makes ATMs vulnerable to physical tampering or theft. This increases the risk of cash loss and security incidents.",
            "Access to ATM room for maintenance purpose was not restricted.": "Unrestricted access allows unauthorized personnel to tamper with ATM equipment. This can result in fraud, theft, or system manipulation.",
            "ATM power back up was not available.": "Lack of UPS or backup power causes ATM outages during power failure. This interrupts services and may lead to operational and financial losses.",
            "System of ATM cash replenishment, adherence to dual control mechanism records is not maintained properly.": "Failure to maintain proper dual control records increases risk of cash mismanagement, theft, or fraud. It reduces accountability during ATM operations.",
            "System of ATM reconciliation status is not maintained.": "Absence of reconciliation records may cause discrepancies in cash handling and reporting. This impacts financial accuracy and audit compliance.",
            "ATM machine does not preserve journal print for future reference.": "Lack of transaction records makes investigation of disputes or fraud incidents difficult. This compromises audit trail and accountability.",
            "Personal email id is in use.": "Use of personal emails increases risk of data leakage and reduces control over sensitive communications. It also violates official communication policies.",
            "Two factor authentication was not enabled for mail. It is recommended to enable Two factor authentication.": "Without 2FA, email accounts remain vulnerable to phishing and unauthorized access. This increases risk of sensitive data compromise and fraud.",
            "Single mail is used by multiple users. Accountability cannot be established in case of any data theft.": "Shared email accounts prevent traceability of actions, making it difficult to identify the responsible person in case of misuse or security incidents.",
            "Bank official email id can be accessed outside of the bank network.": "External access increases the risk of credential compromise and unauthorized data access. It may lead to potential information leakage or phishing attacks.",
            "Other Department or Person is using remote access.": "Unauthorized remote access exposes systems to potential attacks and increases risk of data breaches or misuse of bank resources.",
            "Remote access approval process was not available.": "Absence of approval mechanism leads to unregulated access, reducing control over who can connect remotely. This increases vulnerability to insider threats and security breaches.",
            "Unauthorized applications have been found in the systems.": "Presence of unapproved software can introduce malware, reduce system performance, and violate compliance policies. It increases the risk of data compromise.",
            "Personal Data have been found in the systems.": "Storing personal data on business systems violates data protection policies. It increases risk of privacy breaches, regulatory penalties, and misuse of sensitive information.",
            "Auditor Identity was not verified by organization's employee.": "Failure to verify auditor identity risks unauthorized individuals gaining access to sensitive information, potentially compromising audit integrity.",
            "Windows was not activated in systems.": "Unactivated Windows prevents installation of critical security updates. This leaves systems vulnerable to malware, exploits, and operational instability.",
            "Password was saved in browser at systems.": "Storing passwords in browsers increases risk of credential theft. Unauthorized users can gain system access if the device is compromised.",
            "Password was written on wall, desk, notepad, diary at premises.": "Visible passwords allow easy access to unauthorized personnel. This undermines account security and increases the risk of data breaches.",
            "Asset Movement Register was not available in Branch.": "Without proper tracking, asset movements cannot be verified. This increases the risk of theft, loss, or mismanagement of bank assets.",
            "Dust was present on the network devices and systems.": "Accumulated dust may lead to overheating, hardware failure, and reduced equipment lifespan. It can cause unexpected downtime and affect branch operations.",
            "Auto switchable mode was not available/enable for connectivity.": "Without automatic failover, network or power failure may disrupt operations. This affects CBS, ATM services, and business continuity.",
            "Preventive Maintenance is not carried out for the Systems/UPS/Network Devices.": "Failure to perform preventive maintenance increases the risk of unexpected system and network failures, causing operational disruptions and reduced reliability."
        }
        
        # Now populate data rows starting from Row 4
        if analysis_results:
            # Helper function to normalize unauthorized apps/personal data observations
            def normalize_observation(observation):
                """Normalize observation by removing IP addresses for unauthorized apps/personal data"""
                obs_stripped = observation.strip()
                
                # Check if it matches any of the unauthorized apps/personal data prefixes
                for prefix in unauthorized_apps_personal_data_prefixes:
                    if obs_stripped.startswith(prefix):
                        # Return just the prefix without IP addresses
                        return prefix
                
                # Otherwise return the original observation
                return observation
            
            # Group all non-compliance entries by Column B data
            grouped_issues = {}
            
            for worksheet_data in analysis_results:
                worksheet_name = worksheet_data['worksheet_name']
                branch_name = extract_branch_name_from_worksheet(worksheet_name)
                
                for entry in worksheet_data['non_compliance_entries']:
                    a_data = entry.get('column_a_data', '')
                    b_data = entry['column_b_data']
                    d_data = entry.get('column_d_data', '')
                    row_num = entry['row_number']
                    images = entry.get('images', [])
                    
                    # Normalize the observation (removes IP addresses for unauthorized apps/personal data)
                    normalized_b_data = normalize_observation(b_data)
                    
                    # Create a unique key for this issue using normalized data
                    issue_key = normalized_b_data.strip()
                    
                    if issue_key not in grouped_issues:
                        grouped_issues[issue_key] = {
                            'column_a_data': a_data,
                            'issue_description': normalized_b_data,  # Use normalized version
                            'column_d_data': d_data,
                            'branches': [],
                            'total_occurrences': 0,
                            'row_numbers': [],
                            'images': []  # Store images from all branches
                        }
                    
                    # Add this branch to the list if not already present
                    if branch_name not in grouped_issues[issue_key]['branches']:
                        grouped_issues[issue_key]['branches'].append(branch_name)
                    
                    # Add row number to the list
                    grouped_issues[issue_key]['row_numbers'].append(row_num)
                    grouped_issues[issue_key]['total_occurrences'] += 1
                    
                    # Add images to the list (up to 5 total)
                    if len(grouped_issues[issue_key]['images']) < 5:
                        remaining_slots = 5 - len(grouped_issues[issue_key]['images'])
                        grouped_issues[issue_key]['images'].extend(images[:remaining_slots])
            
            # Sort issues by the minimum row number for each issue
            sorted_issues = sorted(grouped_issues.items(), key=lambda x: min(x[1]['row_numbers']))
            
            # Helper function to check if text starts with any prefix
            def matches_any_prefix(text, prefixes):
                text_stripped = text.strip()
                for prefix in prefixes:
                    if text_stripped.startswith(prefix):
                        return True
                return False
            
            # Separate issues into multiple categories
            cbs_issues = []
            physical_environmental_issues = []
            power_backup_issues_list = []
            user_awareness_issues = []
            maintenance_bc_issues = []
            patch_management_issues = []
            network_security_issues = []
            endpoints_vulnerability_issues = []
            atm_machine_room_issues = []
            email_security_issues = []
            remote_access_issues = []
            unauthorized_apps_issues = []
            important_note_issues = []  # Catch-all for remaining items
            
            for issue_key, issue_data in sorted_issues:
                issue_description = issue_data['issue_description']
                categorized = False
                
                # Check each category in order
                if issue_description.strip() in cbs_access_control_items:
                    cbs_issues.append((issue_key, issue_data))
                    categorized = True
                elif issue_description.strip() in physical_environmental_security_items:
                    physical_environmental_issues.append((issue_key, issue_data))
                    categorized = True
                elif issue_description.strip() in power_backup_items:
                    power_backup_issues_list.append((issue_key, issue_data))
                    categorized = True
                elif issue_description.strip() in user_awareness_items:
                    user_awareness_issues.append((issue_key, issue_data))
                    categorized = True
                elif issue_description.strip() in maintenance_business_continuity_items:
                    maintenance_bc_issues.append((issue_key, issue_data))
                    categorized = True
                elif issue_description.strip() in patch_management_items:
                    patch_management_issues.append((issue_key, issue_data))
                    categorized = True
                elif issue_description.strip() in network_security_items:
                    network_security_issues.append((issue_key, issue_data))
                    categorized = True
                elif issue_description.strip() in endpoints_vulnerability_items:
                    endpoints_vulnerability_issues.append((issue_key, issue_data))
                    categorized = True
                elif issue_description.strip() in atm_machine_room_items:
                    atm_machine_room_issues.append((issue_key, issue_data))
                    categorized = True
                elif issue_description.strip() in email_security_items:
                    email_security_issues.append((issue_key, issue_data))
                    categorized = True
                elif issue_description.strip() in remote_access_items:
                    remote_access_issues.append((issue_key, issue_data))
                    categorized = True
                elif matches_any_prefix(issue_description, unauthorized_apps_personal_data_prefixes):
                    unauthorized_apps_issues.append((issue_key, issue_data))
                    categorized = True
                
                # If not categorized, add to Important Note
                if not categorized:
                    important_note_issues.append((issue_key, issue_data))
            
            # Helper function to sort Important Note items by predefined order
            def sort_important_note_items(issues_list):
                """Sort Important Note items according to predefined order"""
                if not issues_list:
                    return issues_list
                
                # Create a mapping of issue description to its order index
                order_map = {item: idx for idx, item in enumerate(important_note_items_order)}
                
                # Sort issues based on their position in the predefined order
                # Items not in the order list will be placed at the end
                def get_sort_key(issue_tuple):
                    issue_key, issue_data = issue_tuple
                    issue_description = issue_data['issue_description'].strip()
                    # Return the order index if found, otherwise return a large number to put it at the end
                    return order_map.get(issue_description, len(important_note_items_order))
                
                sorted_issues = sorted(issues_list, key=get_sort_key)
                return sorted_issues
            
            # Helper function to add a category section
            def add_category_section(category_name, issues_list, current_row, use_custom_sort=False):
                if not issues_list:
                    return current_row
                
                # Apply custom sorting for Important Note section
                if use_custom_sort and category_name == 'Important Note':
                    issues_list = sort_important_note_items(issues_list)
                
                # Add category header row
                sheet.merge_cells(f'A{current_row}:M{current_row}')
                category_header_cell = sheet[f'A{current_row}']
                category_header_cell.value = category_name
                
                # Apply formatting
                category_header_cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
                category_header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                category_header_cell.fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')
                category_header_cell.border = thin_border
                
                # Apply border to all cells in merged range
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                    cell = sheet[f'{col}{current_row}']
                    cell.border = thin_border
                
                current_row += 1
                
                # Reset Sr. No. to 1 for each new category
                category_sr_no = 1
                
                # Create a mapping for Important Note items to their order numbers
                order_map = {item: idx + 1 for idx, item in enumerate(important_note_items_order)}
                
                # Add data rows for this category
                for issue_key, issue_data in issues_list:
                    column_a_data = issue_data.get('column_a_data', '')
                    issue_description = issue_data['issue_description']
                    column_d_data = issue_data.get('column_d_data', '')
                    branches = issue_data['branches']
                    images_list = issue_data.get('images', [])
                    
                    # For Important Note section, use the predefined order number in Column B
                    if use_custom_sort and category_name == 'Important Note':
                        reference_no = order_map.get(issue_description.strip(), '')
                    else:
                        reference_no = column_a_data
                    
                    # Replace content using all_content_replacement mapping
                    observations_text = all_content_replacement.get(issue_description.strip(), issue_description)
                    
                    # Get impact based on the replaced observations text
                    impact_text = impact_mapping.get(observations_text.strip(), '')
                    
                    # Get recommendation based on the original issue description
                    recommendation_text = recommendation_mapping.get(issue_description.strip(), '')
                    
                    # Populate the row
                    sheet[f'A{current_row}'].value = category_sr_no  # Use category-specific Sr. No.
                    sheet[f'B{current_row}'].value = reference_no
                    sheet.merge_cells(f'C{current_row}:D{current_row}')
                    sheet[f'C{current_row}'].value = observations_text
                    sheet[f'E{current_row}'].value = column_d_data  # Risk Factor (from Column D of source)
                    sheet[f'F{current_row}'].value = impact_text  # Impact (from impact mapping)
                    sheet[f'G{current_row}'].value = '\n'.join(sorted(branches))  # Branches (each on new line)
                    sheet[f'H{current_row}'].value = recommendation_text  # Recommendation column
                    sheet[f'I{current_row}'].value = ''  # POC column 1
                    sheet[f'J{current_row}'].value = ''  # POC column 2
                    sheet[f'K{current_row}'].value = ''  # POC column 3
                    sheet[f'L{current_row}'].value = ''  # POC column 4
                    sheet[f'M{current_row}'].value = ''  # POC column 5
                    
                    # Add images to columns J, K, L, M, I if available (one per column, up to 5 total, in order: J, K, L, M, I)
                    if images_list:
                        poc_columns = ['J', 'K', 'L', 'M', 'I']  # Order: J, K, L, M, I
                        for idx, img_data in enumerate(images_list[:5]):
                            if idx < len(poc_columns):
                                add_images_to_cell(sheet, f'{poc_columns[idx]}{current_row}', [img_data], max_images=1)
                    
                    # Apply formatting
                    for cell_ref in [f'A{current_row}', f'B{current_row}', f'C{current_row}', 
                                    f'D{current_row}', f'E{current_row}', f'F{current_row}', 
                                    f'G{current_row}', f'H{current_row}']:
                        cell = sheet[cell_ref]
                        cell.font = Font(name='Calibri', size=11)
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        cell.border = thin_border
                    
                    # Apply custom borders for POC columns
                    # Column I: Top, Bottom, Left borders only
                    sheet[f'I{current_row}'].font = Font(name='Calibri', size=11)
                    sheet[f'I{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    sheet[f'I{current_row}'].border = Border(
                        left=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Columns J, K, L: Top and Bottom borders only
                    for col in ['J', 'K', 'L']:
                        sheet[f'{col}{current_row}'].font = Font(name='Calibri', size=11)
                        sheet[f'{col}{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        sheet[f'{col}{current_row}'].border = Border(
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                    
                    # Column M: Top, Bottom, Right borders only
                    sheet[f'M{current_row}'].font = Font(name='Calibri', size=11)
                    sheet[f'M{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    sheet[f'M{current_row}'].border = Border(
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Center align Sr. No., Reference No., Risk Factor, and Branches
                    sheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet[f'G{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # Apply conditional formatting to Risk Factor (Column E) based on value
                    risk_cell = sheet[f'E{current_row}']
                    risk_value = str(column_d_data).strip().lower() if column_d_data else ''
                    
                    if 'critical' in risk_value:
                        # Dark Red background for Critical
                        risk_cell.fill = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')
                        risk_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                    elif 'high' in risk_value:
                        # Red background for High
                        risk_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                        risk_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                    elif 'medium' in risk_value:
                        # Orange background for Medium
                        risk_cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                        risk_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                    elif 'low' in risk_value:
                        # Dark Green background for Low
                        risk_cell.fill = PatternFill(start_color='006400', end_color='006400', fill_type='solid')
                        risk_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                    
                    current_row += 1
                    category_sr_no += 1  # Increment category-specific Sr. No.
                
                return current_row  # Only return current_row, not sr_no
            
            # Start populating data from row 4
            current_row = 4
            
            # First, add CBS Access Control issues (already has header in Row 3)
            # Reset Sr. No. to 1 for CBS Access Control section
            category_sr_no = 1
            
            for issue_key, issue_data in cbs_issues:
                column_a_data = issue_data.get('column_a_data', '')
                issue_description = issue_data['issue_description']
                column_d_data = issue_data.get('column_d_data', '')
                branches = issue_data['branches']
                total_occurrences = issue_data['total_occurrences']
                unique_row_numbers = sorted(set(issue_data['row_numbers']))
                images_list = issue_data.get('images', [])
                
                # Replace content for CBS Access Control items
                observations_text = cbs_content_replacement.get(issue_description.strip(), issue_description)
                
                # Get impact based on the replaced observations text
                impact_text = impact_mapping.get(observations_text.strip(), '')
                
                # Get recommendation based on the original issue description
                recommendation_text = recommendation_mapping.get(issue_description.strip(), '')
                
                # Populate the row
                sheet[f'A{current_row}'].value = category_sr_no  # Category-specific Sr. No.
                sheet[f'B{current_row}'].value = column_a_data  # Reference No. (Column A Data from source)
                
                # Merge C and D for Observations
                sheet.merge_cells(f'C{current_row}:D{current_row}')
                sheet[f'C{current_row}'].value = observations_text  # Observations (replaced content)
                
                sheet[f'E{current_row}'].value = column_d_data  # Risk Factor (from Column D of source)
                sheet[f'F{current_row}'].value = impact_text  # Impact (from impact mapping)
                sheet[f'G{current_row}'].value = '\n'.join(sorted(branches))  # Branches (each on new line)
                sheet[f'H{current_row}'].value = recommendation_text  # Recommendation column
                sheet[f'I{current_row}'].value = ''  # POC column 1
                sheet[f'J{current_row}'].value = ''  # POC column 2
                sheet[f'K{current_row}'].value = ''  # POC column 3
                sheet[f'L{current_row}'].value = ''  # POC column 4
                sheet[f'M{current_row}'].value = ''  # POC column 5
                
                # Add images to columns J, K, L, M, I if available (one per column, up to 5 total, in order: J, K, L, M, I)
                if images_list:
                    poc_columns = ['J', 'K', 'L', 'M', 'I']  # Order: J, K, L, M, I
                    for idx, img_data in enumerate(images_list[:5]):
                        if idx < len(poc_columns):
                            add_images_to_cell(sheet, f'{poc_columns[idx]}{current_row}', [img_data], max_images=1)
                
                # Apply formatting to data row
                data_cells = [f'A{current_row}', f'B{current_row}', f'C{current_row}', 
                             f'D{current_row}', f'E{current_row}', f'F{current_row}', 
                             f'G{current_row}', f'H{current_row}']
                
                for cell_ref in data_cells:
                    cell = sheet[cell_ref]
                    cell.font = Font(name='Calibri', size=11)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.border = thin_border
                
                # Apply custom borders for POC columns
                # Column I: Top, Bottom, Left borders only
                sheet[f'I{current_row}'].font = Font(name='Calibri', size=11)
                sheet[f'I{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                sheet[f'I{current_row}'].border = Border(
                    left=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Columns J, K, L: Top and Bottom borders only
                for col in ['J', 'K', 'L']:
                    sheet[f'{col}{current_row}'].font = Font(name='Calibri', size=11)
                    sheet[f'{col}{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    sheet[f'{col}{current_row}'].border = Border(
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                # Column M: Top, Bottom, Right borders only
                sheet[f'M{current_row}'].font = Font(name='Calibri', size=11)
                sheet[f'M{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                sheet[f'M{current_row}'].border = Border(
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Center align Sr. No., Reference No., Risk Factor, and Branches
                sheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet[f'G{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Apply conditional formatting to Risk Factor (Column E) based on value
                risk_cell = sheet[f'E{current_row}']
                risk_value = str(column_d_data).strip().lower() if column_d_data else ''
                
                if 'critical' in risk_value:
                    # Dark Red background for Critical
                    risk_cell.fill = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')
                    risk_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                elif 'high' in risk_value:
                    # Red background for High
                    risk_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    risk_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                elif 'medium' in risk_value:
                    # Orange background for Medium
                    risk_cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                    risk_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                elif 'low' in risk_value:
                    # Dark Green background for Low
                    risk_cell.fill = PatternFill(start_color='006400', end_color='006400', fill_type='solid')
                    risk_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                
                current_row += 1
                category_sr_no += 1
            
            # Add all category sections (each resets Sr. No. to 1)
            current_row = add_category_section('PHYSICAL AND ENVIRONMENTAL SECURITY', physical_environmental_issues, current_row, False)
            current_row = add_category_section('POWER BACK UP', power_backup_issues_list, current_row, False)
            current_row = add_category_section('USER AWARENESS', user_awareness_issues, current_row, False)
            current_row = add_category_section('MAINTENANCE AND BUSINESS CONTINUITY CONTROLS', maintenance_bc_issues, current_row, False)
            current_row = add_category_section('PATCH MANAGEMENT', patch_management_issues, current_row, False)
            current_row = add_category_section('NETWORK SECURITY', network_security_issues, current_row, False)
            current_row = add_category_section('ENDPOINTS VULNERABILITY', endpoints_vulnerability_issues, current_row, False)
            current_row = add_category_section('ATM MACHINE ROOM', atm_machine_room_issues, current_row, False)
            current_row = add_category_section('EMAIL-SECURITY', email_security_issues, current_row, False)
            current_row = add_category_section('REMOTE ACCESS', remote_access_issues, current_row, False)
            current_row = add_category_section('UNAUTHORIZED APPLICATIONS / PERSONAL DATA', unauthorized_apps_issues, current_row, False)
            current_row = add_category_section('Important Note', important_note_issues, current_row, True)  # Use custom sorting
        
        # Return the workbook
        return workbook
        
    except Exception as e:
        print(f"Error generating Excel report: {e}")
        raise e

def generate_console_report(analysis_results):
    """Generate console-style report from analysis results grouped by Column B data"""
    try:
        report_lines = []
        report_lines.append("=" * 80)
        report_lines.append("BRANCH CONSOLE - NON-COMPLIANCE ANALYSIS REPORT")
        report_lines.append("=" * 80)
        report_lines.append(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report_lines.append("")
        
        if not analysis_results:
            report_lines.append("No Non-Compliance entries found in any worksheet.")
            report_lines.append("")
        else:
            # Group all non-compliance entries by Column B data
            grouped_issues = {}
            
            for worksheet_data in analysis_results:
                worksheet_name = worksheet_data['worksheet_name']
                branch_name = extract_branch_name_from_worksheet(worksheet_name)
                
                for entry in worksheet_data['non_compliance_entries']:
                    a_data = entry.get('column_a_data', '')
                    b_data = entry['column_b_data']
                    d_data = entry.get('column_d_data', '')
                    row_num = entry['row_number']
                    cell_ref = entry['cell_reference']
                    e_data = entry['column_e_data']
                    
                    # Create a unique key for this issue
                    issue_key = b_data.strip()
                    
                    if issue_key not in grouped_issues:
                        grouped_issues[issue_key] = {
                            'column_a_data': a_data,
                            'issue_description': b_data,
                            'column_d_data': d_data,
                            'branches': [],
                            'total_occurrences': 0,
                            'row_numbers': []  # Store all row numbers for this issue
                        }
                    
                    # Add this branch to the list if not already present
                    if branch_name not in grouped_issues[issue_key]['branches']:
                        grouped_issues[issue_key]['branches'].append(branch_name)
                    
                    # Add row number to the list
                    grouped_issues[issue_key]['row_numbers'].append(row_num)
                    grouped_issues[issue_key]['total_occurrences'] += 1
            
            total_issues = len(grouped_issues)
            total_non_compliance = sum(issue['total_occurrences'] for issue in grouped_issues.values())
            
            report_lines.append(f"SUMMARY: Found {total_issues} unique Non-Compliance issues with {total_non_compliance} total occurrences")
            report_lines.append("")
            report_lines.append("-" * 80)
            
            # Sort issues by the minimum row number for each issue
            sorted_issues = sorted(grouped_issues.items(), key=lambda x: min(x[1]['row_numbers']))
            
            for issue_key, issue_data in sorted_issues:
                column_a_data = issue_data.get('column_a_data', '')
                issue_description = issue_data['issue_description']
                column_d_data = issue_data.get('column_d_data', '')
                branches = issue_data['branches']
                total_occurrences = issue_data['total_occurrences']
                
                # Get unique row numbers and sort them
                unique_row_numbers = sorted(set(issue_data['row_numbers']))
                
                report_lines.append("")
                # Add Column A Data if present
                if column_a_data and str(column_a_data).strip():
                    report_lines.append(f"Column A Data: {column_a_data}")
                
                report_lines.append(f"Column B Data: {issue_description}")
                
                # Add Column D Data if present
                if column_d_data and str(column_d_data).strip():
                    report_lines.append(f"Column D Data: {column_d_data}")
                
                report_lines.append(f"Row Numbers: {', '.join(map(str, unique_row_numbers))}")
                report_lines.append(f"Total Occurrences: {total_occurrences}")
                report_lines.append(f"Branch Names: {', '.join(sorted(branches))}")
                report_lines.append("-" * 40)
        
        report_lines.append("=" * 80)
        report_lines.append("END OF REPORT")
        report_lines.append("=" * 80)
        
        return "\n".join(report_lines)
        
    except Exception as e:
        print(f"Error generating console report: {e}")
        raise e

@branch_console_bp.route('/process_branch_console', methods=['POST'])
@login_required
def process_branch_console():
    print("=" * 50)
    print("BRANCH CONSOLE FORM SUBMISSION RECEIVED!")
    print("=" * 50)
    
    try:
        # Check if Excel file is present in the request
        if 'excelFile' not in request.files:
            flash('Excel file is required!', 'error')
            return redirect(url_for('audit_dashboard'))
        
        excel_file = request.files['excelFile']
        
        # Check if file is selected
        if excel_file.filename == '':
            flash('Please select an Excel file!', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Validate Excel file extension
        if not allowed_file(excel_file.filename, ALLOWED_EXCEL_EXTENSIONS):
            flash('Invalid Excel file format! Please upload .xlsx or .xls files only.', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Create temporary file that won't be automatically deleted
        temp_excel_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_excel_file.close()  # Close the file handle to avoid locking issues
        
        try:
            # Secure the filename
            excel_filename = secure_filename(excel_file.filename)
            
            # Save uploaded file to temporary file
            excel_file.save(temp_excel_file.name)
            
            print(f"Excel file saved to: {temp_excel_file.name}")
            
            # Get the total number of worksheets in the uploaded file
            uploaded_workbook = load_workbook(temp_excel_file.name, data_only=True)
            total_worksheets = len(uploaded_workbook.sheetnames)
            print(f"Total worksheets in uploaded file: {total_worksheets}")
            
            # Analyze the Excel file for Non-Compliance entries
            analysis_results = analyze_excel_file(temp_excel_file.name)
            
            # Generate console report (for debugging)
            console_report = generate_console_report(analysis_results)
            
            # Print the report to console (for debugging)
            print("\n" + "=" * 50)
            print("CONSOLE REPORT OUTPUT:")
            print("=" * 50)
            print(console_report)
            print("=" * 50)
            
            # Generate Excel report
            output_workbook = generate_excel_report(analysis_results, total_worksheets)
            
            # Create temporary file for output Excel
            temp_output_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_output_file.close()
            
            # Save the workbook to temporary file
            output_workbook.save(temp_output_file.name)
            print(f"Excel report saved to: {temp_output_file.name}")
            
            # Generate filename with timestamp
            output_filename = f"Branch_Console_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Send the file for download
            response = send_file(
                temp_output_file.name,
                as_attachment=True,
                download_name=output_filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Schedule cleanup of temporary output file
            def cleanup_output_file():
                try:
                    if os.path.exists(temp_output_file.name):
                        os.unlink(temp_output_file.name)
                        print(f"Cleaned up temporary output file: {temp_output_file.name}")
                except Exception as e:
                    print(f"Error cleaning up temporary output file: {e}")
            
            response.call_on_close(cleanup_output_file)
            
            return response
            
        except Exception as e:
            print(f"Error processing Excel file: {e}")
            flash(f'Error processing Excel file: {str(e)}', 'error')
            return redirect(url_for('audit_dashboard'))
        
        finally:
            # Clean up temporary file
            try:
                if os.path.exists(temp_excel_file.name):
                    os.unlink(temp_excel_file.name)
                    print(f"Cleaned up temporary file: {temp_excel_file.name}")
            except Exception as e:
                print(f"Error cleaning up temporary file: {e}")
    
    except Exception as e:
        print(f"Error in process_branch_console: {e}")
        flash(f'An error occurred: {str(e)}', 'error')
        return redirect(url_for('audit_dashboard'))
