# nmap_routes.py
from flask import Blueprint, request, send_file, make_response, jsonify, session
from flask_login import current_user
import re
import io
import pandas as pd
import zipfile
import os
import math
from io import BytesIO
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# Create a Blueprint for nmap routes
nmap_bp = Blueprint('nmap', __name__)

def generate_dynamic_filename(organization, end_date):
    """
    Generate static filename for Infrastructure VAPT First Audit.
    Format: Infrastructure_VAPT_First_Audit
    """
    return "Infrastructure_VAPT_First_Audit_Worksheet.xlsx"
        

@nmap_bp.route('/check_vulnerabilities', methods=['POST'])
def check_vulnerabilities():
    """Return both matched and unmatched vulnerabilities with full catalog details for merge management."""
    if 'nmapFiles' not in request.files or 'nessusFiles' not in request.files:
        return jsonify({"error": "Both Nmap and Nessus files are required"}), 400
    
    nmap_file = request.files['nmapFiles']
    nessus_file = request.files['nessusFiles']
    
    if nmap_file.filename == '' or nessus_file.filename == '':
        return jsonify({"error": "No file selected"}), 400
    
    try:
        # Process Nessus files to check vulnerabilities
        nessus_dataframes = process_nessus_zip(nessus_file)
        
        if nessus_dataframes:
            combined_nessus = pd.concat(nessus_dataframes, ignore_index=True)
            
            # Don't store the full Nessus data in session - it's too large
            # We'll re-process from files during report generation
            
            # Calculate matched and unmatched vulnerabilities
            matched_groups = []
            unmatched_vulnerabilities = []
            
            try:
                # Filter by valid risks and normalize
                valid_risks = ['low', 'medium', 'high', 'critical']
                df_filtered = combined_nessus.copy()
                df_filtered['Risk'] = df_filtered['Risk'].astype(str).str.lower().str.strip()
                df_filtered = df_filtered[df_filtered['Risk'].isin(valid_risks)]

                # Get unique vulnerability names from Name column
                unique_vulnerabilities_list = df_filtered['Name'].dropna().drop_duplicates().astype(str).str.strip().tolist()
                unique_vulnerabilities = set(unique_vulnerabilities_list)
                
                # Load catalog to get matched vulnerabilities with full details
                catalog_path = "static/Formats_and_Catalog/Infrastructure VAPT Catalog.xlsx"
                if os.path.exists(catalog_path):
                    try:
                        catalog_df = pd.read_excel(catalog_path, sheet_name=0)
                    except Exception as e:
                        print(f"Error reading catalog file for vulnerability check: {e}")
                        catalog_df = None
                    
                    if catalog_df is not None and 'Vulnerabilities in this group' in catalog_df.columns:
                        matched_vulnerability_names = set()
                        
                        # Build matched groups with catalog details
                        for idx, row in catalog_df.iterrows():
                            vulnerabilities_in_group = str(row.get('Vulnerabilities in this group', '')).strip()
                            if pd.isna(vulnerabilities_in_group) or vulnerabilities_in_group == '':
                                continue
                            
                            # Split vulnerabilities by newlines
                            vuln_list = [v.strip() for v in vulnerabilities_in_group.split('\n') if v.strip()]
                            
                            # Find which vulnerabilities from Excel match this catalog group
                            matched_vulns_in_group = []
                            for vulnerability in unique_vulnerabilities:
                                # Use only first 170 characters for matching (to handle long vulnerability names)
                                vuln_short = str(vulnerability)[:170]
                                escaped_vulnerability = re.escape(vuln_short)
                                pattern = rf'(?:\n|\r\n|\A){escaped_vulnerability}'
                                if re.search(pattern, vulnerabilities_in_group, re.IGNORECASE):
                                    matched_vulns_in_group.append(vulnerability)
                                    matched_vulnerability_names.add(vulnerability)
                            
                            # If any vulnerabilities matched this catalog group, add it
                            # Store only essential data to keep session size small
                            if matched_vulns_in_group:
                                matched_groups.append({
                                    'catalog_id': int(idx),
                                    'group_name': str(row.get('Name of Vulnerability', ''))[:200],
                                    'risk_factor': str(row.get('Risk Factor', ''))[:20],
                                    'cvss_score': str(row.get('CVSS Score', ''))[:10],
                                    'matched_vulnerabilities': matched_vulns_in_group
                                })
                        
                        # Calculate unmatched vulnerabilities
                        unmatched_vulnerabilities = sorted(list(unique_vulnerabilities - matched_vulnerability_names))
                        
            except Exception as e:
                print(f"Error calculating vulnerabilities: {e}")
                import traceback
                traceback.print_exc()
                return jsonify({"error": f"Error processing vulnerabilities: {str(e)}"}), 500
            
            # Initialize merge state in session with size limits
            # Limit unmatched vulnerabilities list to prevent session overflow
            unmatched_limited = unmatched_vulnerabilities[:100] if len(unmatched_vulnerabilities) > 100 else unmatched_vulnerabilities
            
            session['vulnerability_merge_state'] = {
                'matched_groups': matched_groups[:50],  # Limit to 50 groups max
                'unmatched_vulnerabilities': unmatched_limited,
                'merge_operations': [],  # Track merge operations for undo
                'new_group_details': {}  # Store full details for new groups separately
            }
            
            if len(unmatched_vulnerabilities) > 100:
                print(f"⚠️ Warning: {len(unmatched_vulnerabilities)} unmatched vulnerabilities found. Limited to 100 in session.")
            
            # Return matched groups and unmatched vulnerabilities
            return jsonify({
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_vulnerabilities
            })
        else:
            return jsonify({"error": "No Nessus data found"}), 400
            
    except Exception as e:
        print(f"Error checking vulnerabilities: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error processing files: {str(e)}"}), 500


@nmap_bp.route('/submit_vulnerability_details', methods=['POST'])
def submit_vulnerability_details():
    """Handle submission of user-provided details for unmatched vulnerabilities."""
    try:
        data = request.get_json()
        
        if not data or 'vulnerability_details' not in data:
            return jsonify({"error": "No vulnerability details provided"}), 400
        
        vulnerability_details = data['vulnerability_details']
        
        # Validate required fields
        required_fields = ['vulnerabilityName', 'riskFactor', 'cveId', 'cvssScore', 
                          'auditObservation', 'impact', 'recommendation', 'referenceLink']
        
        for vuln_name, details in vulnerability_details.items():
            for field in required_fields:
                if field not in details:
                    return jsonify({"error": f"Missing required field '{field}' for vulnerability '{vuln_name}'"}), 400
        
        # Update catalog with all vulnerabilities (merged and separate)
        update_catalog_with_vulnerabilities(vulnerability_details)
        
        # Store the details in session for use in report generation
        session['unmatched_vulnerability_details'] = vulnerability_details
        
        return jsonify({"success": True, "message": "Vulnerability details saved successfully"})
        
    except Exception as e:
        print(f"Error saving vulnerability details: {e}")
        return jsonify({"error": f"Error saving vulnerability details: {str(e)}"}), 500


@nmap_bp.route('/merge_with_matched', methods=['POST'])
def merge_with_matched():
    """Merge an unmatched vulnerability with an existing matched group."""
    try:
        data = request.get_json()
        
        if not data or 'unmatched_vulnerability' not in data or 'target_group_id' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        unmatched_vuln = data['unmatched_vulnerability']
        target_group_id = data['target_group_id']
        
        # Get current merge state from session
        merge_state = session.get('vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Find the target matched group
        matched_groups = merge_state.get('matched_groups', [])
        target_group = None
        target_index = None
        
        for idx, group in enumerate(matched_groups):
            if group['catalog_id'] == target_group_id:
                target_group = group
                target_index = idx
                break
        
        if target_group is None:
            return jsonify({"error": "Target group not found"}), 404
        
        # Add the unmatched vulnerability to the matched group
        if unmatched_vuln not in target_group['matched_vulnerabilities']:
            target_group['matched_vulnerabilities'].append(unmatched_vuln)
        
        # Remove from unmatched list
        unmatched_list = merge_state.get('unmatched_vulnerabilities', [])
        if unmatched_vuln in unmatched_list:
            unmatched_list.remove(unmatched_vuln)
        
        # Record the merge operation for undo
        merge_state['merge_operations'].append({
            'type': 'merge_with_matched',
            'unmatched_vulnerability': unmatched_vuln,
            'target_group_id': target_group_id,
            'timestamp': datetime.now().isoformat()
        })
        
        # Update session
        session['vulnerability_merge_state'] = merge_state
        session.modified = True
        
        return jsonify({
            "success": True,
            "message": f"Vulnerability merged successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_list
            }
        })
        
    except Exception as e:
        print(f"Error merging with matched: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error merging: {str(e)}"}), 500


@nmap_bp.route('/merge_with_unmatched', methods=['POST'])
def merge_with_unmatched():
    """Merge multiple unmatched vulnerabilities into a new group."""
    try:
        data = request.get_json()
        
        if not data or 'vulnerabilities' not in data or 'vulnerability_details' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        vulnerabilities_to_merge = data['vulnerabilities']  # List of vulnerability names
        vulnerability_details = data['vulnerability_details']  # Group details
        
        # Validate required fields
        required_fields = ['vulnerabilityName', 'riskFactor', 'cveId', 'cvssScore', 
                          'auditObservation', 'impact', 'recommendation', 'referenceLink']
        
        for field in required_fields:
            if field not in vulnerability_details:
                return jsonify({"error": f"Missing required field '{field}'"}), 400
        
        # Get current merge state from session
        merge_state = session.get('vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Create a new matched group from the merged vulnerabilities (minimal data for session)
        new_group_id = -len(merge_state.get('matched_groups', [])) - 1
        new_group = {
            'catalog_id': new_group_id,  # Negative ID for new groups
            'group_name': vulnerability_details['vulnerabilityName'][:200],
            'risk_factor': vulnerability_details['riskFactor'][:20],
            'cvss_score': vulnerability_details['cvssScore'][:10],
            'matched_vulnerabilities': vulnerabilities_to_merge,
            'is_new_group': True  # Flag to indicate this should be added to catalog
        }
        
        # Add to matched groups
        matched_groups = merge_state.get('matched_groups', [])
        matched_groups.append(new_group)
        
        # Store full details separately to save session space
        new_group_details = merge_state.get('new_group_details', {})
        new_group_details[str(new_group_id)] = vulnerability_details
        merge_state['new_group_details'] = new_group_details
        
        # Remove from unmatched list
        unmatched_list = merge_state.get('unmatched_vulnerabilities', [])
        for vuln in vulnerabilities_to_merge:
            if vuln in unmatched_list:
                unmatched_list.remove(vuln)
        
        # Record the merge operation for undo
        merge_state['merge_operations'].append({
            'type': 'merge_with_unmatched',
            'vulnerabilities': vulnerabilities_to_merge,
            'new_group_id': new_group['catalog_id'],
            'timestamp': datetime.now().isoformat()
        })
        
        # Update session
        session['vulnerability_merge_state'] = merge_state
        session.modified = True
        
        # Update catalog with new vulnerability group
        # Add the merged vulnerabilities list to the details
        vulnerability_details_with_merge = vulnerability_details.copy()
        vulnerability_details_with_merge['isMerged'] = True
        vulnerability_details_with_merge['mergedVulnerabilities'] = vulnerabilities_to_merge
        
        update_catalog_with_vulnerabilities({
            vulnerability_details['vulnerabilityName']: vulnerability_details_with_merge
        })
        
        return jsonify({
            "success": True,
            "message": "Vulnerabilities merged into new group successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_list
            }
        })
        
    except Exception as e:
        print(f"Error merging unmatched vulnerabilities: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error merging: {str(e)}"}), 500


@nmap_bp.route('/add_vulnerability_details', methods=['POST'])
def add_vulnerability_details():
    """Add details for a single unmatched vulnerability."""
    try:
        data = request.get_json()
        
        if not data or 'vulnerability_name' not in data or 'vulnerability_details' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        vulnerability_name = data['vulnerability_name']
        vulnerability_details = data['vulnerability_details']
        
        # Validate required fields
        required_fields = ['vulnerabilityName', 'riskFactor', 'cveId', 'cvssScore', 
                          'auditObservation', 'impact', 'recommendation', 'referenceLink']
        
        for field in required_fields:
            if field not in vulnerability_details:
                return jsonify({"error": f"Missing required field '{field}'"}), 400
        
        # Get current merge state from session
        merge_state = session.get('vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Create a new matched group for this single vulnerability (minimal data)
        new_group_id = -len(merge_state.get('matched_groups', [])) - 1
        new_group = {
            'catalog_id': new_group_id,
            'group_name': vulnerability_details['vulnerabilityName'][:200],
            'risk_factor': vulnerability_details['riskFactor'][:20],
            'cvss_score': vulnerability_details['cvssScore'][:10],
            'matched_vulnerabilities': [vulnerability_name],
            'is_new_group': True
        }
        
        # Add to matched groups
        matched_groups = merge_state.get('matched_groups', [])
        matched_groups.append(new_group)
        
        # Store full details separately
        new_group_details = merge_state.get('new_group_details', {})
        new_group_details[str(new_group_id)] = vulnerability_details
        merge_state['new_group_details'] = new_group_details
        
        # Remove from unmatched list
        unmatched_list = merge_state.get('unmatched_vulnerabilities', [])
        if vulnerability_name in unmatched_list:
            unmatched_list.remove(vulnerability_name)
        
        # Record the operation for undo
        merge_state['merge_operations'].append({
            'type': 'add_details',
            'vulnerability': vulnerability_name,
            'new_group_id': new_group['catalog_id'],
            'timestamp': datetime.now().isoformat()
        })
        
        # Update session
        session['vulnerability_merge_state'] = merge_state
        session.modified = True
        
        # Update catalog (single vulnerability - not merged)
        vulnerability_details_single = vulnerability_details.copy()
        vulnerability_details_single['isMerged'] = False
        vulnerability_details_single['actualVulnerabilityName'] = vulnerability_name  # Store the actual vulnerability name
        
        update_catalog_with_vulnerabilities({
            vulnerability_details['vulnerabilityName']: vulnerability_details_single
        })
        
        return jsonify({
            "success": True,
            "message": "Vulnerability details added successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_list
            }
        })
        
    except Exception as e:
        print(f"Error adding vulnerability details: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error adding details: {str(e)}"}), 500


@nmap_bp.route('/merge_matched_groups', methods=['POST'])
def merge_matched_groups():
    """Merge two matched groups together."""
    try:
        data = request.get_json()
        
        if not data or 'source_group_id' not in data or 'target_group_id' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        source_group_id = data['source_group_id']
        target_group_id = data['target_group_id']
        
        # Get current merge state from session
        merge_state = session.get('vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Find both groups
        matched_groups = merge_state.get('matched_groups', [])
        source_group = None
        target_group = None
        source_index = None
        
        for idx, group in enumerate(matched_groups):
            if group['catalog_id'] == source_group_id:
                source_group = group
                source_index = idx
            elif group['catalog_id'] == target_group_id:
                target_group = group
        
        if source_group is None or target_group is None:
            return jsonify({"error": "One or both groups not found"}), 404
        
        # Merge source group vulnerabilities into target group
        for vuln in source_group['matched_vulnerabilities']:
            if vuln not in target_group['matched_vulnerabilities']:
                target_group['matched_vulnerabilities'].append(vuln)
        
        # Remove source group
        matched_groups.pop(source_index)
        
        # Record the merge operation for undo
        merge_state['merge_operations'].append({
            'type': 'merge_matched_groups',
            'source_group_id': source_group_id,
            'target_group_id': target_group_id,
            'source_group_data': source_group,  # Save for undo
            'timestamp': datetime.now().isoformat()
        })
        
        # Update session
        session['vulnerability_merge_state'] = merge_state
        session.modified = True
        
        return jsonify({
            "success": True,
            "message": "Groups merged successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": merge_state.get('unmatched_vulnerabilities', [])
            }
        })
        
    except Exception as e:
        print(f"Error merging matched groups: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error merging groups: {str(e)}"}), 500


@nmap_bp.route('/undo_last_merge', methods=['POST'])
def undo_last_merge():
    """Undo the last merge operation."""
    try:
        # Get current merge state from session
        merge_state = session.get('vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        merge_operations = merge_state.get('merge_operations', [])
        if not merge_operations:
            return jsonify({"error": "No operations to undo"}), 400
        
        # Get the last operation
        last_operation = merge_operations.pop()
        operation_type = last_operation['type']
        
        matched_groups = merge_state.get('matched_groups', [])
        unmatched_list = merge_state.get('unmatched_vulnerabilities', [])
        
        # Undo based on operation type
        if operation_type == 'merge_with_matched':
            # Remove vulnerability from matched group and add back to unmatched
            target_group_id = last_operation['target_group_id']
            unmatched_vuln = last_operation['unmatched_vulnerability']
            
            for group in matched_groups:
                if group['catalog_id'] == target_group_id:
                    if unmatched_vuln in group['matched_vulnerabilities']:
                        group['matched_vulnerabilities'].remove(unmatched_vuln)
                    break
            
            if unmatched_vuln not in unmatched_list:
                unmatched_list.append(unmatched_vuln)
                unmatched_list.sort()
        
        elif operation_type == 'merge_with_unmatched':
            # Remove the new group and restore vulnerabilities to unmatched
            new_group_id = last_operation['new_group_id']
            vulnerabilities = last_operation['vulnerabilities']
            
            # Remove the new group
            matched_groups = [g for g in matched_groups if g['catalog_id'] != new_group_id]
            
            # Add back to unmatched
            for vuln in vulnerabilities:
                if vuln not in unmatched_list:
                    unmatched_list.append(vuln)
            unmatched_list.sort()
        
        elif operation_type == 'add_details':
            # Remove the new group and restore vulnerability to unmatched
            new_group_id = last_operation['new_group_id']
            vulnerability = last_operation['vulnerability']
            
            # Remove the new group
            matched_groups = [g for g in matched_groups if g['catalog_id'] != new_group_id]
            
            # Add back to unmatched
            if vulnerability not in unmatched_list:
                unmatched_list.append(vulnerability)
                unmatched_list.sort()
        
        elif operation_type == 'merge_matched_groups':
            # Restore source group and remove merged vulnerabilities from target
            source_group_data = last_operation['source_group_data']
            target_group_id = last_operation['target_group_id']
            
            # Add source group back
            matched_groups.append(source_group_data)
            
            # Remove source group's vulnerabilities from target group
            for group in matched_groups:
                if group['catalog_id'] == target_group_id:
                    for vuln in source_group_data['matched_vulnerabilities']:
                        if vuln in group['matched_vulnerabilities']:
                            # Only remove if it was originally from source group
                            if vuln not in source_group_data.get('all_catalog_vulnerabilities', []):
                                continue
                            group['matched_vulnerabilities'].remove(vuln)
                    break
        
        # Update session
        merge_state['matched_groups'] = matched_groups
        merge_state['unmatched_vulnerabilities'] = unmatched_list
        merge_state['merge_operations'] = merge_operations
        session['vulnerability_merge_state'] = merge_state
        session.modified = True
        
        return jsonify({
            "success": True,
            "message": "Last operation undone successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_list
            }
        })
        
    except Exception as e:
        print(f"Error undoing merge: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error undoing: {str(e)}"}), 500


def extract_poc_images(evidence_files):
    """Extract POC images from evidence files zip and return mapping of vulnerability names to image paths"""
    poc_mapping = {}
    
    if not evidence_files or evidence_files.filename == '':
        return poc_mapping
    
    try:
        import zipfile
        from io import BytesIO
        from datetime import datetime
        
        # Read the zip file
        zip_data = evidence_files.read()
        
        with zipfile.ZipFile(BytesIO(zip_data), 'r') as zip_ref:
            file_list = zip_ref.namelist()
            
            # Find ALL POC folder
            poc_folder = None
            for file_path in file_list:
                if 'ALL POC' in file_path and file_path.endswith('/'):
                    poc_folder = file_path
                    break
            
            if not poc_folder:
                print("ALL POC folder not found in evidence files")
                return poc_mapping
            
            # Extract images from ALL POC folder
            image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff']
            
            # Prepare timestamped temp folder
            ts = datetime.now().strftime('%Y%m%d%H%M%S')
            temp_dir = f"temp_poc_images_{ts}"
            os.makedirs(temp_dir, exist_ok=True)
            
            for file_path in file_list:
                if file_path.startswith(poc_folder) and not file_path.endswith('/'):
                    # Get filename without extension
                    filename = os.path.basename(file_path)
                    name_without_ext = os.path.splitext(filename)[0]
                    
                    # Check if it's an image file
                    if any(filename.lower().endswith(ext) for ext in image_extensions):
                        # Extract the file to a temporary location
                        try:
                            with zip_ref.open(file_path) as f:
                                image_data = f.read()
                            
                            # Create temporary file
                            temp_file_path = os.path.join(temp_dir, filename)
                            
                            with open(temp_file_path, 'wb') as temp_file:
                                temp_file.write(image_data)
                            
                            # Map vulnerability name to image path
                            poc_mapping[name_without_ext] = temp_file_path
                            
                        except Exception as e:
                            print(f"Error extracting image {filename}: {e}")
                            continue
            
            print(f"Extracted {len(poc_mapping)} POC images")
            
    except Exception as e:
        print(f"Error extracting POC images: {e}")
    
    return poc_mapping

def remove_last_number_from_name(name):
    """
    Remove the last number from a name if it exists at the end.
    Example: "SSL Cert Expire 1" -> "SSL Cert Expire"
             "SSL Cert Expire 123" -> "SSL Cert Expire"
             "SSL 1 Cert Expire" -> "SSL 1 Cert Expire" (no change, number not at end)
    """
    if not name:
        return name
    
    # Use regex to find trailing numbers at the end of the string
    # This matches one or more digits at the end, optionally preceded by whitespace
    pattern = r'\s*\d+$'
    result = re.sub(pattern, '', name)
    return result.strip()

def extract_trailing_number(name):
    """
    Extract the trailing number from a name if it exists at the end.
    Returns the number as an integer, or 0 if no trailing number exists.
    Example: "SSL Cert Expire 1" -> 1
             "SSL Cert Expire 123" -> 123
             "SSL Cert Expire" -> 0
             "SSL 1 Cert Expire" -> 0 (number not at end)
    """
    if not name:
        return 0
    
    # Use regex to find trailing numbers at the end of the string
    pattern = r'\s*(\d+)$'
    match = re.search(pattern, name)
    if match:
        return int(match.group(1))
    return 0

def get_image_sort_key(image_name):
    """
    Get sort key for image name.
    Returns (has_number, number_value) where:
    - has_number: 0 if no trailing number, 1 if has trailing number
    - number_value: the trailing number (0 if no number)
    This ensures images without numbers come first, then sorted by number.
    """
    trailing_number = extract_trailing_number(image_name)
    has_number = 1 if trailing_number > 0 else 0
    return (has_number, trailing_number)

def normalize_vulnerability_name_for_filename(vuln_name):
    """
    Normalize vulnerability name to be compatible with file names by replacing
    invalid characters with dashes.
    
    Args:
        vuln_name (str): Original vulnerability name
        
    Returns:
        str: Normalized vulnerability name safe for file names
    """
    if not vuln_name:
        return vuln_name
    
    # Characters that are not allowed in file names
    invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
    
    normalized_name = str(vuln_name)
    
    # Replace invalid characters with dashes
    for char in invalid_chars:
        normalized_name = normalized_name.replace(char, '-')
    
    # Remove multiple consecutive dashes and trim
    normalized_name = re.sub(r'-+', '-', normalized_name)
    normalized_name = normalized_name.strip('-')
    
    return normalized_name

def load_external_images_by_name(image_directory="static/uploads"):
    """
    Load external image files and match them with vulnerability names using normalized names.
    
    Args:
        image_directory (str): Directory containing image files
        
    Returns:
        dict: Dictionary mapping vulnerability names to image data
    """
    import os
    import glob
    
    external_images = {}
    
    if not os.path.exists(image_directory):
        print(f"Image directory '{image_directory}' not found")
        return external_images
    
    # Supported image extensions
    image_extensions = ['*.jpg', '*.jpeg', '*.png', '*.bmp', '*.gif', '*.tiff']
    
    # Get all image files
    image_files = []
    for ext in image_extensions:
        image_files.extend(glob.glob(os.path.join(image_directory, ext)))
        image_files.extend(glob.glob(os.path.join(image_directory, ext.upper())))
    
    print(f"Found {len(image_files)} external image files in {image_directory}")
    
    for image_path in image_files:
        try:
            # Get filename without extension
            filename = os.path.splitext(os.path.basename(image_path))[0]
            
            # Read image data
            with open(image_path, 'rb') as f:
                image_data = f.read()
            
            # Store under the filename (which should be normalized)
            external_images[filename] = external_images.get(filename, [])
            external_images[filename].append(image_data)
            
            print(f"Loaded external image: {filename}")
            
        except Exception as e:
            print(f"Error loading external image {image_path}: {e}")
            continue
    
    return external_images

def insert_image_to_excel(excel_file, sheet_name, image_path, cell="B2"):
    """
    Insert a resized image into an Excel file at a specified cell.
    
    Args:
        excel_file (str): Path to the Excel file to create/save.
        sheet_name (str): Name of the sheet where image will be placed.
        image_path (str): Path to the image file (jpg, png, bmp, etc.).
        cell (str): Cell location to place the image.
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Load the image
    img = Image(image_path)

    # Get original dimensions and reduce by 50x
    original_width = img.width
    original_height = img.height
    
    # Calculate 50x reduced dimensions (ensure minimum 1 pixel for visibility)
    new_width = max(1, original_width // 25)
    new_height = max(1, original_height // 25)
    
    img.width = new_width
    img.height = new_height

    # Add image to cell
    ws.add_image(img, cell)

    # Save workbook
    wb.save(excel_file)
    print(f"Resized image ({new_width}x{new_height}) inserted into {excel_file} at {sheet_name}:{cell} (reduced from {original_width}x{original_height})")

def insert_poc_images_to_excel(excel_path, poc_mapping, vulnerabilities_data):
    """Insert POC images directly into Excel using openpyxl with 30x30 pixel size
    Returns a set of row numbers that have POC objects for border formatting"""
    rows_with_objects = set()  # Track which rows have POC objects
    
    try:
        # Load external images and merge with existing poc_mapping
        external_images = load_external_images_by_name()
        
        # Convert external image data to temporary files and add to poc_mapping
        import tempfile
        import os
        
        # Track temporary files created for cleanup
        temp_files_created = []
        
        for vuln_name, image_data_list in external_images.items():
            for i, image_data in enumerate(image_data_list):
                try:
                    # Create temporary file for the image
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
                        tmp.write(image_data)
                        temp_path = tmp.name
                    
                    # Add to poc_mapping with the normalized name
                    poc_mapping[vuln_name] = temp_path
                    temp_files_created.append(temp_path)
                    print(f"Added external image to Excel processing: {vuln_name}")
                    
                except Exception as e:
                    print(f"Error processing external image {vuln_name}: {e}")
                    continue
        from openpyxl import load_workbook
        from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
        
        # Load the existing workbook
        wb = load_workbook(excel_path)
        
        # Get the Infra_VAPT worksheet
        if "Infra_VAPT" not in wb.sheetnames:
            print("Infra_VAPT worksheet not found")
            return rows_with_objects
        
        ws = wb["Infra_VAPT"]
        
        # Find POC columns by looking at the merged header in first row
        # We're looking for the merged "POC" header which spans L, M, N, O, P, Q, R columns
        poc_col_start = None
        poc_col_end = None
        
        # Check for merged cells in row 1
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == 1 and merged_range.max_row == 1:
                # Check if this merged cell contains "POC"
                first_cell = ws.cell(row=1, column=merged_range.min_col)
                if first_cell.value and str(first_cell.value).strip() == "POC":
                    poc_col_start = merged_range.min_col
                    poc_col_end = merged_range.max_col
                    break
        
        if not poc_col_start or not poc_col_end:
            print("POC columns not found in worksheet")
            return rows_with_objects
        
        # Define the column order for image insertion: M, N, O, P, Q, R, L
        # POC columns are: L, M, N, O, P, Q, R (7 columns total)
        # Column indices (0-based): L=poc_col_start, M=poc_col_start+1, ..., R=poc_col_end
        # Insertion order: M, N, O, P, Q, R, L
        image_columns = [
            poc_col_start + 1,  # M
            poc_col_start + 2,  # N
            poc_col_start + 3,  # O
            poc_col_start + 4,  # P
            poc_col_start + 5,  # Q
            poc_col_start + 6,  # R (should be poc_col_end)
            poc_col_start       # L
        ]
        
        print(f"Found POC columns from column {poc_col_start} to {poc_col_end}")
        
        # Process each row and match vulnerabilities with POC images
        for row in range(2, ws.max_row + 1):
            vulnerability_cell = ws.cell(row=row, column=2)  # Vulnerabilities column (column B)
            vulnerability_text = str(vulnerability_cell.value) if vulnerability_cell.value else ""
            
            if vulnerability_text:
                # Split vulnerabilities (they might be on separate lines)
                vulnerabilities = [v.strip() for v in vulnerability_text.split('\n') if v.strip()]
                
                # Find all matching POC images for this row
                matching_images = []
                for vuln in vulnerabilities:
                    # Normalize the vulnerability name for filename matching
                    normalized_vuln = normalize_vulnerability_name_for_filename(vuln)
                    
                    # Use only first 170 characters for matching (to handle long vulnerability names)
                    vuln_short = vuln[:170].lower().strip()
                    normalized_vuln_short = normalized_vuln[:170].lower().strip()
                    
                    for image_name, image_path in poc_mapping.items():
                        # Remove last number from image name for matching (ignore trailing numbers)
                        image_name_without_last_number = remove_last_number_from_name(image_name)
                        
                        # Use only first 170 characters for matching (to handle long names)
                        image_name_short = image_name[:170].lower().strip()
                        image_name_without_last_number_short = image_name_without_last_number[:170].lower().strip()
                        
                        # Check if vulnerability name matches image name
                        # 1. Try exact match (original image name)
                        # 2. Try match after removing last number from image name
                        # 3. Try substring matches for both cases
                        matches = False
                        
                        # Exact match with original image name
                        if (vuln_short == image_name_short or 
                            normalized_vuln_short == image_name_short):
                            matches = True
                        
                        # Exact match with image name without last number
                        if (vuln_short == image_name_without_last_number_short or
                            normalized_vuln_short == image_name_without_last_number_short):
                            matches = True
                        
                        # Substring matches (vulnerability in image or image in vulnerability)
                        if (vuln_short in image_name_short or image_name_short in vuln_short or
                            vuln_short in image_name_without_last_number_short or 
                            image_name_without_last_number_short in vuln_short):
                            matches = True
                        
                        if matches:
                            if image_path not in [img[1] for img in matching_images]:  # Avoid duplicates
                                # Store image name along with path for sorting
                                matching_images.append((vuln, image_path, image_name))
                
                # Sort matching images: first images without trailing numbers, then by trailing number (ascending)
                if matching_images:
                    # Sort by image name: no number first, then by number value
                    matching_images.sort(key=lambda x: get_image_sort_key(x[2]))  # x[2] is the image_name
                    # Remove image_name from tuple (keep only vuln and image_path)
                    matching_images = [(vuln, image_path) for vuln, image_path, image_name in matching_images]
                
                if matching_images:
                    try:
                        # Distribute images across columns: M, N, O, P, Q, R, L
                        num_images_to_insert = min(len(matching_images), 7)  # Max 7 images (one per column)
                        
                        for img_idx in range(num_images_to_insert):
                            vuln, matching_image = matching_images[img_idx]
                            col_idx = image_columns[img_idx]
                            
                            if os.path.exists(matching_image):
                                try:
                                    # Load the image
                                    img = Image(matching_image)
                                    
                                    # Get original dimensions
                                    original_width = img.width
                                    original_height = img.height
                                    
                                    # Resize image to a fixed height of 30px and proportional width
                                    if hasattr(img, 'width') and hasattr(img, 'height') and img.height:
                                        scale_factor = 30 / float(img.height)
                                        img.height = 30
                                        img.width = max(1, int(round(float(img.width) * scale_factor)))
                                    
                                    # Get cell reference (e.g., "M2", "N3", etc.)
                                    from openpyxl.utils import get_column_letter
                                    col_letter = get_column_letter(col_idx)
                                    cell_ref = f"{col_letter}{row}"
                                    
                                    # Insert image at the cell
                                    ws.add_image(img, cell_ref)
                                    
                                    print(f"✅ Inserted image {img_idx + 1} at {cell_ref} for vulnerability: {vuln} (original {original_width}x{original_height} → resized to {img.width}x{img.height})")
                                    
                                except Exception as e:
                                    print(f"⚠️ Error inserting image at column {col_idx}, row {row}: {e}")
                        
                        # Track this row as having POC objects
                        if num_images_to_insert > 0:
                            rows_with_objects.add(row)
                            
                    except Exception as e:
                        print(f"Error adding images for row {row}: {e}")
                        continue
        
        # Apply custom borders to ALL POC columns (L, M, N, O, P, Q, R)
        from openpyxl.styles import Border, Side
        
        # Define border styles for each POC column type
        # L column: left, above, below (NOT right)
        left_border = Border(
            left=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # M, N, O, P, Q columns: above and below only (NOT left or right)
        middle_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # R column: above, below, right (NOT left)
        right_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            right=Side(style='thin')
        )
        
        # Find all rows that are part of the data table (have content in any column)
        table_rows = set()
        
        # Always include header row
        table_rows.add(1)
        
        # Check all rows from 2 onwards to find data rows
        for row in range(2, ws.max_row + 1):
            has_content = False
            
            # Check if any cell in this row has content (excluding POC columns)
            for col in range(1, poc_col_start):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip() != "":
                    has_content = True
                    break
            
            if has_content:
                table_rows.add(row)
        
        # Apply custom borders to all table rows for each POC column
        for row_num in sorted(table_rows):
            # L column (poc_col_start): left, top, bottom
            ws.cell(row=row_num, column=poc_col_start).border = left_border
            
            # M, N, O, P, Q columns (middle columns): top, bottom only
            for col_idx in range(poc_col_start + 1, poc_col_end):
                ws.cell(row=row_num, column=col_idx).border = middle_border
            
            # R column (poc_col_end): top, bottom, right
            ws.cell(row=row_num, column=poc_col_end).border = right_border
        
        # Set row height to ~50px (≈37.5 points) for rows with images to fit 30px images comfortably
        for row_num in rows_with_objects:
            # Row height in Excel: 1 point ≈ 1.33 pixels, so 50px ≈ 37.5 points
            ws.row_dimensions[row_num].height = 37.5
        
        print(f"Identified {len(table_rows)} table rows total")
        print(f"Applied custom borders to POC columns for {len(table_rows)} rows")
        print(f"Set row height to 60px (45 units) for {len(rows_with_objects)} rows with images")
        
        # Save the workbook
        wb.save(excel_path)
        
        # Clean up temporary files created for external images
        for temp_file_path in temp_files_created:
            if os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                    print(f"Cleaned up temporary file: {temp_file_path}")
                except Exception as e:
                    print(f"Error cleaning up temporary file {temp_file_path}: {e}")
        
        print("POC images added successfully with 30x reduction from original dimensions")
        print(f"Added images to {len(rows_with_objects)} rows, distributed across columns M, N, O, P, Q, R, L")
        print(f"Applied custom borders to POC columns for {len(table_rows)} rows")
        return rows_with_objects
        
    except Exception as e:
        print(f"Error inserting POC images: {e}")
        import traceback
        traceback.print_exc()
        return rows_with_objects

def update_catalog_with_vulnerabilities(vulnerability_details):
    """Update the Infrastructure VAPT Catalog with both merged and separate vulnerabilities."""
    try:
        catalog_path = "static/Formats_and_Catalog/Infrastructure VAPT Catalog.xlsx"
        
        if not os.path.exists(catalog_path):
            print(f"Catalog file not found at: {catalog_path}")
            return
        
        # Read the existing catalog with error handling - now reading Sheet2 (index 1)
        try:
            catalog_df = pd.read_excel(catalog_path, sheet_name=1)  # Changed from sheet_name=0 to sheet_name=1 for Sheet2
        except Exception as e:
            print(f"Error reading catalog file: {e}")
            print(f"Catalog file may be corrupted. Please check: {catalog_path}")
            return
        
        # Process all vulnerabilities (both merged and separate)
        for vuln_name, details in vulnerability_details.items():
            if details.get('isMerged', False):
                # Handle merged vulnerabilities
                merged_vulns = details.get('mergedVulnerabilities', [])
                if merged_vulns:
                    # Truncate vulnerability names to first 170 characters for catalog storage
                    truncated_merged_vulns = [str(v)[:170] for v in merged_vulns]
                    
                    # Create a new row for the merged vulnerability group
                    new_row = {
                        'Sr No': len(catalog_df) + 1,
                        'Name of Vulnerability': details.get('vulnerabilityName', ''),
                        'Risk Factor': details.get('riskFactor', ''),
                        'CVE ID': details.get('cveId', ''),
                        'CVSS': details.get('cvssScore', ''),
                        'Audit Observation': details.get('auditObservation', ''),
                        'Impact': details.get('impact', ''),
                        'Recommendation/Countermeasure': details.get('recommendation', ''),
                        'Affected System': '',  # Empty as requested
                        'Reference Link': details.get('referenceLink', ''),
                        'Vulnerabilities in this group': '\n'.join(truncated_merged_vulns),
                        'User_name': current_user.employee_name if current_user.is_authenticated else 'Unknown',
                        'Time_stamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    # Add the new row to the catalog
                    catalog_df = pd.concat([catalog_df, pd.DataFrame([new_row])], ignore_index=True)
            else:
                # Handle separate (non-merged) vulnerabilities
                # Get the actual vulnerability name (might be different from group name)
                actual_vuln_name = details.get('actualVulnerabilityName', vuln_name)
                # Truncate to first 170 characters for catalog storage
                actual_vuln_name_short = str(actual_vuln_name)[:170]
                
                new_row = {
                    'Sr No': len(catalog_df) + 1,
                    'Name of Vulnerability': details.get('vulnerabilityName', vuln_name),  # User-provided group name
                    'Risk Factor': details.get('riskFactor', ''),
                    'CVE ID': details.get('cveId', ''),
                    'CVSS': details.get('cvssScore', ''),
                    'Audit Observation': details.get('auditObservation', ''),
                    'Impact': details.get('impact', ''),
                    'Recommendation/Countermeasure': details.get('recommendation', ''),
                    'Affected System': '',  # Empty as requested
                    'Reference Link': details.get('referenceLink', ''),
                    'Vulnerabilities in this group': actual_vuln_name_short,  # Truncated to 170 chars
                    'User_name': current_user.employee_name if current_user.is_authenticated else 'Unknown',
                    'Time_stamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                # Add the new row to the catalog
                catalog_df = pd.concat([catalog_df, pd.DataFrame([new_row])], ignore_index=True)
        
        # Save the updated catalog to Sheet2
        # Read all existing sheets first
        try:
            # Read all sheets to preserve existing data
            all_sheets = pd.read_excel(catalog_path, sheet_name=None)
            
            # Update Sheet2 with our new data
            all_sheets['Sheet2'] = catalog_df
            
            # Write all sheets back to the file
            with pd.ExcelWriter(catalog_path, engine='openpyxl') as writer:
                for sheet_name, sheet_data in all_sheets.items():
                    sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            print(f"Error saving catalog with multiple sheets: {e}")
            # Fallback: just save Sheet2
            catalog_df.to_excel(catalog_path, sheet_name='Sheet2', index=False)
        merged_count = len([v for v in vulnerability_details.values() if v.get('isMerged', False)])
        separate_count = len([v for v in vulnerability_details.values() if not v.get('isMerged', False)])
        print(f"Updated catalog with {merged_count} merged vulnerability groups and {separate_count} separate vulnerabilities")
        
    except Exception as e:
        print(f"Error updating catalog: {e}")
        import traceback
        traceback.print_exc()

    
@nmap_bp.route('/process_first_audit_report', methods=['POST'])
def process_nmap_files():
    if 'nmapFiles' not in request.files or 'nessusFiles' not in request.files:
        return "Both Nmap and Nessus files are required", 400
    
    nmap_file = request.files['nmapFiles']
    nessus_file = request.files['nessusFiles']
    evidence_file = request.files.get('evidenceFiles')  # Get evidence files if present
    
    if nmap_file.filename == '' or nessus_file.filename == '':
        return "No file selected", 400
    
    # Capture form metadata
    form_metadata = {
        'organization': request.form.get('organization', ''),
        'otherOrganization': request.form.get('otherOrganization', ''),
        'city': request.form.get('city', ''),
        'otherCity': request.form.get('otherCity', ''),
        'state': request.form.get('state', ''),
        'startDate': request.form.get('startDate', ''),
        'endDate': request.form.get('endDate', ''),
        'preparedByTitle': request.form.get('preparedByTitle', ''),
        'preparedByName': request.form.get('preparedByName', ''),
        'auditeeTitle': request.form.get('auditeeTitle', ''),
        'auditeeName': request.form.get('auditeeName', ''),
        'designation': request.form.get('designation', ''),
        'bankEmails': request.form.getlist('bankEmail[]'),
        'teamNames': request.form.getlist('teamName[]'),
        'teamDesignations': request.form.getlist('teamDesignation[]'),
        'teamEmails': request.form.getlist('teamEmail[]'),
        'teamQualifications': request.form.getlist('teamQualification[]'),
        'teamCertified': []
    }
    
    # Handle indexed teamCertified radio buttons
    team_certified = []
    i = 0
    while True:
        certified_value = request.form.get(f'teamCertified[{i}]')
        if certified_value is None:
            break
        team_certified.append(certified_value)
        i += 1
    form_metadata['teamCertified'] = team_certified
    
    # Process both files
    nmap_data = process_nmap_zip(nmap_file)
    nessus_dataframes = process_nessus_zip(nessus_file)
    
    # Initialize unmatched_count early
    unmatched_count = 0
    
    # Create Excel file in memory with nan_inf_to_errors option
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
        workbook = writer.book
        
        # Define formats
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'center',
            'fg_color': "#1376d1",
            'font_color': 'white',
            'border': 1
        })  
        
        cell_format = workbook.add_format({
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'center',
            'border': 1
        })
        
        # Create Meta_Data worksheet as the first worksheet
        create_meta_data_worksheet(workbook, form_metadata, header_format, cell_format)
        
        # Process Nmap data
        if nmap_data:
            df_nmap = pd.DataFrame(nmap_data)
            df_nmap.to_excel(writer, sheet_name='Nmap Files', index=False, header=False)
            
            worksheet_nmap = writer.sheets['Nmap Files']
            worksheet_nmap.set_column('A:F', 20)
            
            # Apply formats to Nmap worksheet
            for row_num, row_data in enumerate(nmap_data):
                fmt = header_format if row_data == ["HOST", "PORT", "SERVICE", "HOST", "PORT", "SERVICE"] else cell_format
                for col_num, value in enumerate(row_data):
                    worksheet_nmap.write(row_num, col_num, value, fmt)
            
            # Track IP positions for proper merging
            ip_positions = {"A": {}, "D": {}}
            
            # First pass: identify all IP positions
            for row_num in range(1, len(nmap_data)):
                ip1 = nmap_data[row_num][0]
                ip2 = nmap_data[row_num][3]
                
                if ip1 and ip1 != "-":
                    if ip1 not in ip_positions["A"]:
                        ip_positions["A"][ip1] = []
                    ip_positions["A"][ip1].append(row_num)
                
                if ip2 and ip2 != "-":
                    if ip2 not in ip_positions["D"]:
                        ip_positions["D"][ip2] = []
                    ip_positions["D"][ip2].append(row_num)
            
            # Second pass: merge contiguous IP cells
            for col_letter, ip_dict in ip_positions.items():
                col_index = ord(col_letter) - ord('A')
                
                for ip, row_nums in ip_dict.items():
                    row_nums.sort()
                    
                    # Group contiguous rows
                    groups = []
                    current_group = [row_nums[0]]
                    
                    for i in range(1, len(row_nums)):
                        if row_nums[i] == row_nums[i-1] + 1:
                            current_group.append(row_nums[i])
                        else:
                            groups.append(current_group)
                            current_group = [row_nums[i]]
                    
                    groups.append(current_group)
                    
                    # Merge each contiguous group
                    for group in groups:
                        if len(group) > 1:
                            start_row = group[0]
                            end_row = group[-1]
                            worksheet_nmap.merge_range(start_row, col_index, end_row, col_index, ip, cell_format)
            
            # Handle placeholder "-" IP merging
            has_placeholder = any("-" in row for row in nmap_data)
            if has_placeholder:
                # Find the row where the placeholder "-" appears in column D
                for row_num in range(1, len(nmap_data)):
                    if nmap_data[row_num][3] == "-":
                        # Count how many consecutive rows have "-" in column D
                        placeholder_rows = []
                        current_row = row_num
                        while current_row < len(nmap_data) and nmap_data[current_row][3] == "-":
                            placeholder_rows.append(current_row)
                            current_row += 1
                        
                        # Merge the placeholder cells in column D
                        if len(placeholder_rows) > 1:
                            start_row = placeholder_rows[0]
                            end_row = placeholder_rows[-1]
                            worksheet_nmap.merge_range(start_row, 3, end_row, 3, "-", cell_format)
                        break
            
            # Merge only empty PORT/SERVICE cells (not "Filtered" entries)
            columns_to_merge = [1, 2, 4, 5]
            for col in columns_to_merge:
                merge_start = None
                prev_value = None

                for row in range(1, len(nmap_data)):
                    current_value = nmap_data[row][col]
                    
                    # Only merge truly empty cells, not "Filtered" entries
                    if current_value == "" and merge_start is None:
                        merge_start = row - 1
                        prev_value = nmap_data[merge_start][col]
                    # End merge when we hit a non-empty cell (including "Filtered")
                    elif current_value != "" and merge_start is not None:
                        if row - 1 > merge_start:
                            worksheet_nmap.merge_range(merge_start, col, row - 1, col, prev_value, cell_format)
                        merge_start = None
                        prev_value = None
                
                # Handle merge at the end of the data
                if merge_start is not None and len(nmap_data) - 1 > merge_start:
                    worksheet_nmap.merge_range(merge_start, col, len(nmap_data) - 1, col, prev_value, cell_format)
        
        # Process Nessus data
        if nessus_dataframes:
            combined_nessus = pd.concat(nessus_dataframes, ignore_index=True)

            # Convert INF values to string representation
            for col in combined_nessus.columns:
                if combined_nessus[col].dtype == 'float64':
                    combined_nessus[col] = combined_nessus[col].apply(
                        lambda x: 'INF' if math.isinf(x) else x
                    )
            
            worksheet_nessus = workbook.add_worksheet("Nessus CSV Files")
            writer.sheets["Nessus CSV Files"] = worksheet_nessus
            
            # Create cell formats (with and without border)
            wrapped_cell_with_border = workbook.add_format({
                'text_wrap': True,
                'valign': 'vtop',
                'align': 'center',
                'border': 1
            })
            wrapped_cell_no_border = workbook.add_format({
                'text_wrap': True,
                'valign': 'vtop',
                'align': 'center'
            })
            
            # Create header formats (with and without border)
            wrapped_header_with_border = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'vcenter',
                'align': 'center',
                'fg_color': "#1376d1",
                'font_color': 'white',
                'border': 1
            })
            wrapped_header_no_border = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'vcenter',
                'align': 'center',
                'fg_color': "#1376d1",
                'font_color': 'white'
            })
            
            # Write headers with border only for columns A-N (0..13)
            for col_num, value in enumerate(combined_nessus.columns.values):
                header_fmt = wrapped_header_with_border if col_num <= 13 else wrapped_header_no_border
                worksheet_nessus.write(0, col_num, value, header_fmt)
            
            # Write data rows; apply border only for columns A-N (0..13)
            for row_num in range(len(combined_nessus)):
                for col_num in range(len(combined_nessus.columns)):
                    value = combined_nessus.iat[row_num, col_num]
                    cell_fmt = wrapped_cell_with_border if col_num <= 13 else wrapped_cell_no_border
                    
                    # Handle different data types properly
                    if pd.isna(value):
                        worksheet_nessus.write(row_num + 1, col_num, '', cell_fmt)
                    elif isinstance(value, (int, float)):
                        if isinstance(value, float) and math.isinf(value):
                            worksheet_nessus.write(row_num + 1, col_num, 'INF', cell_fmt)
                        else:
                            worksheet_nessus.write_number(row_num + 1, col_num, value, cell_fmt)
                    elif isinstance(value, str):
                        worksheet_nessus.write_string(row_num + 1, col_num, str(value), cell_fmt)
                    elif value is None:
                        worksheet_nessus.write_string(row_num + 1, col_num, 'None', cell_fmt)
                    else:
                        worksheet_nessus.write(row_num + 1, col_num, str(value), cell_fmt)
            
            # Set specific column widths for Nessus worksheet
            column_widths = [20, 20, 25, 15, 25, 20, 15, 40, 40, 80, 40, 30, 80, 40, 20]
            
            for col_num, width in enumerate(column_widths):
                if col_num < len(combined_nessus.columns):
                    worksheet_nessus.set_column(col_num, col_num, width)
            
            if len(combined_nessus.columns) > len(column_widths):
                for col_num in range(len(column_widths), len(combined_nessus.columns)):
                    worksheet_nessus.set_column(col_num, col_num, 25, wrapped_cell_no_border)
            
            # Set default row height only (no global format to avoid borders on all columns)
            worksheet_nessus.set_default_row(17)
            
                        # Calculate unmatched vulnerabilities here while we still have access to the data
            try:
                # Get unique vulnerability names from Name column
                unique_vulnerabilities = set()
                for vulnerabilities in combined_nessus['Name'].dropna():
                    unique_vulnerabilities.add(str(vulnerabilities).strip())
                
                # Load catalog to get matched vulnerabilities
                catalog_path = "static/Formats_and_Catalog/Infrastructure VAPT Catalog.xlsx"
                if os.path.exists(catalog_path):
                    try:
                        catalog_df = pd.read_excel(catalog_path, sheet_name=0)
                    except Exception as e:
                        print(f"Error reading catalog file for vulnerability calculation: {e}")
                        catalog_df = None
                    if catalog_df is not None and 'Vulnerabilities in this group' in catalog_df.columns:
                        catalog_vulnerabilities = set()
                        for vulnerabilities in catalog_df['Vulnerabilities in this group'].dropna():
                            vuln_list = [v.strip() for v in str(vulnerabilities).split(',') if v.strip()]
                            catalog_vulnerabilities.update(vuln_list)
                        
                        matched_vulnerabilities = unique_vulnerabilities.intersection(catalog_vulnerabilities)
                        unmatched_vulnerabilities = unique_vulnerabilities - matched_vulnerabilities
                        unmatched_count = len(unmatched_vulnerabilities)
            except Exception as e:
                print(f"Error calculating unmatched vulnerabilities: {e}")
    
            # Create Scope worksheet from Nessus data
            create_scope_worksheet(workbook, combined_nessus, header_format, cell_format)
            
            # Create Summary worksheet from Nessus data
            create_summary_worksheet(workbook, combined_nessus, header_format, cell_format)
            
            # Get unmatched vulnerability details from session (provided by user)
            unmatched_details_mapping = {}
            try:
                # First try to get from session (new method)
                if 'unmatched_vulnerability_details' in session:
                    session_details = session.get('unmatched_vulnerability_details', {})
                    for vuln_name, details in session_details.items():
                        unmatched_details_mapping[vuln_name] = {
                            'vulnerabilityName': str(details.get('vulnerabilityName', vuln_name)).strip(),
                            'riskFactor': str(details.get('riskFactor', '')).strip(),
                            'cveId': str(details.get('cveId', '')).strip(),
                            'cvss': str(details.get('cvssScore', '')).strip(),
                            'auditObservation': str(details.get('auditObservation', '')).strip(),
                            'impact': str(details.get('impact', '')).strip(),
                            'recommendation': str(details.get('recommendation', '')).strip(),
                            'referenceLink': str(details.get('referenceLink', '')).strip(),
                            'isMerged': details.get('isMerged', False),
                            'mergedVulnerabilities': details.get('mergedVulnerabilities', [])
                        }
                else:
                    # Fallback to old method (form data)
                    details_json = request.form.get('unmatched_vuln_details')
                    if details_json:
                        details_list = json.loads(details_json)
                        if isinstance(details_list, list):
                            for item in details_list:
                                if not isinstance(item, dict):
                                    continue
                                original_name = str(item.get('vulnerabilityOriginal', '')).strip()
                                edited_name = str(item.get('vulnerabilityName', '')).strip()
                                vuln_name = original_name or edited_name
                                if not vuln_name:
                                    continue
                                unmatched_details_mapping[vuln_name] = {
                                    'vulnerabilityName': edited_name or vuln_name,
                                    'riskFactor': str(item.get('riskFactor', '')).strip(),
                                    'cveId': str(item.get('cveId', '')).strip(),
                                    'cvss': str(item.get('cvss', '')).strip(),
                                    'auditObservation': str(item.get('auditObservation', '')).strip(),
                                    'impact': str(item.get('impact', '')).strip(),
                                    'recommendation': str(item.get('recommendation', '')).strip(),
                                    'referenceLink': str(item.get('referenceLink', '')).strip(),
                                }
            except Exception as e:
                print(f"Failed parsing unmatched_vuln_details: {e}")

            # Create Infra_VAPT worksheet from Summary data
            create_infra_vapt_worksheet(workbook, combined_nessus, header_format, cell_format, unmatched_details_mapping)
            

    output.seek(0)
    
    # If evidence files are provided, save Excel file temporarily and add POC images
    if evidence_file and evidence_file.filename != '':
        try:
            # Save Excel file temporarily
            temp_excel_path = "temp_combined_scan_results.xlsx"
            with open(temp_excel_path, 'wb') as temp_file:
                temp_file.write(output.getvalue())
            
            # Extract POC images
            poc_mapping = extract_poc_images(evidence_file)
            
            if poc_mapping:
                # Insert POC images using openpyxl
                insert_poc_images_to_excel(temp_excel_path, poc_mapping, combined_nessus)
                
                # Read the updated file
                with open(temp_excel_path, 'rb') as updated_file:
                    output = BytesIO(updated_file.read())
            
            # Clean up temporary file
            if os.path.exists(temp_excel_path):
                os.remove(temp_excel_path)
                
        except Exception as e:
            print(f"Error processing POC images: {e}")
            # Continue with original file if POC processing fails
    
    # Generate dynamic filename based on organization and end date
    dynamic_filename = generate_dynamic_filename(
        form_metadata.get('organization', ''),
        form_metadata.get('endDate', '')
    )
    
    # Create response
    response = make_response(send_file(
        output,
        as_attachment=True,
        download_name=dynamic_filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ))
    
    # Add custom header with unmatched count
    response.headers['X-Unmatched-Vulnerabilities'] = str(unmatched_count)
    # Add filename to headers for frontend to use
    response.headers['X-Filename'] = dynamic_filename
    
    # Clean up timestamped temp_poc_images folders created during this run
    try:
        import glob
        import shutil
        for d in glob.glob("temp_poc_images_*"):
            try:
                shutil.rmtree(d)
            except Exception:
                pass
        print("Successfully cleaned temp_poc_images_* folders")
    except Exception as e:
        print(f"Error deleting temp_poc_images_* folders: {e}")
    
    return response
    

def convert_risk_to_camelcase(risk_value):
    """Convert risk value to CamelCase format"""
    if not risk_value:
        return ""
    
    risk_str = str(risk_value).strip().lower()
    if risk_str == 'critical':
        return 'Critical'
    elif risk_str == 'high':
        return 'High'
    elif risk_str == 'medium':
        return 'Medium'
    elif risk_str == 'low':
        return 'Low'
    else:
        # Return original value with first letter capitalized
        return risk_str.capitalize()

def create_infra_vapt_worksheet(workbook, combined_nessus, header_format, cell_format, unmatched_details_mapping=None, evidence_files=None):
    """Create Infra_VAPT worksheet with data from Summary's Name column and match with catalog (using merge state)"""
    # Check if required columns exist
    required_columns = ['Name', 'Host', 'Branch Name', 'Risk']
    missing_cols = [col for col in required_columns if col not in combined_nessus.columns]
    if missing_cols:
        print(f"Cannot create Infra_VAPT worksheet - missing columns: {missing_cols}")
        return
    
    try:
        # Get merge state from session
        merge_state = session.get('vulnerability_merge_state', None)
        use_merge_state = merge_state is not None and 'matched_groups' in merge_state
        
        if use_merge_state:
            print("📊 Using merge state from session for vulnerability grouping")
            merged_groups_from_session = merge_state.get('matched_groups', [])
            new_group_details_dict = merge_state.get('new_group_details', {})
        else:
            print("📊 No merge state found - using standard catalog matching")
            merged_groups_from_session = None
            new_group_details_dict = {}
        # Create color formats for risk factors
        critical_format = workbook.add_format({
            'bg_color': '#8B0000',  # Dark Red
            'font_color': 'white',
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        high_format = workbook.add_format({
            'bg_color': '#FF0000',  # Red
            'font_color': 'white',
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        medium_format = workbook.add_format({
            'bg_color': '#FFA500',  # Orange
            'font_color': 'white',
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        low_format = workbook.add_format({
            'bg_color': '#008000',  # Green
            'font_color': 'white',
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # POC column format with border (for cells with objects)
        poc_format = workbook.add_format({
            'border': 1,
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'center'
        })
        
        # POC column format without border (for empty cells)
        poc_format_no_border = workbook.add_format({
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'center'
        })
        
        # Left-aligned format for specific columns
        left_align_format = workbook.add_format({
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'left',
            'border': 1
        })
        
        # Filter only low, medium, high, critical vulnerabilities
        valid_risks = ['low', 'medium', 'high', 'critical']
        df_filtered = combined_nessus.copy()
        df_filtered['Risk'] = df_filtered['Risk'].astype(str).str.lower().str.strip()
        df_filtered = df_filtered[df_filtered['Risk'].isin(valid_risks)]
        
        # Get unique vulnerability names from Name column
        unique_vulnerabilities = df_filtered['Name'].drop_duplicates().tolist()
        
        # Load the catalog file
        catalog_path = os.path.join('static', 'Formats_and_Catalog', 'Infrastructure VAPT Catalog.xlsx')
        
        # Initialize catalog_df and headers variables
        catalog_df = None
        catalog_headers = []
        headers = ["Sr.No", "Vulnerabilities"]
        
        # Try to load catalog file if it exists
        if os.path.exists(catalog_path):
            try:
                catalog_df = pd.read_excel(catalog_path)
                
                # Check if the required column exists in catalog
                if 'Vulnerabilities in this group' in catalog_df.columns:
                    # Define columns to exclude from catalog
                    exclude_columns = ['Sr No', 'Vulnerabilities in this group', 'Affected System']
                    
                    # Get catalog headers excluding the specified columns
                    catalog_headers = [col for col in catalog_df.columns.tolist() if col not in exclude_columns]
                    
                    # Find the positions of key columns
                    rec_countermeasure_col = None
                    reference_link_col = None
                    audit_observation_col = None
                    risk_factor_col = None
                    
                    for i, col_name in enumerate(catalog_headers):
                        if 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                            rec_countermeasure_col = i
                        if 'reference' in col_name.lower() and 'link' in col_name.lower():
                            reference_link_col = i
                        if 'audit' in col_name.lower() and 'observation' in col_name.lower():
                            audit_observation_col = i
                        if 'risk' in col_name.lower() and 'factor' in col_name.lower():
                            risk_factor_col = i
                    
                    # Insert "Affected Systems" column after Recommendation/Countermeasure and before Reference Link
                    if rec_countermeasure_col is not None and reference_link_col is not None:
                        # Insert after Recommendation/Countermeasure
                        insert_position = rec_countermeasure_col + 1
                    elif rec_countermeasure_col is not None:
                        # If only Recommendation/Countermeasure exists, insert after it
                        insert_position = rec_countermeasure_col + 1
                    elif reference_link_col is not None:
                        # If only Reference Link exists, insert before it
                        insert_position = reference_link_col
                    else:
                        # If neither exists, insert at the end
                        insert_position = len(catalog_headers)
                    
                    # Create worksheet headers - insert "Affected Systems" and POC columns at the correct position
                    # Add catalog headers up to the insertion point
                    headers.extend(catalog_headers[:insert_position])
                    # Add "Affected Systems" column
                    headers.append("Affected Systems")
                    # Add remaining catalog headers
                    headers.extend(catalog_headers[insert_position:])
            except Exception as e:
                print(f"Error reading catalog file for Infra_VAPT worksheet: {e}")
                print(f"Catalog file may be corrupted. Will create worksheet with basic headers only.")
        
        # Add POC columns at the end - Now we need L, M, N, O, P, Q, R (7 columns)
        # Only write headers for individual columns, we'll merge them later
        headers.extend(["POC_L", "POC_M", "POC_N", "POC_O", "POC_P", "POC_Q", "POC_R"])
        
        # Create Infra_VAPT worksheet (always create, even if empty)
        worksheet_infra = workbook.add_worksheet("Infra_VAPT")
        
        for col_num, header in enumerate(headers):
            worksheet_infra.write(0, col_num, header, header_format)
        
        # Find POC column positions (first POC column is L)
        poc_col_start = None
        poc_col_end = None
        for i, header in enumerate(headers):
            if header == "POC_L":
                poc_col_start = i
            if header == "POC_R":
                poc_col_end = i
                break
        
        # Merge cells L1:R1 and write "POC" with same format as other headers
        if poc_col_start is not None and poc_col_end is not None:
            worksheet_infra.merge_range(0, poc_col_start, 0, poc_col_end, "POC", header_format)
        
        # Set specific column widths as requested (always set, even if no vulnerabilities)
        # Expected columns: Sr.No, Vulnerabilities, Risk Factor, CVE ID, CVSS, Audit Observation, Impact, 
        #                   Recommendation/Countermeasure, Affected Systems, Reference Link, Status, POC columns
        column_widths = [7, 35, 30, 15, 20, 10, 60, 60, 60, 40, 50, 30]  # First 12 columns (before POC)
        
        # Apply widths to columns before POC columns
        # Determine how many columns we should set (minimum of column_widths length or columns before POC)
        num_cols_to_set = min(len(column_widths), poc_col_start if poc_col_start is not None else len(headers))
        
        for i in range(num_cols_to_set):
            if i < len(column_widths):
                worksheet_infra.set_column(i, i, column_widths[i])
        
        # Set POC columns (L, M, N, O, P, Q, R) to width 16 without border initially
        if poc_col_start is not None and poc_col_end is not None:
            for col_idx in range(poc_col_start, poc_col_end + 1):
                worksheet_infra.set_column(col_idx, col_idx, 16, poc_format_no_border)
        
        print(f"📏 Set column widths: {num_cols_to_set} regular columns + {poc_col_end - poc_col_start + 1 if poc_col_start and poc_col_end else 0} POC columns")
        
        # Create format for green text, bold, centered, with border (for no vulnerabilities message)
        no_data_format = workbook.add_format({
            'font_color': '#008000',  # Green color
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'text_wrap': True
        })
        
        # If no vulnerabilities, add message and return after creating headers and setting column widths
        if not unique_vulnerabilities:
            print("No vulnerabilities found for Infra_VAPT worksheet - creating worksheet with headers only")
            # Merge A3 to R3 (row 2, columns 0 to 17) and add message
            # R is column 17 (0-indexed: A=0, B=1, ..., R=17)
            worksheet_infra.merge_range(2, 0, 2, 17, "No vulnerabilities were identified during the audit.", no_data_format)
            return
        
        # If catalog file was not found or couldn't be read, add message and return after creating headers and setting column widths
        if catalog_df is None or 'Vulnerabilities in this group' not in catalog_df.columns:
            if not os.path.exists(catalog_path):
                print(f"Catalog file not found at: {catalog_path} - creating worksheet with headers only")
            else:
                print("Catalog file does not contain 'Vulnerabilities in this group' column - creating worksheet with headers only")
            # Merge A3 to R3 and add message
            worksheet_infra.merge_range(2, 0, 2, 17, "No vulnerabilities were identified during the audit.", no_data_format)
            return
        
        # Track matched catalog entries to avoid duplicates
        matched_catalog_indices = set()
        matched_vulnerabilities = set()
        vulnerability_groups = {}  # Store which vulnerabilities belong to which catalog group
        vulnerability_affected_systems = {}  # Store affected systems for each vulnerability
        vulnerability_risks = {}  # Store the highest risk for each vulnerability group
        catalog_risk_values = {}  # Store the risk factor from catalog for each catalog entry
        
        # First pass: Collect affected systems for ALL vulnerabilities
        for vulnerability in unique_vulnerabilities:
            # Get all affected systems for this vulnerability (group by Branch Name)
            vuln_data = df_filtered[df_filtered['Name'] == vulnerability]
            branch_hosts = {}
            max_risk = 'low'  # Default to lowest risk
            
            for _, row in vuln_data.iterrows():
                branch = str(row['Branch Name']).strip()
                host = str(row['Host']).strip()
                risk = str(row['Risk']).lower().strip()
                
                # Track the highest risk for this vulnerability
                risk_levels = {'critical': 4, 'high': 3, 'medium': 2, 'low': 1}
                if risk_levels.get(risk, 0) > risk_levels.get(max_risk, 0):
                    max_risk = risk
                
                if branch and host:
                    if branch not in branch_hosts:
                        branch_hosts[branch] = set()
                    branch_hosts[branch].add(host)
            
            # Format the affected systems data
            formatted_systems = []
            for branch, hosts in sorted(branch_hosts.items()):
                formatted_systems.append(branch)
                for host in sorted(hosts):
                    formatted_systems.append(host)
                formatted_systems.append("")  # Add empty line between branches
            
            # Remove the last empty line if it exists
            if formatted_systems and formatted_systems[-1] == "":
                formatted_systems = formatted_systems[:-1]
            
            vulnerability_affected_systems[vulnerability] = formatted_systems
            vulnerability_risks[vulnerability] = max_risk
        
        # Use merge state if available, otherwise use standard catalog matching
        if use_merge_state and merged_groups_from_session:
            print(f"✓ Using {len(merged_groups_from_session)} merged groups from session")
            
            # Build vulnerability_groups dict from merge state
            for group in merged_groups_from_session:
                catalog_idx = group['catalog_id']
                matched_catalog_indices.add(catalog_idx)
                
                # Add all vulnerabilities in this group
                vulnerability_groups[catalog_idx] = group['matched_vulnerabilities']
                for vuln in group['matched_vulnerabilities']:
                    matched_vulnerabilities.add(vuln)
                
                # Store risk factor
                catalog_risk_values[catalog_idx] = str(group.get('risk_factor', '')).upper().strip()
        else:
            # Standard catalog matching (original logic)
            print("✓ Using standard catalog matching")
            for vulnerability in unique_vulnerabilities:
                # Use only first 170 characters for matching (to handle long vulnerability names)
                vuln_short = str(vulnerability)[:170]
                # Escape special characters in vulnerability name for regex
                escaped_vulnerability = re.escape(vuln_short)
                
                # Create pattern that matches vulnerability (first 170 chars only)
                pattern = rf'(?:\n|\r\n|\A){escaped_vulnerability}'
                
                # Find matching rows in catalog
                matching_rows = catalog_df[
                    catalog_df['Vulnerabilities in this group'].str.contains(
                        pattern, 
                        case=False, 
                        na=False,
                        regex=True
                    )
                ]
                
                if not matching_rows.empty:
                    # Store the catalog index and the vulnerability
                    catalog_idx = matching_rows.index[0]
                    matched_catalog_indices.add(catalog_idx)
                    matched_vulnerabilities.add(vulnerability)
                    
                    # Group vulnerabilities by catalog entry
                    if catalog_idx not in vulnerability_groups:
                        vulnerability_groups[catalog_idx] = []
                    vulnerability_groups[catalog_idx].append(vulnerability)
                    
                    # Store the risk factor from catalog
                    if risk_factor_col is not None:
                        catalog_risk_value = catalog_df.loc[catalog_idx, catalog_headers[risk_factor_col]]
                        if pd.isna(catalog_risk_value):
                            catalog_risk_value = ""
                        catalog_risk_values[catalog_idx] = str(catalog_risk_value).upper().strip()
        
        # Sort catalog indices by risk factor (critical, high, medium, low, then others)
        risk_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
        
        def get_risk_priority(catalog_idx):
            risk_value = catalog_risk_values.get(catalog_idx, "")
            return risk_order.get(risk_value, 4)  # Default to 4 for unknown risks
        
        sorted_catalog_indices = sorted(matched_catalog_indices, key=get_risk_priority)
        
        # Second pass: Write data to worksheet, grouping vulnerabilities by catalog entry
        row_num = 1
        
        for catalog_idx in sorted_catalog_indices:
            # Get catalog data - either from actual catalog or from merge state (for new groups)
            if use_merge_state and catalog_idx < 0:
                # This is a new group from merge state
                group_data = next((g for g in merged_groups_from_session if g['catalog_id'] == catalog_idx), None)
                if not group_data:
                    continue
                
                # Get full details from the separate dictionary
                full_details = new_group_details_dict.get(str(catalog_idx), {})
                
                # Create a pseudo catalog row matching actual catalog column names
                catalog_row = pd.Series({
                    'Name of Vulnerability': group_data.get('group_name', ''),
                    'Risk Factor': full_details.get('riskFactor', group_data.get('risk_factor', '')),
                    'CVE ID': full_details.get('cveId', 'N/A'),
                    'CVSS': full_details.get('cvssScore', group_data.get('cvss_score', '')),  # Note: 'CVSS' not 'CVSS Score'
                    'Audit Observation': full_details.get('auditObservation', ''),
                    'Impact': full_details.get('impact', ''),
                    'Recommendation/Countermeasure': full_details.get('recommendation', ''),  # Note: 'Recommendation/Countermeasure'
                    'Reference Link': full_details.get('referenceLink', ''),
                    'Affected System': '',  # Will be populated from scan data
                    'Vulnerabilities in this group': ''  # Not needed for worksheet generation
                })
            else:
                # Standard catalog entry
                catalog_row = catalog_df.loc[catalog_idx]
            
            vulnerabilities_list = vulnerability_groups.get(catalog_idx, [])
            
            # Determine the highest risk for this group of vulnerabilities
            group_max_risk = 'low'
            risk_levels = {'critical': 4, 'high': 3, 'medium': 2, 'low': 1}
            for vuln in vulnerabilities_list:
                if risk_levels.get(vulnerability_risks.get(vuln, 'low'), 0) > risk_levels.get(group_max_risk, 0):
                    group_max_risk = vulnerability_risks.get(vuln, 'low')
            
            # Collect all affected systems for all vulnerabilities in this group
            all_affected_systems = []
            branch_hosts_combined = {}
            
            for vuln in vulnerabilities_list:
                if vuln in vulnerability_affected_systems:
                    # Parse the formatted systems to combine branches and hosts
                    current_branch = None
                    for line in vulnerability_affected_systems[vuln]:
                        if line and not line.startswith(('192.168.', '10.', '172.')) and not re.match(r'\d+\.\d+\.\d+\.\d+', line):
                            # This is a branch name
                            current_branch = line
                            if current_branch not in branch_hosts_combined:
                                branch_hosts_combined[current_branch] = set()
                        elif line and current_branch:
                            # This is a host IP for the current branch
                            branch_hosts_combined[current_branch].add(line)
            
            # Format the combined affected systems data
            for branch, hosts in sorted(branch_hosts_combined.items()):
                all_affected_systems.append(branch)
                for host in sorted(hosts):
                    all_affected_systems.append(host)
                all_affected_systems.append("")  # Add empty line between branches
            
            # Remove the last empty line if it exists
            if all_affected_systems and all_affected_systems[-1] == "":
                all_affected_systems = all_affected_systems[:-1]
            
            # Write serial number
            worksheet_infra.write(row_num, 0, row_num, cell_format)
            
            # Write vulnerabilities (each on a new line)
            vulnerabilities_str = "\n".join(sorted(vulnerabilities_list))
            worksheet_infra.write(row_num, 1, vulnerabilities_str, cell_format)
            
            # Write catalog data up to the insertion point
            col_idx = 2
            for i in range(insert_position):
                col_name = catalog_headers[i]
                value = catalog_row[col_name]
                if pd.isna(value):
                    value = ""
                
                # Special handling for Risk Factor column - USE CATALOG VALUE INSTEAD OF SCAN RISK
                if risk_factor_col is not None and i == risk_factor_col:
                    # Use the risk factor from the catalog, not from the scan results
                    catalog_risk_value = str(value).strip() if value else ""
                    
                    # Apply color formatting based on the catalog risk level
                    if catalog_risk_value.upper() == 'CRITICAL':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        worksheet_infra.write(row_num, col_idx, camelcase_value, critical_format)
                    elif catalog_risk_value.upper() == 'HIGH':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        worksheet_infra.write(row_num, col_idx, camelcase_value, high_format)
                    elif catalog_risk_value.upper() == 'MEDIUM':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        worksheet_infra.write(row_num, col_idx, camelcase_value, medium_format)
                    elif catalog_risk_value.upper() == 'LOW':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        worksheet_infra.write(row_num, col_idx, camelcase_value, low_format)
                    else:
                        # If catalog doesn't have a risk factor, use the highest from scan results
                        camelcase_value = convert_risk_to_camelcase(group_max_risk)
                        if group_max_risk == 'critical':
                            worksheet_infra.write(row_num, col_idx, camelcase_value, critical_format)
                        elif group_max_risk == 'high':
                            worksheet_infra.write(row_num, col_idx, camelcase_value, high_format)
                        elif group_max_risk == 'medium':
                            worksheet_infra.write(row_num, col_idx, camelcase_value, medium_format)
                        elif group_max_risk == 'low':
                            worksheet_infra.write(row_num, col_idx, camelcase_value, low_format)
                        else:
                            worksheet_infra.write(row_num, col_idx, camelcase_value, cell_format)
                else:
                    # Special handling for Audit Observation column
                    if audit_observation_col is not None and i == audit_observation_col:
                        # Only add the line if there are multiple vulnerabilities
                        if len(vulnerabilities_list) > 1:
                            observation_text = "It was observed that the hosts are affected by multiple vulnerabilities, which are listed below.\n\n"
                            observation_text += vulnerabilities_str
                            
                            if value:
                                value = f"{value}\n\n{observation_text}"
                            else:
                                value = observation_text
                        # For single vulnerability, keep original content or add basic text
                        elif not value:
                            value = "It was observed that the host is affected by a vulnerability."
                    
                    # Special handling for CVE ID column - fill with "N/A" if empty
                    if col_name == 'CVE ID' and value == "":
                        value = "N/A"
                    
                    # Apply left alignment for specific columns
                    if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                        worksheet_infra.write(row_num, col_idx, str(value), left_align_format)
                    else:
                        worksheet_infra.write(row_num, col_idx, str(value), cell_format)
                
                col_idx += 1
            
            # Write affected systems (formatted with branch names and IPs on separate lines)
            affected_systems_str = "\n".join(all_affected_systems)
            worksheet_infra.write(row_num, col_idx, affected_systems_str, cell_format)
            col_idx += 1
            
            # Write remaining catalog data
            for i in range(insert_position, len(catalog_headers)):
                col_name = catalog_headers[i]
                value = catalog_row[col_name]
                if pd.isna(value):
                    value = ""
                
                # Special handling for Risk Factor column - USE CATALOG VALUE INSTEAD OF SCAN RISK
                if risk_factor_col is not None and i == risk_factor_col:
                    # Use the risk factor from the catalog, not from the scan results
                    catalog_risk_value = str(value).strip() if value else ""
                    
                    # Apply color formatting based on the catalog risk level
                    if catalog_risk_value.upper() == 'CRITICAL':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        worksheet_infra.write(row_num, col_idx, camelcase_value, critical_format)
                    elif catalog_risk_value.upper() == 'HIGH':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        worksheet_infra.write(row_num, col_idx, camelcase_value, high_format)
                    elif catalog_risk_value.upper() == 'MEDIUM':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        worksheet_infra.write(row_num, col_idx, camelcase_value, medium_format)
                    elif catalog_risk_value.upper() == 'LOW':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        worksheet_infra.write(row_num, col_idx, camelcase_value, low_format)
                    else:
                        # If catalog doesn't have a risk factor, use the highest from scan results
                        camelcase_value = convert_risk_to_camelcase(group_max_risk)
                        if group_max_risk == 'critical':
                            worksheet_infra.write(row_num, col_idx, camelcase_value, critical_format)
                        elif group_max_risk == 'high':
                            worksheet_infra.write(row_num, col_idx, camelcase_value, high_format)
                        elif group_max_risk == 'medium':
                            worksheet_infra.write(row_num, col_idx, camelcase_value, medium_format)
                        elif group_max_risk == 'low':
                            worksheet_infra.write(row_num, col_idx, camelcase_value, low_format)
                        else:
                            worksheet_infra.write(row_num, col_idx, camelcase_value, cell_format)
                else:
                    # Special handling for Audit Observation column
                    if audit_observation_col is not None and i == audit_observation_col:
                        # Only add the line if there are multiple vulnerabilities
                        if len(vulnerabilities_list) > 1:
                            observation_text = "It was observed that the hosts are affected by multiple vulnerabilities, which are listed below.\n\n"
                            observation_text += vulnerabilities_str
                            
                            if value:
                                value = f"{value}\n\n{observation_text}"
                            else:
                                value = observation_text
                        # For single vulnerability, keep original content or add basic text
                        elif not value:
                            value = "It was observed that the host is affected by a vulnerability."
                    
                    # Special handling for CVE ID column - fill with "N/A" if empty
                    if col_name == 'CVE ID' and value == "":
                        value = "N/A"
                    
                    # Apply left alignment for specific columns
                    if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                        worksheet_infra.write(row_num, col_idx, str(value), left_align_format)
                    else:
                        worksheet_infra.write(row_num, col_idx, str(value), cell_format)
                
                col_idx += 1
            
            # Write POC columns at the end (borders added conditionally by image insertion)
            for col_idx in range(poc_col_start, poc_col_end + 1):
                worksheet_infra.write(row_num, col_idx, "", poc_format_no_border)
            
            row_num += 1
        
        # Handle unmatched vulnerabilities - including merged ones
        unmatched_vulnerabilities = set(unique_vulnerabilities) - matched_vulnerabilities
        if unmatched_vulnerabilities:
            print(f"Unmatched vulnerabilities: {unmatched_vulnerabilities}")
            
            # Process merged vulnerabilities first
            merged_vuln_processed = set()
            for vulnerability in unmatched_vulnerabilities:
                if vulnerability in unmatched_details_mapping:
                    details = unmatched_details_mapping[vulnerability]
                    if details.get('isMerged', False):
                        merged_vulns = details.get('mergedVulnerabilities', [])
                        if merged_vulns:
                            # Process merged vulnerability group
                            merged_vuln_processed.update(merged_vulns)
                            
                            # Create vulnerabilities string for "Vulnerabilities in this group" column
                            vulnerabilities_in_group = "\n".join(merged_vulns)
                            
                            worksheet_infra.write(row_num, 0, row_num, cell_format)
                            
                            # Use the first vulnerability name as the main name
                            main_name = merged_vulns[0] if merged_vulns else str(vulnerability)
                            worksheet_infra.write(row_num, 1, main_name, cell_format)
                            
                            # Write catalog data with merged vulnerability details
                            col_idx = 2
                            for i in range(insert_position):
                                col_name = catalog_headers[i]
                                value_to_write = ""
                                
                                if risk_factor_col is not None and i == risk_factor_col:
                                    value_to_write = str(details.get('riskFactor', '')).upper()
                                elif audit_observation_col is not None and i == audit_observation_col:
                                    value_to_write = details.get('auditObservation', '')
                                elif col_name == 'CVE ID':
                                    value_to_write = details.get('cveId', '') or "N/A"
                                elif 'cvss' in col_name.lower():
                                    value_to_write = details.get('cvss', '')
                                elif 'impact' in col_name.lower():
                                    value_to_write = details.get('impact', '')
                                elif 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                                    value_to_write = details.get('recommendation', '')
                                elif 'reference' in col_name.lower() and 'link' in col_name.lower():
                                    value_to_write = details.get('referenceLink', '')
                                
                                # Apply risk color if applicable
                                if risk_factor_col is not None and i == risk_factor_col and value_to_write:
                                    catalog_risk_value = str(value_to_write).upper().strip()
                                    if catalog_risk_value == 'CRITICAL':
                                        worksheet_infra.write(row_num, col_idx, catalog_risk_value, critical_format)
                                    elif catalog_risk_value == 'HIGH':
                                        worksheet_infra.write(row_num, col_idx, catalog_risk_value, high_format)
                                    elif catalog_risk_value == 'MEDIUM':
                                        worksheet_infra.write(row_num, col_idx, catalog_risk_value, medium_format)
                                    elif catalog_risk_value == 'LOW':
                                        worksheet_infra.write(row_num, col_idx, catalog_risk_value, low_format)
                                    else:
                                        worksheet_infra.write(row_num, col_idx, catalog_risk_value, cell_format)
                                else:
                                    # Apply left alignment for specific columns
                                    if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                                        worksheet_infra.write(row_num, col_idx, value_to_write, left_align_format)
                                    else:
                                        worksheet_infra.write(row_num, col_idx, value_to_write, cell_format)
                                col_idx += 1
                            
                            # Write empty affected systems for merged vulnerabilities
                            worksheet_infra.write(row_num, col_idx, "", cell_format)
                            col_idx += 1
                            
                            # Write remaining catalog columns
                            for i in range(insert_position, len(catalog_headers)):
                                col_name = catalog_headers[i]
                                value_to_write = ""
                                
                                if risk_factor_col is not None and i == risk_factor_col:
                                    value_to_write = str(details.get('riskFactor', '')).upper()
                                elif audit_observation_col is not None and i == audit_observation_col:
                                    value_to_write = details.get('auditObservation', '')
                                elif col_name == 'CVE ID':
                                    value_to_write = details.get('cveId', '') or "N/A"
                                elif 'cvss' in col_name.lower():
                                    value_to_write = details.get('cvss', '')
                                elif 'impact' in col_name.lower():
                                    value_to_write = details.get('impact', '')
                                elif 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                                    value_to_write = details.get('recommendation', '')
                                elif 'reference' in col_name.lower() and 'link' in col_name.lower():
                                    value_to_write = details.get('referenceLink', '')
                                elif col_name == 'Vulnerabilities in this group':
                                    value_to_write = vulnerabilities_in_group
                                
                                # Apply risk color if applicable
                                if risk_factor_col is not None and i == risk_factor_col and value_to_write:
                                    catalog_risk_value = str(value_to_write).upper().strip()
                                    if catalog_risk_value == 'CRITICAL':
                                        worksheet_infra.write(row_num, col_idx, catalog_risk_value, critical_format)
                                    elif catalog_risk_value == 'HIGH':
                                        worksheet_infra.write(row_num, col_idx, catalog_risk_value, high_format)
                                    elif catalog_risk_value == 'MEDIUM':
                                        worksheet_infra.write(row_num, col_idx, catalog_risk_value, medium_format)
                                    elif catalog_risk_value == 'LOW':
                                        worksheet_infra.write(row_num, col_idx, catalog_risk_value, low_format)
                                    else:
                                        worksheet_infra.write(row_num, col_idx, catalog_risk_value, cell_format)
                                else:
                                    # Apply left alignment for specific columns
                                    if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                                        worksheet_infra.write(row_num, col_idx, value_to_write, left_align_format)
                                    else:
                                        worksheet_infra.write(row_num, col_idx, value_to_write, cell_format)
                                col_idx += 1
                            
                            # Write POC columns at the end (borders added conditionally by image insertion)
                            for col_idx in range(poc_col_start, poc_col_end + 1):
                                worksheet_infra.write(row_num, col_idx, "", poc_format_no_border)
                            
                            row_num += 1
            
            # Process individual unmatched vulnerabilities (not merged)
            for vulnerability in unmatched_vulnerabilities:
                if vulnerability not in merged_vuln_processed:
                    # Get affected systems for this unmatched vulnerability
                    affected_systems = vulnerability_affected_systems.get(vulnerability, [])
                    affected_systems_str = "\n".join(affected_systems)
                    
                    worksheet_infra.write(row_num, 0, row_num, cell_format)
                    # Use edited vulnerability name if provided
                    if unmatched_details_mapping and vulnerability in unmatched_details_mapping:
                        edited_name = unmatched_details_mapping[vulnerability].get('vulnerabilityName') or str(vulnerability)
                    else:
                        edited_name = str(vulnerability)
                    worksheet_infra.write(row_num, 1, edited_name, cell_format)
                    
                    # Write empty values for catalog columns up to the insertion point
                    col_idx = 2
                    for i in range(insert_position):
                        col_name = catalog_headers[i]
                        value_to_write = ""
                        if unmatched_details_mapping and vulnerability in unmatched_details_mapping:
                            details = unmatched_details_mapping[vulnerability]
                            if risk_factor_col is not None and i == risk_factor_col:
                                value_to_write = str(details.get('riskFactor', '')).upper()
                            elif audit_observation_col is not None and i == audit_observation_col:
                                value_to_write = details.get('auditObservation', '')
                            elif col_name == 'CVE ID':
                                value_to_write = details.get('cveId', '') or "N/A"
                            elif 'cvss' in col_name.lower():
                                value_to_write = details.get('cvss', '')
                            elif 'impact' in col_name.lower():
                                value_to_write = details.get('impact', '')
                            elif 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                                value_to_write = details.get('recommendation', '')
                            elif 'reference' in col_name.lower() and 'link' in col_name.lower():
                                value_to_write = details.get('referenceLink', '')
                        # Apply risk color if applicable
                        if risk_factor_col is not None and i == risk_factor_col and value_to_write:
                            catalog_risk_value = str(value_to_write).upper().strip()
                            if catalog_risk_value == 'CRITICAL':
                                worksheet_infra.write(row_num, col_idx, catalog_risk_value, critical_format)
                            elif catalog_risk_value == 'HIGH':
                                worksheet_infra.write(row_num, col_idx, catalog_risk_value, high_format)
                            elif catalog_risk_value == 'MEDIUM':
                                worksheet_infra.write(row_num, col_idx, catalog_risk_value, medium_format)
                            elif catalog_risk_value == 'LOW':
                                worksheet_infra.write(row_num, col_idx, catalog_risk_value, low_format)
                            else:
                                worksheet_infra.write(row_num, col_idx, catalog_risk_value, cell_format)
                        else:
                            worksheet_infra.write(row_num, col_idx, value_to_write, cell_format)
                        col_idx += 1
                    
                    # Write affected systems
                    worksheet_infra.write(row_num, col_idx, affected_systems_str, cell_format)
                    col_idx += 1
                    
                    # Write remaining catalog columns, filling from details when possible
                    for i in range(insert_position, len(catalog_headers)):
                        col_name = catalog_headers[i]
                        value_to_write = ""
                        if unmatched_details_mapping and vulnerability in unmatched_details_mapping:
                            details = unmatched_details_mapping[vulnerability]
                            if risk_factor_col is not None and i == risk_factor_col:
                                value_to_write = str(details.get('riskFactor', '')).upper()
                            elif audit_observation_col is not None and i == audit_observation_col:
                                value_to_write = details.get('auditObservation', '')
                            elif col_name == 'CVE ID':
                                value_to_write = details.get('cveId', '') or "N/A"
                            elif 'cvss' in col_name.lower():
                                value_to_write = details.get('cvss', '')
                            elif 'impact' in col_name.lower():
                                value_to_write = details.get('impact', '')
                            elif 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                                value_to_write = details.get('recommendation', '')
                            elif 'reference' in col_name.lower() and 'link' in col_name.lower():
                                value_to_write = details.get('referenceLink', '')
                        # Apply risk color if applicable
                        if risk_factor_col is not None and i == risk_factor_col and value_to_write:
                            catalog_risk_value = str(value_to_write).upper().strip()
                            if catalog_risk_value == 'CRITICAL':
                                worksheet_infra.write(row_num, col_idx, catalog_risk_value, critical_format)
                            elif catalog_risk_value == 'HIGH':
                                worksheet_infra.write(row_num, col_idx, catalog_risk_value, high_format)
                            elif catalog_risk_value == 'MEDIUM':
                                worksheet_infra.write(row_num, col_idx, catalog_risk_value, medium_format)
                            elif catalog_risk_value == 'LOW':
                                worksheet_infra.write(row_num, col_idx, catalog_risk_value, low_format)
                            else:
                                worksheet_infra.write(row_num, col_idx, catalog_risk_value, cell_format)
                        else:
                            worksheet_infra.write(row_num, col_idx, value_to_write, cell_format)
                        col_idx += 1
                    
                    # Write POC columns at the end (borders added conditionally by image insertion)
                    for col_idx in range(poc_col_start, poc_col_end + 1):
                        worksheet_infra.write(row_num, col_idx, "", poc_format_no_border)
                    
                    row_num += 1
        
        
        # Set fixed row height of ~50px (≈37.5 points) for all content rows from second row onward
        for r in range(1, row_num):
            worksheet_infra.set_row(r, 37.5)
        
        print(f"Created Infra_VAPT worksheet with {row_num - 1} rows")
        print(f"Matched vulnerabilities: {len(matched_vulnerabilities)}, Unmatched: {len(unmatched_vulnerabilities)}")
        print(f"Unique catalog entries used: {len(sorted_catalog_indices)}")
        
    except Exception as e:
        print(f"Error creating Infra_VAPT worksheet: {e}")
        import traceback
        traceback.print_exc()

def create_scope_worksheet(workbook, combined_nessus, header_format, cell_format):
    """Create Scope worksheet with Branch Name and Host data from Nessus"""
    if 'Branch Name' in combined_nessus.columns and 'Host' in combined_nessus.columns:
        # Extract unique Branch Name and Host combinations
        scope_data = combined_nessus[['Branch Name', 'Host']].drop_duplicates()
        
        # Group by Branch Name and collect all unique Hosts
        branch_ip_mapping = {}
        for branch_name, group in scope_data.groupby('Branch Name'):
            ips = sorted([str(ip).strip() for ip in group['Host'].unique() if pd.notna(ip) and str(ip).strip()])
            if ips:
                branch_ip_mapping[branch_name] = ips
        
        # Create Scope worksheet
        worksheet_scope = workbook.add_worksheet("Scope")
        
        # Write headers
        worksheet_scope.write(0, 0, "Sr.No", header_format)
        worksheet_scope.write(0, 1, "BRANCH NAME", header_format)
        worksheet_scope.write(0, 2, "HOST", header_format)
        
        row = 1
        merge_ranges = {}
        serial_number = 1
        
        # Write branch names and IPs with serial numbers
        for branch_name, ips in branch_ip_mapping.items():
            if not ips:
                continue
            
            start_row = row
            worksheet_scope.write(row, 0, serial_number, cell_format)
            worksheet_scope.write(row, 1, branch_name, cell_format)
            
            for ip in ips:
                worksheet_scope.write(row, 2, ip, cell_format)
                row += 1
            
            end_row = row - 1
            
            # Only merge if there are multiple IPs for this branch
            if end_row > start_row:
                merge_ranges[branch_name] = {
                    'start_row': start_row,
                    'end_row': end_row,
                    'serial_number': serial_number
                }
            
            serial_number += 1
        
        # Apply merging for branch names and serial numbers
        for branch_name, merge_data in merge_ranges.items():
            start_row = merge_data['start_row']
            end_row = merge_data['end_row']
            serial_number = merge_data['serial_number']
            
            if start_row != end_row:
                worksheet_scope.merge_range(
                    start_row, 0, end_row, 0,
                    serial_number,
                    cell_format
                )
                worksheet_scope.merge_range(
                    start_row, 1, end_row, 1,
                    branch_name,
                    cell_format
                )
        
        # Set column widths
        worksheet_scope.set_column('A:A', 8)
        worksheet_scope.set_column('B:B', 30)
        worksheet_scope.set_column('C:C', 20)

def create_summary_worksheet(workbook, combined_nessus, header_format, cell_format):
    """Create Summary worksheet with Name, Branch Name, and Host data from Nessus"""
    # Check if required columns exist
    required_columns = ['Name', 'Host', 'Branch Name', 'Risk']
    missing_cols = [col for col in required_columns if col not in combined_nessus.columns]
    if missing_cols:
        print(f"Cannot create Summary worksheet - missing columns: {missing_cols}")
        return
    
    try:
        # Filter only low, medium, high, critical vulnerabilities
        valid_risks = ['low', 'medium', 'high', 'critical']
        df_filtered = combined_nessus.copy()
        df_filtered['Risk'] = df_filtered['Risk'].astype(str).str.lower().str.strip()
        df_filtered = df_filtered[df_filtered['Risk'].isin(valid_risks)]
        
        # Select required columns
        df_summary = df_filtered[['Name', 'Branch Name', 'Host']].copy()
        
        # Clean data
        df_summary['Name'] = df_summary['Name'].astype(str).str.strip()
        df_summary['Branch Name'] = df_summary['Branch Name'].astype(str).str.strip()
        df_summary['Host'] = df_summary['Host'].astype(str).str.strip()
        
        # Remove rows with empty values
        df_summary = df_summary.dropna(subset=['Name', 'Host'])
        df_summary = df_summary[(df_summary['Name'] != '') & (df_summary['Host'] != '')]
        
        # Create Summary worksheet (always create, even if empty)
        worksheet_summary = workbook.add_worksheet("Summary")
        
        # Write headers with Sr.No as first column
        headers = ["Sr.No", "Name", "Branch Name", "Host"]
        for col_num, header in enumerate(headers):
            worksheet_summary.write(0, col_num, header, header_format)
        
        # Set column widths as requested (always set, even if no data)
        worksheet_summary.set_column('A:A', 8)  # Sr.No column
        worksheet_summary.set_column('B:B', 110)  # Name column
        worksheet_summary.set_column('C:C', 50)  # Branch Name column
        worksheet_summary.set_column('D:D', 40)  # Host column
        
        # If no data, add message and return after creating headers and setting column widths
        if df_summary.empty:
            print("No data available for Summary worksheet after filtering - creating worksheet with headers only")
            # Create format for green text, bold, centered, with border
            no_data_format = workbook.add_format({
                'font_color': '#008000',  # Green color
                'bold': True,
                'align': 'center',
                'valign': 'vcenter',
                'border': 1,
                'text_wrap': True
            })
            # Merge A3 to D3 and add message
            worksheet_summary.merge_range(2, 0, 2, 3, "No vulnerabilities were identified during the audit.", no_data_format)
            return
        
        # Deduplicate based on Name + Host combination (vulnerability + IP)
        # Keep only unique combinations of vulnerability and IP
        df_summary = df_summary.drop_duplicates(subset=['Name', 'Host'], keep='first')
        
        # Sort data by Name, Branch Name, and Host to group similar entries
        df_summary = df_summary.sort_values(by=['Name', 'Branch Name', 'Host'])
        
        # Write data and prepare for merging
        row = 1
        current_vulnerability = None
        current_branch = None
        name_merge_start = 1
        branch_merge_start = 1
        serial_counter = 1  # Initialize serial number counter
        
        for _, row_data in df_summary.iterrows():
            name = str(row_data['Name'])
            branch = str(row_data['Branch Name'])
            host = str(row_data['Host'])
            
            # If we're starting a new vulnerability
            if name != current_vulnerability:
                # Merge previous vulnerability name cells if needed
                if current_vulnerability is not None and row > name_merge_start:
                    worksheet_summary.merge_range(
                        name_merge_start, 1, row - 1, 1,  # Name column (column B)
                        current_vulnerability,
                        cell_format
                    )
                    # Also merge serial numbers for this vulnerability
                    worksheet_summary.merge_range(
                        name_merge_start, 0, row - 1, 0,  # Sr.No column (column A)
                        serial_counter - 1,  # Use the previous serial number
                        cell_format
                    )
                current_vulnerability = name
                name_merge_start = row
                current_branch = None  # Reset branch tracking
                serial_counter += 1  # Increment serial number for new vulnerability
            
            # If we're starting a new branch for this vulnerability
            if branch != current_branch:
                # Merge previous branch name cells if needed
                if current_branch is not None and row > branch_merge_start:
                    worksheet_summary.merge_range(
                        branch_merge_start, 2, row - 1, 2,  # Branch Name column (column C)
                        current_branch,
                        cell_format
                    )
                current_branch = branch
                branch_merge_start = row
            
            # Write serial number (same for all rows of this vulnerability)
            worksheet_summary.write(row, 0, serial_counter - 1, cell_format)  # Sr.No column
            worksheet_summary.write(row, 1, name, cell_format)  # Name column
            worksheet_summary.write(row, 2, branch, cell_format)  # Branch Name column
            worksheet_summary.write(row, 3, host, cell_format)  # Host column
            
            row += 1
        
        # Merge the last group of same vulnerability names
        if current_vulnerability is not None and row > name_merge_start:
            worksheet_summary.merge_range(
                name_merge_start, 1, row - 1, 1,  # Name column (column B)
                current_vulnerability,
                cell_format
            )
            # Merge the last group of serial numbers
            worksheet_summary.merge_range(
                name_merge_start, 0, row - 1, 0,  # Sr.No column (column A)
                serial_counter - 1,
                cell_format
            )
        
        # Merge the last group of same branch names
        if current_branch is not None and row > branch_merge_start:
            worksheet_summary.merge_range(
                branch_merge_start, 2, row - 1, 2,  # Branch Name column (column C)
                current_branch,
                cell_format
            )
        
        # Column widths are already set earlier (after headers are written)
        
    except Exception as e:
        print(f"Error creating Summary worksheet: {e}")

def create_meta_data_worksheet(workbook, form_metadata, header_format, cell_format):
    """Create Meta_Data worksheet with form information from user input"""
    try:
        # Create Meta_Data worksheet
        worksheet_meta = workbook.add_worksheet("Meta_Data")
        
        # Define the data structure for the metadata
        metadata_sections = [
            {
                'title': 'ORGANIZATION INFORMATION',
                'data': [
                    ('Organization Name', form_metadata.get('organization', '') or form_metadata.get('otherOrganization', '')),
                    ('City', form_metadata.get('city', '') or form_metadata.get('otherCity', '')),
                    ('State', form_metadata.get('state', ''))
                ]
            },
            {
                'title': 'AUDIT PERIOD',
                'data': [
                    ('Start Date', form_metadata.get('startDate', '')),
                    ('End Date', form_metadata.get('endDate', ''))
                ]
            },
            {
                'title': 'REPORT PREPARED BY',
                'data': [
                    ('Name', f"{form_metadata.get('preparedByTitle', '')} {form_metadata.get('preparedByName', '')}".strip()),
                ]
            },
            {
                'title': 'AUDITEE DETAILS',
                'data': [
                    ('Name', f"{form_metadata.get('auditeeTitle', '')} {form_metadata.get('auditeeName', '')}".strip()),
                    ('Designation', form_metadata.get('designation', ''))
                ]
            }
        ]
        
        # Add Bank Email Addresses section
        bank_emails = form_metadata.get('bankEmails', [])
        if bank_emails:
            bank_email_data = []
            for i, email in enumerate(bank_emails, 1):
                if email.strip():
                    bank_email_data.append((f'Email {i}', email.strip()))
            
            if bank_email_data:
                metadata_sections.append({
                    'title': 'BANK EMAIL ADDRESSES',
                    'data': bank_email_data
                })
        
        # Add Auditing Team section
        team_names = form_metadata.get('teamNames', [])
        team_designations = form_metadata.get('teamDesignations', [])
        team_emails = form_metadata.get('teamEmails', [])
        team_qualifications = form_metadata.get('teamQualifications', [])
        team_certified = form_metadata.get('teamCertified', [])
        
        if team_names:
            # Add team member details as separate entries for each member
            for i in range(len(team_names)):
                if team_names[i].strip():
                    team_member_data = [
                        (f'Team Member {i+1} - Name', team_names[i].strip()),
                        (f'Team Member {i+1} - Designation', team_designations[i] if i < len(team_designations) else ''),
                        (f'Team Member {i+1} - Email', team_emails[i] if i < len(team_emails) else ''),
                        (f'Team Member {i+1} - Qualification', team_qualifications[i] if i < len(team_qualifications) else ''),
                        (f'Team Member {i+1} - Certified', 
                         'Yes' if team_certified[i].lower() == 'yes' else 'No' if team_certified[i].lower() == 'no' else team_certified[i] 
                         if i < len(team_certified) else '')
                    ]
                    
                    metadata_sections.append({
                        'title': f'AUDITING TEAM MEMBER {i+1}',
                        'data': team_member_data
                    })
        
        # Write data to worksheet
        row = 0
        
        for section in metadata_sections:
            # Write section title
            worksheet_meta.write(row, 0, section['title'], header_format)
            worksheet_meta.write(row, 1, '', header_format)  # Empty cell for formatting
            row += 1
            
            # Write section data
            for field_name, field_value in section['data']:
                worksheet_meta.write(row, 0, field_name, cell_format)
                worksheet_meta.write(row, 1, field_value, cell_format)
                row += 1
            
            # Add empty row after each section
            row += 1
        
        # Set column widths
        worksheet_meta.set_column('A:A', 30)  # Field names column
        worksheet_meta.set_column('B:B', 60)  # Values column
        
        # Set row heights for better readability
        for r in range(row):
            worksheet_meta.set_row(r, 20)
        
        print(f"Created Meta_Data worksheet with {row} rows")
        
    except Exception as e:
        print(f"Error creating Meta_Data worksheet: {e}")
        import traceback
        traceback.print_exc()

def process_nmap_zip(file):
    """Process Nmap zip file and return data for Excel"""
    if file and file.filename.endswith('.zip'):
        zip_data = file.read()
        ip_ports = {}
        
        with zipfile.ZipFile(BytesIO(zip_data), 'r') as zip_ref:
            file_list = zip_ref.namelist()
            
            for file_name in file_list:
                if file_name.endswith('/'):
                    continue
                
                try:
                    with zip_ref.open(file_name) as f:
                        file_content = f.read().decode('utf-8', errors='ignore')
                    
                    ip_pattern = r"Nmap scan report for (?:[a-zA-Z0-9.-]+ )?\(?(\d+\.\d+\.\d+\.\d+)\)?"
                    port_state_pattern = r"(\d+)/(tcp|udp)\s+(open|filtered|closed|unfiltered)\s+([\w-]*)"
                    filtered_ports_pattern = r"Not shown: (\d+) filtered tcp ports"
                    
                    lines = file_content.splitlines()
                    current_ip = None
                    has_filtered_ports = False

                    for line in lines:
                        ip_match = re.search(ip_pattern, line, re.IGNORECASE)
                        if ip_match:
                            current_ip = ip_match.group(1)
                            if current_ip not in ip_ports:
                                ip_ports[current_ip] = []
                            has_filtered_ports = False
                            continue
                        
                        if not current_ip:
                            continue
                        
                        # Check for filtered ports message
                        filtered_match = re.search(filtered_ports_pattern, line)
                        if filtered_match:
                            has_filtered_ports = True
                            continue
                        
                        port_match = re.search(port_state_pattern, line)
                        if port_match:
                            port = port_match.group(1)
                            state = port_match.group(3)
                            service = port_match.group(4) or state
                            
                            if (port, service) not in ip_ports[current_ip]:
                                ip_ports[current_ip].append((port, service))
                    
                    # If no open ports found but filtered ports detected, add "Filtered" entry
                    if current_ip and has_filtered_ports and len(ip_ports[current_ip]) == 0:
                        ip_ports[current_ip].append(("Filtered", "Filtered"))

                    if file_name.endswith('.csv'):
                        try:
                            with zip_ref.open(file_name) as f:
                                csv_content = f.read()
                            
                            df = pd.read_csv(io.BytesIO(csv_content), 
                                        on_bad_lines="skip", 
                                        encoding="utf-8")
                            
                            if all(col in df.columns for col in ['host', 'port', 'service']):
                                for _, row in df.iterrows():
                                    ip = str(row['host']).strip()
                                    port = str(row['port']).strip()
                                    service = str(row['service']).strip()
                                    
                                    if ip and port and service:
                                        if ip not in ip_ports:
                                            ip_ports[ip] = []
                                        if (port, service) not in ip_ports[ip]:
                                            ip_ports[ip].append((port, service))
                        except Exception as e:
                            print(f"CSV processing error in {file_name}: {e}")
                            continue

                except Exception as e:
                    print(f"Error processing file {file_name}: {e}")
                    continue

        # Prepare data for Excel
        all_ips = sorted(ip_ports.keys())
        has_placeholder = False
        
        if len(all_ips) % 2 != 0:
            all_ips.append("-")
            ip_ports["-"] = [("-", "-")]
            has_placeholder = True

        data = []
        for i in range(0, len(all_ips), 2):
            ip1 = all_ips[i]
            ip2 = all_ips[i + 1]
            ports1 = ip_ports.get(ip1, [])
            ports2 = ip_ports.get(ip2, [])
            
            # If no ports found for an IP, add "Filtered" entry
            if len(ports1) == 0:
                ports1 = [("Filtered", "Filtered")]
            if len(ports2) == 0:
                ports2 = [("Filtered", "Filtered")]
            
            max_ports = max(len(ports1), len(ports2))

            data.append(["HOST", "PORT", "SERVICE", "HOST", "PORT", "SERVICE"])

            for j in range(max_ports):
                data.append([
                    ip1, 
                    ports1[j][0] if j < len(ports1) else "", 
                    ports1[j][1] if j < len(ports1) else "",
                    ip2, 
                    ports2[j][0] if j < len(ports2) else "", 
                    ports2[j][1] if j < len(ports2) else ""
                ])

        return data
    
    return None

def process_nessus_zip(file):
    """Process Nessus zip file and return list of DataFrames"""
    if file and file.filename.endswith('.zip'):
        zip_data = file.read()
        all_nessus_data = []
        
        with zipfile.ZipFile(BytesIO(zip_data), 'r') as zip_ref:
            file_list = zip_ref.namelist()
            nessus_files = [f for f in file_list if f.endswith('.csv') and not f.endswith('/')]
            
            for file_name in nessus_files:
                try:
                    base_name = os.path.basename(file_name)
                    first_part = base_name.split('_')[0]
                    

                    device_types = ["Server", "Switch", "Router", 'Firewall', 'Workstations', 'Access Points', 'CCTV']
                    found_device_type = None

                    for device_type in device_types:
                        if device_type in first_part or device_type.lower() in first_part.lower():
                            found_device_type = device_type
                            break

                    if found_device_type:
                        cleaned_name = first_part.replace(found_device_type, "").replace(found_device_type.lower(), "")
                        cleaned_name = cleaned_name.strip('_').strip()
                        
                        if cleaned_name:
                            branch_name = f"{cleaned_name} {found_device_type}"
                        else:
                            branch_name = found_device_type
                    else:
                        first_word = first_part.split(' ')[0]
                        branch_name = f"{first_word} Branch"

                    with zip_ref.open(file_name) as f:
                        csv_data = io.BytesIO(f.read())
                    
                    df = pd.read_csv(csv_data, 
                                   on_bad_lines='skip', 
                                   encoding='utf-8',
                                   keep_default_na=False,
                                   na_values=[])
                    
                    df['Branch Name'] = branch_name
                    all_nessus_data.append(df)
                
                except Exception as e:
                    print(f"Error processing Nessus file {file_name}: {e}")
                    continue

        return all_nessus_data
    
    return None