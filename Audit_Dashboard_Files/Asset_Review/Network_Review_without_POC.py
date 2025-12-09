import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

def create_network_review_excel(form_data=None):
    """
    Create Network Review Excel file with specified format and data
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Network Review"
    
    # Define the questions list
    questions = [
        "Whether the Secondary Connectivity is available or not?",
        "Whether the Core Switch is manageable or not?",
        "Is HO router in HA (High Availability) mode?",
        "Is the IP pool decided or not?",
        "Is the IP range for different branches the same or not?",
        "Is ATM connected to the branch switch or not?",
        "Are CBS and Internet working on the same system or not?",
        "Is a firewall available?",
        "Are network data monitoring tools available?",
        "Whether Secondary Connectivity is on auto switchable mode or not?",
        "Is WIFI available or not?",
        "How is the Data transferred from CTS infrastructure to CBS infrastructure?",
        "Is there a central log collector?",
        "Is there adequate redundancy for Power Failure? Do critical appliances have dual power cords?",
        "Is there an adequate cooling mechanism available?",
        "Is there any single point of failure?",
        "Whether a firewall is in HA mode or not?",
        "Is VLAN implemented on the networks?",
        "How is the connectivity placed in rural area branches?",
        "Do they use VPNs or connect via direct internet?",
        "Is the server farm on a different VLAN?",
        "Is the ATM network isolated?",
        "Are unused ports in the switch open and in use?",
        "Are Network IDS/IPS in place?",
        "Is there an NTP server configured?",
        "Is Branch to HO ping allowed?",
        "Is Branch to Branch ping allowed?",
        "Are all cables well tagged and organized?",
        "Is there a mechanism to monitor the availability status of important network devices/servers?",
        "Is there a risk of water entering the data center by any means (e.g., AC Vents)?",
        "Is Network Segmentation done?",
        "Is DHCP enabled in the bank network?",
        "Is Network Diagram available?"
    ]
    
    # Define Risk Factor data for E2-E34
    risk_factors = [
        "Critical", "High", "High", "High", "High", "High", "High", "High", "High", "High",
        "High", "High", "High", "High", "High", "High", "High", "High", "High", "High",
        "High", "High", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium",
        "Low"
    ]
    
    # Define comprehensive response data for all 33 questions
    question_responses = {
        1: {
            "compliance": {
                "a": "Compliance",
                "b": "Secondary connectivity was available.",
                "d": "It was observed that the bank had active Primary and Secondary connectivity in place for CBS architecture. In case the primary link went down, services continued seamlessly via secondary connectivity, ensuring uninterrupted banking operations.",
                "e": "Since secondary connectivity was available, there was no negative impact. The bank ensured availability and business continuity under CIA triad principles.",
                "f": "Secondary connectivity was already implemented, so no further recommendation."
            },
            "non_compliance": {
                "a": "Non-Compliance",
                "b": "Secondary connectivity was not available.",
                "d": "In some branches only one connectivity was available for CBS architecture. If that connectivity went down, branch services were interrupted.",
                "e": "In case of compromise of primary connectivity, the bank faced availability issues, and services were interrupted.",
                "f": "It is recommended to have active Primary and Secondary connectivity for CBS architecture."
            },
            "not_applicable": {
                "a": "Not Applicable",
                "b": "Not Applicable",
                "d": "Not Applicable",
                "e": "Not Applicable",
                "f": "Not Applicable"
            }
        },
        2: {
            "compliance": {
                "a": "Compliance",
                "b": "Core switch was manageable.",
                "d": "It was observed that the bank's core switch was configured with proper management access controls. Authorized personnel could manage, configure, and monitor the device securely.",
                "e": "As the core switch was manageable, the bank ensured smooth network administration, monitoring, and secure configuration management.",
                "f": "Since the core switch was manageable with proper controls, no additional recommendation is required."
            },
            "non_compliance": {
                "a": "Non-Compliance",
                "b": "Core switch was not manageable.",
                "d": "It was observed that the core switch was not manageable due to missing or improper configuration, making monitoring and administration difficult.",
                "e": "If the core switch is not manageable, the bank risks losing visibility, control, and timely troubleshooting, leading to potential downtime or misconfigurations.",
                "f": "It is recommended to configure the core switch with secure management access for effective monitoring and administration."
            },
            "not_applicable": {
                "a": "Not Applicable",
                "b": "Not Applicable",
                "d": "Not Applicable",
                "e": "Not Applicable",
                "f": "Not Applicable"
            }
        },
        3: {
            "compliance": {
                "a": "Compliance",
                "b": "HO router was in HA mode.",
                "d": "It was observed that the HO router was configured in High Availability (HA) mode, ensuring failover and redundancy in case one router failed.",
                "e": "Since HA mode was implemented, the bank ensured high availability, resilience, and minimized risk of downtime.",
                "f": "No additional recommendation required as HA mode was already in place."
            },
            "non_compliance": {
                "a": "Non-Compliance",
                "b": "HO router was not in HA mode.",
                "d": "It was observed that the HO router was operating without HA configuration, meaning router failure could cause service interruption.",
                "e": "Without HA, a router failure could lead to complete service disruption at HO, affecting CBS operations.",
                "f": "It is recommended to configure HO router in High Availability mode for redundancy and business continuity."
            },
            "not_applicable": {
                "a": "Not Applicable",
                "b": "Not Applicable",
                "d": "Not Applicable",
                "e": "Not Applicable",
                "f": "Not Applicable"
            }
        },
        4: {
            "compliance": {
                "a": "Compliance",
                "b": "IP pool was decided.",
                "d": "It was observed that the IP pool was properly decided, documented, and allocated for branch, ATM, and CBS systems, ensuring structured network management.",
                "e": "A decided IP pool ensured structured IP management, reduced conflicts, and supported smooth network operations.",
                "f": "No recommendation required as the IP pool was already decided and implemented."
            },
            "non_compliance": {
                "a": "Non-Compliance",
                "b": "IP pool was not decided.",
                "d": "It was observed that the IP pool was not clearly defined, leading to inconsistent allocation of IPs across devices and branches.",
                "e": "Without a defined IP pool, IP conflicts and mismanagement could occur, leading to connectivity and service issues.",
                "f": "It is recommended to define and document a structured IP pool for proper allocation and management."
            },
            "not_applicable": {
                "a": "Not Applicable",
                "b": "Not Applicable",
                "d": "Not Applicable",
                "e": "Not Applicable",
                "f": "Not Applicable"
            }
        },
        5: {
            "compliance": {
                "a": "Compliance",
                "b": "Different IP ranges were used for branches.",
                "d": "It was observed that the bank had allocated different IP ranges for different branches, ensuring proper segregation and management of the network.",
                "e": "Different IP ranges for branches ensured structured addressing, minimized conflicts, and supported efficient troubleshooting.",
                "f": "No recommendation required as separate IP ranges were already in place."
            },
            "non_compliance": {
                "a": "Non-Compliance",
                "b": "Same IP ranges were used for branches.",
                "d": "It was observed that some branches were using the same IP ranges, which could lead to conflicts and difficulties in network management.",
                "e": "Using the same IP ranges across branches could cause IP conflicts, routing errors, and increased downtime during troubleshooting.",
                "f": "It is recommended to assign unique IP ranges to each branch to avoid conflicts and ensure smooth operations."
            },
            "not_applicable": {
                "a": "Not Applicable",
                "b": "Not Applicable",
                "d": "Not Applicable",
                "e": "Not Applicable",
                "f": "Not Applicable"
            }
        }
    }
    
    # Add remaining questions (6-32) - I'll add them in chunks to keep the response manageable
    # Questions 6-10
    question_responses.update({
        6: {
            "compliance": {
                "a": "Compliance",
                "b": "ATM was connected to the branch switch.",
                "d": "It was observed that the branch ATMs were connected directly to the branch switch through secure VLAN configuration, ensuring proper network segregation and monitoring.",
                "e": "ATM connectivity via the branch switch ensured better monitoring, traffic segregation, and enhanced security.",
                "f": "No additional recommendation required as ATMs were connected securely through the branch switch."
            },
            "non_compliance": {
                "a": "Non-Compliance",
                "b": "ATM was not connected to the branch switch.",
                "d": "It was observed that ATMs were connected through alternate or unsecured methods, bypassing the branch switch, leading to weak control and monitoring gaps.",
                "e": "If ATMs are not connected through the branch switch, it increases risks of misrouting, data leakage, and unauthorized access.",
                "f": "It is recommended to connect ATMs through the branch switch with VLAN segregation to ensure secure monitoring and control."
            },
            "not_applicable": {
                "a": "Not Applicable",
                "b": "Not Applicable",
                "d": "Not Applicable",
                "e": "Not Applicable",
                "f": "Not Applicable"
            }
        },
        7: {
            "compliance": {
                "a": "Compliance",
                "b": "CBS and Internet were on different systems.",
                "d": "It was observed that CBS applications and Internet browsing were separated at the system level. Dedicated machines were used for CBS, and Internet was restricted to specific terminals.",
                "e": "Separate systems reduced the risk of malware propagation and safeguarded CBS from Internet-related vulnerabilities.",
                "f": "No recommendation required since CBS and Internet systems were already separated."
            },
            "non_compliance": {
                "a": "Non-Compliance",
                "b": "CBS and Internet were on the same system.",
                "d": "It was observed that CBS operations and Internet access were available on the same machine, increasing exposure to threats such as malware, phishing, and data leakage.",
                "e": "An attacker can create a backdoor by downloading a virus, spyware, keylogger, malware, etc. while the internet is running. Thus it will compromise the CBS application and that data could be moved to the attacker through the backdoor compromising the CIA triad. If the system is compromised and becomes vulnerable, the whole Network of the branch and CBS Application can be infected. Cyber attacks can cause substantial financial losses for the customer as well as the banks through false transactions. Hackers may begin targeting specific customers that are part of the organization. This may result in customer identity theft.",
                "f": "It is recommended to segregate CBS and Internet operations across different systems to ensure security."
            },
            "not_applicable": {
                "a": "Not Applicable",
                "b": "Not Applicable",
                "d": "Not Applicable",
                "e": "Not Applicable",
                "f": "Not Applicable"
            }
        },
        8: {
            "compliance": {
                "a": "Compliance",
                "b": "Firewall was available.",
                "d": "It was observed that the bank had deployed a firewall at the perimeter level, configured with rules and monitoring features to secure internal and external communications.",
                "e": "With a firewall, the bank ensured controlled network traffic, protection from external threats, and compliance with security standards.",
                "f": "No recommendation required as firewall was already implemented."
            },
            "non_compliance": {
                "a": "Non-Compliance",
                "b": "Firewall was not available.",
                "d": "It was observed that no firewall was in place, leaving the internal network exposed to unauthorized access and external attacks.",
                "e": "Because the firewall is not present in the bank’s network infrastructure, it is not possible to restrict the internet and Port Blocking which provides unnecessary services. Everyone can gain access to the bank's network, and there won't be a way to monitor potential threats and untrustworthy traffic. An attacker can steal your data, leak it to the public, encrypt it and hold it for ransom, or simply delete it. Without adequate protection, malicious criminals can effectively shut your business down.",
                "f": "It is recommended to deploy a firewall with proper configuration and monitoring to secure bank's network."
            },
            "not_applicable": {
                "a": "Not Applicable",
                "b": "Not Applicable",
                "d": "Not Applicable",
                "e": "Not Applicable",
                "f": "Not Applicable"
            }
        },
        9: {
            "compliance": {
                "a": "Compliance",
                "b": "Network monitoring tools were available.",
                "d": "It was observed that the bank had implemented network data monitoring tools (like SIEM/NMS) to track logs, monitor traffic, and detect anomalies in real-time.",
                "e": "Availability of monitoring tools allowed quick detection of threats, improved response time, and enhanced network visibility.",
                "f": "No recommendation required as network monitoring tools were already implemented."
            },
            "non_compliance": {
                "a": "Non-Compliance",
                "b": "Network monitoring tools were not available.",
                "d": "It was observed that no network monitoring tools were deployed, leading to lack of visibility over network activities and potential delays in detecting incidents.",
                "e": "Without monitoring tools, suspicious activities may go undetected, increasing risks of breaches and downtime.",
                "f": "It is recommended to implement and regularly use network data monitoring tools for visibility and proactive detection of threats."
            },
            "not_applicable": {
                "a": "Not Applicable",
                "b": "Not Applicable",
                "d": "Not Applicable",
                "e": "Not Applicable",
                "f": "Not Applicable"
            }
        },
        10: {
            "compliance": {
                "a": "Compliance",
                "b": "Auto-switching was enabled for secondary connectivity.",
                "d": "It was observed that the secondary connectivity was configured in auto-switch mode, ensuring immediate failover without manual intervention during primary link failure.",
                "e": "With auto-switching, the bank achieved seamless continuity of CBS operations during link failures.",
                "f": "No recommendation required as auto-switching was already enabled."
            },
            "non_compliance": {
                "a": "Non-Compliance",
                "b": "Auto-switching was not enabled.",
                "d": "It was observed that secondary connectivity required manual switching, causing delays in service restoration during primary link failure.",
                "e": "Without auto-switching, branches may face downtime and service interruptions until manual intervention occurs.",
                "f": "It is recommended to enable auto-switching between primary and secondary connectivity for uninterrupted operations."
            },
            "not_applicable": {
                "a": "Not Applicable",
                "b": "Not Applicable",
                "d": "Not Applicable",
                "e": "Not Applicable",
                "f": "Not Applicable"
            }
        }
    })
    
    # Add remaining questions (11-32) - I'll create a helper function to add them efficiently
    def add_remaining_questions():
        remaining_questions = {
            11: {
                "compliance": {
                    "a": "Compliance",
                    "b": "Wi-Fi was not available.",
                    "d": "It was observed that Wi-Fi was disabled in the branch network, preventing unauthorized wireless access and ensuring the internal network remained secure.",
                    "e": "Disabling Wi-Fi ensured network security and minimized the risk of unauthorized access.",
                    "f": "No recommendation required as Wi-Fi was already disabled."
                },
                "non_compliance": {
                    "a": "Non-Compliance",
                    "b": "Wi-Fi was available in the bank infrastructure.",
                    "d": "It was observed that Wi-Fi was available in the branch, allowing potential unauthorized devices to connect and exploit the network, leading to risks like port scanning, DoS attacks, and malware propagation.",
                    "e": "Active Wi-Fi increased the risk of cyberattacks and exposure of branch systems to external threats.",
                    "f": "It is recommended to disable Wi-Fi in branch networks as per NABARD guidelines."
                },
                "not_applicable": {
                    "a": "Not Applicable",
                    "b": "Not Applicable",
                    "d": "Not Applicable",
                    "e": "Not Applicable",
                    "f": "Not Applicable"
                }
            },
            12: {
                "compliance": {
                    "a": "Compliance",
                    "b": "Data transfer used trusted media.",
                    "d": "It was observed that data transfer from CTS to CBS infrastructure was performed using trusted USB drives or secure channels, minimizing risks of malware infection.",
                    "e": "Trusted media ensured data integrity and security between CTS and CBS systems.",
                    "f": "No recommendation required as trusted media was already in use."
                },
                "non_compliance": {
                    "a": "Non-Compliance",
                    "b": "Data was transferred through untrusted Pen drives.",
                    "d": "It was observed that data was moved using untrusted USB drives, which could carry malware or malicious code, compromising CBS systems.",
                    "e": "Untrusted drives could infect systems, causing data breaches, malware attacks, and operational disruption.",
                    "f": "It is recommended to use trusted USB drives or secure channels for data transfer."
                },
                "not_applicable": {
                    "a": "Not Applicable",
                    "b": "Not Applicable",
                    "d": "Not Applicable",
                    "e": "Not Applicable",
                    "f": "Not Applicable"
                }
            },
            13: {
                "compliance": {
                    "a": "Compliance",
                    "b": "Central log collector was available.",
                    "d": "It was observed that a central log collector was implemented to gather logs from all critical systems, enabling proper monitoring and forensic investigation if required.",
                    "e": "Central log collection ensured traceability, accountability, and support for cyber investigations.",
                    "f": "No recommendation required as central log collector was already implemented."
                },
                "non_compliance": {
                    "a": "Non-Compliance",
                    "b": "Central log collector was not available.",
                    "d": "It was observed that no central log collector existed, making it difficult to track malicious activities and hindering cyber forensic investigations.",
                    "e": "Absence of central logging increased risk of undetected fraud, attacks, and delayed incident response.",
                    "f": "It is recommended to implement a central log collector for all critical systems."
                },
                "not_applicable": {
                    "a": "Not Applicable",
                    "b": "Not Applicable",
                    "d": "Not Applicable",
                    "e": "Not Applicable",
                    "f": "Not Applicable"
                }
            },
            14: {
                "compliance": {
                    "a": "Compliance",
                    "b": "Adequate power redundancy was available.",
                    "d": "It was observed that critical appliances had dual power cords and UPS/generator backups, ensuring continuity during power failures.",
                    "e": "Power redundancy ensured uninterrupted operations even during power failures.",
                    "f": "No recommendation required as power redundancy was already implemented."
                },
                "non_compliance": {
                    "a": "Non-Compliance",
                    "b": "Secondary power connectivity was not available; critical appliances did not have dual power cords.",
                    "d": "Critical appliances lacked backup power sources, causing interruptions in banking operations during power outages.",
                    "e": "Lack of power redundancy led to operational disruptions and potential service unavailability.",
                    "f": "It is recommended to implement UPS/generator backups and dual power cords for critical appliances."
                },
                "not_applicable": {
                    "a": "Not Applicable",
                    "b": "Not Applicable",
                    "d": "Not Applicable",
                    "e": "Not Applicable",
                    "f": "Not Applicable"
                }
            },
            15: {
                "compliance": {
                    "a": "Compliance",
                    "b": "Adequate cooling mechanism was available.",
                    "d": "It was observed that server rooms had proper AC and cooling mechanisms to maintain safe operating temperatures.",
                    "e": "Adequate cooling ensured longevity and reliable performance of critical equipment.",
                    "f": "No recommendation required as adequate cooling was already implemented."
                },
                "non_compliance": {
                    "a": "Non-Compliance",
                    "b": "Proper cooling mechanism was not available.",
                    "d": "ACs and central cooling were absent, risking overheating and potential damage to servers and networking equipment.",
                    "e": "Lack of cooling could cause hardware failure, downtime, and data loss.",
                    "f": "It is recommended to install proper cooling mechanisms in server and critical equipment rooms."
                },
                "not_applicable": {
                    "a": "Not Applicable",
                    "b": "Not Applicable",
                    "d": "Not Applicable",
                    "e": "Not Applicable",
                    "f": "Not Applicable"
                }
            }
        }
        return remaining_questions
    
    # Add questions 11-15
    question_responses.update(add_remaining_questions())
    
    # Add questions 16-32 with similar structure (abbreviated for space)
    remaining_questions_16_32 = {
        16: {"compliance": {"a": "Compliance", "b": "No single point of failure existed.", "d": "It was observed that core switches and routers were in HA mode, ensuring redundancy and avoiding service disruption.", "e": "Redundancy ensured uninterrupted critical operations.", "f": "No recommendation required as redundancy was already implemented."}, "non_compliance": {"a": "Non-Compliance", "b": "There was a single point of failure in the network.", "d": "Core switch and router were not in HA mode, and failure could interrupt branch or HO services.", "e": "Single points of failure risked downtime, disrupting productivity and critical banking operations.", "f": "It is recommended to implement HA for core switch and router."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        17: {"compliance": {"a": "Compliance", "b": "Firewall was in HA mode.", "d": "It was observed that a backup firewall was deployed for HA, ensuring uninterrupted traffic filtering.", "e": "HA firewall ensured continuous network protection and minimized downtime.", "f": "No recommendation required as HA firewall was already implemented."}, "non_compliance": {"a": "Non-Compliance", "b": "Firewall was not in HA mode.", "d": "Without HA, firewall failure could disrupt all network traffic, allowing potential attacks on bank systems.", "e": "Lack of HA could lead to network outage and increased security risks.", "f": "It is recommended to deploy a firewall in HA mode."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        18: {"compliance": {"a": "Compliance", "b": "VLAN was implemented.", "d": "VLANs were configured to segregate different departments and server farms, improving security and network performance.", "e": "VLANs prevented lateral movement of threats and improved network efficiency.", "f": "No recommendation required as VLAN was already implemented."}, "non_compliance": {"a": "Non-Compliance", "b": "VLAN was not implemented.", "d": "All departments shared the same network, risking malware propagation and cross-department compromise.", "e": "Absence of VLAN increased risk of malware spread and unauthorized access.", "f": "It is recommended to enable VLANs for department and server segregation."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        19: {"compliance": {"a": "Compliance", "b": "Secondary connectivity was available.", "d": "Rural branches had both primary and secondary connectivity for CBS architecture, ensuring uninterrupted services.", "e": "Redundant connectivity ensured continuous services and business operations.", "f": "No recommendation required as secondary connectivity was already available."}, "non_compliance": {"a": "Non-Compliance", "b": "Secondary connectivity was not available for CBS architecture.", "d": "Rural branches relied solely on primary connectivity, leading to service interruptions during failures.", "e": "Lack of secondary connectivity compromised availability and CIA triad principles.", "f": "It is recommended to implement secondary connectivity for CBS architecture in rural branches."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        20: {"compliance": {"a": "Compliance", "b": "Internet connections were via VPN.", "d": "It was observed that all branch and HO internet traffic was routed through VPNs, ensuring encrypted and secure communication.", "e": "VPN usage secured communications and reduced exposure to cyber threats.", "f": "No recommendation required as VPNs were already in use."}, "non_compliance": {"a": "Non-Compliance", "b": "Internet connection was direct, not through VPN.", "d": "Direct internet connectivity increased risk of exposure to malware, DDoS, and spoofing attacks.", "e": "Direct connection left network vulnerable to attacks and unauthorized access.", "f": "It is recommended to route all internet traffic through VPNs for secure communication."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        21: {"compliance": {"a": "Compliance", "b": "Server farm was isolated with a dedicated VLAN.", "d": "It was observed that the server farm was segregated with a dedicated VLAN, ensuring that traffic between servers and head office systems remained secure and separate.", "e": "VLAN isolation prevents lateral attacks and improves network security.", "f": "No recommendation required as VLAN isolation was already implemented."}, "non_compliance": {"a": "Non-Compliance", "b": "Server farm was not isolated.", "d": "VLAN was not implemented for the server farm, and the same IP range as HO systems was used. This could allow DoS attacks or unauthorized access between server and HO systems.", "e": "Lack of VLAN exposes server farm to attacks, including DoS, malware propagation, and unauthorized access.", "f": "It is recommended to create VLANs for server farm segregation using manageable switches."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        22: {"compliance": {"a": "Compliance", "b": "ATM network was segregated.", "d": "ATM networks were logically isolated from branch networks using VLANs, preventing direct access between branch devices and ATMs.", "e": "Isolation ensures ATM security and reduces risk of network compromise.", "f": "No recommendation required as isolation was implemented."}, "non_compliance": {"a": "Non-Compliance", "b": "ATM was directly connected to the branch switch.", "d": "Direct connection allowed branch devices and ATMs to communicate, making network devices discoverable and vulnerable to DoS or ping-sweep attacks.", "e": "Direct connection increases risk of attacks on ATMs and branch systems.", "f": "It is recommended to segregate ATM network using VLANs on L3 switches."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        23: {"compliance": {"a": "Compliance", "b": "Unused ports were disabled.", "d": "It was observed that all unused ports in network switches were disabled, preventing unauthorized access.", "e": "Disabled unused ports enhance network security.", "f": "No recommendation required as unused ports were disabled."}, "non_compliance": {"a": "Non-Compliance", "b": "Unused ports in the switch were open and in use.", "d": "Open unused ports could be exploited by attackers to gain unauthorized access or deploy malware, risking data and systems.", "e": "Open ports increase risk of network compromise and unauthorized access.", "f": "It is recommended to disable all unused ports in network switches."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        24: {"compliance": {"a": "Compliance", "b": "Network IDS/IPS were implemented.", "d": "IDS/IPS systems were deployed to detect and prevent unusual network activities, ensuring proactive threat management.", "e": "IDS/IPS ensures network threat detection and prevention.", "f": "No recommendation required as IDS/IPS were implemented."}, "non_compliance": {"a": "Non-Compliance", "b": "Network IDS/IPS were not available.", "d": "Without IDS/IPS, attacks like flooding, protocol abuse, or DoS could go undetected, leaving the network vulnerable.", "e": "Absence of IDS/IPS increases risk of undetected attacks and potential network disruption.", "f": "It is recommended to implement IDS/IPS to monitor and prevent suspicious network activities."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        25: {"compliance": {"a": "Compliance", "b": "NTP server was configured.", "d": "All systems were synchronized to an NTP server, ensuring accurate timestamps for logs and security events.", "e": "Accurate timestamps ensure reliable forensic analysis and operational efficiency.", "f": "No recommendation required as NTP server was implemented."}, "non_compliance": {"a": "Non-Compliance", "b": "NTP server was not configured.", "d": "Lack of NTP synchronization could lead to inaccurate logs, affecting forensic investigations and traceability during incidents.", "e": "Unsynchronized systems may compromise incident investigation and time-sensitive operations.", "f": "It is recommended to implement an NTP server for all systems."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        26: {"compliance": {"a": "Compliance", "b": "Branch to HO ping was disabled.", "d": "ICMP ping from branch to HO was blocked, preventing unauthorized network discovery.", "e": "Blocking ping ensures network security and prevents reconnaissance.", "f": "No recommendation required as ping was disabled."}, "non_compliance": {"a": "Non-Compliance", "b": "Branch to Head Office ping was allowed.", "d": "Allowing ping exposes HO devices to branch network users, risking DoS attacks or network mapping by attackers.", "e": "Allowing ping increases the chance of targeted attacks and network compromise.", "f": "It is recommended to disable branch to HO ping."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        27: {"compliance": {"a": "Compliance", "b": "Branch to branch ping was disabled.", "d": "ICMP ping between branches was blocked to prevent network mapping and DoS attacks.", "e": "Disabled ping prevents reconnaissance and attacks.", "f": "No recommendation required as ping was disabled."}, "non_compliance": {"a": "Non-Compliance", "b": "Branch to branch ping was allowed.", "d": "Ping between branches allowed attackers to discover network devices and perform DoS attacks.", "e": "Allowed ping increases risk of network disruption and DoS attacks.", "f": "It is recommended to disable branch-to-branch ping."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        28: {"compliance": {"a": "Compliance", "b": "Cables were properly tagged and organized.", "d": "Cable tagging and organization were in place, allowing easy identification and maintenance.", "e": "Organized cables reduce errors and improve maintenance efficiency.", "f": "No recommendation required as cables were properly tagged."}, "non_compliance": {"a": "Non-Compliance", "b": "Cables were not well tagged and organized.", "d": "Poor tagging made it difficult to identify cables, increasing troubleshooting time and risk of incorrect connections.", "e": "Unorganized cables increase downtime risk and maintenance complexity.", "f": "It is recommended to tag all cables in server racks and network devices."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        29: {"compliance": {"a": "Compliance", "b": "Network monitoring tool was in place.", "d": "Tools were implemented to monitor server and network device availability, allowing proactive issue resolution.", "e": "Monitoring ensures reliability, availability, and timely resolution of issues.", "f": "No recommendation required as monitoring tools were implemented."}, "non_compliance": {"a": "Non-Compliance", "b": "No mechanism to monitor network/servers was available.", "d": "Without monitoring, network failures, downtime, and service crashes could not be detected or addressed proactively.", "e": "Lack of monitoring increases downtime risk and reduces operational efficiency.", "f": "It is recommended to implement network and server monitoring tools."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        30: {"compliance": {"a": "Compliance", "b": "No risk of water entering the data center.", "d": "It was observed that water ingress risks were mitigated with proper sealing and sensors.", "e": "Mitigated risk ensures uninterrupted operations and protects hardware.", "f": "No recommendation required as water risk was mitigated."}, "non_compliance": {"a": "Non-Compliance", "b": "There are chances of water entering the data center.", "d": "Water ingress could damage critical devices, disrupt operations, and cause data loss.", "e": "Water ingress could cause critical device failure and operational disruption.", "f": "Implement water leakage sensors and warning systems."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        31: {"compliance": {"a": "Compliance", "b": "Network segmentation was implemented.", "d": "Public and internal networks were segregated, reducing the risk of unauthorized access and malware spread.", "e": "Segmentation improves security and prevents lateral movement of attacks.", "f": "No recommendation required as segmentation was implemented."}, "non_compliance": {"a": "Non-Compliance", "b": "Network segmentation was not established.", "d": "It was observed that network segmentation has not been established. All departments are connected within the same network segment, without any segregation between user systems, server infrastructure, and administrative networks.", "e": "Lack of segmentation increases attack surface and vulnerability to network compromise.", "f": "The bank should implement proper network segmentation by separating critical systems, servers, and user networks using VLANs, firewalls, and appropriate access control policies. Each department should operate within its designated network segment, with communication between segments allowed only through secured and monitored channels"}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        32: {"compliance": {"a": "Compliance", "b": "DHCP is not enabled in the bank’s network.", "d": "Dynamic Host Configuration Protocol (DHCP) is not enabled in the bank’s network; IP addresses are assigned manually.", "e": "Provides greater control over IP address allocation and reduces the risk of unauthorized devices connecting to the network.", "f": "The bank should maintain proper documentation of IP assignments and periodically review network configurations to ensure accuracy, avoid conflicts, and maintain network security."}, "non_compliance": {"a": "Non-Compliance", "b": "DHCP is enabled in the bank’s network.", "d": "It was observed that Dynamic Host Configuration Protocol (DHCP) is enabled across the bank’s internal network, allowing systems to obtain IP addresses automatically.", "e": "Uncontrolled DHCP usage can lead to unauthorized network access or rogue DHCP server attacks, where malicious devices assign incorrect IP configurations to divert or intercept network traffic. This may result in loss of network integrity, denial of service, or unauthorized access to sensitive systems and data, compromising the bank’s overall network security.", "f": "The bank should restrict DHCP services in bank network. Proper network access control mechanisms should be implemented to prevent rogue devices from connecting."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
        33: {"compliance": {"a": "Compliance", "b": "Network diagram was available.", "d": "A service-wise network diagram was maintained at HO, showing components, interconnections, and responsibilities, aiding in planning and management.", "e": "Availability of network diagram aids effective planning, troubleshooting, and compliance with Cyber Security Resilience guidelines.", "f": "No recommendation required as diagram was available."}, "non_compliance": {"a": "Non-Compliance", "b": "Network diagram was not available.", "d": "Lack of network diagram makes planning, maintenance, and incident response difficult, impacting resilience.", "e": "Absence leads to inaccurate planning, poor incident response, and operational inefficiencies.", "f": "Prepare a service-wise network diagram and maintain it at HO."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}}
    }
    
    question_responses.update(remaining_questions_16_32)
    
    # Define header row
    headers = [
        "Sr. No.",
        "Questionnaire/ Points",
        "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)",
        "Risk Factor",
        "Observation",
        "Impact",
        "Recommendation"
    ]
    
    # Define column widths
    column_widths = {
        'A': 10,
        'B': 50,
        'C': 20,
        'D': 30,
        'E': 20,
        'F': 50,
        'G': 50,
        'H': 50
    }
    
    # Set column widths
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define risk factor colors
    risk_colors = {
        'Critical': '8B0000',  # Dark red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }
    
    # Set header row (A1-H1)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        # Apply formatting to header cells
        cell.font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Add data rows (A2-H33)
    for i, question in enumerate(questions, 2):
        question_num = i - 1  # Question number (1-32)
        
        # Sr. No. (A2-A33)
        sr_cell = ws.cell(row=i, column=1, value=question_num)
        sr_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sr_cell.border = thin_border
        
        # Questionnaire/ Points (B2-B33)
        question_cell = ws.cell(row=i, column=2, value=question)
        question_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        question_cell.border = thin_border
        
        # Get user input for this question (if form_data is provided)
        user_input = None
        if form_data:
            # Map form field names to question numbers
            field_mapping = {
                1: 'secondaryConnectivity',
                2: 'coreSwitchManageable',
                3: 'hoRouterHA',
                4: 'ipPoolDecided',
                5: 'ipRangeSame',
                6: 'atmConnected',
                7: 'cbsInternetSame',
                8: 'firewallAvailable',
                9: 'networkMonitoringTools',
                10: 'secondaryConnectivityAuto',
                11: 'wifiAvailable',
                12: 'dataTransferMethod',
                13: 'centralLogCollector',
                14: 'powerRedundancy',
                15: 'coolingMechanism',
                16: 'singlePointFailure',
                17: 'firewallHA',
                18: 'vlanImplemented',
                19: 'ruralConnectivity',
                20: 'vpnOrDirect',
                21: 'serverFarmVlan',
                22: 'atmNetworkIsolated',
                23: 'unusedPortsOpen',
                24: 'networkIdsIps',
                25: 'ntpServerConfigured',
                26: 'branchToHoPing',
                27: 'branchToBranchPing',
                28: 'cablesTagged',
                29: 'deviceMonitoring',
                30: 'waterRisk',
                31: 'networkSegmentation',
                32: 'dhcpEnabled',
                33: 'networkDiagram'
            }
            
            field_name = field_mapping.get(question_num)
            if field_name and field_name in form_data:
                user_input = form_data[field_name]
        
        # Populate cells based on user input
        if user_input and question_num in question_responses:
            # Get the response data for this question
            response_data = question_responses[question_num]
            
            # Determine which response set to use based on user input
            if user_input.lower() == 'compliance':
                response_set = response_data['compliance']
            elif user_input.lower() == 'non-compliance':
                response_set = response_data['non_compliance']
            else:  # Not Applicable or any other value
                response_set = response_data['not_applicable']
            
            # Populate columns C, D, F, G, H with response data
            for col_letter, value in response_set.items():
                col_num = ord(col_letter) - ord('a') + 3  # Convert 'a' to 3, 'b' to 4, etc.
                cell = ws.cell(row=i, column=col_num, value=value)
                
                # Apply different alignment based on column
                # 'a' = Column C (center), 'b' = Column D (middle), 'd' = Column F (middle), 'e' = Column G (middle), 'f' = Column H (middle)
                if col_letter in ['b', 'd', 'e', 'f']:  # Columns D, F, G, H - middle align only
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:  # Column C - center align
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                cell.border = thin_border
        else:
            # If no form data or question not found, populate with empty cells
            for col in range(3, 9):  # Columns C to H
                empty_cell = ws.cell(row=i, column=col, value="")
                
                # Apply different alignment based on column
                if col in [4, 6, 7, 8]:  # Columns D, F, G, H - middle align only
                    empty_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:  # Column C - center align
                    empty_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                empty_cell.border = thin_border
        
        # Risk Factor (E2-E34) with color coding
        risk_factor = risk_factors[i-2]  # Get risk factor for this row
        risk_cell = ws.cell(row=i, column=5, value=risk_factor)
        risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
        risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
        risk_cell.border = thin_border
    
    # Set row height for better appearance with wrap text
    for row in range(1, 34):
        ws.row_dimensions[row].height = 30  # Increased height to accommodate wrapped text
    
    # Create output directory if it doesn't exist
    output_dir = "static/uploads"
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate filename
    filename = "Network Review.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # Save the workbook
    wb.save(filepath)
    
    return filepath, filename

def cleanup_file(filepath):
    """
    Delete the file from the uploads folder after download
    """
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            print(f"File cleaned up: {filepath}")
    except Exception as e:
        print(f"Error cleaning up file {filepath}: {str(e)}")

if __name__ == "__main__":
    # Test the function
    filepath, filename = create_network_review_excel()
    print(f"Excel file created: {filepath}")
    print(f"Filename: {filename}")
