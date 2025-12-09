import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_others_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Others"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # Others Questions
    questions = [
        "Whether the bank performs due diligence of vendors before on-boarding.",
        "Whether the bank has signed agreements with all vendors.",
        "Whether the bank has a fraud prevention mechanism.",
        "How is data encryption and access control managed for databases.",
        "How is data encryption and access control managed for data in transit.",
        "Whether the bank has an IS (Information Security) policy.",
        "Whether the bank has a Cyber Security policy.",
        "What criteria are used to evaluate and select software products and vendors.",
        "Is there a documented process for product and vendor selection.",
        "Do the contracts include all necessary provisions/clauses.",
        "Whether the bank has an outsourcing policy.",
        "What security standards or frameworks are in place in the bank (e.g., ISO 27001, NIST).",
        "Whether the bank has an IT policy.",
        "Whether Board members/Administrator/CEO of the bank have attended any programme on Cyber Security.",
        "Whether the bank has a User Awareness policy.",
        "Are employees and customers educated about security best practices.",
        "Whether awareness programmes are conducted for all staff on cyber security and information security."
    ]

    # Risk Factors
    risk_factors = [
        "Critical", "Critical", "Critical", "Critical", "Critical", "Critical", "Critical",
        "High", "High", "High", "High", "High", "High", "High",
        "Medium", "Medium", "Medium"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "vendorDueDiligence": 1,
        "vendorAgreements": 2,
        "fraudPrevention": 3,
        "databaseEncryption": 4,
        "dataTransitEncryption": 5,
        "isPolicy": 6,
        "cyberSecurityPolicy": 7,
        "vendorSelectionCriteria": 8,
        "vendorSelectionProcess": 9,
        "contractProvisions": 10,
        "outsourcingPolicy": 11,
        "securityStandards": 12,
        "itPolicy": 13,
        "boardCyberTraining": 14,
        "userAwarenessPolicy": 15,
        "securityEducation": 16,
        "awarenessPrograms": 17
    }

    # Question responses data - First 8 questions (46.1 to 46.8)
    question_responses = {
        1: {  # vendorDueDiligence
            'compliance': {'a': 'Compliance', 'b': 'Vendor due diligence performed.', 'd': 'The bank assesses vendors\' financial stability, compliance, technical capability, and security posture prior to onboarding.', 'f': 'Reduces the risk of vendor-related operational and security incidents.', 'h': 'Periodically review due diligence criteria and update based on emerging risks.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The bank does not conduct due diligence of vendors before onboarding.', 'd': 'The bank does not consistently conduct background checks, financial stability assessments, or security evaluations before engaging vendors.', 'f': 'Lack of due diligence may expose the bank to potential risks related to the vendor\'s security, compliance, and performance capabilities.', 'h': '''To address this weakness, the bank should consider implementing the following recommendations:

Establish Vendor Due Diligence Procedures: Create a formal process for evaluating and verifying the qualifications, reputation, security, and compliance of potential vendors before onboarding.

Document Vendor Selection Criteria: Clearly define the criteria for selecting vendors, such as their financial stability, track record, and ability to meet the bank's specific needs. This will help ensure consistent and well-informed vendor choices.'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # vendorAgreements
            'compliance': {'a': 'Compliance', 'b': 'Agreements signed with all vendors.', 'd': 'Formal contracts exist for every vendor engagement, covering scope, SLAs, confidentiality, compliance, and security responsibilities.', 'f': 'Minimizes legal disputes and ensures accountability of vendor actions.', 'h': 'Review vendor contracts periodically and ensure they reflect updated policies and regulatory requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The bank has not signed agreement with all vendors.', 'd': 'Some vendors operate without signed contracts, leading to unclear responsibilities and potential disputes.', 'f': 'If a bank has not signed agreements with all its vendors, it exposes itself to a significant security and compliance vulnerability. This can lead to unauthorized access, data breaches, and regulatory penalties, posing a substantial financial and reputational risk to the bank.', 'h': 'The bank shall sign formal agreements with all vendors and agreement shall cover all necessary provision/clauses related to respective service. This helps in minimizing legal and operational risks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # fraudPrevention
            'compliance': {'a': 'Compliance', 'b': 'Fraud prevention mechanism implemented.', 'd': 'Bank has implemented controls such as monitoring, anomaly detection, alerts, and audit mechanisms to detect and prevent fraud.', 'f': 'Reduces likelihood of fraud and strengthens internal control environment.', 'h': 'Continuously enhance fraud detection mechanisms using analytics and periodic review of controls.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The bank does not have an effective fraud prevention mechanism in place.', 'd': 'No formal controls, monitoring, or alerts exist to detect suspicious or fraudulent activities by vendors or internal personnel.', 'f': 'The absence of a fraud prevention mechanism may expose the bank to a higher risk of financial losses and reputational damage due to fraudulent activities.', 'h': "The bank should establish a robust fraud prevention mechanism that includes preventive measures, early detection, and response protocols. This would help mitigate fraud-related risks and enhance the bank's overall security. Communicating these measures to staff and customers is also crucial for raising awareness and vigilance against potential fraud."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # databaseEncryption
            'compliance': {'a': 'Compliance', 'b': 'Database encryption and access control implemented.', 'd': 'Databases use strong encryption (e.g., AES-256), and access is granted based on roles with least privilege principles.', 'f': 'Ensures confidentiality, integrity, and compliance with data protection regulations.', 'h': 'Periodically review access rights and encryption mechanisms to maintain security posture.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The bank does not have data encryption mechanisms or sufficient access controls in place for its database.', 'd': 'Sensitive data stored in databases is not fully encrypted, and user access is not restricted based on roles.', 'f': 'Without effective data encryption and access control measures, the bank\'s sensitive information is vulnerable to unauthorized access and potential data breaches, putting customer trust at risk and potentially leading to non-compliance with data protection regulations.', 'h': "The bank should implement robust data encryption and access control mechanisms for its database, ensuring that only authorized personnel can access and modify data. This should be accompanied by regular security audits and staff training to maintain a secure data environment."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # dataTransitEncryption
            'compliance': {'a': 'Compliance', 'b': 'Data in transit encrypted and access-controlled.', 'd': 'Bank uses TLS/SSL encryption and secure authentication methods to protect data in transit between clients, servers, and third-party services.', 'f': 'Ensures confidentiality and integrity of transmitted data.', 'h': 'Regularly test encryption mechanisms and review configurations to address new vulnerabilities.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Data is not encrypted in transit, and there are no sufficient controls to protect data during transmission.', 'd': 'Sensitive information transmitted over networks is unencrypted or lacks strict access controls.', 'f': 'Unencrypted data transmission exposes sensitive information to interception or manipulation by unauthorized parties during transit. This can lead to data breaches, financial loss, and non-compliance with regulatory security requirements.', 'h': 'The bank should establish and document a robust data encryption and access control framework for data in transit. This framework should align with best practices and regulatory requirements to ensure the confidentiality and integrity of sensitive data during transmission.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # isPolicy
            'compliance': {'a': 'Compliance', 'b': 'Information Security policy exists.', 'd': 'A comprehensive IS policy is implemented covering network security, user access, data protection, incident response, and compliance.', 'f': 'Provides a strong security governance framework and regulatory compliance.', 'h': 'Review and update the IS policy periodically to incorporate new threats and regulatory requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The bank does not have an IS (Information Security) policy.', 'd': 'It was observed that the bank lacks a well-defined Information Security (IS) policy.', 'f': 'The absence of a clear IS policy can result in inadequate protection of sensitive data, increased cybersecurity risks, and non-compliance with industry standards and regulations.', 'h': 'It is important for the bank to establish a robust IS policy that addresses data protection, cybersecurity measures, access controls, and compliance requirements. This policy should be communicated to all staff and regularly updated to adapt to evolving threats.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # cyberSecurityPolicy
            'compliance': {'a': 'Compliance', 'b': 'Cyber Security policy implemented.', 'd': 'Bank maintains a robust Cyber Security policy that covers preventive, detective, and corrective controls against cyber threats.', 'f': 'Enhances resilience against cyber attacks and ensures compliance with regulatory requirements.', 'h': 'Regularly review and test the Cyber Security policy for effectiveness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The bank does not have a Cyber Security policy.', 'd': "It was observed that the bank lacks cyber security policy.  The absence of a Cyber Security policy indicates a potential vulnerability in the bank's security infrastructure.", 'f': 'Without a cyber security policy, the bank may struggle to effectively mitigate cyber threats, safeguard sensitive information, and respond to security incidents.', 'h': 'It is crucial for the bank to develop and implement a comprehensive Cyber Security policy that outlines security measures, incident response procedures, and ongoing risk assessments. This policy should align with industry best practices and regulatory requirements.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # vendorSelectionCriteria
            'compliance': {'a': 'Compliance', 'b': 'Clear criteria for selection exists.', 'd': 'Vendor and software evaluation uses documented criteria including security, compliance, functionality, performance, cost, and support.', 'f': 'Ensures reliable and secure products and services are procured.', 'h': 'Periodically review and refine criteria based on evolving threats and business requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The bank does not have the criteria to evaluate and select software products and vendors.', 'd': 'Decisions on selecting vendors or software are ad-hoc without documented technical, financial, and security evaluation criteria.', 'f': 'If the bank doesn\'t have a careful way to pick software, they could rush into decisions, which would waste money and not fit with their plans. If they don\'t evaluate things consistently, they might end up choosing companies that can\'t keep their data safe, follow the rules, or work well.', 'h': "The bank should establish a formalized procedure for evaluating and selecting software products and vendors. This process should include predefined criteria aligned with the bank's strategic objectives and should involve relevant stakeholders from various departments, including IT, security, compliance, and procurement. Additionally, documentation of each evaluation and selection should be maintained for transparency and accountability."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # vendorSelectionProcess
            'compliance': {'a': 'Compliance', 'b': 'Documented process exists.', 'd': 'A formal workflow is in place to evaluate, approve, and onboard vendors and software products, with records maintained for audit purposes.', 'f': 'Ensures transparency, accountability, and alignment with regulatory and operational requirements.', 'h': 'Periodically audit the process to ensure adherence and update the workflow for new risks or business requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'There was no documented process for product and vendor selection.', 'd': 'It was observed that there was no documented process for product and vendor selection, making it difficult to ensure consistent and informed decision-making.', 'f': 'A lack of a structured process in the bank can lead to inconsistent decisions, inefficient resource allocation, and potential risks. This means they might make unreliable choices, waste resources, and face possible problems.', 'h': 'The bank needs to establish a documented procedure for choosing products and vendors in order to ensure structured and informed decision-making.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # contractProvisions
            'compliance': {'a': 'Compliance', 'b': 'Contracts include necessary provisions.', 'd': 'All vendor and service agreements contain clearly defined clauses addressing data security, confidentiality, regulatory compliance, service levels, and liability.', 'f': 'Provides legal protection and ensures accountability of vendors.', 'h': 'Periodically review contracts to ensure clauses remain relevant to evolving regulations and security requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The contracts do not include all necessary provisions/clauses.', 'd': 'It was observed some necessary provisions/clauses are missing in the agreement between bank and vendor.', 'f': 'Incomplete or inadequate contract provisions can lead to misunderstandings, disputes, and unaddressed risks, potentially exposing the bank to legal and operational challenges.', 'h': 'The bank shall sign formal agreements with all vendors and agreement shall cover all necessary provision/clauses related to respective service. This helps in minimizing legal and operational risks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # outsourcingPolicy
            'compliance': {'a': 'Compliance', 'b': 'Outsourcing policy exists.', 'd': 'A formal outsourcing policy defines processes, roles, and responsibilities, including vendor selection, monitoring, and termination procedures.', 'f': 'Ensures consistency, regulatory compliance, and effective management of outsourced services.', 'h': 'Regularly review and update the policy to address new outsourcing risks and regulatory changes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The bank lacks a formal outsourcing policy.', 'd': 'The bank lacks a formal policy governing outsourcing of IT services and critical operations.', 'f': 'Absence of an outsourcing policy may lead to unregulated third-party engagements, increasing the risk of data leakage, non-compliance with regulatory guidelines, and lack of accountability in vendor management.', 'h': 'Develop and implement an outsourcing policy outlining governance, risk assessment, contract management, monitoring, and compliance requirements.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # securityStandards
            'compliance': {'a': 'Compliance', 'b': 'Security standards/frameworks implemented.', 'd': 'The bank follows recognized frameworks such as ISO 27001, NIST, or equivalent for information and cyber security practices.', 'f': 'Strengthens the security posture, ensures regulatory compliance, and provides a structured approach to risk management.', 'h': 'Periodically review adherence to frameworks and conduct audits to ensure compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "The bank doesn't have any security standards or frameworks in place (e.g., ISO 27001, NIST).", 'd': 'The bank operates without aligning its IS/Cyber Security practices to any formal standard or framework.', 'f': 'The lack of transparency regarding security standards or frameworks raises concerns about the bank\'s commitment to robust cybersecurity practices and adherence to internationally recognized security standards.', 'h': "It is essential for the bank to establish a comprehensive IT policy that outlines guidelines, standards, and procedures for IT management, security, and compliance. This policy should address the bank's specific needs, align with industry standards, and promote efficient and secure IT practices."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # itPolicy
            'compliance': {'a': 'Compliance', 'b': 'IT policy exists.', 'd': 'A formal IT policy defines standards for hardware, software, network, and information security management, ensuring alignment with regulatory and business requirements.', 'f': 'Provides clear guidance, reduces IT risks, and ensures regulatory compliance.', 'h': 'Review and update the IT policy periodically to address evolving technology and security risks.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The bank does not have an IT policy.', 'd': 'There is no formal IT policy covering hardware, software, network, and information security practices across the bank.', 'f': 'The absence of a clear IT policy can lead to inconsistencies, security gaps, and inefficient IT management, potentially exposing the bank to various risks.', 'h': 'It is crucial to conduct/attend cybersecurity training for all relevant bank personnel, including the Board, Administrator, and CEO. This training should focus on raising awareness of cyber threats, understanding cybersecurity policies and procedures. Regular updates and ongoing education are essential to keep pace with evolving cyber risks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # boardCyberTraining
            'compliance': {'a': 'Compliance', 'b': 'Key management trained in cyber security.', 'd': 'Board members, CEO, and administrators have attended certified cyber security programs or workshops.', 'f': 'Enhances informed decision-making and strengthens cyber risk governance.', 'h': 'Conduct periodic refresher programs to maintain awareness of evolving threats and compliance requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Board members, Administrator, or CEO of the bank have not attended any programs on Cyber Security.', 'd': 'The bank has not provided information regarding the specific security standards or frameworks   (e.g., ISO 27001, NIST) in place, making it unclear what specific security measures and best practices are being followed.', 'f': 'The lack of cybersecurity training for Board members, administrators or the CEO poses a potential security risk, increased vulnerability to cyberattacks, and knowledge gap.', 'h': "It is recommended that the bank clearly define and communicate the security standards or frameworks it follows. Implementing recognized standards such as ISO 27001 or NIST can enhance the bank's security posture and demonstrate its dedication to safeguarding sensitive information and data."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # userAwarenessPolicy
            'compliance': {'a': 'Compliance', 'b': 'User awareness policy implemented.', 'd': 'The bank maintains a policy guiding staff and customers on cyber security awareness, safe IT practices, and reporting mechanisms for incidents.', 'f': 'Reduces risk of cyber incidents caused by human error and enhances overall security culture.', 'h': 'Review and update the policy periodically to reflect emerging threats and technology changes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The bank does not have a User Awareness policy.', 'd': 'No documented policy exists to guide staff and customers on safe IT practices, cyber threats, and security responsibilities.', 'f': 'The absence of comprehensive awareness programs for all staff increases the risk of security incidents and data breaches, as employees may not be adequately informed about cyber threats and how to protect sensitive information. This knowledge gap could lead to security vulnerabilities, making the bank more susceptible to cyberattacks and potential financial and reputational damage.', 'h': "It is recommended that the bank promptly initiates regular and mandatory awareness programs on cyber security and information security for all staff. These programs should cover essential topics, best practices, and guidelines to ensure a well-informed. This proactive approach will help enhance the bank's overall security posture and mitigate potential security risks."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # securityEducation
            'compliance': {'a': 'Compliance', 'b': 'Employees and customers trained.', 'd': 'Regular awareness programs, emails, workshops, and training sessions are conducted to educate staff and customers on security best practices.', 'f': 'Strengthens security culture and reduces risk of incidents due to human factors.', 'h': 'Continuously assess training effectiveness and update materials for emerging threats.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Awareness programs on cyber security and information security have not been conducted for all staff.', 'd': 'It was observed that there is a lack of awareness and education regarding security best practices among both employees and customers.', 'f': 'The lack of education on security best practices increases the risk of security incidents, data breaches, and other cyber threats. This could result in financial frauds/losses, reputational damage, and a loss of trust among customers.', 'h': "Implement comprehensive education and training programs for both employees and customers to ensure they are well-informed about security best practices. This will help mitigate security risks and enhance the bank's overall security posture."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # awarenessPrograms
            'compliance': {'a': 'Compliance', 'b': 'Awareness programs conducted.', 'd': 'The bank regularly organizes structured awareness programs covering cyber threats, secure IT practices, and incident response for all staff.', 'f': 'Enhances employee vigilance, reduces human-related security incidents, and ensures regulatory compliance.', 'h': 'Evaluate effectiveness through periodic assessments, quizzes, or simulated phishing exercises.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Awareness programs on cyber security and information security have not been conducted for all staff.', 'd': 'It was observed that awareness programs on cyber security and information security have not been conducted for all staff. There were no regular and mandatory awareness programs conducted for staff.', 'f': 'The absence of comprehensive awareness programs for all staff increases the risk of security incidents and data breaches, as employees may not be adequately informed about cyber threats and how to protect sensitive information. This knowledge gap could lead to security vulnerabilities, making the bank more susceptible to cyberattacks and potential financial and reputational damage.', 'h': "It is recommended that the bank promptly initiates regular and mandatory awareness programs on cyber security and information security for all staff. These programs should cover essential topics, best practices, and guidelines to ensure a well-informed. This proactive approach will help enhance the bank's overall security posture and mitigate potential security risks."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Others Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
