import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_linux_server_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Linux Server Logical Review"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Linux Server Logical Review Questions
    questions = [
        "Is there a separate room for the server?",
        "Is Anti-Virus installed on server?",
        "Are physical controls established for Anti-Virus Server?",
        "Is Internet allowed on servers?",
        "Is NIS Server installed?",
        "Is firewalld installed?",
        "Is SELinux policy configured?",
        "Is telnet client not installed?",
        "Is nodev option set on /tmp partition?",
        "Are updates, patches, and additional security software installed?",
        "Is SETroubleshoot installed?",
        "Is rsh client not installed?",
        "Is NIS Client not installed?",
        "Is NIS Server not installed?",
        "Is tftp server installed?",
        "Is telnet-server not installed?",
        "Is talk client not installed?",
        "Is xinetd installed?",
        "Is DHCP Server installed?",
        "Is LDAP server installed?",
        "Is IMAP and POP3 server installed?",
        "Is DNS Server installed?",
        "Is FTP Server installed?",
        "Is HTTP server installed?",
        "Is packet redirect sending disabled - files 'net.ipv4.conf.all.send_redirects = 0'?",
        "Is secure ICMP redirects not accepted - 'net.ipv4.conf.all.secure_redirects = 0'?",
        "Are suspicious packets logged - 'net.ipv4.conf.all.log_martians = 1'?",
        "Is SSH access limited?",
        "Does the server have adequate space for operational requirements?",
        "Is the Server room visible from a distance, but not easily accessible?",
        "Is the Server room away from the basement, water/drainage systems?",
        "Is /tmp configured?",
        "Is noexec option set on /tmp partition?",
        "Is SSH MaxAuthTries set to 4 or less?",
        "Does /home partition include the nodev option?",
        "Is NTP configured?",
        "Is mail transfer agent configured for local-only mode?",
        "Is rsyslog installed?",
        "Is USB Storage - modprobe disabled?",
        "Is USB Storage - lsmod disabled?",
        "Is last logged in user display disabled - disable user list?",
        "Is Samba uninstalled?",
        "Is SSH root login disabled?",
        "Is SSH Idle Timeout Interval configured - ClientAliveCountMax?",
        "Is SSH LoginGraceTime set to one minute or less?",
        "Are password creation requirements configured - dcredit?",
        "Is password expiration 365 days or less - login.defs?",
        "Is inactive password lock 30 days or less - /etc/default/useradd?",
        "Is nosuid option set on /tmp partition?",
        "Is SSH X11 forwarding disabled?",
        "Is SSH warning banner configured?",
        "Is lockout for failed password attempts configured - password-auth 'auth sufficient pam_unix.so'?",
        "Is password reuse limited?",
        "Is minimum days between password changes configured - /etc/login.defs?"
    ]

    # Risk Factors (provided by user)
    risk_factors = [
        "Critical", "Critical", "Critical", "Critical", "Critical", "Critical", "Critical", "Critical",
        "High", "High", "High", "High", "High", "High", "High", "High", "High", "High", "High", "High",
        "High", "High", "High", "High", "High", "High", "High", "High", "Medium", "Medium", "Medium",
        "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium",
        "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Low", "Low",
        "Low", "Low", "Low", "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "separateServerRoom": 1,
        "antivirusInstalled": 2,
        "physicalControlsAntivirus": 3,
        "internetAllowedServers": 4,
        "nisServerInstalled": 5,
        "firewalldInstalled": 6,
        "selinuxConfigured": 7,
        "telnetClientNotInstalled": 8,
        "nodevOptionTmp": 9,
        "updatesPatchesInstalled": 10,
        "setroubleshootInstalled": 11,
        "rshClientNotInstalled": 12,
        "nisClientNotInstalled": 13,
        "nisServerNotInstalled": 14,
        "tftpServerInstalled": 15,
        "telnetServerNotInstalled": 16,
        "talkClientNotInstalled": 17,
        "xinetdInstalled": 18,
        "dhcpServerInstalled": 19,
        "ldapServerInstalled": 20,
        "imapPop3ServerInstalled": 21,
        "dnsServerInstalled": 22,
        "ftpServerInstalled": 23,
        "httpServerInstalled": 24,
        "packetRedirectDisabled": 25,
        "secureIcmpRedirectsNotAccepted": 26,
        "suspiciousPacketsLogged": 27,
        "sshAccessLimited": 28,
        "adequateServerSpace": 29,
        "serverRoomVisible": 30,
        "serverRoomAwayFromWater": 31,
        "tmpConfigured": 32,
        "noexecOptionTmp": 33,
        "sshMaxAuthTries": 34,
        "homeNodevOption": 35,
        "ntpConfigured": 36,
        "mailTransferAgentLocalOnly": 37,
        "rsyslogInstalled": 38,
        "usbStorageModprobeDisabled": 39,
        "usbStorageLsmodDisabled": 40,
        "lastLoggedUserDisplayDisabled": 41,
        "sambaUninstalled": 42,
        "sshRootLoginDisabled": 43,
        "sshIdleTimeoutConfigured": 44,
        "sshLoginGraceTime": 45,
        "passwordCreationRequirements": 46,
        "passwordExpiration365Days": 47,
        "inactivePasswordLock30Days": 48,
        "nosuidOptionTmp": 49,
        "sshX11ForwardingDisabled": 50,
        "sshWarningBannerConfigured": 51,
        "lockoutFailedPasswordAttempts": 52,
        "passwordReuseLimited": 53,
        "minimumDaysPasswordChanges": 54
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Question responses data
    question_responses = {
        1: {  # separateServerRoom
            'compliance': {'a': 'Compliance', 'b': 'Dedicated server room available.', 'd': 'Servers are housed in a separate, access-controlled room. Only authorized personnel can enter, reducing physical access risks.', 'f': 'Protects servers from unauthorized access, theft, or accidental damage. Ensures critical infrastructure is secured and operational continuity maintained.', 'h': 'Regularly review access logs and maintain server room security protocols to ensure only authorized personnel have access.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'It was observed that there was no separate room for the server. ', 'd': 'Servers are placed in a shared area or common room without physical segregation. This allows unrestricted access to personnel.', 'f': '''1. Inadequate Temperature Control:-
When the temperature around and within the server and networking equipment becomes too high the server will shut down and there will be loss of data.
2. Imbalanced Moisture Levels:-
High humidity can result in rust, corrosion, short-circuiting, and even the growth of fungus that can attack the machinery. Too little moisture in the air is also a concern, as an exceedingly dry environment can result in electrostatic discharge, which can cause system malfunction and damage.
Also, there is a risk of dust and temperature interference.''', 'h': 'It is recommended to have a separate room for the server. When you address the specific server facility threats that most often destroy data assets, you can minimize risk dramatically.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # antivirusInstalled
            'compliance': {'a': 'Compliance', 'b': 'Antivirus installed.', 'd': 'Enterprise antivirus software is installed and actively protecting servers. Real-time monitoring ensures threats are detected promptly.', 'f': 'Reduces the risk of malware infections, data loss, and operational disruption. Enhances overall system security.', 'h': 'Periodically audit antivirus updates and scan logs to verify continuous protection.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Antivirus was not installed on the server.', 'd': 'Servers do not have antivirus software installed, leaving them vulnerable to malware, ransomware, and other threats.', 'f': 'Increased risk of compromise, data loss, and service disruption. Malware infections can spread to other systems on the network.', 'h': 'Install enterprise-grade antivirus on all servers and configure automatic updates and real-time scanning.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # physicalControlsAntivirus
            'compliance': {'a': 'Compliance', 'b': 'Physical controls implemented.', 'd': 'Antivirus server is in a locked, monitored area with access restricted to authorized personnel.', 'f': 'Protects server from unauthorized tampering and ensures antivirus remains operational.', 'h': 'Conduct periodic checks to verify physical controls are effective and access logs are reviewed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Physical controls were not established for anti-virus server.', 'd': 'There are no locks, restricted access, or surveillance for the antivirus server. Unauthorized personnel may tamper with server or antivirus settings.', 'f': 'The lack of physical controls for an antivirus server can have serious security implications. Without proper physical controls, unauthorized individuals may be able to gain access to the server, potentially compromising sensitive information and allowing malware to spread throughout the network.', 'h': 'Implement physical controls such as locked racks, restricted access, and CCTV monitoring for critical servers.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # internetAllowedServers
            'compliance': {'a': 'Compliance', 'b': 'Internet access restricted on servers.', 'd': 'Servers are isolated from the internet or have restricted outbound access through controlled firewalls.', 'f': 'Reduces exposure to external attacks and malware. Ensures servers are less susceptible to unauthorized access.', 'h': 'Regularly audit server network connections to confirm restricted and secure access.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Internet was allowed on servers without any restriction.', 'd': 'Servers are directly connected to the internet without proper segmentation or firewall restrictions.', 'f': 'Full access to the internet on the server will sometimes create a critical problem if some malicious script is downloaded on the server from the internet that will remove or encrypt all the sensitive data on the server and can directly gain access to CBS. As the Internet is not restricted, any malicious activity could be performed through the internet. Also, some social media websites can be accessed by the employee which will affect the bank\'s productivity, and using those sites unintentionally users can click on a malicious link that can download viruses, worms, or any malware that will affect all bank networks.', 'h': 'Restrict server internet access to necessary services only, preferably through secure gateways and firewalls.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # nisServerInstalled
            'compliance': {'a': 'Compliance', 'b': 'NIS Server not installed.', 'd': 'NIS server is not present, reducing potential attack vectors related to centralized authentication vulnerabilities.', 'f': 'Limits exposure to unnecessary services and enhances server security.', 'h': 'Periodically review installed services to ensure only required services are active.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The NIS server was installed on the server.', 'd': 'NIS server is active on the system even though centralized authentication is not required, increasing the attack surface.', 'f': 'The NIS service is inherently an insecure system that has been vulnerable to DOS attacks, and buffer overflows and has poor authentication for querying NIS maps.', 'h': 'Uninstall or disable unnecessary NIS server services to reduce exposure to attacks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # firewalldInstalled
            'compliance': {'a': 'Compliance', 'b': 'firewalld installed and active.', 'd': 'firewalld is installed and configured with rules to control inbound and outbound traffic.', 'f': 'Protects servers from unauthorized network access, mitigating potential attacks and network-based threats.', 'h': 'Regularly review firewall rules and update them as per changing security requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Firewall was not installed.', 'd': 'No firewall management service is installed, leaving the server without controlled traffic filtering.', 'f': "Firewall is a crucial component for managing firewall rules and filtering network traffic, enhancing the system's security. Without Firewall, the system lacks a robust firewall configuration, making it more susceptible to unauthorized access and potential cyber-attacks.", 'h': '''Changing firewall settings while connected over the network can result in being locked out of the system. It is recommended to run the following command to install Firewall and iptables:
# yum install firewall iptables'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # selinuxConfigured
            'compliance': {'a': 'Compliance', 'b': 'SELinux configured and active.', 'd': 'SELinux is in enforcing mode with policies configured to control access to system resources.', 'f': 'Limits exploitation by malicious users or processes and protects critical system files.', 'h': 'Review and update SELinux policies regularly to cover new services or applications.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'SELinux policy is not configured.', 'd': 'SELinux is disabled or running in permissive mode, providing no mandatory access control enforcement.', 'f': 'SELinux (Security-Enhanced Linux) is a critical security feature that enforces mandatory access controls, restricting the actions of processes and users. Without SELinux policy configuration, the system loses an essential layer of protection, leaving it more vulnerable to unauthorized access, privilege escalation, and potential exploits.', 'h': '''It is recommended to configure SELinux to meet or exceed the default targeted policy, which constrains daemons and system software only.  Edit the /etc/selinux/config file to set the SELINUXTYPE parameter:
SELINUXTYPE=targeted'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # telnetClientNotInstalled
            'compliance': {'a': 'Compliance', 'b': 'Telnet client not installed.', 'd': 'Telnet client is absent, preventing insecure remote login over unencrypted channels.', 'f': 'Reduces risk of credential interception and unauthorized access.', 'h': 'Periodically verify installed packages to ensure insecure clients are not present.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Telnet client was installed.', 'd': 'Telnet client is installed, allowing unencrypted remote login over the network.', 'f': 'The telnet protocol is insecure and unencrypted. The use of an unencrypted transmission medium could allow a user with access to sniff network traffic the ability to steal credentials. The ssh package provides an encrypted session and stronger security and is included in most Linux distributions.', 'h': '''Many insecure service clients are used as troubleshooting tools and in testing environments. Uninstalling them can inhibit capability to test and troubleshoot. If they are required it is advisable to remove the clients after use to prevent accidental or intentional misuse. It is recommended to run the following command to remove the telnet package:
# yum remove telnet'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # nodevOptionTmp
            'compliance': {'a': 'Compliance', 'b': 'nodev option set on /tmp.', 'd': '/tmp partition is mounted with nodev, restricting creation of device files.', 'f': 'Reduces local privilege escalation risk and prevents execution of malicious device nodes.', 'h': 'Regularly check mount options to ensure nodev remains configured.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Nodev option was not set on /tmp partition.', 'd': '/tmp partition allows device files to be created, which could be exploited by attackers to escalate privileges.', 'f': 'Without the "nodev" option, users can create and execute device files in the /tmp directory, potentially leading to security vulnerabilities. This can allow malicious actors to exploit the system by executing arbitrary code and gaining unauthorized access.', 'h': '''Since the /tmp filesystem is not intended to support devices, set this option to ensure that users cannot attempt to create block or character special devices in /tmp .
Edit the /etc/fstab file OR the /etc/systemd/system/local-fs.target.wants/tmp.mount file:
IF /etc/fstab is used to mount /tmp
Edit the /etc/fstab file and add nodev to the fourth field (mounting options) for the /tmp partition. See the fstab(5) manual page for more information.
Run the following command to remount /tmp:

# mount -o remount,nodev /tmp

OR if systemd is used to mount /tmp:
Edit /etc/systemd/system/local-fs.target.wants/tmp.mount to add nodev to the /tmp mount options:

[Mount]
Options=mode=1777,strictatime,noexec,nodev,nosuid

Run the following command to restart the systemd daemon:

#  systemctl daemon-reload

Run the following command to restart tmp.mount

# systemctl restart tmp.mount

'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # updatesPatchesInstalled
            'compliance': {'a': 'Compliance', 'b': 'Updates and patches applied.', 'd': 'OS and security software are up to date, reducing the risk of exploitation.', 'f': 'Enhances system security and minimizes potential attack vectors.', 'h': 'Schedule regular updates and monitor patch compliance for all servers.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Updates, patches, and additional security software are not installed.', 'd': 'Servers are missing critical OS updates, security patches, or additional security software.', 'f': 'The system becomes vulnerable to known security vulnerabilities and exploits, increasing the risk of data breaches and unauthorized access. The lack of timely updates leaves the system exposed to potential cyber threats, such as malware and ransomware.', 'h': '''Newer patches may contain security enhancements that would not be available through the latest full update. As a result, it is recommended that the latest software patches be used to take advantage of the latest functionality. As with any software installation, organizations need to determine if a given update meets their requirements and verify the compatibility and supportability of any additional software against the update revision that is selected.
Use your package manager to update all packages on the system according to site policy.
The following command will install all available packages

# yum update'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # setroubleshootInstalled
            'compliance': {'a': 'Compliance', 'b': 'SETroubleshoot installed.', 'd': 'SETroubleshoot is installed and actively captures SELinux alerts for administrator review.', 'f': 'Facilitates quick identification and resolution of SELinux policy violations, improving system security.', 'h': 'Regularly review SELinux alerts and logs to ensure policies are functioning correctly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Troubleshoot was installed.', 'd': 'SELinux alerts are not captured or reported due to the absence of SETroubleshoot.', 'f': 'The Troubleshoot service notifies desktop users of SELinux denials through a user-friendly interface. The service provides important information around configuration errors, unauthorized intrusions, and other potential errors. The Troubleshoot service is an unnecessary daemon to have running on a server, especially if X Windows is disabled.', 'h': '''It is recommended to run the following command to Uninstall troubleshoot:

# yum remove troubleshoot'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # rshClientNotInstalled
            'compliance': {'a': 'Compliance', 'b': 'rsh client not installed.', 'd': 'rsh client is absent, ensuring that insecure remote login protocols are not used.', 'f': 'Prevents credential interception and strengthens remote access security.', 'h': 'Periodically verify installed packages to prevent installation of insecure clients.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The rsh client was installed.', 'd': 'The rsh client is present, allowing unencrypted remote access to other systems.', 'f': 'Many insecure service clients are used as troubleshooting tools and in testing environments. Uninstalling them can inhibit capability to test and troubleshoot. If they are required it is advisable to remove the clients after use to prevent accidental or intentional misuse.', 'h': '''These legacy clients contain numerous security exposures and have been replaced with the more secure SSH package. Even if the server is removed, it is best to ensure the clients are also removed to prevent users from inadvertently attempting to use these commands and therefore exposing their credentials. Note that removing the rsh package removes the clients for rsh , rcp and rlogin . It is recommended to run the following command to remove the rsh package:

# yum remove rsh'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # nisClientNotInstalled
            'compliance': {'a': 'Compliance', 'b': 'NIS client not installed.', 'd': 'Absence of NIS client ensures no unnecessary network authentication services are active.', 'f': 'Reduces attack surface and enhances overall system security.', 'h': 'Conduct periodic service audits to ensure only required services are installed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'NIS Client was installed.', 'd': 'NIS client is active even when not required, creating potential attack vectors.', 'f': 'The NIS service is inherently an insecure system that has been vulnerable to DOS attacks, buffer overflows and has poor authentication for querying NIS maps. NIS generally has been replaced by such protocols as Lightweight Directory Access Protocol (LDAP).', 'h': '''Many insecure service clients are used as troubleshooting tools and in testing environments. Uninstalling them can inhibit capability to test and troubleshoot. If they are required it is advisable to remove the clients after use to prevent accidental or intentional misuse. It is recommended to run the following command to remove the ypbind package:

# yum remove ypbind'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # nisServerNotInstalled
            'compliance': {'a': 'Compliance', 'b': 'NIS server not installed.', 'd': 'NIS server is absent, minimizing unnecessary authentication-related vulnerabilities.', 'f': 'Limits exposure to potential attacks targeting NIS services.', 'h': 'Regularly verify installed services to ensure only essential servers are active.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'NIS Server was installed.', 'd': 'NIS server is running despite not being required, increasing potential vulnerabilities.', 'f': 'The NIS service is inherently an insecure system that has been vulnerable to DOS attacks, buffer overflows and has poor authentication for querying NIS maps. NIS generally has been replaced by such protocols as Lightweight Directory Access Protocol (LDAP).', 'h': '''It is recommended that the ypserv package be removed, and if required a more secure services be used.
Run the following command to remove ypserv:

# yum remove ypserv'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # tftpServerInstalled
            'compliance': {'a': 'Compliance', 'b': 'TFTP server not installed.', 'd': 'Absence of TFTP ensures no unsecured file transfer services are available on the server.', 'f': 'Reduces risk of unauthorized file access or data leakage.', 'h': 'Conduct regular audits to confirm only necessary services are installed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The tftp server was installed.', 'd': 'TFTP service is active, allowing unsecured file transfer without authentication.', 'f': 'It poses Security vulnerability due to its lack of authentication and encryption mechanisms, making it susceptible to unauthorized access and data interception. This could lead to potential security breaches and compromise sensitive files on the server. Moreover, the presence of TFTP increases the attack surface, providing an additional entry point for malicious actors.', 'h': '''If TFTP is required for operational support (such as the transmission of router configurations) its use must be documented with the Information System Security Officer (ISSO), restricted to only authorized personnel, and have access control rules established.
It is recommended to unistall tftp if not required using following command:

# yum remove tftp-server'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # telnetServerNotInstalled
            'compliance': {'a': 'Compliance', 'b': 'Telnet server not installed.', 'd': 'Absence of telnet-server prevents insecure remote access over unencrypted channels.', 'f': 'Strengthens remote access security and prevents potential credential theft.', 'h': 'Periodically check installed services to ensure insecure protocols remain absent.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Telnet-server was installed.', 'd': 'Telnet server allows unencrypted remote access to the system.', 'f': 'The telnet protocol is insecure and unencrypted. The use of an unencrypted transmission medium could allow a user with access to sniff network traffic the ability to steal credentials.', 'h': '''The ssh package provides an encrypted session and stronger security.

Run the following command to remove the telnet-server package:

# yum remove telnet-server'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # talkClientNotInstalled
            'compliance': {'a': 'Compliance', 'b': 'Talk client not installed.', 'd': 'Talk client is absent, eliminating potential misuse of chat services.', 'f': 'Reduces unnecessary attack vectors and internal data leakage risks.', 'h': 'Regularly review installed clients and remove unused software.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The talk client was installed.', 'd': 'Talk client is available, which could allow unmonitored communication and potential misuse.', 'f': 'The software presents a security risk as it uses unencrypted protocols for communication. Many insecure service clients are used as troubleshooting tools and in testing environments. Uninstalling them can inhibit capability to test and troubleshoot. If they are required it is advisable to remove the clients after use to prevent accidental or intentional misuse.', 'h': '''It is recommended to run the following command to remove the talk package:

# yum remove talk'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        18: {  # xinetdInstalled
            'compliance': {'a': 'Compliance', 'b': 'xinetd not installed.', 'd': 'xinetd absence ensures no unnecessary network services are running.', 'f': 'Reduces system exposure to attacks targeting unnecessary services.', 'h': 'Periodically audit server services and disable unused network daemons.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The xinetd was installed.', 'd': 'xinetd super-server is active, potentially hosting unnecessary network services.', 'f': 'Unnecessary software increases the attack surface of the system, potentially exposing it to security vulnerabilities that may otherwise be avoided. Additionally, the presence of xinetd without any associated services might confuse or complicate system administration, making it harder to manage and maintain the network infrastructure efficiently.', 'h': '''If there are no xinetd services required, it is recommended that the package be removed to reduce the attack surface are of the system. It is recommended to run the following command to remove xinetd:

# yum remove xinetd'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        19: {  # dhcpServerInstalled
            'compliance': {'a': 'Compliance', 'b': 'DHCP server not installed.', 'd': 'DHCP server absence ensures IP address allocation is controlled and authorized.', 'f': 'Prevents unauthorized network access and IP conflicts.', 'h': 'Review network services regularly to ensure only necessary services are active.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The DHCP Server was installed.', 'd': 'DHCP service is active on the server without operational requirement, potentially assigning IPs improperly.', 'f': 'Improperly configured DHCP servers can lead to IP address conflicts and disrupt network connectivity. Moreover, DHCP server failures or mismanagement can cause significant downtime and hamper the productivity of connected devices and users. Additionally, if security measures are not adequately implemented, malicious actors might exploit the DHCP infrastructure to launch attacks or gain unauthorized access to the network, compromising its overall security.', 'h': '''Unless a system is specifically set up to act as a DHCP server, it is recommended that the dhcp package be removed to reduce the potential attack surface.
It is recommended to run the following command to remove dhcp:

# yum remove dhcp'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        20: {  # ldapServerInstalled
            'compliance': {'a': 'Compliance', 'b': 'LDAP server not installed.', 'd': 'Absence of LDAP server reduces exposure to potential attacks on directory services.', 'f': 'Protects sensitive authentication data and limits attack vectors.', 'h': 'Regularly verify services to ensure only necessary servers are operational.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The LDAP server was installed.', 'd': 'LDAP service is running on the server without proper justification or security controls.', 'f': 'Having an unused LDAP server may introduce security risks, as it adds another potential point of vulnerability to the system. Unpatched or misconfigured LDAP servers could be exploited by attackers, posing a threat to the overall security of the system and network.', 'h': '''If the system will not need to act as an LDAP server, it is recommended that the software be removed to reduce the potential attack surface.
It is recommended to run the following command to remove openldap-servers:

# yum remove openldap-servers'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        21: {  # imapPop3ServerInstalled
            'compliance': {'a': 'Compliance', 'b': 'IMAP/POP3 server not installed.', 'd': 'IMAP and POP3 services are absent, preventing unrequired email services from running.', 'f': 'Reduces risk of email-related attacks and data leakage.', 'h': 'Periodically audit installed services to ensure unnecessary email protocols are not active.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The IMAP and POP3 server were installed.', 'd': 'IMAP and POP3 services are active without operational need, exposing email services to external attacks.', 'f': '''These servers increase the attack surface of the system, potentially exposing it to email-related security vulnerabilities and attacks. Malicious actors could exploit these vulnerabilities to compromise sensitive data or launch email-based attacks such as phishing or spamming. The lack of proper configuration and security measures on these servers might lead to unauthorized access, data leaks, or data loss, posing significant risks to the confidentiality and integrity of email communications.''', 'h': '''Unless POP3 and/or IMAP servers are to be provided by this system, it is recommended that the package be removed to reduce the potential attack surface.
It is recommended to run the following command to remove dovecot:

# yum remove dovecot'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        22: {  # dnsServerInstalled
            'compliance': {'a': 'Compliance', 'b': 'DNS server not installed.', 'd': 'Absence of DNS server ensures no unauthorized DNS services are active on the server.', 'f': 'Reduces attack surface and prevents DNS-related vulnerabilities.', 'h': 'Regularly verify services and remove unneeded network daemons.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'DNS Server was installed.', 'd': 'DNS service is running without proper requirement, which could be exploited for DNS-based attacks.', 'f': 'Misconfigurations or vulnerabilities in the DNS server could lead to unauthorized access and manipulation of DNS records, potentially redirecting users to malicious websites or causing service disruptions. DNS cache poisoning attacks might occur, leading to incorrect DNS resolutions and exposing users to phishing or other cyber threats.', 'h': '''Unless a system is specifically designated to act as a DNS server, it is recommended that the package be removed to reduce the potential attack surface.
It is recommended to run the following command to remove bind:

# yum remove bind'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        23: {  # ftpServerInstalled
            'compliance': {'a': 'Compliance', 'b': 'FTP server not installed.', 'd': 'Absence of FTP service ensures unencrypted file transfers are not allowed, reducing risk.', 'f': 'Prevents unauthorized access and protects sensitive data.', 'h': 'Periodically audit installed services to ensure no insecure file transfer protocols exist.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'FTP Server was installed.', 'd': 'FTP service allows unencrypted file transfers, exposing credentials and data to interception.', 'f': 'FTP does not protect the confidentiality of data or authentication credentials. It is recommended SFTP be used if file transfer is required. Unless there is a need to run the system as a FTP server (for example, to allow anonymous downloads), it is recommended that the package be removed to reduce the potential attack surface.', 'h': '''It is recommended to run the following command to remove vsftpd:

# yum remove vsftpd'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        24: {  # httpServerInstalled
            'compliance': {'a': 'Compliance', 'b': 'HTTP server not installed.', 'd': 'HTTP services are absent, eliminating unnecessary exposure to web-based attacks.', 'f': 'Reduces attack surface and prevents unauthorized access via web applications.', 'h': 'Conduct regular checks to ensure only required services are active.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'HTTP server was installed.', 'd': 'HTTP service is active without business need, potentially hosting vulnerable web applications.', 'f': 'If not properly configured, attackers may exploit weaknesses in the server\'s settings to gain unauthorized access or launch DDoS attacks. Inadequate security measures may lead to data breaches or unauthorized access to sensitive information transmitted over HTTP connections, potentially compromising user privacy and confidentiality.', 'h': '''Unless there is a need to run the system as a web server, it is recommended that the package be removed to reduce the potential attack surface.
It is recommended to run the following command to remove httpd:

# yum remove httpd'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        25: {  # packetRedirectDisabled
            'compliance': {'a': 'Compliance', 'b': 'Packet redirect sending disabled.', 'd': 'System is configured to block ICMP redirect sending, preventing manipulation of routing information.', 'f': 'Secures network routing and mitigates risk of traffic interception or redirection attacks.', 'h': 'Verify network configurations periodically to ensure ICMP redirects remain disabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "The packet redirect sending was not disabled - files 'net.ipv4.conf.all.send_redirects = 0'.", 'd': 'The system is configured to send ICMP redirects, which can be exploited for network traffic manipulation.', 'f': 'An attacker could potentially exploit this feature to manipulate the routing of network traffic and redirect it to unauthorized destinations. This could lead to man-in-the-middle attacks, where sensitive data may be intercepted or modified without the knowledge of the communicating parties.', 'h': '''It is recommended to set the following parameters in /etc/sysctl.conf or a /etc/sysctl.d/* file:

net.ipv4.conf.all.send_redirects = 0
net.ipv4.conf.default.send_redirects = 0

Run the following commands to set the active kernel parameters:

# sysctl -w net.ipv4.conf.all.send_redirects=0
# sysctl -w net.ipv4.conf.default.send_redirects=0
# sysctl -w net.ipv4.route.flush=1'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        26: {  # secureIcmpRedirectsNotAccepted
            'compliance': {'a': 'Compliance', 'b': 'Secure ICMP redirects not accepted.', 'd': 'System blocks acceptance of secure ICMP redirects, mitigating routing manipulation risks.', 'f': 'Strengthens network security and prevents potential man-in-the-middle attacks.', 'h': 'Regularly audit kernel network settings to maintain secure configuration.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "The secure ICMP redirects was accepted - 'net.ipv4.conf.all.secure_redirects = 0'.", 'd': 'Server accepts secure ICMP redirects, which could be exploited to alter routing tables maliciously.', 'f': 'It is still possible for even known gateways to be compromised. Setting net.ipv4.conf.all.secure_redirects to 0 protects the system from routing table updates by possibly compromised known gateways.', 'h': '''It is recommended to set the following parameters in /etc/sysctl.conf or a /etc/sysctl.d/* file:

net.ipv4.conf.all.secure_redirects = 0
net.ipv4.conf.default.secure_redirects = 0

Run the following commands to set the active kernel parameters:

# sysctl -w net.ipv4.conf.all.secure_redirects=0
# sysctl -w net.ipv4.conf.default.secure_redirects=0
# sysctl -w net.ipv4.route.flush=1'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        27: {  # suspiciousPacketsLogged
            'compliance': {'a': 'Compliance', 'b': 'Suspicious packets logged.', 'd': 'System logs all suspicious packets with invalid source addresses for monitoring.', 'f': 'Helps in detecting potential attacks and enhancing network security monitoring.', 'h': 'Periodically review log files and configure alerting for critical events.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Suspicious packets are not logged - 'net.ipv4.conf.all.log_martians = 1'.", 'd': 'System does not log packets with invalid source addresses, missing potential attack attempts.', 'f': 'Without logging such packets, it becomes difficult to detect and investigate potential security breaches or malicious activities on the network, hindering incident response efforts. The lack of logged information about martian packets can impede network troubleshooting, making it harder to identify and resolve network configuration issues or misbehaving devices.', 'h': '''It is recommended to set the following parameters in /etc/sysctl.conf or a /etc/sysctl.d/* file:

net.ipv4.conf.all.log_martians = 1
net.ipv4.conf.default.log_martians = 1

Run the following commands to set the active kernel parameters:

# sysctl -w net.ipv4.conf.all.log_martians=1
# sysctl -w net.ipv4.conf.default.log_martians=1
# sysctl -w net.ipv4.route.flush=1'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        28: {  # sshAccessLimited
            'compliance': {'a': 'Compliance', 'b': 'SSH access limited.', 'd': 'SSH access is restricted to specific IP addresses or subnets, with authentication controls in place.', 'f': 'Reduces risk of brute-force attacks and unauthorized server access.', 'h': 'Regularly review SSH access controls and authentication settings.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'SSH access was not limited.', 'd': 'SSH allows login from any source without restrictions, exposing servers to brute-force attacks.', 'f': "If SSH access is not limited the bank won't be able to limit which users and group can access the system via SSH.", 'h': '''It is recommended to edit the /etc/ssh/sshd_config file to set one or more of the parameter as follows:

AllowUsers <user list>

OR

Allow Groups <group list>

OR

Deny Users <user list>

OR

Deny Groups <group list>

Default Value:

None'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        29: {  # adequateServerSpace
            'compliance': {'a': 'Compliance', 'b': 'Adequate disk space available.', 'd': 'Servers have sufficient storage for applications, logs, and operational requirements.', 'f': 'Ensures smooth operation of applications and reduces risk of downtime.', 'h': 'Continuously monitor disk usage and plan for capacity expansion as needed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The server does not have adequate space for operational requirements.', 'd': 'Server storage is near capacity, risking application failures and data storage issues.', 'f': 'As the Server has no adequate space for operational requirements there is a risk of dust and temperature interference. Server hardware and related components require specific components to perform optimally, such as adequate cooling, moisture removal, and protection from excessive temperatures. Server rooms that are too hot or cold could cause hardware to malfunction leading to downtime.', 'h': 'It is recommended that the Server must have adequate space for operational requirements.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        30: {  # serverRoomVisible
            'compliance': {'a': 'Compliance', 'b': 'Server room secure and visible.', 'd': 'Server room location allows visibility for monitoring while preventing easy access by unauthorized personnel.', 'f': 'Enhances physical security and protects critical systems from unauthorized access.', 'h': 'Maintain proper access control and surveillance for continuous security.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The server room was not visible from a distance and was easily accessible.', 'd': 'Server room can be easily accessed by personnel or visitors, lacking proper physical barriers.', 'f': 'As the Server room was not visible from a distance, there exists a high chance of Insider Threat. If a server room is accessible easily to anyone, it is vulnerable to unauthorised access.', 'h': 'It is recommended to keep the Server room visible from a distance and accessible for only authorized personnel such as the IT dept.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        31: {  # serverRoomAwayFromWater
            'compliance': {'a': 'Compliance', 'b': 'Server room safely located.', 'd': 'Server room is situated away from basements and water systems, minimizing water-related risks.', 'f': 'Reduces risk of flooding or water damage to critical infrastructure.', 'h': 'Maintain periodic checks for nearby hazards and ensure preventive measures are in place.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A server room was near to the basement and water/drainage systems.', 'd': 'Server room is located close to water pipes or drainage, increasing the risk of flooding or water damage.', 'f': 'If water somehow gets into the server room, the equipment present inside will not only soak up water but also increases the risk of static electricity. This will damage all the pieces of equipment present inside.', 'h': 'It is recommended that the Server room should always be away from the basement and water/drainage systems.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        32: {  # tmpConfigured
            'compliance': {'a': 'Compliance', 'b': '/tmp configured securely.', 'd': '/tmp partition is mounted with proper security options to restrict malicious activities.', 'f': 'Enhances system security and prevents execution of unauthorized code.', 'h': 'Periodically verify /tmp mount options and adjust as needed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': '/ tmp was not configured.', 'd': '/tmp is mounted without proper security options like nodev, noexec, or nosuid.', 'f': 'Since the /tmp directory is intended to be world-writable, there is a risk of resource exhaustion if it is not bound to a separate partition.', 'h': '''It is recommended to create or update an entry for /tmp in either /etc/fstab OR in a systemd tmp.mount file:
If /etc/fstab is used: configure /etc/fstab as appropriate.
_ Example:_

tmpfs/tmptmpfs     defaults,rw,nosuid,nodev,noexec,relatime  0 0

Run the following command to remount /tmp

# mount -o remount,noexec,nodev,nosuid /tmp

OR if systemd tmp.mount file is used: run the following command to create the file /etc/systemd/system/tmp.mount if it doesn't exist:

# [ ! -f /etc/systemd/system/tmp.mount ] && cp -v /usr/lib/systemd/system/tmp.mount /etc/systemd/system/

Edit the file /etc/systemd/system/tmp.mount:

[Mount]
What=tmpfs
Where=/tmp
Type=tmpfs
Options=mode=1777,strictatime,noexec,nodev,nosuid

Run the following command to reload the systemd daemon:

# systemctl daemon-reload

Run the following command to unmask and start tmp.mount:

# systemctl --now unmask tmp.mount'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        33: {  # noexecOptionTmp
            'compliance': {'a': 'Compliance', 'b': 'noexec option set on /tmp.', 'd': '/tmp partition is mounted with noexec, preventing execution of unauthorized scripts.', 'f': 'Reduces risk of local privilege escalation attacks.', 'h': 'Review /tmp mount options regularly to ensure compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Noexec option was not set on /tmp partition.', 'd': 'Users can execute scripts from /tmp, potentially running malicious code.', 'f': 'Without the "noexec" option, users can execute binaries and scripts from the /tmp directory, increasing the risk of running malicious code. This poses a security threat as it allows potential attackers to exploit the system by executing unauthorized scripts or binaries.', 'h': '''Since the /tmp filesystem is only intended for temporary file storage, set this option to ensure that users cannot run executable binaries from /tmp.
Edit the /etc/fstab file OR the /etc/systemd/system/local-fs.target.wants/tmp.mount file:
IF /etc/fstab is used to mount /tmp
Edit the /etc/fstab file and add noexec to the fourth field (mounting options) for the /tmp partition. See the fstab(5) manual page for more information.
Run the following command to remount /tmp:

# mount -o remount,noexec /tmp

OR if systemd is used to mount /tmp:_
Edit /etc/systemd/system/local-fs.target.wants/tmp.mount to add noexec to the /tmp mount options:

[Mount]
Options=mode=1777,strictatime,noexec,nodev,nosuid

Run the following command to restart the systemd daemon:

#  systemctl daemon-reload

Run the following command to restart tmp.mount

# systemctl restart tmp.mount'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        34: {  # sshMaxAuthTries
            'compliance': {'a': 'Compliance', 'b': 'MaxAuthTries set appropriately.', 'd': 'SSH allows a maximum of 4 authentication attempts, limiting brute-force attack opportunities.', 'f': 'Reduces risk of unauthorized server access via repeated login attempts.', 'h': 'Regularly review SSH configuration to maintain secure settings.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'SSH MaxAuthTries was not set to 4 or less.', 'd': 'SSH allows more than 4 authentication attempts, increasing exposure to brute-force attacks.', 'f': 'If this parameter is not properly configured and set to a low value, such as 4 or less, it increases the vulnerability to brute-force attacks. Attackers could repeatedly attempt different login credentials until they successfully guess the correct combination, potentially gaining unauthorized access to the system. Therefore, not enforcing a lower value for SSH MaxAuthTries can significantly weaken the security posture of the system and expose it to a higher risk of unauthorized access.', 'h': '''Setting the MaxAuthTries parameter to a low number will minimize the risk of successful brute force attacks to the SSH server. While the recommended setting is 4, set the number based on site policy.
Edit the /etc/ssh/sshd_config file to set the parameter as follows:

MaxAuthTries 4

Default Value:

MaxAuthTries 6'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        35: {  # homeNodevOption
            'compliance': {'a': 'Compliance', 'b': 'nodev option set on /home.', 'd': '/home partition is mounted with nodev, preventing unauthorized device file creation.', 'f': 'Reduces risk of privilege escalation via device files.', 'h': 'Periodically verify /home mount options for continued compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The /home partition did not include the nodev option.', 'd': 'Device files can be created on /home, increasing potential exploitation by local users.', 'f': 'Without this option, users may be able to create and execute device files in the /home directory, which poses security risks. This can potentially lead to unauthorized code execution and system compromise by malicious actors.', 'h': '''Setting the MaxAuthTries parameter to a low number will minimize the risk of successful brute force attacks to the SSH server. While the recommended setting is 4, set the number based on site policy.
Edit the /etc/ssh/sshd_config file to set the parameter as follows:

MaxAuthTries 4

Default Value:

MaxAuthTries 6'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        36: {  # ntpConfigured
            'compliance': {'a': 'Compliance', 'b': 'NTP configured.', 'd': 'Servers are synchronized with a trusted NTP source, ensuring consistent system time.', 'f': 'Improves log accuracy, authentication, and audit reliability.', 'h': 'Periodically verify NTP synchronization and adjust server configuration as needed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The ntp was not configured.', 'd': 'Servers are not synchronized with a reliable time source, causing time discrepancies.', 'f': 'The absence of NTP means the system will not automatically synchronize its time with reliable time servers, leading to time drift and inconsistencies. This can cause problems in log entries, event sequencing, and coordination between network components. Additionally, the restriction "-4" further limits the use of IPv4 addresses for time synchronization, potentially hindering communication with NTP servers and exacerbating time accuracy issues.', 'h': '''If ntp is in use on the system proper configuration is vital to ensuring time synchronization is working properly.
It is recommended to add or edit restrict lines in /etc/ntp.conf to match the following:

restrict -4 default kod nomodify notrap nopeer noquery
restrict -6 default kod nomodify notrap nopeer noquery

Add or edit server or pool lines to /etc/ntp.conf as appropriate:

server <remote-server>

Add or edit the OPTIONS in /etc/sysconfig/ntpd to include '-u ntp:ntp':

OPTIONS='-u ntp:ntp'

Reload the systemd daemon:

systemctl daemon-reload

Enable and start the ntp service:

systemctl --now enable ntpd'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        37: {  # mailTransferAgentLocalOnly
            'compliance': {'a': 'Compliance', 'b': 'MTA restricted to local-only.', 'd': 'Mail transfer agent is restricted to local delivery, preventing misuse for external communications.', 'f': 'Reduces risk of spam relay and abuse of mail services.', 'h': 'Review MTA configuration periodically to ensure local-only mode remains active.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The mail transfer agent was not configured for local-only mode on the server.', 'd': 'MTA is configured to send or relay emails externally, increasing risk of misuse.', 'f': 'One significant concern is the potential for unauthorized external access to the MTA, making it susceptible to abuse by spammers or other malicious entities. This can result in the server being used as an open relay, leading to reputation damage, blacklisting, and an increased risk of email-related security incidents. Moreover, without local-only mode, the MTA may accept and relay emails from unknown sources, contributing to an overload of resources, increased network traffic, and reduced server performance.', 'h': '''It is recommended to configure the  mail transfer agent for local-only mode on the server.
Edit /etc/postfix/main.cf and add the following line to the RECEIVING MAIL section. If the line already exists, change it to look like the line below:

inet_interfaces = loopback-only

Run the following command to restart postfix:

# systemctl restart postfix
'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        38: {  # rsyslogInstalled
            'compliance': {'a': 'Compliance', 'b': 'rsyslog installed.', 'd': 'rsyslog service is active and logs critical system events for monitoring and auditing.', 'f': 'Enhances system security monitoring and assists in timely incident response.', 'h': 'Regularly review log files and ensure rsyslog is functioning correctly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The rsyslog was not installed.', 'd': 'Logging service is missing, resulting in insufficient capture of system events.', 'f': 'Without rsyslog, the server would lack a centralized mechanism for capturing and managing log information, making it challenging to monitor system activities, troubleshoot issues, and identify security breaches.', 'h': '''The security enhancements of rsyslog such as connection-oriented (i.e. TCP) transmission of logs, the option to log to database formats, and the encryption of log data en route to a central logging server) justify installing and configuring the package.

It is recommended to run the following command to install rsyslog:

# yum install rsyslog'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        39: {  # usbStorageModprobeDisabled
            'compliance': {'a': 'Compliance', 'b': 'USB storage module disabled.', 'd': 'modprobe for USB storage devices is disabled, preventing unauthorized media usage.', 'f': 'Reduces risk of malware introduction and data exfiltration through removable devices.', 'h': 'Periodically verify kernel module settings to maintain security.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'USB Storage - modprobe was not Disabled.', 'd': 'Kernel module loading for USB storage devices is allowed, enabling potential malware introduction.', 'f': 'When USB storage is not disabled, it allows users to connect external storage devices such as USB flash drives or external hard drives, which can be a potential security risk. Malicious actors could exploit this vulnerability to introduce malware, exfiltrate sensitive data, or compromise the system\'s integrity.', 'h': '''It is recommended to restrict USB access on the system will decrease the physical attack surface for a device and diminish the possible vectors to introduce malware.
Edit or create a file in the /etc/modprobe.d/ directory ending in .conf
Example: vim /etc/modprobe.d/usb_storage.conf
Add the following line:

install usb-storage /bin/true

Run the following command to unload the usb-storage module:

rmmod usb-storage'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        40: {  # usbStorageLsmodDisabled
            'compliance': {'a': 'Compliance', 'b': 'USB storage modules not loaded.', 'd': 'USB storage modules are disabled, preventing the system from recognizing removable media.', 'f': 'Mitigates risk of malware introduction and unauthorized data transfer via USB devices.', 'h': 'Regularly audit loaded kernel modules to ensure USB storage remains disabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'USB Storage was not disabled.', 'd': 'USB storage drivers are loaded, allowing the system to recognize and mount USB devices.', 'f': 'It could have significant implications for data security and information protection. Allowing unrestricted access to USB storage devices presents a potential vulnerability, as it increases the risk of unauthorized data transfer and potential data breaches. Malicious actors could easily introduce malware or steal sensitive information by connecting infected USB drives to vulnerable systems.', 'h': '''It is recommended to edit or create a file in the /etc/modprobe.d/ directory ending in .conf
Example: vim /etc/modprobe.d/usb_storage.conf
Add the following line:

install usb-storage /bin/true

Run the following command to unload the usb-storage module:

rmmod usb-storage'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        41: {  # lastLoggedUserDisplayDisabled
            'compliance': {'a': 'Compliance', 'b': 'Last logged-in user display disabled.', 'd': 'Login screen does not show previously logged-in users, protecting usernames from exposure.', 'f': 'Reduces risk of unauthorized access and credential guessing attacks.', 'h': 'Periodically verify login settings to maintain secure configuration.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Last logged in user display was not disabled - disable user list.', 'd': 'Login screen displays previously logged-in usernames, which could assist attackers in guessing credentials.', 'f': 'It compromises user privacy and increases the risk of unauthorized access to accounts. By revealing the last logged-in user, potential attackers gain valuable information that could aid them in their malicious activities, such as social engineering or targeted attacks. it can lead to personal or sensitive data exposure, as unauthorized individuals might have access to user profiles and information.', 'h': '''It is recommended to edit or create the file /etc/dconf/profile/gdm and add the following:

user-db:user
system-db:gdm
file-db:/usr/share/gdm/greeter-dconf-defaults

Edit or create the file /etc/dconf/db/gdm.d/ and add the following: (This is typically /etc/dconf/db/gdm.d/00-login-screen)

[org/gnome/login-screen]
# Do not show the user list
disable-user-list=true

Run the following command to update the system databases:

# dconf update'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        42: {  # sambaUninstalled
            'compliance': {'a': 'Compliance', 'b': 'Samba uninstalled.', 'd': 'Samba service is not installed, reducing unnecessary exposure of file-sharing services.', 'f': 'Enhances system security and prevents unauthorized access to shared files.', 'h': 'Regularly verify installed packages to ensure no unnecessary services are active.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Samba is installed.', 'd': 'Samba service is active, potentially exposing file-sharing services to unauthorized access.', 'f': 'Improper configuration or security vulnerabilities in Samba can lead to unauthorized access, potentially exposing sensitive data to unauthorized users or external attackers.', 'h': '''It is recommended to If there is no need to mount directories and file systems to Windows systems, then this package can be removed to reduce the potential attack surface. Run the following command to remove samba:

# yum remove samba'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        43: {  # sshRootLoginDisabled
            'compliance': {'a': 'Compliance', 'b': 'SSH root login disabled.', 'd': 'Root user cannot log in directly via SSH, requiring privilege escalation from a normal user.', 'f': 'Reduces risk of unauthorized full system access through SSH.', 'h': 'Regularly review SSH configuration to maintain secure login practices.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'SSH root login was enabled.', 'd': 'Root user can log in directly via SSH, increasing risk of full system compromise.', 'f': 'It increases the risk of unauthorized access, as attackers often target the root account as the primary point of entry. This could lead to full control of the system, data breaches, and potential data loss.', 'h': '''It is recommended to disallowing root logins over SSH requires system admins to authenticate using their own individual account, then escalating to root via sudo. This in turn limits opportunity for non-repudiation and provides a clear audit trail in the event of a security incidentEdit the /etc/ssh/sshd_config file to set the parameter as follows:

PermitRootLogin no

Default Value:

PermitRootLogin without-password'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        44: {  # sshIdleTimeoutConfigured
            'compliance': {'a': 'Compliance', 'b': 'SSH idle timeout configured.', 'd': 'Inactive SSH sessions are automatically terminated after a defined period, reducing risk.', 'f': 'Prevents unauthorized access through idle sessions.', 'h': 'Periodically verify SSH timeout settings to ensure compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'SSH Idle Timeout Interval was not configured - ClientAliveCountMax.', 'd': 'SSH sessions do not automatically terminate after inactivity, leaving sessions open for misuse.', 'f': 'The ClientAliveCountMax parameter specifies the number of server-initiated messages sent to the client to check for activity during an idle session. Without a properly set timeout, inactive SSH sessions may remain open indefinitely, potentially exposing the system to security risks. Attackers could exploit idle sessions left unattended, leading to unauthorized access or privilege escalation.', 'h': '''It is recommended to set the following parameters in /etc/sysctl.conf or a /etc/sysctl.d/* file:

net.ipv4.conf.all.log_martians = 1
net.ipv4.conf.default.log_martians = 1

Run the following commands to set the active kernel parameters:

# sysctl -w net.ipv4.conf.all.log_martians=1
# sysctl -w net.ipv4.conf.default.log_martians=1
# sysctl -w net.ipv4.route.flush=1'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        45: {  # sshLoginGraceTime
            'compliance': {'a': 'Compliance', 'b': 'SSH LoginGraceTime configured securely.', 'd': 'Login grace period is limited to one minute, reducing opportunity for brute-force attacks.', 'f': 'Enhances SSH access security and reduces risk of unauthorized login.', 'h': 'Regularly review SSH configuration to maintain secure timeout settings.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'SSH LoginGraceTime was not set to one minute or less.', 'd': 'Longer login grace period allows attackers more time to attempt brute-force attacks.', 'f': 'When this parameter is left unset or not adjusted appropriately, SSH sessions may remain open indefinitely, even if they are inactive. As a result, unauthorized users or attackers could potentially gain prolonged access to the system, increasing the risk of unauthorized access or malicious activities.', 'h': '''It is recommended to edit the /etc/ssh/sshd_config file to set the parameter as follows:

LoginGraceTime 60

Default Value:

LoginGraceTime 2m'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        46: {  # passwordCreationRequirements
            'compliance': {'a': 'Compliance', 'b': 'Password complexity enforced.', 'd': 'Password creation policies require digits, special characters, and complexity rules.', 'f': 'Reduces risk of account compromise through weak passwords.', 'h': 'Regularly review and enforce password complexity policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Password creation requirements are not configured.', 'd': 'Passwords may be created without sufficient complexity, allowing weak passwords.', 'f': 'Weak passwords can compromise the security of user accounts and the systems they access by Increased risk of password guessing or cracking, password guessing or cracking, account takeover and data breaches.', 'h': '''It is recommended to edit the file /etc/security/pwquality.conf and add or modify the following line for password length to conform to site policy

minlen = 14

Edit the file /etc/security/pwquality.conf and add or modify the following line for password complexity to conform to site policy

minclass = 4

OR

dcredit = -1
ucredit = -1
ocredit = -1
lcredit = -1

Edit the /etc/pam.d/password-auth and /etc/pam.d/system-auth files to include the appropriate options for pam_pwquality.so and to conform to site policy:

password requisite pam_pwquality.so try_first_pass retry=3
'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        47: {  # passwordExpiration365Days
            'compliance': {'a': 'Compliance', 'b': 'Password expiration enforced.', 'd': 'Passwords expire within 365 days, requiring periodic updates from users.', 'f': 'Limits risk of long-term password compromise.', 'h': 'Audit password expiration policies regularly for compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Password expiration is not set to 365 days or less.', 'd': 'Passwords can be used indefinitely, increasing the likelihood of compromise.', 'f': "The window of opportunity for an attacker to leverage compromised credentials via a brute force attack, using already compromised credentials, or gaining the credentials by other means, can be limited by the age of the password. Therefore, reducing the maximum age of a password can also reduce an attacker's window of opportunity.", 'h': '''It is recommended to set the PASS_MAX_DAYS parameter to conform to site policy in /etc/login.defs :

PASS_MAX_DAYS 365

Modify user parameters for all users with a password set to match:

# chage --maxdays 365 <user>'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        48: {  # inactivePasswordLock30Days
            'compliance': {'a': 'Compliance', 'b': 'Inactive account lock configured.', 'd': 'User accounts are automatically locked after 30 days of inactivity.', 'f': 'Reduces risk of dormant account exploitation.', 'h': 'Periodically review inactive accounts and lock as per policy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Inactive password lock is not set to 30 days or less.', 'd': 'User accounts remain active indefinitely even after inactivity, increasing attack surface.', 'f': 'Inactive accounts pose a threat to system security since the users are not logging in to notice failed login attempts or other anomalies.', 'h': '''It is recommended that accounts that are inactive for 30 days after password expiration be disabled. 
Run the following command to set the default password inactivity period to 30 days:

# useradd -D -f 30

Modify user parameters for all users with a password set to match:

# chage --inactive 30 <user>
'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        49: {  # nosuidOptionTmp
            'compliance': {'a': 'Compliance', 'b': 'nosuid option set on /tmp.', 'd': '/tmp partition is mounted with nosuid, preventing execution of setuid programs.', 'f': 'Enhances system security by mitigating privilege escalation risk.', 'h': 'Verify /tmp mount options periodically to maintain compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Nosuid option was not set on /tmp partition.', 'd': '/tmp allows setuid programs to run, which can be exploited for privilege escalation.', 'f': 'Running out of /tmp space is a problem regardless of what kind of filesystem lies under it, but in a default installation a disk-based /tmp will essentially have the whole disk available, as it only creates a single / partition. On the other hand, a RAM-based /tmp as with tmpfs will almost certainly be much smaller, which can lead to applications filling up the filesystem much more easily.', 'h': '''Since the /tmp filesystem is only intended for temporary file storage, set this option to ensure that users cannot create setuid files in /tmp.
IF /etc/fstab is used to mount /tmp
Edit the /etc/fstab file and add nosuid to the fourth field (mounting options) for the /tmp partition. See the fstab(5) manual page for more information.
Run the following command to remount /tmp :

# mount -o remount,nosuid /tmp

OR if systemd is used to mount /tmp:
Edit /etc/systemd/system/local-fs.target.wants/tmp.mount to add nosuid to the /tmp mount options:

[Mount]
Options=mode=1777,strictatime,noexec,nodev,nosuid

Run the following command to restart the systemd daemon:

#  systemctl daemon-reload

Run the following command to restart tmp.mount:

# systemctl restart tmp.mount'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        50: {  # sshX11ForwardingDisabled
            'compliance': {'a': 'Compliance', 'b': 'SSH X11 forwarding disabled.', 'd': 'Remote GUI forwarding is blocked, reducing potential exploitation of SSH sessions.', 'f': 'Enhances SSH security and prevents misuse of forwarded sessions.', 'h': 'Review SSH configuration regularly to ensure forwarding is disabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'SSH X11 forwarding was not disabled.', 'd': 'X11 forwarding allows remote GUI applications to run, which may be exploited by attackers.', 'f': 'There is a small risk that the remote X11 servers of users who are logged in via SSH with X11 forwarding could be compromised by other users on the X11 server. Note that even if X11 forwarding is disabled, users can always install their own forwarders.', 'h': '''It is recommended to X11Forwarding parameter provides the ability to tunnel X11 traffic through an existing SSH shell session to enable remote graphic connections. Disable X11 forwarding unless there is an operational requirement to use X11 applications directly.

Edit the /etc/ssh/sshd_config file to set the parameter as follows:

X11Forwarding no'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        51: {  # sshWarningBannerConfigured
            'compliance': {'a': 'Compliance', 'b': 'SSH warning banner configured.', 'd': 'Banner is displayed before login, informing users about access restrictions.', 'f': 'Serves as legal notice and enhances awareness of unauthorized access policies.', 'h': 'Periodically verify SSH banners for correctness and visibility.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'SSH warning banner was not configured.', 'd': 'No warning or legal notice is displayed prior to login, which is important for compliance and deterrence.', 'f': 'SSH warning banners typically contain a warning or disclaimer about the system, such as a notice that unauthorized access is prohibited or that monitoring is in effect. Not configuring an SSH warning banner can cause Reduced security leading to accidental or intentional security breaches, such as password guessing or unauthorized access. A well-crafted warning banner can be a useful tool for communicating important information to users, such as emergency contact numbers, system downtime schedules, or instructions for reporting security incidents. Without a banner, this information may be harder to disseminate effectively.', 'h': '''It is recommended to banners are used to warn connecting users of the particular site's policy regarding connection. Presenting a warning message prior to the normal user login may assist the prosecution of trespassers on the computer system. Edit the /etc/ssh/sshd_config file to set the parameter as follows:

Banner /etc/issue.net'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        52: {  # lockoutFailedPasswordAttempts
            'compliance': {'a': 'Compliance', 'b': 'Failed login lockout configured.', 'd': 'Accounts are locked after defined failed login attempts, reducing brute-force attack risk.', 'f': 'Enhances account security and mitigates unauthorized access attempts.', 'h': 'Regularly review PAM settings to ensure lockout policies are active.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Lockout for failed password attempts were not configured.', 'd': 'Accounts do not lock after multiple failed login attempts, exposing systems to brute-force attacks.', 'f': 'When lockout policies are not in place, malicious actors have a greater opportunity to launch brute-force attacks on user accounts. These attackers can repeatedly attempt various combinations of passwords until they successfully gain unauthorized access. Without a lockout mechanism, there is no deterrent to prevent these attackers from trying numerous passwords, significantly increasing the risk of successful breaches.', 'h': '''It is recommended to Edit the files /etc/pam.d/system-auth and /etc/pam.d/password-auth and add the following lines:
Modify the deny= and unlock_time= parameters to conform to local site policy, Not to be greater than deny=5
To use pam_faillock.so module, add the following lines to the auth section:

auth        required      pam_faillock.so preauth silent audit deny=5 unlock_time=900
auth        [default=die] pam_faillock.so authfail audit deny=5 unlock_time=900

The auth sections should look similar to the following example:
Note: The ordering on the lines in the auth section is important. The preauth line needs to below the line auth required pam_env.so and above all password validation lines. The authfail line needs to be after all password validation lines such as pam_sss.so. Incorrect order can cause you to be locked out of the system
Example:

auth        required      pam_env.so
auth        required      pam_faillock.so preauth silent audit deny=5 unlock_time=900 # <- Under 'auth required pam_env.so'
auth        sufficient    pam_unix.so nullok try_first_pass
auth        [default=die] pam_faillock.so authfail audit deny=5 unlock_time=900 # <- Last auth line before 'auth requisite  pam_succeed_if.so'
auth        requisite     pam_succeed_if.so uid >= 1000 quiet_success
auth        required      pam_deny.so

Add the following line to the account section:

account     required      pam_faillock.so

Example:

account     required     pam_faillock.so
account     required     pam_unix.so
account     sufficient   pam_localuser.so
account     sufficient   pam_pam_succeed_if.so uid < 1000 quiet
account     required     pam_permit.so

OR
To use the pam_tally2.so module, add the following line to the auth section:

auth        required      pam_tally2.so deny=5 onerr=fail unlock_time=900

The auth sections should look similar to the following example:
Note: The ordering on the lines in the auth section is important. the additional line needs to below the line auth required pam_env.so and above all password validation lines.
Example:

auth        required      pam_env.so
auth        required      pam_tally2.so deny=5 onerr=fail unlock_time=900 # <- Under 'auth required pam_env.so'
auth        sufficient    pam_unix.so nullok try_first_pass
auth        requisite     pam_succeed_if.so uid >= 1000 quiet_success
auth        required      pam_deny.so

Add the following line to the account section:

account     required      pam_tally2.so

Example:

account     required     pam_tally2.so
account     required     pam_unix.so
account     sufficient   pam_localuser.so
account     sufficient   pam_pam_succeed_if.so uid < 1000 quiet
account     required     pam_permit.so'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        53: {  # passwordReuseLimited
            'compliance': {'a': 'Compliance', 'b': 'Password reuse restricted.', 'd': 'Users are prevented from reusing previous passwords, enforcing better password hygiene.', 'f': 'Reduces risk of account compromise through predictable passwords.', 'h': 'Periodically audit password policies to ensure reuse restriction is enforced.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Password reuse was not limited.', 'd': 'Users can reuse old passwords, weakening overall password security.', 'f': 'Forcing users not to reuse their past 5 passwords make it less likely that an attacker will be able to guess the password. Without this critical security measure, users are free to repeatedly use the same password for multiple accounts, which increases the vulnerability of their credentials to cyberattacks. And attackers can leverage automated tools to systematically try numerous username and password combinations until they find a match.', 'h': '''It Is recommended to edit both the /etc/pam.d/password-auth and /etc/pam.d/system-auth files to include the remember option and conform to site policy as shown:
Note: Add or modify the line containing the pam_pwhistory.so after the first occurrence of password requisite:

password    required      pam_pwhistory.so remember=5

Example: (Second line is modified)

password    requisite     pam_pwquality.so try_first_pass local_users_only authtok_type=
password    required      pam_pwhistory.so use_authtok remember=5 retry=3
password    sufficient    pam_unix.so sha512 shadow nullok try_first_pass use_authtok
password    required      pam_deny.so

Additional Information:

This setting only applies to local accounts.

This option is configured with the remember=n module option in /etc/pam.d/system-auth and /etc/pam.d/password-auth

This option can be set with either one of the two following modules:

pam_pwhistory.so - This is the newer recommended method included in the remediation section.

pam_unix.so - This is the older method, and is included in the audit to account for legacy configurations.'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        54: {  # minimumDaysPasswordChanges
            'compliance': {'a': 'Compliance', 'b': 'Minimum password age configured.', 'd': 'Users must wait a defined period before changing passwords, ensuring proper password rotation.', 'f': 'Enhances password policy enforcement and reduces potential circumvention.', 'h': 'Review login.defs periodically to maintain secure password rotation policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The minimum days between password changes was not configured.', 'd': 'Users can change passwords repeatedly in a short period, potentially bypassing password history restrictions.', 'f': 'It may lead to reduced password security as users could change their passwords frequently, potentially opting for weaker ones. This increases the risk of unauthorized access to sensitive accounts and data. Moreover, the absence of a minimum days setting can make the system more susceptible to brute-force attacks, where attackers repeatedly attempt to guess passwords until they succeed.', 'h': '''It is recommended to set the PASS_MIN_DAYS parameter to 1 in /etc/login.defs :

PASS_MIN_DAYS 1

Modify user parameters for all users with a password set to match:

# chage --mindays 1 <user>

Additional Information:

Red Hat Enterprise Linux 7 Security Technical Implementation Guide

Version 3, Release: 4 Benchmark Date: 23 Jul 2021

Vul ID: V-204418

Rule ID: SV-204418r603261_rule

STIG ID: RHEL-07-010230

Severity: CAT II

Vul ID: V-204419

Rule ID: SV-204419r603261_rule

STIG ID: RHEL-07-010240

Severity: CAT II'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Populate data based on user input
    for i, question in enumerate(questions, 2):
        # Get user input for this question
        question_num = i - 1
        user_input = None
        
        if form_data:
            # Find the corresponding form field
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        # Get response data
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            # Populate columns C, D, F, G, H
            ws.cell(row=i, column=3, value=response_data['a'])  # Compliance/Non-Compliance/Not Applicable
            ws.cell(row=i, column=4, value=response_data['b'])  # Observation (Short/Brief)
            ws.cell(row=i, column=6, value=response_data['d'])  # Observation
            ws.cell(row=i, column=7, value=response_data['f'])  # Impact
            ws.cell(row=i, column=8, value=response_data['h'])  # Recommendation
            
            # Apply alignment to these columns
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        
        # Populate Risk Factor (column E) with color coding
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            # Apply color coding
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Only apply default font if it's not the Risk Factor column (column 5)
                if col != 5:  # Don't override Risk Factor column formatting
                    cell.font = Font(name='Calibri', size=11)
            # Set row height for wrapped text
            ws.row_dimensions[row].height = 30
    
    # Save file
    filename = "Linux Server Logical Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    """
    Delete the generated Excel file after download
    """
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
