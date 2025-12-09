import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_penetration_testing_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Penetration Testing"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # Penetration Testing Questions
    questions = [
        "Is it ensured that products/services using the Internet for connectivity or communications have undergone a successful penetration test prior to production implementation?",
        "Is there an intrusion detection system in place for all external IP connections?"
    ]

    # Risk Factors
    risk_factors = [
        "High", "High"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "ptInternetConnectivity": 1,
        "ptIntrusionDetection": 2
    }

    # Question responses data
    question_responses = {
        1: {  # ptInternetConnectivity
            'compliance': {'a': 'Compliance', 'b': 'Penetration testing performed.', 'd': 'All internet-facing products/services are tested through a structured penetration testing process prior to production, with findings addressed appropriately.', 'f': 'Reduces the likelihood of security incidents, ensures the system\'s resilience, and maintains regulatory compliance.', 'h': 'Maintain records of all penetration tests and remediation activities; perform periodic retesting to ensure ongoing security.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Penetration tests were not conducted prior to production implementation for products/services that use the Internet for connectivity or communications.', 'd': 'Products or services using internet connectivity are deployed to production without prior penetration testing, leaving potential vulnerabilities unaddressed.', 'f': 'If the penetration tests are not conducted before production implementation then the service/products could be buggy and vulnerable to publicly known exploits. The attackers can use the publicly known exploits and exploit the vulnerability present in the product/service to cause damage to the bank or steal confidential information.', 'h': 'Ensure all internet-facing products/services undergo formal penetration testing before production deployment, with remediation of identified vulnerabilities documented and validated.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # ptIntrusionDetection
            'compliance': {'a': 'Compliance', 'b': 'IDS implemented for external IPs.', 'd': 'An intrusion detection system monitors all external IP connections, providing real-time alerts for suspicious activities.', 'f': 'Enhances network security, allows timely response to potential threats, and ensures compliance with regulatory security standards.', 'h': 'Periodically review IDS configurations and alerts, update signatures, and conduct incident response drills to maintain effectiveness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The intrusion detection system was not present.', 'd': 'External IP connections to the bank\'s systems are not monitored using an intrusion detection system (IDS), leaving unauthorized or malicious activity undetected.', 'f': "If the intrusion detection system is not in place then the bank will not know if an attacker has entered the network of the bank. The attacker can cause significant damage to the bank's services by entering their network. If the IPS is not present then the bank will not be alerted of the intrusion of a malicious attacker and the attack would not be prevented.", 'h': 'Deploy an intrusion detection system for all external connections to detect and alert on suspicious activities, and ensure logs are reviewed and acted upon promptly.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Penetration Testing Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
