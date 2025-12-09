import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_access_control_os_excel(form_data=None):
    """
    Create Access Control – OS Level Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Access Control – OS Level"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Access Control – OS Level Questions
    questions = [
        "Is access to the systems only through password-protected user IDs?",
        "Is unrestricted access to the systems provided only to the System Administrator?",
        "Is administration level access restricted to authorized and limited persons?",
        "Does the Operating System (OS) allot a unique user identity (ID) for all users?",
        "Does the OS prompt for a change of the user password after the lapse of the specified period?",
        "Is a record maintained and authenticated regarding the installation of the Operating System, its up-gradation, re-installation, and maintenance?",
        "Is a register maintained in respect of all the OS level users, giving details such as the date of creation, suspension, cancellation, access rights granted, and the purpose of creation, etc?",
        "Are users created for audit/maintenance purposes disabled immediately after the work is over?",
        "Are all the security features available in the OS enabled/taken advantage of as far as possible for ensuring better security?",
        "Is administration access available to officials who are under notice period, retiring shortly, or under disciplinary action?",
        "Does the OS provide for loading of virus prevention software and is it implemented?",
        "Does the department review the number of OS-level users periodically?"
    ]

    # Risk Factors (provided by user)
    risk_factors = [
        "High", "High", "High", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Low", "Low", "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "accessControlPasswordProtected": 1,
        "accessControlUnrestrictedAccess": 2,
        "accessControlAdminRestricted": 3,
        "accessControlUniqueUserID": 4,
        "accessControlPasswordChangePrompt": 5,
        "accessControlOSInstallationRecord": 6,
        "accessControlUserRegister": 7,
        "accessControlTemporaryUsersDisabled": 8,
        "accessControlSecurityFeaturesEnabled": 9,
        "accessControlExitingPersonnelAccess": 10,
        "accessControlVirusPrevention": 11,
        "accessControlUserReview": 12
    }

    # Question responses data
    question_responses = {
        1: {  # accessControlPasswordProtected
            'compliance': {'a': 'Compliance', 'b': 'Systems accessed through password-protected user IDs.', 'd': 'Only password-protected user IDs are used for system access.', 'f': 'Enhances security by restricting access to authorized personnel.', 'h': 'Periodically review access policies and enforce strong authentication.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Systems accessible without password-protected user IDs.', 'd': 'Access to the system is not restricted to password-protected user IDs.', 'f': 'Unauthorized users can gain access to the system, compromising confidentiality, integrity, and availability of data.', 'h': 'Ensure all system access is through password-protected user IDs and enforce strong password policies.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # accessControlUnrestrictedAccess
            'compliance': {'a': 'Compliance', 'b': 'Only System Administrator has unrestricted access.', 'd': 'Unrestricted system access is provided exclusively to the System Administrator.', 'f': 'Reduces risk of unauthorized system changes and ensures accountability.', 'h': 'Periodically audit access privileges to maintain compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Unrestricted access to the systems was also provided for normal users.', 'd': 'Unrestricted access to critical systems is provided to users other than System Administrator.', 'f': 'Unauthorized users can perform critical operations, modify configurations, or access sensitive information.', 'h': 'Restrict unrestricted access to only the System Administrator and maintain an access control matrix.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # accessControlAdminRestricted
            'compliance': {'a': 'Compliance', 'b': 'Admin access restricted.', 'd': 'Administrative access is restricted to authorized and limited personnel.', 'f': 'Minimizes risk of unauthorized changes and strengthens accountability.', 'h': 'Review admin access periodically and revoke unnecessary privileges.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Admin access not restricted.', 'd': 'Administrative privileges are provided to multiple users without proper authorization.', 'f': 'Multiple users with admin rights increase the risk of malicious activity, accidental misconfiguration, and security breaches.', 'h': 'Limit administrative access to authorized personnel only and maintain logs of admin activity.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # accessControlUniqueUserID
            'compliance': {'a': 'Compliance', 'b': 'The Operating System (OS) allots a unique user identity (ID) for all users.', 'd': 'Unique user id was created for operating system in Branch or HO.', 'f': 'This ensures accountability and traceability of all user activities, enabling effective monitoring, audit trails, and identification of unauthorized access or actions.', 'h': 'Maintain a user ID register and periodically review for duplicates.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Unique user id was not used for operating system in Branch or HO.', 'd': 'The OS does not assign unique user IDs for all users; some accounts share IDs.', 'f': 'Shared or non-unique IDs hinder accountability and audit trails, making it difficult to track user activity.', 'h': 'Ensure that every user is assigned a unique ID for system access.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # accessControlPasswordChangePrompt
            'compliance': {'a': 'Compliance', 'b': 'Password change prompt configured.', 'd': 'OS prompts users to change passwords after the specified period.', 'f': 'Reduces the likelihood of password compromise and strengthens access security.', 'h': 'Periodically review and enforce password policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Change of the user password after the lapse of specified period not defined in system.', 'd': 'The OS does not prompt users to change passwords after the defined period.', 'f': 'Users may continue to use old passwords, increasing the risk of compromise through password guessing or brute force attacks.', 'h': 'It is recommended that the system be configured to enforce password expiry after a defined period, in line with the organization’s security policy.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # accessControlOSInstallationRecord
            'compliance': {'a': 'Compliance', 'b': 'OS installation and maintenance records maintained.', 'd': 'All OS installations, upgrades, and maintenance activities are recorded and authenticated.', 'f': 'Enhances traceability and supports audits and forensic investigations.', 'h': 'Periodically verify records for completeness and accuracy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The record was not maintained and authenticated regarding the installation of the Operating System, its up-gradation, re-installation, and maintenance.', 'd': 'No authenticated record exists for OS installation, upgrades, reinstallation, or maintenance.', 'f': 'Lack of records makes tracking system changes difficult, hindering audit and forensic analysis. It increases the risk of unauthorized modifications, configuration errors, and potential security breaches.', 'h': 'Maintain authenticated records of OS installations, updates, and maintenance activities.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # accessControlUserRegister
            'compliance': {'a': 'Compliance', 'b': 'OS user register maintained.', 'd': 'A register exists documenting all OS-level user details.', 'f': 'Ensures proper tracking and accountability for user access.', 'h': 'Review the register periodically for accuracy and completeness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'OS user register not maintained.', 'd': 'No register exists capturing details of OS-level users including creation, suspension, access rights, and purpose.', 'f': 'Lack of documentation prevents proper user management, accountability, and auditing of system access.', 'h': 'Maintain a comprehensive register for all OS-level users and update it regularly.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # accessControlTemporaryUsersDisabled
            'compliance': {'a': 'Compliance', 'b': 'Users created for audit or maintenance purposes are disabled immediately after the completion of the assigned work.', 'd': 'All users created for audit/maintenance purposes are disabled after work completion.', 'f': 'This control minimizes the risk of unauthorized access and misuse of privileged accounts, ensuring system security and maintaining integrity of the IT environment.', 'h': 'Regularly audit temporary accounts for compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Users created for audit/maintenance purposes were not disabled immediately after the work was over.', 'd': 'It was observed that users created for audit or maintenance tasks remain active even after the work is completed.', 'f': 'Active temporary accounts can be exploited for unauthorized access, compromising system security.', 'h': 'Immediately disable or remove temporary accounts after completion of their task.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # accessControlSecurityFeaturesEnabled
            'compliance': {'a': 'Compliance', 'b': 'OS security features fully enabled.', 'd': 'All relevant OS security features are enabled to enhance system security.', 'f': 'Reduces risk of compromise and improves protection against attacks.', 'h': 'Periodically review OS security settings to ensure continued protection.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'All the security features available in the OS were not enabled.', 'd': 'Not all available security features in the OS are enabled or utilized.', 'f': 'Incomplete use of OS security features exposes the system to vulnerabilities and attacks.', 'h': 'Enable and configure all relevant OS security features according to best practices and organizational policy.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # accessControlExitingPersonnelAccess
            'compliance': {'a': 'Compliance', 'b': 'Admin access restricted for exiting personnel.', 'd': 'Administrative privileges are revoked for personnel who are leaving or under disciplinary review.', 'f': 'Reduces risk of misuse or unauthorized access by exiting employees.', 'h': 'Periodically review access rights of personnel nearing exit.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Administration access was available to the officials who are under notice period, retiring shortly, or under disciplinary action.', 'd': 'It was observed that administration access is still available to users under notice, retiring, or under disciplinary action.', 'f': 'Providing administrative access to officials who are under notice period, retiring soon, or under disciplinary action poses a serious security risk. Such users may misuse their elevated privileges to alter, delete, or leak sensitive information, intentionally or unintentionally.', 'h': 'Immediately revoke administrative access for personnel under notice, retiring, or under disciplinary action.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # accessControlVirusPrevention
            'compliance': {'a': 'Compliance', 'b': 'The Operating System (OS) supports the installation of antivirus software, and virus prevention measures are implemented across all systems.', 'd': 'OS has antivirus software installed and configured properly.', 'f': 'This ensures protection against malware, viruses, and other malicious threats, enhancing system reliability, data integrity, and overall cybersecurity posture.', 'h': 'Ensure regular virus definition updates and periodic system scans.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Virus prevention software not implemented.', 'd': 'OS does not have virus prevention software installed or enabled.', 'f': 'Lack of antivirus protection increases the risk of malware infection, data compromise, and system instability.', 'h': 'Install and configure approved antivirus/anti-malware software and ensure regular updates.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # accessControlUserReview
            'compliance': {'a': 'Compliance', 'b': 'The department conducts periodic reviews of OS-level user accounts.', 'd': 'Department regularly reviews OS-level user accounts to ensure only authorized users exist.', 'f': 'Enhances access control and minimizes risk of unauthorized access.', 'h': 'Continue periodic reviews and maintain proper documentation.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The department does not review the number of OS-level users periodically.', 'd': 'No periodic review of OS-level user accounts is conducted by the department.', 'f': 'Inactive or unauthorized accounts may remain, increasing the risk of unauthorized access and security breaches.', 'h': 'Conduct periodic reviews of OS-level users and remove or disable unnecessary accounts.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate data based on user input
    for i, question in enumerate(questions, 2):
        # Get user input for this question
        question_num = i - 1
        user_input = None
        
        if form_data:
            # Find the corresponding form field
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        # Get response data
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            # Populate columns C, D, F, G, H
            ws.cell(row=i, column=3, value=response_data['a'])  # Compliance/Non-Compliance/Not Applicable
            ws.cell(row=i, column=4, value=response_data['b'])  # Observation (Short/Brief)
            ws.cell(row=i, column=6, value=response_data['d'])  # Observation
            ws.cell(row=i, column=7, value=response_data['f'])  # Impact
            ws.cell(row=i, column=8, value=response_data['h'])  # Recommendation
            
            # Apply alignment to these columns
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        
        # Populate Risk Factor (column E) with color coding
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            # Apply color coding
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Only apply default font if it's not the Risk Factor column (column 5)
                if col != 5:  # Don't override Risk Factor column formatting
                    cell.font = Font(name='Calibri', size=11)
            # Set row height for wrapped text
            ws.row_dimensions[row].height = 30
    
    # Save file
    filename = "Access Control OS Level Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    """
    Delete the generated Excel file after download
    """
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
