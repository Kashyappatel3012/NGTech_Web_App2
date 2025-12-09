import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_user_account_maintenance_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "User Account Maintenance"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # User Account Maintenance Questions
    questions = [
        "Is every user ID at the operating system level created only after specific approval from the Branch Manager/Department Head in writing on a request form signed by the respective user?",
        "Are all user IDs protected with passwords?",
        "Is the branch/office maintaining a user profile register apart from the approved request forms?",
        "Do the operating system user IDs have security equivalence to Super User?",
        "Are all default system login accounts disabled?"
    ]

    # Risk Factors
    risk_factors = [
        "High", "High", "Medium", "Medium", "Medium"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "uamUserIDApproval": 1,
        "uamPasswordProtection": 2,
        "uamUserProfileRegister": 3,
        "uamSuperUserPrivileges": 4,
        "uamDefaultAccountsDisabled": 5
    }

    # Question responses data
    question_responses = {
        1: {  # uamUserIDApproval
            'compliance': {'a': 'Compliance', 'b': 'User IDs created with proper approval.', 'd': 'All operating system user IDs are created only after obtaining written approval from the Branch Manager/Department Head, with the request forms signed by the respective users.', 'f': 'Ensures accountability, traceability, and reduces the risk of unauthorized system access.', 'h': 'Maintain records of approvals for audit purposes and periodically review user ID creation processes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'User IDs created without proper approval.', 'd': 'Some operating system user IDs were found created without documented approval from the Branch Manager or Department Head, and without signed request forms.', 'f': 'Increases the risk of unauthorized access, weak accountability, and potential misuse of critical systems.', 'h': 'Ensure all operating system user IDs are created only after receiving written approval on a signed request form from the responsible authority.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # uamPasswordProtection
            'compliance': {'a': 'Compliance', 'b': 'User IDs protected with passwords.', 'd': 'All operating system user IDs are secured with strong, unique passwords, in accordance with organizational password policy.', 'f': 'Reduces the risk of unauthorized access and strengthens system security.', 'h': 'Periodically review and enforce password policies to ensure compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'User IDs not password protected.', 'd': 'Some operating system user accounts were found without password protection or using default weak passwords.', 'f': 'Exposes systems to unauthorized access, data breaches, and potential misuse by internal or external actors.', 'h': 'Enforce mandatory password protection for all user IDs and implement strong password policies including complexity and periodic changes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # uamUserProfileRegister
            'compliance': {'a': 'Compliance', 'b': 'User profile register maintained.', 'd': 'Each branch/office maintains a user profile register in addition to approved request forms, documenting user IDs, roles, and access privileges.', 'f': 'Supports auditability, accountability, and monitoring of user access.', 'h': 'Periodically update the register and reconcile it with the actual system accounts.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': ' The user profile register was not maintained.', 'd': 'Branches/offices were found not maintaining a separate user profile register that lists all operating system users along with their roles and permissions.', 'f': 'Makes auditing and tracking user access difficult, increasing the risk of undetected unauthorized activities.', 'h': 'Maintain a detailed user profile register for all operating system users, recording roles, permissions, and approval references.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # uamSuperUserPrivileges
            'compliance': {'a': 'Compliance', 'b': 'Super User privileges restricted.', 'd': 'Only authorized personnel have operating system accounts with Super User privileges, with access granted based on role and requirement.', 'f': 'Ensures critical system operations are controlled and reduces the risk of misuse or errors.', 'h': 'Periodically review privileged accounts and monitor their activities for compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Unrestricted Super User privileges granted.', 'd': 'Certain operating system user IDs were found with security privileges equivalent to Super User without proper authorization or business need.', 'f': 'Increases the risk of accidental or intentional system changes, data modification, or deletion, compromising system integrity.', 'h': 'Restrict Super User privileges only to authorized personnel with documented business justification and periodic review.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # uamDefaultAccountsDisabled
            'compliance': {'a': 'Compliance', 'b': 'Default accounts disabled.', 'd': 'All default system login accounts have been disabled, ensuring only authorized users can access the system.', 'f': 'Strengthens system security and minimizes potential attack vectors.', 'h': 'Regularly review the system for any newly created default accounts and disable them immediately.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'All default system login accounts were not disabled.', 'd': 'Some default system login accounts were found enabled, potentially allowing unauthorized access or exploitation.', 'f': 'The default system login user Id and password can be easily guessed and compromised by an attacker and it can be used to perform malicious activities.', 'h': 'Disable all default system accounts and ensure only authorized accounts exist for system access.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "User Account Maintenance Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
