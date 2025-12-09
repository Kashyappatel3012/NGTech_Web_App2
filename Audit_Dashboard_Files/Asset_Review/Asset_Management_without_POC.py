import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_asset_management_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Management"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # Asset Management Questions
    questions = [
        "Whether IT assets have been documented in a separate Inventory Register or Asset Inventory Management tool/software?",
        "Does the inventory include minimum details of the IT assets (e.g., hardware, software, network devices, key personnel, services, associated business applications, criticality of the IT asset such as High/Medium/Low, etc.)?",
        "Whether the IT assets are updated and reviewed periodically?",
        "Whether the IT assets' criticality is updated and reviewed regularly?",
        "Whether the bank has a centralized authorized software inventory/register?",
        "Are All hardware's are Labels properly?",
        "Does the bank have a centralized inventory of authorized devices and other related network devices connected to the bank's network (within/outside bank premises)?",
        "Does the bank have an Asset Management Policy?",
        "Is the Asset Inventory accessed only by authorized users?"
    ]

    # Risk Factors
    risk_factors = [
        "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "assetInventoryRegister": 1,
        "assetInventoryDetails": 2,
        "assetPeriodicUpdate": 3,
        "assetCriticalityReview": 4,
        "assetSoftwareInventory": 5,
        "assetHardwareLabels": 6,
        "assetDeviceInventory": 7,
        "assetManagementPolicy": 8,
        "assetAuthorizedAccess": 9
    }

    # Question responses data
    question_responses = {
        1: {  # assetInventoryRegister
            'compliance': {'a': 'Compliance', 'b': 'IT assets documented systematically.', 'd': 'All IT hardware, software, network devices, and other critical IT assets are recorded in a centralized register or asset management tool.', 'f': 'Facilitates monitoring, tracking, and management of IT assets efficiently.', 'h': 'Periodically reconcile the register with actual assets to ensure completeness and accuracy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'IT Assets have been not documented in a separate Inventory Register or Asset inventory Management tool/Software.', 'd': 'It was observed that the IT assets within the organization have not been documented in a separate Inventory Register or Asset Inventory Management tool.', 'f': 'Difficult to track, manage, or account for IT assets, increasing the risk of asset loss, mismanagement, and inefficient resource allocation.', 'h': 'Implement a centralized Asset Inventory Management system or register capturing all IT assets, ensuring regular updates and accountability.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # assetInventoryDetails
            'compliance': {'a': 'Compliance', 'b': 'Inventory captures all critical details.', 'd': 'The asset register/tool includes comprehensive details for each IT asset, including hardware/software type, criticality, owner, associated applications, and service level information.', 'f': 'Supports risk assessment, security planning, and effective resource management.', 'h': 'Regularly review and validate asset details for accuracy and completeness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Inventory does not  includes minimum detail of the IT Asset.', 'd': 'It was observed that the inventory of IT assets does not include the minimum detail of the IT Asset (viz., hardware/software/network devices, key personnel, services, Associated business applications, Criticality of the IT asset (for example, High/Medium/Low)etc.)', 'f': "If the inventory does not include minimum details of the IT asset, there can be several impacts, including incomplete visibility into the organization's IT infrastructure, Difficulty in tracking assets, Increased security risks and Inefficient resource management.  It can be difficult to detect security threats and vulnerabilities without knowing what assets are on the network and their configuration.", 'h': 'It is recommended to included minimum detail of the IT Asset (viz., hardware/software/network devices, key personnel, services, Associated business applications, Criticality of the IT asset (for example, High/Medium/Low)etc.) in the Inventory of Assets.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # assetPeriodicUpdate
            'compliance': {'a': 'Compliance', 'b': 'Asset inventory reviewed periodically.', 'd': 'IT assets are reviewed and updated at defined intervals, capturing newly procured devices, retired assets, and changes in asset status.', 'f': 'Ensures accurate asset records, effective resource utilization, and improved security control.', 'h': 'Continue periodic audits and reconciliations to maintain inventory integrity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'IT assets are not updated and reviewed periodically.', 'd': 'It was observed that the IT assets are not updated and reviewed periodically. This lack of regular maintenance and oversight could potentially lead to security vulnerabilities, performance issues, and outdated technology.', 'f': "Failing to update and review IT assets periodically can have significant impacts on an organization's information security posture including Increased security risks, Decreased performance and availability and Inefficient resource management.", 'h': 'It is recommended that IT assets should be updated and reviewed periodically for maintaining a secure and efficient IT infrastructure.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # assetCriticalityReview
            'compliance': {'a': 'Compliance', 'b': 'Criticality reviewed regularly.', 'd': 'The criticality of all IT assets is periodically reviewed, ensuring alignment with business needs, compliance requirements, and operational priorities.', 'f': 'Enhances risk management and prioritization for maintenance, security, and disaster recovery planning.', 'h': 'Conduct periodic workshops with IT and business units to validate criticality classifications.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'IT assets criticality is not updated and reviewed regularly.', 'd': 'Asset criticality classifications (High/Medium/Low) are outdated or missing, failing to reflect current business priorities or operational impact.', 'f': 'This lack of updated criticality assessments can lead to misaligned priorities in addressing security incidents and managing resources efficiently. Regularly reviewing and adjusting the criticality levels based on evolving business needs and security risks is essential for proper incident response planning and resource allocation. By maintaining accurate criticality assessments, the organization can better prioritize its efforts to protect high-value assets, reduce potential threats, and enhance overall cybersecurity measures.', 'h': "It is recommended that Criticality should be assigned for critically high devices and defined for the assets with Low/Medium criticality based on the importance of an IT asset to the organization's operations, the potential impact of a failure or breach, and the cost of replacing or repairing the asset. Criticality helps to focus organisation's resources and efforts on the most critical assets."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # assetSoftwareInventory
            'compliance': {'a': 'Compliance', 'b': 'Centralized software inventory maintained.', 'd': 'All approved and licensed software is recorded in a centralized register with associated license and usage information.', 'f': 'Ensures compliance with licensing agreements, reduces legal risk, and enables effective software management.', 'h': 'Regularly audit software installations against the centralized inventory.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The bank does not have a Centralised authorised software inventory/ Register.', 'd': 'It was observed that the bank does not have Centralised authorised software inventory/ Register.', 'f': 'Not having a centralized authorized software inventory/register can make it difficult to track software purchases and negotiate better contracts with vendors and installation of Unlicensed or outdated software can lead to decreased performance and availability, and security branches.', 'h': 'It is recommended that the bank should have a centralised authorised software inventory/ Register to help bank to ensure that they are using authorized and licensed software across their organization.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # assetHardwareLabels
            'compliance': {'a': 'Compliance', 'b': 'All hardware are labeled as per the inventory.', 'd': 'All hardware assets are properly labeled in accordance with the inventory records.', 'f': 'Facilitates easy identification, tracking, and management of IT assets, ensuring accuracy in asset inventory and accountability.', 'h': 'The bank should continue periodic verification of asset labels against inventory records to maintain consistency and detect any discrepancies promptly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'All hardware are not labeled as per the inventory.', 'd': 'It was observed that not all hardware assets within the branch and Head Office are properly labeled with unique identification tags.', 'f': 'Improper labeling of hardware assets may lead to ineffective asset tracking, misplacement, or unauthorized movement of equipment. It also creates challenges in performing inventory verification, maintenance, and audit reconciliation, increasing the risk of asset misuse or loss.', 'h': 'The bank should ensure that all hardware assets are clearly and permanently labeled with unique asset identification numbers that correspond to entries in the asset register.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # assetDeviceInventory
            'compliance': {'a': 'Compliance', 'b': 'Centralized device inventory exists.', 'd': 'All authorized network devices and endpoints, whether on-premises or remote, are recorded in a centralized inventory with appropriate ownership and access details.', 'f': 'Enhances security posture and ensures effective management of network assets.', 'h': 'Periodically reconcile network devices with the inventory to detect unauthorized devices.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "The bank doesn't have a centralized inventory of authorised devices and other related network devices connected to banks network.", 'd': 'Network devices and endpoints, including remote or branch devices, are not consistently inventoried or authorized, leading to potential unauthorized access points.', 'f': 'Without a centralized inventory, it becomes challenging to track and manage the devices connected to the network effectively. This deficiency increases the likelihood of unauthorized devices accessing the network, potential security breaches, and difficulties in identifying and mitigating security incidents.', 'h': 'It is recommended to establish a centralized inventory system to track and manage authorized devices.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # assetManagementPolicy
            'compliance': {'a': 'Compliance', 'b': 'Asset Management Policy exists.', 'd': 'A formal policy defines procedures for inventory management, asset lifecycle, criticality classification, and access controls for IT assets.', 'f': 'Standardizes asset management, improves accountability, and supports compliance requirements.', 'h': 'Review and update the policy periodically to reflect changes in IT infrastructure and business priorities.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The bank does not have any asset management policy.', 'd': 'There is no documented policy outlining roles, responsibilities, processes, and procedures for managing IT assets across the bank.', 'f': 'Inconsistent asset management practices, higher risk of asset misplacement, untracked software installations, and operational inefficiencies.', 'h': 'Develop and implement a formal Asset Management Policy, covering procurement, assignment, maintenance, usage, retirement, and disposal.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # assetAuthorizedAccess
            'compliance': {'a': 'Compliance', 'b': 'Access restricted to authorized personnel.', 'd': 'Role-based access controls are in place, ensuring that only authorized users can modify or update the asset register, while others may have read-only access if required.', 'f': 'Protects integrity of asset data, ensures accountability, and minimizes risk of unauthorized changes.', 'h': 'Periodically review user access rights and conduct audits to verify adherence to access control policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Asset Inventory is accessed  by unauthorised users.', 'd': 'The asset inventory is accessible to multiple personnel without role-based access controls, increasing the risk of tampering or inadvertent modifications.', 'f': 'Unauthorized changes to asset data can lead to mismanagement, inaccurate reporting, and increased operational risk.', 'h': 'Implement role-based access controls to ensure only authorized personnel can view or modify the asset inventory.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Asset Management Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
