import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_logical_access_controls_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Logical Access Controls"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # Logical Access Controls Questions
    questions = [
        "Is access to the operating system command prompt disabled for general users in the branch/office?",
        "Are some or more system administration-related activities, driven through a menu-based utility, assigned to any user ID?",
        "Can any user other than the Super User modify the system activity log file?"
    ]

    # Risk Factors
    risk_factors = [
        "High", "High", "High"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "lacCommandPromptDisabled": 1,
        "lacAdminMenuUtility": 2,
        "lacSystemLogModification": 3
    }

    # Question responses data
    question_responses = {
        1: {  # lacCommandPromptDisabled
            'compliance': {'a': 'Compliance', 'b': 'OS command prompt restricted.', 'd': 'General users do not have access to the operating system command prompt; only authorized users can execute system commands.', 'f': 'Reduces the risk of unauthorized or accidental system changes, ensuring system stability and security.', 'h': 'Periodically audit user permissions to ensure access restrictions are enforced.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Access to the Operating system command prompt was not disabled.', 'd': 'General users in the branch/office were found to have access to the operating system command prompt, allowing execution of system-level commands.', 'f': 'As the access to the Operating system command prompt is not disabled, a user with malicious intent can run commands to execute viruses and malware which might corrupt the OS of the system or the malicious users can elevate their privileges to carry out malicious activities.', 'h': 'Restrict access to the OS command prompt for general users and allow only authorized personnel based on business needs.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # lacAdminMenuUtility
            'compliance': {'a': 'Compliance', 'b': 'Admin functions restricted.', 'd': 'Only authorized user IDs have access to system administration activities via menu-based utilities, ensuring proper segregation of duties.', 'f': 'Ensures operational control, minimizes errors, and strengthens system security.', 'h': 'Regularly review user privileges to ensure compliance with access policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'System administration-related activities were driven through the normal user.', 'd': 'Certain users were found to have access to system administration functions via menu-based utilities without proper authorization.', 'f': 'Users with normal privileges can also carry out administration-related activities due to improper authorization implementation. If normal users are given administrative privileges then the normal user can also  perform the administrative activities which can compromises the confidentiality, integrity , availability of the bank system.', 'h': 'Limit access to system administration utilities strictly to authorized user IDs with documented approval.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # lacSystemLogModification
            'compliance': {'a': 'Compliance', 'b': 'System logs restricted to Super User.', 'd': 'System activity log files can only be modified by Super User accounts; other users have read-only access or no access.', 'f': 'Preserves the integrity of audit trails, ensures accountability, and supports effective monitoring of system activities.', 'h': 'Periodically verify log file permissions and monitor attempts to modify logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Users other than superusers can modify the system log activity.', 'd': 'Some user IDs other than the Super User were found capable of modifying or deleting system activity log files.', 'f': 'As the normal user can modify user log activity, a user with malicious intent can modify the logs to misguide forensic investigators. Hence it will be difficult for investigators to find the root cause of the incident.', 'h': 'Restrict modification of system activity logs strictly to Super User accounts and enforce monitoring for any access attempts.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Logical Access Controls Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
