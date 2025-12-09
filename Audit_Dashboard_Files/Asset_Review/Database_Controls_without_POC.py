import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_database_controls_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Database Controls"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # Database Controls Questions
    questions = [
        "Is there a backup schedule?",
        "If entries in the database are updated/deleted due to any exceptional circumstances, are they approved in writing and recorded?",
        "Is there joint responsibility of the user department and the IT department for administration of mission-critical databases?",
        "Does the IT department have laid down standards/conventions for database creation, storage, naming, and archival?",
        "Is the use of triggers and large queries monitored to prevent overloading of the database and consequent system failure?",
        "Is the creation of users restricted and need-based?",
        "Is the database configured to ensure audit trails, logging of user sessions, and session auditing?",
        "Are batch error logs reviewed and corrective action taken by the administrator periodically?",
        "In cases where customer data is provided to external service providers, does the bank have confidentiality undertakings from these service providers?",
        "Are standard sets of database control reports designed?",
        "In cases when data is migrated from one system to another, is the user department verified and satisfied with the accuracy of the information migrated?",
        "Does the system administrator periodically review the list of users of the database?",
        "Are databases periodically retrieved from the backup in the test environment to ensure accuracy in the physical environment?",
        "Are inactive users deactivated?"
    ]

    # Risk Factors
    risk_factors = [
        "Critical", "High", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", 
        "Medium", "Medium", "Medium", "Medium", "Medium", "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "dcBackupSchedule": 1,
        "dcExceptionalUpdates": 2,
        "dcJointResponsibility": 3,
        "dcDatabaseStandards": 4,
        "dcTriggersQueriesMonitoring": 5,
        "dcUserCreationRestricted": 6,
        "dcAuditTrails": 7,
        "dcBatchErrorLogs": 8,
        "dcConfidentialityUndertakings": 9,
        "dcControlReports": 10,
        "dcDataMigrationVerification": 11,
        "dcUserListReview": 12,
        "dcBackupRestorationTest": 13,
        "dcInactiveUsersDeactivated": 14
    }

    # Question responses data
    question_responses = {
        1: {  # dcBackupSchedule
            'compliance': {'a': 'Compliance', 'b': 'Backup schedule implemented.', 'd': 'All critical systems and databases follow a formal backup schedule, with clearly defined responsibilities, frequency, and retention guidelines.', 'f': 'Ensures timely recovery of data in case of failure or disaster, reduces downtime, and supports business continuity.', 'h': 'Periodically review and test backups to validate integrity and adherence to the schedule.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The backup schedule was not available.', 'd': 'No formal or documented backup schedule exists for critical databases and systems, and backups were performed irregularly without consistency.', 'f': 'Increases the risk of severe data loss during system failures, disasters, or cyberattacks. Recovery without a proper schedule may take longer and compromise business continuity.', 'h': 'Implement a formal, documented backup schedule for all critical systems and databases, specifying frequency, storage location, retention period, and responsible personnel. Conduct periodic validation and testing of backups to ensure recoverability.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # dcExceptionalUpdates
            'compliance': {'a': 'Compliance', 'b': 'Exceptional updates/deletions documented and approved.', 'd': 'All updates or deletions due to exceptional circumstances are approved in writing, recorded in a central register, and traceable to authorized personnel.', 'f': 'Enhances accountability, maintains data integrity, and ensures traceability for audit and regulatory purposes.', 'h': 'Conduct periodic reviews of exceptional changes to verify proper authorization and documentation.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The database that was updated/deleted due to any exceptional circumstances was not approved in writing and recorded.', 'd': 'Database entries updated or deleted under special circumstances were performed without written approval or proper documentation, leaving no audit trail.', 'f': 'Increases the risk of unauthorized modifications, fraudulent activities, and data inconsistencies, which can affect decision-making and regulatory compliance.', 'h': 'Ensure all exceptional changes to database entries are approved in writing, properly documented, and logged in a centralized register for audit purposes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # dcJointResponsibility
            'compliance': {'a': 'Compliance', 'b': 'Joint responsibility established.', 'd': 'Both the IT and user departments share administration responsibilities, ensuring proper checks and balances.', 'f': 'Improves operational control, prevents mismanagement, and ensures accountability for database administration.', 'h': 'Periodically review roles, responsibilities, and workflows to maintain effective segregation of duties.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'It was observed that responsibility  was not shared between user department and the IT Department for administration of mission critical databases.', 'd': 'It was observed that responsibility  was not shared between user department and the IT Department for administration of mission critical databases.', 'f': "There might be a lack of accountability and clarity in managing these databases, resulting in potential data inconsistencies, errors, or security vulnerabilities. Miscommunication and misunderstandings between the two departments may lead to delays in critical database tasks, affecting the organization's operations and decision-making processes. Moreover, without a collaborative approach, it may be challenging to establish and adhere to best practices for database management, leading to suboptimal performance and increased risks of data loss or breaches.", 'h': 'It is recommended that responsibility  must be shared between user department and the IT Department for administration of mission critical databases.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # dcDatabaseStandards
            'compliance': {'a': 'Compliance', 'b': 'Database standards implemented.', 'd': 'IT department follows documented standards for database creation, naming, storage, and archival, ensuring consistency and structured management.', 'f': 'Facilitates effective management, improves operational efficiency, and reduces risk of errors or mismanagement.', 'h': 'Review and update standards periodically to align with business and regulatory requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "IT department hasn't a laid down standards/conventions for database creation, storage, naming, and archival.", 'd': "It was observed that the IT department hasn't a laid down standards/conventions for database creation, storage, naming, and archival.", 'f': 'If the standards are not maintained for database creation, storage, archival, and naming, then different persons working in the IT team will have a hard time maintaining sources and appending changes to the database.  As all members will use different conventions, they will spend more time on searching resources and appending changes rather than solving the issues or improving the database.', 'h': 'It is recommended that the IT department should have a laid down standards/conventions for database creation, storage, naming, and archival for ease of use.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # dcTriggersQueriesMonitoring
            'compliance': {'a': 'Compliance', 'b': 'Triggers and queries monitored.', 'd': 'Triggers and large queries are actively monitored, and preventive measures are in place to avoid overloading or performance issues.', 'f': 'Ensures stable database performance, prevents downtime, and maintains operational efficiency.', 'h': 'Continuously monitor and review query logs to detect and prevent potential overload situations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Triggers and large queries were not monitored to prevent overloading of the database and consequent system failure.', 'd': 'It was observed that triggers and large queries were not monitored to prevent overloading of database and consequent system failure.', 'f': 'If the triggers and large queries overload the database or system then the system can be slowed down for a while until the query is processed. It can make the system unresponsive or freeze it if it cannot handle the request, which will be similar to Denial Of  Service attacks.', 'h': 'It is recommended that triggers and large queries should be monitored to prevent overloading of the database and consequent system failure to ensure system availability.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # dcUserCreationRestricted
            'compliance': {'a': 'Compliance', 'b': 'User creation controlled.', 'd': 'User accounts are created strictly on a need-to-access basis, with proper approvals documented.', 'f': 'Enhances database security and ensures that only authorized personnel have access to critical resources.', 'h': 'Conduct periodic user access reviews and deactivate accounts no longer required.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The creation of users was not restricted and need-based.', 'd': 'It was observed that the creation of users was not restricted and need-based.', 'f': 'If users are not structured based on need and the creation of users is not restricted and need-based, anyone can perform malicious activities in the system by creating a new user id and a virus may enter the system. Due to this the system is likely to get corrupted and important data may get lost.', 'h': 'It is recommended that the creation of users should be restricted and need-based.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # dcAuditTrails
            'compliance': {'a': 'Compliance', 'b': 'Audit trails implemented.', 'd': 'Databases maintain detailed audit trails and user session logs, which are periodically reviewed to detect anomalies or unauthorized activities.', 'f': 'Strengthens accountability, supports regulatory compliance, and enables early detection of suspicious activities.', 'h': 'Regularly audit logs and monitor access patterns for compliance and security.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The database was not configured to ensure audit trails, logging of user sessions, and session auditing.', 'd': 'It was observed that the database was not configured to ensure audit trails, logging of user sessions, and session auditing.', 'f': 'If the database is not configured for session audit trails then it will be hard to know which user accessed the database resources and the other details like when and from where. If the user session audit trails are present then it will be easy to track which user made the changes, accessed the database at a particular time during forensic investigation.', 'h': 'It is recommended that the database should be configured to ensure audit trails, logging of user sessions, and session auditing.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # dcBatchErrorLogs
            'compliance': {'a': 'Compliance', 'b': 'Batch error logs reviewed regularly.', 'd': 'Administrators periodically review batch error logs and take appropriate corrective actions to resolve issues promptly.', 'f': 'Ensures timely resolution of errors, maintains data accuracy, and supports uninterrupted operations.', 'h': 'Maintain records of corrective actions and conduct periodic reviews to improve error management processes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Batch Error logs were not reviewed periodically by the administrator and corrective actions were not taken.', 'd': 'It was observed that batch error logs were not reviewed periodically by the administrator, and corrective actions were not taken.', 'f': 'If the batch error logs are not reviewed periodically then it will be hard for the admin to resolve the issue and enhance the productivity of the bank operations. The Logs will help in debugging the error.', 'h': 'It is recommended that batch error Logs should be reviewed and corrective action must be taken by the Administrator periodically.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # dcConfidentialityUndertakings
            'compliance': {'a': 'Compliance', 'b': 'Confidentiality undertakings in place.', 'd': 'Formal confidentiality agreements exist for all external service providers, ensuring protection of sensitive customer data.', 'f': 'Minimizes risk of data breaches, enhances customer trust, and ensures compliance with regulatory requirements.', 'h': 'Periodically review and update agreements to align with evolving regulatory and business requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Customer data was provided to external service providers, without confidentiality undertakings from these service providers.', 'd': 'It was observed that customer data was provided to external service providers, without confidentiality undertakings from these service providers.', 'f': ' If the external service provider does not sign a confidentiality undertaking and the bank is sharing customer data with this service provider. Then customer data can be sold by the service provider, which can be used to conduct phishing attacks on the customers.', 'h': 'It is recommended that customer data should be provided to external service providers only after having confidentiality undertakings from these service providers.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # dcControlReports
            'compliance': {'a': 'Compliance', 'b': 'Standard control reports implemented.', 'd': 'Predefined reports exist to monitor database activities, providing timely insights into updates, access, and anomalies.', 'f': 'Enhances control over database operations, improves early detection of issues, and supports compliance requirements.', 'h': 'Periodically review report content and effectiveness to ensure relevance and comprehensiveness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Standard set of database control reports were not designed.', 'd': 'There are no predefined database control reports to monitor critical activities, such as access, updates, or exceptions.', 'f': 'Without standard database control reports, it can be difficult for organizations to gain visibility into their databases. This lack of visibility can make it challenging to identify potential issues or areas of concern, such as security breaches or performance bottlenecks. It can also lead to resulting in inconsistent reporting across the organization.', 'h': 'Develop and implement standardized database control reports that cover user activity, system exceptions, and critical updates. Review reports regularly.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # dcDataMigrationVerification
            'compliance': {'a': 'Compliance', 'b': 'Data migration verified.', 'd': 'All migrated data is verified by the user department for accuracy and completeness, with discrepancies addressed before implementation.', 'f': 'Ensures reliability of migrated data, reduces operational risk, and maintains regulatory compliance.', 'h': 'Retain validation documentation for audit purposes and perform periodic reconciliation.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Data migration not verified.', 'd': 'Data migrated between systems is not validated by the user department, leaving room for inaccuracies and missing information.', 'f': 'If the user department does not verify the accuracy of the data that is being migrated from one system to another then it can affect the integrity, and availability of data. Because in migration the data can get corrupted. The inaccurate data can cause more problems for the bank.', 'h': 'Ensure that all migrated data is verified by the user department for accuracy, completeness, and consistency before going live. Maintain documented validation evidence.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # dcUserListReview
            'compliance': {'a': 'Compliance', 'b': 'User list reviewed periodically.', 'd': 'System administrator conducts periodic reviews of the database user list, ensuring only active and authorized users have access.', 'f': 'Reduces unauthorized access risk, maintains proper access control, and enhances overall database security.', 'h': 'Document review outcomes and corrective actions for audit purposes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The system administrator does not periodically review the list of users of the database.', 'd': 'The system administrator does not periodically review the list of database users, leading to potentially inactive or unauthorized accounts remaining active.', 'f': "If the administrator does not review the list of the users of the database, then an administrator will not know the inactive users or rogue users' unauthorizly access to the database.", 'h': 'Implement a periodic review of all database users, deactivate inactive accounts, and remove unauthorized users promptly.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # dcBackupRestorationTest
            'compliance': {'a': 'Compliance', 'b': 'Backup restoration tested.', 'd': 'Databases are regularly restored from backups in a test environment to confirm that backups are accurate and complete.', 'f': 'Ensures reliability of backups, minimizes risk of data loss, and supports business continuity.', 'h': 'Maintain a schedule for backup restoration testing and document results for audit purposes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Databases were not periodically retrieved from the backup in the test environment.', 'd': 'Databases are not restored from backups in a test environment, so the reliability and integrity of backups are unverified.', 'f': 'Increases the risk of data loss or corruption during recovery, potentially causing prolonged system downtime.', 'h': 'Periodically restore databases from backups in a test environment to validate accuracy, completeness, and recoverability.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # dcInactiveUsersDeactivated
            'compliance': {'a': 'Compliance', 'b': 'Inactive users deactivated.', 'd': 'Inactive database user accounts are routinely identified and deactivated to prevent unauthorized access.', 'f': 'Reduces the risk of security breaches and ensures compliance with internal access control policies.', 'h': 'Document deactivation activities and review procedures periodically to maintain security hygiene.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Inactive users were not deactivated.', 'd': 'User accounts that are no longer active remain enabled in the database, potentially providing unauthorized access.', 'f': "If the inactive users are not deactivated then those users can be used by the attackers to attack the bank's system. By deactivating the inactive users it will be easy for the bank to monitor the active users and their activities.", 'h': 'It is recommended that inactive users should be deactivated so that they cannot be used in malicious activities.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Database Controls Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
