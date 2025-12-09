import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_maintenance_patches_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Maintenance and App Patches"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # Maintenance and App Patches Questions
    questions = [
        "Is the system configuration such as memory, clock speed, hard disk size, and OS version as per the order or terms stipulated by CPPD / IT Department?",
        "Is there a well-defined patch management process in place?",
        "Are all the installed patches properly applied and running on the operating system?",
        "Is a patch management process available for the server operating system?",
        "Are the latest updates/patches released by the OS vendor regularly applied?",
        "Is a patch management register maintained and updated properly?",
        "Is a patch management process available for the system operating system?",
        "Is there a well-defined patch management policy and procedure in place?"
    ]

    # Risk Factors
    risk_factors = [
        "High", "High", "High", "High", "Medium", "Medium", "Medium", "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "mpSystemConfiguration": 1,
        "mpPatchManagementProcess": 2,
        "mpPatchesApplied": 3,
        "mpServerPatchProcess": 4,
        "mpLatestUpdates": 5,
        "mpPatchRegister": 6,
        "mpSystemPatchProcess": 7,
        "mpPatchPolicy": 8
    }

    # Question responses data
    question_responses = {
        1: {  # mpSystemConfiguration
            'compliance': {'a': 'Compliance', 'b': 'System configurations meet prescribed standards.', 'd': 'All systems were configured according to CPPD/IT Department requirements, with updated hardware and compatible OS versions.', 'f': 'Ensures smooth performance and compatibility with enterprise applications.', 'h': 'Continue maintaining standard configurations and document system upgrades regularly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The System configuration was not as per the terms stipulated by CPPD/IT Department.', 'd': 'Certain systems were found with outdated OS versions, insufficient RAM, and hardware below CPPD/IT Department standards.', 'f': 'If System configuration such as memory, clock speed, hard disk size, OS version, etc. are not as per the specified conditions, the bank may suffer many losses, including data loss because of unsuccessful retrieval of data due to the small size of the hard disk. An outdated version of OS will lead to attacks and cause damage to the system.', 'h': 'It is recommended to configure systems such as Memory, Clock speed, Hard Disk size, OS version, etc, as per order or terms stipulated by CPPD / IT Dept.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # mpPatchManagementProcess
            'compliance': {'a': 'Compliance', 'b': 'Well-defined patch management process implemented.', 'd': 'The organization follows a documented patch management policy covering patch testing, approval, and deployment timelines.', 'f': 'Reduces security risks and ensures consistent patch deployment.', 'h': 'Continue to review and update the patch policy quarterly to align with evolving security threats.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A well defined patch management process was in not implemented.', 'd': 'Patch deployment was performed inconsistently without documented procedures or defined timelines.', 'f': 'Without regular and timely patching, the system remains vulnerable to known security vulnerabilities, increasing the risk of successful cyber attacks and data breaches.  the absence of a patch management process may lead to delayed or missed critical updates, potentially causing performance issues, system crashes, or even prolonged downtime. Without a structured patch management process, it becomes challenging to prioritize and track patch installations across multiple systems, resulting in inconsistencies and gaps in security coverage.', 'h': 'Develop and implement a formal patch management policy outlining roles, responsibilities, approval workflow, and patch verification steps.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # mpPatchesApplied
            'compliance': {'a': 'Compliance', 'b': 'All OS patches applied successfully.', 'd': 'Verification reports confirmed that all systems were updated with the latest patches and functioning properly.', 'f': 'Strengthens the overall system security and reduces exploit exposure.', 'h': 'Maintain automated patch verification reports after each deployment cycle.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'All the patches installed are not applied and running on the OS.', 'd': 'Some systems had pending or partially applied patches that were not installed successfully due to system restarts or errors.', 'f': 'The unpatched vulnerabilities can be exploited by attackers to gain unauthorized access to the system, steal data or launch other types of attacks. Some patches are designed to improve system performance, and without them, the system may not perform optimally. unpatched vulnerabilities can lead to system crashes or downtime, disrupting business operations, reducing productivity, and causing financial losses.', 'h': 'Use automated tools to monitor patch installation and maintain a verification log for each deployment.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # mpServerPatchProcess
            'compliance': {'a': 'Compliance', 'b': 'Server patch management process implemented.', 'd': 'Server OS patching is managed through a defined process with pre-deployment testing and documentation.', 'f': 'Ensures stable and secure server operation post patch application.', 'h': 'Continue following the established server patching process and maintain audit trails of updates.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A patch management process is not available for Server OS.', 'd': 'Server patching activities were carried out manually without formal documentation, test environment, or approval workflow.', 'f': 'The unpatched vulnerabilities can be exploited by attackers to gain unauthorized access to the server, steal data or launch other types of attacks. Some patches are designed to improve system performance, and without them, the system may not perform optimally. Unpatched vulnerabilities can lead to server crashes or downtime, disrupting business operations, reducing productivity, and causing financial losses.', 'h': 'Establish a dedicated server patch management procedure including testing, approval, and rollback mechanisms.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # mpLatestUpdates
            'compliance': {'a': 'Compliance', 'b': 'Vendor patches applied regularly.', 'd': 'All systems were configured to automatically receive and install the latest vendor patches.', 'f': 'Minimizes risk of attacks exploiting known OS vulnerabilities.', 'h': 'Continue to monitor patch status and verify updates monthly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The patches released by the OS vendor were not applied.', 'd': 'It was observed that the patches released by the OS vendor were not applied.', 'f': 'An attacker can exploit the existing vulnerability of this system. Also, an attacker can perform a Ransomware attack on this system, as the latest updates / patches released by the OS vendor were not applied, and all the vulnerabilities are available on Google. It is easy for an attacker to compromise the system and reach to Datacentre.', 'h': 'It is recommended to implement the new patch released by the vendor for preventive maintenance, which is necessary to keep machines up-to-date, stable, and safe from malware and other threats.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # mpPatchRegister
            'compliance': {'a': 'Compliance', 'b': 'Patch register maintained and updated.', 'd': 'A centralized patch management register is maintained, recording patch release and installation details.', 'f': 'Enhances visibility and ensures complete tracking of patch status.', 'h': 'Continue to update the register after every patch cycle and verify accuracy through audits.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A patch management register is not maintained by the bank.', 'd': 'A patch management register is not maintained by the bank.', 'f': 'Without a patch management register, the bank may not be aware of which systems and applications have been patched and which vulnerabilities remain unaddressed. This can increase the risk of security breaches and data theft. Also, a lack of a patch management register can result in a disorganized and inefficient patching process.', 'h': 'It is recommended to implement a patch management register ad regularly update it.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # mpSystemPatchProcess
            'compliance': {'a': 'Compliance', 'b': 'OS patching process implemented.', 'd': 'A defined process ensures consistent and timely OS patching across all systems in the network.', 'f': 'Maintains consistent system protection across all departments.', 'h': 'Continue using the centralized deployment mechanism and perform random validation checks.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A patch management process is not available for system OS.', 'd': 'It was observed that a patch management process is not available for system OS.', 'f': 'The unpatched vulnerabilities can be exploited by attackers to gain unauthorized access to the system, steal data or launch other types of attacks. Some patches are designed to improve system performance, and without them, the system may not perform optimally. Unpatched vulnerabilities can lead to system crashes or downtime, disrupting business operations, reducing productivity, and causing financial losses.', 'h': 'It is recommended to implement a patch management process  for Endpoint system OS.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # mpPatchPolicy
            'compliance': {'a': 'Compliance', 'b': 'Formal patch management policy and procedure in place.', 'd': 'A bank-approved Patch Management Policy and SOP define patch deployment schedules, testing procedures, and verification responsibilities.', 'f': 'Ensures structured, auditable, and secure patch management.', 'h': 'Review and update the policy annually to align with current security frameworks.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A well defined patch management Policy and Procedure was not in place.', 'd': 'It was observed that a well defined patch management Policy and Procedure was not in place.', 'f': 'Without clear guidelines, there may be delays or inconsistencies in applying critical security updates, leaving the system vulnerable to known exploits and cyberattacks.  The absence of a formal procedure can hinder coordination and communication among IT teams, making it harder to track patching progress and identify potential issues.', 'h': 'Draft and approve a formal Patch Management Policy and SOP outlining scope, frequency, approval, testing, and exception handling.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Maintenance and App Patches Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
