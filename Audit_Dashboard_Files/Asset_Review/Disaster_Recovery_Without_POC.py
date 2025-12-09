import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_disaster_recovery_excel(form_data=None):
    """
    Create Disaster Recovery Excel file with comprehensive data based on user input
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Disaster Recovery"
    
    # Define all 47 questions
    questions = [
        "Is there a separate room for the server?",
        "Is the Internet allowed or not?",
        "Which OS version is currently running?",
        "Are event logs for server OS present up to date?",
        "Is Windows OS genuine or not?",
        "Is the currently running OS outdated or not?",
        "Is password policy present for server OS authentication?",
        "Is antivirus available?",
        "Is server OS updated?",
        "How many CCTV cameras are present in the server room? (Recommended two)",
        "Is antivirus expired?",
        "Is antivirus configured properly?",
        "Are USB devices allowed or not?",
        "Are logs maintained for remote access?",
        "Which application is being used for remote connection?",
        "When is the last DR drill conducted?",
        "Does server have adequate space for operational requirements?",
        "Is the server room visible from a distance and easily accessible?",
        "Is the server room away from the basement and water/drainage systems?",
        "Is the server close to the UPS room?",
        "Is smoking, eating, and drinking prohibited in the server room to prevent spillage of food or liquid into sensitive computer equipment?",
        "How many persons are authorized to access the server room?",
        "Is a biometric or proximity card machine present for access in the server room?",
        "Are biometric logs or proximity card present and maintained to the current date?",
        "Are cables for network devices properly structured or not?",
        "Is the server rack open or not?",
        "Is wire tagging available?",
        "Are cables for CCTV cameras concealed or not?",
        "Is there CCTV footage backup available for at least 30 days or not?",
        "How many fire extinguishers are available?",
        "Is an automatic fire extinguisher present or not?",
        "Is a hygrometer present or not?",
        "Is smoke detector available or not?",
        "Is a fire alarm present or not?",
        "Is antivirus updated or not?",
        "Is unauthorized software present or not?",
        "Is remote access taken by IT team or any other?",
        "At which location the DC and DR present?",
        "Are logs for DR drills properly maintained or not?",
        "Is the server room neat and clean to ensure a dust-free environment?",
        "Is a server room access log registered, present, and maintained?",
        "Are network devices placed properly in the server rack?",
        "Is there proper cooling present in the server room?",
        "How many A/Cs are present in the server room?",
        "Is switching ON/OFF of AC automatic or manual?",
        "Is there any leakage in server room walls or any other dangerous cause affecting servers in the room?",
        "Is power backup present or not?"
    ]
    
    # Risk factors for each question
    risk_factors = [
        "Critical", "Critical", "High", "High", "High", "High", "High", "High", "High", "High",
        "High", "High", "High", "High", "High", "High", "Medium", "Medium", "Medium", "Medium",
        "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium",
        "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Low",
        "Low", "Low", "Low", "Low", "Low", "Low", "Low"
    ]
    
    # Field mapping for form data
    question_mapping = {
        'separateServerRoom': 1, 'internetAllowed': 2, 'osVersion': 3, 'eventLogsUpToDate': 4,
        'windowsOSGenuine': 5, 'osOutdated': 6, 'passwordPolicyPresent': 7, 'antivirusAvailable': 8,
        'serverOSUpdated': 9, 'cctvCamerasCount': 10, 'antivirusExpired': 11, 'antivirusConfigured': 12,
        'usbDevicesAllowed': 13, 'remoteAccessLogs': 14, 'remoteConnectionApp': 15, 'lastDRDrill': 16,
        'adequateSpace': 17, 'serverRoomVisible': 18, 'serverRoomAwayFromBasement': 19, 'serverCloseToUPS': 20,
        'smokingEatingProhibited': 21, 'authorizedPersonsCount': 22, 'biometricProximityPresent': 23,
        'biometricLogsMaintained': 24, 'cablesStructured': 25, 'serverRackOpen': 26, 'wireTaggingAvailable': 27,
        'cctvCablesConcealed': 28, 'cctvBackup30Days': 29, 'fireExtinguishersCount': 30, 'automaticFireExtinguisher': 31,
        'hygrometerPresent': 32, 'smokeDetectorAvailable': 33, 'fireAlarmPresent': 34, 'antivirusUpdated': 35,
        'unauthorizedSoftwarePresent': 36, 'remoteAccessByITTeam': 37, 'dcAndDRLocation': 38, 'drDrillLogsMaintained': 39,
        'serverRoomClean': 40, 'serverRoomAccessLog': 41, 'networkDevicesPlacedProperly': 42, 'properCoolingPresent': 43,
        'acCount': 44, 'acAutomaticManual': 45, 'wallLeakagePresent': 46, 'powerBackupPresent': 47
    }
    
    # Comprehensive response data for Disaster Recovery questions
    # This file contains all 47 questions with detailed compliance/non-compliance responses
    DISASTER_RECOVERY_RESPONSES = {
        1: {  # separateServerRoom
            'compliance': {'a': 'Compliance', 'b': 'A dedicated server room is available.', 'd': 'The server is placed in a separate room designed to control temperature, humidity, and dust.', 'f': 'Ensures stable environmental conditions, reducing risks of overheating, moisture damage, and dust interference.', 'h': 'No recommendation required.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'There was no separate room for the server. ', 'd': 'Server is not housed in a dedicated room, exposing equipment to environmental threats.', 'f': '''1. Inadequate Temperature Control:-
When the temperature around and within the server and networking equipment becomes too high the server will shut down and there will be loss of data.
2. Imbalanced Moisture Levels:-
High humidity can result in rust, corrosion, short-circuiting, and even the growth of fungus that can attack the machinery. Too little moisture in the air is also a concern, as an exceedingly dry environment can result in electrostatic discharge, which can cause system malfunction and damage.
Also, there is a risk of dust and temperature interference.''', 'h': 'Provide a separate server room to protect equipment and minimize operational risks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # internetAllowed
            'compliance': {'a': 'Compliance', 'b': 'Internet access restricted as per operational need.', 'd': 'Servers access only required websites for operations like AV updates or core banking connectivity.', 'f': 'Reduces risk of malware, ransomware, and data breaches.', 'h': 'Continue limiting internet access.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Internet was allowed on servers without any restriction.', 'd': 'Servers have unrestricted internet access, including social media and public websites.', 'f': 'Full access to the internet on the server will sometimes create a critical problem if some malicious script is downloaded on the server from the internet that will remove or encrypt all the sensitive data on the server and can directly gain access to CBS. As the Internet is not restricted, any malicious activity could be performed through the internet. Also, some social media websites can be accessed by the employee which will affect the banks productivity, and using those sites unintentionally users can click on a malicious link that can download viruses, worms, or any malware that will affect all bank networks.', 'h': 'Restrict internet access based on server operational requirements.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # osVersion
            'compliance': {'a': 'Compliance', 'b': 'Updated, supported OS installed.', 'd': 'Latest version of Windows OS is installed and patched.', 'f': 'Reduces vulnerability to malware and cyber-attacks.', 'h': 'Maintain OS updates regularly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Windows 7 OS was installed on the DR server.', 'd': 'It was observed that an outdated version of OS was in use. Windows 7 was outdated in January 2020, yet it was running in the bank. Microsoft will not release security patches for this version of OS, so there is a need to upgrade the version of Windows OS.', 'f': 'An attacker can exploit the existing vulnerability of this system. Also, an attacker can perform a Ransomeware attack on this system, as the patch is not released for this OS, and all the vulnerabilities are available on Google. It is easy for an attacker to compromise the system and reach to Disaster Recovery site.', 'h': 'Upgrade OS to the latest supported version.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # eventLogsUpToDate
            'compliance': {'a': 'Compliance', 'b': 'Event logs up-to-date and regularly reviewed.', 'd': 'Event logs updated, monitored, and aligned with latest patches.', 'f': 'Provides visibility into system events and security incidents.', 'h': 'Continue maintaining up-to-date logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Event logs for server OS was not present up to date', 'd': 'It was observed that Oevent logs for server OS were not present up to date. The latest security patches were not updated on any of the servers, and some systems were not patched last some time. It should be up-to-date with the latest patch.', 'f': 'Reduces detection capability for attacks and impairs system security auditing.', 'h': 'Update and maintain event logs regularly.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # windowsOSGenuine
            'compliance': {'a': 'Compliance', 'b': 'Genuine Windows OS installed.', 'd': 'Licensed Windows OS in use with regular updates.', 'f': 'Reduces exposure to malware, hacking, and operational risks.', 'h': 'No action required.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Windows OS was not genuine.', 'd': 'It was observed that windows OS was not genuine. Bank was using a non-genuine copy of Windows.', 'f': 'Vulnerable to malware, viruses, Trojans, and unauthorized access; operational risks increase.', 'h': 'Install genuine Windows OS on all servers.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # osOutdated
            'compliance': {'a': 'Compliance', 'b': 'Up-to-date OS installed.', 'd': 'OS is fully patched and supported.', 'f': 'Mitigates risk of cyber-attacks and ensures security features are available.', 'h': 'Maintain latest OS versions.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Operating System was outdated.', 'd': 'It was observed that OS was outdated in some of the servers. Outdated OS might not be able to withstand an up-to-date cyber-attack. Bank systems will be more vulnerable to ransomware attacks, malware, and data breaches.', 'f': 'System more vulnerable to malware, ransomware, and attacks exploiting known vulnerabilities.', 'h': 'Upgrade OS to latest supported version.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # passwordPolicyPresent
            'compliance': {'a': 'Compliance', 'b': 'Password policy configured and enforced.', 'd': 'Strong passwords enforced for server OS accounts; meets bank policy.', 'f': 'Reduces risk of brute-force attacks and unauthorized access.', 'h': 'Continue enforcing strong password policy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The password policy was not configured.', 'd': 'No password policy defined for server OS authentication.', 'f': 'Easy for attackers to gain access via brute-force or credential guessing.', 'h': 'Configure password policy per bank security requirements.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # antivirusAvailable
            'compliance': {'a': 'Compliance', 'b': 'Antivirus installed and operational.', 'd': 'All servers protected by up-to-date antivirus.', 'f': 'Protects servers from malware, viruses, and cyber-attacks.', 'h': 'Continue monitoring and updating antivirus.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Antivirus was not available.', 'd': 'Servers not protected by antivirus; critical systems vulnerable.', 'f': 'High risk of virus infection, malware attacks, and network compromise.', 'h': 'Install antivirus on all servers immediately.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # serverOSUpdated
            'compliance': {'a': 'Compliance', 'b': 'Server OS fully updated.', 'd': 'All security patches applied; OS is supported.', 'f': 'Reduces vulnerability to cyber-attacks and ensures compliance with security policies.', 'h': 'Continue regular updates.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Server Windows OS was outdated.', 'd': 'Outdated OS running, missing security patches.', 'f': 'Servers vulnerable to attacks like ransomware; DR site at risk.', 'h': 'Update server OS to latest version.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # cctvCamerasCount
            'compliance': {'a': 'Compliance', 'b': 'Two CCTV cameras are installed in the server room as per the recommended standard.', 'd': 'Cameras positioned to cover all critical areas; no blind spots.', 'f': 'Continuous monitoring and video evidence available.', 'h': 'Maintain CCTV coverage and check functionality regularly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Only one CCTV camera was present.', 'd': 'Single camera with improper cable management; blind spots exist.', 'f': 'Limited monitoring; risk of missing critical incidents.', 'h': 'Install minimum two CCTV cameras to ensure full coverage.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # antivirusExpired
            'compliance': {'a': 'Compliance', 'b': 'Antivirus license is active and valid.', 'd': 'All servers are protected by valid and updated antivirus licenses.', 'f': 'Ensures protection against latest malware, viruses, and cyber threats.', 'h': 'Continue renewing antivirus licenses on time.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Expired license Antivirus was in use.', 'd': 'Antivirus on servers is expired and not providing updated protection.', 'f': 'Protection against new threats is compromised; servers vulnerable to viruses, malware, and ransomware.', 'h': 'Install licensed antivirus on all servers; use enterprise antivirus solution and maintain valid licenses.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # antivirusConfigured
            'compliance': {'a': 'Compliance', 'b': 'Antivirus configured and operational.', 'd': 'Antivirus settings are properly configured to detect malware, update signatures automatically, and perform regular scans.', 'f': 'Maximizes protection and reduces risk of infection.', 'h': 'Continue regular monitoring and configuration checks.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The antivirus was not configured properly. ', 'd': 'Antivirus installed but not configured; updates and scans not properly set.', 'f': 'Inability to detect latest viruses or malware; increased vulnerability to attacks.', 'h': 'Properly configure antivirus for automatic updates, regular scans, and threat detection.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # usbDevicesAllowed
            'compliance': {'a': 'Compliance', 'b': 'USB disabled or restricted.', 'd': 'USB ports are disabled or only trusted USB devices are allowed for operational use.', 'f': 'Reduces risk of malware, keyloggers, and data exfiltration.', 'h': 'Continue controlling USB access.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'USB was enabled in all the Servers.', 'd': 'USB devices enabled, allowing potential malware injection or data theft.', 'f': 'Increased risk of cyber-attacks like USB malware, data exfiltration, and ransomware.', 'h': 'Disable USB access on servers; allow only trusted devices if necessary.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # remoteAccessLogs
            'compliance': {'a': 'Compliance', 'b': 'Logs for remote access maintained.', 'd': 'Detailed logs available with user ID, date, time, reason, and designation.', 'f': 'Enables auditing, accountability, and forensic investigation.', 'h': 'Continue maintaining detailed remote access logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Logs were not maintained for Remote Access.', 'd': 'No records of who accessed the servers remotely, when, and for what purpose.', 'f': 'Difficult to audit remote access; accountability cannot be established.', 'h': 'Maintain detailed logs for remote access, including date, time, user, reason, and designation.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # remoteConnectionApp
            'compliance': {'a': 'Compliance', 'b': 'Paid and updated remote connection software used.', 'd': 'Secure and updated applications used for remote server access; logs generated for all sessions.', 'f': 'Reduces risk of unauthorized access, malware, and data breaches.', 'h': 'Maintain latest software versions for remote connections.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The application being used for remote connection were open source and outdated version.', 'd': 'Outdated remote access applications increase the risk of malware, keyloggers, and backdoors.', 'f': 'Servers vulnerable to cyber-attacks and malicious exploitation; audit trails may be incomplete.', 'h': 'Use secure, paid, and updated remote connection software; implement strong access controls.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # lastDRDrill
            'compliance': {'a': 'Compliance', 'b': 'DR drill conducted recently and logs maintained.', 'd': 'DR drill conducted at least twice a year; records and reports properly maintained.', 'f': 'Ensures preparedness for disasters and system recovery.', 'h': 'Continue conducting regular DR drills and maintaining logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'DR drill was not conducted for the acceptable period.', 'd': 'Last DR drill conducted over a year ago; records not properly maintained.', 'f': 'Inability to ensure system recovery during actual disasters; operational risk remains high.', 'h': 'Conduct DR drills at least twice per year; maintain detailed logs and reports.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # adequateSpace
            'compliance': {'a': 'Compliance', 'b': 'Server room has adequate space.', 'd': 'Servers arranged with proper spacing, airflow, and access for maintenance.', 'f': 'Reduces hardware malfunction risk due to heat, dust, or restricted access.', 'h': 'No action required.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The server does not have adequate space for operational requirements.', 'd': 'It was observed that the Server has no adequate space for operational requirements.', 'f': 'As the Server has no adequate space for operational requirements there is a risk of dust and temperature interference. Server hardware and related components require specific components to perform optimally, such as adequate cooling, moisture removal, and protection from excessive temperatures. Server rooms that are too hot or cold could cause hardware to malfunction leading to downtime.', 'h': 'It is recommended that the Server must have adequate space for operational requirements.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        18: {  # serverRoomVisible
            'compliance': {'a': 'Compliance', 'b': 'Server room visible and restricted.', 'd': 'Only authorized personnel can access; room is monitored.', 'f': 'Reduces insider threat and unauthorized access.', 'h': 'Maintain visibility and restricted access.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The server room was not visible from a distance and was easily accesible.', 'd': 'Room hidden from distance; anyone can enter easily.', 'f': 'Higher risk of insider threat and unauthorized access.', 'h': 'Make server room visible from a distance and restrict access to authorized personnel only.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        19: {  # serverRoomAwayFromBasement
            'compliance': {'a': 'Compliance', 'b': 'Server room is located away from basement and water/drainage systems.', 'd': 'Servers are positioned to avoid water damage risks; safe from flooding or leaks.', 'f': 'Minimizes risk of water damage, corrosion, and electrical hazards.', 'h': 'Continue maintaining location away from basement/water systems.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A server room was near to the basement and water/drainage systems.', 'd': 'Server room located close to basement and water/drainage systems.', 'f': 'Risk of water damage, equipment corrosion, short-circuiting, and static electricity hazards.', 'h': 'Relocate server room or implement protective barriers to avoid water hazards.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        20: {  # serverCloseToUPS
            'compliance': {'a': 'Compliance', 'b': 'Server is not in close to UPS room.', 'd': 'Servers and UPS batteries separated to prevent physical hazards.', 'f': 'Reduces risk of fire, explosion, or heat damage affecting servers.', 'h': 'Maintain separation between UPS and server room.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Servers were near the UPS room.', 'd': 'Servers positioned near UPS batteries in the DR room.', 'f': 'Batteries can explode or overheat, potentially damaging servers.', 'h': 'Relocate servers away from UPS batteries to reduce physical hazards.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        21: {  # smokingEatingProhibited
            'compliance': {'a': 'Compliance', 'b': 'Instructions for no smoking, eating, or drinking present.', 'd': 'Clear signage and enforcement of rules prevent contamination and spillage.', 'f': 'Reduces risk of liquid or food spillage causing equipment damage.', 'h': 'Continue enforcing these rules strictly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Instructions were not available saying that Smoking, eating, and drinking are prohibited in the server room.', 'd': 'No visible instructions in the server room; anyone can smoke, eat, or drink.', 'f': 'Increased risk of spillage, equipment damage, and hardware malfunction.', 'h': 'Place clear instructions and restrict smoking, eating, and drinking in server room.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        22: {  # authorizedPersonsCount
            'compliance': {'a': 'Compliance', 'b': 'Only limited authorized personnel have access.', 'd': 'Access strictly controlled to IT department and critical staff.', 'f': 'Reduces insider threats and unauthorized access.', 'h': 'Continue limiting access to authorized personnel only.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Numerous persons were authorized to access the server room.', 'd': 'Many people have access to the server room.', 'f': 'Increased risk of unauthorized access, insider threats, and data compromise.', 'h': 'Restrict access to essential personnel only.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        23: {  # biometricProximityPresent
            'compliance': {'a': 'Compliance', 'b': 'Biometric or proximity access control installed.', 'd': 'Server room access tracked, attendance monitored, and unauthorized entry prevented.', 'f': 'Improves security, accountability, and traceability of personnel entering the room.', 'h': 'Continue maintaining access control devices and logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A Biometric or Proximity Card machine for access in the server room was not available.', 'd': 'No access control system installed in server room.', 'f': 'Difficult to track who enters/exits; unauthorized personnel can access servers.', 'h': 'Install biometric or proximity card system to control and monitor access.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        24: {  # biometricLogsMaintained
            'compliance': {'a': 'Compliance', 'b': 'Biometric or proximity logs maintained regularly.', 'd': 'Logs up-to-date and available for auditing or forensic purposes.', 'f': 'Provides traceability of access events; helps in investigations if incidents occur.', 'h': 'Continue maintaining and monitoring logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Biometric logs or Proximity Card were not present and not maintained.', 'd': 'Biometric/proximity logs not present or updated.', 'f': 'No traceability of personnel access; complicates forensic investigations.', 'h': 'Maintain logs regularly and ensure they are up-to-date.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        25: {  # cablesStructured
            'compliance': {'a': 'Compliance', 'b': 'Network cables structured properly.', 'd': 'Organized cables allow easy maintenance, troubleshooting, and device replacement.', 'f': 'Reduces risk of accidental disconnections or equipment failure; easier network management.', 'h': 'Maintain structured cabling practices.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The cables of Network Devices were not structured properly in the server rack.', 'd': 'Network device cables tangled and unorganized in the rack.', 'f': 'Hard to maintain devices; increases risk of accidental damage and downtime.', 'h': 'Properly structure and label all network cables in server racks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        26: {  # serverRackOpen
            'compliance': {'a': 'Compliance', 'b': 'Server rack locked and secured.', 'd': 'Only authorized personnel have key access; physical security ensured.', 'f': 'Reduces risk of unauthorized access and physical tampering.', 'h': 'Continue keeping server racks locked and assign key accountability.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The server rack was not locked.', 'd': 'Server rack open and accessible to anyone in the server room.', 'f': 'Critical systems exposed to unauthorized access and physical tampering.', 'h': 'Lock server racks; assign key responsibility to accountable personnel.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        27: {  # wireTaggingAvailable
            'compliance': {'a': 'Compliance', 'b': 'Wire tagging is available for all network cables.', 'd': 'All cables for network devices in the server rack are properly tagged.', 'f': 'Simplifies identification, maintenance, and troubleshooting of network devices.', 'h': 'Maintain proper tagging of all network cables.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Wire tagging was not available for the cables of Network Devices in the server rack.', 'd': 'Wire tagging missing for network device cables in server rack.', 'f': 'Difficult to identify cables, manage, or maintain network devices; increases risk of errors during maintenance.', 'h': 'Tag all cables properly for all networking devices.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        28: {  # cctvCablesConcealed
            'compliance': {'a': 'Compliance', 'b': 'CCTV cables properly concealed.', 'd': 'Concealed cables reduce risk of tampering or accidental damage.', 'f': 'Ensures CCTV reliability and continuous monitoring.', 'h': 'Maintain concealed cabling.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'CCTV camera cables were not properly concealed.', 'd': 'CCTV cables exposed and improperly routed.', 'f': 'Risk of cutting/damaging cables, loss of video evidence during forensic investigations.', 'h': 'Properly conceal CCTV camera cables.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        29: {  # cctvBackup30Days
            'compliance': {'a': 'Compliance', 'b': '30 days of CCTV footage backup available.', 'd': 'CCTV backups maintained for minimum required period per NABARD guidelines.', 'f': 'Enables forensic investigations and audit trail review.', 'h': 'Continue maintaining at least 30 days of backup.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': '30 days CCTV Footage backup was not available.', 'd': 'CCTV backup maintained for less than 30 days.', 'f': 'Forensic investigation may be hindered if incident occurs beyond backup period.', 'h': 'Maintain at least 30 days of CCTV footage backup.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        30: {  # fireExtinguishersCount
            'compliance': {'a': 'Compliance', 'b': 'Adequate fire extinguishers present.', 'd': 'Fire extinguishers available in sufficient quantity in DR/server room.', 'f': 'Reduces fire hazards and equipment damage risk.', 'h': 'Maintain sufficient fire extinguishers and check regularly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Fire extinguisher was not available.', 'd': 'It was observed that the Fire extinguisher was not available. A server room fire can be one of the worst disasters that can befall a business. ', 'f': 'Unable to control fire; high risk of data and hardware loss, personal injury, and operational downtime.', 'h': 'Install fire extinguishers in server/DR site.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        31: {  # automaticFireExtinguisher
            'compliance': {'a': 'Compliance', 'b': 'Automatic fire extinguisher installed.', 'd': 'Rapid fire response ensured even during non-banking hours.', 'f': 'Protects servers and DR site from fire damage.', 'h': 'Continue maintaining automatic fire extinguishers.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The automatic fire extinguisher was not present in the server room. ', 'd': 'It was observed that there was no Automatic fire extinguisher present in the Disaster Recovery. At night if the data recovery site is caught on fire, and no one is present at the bank, it will destroy the whole disaster recovery site.', 'f': 'Fire during unattended hours could destroy entire DR site.', 'h': 'Install automatic fire extinguishers in DR/server room.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        32: {  # hygrometerPresent
            'compliance': {'a': 'Compliance', 'b': 'Hygrometer installed and monitored.', 'd': 'Humidity levels maintained between 45–55% to prevent static or condensation damage.', 'f': 'Reduces risk of corrosion, electrical malfunction, and equipment damage.', 'h': 'Continue monitoring humidity with a hygrometer.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A hygrometer was not present in the server room.', 'd': 'It was observed that the Hygrometer was not present in the Disaster Recovery Site.', 'f': 'Humidity unmonitored; risk of condensation, static discharge, and damage to critical equipment.', 'h': 'Install hygrometer and maintain relative humidity 45–55%.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        33: {  # smokeDetectorAvailable
            'compliance': {'a': 'Compliance', 'b': 'Smoke detector installed.', 'd': 'Provides early fire detection and warning.', 'f': 'Enhances server room safety and reduces damage risk.', 'h': 'Maintain smoke detector functionality.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The smoke Detector was not available.', 'd': 'It was observed that the Smoke Detector was not available.', 'f': 'Fire may go undetected; increased risk of severe damage and downtime.', 'h': 'Install smoke detector in server/DR room.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        34: {  # fireAlarmPresent
            'compliance': {'a': 'Compliance', 'b': 'Fire alarm installed.', 'd': 'Fire alarms alert personnel during incidents.', 'f': 'Enables timely evacuation and fire mitigation.', 'h': 'Maintain fire alarm functionality and testing.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Fire alarm was not present in the disaster recovery site.', 'd': 'It was observed that the fire alarm was not present in the DR room. In case of a fire event in the room, the bank will not be able to identify and prevent it.', 'f': 'Fire may not be identified quickly, causing major damage and data loss.', 'h': 'Install fire alarm system in DR/server room.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        35: {  # antivirusUpdated
            'compliance': {'a': 'Compliance', 'b': 'Antivirus updated regularly.', 'd': 'Antivirus definitions and signatures are up-to-date on all servers.', 'f': 'Ensures protection against latest malware, ransomware, and viruses.', 'h': 'Continue regular updates of antivirus software.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Antivirus was not updated.', 'd': 'It was observed that Antivirus was not updated.', 'f': 'Servers are vulnerable to newly created viruses, malware, and cyber-attacks.', 'h': 'Regularly update antivirus to include latest virus signatures and patches.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        36: {  # unauthorizedSoftwarePresent
            'compliance': {'a': 'Compliance', 'b': 'No unauthorized software installed.', 'd': 'Only approved software is installed on servers.', 'f': 'Reduces risk of data breaches and security vulnerabilities.', 'h': 'Maintain strict software authorization policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Unauthorized software was present.', 'd': 'It was observed that many of the unauthorized software were present like VLC media player, Team viewer 7, Outdated version of Google Chrome web browser, Any desk, Skype, ArcSoft PhotoStudio, Crazy Browser 3.1.0.', 'f': 'Unauthorized software can create backdoors, data leakage, and increase cybersecurity risks.', 'h': 'Uninstall unauthorized software and restrict installations.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        37: {  # remoteAccessByITTeam
            'compliance': {'a': 'Compliance', 'b': 'Remote access responsibility defined.', 'd': 'Access logs and user responsibilities are maintained and auditable.', 'f': 'Clear accountability for all remote actions on servers.', 'h': 'Continue monitoring and controlling remote access.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Responsibility for Remote Access to the servers cannot be defined.', 'd': 'It was observed that the Responsibility for Remote Access to the servers cannot be defined. Anyone in the IT department can get remote access to any server.', 'f': 'No accountability; cannot trace actions to a specific user; increases risk of unauthorized activities.', 'h': 'Create separate user accounts for all server access; define responsibilities clearly.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        38: {  # dcAndDRLocation
            'compliance': {'a': 'Compliance', 'b': 'DC and DR located in separate seismic zones.', 'd': 'Disaster Recovery and Data Center are geographically separated.', 'f': 'Ensures availability of services in case of disaster.', 'h': 'Maintain separate locations for DC and DR.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'DC and DR were present in the same location.', 'd': 'It was observed that DR and DR were in the same location.', 'f': 'Single-point disaster may impact both primary and recovery systems.', 'h': 'Place DR in a different seismic zone to ensure operational continuity.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        39: {  # drDrillLogsMaintained
            'compliance': {'a': 'Compliance', 'b': 'DR drill logs properly maintained.', 'd': 'All records of DR drills including success/failure are maintained.', 'f': 'Helps track issues and measure drill effectiveness.', 'h': 'Continue maintaining detailed DR drill logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Logs for DR drills were not maintained properly.', 'd': 'It was observed that logs for DR drills were not properly maintained.', 'f': 'Difficult to identify issues; cannot evaluate DR drill success; audit trail missing.', 'h': 'Maintain complete logs for all DR drills with dates, actions, and outcomes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        40: {  # serverRoomClean
            'compliance': {'a': 'Compliance', 'b': 'Server room clean and dust-free.', 'd': 'Regular cleaning ensures optimal equipment operation.', 'f': 'Reduces risk of hardware failure due to dust accumulation.', 'h': 'Maintain cleanliness regularly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The server room was not neat and clean.', 'd': 'It was observed that the Server room was not neat and clean to ensure a dust-free environment. ', 'f': 'Dust buildup can slow cooling, damage equipment, and increase maintenance issues.', 'h': 'Ensure regular cleaning of server room to maintain dust-free environment.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        41: {  # serverRoomAccessLog
            'compliance': {'a': 'Compliance', 'b': 'Access logs maintained.', 'd': 'Registers of all server room entries and exits are maintained.', 'f': 'Enables traceability and accountability for physical access.', 'h': 'Continue maintaining access logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The server room access log was not registered, not present, and therefore not maintained.', 'd': 'No register or logs for server room access.', 'f': 'Cannot track who accessed the server; increases risk of insider threat.', 'h': 'Register and maintain server room access logs consistently.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        42: {  # networkDevicesPlacedProperly
            'compliance': {'a': 'Compliance', 'b': 'Network devices properly arranged.', 'd': 'All network equipment organized in server rack, secured, and accessible.', 'f': 'Simplifies management, maintenance, and reduces accidental disconnections.', 'h': 'Maintain proper placement of all network devices.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Network devices were not placed properly in a server rack.', 'd': 'Devices not properly placed in server racks; racks unlocked.', 'f': 'Easier physical tampering; poor cable management; maintenance challenges.', 'h': 'Place devices properly, lock server racks, assign key responsibility.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        43: {  # properCoolingPresent
            'compliance': {'a': 'Compliance', 'b': 'Proper cooling maintained.', 'd': 'Adequate A/Cs with automatic temperature control are installed.', 'f': 'Maintains optimal temperature for servers, prevents overheating, and ensures uninterrupted operations.', 'h': 'Continue monitoring temperature and maintaining proper cooling systems.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Proper cooling was not present in the server room', 'd': 'Only one A/C present with no automatic ON/OFF control; improper cooling in server room.', 'f': 'Overheating may cause server shutdown, data loss, or hardware failure.', 'h': 'Install sufficient A/C units with automatic temperature control for proper cooling.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        44: {  # acCount
            'compliance': {'a': 'Compliance', 'b': 'Adequate number of A/Cs installed.', 'd': 'Number of A/Cs sufficient to maintain optimal server room temperature.', 'f': 'Prevents temperature-related hardware failure and downtime.', 'h': 'Ensure proper number of A/Cs for load and area.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Insufficient A/Cs were present in the server room', 'd': 'Only one A/C present in the server room.', 'f': 'Overheating risk; servers may shut down under high temperature conditions.', 'h': 'Install additional A/Cs as per server room requirements.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        45: {  # acAutomaticManual
            'compliance': {'a': 'Compliance', 'b': 'Automatic ON/OFF control implemented.', 'd': 'A/C system set to automatically regulate temperature based on server room conditions.', 'f': 'Reduces risk of overheating and ensures energy efficiency.', 'h': 'Continue automatic temperature regulation.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The A/C controls are operated manually.', 'd': 'A/Cs switched manually; no automation present.', 'f': 'High risk of temperature fluctuations; potential server damage.', 'h': 'Implement automatic A/C control to maintain constant temperature.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        46: {  # wallLeakagePresent
            'compliance': {'a': 'Compliance', 'b': 'No wall leakage detected.', 'd': 'Server room walls intact, no visible damage or leakage.', 'f': 'Reduces risk of water damage, electrical hazards, or equipment failure.', 'h': 'Continue regular inspections for wall integrity and leaks.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The server room walls had leakage.', 'd': 'It was observed that  Server room walls had leakage.', 'f': 'These threats can damage equipment, force hardware to shut down, and slow performance. One of the most frightening dangers that a water leak can lead to is a fire. If leaking water reaches bank electrical supply or loose wires, it can cause your electricity to short circuit and create a spark which can then ignite a fire.', 'h': 'Take preventive measures to repair leaks and protect servers.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        47: {  # powerBackupPresent
            'compliance': {'a': 'Compliance', 'b': 'UPS/Power backup available.', 'd': 'Adequate power backup installed to support critical server operations during outages.', 'f': 'Ensures uninterrupted operations and protects against data loss.', 'h': 'Regularly maintain and test power backup systems.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Power backup was not available.', 'd': 'It was observed that power backup was not available.  Disaster Recovery site in banking and finance require robust contingency power systems to match the size of their operations to retain power for critical operations under adverse conditions.', 'f': 'The  Disaster Recovery constitutes the critical load in the daily operations of the organization and plays the role of the backbone of financial data computation, transaction system, and records storage. Servers, storages, routers, and switches are the key IT equipment in the disaster recovery site. As power backup was not available, the cost of downtime due to critical load failure will be extreme.', 'h': 'Install high-capacity UPS or alternative power backup solutions for disaster recovery site.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
         }
    }
     
     # Assign the comprehensive response data
    question_responses = DISASTER_RECOVERY_RESPONSES
     
     # Default responses for questions not yet fully defined
    default_responses = {
         'compliance': {'a': 'Compliance', 'b': 'Compliance observed.', 'd': 'System meets requirements.', 'f': 'Risk mitigated.', 'h': 'Continue current practices.'},
         'non_compliance': {'a': 'Non-Compliance', 'b': 'Non-compliance observed.', 'd': 'System does not meet requirements.', 'f': 'Risk identified.', 'h': 'Immediate action required.'},
         'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
     }
     
     # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 50
    
    # Header row
    headers = ['Sr. No.', 'Questionnaire/ Points', 'Compliance/Non-Compliance/Not Applicable', 
               'Observation (Short/Brief)', 'Risk Factor', 'Observation', 'Impact', 'Recommendation']
    
    # Apply header formatting
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Populate Sr. No. and Questions
    for i, question in enumerate(questions, 2):
        ws.cell(row=i, column=1, value=i-1)  # Sr. No.
        ws.cell(row=i, column=2, value=question)  # Questions
    
    # Process form data and populate responses
    for field_name, question_num in question_mapping.items():
        if form_data and field_name in form_data:
            user_input = form_data[field_name]
            row = question_num + 1  # +1 because row 1 is header
            
            if user_input == 'Not Applicable':
                # For Not Applicable, populate all columns with "Not Applicable"
                ws.cell(row=row, column=3, value='Not Applicable')
                ws.cell(row=row, column=4, value='Not Applicable')
                ws.cell(row=row, column=6, value='Not Applicable')
                ws.cell(row=row, column=7, value='Not Applicable')
                ws.cell(row=row, column=8, value='Not Applicable')
            else:
                # Get response data based on user input
                response_data = question_responses.get(question_num, default_responses)
                
                if user_input == 'Compliance':
                    compliance_data = response_data.get('compliance', default_responses['compliance'])
                    ws.cell(row=row, column=3, value=compliance_data.get('a', 'Compliance'))
                    ws.cell(row=row, column=4, value=compliance_data.get('b', ''))
                    ws.cell(row=row, column=6, value=compliance_data.get('d', ''))
                    ws.cell(row=row, column=7, value=compliance_data.get('f', ''))
                    ws.cell(row=row, column=8, value=compliance_data.get('h', 'No recommendation required.'))
                elif user_input == 'Non-Compliance':
                    non_compliance_data = response_data.get('non_compliance', default_responses['non_compliance'])
                    ws.cell(row=row, column=3, value=non_compliance_data.get('a', 'Non-Compliance'))
                    ws.cell(row=row, column=4, value=non_compliance_data.get('b', ''))
                    ws.cell(row=row, column=6, value=non_compliance_data.get('d', ''))
                    ws.cell(row=row, column=7, value=non_compliance_data.get('f', ''))
                    ws.cell(row=row, column=8, value=non_compliance_data.get('h', 'Immediate action required.'))
    
    # Apply alignment to all data cells
    for row in range(2, 49):  # Rows 2-48 for data
        for col in range(1, 9):  # Columns A-H
            cell = ws.cell(row=row, column=col)
            if col in [2, 4, 6, 7, 8]:  # Columns B, D, F, G, H - left horizontal, center vertical (middle align)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            elif col == 3:  # Column C - center alignment
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:  # Other columns (A, E) - center alignment
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Populate Risk Factor column (E) with color coding
    for i, risk_factor in enumerate(risk_factors, 2):
        cell = ws.cell(row=i, column=5, value=risk_factor)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)  # White text, bold
        if risk_factor == 'Critical':
            cell.fill = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')  # Dark Red
        elif risk_factor == 'High':
            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
        elif risk_factor == 'Medium':
            cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # Orange
        elif risk_factor == 'Low':
            cell.fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')  # Green
    
    # Apply borders to all cells with content
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for row in range(1, 49):  # Header + 47 data rows (1-48)
        for col in range(1, 9):  # Columns A-H
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
    
    # Set row height for wrapped text
    for row in range(2, 49):  # Data rows 2-48
        ws.row_dimensions[row].height = 30
    
    # Save the file
    output_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'static', 'uploads')
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate filename
    filename = "Disaster Recovery Review.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # Save the workbook
    wb.save(filepath)
    
    print(f"Excel file created: {filepath}")
    print(f"Filename: {filename}")
    
    return filepath, filename

def cleanup_file(filepath):
    """
    Delete the file from the uploads folder after download
    """
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            print(f"File cleaned up: {filepath}")
    except Exception as e:
        print(f"Error cleaning up file {filepath}: {str(e)}")

if __name__ == "__main__":
    # Test the function
    create_disaster_recovery_excel()
