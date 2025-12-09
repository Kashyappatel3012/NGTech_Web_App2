import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_remote_access_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Remote Access"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # Remote Access Questions
    questions = [
        "Whether remote access logs are maintained and reviewed periodically?",
        "Is remote logon through services such as FTP, Telnet disabled?",
        "Whether remote access is disabled when not in use?",
        "Are proper approvals taken before granting exceptional remote access to users?"
    ]

    # Risk Factors
    risk_factors = [
        "High", "High", "Medium", "Medium"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "raCppdGuidelines": 1,
        "raInsecureServicesDisabled": 2,
        "raUserIdentifiable": 3,
        "raExceptionalApproval": 4
    }

    # Question responses data
    question_responses = {
        1: {  # raCppdGuidelines
            'compliance': {'a': 'Compliance', 'b': 'Remote logon complies with guidelines.', 'd': 'Remote access is strictly controlled, implemented, and monitored as per CPPD/IT Department guidelines, ensuring secure connections.', 'f': 'Reduces the risk of unauthorized access and ensures compliance with organizational IT security policies.', 'h': 'Periodically review remote access policies and monitor logs to maintain secure access.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Logs are not generated.', 'd': 'Logs are not generated for remote access activities', 'f': 'In the absence of log generation, it becomes impossible to trace user activities or detect unauthorized access attempts. This creates a significant gap in accountability and incident response, increasing the risk of undetected security breaches or misuse of remote access.', 'h': 'The bank should configure remote access systems to generate and securely store detailed logs of all user activities, including successful and failed login attempts. These logs should be periodically reviewed and integrated with the SIEM tool for real-time monitoring and alerting.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # raInsecureServicesDisabled
            'compliance': {'a': 'Compliance', 'b': 'Insecure remote services disabled.', 'd': 'FTP, Telnet, and other insecure remote login methods are disabled, with secure protocols enforced for remote access.', 'f': 'Prevents exposure of credentials and protects data from interception during remote access.', 'h': 'Regularly review and audit remote services to ensure no insecure access channels exist.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The remote logon through services such as FTP, telnet was not disabled.', 'd': 'Remote logins using insecure services like FTP and Telnet are still enabled, exposing credentials and data to interception.', 'f': 'Credential information (usernames and passwords) submitted through telnet is not encrypted and is therefore vulnerable to identity theft. Ftp is vulnerable to Packet Capture/Sniffing, Brute Force Attacks, and Port Stealing.', 'h': 'Disable insecure remote services and replace them with secure alternatives like SFTP, SCP, or SSH.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # raUserIdentifiable
            'compliance': {'a': 'Compliance', 'b': 'Remote users identifiable.', 'd': 'All remote logins capture terminal IDs and IP addresses, enabling traceability and accountability for remote activities.', 'f': 'Enhances security monitoring, supports incident investigation, and ensures compliance with access control policies.', 'h': 'Periodically review remote access logs and correlate with user activities to detect anomalies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Remote access is not disabled when not in use.', 'd': 'Remote logins do not capture or log terminal IDs or IP addresses, making it difficult to trace user activities.', 'f': 'If remote access is not disabled when not in use, inactive or unnecessary accounts remain exposed to external threats. This increases the risk of unauthorized access, credential misuse, or exploitation by malicious actors, potentially leading to data breaches or compromise of critical banking systems.', 'h': 'It is recommended that Remote Desktop Protocol (RDP) be disabled on all systems when not in use. Remote access should be enabled only for authorized users based on specific requirements and approved by the competent authority.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # raExceptionalApproval
            'compliance': {'a': 'Compliance', 'b': 'Exceptional remote access properly approved.', 'd': 'All exceptions for remote access are approved by authorized personnel, documented, and reviewed periodically for relevance.', 'f': 'Controls the risk of unauthorized access while allowing necessary operational flexibility.', 'h': 'Maintain records of approvals and review periodically to validate the need for continued access.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Proper approval were not taken before granting Exceptional Remote Access to the users.', 'd': 'Users are provided remote access beyond standard permissions without documented authorization from appropriate authorities.', 'f': 'Granting remote access to users without proper approval can increase the risk of security breaches. If remote access is granted to users who should not have it, they may be able to access sensitive data or systems and compromise security.', 'h': 'Ensure all exceptional remote access is formally approved, documented, and periodically reviewed for necessity.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Remote Access Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
