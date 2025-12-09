import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_domain_controller_excel(form_data=None):
    """
    Create Excel file for Domain Controller (AD) Assessment
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Domain Controller (AD)"
    
    # Define questions
    questions = [
        "Are scheduled tasks set to Disabled for server operators on the domain controller?",
        "Is Domain controller: LDAP server signing requirements set to 'Require signing' or not?",
        "Is Digitally encrypt or sign secure channel data set to Enabled or not?",
        "Is the AD server patched regularly?",
        "Is Domain member Required strong (Windows 2000 or later) session key set to Enabled?",
        "Is the Minimum password length set to 14 or more characters?",
        "Is there any validation to follow the standard password policy?",
        "Are regular VAPT and third-party audits done for the AD server?",
        "Is the Password of the domain administrator/controller user complex and never shared with anyone?",
        "Is domain credential caching disabled?",
        "Is the password displayed in encrypted mode?",
        "Are Access computers from the network configured?",
        "Is 'Act as part of the operating system' set to 'No One'?",
        "Is Back up files and directories set to Administrators/ IT team only or not?",
        "For Interactive logon: Is Domain Controller Authentication required to unlock the workstation?",
        "Is Microsoft network client: Digitally sign communications set to Enabled?",
        "Is there only one domain administrator/controller user in the list?",
        "Is Network access: Allow anonymous SID/Name translation set to Disabled or not?",
        "Are event audit logs enabled on the server?",
        "Is Network security: Allow Local System to use computer identity for NTLM set to Enabled?",
        "Is the vssadmin tool disabled or restricted on the AD server?",
        "Whether Access control is Configured or not for the AD server?",
        "Is Network security: Allow Local System NULL session fallback set to Disabled?",
        "Is Interactive logon text Message configured for users attempting to log on?",
        "Can the local accounts and guest users log in through remote log-on protocols?",
        "Is Generated security audit set to LOCAL SERVICE, NETWORK SERVICE?"
    ]
    
    # Risk factors for each question
    risk_factors = [
        'Critical', 'Critical', 'Critical', 'High', 'High', 'High', 'High', 'High', 'High', 'High',
        'High', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium',
        'Medium', 'Medium', 'Medium', 'Low', 'Low', 'Low'
    ]
    
    # Question mapping for form fields
    question_mapping = [
        'scheduledTasks', 'ldapSigning', 'secureChannelEncryption', 'adServerPatched', 'strongSessionKey',
        'passwordLength', 'passwordValidation', 'vaptAudits', 'adminPasswordComplex', 'credentialCaching',
        'passwordEncrypted', 'networkAccess', 'actAsOS', 'backupPermissions', 'domainAuthUnlock',
        'smbSigning', 'singleAdmin', 'anonymousSID', 'auditLogs', 'ntlmComputerIdentity',
        'vssadminRestricted', 'accessControl', 'nullSessionFallback', 'logonMessage', 'remoteLogonLocal',
        'securityAuditGeneration'
    ]
    
    # Set column widths
    ws.column_dimensions['A'].width = 10  # Sr. No.
    ws.column_dimensions['B'].width = 50  # Questions
    ws.column_dimensions['C'].width = 20  # Compliance/Non-Compliance/Not Applicable
    ws.column_dimensions['D'].width = 30  # Observation (Short/Brief)
    ws.column_dimensions['E'].width = 20  # Risk Factor
    ws.column_dimensions['F'].width = 50  # Observation
    ws.column_dimensions['G'].width = 50  # Impact
    ws.column_dimensions['H'].width = 50  # Recommendation
    
    # Header row
    headers = ['Sr. No.', 'Questionnaire/Points', 'Compliance/Non-Compliance/Not Applicable', 
               'Observation (Short/Brief)', 'Risk Factor', 'Observation', 'Impact', 'Recommendation']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Populate Sr. No. and Questions
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }
    
    # Question responses data
    question_responses = {
        1: {  # scheduledTasks
            'compliance': {'a': 'Compliance', 'b': 'Scheduled tasks are disabled for server operators on the domain controller.', 'd': 'Server Operators are restricted from scheduling automated tasks, ensuring that only authorized administrators can modify or run scheduled activities on the domain controller.', 'f': 'Prevents unauthorized changes or execution of scripts that may impact the server\'s stability or security.', 'h': 'Continue to keep scheduled tasks disabled for Server Operators through Group Policy configuration.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Domain controller allows Server Operators to schedule tasks.', 'd': 'It was found that the Domain controller Allows server operators to schedule tasks. This policy setting determines whether members of the Server Operators group are allowed to submit jobs utilizing the AT schedule facility. An AT Service Account can be modified to select a different account rather than the LOCAL SYSTEM account.', 'f': 'Unauthorized scheduled tasks could run with elevated privileges, potentially compromising system security. Misconfigured AT Service Accounts could lead to privilege escalation or malicious code execution.', 'h': r'''It is recommended that scheduled tasks should be set to Disabled for server operators on the domain controller. To establish the recommended configuration via GP, set the following UI path to Disabled:-
Computer Configuration\Policies\Windows Settings\Security Settings\Local Policies\Security Options\Domain controller: Allow server operators to schedule task  '''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # ldapSigning
            'compliance': {'a': 'Compliance', 'b': 'LDAP server signing requirements are set to \'Require signing\'.', 'd': 'The domain controller enforces LDAP signing, ensuring that all communication between clients and servers is authenticated and protected from tampering.', 'f': 'Protects directory data from replay and man-in-the-middle attacks.', 'h': 'Maintain LDAP signing configuration to ensure secure communication across the domain network.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'LDAP server signing requirements were not enabled.', 'd': 'It was observed that the Domain controller: LDAP server signing requirements were not set to Require signing on the LDAP server. This policy setting determines whether the Lightweight Directory Access Protocol (LDAP) server requires LDAP clients to negotiate data signing.', 'f': 'Unsigned network traffic is vulnerable to replay and man-in-the-middle attacks, allowing attackers to impersonate users or modify LDAP queries to gain unauthorized access.', 'h': r'''It is recommended to set Domain controller: LDAP server signing requirement to 'Require signing'. To establish the recommended configuration via GP, set the following UI path to Require signing:-
Computer Configuration\Policies\Windows Settings\Security Settings\Local Policies\Security Options\Domain controller: LDAP server signing requirements '''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # secureChannelEncryption
            'compliance': {'a': 'Compliance', 'b': 'Secure channel data encryption is enabled.', 'd': 'Secure channel traffic between domain members and domain controllers is encrypted, ensuring confidentiality and authenticity of transmitted data.', 'f': 'Prevents unauthorized interception and data tampering over the network.', 'h': 'Continue to enforce encryption and signing for all secure channel data.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Secure channel encryption/signing was disabled.', 'd': 'It was observed that Digitally encrypting or signing secure channel data was set to disable. This policy setting determines whether all secure channel traffic that is initiated by the domain member must be signed or encrypted. If a system is set to always encrypt or sign secure channel data, it cannot establish a secure channel with a domain controller that is not capable of signing or encrypting all secure channel traffic, because all secure channel data must be signed and encrypted.', 'f': 'Attackers can intercept or modify unencrypted traffic to gain sensitive information or credentials, leading to domain compromise.', 'h': r'''It is recommended to enable Digitally encrypt or sign secure channel data. To establish the recommended configuration via GP, set the following UI path to Enabled:-
Computer Configuration\Policies\Windows Settings\Security Settings\Local Policies\Security Options\Domain member: Digitally encrypt or sign secure channel data (always)'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # adServerPatched
            'compliance': {'a': 'Compliance', 'b': 'AD server is regularly patched.', 'd': 'Regular security updates are applied to the AD server to protect against vulnerabilities and ensure software stability.', 'f': 'Reduces risk of exploitation through known vulnerabilities.', 'h': 'Maintain monthly patch management and compliance tracking.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'AD server is not patched regularly.', 'd': 'The patch is a set of changes to a computer program or its supporting data designed to update, fix, or improve it. This includes fixing security vulnerabilities and other bugs, with such patches usually being called bug fixes, the patch improves the functionality and usability performance of the program. It was observed that the AD server was not patched regularly.', 'f': 'Running outdated software increases the risk of compromise through known exploits, leading to data breaches or domain-wide attacks.', 'h': 'It is recommended to patch regularly and check for security updates regularly.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # strongSessionKey
            'compliance': {'a': 'Compliance', 'b': 'Strong session key is enabled.', 'd': 'Secure channel communications use 128-bit encryption, ensuring data integrity and protection against interception.', 'f': 'Enhances domain authentication security and prevents session hijacking.', 'h': 'Continue to enforce strong session key encryption for domain members.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Strong session key is disabled.', 'd': 'It was observed that the domain member strong(Windows 2000 or later) session key was disabled. When this policy setting is enabled, a secure channel can only be established with domain controllers that are capable of encrypting secure channel data with a strong (128-bit) session key. To enable this policy setting, all domain controllers in the domain must be able to encrypt secure channel data with a strong key. ', 'f': 'Attackers may intercept or alter communications, leading to credential theft or unauthorized data access.', 'h': r'''Session keys that are used to establish secure channel communications between domain controllers and member computers are much stronger starting with Windows 2000. So, it is recommended to enable the domain member strong session key. To establish the recommended configuration via GP, set the following UI path to Enabled: 
Computer Configuration\Policies\Windows Settings\Security Settings\Local Policies\Security Options\Domain member: Require strong (Windows 2000 or later) session key'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # passwordLength
            'compliance': {'a': 'Compliance', 'b': 'Minimum password length is set to 14 characters.', 'd': 'Password policy enforces a minimum of 14 characters, reducing the likelihood of successful brute-force attacks.', 'f': 'Strengthens user authentication and password resilience.', 'h': 'Maintain current configuration to align with security best practices.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Password length was less than 14 characters.', 'd': 'This policy setting determines the least number of characters that make up a password for a user account. The ideal value for the Minimum password length setting is 14 characters, however, the bank can adjust this value to meet the business requirements. It was observed that the password length was less than 14 characters.', 'f': 'Weak passwords can be cracked easily, enabling attackers to gain unauthorized access to user accounts and sensitive systems.', 'h': 'It is recommended to teach users about passphrases, which are often easier to remember and, due to the larger number of character combinations, much harder to discover. Also, the minimum password length should be set to 14 or more characters.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # passwordValidation
            'compliance': {'a': 'Compliance', 'b': 'Strong password validation is enforced.', 'd': 'Password creation requires users to include uppercase, lowercase, numeric, and special characters to ensure strength.', 'f': 'Prevents weak password creation and mitigates credential-based attacks.', 'h': 'Continue enforcing validation checks for strong password creation.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Strong password validation was not enforced.', 'd': 'It was observed that a strong password policy was not maintained and validation to obtain a stronger password was not followed. Users need to set the password as per validation by including alphanumeric, special characters, and a combination of the upper and lower case to follow the standard password policy.', 'f': 'Weak passwords can be brute-forced or guessed easily, leading to unauthorized access, credential theft, and potential domain compromise.', 'h': 'It is recommended that Strong passwords are used for authentication and use standard validation which allows only strong passwords.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # vaptAudits
            'compliance': {'a': 'Compliance', 'b': 'Regular VAPT and third-party audits are performed.', 'd': 'The AD server undergoes periodic security assessments and external audits to identify vulnerabilities.', 'f': 'Improves overall AD security posture and ensures continuous compliance.', 'h': 'Continue scheduled VAPT and audit reviews quarterly or biannually.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'VAPT and third-party audits were not performed.', 'd': 'It was observed that VAPT and third-party audits are not performed on the Active Directory server. VAPT is a term often used to describe security testing that is designed to identify and help address cyber security vulnerabilities. This includes automated vulnerability assessments to human-led penetration testing and full-scale red team simulated cyber-attacks.', 'f': 'Unidentified vulnerabilities may be exploited by attackers, leading to privilege escalation or domain compromise.', 'h': 'It is recommended to perform VAPT and third-party audits on the Active Directory server to know the security posture of the AD server to protect against cyber attacks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # adminPasswordComplex
            'compliance': {'a': 'Compliance', 'b': 'Domain admin password is complex and confidential.', 'd': 'The password adheres to complexity standards and is not shared among employees, ensuring proper privilege separation.', 'f': 'Prevents unauthorized access and strengthens admin account protection.', 'h': 'Maintain strict password confidentiality and complexity enforcement.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Admin password was weak and shared.', 'd': 'It was observed that the password of the domain administrator/controller user was not complex, and it does not follow the standard password policy. ', 'f': 'Shared credentials increase the risk of misuse and unauthorized access, leading to data breaches or privilege escalation.', 'h': 'It  is  recommended to configure the Password Complexity option to a minimum of 12 characters combinations of letters, numbers, and characters and  never shared with anyone.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # credentialCaching
            'compliance': {'a': 'Compliance', 'b': 'Domain credential caching is disabled.', 'd': 'Systems are configured to require verification with the domain controller for each login, preventing reuse of cached credentials.', 'f': 'Reduces risk of unauthorized offline access to domain accounts.', 'h': 'Continue maintaining disabled credential caching on all systems.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Credential caching was not disabled.', 'd': r'''It was observed that credential caching was not disabled. 
A credential cache holds Kerberos credentials while they remain valid and, generally, while the user's session lasts, so that authenticating to service multiple times (e.g., connecting to a web or mail server more than once) doesn't require contacting the KDC every time.''', 'f': 'Attackers can use cached credentials to log in offline, bypassing authentication checks and compromising system integrity.', 'h': r'''It is recommended to disable the credential cache which holds the Kerberos key:
Configuration\Windows Settings\Security Settings\Local Policies\Security Options GPO container.'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # passwordEncrypted
            'compliance': {'a': 'Compliance', 'b': 'Password displayed in encrypted mode.', 'd': 'It was verified that all user and device passwords are masked or encrypted during configuration and transmission, preventing exposure to unauthorized individuals.', 'f': 'Protects credentials from interception or unauthorized access.', 'h': 'Continue to enforce encrypted password visibility across all management interfaces.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Password not displayed in encrypted mode.', 'd': 'The password displayed in encrypted mode is meant to prevent someone from looking over your shoulder and seeing the password. It was observed that the password was not displayed in encrypted mode.', 'f': 'Plain text passwords can be easily viewed or captured by unauthorized users, allowing them to gain administrative access and compromise critical systems or data integrity. This weakens the organization\'s security posture and exposes sensitive infrastructure.', 'h': 'It is recommended to secure the communication by implementing HTTPS.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # networkAccess
            'compliance': {'a': 'Compliance', 'b': 'All computers properly configured within the network domain.', 'd': 'It was confirmed that all systems are joined to the domain, restricting network access only to authenticated users and devices through centralized control.', 'f': 'Ensures controlled network access and reduces unauthorized entry points.', 'h': 'Continue to maintain centralized network access configuration for all connected systems.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Some computers not configured within the network domain.', 'd': 'This policy setting allows other users on the network to connect to the computer and is required by various network protocols that include Server Message Block (SMB)based protocols, NetBIOS, Common Internet File System (CIFS), and Component Object Model Plus (COM+). It was observed that some of the computers were not in the domain.', 'f': 'Non-domain computers can create weak links, leading to unauthorized access, data leakage, or malware propagation. Attackers can exploit these systems to pivot into the internal network.', 'h': r'''To establish the recommended configuration via GP, configure the following UI path: 
Computer Configuration\Policies\Windows Settings\Security Settings\Local Policies\User Rights Assignment\Access this computer from the network '''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # actAsOS
            'compliance': {'a': 'Compliance', 'b': '"Act as part of the operating system" privilege disabled for all users.', 'd': 'It was verified that no user accounts are granted the "Act as part of the operating system" privilege, ensuring restricted privilege escalation capabilities.', 'f': 'Prevents unauthorized processes from gaining full control over system resources.', 'h': 'Continue to restrict this privilege to maintain least privilege principles.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': '"Act as part of the operating system" privilege enabled.', 'd': 'This policy setting allows a process to assume the identity of any user and thus gain access to the resources that the user is authorized to access. It has system privileges. It was observed that acting as a part of the operating system was enabled.', 'f': 'This could enable attackers to gain complete control of the system, hide their actions, and manipulate critical security configurations, posing a severe threat to integrity and availability.', 'h': r'''It is recommended to not give this permission to administrators, when needed these privileges use the local system account for the task. To establish the recommended configuration via GP, set the following UI path to No One:- 
Computer Configuration\Policies\Windows Settings\Security Settings\Local Policies\User Rights Assignment\Act as part of the operating system  
'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # backupPermissions
            'compliance': {'a': 'Compliance', 'b': 'Backup permissions restricted to Administrators/IT team only.', 'd': 'It was verified that only authorized IT personnel have rights to back up or restore critical data, ensuring confidentiality and proper control of backups.', 'f': 'Prevents unauthorized data access or transfer during backup operations.', 'h': 'Continue to restrict backup rights to authorized administrators only.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Backup permissions not restricted to Administrators/IT team.', 'd': r'''This policy setting allows users to circumvent file and directory permissions to back up the system. This user right is enabled only when an application (such as NTBACKUP) attempts to access a file or directory through the NTFS file system backup application programming interface (API). Otherwise, the assigned file and directory permissions apply.  
It was observed that backup of files/directories was not configured to only Administrators/ IT team.''', 'f': 'Unauthorized backup rights can lead to data theft, manipulation, or exposure of confidential data outside the network, leading to operational and reputational risks.', 'h': r'''It is tecommended that Back up files and directories should be set to Administrators/ IT team only. To establish the recommended configuration via GP, set the following UI path to Administrators/ IT team. 
Computer Configuration\Policies\Windows Settings\Security Settings\Local Policies\User Rights Assignment\Back up files and directories'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # domainAuthUnlock
            'compliance': {'a': 'Compliance', 'b': 'Domain Controller Authentication required to unlock workstation.', 'd': 'It was verified that systems require domain controller authentication before unlocking, ensuring validation against the centralized Active Directory.', 'f': 'Prevents unauthorized access to locked systems using cached credentials.', 'h': 'Continue enforcing centralized authentication for workstation unlocks.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Domain Controller Authentication not required to unlock workstation.', 'd': 'It was observed that domain controller authentication was not required to unlock the workstation. This setting determines whether it is necessary to contact a domain controller to unlock a computer. If you enable this setting, a domain controller must authenticate the domain account that is being used to unlock the computer. If you disable this setting, logon information confirmation with a domain controller is not required for a user to unlock the computer. ', 'f': 'This increases the risk of unauthorized system access through disabled or compromised accounts, allowing malicious users to log in without validation, potentially exposing confidential information.', 'h': r'''It is recommended that, For Interactive logon: Domain Controller Authentication required to unlock the workstation should be enabled. To implement the recommended configuration via GP, set the following UI path to Enabled: 
Computer Configuration\Policies\Windows Settings\Security Settings\Local Policies\Security Options\Interactive logon: Require Domain Controller Authentication to unlock workstation'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # smbSigning
            'compliance': {'a': 'Compliance', 'b': 'Digitally sign communications option enabled.', 'd': 'It was verified that Microsoft network client digitally signs communications, ensuring secure and trusted connections between SMB clients and servers.', 'f': 'Protects data integrity and authenticity during network communication.', 'h': 'Continue to enforce SMB signing for all clients and servers to maintain secure data exchange.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Digitally sign communications option disabled.', 'd': 'It was observed that digitally sign communications options were disabled. This policy setting determines whether packet signing is required by the SMB client,server component.', 'f': 'Without digital signing, attackers can intercept and modify SMB packets, perform session hijacking, and tamper with data during transfer. This can lead to unauthorized access and compromise of the integrity and confidentiality of files.', 'h': 'Enable digital signing for all Microsoft network clients to ensure communication authenticity. Implement SMB signing or IPsec-based encryption to protect against tampering and replay attacks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # singleAdmin
            'compliance': {'a': 'Compliance', 'b': 'Multiple domain administrator accounts available for redundancy.', 'd': 'It was confirmed that there are at least two domain administrator accounts — one for daily administrative use and one for emergency access — ensuring continuity in case of account lockout or failure.', 'f': 'Ensures uninterrupted administrative operations and secure segregation of duties.', 'h': 'Continue to maintain multiple secure admin accounts with proper access control.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Only one domain administrator account found.', 'd': 'It was observed that the organization operates with a single domain administrator account. This creates a single point of failure for all administrative operations and system management tasks.', 'f': 'If the lone admin account becomes locked, compromised, or deleted, critical domain management and recovery operations would be halted, severely affecting system availability and security.', 'h': 'Create at least two domain administrator accounts — one for routine operations and another for emergency access — with strict privilege separation and auditing enabled.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        18: {  # anonymousSID
            'compliance': {'a': 'Compliance', 'b': 'Anonymous SID/Name translation disabled.', 'd': 'It was verified that anonymous SID-to-name translation requests are blocked, ensuring that only authenticated users can query sensitive SID or account name data.', 'f': 'Prevents attackers from enumerating privileged account names.', 'h': 'Maintain disabled anonymous SID/Name translation settings across all domain controllers.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Anonymous SID/Name translation enabled.', 'd': 'It was observed that anonymous SID/Name translation was not disabled.This policy setting determines whether an anonymous user can request security identifier (SID) attributes for another user, or use a SID to obtain its corresponding user name.', 'f': 'Attackers can use SID enumeration to identify high-value targets such as domain admin accounts, enabling password-guessing or brute-force attacks to gain elevated access. This compromises overall domain security.', 'h': r'''It is recommended to set Allow anonymous SID/Name translation  to Disabled. To establish the recommended configuration via GP, set the following UI path to Disabled: 
Computer Configuration\Policies\Windows Settings\Security Settings\Local Policies\Security Options\Network access: Allow anonymous SID/Name translation'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        19: {  # auditLogs
            'compliance': {'a': 'Compliance', 'b': 'Event audit logs enabled and monitored.', 'd': 'It was confirmed that event audit logging is active, capturing successful and failed logon attempts, configuration changes, and administrative actions for forensic and compliance purposes.', 'f': 'Supports incident detection, investigation, and compliance reporting.', 'h': 'Continue maintaining detailed audit logs with regular monitoring and secure retention.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Event audit logs were not enabled.', 'd': 'It was observed that event audit logs were not enabled. This log report contains data when a user attempted to log on or log off from the system. These events occur on the accessing computer. For interactive logins, the generation of these events occurs on the computer that is logged on. If a network logon takes place to access a share, these events generate on the computer that hosts the accessed resource.', 'f': 'Absence of audit logs hinders the detection of security incidents, unauthorized access, or malicious activity. It also limits post-incident investigation and weakens accountability within critical systems.', 'h': r'''It is recommended that event audit logs to be enabled on the server.To establish the recommended configuration via GP, set the following UI path to Success and Failure:
Computer Configuration\Policies\Windows Settings\Security Settings\Advanced Audit Policy Configuration\Audit Policies\Logon/Logoff\Audit Logon'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        20: {  # ntlmComputerIdentity
            'compliance': {'a': 'Compliance', 'b': 'Local System allowed to use computer identity for NTLM.', 'd': 'It was verified that the policy is enabled, allowing Local System services to use computer identity for NTLM authentication during network communication.', 'f': 'Ensures authenticated communication between system services and domain resources.', 'h': 'Continue enforcing NTLM computer identity authentication for Local System services.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Local System not allowed to use computer identity for NTLM.', 'd': 'It was observed that Allow Local System to use computer identity for NTLM set to disabled. When enabled, this policy setting causes Local System services that use Negotiate to use the computer identity when NTLM authentication is selected by the negotiation. This policy is supported on at least Windows 7 or Windows Server 2008 R2.', 'f': 'Anonymous authentication weakens network trust boundaries, enabling potential impersonation or data interception attacks. This may lead to unauthorized access or privilege escalation.', 'h': r'''it is recommended to Allow Local System to use computer identity for NTLM should be set to enabled for network security.To establish the recommended configuration via GP, set the following UI path to Enabled: 
Computer Configuration\Policies\Windows Settings\Security Settings\Local Policies\Security Options\Network security: Allow Local System to use computer identity for NTLM'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        21: {  # vssadminRestricted
            'compliance': {'a': 'Compliance', 'b': 'vssadmin is disabled or access to it is restricted.', 'd': 'Use of vssadmin.exe is restricted to authorized administrative accounts and/or the executable is renamed/blocked, preventing routine misuse.', 'f': 'Limits attacker ability to delete shadow copies and reduces ransomware impact.', 'h': 'Continue to restrict or control vssadmin usage and monitor any attempts to run it.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The vssadmin tool was used for the shadow copy of drives.', 'd': "It was observed that vssadmin tool was used for the shadow copy of drives. Vssadmin is a default Windows process that manipulates volume's shadow copies of a given computer. These shadow copies are often used as backups, and they can be used to restore or revert files to a previous state if they are corrupted or lost for some reason. ", 'f': 'Attackers or ransomware can delete all shadow copies quickly (e.g., vssadmin delete shadows /all /quiet), removing local recovery options and forcing reliance on offsite backups — greatly increasing recovery time and potential data loss. Successful misuse can escalate an incident from recoverable to catastrophic.', 'h': 'Restrict vssadmin to a very small set of trusted admin accounts (or rename/deny execute via application control), log and alert all executions, and ensure off-server backup copies exist so that shadow-copy deletion cannot prevent recovery.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        22: {  # accessControl
            'compliance': {'a': 'Compliance', 'b': 'ACLs configured for Active Directory objects.', 'd': 'Access Control Lists are applied to AD objects to permit/deny access by role, minimizing unnecessary access to sensitive directory objects.', 'f': 'Prevents unauthorized reads/changes to AD objects; supports least-privilege access.', 'h': 'Maintain ACL hygiene and periodically review permissions.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'ACLs not configured or weakly applied on AD objects.', 'd': 'It was observed that ACL was not configured for the Ad server. These ACLs are used for access rules to permit and deny access to objects residing in the active directory. ACL is Access Control List which has data about objects and who can access those objects.', 'f': 'Broad or missing ACLs enable unauthorized users to read or change sensitive AD objects (groups, service accounts, trusts), facilitating stealthy privilege escalation, lateral movement, and data exfiltration. Recovery and containment become more complex after compromise.', 'h': 'Implement least-privilege ACLs across AD, perform a permissions audit to remove excessive rights, document who needs access and why, and enforce changes via controlled requests and reviews.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        23: {  # nullSessionFallback
            'compliance': {'a': 'Compliance', 'b': 'NULL session fallback disabled; Local System uses computer identity for NTLM.', 'd': 'Systems authenticate using the computer identity and do not fall back to NULL sessions, ensuring secure signing/encryption.', 'f': 'Reduces risk of anonymous/weakly-protected sessions and data exposure.', 'h': 'Keep NULL fallback disabled and ensure NTLM computer identity is enforced.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'NULL session fallback is enabled; Local System allows fallback.', 'd': 'It was observed that the Local system allows NULL session fallback for encrypting and signing the application data that provides no protection.', 'f': 'NULL session fallback permits session keys that provide no cryptographic protection, exposing sensitive traffic to interception, spoofing, and manipulation. Attackers can exploit this to tamper with signed data, weaken authentication, or exfiltrate information without raising obvious alarms.', 'h': 'Disable NULL session fallback and ensure Local System uses the machine identity for NTLM; test services for compatibility and update any legacy applications that rely on NULL sessions.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        24: {  # logonMessage
            'compliance': {'a': 'Compliance', 'b': 'Interactive logon message configured and displayed.', 'd': 'A login message is presented to users at logon, communicating accepted use and security notice.', 'f': 'Reinforces policy and provides legal/awareness notice prior to authentication.', 'h': 'Keep and periodically review the logon message.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'No interactive logon message configured.', 'd': r'''It was observed that no text message was configured at the time of login. This policy setting specifies a text message that displays to users when they log on. 
Any warning that you display should first be approved by your organization's legal and human resources representatives.''', 'f': 'Absence of a logon warning reduces user awareness of monitoring and acceptable use, and may weaken legal standing in incident investigations; attackers receive no deterrent message and users may be unaware of expected behavior.', 'h': r'''It is recommended to enable Interactive logon text Message configured for users attempting to log on. To establish the recommended configuration via GP, configure the following UI path:
Computer Configuration\Policies\Windows Settings\Security Settings\Local Policies\Security Options\Interactive logon: Message text for users attempting to log on'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        25: {  # remoteLogonLocal
            'compliance': {'a': 'Compliance', 'b': 'Remote logon denied for local/guest accounts.', 'd': 'Policies deny remote logon rights to local and guest accounts; only authorized domain accounts can access RDS/remote console.', 'f': 'Reduces attack vectors and remote account misuse.', 'h': 'Continue to restrict remote logon to authorized administrative/group accounts only.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Local and guest accounts allowed remote logon; no deny policy present.', 'd': 'It was observed that there was no configuration to deny the remote logon of the users and local accounts and guest users can log in through remote log-on prortocols.', 'f': 'Allowing local/guest remote logons significantly increases the risk of brute-force or credential-sharing attacks, enabling unauthorized remote console access. Attackers can use weaker or shared local credentials to gain footholds, move laterally, and deploy malware, complicating detection and response.', 'h': r'''It is recommended that local accounts and guest users should not be allowed to log in through remote log-on protocols. To establish the recommended configuration via Group Policy, set the following UI path to include Guests, Local accounts: 
Computer Configuration\Policies\Windows Settings\Security Settings\Local Policies\User Rights Assignment\Deny log on through Remote Desktop Services '''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        26: {  # securityAuditGeneration
            'compliance': {'a': 'Compliance', 'b': 'Security audit generation restricted to Local Service and Network Service.', 'd': 'Only expected service accounts are permitted to generate security audits, preventing misuse by arbitrary accounts.', 'f': 'Protects integrity of Security log and audit trail.', 'h': 'Maintain this restricted setting and monitor for changes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Generated security audit rights are granted broadly (not limited to Local/Network Service).', 'd': 'It was observed that other services except for local service, and network services have the right to generate audit logs. This policy setting determines which users or processes can generate audit records in the Security log. By default, only network service and local service can generate security audit logs.', 'f': 'Broad audit generation rights enable attackers or malicious software to flood or corrupt the Security log, erase evidence, or mask intrusions by generating excessive or misleading entries; this undermines incident detection and forensic analysis and may cause log-driven failures.', 'h': 'Restrict audit generation rights to Local Service and Network Service only, remove unnecessary accounts from this privilege, and monitor Security log write activity; implement log integrity and centralized log collection to detect tampering.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Apply formatting to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for col in range(1, 9):  # A to H
        cell = ws.cell(row=1, column=col)
        cell.border = thin_border
    
    # Process each question
    for i, question in enumerate(question_mapping, 2):  # Start from row 2
        # Get user input for this question
        user_input = form_data.get(question, 'not_applicable') if form_data else 'not_applicable'
        
        # Get response data
        response_data = question_responses.get(i-1, {})
        user_response = response_data.get(user_input, {})
        
        # Populate columns based on user input
        # Column C: Compliance/Non-Compliance/Not Applicable
        ws.cell(row=i, column=3, value=user_response.get('a', 'Not Applicable'))
        
        # Column D: Observation (Short/Brief)
        ws.cell(row=i, column=4, value=user_response.get('b', 'Not Applicable'))
        
        # Column F: Observation
        ws.cell(row=i, column=6, value=user_response.get('d', 'Not Applicable'))
        
        # Column G: Impact
        ws.cell(row=i, column=7, value=user_response.get('f', 'Not Applicable'))
        
        # Column H: Recommendation
        ws.cell(row=i, column=8, value=user_response.get('h', 'Not Applicable'))
        
        # Column E: Risk Factor with color coding
        risk_factor = risk_factors[i-2]  # risk_factors is 0-indexed
        risk_cell = ws.cell(row=i, column=5, value=risk_factor)
        risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
        risk_cell.fill = PatternFill(start_color=risk_colors.get(risk_factor, '808080'), end_color=risk_colors.get(risk_factor, '808080'), fill_type='solid')
        risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply formatting to all data rows
    for row in range(2, len(question_mapping) + 2):  # Rows 2 to 27
        for col in range(1, 9):  # Columns A to H
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Special alignment for column A (Sr. No.), C (Compliance status), and E (Risk Factor)
            if col == 1 or col == 3 or col == 5:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Set row height for wrapped text
            ws.row_dimensions[row].height = 30
    
    # Save file
    filename = "Domain Controller AD Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    """
    Delete the generated Excel file after download
    """
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            print(f"File {filepath} deleted successfully")
    except Exception as e:
        print(f"Error deleting file {filepath}: {e}")
