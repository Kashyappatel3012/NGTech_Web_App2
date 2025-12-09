import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_ho_win_server_excel(form_data=None):
    """
    Create Excel file for HO Win_Server Logical Review Assessment
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "HO Win_Server Logical Review"
    
    # Define questions
    questions = [
        "Is Antivirus installed on the server?",
        "Are physical controls established for Antivirus Server?",
        "Is the Internet allowed on servers?",
        "Is 'Storing passwords using reversible encryption' set to 'Disabled'?",
        "Is 'Account lockout duration' set to '15 or more minutes'?",
        "Is 'Accounts: Administrator account status' set to 'Disabled'?",
        "Is 'Do not allow password expiration time longer than required by policy' set to 'Enabled'?",
        "Is 'MSS: (AutoAdminLogon) Enable Automatic Logon' set to 'Disabled'?",
        "Is 'Windows Firewall: Domain: Outbound connections' set to 'Allow (default)'?",
        "Is 'Allow anonymous SID/Name translation' set to 'Disabled'?",
        "Is 'Enforce password history' set to '24 or more passwords'?",
        "Is 'Password must meet complexity requirements' set to 'Enabled'?",
        "Is the 'Minimum password age' set to '1 or more day(s)'?",
        "Is all data between client and server encrypted using AES or other techniques?",
        "Is the server ported with the latest versions of patches and service pack?",
        "Is the server up to date?",
        "Are Weak and Guessable passwords in use?",
        "Are Administrator Rights in use?",
        "Whether USB Access is disabled?",
        "Is the Authentication policy defined?",
        "Is the restore files and directories policy defined?",
        "Is 'Access Credential Manager as a trusted caller' set to 'No One'?",
        "Is 'Allow log on locally' configured to authorized users only?",
        "Is 'Allow log on through Remote Desktop Services' configured for authorized users only?",
        "Is 'Accounts: Limit local account use of blank passwords to console logon only' set to 'Enabled'?",
        "Is 'Firewall: Domain: Logging: Log successful connections' set to 'Yes'?",
        "Is there a separate room for the server?",
        "Is Server have adequate space for operational requirements?",
        "Is the Server room visible from a distance, but not easily accessible?",
        "Is the Server room away from the basement, and water/drainage systems?",
        "Is (Disable IP Source Routing) 'IP source routing protection level (protects against packet spoofing)' set to 'Enabled'?",
        "Is IPv6 disabled?",
        "Is 'Turn off background refresh of Group Policy' set to 'Disabled'?",
        "Is 'Prevent enabling lock screen camera' set to 'Enabled'?",
        "Is 'Audit Logon' set to 'Success and Failure'?",
        "Is 'Switch to the secure desktop when prompting for elevation' set to 'Enabled'?",
        "Is 'Run all administrators in Admin Approval Mode' set to 'Enabled'?",
        "Is 'Force logoff when logon hours expire' set to 'Enabled'?",
        "Is 'Interactive logon: Prompt user to change the password before expiration' set to 'between 5 and 14 days'?",
        "Is 'Interactive logon: Do not display last user name' set to 'Enabled'?",
        "Is 'Load and unload device drivers' set to 'Administrators'?",
        "Is 'Debug programs' set to 'Administrators' only'?"
    ]
    
    # Risk factors for each question
    risk_factors = [
        'Critical', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High',
        'High', 'High', 'High', 'High', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium',
        'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium',
        'Medium', 'Low', 'Low', 'Low', 'Low', 'Low', 'Low', 'Low', 'Low', 'Low', 'Low', 'Low'
    ]
    
    # Mapping of form field names to question numbers
    question_mapping = {
        "antivirusInstalled": 1,
        "physicalControlsAntivirus": 2,
        "internetAllowedServers": 3,
        "reversibleEncryptionDisabled": 4,
        "accountLockoutDuration": 5,
        "administratorAccountDisabled": 6,
        "passwordExpirationPolicy": 7,
        "autoAdminLogonDisabled": 8,
        "firewallOutboundConnections": 9,
        "anonymousSIDTranslationDisabled": 10,
        "enforcePasswordHistory": 11,
        "passwordComplexityEnabled": 12,
        "minimumPasswordAge": 13,
        "dataEncryptionAES": 14,
        "serverPatched": 15,
        "serverUpToDate": 16,
        "weakPasswordsInUse": 17,
        "administratorRightsInUse": 18,
        "usbAccessDisabled": 19,
        "authenticationPolicyDefined": 20,
        "restoreFilesPolicyDefined": 21,
        "credentialManagerAccess": 22,
        "allowLogonLocally": 23,
        "allowLogonRDS": 24,
        "blankPasswordsConsoleOnly": 25,
        "firewallLoggingEnabled": 26,
        "separateServerRoom": 27,
        "adequateServerSpace": 28,
        "serverRoomVisible": 29,
        "serverRoomAwayFromWater": 30,
        "ipSourceRoutingEnabled": 31,
        "ipv6Disabled": 32,
        "groupPolicyRefreshDisabled": 33,
        "lockScreenCameraEnabled": 34,
        "auditLogonEnabled": 35,
        "secureDesktopElevation": 36,
        "adminApprovalModeEnabled": 37,
        "forceLogoffEnabled": 38,
        "passwordExpirationPrompt": 39,
        "hideLastUsername": 40,
        "loadUnloadDrivers": 41,
        "debugPrograms": 42
    }
    
    # Set column widths
    ws.column_dimensions['A'].width = 10  # Sr. No.
    ws.column_dimensions['B'].width = 50  # Questions
    ws.column_dimensions['C'].width = 20  # Compliance/Non-Compliance/Not Applicable
    ws.column_dimensions['D'].width = 30  # Observation (Short/Brief)
    ws.column_dimensions['E'].width = 20  # Risk Factor
    ws.column_dimensions['F'].width = 50  # Observation
    ws.column_dimensions['G'].width = 50  # Impact
    ws.column_dimensions['H'].width = 50  # Recommendation
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header row
    headers = ['Sr. No.', 'Questionnaire/Points', 'Compliance/Non-Compliance/Not Applicable', 
               'Observation (Short/Brief)', 'Risk Factor', 'Observation', 'Impact', 'Recommendation']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Populate Sr. No. and Questions
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }
    
    # Question responses data
    question_responses = {
        1: {  # antivirusInstalled
            'compliance': {'a': 'Compliance', 'b': 'Antivirus installed and running.', 'd': 'Antivirus is installed on the server and actively running.', 'f': 'Protects servers from malware, trojans, ransomware, and other attacks, ensuring confidentiality, integrity, and availability of critical bank systems.', 'h': 'Ensure antivirus definitions are updated regularly and periodic scans are scheduled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Antivirus was not installed on the server. ', 'd': 'Antivirus was not installed on the server.', 'f': 'As the Antivirus is not installed, it is not possible to protect against various malicious activities like malware, virus, etc. Servers are considered a critical asset of the bank network so the unavailability of these systems is a major threat to the organization.', 'h': 'It is recommended to install Antivirus on all the servers.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # physicalControlsAntivirus
            'compliance': {'a': 'Compliance', 'b': 'Physical controls established.', 'd': 'Physical access to the Antivirus server is restricted to authorized personnel.', 'f': 'Reduces the risk of unauthorized tampering and improves overall server security.', 'h': 'Maintain physical access logs and periodically review access permissions.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Physical controls were not established for Antivirus Server.', 'd': 'Physical controls were not established for Antivirus Server.', 'f': 'As the Antivirus server is openly accessible and used for other services any user can modify the Antivirus configuration. A malicious user could enable removable media or modify settings, compromising security.', 'h': 'It is recommended to use the dedicated system for Antivirus, and no other services should be running on this server. It should be physically secure so that only authorized persons can access it.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # internetAllowedServers
            'compliance': {'a': 'Compliance', 'b': 'Internet restricted on servers.', 'd': 'Internet access is restricted on servers hosting critical banking applications.', 'f': 'Reduces exposure to external threats and prevents malware infections on critical infrastructure.', 'h': 'Regularly review firewall and network rules to maintain server isolation.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Internet was allowed on servers.', 'd': 'It was observed that the Internet was allowed on servers. Servers are the most critical part as far as infrastructure is concerned. If the internet runs on the server on which applications are hosted, it will become more accessible to attackers.', 'f': 'If the internet is running on the server on which applications like CBS and CTS are hosted, it might be possible that viruses, warms, key loggers, and ransomware can be installed on the server, which will directly affect the whole network and can compromise the CIA triad.', 'h': 'It is recommended to separate the server and restrict the internet on that servers. As per RBI guidelines, the internet should not be running on the servers on which applications are hosted.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # reversibleEncryptionDisabled
            'compliance': {'a': 'Compliance', 'b': 'Reversible encryption disabled.', 'd': 'Passwords are stored using irreversible encryption on the server.', 'f': 'Prevents attackers from easily decrypting stored passwords and protects user credentials.', 'h': 'Periodically review password storage policies to ensure encryption remains secure.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Storing of passwords using reversible encryption' was set to  'Enabled'.", 'd': 'Storing of passwords using reversible encryption was set to \'Enabled\'.', 'f': 'If the attacker gets hold of the encryption key, passwords can be decrypted easily, risking account compromise and sensitive data exposure.', 'h': '''It is recommended to Establish the configuration via GP, and set the following UI path to Disabled:-
Computer Configuration\\Policies\\Windows Settings\\Security Settings\\Account Policies\\Password Policy\\Store passwords using reversible encryption  
'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # accountLockoutDuration
            'compliance': {'a': 'Compliance', 'b': 'Account lockout duration set properly.', 'd': '\'Account lockout duration\' is configured to 15 or more minutes.', 'f': 'Reduces the risk of brute-force attacks and prevents unauthorized access.', 'h': 'Periodically test lockout policy to ensure effectiveness and compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Account lockout duration' was not set to '15 or more minutes'.", 'd': '\'Account lockout duration\' was not set to \'15 or more minutes\'.', 'f': 'Without proper lockout duration, attackers can perform brute-force attacks to guess passwords and gain unauthorized access to user accounts.', 'h': '''It is recommended that you Establish configuration via GP, and set the following UI path to 15 or more minutes:-
Computer Configuration\\Policies\\Windows Settings\\Security Settings\\Account Policies\\Account Lockout Policy\\Account lockout duration '''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # administratorAccountDisabled
            'compliance': {'a': 'Compliance', 'b': 'Administrator account disabled.', 'd': 'Administrator account status is set to \'Disabled\'.', 'f': 'Prevents unauthorized or unintended use of administrative privileges, reducing risk of insider attacks and system compromise.', 'h': 'Periodically review administrative accounts to ensure no unauthorized accounts are enabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Administrator account status' was not set to 'Disabled'.", 'd': 'Administrator account status was not set to \'Disabled\', which allows all administrative rights to the user.', 'f': 'Users with administrative rights can install malware, spyware, or trojans and change system configurations, compromising the CIA triad of information security.', 'h': '''It is recommended that 'Accounts: Administrator account status set to Disable to protect the server from misuse by the insider threat.'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # passwordExpirationPolicy
            'compliance': {'a': 'Compliance', 'b': 'Password expiration enforced.', 'd': 'Password expiration time is enforced as per policy.', 'f': 'Reduces risk of password compromise and limits exposure to brute-force or credential theft attacks.', 'h': 'Ensure users are notified before password expiry and enforce regular password updates.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Do not allow password expiration time longer than required by policy' was set to 'Disabled'.", 'd': '\'Do not allow password expiration time longer than required by policy\' was set to \'Disabled\'.', 'f': 'If passwords are used longer than intended, leaked credentials can be exploited to access accounts, compromising server security.', 'h': '''It is recommended to set 'Do not allow password expiration time longer than required by policy' to 'Enabled'.'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # autoAdminLogonDisabled
            'compliance': {'a': 'Compliance', 'b': 'Automatic logon disabled.', 'd': '\'MSS: (AutoAdminLogon) Enable Automatic Logon\' is set to \'Disabled\'.', 'f': 'Requires explicit authentication for every login, reducing risk of unauthorized access via physical or remote attacks.', 'h': 'Periodically audit logon policies to ensure automatic logon remains disabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "MSS: (AutoAdminLogon) Enable Automatic Logon was set to 'Enabled'.", 'd': '\'MSS: (AutoAdminLogon) Enable Automatic Logon\' was set to \'Enabled\'.', 'f': 'Physical access to the server allows anyone to log in automatically. Passwords are stored in plaintext in the registry, exposing the system to attackers.', 'h': "It is recommended to set MSS: (AutoAdminLogon) Enable Automatic Logon set to 'Disabled'."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # firewallOutboundConnections
            'compliance': {'a': 'Compliance', 'b': 'Outbound connections set to allow.', 'd': '\'Windows Firewall: Domain: Outbound connections\' is configured to \'Allow (default)\'.', 'f': 'Ensures normal network operation while maintaining firewall security controls.', 'h': 'Review firewall rules periodically to avoid misconfigurations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Windows Firewall Domain Outbound connections' was not set to 'Allow (default)'.", 'd': '\'Windows Firewall: Domain: Outbound connections\' was not set to \'Allow (default)\'.', 'f': 'Blocking outbound connections can lead to constant user interruptions and provides minimal security; attackers can reconfigure firewalls if compromised.', 'h': "It is recommended that 'Windows Firewall: Domain: Outbound connections'  should be set to 'Allow (default)."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # anonymousSIDTranslationDisabled
            'compliance': {'a': 'Compliance', 'b': 'Anonymous SID/Name translation disabled.', 'd': '\'Allow anonymous SID/Name translation\' is set to \'Disabled\'.', 'f': 'Prevents enumeration of user accounts by unauthorized users and reduces attack surface.', 'h': 'Periodically audit this setting and ensure all systems comply with the policy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Allow anonymous SID/Name translation' was set to 'Enabled'.", 'd': '\'Allow anonymous SID/Name translation\' was set to \'Enabled\'.', 'f': 'Local users can discover Administrator account names and initiate password guessing attacks, increasing risk of account compromise.', 'h': "It is recommended to disable 'Allow anonymous SID/Name translation'  setting to prevent unauthenticated users from obtaining user names that are associated with their respective SIDs."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # enforcePasswordHistory
            'compliance': {'a': 'Compliance', 'b': 'Password history enforced.', 'd': '\'Enforce password history\' is configured to 24 or more passwords.', 'f': 'Prevents reuse of old passwords, reducing risk of credential-based attacks.', 'h': 'Periodically audit password history settings for compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Enforce password history' was not set to '24 or more passwords.", 'd': '\'Enforce password history\' was not set to \'24 or more passwords\'.', 'f': 'If password history is not enabled then the user can reuse their old passwords again and again which is very dangerous. If the credentials somehow get leaked then the attacker can use that credentials to access the user account to compromise the security of the server.', 'h': '''It is recommeded to set 'Enforce password history'  to '24 or more passwords. Establish the configuration via GP and set the following UI path to 24 or more passwords:-
Computer Configuration\\Policies\\Windows Settings\\Security Settings\\Account Policies\\Password Policy\\Enforce password history '''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # passwordComplexityEnabled
            'compliance': {'a': 'Compliance', 'b': 'Password complexity enabled.', 'd': '\'Password must meet complexity requirements\' is set to \'Enabled\'.', 'f': 'Ensures strong passwords are used, mitigating brute-force and guessing attacks.', 'h': 'Regularly monitor password policies to enforce complexity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Password must meet complexity requirements' was set to 'Disabled'.", 'd': '\'Password must meet complexity requirements\' was set to \'Disabled\'.', 'f': 'If the users use weak and easy-to-remember passwords then it will be easy for an attacker to brute force or guess the passwords to compromise the user account. ', 'h': '''It is recommended to establish the recommended configuration via GP, set the following UI path to Enabled:-              
Computer Configuration\\Policies\\Windows Settings\\Security Settings\\Account Policies\\Password Policy\\Password must meet complexity requirements '''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # minimumPasswordAge
            'compliance': {'a': 'Compliance', 'b': 'Minimum password age set.', 'd': '\'Minimum password age\' is configured to 1 or more day(s).', 'f': 'Prevents rapid password changes that could bypass password history, improving security.', 'h': 'Audit periodically to ensure minimum password age policy is enforced.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Minimum password age' was not set to '1 or more day(s)'. ", 'd': '\'Minimum password age\' was not set to \'1 or more day(s)\'.', 'f': 'Users can repeatedly change passwords to bypass password history policy, enabling attackers to reuse old passwords.', 'h': '''It is recommended to establish the recommended configuration via GP, and set the following UI path to 1 or more day(s):-
Computer Configuration\\Policies\\Windows Settings\\Security Settings\\Account Policies\\Password Policy\\Minimum password age'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # dataEncryptionAES
            'compliance': {'a': 'Compliance', 'b': 'Data encrypted.', 'd': 'All data between client and server is encrypted using AES or other secure techniques.', 'f': 'Protects integrity and confidentiality of information in transit.', 'h': 'Regularly review encryption protocols and certificates to maintain security.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'All data between client and server was not encrypted using AES or other techniques.', 'd': 'All data between client and server was not encrypted using AES or other techniques.', 'f': 'Attackers can intercept and modify data in transit, compromising confidentiality and integrity.', 'h': 'It is recommended that all data between client and server is encrypted using AES or other techniques.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # serverPatched
            'compliance': {'a': 'Compliance', 'b': 'Server fully patched.', 'd': 'The server is updated with the latest patches and service packs.', 'f': 'Reduces risk from known vulnerabilities and improves system stability.', 'h': 'Implement automated patch management and periodic verification.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The server was not ported with the latest versions of patches and service packs. ', 'd': 'The server was not ported with the latest versions of patches and service packs.', 'f': 'Unpatched vulnerabilities can be exploited by attackers to compromise server security and performance.', 'h': 'It is recommended that the server should be ported with the latest versions of patches and service packs.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # serverUpToDate
            'compliance': {'a': 'Compliance', 'b': 'Server up to date.', 'd': 'The server is updated with the latest patches and service packs.', 'f': 'Ensures that known vulnerabilities are mitigated and reduces the risk of exploitation by attackers.', 'h': 'Continuously monitor patch releases and verify updates are applied in a timely manner.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The server was not up to date. ', 'd': 'The server was not up to date.', 'f': 'If the server is not up to date, then the existing vulnerabilities of the server can be used by an attacker to compromise the security of the server.', 'h': 'It is recommended that the server should be up to date. Enable the Auto Update option, or create a centralized patch update system such as WSUS. '},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # weakPasswordsInUse
            'compliance': {'a': 'Compliance', 'b': 'Strong passwords in use.', 'd': 'All accounts are configured with strong passwords in line with the password policy.', 'f': 'Reduces the risk of brute force attacks and unauthorized access.', 'h': 'Periodically audit password strength and enforce password complexity policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Weak and Guessable passwords were in use.', 'd': 'Weak and Guessable passwords were in use.', 'f': 'As weak passwords are being used, the risk of brute force attack is high. Any person with malicious intent can access systems by guessing or brute-forcing the password and can modify configurations to disable security controls.', 'h': 'It is recommended to use strong passwords defined by the Global Password Policy. Strong passwords have alphanumeric characters, symbols, and digits and minimum length of 8 characters.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        18: {  # administratorRightsInUse
            'compliance': {'a': 'Compliance', 'b': 'Admin rights controlled.', 'd': 'Administrative privileges are limited and granted on a need-to-know basis.', 'f': 'Minimizes risk of malicious activity and accidental system misconfiguration.', 'h': 'Maintain a privileged access management system and periodically review admin accounts.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': ' Administrator rights were in use.', 'd': 'Administrator rights were in use without defined privileged access.', 'f': 'Users with unrestricted admin rights can install malicious software, change system configurations, and compromise the CIA triad.', 'h': '''It is recommended to create a standard user on the server so that for requirements of any changes it should ask the administrator's password and the administrator can validate the activity. '''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        19: {  # usbAccessDisabled
            'compliance': {'a': 'Compliance', 'b': 'USB Access disabled.', 'd': 'USB access is disabled on all critical servers.', 'f': 'Prevents introduction of malware and unauthorized data transfers through removable media.', 'h': 'Regularly audit server settings to ensure USB ports remain disabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'USB Access was enabled.', 'd': 'USB Access was enabled on the server.', 'f': 'Malicious users can inject viruses, malware, or steal sensitive data via USB devices, compromising the confidentiality and integrity of data.', 'h': 'It is recommended to disable USB Access.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        20: {  # authenticationPolicyDefined
            'compliance': {'a': 'Compliance', 'b': 'Authentication policy defined.', 'd': 'Authentication policy is implemented and enforced across all servers.', 'f': 'Ensures that only authorized users can access critical resources, maintaining the confidentiality, integrity, and availability of systems.', 'h': 'Review and update authentication policies periodically to address emerging threats.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Authentication policy was not defined.', 'd': 'The Authentication policy was not defined.', 'f': 'Weak authentication and authorization can be exploited to gain unauthorized access to sensitive information and compromise operations.', 'h': 'It is recommended to use define Authentication policy and role-based authorization to restrict unauthorized access.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        21: {  # restoreFilesPolicyDefined
            'compliance': {'a': 'Compliance', 'b': 'Restore files policy defined.', 'd': 'Restore files and directories policy is defined and implemented.', 'f': 'Ensures timely recovery of critical data in case of system failure or malicious activity.', 'h': 'Periodically test backup and restoration procedures to ensure effectiveness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Restore files and directories policy was not defined.', 'd': 'Restore files and directories policy was not defined.', 'f': 'Backup should be taken for files stored in bank systems. If critical data is lost, it will be difficult to retrieve it. If an attacker compromises the system, the bank may be unable to recover critical data, compromising the CIA triad.', 'h': 'It is recommended that the bank clearly define roles for backup and a data restoration policy for files and directories.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        22: {  # credentialManagerAccess
            'compliance': {'a': 'Compliance', 'b': 'Access Credential Manager restricted.', 'd': '\'Access Credential Manager as a trusted caller\' is set to \'No One\'.', 'f': 'Prevents unauthorized access to saved credentials, enhancing system security.', 'h': 'Periodically review user rights to ensure this setting is maintained.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Access Credential Manager as a trusted caller' was not set to 'No One'.", 'd': '\'Access Credential Manager as a trusted caller\' was not set to \'No One\'.', 'f': 'Users\' saved credentials might be compromised if this right is assigned to unauthorized entities, risking system security during backup and restore operations.', 'h': '''It is recommended to set 'Access Credential Manager as a trusted caller'  to 'No One'. Establish the recommended configuration via GP, set the following UI path to No One:-
Computer Configuration\\Policies\\Windows Settings\\Security Settings\\Local Policies\\User Rights Assignment\\Access Credential Manager as a trusted caller '''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        23: {  # allowLogonLocally
            'compliance': {'a': 'Compliance', 'b': 'Allow log on locally restricted.', 'd': '\'Allow log on locally\' is configured for authorized users only.', 'f': 'Reduces risk of unauthorized console access and privilege escalation attacks.', 'h': 'Periodically review local logon permissions.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Allow log on locally' was not configured to authorized users only. ", 'd': '\'Allow log on locally\' was not configured for authorized users only.', 'f': 'Unauthorized users could log on to the console and run malicious software to elevate privileges.', 'h': "It is recommended to Configure 'Allow log on locally' for authorized users only."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        24: {  # allowLogonRDS
            'compliance': {'a': 'Compliance', 'b': 'RDS logon restricted.', 'd': '\'Allow log on through Remote Desktop Services\' is configured for authorized users only.', 'f': 'Prevents unauthorized remote access and reduces risk of privilege escalation.', 'h': 'Regularly audit RDS logon permissions.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Allow log on through Remote Desktop Services'  was not configured for authorized users only.", 'd': '\'Allow log on through Remote Desktop Services\' was not configured for authorized users only.', 'f': 'Unauthorized users could remotely access the system and execute malicious activities or elevate privileges.', 'h': "It is recommended to configure 'Allow log on through Remote Desktop Services' for authorized users only. "},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        25: {  # blankPasswordsConsoleOnly
            'compliance': {'a': 'Compliance', 'b': 'Blank passwords restricted.', 'd': '\'Accounts: Limit local account use of blank passwords to console logon only\' is set to \'Enabled\'.', 'f': 'Reduces risk of unauthorized access through weak accounts.', 'h': 'Periodically verify that no local accounts have blank passwords.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Accounts: Limit local account use of blank passwords to console logon only' was set to 'Disabled' .", 'd': '\'Accounts: Limit local account use of blank passwords to console logon only\' was set to \'Disabled\'.', 'f': 'Attackers could exploit accounts with blank passwords to gain unauthorized access to critical systems.', 'h': "It is recommended to Enable 'Accounts: Limit local account use of blank passwords to console logon only' so that users cannot use blank passwords. "},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        26: {  # firewallLoggingEnabled
            'compliance': {'a': 'Compliance', 'b': 'Firewall logging enabled.', 'd': '\'Firewall Domain Logging: Log successful connections\' is set to \'Yes\'.', 'f': 'Logs provide visibility for troubleshooting and forensic investigations.', 'h': 'Regularly review firewall logs for unusual activity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Firewall Domain Logging: Log successful connections' was not set to 'Yes'.", 'd': '\'Firewall Domain Logging: Log successful connections\' was not set to \'Yes\'.', 'f': 'Successful connection events are not recorded, making it difficult to investigate system issues or unauthorized activities.', 'h': "It is recommended that Firewall Domain Logging Log successful connections' is set to 'Yes'. Use this option to log in when  Firewall with Advanced Security allows an inbound connection."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        27: {  # separateServerRoom
            'compliance': {'a': 'Compliance', 'b': 'Separate server room available.', 'd': 'Server room is dedicated, secured, and equipped with proper environmental controls.', 'f': 'Protects servers from environmental damage and maintains operational stability.', 'h': 'Maintain server room environmental monitoring and access control.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Separate room was not available for the server.', 'd': 'No separate room was available for the server.', 'f': 'Lack of dedicated space can lead to temperature, humidity, and dust issues, which may damage equipment and compromise data.', 'h': 'It is recommended to have a separate room for the server with a proper cooling mechanism.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        28: {  # adequateServerSpace
            'compliance': {'a': 'Compliance', 'b': 'Adequate server space available.', 'd': 'Server room has sufficient space for all operational requirements.', 'f': 'Ensures proper airflow, cooling, and maintenance access.', 'h': 'Periodically review server layout for optimal space management.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The server has no adequate space for operational requirements.', 'd': 'The server has no adequate space for operational requirements.', 'f': 'Lack of space increases dust accumulation, improper airflow, and overheating risk, which can degrade server performance.', 'h': 'It is recommended that the Server must have adequate space for operational requirements.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        29: {  # serverRoomVisible
            'compliance': {'a': 'Compliance', 'b': 'Server room secure and visible.', 'd': 'Server room is visible from a distance and access is restricted.', 'f': 'Reduces risk of unauthorized access and physical tampering.', 'h': 'Maintain access control logs and review periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Server Room was not visible from distance and it was easily accessible. ', 'd': 'Server room was not visible from a distance and easily accessible.', 'f': 'Unauthorized personnel can access servers, tamper with equipment, or steal confidential data, compromising the CIA triad.', 'h': 'It is recommended to keep the server room visible from a distance and accessible to authorized personnel only i.e., the IT department, and administrators.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        30: {  # serverRoomAwayFromWater
            'compliance': {'a': 'Compliance', 'b': 'Server room safe from water/drainage.', 'd': 'Server room is located on higher floors or safe areas away from water/drainage.', 'f': 'Protects servers from water damage and electrical hazards.', 'h': 'Regularly inspect server room location for environmental risks.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The server room was near the basement, and water/drainage systems.', 'd': 'Server room was near the basement and water/drainage systems.', 'f': 'Water ingress can damage servers and create electrical hazards, risking data loss and physical harm.', 'h': 'It is recommended that the Server room must be away from the basement, and water/drainage systems.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        31: {  # ipSourceRoutingEnabled
            'compliance': {'a': 'Compliance', 'b': 'IP source routing protection enabled.', 'd': 'IP source routing protection level is set to \'Enabled\'.', 'f': 'Protects network from packet spoofing and prevents attackers from specifying packet routes to bypass security.', 'h': 'Periodically verify source routing protection settings on all servers.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "IP source routing protection level (protect against packet spoofing)' was not set to 'Enabled'.", 'd': 'IP source routing protection level (protects against packet spoofing) was not set to \'Enabled\'.', 'f': 'Attackers could use source-routed packets to obscure their identity and location, potentially bypassing network security measures.', 'h': '''It is recommended to establish the configuration via GP, set the following UI path to Enabled: Highest protection, source routing is completely disabled:-

Computer Configuration\\Policies\\Administrative Templates\\MSS (Legacy)\\MSS: (DisableIPSourceRouting) IP source routing protection level (protects against packet spoofing)

Note: This Group Policy path does not exist by default. An additional Group Policy template (MSS-legacy.admx/adml) is required, it is included with Microsoft Security Compliance Manager (SCM), or available from this TechNet blog post:-

https://blogs.technet.microsoft.com/secguide/2016/10/02/the-mss-settings/  '''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        32: {  # ipv6Disabled
            'compliance': {'a': 'Compliance', 'b': 'IPv6 disabled.', 'd': 'IPv6 is disabled in internal networks.', 'f': 'Prevents potential IP conflicts and strengthens control over internal addressing.', 'h': 'Review network requirements before enabling IPv6.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'IPv6 was Enabled.', 'd': 'IPv6 was enabled.', 'f': 'Automatic IP assignments from IPv6 may cause conflicts in internal bank networks, especially if static IPs are used, potentially weakening internal network control.', 'h': 'It is recommended to disable IPv6 if it is not needed in the internal network of the bank if network is small in size.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        33: {  # groupPolicyRefreshDisabled
            'compliance': {'a': 'Compliance', 'b': 'Background refresh of GP controlled.', 'd': '\'Turn off background refresh of Group Policy\' is set to \'Enabled\'.', 'f': 'Prevents disruptive GP updates during user sessions and ensures stable system configuration.', 'h': 'Periodically verify that group policy refresh settings are maintained.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Turn off background refresh of Group Policy' was set to 'Disabled'.", 'd': '\'Turn off background refresh of Group Policy\' was set to \'Disabled\'.', 'f': 'Group policies update during user sessions, which can interrupt workflows or cause configuration conflicts.', 'h': '''It is recommended to set 'Turn off background refresh of Group Policy' to 'Enabled'. '''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        34: {  # lockScreenCameraEnabled
            'compliance': {'a': 'Compliance', 'b': 'Lock screen camera restricted.', 'd': '\'Prevent enabling lock screen camera\' is set to \'Enabled\'.', 'f': 'Prevents unauthorized camera access and protects sensitive information.', 'h': 'Ensure camera restrictions are enforced via GP.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Prevent enabling lock screen camera' was set to 'Disabled'. ", 'd': '\'Prevent enabling lock screen camera\' was set to \'Disabled\'.', 'f': 'Users can enable the lock screen camera, increasing the risk of unauthorized surveillance or data capture.', 'h': "It is recommended to set 'Prevent enabling lock screen camera' to 'Enabled'."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        35: {  # auditLogonEnabled
            'compliance': {'a': 'Compliance', 'b': 'Audit logon configured.', 'd': '\'Audit Logon\' is set to \'Success and Failure\'.', 'f': 'Ensures proper logging of all login attempts for security monitoring and forensic analysis.', 'h': 'Regularly review audit logs for anomalies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Audit Logon' was not set to 'Success and Failure'.", 'd': '\'Audit Logon\' was not set to \'Success and Failure\'.', 'f': 'Security incidents may go undetected or lack forensic evidence, reducing the ability to investigate and respond to attacks.', 'h': '''It is recommended that 'Audit Logon' should be set to 'Success and Failure'. '''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        36: {  # secureDesktopElevation
            'compliance': {'a': 'Compliance', 'b': 'Secure desktop enabled for elevation.', 'd': '\'Switch to the secure desktop when prompting for elevation\' is set to \'Enabled\'.', 'f': 'Prevents spoofing of elevation prompts and protects user credentials.', 'h': 'Periodically verify UAC settings on all servers.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Switch to the secure desktop when prompting for elevation' was set to 'Disabled'.", 'd': '\'Switch to the secure desktop when prompting for elevation\' was set to \'Disabled\'.', 'f': 'Elevation prompts can be spoofed, leading to credential theft or accidental approval of malicious actions.', 'h': '''It is recommended that 'Switch to the secure desktop when prompting for elevation' is set to 'Enabled'.'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        37: {  # adminApprovalModeEnabled
            'compliance': {'a': 'Compliance', 'b': 'Admin Approval Mode enabled.', 'd': '\'Run all administrators in Admin Approval Mode\' is set to \'Enabled\'.', 'f': 'Enforces UAC, reducing risk from malware or unauthorized administrative actions.', 'h': 'Regularly verify UAC configuration on servers.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Run all administrators in Admin Approval Mode' was set to 'Disabled'.", 'd': '\'Run all administrators in Admin Approval Mode\' was set to \'Disabled\'.', 'f': 'If this setting is disabled, UAC will not be used and any security benefits and risk mitigations that are dependent on UAC will not be present in the system.', 'h': '''It is recommended to configure the policy value for 
Computer Configuration -> Windows Settings -> Security Settings -> Local Policies -> Security Options -> 'User Account Control- Run all administrators in Admin Approval Mode' to 'Enabled'.'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        38: {  # forceLogoffEnabled
            'compliance': {'a': 'Compliance', 'b': 'Force logoff enabled.', 'd': '\'Force logoff when logon hours expire\' is set to \'Enabled\'.', 'f': 'Ensures users are logged off automatically, reducing unauthorized access risk.', 'h': 'Periodically review logon hour enforcement policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Force logoff when logon hours expire' was set to 'Disabled'.", 'd': '\'Force logoff when logon hours expire\' was set to \'Disabled\'.', 'f': 'Users may remain connected outside permitted hours, allowing attackers to exploit unattended sessions.', 'h': "It is recommended to enable 'Force logoff when logon hours expire'."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        39: {  # passwordExpirationPrompt
            'compliance': {'a': 'Compliance', 'b': 'Password expiration prompt configured.', 'd': '\'Interactive logon: Prompt user to change the password before expiration\' is set to \'between 5 and 14 days\'.', 'f': 'Users are notified in advance, preventing inadvertent lockouts.', 'h': 'Regularly check password expiration notifications.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Interactive logon: Prompt user to change the password before expiration' was not set to between 5 and 14 days.", 'd': '\'Interactive logon: Prompt user to change the password before expiration\' was not set to \'between 5 and 14 days\'.', 'f': 'Users may be locked out or unaware of expiring passwords, affecting productivity and access.', 'h': '''It is recommended that user passwords must be configured to expire periodically, and 'Interactive logon: Prompt user to change the password before expiration' should be set to 'between 5 and 14 days.'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        40: {  # hideLastUsername
            'compliance': {'a': 'Compliance', 'b': 'Last user name hidden.', 'd': '\'Interactive logon: Do not display last user name\' is set to \'Enabled\'.', 'f': 'Enhances security by preventing attackers from learning valid usernames.', 'h': 'Ensure setting is enforced via GP.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Interactive logon: Do not display last user name' was set to 'Disabled'.", 'd': '\'Interactive logon: Do not display last user name\' was set to \'Disabled\'.', 'f': 'Attackers can view the last logged-on user, increasing the risk of password guessing attacks.', 'h': "It is recommended that 'Interactive logon: Do not display last user name' is set to 'Enabled'."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        41: {  # loadUnloadDrivers
            'compliance': {'a': 'Compliance', 'b': 'Device driver rights restricted.', 'd': '\'Load and unload device drivers\' is set to Administrators.', 'f': 'Prevents unauthorized driver installation, protecting system stability.', 'h': 'Regularly audit user rights to ensure only administrators can load/unload drivers.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Load and unload device drivers' was not set to 'Administrators'. ", 'd': '\'Load and unload device drivers\' was not set to \'Administrators\'.', 'f': 'Unauthorized users could load malicious drivers, compromising system integrity and security.', 'h': '''It is recommended that configuration via GP, set the following UI path to Administrators:-
Computer Configuration\\Policies\\Windows Settings\\Security Settings\\Local Policies\\User Rights Assignment\\Load and unload device drivers  
'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        42: {  # debugPrograms
            'compliance': {'a': 'Compliance', 'b': 'Debug program rights restricted.', 'd': '\'Debug programs\' is set to Administrators only.', 'f': 'Prevents misuse of debugging tools and protects sensitive system information.', 'h': 'Audit user rights periodically to maintain proper restrictions.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': '''Debug programs' was not set to 'Administrators' only.''', 'd': '\'Debug programs\' was not set to \'Administrators\' only\'.', 'f': 'If you revoke this user right, no one will be able to debug programs. However, sometimes programs may contain sensitive information in errors and during the time of debugging, so if every user can debug the program then the attacker can use this sensitive information to attack the program.', 'h': "It is recommended that  'Debug programs' should be set to 'Administrators' only."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Populate data based on user input
    for i, question in enumerate(questions, 2):
        # Get user input for this question
        question_num = i - 1
        user_input = None
        
        if form_data:
            # Find the corresponding form field
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        # Get response data
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            # Populate columns C, D, F, G, H
            ws.cell(row=i, column=3, value=response_data['a'])  # Compliance/Non-Compliance/Not Applicable
            ws.cell(row=i, column=4, value=response_data['b'])  # Observation (Short/Brief)
            ws.cell(row=i, column=6, value=response_data['d'])  # Observation
            ws.cell(row=i, column=7, value=response_data['f'])  # Impact
            ws.cell(row=i, column=8, value=response_data['h'])  # Recommendation
            
            # Apply alignment to these columns
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        
        # Populate Risk Factor (column E) with color coding
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            # Apply color coding
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Only apply default font if it's not the Risk Factor column (column 5)
                if col != 5:  # Don't override Risk Factor column formatting
                    cell.font = Font(name='Calibri', size=11)
            # Set row height for wrapped text
            ws.row_dimensions[row].height = 30
    
    # Save file
    filename = "HO Win Server Logical Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    try:
        # Save the workbook
        wb.save(filepath)
        # Close the workbook to ensure it's properly saved
        wb.close()
        return filepath, filename
    except Exception as e:
        # Close workbook even if save fails
        try:
            wb.close()
        except:
            pass
        raise Exception(f"Error saving Excel file: {str(e)}")

def cleanup_file(filepath):
    """
    Delete the generated Excel file after download
    """
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            print(f"File {filepath} deleted successfully")
    except Exception as e:
        print(f"Error deleting file {filepath}: {e}")
