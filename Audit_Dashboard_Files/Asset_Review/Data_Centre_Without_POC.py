import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_data_centre_excel(form_data=None):
    """
    Create Data Centre Excel file with questions and responses
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Center Review"
    
    # Define questions for column B
    questions = [
        "Is there a separate room for the server?",
        "Is the Internet allowed or not?",
        "Which OS version is currently running?",
        "Are event logs for server OS present and up to date?",
        "Is Windows OS genuine or not?",
        "Is the currently running OS outdated or not?",
        "Is a password policy present for server OS authentication?",
        "Is antivirus available?",
        "Is server OS updated?",
        "How many CCTV cameras are present in the server room? (Recommended two)",
        "Is antivirus expired?",
        "Is antivirus configured properly?",
        "Are USB devices allowed or not?",
        "Are logs maintained for remote access?",
        "Which application is being used for remote connection?",
        "When was the last DR drill conducted?",
        "Does the server have adequate space for operational requirements?",
        "Is the server room visible from a distance and easily accessible?",
        "Is the server room away from the basement and water/drainage systems?",
        "Is the server close to the UPS room?",
        "Is smoking, eating, and drinking prohibited in the server room to prevent spillage of food or liquid into sensitive computer equipment?",
        "How many persons are authorized to access the server room?",
        "Is a biometric or proximity card machine present for access in the server room?",
        "Are biometric logs or proximity card logs present and maintained to the current date?",
        "Are cables for network devices properly structured or not?",
        "Is the server rack open or not?",
        "Is wire tagging available?",
        "Are cables for CCTV cameras concealed or not?",
        "Is CCTV footage backup available for at least 30 days or not?",
        "How many fire extinguishers are available?",
        "Is an automatic fire extinguisher present or not?",
        "Is a hygrometer present or not?",
        "Is a smoke detector available or not?",
        "Is a fire alarm present or not?",
        "Is antivirus updated or not?",
        "Is unauthorized software present or not?",
        "Is remote access taken by the IT team or any other?",
        "At which location is DR present?",
        "Are logs for DR drills properly maintained or not?",
        "Is the server room neat and clean to ensure a dust-free environment?",
        "Is a server room access log registered, present, and maintained?",
        "Are network devices placed properly in the server rack?",
        "Is the server room adequately cooled, and how is the air-conditioning system configured in terms of number of units and automatic or manual operation?",
        "Is there any leakage in server room walls or any other dangerous cause affecting servers in the room?",
        "Is power backup present or not?"
    ]
    
    # Define risk factors for column E
    risk_factors = [
        "Critical", "Critical", "High", "High", "High", "High", "High", "High", "High", "High",
        "High", "High", "High", "High", "High", "High", "High", "Medium", "Medium", "Medium",
        "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium",
        "Low", "Low", "Low", "Low", "Low", "Low"

    ]
    
    # Define headers for row 1
    headers = [
        "Sr. No.",
        "Questionnaire/ Points",
        "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)",
        "Risk Factor",
        "Observation",
        "Impact",
        "Recommendation"
    ]
    
    # Define column widths
    column_widths = {
        'A': 10,
        'B': 50,
        'C': 20,
        'D': 30,
        'E': 20,
        'F': 50,
        'G': 50,
        'H': 50
    }
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define risk factor colors
    risk_colors = {
        'Critical': '8B0000',  # Dark red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }
    
    # Set column widths
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Populate header row (A1-H1)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Populate questions and responses (A2-H46)
    for i in range(2, 47):  # Rows 2-46 (45 questions)
        question_num = i - 1
        
        # Sr. No. (A2-A46)
        sr_no_cell = ws.cell(row=i, column=1, value=question_num)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sr_no_cell.border = thin_border
        
        # Questionnaire/Points (B2-B46)
        question_cell = ws.cell(row=i, column=2, value=questions[i-2])
        question_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        question_cell.border = thin_border
        
        # Set row height for content rows
        ws.row_dimensions[i].height = 30
        
        # Get user input for this question
        user_input = None
        if form_data:
            # Map question numbers to form field names
            question_mapping = {
                1: 'separateServerRoom',
                2: 'internetAllowed',
                3: 'osVersion',
                4: 'eventLogsPresent',
                5: 'windowsOSGenuine',
                6: 'osOutdated',
                7: 'passwordPolicyPresent',
                8: 'antivirusAvailable',
                9: 'serverOSUpdated',
                10: 'cctvCamerasCount',
                11: 'antivirusExpired',
                12: 'antivirusConfigured',
                13: 'usbDevicesAllowed',
                14: 'remoteAccessLogs',
                15: 'remoteConnectionApp',
                16: 'lastDRDrill',
                17: 'serverAdequateSpace',
                18: 'serverRoomVisible',
                19: 'serverRoomAwayFromBasement',
                20: 'serverCloseToUPS',
                21: 'smokingProhibited',
                22: 'authorizedPersonsCount',
                23: 'biometricAccessPresent',
                24: 'biometricLogsPresent',
                25: 'cablesStructured',
                26: 'serverRackOpen',
                27: 'wireTaggingAvailable',
                28: 'cctvCablesConcealed',
                29: 'cctvBackupAvailable',
                30: 'fireExtinguishersCount',
                31: 'automaticFireExtinguisher',
                32: 'hygrometerPresent',
                33: 'smokeDetectorAvailable',
                34: 'fireAlarmPresent',
                35: 'antivirusUpdated',
                36: 'unauthorizedSoftwarePresent',
                37: 'remoteAccessByITTeam',
                38: 'drLocation',
                39: 'drDrillLogsMaintained',
                40: 'serverRoomClean',
                41: 'serverRoomAccessLog',
                42: 'networkDevicesPlacedProperly',
                43: 'serverRoomCooling',
                44: 'serverRoomLeakage',
                45: 'powerBackupPresent'
            }
            
            field_name = question_mapping.get(question_num)
            if field_name:
                user_input = form_data.get(field_name)
        
        # Define comprehensive question responses for all 45 questions
        question_responses = {
            1: {
                "compliance": {
                    "a": "Compliance",
                    "b": "A dedicated server room is available.",
                    "d": "A separate room is maintained for the servers, providing controlled access, temperature, humidity, and dust-free environment to ensure the proper functioning of servers.",
                    "e": "Dedicated server room minimizes risks from environmental hazards, unauthorized access, and operational failures.",
                    "f": "No recommendation required as a separate server room is available."
                },
                "non_compliance": {
                    "a": "Non-Compliance",
                    "b": "There was no separate room for the server. ",
                    "d": "Servers are located in shared areas without controlled environmental or access conditions, increasing the risk of damage and operational failures.",
                    "e": "Risk of data loss, hardware failure, and downtime due to temperature fluctuations, moisture, dust, or unauthorized access.",
                    "f": "Establish a dedicated server room with access control, temperature, and humidity regulation."
                },
                "not_applicable": {
                    "a": "Not Applicable",
                    "b": "Not Applicable",
                    "d": "Not Applicable",
                    "e": "Not Applicable",
                    "f": "Not Applicable"
                }
            },
            2: {
                "compliance": {
                    "a": "Compliance",
                    "b": "Internet access on servers is restricted.",
                    "d": "Internet is restricted to essential services only, minimizing exposure to external threats while allowing necessary updates and connectivity.",
                    "e": "Reduced risk of malware, ransomware, and unauthorized data exfiltration.",
                    "f": "No recommendation required as internet is restricted based on process needs."
                },
                "non_compliance": {
                    "a": "Non-Compliance",
                    "b": "Internet was allowed on servers without restriction.",
                    "d": "Servers were exposed to the full internet, increasing the risk of malware download, ransomware, and unauthorized access to CBS systems. Employees could inadvertently access malicious sites.",
                    "e": "High probability of malware infection, ransomware attacks, productivity loss, and unauthorized access to sensitive systems.",
                    "f": "Restrict internet access to only required processes and implement firewall/URL filtering controls."
                },
                "not_applicable": {
                    "a": "Not Applicable",
                    "b": "Not Applicable",
                    "d": "Not Applicable",
                    "e": "Not Applicable",
                    "f": "Not Applicable"
                }
            },
            3: {
                "compliance": {
                    "a": "Compliance",
                    "b": "Latest OS version is running on the servers.",
                    "d": "Servers are running an updated and supported version of the OS with regular security patches applied.",
                    "e": "Ensures protection against known vulnerabilities, malware, and ransomware.",
                    "f": "No recommendation required as OS is up to date."
                },
                "non_compliance": {
                    "a": "Non-Compliance",
                    "b": "Outdated Windows OS version was installed on the DC server.",
                    "d": "Outdated Windows 7 OS is in use, which is no longer supported by Microsoft and does not receive security patches.",
                    "e": "Servers are vulnerable to exploitation, ransomware, and other cyber-attacks.",
                    "f": "Upgrade all servers to a supported version of Windows OS with regular patching."
                },
                "not_applicable": {
                    "a": "Not Applicable",
                    "b": "Not Applicable",
                    "d": "Not Applicable",
                    "e": "Not Applicable",
                    "f": "Not Applicable"
                }
            },
            4: {
                "compliance": {
                    "a": "Compliance",
                    "b": "Event logs are maintained and up to date.",
                    "d": "Server OS logs are regularly updated, capturing all relevant events, security alerts, and system changes.",
                    "e": "Enables proper monitoring, forensic investigation, and compliance with security standards.",
                    "f": "No recommendation required as event logs are up to date."
                },
                "non_compliance": {
                    "a": "Non-Compliance",
                    "b": "Event logs for server OS was not present up to date",
                    "d": "Logs were outdated or missing, reducing visibility of system events, security incidents, and patch status.",
                    "e": "Lack of logs prevents timely detection of attacks and impedes forensic investigation.",
                    "f": "Ensure event logs are regularly updated and maintained for all servers."
                },
                "not_applicable": {
                    "a": "Not Applicable",
                    "b": "Not Applicable",
                    "d": "Not Applicable",
                    "e": "Not Applicable",
                    "f": "Not Applicable"
                }
            },
            5: {
                "compliance": {
                    "a": "Compliance",
                    "b": "Genuine Windows OS is installed.",
                    "d": "All servers run genuine Windows OS, ensuring full access to security updates and Microsoft support.",
                    "e": "Protection against malware, viruses, and system exploits.",
                    "f": "No recommendation required as OS is genuine."
                },
                "non_compliance": {
                    "a": "Non-Compliance",
                    "b": "Windows OS was not genuine.",
                    "d": "Servers run non-genuine Windows, which may contain embedded malware and do not receive security patches.",
                    "e": "Increased risk of hacking, viruses, data theft, and IT system paralysis.",
                    "f": "Install genuine Windows OS on all servers."
                },
                "not_applicable": {
                    "a": "Not Applicable",
                    "b": "Not Applicable",
                    "d": "Not Applicable",
                    "e": "Not Applicable",
                    "f": "Not Applicable"
                }
            },
            6: {
                "compliance": {
                    "a": "Compliance",
                    "b": "Operating System is up-to-date.",
                    "d": "All servers are running the latest OS version with security patches applied, ensuring resistance against modern cyber threats.",
                    "e": "Reduces vulnerability to malware, ransomware, and unauthorized access.",
                    "f": "No recommendation required as OS is updated."
                },
                "non_compliance": {
                    "a": "Non-Compliance",
                    "b": "Operating System was outdated.",
                    "d": "Some servers are running outdated OS versions which may not support the latest security patches and updates.",
                    "e": "Increases risk of ransomware, malware attacks, and data breaches; system vulnerabilities can be easily exploited.",
                    "f": "Upgrade all outdated OS versions to the latest supported version to ensure security and stability."
                },
                "not_applicable": {
                    "a": "Not Applicable",
                    "b": "Not Applicable",
                    "d": "Not Applicable",
                    "e": "Not Applicable",
                    "f": "Not Applicable"
                }
            },
            7: {
                "compliance": {
                    "a": "Compliance",
                    "b": "Password policy is configured for server OS authentication.",
                    "d": "Strong password policies are implemented, including complexity, rotation, and account lockout settings.",
                    "e": "Reduces risk of unauthorized access via brute-force attacks or password guessing.",
                    "f": "No recommendation required as password policy is configured."
                },
                "non_compliance": {
                    "a": "Non-Compliance",
                    "b": "The password policy was not configured.",
                    "d": "No password policy is defined for server OS authentication. Users may use weak or default passwords.",
                    "e": "High risk of unauthorized access through brute-force attacks or password theft.",
                    "f": "Configure a strong password policy with complexity, rotation, and lockout features."
                },
                "not_applicable": {
                    "a": "Not Applicable",
                    "b": "Not Applicable",
                    "d": "Not Applicable",
                    "e": "Not Applicable",
                    "f": "Not Applicable"
                }
            },
            8: {
                "compliance": {
                    "a": "Compliance",
                    "b": "Antivirus is installed on all servers.",
                    "d": "Antivirus software is deployed on all servers and endpoints to prevent, detect, and remove malware.",
                    "e": "Reduces the likelihood of virus infection, malware propagation, and cyber-attacks.",
                    "f": "No recommendation required as antivirus is available."
                },
                "non_compliance": {
                    "a": "Non-Compliance",
                    "b": "Antivirus was not available.",
                    "d": "Servers do not have antivirus software installed, leaving systems vulnerable to malware and destructive viruses.",
                    "e": "Malware infections can compromise servers, disrupt banking operations, and result in data loss.",
                    "f": "Install antivirus software on all servers and critical systems to prevent malware and cyber-attacks."
                },
                "not_applicable": {
                    "a": "Not Applicable",
                    "b": "Not Applicable",
                    "d": "Not Applicable",
                    "e": "Not Applicable",
                    "f": "Not Applicable"
                }
            },
            9: {
                "compliance": {
                    "a": "Compliance",
                    "b": "Server OS is up-to-date.",
                    "d": "Server OS is patched regularly and updated to the latest version, ensuring protection against known vulnerabilities.",
                    "e": "Enhances server security and reduces risk from cyber-attacks.",
                    "f": "No recommendation required as OS is updated."
                },
                "non_compliance": {
                    "a": "Non-Compliance",
                    "b": "Server Windows OS was outdated.",
                    "d": "Server is running an outdated OS version with no current security patches applied.",
                    "e": "Increased risk of ransomware and unauthorized access, making the server vulnerable to exploits.",
                    "f": "Update server OS to the latest supported version with all security patches."
                },
                "not_applicable": {
                    "a": "Not Applicable",
                    "b": "Not Applicable",
                    "d": "Not Applicable",
                    "e": "Not Applicable",
                    "f": "Not Applicable"
                }
            },
            10: {
                "compliance": {
                    "a": "Compliance",
                    "b": "Two or more CCTV cameras are installed.",
                    "d": "CCTV coverage ensures all areas of the server room are monitored, with concealed wiring for tamper-proof operation.",
                    "e": "Provides surveillance, enhances security, and supports forensic investigations.",
                    "f": "No recommendation required as CCTV coverage is adequate."
                },
                "non_compliance": {
                    "a": "Non-Compliance",
                    "b": "Only one CCTV camera was present.",
                    "d": "Only one camera is installed, and its cables are not properly concealed.",
                    "e": "Blind spots exist, and if the camera fails, there will be no backup coverage; CCTV footage may not be reliable during incidents.",
                    "f": "Install at least two cameras and conceal cables properly for full server room coverage."
                },
                "not_applicable": {
                    "a": "Not Applicable",
                    "b": "Not Applicable",
                    "d": "Not Applicable",
                    "e": "Not Applicable",
                    "f": "Not Applicable"
                }
            }
        }
        
        # Add remaining questions 11-45 with similar structure
        remaining_questions_11_45 = {
            11: {"compliance": {"a": "Compliance", "b": "Antivirus license is active and valid.", "d": "Antivirus subscription is current, providing latest virus definitions and malware protection.", "e": "Ensures continuous protection against new and emerging threats.", "f": "No recommendation required as antivirus is active."}, "non_compliance": {"a": "Non-Compliance", "b": "Expired license Antivirus was in use.", "d": "Antivirus license has expired, preventing updates and leaving servers unprotected.", "e": "Servers are vulnerable to the latest malware and viruses, increasing risk of compromise.", "f": "Renew antivirus licenses and implement an enterprise solution for continuous protection."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            12: {"compliance": {"a": "Compliance", "b": "Antivirus is properly configured.", "d": "Antivirus is correctly set up to scan all files, emails, and network traffic regularly.", "e": "Enhances detection and prevention of threats, ensuring servers remain secure.", "f": "No recommendation required as antivirus is properly configured."}, "non_compliance": {"a": "Non-Compliance", "b": "The antivirus was not configured properly.", "d": "Antivirus settings are incomplete or misconfigured, reducing effectiveness against threats.", "e": "If the Antivirus is not properly configured it will not recognize the most current viruses that have been created and threats, either by its malware signature or its behavior.", "f": "It is recommended to properly configure Antivirus."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            13: {"compliance": {"a": "Compliance", "b": "USB access is disabled on servers.", "d": "USB ports are disabled or controlled, preventing unauthorized data transfers and malware introduction.", "e": "Reduces risk of data exfiltration and malware infection from removable media.", "f": "No recommendation required as USB ports are restricted."}, "non_compliance": {"a": "Non-Compliance", "b": "USB was enabled in all the servers.", "d": "USB access is enabled on all servers, allowing potential malware injection or data theft.", "e": "High risk of cyber-attacks like Rubber Ducky, USB Killer, and data exfiltration, compromising critical banking data.", "f": "Disable USB on servers or implement trusted USB controls for authorized devices only."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            14: {"compliance": {"a": "Compliance", "b": "Logs for remote access are maintained.", "d": "Detailed records are maintained for all remote access, including date, time, user, reason, and designation.", "e": "Ensures accountability, traceability, and ease of auditing in case of incidents.", "f": "No recommendation required as remote access logging is maintained."}, "non_compliance": {"a": "Non-Compliance", "b": "Logs were not maintained for remote access.", "d": "No records are available for remote access to servers, including who accessed, when, and for what purpose.", "e": "Difficult to audit remote access trails; accountability cannot be established if any incident occurs, increasing operational and security risk.", "f": "Maintain detailed logs for all remote access, including date, time, user, designation, and purpose."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            15: {"compliance": {"a": "Compliance", "b": "Paid and updated remote access applications are used.", "d": "Remote connection software is up-to-date, licensed, and secure, preventing unauthorized access or malware injection.", "e": "Reduces the risk of malware, keyloggers, and unauthorized entry via remote access.", "f": "No recommendation required as secure remote access software is in use."}, "non_compliance": {"a": "Non-Compliance", "b": "The application being used for remote connection were open source and outdated version.", "d": "Open-source and outdated remote access applications are in use, which may contain vulnerabilities exploitable by attackers.", "e": "Possible installation of malware or trojans creating backdoors; lack of logs may prevent detection of malicious activity.", "f": "Use paid, updated, and secure remote access software; implement strong controls and logging for remote connections."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            16: {"compliance": {"a": "Compliance", "b": "DR drills are conducted regularly (at least twice a year).", "d": "Bank maintains records of DR drills including outcomes, ensuring systems can be restored in case of disaster.", "e": "Ensures preparedness, reduces downtime, and ensures data availability during a disaster.", "f": "No recommendation required as DR drills are conducted as per schedule."}, "non_compliance": {"a": "Non-Compliance", "b": "DR drill was not conducted for the acceptable period.", "d": "Last DR drill was conducted more than a year ago; no proper records maintained.", "e": "Uncertainty in system restoration during actual disaster; higher risk of prolonged downtime and data loss.", "f": "Conduct DR drills at least twice a year; maintain records and reports of the drill outcomes."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            17: {"compliance": {"a": "Compliance", "b": "Server room has adequate space for operational requirements.", "d": "Servers are arranged with sufficient spacing, proper airflow, and access for maintenance.", "e": "Reduces hardware malfunction risk due to heat, dust, or restricted access; ensures optimal operation.", "f": "No recommendation required as server space is adequate."}, "non_compliance": {"a": "Non-Compliance", "b": "The server does not have adequate space for operational requirements.", "d": "Servers are crowded with limited space for airflow, cooling, and maintenance.", "e": "Increased risk of hardware failures, overheating, dust accumulation, and reduced operational efficiency.", "f": "Provide adequate space around servers for optimal cooling, maintenance, and operational efficiency."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            18: {"compliance": {"a": "Compliance", "b": "Server room is visible and access is restricted.", "d": "The server room is located in a visible area with controlled access for authorized personnel only.", "e": "Reduces risk of insider threats and unauthorized physical access; ensures better monitoring.", "f": "No recommendation required as server room is secure and visible."}, "non_compliance": {"a": "Non-Compliance", "b": "The server room was not visible from a distance and was easily accesible.", "d": "The server room was not visible from a distance and was easily accessible to anyone.", "e": "High risk of insider threats and unauthorized access to critical systems.", "f": "Ensure visibility from a distance and restrict access to authorized personnel only."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            19: {"compliance": {"a": "Compliance", "b": "Server room is located away from basement and water sources.", "d": "The server room is positioned on higher floors or areas away from potential water or drainage hazards.", "e": "Minimizes risk of water damage, electrical hazards, and equipment failure.", "f": "No recommendation required as server room location is safe."}, "non_compliance": {"a": "Non-Compliance", "b": "A server room was near to the basement and water/drainage systems.", "d": "Server room is located near the basement and water/drainage systems.", "e": "Risk of flooding, water damage, and static electricity; can damage sensitive equipment.", "f": "Relocate server room or ensure it is away from basements and water/drainage systems."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            20: {"compliance": {"a": "Compliance", "b": "Servers are separate from UPS room.", "d": "UPS and server rooms are in separate locations to minimize physical hazards.", "e": "Reduces risk of damage from battery hazards, temperature fluctuations, and electrical interference.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "Servers were near the UPS room. ", "d": "UPS and batteries are stored close to servers in the same room.", "e": "It may cause damage to DC as batteries can be explosive that can damage the whole DC. All the firewalls in the world won't stop the physical dangers that are inside your building right now. These menaces include temperature, humidity, vibration, water leaks, and intrusion. These threats can damage equipment, force hardware to shut down, and slow performance", "f": "Maintain separation between servers and UPS/battery storage."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            21: {"compliance": {"a": "Compliance", "b": "Clear instructions prohibiting smoking, eating, and drinking are displayed.", "d": "Server room has signs prohibiting smoking, eating, and drinking, followed by staff.", "e": "Prevents spillage, fire hazards, and equipment damage.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "Instructions were not available saying that Smoking, eating, and drinking are prohibited in the server room.", "d": "It was observed tha there were no instructions available saying that Smoking, eating, and drinking are prohibited in the server room.", "e": "As no instructions were available saying that Smoking, eating, and drinking are prohibited in the server room. Anyone with access to the server room can carry out the smoking, eating, and drinking  inside the server room, making the equipment present inside vulnerable to the spillage of food or liquid into sensitive computer equipment and damage. These threats can damage equipment, force hardware to shut down, and slow performance.", "f": "It is recommended that Smoking, eating, and drinking prohibited in the server room instructions must be available to prevent the spillage of food or liquid into sensitive computer equipment."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            22: {"compliance": {"a": "Compliance", "b": "Only essential personnel have access.", "d": "Access to server room is limited to IT department and critical staff.", "e": "Reduces risk of insider threats and unauthorized access.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "Numerous persons were authorized to access the server room.", "d": "Many individuals have access to the server room.", "e": "Higher chance of insider threats or unauthorized access; critical systems vulnerable.", "f": "Limit access to only essential and authorized personnel."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            23: {"compliance": {"a": "Compliance", "b": "Biometric/Proximity Card machine installed and functional.", "d": "Biometric or proximity card system is used for controlled access.", "e": "Ensures accountability, tracks attendance, and prevents unauthorized entry.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "A Biometric or Proximity Card machine for access in the server room was not available.", "d": "No Biometric or Proximity Card system installed in server room.", "e": "Hard to track access, attendance, or maintain accountability; risk of unauthorized entry.", "f": "Install Biometric or Proximity Card access system."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            24: {"compliance": {"a": "Compliance", "b": "Logs are maintained and updated regularly.", "d": "Biometric and proximity card access logs are recorded and kept up-to-date.", "e": "Facilitates forensic investigations and accountability.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "Logs not maintained.", "d": "Biometric logs or Proximity Card were not present and not maintained.", "e": "Forensic investigations impossible; unauthorized access cannot be traced.", "f": "Maintain and regularly update biometric/proximity access logs."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            25: {"compliance": {"a": "Compliance", "b": "Cables structured properly.", "d": "Network cables organized with proper routing, labeling, and management.", "e": "Simplifies maintenance, troubleshooting, and reduces equipment damage risk.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "The cables of Network Devices were not structured properly in the server rack.", "d": "It was observed that the cables of the network devices were not structured properly in the server rack.", "e": "As the cables for the Network Devices are not structured properly in the server rack, it will be hard to manage and maintain the available devices and install new devices in the server rack. ", "f": "It is recommended to structure cables properly for all the networking devices in the server rack."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            26: {"compliance": {"a": "Compliance", "b": "Server rack is locked and access controlled.", "d": "All critical server racks are locked and keys held by authorized personnel.", "e": "Prevents unauthorized physical access and tampering.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "The server rack was not locked.", "d": "Server racks are open and accessible to anyone.", "e": "Critical systems vulnerable to physical intrusion.", "f": "Lock all server racks and assign responsibility for keys to authorized personnel."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            27: {"compliance": {"a": "Compliance", "b": "All network cables properly tagged.", "d": "Each network cable labeled with identification and function.", "e": "Facilitates management, maintenance, and troubleshooting.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "wire tagging was not available for the cables of Network Devices in the server rack.", "d": "It was observed that wire tagging was not present for the cables of the network devices in the server rack.", "e": "As the tagging of cables for the Network Devices was partially available in the server rack, it will be hard to identify the cable and its work. Thus, it will be hard to manage and maintain the available devices.", "f": "It is recommended to tag all the cables properly for all the networking devices in the server rack."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            28: {"compliance": {"a": "Compliance", "b": "CCTV cables concealed.", "d": "All CCTV cables properly routed and hidden to prevent tampering.", "e": "Reduces risk of disabling cameras; ensures continuous surveillance.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "CCTV camera cables were not properly concealed.", "d": "CCTV camera cables exposed and unprotected.", "e": "Risk of cable cutting, disabling CCTV, and losing evidence.", "f": "Conceal all CCTV cables to protect surveillance."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            29: {"compliance": {"a": "Compliance", "b": "CCTV footage backed up for 30+ days.", "d": "Bank maintains CCTV footage backup for 30 days as per NABARD guidelines.", "e": "Supports forensic investigations and incident tracking.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "30 days CCTV Footage backup was not available.", "d": "Bank maintains CCTV backup for less than 30 days.", "e": "Inadequate backup limits forensic investigations and evidence availability.", "f": "Maintain CCTV footage backup for at least 30 days."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            30: {"compliance": {"a": "Compliance", "b": "Fire extinguishers available and functional.", "d": "Fire extinguishers are installed in server room and maintained regularly.", "e": "Mitigates fire hazards; ensures quick response in emergencies.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "Fire extinguishers not available.", "d": "No fire extinguishers installed in server room.", "e": "Fire incidents may cause severe damage, data loss, and business disruption.", "f": "Install and maintain adequate fire extinguishers in server room."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            31: {"compliance": {"a": "Compliance", "b": "Automatic fire extinguisher installed in the server room.", "d": "The Data Centre is equipped with an automatic fire extinguisher system for rapid response.", "e": "Provides immediate fire suppression, minimizing damage to critical equipment.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "The automatic fire extinguisher was not present in the server room.", "d": "No automatic fire extinguisher installed in the Data Centre.", "e": "Delay in fire suppression could lead to extensive damage if a fire occurs during non-banking hours.", "f": "Install automatic fire extinguishers in the Data Centre for rapid fire response."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            32: {"compliance": {"a": "Compliance", "b": "Hygrometer installed in server room.", "d": "Hygrometer installed and relative humidity monitored between 45–55%.", "e": "Maintains optimal humidity; prevents condensation, corrosion, and static electricity buildup.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "A hygrometer was not present in the server room.", "d": "No hygrometer installed in the Data Centre to monitor humidity levels.", "e": "High risk of condensation, corrosion, static electricity, and equipment damage.", "f": "Install a hygrometer and maintain relative humidity between 45–55%."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            33: {"compliance": {"a": "Compliance", "b": "Smoke detector installed and functional.", "d": "Smoke detector installed and connected to fire alarm system.", "e": "Early fire detection reduces risk of damage and downtime.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "The smoke detector not available.", "d": "No smoke detector installed in the Data Centre.", "e": "A smoke detector is an electronic fire-protection device that automatically senses the presence of smoke, as a key indicator of fire, and sounds a warning to building occupants, in case of a fire event in the data center. As a smoke detector was not present in the data center, the bank will not be able to detect/prevent a fire event.", "f": "Install smoke detectors in the Data Centre."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            34: {"compliance": {"a": "Compliance", "b": "Fire alarm installed and functional.", "d": "Fire alarm system installed and operational, connected to detectors.", "e": "Alerts staff to fire incidents promptly, enabling quick mitigation.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "Fire alarm was not present in the data center.", "d": "No fire alarm in the Data Centre.", "e": "The fire alarm system can be set off automatically by smoke detectors, heat detectors, or manually. The optical smoke detector detects the smoke using light sensors (infrared LED). When smoke particles pass through the chamber of the optical detector, it scatters light that triggers the alarm. As the fire alarm was not present in the DC room. In case of a fire event in the room, the bank will not be able to identify and prevent it. ", "f": "Install a fire alarm system in the Data Centre."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            35: {"compliance": {"a": "Compliance", "b": "Antivirus regularly updated with latest signatures.", "d": "Antivirus is up-to-date and provides protection against current malware and viruses.", "e": "Reduces risk of cyber-attacks and malware infections.", "f": "Continue regular updates."}, "non_compliance": {"a": "Non-Compliance", "b": " The Antivirus was not updated.", "d": "Antivirus in the DC is outdated and missing latest virus definitions.", "e": "Servers are vulnerable to new malware, ransomware, and cyber-attacks.", "f": "It is recommended to always update Antivirus for the new signatures of viruses or malware, as the Antivirus updates contain the latest files needed to combat new viruses and protect your server."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            36: {"compliance": {"a": "Compliance", "b": "No unauthorized software installed.", "d": "Only approved software is installed and regularly audited.", "e": "Reduces risk of malware, data breaches, and vulnerability exploits.", "f": "Continue monitoring for unauthorized software."}, "non_compliance": {"a": "Non-Compliance", "b": "Unauthorized software was present. ", "d": "It was observed that many of the unauthorized software were present like VLC media player, Team viewer 7, Outdated version of Google Chrome web browser, Any desk, Skype, ArcSoft PhotoStudio, Crazy Browser 3.1.0. ", "e": "Increases risk of malware, unauthorized access, and data breaches.", "f": "Uninstall all unauthorized software and restrict installation rights."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            37: {"compliance": {"a": "Compliance", "b": "Remote access responsibility clearly defined per user.", "d": "Each IT user has a separate account with defined responsibilities for remote access.", "e": "Accountability is maintained; any action can be traced to the responsible user.", "f": "Continue enforcing defined remote access roles."}, "non_compliance": {"a": "Non-Compliance", "b": "Responsibility for Remote Access to the servers cannot be defined.", "d": "Any IT staff can remotely access servers using shared credentials.", "e": "If anyone in the IT department takes remote access to the servers by using the same user accounts, the bank will not be able to define the responsibility of any specific user for the specific action or process done on the servers.", "f": "Create individual user accounts for remote access to define responsibility."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            38: {"compliance": {"a": "Compliance", "b": "DR located in a different seismic zone than DC.", "d": "Disaster Recovery site is geographically separate from Data Centre to avoid simultaneous risks.", "e": "Ensures continuity of services in case of disaster affecting DC.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "DC and DR were present in the same location.", "d": "DC and DR are in the same location and seismic zone.", "e": "Disaster affecting DC will also impact DR, risking service continuity.", "f": "Relocate DR to a separate seismic zone."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            39: {"compliance": {"a": "Compliance", "b": "DR drill logs properly maintained.", "d": "DR drill outcomes and steps recorded and stored systematically.", "e": "Enables performance review and improvement in disaster readiness.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "Logs for DR drills were not maintained properly.", "d": "DR drill logs incomplete or absent.", "e": "Bank cannot measure DR drill success or identify failures.", "f": "Maintain proper DR drill logs for evaluation and compliance."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            40: {"compliance": {"a": "Compliance", "b": "Server room neat and dust-free.", "d": "Regular cleaning ensures a dust-free environment.", "e": "Prevents fan clogging, overheating, and equipment degradation.", "f": "Continue regular cleaning."}, "non_compliance": {"a": "Non-Compliance", "b": "The server room was not neat and clean.", "d": "It was observed that the Server room was not neat and clean to ensure a dust-free environment.", "e": "Without regular server room cleaning, dirt will build up to create grime on server fans, causing your cooling systems to slow down.", "f": "It is recommended that the Server room should be neat and clean to ensure a dust-free environment."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            41: {"compliance": {"a": "Compliance", "b": "Access logs registered and maintained.", "d": "All entries/exits recorded in access register and verified.", "e": "Tracks visitors; improves accountability and forensic readiness.", "f": "Continue maintaining access logs."}, "non_compliance": {"a": "Non-Compliance", "b": "The server room access log was not registered, not present, and therefore not maintained. ", "d": "Server room access register not present or maintained.", "e": "Unauthorized entry cannot be traced; forensic investigation hindered.", "f": "Register, maintain, and update server room access logs."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            42: {"compliance": {"a": "Compliance", "b": "Network devices properly arranged in server rack.", "d": "Devices are secured, labeled, and organized with proper airflow.", "e": "Prevents physical intrusion and ensures maintainability.", "f": "Continue proper placement."}, "non_compliance": {"a": "Non-Compliance", "b": "Network devices were not placed properly in a server rack. ", "d": "It was observed that the network devices are not placed properly in server rack.", "e": "As network devices are not placed properly in server racks, the critical systems are openly accessible to any physical intrusion. As the physical security is not implemented properly, fewer obstacles are placed in the way of a potential attacker, and physical sites cannot provide security against accidents and attacks.", "f": "It is recommended that network devices are placed properly in the server rack. Keep the server rack locked, and the responsibility of the key should be given to accountable personnel."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            43: {"compliance": {"a": "Compliance", "b": "Server room adequately cooled with automatic A/C.", "d": "Multiple A/C units with automatic control maintain optimal temperature.", "e": "Prevents overheating and ensures uninterrupted server operation.", "f": "Continue proper cooling management."}, "non_compliance": {"a": "Non-Compliance", "b": "Only one A/C was present in the server and no Switching ON/OFF of AC Automatic therefore No proper cooling was present in the server room.", "d": "It was observed that Only one A/C was present in the server and no Switching ON/OFF of AC Automatic therefore No proper cooling was present in the server room.", "e": "When the temperature around and within the server and networking equipment becomes too high the server will shut down and there will be a loss of data.", "f": "It is recommended that enough A/C, proper cooling, and Automatic A/C should be available in the server room."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            44: {"compliance": {"a": "Compliance", "b": "No leakage detected.", "d": "Server room walls intact with no water ingress.", "e": "Reduces risk of water damage, short circuits, and fire hazards.", "f": "Continue regular inspection."}, "non_compliance": {"a": "Non-Compliance", "b": "The server room walls had leakage.", "d": "It was observed that  Server room walls had leakage.", "e": "These threats can damage equipment, force hardware to shut down, and slow performance. One of the most frightening dangers that a water leak can lead to is a fire. If leaking water reaches bank electrical supply or loose wires, it can cause your electricity to short circuit and create a spark which can then ignite a fire.", "f": "It is recommended to take preventive measures for server room wall leakage."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}},
            45: {"compliance": {"a": "Compliance", "b": "Power backup available and operational.", "d": "UPS or generator installed to provide uninterrupted power.", "e": "Ensures continuous operations and prevents data loss during outages.", "f": "No recommendation required."}, "non_compliance": {"a": "Non-Compliance", "b": "Power backup was not available.", "d": "It was observed that power backup was not available. Data centers in banking and finance require robust contingency power systems to match the size of their operations to retain power for critical operations under adverse conditions.", "e": "The data center constitutes the critical load in the daily operations of the organization and plays the role of the backbone of financial data computation, transaction system, and records storage. Servers, storages, routers, and switches are the key IT equipment in the datacenters. As power backup was not available, the cost of downtime due to critical load failure will be extreme.", "f": "It is recommended to have high power efficiency and ultimate availability of the UPS system."}, "not_applicable": {"a": "Not Applicable", "b": "Not Applicable", "d": "Not Applicable", "e": "Not Applicable", "f": "Not Applicable"}}
        }
        
        # Merge all questions
        question_responses.update(remaining_questions_11_45)
        
        # Populate columns C, D, F, G, H with response data
        if user_input and question_num in question_responses:
            response_data = question_responses[question_num]
            if user_input.lower() == 'compliance':
                response_set = response_data['compliance']
            elif user_input.lower() == 'non-compliance':
                response_set = response_data['non_compliance']
            else:  # Not Applicable or any other value
                response_set = response_data['not_applicable']
            
            # Populate columns C, D, F, G, H with response data
            for col_letter, value in response_set.items():
                col_num = ord(col_letter) - ord('a') + 3  # Convert 'a' to 3, 'b' to 4, etc.
                cell = ws.cell(row=i, column=col_num, value=value)
                
                # Apply different alignment based on column
                # 'a' = Column C (center), 'b' = Column D (middle), 'd' = Column F (middle), 'e' = Column G (middle), 'f' = Column H (middle)
                if col_letter in ['b', 'd', 'e', 'f']:  # Columns D, F, G, H - middle align only
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:  # Column C - center align
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                cell.border = thin_border
        else:
            # If no form data or question not found, populate with empty cells
            for col in range(3, 9):  # Columns C to H
                empty_cell = ws.cell(row=i, column=col, value="")
                
                # Apply different alignment based on column
                if col in [4, 6, 7, 8]:  # Columns D, F, G, H - middle align only
                    empty_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:  # Column C - center align
                    empty_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                empty_cell.border = thin_border
        
        # Risk Factor (E2-E46) with color coding
        risk_factor = risk_factors[i-2]  # Get risk factor for this row
        risk_cell = ws.cell(row=i, column=5, value=risk_factor)
        risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
        risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
        risk_cell.border = thin_border
    
    # Save the file
    filename = "Data Centre Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure the uploads directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    
    print(f"Excel file created: {filepath}")
    print(f"Filename: {filename}")
    
    return filepath, filename

def cleanup_file(filepath):
    """
    Delete the file from the uploads folder after download
    """
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            print(f"File cleaned up: {filepath}")
    except Exception as e:
        print(f"Error cleaning up file {filepath}: {str(e)}")

if __name__ == "__main__":
    # Test the function
    create_data_centre_excel()
