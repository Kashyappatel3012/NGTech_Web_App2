import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def create_firewall_excel(form_data=None):
    """
    Create Excel file for Firewall Assessment
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Firewall"
    
    # Define questions
    questions = [
        "Is the firewall in HA (High Availability) mode?",
        "Is the firewall rule set to any any any?",
        "Are all Internet connections routed through a firewall?",
        "Are the firewall rules scanned through antivirus or not?",
        "Is ATP (Advanced Threat Protection) enabled or not?",
        "Is USB access allowed in the firewall?",
        "Is web filtering enabled or not?",
        "Is web-based proxy enabled?",
        "Are ICMP requests disabled, or is the ping flooding feature enabled?",
        "Does the GUI only have HTTPS access, and CLI only have SSH access?",
        "Is audit trail enabled on the firewall to log changes made to rule base settings, and are the logged entries approved by higher authorities in the IT Department?",
        "Are default credentials in use, or are accounts named 'Admin', 'root', or 'Administrator' for administrator login?",
        "Are DoS and spoof protection enabled or not?",
        "Is the firmware up to date?",
        "Does the firewall have a valid support license available?",
        "Is a firewall present in branches?",
        "Are system logs configured?",
        "How frequently is backup done for the firewall configuration?",
        "Is the administrative portal login accessed only by whitelisted IP addresses?",
        "Is a Syslog server present or not?",
        "Is administrator access portal HTTPS on TLS 1.2 or greater?",
        "Are HTTP logins disabled, and redirected to HTTPS?",
        "Are there multiple admin accounts or not?",
        "Is VPN (Virtual Private Network) in use? If yes, is a secured channel used?",
        "Is NTP (Network Time Protocol) configured?",
        "Is the firewall physically secured or not?",
        "Are alerts being monitored regularly or not?",
        "Are password policies configured as per the organization's password management policy?",
        "Are unused ports closed on the firewall?",
        "Are system administrators monitoring the logs produced by the Intrusion Detection System?",
        "Do administrators have two-factor authentication enabled?",
        "Is a login disclaimer message set or not?",
        "Is ACL configured or not?",
        "Are login timeout settings as per organization policies?"
    ]
    
    # Risk factors for each question
    risk_factors = [
        'Critical', 'Critical', 'Critical', 'High', 'High', 'High', 'High', 'High', 'High', 'High',
        'High', 'High', 'High', 'High', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium',
        'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low',
        'Low', 'Low', 'Low', 'Low'
    ]
    
    # Question mapping for form fields
    question_mapping = [
        'firewallHA', 'firewallRuleAny', 'internetThroughFirewall', 'firewallAntivirus', 'firewallATP',
        'firewallUSB', 'webFiltering', 'webProxy', 'icmpDisabled', 'guiHttpsCliSSH',
        'auditTrail', 'defaultCredentials', 'dosSpoofProtection', 'firmwareUpdated', 'supportLicense',
        'firewallBranches', 'systemLogs', 'backupFrequency', 'adminWhitelistIP', 'syslogServer',
        'adminHttpsTls', 'httpLoginDisabled', 'multipleAdminAccounts', 'vpnSecure', 'ntpConfigured',
        'firewallPhysicallySecured', 'alertsMonitored', 'passwordPolicies', 'unusedPortsClosed', 'idsLogsMonitored',
        'twoFactorAuth', 'loginDisclaimer', 'aclConfigured', 'loginTimeout'
    ]
    
    # Set column widths
    ws.column_dimensions['A'].width = 10  # Sr. No.
    ws.column_dimensions['B'].width = 50  # Questions
    ws.column_dimensions['C'].width = 20  # Compliance/Non-Compliance/Not Applicable
    ws.column_dimensions['D'].width = 30  # Observation (Short/Brief)
    ws.column_dimensions['E'].width = 20  # Risk Factor
    ws.column_dimensions['F'].width = 50  # Observation
    ws.column_dimensions['G'].width = 50  # Impact
    ws.column_dimensions['H'].width = 50  # Recommendation
    
    # Header row
    headers = ['Sr. No.', 'Questionnaire/Points', 'Compliance/Non-Compliance/Not Applicable', 
               'Observation (Short/Brief)', 'Risk Factor', 'Observation', 'Impact', 'Recommendation']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Populate Sr. No. and Questions
    for i, question in enumerate(questions, 2):
        ws.cell(row=i, column=1, value=i-1)  # Sr. No.
        ws.cell(row=i, column=2, value=question)  # Questions
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }
    
    # Question responses data
    question_responses = {
        1: {  # firewallHA
            'compliance': {'a': 'Compliance', 'b': 'Firewall is in HA mode.', 'd': 'Firewall is configured in High Availability (HA) mode to ensure redundancy.', 'f': 'Continuous network protection even if one firewall fails.', 'h': 'Regularly test HA failover to ensure uninterrupted operations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Firewall was not in HA mode.', 'd': 'Only a single firewall is present, with no HA configuration or secondary backup firewall.', 'f': 'If the firewall fails, all inbound/outbound traffic will bypass security policies, exposing critical banking systems to attacks, malware, and unauthorized access, potentially compromising the CIA triad (Confidentiality, Integrity, Availability).', 'h': 'Deploy a secondary firewall and enable HA mode to ensure continuous protection, reduce downtime, and safeguard sensitive banking operations.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # firewallRuleAny
            'compliance': {'a': 'Compliance', 'b': 'Firewall rules are properly configured.', 'd': 'Firewall rules are restricted to authorized sources, destinations, and ports.', 'f': 'Prevents unauthorized access and limits attack vectors.', 'h': 'Periodically audit rules to maintain least-privilege access.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Firewall rule was set to any source, to any destination through any port.', 'd': 'Firewall allows all inbound/outbound connections without restrictions.', 'f': 'Attackers can easily bypass firewall controls, inject malware, exfiltrate sensitive banking data, or perform network reconnaissance. This can lead to financial losses, data breaches, and reputational damage.', 'h': 'Reconfigure firewall rules to restrict access to authorized IPs and specific ports. Implement periodic reviews to remove unnecessary rules and enforce least-privilege principles.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # internetThroughFirewall
            'compliance': {'a': 'Compliance', 'b': 'All Internet traffic is routed through the firewall.', 'd': 'All inbound and outbound connections are monitored by the firewall.', 'f': 'Ensures comprehensive traffic inspection, monitoring, and threat prevention.', 'h': 'Periodically verify routing to ensure full coverage.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'All Internet connections were not routed through a Firewall.', 'd': 'Not all Internet traffic is passing through the firewall for inspection and control.', 'f': 'Unmonitored traffic may carry malware, phishing, or malicious activity undetected. This exposes the bank network to cyberattacks and reduces visibility into ongoing threats, potentially impacting critical operations and regulatory compliance.', 'h': 'Ensure all Internet traffic is routed through the firewall to enable monitoring, filtering, and logging, thereby protecting the network from unauthorized or malicious activity.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # firewallAntivirus
            'compliance': {'a': 'Compliance', 'b': 'Firewall rules are scanned through antivirus.', 'd': 'All traffic passing through the firewall is scanned for malware.', 'f': 'Helps detect and block malware, ensuring network integrity.', 'h': 'Keep antivirus definitions updated and review scanning configurations regularly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Firewall rules were not scanned by Antivirus.', 'd': 'Traffic allowed by firewall rules is not inspected by antivirus for malicious content.', 'f': 'Malware, ransomware, or viruses can pass through unchecked, compromising critical banking systems and sensitive customer data. Lack of inspection increases risk of financial fraud, system downtime, and forensic investigation challenges.', 'h': 'Enable antivirus scanning for all firewall rules. Ensure that real-time updates are applied and traffic inspection covers inbound and outbound data to reduce infection and data breach risks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # firewallATP
            'compliance': {'a': 'Compliance', 'b': 'ATP is enabled on the firewall.', 'd': 'Firewall has Advanced Threat Protection enabled to detect zero-day attacks and sophisticated malware.', 'f': 'Provides proactive protection against unknown threats, reducing risk of system compromise.', 'h': 'Regularly update ATP definitions and review protection logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The ATP is not enabled through firewall.', 'd': 'Advanced Threat Protection feature is disabled on the firewall.', 'f': 'The bank\'s network is exposed to zero-day threats, ransomware, and sophisticated malware. Without ATP, attacks may go undetected until critical systems are compromised, causing potential financial loss, data breach, and reputational damage.', 'h': 'Enable ATP on the firewall to provide real-time protection, including the latest threat signatures, heuristic detection, and mitigation against emerging malware and attack vectors.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # firewallUSB
            'compliance': {'a': 'Compliance', 'b': 'USB access is disabled.', 'd': 'Firewall does not allow USB connections to prevent unauthorized configuration changes.', 'f': 'Prevents malware introduction and tampering through external drives.', 'h': 'Periodically audit device policies to ensure USB remains disabled unless explicitly needed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The USB access was enabled.', 'd': 'USB ports on the firewall are accessible and active.', 'f': 'Attackers or unauthorized personnel can inject malicious files, modify firewall configuration, or extract sensitive data via USB drives. This compromises the integrity and availability of the firewall, putting the bank network at risk of intrusion, data theft, or service disruption.', 'h': 'Immediately disable USB access on the firewall and implement strict physical security controls. Restrict configuration changes to authorized personnel only.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # webFiltering
            'compliance': {'a': 'Compliance', 'b': 'Web filtering is enabled.', 'd': 'Users are restricted from visiting malicious websites through firewall web filtering.', 'f': 'Reduces risk of malware, phishing attacks, and accidental data leakage.', 'h': 'Periodically review and update web filtering rules.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Web filtering was enabled.', 'd': 'Users can access harmful or unverified websites without restriction.', 'f': 'An insider attacker can take advantage of disabled proxy server because if it is enable it will provide increased performance and security. In some cases, they monitor employees use of outside resources. A proxy server works by intercepting connections between sender and receiver. All incoming data enters through one port and is forwarded to the rest of the network via another port.', 'h': 'It is recommended to implement a web-based proxy over the firewall network. It can intercept all requests to the real server to see if it can fulfil the requests themselves.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # webProxy
            'compliance': {'a': 'Compliance', 'b': 'Web-based proxy is enabled.', 'd': 'All requests are filtered at the application layer to improve security and performance.', 'f': 'Provides content caching, security monitoring, and controlled access to resources.', 'h': 'Regularly monitor proxy logs and update rules.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Web-based proxy was disabled.', 'd': 'No proxy server is configured to filter web traffic.', 'f': 'Users may bypass security controls, access malicious sites, or leak sensitive data. Insider attackers can exploit disabled proxy, increasing the chance of data compromise, regulatory non-compliance, and reduced network performance.', 'h': 'Implement a web-based proxy on the firewall to filter, monitor, and secure web traffic. Ensure it intercepts all requests and enforces corporate browsing policies.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # icmpDisabled
            'compliance': {'a': 'Compliance', 'b': 'ICMP is disabled; ping flood protection is enabled.', 'd': 'Network is protected from ICMP-based denial-of-service attacks.', 'f': 'Reduces risk of network disruption and ensures availability of critical systems.', 'h': 'Regularly verify firewall configurations for ICMP and flood protections.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'ICMP request was enabled.', 'd': 'ICMP requests are allowed and ping flood protection is not enabled.', 'f': 'Attackers can exploit ICMP to launch DoS attacks, overwhelming network devices and causing downtime. This could lead to unavailability of banking services, disruption of operations, and potential regulatory violations.', 'h': 'Disable ICMP on the firewall and enable ping flood protection. Monitor ICMP traffic for anomalies to prevent denial-of-service attacks and ensure business continuity.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # guiHttpsCliSSH
            'compliance': {'a': 'Compliance', 'b': 'GUI is accessed via HTTPS only; CLI via SSH only.', 'd': 'All administrative access is encrypted to prevent interception.', 'f': 'Ensures confidentiality and integrity of configuration data during access.', 'h': 'Regularly verify access protocols remain restricted to HTTPS/SSH.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'GUI has not only HTTPS Access and CLI has not only SSH Access.', 'd': 'GUI and CLI allow unencrypted connections (HTTP, Telnet).', 'f': 'Attackers can intercept unencrypted traffic, capture administrative credentials, and modify firewall configurations. This can lead to complete compromise of network security, unauthorized access to sensitive banking systems, and regulatory violations.', 'h': 'Restrict all administrative access to HTTPS for GUI and SSH for CLI only. Regularly audit firewall access logs and enforce strong encryption policies to prevent credential theft.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # auditTrail
            'compliance': {'a': 'Compliance', 'b': 'Audit trail enabled and logged entries approved by IT authorities.', 'd': 'All changes to firewall rule base and configurations are logged and approved.', 'f': 'Ensures traceability of administrative actions and accountability for configuration changes.', 'h': 'Periodically review audit logs and maintain approval workflow.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Audit trail is Disable on the firewall to log the changes made to the rule base settings and logged entries were not approved by higher authorities in the IT Department.', 'd': 'It was observed that Audit trail is Disable on the firewall to log the changes made to the rule base settings and logged entries were not approved by higher authorities in the IT Department.', 'f': 'Audit trails are used to verify and track many types of change in firewall security configuration, inbound and outbound firewall rules and enabling and disabling security features. If the audit trail was not available then  then the bank will not be able to verify and track the user who made changes to the firewall configuration.  If logged entries were not approved by higher authorities in the IT Department the activities performed by the malicious attacker will go unnoticed and bank will not be able to identify the security breach. Thus bank might face financial loss, reputational loss because of cyber attack.', 'h': 'It is recommended that audit trail must be enabled on the firewall to log the changes made to the rule base settings and the logged entries are approved by higher authorities in the IT Department.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # defaultCredentials
            'compliance': {'a': 'Compliance', 'b': 'Default credentials not used; unique admin accounts exist.', 'd': 'Administrative accounts have unique, strong credentials.', 'f': 'Prevents unauthorized access through commonly known defaults.', 'h': 'Periodically review and enforce strong password policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Default credentials were in use.', 'd': 'Firewall administrator accounts are accessed using default usernames and passwords.', 'f': 'Attackers can easily gain full administrative access, modify firewall rules, bypass security, and exfiltrate sensitive data. This exposes critical banking infrastructure to cyberattacks, ransomware, and compliance violations.', 'h': 'Change all default credentials to strong, unpredictable passwords. Implement role-based access control to limit administrative privileges.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # dosSpoofProtection
            'compliance': {'a': 'Compliance', 'b': 'DoS and Spoof protection enabled.', 'd': 'Firewall configured to detect and mitigate DoS attacks and IP spoofing.', 'f': 'Prevents service disruption and protects network from external threats.', 'h': 'Regularly test protection settings against simulated attacks.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'DoS and Spoof protection was disabled.', 'd': 'Firewall cannot prevent DoS or spoofing attacks.', 'f': 'Attackers can overwhelm network devices with DoS attacks, intercept or spoof IP packets, and disrupt banking services. This can lead to financial loss, reputational damage, and regulatory non-compliance.', 'h': 'Enable DoS and Spoof protection immediately. Configure thresholds and monitor network traffic for anomalies to detect attacks early.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # firmwareUpdated
            'compliance': {'a': 'Compliance', 'b': 'Firmware updated to latest version.', 'd': 'Firewall firmware includes latest security patches and performance improvements.', 'f': 'Reduces risk of exploitation through known vulnerabilities.', 'h': 'Schedule regular firmware updates and review release notes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Firmware was not updated.', 'd': 'Firewall firmware has not been updated to the latest version.', 'f': 'Attackers can exploit known vulnerabilities in older firmware versions, bypass security controls, and compromise the network. Legacy firmware may also lack performance optimizations, causing inefficient processing of traffic and delayed threat detection.', 'h': 'Update firmware to the latest version promptly. Ensure updates include security patches, performance improvements, and maintain a rollback plan in case of upgrade failure.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # supportLicense
            'compliance': {'a': 'Compliance', 'b': 'Firewall has a valid support license.', 'd': 'Firewall receives regular updates, patches, and vendor support.', 'f': 'Ensures continuous security updates and access to technical support.', 'h': 'Periodically verify license validity and renew before expiration.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The firewall does not have a valid support license available.', 'd': 'Firewall lacks a valid supported license; updates and vendor support are unavailable.', 'f': 'Firewall will not receive critical security patches, leaving it vulnerable to known exploits. In case of failure or misconfiguration, technical support is unavailable, increasing downtime and potential financial and reputational loss. Attackers can exploit outdated firmware and configuration loopholes to compromise the bank network.', 'h': 'Obtain a valid support license immediately to ensure continuous security updates, vendor assistance, and timely vulnerability mitigation.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # firewallBranches
            'compliance': {'a': 'Compliance', 'b': 'Firewall installed at all branch locations.', 'd': 'Branch networks are protected from unauthorized access and malicious traffic.', 'f': 'Prevents intrusions and malware propagation from branch networks to the core network.', 'h': 'Regularly monitor branch firewalls and ensure they remain updated.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A firewall was not present in branches.  ', 'd': 'Branch networks are unprotected against external threats and unauthorized access.', 'f': 'Branch systems can become entry points for attackers, spreading malware or exploiting vulnerabilities to compromise the bank network. This increases risk of data breaches, financial fraud, and regulatory non-compliance, while reducing visibility over branch-level network activities.', 'h': 'Install hardware or software firewalls in all branches. If costs are high, deploy software-based firewall and antivirus solutions for endpoint protection, and ensure centralized monitoring.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # systemLogs
            'compliance': {'a': 'Compliance', 'b': 'System logs configured and actively monitored.', 'd': 'Firewall activities, rule changes, and user actions are logged for traceability.', 'f': 'Enables monitoring, forensic investigations, and compliance reporting.', 'h': 'Periodically review logs for anomalies and retain them as per policy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'System logs were not configured.', 'd': 'Firewall does not record activity logs.', 'f': 'Without logs, unauthorized or malicious changes to firewall rules go undetected. This hampers forensic investigation, incident response, and regulatory compliance. Lack of visibility can lead to prolonged undetected breaches, increasing the risk of data theft, service disruption, and financial losses.', 'h': 'Configure system logging on the firewall. Ensure logs are stored securely, monitored, and periodically reviewed for any suspicious activity.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        18: {  # backupFrequency
            'compliance': {'a': 'Compliance', 'b': 'Firewall configuration backup taken regularly.', 'd': 'Configurations can be restored in case of failure or misconfiguration.', 'f': 'Minimizes downtime and preserves firewall policy integrity.', 'h': 'Schedule periodic backups and store them securely offsite.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'No backup was taken for the firewall configuration.', 'd': 'Firewall firmware or configuration changes are not backed up.', 'f': 'If the firewall fails, is compromised, or misconfigured, the previous configuration cannot be restored. This leads to extended downtime, security gaps, and potential disruption of banking operations. Attackers could exploit downtime to access the network without restrictions.', 'h': 'Implement regular automated backups using TFTP/SFTP servers. Backup after any configuration changes or firmware updates and test restore procedures periodically.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        19: {  # adminWhitelistIP
            'compliance': {'a': 'Compliance', 'b': 'Only whitelisted IPs can access administrative portal.', 'd': 'Administrative access restricted to authorized locations.', 'f': 'Reduces the attack surface and prevents unauthorized access.', 'h': 'Maintain and update the whitelist as per policy changes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'It was observed that no IP addresses were whitelisted for accessing the administrative portal.', 'd': 'Any system can access the administrative portal without restriction.', 'f': 'Unauthorized users or compromised devices can access the firewall admin portal, modify rules, and disable protections. This could lead to total compromise of the firewall, network intrusion, and potential financial and reputational damage.', 'h': 'Implement IP whitelisting for administrative portal access. Restrict access to authorized locations and maintain a change log for any updates to whitelisted IPs.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        20: {  # syslogServer
            'compliance': {'a': 'Compliance', 'b': 'Syslog server configured and operational.', 'd': 'Logs from firewalls and network devices are centralized for analysis and storage.', 'f': 'Enables forensic investigations, auditing, and trend analysis of network activity.', 'h': 'Regularly review logs and maintain retention as per policy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Syslog server was not present.  ', 'd': 'Logs are not centralized or stored for long-term analysis.', 'f': 'In case of malicious activity, the bank will struggle to collect and correlate logs, severely hampering forensic investigation and timely incident response. This could lead to unresolved breaches, regulatory non-compliance, and potential financial and reputational losses.', 'h': 'Deploy a dedicated Syslog server and configure all network devices to send logs to it. Ensure logs are retained securely and periodically reviewed for anomalies.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        21: {  # adminHttpsTls
            'compliance': {'a': 'Compliance', 'b': 'Admin portal uses HTTPS with TLS 1.2 or above.', 'd': 'Data transmitted between administrators and portal is encrypted.', 'f': 'Protects credentials and sensitive data from interception.', 'h': 'Periodically verify encryption standards and update TLS protocols when necessary.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Administrator access portal running on TLS 1.2 or older versions.', 'd': 'Data transmitted may be vulnerable to interception or downgrade attacks.', 'f': 'Attackers can exploit weak TLS protocols to intercept sensitive information, including admin credentials, leading to unauthorized access to the firewall and potential network compromise. This could result in data breaches, service disruption, and reputational damage.', 'h': 'Upgrade the admin portal to TLS 1.3 or latest standards. Ensure periodic security testing to validate encryption strength and prevent vulnerabilities.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        22: {  # httpLoginDisabled
            'compliance': {'a': 'Compliance', 'b': 'HTTP logins disabled or redirected to HTTPS.', 'd': 'All login sessions are encrypted using secure channels.', 'f': 'Protects against credential interception and man-in-the-middle attacks.', 'h': 'Regularly test redirection and encryption configurations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'HTTP logins were not disabled.', 'd': 'Users can login via unsecured HTTP channels.', 'f': 'Attackers can sniff network traffic to steal credentials and gain unauthorized access. This jeopardizes firewall configuration integrity and network security, increasing the risk of data theft and operational disruption.', 'h': 'Disable HTTP logins entirely and enforce HTTPS redirection. Conduct periodic checks to ensure all administrative access uses encrypted channels.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        23: {  # multipleAdminAccounts
            'compliance': {'a': 'Compliance', 'b': 'Multiple admin accounts exist.', 'd': 'Redundant administrative accounts prevent single point of failure.', 'f': 'Ensures continuity of operations even if one account is locked or compromised.', 'h': 'Periodically review admin accounts and access privileges.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'It was found that only one admin account was created.', 'd': 'Single admin account exists for firewall management.', 'f': 'If the admin account is locked or compromised, critical firewall management operations are disrupted, leading to downtime and potential network exposure. This can affect banking operations and increase the likelihood of a successful cyber attack.', 'h': 'Create multiple admin accounts with controlled privileges to ensure redundancy and continuity of administrative access.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        24: {  # vpnSecure
            'compliance': {'a': 'Compliance', 'b': 'VPN used for remote access.', 'd': 'Data transmitted remotely is encrypted, ensuring confidentiality.', 'f': 'Prevents data interception and protects sensitive information during remote operations.', 'h': 'Ensure VPN is configured with strong encryption and periodically updated.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'VPN was not used to ensure the security of channel.', 'd': 'Remote connections occur over unsecured channels.', 'f': 'Data can be intercepted in transit, exposing confidential information. Attackers can exploit unsecured connections to infiltrate the network, potentially leading to data breaches, financial loss, and regulatory penalties.', 'h': 'Implement a secure VPN solution with robust encryption for all remote access. Regularly test VPN security and access controls.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        25: {  # ntpConfigured
            'compliance': {'a': 'Compliance', 'b': 'NTP configured and synchronized.', 'd': 'Firewall and devices maintain accurate time for logs and operations.', 'f': 'Ensures accurate timestamps for auditing, troubleshooting, and forensic investigations.', 'h': 'Periodically verify NTP synchronization across all critical devices.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'NTP was not configured.', 'd': 'Devices are not synchronized with a trusted time source.', 'f': 'Inaccurate timestamps can make it extremely difficult to correlate events across systems during a security breach. This can hinder forensic investigations, delay incident response, and create compliance issues with regulatory guidelines requiring accurate log timestamps.', 'h': 'Configure NTP on all firewalls and critical network devices to synchronize with a trusted server. Regularly monitor synchronization status to ensure accurate event logging.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        26: {  # firewallPhysicallySecured
            'compliance': {'a': 'Compliance', 'b': 'Firewall placed in a secured area or locked cabinet.', 'd': 'Physical access is restricted to authorized personnel.', 'f': 'Prevents tampering, theft, and unauthorized configuration changes.', 'h': 'Periodically verify physical security measures and access logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Firewall was not secured physically.', 'd': 'Firewall is accessible to unauthorized personnel.', 'f': 'An attacker could physically tamper with the firewall, disconnect network cables, or change configuration settings. This compromises network availability, integrity, and confidentiality, increasing the risk of cyberattacks and operational disruption.', 'h': 'Place all firewalls in locked cabinets or server racks with restricted access. Maintain physical access logs and limit keys to accountable personnel only.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        27: {  # alertsMonitored
            'compliance': {'a': 'Compliance', 'b': 'Alerts monitored regularly.', 'd': 'Security events and anomalies are promptly addressed.', 'f': 'Enables proactive detection and mitigation of threats.', 'h': 'Maintain a schedule for monitoring and reviewing alerts.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Alerts were not being monitored properly.', 'd': 'Security alerts generated by the firewall or monitoring tools are ignored or reviewed irregularly.', 'f': 'Delayed detection of cyber threats can allow attackers to exploit vulnerabilities, compromise sensitive data, and disrupt operations. This increases the risk of financial loss, reputational damage, and non-compliance with regulatory requirements.', 'h': 'Implement a defined process to monitor alerts continuously. Assign responsibilities to IT personnel and ensure timely investigation of all critical alerts.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        28: {  # passwordPolicies
            'compliance': {'a': 'Compliance', 'b': 'Password policies configured as per guidelines.', 'd': 'Users are required to create strong passwords with complexity and expiry rules.', 'f': 'Reduces risk of unauthorized access through weak passwords.', 'h': 'Periodically review password policies for compliance and enforce policy updates.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Password was not configured as per the organizations password management policy.', 'd': 'Weak or default passwords used for firewall administrative accounts.', 'f': 'Weak passwords increase the probability of brute-force attacks, potentially allowing attackers to gain administrative access, manipulate firewall rules, and compromise network security. This could result in data breaches, operational downtime, and regulatory penalties.', 'h': 'Enforce strong password policies including complexity, minimum length, expiry, and history. Audit compliance periodically to prevent unauthorized access.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        29: {  # unusedPortsClosed
            'compliance': {'a': 'Compliance', 'b': 'All unused ports are closed.', 'd': 'Only necessary services are exposed through the firewall.', 'f': 'Minimizes attack surface and reduces potential entry points for attackers.', 'h': 'Periodically audit firewall ports and disable unnecessary services.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The unused ports were open on firewall.', 'd': 'Firewall exposes unnecessary ports that are not in use.', 'f': 'Attackers can exploit open ports to gain unauthorized access, probe services, and perform network reconnaissance. This significantly increases the risk of intrusion, malware infection, and potential compromise of sensitive banking data.', 'h': 'Close all unused ports on the firewall and periodically audit port usage. Ensure only required ports are open for authorized traffic.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        30: {  # idsLogsMonitored
            'compliance': {'a': 'Compliance', 'b': 'IDS logs are monitored regularly.', 'd': 'System administrators review IDS logs to detect suspicious activity.', 'f': 'Helps in early detection and prevention of cyberattacks.', 'h': 'Maintain a schedule for monitoring and reviewing IDS logs to ensure timely response.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The system administrators were not monitoring the logs produced by the Intrusion Detection System.', 'd': 'It was observed that the system administrators were not monitoring the logs produced by the Intrusion Detection System.', 'f': 'Unauthorized intrusions or attacks may go undetected, allowing attackers to move laterally within the network, steal sensitive data, or disrupt operations. This increases the risk of financial loss, reputational damage, and regulatory non-compliance.', 'h': 'Assign responsibility to designated administrators for continuous monitoring of IDS logs. Implement alerting mechanisms for critical events to ensure immediate response.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        31: {  # twoFactorAuth
            'compliance': {'a': 'Compliance', 'b': 'Two-factor authentication (2FA) is enabled.', 'd': 'Administrative access requires an additional authentication factor.', 'f': 'Reduces the risk of unauthorized access even if credentials are compromised.', 'h': 'Regularly review and enforce 2FA for all administrative accounts.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Two factor authentication was not enabled for the administration access.', 'd': 'It was observed that two factor authentication was not enabled for the administration access. If anyone gets credentials, then they can easily access admin portal.', 'f': 'If credentials are leaked or guessed, attackers can gain full control over the firewall and network infrastructure. This could lead to unauthorized configuration changes, data breaches, and compromise of banking operations.', 'h': 'Enable two-factor authentication for all administrative accounts and enforce periodic review of access credentials to enhance security.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        32: {  # loginDisclaimer
            'compliance': {'a': 'Compliance', 'b': 'Login disclaimer is displayed.', 'd': 'Users are informed about authorized access and monitoring policies.', 'f': 'Provides legal warning and deters unauthorized access.', 'h': 'Keep the login disclaimer updated and prominently displayed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The login disclaimer message was not set.', 'd': 'No warning or disclaimer is displayed on the login page.', 'f': 'Unauthorized users may not be aware that their actions are being monitored, and legal protection for the bank is reduced. This could complicate investigations and legal actions in case of a breach.', 'h': 'Configure a login disclaimer on all administrative portals to notify users about authorized use and monitoring policies, enhancing both security and legal compliance.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        33: {  # aclConfigured
            'compliance': {'a': 'Compliance', 'b': 'ACL configured properly.', 'd': 'Only authorized traffic is allowed; unauthorized traffic is denied.', 'f': 'Reduces exposure to external and internal threats.', 'h': 'Periodically audit ACL configurations for compliance and effectiveness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'ACL was not configured.', 'd': 'Firewall allows all traffic without restrictions.', 'f': 'Malicious insiders or external attackers can access sensitive systems and network services freely. This increases the risk of unauthorized access, data exfiltration, and compromise of critical banking operations.', 'h': 'Implement ACLs to restrict access to only authorized IP addresses, ports, and services. Regularly review and update ACLs to maintain security.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        34: {  # loginTimeout
            'compliance': {'a': 'Compliance', 'b': 'Login/session timeout configured.', 'd': 'Sessions automatically terminate after a period of inactivity.', 'f': 'Reduces risk of unauthorized access from unattended sessions.', 'h': 'Periodically review session timeout settings to ensure policy compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Login/Session timeout was not set.', 'd': 'It was observed that login/session timeout was not set.', 'f': 'An attacker can exploit unattended sessions to gain unauthorized access, modify firewall rules, or steal sensitive information, potentially disrupting banking operations and causing regulatory non-compliance.', 'h': 'Configure session timeouts in accordance with organization policy, ideally less than 10 minutes of inactivity, to mitigate risks from abandoned sessions.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    
    # Populate data based on user input
    for i, question_field in enumerate(question_mapping, 1):
        user_input = form_data.get(question_field, 'Not Applicable') if form_data else 'Not Applicable'
        
        if user_input.lower() == 'compliance':
            response_data = question_responses[i]['compliance']
        elif user_input.lower() == 'non-compliance':
            response_data = question_responses[i]['non_compliance']
        else:
            response_data = question_responses[i]['not_applicable']
        
        # Populate columns C, D, F, G, H
        ws.cell(row=i+1, column=3, value=response_data['a'])  # Compliance/Non-Compliance/Not Applicable
        ws.cell(row=i+1, column=4, value=response_data['b'])  # Observation (Short/Brief)
        ws.cell(row=i+1, column=6, value=response_data['d'])  # Observation
        ws.cell(row=i+1, column=7, value=response_data['f'])  # Impact
        ws.cell(row=i+1, column=8, value=response_data['h'])  # Recommendation
    
    # Apply formatting to data cells
    for row in range(2, 36):  # Rows 2-35 for data (34 questions)
        for col in range(1, 9):  # Columns A-H
            cell = ws.cell(row=row, column=col)
            if col in [2, 4, 6, 7, 8]:  # Columns B, D, F, G, H - left horizontal, center vertical
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            elif col == 3:  # Column C - center alignment
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:  # Other columns (A, E) - center alignment
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Populate Risk Factor column (E) with color coding
    for i, risk_factor in enumerate(risk_factors, 2):
        cell = ws.cell(row=i, column=5, value=risk_factor)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)  # White text, bold
        if risk_factor == 'Critical':
            cell.fill = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')  # Dark Red
        elif risk_factor == 'High':
            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
        elif risk_factor == 'Medium':
            cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # Orange
        elif risk_factor == 'Low':
            cell.fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')  # Green
    
    # Apply borders to all cells with content
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for row in range(1, 36):  # Header + 34 data rows (1-35)
        for col in range(1, 9):  # Columns A-H
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
    
    # Set row height for wrapped text
    for row in range(2, 36):  # Data rows 2-35
        ws.row_dimensions[row].height = 30
    
    # Save the file
    filename = "Firewall Review.xlsx"
    filepath = os.path.join(os.path.dirname(__file__), '..', '..', 'static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    """
    Delete the generated file after download
    """
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            print(f"File cleaned up: {filepath}")
    except Exception as e:
        print(f"Error cleaning up file {filepath}: {e}")
