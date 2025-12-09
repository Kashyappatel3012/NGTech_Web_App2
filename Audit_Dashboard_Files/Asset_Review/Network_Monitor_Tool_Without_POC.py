import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_network_monitor_tool_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Network Monitor Tool"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # Network Monitor Tool Questions
    questions = [
        "Does the bank have a Network Monitoring Tool implemented?",
        "Have the Network data monitoring tools (e.g., sniffers, data scopes, and probes) utilized by the product/service been approved by the Bank's IT Department?"
    ]

    # Risk Factors
    risk_factors = [
        "High", "High"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "nmtNetworkMonitoringTool": 1,
        "nmtToolsApproved": 2
    }

    # Question responses data
    question_responses = {
        1: {  # nmtNetworkMonitoringTool
            'compliance': {'a': 'Compliance', 'b': 'Network monitoring tool implemented.', 'd': 'The bank has implemented a robust network monitoring tool that provides real-time visibility into network performance, security, and availability. Alerts and reports are regularly reviewed by the IT team to ensure timely detection of anomalies.', 'f': 'Enhances proactive network management, minimizes downtime, and strengthens incident detection and response.', 'h': 'Continue periodic calibration of monitoring thresholds and ensure the tool\'s integration with SIEM and firewall logs for unified visibility.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Network monitoring tool not implemented.', 'd': 'It was observed that the bank does not have an active network monitoring system in place to continuously track bandwidth usage, detect anomalies, or identify security threats. Network activities are being monitored manually or reactively after incidents occur.', 'f': 'Lack of automated network monitoring may lead to delayed detection of network failures, intrusions, or unauthorized activities, increasing the risk of data breaches or downtime.', 'h': 'Implement an enterprise-grade Network Monitoring Tool (e.g., SolarWinds, Nagios, PRTG, or ManageEngine Manager) to ensure continuous visibility, real-time alerts, and performance optimization of network infrastructure.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # nmtToolsApproved
            'compliance': {'a': 'Compliance', 'b': 'Network monitoring tools formally approved.', 'd': 'All network monitoring tools and probes have been reviewed and formally approved by the Bank\'s IT Department. Documentation, version control, and vendor support agreements are maintained as per IT governance standards.', 'f': 'Reduces operational and security risks by ensuring authorized and standardized network monitoring practices.', 'h': 'Periodically review tool approvals and configurations to ensure alignment with updated security policies and compliance standards.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Network monitoring tools not approved by IT Department.', 'd': 'It was observed that certain network data monitoring tools (sniffers, probes) were in use without documented approval or verification by the Bank\'s IT Department. These tools lacked authorization records and were not listed in the official IT inventory.', 'f': 'Unapproved tools can pose a security threat by capturing sensitive data packets, exposing confidential information, or creating unmonitored network entry points.', 'h': 'Immediately review, document, and remove unauthorized tools. Enforce a formal approval process for any network monitoring utilities, ensuring all tools are verified, approved, and logged in the IT asset management register.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Network Monitor Tool Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
