import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_data_input_control_excel(form_data=None):
    """
    Create Data Input Control Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Input Control"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data Input Control Questions
    questions = [
        "Is the history of signatures scanned available in the system?",
        "Is the entire stock of cheque books fed into the system?",
        "Are the issued cheque books entered and confirmed in the system on a day-to-day basis?"
    ]

    # Risk Factors
    risk_factors = [
        "High",
        "Medium", 
        "Medium"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "dicSignatureHistory": 1,
        "dicChequeBookStock": 2,
        "dicDailyChequeBookEntry": 3
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Question responses data
    question_responses = {
        1: {  # dicSignatureHistory
            'compliance': {'a': 'Compliance', 'b': 'Signature history available digitally.', 'd': 'The system maintains scanned records of all customer signatures, providing an accessible audit trail for transaction verification and dispute resolution.', 'f': 'Enhances operational efficiency, reduces the risk of fraudulent transactions, and strengthens audit and compliance capabilities.', 'h': 'Regularly backup signature records and monitor access controls to prevent unauthorized use.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'History of scanned signatures was not available in the system.', 'd': 'The bank does not maintain scanned digital records of customer signatures in the system, or records are incomplete and not easily accessible.', 'f': 'Lack of digital signature history increases the risk of unauthorized transactions, delays verification processes, and complicates audit trails for dispute resolution.', 'h': 'Implement a system to capture and maintain scanned signature history for all customers, ensuring accessibility, integrity, and security for verification and auditing purposes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # dicChequeBookStock
            'compliance': {'a': 'Compliance', 'b': 'Entire stock of cheque books recorded in the system.', 'd': 'All cheque books, including new stock, issued, and reserved ones, are accurately recorded in the system, allowing proper tracking and management.', 'f': 'Prevents unauthorized issuance, ensures accountability, and supports proper inventory management and audits.', 'h': 'Periodically reconcile system records with physical stock to maintain accuracy and integrity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The entire stock of cheque books was not fed to the system.', 'd': 'The bank\'s system does not contain records of the entire stock of cheque books, or entries are incomplete or inconsistent.', 'f': 'Missing or inaccurate records can lead to mismanagement, unauthorized issuance, or fraudulent use of cheque books, affecting operational and financial control.', 'h': 'Ensure that the full stock of cheque books is recorded in the system, with proper reconciliation against physical inventory.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # dicDailyChequeBookEntry
            'compliance': {'a': 'Compliance', 'b': 'Issued cheque books updated daily.', 'd': 'Daily issuance of cheque books is promptly recorded and confirmed in the system, maintaining accurate and up-to-date records for all branches.', 'f': 'Enhances operational control, reduces risk of fraud, and ensures proper tracking for audit and regulatory purposes.', 'h': 'Implement automated alerts or checks to ensure daily updates are completed without delay.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The cheque books issued were not entered and confirmed in the system on day-to-day basis.', 'd': 'The bank does not consistently update the system with daily issued cheque books, or confirmations are delayed, leading to incomplete transaction and inventory records.', 'f': 'Delays in recording issued cheque books may result in reconciliation errors, risk of fraud, and inaccurate reporting for audits.', 'h': 'Update and confirm all issued cheque books in the system on a daily basis, ensuring alignment between physical issuance and digital records.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }

    # Populate data based on user input
    for i, question in enumerate(questions, 2):
        # Get user input for this question
        question_num = i - 1
        user_input = None
        
        if form_data:
            # Find the corresponding form field
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Data Input Control Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
