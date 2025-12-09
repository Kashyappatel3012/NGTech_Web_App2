import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_purging_data_files_excel(form_data=None):
    """
    Create Purging of Data Files Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Purging of Data Files"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Purging of Data Files Questions
    questions = [
        "Is purged backup media kept properly under safe custody?",
        "Is the purging activity recorded and maintained in a register?",
        "Is access to purged data restricted?"
    ]

    # Risk Factors
    risk_factors = [
        "High",
        "Medium", 
        "Medium"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "pdfBackupMediaCustody": 1,
        "pdfPurgingActivityRecorded": 2,
        "pdfAccessRestricted": 3
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Question responses data
    question_responses = {
        1: {  # pdfBackupMediaCustody
            'compliance': {'a': 'Compliance', 'b': 'Purged backup media stored securely.', 'd': 'All purged or retired backup media are stored under secure custody with restricted access and appropriate logging to ensure accountability.', 'f': 'Minimizes the risk of data leakage, protects sensitive information, and supports compliance with regulatory requirements.', 'h': 'Periodically audit storage practices and access logs to ensure continued secure custody of purged media.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Purged backup media was not kept properly under safe custody.', 'd': 'Backup media that have been purged or retired are not kept in a secure location, leaving them vulnerable to unauthorized access or theft.', 'f': 'If purged backup media is not kept in safe custody, then anyone can access those files or data. If data gets infected with malicious files, then there is the possibility that the system will get a compromise. Thus, the CIA triad will get compromised.', 'h': 'Store purged backup media in a secure, access-controlled area with proper logging and physical safeguards to prevent unauthorized access.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # pdfPurgingActivityRecorded
            'compliance': {'a': 'Compliance', 'b': 'Purging activities documented.', 'd': 'All purging activities are properly recorded in a register with details of the media, responsible personnel, and authorization, providing a clear audit trail.', 'f': 'Ensures accountability, supports audits, and reduces the risk of data loss or unauthorized deletion.', 'h': 'Review the purging register regularly to confirm accuracy and compliance with data retention policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Purging activity was not recorded and maintained in a register.', 'd': 'Activities related to purging backup media or data are not recorded in a register, or records are incomplete and inconsistent.', 'f': 'If the bank does not maintain a register for purging activity it is difficult to identify the malicious file’s point of origin. The user responsible for deleting data would not be found.', 'h': 'Maintain a detailed register of all purging activities, including date, media details, person responsible, and authorization for the activity.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # pdfAccessRestricted
            'compliance': {'a': 'Compliance', 'b': 'Access to purged data restricted.', 'd': 'Access to purged data is limited to authorized personnel, and physical or logical safeguards are in place to prevent misuse.', 'f': 'Protects sensitive information, ensures regulatory compliance, and mitigates risks of data breaches.', 'h': 'Periodically review access permissions and logs to verify that only authorized personnel can handle purged data.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Access to purged data was not restricted.', 'd': 'Individuals may access data from backup media that should have been purged, due to weak access controls or improper storage practices.', 'f': 'If purged backup media is not kept in safe custody, then anyone can access those files or data. If data gets infected with malicious files, then there is the possibility that the system will get a compromise. Thus, the CIA triad will get compromised.', 'h': 'Restrict access to purged data and retired media to authorized personnel only, with proper logging and supervision.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }

    # Populate data based on user input
    for i, question in enumerate(questions, 2):
        # Get user input for this question
        question_num = i - 1
        user_input = None
        
        if form_data:
            # Find the corresponding form field
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Purging of Data Files Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
