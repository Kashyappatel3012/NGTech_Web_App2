import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_amc_excel(form_data=None):
    """
    Create AMC Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "AMC"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # AMC Questions
    questions = [
        "Is access for maintenance purposes granted only after verifying the identity of the service person?",
        "Are stamped agreements for maintenance contracts executed and available?",
        "Are activities carried out during maintenance reported in the registers and duly authenticated?",
        "Are contract renewal rates maintained in the register?",
        "Is the maintenance staff support available in time?"
    ]

    # Risk Factors
    risk_factors = [
        "High",
        "Medium", 
        "Medium",
        "Medium",
        "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "amcIdentityVerification": 1,
        "amcStampedAgreements": 2,
        "amcActivityReporting": 3,
        "amcContractRenewalRates": 4,
        "amcStaffSupportAvailability": 5
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Question responses data
    question_responses = {
        1: {  # amcIdentityVerification
            'compliance': {'a': 'Compliance', 'b': 'Access for maintenance purposes is granted only after proper verification of the identity and authorization.', 'd': 'All maintenance personnel are verified before access is granted, with proper documentation and approval recorded in access logs.', 'f': 'This ensures that only verified and authorized individuals perform maintenance activities, thereby reducing the risk of unauthorized access, data tampering, and system compromise.', 'h': 'Regularly audit access logs and ensure procedures are followed consistently.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Access for maintenance purposes was granted without verifying the identity of the service person.', 'd': 'Maintenance staff are allowed access to server rooms, UPS rooms, and other critical areas without proper verification of their identity or authorization.', 'f': 'Unauthorized personnel may gain access to sensitive areas, increasing the risk of theft, sabotage, or compromise of critical systems and data.', 'h': 'Implement strict identity verification procedures for all maintenance personnel, including checking ID cards, access passes, and authorization letters before granting entry. Maintain access logs for audits.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # amcStampedAgreements
            'compliance': {'a': 'Compliance', 'b': 'Maintenance agreements executed and available.', 'd': 'All maintenance contracts are formally executed, stamped, and accessible, ensuring clarity on responsibilities, scope, and timelines.', 'f': 'Strengthens accountability, ensures timely maintenance, and reduces potential disputes with service providers.', 'h': 'Periodically review and update contracts to reflect changes in service requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Stamped agreements for maintenance contracts were not available.', 'd': 'Maintenance contracts or service agreements are not formally executed, stamped, or maintained, leading to ambiguity regarding terms of service and accountability.', 'f': 'Absence of formal agreements may cause disputes, lack of accountability, delayed service, and potential financial loss during maintenance activities.', 'h': 'Ensure all maintenance contracts are formally executed, stamped, and stored securely. Include clauses defining scope, timelines, service levels, and responsibilities.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # amcActivityReporting
            'compliance': {'a': 'Compliance', 'b': 'Maintenance activities properly documented.', 'd': 'All maintenance activities are recorded in registers and authenticated by responsible personnel, providing a clear audit trail.', 'f': 'Ensures accountability, enables tracking of maintenance history, and supports compliance during audits.', 'h': 'Regularly review maintenance logs to verify completeness and accuracy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Activities carried out during maintenance were not reported in the registers.', 'd': 'Maintenance tasks performed by staff or service providers are not documented in registers or authenticated by responsible authorities.', 'f': 'Lack of proper documentation may lead to unverified work, missed maintenance, unresolved issues, and difficulty during audits or investigations.', 'h': 'Maintain detailed logs of all maintenance activities, including date, time, person performing the work, and authentication by the supervisor.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # amcContractRenewalRates
            'compliance': {'a': 'Compliance', 'b': 'Contract renewal rates maintained.', 'd': 'All maintenance contract renewal rates are recorded in a register, ensuring visibility and accountability for cost management.', 'f': 'Facilitates timely renewals, prevents unauthorized charges, and ensures budget compliance.', 'h': 'Update the register regularly and reconcile with financial records.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Contract renewal rates not recorded.', 'd': 'Renewal rates, terms, and schedules for maintenance contracts are not systematically maintained in a register or database.', 'f': 'May lead to financial discrepancies, missed renewals, or unauthorized cost escalations.', 'h': 'Maintain a dedicated register documenting contract renewal rates, renewal dates, and terms. Review periodically to avoid lapses or disputes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # amcStaffSupportAvailability
            'compliance': {'a': 'Compliance', 'b': 'Maintenance support available as required.', 'd': 'Maintenance staff or vendor support is available promptly, with clear escalation procedures and defined service level agreements (SLAs).', 'f': 'Reduces downtime, ensures timely resolution of issues, and maintains operational continuity.', 'h': 'Periodically test response times and review SLAs to ensure support is consistently available when needed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The maintenance staff support was not available in time.', 'd': 'Maintenance staff or vendor support is not readily available during breakdowns or scheduled maintenance, causing prolonged downtime.', 'f': 'Delays in maintenance can affect critical IT operations, resulting in service disruption, potential financial loss, and compromised system reliability.', 'h': 'Ensure availability of trained maintenance personnel or service provider support within agreed SLAs, including emergency contact mechanisms for critical failures.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }

    # Populate data based on user input
    for i, question in enumerate(questions, 2):
        # Get user input for this question
        question_num = i - 1
        user_input = None
        
        if form_data:
            # Find the corresponding form field
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "AMC Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
