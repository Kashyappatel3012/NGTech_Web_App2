import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_core_switch_excel(form_data=None):
    """
    Create Excel file for Core Switch Assessment
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Core Switch"
    
    # Define questions
    questions = [
        "Is Inbound TCP Connection Keep Alives message disabled?",
        "Is telnet service enabled?",
        "Does the GUI only has HTTPS Access? And CLI only has SSH Access?",
        "Is USB access allowed in switch?",
        "Are CDP (Cisco Discovery Protocol) Packets enabled?",
        "Does Cisco enable TCP and UDP Small Servers by default?",
        "Does default credentials are in use? Or are accounts name by the 'Admin' or 'root' or 'Administrator' to access administrator login?",
        "Are unused ports closed on the switch?",
        "Is the Switch Physically Secured or not?",
        "Is NTP (Network Time Protocol) configured?",
        "Are HTTP logins disabled or not? And whether it redirected to HTTPS?",
        "Is the firmware up to date?",
        "Is an administrative portal login accessed only by Whitelisted IP Addresses?",
        "Is Admin lockout set to 3 or more failed login attempts or as per the organisation's policy?",
        "Is logging enabled? Is the Syslog server implemented?",
        "Are Password Policies configured as per the organization's password management policy?",
        "How frequent is backup done for the switch configuration?",
        "Does the core switch have a secure password enabled?",
        "Was the BOOTP not disabled?",
        "Is the login disclaimer message set or not?",
        "Is login timeout as per organization policies?",
        "Are ICMP IP unreachable messages enabled on the interface mgmt0?",
        "Are Domain lookups disabled and DNS servers configured?",
        "Is Packet Assembler / Disassembler (PAD) service-disabled?",
        "Is Maintenance Operations Protocol (MOP) disabled on all ethernet interfaces?",
        "Are HTTPS and SSH ports changed to nonstandard ports?",
        "Are Directed broadcasts enabled by default?",
        "Does the device have support for ICMP redirects enabled on the network interface mgmt0?",
        "Does the Auditor identified any interface that had Proxy ARP enabled, mgmt0?"
    ]
    
    # Risk factors for each question
    risk_factors = [
        'High', 'High', 'High', 'High', 'High', 'High', 'Medium', 'Medium', 'Medium', 'Medium',
        'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Low', 'Low',
        'Low', 'Low', 'Low', 'Low', 'Low', 'Low', 'Low', 'Low', 'Low'
    ]
    
    # Question mapping for form fields
    question_mapping = [
        'tcpKeepAlives', 'telnetEnabled', 'guiHttpsCliSSH', 'usbAccess', 'cdpEnabled',
        'tcpUdpSmallServers', 'defaultCredentials', 'unusedPortsClosed', 'physicallySecured', 'ntpConfigured',
        'httpLoginsDisabled', 'firmwareUpdated', 'adminWhitelistIP', 'adminLockout', 'loggingEnabled',
        'passwordPolicies', 'backupFrequency', 'securePassword', 'bootpDisabled', 'loginDisclaimer',
        'loginTimeout', 'icmpUnreachable', 'domainLookups', 'padService', 'mopDisabled',
        'httpsSshPorts', 'directedBroadcasts', 'icmpRedirects', 'proxyArp'
    ]
    
    # Set column widths
    ws.column_dimensions['A'].width = 10  # Sr. No.
    ws.column_dimensions['B'].width = 50  # Questions
    ws.column_dimensions['C'].width = 20  # Compliance/Non-Compliance/Not Applicable
    ws.column_dimensions['D'].width = 30  # Observation (Short/Brief)
    ws.column_dimensions['E'].width = 20  # Risk Factor
    ws.column_dimensions['F'].width = 50  # Observation
    ws.column_dimensions['G'].width = 50  # Impact
    ws.column_dimensions['H'].width = 50  # Recommendation
    
    # Header row
    headers = ['Sr. No.', 'Questionnaire/Points', 'Compliance/Non-Compliance/Not Applicable', 
               'Observation (Short/Brief)', 'Risk Factor', 'Observation', 'Impact', 'Recommendation']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Populate Sr. No. and Questions
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }
    
    # Question responses data
    question_responses = {
        1: {  # tcpKeepAlives
            'compliance': {'a': 'Compliance', 'b': 'Inbound TCP Keep-Alives message disabled.', 'd': 'TCP Keep-Alive feature was disabled, ensuring unused sessions do not remain active unnecessarily.', 'f': 'Prevents unauthorized persistence of idle sessions, enhancing session security.', 'h': 'Continue monitoring TCP configurations to ensure security best practices remain enforced.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Inbound TCP Connection Keep Alives message was enabled.', 'd': 'TCP Keep-Alive feature was enabled, allowing idle sessions to persist longer.', 'f': 'An attacker could attempt a DoS by exhausting the number of possible connections. The TCP keep-alive mechanism does not replace the active OOS check, which should be configured as usual even if the TCP keep-alive feature is enabled.', 'h': 'Disable inbound TCP Keep-Alive messages to reduce idle connection risks and improve security posture.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # telnetEnabled
            'compliance': {'a': 'Compliance', 'b': 'Telnet disabled.', 'd': 'Telnet service was disabled, and SSH was used for remote management.', 'f': 'Prevents unencrypted communication and enhances secure administrative access.', 'h': 'Maintain SSH-only remote access policy and ensure no reactivation of Telnet.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Telnet service was enabled.', 'd': 'Telnet access was enabled on network switches.', 'f': 'Telnet transmits credentials in plaintext, exposing them to interception during transmission. Attackers can gain unauthorized administrative access, leading to configuration tampering or service disruption.', 'h': 'Disable Telnet service and use SSH (Secure Shell) for all remote device administration to ensure encrypted communication.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # guiHttpsCliSSH
            'compliance': {'a': 'Compliance', 'b': 'GUI via HTTPS and CLI via SSH configured.', 'd': 'The switch management interface enforces HTTPS for GUI and SSH for CLI access.', 'f': 'Ensures encrypted communication and prevents credential leakage during remote management.', 'h': 'Regularly review access protocols to confirm HTTPS and SSH enforcement.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'GUI does not have only HTTPS access and CLI does not have only SSH access enabled.', 'd': 'It was observed that the GUI (Graphical User Interface) does not have only HTTPS access enabled, and the CLI (Command-Line Interface) does not have only SSH access enabled. This configuration poses security risks as it leaves the system susceptible to potential unauthorized access and data breaches.', 'f': 'Without HTTPS-only access for the GUI and SSH-only access for CLI, sensitive data transmitted through the web interface may be susceptible to eavesdropping and man-in-the-middle attacks, potentially leading to data breaches and privacy violations. This can result in unauthorized configuration changes, data manipulation, and even complete system compromise.', 'h': 'Disable HTTP/Telnet access and enforce HTTPS/SSH as the only allowed management protocols.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # usbAccess
            'compliance': {'a': 'Compliance', 'b': 'USB access disabled.', 'd': 'USB access was restricted to prevent unauthorized data transfer or firmware installation.', 'f': 'Protects the device from unauthorized data injection or malware infection.', 'h': 'Continue enforcing USB restrictions as part of hardware security.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'USB was accessible in switch port.', 'd': 'USB ports were enabled, allowing external storage devices to connect.', 'f': ' An attacker can attach malicious USB to the switch that can bypass the security and can place virus, malware, change the route of it. Also, he can interrupt the transaction and business process.', 'h': 'It is recommended to disable USB access in switch.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # cdpEnabled
            'compliance': {'a': 'Compliance', 'b': 'CDP Packets disabled.', 'd': 'Cisco Discovery Protocol was disabled to prevent exposure of device information.', 'f': 'Reduces information leakage that could assist in network reconnaissance.', 'h': 'Maintain CDP disabled unless required for critical network discovery.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'CDP (Cisco Discovery Protocol) were enabled on all interfaces.', 'd': 'CDP was enabled and broadcasting network topology information.', 'f': 'CDP packets were enabled on all interfaces by default on switches. They are transmitted in clear text which allows an attacker to analyze the packets and gain a wealth of information about the network device. An attacker can use this information to execute a known vulnerability against the device platform.', 'h': 'It is recommended that Disable CDP on non-management interfaces.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # tcpUdpSmallServers
            'compliance': {'a': 'Compliance', 'b': 'TCP/UDP small servers disabled.', 'd': 'Small TCP and UDP services were disabled on Cisco switches.', 'f': 'Reduces unnecessary attack surface and prevents exploitation of unused services.', 'h': 'Continue monitoring configurations to ensure small servers remain disabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Cisco enables TCP and UDP Small Servers by default.', 'd': 'It was observed that Cisco enables TCP and UDP Small Servers by default.Cisco devices provide a set of simple servers which are collectively known as TCP small servers and User Datagram Protocol (UDP) small servers. The services provide little functionality and include chargen, echo and daytime. Cisco IOS version 11.2 and older enable these services by default; newer IOS versions explicitly require them to be started.', 'f': 'Each running service increases the chances of an attacker being able to identify the device and successfully compromise it. It is good security practice to disable all unused services. These services, especially their User Datagram Protocol (UDP) versions, are infrequently used for legitimate purposes. However, they have been used to launch denial of service attacks that would otherwise be prevented by packet filtering.', 'h': '''It is recommended that if not required, TCP and UDP small servers should be explicitly disabled. TCP and UDP small services are rarely used and are disabled by default in newer versions of Cisco IOS. TCP small servers can be disabled with the following IOS command:-
no service tcp-small-servers

UDP small servers can be disabled with the following IOS command:-
no service udp-small-servers'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # defaultCredentials
            'compliance': {'a': 'Compliance', 'b': 'Default credentials not in use.', 'd': 'Custom administrative usernames and strong passwords were configured.', 'f': 'Prevents unauthorized access using default or commonly known credentials.', 'h': 'Periodically audit admin accounts and enforce complex credential policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Default credentials were in use in the name of account with Admin.', 'd': 'It was observed that Default credentials like username "Admin” for accessing administrator panel of switch was used.', 'f': 'If not changed, default credentials make an organization more vulnerable to potential cyber-attacks. Attackers can easily obtain these standard login details, allowing them access to the devices on your network – usually with admin rights – and leaving them open to takeover.', 'h': 'It is recommended to follow standard Username so that user enumeration attack will not be achieved by an attacker who would try to gain access to networking devices.  Also, do not use Default credentials.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # unusedPortsClosed
            'compliance': {'a': 'Compliance', 'b': 'All unused ports disabled.', 'd': 'Unused switch ports were administratively shut down to prevent unauthorized connections.', 'f': 'Reduces risk of rogue device connection or unauthorized network access.', 'h': 'Continue regular port audits and maintain port security settings.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'It was found that the unused ports were open on switch.', 'd': 'It was observed that ports were open. So any one can directy access the switch by attaching LAN cable to it.', 'f': 'An attacker can intrude the network by unauthorizedly accessing the firewall with the help unused port and gather sensitive information to commit bank frauds.', 'h': 'It is recommended to disable unused ports on switch. It is simple to make configuration changes to multiple ports on a switch. If a range of ports must be configured, use the interface range command to disable ports.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # physicallySecured
            'compliance': {'a': 'Compliance', 'b': 'Switch physically secured.', 'd': 'The switch was placed in a locked server rack with restricted access.', 'f': 'Prevents unauthorized physical access or tampering with networking equipment.', 'h': 'Maintain restricted access to network racks and enforce visitor logging.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Core switch was not secured physically.', 'd': 'The switch was openly accessible without rack lock or access restriction.', 'f': 'Unauthorized individuals could disconnect cables, reset configurations, or install rogue devices, compromising network integrity and availability.', 'h': 'Secure the switch within a locked rack and implement access control policies for authorized personnel only.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # ntpConfigured
            'compliance': {'a': 'Compliance', 'b': 'NTP configured.', 'd': 'Switch synchronized with reliable NTP server for accurate event timestamping.', 'f': 'Ensures proper log correlation and facilitates accurate incident investigation.', 'h': 'Periodically verify NTP synchronization and maintain redundancy in time servers.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'NTP was not configured.', 'd': 'It was observed that the NTP was not configured.  NTP is intended to synchronize all participating computers within a few milliseconds of Coordinated Universal Time (UTC). to mitigate the effects of variable network latency.', 'f': 'A more insidious effect of weak time keeping is that it damages the ability to investigate security breaches and other kinds of system problems. Hackers, for example, will often exploit backdoor, and proxy computers when mounting and attacking- both to hide their tracks and to exploit whatever opportunities (like NTP System privileges ) the hacker encounters along the way. Finding these stopping-off points is critical for shutting the door to future attacks and requires precise measurements of time in order to reconstruct the exact sequence of events. log file and application time stamp obviously become essential pieces of evidence.', 'h': 'It is recommended to configure NTP for synchronization which helps at the time of forensics to match the logs as per global time.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # httpLoginsDisabled
            'compliance': {'a': 'Compliance', 'b': 'HTTP logins were disabled, and redirection to HTTPS was configured.', 'd': 'The device was configured to disallow plain HTTP access, automatically redirecting all login attempts to HTTPS, ensuring secure communication channels.', 'f': 'Data transmission is encrypted, protecting credentials and administrative sessions from interception.', 'h': 'No action required as HTTPS redirection is properly configured.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'HTTP logins were enabled, and no HTTPS redirection was configured.', 'd': 'The switch allowed login access via plain HTTP without encryption, making the credentials vulnerable to sniffing or man-in-the-middle attacks.', 'f': 'Exposes sensitive credentials during transmission, increasing risk of unauthorized access and network compromise.', 'h': 'Disable HTTP logins and enforce HTTPS redirection to ensure encrypted administrator access and data confidentiality.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # firmwareUpdated
            'compliance': {'a': 'Compliance', 'b': 'Firmware is up to date.', 'd': 'The switch runs the latest vendor-recommended firmware version, ensuring patched vulnerabilities and enhanced stability.', 'f': 'Ensures protection against known vulnerabilities and maintains operational stability.', 'h': 'Continue timely patch management and firmware updates as per vendor advisories.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Switch is not Updated to the highest firmware version available.', 'd': 'The switch runs an outdated firmware version containing known vulnerabilities that may be exploited by attackers.', 'f': 'As, the firmware was not updated, an attacker can exploit the existing bug to compromise the security of banks digital infrastructure. The older firmware may not utilize the hardware properly to enhance the perfomance of the device.', 'h': 'Upgrade to the latest stable firmware release and implement a regular patch management schedule to maintain security.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # adminWhitelistIP
            'compliance': {'a': 'Compliance', 'b': 'Admin login restricted to whitelisted IPs.', 'd': 'Access to the administrative interface is allowed only from predefined whitelisted IP addresses, reducing unauthorized login attempts.', 'f': 'Limits administrative access exposure to trusted systems, minimizing external threats.', 'h': 'Maintain and periodically review the whitelist to ensure only authorized IPs have access.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Administrative portal login accessed by any of the organizations computers.', 'd': 'The administrative interface is accessible from all network segments, increasing the risk of brute-force or unauthorized login attempts.', 'f': 'Any bank employee who does not have higher privileges can access the administrative portal and make changes in rules and policies from any computer.', 'h': 'Implement IP whitelisting for the administrative portal and limit access to internal trusted systems only.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # adminLockout
            'compliance': {'a': 'Compliance', 'b': 'Admin lockout configured after 3 failed attempts.', 'd': 'The switch is configured to lock administrator accounts after consecutive failed login attempts as per the organization\'s security policy.', 'f': 'Prevents brute-force login attacks and enhances account protection.', 'h': 'Continue enforcing lockout policy and monitor logs for repeated failed login attempts.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'There was no lockout functionality available.', 'd': 'The switch does not lock out accounts after multiple failed login attempts, leaving it open to brute-force attacks.', 'f': 'Allows repeated login attempts without restriction, increasing the likelihood of unauthorized access.', 'h': 'Configure admin lockout policy after 3 failed login attempts or as per organizational standards to mitigate brute-force threats.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # loggingEnabled
            'compliance': {'a': 'Compliance', 'b': 'Logging enabled and Syslog server configured.', 'd': 'The switch logs all administrative and system events to a centralized Syslog server for monitoring and auditing.', 'f': 'Ensures accountability, event tracking, and timely incident detection.', 'h': 'Continue periodic verification of Syslog server connectivity and log retention.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Logging was not enabled.', 'd': 'The switch does not forward logs to a centralized Syslog server, reducing visibility into system events and potential anomalies.', 'f': 'Weakens incident response capabilities and delays threat detection, increasing operational risk.', 'h': 'Enable logging and integrate the switch with a centralized Syslog server to maintain audit trails and monitor activities effectively.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # passwordPolicies
            'compliance': {'a': 'Compliance', 'b': 'Password policies configured as per standards.', 'd': 'The switch enforces complex password requirements, expiration, and reuse restrictions aligned with the organization\'s security policy.', 'f': 'Enhances account security and reduces unauthorized access risks.', 'h': 'Maintain and periodically review password policy configurations to align with current standards.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Password Policies were not configured as per the organization's password management policy.", 'd': 'The device lacks complex password enforcement, allowing weak credentials to be set for administrative accounts.', 'f': 'Increases susceptibility to credential theft and unauthorized administrative access.', 'h': 'Enforce a password policy requiring complexity, expiry, and reuse restrictions as per organizational and regulatory standards.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # backupFrequency
            'compliance': {'a': 'Compliance', 'b': 'Regular backups are performed.', 'd': 'Switch configuration backups are performed regularly and stored securely to ensure quick restoration when required.', 'f': 'Ensures business continuity and minimizes downtime during failures or configuration errors.', 'h': 'Continue performing scheduled backups and verify restoration procedures periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'No backup was taken for the switch configuration.', 'd': 'No evidence of scheduled configuration backups was found, leaving the network vulnerable to data loss during hardware failure.', 'f': 'Results in configuration loss, extended downtime, and difficulty in restoring network functionality.', 'h': 'It is recommended to use a TFTP Server to Backup and Restore a configuration, and take backup on a periodical basis or after changing or updating the firmware.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        18: {  # securePassword
            'compliance': {'a': 'Compliance', 'b': 'Secure password configured for core switch access.', 'd': 'The core switch uses a strong, complex password meeting organizational and industry-standard requirements, protecting administrative access.', 'f': 'Prevents unauthorized login attempts and secures network control plane.', 'h': 'Maintain strong password standards and rotate credentials periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A secure password was not enabled for core switch.', 'd': 'The core switch is using weak, default, or easily guessable passwords, exposing it to unauthorized access attempts.', 'f': 'May allow attackers to gain administrative control, modify configurations, or disrupt network operations. Compromise of the core switch could lead to complete network compromise.', 'h': 'Implement strong, complex passwords for administrative access, aligned with organizational password policy, and ensure periodic password rotation and monitoring.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        19: {  # bootpDisabled
            'compliance': {'a': 'Compliance', 'b': 'BOOTP service disabled.', 'd': 'The switch has BOOTP functionality disabled to prevent unauthorized device bootstrapping or remote configuration.', 'f': "Eliminates exposure to unauthorized automatic configurations.", 'h': '''Continue maintaining BOOTP service disabled unless specifically required.'''},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'BOOTP service enabled.', 'd': 'BOOTP remains active on the switch, allowing remote configuration requests from unknown devices.', 'f': 'May allow malicious users to load unauthorized configurations or hijack network initialization processes, leading to misconfigurations or compromise.', 'h': 'Disable BOOTP unless explicitly required for a controlled setup. Ensure only DHCP or static IP assignment methods are used with strict access control.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        20: {  # loginDisclaimer
            'compliance': {'a': 'Compliance', 'b': 'Login banner/disclaimer message configured.', 'd': 'The switch displays an authorized access warning banner before login, fulfilling organizational security and legal requirements.', 'f': 'Provides deterrence to unauthorized access and ensures compliance with regulatory security guidelines.', 'h': 'No recommendation needed; maintain the disclaimer text as per policy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Login disclaimer message not configured.', 'd': 'The switch does not display a security warning before login, failing to inform users of authorized access policies.', 'f': 'Increases legal and compliance risks during security incidents, as unauthorized access attempts lack formal deterrence and traceability.', 'h': 'Configure a proper login banner message that clearly states "Authorized Access Only" and references organizational and legal implications.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        21: {  # loginTimeout
            'compliance': {'a': 'Compliance', 'b': 'Login timeout configured as per policy.', 'd': 'The switch disconnects idle sessions after a specified timeout period, reducing session hijacking risks.', 'f': 'Minimizes risk of unauthorized access due to unattended sessions.', 'h': 'Maintain timeout configuration as per organization policy and verify periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Time limit was not set for the ideal condition.', 'd': 'It was observed that the feature of login timeout after a specific duration was not set. The switch is the core part of a bank’s network architecture. Its admin panel can be accessed by an unauthorized person if the Session time-out functionality is not configured by the bank.', 'f': 'An attacker can utilize this time to perform malicious activities and cause  harm to the network infrastructure, when an authorize user leave his system without logging out.', 'h': 'It is recommended to set login/session time out as per the organization policy'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        22: {  # icmpUnreachable
            'compliance': {'a': 'Compliance', 'b': 'ICMP unreachable messages disabled as per standard.', 'd': 'The switch is configured to block ICMP unreachable messages, preventing network reconnaissance activities.', 'f': 'Reduces exposure to external mapping or scanning attempts.', 'h': 'Maintain ICMP unreachable message suppression to minimize attack surface.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'ICMP IP unreachable messages were enabled on the interface mgmt0.', 'd': 'The interface mgmt0 responds with ICMP unreachable messages, disclosing active IP ranges and network configurations.', 'f': 'Enables attackers to map internal IPs and perform reconnaissance, increasing attack success probability.', 'h': 'Disable ICMP unreachable messages on management interfaces to limit information disclosure and network profiling risks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        23: {  # domainLookups
            'compliance': {'a': 'Compliance', 'b': 'Domain lookup disabled; DNS configured properly.', 'd': 'Domain lookup is disabled to prevent misdirected name resolutions, and legitimate DNS servers are defined for controlled resolution.', 'f': 'Reduces processing delays and mitigates DNS misuse or redirection risks.', 'h': 'Continue maintaining DNS configuration and domain lookup settings as per standards.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Domain lookups were not disabled, and  DNS servers were  not configured.', 'd': 'The switch performs domain lookups by default and lacks a defined DNS configuration, causing resolution delays and possible misuse.', 'f': 'An attacker who could be able to capture network traffic would monitor DNS queries from the Switch. Furthermore, devices can connect to Telnet servers by supplying only the hostname or IP address of the server. A mistyped  command could be interpreted as an attempt to connect to a Telnet server and broadcast on the network.', 'h': '''It is recommended that domain lookups should be disabled. Domain lookups can be disabled with the following command:-
no ip domain-lookup
 
If domain lookups are required, we recommend that DNS should be configured. DNS can be configured with the following command:
ip name-server {IP address}'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        24: {  # padService
            'compliance': {'a': 'Compliance', 'b': 'PAD service disabled.', 'd': 'The switch has PAD (X.25 protocol) service disabled as it is obsolete and not required in the current environment.', 'f': 'Reduces unnecessary service exposure and attack vectors.', 'h': 'Continue maintaining PAD service disabled unless explicitly required for legacy integration.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'PAD service enabled.', 'd': 'The switch still has PAD (X.25 protocol) service enabled, though it is outdated and unused in modern banking networks.', 'f': 'Exposes the network to unnecessary and potentially exploitable services, increasing the attack surface for legacy protocol attacks or misconfigurations.', 'h': 'Disable PAD service immediately to minimize exposure to legacy vulnerabilities and comply with secure network configuration standards.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        25: {  # mopDisabled
            'compliance': {'a': 'Compliance', 'b': 'MOP disabled on all ethernet interfaces.', 'd': 'The switch has Maintenance Operations Protocol (MOP) disabled, preventing unauthorized device maintenance communications.', 'f': 'Protects against unauthorized access attempts through MOP.', 'h': 'Maintain MOP disabled and periodically verify it across all interfaces.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'MOP enabled on ethernet interfaces.', 'd': 'MOP protocol remains active, allowing network devices to advertise and exchange management information over Ethernet.', 'f': 'Attackers could exploit the MOP service for reconnaissance or inject unauthorized management commands into the network.', 'h': 'Disable MOP protocol on all interfaces unless specifically required; perform configuration audits to confirm its status.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        26: {  # httpsSshPorts
            'compliance': {'a': 'Compliance', 'b': 'HTTPS/SSH ports changed from default.', 'd': 'Administrative services use non-standard ports to reduce exposure to automated scans and brute-force attacks.', 'f': 'Adds an extra layer of obscurity, reducing risk of common exploit attempts.', 'h': 'Continue using non-standard ports and document port mappings securely.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'HTTPS/SSH ports set to default.', 'd': 'The switch continues to use default ports (443/22) for HTTPS and SSH, making it easier for attackers to locate and target admin services.', 'f': 'Increases likelihood of targeted brute-force or exploit attempts since these ports are easily discovered during network scans.', 'h': 'Change administrative access ports to non-standard ports and apply IP-based access restrictions for enhanced protection.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        27: {  # directedBroadcasts
            'compliance': {'a': 'Compliance', 'b': 'Directed broadcasts disabled.', 'd': 'Directed broadcast functionality has been disabled to prevent amplification attacks.', 'f': 'Mitigates risk of Smurf and DDoS amplification attacks.', 'h': 'Keep directed broadcasts disabled and verify configuration periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Directed broadcasts enabled.', 'd': 'The switch is configured to allow directed broadcasts, which can be abused for network amplification attacks.', 'f': 'Attackers can use the switch to flood networks with spoofed ICMP requests, leading to Denial-of-Service conditions and degraded performance.', 'h': 'Disable directed broadcasts immediately on all interfaces to prevent misuse in network amplification or DDoS attacks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        28: {  # icmpRedirects
            'compliance': {'a': 'Compliance', 'b': 'ICMP redirects disabled.', 'd': 'The switch is configured to block ICMP redirect messages to prevent route manipulation by attackers.', 'f': 'Protects routing integrity and mitigates man-in-the-middle (MITM) risks.', 'h': 'Maintain ICMP redirect disabled on management interfaces.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'ICMP redirects enabled on mgmt0.', 'd': 'The management interface responds to ICMP redirect messages, potentially altering routing tables based on forged packets.', 'f': 'Attackers could manipulate routing paths, perform MITM attacks, or redirect traffic for packet sniffing and data interception.', 'h': 'Disable ICMP redirect support on all interfaces, especially management ones, to safeguard routing stability and prevent redirection abuse.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        29: {  # proxyArp
            'compliance': {'a': 'Compliance', 'b': 'Proxy ARP disabled on mgmt0.', 'd': 'The switch\'s management interface does not respond to Proxy ARP requests, maintaining strict ARP control.', 'f': 'Reduces ARP spoofing and unintended routing risks.', 'h': 'Maintain Proxy ARP disabled on management and internal interfaces.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Auditor identified one interface that had Proxy ARP enabled, mgmt0.', 'd': "It was observed that Auditor identified one interface that had Proxy ARP enabled, mgmt0. Address Resolution Protocol (ARP) is a protocol that network hosts use to translate network addresses into media addresses. Under normal circumstances, ARP packets are confined to the sender's network segment. However, a router with Proxy ARP enabled on network interfaces can act as a proxy for ARP, responding to queries and acting as an intermediary.", 'f': '''A router that acts as a proxy for ARP requests will extend layer two access across multiple network segments, breaking perimeter security.
Hosts have no idea of the physical details of their network and assume it to be a flat network in which they can reach any destination simply by sending an ARP request. But using ARP for everything has disadvantages. These are some of the disadvantages:-
•  It increases the amount of ARP traffic on your segment.
•  Hosts need larger ARP tables to handle IP-to-MAC address mappings.
•  Security can be undermined. A machine can claim to be another to intercept packets, an act called "spoofing."
•  It does not work for networks that do not use ARP for address resolution.
•  It does not generalize to all network topologies. For example, more than one router that connects two physical networks.''', 'h': '''It is recommended if not required, to disable Proxy ARP on all interfaces. Proxy ARP can be disabled on each interface with the following IOS command:- 
no ip proxy-arp'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Apply formatting to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for col in range(1, 9):  # A to H
        cell = ws.cell(row=1, column=col)
        cell.border = thin_border
    
    # Process each question
    for i, question in enumerate(question_mapping, 2):  # Start from row 2
        # Get user input for this question
        user_input = form_data.get(question, 'not_applicable') if form_data else 'not_applicable'
        
        # Get response data
        response_data = question_responses.get(i-1, {})
        user_response = response_data.get(user_input, {})
        
        # Populate columns based on user input
        # Column C: Compliance/Non-Compliance/Not Applicable
        ws.cell(row=i, column=3, value=user_response.get('a', 'Not Applicable'))
        
        # Column D: Observation (Short/Brief)
        ws.cell(row=i, column=4, value=user_response.get('b', 'Not Applicable'))
        
        # Column F: Observation
        ws.cell(row=i, column=6, value=user_response.get('d', 'Not Applicable'))
        
        # Column G: Impact
        ws.cell(row=i, column=7, value=user_response.get('f', 'Not Applicable'))
        
        # Column H: Recommendation
        ws.cell(row=i, column=8, value=user_response.get('h', 'Not Applicable'))
        
        # Column E: Risk Factor with color coding
        risk_factor = risk_factors[i-2]  # risk_factors is 0-indexed
        risk_cell = ws.cell(row=i, column=5, value=risk_factor)
        risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
        risk_cell.fill = PatternFill(start_color=risk_colors.get(risk_factor, '808080'), end_color=risk_colors.get(risk_factor, '808080'), fill_type='solid')
        risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply formatting to all data rows
    for row in range(2, len(question_mapping) + 2):  # Rows 2 to 30
        for col in range(1, 9):  # Columns A to H
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Special alignment for column A (Sr. No.), C (Compliance status), and E (Risk Factor)
            if col == 1 or col == 3 or col == 5:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Set row height for wrapped text
            ws.row_dimensions[row].height = 30
    
    # Save file
    filename = "Core Switch Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    """
    Delete the generated Excel file after download
    """
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            print(f"File {filepath} deleted successfully")
    except Exception as e:
        print(f"Error deleting file {filepath}: {e}")