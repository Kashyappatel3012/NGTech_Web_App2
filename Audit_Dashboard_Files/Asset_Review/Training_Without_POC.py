import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_training_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Training"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # Training Questions
    questions = [
        "Are users given adequate training on the application systems' functionalities?",
        "Are technical personnel given adequate training in the technical details of the application system to provide necessary troubleshooting/help to users?",
        "Are users aware of the steps to be carried out in case of contingency due to the unavailability of systems?"
    ]

    # Risk Factors
    risk_factors = [
        "Medium", "Medium", "Medium"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "trUserApplicationTraining": 1,
        "trTechnicalPersonnelTraining": 2,
        "trContingencyAwareness": 3
    }

    # Question responses data
    question_responses = {
        1: {  # trUserApplicationTraining
            'compliance': {'a': 'Compliance', 'b': 'Users adequately trained.', 'd': 'Users are provided structured training sessions covering all aspects of the application system, ensuring they understand how to use the system efficiently.', 'f': 'Reduces operational errors, improves system utilization, and enhances overall productivity.', 'h': 'Periodically evaluate user knowledge and update training programs to cover new features or system updates.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The users were not given adequate training on the application system functionalities.', 'd': 'End-users have not received sufficient training on how to operate the application systems, resulting in improper or inefficient use of features.', 'f': "If the user does not have training about the application functionalities then the user will not be able to properly utilize the application, troubleshoot it, and detect abnormal behaviour in the application. An unskilled user might delete critical data unintentionally, which will impact the bank's productivity.", 'h': 'Conduct comprehensive training sessions for all users on application functionalities, maintain training records, and provide refresher courses periodically.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # trTechnicalPersonnelTraining
            'compliance': {'a': 'Compliance', 'b': 'Technical staff well-trained.', 'd': 'Technical personnel receive detailed and ongoing training on the system\'s technical aspects, enabling them to support users and resolve issues efficiently.', 'f': 'Ensures smooth system operations, timely issue resolution, and enhances overall user satisfaction.', 'h': 'Maintain updated training records and conduct regular skill assessments for technical staff.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Technical persons were not adequately trained in the technical details of the application system.', 'd': 'Technical personnel lack in-depth knowledge of the application system, limiting their ability to troubleshoot issues or assist users effectively.', 'f': "When technical persons are not given adequate training in technical details of the application then they might not be able to solve common issues with the application, Which will result in a delay in the bank's operations. They might not be able to provide necessary troubleshooting/help to the users.", 'h': 'Provide technical staff with detailed training on system architecture, troubleshooting procedures, and maintenance practices, with periodic knowledge updates.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # trContingencyAwareness
            'compliance': {'a': 'Compliance', 'b': 'Users were aware of the steps to be carried out in case of a contingency', 'd': 'Users are trained on contingency procedures and know the steps to follow during system unavailability, including alternate processes or escalation paths.', 'f': 'Minimizes disruption during outages, ensures continuity of critical operations, and reduces operational risk.', 'h': 'Review and update contingency training regularly and document user preparedness exercises.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The users were not aware of the steps to be carried out in case of contingency due to non-availability of systems.', 'd': 'Users are unaware of the actions to take during system outages or failures, leading to confusion and delayed response in business processes.', 'f': "If the users are not aware of steps to be carried out in case of contingency due to the non-availability of the system then the users will not able to continue the bank's day-to-day operation during a catastrophic event, and thus bank may face financial losses.", 'h': 'Train users on contingency procedures, ensure clear documentation is available, and conduct regular drills to simulate system unavailability scenarios.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Training Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
