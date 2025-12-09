import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_fire_protection_excel(form_data=None):
    """
    Create Fire Protection Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Fire Protection"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Fire Protection Questions
    questions = [
        "Is the fire alarm system installed?",
        "Are smoke detectors provided in the server room and other areas of computer installations?",
        "Are smoke detectors tested regularly to ensure they are working properly?",
        "Are fire extinguishers installed at strategic places like the server room, UPS room, and near the nodes and printers?",
        "Are fire extinguishers regularly refilled/maintained?",
        "Does staff know how to use the fire extinguishers?",
        "Is the evacuation plan documented and rehearsed at regular intervals for immediate action in case of a fire outbreak?"
    ]

    # Risk Factors
    risk_factors = [
        "Medium",
        "Medium", 
        "Medium",
        "Medium",
        "Medium",
        "Low",
        "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "fpFireAlarmSystem": 1,
        "fpSmokeDetectors": 2,
        "fpSmokeDetectorsTested": 3,
        "fpFireExtinguishersInstalled": 4,
        "fpFireExtinguishersMaintained": 5,
        "fpStaffTraining": 6,
        "fpEvacuationPlan": 7
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Question responses data
    question_responses = {
        1: {  # fpFireAlarmSystem
            'compliance': {'a': 'Compliance', 'b': 'Fire alarm system installed.', 'd': 'A functioning fire alarm system is installed throughout critical areas, providing automated alerts to personnel and safety teams in case of fire detection.', 'f': 'Early detection reduces damage to critical assets, ensures timely evacuation, and enhances staff safety.', 'h': 'Periodically test the fire alarm system and maintain logs of tests to ensure reliability.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The fire alarm system was not installed in the Bank.', 'd': 'The premises, including the server room and critical IT infrastructure areas, do not have a fire alarm system installed to detect and alert personnel in case of fire.', 'f': 'As the fire alarm is not present in the bank. In case of a fire event in the bank, the bank will not be able to identify and prevent it. The fire can damage the critical assets of the bank and cause harm to employees.', 'h': 'Install a centralized fire alarm system covering all critical areas, including server rooms, UPS rooms, and office spaces, to ensure timely detection and alert in case of fire.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # fpSmokeDetectors
            'compliance': {'a': 'Compliance', 'b': 'Smoke detectors installed.', 'd': 'Smoke detectors are installed in all critical IT areas and connected to the fire alarm system, ensuring early detection and timely alerts.', 'f': 'Helps in preventing large-scale damage and ensures quick evacuation and emergency response.', 'h': 'Conduct regular inspections and functional testing of smoke detectors to ensure continuous operation.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The smoke detector was not present.', 'd': 'Critical areas, including the server room and UPS room, lack smoke detectors to provide early warning of fire incidents.', 'f': 'If the smoke detector is not present in the bank then the employees will not get alerted in case of fire. The fire can damage the critical assets, employees, and documents of the bank, which can incur a huge loss to the bank.', 'h': 'Install smoke detectors in all server rooms, UPS areas, and key IT installation points with automatic alerts to security personnel.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # fpSmokeDetectorsTested
            'compliance': {'a': 'Compliance', 'b': 'Smoke detectors regularly tested.', 'd': 'Smoke detectors undergo periodic functional testing, ensuring reliability and readiness during emergencies.', 'f': 'Guarantees early warning during fire outbreaks, reducing damage and improving safety.', 'h': 'Maintain test logs and conduct surprise checks to ensure detectors remain operational.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Smoke detectors were not tested regularly.', 'd': 'Installed smoke detectors are not tested periodically for functionality, risking failure during actual fire incidents.', 'f': 'If smoke detectors were not tested regularly to ensure that they work properly. In case of a fire event in the bank, the faulty smoke detector will not generate an alarm to alert about the smoke or fire. Thus, fire can cause a lot of damage to bank employees and critical assets of the bank.', 'h': 'Implement a scheduled testing and maintenance program for all smoke detectors, documenting results for audit purposes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # fpFireExtinguishersInstalled
            'compliance': {'a': 'Compliance', 'b': 'Fire extinguishers installed strategically.', 'd': 'Fire extinguishers are placed at server rooms, UPS rooms, network nodes, and other critical areas, ensuring immediate availability during fire incidents.', 'f': 'Facilitates prompt fire suppression and minimizes equipment and data damage.', 'h': 'Conduct regular inspections and ensure extinguishers are visible and accessible.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Fire extinguishers were not installed near UPS and the server room in the Bank.', 'd': 'Fire extinguishers are missing or not placed in critical locations like server rooms, UPS rooms, and near network nodes or printers.', 'f': 'If the fire extinguishers are not near to server room, or UPS room and the fire breaks out due to a short circuit then the critical assets will be damaged and the data will be lost. Fire extinguishers can help in dousing the fire they should not be so close that they are involved in the fire or that a person cannot reach them. The suggested distance from their point of use is between 5 and 15 m.', 'h': 'Install appropriate fire extinguishers (CO2 or FM200 for electrical equipment) at all strategic locations for quick access during emergencies.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # fpFireExtinguishersMaintained
            'compliance': {'a': 'Compliance', 'b': 'Fire extinguishers regularly maintained.', 'd': 'Fire extinguishers are checked, refilled, and serviced periodically, with maintenance records documented.', 'f': 'Ensures readiness of extinguishers and quick response in case of fire.', 'h': 'Continue regular maintenance and update service records for compliance audits.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Fire extinguishers were not regularly refilled/maintained.', 'd': 'Installed fire extinguishers are not checked, refilled, or maintained periodically, potentially rendering them ineffective during emergencies.', 'f': 'If fire-extinguishers are not maintained regularly, then whenever a fire incident happens the fire-extinguishers might not function properly, resulting in the spread of fire causing damage to bank property.', 'h': 'Establish a maintenance schedule for all fire extinguishers, including refills, pressure checks, and service records.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # fpStaffTraining
            'compliance': {'a': 'Compliance', 'b': 'Staff trained on fire extinguisher use.', 'd': 'Personnel are trained to handle fire extinguishers effectively, including identifying types of fire and using the appropriate extinguisher.', 'f': 'Reduces risk of injuries and asset damage, ensuring prompt containment of fire incidents.', 'h': 'Conduct refresher training sessions and practical drills periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Staff does not know, how to use the fire extinguishers.', 'd': 'Personnel are not aware or trained in operating fire extinguishers during emergencies, leading to ineffective responses.', 'f': "In case of a fire event, the employees will not be able to use the fire extinguisher due to a lack of knowledge regarding the working of a fire extinguisher. This will result in banks' property being damaged as employees will not be able to prevent the fire spread during a fire event.", 'h': 'Conduct regular staff training and fire drills focused on the proper use of fire extinguishers.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # fpEvacuationPlan
            'compliance': {'a': 'Compliance', 'b': 'Evacuation plan documented and rehearsed.', 'd': 'A formal evacuation plan exists, and staff participate in regular fire drills to practice safe evacuation from all critical areas.', 'f': 'Ensures rapid, organized response to fire, protecting personnel and minimizing asset damage.', 'h': 'Update evacuation plans regularly and conduct unannounced drills to assess staff readiness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The evacuation plan was not available.', 'd': 'No formal evacuation plan exists, or staff have not practiced evacuation procedures for server rooms, offices, or critical IT areas.', 'f': 'As an evacuation plan is not available, one will be confused about what should be done in an emergency. Thus people will panic during a fire emergency and it will be chaotic to take control of the situation to save people from danger. This reduces the efficiency to safely get people away from an area where there is an imminent threat, ongoing threat, or a hazard to lives or property.', 'h': 'Document a detailed evacuation plan and conduct regular fire drills for all personnel, including specific procedures for server rooms and critical areas.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }

    # Populate data based on user input
    for i, question in enumerate(questions, 2):
        # Get user input for this question
        question_num = i - 1
        user_input = None
        
        if form_data:
            # Find the corresponding form field
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Fire Protection Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
