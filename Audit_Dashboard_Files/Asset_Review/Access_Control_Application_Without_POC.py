import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_access_control_application_excel(form_data=None):
    """
    Create Access Control Application Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Access Control Application"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Access Control Application Questions
    questions = [
        "Does the system provide unique user IDs and passwords for all users?",
        "Does the system provide different levels of access?",
        "Is there a dummy user ID created in the system?",
        "Do users write their passwords on walls, desk diaries, etc.? Are they aware of the need for password secrecy?",
        "Does the system prompt for a change of user password after the lapse of a specified period?",
        "Is super user access at the application level not given to staff who are under notice period, retiring shortly, or under disciplinary action?",
        "Does the system allow concurrent login to a single user ID from different nodes?",
        "Does the system log out automatically if the user is inactive for a specified time?",
        "Does the system allow users to cancel their user ID?",
        "Do users share their passwords?",
        "Does the application make use of all the security features available at the application system level?",
        "Is the application system user list periodically reviewed?"
    ]

    # Risk Factors (provided by user)
    risk_factors = [
        "High", "High", "High", "High", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Low", "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "accessControlAppUniqueUserIDs": 1,
        "accessControlAppDifferentAccessLevels": 2,
        "accessControlAppDummyUserID": 3,
        "accessControlAppPasswordSecrecy": 4,
        "accessControlAppPasswordChangePrompt": 5,
        "accessControlAppSuperUserAccess": 6,
        "accessControlAppConcurrentLogin": 7,
        "accessControlAppAutoLogout": 8,
        "accessControlAppUserCancelID": 9,
        "accessControlAppPasswordSharing": 10,
        "accessControlAppSecurityFeatures": 11,
        "accessControlAppUserListReview": 12
    }

    # Question responses data
    question_responses = {
        1: {  # accessControlAppUniqueUserIDs
            'compliance': {'a': 'Compliance', 'b': 'Unique user IDs assigned to all users.', 'd': 'The application provide unique user IDs and passwords for all users. ', 'f': 'Providing unique user IDs and passwords for all users ensures that every individual accessing the application is uniquely identifiable, thereby enhancing accountability and traceability.', 'h': 'Continue maintaining unique credentials and audit logs periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The system does not provide unique user IDs and passwords for all users.', 'd': 'It was observed that the application does not provide unique user IDs and passwords for all users.', 'f': 'Without unique user IDs and passwords, user activities cannot be properly tracked, leading to a lack of accountability. This increases the risk of unauthorized access, data manipulation, and potential misuse of system privileges.', 'h': 'It is recommended that each user is assigned a unique user ID and password to maintain accountability. Shared or generic accounts must be disabled immediately. Additionally, strong password policies and periodic access reviews should be implemented to enhance system security and compliance.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # accessControlAppDifferentAccessLevels
            'compliance': {'a': 'Compliance', 'b': 'Role-based access levels implemented.', 'd': 'The system provides clearly defined access levels based on job roles and responsibilities.', 'f': 'Ensures segregation of duties and reduces insider threat risks.', 'h': 'Review access roles periodically and align with organizational policy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The system  provide the same levels of access.', 'd': 'It was found that the system does not differentiate between user roles, granting excessive permissions to all users.', 'f': 'Lack of access control may lead to unauthorized data modification or access to sensitive information, compromising confidentiality and integrity.', 'h': 'Implement role-based access control (RBAC) ensuring users have permissions only as per their responsibilities.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # accessControlAppDummyUserID
            'compliance': {'a': 'Compliance', 'b': 'No dummy user IDs present.', 'd': 'The system does not contain any dummy or unused user IDs.', 'f': 'Reduces attack surface and prevents unauthorized access through inactive or unknown accounts.', 'h': 'Continue monitoring for unauthorized or test accounts periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'There was a dummy user ID created in the system.', 'd': 'Dummy or test user IDs were identified in the system that are active but not linked to any actual user.', 'f': 'Such accounts can be exploited for unauthorized access, making them a security loophole.', 'h': 'Remove all dummy or unused user IDs and ensure new test accounts are deactivated immediately after use.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # accessControlAppPasswordSecrecy
            'compliance': {'a': 'Compliance', 'b': 'Password secrecy maintained.', 'd': 'Users are aware of password secrecy and do not store passwords in visible locations.', 'f': 'Reduces risk of credential compromise through physical observation.', 'h': 'Continue periodic awareness programs on password handling.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Users write their passwords on the wall and in desk diaries.', 'd': 'Some users were found to be storing passwords in written form near workstations or on diaries.', 'f': 'Visible or easily accessible passwords can lead to unauthorized use and serious data breaches.', 'h': 'Conduct user awareness training on password secrecy and enforce policies against written storage of passwords.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # accessControlAppPasswordChangePrompt
            'compliance': {'a': 'Compliance', 'b': 'Password expiry enforced.', 'd': 'The system prompts users to change passwords after the defined validity period.', 'f': 'Improves overall access security and reduces risk of prolonged credential misuse.', 'h': 'Maintain password rotation and monitor compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'System does not prompt for Change of the user password after the lapse of the specified period.', 'd': 'The system does not prompt users to change their passwords periodically as per policy.', 'f': 'Static passwords increase the likelihood of compromise through brute-force or social engineering.t weakens overall access control and may allow prolonged use of outdated or exposed passwords.', 'h': 'It is recommended to define a System prompt for a change of the user password after the lapse of the specified period.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # accessControlAppSuperUserAccess
            'compliance': {'a': 'Compliance', 'b': 'Super user access restricted for exiting staff.', 'd': 'Application-level administrative rights are promptly revoked for staff under notice or disciplinary action.', 'f': 'Prevents intentional misuse of administrative access and secures sensitive data.', 'h': 'Periodically review admin access and align it with HR exit procedures.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Upper User access in application level was given to staff who is under notice period, retiring shortly, under disciplinary action.', 'd': 'Application-level admin privileges are not revoked from staff under notice or disciplinary action.', 'f': 'If super user access is given to user who is on notice period or else retiring shortly, If the user is not happy with his job or gets disgruntled he/she might be using those privileges to perform malicious activities.', 'h': 'It is recommended that Super User access in application level should not given to staff who is under notice period, retiring shortly, under disciplinary action '},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # accessControlAppConcurrentLogin
            'compliance': {'a': 'Compliance', 'b': 'Concurrent login restricted.', 'd': 'The system restricts simultaneous logins using the same credentials.', 'f': 'Enhances accountability and prevents misuse of user accounts.', 'h': 'Continue enforcing session control and monitor for violations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The system allowed concurrent login to a single user ID from different nodes.', 'd': 'The system allows multiple logins with the same user ID simultaneously from different terminals.', 'f': 'When multiple logins are allowed, accountability is not defined. And for the same machine, an attacker can add the new data entry and approve it, compromising the CIA triad. ', 'h': 'It is recommended that the System should not allow concurrent login to a single user ID from different nodes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # accessControlAppAutoLogout
            'compliance': {'a': 'Compliance', 'b': 'Auto logout configured.', 'd': 'The system logs out inactive users after a specified period.', 'f': 'Prevents misuse of unattended terminals and enhances session security.', 'h': 'Periodically test session timeout functionality for effectiveness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The system does not log out automatically after a specified  time.', 'd': 'The system does not automatically log out inactive users after the defined idle period.', 'f': 'If the system does not automatically log out inactive users, it increases the risk of unauthorized access to sensitive information. Unattended active sessions can be exploited by unauthorized individuals to misuse system privileges.', 'h': 'The bank should configure the system to automatically log out users after a defined period of inactivity as per the security policy. This control helps prevent unauthorized access through unattended sessions.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # accessControlAppUserCancelID
            'compliance': {'a': 'Compliance', 'b': 'User ID cancellation restricted.', 'd': 'Only system administrators have the privilege to cancel or delete user IDs.', 'f': 'Ensures control over user management and prevents unauthorized deletion.', 'h': 'Maintain strict administrative oversight for user account changes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The system allows the users to cancel their own user ID.', 'd': 'The system allows users to disable or delete their own IDs, bypassing administrative controls.', 'f': 'Allowing users to cancel their own user IDs poses a serious security risk, as it may lead to unauthorized disabling of accounts and loss of accountability. Malicious users could exploit this to avoid detection or disrupt operations.', 'h': 'It is recommended that the ability to create, modify, or cancel user IDs be restricted to authorized administrators only. Proper approval and documentation should be required for any user ID deactivation.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # accessControlAppPasswordSharing
            'compliance': {'a': 'Compliance', 'b': 'Passwords not shared.', 'd': 'Users maintain password confidentiality and do not share them with others.', 'f': 'Ensures accountability and maintains access integrity.', 'h': 'Continue enforcing password secrecy through periodic awareness and monitoring.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Users share their passwords. ', 'd': 'Some users share passwords among team members for convenience.', 'f': 'Sharing of passwords among users leads to a complete loss of accountability, as individual activities cannot be traced to a specific user. It increases the risk of unauthorized access, data manipulation, and potential misuse of system privileges.', 'h': 'It is recommended that users be strictly instructed not to share their passwords under any circumstances.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # accessControlAppSecurityFeatures
            'compliance': {'a': 'Compliance', 'b': 'Application security features fully utilized.', 'd': 'All available application-level security features are enabled and operational.', 'f': 'Strengthens application defense against threats and minimizes vulnerability exposure.', 'h': 'Review application configurations regularly to ensure ongoing compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Application security features not fully utilized.', 'd': 'Some built-in application security features such as logging, encryption, or role management are not enabled.', 'f': 'Unused features leave the application vulnerable to exploits and unauthorized access.', 'h': 'It is recommended to enable all the security features available in the Application to be taken advantage of as far as possible for ensuring better security.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # accessControlAppUserListReview
            'compliance': {'a': 'Compliance', 'b': 'Application user list reviewed periodically.', 'd': 'The department performs periodic reviews of application system users.', 'f': 'Ensures access is limited to authorized and active users only.', 'h': 'Continue regular access reviews to maintain proper access control hygiene.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The application system user list was not reviewed periodically.', 'd': 'The department does not periodically review the list of users within the application system.', 'f': 'Inactive or unauthorized users may retain access, increasing risk of data compromise.', 'h': 'Conduct regular reviews of the application user list and remove obsolete or inactive accounts.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate data based on user input
    for i, question in enumerate(questions, 2):
        # Get user input for this question
        question_num = i - 1
        user_input = None
        
        if form_data:
            # Find the corresponding form field
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        # Get response data
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            # Populate columns C, D, F, G, H
            ws.cell(row=i, column=3, value=response_data['a'])  # Compliance/Non-Compliance/Not Applicable
            ws.cell(row=i, column=4, value=response_data['b'])  # Observation (Short/Brief)
            ws.cell(row=i, column=6, value=response_data['d'])  # Observation
            ws.cell(row=i, column=7, value=response_data['f'])  # Impact
            ws.cell(row=i, column=8, value=response_data['h'])  # Recommendation
            
            # Apply alignment to these columns
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        
        # Populate Risk Factor (column E) with color coding
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            # Apply color coding
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Only apply default font if it's not the Risk Factor column (column 5)
                if col != 5:  # Don't override Risk Factor column formatting
                    cell.font = Font(name='Calibri', size=11)
            # Set row height for wrapped text
            ws.row_dimensions[row].height = 30
    
    # Save file
    filename = "Access Control Application Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    """
    Delete the generated Excel file after download
    """
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
