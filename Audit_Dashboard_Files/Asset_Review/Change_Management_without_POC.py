import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_change_management_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Change Management"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # Change Management Questions
    questions = [
        "Are changes made with proper approval from authorities and communicated to all the respective personnel that may get affected by the changes made?",
        "Are changes managed and recorded properly?",
        "Is there a proper change management process in place for all types of changes made in the Data Centre?",
        "Are approvals taken for major changes made and documented?",
        "Is a well-defined Change Management Policy and Procedure in place?",
        "Is there an approval form/document available for requesting changes?",
        "Are changes documented in a register with proper date, time, device, and types of changes made?",
        "Are all types of changes made, including changes in critical devices, systems, servers, networking devices, databases, and applications, recorded in the Change Management Register?"
    ]

    # Risk Factors
    risk_factors = [
        "High", "High", "High", "High", "Medium", "Medium", "Medium", "Medium"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "changeApprovalCommunication": 1,
        "changeManagedRecorded": 2,
        "changeManagementProcess": 3,
        "changeMajorApprovals": 4,
        "changePolicyProcedure": 5,
        "changeApprovalForm": 6,
        "changeDocumentedRegister": 7,
        "changeAllTypesRecorded": 8
    }

    # Question responses data
    question_responses = {
        1: {  # changeApprovalCommunication
            'compliance': {'a': 'Compliance', 'b': 'All changes approved and communicated.', 'd': 'Any change to critical systems, applications, or network components is first reviewed and approved by the relevant authority and communicated to all impacted teams before implementation.', 'f': 'Enhances operational coordination, reduces risk of errors, and ensures transparency.', 'h': 'Regularly audit communication logs and approval records to ensure adherence to policy and continuous improvement.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Changes  made without proper approval and are not communicated.', 'd': 'It was observed that changes are not made with proper approval from authorities and and are not communicated to all the respective personnel that may get affected by the changes made.', 'f': 'Unauthorized modifications may lead to operational disruptions and compromise system integrity, potentially causing significant downtime and financial losses. Moreover, the lack of clear communication can result in confusion and resistance from affected stakeholders, hindering the successful implementation of changes and affecting employee morale. Without proper approvals and communication, it becomes challenging to assess the potential risks and benefits of changes, making it difficult to make informed decisions.', 'h': 'It is recommended to obtain proper approval from designated authorities before implementing any modifications to systems or processes. Additionally, it is essential to establish clear communication channels to inform all relevant personnel about impending changes and their potential impact. Regular communication updates, training sessions, and documentation should be utilized to ensure transparency and understanding among all stakeholders.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # changeManagedRecorded
            'compliance': {'a': 'Compliance', 'b': 'Changes systematically recorded.', 'd': 'All system, application, and network changes are documented in a centralized register, capturing relevant details for traceability and audit purposes.', 'f': 'Enables accountability, facilitates troubleshooting, and strengthens audit compliance.', 'h': 'Conduct periodic audits to ensure the completeness and accuracy of change records.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Changes are not managed and recorded properly.', 'd': "It was observed that there wasn't a proper change management process and changes are not recorded.", 'f': 'It exposes the organization to potential risks, such as unauthorized system modifications, leading to operational disruptions and compromising security. The lack of clear policies and procedures makes it challenging to assess and control changes effectively, leading to conflicts and compatibility issues. Furthermore, without proper documentation, it becomes difficult to identify the root causes of problems and measure the impact of modifications, hindering learning from past experiences and impeding progress', 'h': 'It is recommended to establish effective change management processes and implement proper record-keeping practices. This should include the establishment of clear policies and procedures to govern all changes made to systems and software. A designated change management team should be assigned to oversee and approve modifications, ensuring they are properly documented and tracked throughout the process.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # changeManagementProcess
            'compliance': {'a': 'Compliance', 'b': 'Change management process established.', 'd': 'All data center changes follow a structured process, including planning, risk assessment, testing, approval, execution, and documentation.', 'f': 'Ensures that changes are executed in a controlled manner, minimizing risk to operations and security.', 'h': 'Regularly review and refine the process based on incident reports and lessons learned.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A proper change management process was not defined for all types of changes made in the Data centre.', 'd': 'It was observed that A proper change management process was not defined for all types of changes made in the Data centre. The lack of a standardized approach to manage changes introduces the potential for errors, miscommunication, and inconsistencies in handling modifications to critical systems. ', 'f': 'Changes that are made without proper planning, testing, and approval can result in system failures, data breaches, and other security vulnerabilities.  It can lead to increased risk, disruption of operations, inefficient resource utilization, and a lack of accountability. The lack of a structured approach to managing changes can also result in disruption of normal operations, causing delays, confusion, and loss of productivity.without a formal change management process, accountability can be difficult to establish, making it challenging to determine who is responsible for any issues that arise.', 'h': 'It is recommended to implement  a proper chnage management process  for all types of changes made in the Data centre.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # changeMajorApprovals
            'compliance': {'a': 'Compliance', 'b': 'Major changes formally approved.', 'd': 'All major changes, including system upgrades and network reconfigurations, receive documented approval from appropriate authorities before execution.', 'f': 'Reduces operational risk and ensures accountability and regulatory compliance.', 'h': 'Periodically review major change approvals to confirm adherence to organizational policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Major changes executed without formal approval.', 'd': 'Critical upgrades or configuration changes were implemented without documented approval from designated authorities, such as the IT Head or Change Advisory Board (CAB).', 'f': 'Unauthorized major changes could result in system instability, potential downtime, or regulatory non-compliance.', 'h': 'Establish strict approval protocols for all major changes and maintain formal records of approvals in a centralized repository.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # changePolicyProcedure
            'compliance': {'a': 'Compliance', 'b': 'Policy and procedure exist.', 'd': 'A documented Change Management Policy defines roles, responsibilities, workflows, and control measures for all change activities.', 'f': 'Standardizes processes, reduces risk of unauthorized changes, and ensures audit readiness.', 'h': 'Periodically review and update the policy to align with new technology and business requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Change managagement  policies and procedures were not defined.', 'd': 'The bank lacks a documented change management policy guiding approvals, communication, documentation, and testing of changes.', 'f': 'Without proper change management, the organization is at risk of unauthorized and uncontrolled modifications to systems and software, potentially leading to operational disruptions and security vulnerabilities. The absence of clear policies and procedures also hampers the ability to assess and mitigate the impact of changes on the overall system', 'h': 'It is recommended to promptly establish clear change management policies and procedures within the organization. Defining structured processes for managing changes to the IT environment will help prevent unauthorized and uncontrolled modifications, ensuring better system stability and security.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # changeApprovalForm
            'compliance': {'a': 'Compliance', 'b': 'Change request form is in use.', 'd': 'All change requests are submitted using a standardized form that documents the scope, impact, approvals, and risk assessment.', 'f': 'Ensures accountability, traceability, and audit compliance.', 'h': 'Ensure all personnel are trained to submit requests via the formal form consistently.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A approval form/document was not available for requesting changes.', 'd': 'Change requests are submitted informally or verbally, with insufficient detail regarding scope, impact, and required approvals.', 'f': 'If an approval form/document is not available for requesting all types of changes, it can cause inconsistencies in the way changes are requested and approved, miscommunication between the requester and approver, increased risk of unauthorized changes being made and establish accountability for changes made within the organization.', 'h': 'It is recommended to implement Approval form/document for requesting  all types of changes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # changeDocumentedRegister
            'compliance': {'a': 'Compliance', 'b': 'Changes fully documented.', 'd': 'All changes, including critical and routine updates, are logged with comprehensive details in a centralized register.', 'f': 'Enhances traceability, supports audits, and facilitates root-cause analysis for any incidents.', 'h': 'Periodically audit the register for completeness and accuracy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Changes made are not documented in a register.', 'd': 'It was observed that A register is not present for recording the changes made.', 'f': 'Not documenting changes in a register with proper date, time, device, and types of changes made can have a range of negative impacts, including difficulties in tracking changes, increased risk of errors, compliance and regulatory issues, and difficulties in communicating changes.', 'h': 'It is recommended to implement a register for recording the changes made with proper date, time,device and types of changes made.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # changeAllTypesRecorded
            'compliance': {'a': 'Compliance', 'b': 'All changes recorded.', 'd': 'Every change, whether minor or critical, is recorded in the Change Management Register with all required details.', 'f': 'Ensures full traceability, operational accountability, and audit compliance.', 'h': 'Periodically review the register and cross-check with implemented changes to identify any missing entries.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'All type of changes made are not recorded into the Change Management Register.', 'd': 'It was observed that all type of changes made including changes in Critical devices, Systems, Severs, Networking devices, Databases Applications were not  recorded into the Change Management Register.', 'f': 'Not recording all types of changes made, including changes in critical devices, systems, servers, networking devices, databases, and applications in a change management register can cause Increased risk of security breaches,Difficulty in troubleshooting, Compliance and regulatory issues,Lack of transparency and accountability and Inefficient use of resources.', 'h': 'It is recommended to record all type of changes made including changes in Critical devices, Systems, Severs, Networking devices, Databases Applications in Change Management Register.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Change Management Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
