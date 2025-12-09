import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_pam_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "PAM"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # PAM Questions
    questions = [
        "Is Access to all servers, databases, network devices, and critical devices managed through PAM solution?",
        "Is Access to all servers, databases, network devices, and critical devices exclusively granted through PAM solution only?",
        "Are exclusions to PAM solution approved by the respective authority and granted according to need-to-know basis?",
        "Is PAM solution compatible with other assets?",
        "Is PAM solution available in HA mode?",
        "Whether Logs server is configured for PAM?",
        "Does system logs contain adequate information for identification purpose?",
        "Check whether Video logs are enabled for PAM user sessions?",
        "Check whether Two Factor Authentication is available for PAM Access?",
        "Are there multiple admin accounts to avoid lockout?",
        "Are unused accounts disabled?",
        "Whether PAM solution is updated?",
        "Check whether PAM password policy is configured as per organizational password policy?",
        "Check whether password policy for devices integrated into PAM is managed through PAM?",
        "Check whether 2FA Access is configured properly?",
        "Whether configuration backup of PAM solution is taken on a regular basis?",
        "Check whether PAM solution is integrated with Active Directory?",
        "Whether file type permissions are configured properly through PAM?",
        "Does the password expire periodically (preferably after 30 days)?",
        "Does PAM admin monitor the users' session activity?",
        "Whether user access policy is configured properly?",
        "Whether 2FA is enabled in all user accounts?",
        "Is there a backup process for storage of usernames and passwords used for accessing PAM?"
    ]

    # Risk Factors
    risk_factors = [
        "High", "High", "High", "High", "High", "High", "High", "High", "Medium", "Medium",
        "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium",
        "Medium", "Medium", "Medium", "Medium"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "pamAccessManaged": 1,
        "pamExclusiveAccess": 2,
        "pamExclusionsApproved": 3,
        "pamCompatible": 4,
        "pamHaMode": 5,
        "pamLogsServer": 6,
        "pamAdequateLogs": 7,
        "pamVideoLogs": 8,
        "pamTwoFactorAuth": 9,
        "pamMultipleAdmin": 10,
        "pamUnusedAccountsDisabled": 11,
        "pamSolutionUpdated": 12,
        "pamPasswordPolicy": 13,
        "pamDevicePasswordPolicy": 14,
        "pam2FAConfigured": 15,
        "pamConfigBackup": 16,
        "pamAdIntegration": 17,
        "pamFilePermissions": 18,
        "pamPasswordExpiry": 19,
        "pamSessionMonitoring": 20,
        "pamUserAccessPolicy": 21,
        "pam2FAEnabledAllUsers": 22,
        "pamCredentialBackup": 23
    }

    # Question responses data - First 11 questions
    question_responses = {
        1: {  # pamAccessManaged
            'compliance': {'a': 'Compliance', 'b': 'Access managed through PAM solution.', 'd': 'All privileged access to servers, databases, and network devices is controlled and monitored via the PAM system, ensuring centralized management.', 'f': 'Enhances security, accountability, and auditing of privileged user actions.', 'h': 'Periodically review PAM logs and access policies to maintain effectiveness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Access to servers, databases, network devices, and critical devices is not managed through PAM.', 'd': 'It was observed that access to servers, databases, network devices, and critical devices is not managed through PAM. This deficiency raises significant security risks, as there is no centralized control or monitoring over privileged access. Without proper access management, the organization is susceptible to unauthorized access and potential breaches. The absence of PAM also hinders the enforcement of least privilege principles and the ability to track user activities, posing a challenge to maintaining a comprehensive audit trail.', 'f': 'Without PAM, privileged users (such as system administrators or IT staff) may have unrestricted access to sensitive systems and data, which increases the risk of data breaches, cyber attacks, and other security incidents. Unmanaged privileged access can also result in unauthorized changes to systems, accidental data loss or corruption, and compliance violations.', 'h': 'It is recommended to manage access to all servers, databases network devices and critical devices are through PAM solution. . By adopting PAM, the bank can establish centralized control and monitoring over privileged access, reducing the risk of unauthorized entry and potential security breaches.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # pamExclusiveAccess
            'compliance': {'a': 'Compliance', 'b': 'Exclusive access through PAM solution.', 'd': 'All administrative and privileged access is strictly enforced via PAM, eliminating bypass options.', 'f': 'Provides full visibility, accountability, and control over privileged sessions.', 'h': 'Regularly audit access paths to ensure exclusive use of the PAM system.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Access to all servers, databases, network devices, and critical devices is not exclusively granted only through the PAM solution.', 'd': 'It was observed that access to all servers, databases, network devices, and critical devices is not exclusively granted through a PAM solution. Bank employees also use Remote Desktop Protocol service and other means like Putty/browser to connect to the devices and servers.', 'f': 'Access to critical devices and sensitive data should be managed through a PAM solution to ensure that only authorized users can access these resources, and all access is audited and monitored. Unmanaged privileged access can lead to compliance violations, legal issues, and reputational damage. Without PAM, there is a lack of centralized control and monitoring over privileged access, which increases the risk of unauthorized entry and potential security breaches.', 'h': 'It is recommended to access all servers, databases, network devices and critical devices exclusively through PAM solution only.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # pamExclusionsApproved
            'compliance': {'a': 'Compliance', 'b': 'PAM exclusions authorized.', 'd': 'Any exceptions to PAM access are formally approved and documented, following the principle of least privilege.', 'f': 'Reduces risk of privilege misuse while accommodating necessary operational requirements.', 'h': 'Review and revalidate all PAM exceptions periodically to ensure continued appropriateness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Exclusions to the PAM solution are not approved by the respective authority and are not granted on a need-to-know basis.', 'd': 'Certain users have privileged access outside PAM without proper authorization or documentation.', 'f': 'Without proper authorization, there is a heightened risk of unauthorized access to critical resources and sensitive data. This deficiency undermines the effectiveness of the PAM solution, potentially leading to security breaches and data leaks. Lack of adherence to need-to-know principles hampers access control, making it challenging to track and manage user privileges.', 'h': 'It is recommended to establish a robust approval process for PAM exclusions and enforce access based on the principle of least privilege. This will enhance security, reduce unauthorized access risk, and safeguard critical resources and data.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # pamCompatible
            'compliance': {'a': 'Compliance', 'b': 'PAM solution compatible with all critical assets.', 'd': 'PAM system is fully integrated and interoperable with all servers, databases, and network devices, ensuring seamless management of privileged access.', 'f': 'Enables comprehensive monitoring, control, and auditing of privileged sessions across all assets.', 'h': 'Test PAM integrations periodically after updates or additions of new assets.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'PAM solution was not compatible with other assets.', 'd': 'The PAM solution is unable to integrate effectively with certain servers, databases, or network devices, leading to partial management of privileged access.', 'f': 'Authentication problems and access management inefficiencies may arise, leading to operational disruptions and potential security vulnerabilities. Incompatibility can impede the effective integration of the PAM solution with existing systems and applications, hindering the implementation of robust access controls and centralized monitoring. This deficiency may result in increased risk of unauthorized access and difficulties in tracking privileged user activities.', 'h': "It is recommended to address the PAM solution's compatibility issues promptly. Conduct a thorough assessment and consider a PAM solution that seamlessly integrates with existing systems. Regularly monitor for updates and patches to ensure continued compatibility with evolving technologies, enhancing access management, security, and operational efficiency."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # pamHaMode
            'compliance': {'a': 'Compliance', 'b': 'PAM solution deployed in HA mode.', 'd': 'PAM system is configured with redundancy and failover mechanisms to provide uninterrupted access to privileged accounts.', 'f': 'Ensures business continuity and reliable management of administrative access.', 'h': 'Regularly test failover scenarios to verify HA functionality.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'PAM solution was not in HA mode.', 'd': 'The PAM system is deployed without redundancy, creating a single point of failure that may disrupt access to critical systems if it goes offline.', 'f': 'PAM solution becomes a single point of failure, susceptible to downtime and service disruptions in the event of hardware failures, network outages, or other unforeseen incidents.', 'h': 'It is recommended to configure the PAM solution in High Availability mode to ensure system resilience, availability, and uninterrupted privileged access management operations.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # pamLogsServer
            'compliance': {'a': 'Compliance', 'b': 'PAM logs centralized on log server.', 'd': 'All PAM activities, including user logins, session starts, and command execution, are logged centrally on a secure log server.', 'f': 'Improves auditability, monitoring, and accountability of privileged user actions.', 'h': 'Regularly review log retention policies and monitor for anomalies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Log server was not configured for PAM.', 'd': 'Privileged access activities are not being sent to a dedicated log server, and logs are stored locally on the PAM server or individual devices.', 'f': 'Without a Log server, there is a lack of centralized and secure storage for logs, hindering the organization\'s ability to monitor user activities, detect security incidents, and perform comprehensive audits. The absence of centralized logs also limits the organization\'s ability to investigate and mitigate potential threats effectively.', 'h': 'It is recommended to promptly configure a Log server for the PAM solution. Implementing a Log server will centralize and secure logs, enabling effective monitoring of user activities and timely detection of security incidents.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # pamAdequateLogs
            'compliance': {'a': 'Compliance', 'b': 'Detailed logs maintained.', 'd': 'System logs include all necessary information such as user ID, session timestamps, executed commands, and affected resources.', 'f': 'Enhances traceability, forensic capability, and regulatory compliance.', 'h': 'Periodically audit logs to ensure accuracy and completeness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'System logs did not contain adequate information for identification purposes.', 'd': 'PAM logs do not capture user identifiers, session timestamps, or executed commands comprehensively, hindering accountability.', 'f': 'Insufficient details in the logs hinder the organization\'s ability to effectively track and identify user activities and potential threats. This deficiency compromises the accuracy of incident investigations, forensic analysis, and security monitoring. Without comprehensive information in the logs, identifying the source of security breaches or unauthorized access becomes challenging, leading to delayed incident response and potential data breaches.', 'h': 'It is recommended to enhance the logging configuration to include adequate information in the system logs for identification purposes. By capturing relevant details such as user identities, timestamps, and actions performed, the organization can improve incident response capabilities and security monitoring.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # pamVideoLogs
            'compliance': {'a': 'Compliance', 'b': 'PAM video session logging enabled.', 'd': 'All privileged sessions are recorded via video logs capturing user activity, keystrokes, and commands for audit purposes.', 'f': 'Strengthens accountability, provides visual audit trails, and aids in forensic investigations.', 'h': 'Regularly verify video logging functionality and storage integrity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Video logs were not enabled for PAM user sessions.', 'd': 'Privileged user sessions are not recorded, resulting in lack of visual evidence of actions performed on critical systems.', 'f': 'It limits visibility, hindering the monitoring of privileged user activities and detection of unauthorized actions. Accountability is compromised due to the inability to attribute actions to individuals. Without video logs, the organization may face challenges in investigating suspicious behaviour or policy violations.', 'h': 'It is recommended to enable video logs for PAM user sessions. By implementing video logs, the organization can enhance its security monitoring and incident response capabilities. Video logs provide visual records of user actions, enabling more effective investigation of suspicious behaviour and policy violations.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # pamTwoFactorAuth
            'compliance': {'a': 'Compliance', 'b': '2FA enabled for PAM access.', 'd': 'All privileged users are required to authenticate using two factors (e.g., password + token or OTP) before accessing the PAM system.', 'f': 'Provides an additional security layer, significantly reducing the risk of unauthorized access.', 'h': 'Periodically test 2FA mechanisms and review access policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Two-factor authentication was not available in PAM.', 'd': 'Privileged users can access PAM system using only a username and password without additional authentication factors.', 'f': 'Without 2FA, the authentication process relies solely on a single factor, making it easier for attackers to gain unauthorized access if credentials are compromised. This increases the risk of privilege escalation and data breaches. The absence of Two-Factor Authentication (2FA) in the PAM configuration increases the system\'s vulnerability to unauthorized access, data breaches, and potential security threats.', 'h': 'Implement Two-Factor Authentication (2FA) for all PAM logins to enhance security.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # pamMultipleAdmin
            'compliance': {'a': 'Compliance', 'b': 'Multiple admin accounts maintained.', 'd': 'Multiple authorized administrative accounts are available with appropriate roles and permissions to ensure continuous PAM management.', 'f': 'Prevents operational disruption and ensures continuity of privileged access management.', 'h': 'Review and manage admin accounts periodically to avoid excessive privileges.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'It was found that only one admin account was created.', 'd': 'Only one administrative account exists for PAM management, creating a risk of complete lockout if credentials are lost or compromised.', 'f': 'If only one admin account is there, then if the admin account gets locked out, it will affect the bank\'s productivity and day-to-day operations. Thus, the bank may face financial losses due to delays in daily operations. It results in a single point of failure, increasing the risk of unauthorized access and potentially hindering critical administrative tasks. The absence of multiple accounts with varying privileges poses challenges for accountability and auditing.', 'h': 'Maintain multiple administrative accounts with proper authorization and role segregation to prevent lockout scenarios.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # pamUnusedAccountsDisabled
            'compliance': {'a': 'Compliance', 'b': 'Unused accounts disabled.', 'd': 'Accounts that are inactive or no longer required are disabled immediately in PAM and associated systems.', 'f': 'Reduces the attack surface and strengthens overall access control.', 'h': 'Maintain periodic audits to verify all unused accounts are disabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Unused accounts were not disabled.', 'd': 'Privileged or standard accounts that are no longer required remain active in the PAM system or on integrated devices.', 'f': 'The failure to disable unused accounts in a Privileged Access Management (PAM) system can result in increased security risks, unauthorized access, lack of accountability, and resource wastage.', 'h': 'Regularly review all accounts in PAM and integrated systems and promptly disable any unused accounts.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # pamSolutionUpdated
            'compliance': {'a': 'Compliance', 'b': 'PAM solution updated.', 'd': 'PAM is running the latest version, with all critical patches applied as per vendor guidance.', 'f': 'Ensures protection against known vulnerabilities and maintains system integrity.', 'h': 'Continuously monitor vendor releases and apply updates promptly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'PAM solution was not updated.', 'd': 'The PAM software is running outdated versions or missing critical security patches.', 'f': 'Outdated PAM solutions might contain known vulnerabilities that attackers can exploit to gain unauthorized access or compromise user accounts. As new security threats and attack vectors emerge over time, failing to update the PAM solution leaves the system exposed to potential risks. Additionally, compatibility issues with other system components or software may arise, leading to reduced performance and functionality.', 'h': 'Keep the PAM solution updated with the latest security patches and software versions as released by the vendor.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # pamPasswordPolicy
            'compliance': {'a': 'Compliance', 'b': 'PAM password policy aligned with organizational standards.', 'd': 'PAM enforces password rules consistent with the organization\'s policy, including complexity, minimum length, expiration, and reuse prevention.', 'f': 'Strengthens privileged account security and reduces the risk of unauthorized access.', 'h': 'Periodically review PAM password policy to ensure continued compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'PAM password policy was not configured as per the organizational password policy.', 'd': 'Password policies in PAM do not comply with the organization\'s defined standards, including complexity, expiration, or reuse restrictions.', 'f': 'A weak PAM password policy can increase the risk of password-related security incidents such as unauthorized access, data breaches, and cyber attacks. Weak passwords are easy to guess or crack, and can compromise the security of critical systems and data.', 'h': 'Configure PAM password policies to enforce organizational password standards, including complexity, expiration, and reuse restrictions.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # pamDevicePasswordPolicy
            'compliance': {'a': 'Compliance', 'b': 'Device password policies managed through PAM.', 'd': 'PAM enforces and rotates passwords for all integrated devices according to organizational policy.', 'f': 'Enhances security of critical systems by standardizing password management.', 'h': 'Continuously monitor device integrations to ensure all systems are included in PAM management.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Password policy for devices integrated into PAM was not managed through PAM.', 'd': 'Integrated systems\' passwords are not controlled or rotated through PAM, leading to inconsistent password management.', 'f': 'Without a centralized password policy, users may choose weak passwords that are easy to guess or crack, leaving the system vulnerable to brute-force attacks. Users may reuse passwords across multiple devices and systems, and password policy for that device may not be centralized and standardized across the organization.', 'h': 'Ensure that all integrated devices\' password policies are centrally managed and enforced through PAM.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # pam2FAConfigured
            'compliance': {'a': 'Compliance', 'b': '2FA properly enforced.', 'd': 'All PAM users are required to authenticate using two factors (e.g., password + OTP/token), and enforcement is verified regularly.', 'f': 'Significantly reduces risk of unauthorized privileged access.', 'h': 'Periodically test 2FA functionality and maintain audit records of authentication logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Two-factor authentication was not configured for PAM.', 'd': 'Two-Factor Authentication for PAM access is either not configured correctly or is bypassable.', 'f': 'If an attacker gets the credential of the user account and 2FA is not enabled, then the attacker can easily take over the user\'s account and perform financial fraud.', 'h': 'Configure 2FA properly for all PAM users and test periodically to ensure it cannot be bypassed.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # pamConfigBackup
            'compliance': {'a': 'Compliance', 'b': 'PAM configuration backups performed regularly.', 'd': 'PAM settings are backed up periodically, and backup integrity is verified to ensure recovery capability.', 'f': 'Protects against accidental loss, corruption, or misconfiguration of PAM settings.', 'h': 'Maintain and periodically test backup procedures to ensure continuity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Configuration backup of PAM solution was not taken on a regular basis.', 'd': 'The PAM solution lacks routine backups, or backups are inconsistent and not verified for integrity.', 'f': 'Failing to take regular backups in a PAM solution system can result in data loss, inability to restore the system, limited recovery options, and increased recovery time and effort.', 'h': 'Schedule and automate regular PAM configuration backups and periodically test restoration processes to ensure recoverability.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # pamAdIntegration
            'compliance': {'a': 'Compliance', 'b': 'PAM integrated with Active Directory.', 'd': 'User accounts, groups, and authentication policies are synchronized with AD for centralized management.', 'f': 'Ensures consistency in access control, efficient user management, and faster de-provisioning.', 'h': 'Regularly verify the AD-PAM integration for synchronization accuracy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'PAM was not integrated with Active Directory.', 'd': 'User provisioning, authentication, and access control are not synchronized with Active Directory, resulting in manual management.', 'f': 'Not integrating a PAM solution with Active Directory can result in limited centralized user management, lack of single sign-on capability, inefficient access management, inconsistent user data, and increased administrative overhead.', 'h': 'Integrate PAM with Active Directory to centralize authentication and access management.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        18: {  # pamFilePermissions
            'compliance': {'a': 'Compliance', 'b': 'File type permissions managed through PAM.', 'd': 'All critical files and directories have access permissions enforced via PAM according to organizational policy.', 'f': 'Reduces risk of unauthorized data access or modification.', 'h': 'Periodically audit file permissions and PAM policies for accuracy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'File type permissions were not configured properly through PAM.', 'd': 'File type or resource access permissions are not managed through PAM, allowing potential unauthorized access or modification.', 'f': 'The impact of file type permissions not being configured properly through PAM can be significant for the organization\'s security and data integrity. Improper file type permissions may allow unauthorized users to access, modify, or delete critical files, leading to potential data breaches and unauthorized changes. Without proper configuration, sensitive information becomes vulnerable to unauthorized disclosure, and system stability may be compromised.', 'h': 'Configure and enforce file type permissions centrally through PAM to ensure consistent access control.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        19: {  # pamPasswordExpiry
            'compliance': {'a': 'Compliance', 'b': 'PAM passwords configured to expire.', 'd': 'All PAM user passwords follow the organization\'s defined expiration policy, requiring periodic changes.', 'f': 'Minimizes risk of password compromise and ensures adherence to security standards.', 'h': 'Periodically review password expiration settings and compliance reports.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Passwords did not expire after 30 days.', 'd': 'Privileged accounts in PAM have passwords that do not follow a periodic expiration policy.', 'f': 'Passwords that don\'t expire after 30 days can have a significant impact on the security of a system or network, including increased risk of password compromise and reduced accountability.', 'h': 'Configure passwords in PAM to expire periodically (preferably every 30 days) and enforce mandatory updates.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        20: {  # pamSessionMonitoring
            'compliance': {'a': 'Compliance', 'b': 'PAM session activity monitored.', 'd': 'All privileged sessions are logged and monitored by PAM administrators for compliance and security.', 'f': 'Enhances accountability and helps detect unauthorized or suspicious activities.', 'h': 'Continue regular monitoring and maintain audit trails for all sessions.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'PAM admin did not monitor user session activity.', 'd': 'Administrators do not actively monitor privileged sessions, and there is no review of session logs.', 'f': 'The lack of session monitoring hinders the organization\'s ability to detect potential security incidents and unauthorized access. Monitoring user session activity is essential for proactive threat detection and response, allowing the organization to identify suspicious behaviour or policy violations. Without proper monitoring, the organization may be unaware of malicious activities, putting sensitive data and assets at risk.', 'h': 'Implement real-time monitoring and logging of all PAM sessions, and conduct periodic reviews.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        21: {  # pamUserAccessPolicy
            'compliance': {'a': 'Compliance', 'b': 'User access policy properly configured.', 'd': 'PAM enforces access control policies consistently, assigning permissions based on roles and need-to-know basis.', 'f': 'Ensures that users have appropriate access rights and reduces the risk of privilege abuse.', 'h': 'Periodically review and update access policies as roles or organizational requirements change.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'User access policy was not configured properly.', 'd': 'PAM user roles, access levels, or permissions are inconsistently applied or do not reflect organizational policy.', 'f': 'If user access policy is not configured properly, it can lead to security breaches, as users with excessive privileges can access sensitive data, systems, or applications they should not be able to, and may cause accidental deletion or modification of data. Also, users may be unable to access the resources they need to do their jobs, leading to productivity losses and increased support costs.', 'h': 'Review and enforce user access policies centrally in PAM, ensuring compliance with organizational standards.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        22: {  # pam2FAEnabledAllUsers
            'compliance': {'a': 'Compliance', 'b': '2FA is enabled in all user accounts.', 'd': 'Two-Factor Authentication (2FA) is enabled for all user accounts to enhance login security.', 'f': 'Adds an extra layer of protection against unauthorized access, even if user credentials are compromised.', 'h': 'The bank should periodically review 2FA configurations, ensure coverage across all critical applications, and promote user awareness on secure authentication practices.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': '2FA is not enabled in all user accounts.', 'd': 'It was observed that two-factor authentication (2FA) is not enabled for all user accounts, particularly for some internal applications and administrative portals.', 'f': 'Absence of 2FA for all user accounts increases the risk of unauthorized access in case of compromised credentials. Attackers may exploit weak or stolen passwords to gain access to critical systems, leading to data breaches, financial fraud, or disruption of banking operations.', 'h': 'The bank should ensure that two-factor authentication is implemented for all user accounts, especially for privileged and remote access users. The 2FA mechanism should be integrated with the centralized identity management system and monitored for compliance.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        23: {  # pamCredentialBackup
            'compliance': {'a': 'Compliance', 'b': 'PAM credential backup maintained.', 'd': 'All PAM usernames and passwords are securely backed up, with access restricted and encryption enforced.', 'f': 'Prevents administrative disruption and ensures continuity of privileged access management.', 'h': 'Regularly test backup restoration and update backup procedures as needed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'No backup process was defined for the storage of usernames and passwords used for accessing PAM.', 'd': 'Usernames and passwords for PAM accounts are not backed up, or backups are not stored securely.', 'f': 'Not having a backup process for the storage of usernames and passwords used for accessing PAM raises the risks of data loss, service disruptions, unauthorized access, operational disruptions, and increased security breach risks. Without proper backups, any accidental deletion or hardware failure could lead to the permanent loss of critical authentication data, hindering user access and administrative tasks.', 'h': 'Maintain secure, encrypted backups of all PAM credentials and verify restore procedures regularly.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "PAM Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
