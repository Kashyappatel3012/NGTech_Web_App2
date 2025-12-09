import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_mail_messaging_excel(form_data=None):
    """
    Create Excel file for Mail and Messaging Assessment
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Mail and Messaging"
    
    # Define questions
    questions = [
        "Is Public Domain Email used?",
        "Is Two-step verification present?",
        "Is Device Approval configured?",
        "Is DKIM, and DMARC implemented in the Mail domain?",
        "Is Email Spoofing possible from the external and internal domains?",
        "Is virus & spam protection Enabled?",
        "Is mail monitoring done?",
        "Is Directory Harvest prevention Enabled?",
        "Is there sender authentication for all approved senders?",
        "Are all incoming emails scanned for malware?",
        "Whether all e-mails sent and received by employees via Bank's network are treated as Bank records?",
        "Whether there are procedures, which require that all the incoming e-mail messages be scanned for the virus to prevent virus infection to the Bank's network?",
        "Is IMAP/POP access disabled?",
        "Is Sender Policy Framework (SPF) enabled?",
        "Are Add-ons restricted for email and file sharing services?",
        "Whether Single mail is used by multiple user?",
        "Whether password expiry date is defined?",
        "Can less secure apps access user accounts?",
        "Is Early phishing detection disabled?",
        "Is an Unintended external reply warning off?",
        "Is Inbound and outbound filtering available?",
        "Is the Super Admin account used for daily administration?",
        "Are multiple admin accounts set up to avoid any admin lockout?",
        "Is logging enabled as per the organization's logging policy?",
        "Is automatic Email forwarding Disabled?",
        "Is the banner set for the external emails for users?",
        "Is VAPT done regularly for the mail exchange server hosted internally?",
        "Are mails outside of the domain policy enforced for internal users?",
        "Whether NTP is configured?",
        "Are mails/domains configured to bypass spam filters?",
        "Are all login activities periodically reviewed?",
        "Do all admin accounts have recovery options configured to recover from forgotten passwords, etc?",
        "Is admin email configured to receive alert notifications?",
        "Is Access from third-party services/applications restricted to drive?",
        "Is external calendar sharing limited to authorized users only?",
        "Is sharing option for Google Drive and OneDrive limited to authorized domains only?",
        "Is Anonymous file sharing disabled?",
        "Is the Bypass spam filters option off for internal senders?",
        "Whether e-mail ID allotted to a user prevented from being used by another user?",
        "Whether all e-mails are identified with a user's name or e-mail ID to facilitate tracking?"
    ]
    
    # Risk factors for each question
    risk_factors = [
        'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'Medium',
        'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium',
        'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Low',
        'Low', 'Low', 'Low', 'Low', 'Low', 'Low', 'Low', 'Low', 'Low', 'Low'
    ]
    
    # Question mapping for form fields
    question_mapping = [
        'publicDomainEmail', 'twoStepVerification', 'deviceApproval', 'dkimDmarc', 'emailSpoofing',
        'virusSpamProtection', 'mailMonitoring', 'directoryHarvestPrevention', 'senderAuthentication', 'malwareScanning',
        'bankRecords', 'virusScanningProcedures', 'imapPopDisabled', 'spfEnabled', 'addonsRestricted',
        'singleMailMultipleUsers', 'passwordExpiry', 'lessSecureApps', 'phishingDetection', 'externalReplyWarning',
        'inboundOutboundFiltering', 'superAdminDailyUse', 'multipleAdminAccounts', 'loggingEnabled', 'autoEmailForwarding',
        'externalEmailBanner', 'vaptRegular', 'domainPolicyEnforced', 'ntpConfigured', 'spamFilterBypass',
        'loginActivityReview', 'adminRecoveryOptions', 'adminAlertNotifications', 'thirdPartyAccessRestricted', 'externalCalendarSharing',
        'driveSharingRestricted', 'anonymousFileSharing', 'internalSenderBypass', 'uniqueEmailPerUser', 'emailIdentification'
    ]
    
    # Set column widths
    ws.column_dimensions['A'].width = 10  # Sr. No.
    ws.column_dimensions['B'].width = 50  # Questions
    ws.column_dimensions['C'].width = 20  # Compliance/Non-Compliance/Not Applicable
    ws.column_dimensions['D'].width = 30  # Observation (Short/Brief)
    ws.column_dimensions['E'].width = 20  # Risk Factor
    ws.column_dimensions['F'].width = 50  # Observation
    ws.column_dimensions['G'].width = 50  # Impact
    ws.column_dimensions['H'].width = 50  # Recommendation
    
    # Header row
    headers = ['Sr. No.', 'Questionnaire/Points', 'Compliance/Non-Compliance/Not Applicable', 
               'Observation (Short/Brief)', 'Risk Factor', 'Observation', 'Impact', 'Recommendation']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Populate Sr. No. and Questions
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions
    
    # Risk colors
    risk_colors = {
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }
    
    # Question responses data
    question_responses = {
        1: {  # publicDomainEmail
            'compliance': {'a': 'Compliance', 'b': 'Bank Specific domain email enforced.', 'd': 'All email communication is via Bank domain email accounts.', 'f': 'Reduces risk of phishing, malware, and unauthorized access to official communication.', 'h': 'Maintain private domain emails for all staff and regularly review access policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Public Domain Email was used.', 'd': 'It was observed that users are using public domain emails for official communication.', 'f': 'If public domain email is used, then users can get illegitimate mails that might contain various malicious links and attachments. Users are more likely to open an email when they think it has been sent by a legitimate or familiar source. Hackers continue to Public Domain Email.', 'h': 'It is recommended to use the private domain to easily manage email accounts and enforce organizational policies for the protection of email accounts.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # twoStepVerification
            'compliance': {'a': 'Compliance', 'b': 'Two-step verification enabled.', 'd': 'Email accounts are secured with 2FA.', 'f': 'Adds an extra layer of security, reducing the likelihood of account compromise even if passwords are leaked.', 'h': 'Enforce 2FA for all email accounts and periodically review authentication methods.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Two-step verification was not present. ', 'd': 'It was observed that two-step verification is not configured on email accounts.', 'f': 'If two-factor authentication is not present, an attacker can easily use the leaked credentials or crack the password to access that email account.', 'h': '''It is recommended to use two-factor authentications. With 2-Step Verification, you add an extra layer of security to your account. After one set it up, signing in the account id done in two steps using:-
 Something you know (your password)
 Something you have (like your phone or a security key dongle ex: Yubico) '''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # deviceApproval
            'compliance': {'a': 'Compliance', 'b': 'Device approval configured.', 'd': 'Email accounts can only be accessed from approved devices.', 'f': 'Enhances account security and prevents unauthorized access from untrusted devices.', 'h': 'Maintain and periodically review approved device lists.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Device approval was not configured.', 'd': 'Device approval is not enabled for email access.', 'f': 'Unauthorized devices can access email accounts if credentials are compromised, leading to potential data breaches.', 'h': 'Configure device approval policies to allow only trusted devices to access email accounts.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # dkimDmarc
            'compliance': {'a': 'Compliance', 'b': 'DKIM and DMARC implemented.', 'd': 'Mail domain has DKIM and DMARC records configured.', 'f': 'Protects against email spoofing and phishing attacks, improving trust in organizational emails.', 'h': 'Continuously monitor DKIM/DMARC reports to detect and respond to unauthorized email activity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'DKIM, and DMARC were not implemented in the Mail domain. ', 'd': 'DKIM and DMARC records are not configured in the mail domain.', 'f': 'If DKIM, and DMARC are not implemented, then the emails categorized as spam mails can be easily spoofed by the spammer to use for spam or phishing emails.', 'h': 'It is recommended to Implement DKIM-DMARC policies in the Mail domain. We recommend slowly deploying DMARC policies. Start with monitoring your traffic and looking for deviations in the reports, such as unsigned messages or spoofed messages.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # emailSpoofing
            'compliance': {'a': 'Compliance', 'b': 'Email spoofing blocked.', 'd': 'Internal and external email spoofing is prevented.', 'f': 'Reduces risk of phishing attacks and protects users from impersonation attempts.', 'h': 'Maintain anti-spoofing mechanisms and regularly audit email domain configurations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Email Spoofing was possible from external and internal domain.', 'd': 'Email spoofing was possible from both external and internal domains.', 'f': 'Attackers can impersonate legitimate entities to perform phishing or social engineering attacks, leading to credential theft or data compromise.', 'h': 'Implement SPF, DKIM, DMARC, subdomains, and content filters to prevent spoofing.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # virusSpamProtection
            'compliance': {'a': 'Compliance', 'b': 'Virus & spam protection enabled.', 'd': 'All emails are scanned for malware and spam.', 'f': 'Protects users from phishing, malware, and ransomware attacks, preserving data integrity.', 'h': 'Maintain updated virus and spam definitions and monitor email security reports.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Virus & spam protection was Disabled.', 'd': 'Virus and spam protection is not enabled in the email system.', 'f': 'If virus and spam protection are not enabled then bank employees can fall victim to phishing attacks, they may download malicious attachments which may contain viruses, trojans, and spyware which can compromise the CIA triad of information security.', 'h': 'Enable virus and spam protection to scan all incoming and outgoing emails for threats.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # mailMonitoring
            'compliance': {'a': 'Compliance', 'b': 'Mail monitoring performed.', 'd': 'Email accounts are monitored for suspicious activities.', 'f': 'Early detection of anomalies allows timely response and prevents security incidents.', 'h': 'Regularly review monitoring logs and take corrective actions promptly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Mail monitoring was not done.', 'd': 'Email activity is not monitored for suspicious behavior.', 'f': 'If mail monitoring is not done then bank admins will not be alerted of any suspicious activity that occurs by using the email accounts. They will not know if any email accounts are compromised or not.', 'h': 'It is recommended that mail monitoring should be done. It will prevent mischief and control the quality of mails that are being sent and received by the email users.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # directoryHarvestPrevention
            'compliance': {'a': 'Compliance', 'b': 'Directory Harvest prevention enabled.', 'd': 'DHA prevention is configured.', 'f': 'Prevents attackers from enumerating valid email accounts, reducing phishing/spam risk.', 'h': 'Continuously monitor and adjust DHA settings to maintain effectiveness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Directory Harvest prevention was Disabled.', 'd': 'Directory Harvest Attack (DHA) prevention is not configured.', 'f': 'The spammers can get to know the valid email address which can be used by them to perform phishing/spam attacks on the valid email addresses.', 'h': 'It is recommended to have Directory Harvest prevention Enabled.  DHA prevention works by limiting the number of unsuccessful email delivery attempts, which can help prevent spammers from using automated tools to harvest valid email addresses from your server.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # senderAuthentication
            'compliance': {'a': 'Compliance', 'b': 'Sender authentication enabled.', 'd': 'All approved senders are authenticated.', 'f': 'Ensures that only verified senders communicate, mitigating phishing/spoofing risks.', 'h': 'Periodically review authentication records to ensure continued compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Sender authentication was not available for all approved senders. ', 'd': 'Approved senders are not authenticated.', 'f': 'Not having sender authentication increases the risk of spoofing and phishing/whaling attacks. Email received from unauthenticated whitelisted domains will not be filtered for spam. This might result in spoofing and phishing/whaling attacks leading to account compromise of email users.', 'h': 'It is recommended that sender authentication must be available for all approved senders.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # malwareScanning
            'compliance': {'a': 'Compliance', 'b': 'Emails scanned for malware.', 'd': 'All incoming emails are scanned before delivery.', 'f': 'Reduces risk of malware infection and protects network infrastructure.', 'h': 'Maintain regular updates of scanning engines and malware definitions.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'All incoming emails were not scanned for malware.', 'd': 'Not all incoming emails are scanned for malware.', 'f': 'If incoming emails are not scanned for malware then bank employees can unknowingly download malicious attachments which can compromise their machine and also infect the whole network. The attachments may contain a virus, spyware, keylogger, or trojan which can affect the CIA triad of information security.', 'h': 'It is recommended that all incoming emails are scanned for malware. The sandbox allows admins to add an extra layer of protection on top of the standard attachment scans for known viruses and malware.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # bankRecords
            'compliance': {'a': 'Compliance', 'b': 'All emails treated as Bank records.', 'd': 'Emails are recorded and stored as Bank records.', 'f': 'Ensures accountability and provides a reliable audit trail for forensic investigations.', 'h': 'Regularly review email archiving processes to ensure all records are securely stored and accessible for investigations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'All e-mails sent and received by employees via Bank’s network were not treated as Bank records.', 'd': 'All e-mails sent and received by employees via Bank\'s network were not maintained as official Bank records.', 'f': 'Makes it difficult to track suspicious activity, perform forensic analysis, or investigate incidents effectively.', 'h': 'All e-mails sent and received via Bank\'s network should be treated as official Bank records and stored securely for audit and investigation purposes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # virusScanningProcedures
            'compliance': {'a': 'Compliance', 'b': 'Virus scanning implemented.', 'd': 'All incoming and outgoing emails are scanned for viruses.', 'f': 'Reduces risk of malware infection, protecting network integrity and sensitive data.', 'h': 'Regularly update virus definitions and monitor scan results for unusual activity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Inbound and outbound filtering to prevent virus infection to the Bank’s network was not available.', 'd': 'Inbound and outbound virus filtering was not configured.', 'f': 'Users may download malicious attachments or links, potentially infecting systems and compromising data integrity, confidentiality, and availability.', 'h': 'Configure inbound and outbound email scanning rules to prevent virus infections and block suspicious attachments.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # imapPopDisabled
            'compliance': {'a': 'Compliance', 'b': 'IMAP/POP access disabled.', 'd': 'IMAP and POP protocols are disabled for all users.', 'f': 'Minimizes unauthorized access risk via email protocols.', 'h': 'Monitor for any attempts to bypass this control and ensure only approved access methods are used.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'IMAP/POP access was enabled.', 'd': 'IMAP and POP protocols were enabled, allowing users to access emails outside the secure interface.', 'f': 'Attackers can use these protocols to download mailboxes and gain unauthorized access to sensitive information.', 'h': 'Disable IMAP and POP access for all users; enforce secure webmail or approved clients.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # spfEnabled
            'compliance': {'a': 'Compliance', 'b': 'SPF enabled.', 'd': 'SPF record is configured and active for all domains.', 'f': 'Protects against spoofed emails and improves email authenticity.', 'h': 'Periodically review SPF records and logs to ensure effectiveness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Sender Policy Framework (SPF) was Disabled.', 'd': 'SPF record was not configured for the Bank\'s domain.', 'f': 'Increases risk of email spoofing, allowing attackers to impersonate Bank email addresses.', 'h': 'Enable SPF records on the Bank\'s domain to validate legitimate sending servers and prevent spoofing.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # addonsRestricted
            'compliance': {'a': 'Compliance', 'b': 'Add-ons restricted.', 'd': 'Only approved add-ons are allowed for email and file sharing services.', 'f': 'Reduces risk of data leaks or compromise through insecure software.', 'h': 'Periodically review add-on permissions and enforce strict installation policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Addons were not restricted to email and file sharing services.', 'd': 'Users could install add-ons without restrictions, including outdated or insecure ones.', 'f': 'Vulnerable add-ons can access sensitive data, causing data breaches or unauthorized data modification.', 'h': 'Restrict installation of unverified or outdated add-ons; allow only secure, verified add-ons approved by IT policy.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # singleMailMultipleUsers
            'compliance': {'a': 'Compliance', 'b': 'Unique email per user.', 'd': 'Each employee has a personal email account for banking operations.', 'f': 'Enhances accountability and traceability for all email activities.', 'h': 'Regularly audit email usage to ensure no account sharing occurs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Single mail used by multiple users.', 'd': 'Multiple users shared a single email account for banking operations.', 'f': 'Accountability is unclear; difficult to identify responsible users during investigations or forensic analysis.', 'h': 'Assign unique email accounts to each user and maintain policies defining user responsibilities.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # passwordExpiry
            'compliance': {'a': 'Compliance', 'b': 'Password expiry defined.', 'd': 'Passwords expire periodically according to Bank policy.', 'f': 'Reduces risk of long-term credential misuse.', 'h': 'Maintain and review password policies for strength and expiry compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The password expiry date was not defined. ', 'd': 'Passwords do not expire periodically.', 'f': 'Stolen or leaked credentials can be reused indefinitely, increasing account compromise risk.', 'h': 'Enforce periodic password expiration (e.g., every 45–60 days) and ensure strong password policies.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        18: {  # lessSecureApps
            'compliance': {'a': 'Compliance', 'b': 'Less secure apps blocked.', 'd': 'Access from less secure apps is disabled.', 'f': 'Enhances account security and reduces exposure to attacks.', 'h': 'Periodically review app access logs to ensure compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Less secure apps can access user accounts.', 'd': 'Users could access accounts through less secure applications.', 'f': 'The attacker tries to access a user’s account using a less secure app to compromise the email account of the user for malicious purposes. ', 'h': 'It is recommended to disable access to less secure apps for all users. Banks can block sign-in attempts from some apps or devices that are less secure. '},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        19: {  # phishingDetection
            'compliance': {'a': 'Compliance', 'b': 'Early phishing detection enabled.', 'd': 'Anti-phishing mechanisms are active.', 'f': 'Reduces risk of sensitive data compromise and email-based attacks.', 'h': 'Regularly update phishing rules and monitor alerts for suspicious activity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Early phishing detection was disabled.', 'd': 'Anti-phishing mechanisms were not configured.', 'f': 'Attackers use Phishing mail to obtain sensitive information such as usernames, passwords, account information, etc for personal gains.', 'h': 'Enable advanced anti-phishing features to detect malicious emails and links proactively.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        20: {  # externalReplyWarning
            'compliance': {'a': 'Compliance', 'b': 'External reply warning enabled.', 'd': 'Users are warned when replying to external emails.', 'f': 'Prevents accidental data leaks and enhances information confidentiality.', 'h': 'Regularly test and validate the external reply warning feature.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'An Unintended external reply warning was off.', 'd': 'No warning is provided when replying to emails outside the organization.', 'f': 'Bank will not be able to get a warning to protect users from an unintentional leak of internal data. No such warning appears when mail is sent to an external domain. Users can unknowingly share confidential data with external users which will compromise the confidentiality of the data.', 'h': 'It is recommended to turn on the Unintended external reply warning option when your users respond to an email message, the recipients may be external to your organization. You can set up external recipient alerts to protect your users from unintentionally sharing information externally.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        21: {  # inboundOutboundFiltering
            'compliance': {'a': 'Compliance', 'b': 'Filtering available.', 'd': 'Inbound and outbound email filtering is active.', 'f': 'Reduces risk of malware, phishing, and spam threats.', 'h': 'Periodically review filtering rules to ensure effectiveness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Inbound and outbound filtering was not available.  ', 'd': 'Inbound and outbound filtering for emails was not configured.', 'f': "If inbound and outbound filtering is not available then the users can download harmful files from untrusted emails which can have trojans, and viruses which can hamper the bank's business and compromise the CIA triad of information security.", 'h': 'It is recommended to deploy an email filtering solution that can filter both inbound and outbound messages to protect users and customers.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        22: {  # superAdminDailyUse
            'compliance': {'a': 'Compliance', 'b': 'Super Admin restricted.', 'd': 'Super Admin accounts are not used for daily administration.', 'f': 'Reduces the risk of privilege escalation and compromise.', 'h': 'Regularly review Super Admin activity logs to detect misuse.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Super Admin account was used for daily administration.', 'd': 'Super Admin account was used for routine email and system administration.', 'f': 'Allowing a systems administrator, especially one with Domain Administrator privileges, to access his/her e-mail and the Internet via their administrative account makes it easier for attackers to introduce malware via a phishing attack or gain those credentials by using impersonation.', 'h': 'It is recommended that the Super Admin account is not used for daily administration. The bank should implement policies that will secure your super admin accounts and make users less likely to use them for day-to-day operations.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        23: {  # multipleAdminAccounts
            'compliance': {'a': 'Compliance', 'b': 'Multiple admin accounts configured.', 'd': 'Redundant admin accounts exist to prevent lockouts.', 'f': 'Ensures uninterrupted system management and operational continuity.', 'h': 'Periodically test admin account redundancy to confirm availability.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Multiple admin accounts were not setup to avoid any admin lockout.', 'd': 'Only one admin account exists; no redundancy to prevent lockout.', 'f': "A locked account cannot be used until an administrator unlocks it. If the administrator account gets locked, it will be hard for an admin to manage the user account or bank's day-to-day operations. Thus, a bank may face productivity and financial loss. ", 'h': 'It is recommended that multiple admin accounts should be setup to avoid any admin lockout.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        24: {  # loggingEnabled
            'compliance': {'a': 'Compliance', 'b': 'Logging enabled.', 'd': 'Logging is active and aligned with Bank policy.', 'f': 'Provides traceability and supports incident response and forensic investigations.', 'h': 'Regularly audit logs for anomalies and retention compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Logging was not enabled as per the organization's logging policy.", 'd': 'Logging is not configured in line with Bank policies.', 'f': 'If Logs settings are not properly configured, Bank will not be able to detect the DoS attack, IP Spoofing attacks, and unauthorized access of data. Without logs the bank will not be able to trace the event trail to catch the attacker.', 'h': "It is recommended that logging should be enabled as per the organization's logging policy."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        25: {  # autoEmailForwarding
            'compliance': {'a': 'Compliance', 'b': 'Auto-forwarding disabled.', 'd': 'Automatic forwarding is disabled.', 'f': 'Protects sensitive information from being sent outside the organization without authorization.', 'h': 'Periodically review forwarding rules for compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Automatic Email  forwarding was Enabled.', 'd': 'Users can forward emails automatically to external addresses.', 'f': 'Increases risk of confidential data leakage if credentials are compromised.', 'h': 'Disable automatic email forwarding to prevent unauthorized data exfiltration.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        26: {  # externalEmailBanner
            'compliance': {'a': 'Compliance', 'b': 'Banner set.', 'd': 'External email banners are configured and visible.', 'f': 'Helps users distinguish external emails, reducing phishing risks.', 'h': 'Ensure banner messages are clear and regularly updated.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The banner  was not set for the external emails for users.', 'd': 'External email warnings or banners were not configured.', 'f': 'Users may fail to recognize external emails, increasing phishing risk.', 'h': 'Configure banners for all external emails to alert users about potential risks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        27: {  # vaptRegular
            'compliance': {'a': 'Compliance', 'b': 'VAPT performed regularly.', 'd': 'Mail server is regularly assessed through VAPT.', 'f': 'Ensures emerging threats are identified and mitigated proactively.', 'h': 'Maintain VAPT reports and remediate issues promptly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'VAPT was not done regularly for the mail exchange server hosted internally.', 'd': 'Mail server has not undergone periodic vulnerability assessment and penetration testing (VAPT).', 'f': 'Attackers can exploit known vulnerabilities, compromising mail server confidentiality, integrity, and availability.', 'h': 'Conduct VAPT periodically on mail servers to identify and remediate vulnerabilities.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        28: {  # domainPolicyEnforced
            'compliance': {'a': 'Compliance', 'b': 'Mails from outside of the domain policy does not enforce for internal users.', 'd': 'Internal users comply with domain email policies.', 'f': 'Reduces risk of unauthorized email-based data leaks.', 'h': 'Monitor email traffic to ensure policy enforcement.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Mails outside of the domain policy are enforced for internal users.', 'd': 'Internal users are allowed to send emails outside the domain policy.', 'f': 'Increases risk of phishing and unauthorized disclosure of sensitive information.', 'h': 'Restrict email sending policies to internal communication only unless authorized.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        29: {  # ntpConfigured
            'compliance': {'a': 'Compliance', 'b': 'NTP configured.', 'd': 'All systems are synchronized with NTP servers.', 'f': 'Accurate timekeeping aids incident investigations and event correlation.', 'h': 'Periodically verify NTP synchronization status.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'NTP is not configured.', 'd': '''It was observed that the NTP server was not configured.  NTP is intended to synchronize all participating computers within a few milliseconds of Coordinated Universal Time (UTC).
to mitigate the effects of variable network latency.''', 'f': 'A more insidious effect of weak timekeeping is that it damages the ability to investigate security breaches and other kinds of system problems. Hackers, for example, will often exploit backdoor, and proxy computers when mounting and attacking- both to hide their tracks and to exploit whatever opportunities (like NTP System privileges ) the hacker encounters along the way. Finding these stopping-off points is critical for shutting the door to future attacks and requires precise measurements of time in order to reconstruct the exact sequence of events. log file and application time stamp obviously become essential pieces of evidence.', 'h': 'It is recommended that Configured  NTP.All Compute Engine virtual machine instances are configured to use the internal Google NTP services.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        30: {  # spamFilterBypass
            'compliance': {'a': 'Compliance', 'b': 'No bypass configured.', 'd': 'All emails are scanned through spam filters.', 'f': 'Reduces risk of phishing, malware, and spam threats.', 'h': 'Regularly review whitelisted domains to ensure security compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Emails/domains were configured to bypass spam filters.', 'd': 'Some emails or domains are allowed to bypass spam filters.', 'f': 'Attackers can use whitelisted domains to send phishing or malicious emails without detection.', 'h': 'Ensure no emails or domains bypass spam filters unless explicitly approved after risk assessment.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        31: {  # loginActivityReview
            'compliance': {'a': 'Compliance', 'b': 'Login activities reviewed.', 'd': 'All login activities are periodically monitored.', 'f': 'Helps detect and respond to unauthorized access promptly.', 'h': 'Implement automated alerts for unusual login activity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'All login activities were not periodically reviewed.', 'd': 'All login activities were not periodically reviewed.', 'f': 'Bank will not detect unauthorized access or malicious activities in a timely manner, increasing risk to systems.', 'h': 'Periodically review all login activities to identify unauthorized access and potential threats.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        32: {  # adminRecoveryOptions
            'compliance': {'a': 'Compliance', 'b': 'Recovery options configured.', 'd': 'All admin accounts have recovery options enabled.', 'f': 'Ensures smooth access recovery and continuity of administration.', 'h': 'Regularly test recovery mechanisms for functionality.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'All admin accounts do not have recovery options configured to recover from forgotten passwords.', 'd': 'Admin accounts do not have recovery options configured.', 'f': 'If the bank Admin Password does not have recovery options configured and in case an attacker takes over the admin account, or the admin forgets the password, then it will be hard for the bank to recover the admin account to manage the mail system of the bank. Thus banks will face delays in day-to-day operations that can result in financial loss for the bank and its customers.', 'h': 'Configure recovery options for all admin accounts to enable password reset and account recovery.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        33: {  # adminAlertNotifications
            'compliance': {'a': 'Compliance', 'b': 'Alerts configured.', 'd': 'Admin email receives security alerts and notifications.', 'f': 'Enables timely response to security incidents.', 'h': 'Periodically review alert settings to ensure coverage of all critical events.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Admin email was not configured to receive alert notifications.', 'd': 'Admin email was not configured to receive alert notifications.', 'f': 'If admin emails are not configured to receive alert notifications, then bank admin will not get necessary security alerts of email accounts for unauthorized logins or suspicious activities happening by using the email accounts of the bank.', 'h': 'It is recommended that the admin email is configured to receive alert notifications. Bank can configure center email alert notifications. Bank can also use a rule to configure other settings, for example, to turn the alert center notification on.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        34: {  # thirdPartyAccessRestricted
            'compliance': {'a': 'Compliance', 'b': 'Third-party access restricted.', 'd': 'Access from third-party apps is restricted to authorized applications only.', 'f': 'Reduces risk of unauthorized data access or leakage.', 'h': 'Periodically review access policies for new apps or updates.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Access to third-party services/applications was not restricted to drive.', 'd': 'Access from third-party services/applications was not restricted.', 'f': 'Users can open Drive files with third-party apps and scripts that use the Drive API. People can create programs or scripts that use the Drive API to access data from your domain.', 'h': 'It is recommended to restrict access from third-party services/applications so that users cannot access data from third-party apps, and APIs, to protect the confidentiality, integrity, and availability of information.  '},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        35: {  # externalCalendarSharing
            'compliance': {'a': 'Compliance', 'b': 'Sharing restricted.', 'd': 'External calendar sharing is limited to authorized personnel.', 'f': 'Prevents unauthorized access to sensitive scheduling information.', 'h': 'Regularly audit calendar sharing permissions.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'External calendar sharing was not limited to authorized users only.', 'd': 'External calendar sharing was not limited to authorized users.', 'f': 'External calendar sharing was not limited to only authorized users, any malicious user can reschedule, cancel, or know about important events.', 'h': 'It is recommended to limit external calendar sharing to authorized persons only.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        36: {  # driveSharingRestricted
            'compliance': {'a': 'Compliance', 'b': 'Sharing restricted.', 'd': 'Sharing is limited to approved domains.', 'f': 'Ensures data confidentiality and reduces risk of leakage.', 'h': 'Periodically review sharing policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Sharing option for Google Drive and OneDrive was not limited to authorized domains only.', 'd': 'File sharing was not limited to authorized domains.', 'f': 'If the file sharing option is not limited to authorized google drive or Onedrive domains then anyone can share files outside these domains, compromising the confidentiality of the files for personal gains.', 'h': 'It is recommended that sharing option for Google Drive and OneDrive should be limited to authorized domains only to protect the confidentiality of the data.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        37: {  # anonymousFileSharing
            'compliance': {'a': 'Compliance', 'b': 'Anonymous sharing disabled.', 'd': 'Anonymous file sharing is disabled.', 'f': 'Protects confidentiality of bank data.', 'h': 'Periodically verify that anonymous access remains disabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Anonymous file sharing was Enabled.', 'd': 'Anonymous file sharing was enabled.', 'f': 'If anonymous file sharing is enabled then anyone can upload files to their account and compromise the confidentiality of the document or share it with malicious attackers for personal gains.', 'h': 'It is recommended to disable the Anonymous file sharing option to protect the confidentiality of the data.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        38: {  # internalSenderBypass
            'compliance': {'a': 'Compliance', 'b': 'Bypass disabled.', 'd': 'Internal emails are filtered through spam detection mechanisms.', 'f': 'Reduces internal propagation of malicious emails.', 'h': 'Regularly audit email filter configurations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Spam filters can be bypassed by internal senders.', 'd': 'Internal senders can bypass spam filters.', 'f': 'If the internal senders can bypass the spam filters then their email can be used for spam mails. Thus, maximizing the chances for spoofing and phishing/whaling attacks.', 'h': 'It is recommended to Turn off Bypassing of the spam filters for messages received from internal senders for all organizational units. By turning this setting off, Bank can make sure that all Bank users’ email are filtered for spam, including mail from internal senders.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        39: {  # uniqueEmailPerUser
            'compliance': {'a': 'Compliance', 'b': 'Unique mail per user.', 'd': 'Each user has a separate email account.', 'f': 'Improves accountability and simplifies auditing.', 'h': 'Maintain user-to-email mapping regularly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Single mail used by multiple users.', 'd': 'Multiple users shared a single email account.', 'f': 'Because of using single mail with multiple user Email system, User’s accountability cannot be defined. Thus, it will be very difficult to track the person responsible for the incident.', 'h': 'It is recommended to provide separate mail for every user so that user accounts can be defined. '},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        40: {  # emailIdentification
            'compliance': {'a': 'Compliance', 'b': 'Emails identifiable.', 'd': 'All emails are tagged with user identifiers.', 'f': 'Facilitates auditing and forensic investigation.', 'h': 'Regularly audit email logs for proper tagging.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'E-mails were not identified with a user’s name or e-mail ID to facilitate tracking.', 'd': 'Emails are not tagged with user identifiers for tracking.', 'f': "If emails are not tracked using the user's name or email id then the bank won't be able to catch the rogue users doing suspicious activities, or the users who violated the organization's policy.", 'h': "It is recommended that all e-mails should be identified with a user’s name or e-mail ID to facilitate tracking."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Apply formatting to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for col in range(1, 9):  # A to H
        cell = ws.cell(row=1, column=col)
        cell.border = thin_border
    
    # Process each question
    for i, question in enumerate(question_mapping, 2):  # Start from row 2
        # Get user input for this question
        user_input = form_data.get(question, 'not_applicable') if form_data else 'not_applicable'
        
        # Get response data
        response_data = question_responses.get(i-1, {})
        user_response = response_data.get(user_input, {})
        
        # Populate columns based on user input
        # Column C: Compliance/Non-Compliance/Not Applicable
        ws.cell(row=i, column=3, value=user_response.get('a', 'Not Applicable'))
        
        # Column D: Observation (Short/Brief)
        ws.cell(row=i, column=4, value=user_response.get('b', 'Not Applicable'))
        
        # Column F: Observation
        ws.cell(row=i, column=6, value=user_response.get('d', 'Not Applicable'))
        
        # Column G: Impact
        ws.cell(row=i, column=7, value=user_response.get('f', 'Not Applicable'))
        
        # Column H: Recommendation
        ws.cell(row=i, column=8, value=user_response.get('h', 'Not Applicable'))
        
        # Column E: Risk Factor with color coding
        risk_factor = risk_factors[i-2]  # risk_factors is 0-indexed
        risk_cell = ws.cell(row=i, column=5, value=risk_factor)
        risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
        risk_cell.fill = PatternFill(start_color=risk_colors.get(risk_factor, '808080'), end_color=risk_colors.get(risk_factor, '808080'), fill_type='solid')
        risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply formatting to all data rows
    for row in range(2, len(question_mapping) + 2):  # Rows 2 to 41
        for col in range(1, 9):  # Columns A to H
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Special alignment for column A (Sr. No.), C (Compliance status), and E (Risk Factor)
            if col == 1 or col == 3 or col == 5:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Set row height for wrapped text
            ws.row_dimensions[row].height = 30
    
    # Save file
    filename = "Mail and Messaging Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    """
    Delete the generated Excel file after download
    """
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            print(f"File {filepath} deleted successfully")
    except Exception as e:
        print(f"Error deleting file {filepath}: {e}")
