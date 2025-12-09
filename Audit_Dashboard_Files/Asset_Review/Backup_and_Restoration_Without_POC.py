import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_backup_restoration_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Backup and Restoration"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # Backup and Restoration Questions
    questions = [
        "Are all the floppies/CDs/tapes purchased for OS software, application software, utility programs, drivers, etc., recorded in a register and properly stored?",
        "Is daily/weekly/monthly and quarterly backup of data taken without fail and available as per requirements?",
        "Is the data integrity verified after the restoration work is completed?",
        "Are backup tapes verified/tested periodically by restoring the data, and is a record maintained?",
        "Are backup tapes properly labeled and numbered?",
        "Are proper storage procedures and facilities in place for backup copies?",
        "Is backup media verified periodically for readability?",
        "Is the latest copy of the backup of software (Operating System, RDBMS, application, etc.) taken and preserved at the user site?",
        "Are hardware, software, operating system, and printer manuals properly labeled and maintained?"
    ]

    # Risk Factors
    risk_factors = [
        "High", "High", "High", "Medium", "Low", "Low", "Low", "Low", "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "brSoftwareMediaRegister": 1,
        "brBackupSchedule": 2,
        "brDataIntegrityVerification": 3,
        "brBackupTapeTesting": 4,
        "brBackupTapeLabeling": 5,
        "brStorageProcedures": 6,
        "brMediaReadabilityVerification": 7,
        "brSoftwareBackupPreservation": 8,
        "brManualsMaintenance": 9
    }

    # Question responses data
    question_responses = {
        1: {  # brSoftwareMediaRegister
            'compliance': {'a': 'Compliance', 'b': 'All floppies, CDs, tapes, and other media related to OS, applications, and utilities are recorded in a register and securely stored.', 'd': 'All physical and digital installation media are logged in a register and kept in a secure cabinet with restricted access.', 'f': 'Ensures traceability, accountability, and legal compliance for licensed software.', 'h': 'Continue maintaining and reviewing the media register periodically for accuracy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The register was not maintained for the floppies/CDs/tapes purchased pertaining to the OS software, application software, utility programs, drivers, etc.', 'd': 'The organization does not maintain an updated register for CDs, tapes, or other installation media, and some software copies are stored in unsecured locations.', 'f': 'It can be risky not to know about their warranty status or availability. It reduces asset availability and increases business risk unnecessarily. Not maintaining and storing assets properly can lead to increased costs, compliance breaches, and poor performance. ', 'h': 'It is recommended to record in a register and properly store all the floppies/CDs/tapes purchased about the OS software, application software, utility programs, drivers, etc.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # brBackupSchedule
            'compliance': {'a': 'Compliance', 'b': 'Backup performed regularly as per policy.', 'd': 'Daily incremental and periodic full backups are performed automatically and verified by the IT team.', 'f': 'Ensures business continuity and data protection against accidental deletion or system failures.', 'h': 'Continue regular monitoring and review of backup logs to ensure consistency.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Back-up of data was not taken.', 'd': 'Backups are not consistently taken according to the defined daily, weekly, or monthly schedule, and some recent backups were unavailable.', 'f': 'Without a good backup regime, we are putting the bank at risk of data loss in the event of an accident or disaster. If data backup is not taken, some important data (customer & bank) might get lost, if the device is damaged or the internal drive gets corrupted. Also, not having a backup can impact the bank by lower productivity/efficiency, Lost partnerships, Reputation damage, and No data for customer services/support.', 'h': 'It is recommended to take daily/weekly/monthly and quarterly back-up of data without fail and it should be available (as per requirement).'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # brDataIntegrityVerification
            'compliance': {'a': 'Compliance', 'b': 'Data integrity verified after restoration.', 'd': 'Data restored from backup is verified for integrity and consistency using automated validation checks.', 'f': 'Confirms data reliability and completeness, ensuring confidence in recovery processes.', 'h': 'Maintain records of validation tests and include integrity checks in all DR drills.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The data integrity was not checked after the restore operation was complete.', 'd': 'Restored data is not validated for accuracy or completeness after recovery testing, leaving potential integrity issues undetected.', 'f': "Integrity protects the system data from intentional or accidental unauthorized changes. If data integrity is not verified it would be difficult to determine if the data maintained in the bank is the same as expected by the users and the data won't be reliable. We won't be able to determine if the integrity of the data has been lost during storage or backup.", 'h': 'Implement a post-restoration validation process comparing checksum or hash values to ensure data integrity.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # brBackupTapeTesting
            'compliance': {'a': 'Compliance', 'b': 'Backup tapes tested and documented.', 'd': 'Backup tapes are periodically tested for readability and restoration success, with test logs properly maintained.', 'f': 'Ensures the reliability and usability of backups during system failures.', 'h': 'Continue periodic testing and maintain restoration logs for audit purposes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Backup tapes were not tested periodically.', 'd': 'The organization does not conduct periodic restoration testing of backup tapes or maintain related records.', 'f': 'Backups that are not tested regularly are essentially useless. Without consistent testing, there is the risk of losing the data, applications, systems, and workloads that the backups contain, potentially with no way to recover them. That is why a comprehensive testing plan is necessary to ensure that the backups will perform as expected in a disaster scenario.', 'h': 'It is recommended that Backup tapes are verified/tested periodically by restoring the data and records should be maintained. When the backup is taken in the bank, it should be checked weekly/monthly.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # brBackupTapeLabeling
            'compliance': {'a': 'Compliance', 'b': 'Backup tapes properly labeled and indexed.', 'd': 'All backup media are sequentially labeled and cataloged in a register, ensuring traceable record management.', 'f': 'Facilitates quick identification and retrieval during restoration.', 'h': 'Continue enforcing standardized labeling and indexing procedures.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Backup tapes were not properly labelled and numbered.', 'd': 'Backup media lack clear labelling and numbering, making it difficult to identify content and version.', 'f': 'Can lead to confusion or delays during data restoration.', 'h': 'Label all backup media with clear identifiers, including backup date, system, and retention period.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # brStorageProcedures
            'compliance': {'a': 'Compliance', 'b': 'Proper storage facility maintained for backups.', 'd': 'Backup media are stored in a dedicated, secure, and environmentally controlled area with access restricted to authorized personnel.', 'f': 'Enhances protection and longevity of backup media.', 'h': 'Continue monitoring environmental conditions and reviewing storage security regularly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Proper storage procedures and facilities are not available for backup copies.', 'd': 'Backup media are stored in areas lacking environmental control or security, increasing risk of damage or unauthorized access.', 'f': 'Storage and backup procedures make custodian data available on devices other than an individual computer, such as on network servers or backup tapes. If backup procedures and facilities are not available for backup copies then, there is a possibility of data loss, and the specific piece of information cannot be quickly/easily accessed. Also, the integrity of data gets compromised as data can be altered at any point because of a lack of controls on storage and backup procedure.', 'h': 'It is recommended that proper storage procedures and facilities should be made available for backup copies.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # brMediaReadabilityVerification
            'compliance': {'a': 'Compliance', 'b': 'Backup media readability checked regularly.', 'd': 'Backup tapes and disks are periodically read and verified using automated tools to ensure continued accessibility.', 'f': 'Guarantees reliable data recovery and minimizes backup corruption risks.', 'h': 'Maintain verification logs and continue proactive monitoring.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Backup media was not verified periodically for readability. ', 'd': 'Periodic checks to confirm media readability are not conducted, leading to unverified data recovery reliability.', 'f': 'Backup media may become unreadable over time, risking data recovery failure.', 'h': 'Implement periodic media verification tests and replace deteriorated media promptly.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # brSoftwareBackupPreservation
            'compliance': {'a': 'Compliance', 'b': 'Latest software backups preserved locally.', 'd': 'The most recent versions of system and application software are backed up and stored at the user location.', 'f': 'Enables rapid recovery and minimizes downtime in the event of system failure.', 'h': 'Continue updating and securing backup copies after every major software change.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The latest copy of the backup of software (Operating System, RDBMS, application, etc.) was not taken and preserved at the user site.  ', 'd': 'The latest copies of OS, RDBMS, and applications are not preserved locally, delaying restoration during outages.', 'f': 'Increases downtime and dependency on external support during recovery.', 'h': 'Keep the latest verified backup copies of all software components securely at the user site for quick restoration.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # brManualsMaintenance
            'compliance': {'a': 'Compliance', 'b': 'Manuals labeled and well-maintained.', 'd': 'All technical manuals are labeled, indexed, and securely stored for quick access during maintenance or audit.', 'f': 'Facilitates efficient troubleshooting and adherence to operational procedures.', 'h': 'Keep manuals updated following any hardware or software changes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Manuals not labeled or updated.', 'd': 'Equipment and software manuals are either misplaced, unlabelled, or outdated, limiting their utility during troubleshooting.', 'f': 'Causes delay in issue resolution and may result in misconfigurations.', 'h': 'Maintain and label all manuals systematically, ensuring they are updated and readily accessible for reference.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Backup and Restoration Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
