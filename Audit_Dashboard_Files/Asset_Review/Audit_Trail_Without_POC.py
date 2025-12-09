import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_audit_trail_excel(form_data=None):
    """
    Create Audit Trail Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Trail"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Audit Trail Questions
    questions = [
        "Does the audit trail report generate the user ID of the operator and official for any addition/modification/deletion of transaction data in the database?",
        "Is the audit trail report generated daily, and are entries scrutinized and verified?",
        "Does the audit trail for products/services record all identification and authentication processes?",
        "Is there a process to log and review all actions performed by system operators, system managers, system engineers, system administrators, security administrators, and highly privileged IDs?",
        "Is a list of canceled entries scrutinized, and are reasons for cancellation recorded?"
    ]

    # Risk Factors
    risk_factors = [
        "High",
        "Medium", 
        "Medium",
        "Medium",
        "Medium"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "atUserIDGeneration": 1,
        "atDailyReportGeneration": 2,
        "atIdentificationAuthentication": 3,
        "atPrivilegedUserActions": 4,
        "atCanceledEntriesReview": 5
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Question responses data
    question_responses = {
        1: {  # atUserIDGeneration
            'compliance': {'a': 'Compliance', 'b': 'Audit trail captures user IDs.', 'd': 'The system logs the user ID for all critical database activities, including additions, modifications, and deletions, providing a clear trail of responsibility.', 'f': 'Enhances accountability, enables forensic analysis, and ensures regulatory compliance.', 'h': 'Periodically review audit trail configurations to confirm all critical activities are captured accurately.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The audit trail report does not generate the user ID of the operator and the official for any addition/modification/deletion of the transaction data affected in the database.', 'd': 'The system fails to record the user ID of operators or officials responsible for additions, modifications, or deletions of transaction data in the database.', 'f': 'Audit trails are used to verify and track many types of transactions, including accounting transactions and trades in brokerage accounts. If the audit trail report has not generated the user ID for any changes made to the transaction database, then the bank will not be able to verify and track the user who made changes to the transaction database.', 'h': 'It is recommended that the audit trail report should generate the user ID of the operator and the official for any addition/modification/deletion of the transaction data affected in the database for forensic purposes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # atDailyReportGeneration
            'compliance': {'a': 'Compliance', 'b': 'Daily audit trail review implemented.', 'd': 'Audit trail reports are generated daily and reviewed by authorized personnel to verify entries and detect any discrepancies or unauthorized actions.', 'f': 'Ensures timely detection of irregularities, improves operational oversight, and strengthens internal controls.', 'h': 'Maintain documentation of daily review findings and periodically audit the review process for effectiveness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'An Audit trail report was not generated daily, and entries were not scrutinized and verified.', 'd': 'Audit trail reports are not generated daily, or generated reports are not scrutinized for anomalies or unauthorized activities.', 'f': 'If audit trail reports are not generated daily and entries are not scrutinized/verified, the fraud or default transactions performed by a malicious attacker or disgruntled employees will get unnoticed. Thus banks will face financial, and reputational loss.', 'h': ' It is recommended that the Audit trail reports are generated daily and entries are scrutinized and verified.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # atIdentificationAuthentication
            'compliance': {'a': 'Compliance', 'b': 'Identification/authentication fully logged.', 'd': 'Audit trails comprehensively record all identification and authentication events for products and services, ensuring complete traceability of user activities.', 'f': 'Strengthens accountability, enhances security monitoring, and facilitates forensic investigations.', 'h': 'Periodically test audit configurations to ensure all relevant authentication events are being captured accurately.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The audit trail for product/service does not record all identification and authentication processes.', 'd': 'Audit trails for products or services do not capture all user identification and authentication activities, such as logins, access attempts, or role validations.', 'f': 'If the audit trail for a product/service does not record all identification and authentication processes, the fraud or default transactions performed by a malicious attacker or disgruntled employees will go unnoticed. In the case of cyber attack there will not be enough evidence to identify the culprit and learn how to protect against it.', 'h': 'It is recommended that the audit trail for product/service records all identification and authentication processes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # atPrivilegedUserActions
            'compliance': {'a': 'Compliance', 'b': 'Privileged user actions logged and reviewed.', 'd': 'All activities performed by highly privileged users are logged, and a review process is in place to detect anomalies or unauthorized actions.', 'f': 'Enhances security oversight, accountability, and ensures regulatory compliance for critical system activities.', 'h': 'Conduct periodic audits of privileged user logs to verify completeness and accuracy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Not all processes performed by system operators, systems managers, system engineers, system administrators, security administrators, and highly privileged IDs are logged and reviewed.', 'd': 'Actions performed by privileged users, including system operators and administrators, are not consistently logged or reviewed, leaving critical activities unmonitored.', 'f': 'If all processes performed by highly privileged IDs are not logged and reviewed, the activities performed by the malicious attacker will go unnoticed and bank will not be able to identify the security breach. Thus bank might face financial loss, reputational loss because of cyber attack.', 'h': 'It is recommended to have a process to log and review all actions performed by systems operators, systems managers, system engineers, system administrators, security administrators, and highly privileged Ids.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # atCanceledEntriesReview
            'compliance': {'a': 'Compliance', 'b': 'Cancellations monitored and recorded.', 'd': 'All canceled entries are logged with detailed reasons, and these records are regularly reviewed to ensure proper authorization and detect irregularities.', 'f': 'Enhances accountability, prevents misuse, and ensures audit readiness.', 'h': 'Maintain a documented review process and periodically verify that all cancellations are properly authorized and recorded.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The cancelled entries were not scrutinized and reasons for cancellation were not recorded.', 'd': 'The system does not maintain a record of canceled entries, or reasons for cancellations are not documented and reviewed.', 'f': 'If cancelled entries are not scrutinized then evidence pertaining to an incident can be lost forever. If the bank wants to analyse the cyber attack or find the root cause of it then it might be very hard to find trail as there is no evidence because cancelled entries were not scrutinized.', 'h': 'It is recommended that the cancelled entries are scrutinized and reasons for cancellation are recorded.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }

    # Populate data based on user input
    for i, question in enumerate(questions, 2):
        # Get user input for this question
        question_num = i - 1
        user_input = None
        
        if form_data:
            # Find the corresponding form field
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Audit Trail Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
