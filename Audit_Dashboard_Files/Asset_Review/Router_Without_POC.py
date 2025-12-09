import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_router_excel(form_data=None):
    """
    Create Excel file for Router Assessment
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Router"
    
    # Define questions
    questions = [
        "Whether Router is used to provide Wi-Fi?",
        "Whether Session timeout was defined?",
        "Is Telnet service Disabled?",
        "Is Idle access/login timeout set to 15 minutes/lower or according to policy?",
        "Is Router in HA (High Availability) Mode?",
        "Is ICMP Redirects Disabled?",
        "Are unused management protocols disabled?",
        "Are unused router interfaces disabled?",
        "Is configuration for new device notification available?",
        "Is IP filtering and MAC filtering enabled?",
        "Is Weak Password in use?",
        "Whether DHCP (Dynamic Host Configuration Protocol) enabled in the router?",
        "Is System logs configured?",
        "Is URL filtering configured?",
        "Is Firmware updated?",
        "Is Default Password in use?",
        "Are gratuitous ARP and Proxy ARP disabled?",
        "Is IP-directed broadcast disabled?",
        "Do Administrators have two-factor authentication enabled?",
        "Are router login IDs and passwords treated as sensitive information and managed by authorized administrators?",
        "Whether USB Access Disabled?",
        "Is IP Source Routing Disabled?",
        "Is Private VLAN in use?",
        "Is Dynamic ARP Inspection performed?",
        "Are Password Policies configured as per the organization's password management policy?",
        "Are Notification Banners Implemented?",
        "Whether NTP (Network Time Protocol) is configured?",
        "Is the User limit for accessing the Wi-Fi Set?",
        "Is the Administrative access to the external (Internet-facing) interfaces disabled?",
        "Is No Service Password-Recovery Disabled?",
        "Are Secure Interactive Management Sessions available?",
        "Are console and auxiliary (AUX) ports configured?",
        "Is Anti-Spoofing Protection Disabled?"
    ]
    
    # Risk factors for each question
    risk_factors = [
        'High', 'High', 'High', 'High', 'High', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium',
        'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium',
        'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Low', 'Low',
        'Low', 'Low', 'Low'
    ]
    
    # Question mapping for form fields
    question_mapping = [
        'routerWifi', 'sessionTimeout', 'telnetDisabled', 'idleTimeout', 'routerHA',
        'icmpRedirects', 'unusedProtocols', 'unusedInterfaces', 'deviceNotification', 'ipMacFiltering',
        'weakPassword', 'dhcpEnabled', 'systemLogs', 'urlFiltering', 'firmwareUpdated',
        'defaultPassword', 'arpDisabled', 'ipDirectedBroadcast', 'twoFactorAuth', 'credentialManagement',
        'usbAccess', 'ipSourceRouting', 'privateVlan', 'dynamicArpInspection', 'passwordPolicies',
        'notificationBanners', 'ntpConfigured', 'wifiUserLimit', 'adminAccessExternal', 'passwordRecovery',
        'secureManagement', 'consoleAuxPorts', 'antiSpoofing'
    ]
    
    # Set column widths
    ws.column_dimensions['A'].width = 10  # Sr. No.
    ws.column_dimensions['B'].width = 50  # Questions
    ws.column_dimensions['C'].width = 20  # Compliance/Non-Compliance/Not Applicable
    ws.column_dimensions['D'].width = 30  # Observation (Short/Brief)
    ws.column_dimensions['E'].width = 20  # Risk Factor
    ws.column_dimensions['F'].width = 50  # Observation
    ws.column_dimensions['G'].width = 50  # Impact
    ws.column_dimensions['H'].width = 50  # Recommendation
    
    # Header row
    headers = ['Sr. No.', 'Questionnaire/Points', 'Compliance/Non-Compliance/Not Applicable', 
               'Observation (Short/Brief)', 'Risk Factor', 'Observation', 'Impact', 'Recommendation']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Populate Sr. No. and Questions
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }
    
    # Question responses data
    question_responses = {
        1: {  # routerWifi
            'compliance': {'a': 'Compliance', 'b': 'Router not used for Wi-Fi.', 'd': 'No Wi-Fi service is provided through the router.', 'f': 'Prevents potential unauthorized network access.', 'h': 'Maintain Wi-Fi disabled on routers.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The router was used to provide the Wi-Fi connection.', 'd': 'The router is actively providing Wi-Fi access to users. This exposes the bank\'s internal network to external access, which is against RBI/NABARD guidelines. Unauthorized users can connect without oversight and exploit vulnerabilities in the Wi-Fi network.', 'f': 'Malicious actors can gain unauthorized access to the network. They may intercept sensitive data, inject malicious content, or manipulate traffic. The presence of Wi-Fi creates multiple attack vectors that can compromise confidentiality, integrity, and availability of critical banking systems.', 'h': 'Disable Wi-Fi on all routers immediately. Restrict connectivity to wired access only, ensure no rogue access points exist, and enforce network isolation for critical systems to prevent unauthorized access and attacks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # sessionTimeout
            'compliance': {'a': 'Compliance', 'b': 'Session timeout defined.', 'd': 'User sessions terminate automatically after the defined period.', 'f': 'Prevents unauthorized use of unattended sessions.', 'h': 'Maintain session timeout as per policy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Session timeout not defined.', 'd': 'No session timeout is configured, allowing users to stay logged in indefinitely. If an authorized user leaves their system unattended, attackers can exploit the session to perform unauthorized actions or access sensitive information.', 'f': 'Attackers can hijack active sessions and perform malicious operations, increasing the risk of data leakage, fraud, or system compromise. Unmonitored sessions create significant exposure for critical banking infrastructure.', 'h': 'Configure a session timeout of 10-15 minutes. Ensure inactive sessions are automatically logged out to prevent unauthorized access and minimize risk of session hijacking.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # telnetDisabled
            'compliance': {'a': 'Compliance', 'b': 'Telnet service disabled.', 'd': 'Telnet is not active, preventing insecure remote access.', 'f': 'Reduces risk of credential theft and network compromise.', 'h': 'Keep Telnet disabled and use SSH for remote management.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Telnet service enabled.', 'd': 'Telnet service is active on the router, transmitting all credentials in plain text. This makes it possible for attackers on the network to capture login information and gain administrative access to critical devices.', 'f': 'Attackers can intercept credentials and take control of network devices, compromising data confidentiality and integrity. The service is highly vulnerable and may allow attackers to perform session hijacking or network intrusion, affecting banking operations.', 'h': 'Disable Telnet immediately and replace it with secure protocols like SSH. Ensure all sensitive data is transmitted over encrypted channels to prevent credential theft and unauthorized access.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # idleTimeout
            'compliance': {'a': 'Compliance', 'b': 'Idle timeout set as per policy.', 'd': 'Users are logged out automatically after idle period.', 'f': 'Reduces unauthorized access via unattended sessions.', 'h': 'Maintain idle timeout according to policy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Idle timeout not set as per policy.', 'd': 'The system does not enforce an idle access timeout. Users remaining logged in without activity can be exploited by attackers to perform unauthorized actions. Unattended sessions may be hijacked, leading to potential compromise of banking systems.', 'f': 'Attackers can misuse idle sessions to gain access to sensitive systems, potentially leading to data theft, fraud, or system disruption. Lack of timeout exposes critical infrastructure to extended attack windows.', 'h': 'Set idle timeout to 5-10 minutes or lower. Regularly review session policies to ensure inactive sessions are automatically terminated, reducing exposure to unauthorized access.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # routerHA
            'compliance': {'a': 'Compliance', 'b': 'Router in HA mode.', 'd': 'Router has redundancy to ensure uninterrupted operations.', 'f': 'Reduces downtime and ensures continuity of critical services.', 'h': 'Maintain router in HA configuration.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Router not in HA mode.', 'd': 'The Head Office router is not configured for High Availability. A single point of failure exists, which can disrupt connectivity across branches and the head office if the router fails. This lack of redundancy exposes operations to unnecessary downtime.', 'f': 'Network failure could halt banking operations, affecting all connected branches. Critical services may be interrupted, reducing productivity and creating financial and operational risk due to lack of failover mechanisms.', 'h': 'Implement HA for the router. Ensure failover configuration is in place to maintain continuity of network services and prevent operational disruption in case of device failure.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # icmpRedirects
            'compliance': {'a': 'Compliance', 'b': 'ICMP Redirects are disabled.', 'd': 'The router does not send ICMP redirect messages, preventing misuse of network routing information.', 'f': 'Reduces the risk of CPU overload and routing attacks.', 'h': 'Keep ICMP Redirects disabled to maintain secure routing.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'ICMP Redirects are enabled by default.', 'd': 'The router is configured to send ICMP redirect messages whenever packets are received on certain interfaces. This behaviour allows external or internal attackers to manipulate routing tables, potentially misdirecting traffic or overwhelming the device.', 'f': 'Attackers can exploit ICMP redirects to reroute traffic, cause network instability, or increase CPU load on the router. This may degrade network performance and create opportunities for man-in-the-middle attacks.', 'h': 'Disable ICMP Redirects immediately. Review router configurations to prevent unintended routing messages and ensure the network operates securely without exposure to routing manipulation attacks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # unusedProtocols
            'compliance': {'a': 'Compliance', 'b': 'Unused management protocols are disabled.', 'd': 'The router only allows required management protocols, reducing attack surface.', 'f': 'Enhances network security by limiting external attack vectors.', 'h': 'Keep all unused management protocols disabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Unused management protocols are enabled.', 'd': 'The router has multiple management protocols enabled that are not required for operations. These protocols can be accessed remotely and may allow attackers to connect from the internet. This exposes critical network devices and data to potential attacks.', 'f': 'Malicious actors can exploit these unnecessary protocols to perform DoS attacks, gain unauthorized access, or compromise the confidentiality, integrity, and availability of network devices. The overall security posture of the network is weakened.', 'h': 'Disable all unused management protocols. Restrict access to only authorized administrators to prevent potential exploitation and reduce the attack surface.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # unusedInterfaces
            'compliance': {'a': 'Compliance', 'b': 'Unused router interfaces are disabled.', 'd': 'Only active interfaces are enabled, preventing unauthorized network access.', 'f': 'Reduces risk of exploitation via unused network ports.', 'h': 'Maintain unused interfaces disabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Unused router interfaces are enabled.', 'd': 'The router has several interfaces active that are not being used for any operational purpose. These open interfaces can be accessed by attackers to infiltrate the network, inject malicious traffic, or launch attacks such as DDoS, increasing exposure to unauthorized activity.', 'f': 'Hackers can exploit these interfaces to steal sensitive data or disrupt services. Unused open ports serve as entry points for malware, phishing, and network-based attacks, threatening the security of critical banking operations.', 'h': 'Disable all unused router interfaces. Restrict traffic to only authorized and necessary connections to reduce potential attack vectors and improve network security.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # deviceNotification
            'compliance': {'a': 'Compliance', 'b': 'New device notification is configured.', 'd': 'Administrators are alerted whenever a new device connects to the network.', 'f': 'Helps in timely detection of unauthorized devices.', 'h': 'Maintain new device notification enabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'New device notification is not configured.', 'd': 'The router does not notify administrators when a new device connects to the network. Without such notifications, unauthorized devices can join the network unnoticed, potentially installing malware or stealing sensitive information. This leaves the network vulnerable to undetected intrusions.', 'f': 'Attackers can connect rogue devices and compromise network security without triggering alerts. This can lead to malware installation, data theft, or exploitation of network resources, increasing operational and reputational risk.', 'h': 'Enable new device notification immediately. Monitor all device connections regularly to detect unauthorized access and respond promptly to potential security incidents.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # ipMacFiltering
            'compliance': {'a': 'Compliance', 'b': 'IP and MAC filtering are enabled.', 'd': 'Access is restricted to authorized devices, preventing unauthorized connections.', 'f': 'Enhances network control and prevents unauthorized internet access.', 'h': 'Maintain IP and MAC filtering enabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'IP and MAC filtering are disabled.', 'd': 'The router does not enforce IP or MAC filtering, allowing any device to connect to the network or internet. Unauthorized devices can access network resources freely, bypassing any intended restrictions, and potentially compromise sensitive data.', 'f': 'Lack of filtering enables unauthorized users to connect, potentially causing data leakage, network congestion, or malicious activity. Without restrictions, attackers can exploit network resources and reduce operational efficiency.', 'h': 'Enable IP and MAC filtering immediately. Implement whitelisting for authorized devices to restrict access and protect network integrity.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # weakPassword
            'compliance': {'a': 'Compliance', 'b': 'Strong passwords are in use.', 'd': 'Passwords meet complexity requirements and are resistant to brute-force attacks.', 'f': 'Reduces risk of unauthorized access and credential theft.', 'h': 'Maintain strong passwords with minimum complexity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Weak password is in use.', 'd': 'The router\'s password is short or easily guessable. This allows attackers to rapidly compromise credentials using brute-force or dictionary attacks. Weak passwords significantly increase the risk of unauthorized access to critical network devices and data.', 'f': 'Attackers can easily gain administrative access, potentially modifying configurations, stealing sensitive information, or disrupting network operations. Weak passwords create a critical vulnerability in the network security posture.', 'h': 'Implement strong passwords immediately, using a combination of alphanumeric characters, symbols, and a minimum of 8–12 characters. Regularly update passwords and avoid defaults or easily guessable strings.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # dhcpEnabled
            'compliance': {'a': 'Compliance', 'b': 'DHCP is disabled.', 'd': 'Only manual IP assignments are allowed, restricting unauthorized devices from connecting automatically.', 'f': 'Enhances control over devices connecting to the network and prevents unauthorized access.', 'h': 'Keep DHCP disabled and allow only manual IP assignment.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'DHCP is enabled on the router.', 'd': 'Any device connecting to the router automatically receives an IP address, allowing unrestricted internet access. Unauthorized devices can connect to the network without administrator knowledge, bypassing intended access restrictions.', 'f': 'Attackers can exploit automatic IP assignment to connect rogue devices, potentially gaining access to sensitive data or network resources. This reduces network control and increases the risk of malicious activities.', 'h': 'Disable DHCP immediately and configure manual IP settings. Only authorized devices should be assigned IP addresses to prevent unauthorized access and improve network security.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # systemLogs
            'compliance': {'a': 'Compliance', 'b': 'System logs are configured.', 'd': 'Logs record user and device activity, aiding monitoring and forensic investigations.', 'f': 'Supports auditing and security investigations, ensuring accountability.', 'h': 'Maintain system logs configured and monitored regularly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'System logs are not configured.', 'd': 'The router does not maintain logs of user or device activities. Any network events, malicious attempts, or policy violations remain unrecorded, leaving no evidence for investigations.', 'f': 'Lack of logging impedes forensic analysis, makes incident response difficult, and increases the risk of undetected breaches or misuse of network resources.', 'h': 'Configure system logs immediately to capture all critical events. Regularly monitor logs to detect anomalies and enable timely incident response.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # urlFiltering
            'compliance': {'a': 'Compliance', 'b': 'URL filtering is configured.', 'd': 'Access to non-business websites is restricted, reducing exposure to malware or distractions.', 'f': 'Prevents access to malicious or non-work-related sites, protecting productivity and security.', 'h': 'Keep URL filtering active and updated regularly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'URL filtering is not configured.', 'd': 'The router does not restrict access to websites, allowing users to browse any site freely. This can lead to accidental or intentional downloading of malware, trojans, or other malicious software, compromising network security.', 'f': 'Users may access harmful or inappropriate websites, leading to malware infections, data breaches, or reduced productivity. The absence of URL restrictions increases the risk of security incidents.', 'h': 'Implement URL filtering immediately, restricting access to only business-critical or authorized websites to safeguard the network and maintain operational efficiency.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # firmwareUpdated
            'compliance': {'a': 'Compliance', 'b': 'Firmware is updated.', 'd': 'Router runs the latest firmware, minimizing exposure to known vulnerabilities.', 'f': 'Ensures protection against known exploits and enhances overall network security.', 'h': 'Continue regular firmware updates.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Firmware is not updated.', 'd': 'The router is running outdated firmware with known vulnerabilities. Attackers can easily exploit these weaknesses to compromise the device or network. The lack of updates increases the likelihood of attacks targeting unpatched flaws.', 'f': 'Unpatched firmware allows attackers to execute known exploits, potentially gaining control over network devices or disrupting operations. The security of connected systems is significantly compromised.', 'h': 'Update firmware immediately and schedule regular updates. This ensures vulnerabilities are patched and the network remains protected against known threats.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # defaultPassword
            'compliance': {'a': 'Compliance', 'b': 'Default passwords are not in use.', 'd': 'Strong, unique passwords are set, preventing unauthorized access.', 'f': 'Reduces risk of unauthorized login and device compromise.', 'h': 'Maintain strong passwords with adequate complexity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Default password is in use.', 'd': 'The router is using the default password, making it extremely easy for attackers to gain administrative access. Any malicious user can log in without effort and modify router settings, potentially affecting all connected systems.', 'f': 'Default passwords allow attackers to fully control network devices, change configurations, or launch attacks on connected systems. This exposes the bank network to severe security risks.', 'h': 'Immediately replace default passwords with strong, unique passwords including alphanumeric characters, symbols, and a minimum of 8–12 characters. Regularly update these passwords to maintain security.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # arpDisabled
            'compliance': {'a': 'Compliance', 'b': 'ARP and Proxy ARP are disabled.', 'd': 'Gratuitous and Proxy ARP are not responding on the network, reducing unnecessary ARP traffic.', 'f': 'Reduces the risk of ARP spoofing, network congestion, and resource exhaustion.', 'h': 'Maintain ARP and Proxy ARP disabled on all interfaces.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Gratuitous ARP and Proxy ARP are enabled.', 'd': 'The router responds to ARP requests for other IPs, sending gratuitous ARP packets on the network. This can lead to excessive ARP traffic, which may affect network performance and stability, and allows attackers to perform ARP poisoning attacks.', 'f': 'Attackers can manipulate ARP tables to intercept traffic or exhaust network resources, potentially enabling man-in-the-middle attacks and service disruption.', 'h': 'Disable Gratuitous ARP and Proxy ARP immediately to prevent ARP-based attacks, reduce network load, and protect data integrity.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        18: {  # ipDirectedBroadcast
            'compliance': {'a': 'Compliance', 'b': 'IP-directed broadcast is disabled.', 'd': 'Router does not allow IP-directed broadcasts, preventing network amplification attacks.', 'f': 'Protects the network from smurf attacks and other broadcast amplification attacks.', 'h': 'Maintain IP-directed broadcast disabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'IP-directed broadcast is enabled.', 'd': 'The router forwards IP-directed broadcast packets to remote subnets. This functionality could be exploited to amplify attacks or reflect traffic to other networks, resulting in denial-of-service conditions.', 'f': 'Enables attackers to perform smurf or amplification attacks, potentially causing network outages, service disruption, or congestion across multiple subnets.', 'h': 'Disable IP-directed broadcasts to mitigate amplification attacks and reduce risk to network availability and stability.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        19: {  # twoFactorAuth
            'compliance': {'a': 'Compliance', 'b': 'Two-factor authentication (2FA) is enabled for administrators.', 'd': 'Administrator accounts are protected by 2FA, adding an extra security layer beyond passwords.', 'f': 'Reduces risk of unauthorized access and potential fraudulent activity.', 'h': 'Maintain 2FA for all administrator accounts.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Two-factor authentication is disabled for administrators.', 'd': 'Administrator accounts rely solely on passwords. If credentials are stolen or guessed, attackers can fully compromise administrative privileges. The lack of 2FA increases the likelihood of unauthorized access to sensitive systems.', 'f': 'Unauthorized users can bypass security measures, access critical systems, and perform fraudulent activities. It compromises the integrity of all administrative actions.', 'h': 'Enable 2FA for all administrator accounts immediately to prevent unauthorized access and reduce the risk of compromise.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        20: {  # credentialManagement
            'compliance': {'a': 'Compliance', 'b': 'Router login IDs and passwords are managed securely by authorized administrators.', 'd': 'Credentials are restricted, reducing the risk of unauthorized access.', 'f': 'Protects critical network configurations and sensitive information from compromise.', 'h': 'Continue secure credential management practices.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Router login IDs and passwords are not treated as sensitive and are unmanaged.', 'd': 'Credentials are accessible without proper controls, allowing unauthorized personnel to access and manipulate router configurations. Sensitive data could be leaked, intentionally or accidentally.', 'f': 'Unauthorized access to credentials can lead to full network compromise, data breaches, or manipulation of critical configurations, threatening confidentiality and operational integrity.', 'h': 'Manage all router login IDs and passwords as sensitive information, restricting access only to authorized administrators and enforcing secure handling procedures.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        21: {  # usbAccess
            'compliance': {'a': 'Compliance', 'b': 'USB access is disabled on the router.', 'd': 'No external devices can connect via USB, minimizing the risk of malware or unauthorized data transfer.', 'f': 'Reduces the risk of USB-borne malware and physical attacks.', 'h': 'Keep USB access disabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'USB access is enabled on the router.', 'd': 'Users can connect USB drives to the router, which could introduce malware, keyloggers, or other malicious tools. This opens a physical attack vector that could compromise the router and connected systems.', 'f': 'Attackers could use USB-based exploits to infect network devices, steal information, or disrupt services, putting critical infrastructure at risk.', 'h': 'Disable USB access immediately to prevent potential malware injection and unauthorized physical access.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        22: {  # ipSourceRouting
            'compliance': {'a': 'Compliance', 'b': 'IP Source Routing is disabled.', 'd': 'Router discards packets with source routing options, preventing routing manipulation.', 'f': 'Ensures network traffic cannot be rerouted maliciously, protecting internal network integrity.', 'h': 'Keep IP Source Routing disabled on all router interfaces.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'IP Source Routing is enabled.', 'd': 'Packets with source routing options are processed by the router. Attackers could spoof source addresses of internal hosts to redirect traffic and bypass security controls.', 'f': 'Enables potential man-in-the-middle attacks, unauthorized network access, and data interception, threatening confidentiality and integrity of internal communications.', 'h': 'Disable IP Source Routing to prevent traffic manipulation and protect internal network communications.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        23: {  # privateVlan
            'compliance': {'a': 'Compliance', 'b': 'Private VLAN is implemented.', 'd': 'Segmentation isolates ports to restrict unnecessary communication between devices.', 'f': 'Prevents compromised devices from affecting the broader network and enhances network security.', 'h': 'Maintain Private VLAN configuration for critical segments.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Private VLAN is not used.', 'd': 'All devices can communicate freely on the network without segmentation, which increases exposure. A compromised device could access or infect multiple systems.', 'f': 'Increases risk of lateral attacks and propagation of malware across the network. Critical systems are more vulnerable to compromise.', 'h': 'Implement Private VLANs to isolate devices and limit the spread of attacks within the network.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        24: {  # dynamicArpInspection
            'compliance': {'a': 'Compliance', 'b': 'Dynamic ARP Inspection (DAI) is configured.', 'd': 'ARP traffic is validated, preventing ARP poisoning on local network segments.', 'f': 'Reduces risk of man-in-the-middle attacks and ARP spoofing.', 'h': 'Maintain DAI on all critical interfaces.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Dynamic ARP Inspection is not performed.', 'd': 'The network does not validate ARP packets, allowing attackers to perform ARP spoofing and redirect traffic to themselves. ARP-based attacks are not detected or prevented.', 'f': 'Attackers can intercept sensitive data, manipulate traffic, or launch man-in-the-middle attacks. Network integrity and confidentiality are compromised.', 'h': 'Enable Dynamic ARP Inspection on all network segments to protect against ARP poisoning attacks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        25: {  # passwordPolicies
            'compliance': {'a': 'Compliance', 'b': 'Password policies are configured per organizational standards.', 'd': 'Users follow strong password rules, including complexity and rotation, ensuring account security.', 'f': 'Reduces the likelihood of brute-force attacks and unauthorized access.', 'h': 'Continue enforcing password policies consistently.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Password policies are not configured as per the organization\'s standards.', 'd': 'Weak or inconsistent passwords are allowed, increasing susceptibility to guessing or brute-force attacks. Users may use easily guessable or reused passwords, exposing critical accounts.', 'f': 'Attackers can compromise user accounts, access sensitive data, and potentially escalate privileges to gain administrative control.', 'h': 'Configure password policies according to organizational guidelines, enforcing minimum length, complexity, and periodic changes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        26: {  # notificationBanners
            'compliance': {'a': 'Compliance', 'b': 'Notification banners are implemented.', 'd': 'Users see security notices before login, informing them of legal or policy requirements.', 'f': 'Increases user awareness and compliance with security policies.', 'h': 'Maintain up-to-date notification banners on all systems.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Notification banners are not implemented.', 'd': 'No Message of the Day (MOTD) or login banners are displayed, so users are not informed about legal notices or security policies. Temporary or critical alerts cannot be communicated.', 'f': 'Users may be unaware of security obligations, increasing the risk of unauthorized or non-compliant actions, and reducing accountability.', 'h': 'Implement notification banners to inform users of legal and security policies before system access.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        27: {  # ntpConfigured
            'compliance': {'a': 'Compliance', 'b': 'NTP is configured.', 'd': 'All devices are synchronized with Coordinated Universal Time (UTC), ensuring consistent timestamps across logs and systems.', 'f': 'Accurate timekeeping supports incident investigation, log correlation, and security auditing.', 'h': 'Maintain NTP configuration for precise network time synchronization.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'NTP is not configured.', 'd': 'Devices have inconsistent timestamps, making it difficult to correlate events across systems. Attackers can exploit time discrepancies to cover tracks and manipulate logs.', 'f': 'Investigating security breaches or reconstructing attack sequences becomes challenging. Forensic evidence from logs may be unreliable, delaying incident response.', 'h': 'Configure NTP on all devices to synchronize clocks, ensuring accurate logs for forensic and operational purposes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        28: {  # wifiUserLimit
            'compliance': {'a': 'Compliance', 'b': 'User limit for Wi-Fi is set.', 'd': 'The number of simultaneous users is restricted according to bank policies, ensuring controlled bandwidth usage.', 'f': 'Prevents network congestion and ensures optimal internet performance for critical banking operations.', 'h': 'Maintain user limits based on bank requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'User limit for Wi-Fi is not set.', 'd': 'Any user can connect to the Wi-Fi without restrictions, potentially overloading the network. Bandwidth may be consumed by non-critical activities.', 'f': 'Internet speed may degrade during peak usage, impacting essential banking operations and productivity. Unauthorized or excessive usage can also increase security risks.', 'h': 'Implement user limits on Wi-Fi access to ensure bandwidth availability and enhance network security.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        29: {  # adminAccessExternal
            'compliance': {'a': 'Compliance', 'b': 'Administrative access is disabled on external interfaces.', 'd': 'Only internal users can access management interfaces, reducing exposure to external threats.', 'f': 'Minimizes risk of compromise via internet-facing devices and strengthens perimeter security.', 'h': 'Maintain administrative access restrictions for external interfaces.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Administrative access to external interfaces is enabled.', 'd': 'Internet-facing interfaces are accessible to all, allowing attackers to attempt unauthorized login and probing.', 'f': 'High risk of compromise, unauthorized access, and potential data theft. External threats can directly affect the bank network.', 'h': 'Disable administrative access to all Internet-facing interfaces to reduce attack surface and protect sensitive systems.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        30: {  # passwordRecovery
            'compliance': {'a': 'Compliance', 'b': 'No Service Password-Recovery is enabled.', 'd': 'Device prevents console users from bypassing password protection, maintaining configuration integrity.', 'f': 'Ensures that unauthorized users cannot reset or access device configurations.', 'h': 'Keep No Service Password-Recovery enabled on all critical devices.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'No Service Password-Recovery is disabled.', 'd': 'Anyone with console access can reset device passwords and access configurations. Malicious users can modify device settings or extract sensitive information from NVRAM.', 'f': 'Unauthorized configuration changes can compromise network security, enable privilege escalation, or cause service disruptions.', 'h': 'Enable No Service Password-Recovery to secure device configurations against unauthorized access.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        31: {  # secureManagement
            'compliance': {'a': 'Compliance', 'b': 'Secure Interactive Management Sessions are available.', 'd': 'All remote management uses encrypted protocols (SSH, HTTPS), preventing credential interception.', 'f': 'Reduces risk of eavesdropping and unauthorized access during administrative tasks.', 'h': 'Continue enforcing secure management sessions across all devices.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Secure Interactive Management Sessions are not available.', 'd': 'Management sessions use unencrypted protocols, exposing credentials and session data to attackers. Privileged access can be intercepted and misused.', 'f': 'If this information is disclosed to a malicious user, the device will become the target of an attack, compromised, and used to perform additional attacks. Anyone with privileged access to a device has the capability for full administrative control of that device.', 'h': 'Implement secure interactive management sessions (SSH, HTTPS) to protect device access and sensitive data.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        32: {  # consoleAuxPorts
            'compliance': {'a': 'Compliance', 'b': 'Console and AUX ports are properly configured.', 'd': 'Physical access is restricted and authentication is enforced, preventing unauthorized usage.', 'f': 'Reduces risk of attackers gaining direct access to routers and network devices.', 'h': 'Maintain proper configuration and access control on all console/AUX ports.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Console and AUX ports are not configured.', 'd': 'Ports are open and can be used for direct access to devices. Attackers may exploit this to bypass network authentication or perform spoofing attacks.', 'f': 'The attacker used to get dial-in access to the router. Spoofing attacks are possible in the bank network. An attacker can connect one end of the console cable to the router and the other end to the serial interface of their PC. Once that is done, an attacker can start your terminal emulation program.', 'h': 'Configure console and AUX ports with authentication and restricted access to prevent direct attacks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        33: {  # antiSpoofing
            'compliance': {'a': 'Compliance', 'b': 'Anti-Spoofing Protection is enabled.', 'd': 'Network validates IP addresses to prevent spoofed packets and unauthorized traffic.', 'f': 'Prevents IP spoofing attacks that can lead to fraud or network misuse.', 'h': 'Maintain Anti-Spoofing Protection on all network interfaces.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Anti-Spoofing Protection is disabled.', 'd': 'Network does not filter spoofed IP packets. Attackers can impersonate internal hosts or external entities to bypass security controls.', 'f': 'IP spoofed attacks from networks and direct administrative control are possible.  Spoofing is the cause of financial crimes related to criminal activities, so there is money laundering when there is a fraud.', 'h': 'Enable Anti-Spoofing Protection to validate IP addresses and protect against spoofing attacks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Apply formatting to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for col in range(1, 9):  # A to H
        cell = ws.cell(row=1, column=col)
        cell.border = thin_border
    
    # Process each question
    for i, question in enumerate(question_mapping, 2):  # Start from row 2
        # Get user input for this question
        user_input = form_data.get(question, 'not_applicable') if form_data else 'not_applicable'
        
        # Get response data
        response_data = question_responses.get(i-1, {})
        user_response = response_data.get(user_input, {})
        
        # Populate columns based on user input
        # Column C: Compliance/Non-Compliance/Not Applicable
        ws.cell(row=i, column=3, value=user_response.get('a', 'Not Applicable'))
        
        # Column D: Observation (Short/Brief)
        ws.cell(row=i, column=4, value=user_response.get('b', 'Not Applicable'))
        
        # Column F: Observation
        ws.cell(row=i, column=6, value=user_response.get('d', 'Not Applicable'))
        
        # Column G: Impact
        ws.cell(row=i, column=7, value=user_response.get('f', 'Not Applicable'))
        
        # Column H: Recommendation
        ws.cell(row=i, column=8, value=user_response.get('h', 'Not Applicable'))
        
        # Column E: Risk Factor with color coding
        risk_factor = risk_factors[i-2]  # risk_factors is 0-indexed
        risk_cell = ws.cell(row=i, column=5, value=risk_factor)
        risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
        risk_cell.fill = PatternFill(start_color=risk_colors.get(risk_factor, '808080'), end_color=risk_colors.get(risk_factor, '808080'), fill_type='solid')
        risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply formatting to all data rows
    for row in range(2, len(question_mapping) + 2):  # Rows 2 to 34
        for col in range(1, 9):  # Columns A to H
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Special alignment for column A (Sr. No.), C (Compliance status), and E (Risk Factor)
            if col == 1 or col == 3 or col == 5:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Set row height for wrapped text
            ws.row_dimensions[row].height = 30
    
    # Save file
    filename = "Router Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    """
    Delete the generated Excel file after download
    """
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            print(f"File {filepath} deleted successfully")
    except Exception as e:
        print(f"Error deleting file {filepath}: {e}")
