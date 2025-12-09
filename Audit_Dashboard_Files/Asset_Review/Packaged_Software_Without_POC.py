import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_packaged_software_excel(form_data=None):
    """
    Create Packaged Software Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Packaged Software"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Packaged Software Questions
    questions = [
        "Is the plan in line with the bank's overall IS strategy?",
        "Has the sponsor approved the requirement document?",
        "Has a document been prepared clearly detailing the requirements?",
        "Does the IT department have a technology standard for product selection?",
        "Does the bank have approved terms and conditions for product licensing agreements?",
        "Does the contract segregate the duties and responsibilities of the bank and the vendor?",
        "Has the IT department taken the required consequential actions for backup, disaster recovery, and performance tuning?",
        "Is there an identified system administrator responsible for managing access, backups, and ensuring database controls?",
        "Does the IT department use a scoring model for evaluating products and vendors?",
        "Do the scoring criteria consider all relevant aspects?",
        "Is there a system for measuring vendors' support against agreed service levels?",
        "Has the requirements document been clearly translated into product acceptance criteria?"
    ]

    # Risk Factors
    risk_factors = [
        "High",
        "High", 
        "High",
        "High",
        "High",
        "High",
        "High",
        "High",
        "Medium",
        "Medium",
        "Medium",
        "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "psPlanISStrategy": 1,
        "psSponsorApproval": 2,
        "psRequirementsDocumented": 3,
        "psTechnologyStandards": 4,
        "psLicensingTerms": 5,
        "psContractSegregation": 6,
        "psBackupDRPerformance": 7,
        "psSystemAdministrator": 8,
        "psScoringModel": 9,
        "psScoringCriteria": 10,
        "psVendorSupportMeasurement": 11,
        "psAcceptanceCriteria": 12
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Question responses data
    question_responses = {
        1: {  # psPlanISStrategy
            'compliance': {'a': 'Compliance', 'b': 'Plan aligned with IS strategy.', 'd': 'The plan has been reviewed to ensure alignment with the bank\'s IS strategy, supporting organizational goals and IT initiatives.', 'f': 'Promotes effective resource utilization and ensures IT initiatives support strategic objectives.', 'h': 'Periodically review projects for alignment with evolving IS strategy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "The plan was not in line with the bank's overall IS strategy.", 'd': 'The project or system plan does not align with the overall Information Systems strategy of the bank, leading to potential conflicts and inefficiencies.', 'f': "If the plan is not in line with the bank's overall IS strategy, it will be difficult for the bank to meet business objectives.Also the security posture of the bank will be a lot weaker if the IS strategy is not followed and bank will be prone to cyber attacks and security breaches often which will harm bank's reputation and bank may face financial loss.", 'h': 'Ensure all plans are reviewed and approved for alignment with the bank\'s IS strategy before initiation.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # psSponsorApproval
            'compliance': {'a': 'Compliance', 'b': 'Requirements document approved.', 'd': 'The requirement document has been reviewed and formally approved by the project sponsor, ensuring clarity and mutual understanding.', 'f': 'Reduces risk of scope changes, ensures business needs are met, and promotes accountability.', 'h': 'Maintain documented approval records for audit purposes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Sponsor did  not approve requirement document of packaged software.', 'd': 'The requirement document lacks formal approval from the designated project sponsor, leading to potential misunderstandings and scope creep.', 'f': "The sponsor's non-approval of the requirement document for the packaged software can lead to project delays, scope misalignment, uncertainty, and potential budget impacts. Clear communication and collaboration are essential to address concerns and achieve mutual agreement for a successful development process.", 'h': 'It is recommended to engage in clear communication and collaboration with the sponsor to address concerns and achieve mutual agreement on the requirement document. This will help avoid project delays, scope misalignment, uncertainty, and potential budget impacts, leading to a successful development process.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # psRequirementsDocumented
            'compliance': {'a': 'Compliance', 'b': 'Requirements documented clearly.', 'd': 'All functional and non-functional requirements are clearly documented, providing an unambiguous guide for development and testing.', 'f': 'Reduces errors, ensures alignment with business objectives, and facilitates validation and testing.', 'h': 'Periodically review requirements documentation for accuracy and completeness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Document was not prepared clearly detailing.', 'd': 'Requirement documents are incomplete, ambiguous, or inconsistent, leading to potential misinterpretation during development.', 'f': 'If the document is not prepared clearly detailing the features or requirements from the application, the application will be incomplete and miss the important features required by the bank.', 'h': 'Prepare detailed, clear, and unambiguous requirement documents reviewed by relevant stakeholders.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # psTechnologyStandards
            'compliance': {'a': 'Compliance', 'b': 'Technology standards in place.', 'd': 'The IT department follows predefined standards for selecting products, ensuring compatibility, quality, and alignment with bank policies.', 'f': 'Ensures consistent, reliable, and secure product selection, reducing operational risks.', 'h': 'Periodically update technology standards to incorporate new technologies and best practices.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The IT department does not have a technology standard for product selection.', 'd': 'The IT department lacks predefined standards for evaluating and selecting products, resulting in ad-hoc decisions.', 'f': "If the IT department does not have a technology standard for product selection, the delivered product will not be the standard one as there won't be a basic criteria to compare and determine the quality of the product. ", 'h': 'It is recommended that the IT Department should have a technology standard for product selection.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # psLicensingTerms
            'compliance': {'a': 'Compliance', 'b': 'Licensing agreements approved.', 'd': 'Product licensing agreements are formally reviewed and approved, ensuring compliance with legal and organizational policies.', 'f': 'Reduces legal and financial risks and ensures authorized use of software.', 'h': 'Maintain an archive of approved licensing agreements and review them periodically for renewals or compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Bank has not approved the terms and conditions for Product Licensing Agreements.', 'd': 'Licensing agreements for products are not formally approved, leading to potential legal and compliance risks.', 'f': 'If the bank has not approved the terms and conditions for Product Licensing Agreements, the responsibility and accountability of the service provider cannot be established.', 'h': 'Ensure all product licensing agreements are reviewed and approved by authorized personnel before execution.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # psContractSegregation
            'compliance': {'a': 'Compliance', 'b': 'Duties/responsibilities segregated.', 'd': 'Contracts clearly define the respective responsibilities of the bank and the vendor, reducing ambiguity and conflict.', 'f': 'Ensures smooth operations, accountability, and clarity in vendor engagements.', 'h': 'Review contracts periodically to ensure roles remain clear and updated.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The agreement does not explicitly separate the duties and responsibilities of the bank and the vendor.', 'd': 'It was observed that the agreement does not explicitly separate the duties and responsibilities of the bank and the vendor.', 'f': 'If the contract does not segregate the duties and responsibilities of the Bank and the Vendor, accountability cannot be established in case of the failure of asset/software or critical event.', 'h': 'Define and document duties and responsibilities clearly in contracts to avoid confusion and ensure accountability.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # psBackupDRPerformance
            'compliance': {'a': 'Compliance', 'b': 'Backup/DR/performance measures implemented.', 'd': 'Adequate backup, disaster recovery, and performance tuning procedures are in place to ensure system availability and reliability.', 'f': 'Enhances business continuity, protects critical data, and ensures optimal system performance.', 'h': 'Periodically test backup and DR processes and monitor system performance for continuous improvement.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The IT department has not taken the necessary consequential steps for backup, disaster recovery, and performance tuning. ', 'd': 'The IT department has not implemented necessary backup, disaster recovery, or performance tuning actions for critical systems.', 'f': 'In case of unavailability of application components, the bank cannot continue its critical operations as required consequential action for Back ups, Disaster Recovery, and Performance Tuning were not taken by the IT Department.', 'h': 'Implement a structured plan for backup, disaster recovery, and system performance tuning, and test regularly.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # psSystemAdministrator
            'compliance': {'a': 'Compliance', 'b': 'Responsible system administrator identified.', 'd': 'A designated system administrator is responsible for managing user access, performing backups, and ensuring database security controls are enforced.', 'f': 'Reduces operational and security risks, ensures accountability, and strengthens database management.', 'h': 'Review administrator responsibilities periodically to ensure compliance with policies and operational requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'There is no recognized system administrator responsible for managing system access, backing up, and ensuring database controls.', 'd': 'It was observed that there is no recognized system administrator responsible for managing system access, backing up, and ensuring database controls.', 'f': 'An unauthenticated user can make use of the privileges of the System Administrator and can access the backend database, and perform malicious activities like deletion, and modification of data.', 'h': 'It is recommended that there should be an identified System Administrator who is responsible for managing access to the system, backup, and ensuring database controls.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # psScoringModel
            'compliance': {'a': 'Compliance', 'b': 'Scoring model used.', 'd': 'The IT department utilizes a formal scoring model that evaluates products and vendors based on multiple objective criteria such as functionality, cost, security, and support.', 'f': 'Ensures objective, consistent, and justified decisions in product and vendor selection.', 'h': 'Periodically review and update the scoring model to incorporate evolving technology and regulatory requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The IT department does not use scoring models to evaluate products and vendors.', 'd': 'The IT department does not employ a formal scoring model to evaluate products or vendors, resulting in subjective selection decisions.', 'f': "If the IT department is not using a scoring model for evaluating the products and vendors, there won't be criteria to conclude and compare to determine the quality of the product and bank will not  be able to choose the best vendors or products which can enhance the bank's productivity.", 'h': 'Implement a structured scoring model for evaluating products and vendors based on defined criteria, ensuring objective and transparent selection.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # psScoringCriteria
            'compliance': {'a': 'Compliance', 'b': 'Scoring criteria comprehensive.', 'd': 'Scoring criteria consider all relevant factors, including performance, security, regulatory compliance, vendor support, and cost implications, ensuring thorough evaluation.', 'f': 'Promotes informed decision-making and reduces the risk of selecting unsuitable products or vendors.', 'h': 'Review and refine scoring criteria periodically to ensure continued relevance and effectiveness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Scoring criteria does not consider all aspects.', 'd': 'The scoring criteria used for product and vendor evaluation do not cover all relevant aspects, such as security, support, compliance, or total cost of ownership.', 'f': "If the scoring criteria are not made considering all the aspects, then the scoring criteria won't give a correct conclusion regarding a product.", 'h': 'Define comprehensive scoring criteria covering all critical aspects including security, functionality, support, compliance, and cost.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # psVendorSupportMeasurement
            'compliance': {'a': 'Compliance', 'b': 'Vendor support monitored.', 'd': 'The IT department tracks and measures vendor performance against agreed SLAs, ensuring prompt support and adherence to contractual commitments.', 'f': 'Improves accountability, reduces operational risks, and ensures service reliability.', 'h': 'Periodically review vendor performance reports and update SLAs as necessary.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'There is no system to measure vendor support with agreed service levels.', 'd': 'There is no system in place to monitor and measure vendor support performance against the agreed service level agreements (SLAs).', 'f': 'Unmonitored vendor performance may lead to delays in issue resolution, service disruptions, and unmet contractual obligations.', 'h': 'Implement a monitoring system to track vendor support performance against SLAs and take corrective action if service levels are not met.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # psAcceptanceCriteria
            'compliance': {'a': 'Compliance', 'b': 'Acceptance criteria defined.', 'd': 'The requirements document has been clearly translated into product acceptance criteria, providing a concrete framework for testing and approval.', 'f': 'Ensures that all products meet business and technical requirements, reducing operational and compliance risks.', 'h': 'Periodically review and update acceptance criteria to reflect changes in requirements or regulatory standards.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Requirements Document was not  translated clearly into product acceptance criteria.', 'd': 'The requirements document has not been clearly translated into product acceptance criteria, leading to ambiguity in testing and approval processes.', 'f': 'Unclear translation of requirements into product acceptance criteria can lead to misunderstandings, testing ambiguities, resource inefficiencies, and potential scope creep. Clear and measurable acceptance criteria are essential to mitigate these negative impacts and ensure a successful project outcome.', 'h': 'Define explicit product acceptance criteria based on the requirements document and ensure all products are tested against these criteria before approval.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }

    # Populate data based on user input
    for i, question in enumerate(questions, 2):
        # Get user input for this question
        question_num = i - 1
        user_input = None
        
        if form_data:
            # Find the corresponding form field
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Packaged Software Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
