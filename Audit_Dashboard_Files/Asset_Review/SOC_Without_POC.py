import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_soc_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "SOC"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # SOC Questions
    questions = [
        "Real Time Protection/Detection available or not?",
        "Real-time Event Correlation configured or not?",
        "Are all critical devices integrated into SOC or not?",
        "Is Admin lockout set to 3 or more failed login attempts or as per the organisation policy?",
        "Are default credentials in use or are accounts named 'Admin', 'root', or 'Administrator' to access administrator login?",
        "Is the software up to date?",
        "Is Syslog server configured or not?",
        "Is Event Log Analyzer configured or not?",
        "Is SSL used?",
        "Are Administrators having Two Factor Authentication enabled?",
        "Does the GUI only have HTTPS Access and CLI only have SSH Access?",
        "Whether Session timeout was defined?",
        "Is logging enabled?",
        "Are Password Policies configured as per the organization's password management policy?",
        "Is NTP configured?",
        "Is File monitoring configured or not?",
        "Whether Log Retention is followed as per agreed T&C & organization policy?",
        "Is an administrative portal login accessed only by Whitelisted IP Addresses?",
        "Is the SIEM tool configured with High Availability (HA) mode?",
        "How frequently is backup done for the SIEM Tool configuration?",
        "Are there multiple admin accounts to avoid lockout?",
        "Is the SIEM Tool integrated with Active Directory for user management?",
        "Are alerts configured or not?",
        "Are login timeouts set as per organization policies?"
    ]

    # Risk Factors
    risk_factors = [
        "Critical", "Critical", "High", "High", "High", "High", "High", "High", "Medium", "Medium",
        "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium",
        "Medium", "Medium", "Medium", "Low", "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "socRealTimeProtection": 1,
        "socEventCorrelation": 2,
        "socCriticalDevicesIntegrated": 3,
        "socAdminLockout": 4,
        "socDefaultCredentials": 5,
        "socSoftwareUpdated": 6,
        "socSyslogServer": 7,
        "socEventLogAnalyzer": 8,
        "socSslUsed": 9,
        "socTwoFactorAuth": 10,
        "socSecureAccess": 11,
        "socSessionTimeout": 12,
        "socLoggingEnabled": 13,
        "socPasswordPolicies": 14,
        "socNtpConfigured": 15,
        "socFileMonitoring": 16,
        "socLogRetention": 17,
        "socAdminPortalWhitelisted": 18,
        "socSiemHaMode": 19,
        "socSiemBackup": 20,
        "socMultipleAdmin": 21,
        "socAdIntegration": 22,
        "socAlertsConfigured": 23,
        "socLoginTimeouts": 24
    }

    # Question responses data - First 11 questions
    question_responses = {
        1: {  # socRealTimeProtection
            'compliance': {'a': 'Compliance', 'b': 'Real-time protection enabled.', 'd': 'Systems are actively monitored for threats, malware, and anomalies in real time, providing immediate alerts.', 'f': 'Enhances security posture by detecting and preventing attacks promptly.', 'h': 'Periodically test real-time protection mechanisms and update signatures or detection rules regularly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Real Time Protection Detection was not available.', 'd': 'Critical systems are not monitored in real time for threats or suspicious activities, leaving delays in threat detection.', 'f': 'Without real-time monitoring, the organization may experience delayed detection and response to security incidents, leaving the system vulnerable to potential cyber-attacks. The lack of immediate threat visibility increases the risk of data breaches and compromises sensitive information.', 'h': 'It is recommended to implement real-time protection detection in the SIEM without delay. Real-time monitoring is essential for immediate threat visibility and prompt detection of security incidents. Configure real-time alerting and automated responses to potential threats to reduce the risk of data breaches and minimize the impact of cyber-attacks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # socEventCorrelation
            'compliance': {'a': 'Compliance', 'b': 'Real-time event correlation implemented.', 'd': 'Security events from all critical devices are collected and correlated, enabling faster detection of complex attack patterns.', 'f': 'Improves incident detection, reduces response time, and enhances overall security monitoring.', 'h': 'Regularly review event correlation rules and adjust for new threats or system changes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Real-time Event Correlation was not configured.', 'd': 'Security events from multiple devices and logs are not correlated in real time, making it difficult to detect complex attacks.', 'f': "Without this capability, the SIEM system may struggle to effectively detect and respond to security incidents as they happen. The lack of real-time correlation hinders the organization's ability to quickly identify potential threats and mitigate them promptly. This delay in incident response increases the risk of data breaches, system disruptions, and potential damage to the organization's reputation.", 'h': "It is recommended to promptly configure real-time event correlation in the SIEM. Real-time correlation will enhance the system's ability to detect and respond to security incidents in real-time, reducing the risk of data breaches and minimizing the impact of cyber-attacks."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # socCriticalDevicesIntegrated
            'compliance': {'a': 'Compliance', 'b': 'All critical devices integrated into SOC.', 'd': 'Servers, network devices, and security appliances are sending logs to the SOC for real-time monitoring and analysis.', 'f': 'Enhances visibility, threat detection, and incident response capabilities.', 'h': 'Periodically audit the SOC integration to ensure all devices remain connected and reporting correctly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'All critical devices are not integrated into SOC.', 'd': 'It was observed that not all critical devices are integrated into the Security Operations Center (SOC). This lack of integration poses a significant security risk as it may result in a blind spot for the SOC team, making it challenging to monitor and respond effectively to potential security incidents on those devices. Integrating all critical devices into the SOC infrastructure is crucial to ensure comprehensive threat visibility and timely detection of any suspicious activities or threats.', 'f': "The impact of not integrating all critical devices into the Security Operations Center (SOC) can be severe. Without comprehensive visibility into these devices, the organization is left vulnerable to potential security threats and breaches. The SOC team may lack crucial insights into suspicious activities or security incidents on these devices, leading to delayed or ineffective responses. This, in turn, increases the risk of undetected attacks, data breaches, and potential damage to the organization's reputation.", 'h': 'It is recommended to  integrate all critical devices into the Security Operations Center (SOC). This integration will empower the SOC team to proactively monitor and respond to potential security incidents on all critical devices promptly.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # socAdminLockout
            'compliance': {'a': 'Compliance', 'b': 'Admin lockout configured.', 'd': 'Administrative accounts lock after the specified number of failed login attempts, as per organizational policy.', 'f': 'Reduces the risk of brute-force attacks and unauthorized access.', 'h': 'Periodically test lockout configurations and adjust policies if necessary.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'User ID does not get lockout after 3 failed consecutive log-on attempts.', 'd': 'It was observed that the User ID does not get locked out after three consecutive failed log-on attempts. This configuration poses a security risk as it leaves the system vulnerable to brute force attacks, where malicious actors can repeatedly try different combinations of passwords until they gain unauthorized access.', 'f': 'Without a lockout mechanism, the system is more susceptible to brute force attacks, where malicious actors can repeatedly attempt to guess passwords until they gain unauthorized access. This increases the risk of successful unauthorized logins, potentially leading to data breaches, privacy violations, and unauthorized manipulation of critical information.', 'h': "It is recommended to implement a lockout policy after three consecutive failed log-on attempts to enhance the system's security. By doing so, the organization can mitigate the risk of brute force attacks and deter potential unauthorized access. The lockout policy should be configured to temporarily block User IDs after a specific number of unsuccessful login attempts, thereby preventing malicious actors from repeatedly trying different passwords."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # socDefaultCredentials
            'compliance': {'a': 'Compliance', 'b': 'Default accounts disabled/renamed.', 'd': 'Administrative accounts follow naming conventions and use strong, unique passwords. Default accounts are removed or disabled.', 'f': 'Enhances system security and reduces exposure to targeted attacks.', 'h': 'Periodically review account configurations and ensure compliance with naming and password policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Default credentials were used.', 'd': 'It was observed that default credentials were used. This poses a significant security risk as default credentials are widely known and easily exploitable by attackers. Using default credentials leaves the system vulnerable to unauthorized access and potential cyber-attacks.', 'f': 'When not changed, default credentials make an organization more vulnerable to potential cyberattacks. Attackers can obtain these standard login details, allowing them access to the devices on your network – usually with admin rights – and leaving them open to takeover. Malicious actors can gain control over the system, compromise sensitive data, and disrupt operations.', 'h': "It is recommended to immediately change all default credentials to enhance the system's security. Changing default passwords to strong, unique, and complex ones will significantly reduce the risk of unauthorized access and potential cyber-attacks."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # socSoftwareUpdated
            'compliance': {'a': 'Compliance', 'b': 'Software up to date.', 'd': 'All critical software and applications are updated with the latest patches and versions.', 'f': 'Reduces the risk of exploitation and enhances system stability and security.', 'h': 'Continue regular patching and monitor vendor advisories for timely updates.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The software was not up to date.', 'd': 'It was observed that the SIEM software is running on outdated versions, which may lack important security patches and enhancements, increasing vulnerability to known vulnerabilities.', 'f': 'Outdated software lacks important security patches and updates, making the SIEM system more susceptible to known vulnerabilities and potential exploits.', 'h': 'It is strongly recommended to upgrade SIEM software.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # socSyslogServer
            'compliance': {'a': 'Compliance', 'b': 'Syslog server configured.', 'd': 'All critical devices and systems forward logs to a centralized Syslog server for monitoring and analysis.', 'f': 'Ensures effective tracking of system events, audit readiness, and faster incident response.', 'h': 'Regularly review Syslog configurations and retention policies to maintain integrity and availability of logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Syslog server was not configured.', 'd': 'It was observed that a Syslog server was not configured in the SIEM (Security Information and Event Management) system. This lack of configuration poses a security risk as it hinders the ability to collect and centralize log data from various sources.', 'f': 'Without a Syslog server, the SIEM lacks centralized log management, making it difficult to effectively monitor and analyse security events from various sources. This hinders the ability to detect and respond to security incidents promptly, potentially leaving the organization vulnerable to prolonged cyber-attacks and data breaches.', 'h': 'It is recommended to promptly configure a Syslog server in the SIEM. A Syslog server will enable centralized log management, allowing the SIEM to collect and analyse security events from diverse sources. Properly configuring the Syslog server will enhance threat detection, incident response, and overall cybersecurity capabilities.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # socEventLogAnalyzer
            'compliance': {'a': 'Compliance', 'b': 'Event Log Analyzer configured.', 'd': 'All relevant security and system events are captured and analyzed to detect anomalies and incidents in real time.', 'f': 'Improves security monitoring, faster incident detection, and reduces manual oversight.', 'h': 'Periodically review analyzer rules and thresholds to ensure relevance to emerging threats.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Event Log Analyzer was not configured.', 'd': "It was observed that the Event Log Analyzer was not configured. This lack of configuration poses a security risk as it hinders the organization's ability to monitor and analyse event logs effectively.", 'f': 'Without this tool, the organization lacks the capability to monitor and analyse event logs efficiently, leading to delayed detection of security incidents and potential threats. The absence of real-time monitoring and analysis hinders incident response, increasing the risk of data breaches and unauthorized access.', 'h': "It is recommended the configure and integrate the tool within the SOC infrastructure to enable comprehensive log collection and analysis. Configure the Event Log Analyzer to enhance the organization's cybersecurity capabilities. Implement real-time monitoring and analysis of event logs to promptly detect and respond to security incidents. Define and customize alerting rules based on the organization's specific security needs and compliance requirements."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # socSslUsed
            'compliance': {'a': 'Compliance', 'b': 'SSL/TLS implemented.', 'd': 'All communication channels for applications and systems are encrypted using SSL/TLS protocols.', 'f': 'Ensures confidentiality, integrity, and security of data in transit.', 'h': 'Regularly review SSL/TLS configurations and certificates to prevent expiration or weak cipher usage.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'SSL was not used in the SIEM Application.', 'd': 'Sensitive communication between clients and servers occurs over unencrypted channels, exposing data to interception.', 'f': 'Without encryption, sensitive data becomes vulnerable to interception, compromising its security. Unauthorized access risks increase as SSL\'s authentication mechanism is missing, enabling attackers to impersonate the server or intercept communication. Data integrity is compromised, as tampering or injection of malicious content becomes possible.', 'h': 'Implement SSL/TLS for all client-server communications and enforce HTTPS on web interfaces.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # socTwoFactorAuth
            'compliance': {'a': 'Compliance', 'b': 'Two-Factor Authentication enabled for admins.', 'd': 'All administrative users authenticate with passwords plus a second factor (e.g., OTP, token) for enhanced security.', 'f': 'Significantly reduces the risk of unauthorized access and strengthens account security.', 'h': 'Regularly audit 2FA enforcement and update policies for new admin accounts or system integrations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Two factor authentication was not enabled for the administration access.', 'd': 'Administrative accounts rely solely on passwords without a secondary authentication factor.', 'f': 'The impact of not enabling Two-factor authentication for administration access could be significant. Without this added layer of security, the system becomes more vulnerable to potential security breaches and unauthorized access. Hackers or malicious actors could exploit this weakness to gain control over critical administrative functions, compromise sensitive data, or disrupt the system\'s operations. This could lead to severe consequences, such as data breaches, financial losses, reputational damage, and legal implications.', 'h': 'Enforce 2FA for all administrative accounts using hardware tokens, OTPs, or authenticator apps.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # socSecureAccess
            'compliance': {'a': 'Compliance', 'b': 'Secure access protocols enforced.', 'd': 'GUI access is restricted to HTTPS, and CLI access is restricted to SSH, ensuring encrypted communication.', 'f': 'Protects credentials and sensitive data from interception.', 'h': 'Periodically validate protocol configurations and disable any accidental fallback to insecure protocols.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'GUI does not have HTTPS Access and CLI does not have SSH Access.', 'd': 'The GUI allows unencrypted HTTP access and the CLI allows Telnet, exposing credentials and data in transit.', 'f': 'An attacker can collect unencrypted passwords from the connection, and all the data is transferred in plain text so that anyone can intercept the request and gain confidential information on the connection, data, and traffic that are going through the firewall, so it is easy for an attacker to gain sensitive information by intercepting traffic.', 'h': 'Restrict GUI access to HTTPS only and CLI access to SSH exclusively, disabling all insecure protocols like HTTP and Telnet.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # socSessionTimeout
            'compliance': {'a': 'Compliance', 'b': 'Session timeout configured.', 'd': 'All user sessions terminate automatically after a defined period of inactivity.', 'f': 'Reduces unauthorized access risks from unattended sessions.', 'h': 'Review session timeout settings periodically and adjust according to security standards.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Session timeout was not defined.', 'd': 'Users\' sessions remain active indefinitely or for long periods without activity.', 'f': 'The purpose of session timeout is to prevent unauthorized access to a user\'s account if they forget to log out or leave their computer unattended. If session timeout is not defined, it can have a significant impact on the security of the system, Unauthorized access, Unauthorized access and Increased risk of fraud.', 'h': 'Implement session timeout according to organizational policy, automatically logging out inactive users.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # socLoggingEnabled
            'compliance': {'a': 'Compliance', 'b': 'Logging enabled.', 'd': 'All relevant system, network, and user activities are logged and stored for audit and monitoring purposes.', 'f': 'Provides traceability and accountability, supporting incident response and compliance requirements.', 'h': 'Regularly audit logging configurations and verify log integrity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Logging was not enabled.', 'd': 'Critical system and user activity are not being recorded, preventing effective monitoring and audit.', 'f': 'Without proper logging, the SIEM lacks the necessary data to effectively detect and respond to security incidents in real-time. This can lead to delayed or missed identification of critical threats, leaving the organization exposed to prolonged cyberattacks and data breaches. Moreover, the absence of detailed logs hinders the ability to conduct thorough incident investigations, making it challenging to understand the scope and nature of security breaches.', 'h': 'Enable logging for all critical systems and ensure logs are centralized, tamper-proof, and regularly reviewed.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # socPasswordPolicies
            'compliance': {'a': 'Compliance', 'b': 'Password policies enforced.', 'd': 'All user accounts comply with organizational password management policies, enforcing minimum complexity, length, and periodic changes.', 'f': 'Reduces risk of unauthorized access due to weak or compromised passwords.', 'h': 'Periodically review and update password policies to align with evolving security standards.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Password was not configured as per the organizations password management policy.', 'd': 'Weak passwords are allowed, and settings like minimum length, complexity, and expiration are not applied.', 'f': 'A weak password policy increases the probability of an attacker having success using brute force and dictionary attacks against user accounts. An attacker who can determine user passwords can take over a user\'s account and potentially access sensitive data in the application.', 'h': 'Configure and enforce password policies consistent with organizational standards, including complexity, length, and expiration rules.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # socNtpConfigured
            'compliance': {'a': 'Compliance', 'b': 'NTP configured.', 'd': 'All servers and network devices synchronize time using NTP, ensuring consistent timestamps across systems.', 'f': 'Supports accurate logging, auditing, and troubleshooting of system events.', 'h': 'Periodically verify NTP synchronization and maintain redundancy for time sources.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'NTP (Network Time Protocol) was not configured.', 'd': 'System clocks are unsynchronized across servers and devices, leading to inconsistent timestamps for logs and transactions.', 'f': 'Without proper time synchronization, system logs and events may have inaccurate timestamps, affecting incident investigation and event correlation. Inconsistent time across network devices can also lead to authentication failures and disrupt critical system operations.', 'h': 'Configure NTP on all critical devices to synchronize time with a reliable and secure time source.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # socFileMonitoring
            'compliance': {'a': 'Compliance', 'b': 'File monitoring enabled.', 'd': 'Critical system and configuration files are continuously monitored, and any unauthorized changes are logged and reported.', 'f': 'Enhances security by detecting tampering and supporting incident investigations.', 'h': 'Periodically review monitoring rules and verify that alerts are correctly generated.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'File monitoring was not configured in SIEM.', 'd': 'Changes to critical system or configuration files are not tracked or logged.', 'f': 'The impact of not configuring file monitoring in the SIEM can be significant. Without file monitoring, the organization lacks visibility into file-level activities, making it challenging to detect unauthorized access, data breaches, and insider threats. This increases the risk of critical data being compromised or leaked without detection.', 'h': 'Enable file integrity monitoring for all critical files and directories, and ensure alerts are triggered for unauthorized changes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # socLogRetention
            'compliance': {'a': 'Compliance', 'b': 'Log retention is maintained as per the organization policy and agreed terms & conditions.', 'd': 'The organization follows its defined log retention policy, ensuring all security and system logs are securely stored for the mandated duration in line with regulatory and internal requirements.', 'f': 'Facilitates effective forensic analysis, compliance audits, and incident investigations. Enhances accountability and traceability of security events.', 'h': 'Continue to review log retention practices periodically to ensure alignment with evolving regulatory and organizational requirements. Validate secure storage and timely archival of logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Log Retention is not followed as per agreed T&C & organization policy.', 'd': 'It was observed that the log retention process is not fully aligned with the organization\'s policy and agreed T&C. While certain critical systems maintain logs for the prescribed duration, other systems were found retaining logs for shorter periods or without documented retention schedules.', 'f': 'Failure to retain logs as per the organization\'s policy and agreed T&C can lead to loss of critical audit trails required for investigating security incidents, operational issues, or compliance reviews. Inadequate log retention may impair forensic analysis, weaken accountability, and result in non-compliance with regulatory and statutory requirements such as RBI and CERT-In guidelines.', 'h': 'The bank should ensure that log retention is implemented consistently across all systems in accordance with organizational policy and regulatory requirements.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        18: {  # socAdminPortalWhitelisted
            'compliance': {'a': 'Compliance', 'b': 'Admin portal restricted to whitelisted IPs.', 'd': 'Only designated IP addresses are allowed to access the administrative portal, preventing unauthorized remote access.', 'f': 'Significantly reduces the attack surface and protects critical system configurations.', 'h': 'Review and update the whitelist periodically to ensure only authorized IPs have access.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'It was observed that IP addresses were not whitelisted for accessing the administrative portal.', 'd': 'No restrictions exist on IP addresses for admin portal login, allowing access from untrusted locations.', 'f': 'Any bank employee who does not have higher privileges can access the administrative portal and make changes in rules and policies from any computer. Without IP whitelisting, the administrative portal is vulnerable to unauthorized access from any location, increasing the risk of unauthorized changes and potential data breaches. Malicious actors could exploit this weakness to gain control over critical administrative functions, compromising sensitive information and system integrity.', 'h': 'Restrict administrative portal access to approved whitelisted IP addresses only.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        19: {  # socSiemHaMode
            'compliance': {'a': 'Compliance', 'b': 'SIEM tool was configured with High Availability (HA) mode', 'd': 'It was observed that the bank\'s Security Information and Event Management (SIEM) solution is configured in High Availability (HA) mode.', 'f': 'Ensures uninterrupted security event collection, correlation, and alerting. Reduces downtime risk and improves reliability of incident detection and response.', 'h': 'Continue maintaining the HA configuration with regular failover testing and monitoring to ensure seamless functionality and resilience during system outages.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'SIEM tool was not configured with High Availability (HA) mode.', 'd': 'It was observed that the bank\'s Security Information and Event Management (SIEM) solution is not configured in High Availability (HA) mode. In the event of a system failure or network disruption, there is a risk that log collection and correlation activities may be interrupted, potentially leading to gaps in monitoring and incident detection.', 'f': 'Lack of HA configuration in the SIEM environment can result in loss of critical security logs, delayed detection of security incidents, and reduced visibility into ongoing cyber threats. This may hinder timely response to security breaches and affect compliance with regulatory monitoring requirements.', 'h': 'The bank should configure the SIEM solution in High Availability (HA) mode to ensure uninterrupted collection, correlation, and analysis of security events.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        20: {  # socSiemBackup
            'compliance': {'a': 'Compliance', 'b': 'Regular SIEM configuration backups performed.', 'd': 'SIEM configurations are backed up as per defined schedule, ensuring recovery in case of system failure.', 'f': 'Preserves monitoring integrity and enables quick restoration after incidents.', 'h': 'Periodically test the backup and restoration process to ensure reliability.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'No backup was taken for the SIEM Tool configuration.', 'd': 'The SIEM tool\'s configuration is not backed up consistently, risking loss of critical monitoring settings.', 'f': 'In the event of system failures or cyber-attacks, critical configuration settings may be lost, leading to prolonged downtime and reduced system functionality. Without backups, it becomes challenging to restore the SIEM to a previous working state, hindering incident investigation and recovery efforts.', 'h': 'Implement a regular backup schedule for SIEM configurations, with offsite storage for redundancy.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        21: {  # socMultipleAdmin
            'compliance': {'a': 'Compliance', 'b': 'Multiple admin accounts configured.', 'd': 'Two or more administrative accounts exist to prevent lockout and ensure uninterrupted system management.', 'f': 'Reduces operational risk and ensures continuity of critical system administration.', 'h': 'Periodically review admin accounts for proper usage and access levels.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'It was found that only one admin account was created.', 'd': 'Only one administrative account exists, risking complete lockout if credentials are lost or compromised.', 'f': 'If only one admin account is there, then if the admin account gets locked out, it will affect the bank\'s productivity and day to day operations. Thus, the bank may face financial losses due to delay of day to day operations. It results in a single point of failure, increasing the risk of unauthorized access and potentially hindering critical administrative tasks. The absence of multiple accounts with varying privileges poses challenges for accountability and auditing.', 'h': 'Maintain at least two administrative accounts to ensure redundancy and operational continuity.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        22: {  # socAdIntegration
            'compliance': {'a': 'Compliance', 'b': 'SIEM integrated with AD.', 'd': 'User authentication and authorization are centrally managed via Active Directory, ensuring consistent access control.', 'f': 'Enhances security by applying uniform user policies and simplifying account management.', 'h': 'Periodically audit user access and AD integration configurations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'SIEM Tool was not integrated with Active Directory for user management.', 'd': 'User management is handled manually or locally, increasing risk of inconsistent access controls.', 'f': 'Without SIEM integration, user provisioning and deprovisioning processes may be manual and prone to errors, leading to potential unauthorized access or delays in removing access for former employees. Inconsistent user management can also hinder proper audit trails and accountability, making it challenging to track user activities accurately.', 'h': 'Integrate the SIEM tool with Active Directory to centralize authentication and enforce consistent user access policies.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        23: {  # socAlertsConfigured
            'compliance': {'a': 'Compliance', 'b': 'Alerts configured.', 'd': 'Critical events, anomalies, and threshold breaches trigger alerts in the SIEM tool, enabling timely response.', 'f': 'Supports proactive incident detection and mitigation.', 'h': 'Review alert configurations periodically and tune thresholds to reduce false positives.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Alerts are not configured in SOC.', 'd': "It was observed that alerts are not configured in the Security Operations Center (SOC). This absence of alerting mechanisms poses a significant risk as it hinders the SOC team's ability to promptly detect and respond to potential security incidents.", 'f': "Without alerts, the SOC team may miss critical security events and incidents in real-time, leading to delayed or inadequate responses. This can result in prolonged exposure to cyber threats, increasing the risk of data breaches, unauthorized access, and potential damage to the organization's assets and reputation. The lack of timely alerts can also hinder incident investigation and make it challenging to identify the root cause of security issues, hindering the organization's ability to prevent similar incidents in the future.", 'h': 'It is recommended to implement a robust alerting system in the Security without delay. Configuring alerts will provide the SOC team with real-time notifications of potential security incidents, enabling them to respond promptly and effectively. Alerts should be set up for critical events, unusual activities, and suspicious behaviour to ensure comprehensive threat detection.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        24: {  # socLoginTimeouts
            'compliance': {'a': 'Compliance', 'b': 'Login timeouts configured as per policy.', 'd': 'User sessions terminate automatically after defined inactivity periods, in line with organizational standards.', 'f': 'Reduces risk of unauthorized access due to unattended sessions.', 'h': 'Periodically review session timeout settings and adjust based on security requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Login/Session timeout was not set as per organization policies.', 'd': "It was observed that for some user accounts, passwords did not comply with the organization's password management policy. Passwords were found to lack complexity, length, and regular expiration. Additionally, instances of password reuse were identified, posing potential security risks.", 'f': 'Weak and easily guessable passwords increase the risk of unauthorized access to user accounts and critical systems. Password reuse compounds the vulnerability, making it easier for attackers to compromise multiple accounts. Such practices leave the organization exposed to potential data breaches, financial losses, and damage to its reputation.', 'h': 'It is recommended to set login/session time out as per the organization policy. It is recommended to set login timeout  less than 10 minutes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "SOC Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
