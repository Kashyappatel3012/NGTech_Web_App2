import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_internal_control_evaluation_excel(form_data=None):
    """
    Create Internal Control Evaluation Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Internal Control Evaluation"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Internal Control Evaluation Questions
    questions = [
        "Are the internal controls in place to provide reliance on the control procedures for the information/evidence gathered during the audit?",
        "Is internal audit conducted by the bank?"
    ]

    # Risk Factors
    risk_factors = [
        "High",
        "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "iceAuditEvidenceReliability": 1,
        "iceInternalAuditConducted": 2
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Question responses data
    question_responses = {
        1: {  # iceAuditEvidenceReliability
            'compliance': {'a': 'Compliance', 'b': 'Robust internal controls in place for audit evidence.', 'd': 'The bank has implemented strong internal controls that ensure the reliability and integrity of information used in audits. This includes segregation of duties, system-generated reports with validations, periodic reconciliations, and supervisory review of critical data. These measures ensure that audit evidence can be trusted and accurately reflects operational and financial realities.', 'f': 'Enables auditors to rely on the evidence for assessing compliance, risk management, and operational efficiency. Reduces the likelihood of undetected errors or misstatements and supports regulatory and internal governance requirements.', 'h': 'Regularly assess and update internal controls to address process changes, new technologies, and evolving regulatory requirements. Conduct training and awareness sessions to ensure staff consistently follow control procedures.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'There were no internal controls for the purpose of placing reliance on the control procedures in support of the information/evidence/ information being gathered as part of the audit.', 'd': 'The bank has weak or incomplete internal controls over the processes that generate, maintain, and validate information used during audits. For example, transaction logs, system-generated reports, or reconciliations may not be properly controlled, reviewed, or verified before being submitted for audit purposes.', 'f': 'Audit evidence may be unreliable, leading to incorrect conclusions about the bank\'s operations, IT systems, or financial health. This may result in undetected errors, potential fraud, regulatory non-compliance, or misstatements in financial and operational reporting.', 'h': 'Establish comprehensive internal controls covering data collection, reporting, and validation processes. This includes approvals, reconciliations, access restrictions, and regular monitoring to ensure the accuracy, completeness, and integrity of audit-related information. Implement periodic reviews of these controls to adapt to process changes or emerging risks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # iceInternalAuditConducted
            'compliance': {'a': 'Compliance', 'b': 'Internal audit conducted regularly.', 'd': 'The bank has a dedicated internal audit function that performs systematic reviews of all critical areas, including IT infrastructure, operational processes, financial reporting, and compliance with regulations. Internal audit findings are documented, reported to management, and tracked until resolution, ensuring that identified risks are mitigated promptly.', 'f': 'Strengthens governance, enhances risk management, and ensures compliance with policies, procedures, and regulatory requirements. Provides management and regulators with assurance on operational efficiency and control effectiveness.', 'h': 'Maintain continuous internal audit programs, periodically review the audit methodology, and ensure auditors are trained in emerging risks and regulatory updates. Implement dashboards or KPIs to track remediation progress effectively.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'No process of Internal Audit is implemented.', 'd': 'The bank either does not conduct internal audits or performs them irregularly without adequate coverage of critical areas such as IT systems, operations, regulatory compliance, and financial transactions.', 'f': 'Key operational and compliance risks remain unassessed, which could lead to undetected fraudulent activities, mismanagement, or regulatory violations. Lack of internal audit also weakens governance and management oversight, potentially exposing the bank to financial loss, operational disruption, or reputational damage.', 'h': 'Establish a structured internal audit function with defined scope, frequency, and coverage for all critical areas. Ensure audits are performed by qualified personnel, with findings documented and communicated to management. Implement follow-up mechanisms to remediate identified issues and improve operational and compliance controls.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }

    # Populate data based on user input
    for i, question in enumerate(questions, 2):
        # Get user input for this question
        question_num = i - 1
        user_input = None
        
        if form_data:
            # Find the corresponding form field
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Internal Control Evaluation Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
