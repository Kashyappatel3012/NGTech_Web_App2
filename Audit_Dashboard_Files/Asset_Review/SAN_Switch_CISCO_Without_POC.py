import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_san_switch_cisco_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "SAN Switch CISCO"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # SAN Switch CISCO Questions
    questions = [
        "Is telnet service enabled?",
        "Does the GUI only have HTTPS access and CLI only have SSH access?",
        "Is USB access allowed on the switch?",
        "Are CDP (Cisco Discovery Protocol) packets enabled?",
        "Are default credentials in use, or are accounts named 'Admin', 'root', or 'Administrator' to access the administrator login?",
        "Does the core switch have a secure password enabled?",
        "Does Cisco enable TCP and UDP Small Servers by default?",
        "Is ICMP redirects enabled on the network interface mgmt0?",
        "Is Proxy ARP enabled on mgmt0?",
        "Are unused ports closed on the switch?",
        "Is the switch physically secured?",
        "Is NTP configured?",
        "Is an administrative portal login accessed only by whitelisted IP addresses?",
        "Is admin lockout set to 3 or more failed login attempts or as per the organization's policy?",
        "Is logging enabled and is the Syslog server implemented?",
        "Are password policies configured as per the organization's password management policy?",
        "How frequently is backup done for the switch configuration?",
        "Is Maintenance Operations Protocol (MOP) disabled on all ethernet interfaces?",
        "Are directed broadcasts enabled by default?",
        "Is inbound TCP connection keep-alives message disabled?",
        "Is BOOTP disabled?",
        "Is the login disclaimer message set?",
        "Is login timeout configured as per organization policies?",
        "Are ICMP IP unreachable messages enabled on the interface mgmt0?",
        "Are domain lookups enabled and DNS servers configured?",
        "Is the Packet Assembler/Disassembler (PAD) service disabled?"
    ]

    # Risk Factors
    risk_factors = [
        "High", "High", "High", "High", "High", "High", "High", "Medium", "Medium", "Medium",
        "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Low",
        "Low", "Low", "Low", "Low", "Low", "Low", "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "sscTelnetService": 1,
        "sscGuiHttpsCliSsh": 2,
        "sscUsbAccess": 3,
        "sscCdpPackets": 4,
        "sscDefaultCredentials": 5,
        "sscSecurePassword": 6,
        "sscTcpUdpSmallServers": 7,
        "sscIcmpRedirects": 8,
        "sscProxyArp": 9,
        "sscUnusedPortsClosed": 10,
        "sscPhysicallySecured": 11,
        "sscNtpConfigured": 12,
        "sscAdminPortalWhitelisted": 13,
        "sscAdminLockout": 14,
        "sscLoggingSyslog": 15,
        "sscPasswordPolicies": 16,
        "sscBackupFrequency": 17,
        "sscMopDisabled": 18,
        "sscDirectedBroadcasts": 19,
        "sscTcpKeepAlives": 20,
        "sscBootpDisabled": 21,
        "sscLoginDisclaimer": 22,
        "sscLoginTimeout": 23,
        "sscIcmpUnreachable": 24,
        "sscDomainLookups": 25,
        "sscPadService": 26
    }

    # Question responses data
    question_responses = {
        1: {  # sscTelnetService
            'compliance': {'a': 'Compliance', 'b': 'Telnet service disabled.', 'd': 'Telnet service is disabled on all devices, and remote management is restricted to SSH or other secure protocols.', 'f': 'Prevents credential interception and enhances secure remote management of network devices.', 'h': 'Periodically audit device configurations to confirm Telnet remains disabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Telnet service was enabled.', 'd': 'It was observed that Telnet service is running on some network devices, allowing remote management using unencrypted credentials over the network.', 'f': 'An attacker who was able to monitor network traffic could capture sensitive information or  authentication credentials. Network packet and password sniffing tools are widely available on the Internet and some of the tools are specifically designed to capture clear-text protocol authentication credentials. However, in a switched environment an attacker may not be able to capture network traffic destined for other devices without employing an attack such as Address Resolution Protocol (ARP) spoofing.', 'h': 'Disable Telnet service immediately and enforce SSH for secure remote management. Review all devices to ensure Telnet is not running.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # sscGuiHttpsCliSsh
            'compliance': {'a': 'Compliance', 'b': 'Secure access protocols enforced.', 'd': 'All network devices are configured to allow GUI access only via HTTPS and CLI access only via SSH. Unencrypted access methods such as HTTP or Telnet are disabled.', 'f': 'Enhances confidentiality and integrity of administrative access.', 'h': 'Maintain secure protocol enforcement and periodically verify configurations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'GUI does not have HTTPS Access and CLI does not have SSH Access.', 'd': 'Some network devices allow HTTP access to the GUI or CLI access without SSH, exposing credentials and management traffic to the network in unencrypted form.', 'f': 'An attacker can collect unencrypted passwords from the connection, and all the data is transferred in plain text so that anyone can intercept the request and gain confidential information on the connection, data, and traffic that are going through the firewall, so it is easy for an attacker to gain sensitive information by intercepting traffic.', 'h': 'Configure all devices to use HTTPS for GUI access and SSH for CLI access. Disable all unencrypted protocols.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # sscUsbAccess
            'compliance': {'a': 'Compliance', 'b': 'USB access disabled on switches.', 'd': 'All switches have USB ports disabled to prevent unauthorized access and mitigate malware risks.', 'f': 'Reduces physical attack surfaces and enhances network security.', 'h': 'Periodically verify USB access settings and maintain device hardening records.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'USB access was enabled.', 'd': 'USB ports on network switches are active, allowing potential unauthorized file transfers or malware injection.', 'f': "Data leakage and any kind of threat are injected into the network through Switch's USB port. There is a proverb that says that A hacker's best friend is that little USB stick you plug into your system. The USB drive can inject malware into the system and the whole business network would be interrupted.", 'h': 'Disable USB access on all switches unless specifically required and document exceptions with approval.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # sscCdpPackets
            'compliance': {'a': 'Compliance', 'b': 'CDP properly controlled.', 'd': 'CDP is disabled on all external-facing interfaces and enabled only on internal trusted networks where necessary.', 'f': 'Protects network topology information from unauthorized disclosure.', 'h': 'Periodically review CDP settings to ensure they align with the security policy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'CDP was enabled on all interfaces.', 'd': 'CDP packets are active on switches and routers, exposing network topology information to anyone with network access.', 'f': '''CDP packets contain information about the sender, such as hardware model information, 
operating system version and IP address details. This information would allow an attacker to gain 
information about the configuration of the network infrastructure.
CDP packets are broadcast to an entire network segment. An attacker could use one of the many 
publicly available tools to capture network traffic and view the leaked information.''', 'h': 'Disable CDP on interfaces facing untrusted networks and limit its use to internal trusted networks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # sscDefaultCredentials
            'compliance': {'a': 'Compliance', 'b': 'Default credentials replaced with secure accounts.', 'd': 'All devices have unique administrator accounts with strong passwords, and default accounts have been disabled or renamed.', 'f': 'Reduces the likelihood of unauthorized access and enhances device security posture.', 'h': 'Continue periodic password audits and maintain an updated account inventory.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Default credentials like username "admin" or "root" were used for administrative login.', 'd': 'Some network devices still use default credentials or accounts named \'Admin\', \'root\', or \'Administrator\', making them easily guessable by attackers.', 'f': 'When not changed, default credentials make an organization more vulnerable to potential cyberattacks. Attackers can obtain these standard login details, allowing them access to the devices on your network usually with admin rights, and leaving them open to takeover.', 'h': 'Replace all default credentials with unique, strong passwords and rename administrator accounts where possible. Implement periodic password rotation and access reviews.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # sscSecurePassword
            'compliance': {'a': 'Compliance', 'b': 'Secure password configured on core switch.', 'd': 'The core switch has a strong, unique password configured in line with the organization\'s security policies, and password change procedures are followed.', 'f': 'Protects the switch from unauthorized access and mitigates network compromise risks.', 'h': 'Continue periodic password reviews and audits to maintain security.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The switch was using a weak password.', 'd': 'It was observed that the core switch is using weak or default passwords, which could be easily guessed or brute-forced by attackers.', 'f': 'As the switch is using a weak password, the attacker can brute force the password and gain the administrator-level privileges.', 'h': 'Set a strong, unique password on the core switch following the bank\'s password policy, and enforce periodic password changes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # sscTcpUdpSmallServers
            'compliance': {'a': 'Compliance', 'b': 'TCP and UDP small servers disabled.', 'd': 'All Cisco devices have TCP and UDP small servers disabled by default, minimizing unnecessary services.', 'f': 'Reduces potential attack vectors and strengthens network device security.', 'h': 'Periodically review device configurations to ensure no legacy services are re-enabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Cisco enables TCP and UDP Small Servers by default.', 'd': 'The default configuration shows TCP and UDP small servers enabled on Cisco devices, which could allow attackers to exploit legacy services.', 'f': 'Each running service increases the chances of an attacker being able to identify the device and successfully compromise it. It is good security practice to disable all unused services. These services, especially their User Datagram Protocol (UDP) versions, are infrequently used for legitimate purposes. They could be used to launch denial of service attacks.', 'h': 'Disable TCP and UDP small servers unless explicitly required for business purposes, and document any exceptions.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # sscIcmpRedirects
            'compliance': {'a': 'Compliance', 'b': 'ICMP redirects disabled on mgmt0.', 'd': 'ICMP redirects are disabled on the management interface mgmt0, ensuring secure and controlled routing.', 'f': 'Mitigates risks associated with traffic redirection and network spoofing attacks.', 'h': 'Periodically verify ICMP settings across all critical interfaces.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'ICMP Redirects were enabled.', 'd': 'ICMP redirect messages are currently enabled on the management interface mgmt0, which could allow attackers to manipulate routing tables.', 'f': 'It might be possible for an attacker to make the switch device send many ICMP to redirect messages, which will result in an elevated CPU load. An attacker could use ICMP redirect messages to network traffic through their own switch, possibly allowing them to monitor network traffic.', 'h': 'Disable ICMP redirects on all management interfaces to prevent unauthorized routing changes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # sscProxyArp
            'compliance': {'a': 'Compliance', 'b': 'Proxy ARP disabled on mgmt0.', 'd': 'Proxy ARP is disabled on mgmt0, reducing the risk of ARP spoofing and unauthorized network access.', 'f': 'Strengthens network security by limiting unauthorized traffic interception opportunities.', 'h': 'Review interface configurations periodically to maintain Proxy ARP disabled.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Proxy ARP was not disabled.', 'd': 'Proxy ARP is enabled on the management interface mgmt0, which can allow attackers to intercept traffic or spoof hosts on the network.', 'f': '''If Proxy ARP is enabled, then it has the following disadvantages:
1. It increases the amount of ARP traffic in your segment.
2. Hosts need larger ARP tables to handle IP-to-MAC address mappings.
3. Security can be undermined. A machine can claim to be another in order to intercept packets, an act called "spoofing."
4. It does not work for networks that do not use ARP for address resolution.
5. It does not generalize to all network topologies. For example, more than one router connects two physical networks.''', 'h': 'Disable Proxy ARP on all management interfaces unless specifically required and documented.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # sscUnusedPortsClosed
            'compliance': {'a': 'Compliance', 'b': 'Unused switch ports closed.', 'd': 'All unused switch ports are administratively shut down and secured with port security to prevent unauthorized access.', 'f': 'Reduces potential entry points for attackers and enhances overall network security.', 'h': 'Conduct periodic port audits to ensure unused ports remain closed and secure.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Unused ports were open on the switch.', 'd': 'Some unused ports on switches are left active, which could allow unauthorized devices to connect and gain network access.', 'f': 'An attacker can intrude on the network by unauthorizedly accessing the switch with the help of an unused port and gathering sensitive information to commit bank frauds.', 'h': 'Disable all unused ports on switches and implement port security measures such as MAC address binding and logging.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # sscPhysicallySecured
            'compliance': {'a': 'Compliance', 'b': 'Switch physically secured.', 'd': 'The switch is housed in a locked, restricted access room or cabinet, limiting physical access to authorized personnel only.', 'f': 'Protects critical network devices from unauthorized physical manipulation or theft.', 'h': 'Periodically verify access logs and the physical security of network infrastructure.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The switch was not secured physically.', 'd': 'The switch is installed in an unsecured area without restricted access, making it vulnerable to tampering or unauthorized physical access.', 'f': 'An attacker can intrude on the network by changing the LAN ports of the switch or sniffing the network. Also, it is vulnerable to side-channel attacks.', 'h': 'Install switches in locked, access-controlled rooms or cabinets, and maintain a log of personnel accessing these areas.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # sscNtpConfigured
            'compliance': {'a': 'Compliance', 'b': 'NTP configured on devices.', 'd': 'All switches and critical network devices are synchronized with a secure, reliable NTP server, ensuring consistent timestamps across logs.', 'f': 'Enables accurate event correlation, troubleshooting, and auditing.', 'h': 'Regularly monitor NTP synchronization and update NTP server configurations as needed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'NTP was not configured.', 'd': 'Switches and network devices are not synchronized with a reliable NTP server, leading to inconsistent timestamps across logs and events.', 'f': 'A more insidious effect of weak timekeeping is that it damages the ability to investigate security breaches and other kinds of system problems. Hackers, for example, will often exploit backdoor, and proxy computers when mounting and attacking- both to hide their tracks and to exploit whatever opportunities (like NTP System privileges ) the hacker encounters along the way. Finding these stopping-off points is critical for shutting the door to future attacks and requires precise measurements of time in order to reconstruct the exact sequence of events. log file and application time stamp obviously become essential pieces of evidence.', 'h': 'Configure NTP on all network devices using a trusted NTP server and periodically verify synchronization.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # sscAdminPortalWhitelisted
            'compliance': {'a': 'Compliance', 'b': 'Admin portal restricted to whitelisted IPs.', 'd': 'Only IP addresses approved by the organization can access the administrative portal, preventing unauthorized access from untrusted sources.', 'f': 'Reduces risk of external attacks and ensures administrative control remains secure.', 'h': 'Periodically review and update the whitelist based on business requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Administrative portal login could be accessed by any of the organization\'s computers.', 'd': 'The admin portal can be accessed from any external IP, increasing the risk of unauthorized login attempts and brute-force attacks.', 'f': 'Any bank employee who does not have higher privileges can access the administrative portal and make changes in rules and policies from any computer.', 'h': 'Restrict access to the administrative portal using IP whitelisting, VPN, or dedicated management networks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # sscAdminLockout
            'compliance': {'a': 'Compliance', 'b': 'Admin lockout configured.', 'd': 'Administrative accounts are locked automatically after a defined number of failed login attempts, aligning with organizational security policy.', 'f': 'Mitigates brute-force attacks and strengthens account security.', 'h': 'Regularly test and review lockout settings to ensure they function correctly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'There was no lockout functionality available.', 'd': 'Administrative accounts are not locked after multiple failed login attempts, leaving the system vulnerable to brute-force attacks.', 'f': 'A brute-force attack can be performed which consists of an attacker submitting many passwords or passphrases with the hope of eventually guessing correctly. The attacker systematically checks all possible passwords and passphrases until the correct one is found.', 'h': 'Configure account lockout after a defined number of failed attempts as per organizational policy, and reset after proper verification.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # sscLoggingSyslog
            'compliance': {'a': 'Compliance', 'b': 'Logging enabled and Syslog server implemented.', 'd': 'All critical network devices generate logs that are sent to a centralized Syslog server, allowing real-time monitoring, analysis, and retention for audit purposes.', 'f': 'Enhances visibility into network activity, enables quick detection of anomalies, and supports incident investigations.', 'h': 'Periodically review Syslog configurations and monitor alerts for any unusual activity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Logging was not enabled and the Syslog server was not implemented.', 'd': 'Network devices are not sending logs to a centralized Syslog server, or logging is disabled, preventing proper monitoring of network activity.', 'f': 'If any malicious activity takes place at the network level and logs are needed for forensic investigation. Then it will be difficult to get logs for investigation, thus affecting the quality of investigation. Also, periodic review of the logs will not be possible if the Syslog server is not present.', 'h': 'Enable logging on all network devices and forward logs to a secure, centralized Syslog server with proper retention and monitoring.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # sscPasswordPolicies
            'compliance': {'a': 'Compliance', 'b': 'Password policies configured correctly.', 'd': 'All network devices enforce password complexity, expiration, and history policies in line with the organization\'s password management standards.', 'f': 'Enhances protection against unauthorized access and brute-force attacks.', 'h': 'Periodically audit device password policies to ensure continued compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Password policies were not configured as per the organization\'s password management policy.', 'd': 'The switch or network device does not enforce password complexity, expiration, or reuse restrictions as per the organization\'s password management policy.', 'f': "A weak password policy increases the probability of an attacker having success using brute force and dictionary attacks against user accounts. An attacker who can determine user passwords can take over a user's account and potentially access sensitive data in the application.", 'h': 'Configure password policies on all devices to enforce complexity, minimum length, expiration, and history rules according to organizational standards.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # sscBackupFrequency
            'compliance': {'a': 'Compliance', 'b': 'Regular switch configuration backups performed.', 'd': 'All critical switch configurations are backed up periodically, with secure storage maintained for quick restoration if required.', 'f': 'Ensures rapid recovery from failures or misconfigurations, minimizing downtime and operational impact.', 'h': 'Test backup restoration periodically and maintain versioned backups.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'No backup was taken for the switch configuration.', 'd': 'Network device configurations are not backed up on a scheduled basis, leaving the system vulnerable to configuration loss in case of device failure.', 'f': 'All the data traffic passes through the Switch. If its firmware crashes and the bank does not have the backup. The bank will lose important data, as well as interrupt the business process.', 'h': 'Schedule automatic, regular backups of all critical switch configurations and store them securely.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        18: {  # sscMopDisabled
            'compliance': {'a': 'Compliance', 'b': 'MOP disabled on all interfaces.', 'd': 'Maintenance Operations Protocol is disabled on all ethernet interfaces, reducing exposure to network-level attacks.', 'f': 'Protects network devices from unauthorized monitoring and reduces attack surface.', 'h': 'Review configurations periodically to ensure MOP remains disabled on non-essential interfaces.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'MOP was not disabled on the Ethernet interfaces.', 'd': 'Maintenance Operations Protocol (MOP) is active on ethernet interfaces, exposing the network device to potential unauthorized monitoring or network disruption.', 'f': 'MOP enables personnel on the local network, or a remote network that is bridged to the local network, to obtain access to a remote console on the router if they possess credentials for the device. This is significant because access to router management is usually protected by IP-based ACLs. As a Layer 2 protocol, MOP allows for the circumvention of this type of ACL, making brute force login attempts possible if account lockout is not enabled. If account lockout is enabled, such attempts could result in a denial of service due to user accounts being locked out. Running unused services increases the chances of an attacker finding a security hole or fingerprinting a device.', 'h': 'Disable MOP on all interfaces unless explicitly required for operational purposes, and document exceptions.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        19: {  # sscDirectedBroadcasts
            'compliance': {'a': 'Compliance', 'b': 'Directed broadcasts disabled.', 'd': 'Directed broadcasts are disabled on all network interfaces, mitigating the risk of amplification-based attacks.', 'f': 'Reduces vulnerability to DoS attacks and ensures network stability.', 'h': 'Periodically verify that directed broadcasts remain disabled and audit device configurations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Directed broadcasts were enabled by default.', 'd': 'Network devices have directed broadcast enabled by default, which could allow amplification attacks or network flooding.', 'f': "A Denial of Service (DoS) attack exists that makes use of network echo requests, known as a smurf attack. An attacker would send an ICMP echo request with the victim's host's IP address spoofed as the source. The hosts on the network would then reply to the echo request, flooding the victim host.", 'h': 'Disable directed broadcasts on all interfaces unless explicitly required for legitimate operational purposes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        20: {  # sscTcpKeepAlives
            'compliance': {'a': 'Compliance', 'b': 'Inbound TCP keep-alives disabled.', 'd': 'TCP keep-alive messages are disabled for inbound connections, reducing unnecessary exposure and idle connections.', 'f': 'Strengthens network security by limiting unnecessary network communication and potential reconnaissance.', 'h': 'Review device TCP settings periodically to maintain secure configurations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Inbound TCP connection keep-alive message was not disabled.', 'd': 'Inbound TCP connection keep-alive messages are active, potentially exposing devices to unnecessary connections or reconnaissance attacks.', 'f': "An attacker could attempt a DoS attack against a device by exhausting the number of possible connections. To perform this attack, the attacker could keep requesting new connections to the device and spoof the source IP addresses. This would then prevent any new legitimate connections to the device from being made as the device awaits the completion of the connection attempts that have already been initiated. This attack would prevent both users and administrators from connecting to the device", 'h': 'Disable inbound TCP keep-alive messages unless specifically required for operational purposes, and monitor connections closely.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        21: {  # sscBootpDisabled
            'compliance': {'a': 'Compliance', 'b': 'BOOTP disabled.', 'd': 'BOOTP is disabled on all devices, reducing exposure to unauthorized network boot attacks.', 'f': 'Ensures secure device booting and minimizes attack surface.', 'h': 'Periodically verify BOOTP status and document exceptions if needed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'BOOTP was not disabled.', 'd': 'BOOTP (Bootstrap Protocol) is active on network devices, which could be exploited by attackers to gain unauthorized access or manipulate network boot processes.', 'f': "BOOTP uses a relay agent, which allows packet forwarding from the local network using standard IP routing, allowing one BOOTP server to serve hosts on multiple subnets. An attacker could use the BOOTP service to download a copy of the router's IOS software.", 'h': 'Disable BOOTP on all devices unless explicitly required for operational purposes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        22: {  # sscLoginDisclaimer
            'compliance': {'a': 'Compliance', 'b': 'Login disclaimer configured.', 'd': 'A standard login disclaimer message is displayed on all device logins, notifying users of authorized access requirements.', 'f': 'Enhances security awareness and provides legal notice to potential unauthorized users.', 'h': 'Review the message periodically to ensure compliance with legal and organizational standards.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The login disclaimer message was not set.', 'd': 'The network device does not display a login disclaimer message warning users of authorized access only.', 'f': 'If the login disclaimer message is not set it will not deter the user from trying to file a lawsuit against you for something covered by your disclaimer. Attackers who have gained access to a device could avoid legal action if no banner is configured to warn against unauthorized access.', 'h': 'Configure a standard login disclaimer on all devices per organization policy.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        23: {  # sscLoginTimeout
            'compliance': {'a': 'Compliance', 'b': 'Login timeout configured.', 'd': 'All network devices enforce login timeout settings as per organizational standards, automatically disconnecting inactive sessions.', 'f': 'Reduces the risk of session hijacking and unauthorized access.', 'h': 'Periodically test and validate timeout configurations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'No time limit was set for idle sessions.', 'd': 'Devices do not automatically terminate inactive sessions according to organizational security policies.', 'f': 'An attacker can utilize this time to perform malicious activities and cause harm to the network infrastructure when an authorized user leaves his system without logging out.', 'h': 'Configure login timeout settings consistent with the organization\'s policy to disconnect inactive sessions automatically.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        24: {  # sscIcmpUnreachable
            'compliance': {'a': 'Compliance', 'b': 'ICMP unreachable messages disabled.', 'd': 'ICMP IP unreachable messages are disabled on mgmt0, reducing the risk of network reconnaissance and unauthorized mapping.', 'f': 'Minimizes exposure to network scanning attacks and enhances device security.', 'h': 'Periodically verify ICMP settings to maintain secure configurations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'ICMP IP unreachable messages were not disabled on the interface.', 'd': 'ICMP IP unreachable messages are active on the management interface, which could be used by attackers for network mapping and reconnaissance.', 'f': 'If the ICMP IP unreachable messages are not disabled then an attacker can use the ICMP type 3 Flood DOS attack. Also, the attacker can gather more information about the network if ICMP IP unreachable messages are not disabled.', 'h': 'Disable ICMP unreachable messages on management interfaces unless operationally required, and monitor network activity.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        25: {  # sscDomainLookups
            'compliance': {'a': 'Compliance', 'b': 'Domain lookups and DNS configured correctly.', 'd': 'All devices have domain lookups enabled and DNS servers configured per organizational standards.', 'f': 'Ensures reliable name resolution and reduces the risk of misrouting or DNS-based attacks.', 'h': 'Periodically validate DNS configurations and maintain approved DNS server lists.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Domain lookup was enabled and DNS server was configured.', 'd': 'Devices either have domain lookups disabled or do not have authorized DNS servers configured, which may affect proper name resolution and lead to misrouted traffic.', 'f': 'An attacker who was able to capture network traffic could monitor DNS queries from the Cisco Switch. Furthermore, Cisco devices can connect to Telnet servers by supplying only the hostname or IP address of the server. A mistyped Cisco command could be interpreted as an attempt to connect to a Telnet server and broadcast on the network', 'h': 'Configure authorized DNS servers and ensure domain lookups are correctly enabled to prevent misrouting and potential attacks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        26: {  # sscPadService
            'compliance': {'a': 'Compliance', 'b': 'PAD service disabled.', 'd': 'PAD service is disabled on all network devices, minimizing exposure to unnecessary or legacy protocol vulnerabilities.', 'f': 'Reduces the attack surface and enhances device security.', 'h': 'Review service configurations periodically to ensure no unauthorized services are active.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Packet Assembler/Disassembler (PAD) service was not disabled.', 'd': 'The Packet Assembler/Disassembler (PAD) service is active, potentially exposing devices to legacy X.25 network attacks or unnecessary services.', 'f': 'Running unused services increases the chances of an attacker finding a security hole or fingerprinting a device. PAD acts like a multiplexer for the terminals. If enabled, it can render the device open to attacks.', 'h': 'Disable PAD services on all devices unless required for operational needs, and document any exceptions.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "SAN Switch CISCO Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
