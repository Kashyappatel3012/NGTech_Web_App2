import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_esxi_server_excel(form_data=None):
    """
    Create ESXi Server Logical Review Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "ESXi Server Logical Review"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ESXi Server Logical Review Questions
    questions = [
        "Is there a separate room for the server?",
        "Is Anti-Virus installed on server?",
        "Are physical controls established for Anti-Virus Server?",
        "Is Internet allowed on servers?",
        "Is SSH authorized_keys file empty?",
        "Is Image Profile VIB acceptance level configured properly?",
        "Are unauthorized kernel modules loaded on the host?",
        "Is ESXi host firewall configured to restrict access to services running on the host?",
        "Does a non-root user account exist for local admin access?",
        "Does the server have adequate space for operational requirements?",
        "Is the server room visible from a distance, but not easily accessible?",
        "Is the server room away from the basement, water/drainage systems?",
        "Are default self-signed certificates for ESXi communication not used?",
        "Are expired and revoked SSL certificates removed from the ESXi server?",
        "Is a centralized location configured to collect ESXi host core dumps?"
    ]

    # Risk Factors (provided by user)
    risk_factors = [
        "Critical", "Critical", "Critical", "Critical", "High", "High", "High", "High", "High", 
        "Medium", "Medium", "Medium", "Medium", "Medium", "Medium"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "esxiSeparateServerRoom": 1,
        "esxiAntivirusInstalled": 2,
        "esxiPhysicalControlsAntivirus": 3,
        "esxiInternetAllowedServers": 4,
        "esxiSshAuthorizedKeysEmpty": 5,
        "esxiImageProfileVibAcceptance": 6,
        "esxiUnauthorizedKernelModules": 7,
        "esxiHostFirewallConfigured": 8,
        "esxiNonRootUserAccount": 9,
        "esxiAdequateServerSpace": 10,
        "esxiServerRoomVisible": 11,
        "esxiServerRoomAwayFromWater": 12,
        "esxiDefaultSelfSignedCertificates": 13,
        "esxiExpiredRevokedCertificates": 14,
        "esxiCentralizedCoreDumps": 15
    }

    # Question responses data
    question_responses = {
        1: {  # esxiSeparateServerRoom
            'compliance': {'a': 'Compliance', 'b': 'Dedicated server room available.', 'd': 'A dedicated server room is available.', 'f': 'Ensures physical protection, environmental control, and restricted access to critical server infrastructure.', 'h': 'Periodically review physical access controls and environmental conditions to maintain compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'No separate server room available.', 'd': 'No separate room is available for the server, and it is placed in a common area.', 'f': '''1. Inadequate Temperature Control:-
When the temperature around and within the server and networking equipment becomes too high the server will shut down and there will be loss of data.
2. Imbalanced Moisture Levels:-
High humidity can result in rust, corrosion, short-circuiting, and even the growth of fungus that can attack the machinery. Too little moisture in the air is also a concern, as an exceedingly dry environment can result in electrostatic discharge, which can cause system malfunction and damage.
Also, there is a risk of dust and temperature interference. ''', 'h': 'It is recommended to have a separate room for the server. When you address the specific server facility threats that most often destroy data assets, you can minimize risk dramatically.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # esxiAntivirusInstalled
            'compliance': {'a': 'Compliance', 'b': 'Anti-virus installed and operational.', 'd': 'Anti-virus is installed on the server.', 'f': 'Protects the server from malware and other threats, reducing the risk of data breaches and system compromise.', 'h': 'Regularly verify anti-virus updates and scan schedules to ensure continuous protection.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Antivirus was not installed on the server.', 'd': 'Anti-virus software is not installed on the server.', 'f': 'As the Antivirus is not installed, it is not possible to protect against various malicious activities like malware, virus, etc. Servers are considered a critical asset of the bank network so the unavailability of these systems is a major threat to the organization.', 'h': 'Install enterprise-grade anti-virus software on all servers and ensure it is centrally managed with automatic updates.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # esxiPhysicalControlsAntivirus
            'compliance': {'a': 'Compliance', 'b': 'Physical controls implemented.', 'd': 'Physical controls are implemented for the Anti-Virus server.', 'f': 'Prevents unauthorized access and modification of the server, ensuring the anti-virus system functions effectively.', 'h': 'Periodically audit physical access to maintain security.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Physical controls were not established for anti-virus server.', 'd': 'Physical controls for the Anti-Virus server are not established, and the server is accessible to multiple personnel.', 'f': 'The lack of physical controls for an antivirus server can have serious security implications. Without proper physical controls, unauthorized individuals may be able to gain access to the server, potentially compromising sensitive information and allowing malware to spread throughout the network. ', 'h': 'Use a dedicated Anti-Virus server with no other services running, and secure it physically so only authorized personnel can access it.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # esxiInternetAllowedServers
            'compliance': {'a': 'Compliance', 'b': 'Internet access restricted on servers.', 'd': 'Servers are isolated from the internet.', 'f': 'Minimizes exposure to external threats and strengthens security of critical banking systems.', 'h': 'Regularly monitor server network settings to ensure isolation remains effective.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Internet was allowed on servers without any restriction.', 'd': 'The Internet is allowed on servers where critical banking applications are hosted.', 'f': "Full access to the internet on the server will sometimes create a critical problem if some malicious script is downloaded on the server from the internet that will remove or encrypt all the sensitive data on the server and can directly gain access to CBS. As the Internet is not restricted, any malicious activity could be performed through the internet. Also, some social media websites can be accessed by the employee which will affect the bank's productivity, and using those sites unintentionally users can click on a malicious link that can download viruses, worms, or any malware that will affect all bank networks.", 'h': 'Restrict internet access on servers hosting banking applications and separate servers from public networks as per RBI guidelines.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # esxiSshAuthorizedKeysEmpty
            'compliance': {'a': 'Compliance', 'b': 'SSH authorized_keys file empty or approved only.', 'd': 'The authorized_keys file is empty or contains only authorized keys.', 'f': 'Prevents unauthorized SSH access, maintaining server security.', 'h': 'Periodically review and audit SSH keys to ensure only authorized access is allowed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The SSH authorized_keys file was not empty.', 'd': '''It was observed that the SSH authorized_keys file was not empty. ESXi hosts come with Secure Shell (SSH), which can be configured to authenticate remote users using public key authentication. For day-to-day operations, the ESXi host should be in lockdown mode with the SSH service disabled. Lockdown mode does not prevent root users from logging in using keys. The presence of a remote user's public key in the /etc/ssh/keys-root/authorized_keys file on an ESXi host identifies the user as trusted, meaning the user is granted access to the host without providing a password.

Disabling authorized_keys access may limit your ability to run unattended remote scripts.
Keeping the authorized_keys file empty prevents users from circumventing the intended restrictions of lockdown mode.''', 'f': 'If the file contains unauthorized or malicious public keys, it can grant unauthorized access to the system. This poses a significant security risk as it allows potential attackers to gain entry into the system using compromised credentials.', 'h': '''It is recommended to regularly review and ensure the SSH authorized_keys file is empty or contains only authorized public keys. To remove all keys from the authorized_keys file, perform the following:
    1. Logon to the ESXi shell as root or another admin user.
    2. Edit the /etc/ssh/keys-root/authorized_keys file.
    3. Remove all keys from the file and save the file.'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # esxiImageProfileVibAcceptance
            'compliance': {'a': 'Compliance', 'b': 'VIB acceptance level properly configured.', 'd': 'The ESXi host\'s Image Profile VIB acceptance level is set correctly.', 'f': 'Ensures only trusted and verified VIBs are installed, reducing risk of host compromise.', 'h': 'Periodically review VIB acceptance levels and updates to maintain security.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Image Profile VIB acceptance level was not configured properly.', 'd': '''It was observed that Image Profile VIB acceptance level was not configured properly. A VIB (vSphere Installation Bundle) is a collection of files that are packaged into an archive. The VIB contains
a signature file that is used to verify the level of trust. The ESXi Image Profile supports four VIB acceptance levels:
1. VMware Certified - VIBs created, tested, and signed by VMware
2. VMware Accepted - VIBs created by a VMware partner but tested and signed by VMware
3. Partner Supported - VIBs created, tested, and signed by a certified VMware partner
4. Community Supported - VIBs that have not been tested by VMware or a VMware partner
The ESXi Image Profile should only allow signed VIBs because an unsigned VIB represents untested code installed on an ESXi host. Also, use of unsigned VIBs will cause hypervisor Secure Boot to fail to configure. Community Supported VIBs do not have digital signatures.''', 'f': 'Improper configuration may lead to the installation of unsigned or unverified VIBs, increasing the risk of compatibility issues, system instability, and potential security vulnerabilities. This can compromise the integrity of the virtualization environment and expose it to unauthorized or malicious software', 'h': '''To protect the security and integrity of your ESXi hosts, do not allow unsigned (CommunitySupported) VIBs to be installed on your hosts. It is recommended to properly configure the Image Profile VIB acceptance level to maintain a secure and reliable virtual infrastructure. To implement the recommended configuration state, run the following PowerCLI command (in the example code, the level is Partner Supported):

# Set the Software AcceptanceLevel for each host<span>
Foreach ($VMHost in Get-VMHost ) {
 $ESXCli = Get-EsxCli -VMHost $VMHost
 $ESXCli.software.acceptance.Set('PartnerSupported')
}'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # esxiUnauthorizedKernelModules
            'compliance': {'a': 'Compliance', 'b': 'Only authorized kernel modules loaded.', 'd': 'The host contains only approved kernel modules.', 'f': 'Reduces risk of malicious code execution and ensures host stability.', 'h': 'Continuously monitor kernel modules for compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Unauthorized kernel modules were  loaded on the host.', 'd': 'It was observed that unauthorized kernel modules are loaded on the host. ESXi hosts by default do not permit the loading of kernel modules that lack valid digital signatures. This feature can be overridden, which would allow unauthorized kernel modules to be loaded.', 'f': 'VMware provides digital signatures for kernel modules. Untested or malicious kernel modules loaded on the ESXi host can put the host at risk for instability and/or exploitation.  These unauthorized modules can introduce vulnerabilities, compromise system stability, and lead to potential system crashes or malfunctions. Additionally, unauthorized kernel modules may bypass security controls and allow malicious actors to gain unauthorized access to the system, potentially leading to data breaches and unauthorized system manipulations.', 'h': '''Secure the host by disabling unsigned modules and removing the offending VIBs from the host. It is recommended to run the following PowerCLI
command-# To disable a module-
$ESXCli = Get-EsxCli -VMHost MyHost
$ESXCli.system.module.set($false, $false, 'MyModuleName')'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # esxiHostFirewallConfigured
            'compliance': {'a': 'Compliance', 'b': 'ESXi firewall configured properly.', 'd': 'Firewall restricts access to only required services on the ESXi host.', 'f': 'Reduces exposure to attacks and unauthorized access attempts.', 'h': 'Audit firewall rules periodically to maintain security.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The ESXi host firewall was not configured to restrict access to services running on the host.', 'd': 'It was observed that the the ESXi host firewall was not configured to restrict access to services running on the host. The ESXi firewall is enabled by default and allows ping (ICMP) and communication with DHCP/DNS clients. Access to services should only be allowed by authorized IP addresses/networks.', 'f': 'Unrestricted access to services running on an ESXi host can expose a host to outside attacks and unauthorized access. Reduce the risk by configuring the ESXi firewall to only allow access from authorized IP addresses and networks.', 'h': '''
It is recommended to promptly configure the ESXi host firewall to restrict access to services running on the host. To properly restrict access to services running on an ESXi host, perform the following from the vSphere web client:
1. Select the host.
2. Go to 'Configure' -> 'System' -> 'Security Profile'.
3. In the 'Firewall' section, select 'Edit...'.
4. For each enabled service, (e.g., ssh, vSphere Web Access, http client) provide the
range of allowed IP addresses.
5. Click 'OK'.'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # esxiNonRootUserAccount
            'compliance': {'a': 'Compliance', 'b': 'Non-root admin account exists.', 'd': 'A non-root user account is available for administrative tasks.', 'f': 'Reduces risk associated with root access and enhances accountability.', 'h': 'Periodically review and restrict root access.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Non-root user account was not exists for local admin access.', 'd': "It was observed that the non-root user account was not exists for local admin access. By default, each ESXi host has a single 'root' admin account that is used for local administration and to connect the host to vCenter Server. Use of this shared account should be limited, and named (non-root) user accounts with admin privileges should be used instead.", 'f': 'The impact of not having a non-root user account for local admin access can be significant. Without a non-root account, all administrative tasks are performed using the root account, which poses security risks. Using the root account for routine tasks increases the likelihood of accidental system modifications or unauthorized changes.', 'h': '''To avoid sharing a common root account, it is recommended on each host to create at least one named user account and assign it full admin privileges, and to use this account in lieu of a shared 'root' account. Limit the use of 'root', including setting a highly complex password for the account, but do not remove the 'root' account.
o create one or more named user accounts (local ESXi user accounts), perform the following using the vSphere client (not the vSphere web client) for each ESXi host:

1.Connect directly to the ESXi host using the vSphere Client.
2.Login as root.
3.Select Manage, then select the Security & Users tab.
4.Select User and view the local users.
5.Add a local user and grant shell access to this user.
6.Select the Host, then select 'Actions' and 'Permissions'.
7.Assign the 'Administrator' role to the user.

Notes:
Even if you add your ESXi host to an Active Directory domain, it is still recommended to add at least one local user account to ensure admins can still login in the event the host ever becomes isolated and unable to access Active Directory.

Adding local user accounts can be automated using Host Profiles.'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # esxiAdequateServerSpace
            'compliance': {'a': 'Compliance', 'b': 'Adequate server space available.', 'd': 'Server has sufficient operational space.', 'f': 'Ensures proper airflow, cooling, and maintenance access, reducing hardware and operational risks.', 'h': 'Regularly review server placement and space utilization.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The server does not have adequate space for operational requirements.', 'd': 'The server does not have sufficient physical or logical space for operational needs.', 'f': 'As the Server has no adequate space for operational requirements there is a risk of dust and temperature interference. Server hardware and related components require specific components to perform optimally, such as adequate cooling, moisture removal, and protection from excessive temperatures. Server rooms that are too hot or cold could cause hardware to malfunction leading to downtime.', 'h': 'Ensure adequate space for server installation, maintenance, and operation. Plan for future expansion and proper ventilation.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # esxiServerRoomVisible
            'compliance': {'a': 'Compliance', 'b': 'The server room is visible from a distance but not easily accessible to unauthorized personnel.', 'd': 'Server room is visible from a distance and only accessible to authorized personnel.', 'f': 'Provides physical security while allowing monitoring of the room.', 'h': 'Continuously enforce access restrictions and monitoring.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The server room was not visible from a distance and was easily accesible.', 'd': 'The server room is not visible from a distance and can be accessed by unauthorized personnel.', 'f': 'Increases risk of physical tampering, theft, and unauthorized access, compromising the confidentiality and integrity of critical systems.', 'h': 'Make server room visible for monitoring but restrict access to authorized personnel only. Implement locks and monitoring mechanisms.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # esxiServerRoomAwayFromWater
            'compliance': {'a': 'Compliance', 'b': 'Server room safely located.', 'd': 'The server room is away from basement and water/drainage systems.', 'f': 'Minimizes risk of environmental damage to servers and data.', 'h': 'Periodically inspect for any water ingress risks.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Server room near basement/water systems.', 'd': 'The server room is located near the basement or water/drainage systems.', 'f': 'Risk of water damage, flooding, and electrical hazards, potentially leading to server failure and data loss.', 'h': 'Relocate or protect the server room from water and drainage hazards. Implement water detection and mitigation systems.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # esxiDefaultSelfSignedCertificates
            'compliance': {'a': 'Compliance', 'b': 'Trusted certificates used.', 'd': 'ESXi host uses CA-signed certificates.', 'f': 'Ensures secure communication and prevents MITM attacks.', 'h': 'Periodically verify and renew certificates before expiration.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The default self-signed certificate for ESXi communication was used.', 'd': 'It was observed that the default self-signed certificate for ESXi communication is used. The default certificates are not signed by a trusted certificate authority (CA) and should be replaced with valid certificates that have been issued by a trusted CA.', 'f': 'Using the default self-signed certificate may increase risk related to man-in-the-middle (MITM) attacks.', 'h': '''Backup and replace the details of the SSL certificate presented by the ESXi host and determine if it is issued by a trusted CA:
Log in to the ESXi Shell, either directly from the DCUI or from an SSH client, as a user with administrator privileges.
In the directory /etc/vmware/ssl, rename the existing certificates using the following commands:

mv rui.crt orig.rui.crt
mv rui.key orig.rui.key

Copy the certificates you want to use to /etc/vmware/ssl.
Rename the new certificate and key to rui.crt and rui.key.
Restart the host after you install the new certificate.

Alternatively, you can put the host into maintenance mode, install the new certificate, use the Direct Console User Interface (DCUI) to restart the management agents, and set the host to exit maintenance mode.'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # esxiExpiredRevokedCertificates
            'compliance': {'a': 'Compliance', 'b': 'Expired/revoked certificates removed.', 'd': 'Only valid SSL certificates are present on the ESXi host.', 'f': 'Ensures secure communication and prevents trust issues.', 'h': 'Periodically review SSL certificate validity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Expired and revoked SSL certificates was not removed from the ESXi server.', 'd': 'It was observed that expired and revoked SSL certificates was not removed from the ESXi server. The ESXi hosts do not have Certificate Revocation List (CRL) checking available, so expired and revoked SSL certificates must be checked and removed manually.', 'f': 'Leaving expired and revoked certificates on your vCenter Server system can compromise your environment.', 'h': '''It is recommended to replace expired and revoked certificates with certificates from a trusted CA. Certificates can be replaced in a number of ways:
Replace a Default ESXi Certificate and Key from the ESXi Shell

1. Log in to the ESXi Shell, either directly from the DCUI or from an SSH client, as a user with administrator privileges.
2. In the directory /etc/vmware/ssl, rename the existing certificates using the following commands:

mv rui.crt orig.rui.crt
mv rui.key orig.rui.key

3. Copy the certificates that you want to use to /etc/vmware/ssl.
4. Rename the new certificate and key to rui.crt and rui.key.
5. Restart the host after you install the new certificate.

Alternatively, you can put the host into maintenance mode, install the new certificate, use the Direct Console User Interface (DCUI) to restart the management agents, and set the
host to exit maintenance mode.

Replace a Default ESI Certificate and Key by Using the vifs Command

1. Back up the existing certificates.
2. Generate a certificate request following the instructions from the certificate
authority.
3. At the command line, use the vifs command to upload the certificate to the
appropriate location on the host.

vifs --server hostname --username username --put rui.crt /host/ssl_cert
vifs --server hostname --username username --put rui.key /host/ssl_key

4. Restart the host.

Alternatively, you can put the host into maintenance mode, install the new certificate, and then use the Direct Console User Interface (DCUI) to restart the management agents. Replace A Default ESI Certificate and Key Using HTTP PUT
1. Back up the existing certificates.
2. In your upload application, process each file as follows:
3. Open the file.
4. Publish the file to one of these locations:

Certificates https://hostname/host/ssl_cert
Keys https://hostname/host/ssl_key

3. The locations /host/ssl_cert and host/ssl_key link to the certificate files in /etc/vmware/ssl.
4. Restart the host.'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # esxiCentralizedCoreDumps
            'compliance': {'a': 'Compliance', 'b': 'Core dumps centralized.', 'd': 'ESXi host core dumps are sent to a centralized location.', 'f': 'Facilitates efficient troubleshooting, incident response, and forensic investigations.', 'h': 'Periodically verify the core dump collection system.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A centralized location is not configured to collect ESXi host core dumps.', 'd': 'It was observed that the centralized location is not configured to collect ESXi host core dumps.The VMware vSphere Network Dump Collector service allows for collecting diagnostic information from a host that experiences a critical fault. This service provides a centralized location for collecting ESXi host core dumps.', 'f': 'When a host crashes, an analysis of the resultant core dump is essential to being able to identify the cause of the crash and determine a resolution.', 'h': '''Installing a centralized dump collector helps ensure that core files are successfully saved and made available in the event an ESXi host should ever panic.
To implement the recommended configuration state, run the following ESXi shell commands:

# Configure remote Dump Collector Server
esxcli system coredump network set -v [VMK#] -i [DUMP_SERVER] -o [PORT]
# Enable remote Dump Collector
esxcli system coredump network set -e true
'''},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate data based on user input
    for i, question in enumerate(questions, 2):
        # Get user input for this question
        question_num = i - 1
        user_input = None
        
        if form_data:
            # Find the corresponding form field
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        # Get response data
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            # Populate columns C, D, F, G, H
            ws.cell(row=i, column=3, value=response_data['a'])  # Compliance/Non-Compliance/Not Applicable
            ws.cell(row=i, column=4, value=response_data['b'])  # Observation (Short/Brief)
            ws.cell(row=i, column=6, value=response_data['d'])  # Observation
            ws.cell(row=i, column=7, value=response_data['f'])  # Impact
            ws.cell(row=i, column=8, value=response_data['h'])  # Recommendation
            
            # Apply alignment to these columns
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        
        # Populate Risk Factor (column E) with color coding
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            # Apply color coding
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Only apply default font if it's not the Risk Factor column (column 5)
                if col != 5:  # Don't override Risk Factor column formatting
                    cell.font = Font(name='Calibri', size=11)
            # Set row height for wrapped text
            ws.row_dimensions[row].height = 30
    
    # Save file
    filename = "ESXi Server Logical Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    """
    Delete the generated Excel file after download
    """
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
