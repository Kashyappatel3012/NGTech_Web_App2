import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_load_balancer_array_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Load Balancer Array"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # Load Balancer Array Questions
    questions = [
        "Is an administrative portal login accessed only by Whitelisted IP Addresses?",
        "Whether the DoS protection is enabled or not?",
        "Are some of the supportive licenses expired?",
        "Is Data Packet Inspection (DPI) configured?",
        "Is NTP configured?",
        "Is the Syslog server implemented?",
        "Are Administrators having two-factor authentications enabled?",
        "Is Email Notification enabled?"
    ]

    # Risk Factors
    risk_factors = [
        "High", "High", "High", "Medium", "Medium", "Medium", "Low", "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "lbaAdminPortalWhitelisted": 1,
        "lbaDosProtection": 2,
        "lbaLicensesExpired": 3,
        "lbaDpiConfigured": 4,
        "lbaNtpConfigured": 5,
        "lbaSyslogServer": 6,
        "lbaTwoFactorAuth": 7,
        "lbaEmailNotification": 8
    }

    # Question responses data
    question_responses = {
        1: {  # lbaAdminPortalWhitelisted
            'compliance': {'a': 'Compliance', 'b': 'Admin portal restricted to whitelisted IPs.', 'd': 'Only approved IP addresses are allowed to access the administrative portal, preventing unauthorized remote logins.', 'f': 'Reduces the risk of external attacks and unauthorized access.', 'h': 'Periodically review and update the whitelist to maintain security.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'An administrative portal login accessed  was not only by Whitelisted IP Addresses', 'd': 'Administrative login can be accessed from any IP address, without limiting access to whitelisted IPs.', 'f': 'Any bank employee who does not have higher privileges can access the administrative portal and make changes in rules and policies from any computer. If the credentials are leaked then an attacker can access the admin portal from anywhere to change the configuration of that device thus an attacker can compromise the availability of that device.', 'h': 'It is recommended to Whitelist IP addresses so that only authorized users can access through the administrative portal login from the authorized computer only.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # lbaDosProtection
            'compliance': {'a': 'Compliance', 'b': 'DoS protection enabled.', 'd': 'DoS/DDoS protection mechanisms are actively configured and monitored to prevent service disruption.', 'f': 'Enhances system availability and resilience against attack.', 'h': 'Regularly review protection settings and test response to simulated attacks.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'DoS and Spoof protection was disabled.', 'd': 'Systems are not configured to detect or prevent Denial of Service attacks, leaving them vulnerable to service disruption.', 'f': "An attacker can perform attacks like Blind Spoofing, DoS attack, Man-in-the-Middle-Attack, and interrupt network traffic, hampering the bank's productivity.", 'h': 'It is recommended to configure and enable DoS and Spoof protection.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # lbaLicensesExpired
            'compliance': {'a': 'Compliance', 'b': 'All licenses valid and up-to-date.', 'd': 'Software and supportive licenses are active, ensuring access to updates, patches, and vendor support.', 'f': 'Ensures operational continuity and compliance with vendor agreements.', 'h': 'Maintain a license tracking system to avoid expirations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Some of the supportive Licences were expired.', 'd': 'Certain software or support licenses have expired, which may prevent updates, support, and critical patches.', 'f': 'The protection against new threats cannot be established with expired licences of various applications.', 'h': 'Renew all critical software and supportive licenses promptly and maintain an updated license inventory.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # lbaDpiConfigured
            'compliance': {'a': 'Compliance', 'b': 'DPI configured.', 'd': 'Data Packet Inspection is active, analyzing traffic for anomalies, threats, and policy compliance.', 'f': 'Enhances network security and prevents malicious activity.', 'h': 'Continuously monitor DPI logs and update inspection rules as required.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Data Packet Inspection (DPI) was not configured.', 'd': 'Network devices are not configured to inspect data packets, reducing visibility into malicious or unauthorized traffic.', 'f': 'If Data packet inspection (DPI) is disabled, it may not be able to perform inspection actions such as alerting, blocking, re-routing, or logging, better bandwidth utilization reporting, and traffic control/quality of service, which will result in low security.', 'h': 'It is recommended that Data Packet Inspection (DPI) should be configured and set to enabled.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # lbaNtpConfigured
            'compliance': {'a': 'Compliance', 'b': 'NTP configured.', 'd': 'All servers and network devices are synchronized with an NTP server, ensuring accurate timestamps.', 'f': 'Supports reliable audit trails, logs, and forensic investigations.', 'h': 'Periodically verify NTP synchronization and connectivity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'NTP server was not configured.', 'd': 'Servers and network devices are not synchronized with a reliable time source, causing inconsistent timestamps in logs and transactions.', 'f': 'A more insidious effect of weak timekeeping is that it damages the ability to investigate security breaches and other kinds of system problems. Hackers, for example, will often exploit backdoor, and proxy computers when mounting and attacking- both to hide their tracks and to exploit whatever opportunities (like NTP System privileges ) the hacker encounters along the way. Finding these stopping-off points is critical for shutting the door to future attacks and requires precise measurements of time in order to reconstruct the exact sequence of events. log file and application time stamp obviously become essential pieces of evidence.', 'h': 'It is recommended to implement an NTP server. By connecting your networked devices to a time server, which receives a signal from a definitive time source, you can enjoy the benefits of precise time in any location, boosting productivity, improving customer service, and synchronizing your operations.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # lbaSyslogServer
            'compliance': {'a': 'Compliance', 'b': 'Syslog server implemented.', 'd': 'Network and system events are logged centrally on a Syslog server for monitoring and auditing purposes.', 'f': 'Enhances visibility, security monitoring, and compliance reporting.', 'h': 'Regularly review logs and implement alerts for anomalous events.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Syslog server was not present.', 'd': 'Centralized logging of network and system events is missing, making it difficult to track and analyze events.', 'f': 'If any malicious activity takes place at the network level and logs are needed for forensic investigation. Then it will be difficult to get logs for investigation, thus affecting the quality of investigation. Also, periodic review of the logs will not be possible if the Syslog server is not present.', 'h': 'It is recommended to configure the Syslog server, so that logs will be available for a long duration and can be utilized when required.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # lbaTwoFactorAuth
            'compliance': {'a': 'Compliance', 'b': '2FA enabled for administrators.', 'd': 'All administrative accounts require an additional authentication factor (OTP, token, or biometric) along with a password.', 'f': 'Significantly reduces risk of unauthorized access.', 'h': 'Regularly verify and test 2FA functionality for all privileged accounts.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Two factor authentication was not enabled for the administration access.', 'd': 'Administrative accounts rely only on passwords, lacking additional authentication factors.', 'f': 'Increased risk of unauthorized access if credentials are compromised.', 'h': 'Enforce 2FA for all administrator accounts to strengthen security.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # lbaEmailNotification
            'compliance': {'a': 'Compliance', 'b': 'Email notifications enabled.', 'd': 'System alerts and notifications are configured to be sent via email for timely awareness and action.', 'f': 'Improves response time and operational security.', 'h': 'Periodically test email alerts to ensure proper delivery and effectiveness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Email notification was not enabled.', 'd': 'Alerts for critical events or failures are not sent via email, limiting timely awareness.', 'f': 'Delays detection and response to potential security or operational issues.', 'h': 'Enable email notifications for all critical system alerts to ensure prompt administrative response.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Load Balancer Array Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
