import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_san_storage_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "SAN Storage"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # SAN Storage Questions
    questions = [
        "Is an administrative portal login accessed only by whitelisted IP addresses?",
        "Is Full Disk Encryption configured?",
        "Is FTP being used?",
        "Is NTP configured?",
        "Is the Syslog server implemented?",
        "Is Email Notification enabled?",
        "Is login timeout as per organization policies?"
    ]

    # Risk Factors
    risk_factors = [
        "High", "High", "Medium", "Medium", "Medium", "Low", "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "ssAdminPortalWhitelisted": 1,
        "ssFullDiskEncryption": 2,
        "ssFtpUsed": 3,
        "ssNtpConfigured": 4,
        "ssSyslogServer": 5,
        "ssEmailNotification": 6,
        "ssLoginTimeout": 7
    }

    # Question responses data
    question_responses = {
        1: {  # ssAdminPortalWhitelisted
            'compliance': {'a': 'Compliance', 'b': 'Admin portal restricted to whitelisted IPs.', 'd': 'Administrative access is limited only to approved IP addresses, preventing unauthorized external access.', 'f': 'Enhances security of critical systems and reduces exposure to remote attacks.', 'h': 'Periodically verify the whitelist and update based on operational changes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': ' An administrative portal login accessed  was not only by Whitelisted IP Addresses.', 'd': 'Administrative login is not restricted to a defined set of trusted IP addresses, allowing access from any network location.', 'f': 'Any bank employee who does not have higher privileges can access the administrative portal and make changes in rules and policies from any computer. If the credentials are leaked then an attacker can access the admin portal from anywhere to change the configuration of that device thus an attacker can compromise the availability of that device.', 'h': 'Restrict administrative portal access to a predefined set of whitelisted IP addresses and review the list periodically.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # ssFullDiskEncryption
            'compliance': {'a': 'Compliance', 'b': 'Full disk encryption enabled.', 'd': 'All critical systems have full disk encryption applied, ensuring stored data is secure and unreadable without proper authentication.', 'f': 'Protects sensitive information from unauthorized access, theft, or loss.', 'h': 'Periodically validate encryption configurations and ensure encryption keys are securely managed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Full Disk Encryption was not configured in the system.', 'd': 'Data on servers or endpoints is stored in plaintext without encryption, making it vulnerable if devices are lost, stolen, or accessed by unauthorized personnel.', 'f': 'Full disk encryption protects the data on your device in the event it is lost or stolen. Without full disk encryption, if the data drive in the computer is removed, the data can be easily read and accessed. An unauthorized user can view and steal sensitive data, and delete and modify data if he gets access to the system. The whole data inside the disk may get compromised.', 'h': 'Implement Full Disk Encryption (FDE) on all critical servers and endpoints to protect data at rest.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # ssFtpUsed
            'compliance': {'a': 'Compliance', 'b': 'FTP not used; secure alternatives implemented.', 'd': 'Data transfers use secure protocols such as SFTP, FTPS, or HTTPS, ensuring encryption during transmission.', 'f': 'Reduces risk of data interception and protects credentials and sensitive information.', 'h': 'Regularly audit file transfer methods to confirm only secure protocols are used.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'FTP service active.', 'd': 'File Transfer Protocol (FTP) is used for data transfer, which is unencrypted and susceptible to eavesdropping or interception.', 'f': 'The FTP protocol is so easy to hack and intercept. It is vulnerable to Brute Force Attack and Port Stealing. A hacker can guess the next open port or use a PORT command to gain access as a middleman and brute force by checking frequently used and repeated passwords until they find the correct one.', 'h': 'Disable FTP and replace with secure alternatives like SFTP, FTPS, or HTTPS for file transfers.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # ssNtpConfigured
            'compliance': {'a': 'Compliance', 'b': 'NTP configured and synchronized.', 'd': 'All devices and servers are synchronized using NTP, ensuring consistent timestamps for logs, audit trails, and system events.', 'f': 'Facilitates accurate monitoring, incident investigation, and compliance with auditing requirements.', 'h': 'Periodically verify NTP synchronization and review time server configurations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Network Time Protocol (NTP) not configured.', 'd': 'Devices and servers are not synchronized to a reliable time source, causing inconsistent timestamps across logs and system events.', 'f': 'A more insidious effect of weak timekeeping is that it damages the ability to investigate security breaches and other kinds of system problems. Hackers, for example, will often exploit backdoor, and proxy computers when mounting and attacking- both to hide their tracks and to exploit whatever opportunities (like NTP System privileges ) the hacker encounters along the way. Finding these stopping-off points is critical for shutting the door to future attacks and requires precise measurements of time in order to reconstruct the exact sequence of events. log file and application time stamp obviously become essential pieces of evidence.', 'h': 'Configure NTP on all servers and critical devices to synchronize system clocks with a reliable time source.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # ssSyslogServer
            'compliance': {'a': 'Compliance', 'b': 'Syslog server implemented.', 'd': 'All relevant logs are centralized on a dedicated Syslog server, ensuring proper collection, retention, and accessibility for auditing and monitoring purposes.', 'f': 'Enhances visibility of system events, improves incident response capability, and supports regulatory compliance.', 'h': 'Regularly review Syslog configurations, log retention policies, and ensure all critical devices are forwarding logs appropriately.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Syslog server was not implemented.', 'd': 'Logs from network devices, servers, and applications are not centralized, making monitoring and audit trail management inconsistent and fragmented.', 'f': 'If any malicious activity takes place at the network level and logs are needed for forensic investigation. Then it will be difficult to get logs for investigation, thus affecting the quality of investigation. Also, periodic review of the logs will not be possible if the Syslog server is not present.', 'h': 'Implement a centralized Syslog server to collect, store, and manage logs from all critical systems, ensuring proper monitoring and audit readiness.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # ssEmailNotification
            'compliance': {'a': 'Compliance', 'b': 'Email notifications enabled.', 'd': 'Email alerts are configured for critical system events, failures, and security incidents, notifying the responsible staff promptly.', 'f': 'Improves incident response times, reduces system downtime, and ensures timely handling of security or operational issues.', 'h': 'Periodically test the notification system to verify timely delivery and effectiveness of alerts.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Email notifications not enabled.', 'd': 'Alerts for critical events, failures, or anomalies are not sent to responsible personnel, causing delays in response to system or security issues.', 'f': 'As the alert policy is not configured, Admin will not get a notification for any suspicious activity. If any malicious activity ensues then an email notification will not be generated and the bank may not be able to take necessary actions to prevent that activity.', 'h': 'Enable email notifications for critical events, including security alerts, system failures, and operational exceptions, to ensure prompt response.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # ssLoginTimeout
            'compliance': {'a': 'Compliance', 'b': 'Login timeout configured as per policy.', 'd': 'Sessions are automatically terminated after the defined period of inactivity, preventing unauthorized access from unattended terminals.', 'f': 'Enhances security by limiting exposure to unauthorized access and ensures compliance with organizational policies.', 'h': 'Periodically review timeout settings and adjust based on evolving security requirements or policy changes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Login timeout not enforced or misconfigured.', 'd': 'Sessions remain active beyond the organization\'s defined idle timeout period, allowing unauthorized access if the user leaves the system unattended.', 'f': 'An attacker can utilize this time to perform malicious activities and cause  harm to the network infrastructure, when an authorize user leave his system without logging out.', 'h': 'Configure session timeout settings as per organizational policies and enforce automatic logout after defined periods of inactivity.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "SAN Storage Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
