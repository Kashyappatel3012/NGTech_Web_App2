import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_nas_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "NAS"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # NAS Questions
    questions = [
        "Is an administrative portal login accessed only by Whitelisted IP Addresses?",
        "Are default credentials in use, or are accounts named 'Admin', 'root', or 'Administrator' to access administrator login?",
        "Is the DoS protection enabled or not?",
        "Is the firmware/DSM up to date?",
        "Is the administrative password set as per the organization's password policy?",
        "Is Admin lockout set to 3 or more failed login attempts or as per the organization policy?",
        "Is the firewall configured?",
        "Is the administrative password expiry set as per the organization's password policy?",
        "Is Security Advisor provided and used?",
        "Is Email Notification enabled?",
        "Are Administrators having two-factor authentications enabled?"
    ]

    # Risk Factors
    risk_factors = [
        "High", "High", "High", "High", "High", "Medium", "Medium", "Medium", "Medium", "Low", "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "nasAdminPortalWhitelisted": 1,
        "nasDefaultCredentials": 2,
        "nasDosProtection": 3,
        "nasFirmwareUpToDate": 4,
        "nasAdminPasswordPolicy": 5,
        "nasAdminLockout": 6,
        "nasFirewallConfigured": 7,
        "nasPasswordExpiry": 8,
        "nasSecurityAdvisor": 9,
        "nasEmailNotification": 10,
        "nasTwoFactorAuth": 11
    }

    # Question responses data
    question_responses = {
        1: {  # nasAdminPortalWhitelisted
            'compliance': {'a': 'Compliance', 'b': 'Administrative portal restricted to whitelisted IPs.', 'd': 'Admin access is allowed only from approved IP addresses, preventing unauthorized or external access attempts.', 'f': 'Reduces the attack surface and mitigates the risk of unauthorized access.', 'h': 'Regularly review and update the whitelist to reflect organizational changes or new administrative requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'An administrative portal login accessed  was not only by Whitelisted IP Addresses', 'd': 'Admin login can be attempted from any IP, without restriction, increasing exposure to unauthorized access attempts.', 'f': 'Any bank employee who does not have higher privileges can access the administrative portal and make changes in rules and policies from any computer. If the credentials are leaked then an attacker can access the admin portal from anywhere to change the configuration of that device thus an attacker can compromise the availability of that device.', 'h': 'Restrict administrative portal access to a predefined list of trusted IP addresses, enforce network-level restrictions, and monitor login attempts.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # nasDefaultCredentials
            'compliance': {'a': 'Compliance', 'b': 'Default credentials not in use.', 'd': 'All administrative accounts use unique identifiers and strong passwords, preventing common automated attack attempts.', 'f': 'Enhances security by reducing exposure to password guessing or brute-force attacks.', 'h': 'Periodically audit administrative accounts to ensure no default or weak credentials exist.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Default credentials were in use in the name of an account with 'Admin'.", 'd': 'Administrative accounts are using default names like \'Admin\' or \'root\', increasing vulnerability to automated attacks.', 'f': 'When not changed, default credentials make an organization more vulnerable to potential cyberattacks. Attackers can obtain these standard login details, allowing them access to the devices on your network usually with admin rights and leaving them open to takeover.', 'h': 'Rename default accounts, enforce strong unique passwords, and remove any unused default accounts.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # nasDosProtection
            'compliance': {'a': 'Compliance', 'b': 'DoS protection enabled.', 'd': 'Systems have DoS protection mechanisms in place to detect and mitigate attack traffic, ensuring availability.', 'f': 'Maintains business continuity and prevents potential service disruptions due to attacks.', 'h': 'Regularly test and update DoS protection rules to counter evolving attack patterns.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'DoS and Spoof protection was disabled.', 'd': 'Systems and applications are exposed to potential Denial-of-Service attacks without mitigation mechanisms.', 'f': "An attacker can perform attacks like Blind Spoofing, DoS attack, Man-in-the-Middle-Attack, and interrupt network traffic, hampering the bank's productivity.", 'h': 'Implement DoS/DDoS protection mechanisms, such as rate limiting, firewalls, and intrusion prevention systems.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # nasFirmwareUpToDate
            'compliance': {'a': 'Compliance', 'b': 'Firmware/DSM up to date.', 'd': 'All devices have the latest firmware/DSM applied, including relevant security updates.', 'f': 'Reduces vulnerabilities and strengthens overall device security.', 'h': 'Continuously monitor vendor updates and apply patches promptly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Firmware was not updated.', 'd': 'Devices are running older firmware/DSM versions that may contain known vulnerabilities.', 'f': "As, the firmware was not updated, an attacker can exploit the existing bug to compromise the security of bank's digital infrastructure. As, the firmware was not updated, an attacker can exploit the existing bug to compromise the security of bank's digital infrastructure. The older firmware may not utilize the hardware properly to enhance the performance of the device.", 'h': 'Ensure all devices are updated with the latest firmware/DSM releases and security patches.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # nasAdminPasswordPolicy
            'compliance': {'a': 'Compliance', 'b': 'Administrative passwords compliant with policy.', 'd': 'All administrative accounts adhere to organizational password standards, ensuring strong and secure credentials.', 'f': 'Reduces the risk of unauthorized access and maintains system integrity.', 'h': 'Periodically audit administrative passwords for compliance and enforce mandatory updates.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "An administrative password was not set as per the organization's password policy.", 'd': 'Admin passwords do not meet the organization\'s requirements for length, complexity, or rotation.', 'f': 'An attacker or malicious user could gain access by authenticating with a weak password. The attacker could enumerate information about the device and network configuration.', 'h': 'Enforce password policies for all administrative accounts, including minimum length, complexity, and periodic changes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # nasAdminLockout
            'compliance': {'a': 'Compliance', 'b': 'Admin lockout policy configured.', 'd': 'Admin accounts are locked after a set number of failed login attempts according to organizational standards.', 'f': 'Prevents brute-force attacks and reduces risk of unauthorized access.', 'h': 'Periodically review and test lockout settings to ensure effectiveness.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Lockout functionality was not available. ', 'd': 'It was observed that the admin panel will not lockout by any consecutive failed attempt making it vulnerable to brute force attack. Admin lockout not set to 3 or more failed login attempts or as per the organization policy.', 'f': 'A brute-force attack can be performed, which consists of an attacker submitting many passwords or passphrases with the hope of eventually guessing correctly. The attacker systematically checks all possible passwords and passphrases until the correct one is found.', 'h': 'Configure account lockout policies to disable admin accounts after a defined number of failed login attempts, as per organizational policy.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # nasFirewallConfigured
            'compliance': {'a': 'Compliance', 'b': 'Firewall configured.', 'd': 'Firewall rules are enforced to permit only authorized traffic and block unauthorized connections.', 'f': 'Enhances network security and protects critical systems from external threats.', 'h': 'Regularly audit firewall configurations and update rules in response to emerging threats.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The firewall was not configured.', 'd': 'It was observed that the firewall was not configured. Firewalls are important for protecting the computer from unwanted access, identifying and blocking unwanted data packets, and helping prevent worms, viruses, and malware.', 'f': 'Firewall misconfiguration that results in unintended access can open the door to breaches, data loss, and stolen or ransomed IP. Without adequate protection, malicious criminals can effectively shut your business down. And that can result in catastrophic damage to your business. attackers or external threats from getting access to your system in the first place.', 'h': 'Implement and regularly review firewall rules to restrict access and monitor traffic as per security policy.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # nasPasswordExpiry
            'compliance': {'a': 'Compliance', 'b': 'Administrative password expiry enforced.', 'd': 'All administrative accounts have password expiry configured according to organizational standards.', 'f': 'Reduces the risk of credential compromise and ensures stronger account security.', 'h': 'Periodically verify compliance and enforce password updates.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Administrative password expiry was not set as per the organization's password policy.", 'd': "It was observed that the administrative password expiry was not set as per the organization's password policy.", 'f': "Password must be changed within 30-60 days. If a bank is not able to change the password in 30-60 days, it increases the probability of an attacker guessing or cracking the password of user accounts. An attacker who can determine user passwords can take over a user's account and potentially access sensitive data in the application.", 'h': 'Enforce periodic password expiration and ensure all administrative accounts comply with the defined rotation policy.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # nasSecurityAdvisor
            'compliance': {'a': 'Compliance', 'b': 'Security Advisor implemented and in use.', 'd': 'Security Advisor or equivalent monitoring tools are actively used to detect vulnerabilities, misconfigurations, and policy deviations.', 'f': 'Enhances proactive risk management and reduces potential security incidents.', 'h': 'Regularly review Security Advisor reports and implement recommended actions promptly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Security Advisor was not used in the bank.', 'd': 'It was observed that Security Advisor was provided but not used by the Bank employees.', 'f': 'It deliver real-time coaching for each employee, teaching them how to identify and remediate cyberattacks and help security teams better measure the ROI and effectiveness of their training initiatives.', 'h': 'Deploy and utilize Security Advisor or equivalent tools to continuously monitor and improve system security posture.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # nasEmailNotification
            'compliance': {'a': 'Compliance', 'b': 'Email notifications enabled.', 'd': 'System is configured to send alerts and notifications for critical events, failures, and policy violations via email.', 'f': 'Ensures timely detection and response to security or operational issues.', 'h': 'Periodically test email alerts to confirm proper delivery and relevance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Email notification was not enabled.', 'd': 'It was observed that email notification was not enabled. Admin emails were not configured to receive alert notifications.', 'f': 'As the alert policy is not configured, Admin will not get a notification for any suspicious activity. If any malicious activity ensues then an email notification will not be generated and the bank may not be able to take necessary actions to prevent that activity.', 'h': 'Enable email notifications for critical alerts to ensure timely awareness and response by administrators.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # nasTwoFactorAuth
            'compliance': {'a': 'Compliance', 'b': '2FA enabled for administrators.', 'd': 'All administrator accounts require a second factor (e.g., OTP, token, or biometric) in addition to a password for login.', 'f': 'Significantly reduces the likelihood of unauthorized access and enhances overall system security.', 'h': 'Regularly verify that 2FA is active and functional for all privileged accounts.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Two-factor authentication (2FA) not enabled for administrators.', 'd': 'It was observed that two-factor authentication was not enabled for administration access. If anyone gets credentials, then they can easily access the admin portal.', 'f': "Two-factor authentication is a security process in which users provide two different authentication factors to verify themselves. This process is done to protect both the user's credentials and the resources the user can access. If it is disabled, an unauthenticated user can access the admin account without the admin person knowing that his account was logged in unauthorizedly.", 'h': 'Enforce 2FA for all administrative accounts to strengthen security and mitigate credential theft risks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "NAS Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
