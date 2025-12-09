import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_business_continuity_planning_excel(form_data=None):
    """
    Create Business Continuity Planning Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Business Continuity Planning"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Business Continuity Planning Questions
    questions = [
        "Has the business continuity plan (BCP) been documented?",
        "Does the BCP cover all levels of disaster, from partial to total destruction of facilities, and provide guidelines to determine the level of recovery necessary?",
        "Is a ready or alternate source of hardware/software available to resume business activity within the shortest possible time after disruption?",
        "Is a reliable backup of data and software available at all times for restoration?",
        "Are BCP tests conducted regularly, and what are the results?",
        "How are identified weaknesses in the BCP addressed?",
        "Is a copy of the BCP stored securely off-site?"
    ]

    # Risk Factors
    risk_factors = [
        "High",
        "High", 
        "Medium",
        "Medium",
        "Medium",
        "Medium",
        "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "bcpDocumented": 1,
        "bcpCoversAllDisasterLevels": 2,
        "bcpAlternateHardwareSoftware": 3,
        "bcpReliableBackup": 4,
        "bcpTestsConducted": 5,
        "bcpWeaknessesAddressed": 6,
        "bcpOffSiteStorage": 7
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Question responses data
    question_responses = {
        1: {  # bcpDocumented
            'compliance': {'a': 'Compliance', 'b': 'BCP documented.', 'd': 'A detailed Business Continuity Plan exists and is formally documented, outlining procedures, responsibilities, and recovery objectives for all critical business processes.', 'f': 'Ensures structured response during disruptions, minimizes downtime, and supports organizational resilience.', 'h': 'Review and update the BCP periodically to reflect changes in business processes or infrastructure.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': ' A business continuity plan was not available.', 'd': 'The organization does not have a formally documented Business Continuity Plan, or the plan is incomplete and lacks critical details.', 'f': 'If a bank has not prepared a business continuity plan properly and cannot address a disruption that has an impact on the business continuity of an organization then it can lead to financial losses because of events like fire, earthquake, or manmade disasters. As banks cannot effectively communicate with their stakeholders during an emergency, Bank might have decreased stability or complete shutdown and face reputation loss, and loss of productivity.', 'h': 'It is recommended to implement a business continuity plan which could be beneficial for the bank if any kind of disaster, as well as destruction, arises at the Bank.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # bcpCoversAllDisasterLevels
            'compliance': {'a': 'Compliance', 'b': 'BCP covers all disaster levels.', 'd': 'The plan includes detailed procedures for partial, significant, and total disasters, with clear guidelines for assessing recovery requirements.', 'f': 'Facilitates effective decision-making during emergencies, reducing downtime and operational impact.', 'h': 'Conduct scenario-based exercises to test effectiveness and refine the plan.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'BCP does not cover all levels of the disaster from partial to total destruction of facilities.', 'd': 'The existing plan lacks guidelines to address different disaster scenarios or does not specify procedures for partial vs. total facility loss.', 'f': 'As BCP does not cover all levels of disaster and does not contain guidelines to help determine the level of recovery necessary, in case of a disruption the bank will not be able to take necessary actions to prevent or face an emergency situation. Thus, the bank may face financial loss, reputation loss, or total shutdown.', 'h': 'It is recommended that BCP should cover all levels of the disaster from partial to total destruction of facilities and contain guidelines to help determine the level of recovery necessary.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # bcpAlternateHardwareSoftware
            'compliance': {'a': 'Compliance', 'b': 'Alternate hardware/software available.', 'd': 'Standby hardware and software resources are ready for use in case of disruptions, ensuring rapid resumption of business activities.', 'f': 'Reduces downtime, minimizes operational losses, and supports continuity of services.', 'h': 'Periodically test the readiness of alternate resources to ensure effectiveness during an actual disruption.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'An alternate source of hardware and software was not available to resume business activities within the shortest possible time after a disruption.', 'd': 'The organization does not maintain ready or standby hardware and software resources for rapid recovery in case of system failure or disaster.', 'f': 'Because of lack of ready or alternate source of hardware/software the bank will not be able to cope from a disaster or resume business activity within the shortest possible time after disruptions. All the critical bank functions will be terminated. Thus bank can face financial , reputational, customer loss.', 'h': 'It is recommended to have a ready or alternate source of hardware/software  to resume business activity within the shortest possible time after disruption.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # bcpReliableBackup
            'compliance': {'a': 'Compliance', 'b': 'Reliable backups maintained.', 'd': 'Regular and secure backups of all critical data and software are maintained, with periodic verification to ensure integrity and recoverability.', 'f': 'Guarantees operational resilience and facilitates quick restoration after disruptions.', 'h': 'Review and test backup procedures regularly to confirm recoverability.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Reliable backup of data and software was not available.', 'd': 'Data and software backups are inconsistent, incomplete, or not maintained regularly, limiting the ability to restore operations effectively.', 'f': 'Because of a lack of backup data and software unavailability the bank will not be able to cope with a disaster as no backup is available which can bring the business back to running. All the critical bank functions will be terminated. Thus the bank can face financial, reputational, and customer losses.', 'h': 'It is recommended to have a reliable backup of data and software available all the time for restoration purposes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # bcpTestsConducted
            'compliance': {'a': 'Compliance', 'b': 'BCP tested regularly.', 'd': 'Periodic BCP testing is conducted, results are documented, and corrective actions are implemented to address identified weaknesses.', 'f': 'Strengthens organizational preparedness, reduces risk during disruptions, and ensures continuous improvement of the plan.', 'h': 'Schedule annual or semi-annual tests and integrate lessons learned into BCP updates.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'There is no information available to assess whether BCP tests are conducted regularly or the results of such tests.', 'd': 'The organization does not perform regular BCP testing, or test results are incomplete and not analyzed to identify gaps.', 'f': 'Potential weaknesses in the plan remain unaddressed, increasing the risk of inadequate response during actual incidents.', 'h': 'Conduct regular BCP tests, document results, and review gaps to enhance preparedness and response effectiveness.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # bcpWeaknessesAddressed
            'compliance': {'a': 'Compliance', 'b': 'Weaknesses addressed promptly.', 'd': 'All identified gaps in the BCP are tracked, root causes analyzed, and corrective actions implemented to strengthen business continuity measures.', 'f': 'Reduces operational risk, ensures rapid recovery, and improves organizational resilience.', 'h': 'Maintain a documented log of issues and corrective actions, and review progress periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The process for addressing identified weaknesses in the BCP is not clearly defined.', 'd': 'Identified gaps or weaknesses in the BCP are not formally tracked, analyzed, or remediated, leaving the organization vulnerable to operational risks.', 'f': 'Unresolved weaknesses may result in inefficient disaster response, prolonged downtime, and increased financial and reputational damage.', 'h': 'Establish a formal process to track, analyze, and remediate weaknesses identified during BCP reviews or testing.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # bcpOffSiteStorage
            'compliance': {'a': 'Compliance', 'b': 'BCP stored securely off-site.', 'd': 'Copies of the BCP are stored in secure off-site locations with controlled access and redundancy, ensuring availability even if primary site is compromised.', 'f': 'Guarantees that business continuity procedures can be accessed during disasters, supporting timely recovery.', 'h': 'Regularly review off-site storage procedures and update BCP copies to reflect the latest version.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A copy of the plan is not stored securely off-site.', 'd': 'No secure off-site copies of the BCP exist, or off-site storage lacks proper access controls and protection.', 'f': 'In case of total site disaster, on-site BCP documents may be lost, delaying recovery and risking business continuity.', 'h': 'Store secure copies of the BCP off-site, with restricted access, redundancy, and regular updates to ensure availability during emergencies.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }

    # Populate data based on user input
    for i, question in enumerate(questions, 2):
        # Get user input for this question
        question_num = i - 1
        user_input = None
        
        if form_data:
            # Find the corresponding form field
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Business Continuity Planning Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
