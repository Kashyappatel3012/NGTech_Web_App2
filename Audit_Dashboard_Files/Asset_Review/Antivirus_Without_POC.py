import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_antivirus_excel(form_data=None):
    """
    Create Excel file for Antivirus Assessment
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Antivirus"
    
    # Define questions
    questions = [
        "Is the antivirus available?",
        "How is the antivirus administrator console accessed? Is access to the administrative console limited to whitelisted computers only?",
        "Is the Antivirus database up to date? Is the virus definition updated automatically?",
        "Does Antivirus have behavior mode enabled?",
        "Is the administrative password set as per the organization's password policy?",
        "Is the bank using an enterprise-grade antivirus that is managed centrally?",
        "Is the Antivirus Application up to date? Or Is Authorized version of Antivirus software with its latest updates available?",
        "Is the Antivirus software configured to check viruses even from the floppy drive / CD ROM drive?",
        "Is USB enabled in the antivirus server?",
        "Does the antivirus have Anti-APT functionality to detect advanced or signature-less threats? Is this feature enabled?",
        "Does the antivirus include HIPS to detect and block any scanning or exploitation attempts? Is this feature enabled?",
        "Does the antivirus provide a firewall to block unauthorized port access? Is the firewall enabled?",
        "Does the antivirus support device control to block USB storage, CD/DVD-ROM, and other external devices? Is this feature enabled?",
        "Is the antivirus configured to quarantine/repair or delete infected files?",
        "Is the patch server configured?",
        "Does the antivirus have any exclusions configured? If yes, are they reviewed and approved? Is there a plan to remove these exclusions?",
        "Is the antivirus integrated with Active Directory for user management?",
        "Are DDOS and Port scanning attacks Enabled?",
        "Is Data Loss Prevention enabled?",
        "Does the antivirus have real-time scanning enabled?",
        "Is the antivirus configured to scan all file extensions? Are zip/compressed files scanned?",
        "Has the bank scheduled periodic (at least weekly) full system scans?",
        "Does the antivirus client have a self-defense module enabled which restricts the user/program from performing changes or disabling them to the client antivirus software?",
        "Is there any mechanism to control web content?",
        "Is the antivirus dashboard monitored regularly for security events?",
        "Is there an NTP server configured?",
        "Is logging enabled and the logs stored on a centralized logging server?",
        "Is the notifications option enabled?",
        "Is application control enabled?"
    ]
    
    # Risk factors for each question
    risk_factors = [
        'Critical', 'Critical', 'Critical', 'High', 'High', 'High', 'High', 'High', 'High', 'High',
        'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'Medium', 'Medium',
        'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'Low'
    ]
    
    # Question mapping for form fields
    question_mapping = [
        'antivirusAvailable', 'adminConsoleAccess', 'databaseUpdated', 'behaviorMode', 'adminPassword',
        'enterpriseGrade', 'antivirusUpdated', 'floppyCdScan', 'usbEnabled', 'antiApt',
        'hipsEnabled', 'firewallEnabled', 'deviceControl', 'quarantineConfig', 'patchServer',
        'exclusionsConfig', 'adIntegration', 'ddosProtection', 'dlpEnabled', 'realtimeScanning',
        'allFileTypes', 'scheduledScans', 'selfDefense', 'webContentControl', 'dashboardMonitoring',
        'ntpServer', 'centralizedLogging', 'notificationsEnabled', 'applicationControl'
    ]
    
    # Set column widths
    ws.column_dimensions['A'].width = 10  # Sr. No.
    ws.column_dimensions['B'].width = 50  # Questions
    ws.column_dimensions['C'].width = 20  # Compliance/Non-Compliance/Not Applicable
    ws.column_dimensions['D'].width = 30  # Observation (Short/Brief)
    ws.column_dimensions['E'].width = 20  # Risk Factor
    ws.column_dimensions['F'].width = 50  # Observation
    ws.column_dimensions['G'].width = 50  # Impact
    ws.column_dimensions['H'].width = 50  # Recommendation
    
    # Header row
    headers = ['Sr. No.', 'Questionnaire/Points', 'Compliance/Non-Compliance/Not Applicable', 
               'Observation (Short/Brief)', 'Risk Factor', 'Observation', 'Impact', 'Recommendation']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Populate Sr. No. and Questions
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }
    
    # Question responses data
    question_responses = {
        1: {  # antivirusAvailable
            'compliance': {'a': 'Compliance', 'b': 'Antivirus installed on all systems.', 'd': 'It was verified that antivirus is installed and operational on all bank systems, providing protection against viruses, malware, and other malicious software.', 'f': 'Ensures systems are safeguarded against malware attacks, maintaining the confidentiality, integrity, and availability of critical information.', 'h': 'Regularly verify antivirus installation and ensure continuous protection on all systems.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Antivirus not installed on some systems.', 'd': 'It was observed that antivirus was not available on certain bank systems, leaving them vulnerable to malicious software, malware, and viruses.', 'f': 'The absence of antivirus software on some systems exposes them to malware, ransomware, and other malicious attacks. This can lead to data loss, system compromise, and unauthorized access to critical information.', 'h': 'It is recommended that licensed antivirus software be installed and activated on all systems to ensure adequate protection against malware and other threats.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # adminConsoleAccess
            'compliance': {'a': 'Compliance', 'b': 'Admin console access restricted to authorized users and whitelisted computers.', 'd': 'It was verified that the antivirus administrator console is accessible only to authorized IT personnel from whitelisted systems, preventing unauthorized modifications.', 'f': 'Prevents misuse or misconfiguration of antivirus settings by unauthorized users.', 'h': 'Maintain strict access control to the antivirus admin console with whitelisted IP addresses.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': ' Antivirus admin console access not restricted to whitelisted systems only.', 'd': 'It was observed that the antivirus administrator console is accessed by the IT department but access is not restricted to whitelisted computers, allowing potential unauthorized access.', 'f': 'Unauthorized users could disable or modify antivirus settings, exposing systems to malware and other threats.', 'h': 'Restrict administrative console access to authorized personnel only and enforce whitelisted IP-based access.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # databaseUpdated
            'compliance': {'a': 'Compliance', 'b': 'Antivirus database updated and definitions auto-updated.', 'd': 'It was verified that antivirus databases are regularly updated and virus definitions are refreshed automatically, ensuring detection of latest threats.', 'f': 'Systems are protected against emerging malware, maintaining security of bank data and infrastructure.', 'h': 'Review and update reports on a periodic basis to ensure that all systems are being updated.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Antivirus database was not updated. The virus definition was not updated automatically.', 'd': 'It was observed that antivirus databases were not updated regularly, and automatic virus definition updates were not enabled.', 'f': 'Outdated antivirus cannot detect new malware or threats, leaving systems exposed to attacks and potential compromise.', 'h': "It is recommended to update antivirus daily and enable automatic updates. It's recommended to check that the antivirus definitions are up to date on critical systems."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # behaviorMode
            'compliance': {'a': 'Compliance', 'b': 'Behavior mode enabled.', 'd': 'It was verified that antivirus behavior mode is enabled, allowing detection of suspicious or unknown activities beyond signature-based scanning.', 'f': 'Enhances malware detection and prevents zero-day threats from compromising systems.', 'h': 'Maintain behavior mode enabled on all antivirus solutions.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Antivirus was not having behavior mode enabled.', 'd': 'It was observed that antivirus behavior mode was not enabled on several systems, limiting the detection to signature-based threats only.', 'f': 'Attackers could bypass signature detection using custom malware, risking data integrity and system security.', 'h': 'Enable behavior mode to enhance protection against unknown and advanced threats.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # adminPassword
            'compliance': {'a': 'Compliance', 'b': 'Admin password compliant with policy.', 'd': 'It was verified that administrative passwords are set as per the organization\'s password policy, ensuring strong and secure credentials.', 'f': 'Reduces risk of unauthorized access to critical systems and maintains overall system security.', 'h': 'Maintain password compliance and enforce periodic password changes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Administrative password was not set as per the organizations password policy.', 'd': 'It was observed that administrative passwords were not set according to the organization\'s password policy, making them weak and guessable.', 'f': 'Increased risk of brute-force or password-guessing attacks, potentially compromising critical systems.', 'h': 'It is recommended that administrative passwords be configured in accordance with the organization’s password policy, ensuring sufficient complexity and length. Passwords should be changed periodically and kept confidential to prevent unauthorized use.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # enterpriseGrade
            'compliance': {'a': 'Compliance', 'b': 'Centralized antivirus server implemented.', 'd': 'It was verified that antivirus is managed centrally, allowing consistent policy enforcement and updates across all systems.', 'f': 'Simplifies management, ensures uniform protection, and reduces the workload on IT staff.', 'h': 'Continue centralized antivirus management to maintain consistent security controls.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Centralized antivirus server was not established for systems having internet access.', 'd': 'It was observed that no centralized antivirus server was established, resulting in standalone deployments without uniform policy enforcement.', 'f': 'Inconsistent protection across systems increases risk of malware infections and complicates antivirus management for IT staff.', 'h': 'Implement a centralized antivirus server to enforce policies, automate updates, and simplify management.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # antivirusUpdated
            'compliance': {'a': 'Compliance', 'b': 'Antivirus application updated to latest version.', 'd': 'It was verified that all systems have the latest version of the antivirus application installed, ensuring up-to-date protection against emerging threats.', 'f': 'Reduces risk of malware bypassing outdated antivirus software and ensures comprehensive threat detection.', 'h': 'Continue updating antivirus software regularly and ensure all endpoints have the latest versions installed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Antivirus was not updated in bank systems.', 'd': 'It was observed that some systems were running outdated antivirus versions, making them unable to detect new viruses and malware effectively.', 'f': 'Systems remain vulnerable to new threats, potentially leading to data compromise, malware infections, and disruption of operations.', 'h': 'Update antivirus software on all systems regularly and configure automatic update checks to ensure the latest version is always in use.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # floppyCdScan
            'compliance': {'a': 'Compliance', 'b': 'Antivirus scans all removable media.', 'd': 'It was verified that antivirus is configured to scan viruses from floppy drives and CD-ROMs, preventing malware from removable media from infecting the system.', 'f': 'Ensures data integrity and prevents propagation of malware through external media.', 'h': 'Maintain antivirus scanning for all removable media and verify configuration periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Antivirus was not configured to check viruses from the floppy drive/CD ROM drive.', 'd': 'It was observed that antivirus was not configured to scan floppy drives or CD-ROMs, allowing potential malware or malicious files to enter systems via removable media.', 'f': 'External media can be used to introduce malware, create backdoors, or compromise system data, risking confidentiality and integrity.', 'h': 'It is recommended that configure antivirus to scan all removable media, including floppy drives and CD-ROMs, to prevent malware infections.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # usbEnabled
            'compliance': {'a': 'Compliance', 'b': 'USB access restricted on antivirus server.', 'd': 'It was verified that USB access is restricted on antivirus servers, preventing unauthorized devices from connecting and introducing threats.', 'f': 'Protects critical servers from data theft, malware injection, and unauthorized access.', 'h': 'Continue restricting USB access on antivirus servers and enforce policy compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'USB enabled on antivirus server.', 'd': 'It was observed that USB ports were enabled on antivirus servers, allowing potential unauthorized devices to connect and introduce malware or steal data.', 'f': 'Increases risk of data leakage, malware infections, and compromise of sensitive server data.', 'h': 'Disable USB access on antivirus servers and allow only authorized removable devices if absolutely necessary.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # antiApt
            'compliance': {'a': 'Compliance', 'b': 'Anti-APT enabled.', 'd': 'It was verified that Anti-APT functionality is enabled, allowing detection of advanced and signature-less threats.', 'f': 'Protects systems against sophisticated attacks that bypass traditional antivirus signatures.', 'h': 'Maintain Anti-APT enabled on all systems.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Anti-APT not enabled.', 'd': 'It was observed that Anti-APT detection was not enabled, leaving systems vulnerable to advanced persistent threats.', 'f': 'Not enabling Anti-APT (Advanced Persistent Threat) protection in the antivirus solution increases the risk of undetected sophisticated attacks targeting critical systems and data. Without this feature, advanced malware, zero-day exploits, and targeted intrusions may bypass traditional antivirus defenses, leading to potential data breaches and system compromise.', 'h': 'Enable Anti-APT in antivirus to detect and mitigate advanced threats effectively.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # hipsEnabled
            'compliance': {'a': 'Compliance', 'b': 'HIPS enabled.', 'd': 'HIPS functionality is enabled, providing protection against scanning and exploitation attempts on endpoints.', 'f': 'Prevents attackers from exploiting vulnerabilities on endpoints.', 'h': 'Keep HIPS enabled across all systems.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'HIPS not enabled.', 'd': 'It was observed that HIPS is not enabled, leaving systems vulnerable to exploitation and attacks.', 'f': 'Exploitation of system vulnerabilities could allow attackers to gain unauthorized access or execute malicious code.', 'h': 'Enable HIPS to detect and block intrusion attempts.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # firewallEnabled
            'compliance': {'a': 'Compliance', 'b': 'Firewall enabled.', 'd': 'Firewall is enabled on all antivirus-protected systems, controlling port access and blocking unauthorized network traffic.', 'f': 'Reduces risk of network-based attacks and unauthorized access.', 'h': 'Maintain firewall enabled and monitor configurations periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Firewall was disabled.', 'd': 'It was observed that firewall was not enabled, allowing unrestricted network access to endpoints.', 'f': 'Systems are exposed to port-based attacks, malware propagation, and unauthorized network access.', 'h': 'Enable firewall on all systems and ensure proper rules are configured to block unauthorized traffic.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # deviceControl
            'compliance': {'a': 'Compliance', 'b': 'Device control enabled.', 'd': 'Device control is enabled to block unauthorized USBs, CDs, and other peripherals from connecting to endpoints.', 'f': 'Prevents malware introduction and data exfiltration through removable devices.', 'h': 'Maintain device control enabled on all endpoints.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Device control feature was not enabled.', 'd': 'It was observed that device control was not enabled, to block USB storage, CD/DVD-ROM, and other external devices.', 'f': 'Not enabling device control to block USB storage, CD/DVD-ROM, and other external devices exposes the system to data leakage and malware infection risks. Unauthorized copying or transfer of sensitive data through external media can lead to data breaches', 'h': 'It is recommended that device control be enabled to restrict or block unauthorized use of USB storage devices, CD/DVD-ROMs, and other external media.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # quarantineConfig
            'compliance': {'a': 'Compliance', 'b': 'The antivirus solution is configured to automatically quarantine, repair, or delete infected files upon detection to prevent the spread of malware.', 'd': 'It was verified that antivirus is configured to quarantine, repair, or delete infected files, isolating malware and preventing it from spreading to other parts of the system.', 'f': 'This configuration ensures timely containment and removal of malicious files, protecting system integrity, minimizing infection impact, and maintaining a secure computing environment.', 'h': 'Maintain antivirus configuration to quarantine, repair, or delete infected files.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Antivirus was not configured to quarantine/repair or delete infected files.', 'd': 'It was observed that antivirus was not configured to quarantine, repair, or delete infected files. Infected files remain accessible, potentially compromising the entire system.', 'f': 'Malware can spread, damage files, and compromise the CIA triad of bank systems.', 'h': 'Configure antivirus to quarantine, repair, or delete infected files to prevent malware propagation.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # patchServer
            'compliance': {'a': 'Compliance', 'b': 'Patch server configured properly.', 'd': 'It was verified that the patch server is configured to deploy updates and patches in real-time across all systems, fixing vulnerabilities and improving system functionality.', 'f': 'Protects systems from known vulnerabilities, ensuring security, reliability, and compliance.', 'h': 'Maintain proper configuration and monitoring of the patch server.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The patch server was not configured properly.', 'd': 'It was observed that the patch server was not configured properly, preventing real-time monitoring and deployment of patches.', 'f': 'Systems remain exposed to vulnerabilities, increasing the risk of exploitation and compromise of sensitive data.', 'h': 'Configure the patch server properly to ensure timely patch deployment and system security.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # exclusionsConfig
            'compliance': {'a': 'Compliance', 'b': 'Antivirus exclusions are reviewed, approved, and periodically monitored.', 'd': 'It was verified that antivirus exclusions are documented, reviewed, and approved by authorized personnel, ensuring that only legitimate files are excluded from scans.', 'f': 'Prevents malware from being unintentionally excluded, maintaining system security and performance.', 'h': 'Maintain a formal approval process and review exclusions periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'There was no authority-approved process to exclude program/software in Antivirus.', 'd': 'It was observed that there was no authority-approved process to exclude program/software in antivirus.The exclusion section allows you to create a list of files and folders that should be skipped during antivirus scans. This includes manually excluded items and items which you chose to ignore at the scan results window or antivirus alert.', 'f': 'It is also important that any unauthorized malicious files are not excluded from the scan otherwise they can compromise the CIA triad of the bank system. It is important to achieve a balance between ensuring a secure and virus-free server environment, while not interfering with the reliability and performance of each server or application. Virus scanning is often a cause of performance issues because a lack of properly configured antivirus exclusions may cause outages of applications and services due to contention or file locking.', 'h': 'It is recommended to implement an official approval process to exclude any program or software  in the Antivirus within the given particular timeline.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # adIntegration
            'compliance': {'a': 'Compliance', 'b': 'Integrated with Active Directory.', 'd': 'It was verified that antivirus is integrated with Active Directory, enabling centralized user management, configuration updates, and patch deployment across all systems.', 'f': 'Simplifies IT management and ensures consistent security policies across the network.', 'h': 'Maintain Active Directory integration for centralized antivirus management.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Antivirus was not integrated with Active Directory.', 'd': 'It was observed that antivirus is not integrated with Active Directory, making it difficult to manage updates, patches, and configurations efficiently.', 'f': 'Increased risk of inconsistent updates, delayed patch deployment, and administrative overhead, potentially compromising security.', 'h': 'It is recommended that the antivirus solution be integrated with Active Directory to enable centralized management and monitoring of all domain-connected systems. This integration ensures uniform policy enforcement, timely updates, and effective incident response.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        18: {  # ddosProtection
            'compliance': {'a': 'Compliance', 'b': 'DDOS and Port scanning protections enabled.', 'd': 'It was verified that antivirus has protections enabled for DDOS and port scanning attacks, preventing network-based threats.', 'f': 'Protects critical systems from service disruption and maintains availability of banking services.', 'h': 'Maintain protection against DDOS and port scanning attacks.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'DDOS and Port scanning attacks were disabled.', 'd': 'It was observed that DDOS and port scanning attacks were disabled. Which will make the antivirus unable to prevent DDOS and port scanning attacks.', 'f': 'As these services are not enabled in the Anti-Virus, it will not protect against DDoS and port scanning attacks. Money, time, clients, and even reputation can be lost in the event of a DDoS attack. Depending on the severity of an attack, resources could be offline for hours or days.', 'h': 'It is recommended to enable protection against port scanning and DDOS attacks to gain additional security levels.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        19: {  # dlpEnabled
            'compliance': {'a': 'Compliance', 'b': 'DLP enabled.', 'd': 'It was verified that DLP is enabled, ensuring sensitive data is not lost, misused, or accessed by unauthorized users.', 'f': 'Protects sensitive data, ensures regulatory compliance, and supports incident response.', 'h': 'Maintain DLP enabled and monitor for anomalies or unauthorized access.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Data loss prevention was not enabled.', 'd': 'It was observed that DLP is not enabled, leaving sensitive data unprotected.', 'f': 'The absence of Data Loss Prevention (DLP) controls increases the risk of unauthorized disclosure or leakage of sensitive and confidential information. Without DLP, users may intentionally or unintentionally transfer critical data through email, external storage, or the internet, leading to financial loss, reputational damage.', 'h': 'Enable DLP in antivirus to protect sensitive information and maintain compliance.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        20: {  # realtimeScanning
            'compliance': {'a': 'Compliance', 'b': 'Real-time scanning enabled.', 'd': 'It was verified that antivirus real-time scanning is enabled, checking files as they are accessed, created, or copied.', 'f': 'Detects threats immediately, preventing malware execution and system compromise.', 'h': 'Maintain real-time scanning enabled and configure proactive scan and failure settings.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Antivirus had the real time scanning set to disabled.', 'd': 'It was observed that real-time scanning is disabled, leaving files unmonitored until scheduled scans occur.', 'f': 'Disabling real-time scanning in antivirus software leaves systems vulnerable to immediate threats such as malware, ransomware, and other malicious files. Without real-time protection, infections may go undetected until manual scans are performed, increasing the risk of data compromise and system downtime.', 'h': 'Enable real-time scanning on all systems and configure proactive scanning and scan failure alerts.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        21: {  # allFileTypes
            'compliance': {'a': 'Compliance', 'b': 'All file types, including compressed files, scanned.', 'd': 'It was verified that antivirus is configured to scan all file extensions, including zip and compressed files, preventing malicious files from bypassing detection.', 'f': 'Detects malware hidden in compressed files, reducing risk of system compromise.', 'h': 'Maintain configuration to scan all file types, including compressed files.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Antivirus was not configured to scan all file extensions such as Zip files/compressed files.', 'd': 'It was observed that antivirus is not configured to scan all file extensions, such as zip/compressed files, allowing potential threats to bypass detection.', 'f': 'Not configuring the antivirus to scan all file extensions, including ZIP and other compressed files, increases the risk of hidden malware going undetected. Attackers often use compressed files to bypass standard antivirus scans, which can lead to system infection, data compromise, or network-wide security incidents.', 'h': 'Configure antivirus to scan all file types, including compressed files, to prevent malware infiltration.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        22: {  # scheduledScans
            'compliance': {'a': 'Compliance', 'b': 'Scheduled full system scans enabled.', 'd': 'It was verified that periodic full system scans are scheduled at least weekly, ensuring all files are regularly checked for malware.', 'f': 'Early detection of dormant threats and improved system security.', 'h': 'Maintain scheduled full system scans and preferably run them during off-business hours.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Scheduled scanning was disabled.', 'd': 'It was observed that scheduled full system scans are disabled, leaving files unchecked for potential threats.', 'f': "As the scheduled scan is not enabled, the scans can not be performed at regular periods. Thus, the detection of any malicious activities is not possible. This will be a major threat to systems and bank's infrastructure.", 'h': 'It is recommended to enable scheduled, periodic(at least weekly)full system scans. Ideally, all systems would be scanned during off-business hours to limit any potential performance hit from affecting a user’s ability to work.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        23: {  # selfDefense
            'compliance': {'a': 'Compliance', 'b': 'Self-defense module enabled.', 'd': 'It was verified that antivirus clients have self-defense modules enabled, preventing users or programs from altering or disabling antivirus settings.', 'f': 'Protects antivirus from tampering, maintaining continuous protection against malware.', 'h': 'Maintain self-defense module enabled on all antivirus clients.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Antivirus client does not have a self-defense module enabled. ', 'd': 'It was observed that the antivirus client does not have a self-defense module enabled. Client antivirus software can make changes raising the risk of a virus entering the system.', 'f': 'Not enabling the self-defense module in the antivirus client exposes the system to the risk of tampering or disabling of antivirus services by malware or unauthorized users. This can lead to loss of protection, undetected infections, and potential compromise of critical systems and data.', 'h': 'Enable antivirus self-defense module to restrict changes or disabling of antivirus software.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        24: {  # webContentControl
            'compliance': {'a': 'Compliance', 'b': 'Web content filtering enabled.', 'd': 'It was verified that category-wise web content filtering is implemented, blocking access to malicious or inappropriate websites.', 'f': 'Prevents access to harmful websites, malware, and content that can disrupt bank operations.', 'h': 'Maintain web content filtering for all users.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Web content filtering was disabled.', 'd': 'It was observed that web content filtering is not implemented, allowing users to access potentially harmful websites.', 'f': "As web-content filtering is not implemented, any user can access any website having pornography, hacking softwares, torrents, etc using the bank’s infrastructure. This will slow down the bank's production, as well as these websites contain malware as well as viruses that can infect bank's system and can  spread into the bank’s network infrastructure.", 'h': 'It is recommended to implement category-wise web content filtering. So, any harmful websites can not be accessed by user.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        25: {  # dashboardMonitoring
            'compliance': {'a': 'Compliance', 'b': 'Antivirus dashboard is monitored regularly for security events and alerts.', 'd': 'It was verified that the antivirus dashboard and logs are monitored regularly to track security events and alerts.', 'f': 'Ensures timely detection and response to security incidents.', 'h': 'Continue regular monitoring of antivirus dashboard and logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Antivirus dashboard or Antivirus logs were not available with the bank.', 'd': 'It was observed that the antivirus dashboard is not monitored regularly for security events.', 'f': 'If the antivirus dashboard is not monitored regularly then alerts of expired antivirus or alerts of suspicious activities will go unnoticed. Thus, the bank will not if the bank infrastructure is under cyber attack.', 'h': 'It is recommended to review antivirus dashboard regularly for security events.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        26: {  # ntpServer
            'compliance': {'a': 'Compliance', 'b': 'NTP server configured.', 'd': 'It was verified that an NTP server is configured, synchronizing system time across all devices accurately.', 'f': 'Supports precise event logging, forensic investigations, and accurate security monitoring.', 'h': 'Maintain NTP server configuration across all systems.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'NTP server was not configured.', 'd': 'It was observed that no NTP server is configured, causing potential time discrepancies across systems.', 'f': 'Inaccurate timestamps affect log analysis, forensic investigations, and security incident reconstruction.', 'h': 'Implement an NTP server to synchronize system clocks and maintain accurate logs.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        27: {  # centralizedLogging
            'compliance': {'a': 'Compliance', 'b': 'Logging enabled with centralized storage.', 'd': 'It was verified that logging is enabled and logs are stored on a centralized server for monitoring and forensic purposes.', 'f': 'Ensures traceability of security events and supports forensic investigations.', 'h': 'Maintain centralized logging and ensure logs are regularly reviewed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Logging was not enabled.', 'd': 'It was observed that logging is not enabled or logs are not stored centrally, leaving event tracking incomplete.', 'f': 'The absence of logging prevents the recording of user and system activities, making it difficult to detect, investigate, or respond to security incidents. Without proper logs, unauthorized access or malicious actions may go unnoticed, increasing the risk of data breaches and compliance violations.', 'h': 'Enable logging and store logs on a centralized server to support monitoring and forensic investigations.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        28: {  # notificationsEnabled
            'compliance': {'a': 'Compliance', 'b': 'Notifications enabled.', 'd': 'It was verified that antivirus notifications are enabled, providing alerts for virus detections and other security events.', 'f': 'Ensures early warning and timely response to threats.', 'h': 'Maintain notification settings enabled on all systems.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The notifications option was disabled.', 'd': 'It was observed that antivirus notifications are disabled, preventing alerts for virus detections and outbreaks.', 'f': 'IT team may not receive early warnings, delaying response to security incidents.', 'h': 'Enable notifications to alert IT personnel of critical antivirus events.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        29: {  # applicationControl
            'compliance': {'a': 'Compliance', 'b': 'Application control enabled.', 'd': 'It was verified that application control is enabled, preventing installation of unauthorized or malicious software.', 'f': 'Protects the bank\'s systems from malware, spyware, and other unauthorized applications.', 'h': 'Maintain application control enabled across all systems.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Application control was not enabled.', 'd': 'It was observed that application control is not enabled, allowing users to install potentially harmful software.', 'f': "If the user installs unauthorized, malicious software which contains spyware, keylogger, malware. It can compromise the security of bank's digital infrastructure.", 'h': 'Enable application control to restrict unauthorized software installations.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Apply formatting to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for col in range(1, 9):  # A to H
        cell = ws.cell(row=1, column=col)
        cell.border = thin_border
    
    # Process each question
    for i, question in enumerate(question_mapping, 2):  # Start from row 2
        # Get user input for this question
        user_input = form_data.get(question, 'not_applicable') if form_data else 'not_applicable'
        
        # Get response data
        response_data = question_responses.get(i-1, {})
        user_response = response_data.get(user_input, {})
        
        # Populate columns based on user input
        # Column C: Compliance/Non-Compliance/Not Applicable
        ws.cell(row=i, column=3, value=user_response.get('a', 'Not Applicable'))
        
        # Column D: Observation (Short/Brief)
        ws.cell(row=i, column=4, value=user_response.get('b', 'Not Applicable'))
        
        # Column F: Observation
        ws.cell(row=i, column=6, value=user_response.get('d', 'Not Applicable'))
        
        # Column G: Impact
        ws.cell(row=i, column=7, value=user_response.get('f', 'Not Applicable'))
        
        # Column H: Recommendation
        ws.cell(row=i, column=8, value=user_response.get('h', 'Not Applicable'))
        
        # Column E: Risk Factor with color coding
        risk_factor = risk_factors[i-2]  # risk_factors is 0-indexed
        risk_cell = ws.cell(row=i, column=5, value=risk_factor)
        risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
        risk_cell.fill = PatternFill(start_color=risk_colors.get(risk_factor, '808080'), end_color=risk_colors.get(risk_factor, '808080'), fill_type='solid')
        risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply formatting to all data rows
    for row in range(2, len(question_mapping) + 2):  # Rows 2 to 30
        for col in range(1, 9):  # Columns A to H
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Special alignment for column A (Sr. No.), C (Compliance status), and E (Risk Factor)
            if col == 1 or col == 3 or col == 5:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Set row height for wrapped text
            ws.row_dimensions[row].height = 30
    
    # Save file
    filename = "Antivirus Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    """
    Delete the generated Excel file after download
    """
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            print(f"File {filepath} deleted successfully")
    except Exception as e:
        print(f"Error deleting file {filepath}: {e}")
