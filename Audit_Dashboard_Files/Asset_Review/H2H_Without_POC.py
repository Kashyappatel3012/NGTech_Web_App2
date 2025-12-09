import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_h2h_excel(form_data=None):
    """
    Create Excel file for H2H Audit Assessment
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "H2H Audit"
    
    # Define questions
    questions = [
        "Is an administrative portal login accessed only by Whitelisted IP Addresses?",
        "Are Administrators having two factor authentications enabled?",
        "Are default credentials in use or Are accounts named by the 'Admin' or 'root' or 'Administrator' to access administrator login?",
        "Is the firewall Configured?",
        "Are APIs using encrypted/tunnel connections, any APIs permitted over unencrypted protocols like HTTP, FTP, or TELNET?",
        "Are UCBs ensuring to reconcile the transactions put through the sponsor bank mandatorily daily and preferably more than once a day to identify unauthorized transactions?",
        "Is 2FA being used for transaction authorization?",
        "Are communications with the sponsor bank done through bank domain email id?",
        "Is 2FA being used for CBS login?",
        "Are Below attached IOCs (Indicators of Compromise) blocked through Anti-Virus?",
        "Are Below attached IOCs (Indicators of Compromise) blocked through Firewall?",
        "Is the Security Advisor provided?",
        "Whether the DoS protection is enabled or not?",
        "Is the firmware/DSM up to date?",
        "Is Admin lockout set to 3 or more failed login attempts or as per the organization policy?",
        "Is the administrative password expiry set as per the organization's password policy?",
        "Is the administrative password set as per the organization's password policy?",
        "Is Bank implementing policies through GPOs for periodic password expiry, password age, and history?",
        "Has the bank employed whitelisting of executables/software?",
        "Is the dongle plugged in only during the time of signing the transactions or not?",
        "Is access to the dongle secured using a PIN/password?",
        "Is the dongle, when not in use, in the safe custody of the owner of the digital signing certificate and not left unattended at the terminal/H2H (Host to Host) client?",
        "Are Email Notifications enabled?"
    ]
    
    # Risk factors for each question
    risk_factors = [
        'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High',
        'High', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium',
        'Medium', 'Medium', 'Low'
    ]
    
    # Question mapping for form fields
    question_mapping = [
        'whitelistedIP', 'admin2FA', 'defaultCredentials', 'firewallConfigured', 'encryptedAPIs',
        'transactionReconciliation', 'transaction2FA', 'bankEmailDomain', 'cbs2FA', 'iocAntivirus',
        'iocFirewall', 'securityAdvisor', 'dosProtection', 'firmwareUpdated', 'adminLockout',
        'passwordExpiry', 'passwordPolicy', 'gpoPolicies', 'executableWhitelisting', 'dongleUsage',
        'donglePIN', 'dongleCustody', 'emailNotifications'
    ]
    
    # Set column widths
    ws.column_dimensions['A'].width = 10  # Sr. No.
    ws.column_dimensions['B'].width = 50  # Questions
    ws.column_dimensions['C'].width = 20  # Compliance/Non-Compliance/Not Applicable
    ws.column_dimensions['D'].width = 30  # Observation (Short/Brief)
    ws.column_dimensions['E'].width = 20  # Risk Factor
    ws.column_dimensions['F'].width = 50  # Observation
    ws.column_dimensions['G'].width = 50  # Impact
    ws.column_dimensions['H'].width = 50  # Recommendation
    
    # Header row
    headers = ['Sr. No.', 'Questionnaire/Points', 'Compliance/Non-Compliance/Not Applicable', 
               'Observation (Short/Brief)', 'Risk Factor', 'Observation', 'Impact', 'Recommendation']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Populate Sr. No. and Questions
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }
    
    # Question responses data
    question_responses = {
        1: {  # whitelistedIP
            'compliance': {'a': 'Compliance', 'b': 'Administrative portal accessible only from whitelisted IPs.', 'd': 'Access to the administrative console is restricted to a predefined set of whitelisted IP addresses belonging to authorized IT staff. Unauthorized IPs cannot reach the login page.', 'f': 'Reduces attack surface by blocking external or untrusted access attempts.', 'h': 'Continue enforcing and periodically reviewing IP whitelist rules to ensure accuracy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Administrative portal login accessed by IP address outside of whitelisted IP addresses.', 'd': 'It was observed that the administrative portal could be accessed from IPs not included in the whitelist. This means users or attackers outside the secured network perimeter can attempt direct login, increasing the exposure of the admin interface.', 'f': 'Without IP whitelisting, attackers can scan and attempt brute-force or credential-stuffing attacks on the admin portal from any location. Compromise of these credentials could grant full administrative access, allowing system takeover, data theft, or injection of malicious code. It also becomes difficult to distinguish between legitimate and suspicious admin logins.', 'h': 'Restrict access to the admin portal strictly via whitelisted IPs belonging to trusted administrative subnets or VPNs. Implement logging, alerting for access attempts from unauthorized IPs, and review whitelist entries regularly to remove obsolete addresses.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # admin2FA
            'compliance': {'a': 'Compliance', 'b': 'Administrator 2FA enabled.', 'd': 'All administrative accounts are protected with strong two-factor authentication mechanisms (e.g., OTP, hardware key, authenticator app).', 'f': 'Prevents unauthorized access even if credentials are compromised.', 'h': 'Maintain 2FA enforcement and review for all privileged accounts.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Administrators do not have two-factor authentications enabled.', 'd': 'It was observed that administrator accounts are only protected by username and password credentials without any additional authentication factor. This significantly weakens account protection against credential theft or brute-force attacks.', 'f': 'Without 2FA, stolen or guessed credentials immediately allow attackers full administrative access. Attackers can alter system configurations, disable security controls, and exfiltrate sensitive data. Compromise of an admin account can lead to total environment takeover and loss of service integrity.', 'h': 'Enforce two-factor authentication (OTP, push-based, or hardware token) for all admin accounts. Integrate 2FA into the IAM solution and enforce it across VPN, web portals, and critical systems.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # defaultCredentials
            'compliance': {'a': 'Compliance', 'b': 'Default credentials removed or renamed.', 'd': 'All default usernames have been renamed, and unique complex passwords are used per device/system.', 'f': 'Prevents attackers from exploiting known default credentials.', 'h': 'Maintain strong unique admin accounts and rotate credentials periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Default credentials were in use in the name of account with Admin.', 'd': 'It was observed that default usernames like "admin" or "root" were active and used for device or portal login, and some still retain default or weak passwords. This indicates credentials were never customized post-deployment.', 'f': 'Default usernames are publicly known, making brute-force or dictionary attacks trivial. Attackers can easily obtain access to administrative interfaces, modify configurations, and install malicious software. This exposes all connected systems to potential compromise and complete administrative control by external actors.', 'h': 'Immediately rename default accounts, enforce strong passwords, disable unused administrative IDs, and monitor all administrative login attempts. Regularly audit devices for any reversion to default credentials.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # firewallConfigured
            'compliance': {'a': 'Compliance', 'b': 'Firewall properly configured and rules reviewed.', 'd': 'Network firewall is configured with restrictive inbound/outbound policies, logging, and periodic rule audits.', 'f': 'Ensures network perimeter is protected from unauthorized or malicious traffic.', 'h': 'Continue periodic rule reviews and configuration backups.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The firewall was not configured properly.', 'd': 'It was observed that firewall filtering rules were weak or improperly defined, allowing unnecessary inbound/outbound connections and unmonitored service ports. Logs were also not effectively monitored for anomalies.', 'f': 'A misconfigured firewall exposes the internal network to intrusion, malware infection, and lateral movement. Attackers can exploit open services, bypass monitoring, and compromise the confidentiality, integrity, and availability (CIA) of banking systems. The entire branch network becomes susceptible to external attacks.', 'h': 'Reconfigure the firewall with least-privilege rules, block unused ports, enable logging and alerts for suspicious activity, and review configurations monthly or after any infrastructure changes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # encryptedAPIs
            'compliance': {'a': 'Compliance', 'b': 'APIs restricted to encrypted protocols only.', 'd': 'All APIs communicate over HTTPS, SSH, or TLS-based tunnels; insecure protocols are blocked at the firewall.', 'f': 'Safeguards sensitive data in transit and mitigates eavesdropping.', 'h': 'Maintain TLS configurations and disable weak cipher suites.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'APIs were re not using encrypted/tunnel connection, Application or API permitted over unencrypted protocols like HTTP, FTP, or TELNET.', 'd': 'It was observed that APIs and applications are accessible through unencrypted protocols (HTTP/FTP/Telnet) instead of secure channels. Data transmissions occur in plaintext, exposing credentials and other sensitive information.', 'f': 'Unencrypted API communication allows attackers to intercept, read, or modify data packets via sniffing or MITM attacks. This compromises confidential financial data and can lead to credential theft, fraud, or system manipulation. It violates standard security baselines (e.g., RBI/CERT-In/ISO 27001).', 'h': 'Disable unencrypted protocols entirely. Configure all APIs and interfaces to enforce HTTPS/TLS and SSH connections only. Deploy certificates from trusted CAs and perform regular vulnerability scans to ensure encryption enforcement.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # transactionReconciliation
            'compliance': {'a': 'Compliance', 'b': 'Daily reconciliation performed and verified.', 'd': 'All sponsor bank transactions are reconciled daily, with exceptions reviewed immediately.', 'f': 'Detects unauthorized transactions and ensures integrity of financial data.', 'h': 'Continue enforcing mandatory daily reconciliation and exception logging.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'UCBs never try to identify the unauthorized transactions that are done through sponsor banks on a daily basis.', 'd': 'It was observed that UCBs do not perform daily or intra-day reconciliation with the sponsor bank. Unauthorized or erroneous transactions remain undetected for extended periods.', 'f': 'Failure to reconcile daily can result in unnoticed fraudulent or duplicate transactions, delayed detection of financial discrepancies, and potential monetary loss. This can also breach regulatory obligations and impact customer trust and audit scores.', 'h': 'Implement automated reconciliation tools to compare sponsor bank transactions daily and raise alerts for mismatches. Conduct manual verification for anomalies and record daily reconciliation reports for audit review.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # transaction2FA
            'compliance': {'a': 'Compliance', 'b': '2FA enforced for transaction authorization.', 'd': 'All transactions require secondary authentication (OTP/token/app approval) before execution.', 'f': 'Strengthens user verification and prevents fraudulent transfers.', 'h': 'Continue enforcing 2FA on all transactional systems.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Two-factor authentications was not available for transaction authorization.', 'd': 'It was observed that transactions are executed after single-factor authentication (username/password), without secondary verification.', 'f': 'Without 2FA, attackers gaining user credentials can perform unauthorized transactions and financial fraud without detection. The absence of multi-factor controls undermines user authenticity and can lead to financial, legal, and reputational damage to the bank.', 'h': 'Enforce 2FA for all financial transactions through OTP, app-based approval, or hardware tokens. Integrate transaction-level verification in CBS or payment gateway systems.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # bankEmailDomain
            'compliance': {'a': 'Compliance', 'b': 'Communications through official bank domain only.', 'd': 'All official correspondence with sponsor banks occurs via bank\'s secured domain.', 'f': 'Prevents phishing and ensures message authenticity.', 'h': 'Continue using official mail domains and enforce SPF/DKIM/DMARC.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Public Mail domain was in use.', 'd': 'It was observed that bank personnel use public domain (e.g., @gmail.com) for communication with sponsor banks.', 'f': 'If communications with the sponsor bank are not conducted through the official bank domain email ID, there is an increased risk of data interception, phishing, and impersonation attacks. Using non-official or public email domains reduces the authenticity and integrity of communication, potentially leading to unauthorized disclosure of sensitive financial information and regulatory non-compliance.', 'h': 'It is recommended that all communications with the sponsor bank be carried out exclusively through the bank’s official domain email IDs to ensure authenticity and security. This helps maintain confidentiality, integrity, and traceability of official correspondence.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # cbs2FA
            'compliance': {'a': 'Compliance', 'b': 'CBS login protected by 2FA.', 'd': 'Core Banking System users authenticate using password plus OTP/token verification.', 'f': 'Enhances CBS access security and mitigates credential theft.', 'h': 'Continue enforcing 2FA for CBS login with periodic review.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Two-factor authentication was not available for CBS.', 'd': 'It was observed that CBS accounts are accessible with only username and password credentials without second-factor verification, weakening authentication.', 'f': 'Unauthorized access to CBS can result in financial data manipulation, fraudulent fund transfers, and leakage of sensitive customer data. Attackers exploiting stolen credentials can control transactions and compromise banking integrity.', 'h': 'Enable 2FA for CBS login using secure OTP or app-based verification methods. Integrate with IAM and monitor login attempts for abnormal behavior.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # iocAntivirus
            'compliance': {'a': 'Compliance', 'b': 'IOCs successfully blocked by Anti-Virus.', 'd': 'The Anti-Virus solution is updated and actively blocking all attached IOCs during periodic scans, ensuring effective malware defense.', 'f': 'Helps prevent malicious code execution and ensures systems remain protected from known threats.', 'h': 'Continue maintaining updated threat signatures and monitoring IOC detections.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'IOCs (Indicators of Compromise) was not blocked by Anti-Virus.', 'd': 'It was observed that the Anti-Virus system failed to detect or block known IOCs, indicating outdated signatures or disabled protection modules on several endpoints.', 'f': 'Failure to block IOCs can lead to malware infections, data breaches, or ransomware attacks. Attackers can exploit unprotected systems to exfiltrate data, spread infections laterally, and compromise critical servers.', 'h': 'Update Anti-Virus definitions immediately and enable real-time scanning. Ensure all endpoints report to a centralized console for IOC monitoring and automated quarantine of infected files.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # iocFirewall
            'compliance': {'a': 'Compliance', 'b': 'Firewall rules effectively blocking IOCs.', 'd': 'Firewall configurations were verified and found to be successfully blocking the listed IOCs through IP, domain, and port-level filtering.', 'f': 'Mitigates external threat communication and command-and-control traffic.', 'h': 'Continue maintaining updated IOC block lists and review firewall rules regularly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'IOCs (Indicators of Compromise) was not blocked through Firewall.', 'd': 'It was observed that firewall filters were not updated with the latest IOCs, allowing potential connections to known malicious IPs or domains.', 'f': 'Unblocked IOCs can allow malware to communicate with external threat actors, leading to data leakage, backdoor access, and persistent infections within the bank\'s network.', 'h': 'Update the firewall IOC list regularly. Implement automated feeds for threat intelligence and enforce strict outbound blocking for known malicious indicators.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # securityAdvisor
            'compliance': {'a': 'Compliance', 'b': 'Security Advisor is provided and in use.', 'd': 'The Security Advisor tool is deployed and utilized for periodic review of vulnerabilities and system hardening recommendations.', 'f': 'Improves overall system posture by continuous monitoring and advisory updates.', 'h': 'Continue using Security Advisor for timely remediation and security enhancement.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Security Advisor provided but not used.', 'd': 'It was observed that the Security Advisor utility, though provided, is not actively used by IT staff for monitoring or implementing suggested actions.', 'f': 'Neglecting Security Advisor insights can leave the environment vulnerable to unpatched risks and missed security improvements. This reduces proactive threat detection and weakens the defense against emerging attacks.', 'h': 'It is recommended to actively use Security Advisor for vulnerability review and configuration management. Ensure that all identified issues are addressed as per priority and track closure for compliance.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # dosProtection
            'compliance': {'a': 'Compliance', 'b': 'DoS protection enabled.', 'd': 'Firewall and network security systems have DoS and DDoS protection mechanisms actively configured to mitigate attack attempts.', 'f': 'Prevents service interruptions and network resource exhaustion.', 'h': 'Maintain and review DoS configurations regularly for optimal protection.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'DoS protection disabled.', 'd': 'It was observed that DoS protection is disabled on network devices, leaving critical services exposed to flooding or resource exhaustion attacks.', 'f': 'Without DoS protection, the bank\'s systems may become inaccessible during attacks, disrupting operations, online banking, and ATM services. It can cause denial of legitimate access and financial loss.', 'h': 'Enable DoS protection on all edge devices and firewalls. Regularly update firmware and test configurations against simulated attacks to ensure resilience.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # firmwareUpdated
            'compliance': {'a': 'Compliance', 'b': 'Firmware is up to date.', 'd': 'All network and security devices have the latest firmware installed with vendor-released patches applied timely.', 'f': 'Minimizes exploitable vulnerabilities and enhances system performance.', 'h': 'Continue maintaining patch compliance and periodic firmware updates.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Firmware was not updated.', 'd': 'It was observed that several devices are running outdated firmware versions missing critical patches and bug fixes.', 'f': 'The vulnerabilities which are already present in the outdated firmware can be led to exploitation and the attacker can be able to perform malicious activity.As firmware carries out the integral functions of hardware, firmware updates bring some alterations in the program, which are necessary to enable the corresponding devices to operate proficiently as well as to fix the bugs for better security. As, the firmware was not updated, an attacker can exploit the bug and interrupt the network.  ', 'h': 'Update all device firmware to the latest stable version. Establish a regular firmware update cycle and maintain documentation for compliance verification.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # adminLockout
            'compliance': {'a': 'Compliance', 'b': 'Admin lockout policy configured.', 'd': 'Account lockout threshold is configured to trigger after three failed login attempts as per the security policy.', 'f': 'Mitigates brute-force and password guessing attacks effectively.', 'h': 'Continue enforcing account lockout as per policy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Account lockout duration and failed login attempts were not set.', 'd': 'It was observed that no account lockout policy exists for administrative users, allowing unlimited login attempts.', 'f': 'Brute force password attacks can be automated to try thousands or even millions of password combinations for any or all user accounts. Limiting the number of failed sign-ins that can be performed nearly eliminates the effectiveness of such attacks. However, it is important to note that a denial-of-service (DoS) attack could be performed on a domain that has an account lockout threshold configured. A malicious user could programmatically attempt a series of password attacks against all users in the organization. If the number of attempts is greater than the value of the Account lockout threshold, the attacker could potentially lock every account.', 'h': 'Set account lockout threshold to a maximum of three failed attempts or per policy. Configure alerting mechanisms for repeated login failures to detect brute-force patterns.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # passwordExpiry
            'compliance': {'a': 'Compliance', 'b': 'Password expiry configured as per policy.', 'd': 'Administrative passwords are configured to expire within the defined organizational policy period (45–60 days).', 'f': 'Reduces risk of credential theft and unauthorized prolonged access.', 'h': 'Continue enforcing periodic password expiry as per the organization\'s password policy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Administrative password expiry was not set as per the organization's password policy.", 'd': 'It was observed that administrative account passwords do not have expiry settings, allowing indefinite use of the same password.', 'f': 'Long-term reuse of passwords increases the risk of compromise from leaked or guessed credentials, enabling persistent unauthorized access.', 'h': "Enforce administrative password expiry in accordance with the organization's password policy. Implement password change reminders and integrate the process with centralized identity management for effective compliance monitoring."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # passwordPolicy
            'compliance': {'a': 'Compliance', 'b': 'Password complexity policy enforced.', 'd': 'Administrative accounts follow strong password standards, including alphanumeric and special characters, meeting organizational security requirements.', 'f': 'Prevents easy password guessing and brute-force attacks.', 'h': 'Continue enforcing password complexity for administrative accounts.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Passwords were not as per the standard password policy.', 'd': 'It was observed that administrative accounts use weak passwords not meeting the minimum complexity defined in the organization\'s policy.', 'f': 'Weak passwords can be easily cracked, giving attackers administrative access to critical systems, leading to configuration tampering and data compromise.', 'h': "It is recommended that the administrative password should be set as per the organization's password policy."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        18: {  # gpoPolicies
            'compliance': {'a': 'Compliance', 'b': 'Password GPO policies implemented.', 'd': 'Group Policy Objects (GPOs) enforce periodic password expiry, history, and minimum password age across the domain.', 'f': 'Ensures uniform password control and minimizes password reuse.', 'h': 'Continue maintaining GPO enforcement for password policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Password expiry, password age, and history policies were not configured through GPOs.', 'd': 'It was observed that password expiry, history, and minimum age are not enforced via GPO, allowing inconsistent password practices across systems.', 'f': 'Absence of centralized password enforcement increases risk of weak or reused passwords, making brute-force and credential reuse attacks more likely.', 'h': 'Configure GPOs to enforce password expiry between 30–90 days and maintain at least 24 password histories. Verify enforcement through periodic audits.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        19: {  # executableWhitelisting
            'compliance': {'a': 'Compliance', 'b': 'Application whitelisting implemented.', 'd': 'Only approved applications are permitted to execute on bank systems using an application control mechanism.', 'f': 'Prevents execution of unauthorized or malicious software.', 'h': 'Continue maintaining application whitelisting and update the list periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The bank has not employed whitelisting of executables/software.', 'd': 'It was observed that executable or software whitelisting has not been implemented, allowing unverified applications to run on critical systems.', 'f': "If the executables/software are not whitelisted the bank will not be able to protect computers and networks from potentially harmful applications, which can compromise the bank's system.Application whitelisting is the practice of specifying an index of approved software applications that are permitted to be present and active on a computer system.", 'h': 'It is recommended that IT organization should use technologies that are built into the host operating system or leverage the capabilities of antivirus to implement the whitelisting of applications.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        20: {  # dongleUsage
            'compliance': {'a': 'Compliance', 'b': 'Dongle connected only during transaction signing.', 'd': 'Digital signing dongle is used strictly for transaction authorization and removed immediately afterward.', 'f': 'Prevents unauthorized or fraudulent use of the dongle.', 'h': 'Continue following secure dongle usage practices during transactions only.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The dongle used for digitally signing the transactions was always connected.', 'd': 'It was observed that the digital signing dongle remains connected to the terminal continuously even when not in use.', 'f': 'Always-connected dongles can be misused by unauthorized users or malware to perform illegal or fraudulent digital signing.', 'h': 'Ensure dongles are plugged in only during transaction signing and stored securely afterward. Implement automatic session disconnection when idle.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        21: {  # donglePIN
            'compliance': {'a': 'Compliance', 'b': 'Dongle access protected by PIN/password.', 'd': 'Each dongle used for digital signing is configured with a PIN or password for secure access.', 'f': 'Prevents unauthorized dongle use and strengthens transaction integrity.', 'h': 'Continue enforcing PIN/password protection for dongles.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Access to the dongle was not secured using a PIN/password.', 'd': 'It was observed that the digital signing dongle lacks any PIN/password protection, allowing unrestricted access.', 'f': 'Unprotected dongles can be misused for unauthorized transactions, document signing, or fraudulent digital activities without traceability.', 'h': 'Configure dongles with strong PIN/password protection and maintain access logs for accountability.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        22: {  # dongleCustody
            'compliance': {'a': 'Compliance', 'b': 'Dongle stored securely under owner\'s custody.', 'd': 'It was verified that all dongles are properly secured with authorized owners when not in use.', 'f': 'Prevents unauthorized dongle use and ensures accountability.', 'h': 'Continue maintaining strict custody of digital signing dongles.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The dongle was not in the safe custody of the owner of the digital signing certificate.', 'd': 'It was observed that dongles were left connected or unattended at terminals, not kept in the safe custody of the respective certificate owners.', 'f': 'Exposed dongles can be used by unauthorized individuals to digitally sign fraudulent transactions or documents, causing regulatory non-compliance.', 'h': 'Keep dongles in locked storage or under personal custody of the certificate owner when not in use. Implement physical access restrictions for terminals.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        23: {  # emailNotifications
            'compliance': {'a': 'Compliance', 'b': 'Email alerts enabled.', 'd': 'Email notifications are configured to alert administrators on security, login, and transaction events.', 'f': 'Facilitates timely response to suspicious activities.', 'h': 'Continue monitoring email alerts and maintain escalation procedures.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Email notifications were not enabled.', 'd': 'It was observed that email notifications are not enabled on the system, preventing admins from receiving alerts about security events or unauthorized access.', 'f': 'Without email notifications, critical security incidents may go unnoticed, delaying response and allowing prolonged unauthorized activity or fraud.', 'h': 'Enable email notifications for system alerts, login failures, and configuration changes. Verify delivery using admin test alerts and maintain logs for audit.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Apply formatting to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for col in range(1, 9):  # A to H
        cell = ws.cell(row=1, column=col)
        cell.border = thin_border
    
    # Process each question
    for i, question in enumerate(question_mapping, 2):  # Start from row 2
        # Get user input for this question
        user_input = form_data.get(question, 'not_applicable') if form_data else 'not_applicable'
        
        # Get response data
        response_data = question_responses.get(i-1, {})
        user_response = response_data.get(user_input, {})
        
        # Populate columns based on user input
        # Column C: Compliance/Non-Compliance/Not Applicable
        ws.cell(row=i, column=3, value=user_response.get('a', 'Not Applicable'))
        
        # Column D: Observation (Short/Brief)
        ws.cell(row=i, column=4, value=user_response.get('b', 'Not Applicable'))
        
        # Column F: Observation
        ws.cell(row=i, column=6, value=user_response.get('d', 'Not Applicable'))
        
        # Column G: Impact
        ws.cell(row=i, column=7, value=user_response.get('f', 'Not Applicable'))
        
        # Column H: Recommendation
        ws.cell(row=i, column=8, value=user_response.get('h', 'Not Applicable'))
        
        # Column E: Risk Factor with color coding
        risk_factor = risk_factors[i-2]  # risk_factors is 0-indexed
        risk_cell = ws.cell(row=i, column=5, value=risk_factor)
        risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
        risk_cell.fill = PatternFill(start_color=risk_colors.get(risk_factor, '808080'), end_color=risk_colors.get(risk_factor, '808080'), fill_type='solid')
        risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply formatting to all data rows
    for row in range(2, len(question_mapping) + 2):  # Rows 2 to 24
        for col in range(1, 9):  # Columns A to H
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Special alignment for column A (Sr. No.), C (Compliance status), and E (Risk Factor)
            if col == 1 or col == 3 or col == 5:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Set row height for wrapped text
            ws.row_dimensions[row].height = 30
    
    # Save file
    filename = "H2H Audit Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    """
    Delete the generated Excel file after download
    """
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            print(f"File {filepath} deleted successfully")
    except Exception as e:
        print(f"Error deleting file {filepath}: {e}")