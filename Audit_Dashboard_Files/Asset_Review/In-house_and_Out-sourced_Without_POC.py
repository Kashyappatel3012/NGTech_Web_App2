import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_inhouse_outsourced_excel(form_data=None):
    """
    Create In-house and Out-sourced Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "In-house and Out-sourced"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # In-house and Out-sourced Questions
    questions = [
        "Is the software designed using pre-defined formats at three levels: program level, application level, and organization level?",
        "Has the IT department adopted any standardized quality processes such as ISO, SEI CMM, etc., for software development?",
        "Is software tested for quality assurance?",
        "Are adequate input validation checks built into data entry programs?",
        "Does the development process consider security requirements as per the approved security policy?",
        "Is every patch/update authorized by a competent authority?",
        "Is the quality assurance team separate from the development team?",
        "Are data/test results preserved for future reference?",
        "Are well-established testing procedures in place?"
    ]

    # Risk Factors
    risk_factors = [
        "High",
        "High", 
        "High",
        "High",
        "High",
        "High",
        "Medium",
        "Medium",
        "Medium"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "ioPreDefinedFormats": 1,
        "ioStandardizedQualityProcesses": 2,
        "ioSoftwareTestedQA": 3,
        "ioInputValidationChecks": 4,
        "ioSecurityRequirements": 5,
        "ioPatchUpdateAuthorization": 6,
        "ioQATeamSeparate": 7,
        "ioTestDataPreserved": 8,
        "ioTestingProcedures": 9
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Question responses data
    question_responses = {
        1: {  # ioPreDefinedFormats
            'compliance': {'a': 'Compliance', 'b': 'Pre-defined formats used.', 'd': 'Software development follows established formats at program, application, and organizational levels, promoting consistency and structured development.', 'f': 'Ensures maintainable, high-quality software that aligns with organizational objectives and eases integration.', 'h': 'Periodically review design templates and update them to reflect evolving best practices.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The software was not designed using pre-designed formats at three levels.', 'd': 'The software development process does not consistently follow standardized design formats at program, application, or organization levels.', 'f': 'If the developer does not use a predefined format for software design, a developer cannot meet the customer requirements. The software might not work properly because it is not designed properly which can be root cause of crtical bugs  which might have severe security implications.', 'h': 'Adopt and enforce pre-defined design formats at all levels to ensure consistency, maintainability, and better alignment with organizational standards.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # ioStandardizedQualityProcesses
            'compliance': {'a': 'Compliance', 'b': 'Standardized quality processes implemented.', 'd': 'IT development processes comply with recognized standards such as ISO or SEI CMM, ensuring consistent software quality and process maturity.', 'f': 'Reduces errors, enhances software reliability, and aligns IT processes with organizational and regulatory expectations.', 'h': 'Conduct periodic audits of quality processes to ensure adherence and continuous improvement.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The IT team was not following quality processes such as ISO, SEI, and CMM.', 'd': 'The IT department does not follow any recognized software quality framework like ISO, SEI CMM, or similar standards.', 'f': 'If the standards such as ISO, and SEI are not followed then the software may be lacking essential qualities like functionality, Reliability, Usability, Efficiency, Maintainability, and Portability which can hinder the productivity of the bank, and also some bugs can be exploited for malicious purposes.', 'h': 'Adopt industry-standard quality frameworks and integrate them into the software development lifecycle for consistent quality and reliability.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # ioSoftwareTestedQA
            'compliance': {'a': 'Compliance', 'b': 'Software tested for QA.', 'd': 'All software undergoes comprehensive QA testing to validate functionality, security, and performance before deployment.', 'f': 'Ensures reliable software, reduces post-deployment defects, and maintains organizational efficiency.', 'h': 'Periodically update testing protocols and maintain detailed test reports for audit purposes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The software was not tested for quality assurance.', 'd': 'Software applications are released without adequate quality assurance testing, or testing is limited and inconsistent.', 'f': "If the software is not tested for quality assurance then the existing bugs or broken functionality can affect the bank's productivity, which can cause financial damage to banks, and also impact the security posture of the bank.", 'h': 'It is recommended to perform quality assurance testing on software to find bugs and fix them as soon as possible so that it does affect the productivity, and security posture of the bank.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # ioInputValidationChecks
            'compliance': {'a': 'Compliance', 'b': 'Input validation implemented.', 'd': 'Data entry programs include comprehensive input validation checks, preventing invalid or malicious data from entering the system.', 'f': 'Enhances data integrity, improves system reliability, and mitigates security risks.', 'h': 'Periodically review and update validation rules to address emerging threats and maintain data accuracy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'There were no adequate input validation checks built into the data entry program.', 'd': 'Data entry programs lack proper input validation, allowing incorrect, malicious, or incomplete data to be processed.', 'f': 'If there are no input validations checks built into the data entry programs, then the attackers can inject malicious payloads using special characters. If the user enters mismatched data then the application will produce error-prone data which will impact the banks day to day business.', 'h': 'Implement rigorous input validation checks to ensure correctness, completeness, and security of data entered into the system.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # ioSecurityRequirements
            'compliance': {'a': 'Compliance', 'b': 'Security requirements incorporated.', 'd': 'Software development incorporates security policies from planning through deployment, including authentication, authorization, and encryption measures.', 'f': 'Reduces vulnerabilities, ensures compliance, and protects sensitive organizational data.', 'h': 'Conduct periodic security audits of software applications to verify compliance with security policies.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Development of the software was not done according to the security requirement of the approved security policy.', 'd': 'Software development does not consistently incorporate security requirements defined in the organization\'s security policy.', 'f': "If the security requirements are not considered as per the approved security policy then the software may pose a security risk to the whole infrastructure. The software might have multiple security vulnerabilities, which may make software unsafe to use because it does not adhere to the bank's security policy. ", 'h': 'Integrate security requirements into every phase of the software development lifecycle and perform security testing prior to deployment.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # ioPatchUpdateAuthorization
            'compliance': {'a': 'Compliance', 'b': 'Patches/updates authorized.', 'd': 'All software patches and updates are approved by competent authority, tested, and properly documented before deployment.', 'f': 'Ensures controlled and safe application updates, reducing operational and security risks.', 'h': 'Maintain a change log and conduct periodic reviews to verify adherence to authorization procedures.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Patch/update was not authorized by a competent authority.', 'd': 'Software patches or updates are applied without formal approval from the competent authority or documentation of authorization.', 'f': 'If the patches are not authorized by a competent authority, then patches may contain malware or malicious program which can impact the Confidentiality, Integrity, and Availablity of the system.', 'h': 'Establish a formal patch management process requiring prior authorization, testing, and documentation for all updates.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # ioQATeamSeparate
            'compliance': {'a': 'Compliance', 'b': 'QA team independent.', 'd': 'The QA team is organizationally and operationally independent from the development team, ensuring unbiased testing and evaluation.', 'f': 'Improves software reliability, detects defects effectively, and enhances overall quality assurance.', 'h': 'Review QA independence periodically to maintain objectivity and ensure high-quality testing outcomes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The quality assurance team was not different from the developer team.', 'd': 'Quality assurance is performed by the same team that develops the software, creating a potential conflict of interest.', 'f': 'If the developers are part of the quality assurance team then, some bugs might be present in the application which may get overlooked. A third-party tester or person can look at the application from a different perspective to uncover hidden bugs.', 'h': 'Establish an independent QA team separate from development to ensure objective evaluation and unbiased testing of software.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # ioTestDataPreserved
            'compliance': {'a': 'Compliance', 'b': 'Test data/results preserved.', 'd': 'All test results and relevant data are stored securely for future reference, with proper documentation for audits and quality reviews.', 'f': 'Provides traceability, supports continuous improvement, and strengthens audit readiness.', 'h': 'Review storage and retention policies periodically to ensure ongoing availability and integrity of test data.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Data/test results were not preserved for future reference.', 'd': 'Test results and related data are discarded or not stored systematically, preventing future review or audit.', 'f': 'Previous test results can help to debug applications and resolve bugs to produce quality software. If the developers do not have test results then it will be difficult for them to analyze the root cause of the error, thus making it difficult to solve the issue to produce high-quality software.', 'h': 'Maintain structured storage of test data and results with proper versioning, retention policies, and access controls.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # ioTestingProcedures
            'compliance': {'a': 'Compliance', 'b': 'Testing procedures formalized.', 'd': 'Standardized testing procedures are documented and implemented, ensuring thorough verification of software functionality, security, and performance.', 'f': 'Enhances software quality, reduces post-deployment defects, and ensures compliance with organizational standards.', 'h': 'Periodically review and update testing procedures to incorporate best practices and address emerging risks.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Well-established testing procedures were not being followed.', 'd': 'Testing procedures are ad-hoc, inconsistent, or not well-documented, leading to incomplete or ineffective testing.', 'f': 'If the testing procedure is not well-established then some bugs can be missed due to improper testing procedures which can delay the production of software, and cause finacial, and productivity loss to the bank.', 'h': 'Develop and enforce structured testing procedures covering unit, integration, system, and security testing, with clear documentation.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }

    # Populate data based on user input
    for i, question in enumerate(questions, 2):
        # Get user input for this question
        question_num = i - 1
        user_input = None
        
        if form_data:
            # Find the corresponding form field
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "In-house and Out-sourced Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
