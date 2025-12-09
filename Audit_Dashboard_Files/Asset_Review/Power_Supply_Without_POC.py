import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_power_supply_excel(form_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Power Supply"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # Power Supply Questions
    questions = [
        "Is there a separate enclosure and locking arrangement for the UPS?",
        "Does the maintenance agency provide battery service regularly?",
        "Does the UPS function properly when electricity fails?",
        "Is there a regular contract for maintenance of the UPS, and is preventive maintenance carried out as per the contract?",
        "Is a record of the tests undertaken maintained to verify the satisfactory functioning of the UPS?"
    ]

    # Risk Factors
    risk_factors = [
        "High", "High", "Medium", "Medium", "Medium"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "psSeparateEnclosure": 1,
        "psBatteryService": 2,
        "psFunctionProperly": 3,
        "psMaintenanceContract": 4,
        "psTestRecords": 5
    }

    # Question responses data
    question_responses = {
        1: {  # psSeparateEnclosure
            'compliance': {'a': 'Compliance', 'b': 'UPS secured in a locked enclosure.', 'd': 'The UPS is placed within a dedicated and locked enclosure, ensuring that only authorized personnel have access.', 'f': 'Enhances physical security, minimizes tampering risk, and ensures uninterrupted power availability.', 'h': 'Continue periodic inspections to ensure locks and enclosures remain secure and functional.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'UPS not secured in a separate enclosure.', 'd': 'The UPS system is installed in an open area without a dedicated enclosure or locking mechanism, making it vulnerable to unauthorized access or accidental interference.', 'f': 'If there is no separate enclosure and locking arrangement for UPS. Then the UPS system might suffer from over heating due to not having separate enclosure and if the room is not locked any unauthorized person can damage the UPS system.', 'h': 'Install the UPS in a separate, restricted enclosure with proper locking arrangements and access logs to prevent unauthorized handling.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # psBatteryService
            'compliance': {'a': 'Compliance', 'b': 'Battery service conducted as per schedule.', 'd': 'The maintenance agency performs timely battery health checks and servicing in accordance with the contract.', 'f': 'Ensures consistent UPS performance and minimizes risks of power disruption during outages.', 'h': 'Continue maintaining servicing records and monitor battery health parameters regularly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The maintenance agency does not provide battery service regularly.', 'd': 'The maintenance agency does not perform scheduled UPS battery servicing or inspection as per the agreement, leading to potential degradation in battery performance.', 'f': 'When maintenance agency service does not provide battery service regularly, the working time of the UPS will be reduced and the failure rate will be increased leading to loss of necessary power backup during an outage.', 'h': 'Enforce a strict maintenance schedule with the agency and maintain service records for accountability.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # psFunctionProperly
            'compliance': {'a': 'Compliance', 'b': 'UPS operates seamlessly during power outages.', 'd': 'The UPS provides uninterrupted power to critical systems during electricity failures, ensuring smooth operations.', 'f': 'Maintains continuity of IT and operational processes during power disruptions.', 'h': 'Continue regular testing and monitoring to ensure reliable UPS performance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'UPS does not functioning properly during power failure.', 'd': 'The UPS failed to supply uninterrupted power during electricity outages, indicating potential faults or improper battery capacity.', 'f': "If UPS is not functioning properly, there will be no power guarantee of the core business system of head office or provincial branches of the bank, which process a high volume of transactions. If there's no UPS/ power backup, the bank will not be able to process transactions. Due to a power outage, productivity will hamper, and the bank will suffer from financial loss.", 'h': 'Conduct a full-load testing of UPS, inspect battery capacity, and ensure timely maintenance and component replacement.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # psMaintenanceContract
            'compliance': {'a': 'Compliance', 'b': 'UPS maintenance carried out under valid contract.', 'd': 'The UPS is covered under a valid AMC, and preventive maintenance is performed as per schedule.', 'f': 'Ensures high system availability and minimizes risk of power-related failures.', 'h': 'Retain AMC documentation and verify completion of maintenance activities periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'There was no regular contract for maintenance of the UPS and thus preventive maintenance was not carried out.', 'd': 'The organization lacks an annual maintenance contract (AMC) for the UPS, or preventive maintenance activities are not being carried out per the defined schedule.', 'f': 'If there is no regular contract for maintenance of the UPS and thus preventive maintenance is not carried. Then the UPS will not work properly and there is a possibility of immediate deterioration.', 'h': 'Establish or renew AMC with the vendor, defining preventive maintenance schedules and performance benchmarks.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # psTestRecords
            'compliance': {'a': 'Compliance', 'b': 'UPS testing records properly maintained.', 'd': 'All UPS performance and battery load tests are documented and maintained systematically for audit and verification purposes.', 'f': 'Provides traceability and assurance of UPS reliability.', 'h': 'Continue documenting all test results and review them quarterly for performance consistency.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A record of the tests undertaken was not maintained.', 'd': 'The organization does not maintain logs or records of UPS testing, making it difficult to verify performance history and reliability.', 'f': "If the record of the tests is not maintained as per the contract, there will be conflict regarding the accountability, if UPS malfunctions. Also, if the bank has test records they will be well aware of the Service Quality of UPS.", 'h': 'Maintain a dedicated register or digital log for all UPS testing, including test dates, outcomes, and corrective actions taken.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:  # Only columns A through H (1-8)
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Power Supply UPS Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
