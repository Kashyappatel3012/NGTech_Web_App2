import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_atm_excel(form_data=None):
    """
    Create Excel file for ATM Assessment
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "ATM"
    
    # Define questions
    questions = [
        "Are there any physical, local attacks that target account data/card readers?",
        "Are there any physical, local attacks that target PINs?",
        "Is the access to sensitive areas and resources in the cabinet restricted or not?",
        "Whether the EPP has valid approval or not?",
        "Whether visual observation of PIN values is possible as the cardholder enters them?",
        "Does the Bank have a policy for the Personal Identification Numbers, used by various sets of customers who access the Banks systems directly using channels like ATMs?",
        "Are there CCTV cameras installed in the proper position?",
        "Whether the USB access is disabled or not?",
        "Whether the OS is up-to-date with patches or not?",
        "Whether the expired or unlicensed version of OS is used?",
        "Is there any unwanted software installed in the ATM?",
        "Whether the installation of rogue software is possible or not?",
        "Whether the logs are available or not?",
        "Whether the ATM is grouted to a wall, pillar, floor, etc or not?",
        "Whether the Cables of ATM and CCTV are concealed or not?",
        "Are the network ports openly accessible?",
        "Whether the ATM is network-segmented or not?",
        "Whether the ATM software is up-to-date with patches or not?",
        "Whether the unauthorized change in logs is possible or not?",
        "Whether the licensed Antivirus is available and up-to-date with the latest database or not?",
        "Do the ATM cabinet properly locked and has two PINs from two different personnel to open the safe?",
        "Whether the preventive maintenance is performed with a defined periodicity or not?",
        "Whether the Do's and Don'ts for the user awareness available or not?",
        "Whether the guard is available during non-banking hours or not?"
    ]
    
    # Risk factors for each question
    risk_factors = [
        'Critical', 'Critical', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'High', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low'
    ]
    
    # Question mapping for form fields
    question_mapping = [
        'physicalAttacksCardReaders', 'physicalAttacksPINs', 'sensitiveAreasRestricted', 'eppValidApproval', 'visualObservationPIN',
        'pinPolicy', 'cctvCamerasInstalled', 'usbAccessDisabled', 'osUpToDate', 'expiredOS',
        'unwantedSoftware', 'rogueSoftwareInstallation', 'logsAvailable', 'atmGrouted', 'cablesConcealed',
        'networkPortsAccessible', 'atmNetworkSegmented', 'atmSoftwareUpdated', 'unauthorizedLogChanges', 'licensedAntivirus',
        'cabinetDualPINs', 'preventiveMaintenance', 'userAwareness', 'guardAvailable'
    ]
    
    # Set column widths
    ws.column_dimensions['A'].width = 10  # Sr. No.
    ws.column_dimensions['B'].width = 50  # Questions
    ws.column_dimensions['C'].width = 20  # Compliance/Non-Compliance/Not Applicable
    ws.column_dimensions['D'].width = 30  # Observation (Short/Brief)
    ws.column_dimensions['E'].width = 20  # Risk Factor
    ws.column_dimensions['F'].width = 50  # Observation
    ws.column_dimensions['G'].width = 50  # Impact
    ws.column_dimensions['H'].width = 50  # Recommendation
    
    # Header row
    headers = ['Sr. No.', 'Questionnaire/Points', 'Compliance/Non-Compliance/Not Applicable', 
               'Observation (Short/Brief)', 'Risk Factor', 'Observation', 'Impact', 'Recommendation']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Populate Sr. No. and Questions
    for i, question in enumerate(questions, 2):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions
    
    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }
    
    # Question responses data
    question_responses = {
        1: {  # physicalAttacksCardReaders
            'compliance': {'a': 'Compliance', 'b': 'No skimming devices detected.', 'd': 'It was verified that ATMs were free from skimming devices or bugs, ensuring card data is not intercepted by unauthorized parties.', 'f': 'Prevents unauthorized access to card data, reducing the risk of fraudulent transactions and financial loss.', 'h': 'Continue regular inspections and maintenance of ATMs to ensure no skimming devices are installed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The skimming device or bugs were placed. ', 'd': 'Skimming devices or bugs were placed on ATMs to capture card information.', 'f': 'Card data can be stolen and used for unauthorized withdrawals, causing financial and reputational damage.', 'h': 'Upgrade all cards with EMV chips, install anti-skimming mechanisms, and conduct regular ATM inspections.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # physicalAttacksPINs
            'compliance': {'a': 'Compliance', 'b': 'PIN entry is secure.', 'd': 'No cameras or PIN-pad overlays were detected that could compromise PIN entry.', 'f': 'Protects customers\' PINs from theft, minimizing risk of unauthorized transactions.', 'h': 'Maintain secure ATM fascias and periodically inspect for tampering.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Physical, local attacks that target PINs were possible.', 'd': 'It was observed that the pinhole camera or other cameras leveraging the ATM surroundings or PIN-pad overlays were placed. A pinhole camera is a simple camera without a lens but with a tiny aperture, effectively a light-proof box with a small hole in one side. PIN pad overlays are devices that sit on top of the pin pad to record typed numbers.', 'f': 'Attackers can combine stolen PINs with card data to conduct fraudulent withdrawals, leading to financial loss and customer trust issues.', 'h': 'Regularly inspect ATMs for hidden cameras and overlays, and implement anti-PIN pad designs or privacy shields.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # sensitiveAreasRestricted
            'compliance': {'a': 'Compliance', 'b': 'Access restricted to authorized personnel.', 'd': 'Only authorized personnel could access sensitive areas such as ATM controllers and cash vaults.', 'f': 'Reduces risk of theft, tampering, or malicious software installation.', 'h': 'Maintain access logs and periodically audit cabinet access controls.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Unauthorized access to sensitive areas and resources in the cabinet was possible.', 'd': 'Unauthorized personnel could access critical ATM components, including the controller and cash vault.', 'f': 'If access to such sensitive areas is not restricted, it may cause financial damage and theft.', 'h': 'Restrict cabinet access to authorized personnel, install electronic access controls, and maintain access audit logs.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # eppValidApproval
            'compliance': {'a': 'Compliance', 'b': 'EPP has a valid approval.', 'd': 'It was verified that all Encrypting PIN Pads (EPPs) had valid regulatory approval.', 'f': 'Ensures PINs are encrypted and protected during entry, reducing the risk of compromise.', 'h': 'Continue using approved EPPs and verify approval periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'EPP contained no valid approval.', 'd': 'An Encrypting PIN Pad are used in Automated teller machines to ensure that the unencrypted PIN is not stored or transmitted anywhere in the rest of the system and thus cannot be revealed accidentally or through manipulations of the system. It was observed that the EPP contained no valid approval.', 'f': 'It will increase the chances of the PIN stealing if the EPP has not a valid approval.', 'h': 'It is recommended to use EPP with valid approval.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # visualObservationPIN
            'compliance': {'a': 'Compliance', 'b': 'PIN entry is visually protected.', 'd': 'ATM design prevents observation of PIN entry.', 'f': 'Mitigates shoulder surfing and protects customers from PIN theft.', 'h': 'Maintain privacy shields and verify ATM design periodically.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The visual observation of PIN values was possible as the cardholder enters them.', 'd': 'Visual observation of PIN values was possible as customers entered them.', 'f': 'Increases risk of shoulder surfing attacks and unauthorized access to customer accounts.', 'h': 'It is recommended that the fascia and cabinet design or the mechanical integration of the EPP should not facilitate the visual observation of PIN values as the cardholder is entering them.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # pinPolicy
            'compliance': {'a': 'Compliance', 'b': 'Policy exists.', 'd': 'The bank has a policy governing the issuance and management of PINs for ATM customers.', 'f': 'Standardizes PIN management and reduces the risk of misuse or theft.', 'h': 'Ensure all staff and customers are aware of the PIN policy and follow best practices.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Bank does not have policy regarding PINs used in ATM', 'd': 'The bank lacks a formal policy for PIN issuance and management.', 'f': 'The lack of a policy for PINs can result in inconsistencies in how PINs are managed and protected, leading to gaps in security that can be exploited by attackers. If the customers are not aware about the best practices for PIN while accessing ATMs they can be victim of cyber crimes, Phishing Campaigns.', 'h': 'It is recommended that the bank shall have a policy for the Personal Identification Numbers, used by various sets of customers who access the Banks systems directly using channels like ATMs.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # cctvCamerasInstalled
            'compliance': {'a': 'Compliance', 'b': 'Cameras properly positioned.', 'd': 'CCTV cameras capture relevant ATM areas and users without obstructing operations.', 'f': 'Enhances security, aids in investigations, and deters criminal activity.', 'h': 'Maintain camera positions and periodically review footage coverage.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'CCameras were not installed in the proper position.', 'd': 'CCTV cameras are not positioned to fully monitor ATM and user areas.', 'f': 'If the cameras are not in the proper position, they will not be able to achieve their goal of providing more security to the ATM. The footage will not capture the faces, unique attributes of the culprits to identify them which can aid in police/forensic investigations  if and incident ensues.', 'h': 'It is recommended to install CCTV cameras in the proper positions. The location for camera installation should be carefully chosen to ensure that images of keypad entries are not recorded. The camera should support the detection of the attachment of alien devices to the fascia and the face of all the ATM users.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # usbAccessDisabled
            'compliance': {'a': 'Compliance', 'b': 'USB access disabled.', 'd': 'ATM USB ports are disabled, preventing unauthorized peripheral connections.', 'f': 'Reduces risk of malware injection or data exfiltration.', 'h': 'Maintain USB access restrictions and periodically verify configurations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'It was observed that USB access was enabled in the ATM.', 'd': 'USB ports are active on ATMs, allowing potential unauthorized device connections.', 'f': 'Malware or malicious code can be introduced via USB, compromising ATM and network security.', 'h': 'Disable USB access on all ATMs and audit periodically to ensure compliance.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # osUpToDate
            'compliance': {'a': 'Compliance', 'b': 'OS fully patched.', 'd': 'Windows OS on ATMs is updated with the latest security patches.', 'f': 'Reduces vulnerabilities, ensuring resilience against known attacks.', 'h': 'Continue regular patch management and monitor updates.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The windows OS was not updated with the latest security patches.', 'd': 'ATMs are running outdated Windows OS without recent security patches.', 'f': 'Exploitable vulnerabilities can allow attackers to compromise systems, disrupting ATM services and causing financial loss.', 'h': 'Apply all security patches promptly and establish a regular patching schedule for ATMs.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # expiredOS
            'compliance': {'a': 'Compliance', 'b': 'Licensed OS in use.', 'd': 'All ATMs are running properly licensed and supported versions of Windows OS.', 'f': 'Ensures support, bug fixes, and security updates are available, reducing vulnerability risk.', 'h': 'Continue monitoring OS licenses and renew as necessary.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Windows OS was outdated. ', 'd': 'Windows OS on ATMs is expired (e.g., Windows 7) and unsupported.', 'f': 'Expired OS lacks updates, leaving ATMs vulnerable to attacks and malware, and no vendor support is available.', 'h': 'Upgrade to a supported, licensed OS immediately to ensure security and maintain compliance.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # unwantedSoftware
            'compliance': {'a': 'Compliance', 'b': 'No unwanted software installed.', 'd': 'Only authorized software required for ATM operations is installed.', 'f': 'Reduces risk of malware infection and operational issues.', 'h': 'Periodically audit installed software to ensure only approved applications are present.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The unwanted software was installed in ATM.', 'd': 'Unwanted software is a program that plays no role in the processes and the OS experience in the ATM It was observed that such software was installed on the ATM.', 'f': 'The software can take the form of a modified browsing experience, lack of control over process and installation, misleading messages, or unauthorized changes to ATM settings.', 'h': 'It is recommended to uninstall all the unwanted software which are not required for ATM procedures.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # rogueSoftwareInstallation
            'compliance': {'a': 'Compliance', 'b': 'Installation restricted.', 'd': 'ATMs prevent installation of unauthorized software, ensuring secure operations.', 'f': 'Protects ATMs from malware, spyware, or ransomware attacks.', 'h': 'Maintain installation restrictions and monitor for policy violations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The installation of rogue software was possible.', 'd': 'Rogue or Malicious Software refers to any malicious program that causes harm to an ATM system or network. It was observed that the installation of such software was possible on the ATM.', 'f': 'Malicious Malware Software can attack an ATM system or network in the form of viruses, worms, Trojans, spyware, adware, or rootkits, by injecting such malware into ATM software, affecting the whole network and interrupting business production.', 'h': 'It is recommended to disable the installation of any software without proper authorization by installing an Antivirus in ATM.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # logsAvailable
            'compliance': {'a': 'Compliance', 'b': 'Logs are maintained on ATM systems.', 'd': 'Logs are available and properly maintained, recording all critical ATM events and transactions.', 'f': 'Enables tracking of suspicious activity, helps in forensic investigations, and ensures accountability.', 'h': 'Continue to maintain logs and ensure periodic review and backup for audit purposes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The logs were not available.', 'd': 'Logs were not available on ATM systems, preventing tracking of system activities.', 'f': 'Administrator cannot investigate incidents or trace malicious activity, leading to delays in identifying attacks.', 'h': 'Ensure all ATM activities are logged and retained on ATM or centralized servers to support investigation and security monitoring.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # atmGrouted
            'compliance': {'a': 'Compliance', 'b': 'ATM properly grouted.', 'd': 'ATMs are installed and grouted as per RBI guidelines to prevent physical theft or removal.', 'f': 'Physical security is ensured, reducing risk of ATM tampering or theft.', 'h': 'Maintain proper grouting and inspect periodically for integrity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The ATM was not grouted to a wall, pillar, floor, etc.', 'd': 'ATM was not grouted to any wall, pillar, or floor.', 'f': 'High risk of physical theft or unauthorized movement of the ATM, violating RBI installation guidelines.', 'h': 'Grout ATMs properly to walls, pillars, or floors to prevent physical security breaches.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # cablesConcealed
            'compliance': {'a': 'Compliance', 'b': 'Cables are concealed and secured.', 'd': 'ATM and CCTV cables are properly concealed to prevent tampering or unauthorized access.', 'f': 'Reduces the risk of network attacks or physical tampering.', 'h': 'Continue proper cable management and periodic inspection for integrity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Cables of the ATM and CCTV were not concealed.', 'd': 'It was observed that the cable of ATM and CCTV were not concealed and secured properly.', 'f': 'It will make easy to an attacker, enter in to the bank’s network and perform Man in the Middle (MiTM) attack. An attacker first tamper with the Internet network (LAN) cable of the ATM and perform MiTM to withdraw money from ATMs.', 'h': 'It is recommended to properly conceal Cables of ATM and CCTV.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # networkPortsAccessible
            'compliance': {'a': 'Compliance', 'b': 'Network ports secured.', 'd': 'Network ports are protected, preventing unauthorized access to ATM network devices.', 'f': 'Minimizes risk of attacks through open ports and preserves ATM network security.', 'h': 'Maintain port security and access control measures.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The network ports were openly accessible.', 'd': 'A physical port is where communication begins or ends on a physical device or unit of equipment. It was observed that the network ports were openly accessible.', 'f': 'As the ATM is a public-facing infrastructure and the network ports are openly accessible. In the case of insufficient network security, a criminal with access to the ATM network through open ports can target available network services, intercept and spoof traffic, and attack network equipment. Criminals can also spoof responses from the processing center or obtain control of the ATM.', 'h': 'It is recommended that network ports should not be openly accessible.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # atmNetworkSegmented
            'compliance': {'a': 'Compliance', 'b': 'ATM network properly segmented.', 'd': 'ATM and branch networks are segmented, reducing cross-network threats.', 'f': 'Prevents compromise of ATM network from affecting branch internal systems.', 'h': 'Maintain network segmentation using VLANs or Layer 3 switches.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The ATM and branch networks were not segmented.', 'd': 'Ping and traffic between ATM and branch systems were possible, indicating no network segmentation.', 'f': 'If an attacker gets into the network, whether in a branch or ATM, can easily discover all the network devices and systems connected to the branch network. Thus, there is a possibility that an attacker can perform a DoS attack or pin-sweep attack bringing down the entire branch network.', 'h': 'It is recommended that segregation should be made for public and internal infrastructure. Layer 3 switch can be used for segmentation by creating a VLAN. Also, its recommended to disable ping between ATM and branch systems. '},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        18: {  # atmSoftwareUpdated
            'compliance': {'a': 'Compliance', 'b': 'ATM software updated.', 'd': 'ATM application and related software are updated with latest patches and security fixes.', 'f': 'Protects against known vulnerabilities and ensures smooth functioning of ATM systems.', 'h': 'Continue timely updates of ATM software and monitor patch deployment.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The ATM software was not up-to-date with patches.', 'd': 'It was observed that the ATM software was not up-to-date with patches. ATM application enables banks to monitor and manage their entire domestic or international ATM networks.', 'f': 'Capability to protect against new generated threats and new features will not be available if applications are either not updated or outdated. An attacker can exploit the existing vulnerability of the applications.', 'h': 'It is recommended to keep ATM applications up-to-date.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        19: {  # unauthorizedLogChanges
            'compliance': {'a': 'Compliance', 'b': 'Logs protected from unauthorized changes.', 'd': 'ATM logs are secured and cannot be modified or deleted without proper authorization.', 'f': 'Ensures integrity of audit trails, enabling accurate forensic investigations and accountability.', 'h': 'Continue to implement log integrity mechanisms and regular monitoring for unauthorized modifications.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The unauthorized change in logs was possible.', 'd': 'Unauthorized changes to ATM logs were possible.', 'f': 'Attackers can modify or delete logs to hide their tracks, making investigation of incidents difficult and increasing the risk of repeated attacks.', 'h': 'Implement controls to prevent log tampering, store logs on centralized servers, and enable alerts for any unauthorized changes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        20: {  # licensedAntivirus
            'compliance': {'a': 'Compliance', 'b': 'Licensed AV installed and updated.', 'd': 'ATM systems have licensed antivirus installed with latest virus definitions.', 'f': 'Provides protection against malware, ransomware, and other cyber threats, ensuring the confidentiality, integrity, and availability of ATM systems.', 'h': 'Maintain licensed antivirus and schedule automatic updates for virus definitions regularly.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Licensed antivirus is not available and it is not up to date with latest database.', 'd': 'It was observed that the AV was not available, Antivirus software (AV software), also known as anti-malware, is a computer program used to prevent, detect, and remove malicious programs.', 'f': 'If licensed antivirus with up to date database  is not available then the system will not be protected against newer threats of malware ,ransomware  which can compromise the confidentiality , integrity, availability of the system. Thus bank may face financial, reputational loss due to cyber attack.', 'h': 'It is recommended to use licensed Antivirus and update it regularly in the ATM to restrict unauthorized access of ATM OS and protect against malwares like virus, worms, keylogger.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        21: {  # cabinetDualPINs
            'compliance': {'a': 'Compliance', 'b': 'Cabinet properly secured with dual PINs.', 'd': 'ATM cabinets are locked securely and require two different personnel to open the safe.', 'f': 'Minimizes risk of internal collusion or theft and ensures strong physical security for cash and ATM components.', 'h': 'Maintain dual control access and periodically audit PIN assignments.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The ATM cabinet was not properly locked and was not configured to require two PINs from two different personnel to open the safe.', 'd': 'Only one person had both PINs to open the ATM safe.', 'f': 'Single-person access increases the risk of theft or collusion, compromising the security of the ATM.', 'h': 'Enforce dual-control mechanism, ensuring at least two authorized personnel are required to access the safe.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        22: {  # preventiveMaintenance
            'compliance': {'a': 'Compliance', 'b': 'Preventive maintenance performed regularly.', 'd': 'ATM preventive maintenance is scheduled, including replenishment of consumables, hardware checks, and software updates.', 'f': 'Ensures smooth functioning of ATM, reduces downtime, and mitigates potential risks due to hardware/software failures.', 'h': 'Continue preventive maintenance as per defined schedules and document all maintenance activities.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'It was observed that the preventive maintenance was not performed with defined periodicity.', 'd': 'Preventive maintenance was not performed with a defined periodicity.', 'f': 'Increases chances of hardware failure, cash jams, and operational disruptions, reducing ATM reliability and security.', 'h': 'Establish and adhere to a preventive maintenance schedule covering hardware, software, and consumables.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        23: {  # userAwareness
            'compliance': {'a': 'Compliance', 'b': 'Do\'s and Don\'ts displayed.', 'd': 'ATM rooms have visible Do\'s and Don\'ts for safe usage.', 'f': 'Educates customers on secure ATM usage, reducing the likelihood of fraud and misuse.', 'h': 'Continue displaying updated user guidance in all ATM locations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Do\'s and Don\'ts for user awarness were not available in some of the branch ATMs.', 'd': 'ATM rooms in some branches lacked Do\'s and Don\'ts for user awareness.', 'f': 'Customers may inadvertently expose themselves to fraud or security risks due to lack of guidance, increasing the risk of financial loss.', 'h': 'Display Do\'s and Don\'ts prominently in all ATMs, educating users on safe practices.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        24: {  # guardAvailable
            'compliance': {'a': 'Compliance', 'b': 'Guard present.', 'd': 'Security personnel are present at ATMs operating 24×7 during non-banking hours.', 'f': 'Deters unauthorized access, vandalism, or physical attacks on ATMs, enhancing security.', 'h': 'Continue security staffing for all ATMs during non-banking hours.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The guard was not available during non-banking hours.', 'd': 'It was observed that the guard was not available during non-banking hours. A vigilant ATM security guard can bring in important safety and security to Bank ATMs', 'f': "Most of the attackers are aimed at ATMs that don't have security guards. So not having a guard during non-banking hours increase the chances of ATM fraud.", 'h': 'It is recommended to keep an ATM security guard during non-banking hours if the ATM is working 24*7.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }
    
    # Apply formatting to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for col in range(1, 9):  # A to H
        cell = ws.cell(row=1, column=col)
        cell.border = thin_border
    
    # Process each question
    for i, question in enumerate(question_mapping, 2):  # Start from row 2
        # Get user input for this question
        user_input = form_data.get(question, 'not_applicable') if form_data else 'not_applicable'
        
        # Get response data
        response_data = question_responses.get(i-1, {})
        user_response = response_data.get(user_input, {})
        
        # Populate columns based on user input
        # Column C: Compliance/Non-Compliance/Not Applicable
        ws.cell(row=i, column=3, value=user_response.get('a', 'Not Applicable'))
        
        # Column D: Observation (Short/Brief)
        ws.cell(row=i, column=4, value=user_response.get('b', 'Not Applicable'))
        
        # Column F: Observation
        ws.cell(row=i, column=6, value=user_response.get('d', 'Not Applicable'))
        
        # Column G: Impact
        ws.cell(row=i, column=7, value=user_response.get('f', 'Not Applicable'))
        
        # Column H: Recommendation
        ws.cell(row=i, column=8, value=user_response.get('h', 'Not Applicable'))
        
        # Column E: Risk Factor with color coding
        risk_factor = risk_factors[i-2]  # risk_factors is 0-indexed
        risk_cell = ws.cell(row=i, column=5, value=risk_factor)
        risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
        risk_cell.fill = PatternFill(start_color=risk_colors.get(risk_factor, '808080'), end_color=risk_colors.get(risk_factor, '808080'), fill_type='solid')
        risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply formatting to all data rows
    for row in range(2, len(question_mapping) + 2):  # Rows 2 to 25
        for col in range(1, 9):  # Columns A to H
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Special alignment for column A (Sr. No.), C (Compliance status), and E (Risk Factor)
            if col == 1 or col == 3 or col == 5:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Set row height for wrapped text
            ws.row_dimensions[row].height = 30
    
    # Save file
    filename = "ATM Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    """
    Delete the generated Excel file after download
    """
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            print(f"File {filepath} deleted successfully")
    except Exception as e:
        print(f"Error deleting file {filepath}: {e}")
