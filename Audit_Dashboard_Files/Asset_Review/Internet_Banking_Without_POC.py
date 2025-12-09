import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_internet_banking_excel(form_data=None):
    """
    Create Internet Banking Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Internet Banking"

    # Define headers
    headers = [
        "Sr. No.", "Questionnaire/Points", "Compliance/Non-Compliance/Not Applicable",
        "Observation (Short/Brief)", "Risk Factor", "Observation", "Impact", "Recommendation"
    ]

    # Define column widths
    column_widths = {
        'A': 10, 'B': 50, 'C': 20, 'D': 30, 'E': 20, 'F': 50, 'G': 50, 'H': 50
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header styling
    header_font = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid') # Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate questions starting from row 2
    start_row = 2

    # Internet Banking Questions
    questions = [
        "Are server and database separate with VLANs?",
        "Does the disaster recoverability testing plan exist?",
        "Is the IB Web server hosted in a Demilitarized Zone (DMZ)?",
        "Is the server and database of internet banking updated and secured?",
        "Is there any validation to follow the standard password policy?",
        "Is the internet banking application the latest and well tested?",
        "Is SSL used?",
        "Is VAPT performed on the IB server and application?",
        "Is WAF or IPS used by the bank?",
        "Is IB running on HTTP TLS version 1.2 or greater?",
        "Are unused services like FTP, NFS, and TELNET disabled?",
        "Are only whitelisted applications allowed to execute on the IB Server?",
        "Are all data between client and server encrypted using AES or other techniques?",
        "Are all PII and sensitive customer information stored in encrypted form in the database?",
        "Is there an access control policy on the web and database server?",
        "Is the server hosting the IB application & DB hardened?",
        "Is operating system security review done on all servers used for internet banking?",
        "Does the software maintain password history (prevent reuse)?",
        "Does the bank have a policy for the Personal Identification Numbers used by customers accessing internet banking?",
        "Is logging enabled?",
        "Is multi-factor authentication present?",
        "Is a load balancer in use?",
        "Has a PCI-DSS audit been done?",
        "Is the password displayed in encrypted mode?",
        "Does the system enforce a minimum password length with alpha, numeric, and special character combinations?",
        "Does the software force the user to change the password at set periodic intervals?",
        "Are digital certificates available?",
        "Is there any browser protection mechanism?",
        "Is user and entity behavior analytics implemented for fraud detection?"
    ]

    # Risk Factors
    risk_factors = [
        "Critical", "Critical", "Critical", "High", "High", "High", "High", "High", "High", "High",
        "High", "High", "High", "High", "High", "High", "Medium", "Medium", "Medium", "Medium",
        "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Medium", "Low", "Low"
    ]

    # Mapping of form field names to question numbers
    question_mapping = {
        "ibServerDatabaseVLAN": 1,
        "ibDisasterRecoveryPlan": 2,
        "ibWebServerDMZ": 3,
        "ibServerDatabaseUpdated": 4,
        "ibPasswordPolicyValidation": 5,
        "ibApplicationLatestTested": 6,
        "ibSSLUsed": 7,
        "ibVAPTPerformed": 8,
        "ibWAFIPSUsed": 9,
        "ibTLSVersion": 10,
        "ibUnusedServicesDisabled": 11,
        "ibApplicationWhitelisting": 12,
        "ibDataEncryptionAES": 13,
        "ibPIIEncrypted": 14,
        "ibAccessControlPolicy": 15,
        "ibServerHardened": 16,
        "ibOSSecurityReview": 17,
        "ibPasswordHistory": 18,
        "ibPINPolicy": 19,
        "ibLoggingEnabled": 20,
        "ibMultiFactorAuth": 21,
        "ibLoadBalancer": 22,
        "ibPCIDSSAudit": 23,
        "ibPasswordEncryptedMode": 24,
        "ibPasswordComplexity": 25,
        "ibPasswordExpiration": 26,
        "ibDigitalCertificates": 27,
        "ibBrowserProtection": 28,
        "ibBehaviorAnalytics": 29
    }

    # Populate questions and Sr. No.
    for i, question in enumerate(questions, start_row):
        # Sr. No. with center alignment
        sr_no_cell = ws.cell(row=i, column=1, value=i-start_row+1)
        sr_no_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=i, column=2, value=question)  # Questions

    # Risk colors
    risk_colors = {
        'Critical': '8B0000',  # Dark Red
        'High': 'FF0000',      # Red
        'Medium': 'FFA500',    # Orange
        'Low': '008000'        # Green
    }

    # Question responses data
    question_responses = {
        1: {  # ibServerDatabaseVLAN
            'compliance': {'a': 'Compliance', 'b': 'Servers and databases isolated with VLANs.', 'd': 'Application servers and databases are hosted on separate VLANs with controlled communication rules.', 'f': 'Limits attack surface and improves overall security posture.', 'h': 'Periodically review VLAN configurations to ensure segmentation remains effective.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The web server and database were running on the same LAN network.', 'd': 'The application and database servers are hosted on the same network segment without VLAN segmentation, allowing unrestricted communication between them.', 'f': "If the web server and database are running on the same LAN network, it would be easy for an attacker to take over the application. If one of them is compromised, then the other one will automatically get compromised, which will lead to interruption of the whole business process and affect the bank's productivity.", 'h': 'Implement network segmentation using VLANs to isolate servers and databases, restricting unnecessary communication paths.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        2: {  # ibDisasterRecoveryPlan
            'compliance': {'a': 'Compliance', 'b': 'Disaster recovery plan exists and tested.', 'd': 'The organization maintains a documented disaster recovery plan and performs scheduled tests to ensure system recoverability.', 'f': 'Ensures minimal downtime and data integrity in case of disasters.', 'h': 'Continuously review and update the disaster recovery plan to reflect system changes.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'A disaster recoverability testing plan was not established.', 'd': 'There is no formal disaster recovery testing plan documented or periodically tested for critical systems.', 'f': "There is a high chance of losing data, assets even employees if a disaster ensues. It will be impossible to recover the data, assets lost in the disaster, which will impact the bank's ability to operate effectively and cause business interruption, loss of clients,financial loss etc.", 'h': 'Develop a formal disaster recovery plan, conduct regular tests, and update it based on test outcomes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        3: {  # ibWebServerDMZ
            'compliance': {'a': 'Compliance', 'b': 'IB Web server hosted in DMZ.', 'd': 'Internet banking web server is deployed in a DMZ with controlled access, isolating it from the internal network.', 'f': 'Enhances security by restricting direct access to internal resources from external users.', 'h': 'Regularly audit DMZ configuration and firewall rules to maintain security.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'IB Web server was not hosted in a DMZ.', 'd': 'Internet banking web server is hosted within the internal network instead of a DMZ, exposing internal systems to external attacks.', 'f': "If the IB Web server is not hosted in a DMZ. We won't be able to restrict access to sensitive data, resources, and servers or access control if an attacker compromises the IB server. An attacker can carry out reconnaissance of potential targets and attack through IP spoofing.", 'h': 'It is recommended that the IB Web server should be hosted in a DMZ.  Any network service that runs as a server requiring communication to an external network or the Internet should be placed in the DMZ.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        4: {  # ibServerDatabaseUpdated
            'compliance': {'a': 'Compliance', 'b': 'Servers and databases updated and secured.', 'd': 'Internet banking servers and databases are regularly patched, updated, and configured according to security best practices.', 'f': 'Reduces vulnerability exposure and strengthens system security.', 'h': 'Continuously monitor patch levels and security configurations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The web server and database used for the IB are not the latest version and are not secured.', 'd': 'It was observed that the web server and database used for the IB are not having the latest version and are not secure.', 'f': 'As the web server and database are the core part to run any application, if an attacker can gain access either a web server or database,by exploiting the vulnerability in outdated version of  server os or database bank can lose the highly confidential data.', 'h': 'Implement a patch management process to ensure servers and databases are regularly updated and hardened.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        5: {  # ibPasswordPolicyValidation
            'compliance': {'a': 'Compliance', 'b': 'Password policy enforced.', 'd': 'The application validates passwords against standard organizational rules, ensuring strong credentials for all users.', 'f': 'Improves account security and reduces the risk of unauthorized access.', 'h': 'Periodically review password enforcement rules and compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'There is no validation to follow standard password policy.', 'd': 'The internet banking application does not validate passwords against the organization\'s standard password rules during user creation or changes.', 'f': "Increases the probability of an attacker guessing or cracking the password of user accounts. An attacker who can determine user passwords can take over a user's account and potentially access sensitive data in the application.   ", 'h': 'Enforce password rules for length, complexity, and expiry, and validate them during account creation and password changes.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        6: {  # ibApplicationLatestTested
            'compliance': {'a': 'Compliance', 'b': 'Latest version and tested.', 'd': 'The internet banking application is up-to-date and has been thoroughly tested for security, functionality, and performance.', 'f': 'Minimizes operational risk and ensures secure, reliable service for users.', 'h': 'Continuously monitor updates and conduct regression testing after upgrades.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The application used for internet banking was not having the latest version.', 'd': 'Internet banking software in use is not the latest version and may not have undergone thorough testing for functionality and security.', 'f': 'Increased risk of vulnerabilities, software bugs, and operational failures affecting end-users.', 'h': 'Upgrade to the latest stable version and conduct rigorous testing for security, functionality, and performance before deployment.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        7: {  # ibSSLUsed
            'compliance': {'a': 'Compliance', 'b': 'SSL/TLS implemented.', 'd': 'Internet banking communication is encrypted using SSL/TLS, ensuring confidentiality and integrity of transmitted data.', 'f': 'Protects sensitive information from interception and enhances user trust.', 'h': 'Regularly renew certificates and verify encryption strength.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'SSL was not configured on the Internet Banking URL.', 'd': 'Internet banking communication occurs over non-secure HTTP instead of HTTPS, leaving data unencrypted during transmission.', 'f': 'Exposes sensitive data such as login credentials and transaction details to interception and man-in-the-middle attacks.', 'h': 'Implement SSL/TLS encryption to secure all communication between clients and the server.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        8: {  # ibVAPTPerformed
            'compliance': {'a': 'Compliance', 'b': 'VAPT performed regularly.', 'd': 'Vulnerability assessments and penetration testing are conducted on the internet banking infrastructure and application, identifying and mitigating security gaps.', 'f': 'Reduces risk exposure and ensures a secure and resilient application environment.', 'h': 'Schedule regular VAPT cycles and ensure remediation of identified vulnerabilities.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'VAPT was not performed on the IB server and application.', 'd': 'Vulnerability Assessment and Penetration Testing (VAPT) has not been performed on the internet banking server or application to identify potential security weaknesses.', 'f': 'Undetected vulnerabilities may exist, increasing the risk of exploitation, data breaches, and unauthorized access.', 'h': 'Conduct periodic VAPT on the IB server and application, document findings, and remediate all critical issues.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        9: {  # ibWAFIPSUsed
            'compliance': {'a': 'Compliance', 'b': 'WAF/IPS implemented.', 'd': 'The bank has deployed WAF/IPS solutions to monitor, detect, and prevent malicious activity targeting the IB application.', 'f': 'Enhances protection against common web attacks and network intrusions.', 'h': 'Periodically review WAF/IPS configurations and update rulesets.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'No IPS or firewall was used for securing internet banking applications.', 'd': 'The bank does not utilize Web Application Firewall (WAF) or Intrusion Prevention System (IPS) to protect the IB application from external attacks.', 'f': 'If an attacker finds that no security measures like WAF and IPS are implemented, he/she can attack the webserver and easily gain access to the server without any hurdle.Thus, the attacker can compromise the CIA triad of information security.', 'h': 'Deploy WAF and/or IPS to detect and block malicious traffic and protect critical infrastructure.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        10: {  # ibTLSVersion
            'compliance': {'a': 'Compliance', 'b': 'TLS 1.2 or higher enforced.', 'd': 'The IB application enforces TLS 1.2 or higher for all communication, ensuring secure data transmission.', 'f': 'Protects sensitive information from interception and aligns with industry security standards.', 'h': 'Regularly test TLS configurations to prevent downgrade attacks.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Internet banking application running on HTTP TLS version 1.0.', 'd': 'The internet banking application uses an older TLS version (below 1.2) that is vulnerable to known attacks.', 'f': 'Any services that currently rely on TLS 1.1 or older will no longer be available. The older TLS versions allow attacks similar to POODLE to breach it, even if a system has fully eradicated the POODLE flaw, it could still be vulnerable to GOLDENDOODLE attacks.', 'h': 'Upgrade to TLS 1.2 or higher and disable deprecated protocols.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        11: {  # ibUnusedServicesDisabled
            'compliance': {'a': 'Compliance', 'b': 'Unused services disabled.', 'd': 'Only required services are enabled on the IB server, and FTP, NFS, TELNET, and other unnecessary services are disabled.', 'f': 'Reduces the attack surface and strengthens server security posture.', 'h': 'Periodically review enabled services to ensure compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'All services like FTP, NFS, and TELNET were running on the IB server.', 'd': 'It was observed that services like FTP, NFS, and TELNET were running on the IB server. which could lead to the RCE Attack.TELNET (Telecommunication Network) and FTP (File Transfer Protocol) both are application layer protocols that create a connection between the remote host and a server. FTP is used for transferring files from one system to another system. ', 'f': 'These services increase the attack surface and may be exploited by attackers to gain unauthorized access.', 'h': 'Disable all unnecessary services and regularly audit server configurations.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        12: {  # ibApplicationWhitelisting
            'compliance': {'a': 'Compliance', 'b': 'Application whitelisting enforced.', 'd': 'The IB server allows execution only of approved applications, preventing untrusted software from running.', 'f': 'Reduces malware risk and enforces operational security.', 'h': 'Update the whitelist regularly and audit execution logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Application whitelisting not enforced.', 'd': 'Any software can execute on the IB server without restrictions, including potentially malicious applications.', 'f': 'An attacker can install any malicious application containing ransomware, or malware on the IB server, which will create a backdoor for an attacker to enter into the system and access the confidential data.Thus it will compromise the CIA triad and bank may face productivity loss, reputational loss, financial loss due to cyber attack.', 'h': 'Implement application whitelisting to allow only authorized software to run on the IB server.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        13: {  # ibDataEncryptionAES
            'compliance': {'a': 'Compliance', 'b': 'Data encrypted in transit.', 'd': 'All communication between client and server is encrypted using AES or equivalent encryption algorithms, ensuring data confidentiality.', 'f': 'Protects sensitive customer data from interception or tampering.', 'h': 'Regularly validate encryption standards and monitor for vulnerabilities.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'All data transmitted between client-server was in plain text.', 'd': 'It was observed that all data transmitted between client-server was transmitted in plain text.', 'f': " If the important content is not encrypted, the most serious impact will be that sensitive or confidential data can be breached. A potentially damaging impact of Sensitive data or other data breaches is potential cyber attacks planned using the stolen data, loss of sensitive data, and  damage to an organization's reputation.", 'h': 'It is recommended to encrypt all the data between client and server using AES or other encryption techniques.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        14: {  # ibPIIEncrypted
            'compliance': {'a': 'Compliance', 'b': 'Customer data encrypted in storage.', 'd': 'All sensitive customer information, including PII, is stored in encrypted form in the database, with secure key management practices in place.', 'f': 'Ensures confidentiality and regulatory compliance while mitigating data breach risks.', 'h': 'Periodically audit encryption mechanisms and key management practices.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The sensitive informaton of customer was stored in plain text in the database.', 'd': 'It was observed that all the PII and sensitive information of customers stored in database was not encrypted.', 'f': "If a hacker gets access to the data in this table, he/she will have a list of all passwords for all users. These passwords are in plain text, which can be used to log in to the system. People often use the same password for many accounts.so, if a hacker gets user names and passwords, he/she can use that information of users' accounts on other sites.", 'h': 'It is recommended to encrypt all the PII and sensitive information of customers before storing it in the database using hashing.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        15: {  # ibAccessControlPolicy
            'compliance': {'a': 'Compliance', 'b': 'Access control policy enforced.', 'd': 'Web and database servers implement RBAC or equivalent access control policies, ensuring users have only the permissions required for their roles.', 'f': 'Protects sensitive systems from unauthorized access and ensures accountability.', 'h': 'Regularly review access rights and update policies as per operational and security requirements.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'No access control policy was established.', 'd': 'It was observed that no access control policy was placed on the web and database server.Access control is a security technique that regulates who or what can view or use resources in a computing environment. Logical access control limits connections to computer networks, system files, servers, data, and data.', 'f': 'If the Access control policy is not implemented, an attacker can gain access from any low privilege user and can access the database and use all the sensitive information of the database and try to escalate the privileges of admin to compromise the database and gain confidential information including username and passwords of users,account details.', 'h': 'It is recommended to configure the access control policy to limit the access to databases and webservers to normal users. Only allow specific systems on the network to access the database server.  Implement PIM/PAM solution for the same.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        16: {  # ibServerHardened
            'compliance': {'a': 'Compliance', 'b': 'Servers hardened.', 'd': 'The IB application and database servers are hardened according to best practices, including secure configuration, patching, and minimal service exposure.', 'f': 'Reduces risk of compromise and strengthens overall security posture.', 'h': 'Periodically review server configurations and maintain hardening standards.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The Web server and database were not hardened. ', 'd': 'The servers hosting the internet banking application and database lack standard hardening measures such as disabling unnecessary services, patching, and secure configuration.', 'f': "As server is not hardened enough an attacker can find any loophole or vulnerable part of server or databse to get in to it and perform  malicious activities which will directly impact to the bank's productivity. The server will be vulnerable to unauthorized access, unauthorized usage and disruptions in service.", 'h': 'Apply industry-standard hardening practices, including OS and database configuration, patch management, and removal of unnecessary services.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        17: {  # ibOSSecurityReview
            'compliance': {'a': 'Compliance', 'b': 'OS security reviews performed.', 'd': 'Operating systems on all IB servers are periodically reviewed for security, ensuring proper patching, configuration, and compliance.', 'f': 'Identifies vulnerabilities early, reducing the risk of exploitation.', 'h': 'Schedule routine security reviews and maintain audit logs.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Operating System Security review was not done on all the servers used for internet banking.', 'd': 'Security reviews or audits of the operating systems on IB servers are not conducted periodically.', 'f': 'If the Operating System Security review is not done then the attacker can use the vulnerability in the existing server OS  to compromise the server and the CIA triad of information security.', 'h': 'It is recommended that Operating System Security review must be done on all the servers used for internet banking.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        18: {  # ibPasswordHistory
            'compliance': {'a': 'Compliance', 'b': 'Password history enforced.', 'd': 'The software enforces password history rules, disallowing reuse of prior passwords to ensure strong account security.', 'f': 'Reduces risk of password-based attacks and ensures compliance with security policies.', 'h': 'Regularly review and update password history settings as per policy.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The application was not maintaining a password history.', 'd': 'The IB application allows users to reuse previous passwords, reducing the effectiveness of password policies.', 'f': 'As the history of the password is not retained, the user can use a repetitive password for his ease.If the older credentials are leaked then the user account will be easily compromised.', 'h': 'It is recommended to implement a process that retains the password history and the user should not be allowed to use the last 3 passwords he used.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        19: {  # ibPINPolicy
            'compliance': {'a': 'Compliance', 'b': 'PIN policy implemented.', 'd': 'A formal PIN policy exists, enforcing secure creation, management, and periodic change of customer PINs.', 'f': 'Enhances security and reduces risk of unauthorized account access.', 'h': 'Periodically review PIN policy compliance and update standards as required.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Banks do not have a policy for the Personal Identification Numbers, used by various sets of customers who access the Bank's systems directly using channels like Internet banking.", 'd': 'No formal policy exists to define the creation, management, or complexity requirements for customer PINs in internet banking.', 'f': 'Without a clear policy, customers may not be educated about the importance of creating strong PINs or be unaware of best practices for protecting their account information. The lack of a PIN policy can hinder fraud detection and prevention efforts, making it harder to identify suspicious activities and unauthorized transactions.', 'h': "It is recommended to create a policy for the Personal Identification Numbers, used by various sets of customers who access the Bank's systems directly using channels like  Internet banking."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        20: {  # ibLoggingEnabled
            'compliance': {'a': 'Compliance', 'b': 'Logging enabled.', 'd': 'All critical actions, transactions, and system events are logged with proper retention policies and monitored for anomalies.', 'f': 'Facilitates incident detection, audit, and forensic analysis.', 'h': 'Regularly review log configurations and monitor for unusual activity.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Logging was not enabled.', 'd': 'The IB application and server activities are not logged or monitored, including user actions, transactions, and system events.', 'f': 'If log settings are not properly configured, bank will not be able to detect the DoS attack, IP Spoofing, and other kinds of attacks. Logs will not be available at the time of forensic investigation to analyze the attack.', 'h': 'Enable comprehensive logging and retain logs according to regulatory requirements, with regular monitoring.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        21: {  # ibMultiFactorAuth
            'compliance': {'a': 'Compliance', 'b': 'Multi-factor authentication enforced.', 'd': 'The application enforces MFA using OTP, token, or app-based verification, ensuring strong authentication.', 'f': 'Significantly reduces the risk of unauthorized access.', 'h': 'Periodically test MFA mechanisms and update configurations as needed.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Multi-factor authentication was not present on the login page.', 'd': 'Internet banking users can access accounts using only a username and password without additional verification factors.', 'f': 'Employees fall for phishing scams and share passwords, and if you’re not using multi-factor authentication (MFA), the bank is  open  wide to attacks. If multi-factor authentication is not available, the bank will not be able to protect against compromised credentials. ', 'h': 'Implement multi-factor authentication (MFA) for all customer and administrative access.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        22: {  # ibLoadBalancer
            'compliance': {'a': 'Compliance', 'b': 'Load balancer implemented.', 'd': 'A load balancer distributes traffic across multiple servers, improving performance, scalability, and redundancy.', 'f': 'Ensures high availability and smooth user experience during peak usage.', 'h': 'Monitor load balancer performance and health of backend servers.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The load balancer was not in use.', 'd': 'Internet banking traffic is handled by a single server without a load balancer, leading to potential performance and availability issues.', 'f': 'If the network traffic is high, the absence of a load balancer might crash the system. The unavailability of a load balancer will affect the availability of applications and websites for users. Modern applications cannot run without load balancers.', 'h': 'Implement a load balancer to distribute traffic, enhance performance, and improve availability.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        23: {  # ibPCIDSSAudit
            'compliance': {'a': 'Compliance', 'b': 'PCI-DSS audit completed.', 'd': 'The bank has undergone a PCI-DSS audit, ensuring proper handling, storage, and processing of payment card data.', 'f': 'Enhances security of cardholder data and maintains regulatory compliance.', 'h': 'Perform regular audits and remediation to maintain ongoing PCI-DSS compliance.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'PCI-DSS audit was not done by the bank. ', 'd': 'No formal PCI-DSS audit has been conducted to validate compliance with standards for processing, storing, or transmitting payment card data.', 'f': 'Merchants who do not comply with PCI DSS and are involved in a credit card breach may be subject to fines, card replacement costs, or incur costly forensic audits. The acquirer in turn passes the fines downstream until it eventually hits the merchant.', 'h': 'Conduct a PCI-DSS audit and remediate gaps to ensure compliance with card data security standards.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        24: {  # ibPasswordEncryptedMode
            'compliance': {'a': 'Compliance', 'b': 'Password encrypted/masked.', 'd': 'Passwords are displayed in masked form (e.g., asterisks) and stored securely using encryption techniques.', 'f': 'Protects user credentials from exposure and unauthorized access.', 'h': 'Periodically verify that all password fields are properly masked and stored securely.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Password was not displayed in encrypted mode.', 'd': 'The system shows user passwords in plaintext during login or in user management screens, exposing them to unauthorized viewing.', 'f': 'If passwords are not displayed in encrypted mode, they can be stolen and used by the unauthenticated user.Attacker can use the password to compromise the user account thus the confidentiality ,  integrity, availability of the user account can be compromised.', 'h': 'Implement encryption or masking for all password fields to ensure passwords are never displayed in plaintext.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        25: {  # ibPasswordComplexity
            'compliance': {'a': 'Compliance', 'b': 'Strong password policy enforced.', 'd': 'The system requires passwords to meet minimum length and include alpha, numeric, and special characters.', 'f': 'Enhances security by reducing the likelihood of password-related attacks.', 'h': 'Regularly review password policy compliance and adjust as per best practices.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': "Password Policies were not configured as per the organization's password management policy.", 'd': 'The system does not enforce a strong password policy; users can create passwords without minimum length or required character combinations.', 'f': 'It was very easy to crack a password through a brute force attack. If short passwords are permitted, security will be reduced because they can be easily cracked with tools that perform either dictionary or brute force attacks.', 'h': 'Enforce password complexity rules requiring a minimum length and a combination of alphabets, numbers, and special characters.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        26: {  # ibPasswordExpiration
            'compliance': {'a': 'Compliance', 'b': 'Password expiration enforced.', 'd': 'Users are required to change passwords periodically according to organizational policy, ensuring old passwords are retired regularly.', 'f': 'Minimizes the risk of long-term credential exposure and enhances account security.', 'h': 'Monitor compliance with password expiration rules and enforce reminders.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'The software did not force the user to change the password at set periodical intervals.', 'd': 'Users are allowed to retain the same password indefinitely without any periodic expiration policy.', 'f': 'The longer a password exists, the higher the likelihood that it will be compromised by a brute force attack, by an attacker gaining general knowledge about the user, or by the user sharing the password. Configuring the Minimum password age policy setting to 1 so that users are never required to change their passwords allows a compromised password to be used by the malicious user for as long as the valid user is authorized to access it.', 'h': "Implement a mandatory password expiration policy that prompts users to change their passwords at defined intervals, in accordance with the bank's password policy."},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        27: {  # ibDigitalCertificates
            'compliance': {'a': 'Compliance', 'b': 'Digital certificates implemented.', 'd': 'All critical servers and applications use valid digital certificates for authentication and secure communication.', 'f': 'Ensures encryption of data in transit and verifies server identity, reducing security risks.', 'h': 'Monitor certificate validity and renew before expiration.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Digital Certificate was not used.', 'd': 'Servers and applications lack digital certificates for secure identification and encryption, affecting SSL/TLS communications.', 'f': 'It is possible to export A1 certificates and remotely utilize them. A3 certificates can be used by more than one user at the same time, allowing adversaries to use stolen certificates.', 'h': 'Install valid digital certificates for all servers and applications and enforce their use for secure communication.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        28: {  # ibBrowserProtection
            'compliance': {'a': 'Compliance', 'b': 'Browser protection mechanisms implemented.', 'd': 'The application enforces CSP, secure headers, and other browser-level security protections to prevent client-side attacks.', 'f': 'Reduces risk of XSS, clickjacking, and other browser-based attacks.', 'h': 'Periodically review and update browser security configurations.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'Browser Protection was not available.', 'd': 'The system does not enforce browser security mechanisms like Content Security Policy (CSP), anti-clickjacking headers, or certificate pinning.', 'f': 'New malware remains active until they are identified by the model. Counterfeit online banking system web pages which prevent the protection from properly loading can be used to make the user input his sensitive data (such as passwords) in an unsafe environment.', 'h': 'Implement browser protection mechanisms including CSP, secure headers, and certificate pinning to protect users.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        },
        29: {  # ibBehaviorAnalytics
            'compliance': {'a': 'Compliance', 'b': 'UEBA implemented.', 'd': 'The system monitors and analyzes user and entity behavior to identify anomalies, unusual transactions, or potential fraud.', 'f': 'Enhances fraud detection capabilities and protects customer and organizational assets.', 'h': 'Continuously update analytics models and review detected anomalies for timely action.'},
            'non_compliance': {'a': 'Non-Compliance', 'b': 'User and entity behavior analytics for fraud detection was not implemented. ', 'd': 'The system does not analyze user or entity behavior to detect anomalies or fraudulent activities.', 'f': "If an attacker tries to perform some malicious activity on the bank server, it will not be detected because UEBA was not implemented. Also,bank won't be able to analyze the file, flow, and packet information for fraud detection.", 'h': 'Implement user and entity behavior analytics (UEBA) to detect suspicious patterns and prevent fraud.'},
            'not_applicable': {'a': 'Not Applicable', 'b': 'Not Applicable', 'd': 'Not Applicable', 'f': 'Not Applicable', 'h': 'Not Applicable'}
        }
    }

    # Populate data based on user input
    for i, question in enumerate(questions, start_row):
        # Get user input for this question
        question_num = i - start_row + 1
        user_input = None
        
        if form_data:
            # Find the corresponding form field
            for field_name, q_num in question_mapping.items():
                if q_num == question_num:
                    user_input = form_data.get(field_name, 'not_applicable')
                    break
        
        if not user_input:
            user_input = 'not_applicable'
        
        if question_num in question_responses:
            response_data = question_responses[question_num].get(user_input, question_responses[question_num]['not_applicable'])
            
            ws.cell(row=i, column=3, value=response_data['a'])
            ws.cell(row=i, column=4, value=response_data['b'])
            ws.cell(row=i, column=6, value=response_data['d'])
            ws.cell(row=i, column=7, value=response_data['f'])
            ws.cell(row=i, column=8, value=response_data['h'])
            
            for col in [3, 4, 6, 7, 8]:
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
        else:
            # Default values for questions not yet added
            ws.cell(row=i, column=3, value="Not Applicable")
            ws.cell(row=i, column=4, value="Not Applicable")
            ws.cell(row=i, column=6, value="Not Applicable")
            ws.cell(row=i, column=7, value="Not Applicable")
            ws.cell(row=i, column=8, value="Not Applicable")
        
        for col in [3, 4, 6, 7, 8]:
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Risk Factor
        if question_num <= len(risk_factors):
            risk_factor = risk_factors[question_num - 1]
            risk_cell = ws.cell(row=i, column=5, value=risk_factor)
            risk_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            risk_cell.border = thin_border
            
            if risk_factor in risk_colors:
                risk_cell.fill = PatternFill(start_color=risk_colors[risk_factor], end_color=risk_colors[risk_factor], fill_type='solid')
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
            else:
                risk_cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
    
    # Apply general formatting to all cells
    for row in range(1, len(questions) + 2):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if row > 1:  # Skip header row
                cell.border = thin_border
                if col in [1, 3, 5]:  # Sr. No., Compliance, Risk Factor - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Other columns - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                if col != 5:  # Don't override Risk Factor font
                    cell.font = Font(name='Calibri', size=11)
            
            ws.row_dimensions[row].height = 30
    
    filename = "Internet Banking Review.xlsx"
    filepath = os.path.join('static', 'uploads', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

def cleanup_file(filepath):
    if os.path.exists(filepath):
        os.remove(filepath)
        print(f"Cleaned up file: {filepath}")
