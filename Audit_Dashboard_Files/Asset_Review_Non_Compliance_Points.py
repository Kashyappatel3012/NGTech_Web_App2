from flask import Blueprint, request, flash, redirect, url_for, Response
from flask_login import login_required
from datetime import datetime
import os
import tempfile
import shutil
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
import pandas as pd
import numpy as np
from io import BytesIO
import re

# Create blueprint
asset_review_non_compliance_bp = Blueprint('asset_review_non_compliance_bp', __name__)

# Allowed file extensions
ALLOWED_EXCEL_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename, allowed_extensions):
    """Check if the uploaded file has an allowed extension"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in allowed_extensions

def should_keep_row(value):
    """Check if the row should be kept based on Column C value"""
    if pd.isna(value) or value is None:
        return False
    
    # Convert to string and normalize
    value_str = str(value).strip()
    
    # If the cell contains multiple entries (like "Compliance/Non-Compliance"), keep the row
    if '/' in value_str or '\\' in value_str or '|' in value_str or ',' in value_str:
        return True
    
    # Valid Non-Compliance variations (case-insensitive)
    non_compliance_variations = [
        'Non-Compliance', 'Non-compliance', 'non-Compliance', 
        'Noncompliance', 'noncompliance', 'NonCompliance'
    ]
    
    # Check if the value matches any variation (case-insensitive)
    return value_str.lower() in [var.lower() for var in non_compliance_variations]

def remove_images_from_excel_for_rows(worksheet, rows_to_remove):
    """Remove images from worksheet for specified rows in POC columns (I, J, K)"""
    try:
        print(f"Removing images from worksheet for rows: {rows_to_remove}")
        
        # Get all images in the worksheet
        images_to_remove = []
        
        # Check each image in the worksheet
        for i, image in enumerate(worksheet._images[:]):  # Create a copy of the list to iterate safely
            try:
                # Get image position using different methods based on openpyxl version
                row_num = None
                col_num = None
                
                # Try different ways to get the anchor position
                if hasattr(image, 'anchor'):
                    anchor = image.anchor
                    
                    # Method 1: Direct access to _from
                    if hasattr(anchor, '_from'):
                        if hasattr(anchor._from, 'row') and hasattr(anchor._from, 'col'):
                            row_num = anchor._from.row + 1  # Convert from 0-based to 1-based indexing
                            col_num = anchor._from.col
                            print(f"Image {i}: Found at row {row_num}, column {col_num} (method 1)")
                    
                    # Method 2: Using anchor coordinates
                    elif hasattr(anchor, 'row') and hasattr(anchor, 'col'):
                        row_num = anchor.row + 1
                        col_num = anchor.col
                        print(f"Image {i}: Found at row {row_num}, column {col_num} (method 2)")
                    
                    # Method 3: Try to get from anchor string
                    elif hasattr(anchor, '_from') and hasattr(anchor._from, '_row') and hasattr(anchor._from, '_col'):
                        row_num = anchor._from._row + 1
                        col_num = anchor._from._col
                        print(f"Image {i}: Found at row {row_num}, column {col_num} (method 3)")
                
                # Check if this image is in POC columns (I=8, J=9, K=10) and row should be removed
                if row_num is not None and col_num is not None:
                    col_letter = chr(65 + col_num) if col_num < 26 else f"{chr(65 + (col_num // 26) - 1)}{chr(65 + col_num % 26)}"
                    
                    if col_num in [8, 9, 10] and row_num in rows_to_remove:
                        images_to_remove.append((i, image, row_num, col_letter))
                        print(f"Marking image {i} for removal at row {row_num}, column {col_letter}")
                    else:
                        print(f"Image {i} at row {row_num}, column {col_letter} - keeping (not in removal list)")
                else:
                    print(f"Image {i}: Could not determine position, skipping")
                    
            except Exception as e:
                print(f"Error processing image {i}: {e}")
        
        # Remove the marked images (in reverse order to maintain indices)
        removed_count = 0
        for i, image, row_num, col_letter in reversed(images_to_remove):
            try:
                worksheet._images.remove(image)
                removed_count += 1
                print(f"Successfully removed image from row {row_num}, column {col_letter}")
            except Exception as e:
                print(f"Error removing image from row {row_num}, column {col_letter}: {e}")
        
        print(f"Successfully removed {removed_count} images from worksheet")
        
    except Exception as e:
        print(f"Error removing images from worksheet: {e}")
        raise e

def reposition_images_after_row_deletion(worksheet, deleted_rows):
    """Reposition remaining images to match new row positions after row deletion"""
    try:
        print(f"Repositioning images after deleting rows: {sorted(deleted_rows)}")
        
        # Sort deleted rows to calculate offset correctly
        sorted_deleted_rows = sorted(deleted_rows)
        
        # Check each remaining image in the worksheet
        for i, image in enumerate(worksheet._images[:]):
            try:
                # Get current image position
                current_row = None
                col_num = None
                
                # Try different ways to get the anchor position
                if hasattr(image, 'anchor'):
                    anchor = image.anchor
                    
                    # Method 1: Direct access to _from
                    if hasattr(anchor, '_from'):
                        if hasattr(anchor._from, 'row') and hasattr(anchor._from, 'col'):
                            current_row = anchor._from.row + 1  # Convert from 0-based to 1-based indexing
                            col_num = anchor._from.col
                    
                    # Method 2: Using anchor coordinates
                    elif hasattr(anchor, 'row') and hasattr(anchor, 'col'):
                        current_row = anchor.row + 1
                        col_num = anchor.col
                    
                    # Method 3: Try to get from anchor string
                    elif hasattr(anchor, '_from') and hasattr(anchor._from, '_row') and hasattr(anchor._from, '_col'):
                        current_row = anchor._from._row + 1
                        col_num = anchor._from._col
                
                # Only reposition images in POC columns (I=8, J=9, K=10)
                if current_row is not None and col_num is not None and col_num in [8, 9, 10]:
                    # Calculate how many rows were deleted before this row
                    rows_deleted_before = sum(1 for deleted_row in sorted_deleted_rows if deleted_row < current_row)
                    
                    # Calculate new row position
                    new_row = current_row - rows_deleted_before
                    
                    if new_row != current_row:
                        col_letter = chr(65 + col_num) if col_num < 26 else f"{chr(65 + (col_num // 26) - 1)}{chr(65 + col_num % 26)}"
                        
                        # Update the image position
                        try:
                            # Method 1: Update _from coordinates
                            if hasattr(anchor, '_from') and hasattr(anchor._from, 'row'):
                                anchor._from.row = new_row - 1  # Convert back to 0-based indexing
                                print(f"Repositioned image {i} from row {current_row} to row {new_row} in column {col_letter} (method 1)")
                            
                            # Method 2: Update anchor coordinates
                            elif hasattr(anchor, 'row'):
                                anchor.row = new_row - 1
                                print(f"Repositioned image {i} from row {current_row} to row {new_row} in column {col_letter} (method 2)")
                            
                            # Method 3: Update _from _row coordinates
                            elif hasattr(anchor, '_from') and hasattr(anchor._from, '_row'):
                                anchor._from._row = new_row - 1
                                print(f"Repositioned image {i} from row {current_row} to row {new_row} in column {col_letter} (method 3)")
                            
                            else:
                                print(f"Could not reposition image {i} at row {current_row} - unknown anchor structure")
                                
                        except Exception as e:
                            print(f"Error repositioning image {i} from row {current_row} to row {new_row}: {e}")
                    else:
                        print(f"Image {i} at row {current_row} in column {col_letter} - no repositioning needed")
                else:
                    print(f"Image {i} - not in POC columns or position could not be determined, skipping")
                    
            except Exception as e:
                print(f"Error processing image {i} for repositioning: {e}")
        
        print("Image repositioning completed")
        
    except Exception as e:
        print(f"Error repositioning images: {e}")
        raise e

def renumber_column_a(worksheet):
    """Renumber Column A (Sr. No.) to show sequential numbers after row deletion"""
    try:
        print("Renumbering Column A (Sr. No.) to show sequential numbers")
        
        # Find the last row with data in the worksheet
        max_row = worksheet.max_row
        
        if max_row <= 1:  # If only header row or empty sheet
            print("No data rows to renumber")
            return
        
        # Start renumbering from row 2 (assuming row 1 is header)
        # Check if row 1 is actually a header by looking for "Sr. No." or similar
        header_row = 1
        if max_row >= 1:
            header_cell = worksheet['A1']
            header_value = str(header_cell.value).lower() if header_cell.value else ""
            if 'sr' in header_value and 'no' in header_value:
                print("Found header row with 'Sr. No.' - starting renumbering from row 2")
                start_row = 2
            else:
                print("No clear header found - starting renumbering from row 1")
                start_row = 1
        else:
            start_row = 1
        
        # Renumber the rows sequentially
        sequence_number = 1
        for row in range(start_row, max_row + 1):
            cell = worksheet[f'A{row}']
            
            # Only renumber if the cell has content or is not empty
            if cell.value is not None and str(cell.value).strip() != '':
                # Store the new sequential number
                new_value = sequence_number
                
                # Set the new sequential number
                cell.value = new_value
                print(f"Renumbered row {row}: {new_value}")
                
                sequence_number += 1
            else:
                print(f"Row {row} has empty Column A - skipping")
        
        print(f"Column A renumbering completed. Total sequential numbers: {sequence_number - 1}")
        
    except Exception as e:
        print(f"Error renumbering Column A: {e}")
        raise e

def process_excel_file(file_path):
    """Process the Excel file to filter rows based on Non-Compliance values in Column C while preserving formatting"""
    try:
        print(f"Processing Excel file: {file_path}")
        
        # Load the Excel file with openpyxl to preserve formatting
        workbook = load_workbook(file_path)
        
        processed_data = {
            'sheets': [],
            'total_rows_before': 0,
            'total_rows_after': 0,
            'non_compliance_rows_found': 0,
            'rows_removed': 0
        }
        
        print(f"Sheet names: {workbook.sheetnames}")
        
        # Process each sheet
        for sheet_name in workbook.sheetnames:
            print(f"\nProcessing sheet: {sheet_name}")
            
            worksheet = workbook[sheet_name]
            
            # Find the last row with data
            max_row = worksheet.max_row
            print(f"Original rows in sheet '{sheet_name}': {max_row}")
            processed_data['total_rows_before'] += max_row
            
            # Check if Column C exists
            if max_row == 0:
                print(f"Warning: Sheet '{sheet_name}' is empty, skipping...")
                continue
            
            # Identify rows to keep (where Column C contains Non-Compliance values or multiple entries)
            rows_to_keep = []
            rows_to_remove = []
            
            # Process rows from bottom to top to avoid index shifting issues when deleting
            for row_num in range(max_row, 0, -1):
                cell_c = worksheet[f'C{row_num}']
                column_c_value = cell_c.value
                
                if should_keep_row(column_c_value):
                    rows_to_keep.append(row_num)
                    processed_data['non_compliance_rows_found'] += 1
                else:
                    rows_to_remove.append(row_num)
                    processed_data['rows_removed'] += 1
            
            print(f"Rows to keep (Non-Compliance or Multiple entries): {len(rows_to_keep)}")
            print(f"Rows to remove: {len(rows_to_remove)}")
            
            # Remove images from POC columns (I, J, K) for rows that will be deleted
            if rows_to_remove:
                print(f"Removing images from POC columns for {len(rows_to_remove)} rows before deleting rows")
                remove_images_from_excel_for_rows(worksheet, rows_to_remove)
            
            # Remove rows that should be deleted (from bottom to top to avoid index issues)
            for row_num in rows_to_remove:
                try:
                    worksheet.delete_rows(row_num)
                    print(f"Deleted row {row_num}")
                except Exception as e:
                    print(f"Error deleting row {row_num}: {e}")
            
            # Reposition remaining images to match new row positions after deletion
            if rows_to_remove:
                print(f"Repositioning remaining images after row deletion")
                reposition_images_after_row_deletion(worksheet, rows_to_remove)
            
            # Renumber Column A (Sr. No.) to show sequential numbers after filtering
            if rows_to_remove:
                print(f"Renumbering Column A (Sr. No.) after row deletion")
                renumber_column_a(worksheet)
            
            # Get final row count
            final_row_count = worksheet.max_row
            processed_data['total_rows_after'] += final_row_count
            
            sheet_data = {
                'name': sheet_name,
                'original_rows': max_row,
                'filtered_rows': final_row_count,
                'rows_removed': len(rows_to_remove),
                'non_compliance_rows': len(rows_to_keep)
            }
            
            processed_data['sheets'].append(sheet_data)
            print(f"Sheet '{sheet_name}' processed: {sheet_data['original_rows']} → {sheet_data['filtered_rows']} rows")
        
        # Save the workbook with all changes
        workbook.save(file_path)
        workbook.close()
        
        return processed_data
        
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        raise e

def apply_compliance_message_and_row_heights(file_path):
    """Apply compliance message and normalize row heights just before download"""
    try:
        print(f"\nApplying compliance message and row height normalization to: {file_path}")
        
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Process each sheet
        for sheet_name in workbook.sheetnames:
            print(f"\nProcessing sheet: {sheet_name}")
            worksheet = workbook[sheet_name]
            
            # Check rows C2 to C10 for "Non-Compliance" word
            # Get the maximum row to check (up to row 10, but not more than actual max_row)
            max_check_row = min(10, worksheet.max_row)
            has_non_compliance = False
            
            for row_num in range(2, max_check_row + 1):  # Rows 2 to 10 (or max_row if less than 10)
                try:
                    cell_c = worksheet[f'C{row_num}']
                    cell_value = cell_c.value
                    
                    if cell_value is not None:
                        # Convert to string and check for "Non-Compliance" (case-insensitive)
                        value_str = str(cell_value).strip()
                        if 'non-compliance' in value_str.lower() or 'noncompliance' in value_str.lower():
                            has_non_compliance = True
                            print(f"Found 'Non-Compliance' in C{row_num}: {value_str}")
                            break
                except Exception as e:
                    print(f"Error checking C{row_num}: {e}")
                    continue
            
            # If no "Non-Compliance" found in C2-C10, add compliance message
            if not has_non_compliance:
                print(f"No 'Non-Compliance' found in C2-C10 for sheet '{sheet_name}' - adding compliance message")
                
                # Merge cells A3 to K3
                worksheet.merge_cells('A3:K3')
                
                # Get the merged cell (A3)
                merged_cell = worksheet['A3']
                
                # Set the text
                merged_cell.value = "As per Auditor's observation All the points are compliance."
                
                # Set font: Calibri, 12 size, bold, green color
                merged_cell.font = Font(name='Calibri', size=12, bold=True, color='008000')  # Green color (RGB: 0, 128, 0)
                
                # Set alignment: center (horizontal and vertical)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Set border: 1px black border to all cells in the merged range (A3 to K3)
                thin_border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                
                # Apply border to all cells in the merged range (A3, B3, C3, D3, E3, F3, G3, H3, I3, J3, K3)
                for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                    cell = worksheet[f'{col_letter}3']
                    cell.border = thin_border
                
                print(f"Compliance message added to merged cells A3:K3 in sheet '{sheet_name}' with borders on all cells")
            else:
                print(f"Sheet '{sheet_name}' has 'Non-Compliance' in C2-C10 - skipping compliance message")
            
            # Check ALL rows A2 to A60 individually - set row height based on each cell's content
            # Check up to row 60, but not more than actual max_row
            max_check_row = min(60, worksheet.max_row)
            
            print(f"Checking rows A2 to A{max_check_row} for content and setting row heights accordingly...")
            
            for row_num in range(2, max_check_row + 1):  # Check ALL rows from 2 to 60 (or max_row)
                try:
                    # Get cell A for this row
                    cell_a = worksheet[f'A{row_num}']
                    
                    # Check if cell A has no content (empty or whitespace only)
                    has_content = False
                    if cell_a.value is not None:
                        value_str = str(cell_a.value).strip()
                        if value_str != '':
                            has_content = True
                    
                    # If cell A has no content, set row height to normal (default)
                    if not has_content:
                        # Set row height to default (normal) - Excel default is around 15 points
                        worksheet.row_dimensions[row_num].height = 15.0
                        print(f"Row {row_num}: Column A is empty - Set height to normal (15 points)")
                    else:
                        # Cell has content - keep current height (or don't modify)
                        print(f"Row {row_num}: Column A has content ('{str(cell_a.value)[:50]}...') - Keeping current height")
                        
                except Exception as e:
                    print(f"Error checking A{row_num} in sheet '{sheet_name}': {e}")
                    # Continue checking next row even if this one failed
                    continue
            
            print(f"Completed checking rows A2 to A{max_check_row} in sheet '{sheet_name}'")
        
        # Save the workbook with all changes
        workbook.save(file_path)
        workbook.close()
        
        print("\nCompliance message and row height normalization completed successfully")
        
    except Exception as e:
        print(f"Error applying compliance message and row heights: {e}")
        raise e

def create_processing_summary_report(processed_data):
    """Create a summary report of the processing results"""
    try:
        # Create a new workbook
        report_wb = Workbook()
        report_wb.remove(report_wb.active)  # Remove default sheet
        
        # Create summary sheet
        summary_sheet = report_wb.create_sheet("Processing Summary")
        
        # Define styles
        header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        subheader_font = Font(name='Calibri', size=12, bold=True)
        data_font = Font(name='Calibri', size=11)
        
        # Add main header
        summary_sheet.cell(row=1, column=1, value="Asset Review Non Compliance Points - Processing Summary")
        summary_sheet.cell(row=1, column=1).font = header_font
        summary_sheet.cell(row=1, column=1).fill = header_fill
        summary_sheet.cell(row=1, column=1).alignment = header_alignment
        summary_sheet.merge_cells('A1:D1')
        
        # Add generation timestamp
        summary_sheet.cell(row=2, column=1, value=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        summary_sheet.cell(row=2, column=1).font = data_font
        
        # Add spacing
        current_row = 4
        
        # Overall Statistics
        summary_sheet.cell(row=current_row, column=1, value="Overall Processing Statistics:")
        summary_sheet.cell(row=current_row, column=1).font = subheader_font
        current_row += 2
        
        stats_data = [
            ("Total Sheets Processed:", len(processed_data['sheets'])),
            ("Total Rows Before Processing:", processed_data['total_rows_before']),
            ("Total Rows After Processing:", processed_data['total_rows_after']),
            ("Total Rows Removed:", processed_data['rows_removed']),
            ("Total Non-Compliance Rows Found:", processed_data['non_compliance_rows_found'])
        ]
        
        for label, value in stats_data:
            summary_sheet.cell(row=current_row, column=1, value=label)
            summary_sheet.cell(row=current_row, column=1).font = data_font
            summary_sheet.cell(row=current_row, column=2, value=value)
            summary_sheet.cell(row=current_row, column=2).font = data_font
            current_row += 1
        
        # Add spacing
        current_row += 2
        
        # Sheet-wise Breakdown
        summary_sheet.cell(row=current_row, column=1, value="Sheet-wise Processing Breakdown:")
        summary_sheet.cell(row=current_row, column=1).font = subheader_font
        current_row += 2
        
        # Headers for breakdown table
        headers = ["Sheet Name", "Original Rows", "Filtered Rows", "Rows Removed", "Non-Compliance Rows"]
        for i, header in enumerate(headers, 1):
            cell = summary_sheet.cell(row=current_row, column=i, value=header)
            cell.font = subheader_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        current_row += 1
        
        # Add sheet data
        for sheet_data in processed_data['sheets']:
            row_data = [
                sheet_data['name'],
                sheet_data['original_rows'],
                sheet_data['filtered_rows'],
                sheet_data['rows_removed'],
                sheet_data['non_compliance_rows']
            ]
            
            for i, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=current_row, column=i, value=value)
                cell.font = data_font
            current_row += 1
        
        # Add processing notes
        current_row += 2
        summary_sheet.cell(row=current_row, column=1, value="Processing Notes:")
        summary_sheet.cell(row=current_row, column=1).font = subheader_font
        current_row += 2
        
        notes = [
            "• Only rows with 'Non-Compliance' values or multiple entries in Column C were retained",
            "• Valid Non-Compliance variations include: Non-Compliance, Non-compliance, non-Compliance, Noncompliance, noncompliance, NonCompliance",
            "• Rows with multiple entries (like 'Compliance/Non-Compliance') were kept regardless of content",
            "• All other single-entry rows were removed along with their associated POC images in columns I, J, K",
            "• Original Excel file formatting and structure were preserved completely",
            "• Processing was performed using openpyxl to maintain all original formatting"
        ]
        
        for note in notes:
            summary_sheet.cell(row=current_row, column=1, value=note)
            summary_sheet.cell(row=current_row, column=1).font = data_font
            current_row += 1
        
        # Set fixed column widths to avoid merged cell issues
        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 15
        summary_sheet.column_dimensions['C'].width = 15
        summary_sheet.column_dimensions['D'].width = 15
        summary_sheet.column_dimensions['E'].width = 20
        
        # Save to BytesIO
        output_buffer = BytesIO()
        report_wb.save(output_buffer)
        report_wb.close()
        
        output_buffer.seek(0)
        excel_data = output_buffer.getvalue()
        output_buffer.close()
        
        print(f"Processing summary report created successfully - Size: {len(excel_data)} bytes")
        return excel_data
        
    except Exception as e:
        print(f"Error creating processing summary report: {e}")
        raise e

@asset_review_non_compliance_bp.route('/test_asset_review_non_compliance_route')
def test_asset_review_non_compliance_route():
    return "Asset Review Non Compliance Points route is working!"

@asset_review_non_compliance_bp.route('/process_asset_review_non_compliance', methods=['POST'])
@login_required
def process_asset_review_non_compliance():
    print("="*50)
    print("ASSET REVIEW NON COMPLIANCE POINTS FORM SUBMISSION RECEIVED!")
    print("="*50)
    
    try:
        # Check if Excel file is present in the request
        if 'excelFile' not in request.files:
            flash('Excel file is required!', 'error')
            return redirect(url_for('audit_dashboard'))
        
        excel_file = request.files['excelFile']
        
        # Check if file is selected
        if excel_file.filename == '':
            flash('Please select an Excel file!', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Validate Excel file extension
        if not allowed_file(excel_file.filename, ALLOWED_EXCEL_EXTENSIONS):
            flash('Invalid Excel file format! Please upload .xlsx or .xls files only.', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Create temporary file
        temp_excel_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_excel_file.close()
        
        try:
            # Secure the filename
            excel_filename = secure_filename(excel_file.filename)
            
            # Save uploaded file to temporary file
            excel_file.save(temp_excel_file.name)
            
            print(f"Excel file saved to: {temp_excel_file.name}")
            
            # Process the Excel file
            print("\n" + "="*60)
            print("PROCESSING EXCEL FILE FOR NON-COMPLIANCE FILTERING")
            print("="*60)
            processed_data = process_excel_file(temp_excel_file.name)
            
            if processed_data['total_rows_before'] > 0:
                print(f"\nSuccessfully processed {processed_data['total_rows_before']} rows from {len(processed_data['sheets'])} sheets")
                print(f"Found {processed_data['non_compliance_rows_found']} Non-Compliance rows")
                print(f"Removed {processed_data['rows_removed']} rows that were not Non-Compliance")
                print(f"Final result: {processed_data['total_rows_after']} rows retained")
                
                # Apply compliance message and row height normalization just before download
                print("\n" + "="*60)
                print("APPLYING COMPLIANCE MESSAGE AND ROW HEIGHT NORMALIZATION")
                print("="*60)
                apply_compliance_message_and_row_heights(temp_excel_file.name)
                
                # Create output filename
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_filename = f"Asset_Review_Non_Compliance_Points.xlsx"
                
                # Create processing summary report
                summary_data = create_processing_summary_report(processed_data)
                
                if summary_data:
                    print(f"Processing summary report created successfully - Size: {len(summary_data)} bytes")
                    
                    # Send the processed Excel file for download
                    print("Sending processed Excel file for download...")
                    try:
                        # Read the processed Excel file
                        with open(temp_excel_file.name, 'rb') as f:
                            excel_data = f.read()
                        
                        response = Response(
                            excel_data,
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            headers={
                                'Content-Disposition': f'attachment; filename="{output_filename}"',
                                'Content-Length': str(len(excel_data)),
                                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                'Cache-Control': 'no-cache, no-store, must-revalidate',
                                'Pragma': 'no-cache',
                                'Expires': '0',
                                'X-Content-Type-Options': 'nosniff',
                                'X-Download-Options': 'noopen',
                                'Content-Transfer-Encoding': 'binary'
                            }
                        )
                        
                        print("Processed Excel file sent successfully!")
                        
                        # Schedule cleanup after response is sent
                        def cleanup_temp_files():
                            try:
                                if os.path.exists(temp_excel_file.name):
                                    os.unlink(temp_excel_file.name)
                                print("Temporary files cleaned up")
                            except Exception as e:
                                print(f"Error cleaning up temp files: {e}")
                        
                        response.call_on_close(cleanup_temp_files)
                        return response
                        
                    except Exception as download_error:
                        print(f"Error sending file: {download_error}")
                        flash(f'Error downloading file: {download_error}', 'error')
                        return redirect(url_for('audit_dashboard'))
                else:
                    flash('Failed to create processing summary report!', 'error')
                    return redirect(url_for('audit_dashboard'))
            else:
                flash('No data found in the Excel file!', 'error')
                return redirect(url_for('audit_dashboard'))
                
        except Exception as e:
            # Clean up the temporary files if there was an error
            try:
                if os.path.exists(temp_excel_file.name):
                    os.unlink(temp_excel_file.name)
            except:
                pass
            raise e
    
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        flash(f'Error processing file: {str(e)}', 'error')
        return redirect(url_for('audit_dashboard'))
