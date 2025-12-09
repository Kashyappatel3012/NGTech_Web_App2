from flask import Blueprint, request, flash, redirect, url_for, send_file
from flask_login import login_required
from datetime import datetime
import os
import tempfile
import shutil
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.cell.cell import MergedCell
import zipfile

# Create blueprint
application_evidence_bp = Blueprint('application_evidence_bp', __name__)

# Allowed file extensions
ALLOWED_EXCEL_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename, allowed_extensions):
    """Check if the uploaded file has an allowed extension"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in allowed_extensions

def extract_and_list_images_from_zip(zip_file_path, extract_to_dir):
    """Extract ZIP file and list all image files contained within it, organized by base number (17.1, 17.2, etc.)"""
    try:
        image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp', '.svg'}
        image_files = []
        extracted_images = {}  # Dictionary to store images by base number (e.g., "17.1", "17.2")
        
        with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            # Get list of all files in the ZIP
            file_list = zip_ref.namelist()
            
            print(f"Total files in ZIP: {len(file_list)}")
            print("="*50)
            print("FILES IN ZIP ARCHIVE:")
            print("="*50)
            
            for file_name in file_list:
                print(f"- {file_name}")
                
                # Check if file is an image
                file_extension = os.path.splitext(file_name)[1].lower()
                if file_extension in image_extensions:
                    image_files.append(file_name)
                    
                    # Extract image to temporary directory
                    try:
                        # Get just the filename without path
                        filename = os.path.basename(file_name)
                        print(f"Extracted filename from path '{file_name}': '{filename}'")
                        
                        # Remove extension to get the number
                        name_without_ext = os.path.splitext(filename)[0]
                        print(f"Filename without extension: '{name_without_ext}'")
                        
                        # Extract base number (e.g., "17.1" from "17.1", "17.1_1", "17.1 3", etc.)
                        base_number = extract_base_number(name_without_ext)
                        
                        if base_number:
                            extracted_path = os.path.join(extract_to_dir, filename)
                            
                            # Extract the file
                            with zip_ref.open(file_name) as source, open(extracted_path, 'wb') as target:
                                target.write(source.read())
                            
                            # Store image with base number as key (e.g., "17.1")
                            if base_number not in extracted_images:
                                extracted_images[base_number] = []
                            extracted_images[base_number].append(extracted_path)
                            print(f"Extracted image {base_number}: {filename}")
                            
                    except Exception as e:
                        print(f"Error extracting {file_name}: {e}")
            
            print("="*50)
            print(f"IMAGE FILES FOUND: {len(image_files)}")
            print(f"EXTRACTED NUMBERED IMAGES: {sum(len(images) for images in extracted_images.values())}")
            print("="*50)
            
            if extracted_images:
                for base_num in sorted(extracted_images.keys(), key=lambda x: float(x.split('.')[1])):
                    images = extracted_images[base_num]
                    print(f"Base number {base_num}: {len(images)} images")
                    for i, path in enumerate(images):
                        print(f"  - Image {i+1}: {os.path.basename(path)}")
            else:
                print("No numbered image files found in the ZIP archive.")
            
            print("="*50)
            
        return extracted_images
        
    except Exception as e:
        print(f"Error extracting ZIP file: {str(e)}")
        return {}

def extract_base_number(filename):
    """Extract the base number (17.1, 17.2, etc.) from filename for image matching"""
    try:
        # The filename parameter is already without extension (passed from extract_and_list_images_from_zip)
        # Extract base number - handle formats like:
        # 17.1, 17.1_1, 17.1 3, 17.1_2, etc.
        # We want to extract "17.1" from all these variations
        
        # First, split by space to get the first part (handles "17.1 3" -> "17.1")
        # split always returns at least one element, so just take the first one
        first_part = filename.split(' ')[0]
        
        # Then, split by underscore to get the base part (handles "17.1_1" -> "17.1")
        # split always returns at least one element, so just take the first one
        base_part = first_part.split('_')[0]
        
        print(f"Processing filename: {filename}")
        print(f"First part: {first_part}")
        print(f"Base part: {base_part}")
        
        # Check if it matches pattern 17.X where X is 1-41
        if '.' in base_part:
            parts = base_part.split('.')
            if len(parts) >= 2:
                prefix = parts[0]
                decimal_part = parts[1]
                
                print(f"Prefix: {prefix}, Decimal part: {decimal_part}")
                
                # Only process if the prefix is '17'
                if prefix == '17':
                    try:
                        decimal_num = int(decimal_part)
                        if 1 <= decimal_num <= 41:
                            base_number = f"17.{decimal_num}"
                            print(f"Extracted base number: {base_number}")
                            return base_number
                        else:
                            print(f"Decimal part {decimal_num} is out of range (1-41)")
                            return None
                    except ValueError:
                        print(f"Invalid decimal part: {decimal_part}")
                        return None
                else:
                    print(f"Ignoring filename {filename} - prefix '{prefix}' is not '17'")
                    return None
            else:
                print(f"Invalid decimal format in {base_part}")
                return None
        else:
            print(f"No decimal point found in {base_part}")
            return None
    except Exception as e:
        print(f"Error parsing filename {filename}: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_question_to_number_mapping():
    """Create mapping from question text to question number (17.1 to 17.41)"""
    mapping = {
        "Verify whether the password is displayed in an encrypted format or not?": "17.1",
        "Verify whether the password rules are implemented on pages like signup, forgot password, and change password?": "17.2",
        "Whether the user is allowed to log in with an old password after a password change?": "17.3",
        "Whether passwords are stored in cookies or not?": "17.4",
        "Is session timeout functionality properly working?": "17.5",
        "Whether the user can upload a file?": "17.6",
        "Whether the web application is using HTTPS (SSL certificate) or not?": "17.7",
        "Does the software lock the user ID after 3 unsuccessful login attempts?": "17.8",
        "Does the software allow the same user to be both maker and checker of the same transaction?": "17.9",
        "Whether proper consistency/concurrency of user inputs is maintained if two users access the same record simultaneously?": "17.10",
        "Whether key data is authorized by an appropriate level of users and kept secure?": "17.11",
        "Are the transactions for the day identifiable?": "17.12",
        "Is there an event log for batch processes?": "17.13",
        "Whether the software maintains an audit trail to trace modification/deletion/addition with user ID?": "17.14",
        "Whether the format, contents, accuracy, and utility of system-generated reports are appropriate?": "17.15",
        "Does the output contain key control information needed to validate accuracy and completeness?": "17.16",
        "Is manual intervention possible when data passes from one process to another?": "17.17",
        "Does the software prevent the same user from entering and verifying the same transaction?": "17.18",
        "If transactions are authorized manually, are there controls to ensure proper authorization?": "17.19",
        "Is the Change Management procedure—including testing & documentation—followed?": "17.20",
        "Verify whether the wrong password policy is defined or not?": "17.21",
        "Whether 2FA is available or not?": "17.22",
        "What happens if a user deletes cookies while on the site?": "17.23",
        "Whether the user gets redirected to a custom error page during functionality failure?": "17.24",
        "Does the application open in a different browser?": "17.25",
        "Whether the application is responsive or not?": "17.26",
        "Whether the error messages display any important information?": "17.27",
        "Whether negative numbers are allowed in numeric fields?": "17.28",
        "Whether a confirmation message is displayed for update and delete operations?": "17.29",
        "Whether a pop-up message is displayed when data reaches the maximum field size?": "17.30",
        "Whether special characters are accepted in input fields?": "17.31",
        "Whether button functionality is available and working correctly?": "17.32",
        "Whether error messages are displayed correctly?": "17.33",
        "Does the software allow the creation of duplicate user IDs with the same name?": "17.34",
        "Is there a mechanism to identify and remove outdated libraries, frameworks, or plugins that introduce vulnerabilities?": "17.35",
        "Is data transmission encrypted using TLS 1.2/1.3 to prevent interception?": "17.36",
        "Is CBS backup ensured? What is the backup duration and whether the backup is secure?": "17.37",
        "Does the software maintain password history?": "17.38",
        "Are all asset movements supported by suitable written authorizations?": "17.39",
        "Whether input validation for name fields is available?": "17.40",
        "Whether input validation for numeric fields is available?": "17.41"
    }
    return mapping

def insert_images_to_excel(excel_path, image_mapping):
    """Insert images into Excel based on column B text matching question numbers"""
    try:
        # Load the existing workbook
        wb = load_workbook(excel_path)
        
        # Get question to number mapping
        question_mapping = create_question_to_number_mapping()
        
        # Process each worksheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            print(f"Processing images for sheet: {sheet_name}")
            
            # Define columns in priority order: J, K, I
            # First image goes to J, second to K, third to I
            columns = ['J', 'K', 'I']
            
            # Scan column B to find questions and match with images
            for row in range(1, ws.max_row + 1):
                # Get cell value from column B (column 2)
                cell_b = ws.cell(row=row, column=2)
                cell_value = cell_b.value
                
                if cell_value and isinstance(cell_value, str):
                    # Clean the text (remove extra spaces, newlines, etc.)
                    question_text = cell_value.strip()
                    
                    # Try to find matching question in mapping
                    matched_question = None
                    question_number = None
                    
                    # Direct match
                    if question_text in question_mapping:
                        matched_question = question_text
                        question_number = question_mapping[question_text]
                    else:
                        # Try partial match (in case of slight variations)
                        for q_text, q_num in question_mapping.items():
                            # Check if the question text contains key parts of the mapping text
                            # or vice versa
                            if question_text in q_text or q_text in question_text:
                                matched_question = q_text
                                question_number = q_num
                                break
                    
                    if question_number and question_number in image_mapping:
                        # Found matching question and images
                        image_paths = image_mapping[question_number]
                        
                        # Take maximum 3 images (one for each column: J, K, I)
                        images_to_process = image_paths[:3]
                        
                        print(f"Row {row}: Found question '{matched_question[:50]}...' -> {question_number}")
                        print(f"  Inserting {len(images_to_process)} images into row {row}")
                        
                        # Insert images into J, K, I columns
                        for i, image_path in enumerate(images_to_process):
                            if i < len(columns):
                                target_cell = f"{columns[i]}{row}"
                                
                                try:
                                    # Load the image
                                    img = Image(image_path)
                                    
                                    # Get original dimensions and resize
                                    original_width = img.width
                                    original_height = img.height
                                    
                                    # Resize to 25px height with aspect ratio maintained
                                    target_height = 25
                                    aspect_ratio = original_width / original_height if original_height else 1
                                    img.height = target_height
                                    img.width = int(target_height * aspect_ratio)
                                    
                                    # Try to center the image by calculating proper offsets
                                    cell_width_pixels = 96  # Column width 12 in pixels
                                    cell_height_pixels = 25  # Row height in pixels
                                    
                                    # Calculate center offsets
                                    center_x = (cell_width_pixels - img.width) / 2
                                    center_y = (cell_height_pixels - img.height) / 2
                                    
                                    # Ensure offsets are positive
                                    center_x = max(0, center_x)
                                    center_y = max(0, center_y)
                                    
                                    # Set image position
                                    img.anchor = target_cell
                                    img.left = int(center_x)
                                    img.top = int(center_y)
                                    
                                    # Add image to worksheet
                                    ws.add_image(img)
                                    
                                    print(f"  Inserted image {i+1} into {target_cell} (size: {img.width}x{img.height})")
                                    
                                except Exception as e:
                                    print(f"  Error inserting image {i+1} into {target_cell}: {e}")
                            else:
                                print(f"  Ignoring image {i+1} - maximum 3 images per question")
        
        # Save the workbook
        wb.save(excel_path)
        total_images = sum(len(images) for images in image_mapping.values())
        print(f"Successfully inserted images into Excel file")
        
    except Exception as e:
        print(f"Error inserting images into Excel: {e}")
        raise e

def process_excel_file(file_path):
    """Process the Excel file to add column I, J, K with borders based on column F data"""
    try:
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Process each worksheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Find the last row with data in column F
            last_row_with_data = 0
            for row in range(1, worksheet.max_row + 1):
                cell_f = worksheet[f'F{row}']
                if cell_f.value is not None and str(cell_f.value).strip() != '':
                    last_row_with_data = row
            
            print(f"Last row with data in column F: {last_row_with_data}")
            
            # If there's data in column F, add columns I, J, and K with borders
            if last_row_with_data > 0:
                # Insert new columns I, J, and K (this will shift existing columns to the right)
                worksheet.insert_cols(9, 3)  # Insert 3 columns starting at position 9 (columns I, J, K)
                
                # Define font style for header
                header_font = Font(
                    name='Calibri',
                    size=12,
                    bold=True,
                    color='FFFFFF'  # White font color
                )
                
                # Define alignment for header
                header_alignment = Alignment(
                    horizontal='center',
                    vertical='center'
                )
                
                # Define background fill for header
                header_fill = PatternFill(
                    start_color='366092',  # Background color
                    end_color='366092',
                    fill_type='solid'
                )
                
                # Define border styles
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Border for I column (no right border)
                i_border = Border(
                    left=Side(style='thin'),
                    right=Side(style=None),  # No right border
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Border for J column (no left or right border)
                j_border = Border(
                    left=Side(style=None),   # No left border
                    right=Side(style=None),  # No right border
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Border for K column (no left border)
                k_border = Border(
                    left=Side(style=None),  # No left border
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Set value and formatting on I2 BEFORE merging (to avoid MergedCell read-only error)
                cell_i2 = worksheet.cell(row=2, column=9)  # I2 is row 2, column 9
                cell_i2.value = 'POC'  # Header
                cell_i2.font = header_font
                cell_i2.alignment = header_alignment
                cell_i2.border = thin_border
                cell_i2.fill = header_fill
                
                # Merge cells I2, J2, and K2 for the header (moved from I1:K1 to I2:K2)
                worksheet.merge_cells('I2:K2')
                
                # Apply borders and formatting to columns I, J, and K from row 1 to last_row_with_data
                # Column I = 9, J = 10, K = 11
                col_mapping = {'I': 9, 'J': 10, 'K': 11}
                
                for row in range(1, last_row_with_data + 1):
                    for col_letter in ['I', 'J', 'K']:
                        # Use worksheet.cell() to get actual cell object (not MergedCell)
                        col_num = col_mapping[col_letter]
                        cell = worksheet.cell(row=row, column=col_num)
                        
                        # Apply appropriate border style based on column
                        if col_letter == 'I':
                            cell.border = i_border  # No right border
                        elif col_letter == 'J':
                            cell.border = j_border  # No left or right border
                        elif col_letter == 'K':
                            cell.border = k_border  # Full border
                        
                        # Add empty content to non-header cells
                        # Skip setting value on MergedCell objects (they are read-only)
                        # Also skip row 2 for merged POC header (I2-K2)
                        if row > 2:  # Skip row 1 (might have A1-K1 merge) and row 2 (I2-K2 merged)
                            # Check if this cell is not a MergedCell before setting value
                            if not isinstance(cell, MergedCell):
                                cell.value = ''
                            # If it's a MergedCell, skip setting value (it's read-only)
                
                # Set column widths for I, J, and K to 12
                for col_letter in ['I', 'J', 'K']:
                    worksheet.column_dimensions[col_letter].width = 12
                
                # Increase all row heights by 10 points
                for row in range(1, worksheet.max_row + 1):
                    current_height = worksheet.row_dimensions[row].height
                    if current_height is None:
                        # Default height is usually around 15 points
                        new_height = 30
                    else:
                        new_height = current_height + 10
                    
                    worksheet.row_dimensions[row].height = new_height
                
                print(f"Added columns I, J, and K with merged header 'POC' in I2:K2, Times New Roman 12pt bold centered, borders from I1 to K{last_row_with_data}, width set to 12")
                print(f"Increased all row heights by 10 points (total rows: {worksheet.max_row})")
        
        # Create a new temporary file for the processed workbook
        processed_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        processed_temp_file.close()
        
        # Save the processed workbook
        workbook.save(processed_temp_file.name)
        workbook.close()
        
        print(f"Processed Excel file saved to: {processed_temp_file.name}")
        return processed_temp_file.name
        
    except Exception as e:
        print(f"Error processing Excel file: {str(e)}")
        raise e

@application_evidence_bp.route('/test_application_evidence_route')
def test_application_evidence_route():
    return "Application Evidence route is working!"

@application_evidence_bp.route('/process_application_evidence', methods=['POST'])
@login_required
def process_application_evidence():
    print("="*50)
    print("APPLICATION EVIDENCE FORM SUBMISSION RECEIVED!")
    print("="*50)
    
    try:
        # Check if both files are present in the request
        if 'excelFile' not in request.files or 'zipFile' not in request.files:
            flash('Both Excel file and ZIP file are required!', 'error')
            return redirect(url_for('audit_dashboard'))
        
        excel_file = request.files['excelFile']
        zip_file = request.files['zipFile']
        
        # Check if files are selected
        if excel_file.filename == '' or zip_file.filename == '':
            flash('Please select both Excel file and ZIP file!', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Validate Excel file extension
        if not allowed_file(excel_file.filename, ALLOWED_EXCEL_EXTENSIONS):
            flash('Invalid Excel file format! Please upload .xlsx or .xls files only.', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Validate ZIP file extension
        if not allowed_file(zip_file.filename, {'zip'}):
            flash('Invalid ZIP file format! Please upload .zip files only.', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Create temporary files that won't be automatically deleted
        temp_excel_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_excel_file.close()  # Close the file handle to avoid locking issues
        
        temp_zip_file = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        temp_zip_file.close()  # Close the file handle to avoid locking issues
        
        try:
            # Secure the filenames
            excel_filename = secure_filename(excel_file.filename)
            zip_filename = secure_filename(zip_file.filename)
            
            # Get original Excel filename for output and add "Review" after filename
            original_excel_filename = excel_file.filename
            # Ensure it has .xlsx extension
            if original_excel_filename:
                # Get filename without path
                original_filename = os.path.basename(original_excel_filename)
                # Split filename and extension
                name_without_ext, file_ext = os.path.splitext(original_filename)
                # Add "Review" before extension (e.g., "abc.xlsx" -> "abc Review.xlsx")
                output_filename = f"{name_without_ext} Review.xlsx"
            else:
                output_filename = "Application Review.xlsx"
            
            print(f"Original Excel filename: {original_excel_filename}")
            print(f"Output filename: {output_filename}")
            
            # Save uploaded files to temporary files
            excel_file.save(temp_excel_file.name)
            zip_file.save(temp_zip_file.name)
            
            print(f"Excel file saved to: {temp_excel_file.name}")
            print(f"ZIP file saved to: {temp_zip_file.name}")
            
            # Create temporary directory for extracted images
            temp_images_dir = tempfile.mkdtemp()
            
            # Extract and list images from ZIP file
            print("\n" + "="*60)
            print("PROCESSING ZIP FILE FOR APPLICATION EVIDENCE IMAGES")
            print("="*60)
            extracted_images = extract_and_list_images_from_zip(temp_zip_file.name, temp_images_dir)
            
            # Process the Excel file to add column I, J, K with borders
            processed_file_path = process_excel_file(temp_excel_file.name)
            
            # Insert images into Excel if any were extracted
            if extracted_images:
                print("\n" + "="*60)
                print("INSERTING APPLICATION EVIDENCE IMAGES INTO EXCEL")
                print("="*60)
                
                # Insert images into Excel (matching column B text to question numbers)
                insert_images_to_excel(processed_file_path, extracted_images)
            else:
                print("No numbered images found to insert into Excel.")
            
            # Output filename already set above with "Review" added
            
            # Verify the file was created successfully
            if os.path.exists(processed_file_path):
                print(f"\nFile processed successfully: {output_filename}")
                print(f"Found {len(extracted_images)} numbered image files in ZIP archive")
                
                # Send the file for download
                response = send_file(
                    processed_file_path,
                    as_attachment=True,
                    download_name=output_filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
                # Schedule the temporary files for deletion after response is sent
                def cleanup_temp_files():
                    try:
                        if os.path.exists(temp_excel_file.name):
                            os.unlink(temp_excel_file.name)
                        if os.path.exists(temp_zip_file.name):
                            os.unlink(temp_zip_file.name)
                        if os.path.exists(processed_file_path):
                            os.unlink(processed_file_path)
                        # Clean up temporary images directory
                        if 'temp_images_dir' in locals() and os.path.exists(temp_images_dir):
                            shutil.rmtree(temp_images_dir)
                    except Exception as e:
                        print(f"Error cleaning up temp files: {e}")
                
                # Use response callback to clean up the files after sending
                response.call_on_close(cleanup_temp_files)
                
                return response
            else:
                flash('Error processing the Excel file!', 'error')
                return redirect(url_for('audit_dashboard'))
                
        except Exception as e:
            # Clean up the temporary files if there was an error
            try:
                if os.path.exists(temp_excel_file.name):
                    os.unlink(temp_excel_file.name)
                if os.path.exists(temp_zip_file.name):
                    os.unlink(temp_zip_file.name)
                # Also clean up processed file if it exists
                if 'processed_file_path' in locals() and os.path.exists(processed_file_path):
                    os.unlink(processed_file_path)
                # Clean up temporary images directory
                if 'temp_images_dir' in locals() and os.path.exists(temp_images_dir):
                    shutil.rmtree(temp_images_dir)
            except:
                pass
            raise e
    
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        flash(f'Error processing file: {str(e)}', 'error')
        return redirect(url_for('audit_dashboard'))
