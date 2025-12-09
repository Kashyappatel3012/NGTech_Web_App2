from flask import Blueprint, request, flash, redirect, url_for, send_file
from flask_login import login_required
from datetime import datetime
import os
import tempfile
import shutil
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.drawing.image import Image
import zipfile

# Create blueprint
internet_banking_evidence_bp = Blueprint('internet_banking_evidence_bp', __name__)

# Allowed file extensions
ALLOWED_EXCEL_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename, allowed_extensions):
    """Check if the uploaded file has an allowed extension"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in allowed_extensions

def extract_and_list_images_from_zip(zip_file_path, extract_to_dir):
    """Extract ZIP file and list all image files contained within it"""
    try:
        image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp', '.svg'}
        image_files = []
        extracted_images = {}  # Dictionary to store images by base number and column
        
        with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            # Get list of all files in the ZIP
            file_list = zip_ref.namelist()
            
            print(f"Total files in ZIP: {len(file_list)}")
            print("="*50)
            print("FILES IN ZIP ARCHIVE:")
            print("="*50)
            
            for file_name in file_list:
                print(f"- {file_name}")
                
                # Check if file is an image
                file_extension = os.path.splitext(file_name)[1].lower()
                if file_extension in image_extensions:
                    image_files.append(file_name)
                    
                    # Extract image to temporary directory
                    try:
                        # Get just the filename without path
                        filename = os.path.basename(file_name)
                        # Remove extension to get the number
                        name_without_ext = os.path.splitext(filename)[0]
                        
                        # Extract base number (ignore everything after space or underscore, but keep decimal point)
                        base_number = extract_base_number(name_without_ext)
                        
                        if base_number:
                            extracted_path = os.path.join(extract_to_dir, filename)
                            
                            # Extract the file
                            with zip_ref.open(file_name) as source, open(extracted_path, 'wb') as target:
                                target.write(source.read())
                            
                            # Store image with base number as key
                            if base_number not in extracted_images:
                                extracted_images[base_number] = []
                            extracted_images[base_number].append(extracted_path)
                            print(f"Extracted image {base_number}: {filename}")
                            
                    except Exception as e:
                        print(f"Error extracting {file_name}: {e}")
            
            print("="*50)
            print(f"IMAGE FILES FOUND: {len(image_files)}")
            print(f"EXTRACTED NUMBERED IMAGES: {sum(len(images) for images in extracted_images.values())}")
            print("="*50)
            
            if extracted_images:
                for base_num, images in sorted(extracted_images.items()):
                    print(f"Base number {base_num}: {len(images)} images")
                    for i, path in enumerate(images):
                        print(f"  - Image {i+1}: {os.path.basename(path)}")
            else:
                print("No numbered image files found in the ZIP archive.")
            
            print("="*50)
            
        return extracted_images
        
    except Exception as e:
        print(f"Error extracting ZIP file: {str(e)}")
        return {}

def extract_base_number(filename):
    """Extract only the decimal part from filename for image placement logic, only if it starts with '18.'"""
    try:
        # Remove everything after space or underscore (but keep the decimal part)
        # 18.1 1 -> 18.1
        # 18.1_2 -> 18.1
        # 18.2 -> 18.2
        # 18.29 -> 18.29
        base_part = filename.split(' ')[0].split('_')[0]
        
        print(f"Processing filename: {filename}")
        print(f"Base part: {base_part}")
        
        # Check if it's a decimal number (contains a dot)
        if '.' in base_part:
            # Split by decimal point to check prefix
            parts = base_part.split('.')
            if len(parts) >= 2:
                prefix = parts[0]
                decimal_part = parts[1]
                
                # Only process if the prefix is '18'
                if prefix == '18':
                    result = int(decimal_part)
                    print(f"Decimal part extracted: {decimal_part} -> {result} (prefix '18' confirmed)")
                    return result
                else:
                    print(f"Ignoring filename {filename} - prefix '{prefix}' is not '18'")
                    return None
            else:
                print(f"Invalid decimal format in {base_part}")
                return None
        else:
            print(f"No decimal point found in {base_part}")
            return None
    except (ValueError, IndexError) as e:
        print(f"Error parsing filename {filename}: {e}")
        return None

def create_image_cell_mapping():
    """Create mapping from decimal numbers to Excel cells for Internet Banking Evidence"""
    mapping = {}
    
    # Define the mapping based on the provided requirements
    # Decimal 1 -> J2, K2, I2 (if one then in J2 if two then J2 and K2 if 3 then J2, K2, I2)
    # Decimal 2 -> J3, K3, I3
    # Decimal 3 -> J4, K4, I4
    # .........
    # Decimal 29 -> J30, K30, I30
    
    # Create mappings for decimal numbers 1 to 29
    for decimal_num in range(1, 30):  # 1 to 29 inclusive
        # For decimal x, row number = x + 1
        # Decimal 1 -> 1 + 1 = 2, Decimal 2 -> 2 + 1 = 3, Decimal 29 -> 29 + 1 = 30
        row_num = decimal_num + 1
        
        # Store the base row number for this decimal
        mapping[decimal_num] = row_num
        print(f"Mapping decimal {decimal_num} to row {row_num}")
    
    return mapping

def insert_images_to_excel(excel_path, image_mapping, cell_mapping):
    """Insert images into Excel cells based on the mapping with J, K, I priority"""
    try:
        # Load the existing workbook
        wb = load_workbook(excel_path)
        
        # Process each worksheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            print(f"Processing images for sheet: {sheet_name}")
            
            # Insert images based on mapping
            for decimal_number, image_paths in image_mapping.items():
                if decimal_number in cell_mapping:
                    row_num = cell_mapping[decimal_number]  # Get the row number for this decimal
                    
                    # Define columns in priority order: J, K, I
                    # First image goes to J, second to K, third to I
                    columns = ['J', 'K', 'I']
                    
                    # Take maximum 3 images (one for each column)
                    images_to_process = image_paths[:3]
                    
                    print(f"Processing {len(images_to_process)} images for decimal {decimal_number} in row {row_num}")
                    
                    for i, image_path in enumerate(images_to_process):
                        if i < len(columns):
                            target_cell = f"{columns[i]}{row_num}"
                            
                            try:
                                # Load the image
                                img = Image(image_path)
                                
                                # Get original dimensions and resize
                                original_width = img.width
                                original_height = img.height
                                
                                # Resize to 25px height with aspect ratio maintained
                                target_height = 25
                                aspect_ratio = original_width / original_height if original_height else 1
                                img.height = target_height
                                img.width = int(target_height * aspect_ratio)
                                
                                # Try to center the image by calculating proper offsets
                                cell_width_pixels = 96  # Column width 12 in pixels
                                cell_height_pixels = 25  # Row height in pixels
                                
                                # Calculate center offsets
                                center_x = (cell_width_pixels - img.width) / 2
                                center_y = (cell_height_pixels - img.height) / 2
                                
                                # Ensure offsets are positive
                                center_x = max(0, center_x)
                                center_y = max(0, center_y)
                                
                                # Set image position
                                img.anchor = target_cell
                                img.left = int(center_x)
                                img.top = int(center_y)
                                
                                # Add image to worksheet
                                ws.add_image(img)
                                
                                print(f"Inserted image {decimal_number}_{i+1} into {target_cell} (size: {img.width}x{img.height}, original: {original_width}x{original_height})")
                                
                            except Exception as e:
                                print(f"Error inserting image {decimal_number}_{i+1} into {target_cell}: {e}")
                        else:
                            print(f"Ignoring image {decimal_number}_{i+1} - maximum 3 images per decimal number")
        
        # Save the workbook
        wb.save(excel_path)
        total_images = sum(len(images) for images in image_mapping.values())
        print(f"Successfully inserted {total_images} images into Excel file")
        
    except Exception as e:
        print(f"Error inserting images into Excel: {e}")
        raise e

def process_excel_file(file_path):
    """Process the Excel file to add column I, J, K with borders based on column F data"""
    try:
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Process each worksheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Find the last row with data in column F
            last_row_with_data = 0
            for row in range(1, worksheet.max_row + 1):
                cell_f = worksheet[f'F{row}']
                if cell_f.value is not None and str(cell_f.value).strip() != '':
                    last_row_with_data = row
            
            print(f"Last row with data in column F: {last_row_with_data}")
            
            # If there's data in column F, add columns I, J, and K with borders
            if last_row_with_data > 0:
                # Insert new columns I, J, and K (this will shift existing columns to the right)
                worksheet.insert_cols(9, 3)  # Insert 3 columns starting at position 9 (columns I, J, K)
                
                # Define font style for header
                header_font = Font(
                    name='Calibri',
                    size=12,
                    bold=True,
                    color='FFFFFF'  # White font color
                )
                
                # Define alignment for header
                header_alignment = Alignment(
                    horizontal='center',
                    vertical='center'
                )
                
                # Define background fill for header
                header_fill = PatternFill(
                    start_color='366092',  # Background color
                    end_color='366092',
                    fill_type='solid'
                )
                
                # Define border styles
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Border for I column (no right border)
                i_border = Border(
                    left=Side(style='thin'),
                    right=Side(style=None),  # No right border
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Border for J column (no left or right border)
                j_border = Border(
                    left=Side(style=None),   # No left border
                    right=Side(style=None),  # No right border
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Border for K column (no left border)
                k_border = Border(
                    left=Side(style=None),  # No left border
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Merge cells I1, J1, and K1 for the header
                worksheet.merge_cells('I1:K1')
                merged_cell = worksheet['I1']
                merged_cell.value = 'POC'  # Header
                merged_cell.font = header_font
                merged_cell.alignment = header_alignment
                merged_cell.border = thin_border
                merged_cell.fill = header_fill
                
                # Apply borders and formatting to columns I, J, and K from row 1 to last_row_with_data
                for row in range(1, last_row_with_data + 1):
                    for col_letter in ['I', 'J', 'K']:
                        cell = worksheet[f'{col_letter}{row}']
                        
                        # Apply appropriate border style based on column
                        if col_letter == 'I':
                            cell.border = i_border  # No right border
                        elif col_letter == 'J':
                            cell.border = j_border  # No left or right border
                        elif col_letter == 'K':
                            cell.border = k_border  # Full border
                        
                        # Add empty content to non-header cells
                        if row > 1:
                            cell.value = ''
                
                # Set column widths for I, J, and K to 12
                for col_letter in ['I', 'J', 'K']:
                    worksheet.column_dimensions[col_letter].width = 12
                
                # Increase all row heights by 10 points
                for row in range(1, worksheet.max_row + 1):
                    current_height = worksheet.row_dimensions[row].height
                    if current_height is None:
                        # Default height is usually around 15 points
                        new_height = 30
                    else:
                        new_height = current_height + 10
                    
                    worksheet.row_dimensions[row].height = new_height
                
                print(f"Added columns I, J, and K with merged header 'POC' in I1:K1, Times New Roman 12pt bold centered, borders from I1 to K{last_row_with_data}, width set to 12")
                print(f"Increased all row heights by 10 points (total rows: {worksheet.max_row})")
        
        # Create a new temporary file for the processed workbook
        processed_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        processed_temp_file.close()
        
        # Save the processed workbook
        workbook.save(processed_temp_file.name)
        workbook.close()
        
        print(f"Processed Excel file saved to: {processed_temp_file.name}")
        return processed_temp_file.name
        
    except Exception as e:
        print(f"Error processing Excel file: {str(e)}")
        raise e

@internet_banking_evidence_bp.route('/test_internet_banking_evidence_route')
def test_internet_banking_evidence_route():
    return "Internet Banking Evidence route is working!"

@internet_banking_evidence_bp.route('/process_internet_banking_evidence', methods=['POST'])
@login_required
def process_internet_banking_evidence():
    print("="*50)
    print("INTERNET BANKING EVIDENCE FORM SUBMISSION RECEIVED!")
    print("="*50)
    
    try:
        # Check if both files are present in the request
        if 'excelFile' not in request.files or 'zipFile' not in request.files:
            flash('Both Excel file and ZIP file are required!', 'error')
            return redirect(url_for('audit_dashboard'))
        
        excel_file = request.files['excelFile']
        zip_file = request.files['zipFile']
        
        # Check if files are selected
        if excel_file.filename == '' or zip_file.filename == '':
            flash('Please select both Excel file and ZIP file!', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Validate Excel file extension
        if not allowed_file(excel_file.filename, ALLOWED_EXCEL_EXTENSIONS):
            flash('Invalid Excel file format! Please upload .xlsx or .xls files only.', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Validate ZIP file extension
        if not allowed_file(zip_file.filename, {'zip'}):
            flash('Invalid ZIP file format! Please upload .zip files only.', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Create temporary files that won't be automatically deleted
        temp_excel_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_excel_file.close()  # Close the file handle to avoid locking issues
        
        temp_zip_file = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        temp_zip_file.close()  # Close the file handle to avoid locking issues
        
        try:
            # Secure the filenames
            excel_filename = secure_filename(excel_file.filename)
            zip_filename = secure_filename(zip_file.filename)
            
            # Save uploaded files to temporary files
            excel_file.save(temp_excel_file.name)
            zip_file.save(temp_zip_file.name)
            
            print(f"Excel file saved to: {temp_excel_file.name}")
            print(f"ZIP file saved to: {temp_zip_file.name}")
            
            # Create temporary directory for extracted images
            temp_images_dir = tempfile.mkdtemp()
            
            # Extract and list images from ZIP file
            print("\n" + "="*60)
            print("PROCESSING ZIP FILE FOR INTERNET BANKING EVIDENCE IMAGES")
            print("="*60)
            extracted_images = extract_and_list_images_from_zip(temp_zip_file.name, temp_images_dir)
            
            # Process the Excel file to add column I, J, K with borders
            processed_file_path = process_excel_file(temp_excel_file.name)
            
            # Insert images into Excel if any were extracted
            if extracted_images:
                print("\n" + "="*60)
                print("INSERTING INTERNET BANKING EVIDENCE IMAGES INTO EXCEL")
                print("="*60)
                
                # Create cell mapping
                cell_mapping = create_image_cell_mapping()
                
                # Insert images into Excel
                insert_images_to_excel(processed_file_path, extracted_images, cell_mapping)
            else:
                print("No numbered images found to insert into Excel.")
            
            # Create the output filename
            output_filename = "Internet Banking Review.xlsx"
            
            # Verify the file was created successfully
            if os.path.exists(processed_file_path):
                print(f"\nFile processed successfully: {output_filename}")
                print(f"Found {len(extracted_images)} numbered image files in ZIP archive")
                
                # Send the file for download
                response = send_file(
                    processed_file_path,
                    as_attachment=True,
                    download_name=output_filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
                # Schedule the temporary files for deletion after response is sent
                def cleanup_temp_files():
                    try:
                        if os.path.exists(temp_excel_file.name):
                            os.unlink(temp_excel_file.name)
                        if os.path.exists(temp_zip_file.name):
                            os.unlink(temp_zip_file.name)
                        if os.path.exists(processed_file_path):
                            os.unlink(processed_file_path)
                        # Clean up temporary images directory
                        if 'temp_images_dir' in locals() and os.path.exists(temp_images_dir):
                            shutil.rmtree(temp_images_dir)
                    except Exception as e:
                        print(f"Error cleaning up temp files: {e}")
                
                # Use response callback to clean up the files after sending
                response.call_on_close(cleanup_temp_files)
                
                return response
            else:
                flash('Error processing the Excel file!', 'error')
                return redirect(url_for('audit_dashboard'))
                
        except Exception as e:
            # Clean up the temporary files if there was an error
            try:
                if os.path.exists(temp_excel_file.name):
                    os.unlink(temp_excel_file.name)
                if os.path.exists(temp_zip_file.name):
                    os.unlink(temp_zip_file.name)
                # Also clean up processed file if it exists
                if 'processed_file_path' in locals() and os.path.exists(processed_file_path):
                    os.unlink(processed_file_path)
                # Clean up temporary images directory
                if 'temp_images_dir' in locals() and os.path.exists(temp_images_dir):
                    shutil.rmtree(temp_images_dir)
            except:
                pass
            raise e
    
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        flash(f'Error processing file: {str(e)}', 'error')
        return redirect(url_for('audit_dashboard'))
