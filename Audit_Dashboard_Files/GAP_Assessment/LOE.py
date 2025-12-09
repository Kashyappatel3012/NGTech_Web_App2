from flask import Blueprint, request, jsonify
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import tempfile
import shutil
import glob

# Create blueprint for LOE
loe_bp = Blueprint('loe', __name__)

def cleanup_loe_files():
    """Clean up old LOE Excel files from uploads directory"""
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            loe_pattern = os.path.join(upload_dir, 'LOE_*.xlsx')
            old_files = glob.glob(loe_pattern)
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

@loe_bp.route('/cleanup_loe_files', methods=['POST'])
def cleanup_loe_files_endpoint():
    """Endpoint to cleanup LOE files after download"""
    try:
        cleanup_loe_files()
        return jsonify({"success": True, "message": "Cleanup completed"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

def create_loe_excel(loe_data):
    """
    Create Excel file for LOE (Level of Exposure) with formatted data
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "LOE"
    
    # Define styles
    header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    title_font = Font(name='Times New Roman', size=12, bold=True, color='000000')
    title_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    title_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Times New Roman', size=12)
    data_alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    data_alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 20
    
    # Merge A1:C1 and add title
    ws.merge_cells('A1:C1')
    ws['A1'] = "Level of Exposure"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = title_alignment
    ws['A1'].border = thin_border
    
    # Add headers in row 2
    ws['A2'] = "Sr. No."
    ws['A2'].font = header_font
    ws['A2'].fill = header_fill
    ws['A2'].alignment = header_alignment
    ws['A2'].border = thin_border
    
    ws['B2'] = "Questions"
    ws['B2'].font = header_font
    ws['B2'].fill = header_fill
    ws['B2'].alignment = header_alignment
    ws['B2'].border = thin_border
    
    ws['C2'] = "Input"
    ws['C2'].font = header_font
    ws['C2'].fill = header_fill
    ws['C2'].alignment = header_alignment
    ws['C2'].border = thin_border
    
    # Define questions with their Sr. No. and text
    questions = [
        ("1", "Is the bank implementing CBS?"),
        ("2", "If yes, are the following modules operational? Please see maker guide (click on hyperlink) to answer questions 2(a) to 2(h)"),
        ("2)(a)", "Core Modules?"),
        ("2)(b)", "Retail Banking Module?"),
        ("2)(c)", "Loan and Advances Module?"),
        ("2)(d)", "Remittances and Services Module?"),
        ("2)(e)", "Reports and MIS Module?"),
        ("2)(f)", "Customer Information Module?"),
        ("2)(g)", "Head Office Module?"),
        ("2)(h)", "Interfaces with outside agencies?"),
        ("3", "Whether direct member of Central Payment System (RTGS / NEFT)?"),
        ("4", "Whether sub member of Central Payment System (RTGS / NEFT)?"),
        ("5", "If sub member Name of Sponsor Bank providing NEFT / RTGS facility?"),
        ("6", "Whether bank is providing Internet View facility to customers?"),
        ("7", "Whether RBI has given approval for internet Transaction facility to your bank?"),
        ("8", "If yes, date of approval?"),
        ("9", "Whether your bank is providing internet transaction facility to customers?"),
        ("10", "Has RBI permitted mobile banking?"),
        ("11", "If yes, date of permission for mobile banking"),
        ("12", "If yes, whether providing Mobile Banking facility to customers? (either through App or Smart phone answer YES)"),
        ("13", "Whether Direct Member of Cheque Truncation System (CTS)?"),
        ("14", "Whether Direct Member of Immediate payment Service (IMPS)?"),
        ("15", "Whether Direct Member of Unified Payment Interface (UPI)?"),
        ("16", "Whether Bank has its own ATM switch?"),
        ("17", "Whether Bank has SWIFT interface?"),
        ("18", "Whether hosting Data center of other banks ?"),
        ("19", "Whether providing software support to other banks? (if providing either directly or through fully owned subsidiaries then answer YES)"),
        ("20", "Whether acting as sponsor bank for other banks (DCCBs/UCBs) for CPS/CTS/UPI/IMPS etc)")
    ]
    
    # Map form field names to question numbers
    field_to_question = {
        'q1': '1',
        'q2': '2',
        'q2_A': '2)(a)',
        'q2_B': '2)(b)',
        'q2_C': '2)(c)',
        'q2_D': '2)(d)',
        'q2_E': '2)(e)',
        'q2_F': '2)(f)',
        'q2_G': '2)(g)',
        'q2_H': '2)(h)',
        'q3': '3',
        'q4': '4',
        'q5': '5',
        'q6': '6',
        'q7': '7',
        'q8': '8',
        'q9': '9',
        'q10': '10',
        'q11': '11',
        'q12': '12',
        'q13': '13',
        'q14': '14',
        'q15': '15',
        'q16': '16',
        'q17': '17',
        'q18': '18',
        'q19': '19',
        'q20': '20'
    }
    
    # Add questions and answers
    current_row = 3
    for sr_no, question_text in questions:
        # Add Sr. No.
        ws.cell(row=current_row, column=1).value = sr_no
        ws.cell(row=current_row, column=1).font = data_font
        ws.cell(row=current_row, column=1).alignment = data_alignment_center
        ws.cell(row=current_row, column=1).border = thin_border
        
        # Add Question
        ws.cell(row=current_row, column=2).value = question_text
        ws.cell(row=current_row, column=2).font = data_font
        ws.cell(row=current_row, column=2).alignment = data_alignment_left
        ws.cell(row=current_row, column=2).border = thin_border
        
        # Find the corresponding form field and add answer
        field_name = None
        for field, q_num in field_to_question.items():
            if q_num == sr_no:
                field_name = field
                break
        
        if field_name:
            answer = loe_data.get(field_name, '')
            
            # Special handling for Q8 and Q11 (dates)
            if field_name == 'q8':
                # Check if Q7 is Yes
                if loe_data.get('q7') == 'Yes' and answer:
                    # Convert date from YYYY-MM-DD to DD/MM/YYYY
                    try:
                        date_obj = datetime.strptime(answer, '%Y-%m-%d')
                        answer = date_obj.strftime('%d/%m/%Y')
                    except:
                        pass
                else:
                    answer = 'NA'
            elif field_name == 'q11':
                # Check if Q10 is Yes
                if loe_data.get('q10') == 'Yes' and answer:
                    # Convert date from YYYY-MM-DD to DD/MM/YYYY
                    try:
                        date_obj = datetime.strptime(answer, '%Y-%m-%d')
                        answer = date_obj.strftime('%d/%m/%Y')
                    except:
                        pass
                else:
                    answer = 'NA'
            
            ws.cell(row=current_row, column=3).value = answer
            ws.cell(row=current_row, column=3).font = data_font
            ws.cell(row=current_row, column=3).alignment = data_alignment_center
            ws.cell(row=current_row, column=3).border = thin_border
        
        current_row += 1
    
    # Save to a temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()
    wb.save(temp_file.name)
    wb.close()
    
    return temp_file.name

@loe_bp.route('/process_loe', methods=['POST'])
def process_loe():
    """
    Process LOE form data and generate Excel file
    """
    try:
        print("\n" + "="*80)
        print("📋 Processing LOE (Level of Exposure)")
        print("="*80)
        
        # Get form data
        form_data = request.form.to_dict()
        
        print("📋 Form Data Received:")
        for key, value in form_data.items():
            print(f"  {key}: {value}")
        
        # Clean up old LOE files before generating new one
        print("\n🧹 Cleaning up old LOE files...")
        cleanup_loe_files()
        
        # Generate Excel file
        print("\n📝 Generating Excel file...")
        excel_file_path = create_loe_excel(form_data)
        print(f"✅ Excel file created: {excel_file_path}")
        
        # Save to static/uploads directory
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"LOE_{timestamp}.xlsx"
        excel_output_path = os.path.join('static', 'uploads', excel_filename)
        
        # Ensure uploads directory exists
        os.makedirs(os.path.dirname(excel_output_path), exist_ok=True)
        
        # Copy the file to uploads directory
        shutil.copy(excel_file_path, excel_output_path)
        
        # Remove temp file
        os.unlink(excel_file_path)
        
        print(f"💾 Excel file saved to: {excel_output_path}")
        print("="*80)
        
        # Return the download URL
        download_url = f"/static/uploads/{excel_filename}"
        return jsonify({
            "success": True,
            "message": "LOE Excel file generated successfully",
            "download_url": download_url,
            "excel_file": "LOE.xlsx"
        }), 200
        
    except Exception as e:
        print(f"\n❌ Error processing LOE: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

