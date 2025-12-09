from flask import Blueprint, request, jsonify, send_file
import json
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import tempfile
import glob

# Create blueprint for VICS Part 5
vics_part5_bp = Blueprint('vics_part5', __name__)

def cleanup_old_vics_files():
    """Clean up old VICS Excel files from uploads directory"""
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            vics_pattern = os.path.join(upload_dir, 'VICS_Part*.xlsx')
            old_files = glob.glob(vics_pattern)
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

def create_vics_part5_excel(vics_data):
    """
    Create Excel file for VICS Part 5 with formatted data
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "VICS Part 5"
    
    # Set column widths
    column_widths = {
        'A': 20,
        'B': 60,
        'C': 20,
        'D': 20,
        'E': 20,
        'F': 60
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Merge A1 and B1
    ws.merge_cells('A1:B1')
    
    # Set header values
    headers = {
        'A1': 'B) Governance & Policy',
        'C1': 'Max Marks',
        'D1': 'Yes/No',
        'E1': 'Marks given by the Auditor',
        'F1': "Auditor's Observation"
    }
    
    # Header styling
    header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')  # Dark Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply headers and styling
    for cell_ref, value in headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Populate data rows
    current_row = 2
    section_number = 13  # Part 5 starts from section 13
    
    # Track rows that need bold formatting
    bold_rows = []
    
    # Track section total rows for grand total calculation
    section_total_rows = []
    
    # Map section keys to their display numbers
    section_mapping = {
        "policies_and_procedures": "13",
        "governance": "14",
        "information_system_audit": "15",
        "chief_information_security_officer": "16"
    }
    
    # Observations for each question
    observations = {
        "13.1": {
            "Implemented": "The bank has a Board-approved Information Security Policy in place.",
            "Not Implemented": "The bank does not have a Board-approved Information Security Policy."
        },
        "13.2": {
            "Implemented": "The IS Policy is documented separately from the IT Policy.",
            "Not Implemented": "IS Policy is merged with or not clearly distinguished from IT Policy."
        },
        "13.3": {
            "Implemented": "The bank has a Board-approved Cyber Security Policy in place.",
            "Not Implemented": "The bank has not obtained Board approval for its Cyber Security Policy."
        },
        "13.4": {
            "Implemented": "Cyber Security Policy is distinct and documented separately from IS Policy.",
            "Not Implemented": "Cyber Security Policy is not distinct or combined with IS Policy."
        },
        "13.5": {
            "Implemented": "SOPs and guidelines are framed and followed to implement IS/IT/Cyber Policies.",
            "Not Implemented": "SOPs or implementation guidelines are not documented or followed."
        },
        "13.6": {
            "Implemented": "IT and IS policies are reviewed annually and updated as required.",
            "Not Implemented": "IT and IS policies are not reviewed periodically."
        },
        "13.7": {
            "Implemented": "Policies are communicated and made accessible to all relevant stakeholders.",
            "Not Implemented": "Policies are not adequately communicated to staff or stakeholders."
        },
        "13.8": {
            "Implemented": "Exceptions are approved by CISO and reported to the Information Security Committee.",
            "Not Implemented": "Policy/process exceptions are not formally approved or reported."
        },
        "14.1": {
            "Implemented": "IT Strategy/IT Sub-Committee of the Board has been established.",
            "Not Implemented": "IT Strategy/IT Sub-Committee of the Board is not formed."
        },
        "14.2": {
            "Implemented": "The committee reviews IT Security performance at regular intervals.",
            "Not Implemented": "Periodic review of IT Security by the committee is not conducted."
        },
        "14.3": {
            "Implemented": "An IT Steering Committee is established to oversee IT operations.",
            "Not Implemented": "IT Steering Committee has not been constituted."
        },
        "14.4": {
            "Implemented": "Regular meetings are held to review IT performance and initiatives.",
            "Not Implemented": "IT Steering Committee meetings are not conducted periodically."
        },
        "14.5": {
            "Implemented": "Information Security Committee is established and functional.",
            "Not Implemented": "Information Security Committee is not set up."
        },
        "14.6": {
            "Implemented": "Committee meetings are conducted at least once every quarter.",
            "Not Implemented": "Committee meetings are not held quarterly as required."
        },
        "15.1": {
            "Implemented": "IS Audit for the due period has been completed as per schedule.",
            "Not Implemented": "IS Audit for the due period is pending or delayed."
        },
        "15.2": {
            "Implemented": "All previous IS audit observations have been addressed and closed.",
            "Not Implemented": "Some IS audit observations remain open or unresolved."
        },
        "15.3": {
            "Implemented": "The Audit Committee regularly reviews IS audit findings and closure progress.",
            "Not Implemented": "Audit Committee does not review IS audit findings periodically."
        },
        "15.4": {
            "Implemented": "The Audit Committee monitors IS assessments, VAPT reports, and ensures timely closure.",
            "Not Implemented": "Monitoring of IS assessments and VAPT closure is not performed by the Audit Committee."
        },
        "15.5": {
            "Implemented": "Compliance with Cyber Security Framework is included in the IS Audit scope.",
            "Not Implemented": "Cyber Security Framework compliance is not covered under IS Audit."
        },
        "16.1": {
            "Implemented": "A qualified CISO has been appointed by the bank.",
            "Not Implemented": "CISO position is vacant or not formally appointed."
        },
        "16.2": {
            "Implemented": "CISO is an active member of the Information Security Committee.",
            "Not Implemented": "CISO is not included in the Information Security Committee."
        },
        "16.3": {
            "Implemented": "CISO belongs to senior management level.",
            "Not Implemented": "CISO designation does not meet senior management criteria."
        },
        "16.4": {
            "Implemented": "CISO's role, responsibilities, and reporting structure are documented.",
            "Not Implemented": "CISO's role and responsibilities are not formally defined or documented."
        }
    }
    
    # Helper function to format date as DD/MM/YYYY
    def format_date(date_str):
        if not date_str:
            return date_str
        try:
            # Parse date from YYYY-MM-DD format
            from datetime import datetime
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            return date_obj.strftime('%d/%m/%Y')
        except:
            return date_str
    
    for section_key, section_data in vics_data["sections"].items():
        section_title = section_data["title"]
        section_num = section_mapping.get(section_key, str(section_number))
        
        # Add section header row
        section_header_row = current_row
        bold_rows.append(section_header_row)  # Track for bold formatting
        
        ws.cell(row=current_row, column=1).value = section_num
        ws.cell(row=current_row, column=2).value = section_title
        ws.cell(row=current_row, column=3).value = ""
        ws.cell(row=current_row, column=4).value = ""
        ws.cell(row=current_row, column=5).value = ""
        ws.cell(row=current_row, column=6).value = ""
        current_row += 1
        
        # Track section totals
        section_total_marks = 0
        section_marks_given = 0
        
        for q_num, q_data in section_data["questions"].items():
            answer = q_data.get("answer", "")
            
            # Main question row
            ws.cell(row=current_row, column=1).value = q_num  # Column A - Question number
            ws.cell(row=current_row, column=2).value = q_data["question"]  # Column B - Question
            ws.cell(row=current_row, column=3).value = q_data["marks"]  # Column C - Max Marks
            
            # Column D - Yes/No
            yes_no = "Yes" if answer == "Implemented" else "No" if answer == "Not Implemented" else ""
            ws.cell(row=current_row, column=4).value = yes_no
            
            # Calculate marks given (full marks if Implemented, 0 if Not Implemented)
            marks_given = q_data["marks"] if answer == "Implemented" else 0
            ws.cell(row=current_row, column=5).value = marks_given  # Column E - Marks given
            
            # Column F - Observation
            observation = observations.get(q_num, {}).get(answer, "")
            ws.cell(row=current_row, column=6).value = observation
            
            # Add to section totals
            if isinstance(q_data["marks"], (int, float)):
                section_total_marks += q_data["marks"]
                section_marks_given += marks_given
            
            current_row += 1
            
            # Sub-questions
            if "sub_questions" in q_data:
                parent_answer = answer
                
                for sub_num, sub_data in q_data["sub_questions"].items():
                    ws.cell(row=current_row, column=1).value = sub_num
                    ws.cell(row=current_row, column=2).value = sub_data['question']
                    ws.cell(row=current_row, column=3).value = ""
                    ws.cell(row=current_row, column=4).value = ""
                    ws.cell(row=current_row, column=5).value = ""
                    
                    # Handle sub-question observations
                    sub_observation = ""
                    sub_answer = sub_data.get('answer', '')
                    
                    if parent_answer == "Not Implemented":
                        # Special cases for Not Implemented
                        if sub_num in ["13.1.1", "13.3.1", "13.6.1", "14.1.1", "14.2.1", "14.3.1", "14.4.1", "14.5.1", "14.6.1", "15.1.1", "15.3.1", "15.4.1", "16.1.1"]:
                            sub_observation = "-"
                        elif sub_num == "13.8.1":
                            sub_observation = f"Total no. of exception: {sub_answer}" if sub_answer else "-"
                        elif sub_num == "15.2.1":
                            sub_observation = f"Total no. of observation (in percentage) still open: {sub_answer}%" if sub_answer else "-"
                        elif sub_num == "16.3.1":
                            sub_observation = f"Designation of CISO: {sub_answer}" if sub_answer else "-"
                        else:
                            sub_observation = "-"
                    elif parent_answer == "Implemented":
                        # Special handling for Implemented
                        if sub_num == "13.1.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of Approval: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "13.3.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of Approval: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "13.6.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last reviewed: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "13.8.1":
                            if sub_answer:
                                sub_observation = f"Total no. of exception: {sub_answer}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "14.1.1":
                            if sub_answer:
                                sub_observation = f"No. of technically competent Board members: {sub_answer}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "14.2.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last reviewed: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "14.3.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of Committee formation: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "14.4.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last review: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "14.5.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of Committee formation: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "14.6.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last review: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "15.1.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last audit: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "15.2.1":
                            sub_observation = "Total no. of observation (in percentage) still open: 0%"
                        elif sub_num == "15.3.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last review: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "15.4.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last review: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "16.1.1":
                            if sub_answer:
                                sub_observation = f"The name of the assigned CISO is {sub_answer}."
                            else:
                                sub_observation = "-"
                        elif sub_num == "16.3.1":
                            if sub_answer:
                                sub_observation = f"Designation of CISO: {sub_answer}."
                            else:
                                sub_observation = "-"
                        else:
                            sub_observation = "-"
                    
                    ws.cell(row=current_row, column=6).value = sub_observation
                    
                    current_row += 1
        
        # Add section total row
        total_row = current_row
        bold_rows.append(total_row)  # Track for bold formatting
        section_total_rows.append(total_row)  # Track for grand total
        
        ws.cell(row=current_row, column=1).value = ""
        ws.cell(row=current_row, column=2).value = f"Total_B.{section_num}"
        ws.cell(row=current_row, column=3).value = section_total_marks
        ws.cell(row=current_row, column=4).value = ""
        ws.cell(row=current_row, column=5).value = section_marks_given
        ws.cell(row=current_row, column=6).value = ""
        
        current_row += 1
        section_number += 1
    
    # Add grand total row
    grand_total_row = current_row
    bold_rows.append(grand_total_row)
    
    # Calculate grand total from section totals
    grand_total_marks_given = sum([ws.cell(row=row, column=5).value for row in section_total_rows])
    
    ws.cell(row=current_row, column=1).value = ""
    ws.cell(row=current_row, column=2).value = "Marks in the category B"
    ws.cell(row=current_row, column=3).value = "30"
    ws.cell(row=current_row, column=4).value = ""
    ws.cell(row=current_row, column=5).value = grand_total_marks_given
    ws.cell(row=current_row, column=6).value = ""
    
    current_row += 1
    
    # Apply borders and alignment to all data cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for row in range(1, current_row):
        for col in range(1, 7):  # Only A-F columns
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            # Header row (row 1) should always be center-aligned
            if row == 1:
                cell.alignment = center_alignment
            # Column B (questions) and F (observations) should be left-aligned for data rows
            elif col == 2 or col == 6:  # Column B and F
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment
            
            # Set font for data rows
            if row > 1:
                # Apply bold formatting to section headers and totals
                if row in bold_rows:
                    cell.font = Font(name='Times New Roman', size=12, bold=True)
                else:
                    cell.font = Font(name='Times New Roman', size=12)
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name

@vics_part5_bp.route('/process_vics_part5', methods=['POST'])
def process_vics_part5():
    """
    Process VICS Part 5 form data and return JSON response
    """
    try:
        print("\n" + "="*80)
        print("🎯 VICS PART 5 - Policies and procedures | Governance | Information System Audit | Chief Information Security Officer (CISO)")
        print("="*80)
        
        # Get form data
        form_data = request.form.to_dict()
        
        # Print all form data for debugging
        print("📋 Form Data Received:")
        for key, value in form_data.items():
            print(f"  {key}: {value}")
        
        # Organize data by sections
        vics_data = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "part": "VICS Part 5: Policies and procedures | Governance | Information System Audit | Chief Information Security Officer (CISO)",
            "sections": {
                "policies_and_procedures": {
                    "title": "Policies and procedures",
                    "questions": {
                        "13.1": {
                            "question": "Whether Bank has Board approved Information Security (IS) Policy?",
                            "marks": 2,
                            "answer": form_data.get('q13_1', ''),
                            "sub_questions": {
                                "13.1.1": {
                                    "question": "Provide date of approval.",
                                    "marks": "-",
                                    "answer": form_data.get('q13_1_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "13.2": {
                            "question": "Whether Board approved IS policy distinct from IT policy is available?",
                            "marks": 1,
                            "answer": form_data.get('q13_2', '')
                        },
                        "13.3": {
                            "question": "Whether Bank has Board approved Cyber Security Policy?",
                            "marks": 2,
                            "answer": form_data.get('q13_3', ''),
                            "sub_questions": {
                                "13.3.1": {
                                    "question": "Provide date of approval.",
                                    "marks": "-",
                                    "answer": form_data.get('q13_3_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "13.4": {
                            "question": "Whether Board approved Cyber Security policy distinct from IS policy is available?",
                            "marks": 1,
                            "answer": form_data.get('q13_4', '')
                        },
                        "13.5": {
                            "question": "Whether Standard Operating Procedures/ guidelines are framed for implementing policies?",
                            "marks": 1,
                            "answer": form_data.get('q13_5', '')
                        },
                        "13.6": {
                            "question": "Whether IT and IS policies are reviewed periodically atleast annually?",
                            "marks": 1,
                            "answer": form_data.get('q13_6', ''),
                            "sub_questions": {
                                "13.6.1": {
                                    "question": "Provide date of last review.",
                                    "marks": "-",
                                    "answer": form_data.get('q13_6_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "13.7": {
                            "question": "Whether policies are communicated to all the stake holders including staff?",
                            "marks": 1,
                            "answer": form_data.get('q13_7', '')
                        },
                        "13.8": {
                            "question": "In case of any exception to policy/ process, whether it is approved by CISO and reported to Information Security Committee?",
                            "marks": 1,
                            "answer": form_data.get('q13_8', ''),
                            "sub_questions": {
                                "13.8.1": {
                                    "question": "Provide total no. of exception.",
                                    "marks": "-",
                                    "answer": form_data.get('q13_8_1', ''),
                                    "type": "number"
                                }
                            }
                        }
                    }
                },
                "governance": {
                    "title": "Governance",
                    "questions": {
                        "14.1": {
                            "question": "Whether IT Strategy Committee or IT Sub-Committee of the Board set up?",
                            "marks": 2,
                            "answer": form_data.get('q14_1', ''),
                            "sub_questions": {
                                "14.1.1": {
                                    "question": "Provide no. of technically competent Board members.",
                                    "marks": "-",
                                    "answer": form_data.get('q14_1_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "14.2": {
                            "question": "Whether IT Strategy Committee or IT Sub-Committee of the Board review performance related to IT Security periodically?",
                            "marks": 1,
                            "answer": form_data.get('q14_2', ''),
                            "sub_questions": {
                                "14.2.1": {
                                    "question": "Provide date of last review.",
                                    "marks": "-",
                                    "answer": form_data.get('q14_2_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "14.3": {
                            "question": "Whether IT Steering Committee set up?",
                            "marks": 1,
                            "answer": form_data.get('q14_3', ''),
                            "sub_questions": {
                                "14.3.1": {
                                    "question": "Provide date of committee formed.",
                                    "marks": "-",
                                    "answer": form_data.get('q14_3_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "14.4": {
                            "question": "Whether IT Steering Committee meets periodically to review IT initiatives and performance?",
                            "marks": 1,
                            "answer": form_data.get('q14_4', ''),
                            "sub_questions": {
                                "14.4.1": {
                                    "question": "Provide date of last review.",
                                    "marks": "-",
                                    "answer": form_data.get('q14_4_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "14.5": {
                            "question": "Whether Information Security Committee set up?",
                            "marks": 2,
                            "answer": form_data.get('q14_5', ''),
                            "sub_questions": {
                                "14.5.1": {
                                    "question": "Provide date of committee formed.",
                                    "marks": "-",
                                    "answer": form_data.get('q14_5_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "14.6": {
                            "question": "Whether the Information Security committee meets periodically at least quarterly?",
                            "marks": 1,
                            "answer": form_data.get('q14_6', ''),
                            "sub_questions": {
                                "14.6.1": {
                                    "question": "Provide date of last review.",
                                    "marks": "-",
                                    "answer": form_data.get('q14_6_1', ''),
                                    "type": "date"
                                }
                            }
                        }
                    }
                },
                "information_system_audit": {
                    "title": "Information System Audit",
                    "questions": {
                        "15.1": {
                            "question": "Whether IS audit is completed for the due period?",
                            "marks": 1,
                            "answer": form_data.get('q15_1', ''),
                            "sub_questions": {
                                "15.1.1": {
                                    "question": "Provide date of last audit.",
                                    "marks": "-",
                                    "answer": form_data.get('q15_1_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "15.2": {
                            "question": "Whether all the observations of last IS audit are closed?",
                            "marks": 2,
                            "answer": form_data.get('q15_2', ''),
                            "sub_questions": {
                                "15.2.1": {
                                    "question": "Provide total no. of observation (in percentage) still open.",
                                    "marks": "-",
                                    "answer": form_data.get('q15_2_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "15.3": {
                            "question": "Whether the Audit Committee of the Board review IS audit findings and its closure status?",
                            "marks": 1.5,
                            "answer": form_data.get('q15_3', ''),
                            "sub_questions": {
                                "15.3.1": {
                                    "question": "Provide date of last review.",
                                    "marks": "-",
                                    "answer": form_data.get('q15_3_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "15.4": {
                            "question": "Whether Audit Committee of Board monitors Information security assessment, VAPT and ensure compliance and closure of issues?",
                            "marks": 1.5,
                            "answer": form_data.get('q15_4', ''),
                            "sub_questions": {
                                "15.4.1": {
                                    "question": "Provide date of last review.",
                                    "marks": "-",
                                    "answer": form_data.get('q15_4_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "15.5": {
                            "question": "Whether compliance review of Cyber Security Framework is part of IS Audit?",
                            "marks": 1,
                            "answer": form_data.get('q15_5', '')
                        }
                    }
                },
                "chief_information_security_officer": {
                    "title": "Chief Information Security Officer (CISO)",
                    "questions": {
                        "16.1": {
                            "question": "Whether the bank has appointed CISO?",
                            "marks": 2,
                            "answer": form_data.get('q16_1', ''),
                            "sub_questions": {
                                "16.1.1": {
                                    "question": "Provide name of the CISO.",
                                    "marks": "-",
                                    "answer": form_data.get('q16_1_1', ''),
                                    "type": "text"
                                }
                            }
                        },
                        "16.2": {
                            "question": "Whether is CISO a member of Information Security Committee?",
                            "marks": 1,
                            "answer": form_data.get('q16_2', '')
                        },
                        "16.3": {
                            "question": "Whether CISO is from senior management? [GM or above for RRBs and StCBs & DGM or above for DCCBs]",
                            "marks": 1,
                            "answer": form_data.get('q16_3', ''),
                            "sub_questions": {
                                "16.3.1": {
                                    "question": "Provide designation of CISO.",
                                    "marks": "-",
                                    "answer": form_data.get('q16_3_1', ''),
                                    "type": "text"
                                }
                            }
                        },
                        "16.4": {
                            "question": "Whether role of CISO is defined and documented?",
                            "marks": 1,
                            "answer": form_data.get('q16_4', '')
                        }
                    }
                }
            }
        }
        
        # Print organized data
        print("\n" + "="*80)
        print("📊 ORGANIZED DATA:")
        print("="*80)
        print(json.dumps(vics_data, indent=2))
        
        # Calculate total marks
        total_marks = 0
        for section_key, section_data in vics_data["sections"].items():
            print(f"\n📌 {section_data['title']}:")
            for q_num, q_data in section_data["questions"].items():
                marks = q_data.get('marks', 0)
                if marks != '-' and marks != 0:
                    total_marks += marks
                answer = q_data.get('answer', 'Not answered')
                print(f"  {q_num}. {q_data['question']}")
                print(f"      Answer: {answer} | Marks: {marks}")
                
                # Print sub-questions if any
                if 'sub_questions' in q_data:
                    for sub_num, sub_data in q_data['sub_questions'].items():
                        sub_answer = sub_data.get('answer', 'Not answered')
                        print(f"    {sub_num}. {sub_data['question']}")
                        print(f"        Answer: {sub_answer} | Type: {sub_data.get('type', 'text')}")
        
        print(f"\n💯 Total Marks for Part 5: {total_marks}")
        print("="*80 + "\n")
        
        # Clean up old VICS files before generating new one
        print("🧹 Cleaning up old VICS files...")
        cleanup_old_vics_files()
        
        # Generate Excel file
        print("📝 Generating Excel file...")
        excel_file_path = create_vics_part5_excel(vics_data)
        print(f"✅ Excel file created: {excel_file_path}")
        
        # Save the generated file to static/uploads directory
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"VICS_Part5_{timestamp}.xlsx"
        static_upload_path = os.path.join('static', 'uploads', excel_filename)
        
        # Ensure the uploads directory exists
        os.makedirs(os.path.dirname(static_upload_path), exist_ok=True)
        
        # Copy the file
        import shutil
        shutil.copy2(excel_file_path, static_upload_path)
        
        # Clean up temp file
        os.unlink(excel_file_path)
        
        # Return JSON response with download URL
        return jsonify({
            "success": True,
            "message": "VICS Part 5 data processed successfully",
            "download_url": f"/static/uploads/{excel_filename}",
            "excel_file": excel_filename
        }), 200
        
    except Exception as e:
        print(f"❌ Error processing VICS Part 5: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"Error processing VICS Part 5: {str(e)}"
        }), 500

