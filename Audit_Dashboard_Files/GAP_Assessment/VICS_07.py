from flask import Blueprint, request, jsonify, send_file
import json
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import tempfile
import glob

# Create blueprint for VICS Part 7
vics_part7_bp = Blueprint('vics_part7', __name__)

def cleanup_old_vics_files():
    """Clean up old VICS Excel files from uploads directory"""
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            vics_pattern = os.path.join(upload_dir, 'VICS_Part*.xlsx')
            old_files = glob.glob(vics_pattern)
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

def create_vics_part7_excel(vics_data):
    """
    Create Excel file for VICS Part 7 with formatted data
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "VICS Part 7"
    
    # Set column widths
    column_widths = {
        'A': 20,
        'B': 60,
        'C': 20,
        'D': 20,
        'E': 20,
        'F': 60
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Merge A1 and B1
    ws.merge_cells('A1:B1')
    
    # Set header values
    headers = {
        'A1': 'D) Cyber Crisis Management',
        'C1': 'Max Marks',
        'D1': 'Yes/No',
        'E1': 'Marks given by the Auditor',
        'F1': "Auditor's Observation"
    }
    
    # Header styling
    header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')  # Dark Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply headers and styling
    for cell_ref, value in headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Populate data rows
    current_row = 2
    section_number = 22  # Part 7 starts from section 22
    
    # Track rows that need bold formatting
    bold_rows = []
    
    # Track section total rows for grand total calculation
    section_total_rows = []
    
    # Map section keys to their display numbers
    section_mapping = {
        "cyber_crisis_incident_management": "22",
        "detection_and_response_to_cyber_incidents": "23",
        "corrective_actions_collection_of_evidence_and_reporting": "24"
    }
    
    # Observations for each question
    observations = {
        "22.1": {
            "Implemented": "The bank has a Board-approved Cyber Crisis Management Plan (CCMP).",
            "Not Implemented": "The bank does not have a Board-approved CCMP."
        },
        "22.2": {
            "Implemented": "The CCMP aligns with the National Cyber Crisis Management Plan and CERT-In framework.",
            "Not Implemented": "The CCMP is not aligned with the National Cyber Crisis Management Plan or CERT-In framework."
        },
        "22.3": {
            "Implemented": "The CCMP is reviewed and updated periodically or at least annually.",
            "Not Implemented": "The CCMP is not reviewed or updated regularly."
        },
        "22.4": {
            "Implemented": "Roles and responsibilities of staff are clearly defined in the CCMP.",
            "Not Implemented": "Staff roles and responsibilities are not defined in the CCMP."
        },
        "22.5": {
            "Implemented": "The CCMP has been communicated to all relevant stakeholders.",
            "Not Implemented": "The CCMP has not been communicated to stakeholders."
        },
        "23.1": {
            "Implemented": "Suspicious events and transactions are reviewed regularly.",
            "Not Implemented": "Suspicious events and transactions are not reviewed regularly."
        },
        "23.2": {
            "Implemented": "Mechanisms and tools are implemented to detect breaches and cyber incidents.",
            "Not Implemented": "No mechanisms or tools are in place for detecting breaches or incidents."
        },
        "23.3": {
            "Implemented": "Daily transaction reconciliation is performed.",
            "Not Implemented": "Daily transaction reconciliation is not performed."
        },
        "23.4": {
            "Implemented": "Daily reviews of chargeback claims for digital channels are conducted.",
            "Not Implemented": "Daily reviews of chargeback claims are not conducted."
        },
        "24.1": {
            "Implemented": "The bank maintains source and RCA of all cyber incidents.",
            "Not Implemented": "The bank does not maintain source or RCA of cyber incidents."
        },
        "24.2": {
            "Implemented": "RCA of cyber incidents is presented to the Board for information.",
            "Not Implemented": "RCA is not placed before the Board."
        },
        "24.3": {
            "Implemented": "Corrective actions identified during RCA are implemented.",
            "Not Implemented": "Corrective actions from RCA are not implemented."
        },
        "24.4": {
            "Implemented": "Core staff are aware of the 6-hour reporting requirement.",
            "Not Implemented": "Core staff are not aware of the reporting requirement."
        }
    }
    
    # Helper function to format date as DD/MM/YYYY
    def format_date(date_str):
        if not date_str:
            return date_str
        try:
            # Parse date from YYYY-MM-DD format
            from datetime import datetime
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            return date_obj.strftime('%d/%m/%Y')
        except:
            return date_str
    
    for section_key, section_data in vics_data["sections"].items():
        section_title = section_data["title"]
        section_num = section_mapping.get(section_key, str(section_number))
        
        # Add section header row
        section_header_row = current_row
        bold_rows.append(section_header_row)  # Track for bold formatting
        
        ws.cell(row=current_row, column=1).value = section_num
        ws.cell(row=current_row, column=2).value = section_title
        ws.cell(row=current_row, column=3).value = ""
        ws.cell(row=current_row, column=4).value = ""
        ws.cell(row=current_row, column=5).value = ""
        ws.cell(row=current_row, column=6).value = ""
        current_row += 1
        
        # Track section totals
        section_total_marks = 0
        section_marks_given = 0
        
        for q_num, q_data in section_data["questions"].items():
            answer = q_data.get("answer", "")
            
            # Main question row
            ws.cell(row=current_row, column=1).value = q_num  # Column A - Question number
            ws.cell(row=current_row, column=2).value = q_data["question"]  # Column B - Question
            ws.cell(row=current_row, column=3).value = q_data["marks"]  # Column C - Max Marks
            
            # Column D - Yes/No
            yes_no = "Yes" if answer == "Implemented" else "No" if answer == "Not Implemented" else ""
            ws.cell(row=current_row, column=4).value = yes_no
            
            # Calculate marks given (full marks if Implemented, 0 if Not Implemented)
            marks_given = q_data["marks"] if answer == "Implemented" else 0
            ws.cell(row=current_row, column=5).value = marks_given  # Column E - Marks given
            
            # Column F - Observation
            observation = observations.get(q_num, {}).get(answer, "")
            ws.cell(row=current_row, column=6).value = observation
            
            # Add to section totals
            if isinstance(q_data["marks"], (int, float)):
                section_total_marks += q_data["marks"]
                section_marks_given += marks_given
            
            current_row += 1
            
            # Sub-questions
            if "sub_questions" in q_data:
                parent_answer = answer
                
                for sub_num, sub_data in q_data["sub_questions"].items():
                    ws.cell(row=current_row, column=1).value = sub_num
                    ws.cell(row=current_row, column=2).value = sub_data['question']
                    ws.cell(row=current_row, column=3).value = ""
                    ws.cell(row=current_row, column=4).value = ""
                    ws.cell(row=current_row, column=5).value = ""
                    
                    # Handle sub-question observations
                    sub_observation = ""
                    sub_answer = sub_data.get('answer', '')
                    
                    if parent_answer == "Not Implemented":
                        # Special cases for Not Implemented
                        sub_observation = "-"
                    elif parent_answer == "Implemented":
                        # Special handling for Implemented
                        if sub_num == "22.1.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of Approval: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "22.3.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last review: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "22.5.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last communication: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "23.3.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last review: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "23.4.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last review: {formatted_date}"
                            else:
                                sub_observation = "-"
                        else:
                            sub_observation = "-"
                    
                    ws.cell(row=current_row, column=6).value = sub_observation
                    
                    current_row += 1
        
        # Add section total row
        total_row = current_row
        bold_rows.append(total_row)  # Track for bold formatting
        section_total_rows.append(total_row)  # Track for grand total
        
        ws.cell(row=current_row, column=1).value = ""
        ws.cell(row=current_row, column=2).value = f"Total_D.{section_num}"
        ws.cell(row=current_row, column=3).value = section_total_marks
        ws.cell(row=current_row, column=4).value = ""
        ws.cell(row=current_row, column=5).value = section_marks_given
        ws.cell(row=current_row, column=6).value = ""
        
        current_row += 1
        section_number += 1
    
    # Add grand total row
    grand_total_row = current_row
    bold_rows.append(grand_total_row)
    
    # Calculate grand total from section totals
    grand_total_marks_given = sum([ws.cell(row=row, column=5).value for row in section_total_rows])
    
    ws.cell(row=current_row, column=1).value = ""
    ws.cell(row=current_row, column=2).value = "Marks in the category D"
    ws.cell(row=current_row, column=3).value = "16"
    ws.cell(row=current_row, column=4).value = ""
    ws.cell(row=current_row, column=5).value = grand_total_marks_given
    ws.cell(row=current_row, column=6).value = ""
    
    current_row += 1
    
    # Apply borders and alignment to all data cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for row in range(1, current_row):
        for col in range(1, 7):  # Only A-F columns
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            # Header row (row 1) should always be center-aligned
            if row == 1:
                cell.alignment = center_alignment
            # Column B (questions) and F (observations) should be left-aligned for data rows
            elif col == 2 or col == 6:  # Column B and F
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment
            
            # Set font for data rows
            if row > 1:
                # Apply bold formatting to section headers and totals
                if row in bold_rows:
                    cell.font = Font(name='Times New Roman', size=12, bold=True)
                else:
                    cell.font = Font(name='Times New Roman', size=12)
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name

@vics_part7_bp.route('/process_vics_part7', methods=['POST'])
def process_vics_part7():
    """
    Process VICS Part 7 form data and return JSON response
    """
    try:
        print("\n" + "="*80)
        print("🎯 VICS PART 7 - Cyber Crisis/ Incident Management | Detection and Response to cyber incidents | Corrective actions, Collection of evidence and Reporting")
        print("="*80)
        
        # Get form data
        form_data = request.form.to_dict()
        
        # Print all form data for debugging
        print("📋 Form Data Received:")
        for key, value in form_data.items():
            print(f"  {key}: {value}")
        
        # Organize data by sections
        vics_data = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "part": "VICS Part 7: Cyber Crisis/ Incident Management | Detection and Response to cyber incidents | Corrective actions, Collection of evidence and Reporting",
            "sections": {
                "cyber_crisis_incident_management": {
                    "title": "Cyber Crisis/ Incident Management",
                    "questions": {
                        "22.1": {
                            "question": "Whether bank has Board approved CCMP?",
                            "marks": 2,
                            "answer": form_data.get('q22_1', ''),
                            "sub_questions": {
                                "22.1.1": {
                                    "question": "Provide date of approval.",
                                    "marks": "-",
                                    "answer": form_data.get('q22_1_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "22.2": {
                            "question": "Whether the CCMP is in line with National Cyber Crisis Management Plan and Cyber Security Assessment Framework, framed by CERT-In, GOI?",
                            "marks": 1,
                            "answer": form_data.get('q22_2', '')
                        },
                        "22.3": {
                            "question": "Whether CCMP updated whenever changes are required based on various internal and external inputs or reviewed atleast annually?",
                            "marks": 1,
                            "answer": form_data.get('q22_3', ''),
                            "sub_questions": {
                                "22.3.1": {
                                    "question": "Provide date of last review.",
                                    "marks": "-",
                                    "answer": form_data.get('q22_3_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "22.4": {
                            "question": "Whether roles and responsibilities of staff is defined in CCMP?",
                            "marks": 1.5,
                            "answer": form_data.get('q22_4', '')
                        },
                        "22.5": {
                            "question": "Whether CCMP has been communicated to the relevant stakeholders?",
                            "marks": 1,
                            "answer": form_data.get('q22_5', ''),
                            "sub_questions": {
                                "22.5.1": {
                                    "question": "Provide date of last such communication.",
                                    "marks": "-",
                                    "answer": form_data.get('q22_5_1', ''),
                                    "type": "date"
                                }
                            }
                        }
                    }
                },
                "detection_and_response_to_cyber_incidents": {
                    "title": "Detection and Response to cyber incidents",
                    "questions": {
                        "23.1": {
                            "question": "Whether suspicious events/ transactions are being reviewed regularly from cyber security perspective?",
                            "marks": 1.5,
                            "answer": form_data.get('q23_1', '')
                        },
                        "23.2": {
                            "question": "Whether any mechanism/controls/tools has been put in place for detection of breaches / incidents?",
                            "marks": 2,
                            "answer": form_data.get('q23_2', '')
                        },
                        "23.3": {
                            "question": "Whether daily reconciliation of transactions between sponsor and client bank is ensured, if applicable?",
                            "marks": 1,
                            "answer": form_data.get('q23_3', ''),
                            "sub_questions": {
                                "23.3.1": {
                                    "question": "Provide date of last review.",
                                    "marks": "-",
                                    "answer": form_data.get('q23_3_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "23.4": {
                            "question": "Whether daily review of chargeback claims for digital payment channels (e.g. ATM, UPI, IMPS) are ensured?",
                            "marks": 1,
                            "answer": form_data.get('q23_4', ''),
                            "sub_questions": {
                                "23.4.1": {
                                    "question": "Provide date of last review.",
                                    "marks": "-",
                                    "answer": form_data.get('q23_4_1', ''),
                                    "type": "date"
                                }
                            }
                        }
                    }
                },
                "corrective_actions_collection_of_evidence_and_reporting": {
                    "title": "Corrective actions, Collection of evidence and Reporting",
                    "questions": {
                        "24.1": {
                            "question": "Whether the bank maintains the source, root cause analysis (RCA) Cyber Incidents?",
                            "marks": 1,
                            "answer": form_data.get('q24_1', '')
                        },
                        "24.2": {
                            "question": "Whether such RCA are placed before the Board for information?",
                            "marks": 1,
                            "answer": form_data.get('q24_2', '')
                        },
                        "24.3": {
                            "question": "Whether actions identified during RCA, are implemented to prevent incidents in future?",
                            "marks": 1,
                            "answer": form_data.get('q24_3', '')
                        },
                        "24.4": {
                            "question": "Whether core staff is aware that cyber incidents have to be reported to NABARD and CERT-In within 06 hours of detection?",
                            "marks": 1,
                            "answer": form_data.get('q24_4', '')
                        }
                    }
                }
            }
        }
        
        # Print organized data
        print("\n" + "="*80)
        print("📊 ORGANIZED DATA:")
        print("="*80)
        print(json.dumps(vics_data, indent=2))
        
        # Calculate total marks
        total_marks = 0
        for section_key, section_data in vics_data["sections"].items():
            print(f"\n📌 {section_data['title']}:")
            for q_num, q_data in section_data["questions"].items():
                marks = q_data.get('marks', 0)
                if marks != '-' and marks != 0:
                    total_marks += marks
                answer = q_data.get('answer', 'Not answered')
                print(f"  {q_num}. {q_data['question']}")
                print(f"      Answer: {answer} | Marks: {marks}")
                
                # Print sub-questions if any
                if 'sub_questions' in q_data:
                    for sub_num, sub_data in q_data['sub_questions'].items():
                        sub_answer = sub_data.get('answer', 'Not answered')
                        print(f"    {sub_num}. {sub_data['question']}")
                        print(f"        Answer: {sub_answer} | Type: {sub_data.get('type', 'text')}")
        
        print(f"\n💯 Total Marks for Part 7: {total_marks}")
        print("="*80 + "\n")
        
        # Clean up old VICS files before generating new one
        print("🧹 Cleaning up old VICS files...")
        cleanup_old_vics_files()
        
        # Generate Excel file
        print("📝 Generating Excel file...")
        excel_file_path = create_vics_part7_excel(vics_data)
        print(f"✅ Excel file created: {excel_file_path}")
        
        # Save the generated file to static/uploads directory
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"VICS_Part7_{timestamp}.xlsx"
        static_upload_path = os.path.join('static', 'uploads', excel_filename)
        
        # Ensure the uploads directory exists
        os.makedirs(os.path.dirname(static_upload_path), exist_ok=True)
        
        # Copy the file
        import shutil
        shutil.copy2(excel_file_path, static_upload_path)
        
        # Clean up temp file
        os.unlink(excel_file_path)
        
        # Return JSON response with download URL
        return jsonify({
            "success": True,
            "message": "VICS Part 7 data processed successfully",
            "download_url": f"/static/uploads/{excel_filename}",
            "excel_file": excel_filename
        }), 200
        
    except Exception as e:
        print(f"❌ Error processing VICS Part 7: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"Error processing VICS Part 7: {str(e)}"
        }), 500

