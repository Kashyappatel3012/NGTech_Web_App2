from flask import Blueprint, request, send_file
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# Create blueprint for Meity Audit Part 2
meity_audit_part2_bp = Blueprint('meity_audit_part2', __name__)

# Questions for Part 2 (26-50)
QUESTIONS_PART2 = [
    "Is USB port usage restricted and controlled across all endpoints?",
    "Is a Data Loss Prevention (DLP) solution implemented to control unauthorized data transfers?",
    "Is the internal network segmented from the DMZ and external zones?",
    "Is the SIEM solution at the data center capable of detecting insider threats, monitoring fraud, and supporting regulatory compliance?",
    "Is Active Directory regularly reviewed for access controls and hygiene?",
    "Are employees educated on phishing attacks, and does IT conduct awareness activities or simulations?",
    "Is there a mechanism to restrict or block unauthorized software installations within the bank's systems?",
    "Are periodic fire drills conducted?",
    "Are access control lists for the bank's networking devices maintained and reviewed regularly?",
    "Are common or shared user IDs avoided for systems and applications, and is their reviewed periodically?",
    "Is RDP disabled on endpoints, allowed only on approved servers, with access restricted to specific IPs and documented server lists?",
    "Are all servers, Systems and network devices hardened according to approved configuration guidelines or documentation?",
    "Is the Disaster Recovery (DR) site situated in a different seismic zone from the primary data center?",
    "Are the bank's IT, Information Security, VAPT, and Cybersecurity policies established and reviewed on an annual basis?",
    "Are IS Audit and VAPT findings addressed and resolved by the bank within the stipulated time period?",
    "Is a CISO appointed with defined responsibilities, and does the CISO quarterly oversee and review the bank's cybersecurity matters?",
    "Does the bank follow and act upon alerts and advisories issued by regulatory authorities?",
    "Is the infrastructure designed to eliminate single points of failure?",
    "Is IDS/IPS placed in the infrastructure network to detect and prevent unauthorized or malicious activities?",
    "Is there an NTP server configured to synchronize system time across all network devices and servers?",
    "Is there any mechanism in place to detect DDoS and port scanning attacks?",
    "Is multi-factor authentication implemented for all critical assets and applications?",
    "Are all passwords set according to the organization's password policy or security standards?",
    "Are login/Session timeouts for all critical assets and applications set according to organizational policy or security standards?",
    "Is system access limited to authorized users, and are normal users prevented from performing administrative tasks?"
]

def create_meity_audit_part2_excel(audit_data):
    """Create Excel file for Meity Audit Part 2"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Meity Audit Part 2"
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 60
    
    # Border formatting
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Header formatting
    header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Create headers
    headers = [
        "Sr. No.",
        "Requirements",
        "Complied Status (Fully Complied/Partially Complied/Not Complied/Not Applicable)",
        "Auditor's Remark"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data formatting
    data_font = Font(name='Times New Roman', size=12)
    
    # Alignment for different columns
    sr_no_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    requirement_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    status_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    remark_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Fill data rows (questions 26-50)
    for idx, question in enumerate(QUESTIONS_PART2, 26):
        row_num = idx - 25 + 1  # Row number in Excel (starts from 2)
        
        # Sr. No. (Column A)
        cell_a = ws.cell(row=row_num, column=1)
        cell_a.value = idx
        cell_a.font = data_font
        cell_a.alignment = sr_no_alignment
        cell_a.border = thin_border
        
        # Requirements (Column B)
        cell_b = ws.cell(row=row_num, column=2)
        cell_b.value = question
        cell_b.font = data_font
        cell_b.alignment = requirement_alignment
        cell_b.border = thin_border
        
        # Complied Status (Column C)
        cell_c = ws.cell(row=row_num, column=3)
        status_value = audit_data.get(f'q{idx}_status', '')
        cell_c.value = status_value
        cell_c.font = data_font
        cell_c.alignment = status_alignment
        cell_c.border = thin_border
        
        # Auditor's Remark (Column D)
        cell_d = ws.cell(row=row_num, column=4)
        remark_value = audit_data.get(f'q{idx}_remark', '')
        cell_d.value = remark_value
        cell_d.font = data_font
        cell_d.alignment = remark_alignment
        cell_d.border = thin_border
    
    # Save to file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Meity_Audit_Part2_{timestamp}.xlsx'
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

@meity_audit_part2_bp.route('/process_meity_audit_part2', methods=['POST'])
def process_meity_audit_part2():
    """Process Meity Audit Part 2 form submission"""
    try:
        print("\n" + "="*80)
        print("🚀 Processing Meity Audit Part 2")
        print("="*80)
        
        # Get form data
        audit_data = {}
        for i in range(26, 51):  # Questions 26-50
            status_key = f'q{i}_status'
            remark_key = f'q{i}_remark'
            
            audit_data[status_key] = request.form.get(status_key, '')
            audit_data[remark_key] = request.form.get(remark_key, '')
            
            print(f"  Q{i}: Status={audit_data[status_key]}, Remark={audit_data[remark_key][:30]}...")
        
        # Create Excel file
        filepath, filename = create_meity_audit_part2_excel(audit_data)
        print(f"\n✅ Excel file created: {filename}")
        print("="*80)
        
        # Return file info as JSON
        from flask import jsonify
        return jsonify({
            'success': True,
            'filename': filename,
            'download_url': f'/static/uploads/{filename}'
        })
    
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        from flask import jsonify
        return jsonify({'success': False, 'error': str(e)}), 500

# Cleanup endpoint
@meity_audit_part2_bp.route('/cleanup_meity_audit_part2', methods=['POST'])
def cleanup_meity_audit_part2():
    """Clean up old Meity Audit Part 2 Excel files"""
    try:
        uploads_dir = os.path.join('static', 'uploads')
        if os.path.exists(uploads_dir):
            for filename in os.listdir(uploads_dir):
                if filename.startswith('Meity_Audit_Part2_') and filename.endswith('.xlsx'):
                    file_path = os.path.join(uploads_dir, filename)
                    try:
                        os.remove(file_path)
                        print(f"Deleted: {filename}")
                    except Exception as e:
                        print(f"Error deleting {filename}: {e}")
        
        from flask import jsonify
        return jsonify({'success': True})
    except Exception as e:
        from flask import jsonify
        return jsonify({'success': False, 'error': str(e)}), 500

