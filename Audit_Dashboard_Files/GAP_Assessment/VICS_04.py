from flask import Blueprint, request, jsonify, send_file
import json
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import tempfile
import glob

# Create blueprint for VICS Part 4
vics_part4_bp = Blueprint('vics_part4', __name__)

def cleanup_old_vics_files():
    """Clean up old VICS Excel files from uploads directory"""
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            vics_pattern = os.path.join(upload_dir, 'VICS_Part*.xlsx')
            old_files = glob.glob(vics_pattern)
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

def create_vics_part4_excel(vics_data):
    """
    Create Excel file for VICS Part 4 with formatted data
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "VICS Part 4"
    
    # Set column widths
    column_widths = {
        'A': 20,
        'B': 60,
        'C': 20,
        'D': 20,
        'E': 20,
        'F': 60
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Merge A1 and B1
    ws.merge_cells('A1:B1')
    
    # Set header values
    headers = {
        'A1': 'A) Info Sec Processes & Controls',
        'C1': 'Max Marks',
        'D1': 'Yes/No',
        'E1': 'Marks given by the Auditor',
        'F1': "Auditor's Observation"
    }
    
    # Header styling
    header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')  # Dark Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply headers and styling
    for cell_ref, value in headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Populate data rows
    current_row = 2
    section_number = 10  # Part 4 starts from section 10
    
    # Track rows that need bold formatting
    bold_rows = []
    
    # Map section keys to their display numbers
    section_mapping = {
        "threat_intelligence_and_vapt_management": "10",
        "secure_mail_and_messaging": "11",
        "awareness_trainings_on_cyber_security": "12"
    }
    
    # Observations for each question
    observations = {
        "10.1": {
            "Implemented": "The bank has a Board-approved VAPT policy.",
            "Not Implemented": "The bank does not have a formally approved VAPT policy by the Board."
        },
        "10.2": {
            "Implemented": "All critical systems and internet-facing assets undergo VAPT at least twice a year.",
            "Not Implemented": "Regular six-monthly VAPT cycles are not being followed for critical systems."
        },
        "10.3": {
            "Implemented": "The bank remediates all critical and high-risk vulnerabilities immediately after identification.",
            "Not Implemented": "Closure of critical and high-risk vulnerabilities is delayed or pending."
        },
        "10.4": {
            "Implemented": "VA/PT observations and status reports are shared periodically with top management.",
            "Not Implemented": "VA/PT reports are not regularly reviewed or presented to top management."
        },
        "10.5": {
            "Implemented": "The bank promptly complies with all NABARD advisories and alerts.",
            "Not Implemented": "The bank has not ensured timely compliance with NABARD advisories."
        },
        "11.1": {
            "Implemented": "The bank uses its own secured domain-based email system for official communication.",
            "Not Implemented": "The bank does not use a domain-specific email system."
        },
        "11.2": {
            "Implemented": "Bank restricts communication from non-bank domain email IDs through email policy controls.",
            "Not Implemented": "Communication through non-bank domain email IDs is not fully restricted."
        },
        "11.3": {
            "Implemented": "Advanced email security measures including DMARC, SPF, and DKIM are implemented and enforced.",
            "Not Implemented": "Email security measures such as DMARC, SPF, or anti-spam filters are not fully enforced."
        },
        "11.4": {
            "Implemented": "DLP strategy is included in the Cyber Security Policy to prevent unauthorized data transfer.",
            "Not Implemented": "No defined DLP strategy is incorporated in the Cyber Security Policy."
        },
        "12.1": {
            "Implemented": "Regular cyber security awareness programmes are conducted for all staff.",
            "Not Implemented": "No formal or regular cyber security awareness programmes are conducted."
        },
        "12.2": {
            "Implemented": "Cyber security guidelines and awareness materials are distributed among all staff.",
            "Not Implemented": "No cyber security guidance document has been distributed to staff."
        },
        "12.3": {
            "Implemented": "CEO/Chairman has attended a cyber security training programme within the stipulated period.",
            "Not Implemented": "CEO/Chairman has not attended any cyber security training within the stipulated period."
        },
        "12.4": {
            "Implemented": "Board members/Administrator have participated in cyber security awareness or training sessions.",
            "Not Implemented": "Board members/Administrator have not attended any cyber security training programmes."
        }
    }
    
    # Helper function to format date as DD/MM/YYYY
    def format_date(date_str):
        if not date_str:
            return date_str
        try:
            # Parse date from YYYY-MM-DD format
            from datetime import datetime
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            return date_obj.strftime('%d/%m/%Y')
        except:
            return date_str
    
    for section_key, section_data in vics_data["sections"].items():
        section_title = section_data["title"]
        section_num = section_mapping.get(section_key, str(section_number))
        
        # Add section header row
        section_header_row = current_row
        bold_rows.append(section_header_row)  # Track for bold formatting
        
        ws.cell(row=current_row, column=1).value = section_num
        ws.cell(row=current_row, column=2).value = section_title
        ws.cell(row=current_row, column=3).value = ""
        ws.cell(row=current_row, column=4).value = ""
        ws.cell(row=current_row, column=5).value = ""
        ws.cell(row=current_row, column=6).value = ""
        current_row += 1
        
        # Track section totals
        section_total_marks = 0
        section_marks_given = 0
        
        for q_num, q_data in section_data["questions"].items():
            answer = q_data.get("answer", "")
            
            # Main question row
            ws.cell(row=current_row, column=1).value = q_num  # Column A - Question number
            ws.cell(row=current_row, column=2).value = q_data["question"]  # Column B - Question
            ws.cell(row=current_row, column=3).value = q_data["marks"]  # Column C - Max Marks
            
            # Column D - Yes/No
            yes_no = "Yes" if answer == "Implemented" else "No" if answer == "Not Implemented" else ""
            ws.cell(row=current_row, column=4).value = yes_no
            
            # Calculate marks given (full marks if Implemented, 0 if Not Implemented)
            marks_given = q_data["marks"] if answer == "Implemented" else 0
            ws.cell(row=current_row, column=5).value = marks_given  # Column E - Marks given
            
            # Column F - Observation
            observation = observations.get(q_num, {}).get(answer, "")
            ws.cell(row=current_row, column=6).value = observation
            
            # Add to section totals
            if isinstance(q_data["marks"], (int, float)):
                section_total_marks += q_data["marks"]
                section_marks_given += marks_given
            
            current_row += 1
            
            # Sub-questions
            if "sub_questions" in q_data:
                parent_answer = answer
                
                for sub_num, sub_data in q_data["sub_questions"].items():
                    ws.cell(row=current_row, column=1).value = sub_num
                    ws.cell(row=current_row, column=2).value = sub_data['question']
                    ws.cell(row=current_row, column=3).value = ""
                    ws.cell(row=current_row, column=4).value = ""
                    ws.cell(row=current_row, column=5).value = ""
                    
                    # Handle sub-question observations based on parent answer and sub-answer
                    sub_observation = ""
                    sub_answer = sub_data.get('answer', '')
                    
                    if parent_answer == "Not Implemented":
                        # Special cases for Not Implemented
                        if sub_num == "10.1.1" or sub_num == "10.2.1" or sub_num == "10.4.1" or sub_num == "11.1.1" or sub_num == "11.3.1" or sub_num == "12.2.1" or sub_num == "12.3.1" or sub_num == "12.4.1":
                            sub_observation = "-"
                        elif sub_num == "10.3.1":
                            sub_observation = f"Average time of closure (in Hrs) of critical/ high points: {sub_answer}" if sub_answer else "-"
                        elif sub_num == "10.5.1":
                            sub_observation = f"The bank has taken action {sub_answer} alerts and advisories sent by NABARD/C-SITE in current financial year." if sub_answer else "-"
                        elif sub_num == "11.2.1":
                            sub_observation = f"No. of non-specific domain email ids active within the bank: {sub_answer}" if sub_answer else "-"
                        elif sub_num == "12.1.1":
                            sub_observation = "No. of program conducted in last half year: 0"
                        else:
                            sub_observation = "-"
                    elif parent_answer == "Implemented":
                        # Special handling for each sub-question when Implemented
                        if sub_num == "10.1.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of Approval: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "10.2.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last VA testing: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "10.3.1":
                            if sub_answer:
                                sub_observation = f"Average time of closure (in Hrs) of critical/ high points: {sub_answer}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "10.4.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last reviewed: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "10.5.1":
                            if sub_answer:
                                sub_observation = f"The bank has taken action {sub_answer} alerts and advisories sent by NABARD/C-SITE in current financial year."
                            else:
                                sub_observation = "-"
                        elif sub_num == "11.1.1":
                            if sub_answer:
                                sub_observation = f"Domain name: {sub_answer}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "11.2.1":
                            sub_observation = "No. of non-specific domain email ids active within the bank: 0"
                        elif sub_num == "11.3.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of Implementation: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "12.1.1":
                            if sub_answer:
                                sub_observation = f"No. of program conducted in last half year: {sub_answer}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "12.2.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last such communication: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "12.3.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last such training: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "12.4.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last such training: {formatted_date}"
                            else:
                                sub_observation = "-"
                        else:
                            sub_observation = "-"
                    
                    ws.cell(row=current_row, column=6).value = sub_observation
                    
                    current_row += 1
        
        # Add section total row
        total_row = current_row
        bold_rows.append(total_row)  # Track for bold formatting
        
        ws.cell(row=current_row, column=1).value = ""
        ws.cell(row=current_row, column=2).value = f"Total_A.{section_num}"
        ws.cell(row=current_row, column=3).value = section_total_marks
        ws.cell(row=current_row, column=4).value = ""
        ws.cell(row=current_row, column=5).value = section_marks_given
        ws.cell(row=current_row, column=6).value = ""
        
        current_row += 1
        section_number += 1
    
    # Apply borders and alignment to all data cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for row in range(1, current_row):
        for col in range(1, 7):  # Only A-F columns
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            # Header row (row 1) should always be center-aligned
            if row == 1:
                cell.alignment = center_alignment
            # Column B (questions) and F (observations) should be left-aligned for data rows
            elif col == 2 or col == 6:  # Column B and F
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment
            
            # Set font for data rows
            if row > 1:
                # Apply bold formatting to section headers and totals
                if row in bold_rows:
                    cell.font = Font(name='Times New Roman', size=12, bold=True)
                else:
                    cell.font = Font(name='Times New Roman', size=12)
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name

@vics_part4_bp.route('/process_vics_part4', methods=['POST'])
def process_vics_part4():
    """
    Process VICS Part 4 form data and return JSON response
    """
    try:
        print("\n" + "="*80)
        print("🎯 VICS PART 4 - Threat Intelligence and VAPT Management | Secure mail and messaging | Awareness / trainings on Cyber Security")
        print("="*80)
        
        # Get form data
        form_data = request.form.to_dict()
        
        # Print all form data for debugging
        print("📋 Form Data Received:")
        for key, value in form_data.items():
            print(f"  {key}: {value}")
        
        # Organize data by sections
        vics_data = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "part": "VICS Part 4: Threat Intelligence and VAPT Management | Secure mail and messaging | Awareness / trainings on Cyber Security",
            "sections": {
                "threat_intelligence_and_vapt_management": {
                    "title": "Threat Intelligence and VAPT Management",
                    "questions": {
                        "10.1": {
                            "question": "Whether Bank has Board approved Vulnerability Assessment and Penetration Testing (VAPT) policy?",
                            "marks": 1,
                            "answer": form_data.get('q10_1', ''),
                            "sub_questions": {
                                "10.1.1": {
                                    "question": "Provide date of approval.",
                                    "marks": "-",
                                    "answer": form_data.get('q10_1_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "10.2": {
                            "question": "Whether critical system in intranet and all the servers/ applications /network devices published over internet (DMZ) are tested every six months?",
                            "marks": 1,
                            "answer": form_data.get('q10_2', ''),
                            "sub_questions": {
                                "10.2.1": {
                                    "question": "Provide date of last VA testing?",
                                    "marks": "-",
                                    "answer": form_data.get('q10_2_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "10.3": {
                            "question": "Whether bank is closing critical/high VA/PT points immediately after the identification?",
                            "marks": 1,
                            "answer": form_data.get('q10_3', ''),
                            "sub_questions": {
                                "10.3.1": {
                                    "question": "Provide average time of closure (in Hrs) of critical/ high points.",
                                    "marks": "-",
                                    "answer": form_data.get('q10_3_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "10.4": {
                            "question": "Whether VA/PT observation and compliance periodically reported to the top management?",
                            "marks": 1,
                            "answer": form_data.get('q10_4', ''),
                            "sub_questions": {
                                "10.4.1": {
                                    "question": "Provide date of last review.",
                                    "marks": "-",
                                    "answer": form_data.get('q10_4_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "10.5": {
                            "question": "Whether bank has ensured compliance to all the advisories and alerts sent by NABARD?",
                            "marks": 1,
                            "answer": form_data.get('q10_5', ''),
                            "sub_questions": {
                                "10.5.1": {
                                    "question": "Provide No. of complied advisories/ alerts in current FY.",
                                    "marks": "-",
                                    "answer": form_data.get('q10_5_1', ''),
                                    "type": "number"
                                }
                            }
                        }
                    }
                },
                "secure_mail_and_messaging": {
                    "title": "Secure mail and messaging",
                    "questions": {
                        "11.1": {
                            "question": "Whether bank specific domain email system in place?",
                            "marks": 1,
                            "answer": form_data.get('q11_1', ''),
                            "sub_questions": {
                                "11.1.1": {
                                    "question": "Provide domain name.",
                                    "marks": "-",
                                    "answer": form_data.get('q11_1_1', ''),
                                    "type": "text"
                                }
                            }
                        },
                        "11.2": {
                            "question": "Whether bank ensures no communication from the non-specific domain email ids?",
                            "marks": 1,
                            "answer": form_data.get('q11_2', ''),
                            "sub_questions": {
                                "11.2.1": {
                                    "question": "Provide No. of non-specific domain email ids active within the bank.",
                                    "marks": "-",
                                    "answer": form_data.get('q11_2_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "11.3": {
                            "question": "Whether anti-phishing, anti-malware, anti-spam and DMARC controls enforced with email solution?",
                            "marks": 1,
                            "answer": form_data.get('q11_3', ''),
                            "sub_questions": {
                                "11.3.1": {
                                    "question": "Provide date of implementation.",
                                    "marks": "-",
                                    "answer": form_data.get('q11_3_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "11.4": {
                            "question": "Whether comprehensive data loss / leakage prevention strategy to safeguard sensitive/confidential business and customer data / information is part of Cyber Security Policy?",
                            "marks": 1,
                            "answer": form_data.get('q11_4', '')
                        }
                    }
                },
                "awareness_trainings_on_cyber_security": {
                    "title": "Awareness / trainings on Cyber Security",
                    "questions": {
                        "12.1": {
                            "question": "Whether awareness programmes conducted for all staff on cyber security?",
                            "marks": 1,
                            "answer": form_data.get('q12_1', ''),
                            "sub_questions": {
                                "12.1.1": {
                                    "question": "Provide no. of programme conducted in last half year.",
                                    "marks": "-",
                                    "answer": form_data.get('q12_1_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "12.2": {
                            "question": "Whether any document prepared and distributed among staff on measures to be taken by them on cyber security?",
                            "marks": 1,
                            "answer": form_data.get('q12_2', ''),
                            "sub_questions": {
                                "12.2.1": {
                                    "question": "Provide date of last such communication.",
                                    "marks": "-",
                                    "answer": form_data.get('q12_2_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "12.3": {
                            "question": "Whether CEO/ Chairman has attended any training programme on Cyber Security?",
                            "marks": 1,
                            "answer": form_data.get('q12_3', ''),
                            "sub_questions": {
                                "12.3.1": {
                                    "question": "Provide date of last such training.",
                                    "marks": "-",
                                    "answer": form_data.get('q12_3_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "12.4": {
                            "question": "Whether Board members/ Administrator of the bank have attended any programme on Cyber Security?",
                            "marks": 1,
                            "answer": form_data.get('q12_4', ''),
                            "sub_questions": {
                                "12.4.1": {
                                    "question": "Provide date of last such training.",
                                    "marks": "-",
                                    "answer": form_data.get('q12_4_1', ''),
                                    "type": "date"
                                }
                            }
                        }
                    }
                }
            }
        }
        
        # Print organized data
        print("\n" + "="*80)
        print("📊 ORGANIZED DATA:")
        print("="*80)
        print(json.dumps(vics_data, indent=2))
        
        # Calculate total marks
        total_marks = 0
        for section_key, section_data in vics_data["sections"].items():
            print(f"\n📌 {section_data['title']}:")
            for q_num, q_data in section_data["questions"].items():
                marks = q_data.get('marks', 0)
                if marks != '-' and marks != 0:
                    total_marks += marks
                answer = q_data.get('answer', 'Not answered')
                print(f"  {q_num}. {q_data['question']}")
                print(f"      Answer: {answer} | Marks: {marks}")
                
                # Print sub-questions if any
                if 'sub_questions' in q_data:
                    for sub_num, sub_data in q_data['sub_questions'].items():
                        sub_answer = sub_data.get('answer', 'Not answered')
                        print(f"    {sub_num}. {sub_data['question']}")
                        print(f"        Answer: {sub_answer} | Type: {sub_data.get('type', 'text')}")
        
        print(f"\n💯 Total Marks for Part 4: {total_marks}")
        print("="*80 + "\n")
        
        # Clean up old VICS files before generating new one
        print("🧹 Cleaning up old VICS files...")
        cleanup_old_vics_files()
        
        # Generate Excel file
        excel_file_path = create_vics_part4_excel(vics_data)
        print(f"\n✅ Excel file generated: {excel_file_path}")
        
        # Save to static/uploads directory for download
        filename = f"VICS_Part_4_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_dir = "static/uploads"
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, filename)
        
        # Copy the temp file to static/uploads
        import shutil
        shutil.copy(excel_file_path, output_path)
        
        # Clean up temp file
        try:
            os.unlink(excel_file_path)
        except:
            pass
        
        return jsonify({
            "success": True,
            "message": "VICS Part 4 data processed successfully",
            "data": vics_data,
            "total_marks": total_marks,
            "excel_file": filename,
            "download_url": f"/static/uploads/{filename}"
        }), 200
        
    except Exception as e:
        print(f"❌ Error processing VICS Part 4: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"Error processing VICS Part 4: {str(e)}"
        }), 500

