from flask import Blueprint, request, send_file
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# Create blueprint for Meity Audit Part 1
meity_audit_part1_bp = Blueprint('meity_audit_part1', __name__)

# Questions for Part 1 (1-25)
QUESTIONS_PART1 = [
    "Has the bank conducted its audit through a CERT-In empanelled agency?",
    "Does the third-party hosting provider furnish compliance reports regularly?",
    "Are externally hosted data and websites audited by a CERT-In empanelled agency?",
    "Is the Board or top management reviewing the corrective actions taken to resolve audit findings?",
    "Are post-change audits conducted after infrastructure updates?",
    "Is the IT asset inventory maintained and updated regularly?",
    "Are annual vulnerability assessments and penetration tests performed?",
    "Are critical vulnerabilities remediated within timelines defined by CERT-In?",
    "Does the bank enforce quarterly access reviews and implement multi-factor authentication (MFA) for privileged accounts on servers?",
    "Is a documented Incident Response Plan available?",
    "Is annual cybersecurity awareness training conducted for all employees?",
    "Are third-party vendors assessed for cybersecurity compliance?",
    "Is the bank's application (e.g. CBS, CTS) ensuring data encryption both at rest and during transmission, in compliance with CERT-In/MeitY guidelines?",
    "Are core banking systems like CBS and internet banking operating on the same system?",
    "Are email security protocols (SPF, DKIM, DMARC) properly configured?",
    "Are mobile banking applications tested against reverse engineering and injection threats?",
    "Is a SOC team deployed for 24/7 monitoring data center operations?",
    "Are audit findings and actions submitted to NABARD in the prescribed format?",
    "Is Wi-Fi secured using WPA3 and segmented from critical infrastructure?",
    "Are Disaster Recovery (DR) drills conducted at least twice a year for Data Center?",
    "Are all Windows OS versions licensed and updated, or are there outdated versions in use?",
    "Is antivirus protection installed and regularly updated on all systems?",
    "Is dual connectivity implemented?",
    "Are Business Continuity Plan (BCP) and Disaster Recovery Plan (DRP) documented and tested?",
    "Are vendor agreements, such as SLA and NDA, in place and reviewed after any changes are made to them?"
]

def create_meity_audit_part1_excel(audit_data):
    """Create Excel file for Meity Audit Part 1"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Meity Audit Part 1"
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 60
    
    # Border formatting
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Header formatting
    header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Create headers
    headers = [
        "Sr. No.",
        "Requirements",
        "Complied Status (Fully Complied/Partially Complied/Not Complied/Not Applicable)",
        "Auditor's Remark"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data formatting
    data_font = Font(name='Times New Roman', size=12)
    
    # Alignment for different columns
    sr_no_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    requirement_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    status_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    remark_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Fill data rows
    for idx, question in enumerate(QUESTIONS_PART1, 1):
        row_num = idx + 1
        
        # Sr. No. (Column A)
        cell_a = ws.cell(row=row_num, column=1)
        cell_a.value = idx
        cell_a.font = data_font
        cell_a.alignment = sr_no_alignment
        cell_a.border = thin_border
        
        # Requirements (Column B)
        cell_b = ws.cell(row=row_num, column=2)
        cell_b.value = question
        cell_b.font = data_font
        cell_b.alignment = requirement_alignment
        cell_b.border = thin_border
        
        # Complied Status (Column C)
        cell_c = ws.cell(row=row_num, column=3)
        status_value = audit_data.get(f'q{idx}_status', '')
        cell_c.value = status_value
        cell_c.font = data_font
        cell_c.alignment = status_alignment
        cell_c.border = thin_border
        
        # Auditor's Remark (Column D)
        cell_d = ws.cell(row=row_num, column=4)
        remark_value = audit_data.get(f'q{idx}_remark', '')
        cell_d.value = remark_value
        cell_d.font = data_font
        cell_d.alignment = remark_alignment
        cell_d.border = thin_border
    
    # Save to file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Meity_Audit_Part1_{timestamp}.xlsx'
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

@meity_audit_part1_bp.route('/process_meity_audit_part1', methods=['POST'])
def process_meity_audit_part1():
    """Process Meity Audit Part 1 form submission"""
    try:
        print("\n" + "="*80)
        print("🚀 Processing Meity Audit Part 1")
        print("="*80)
        
        # Get form data
        audit_data = {}
        for i in range(1, 26):  # Questions 1-25
            status_key = f'q{i}_status'
            remark_key = f'q{i}_remark'
            
            audit_data[status_key] = request.form.get(status_key, '')
            audit_data[remark_key] = request.form.get(remark_key, '')
            
            print(f"  Q{i}: Status={audit_data[status_key]}, Remark={audit_data[remark_key][:30]}...")
        
        # Create Excel file
        filepath, filename = create_meity_audit_part1_excel(audit_data)
        print(f"\n✅ Excel file created: {filename}")
        print("="*80)
        
        # Return file info as JSON
        from flask import jsonify
        return jsonify({
            'success': True,
            'filename': filename,
            'download_url': f'/static/uploads/{filename}'
        })
    
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        from flask import jsonify
        return jsonify({'success': False, 'error': str(e)}), 500

# Cleanup endpoint
@meity_audit_part1_bp.route('/cleanup_meity_audit_part1', methods=['POST'])
def cleanup_meity_audit_part1():
    """Clean up old Meity Audit Part 1 Excel files"""
    try:
        uploads_dir = os.path.join('static', 'uploads')
        if os.path.exists(uploads_dir):
            for filename in os.listdir(uploads_dir):
                if filename.startswith('Meity_Audit_Part1_') and filename.endswith('.xlsx'):
                    file_path = os.path.join(uploads_dir, filename)
                    try:
                        os.remove(file_path)
                        print(f"Deleted: {filename}")
                    except Exception as e:
                        print(f"Error deleting {filename}: {e}")
        
        from flask import jsonify
        return jsonify({'success': True})
    except Exception as e:
        from flask import jsonify
        return jsonify({'success': False, 'error': str(e)}), 500

