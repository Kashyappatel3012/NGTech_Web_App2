from flask import Blueprint, request, jsonify
import os
import tempfile
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import glob
import re

# Create blueprint for Create LOC Worksheet with Bank Input
create_loc_with_bank_input_bp = Blueprint('create_loc_with_bank_input', __name__)

def cleanup_loc_with_bank_input_files():
    """Clean up old LOC with bank input files from uploads directory"""
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            loc_pattern = os.path.join(upload_dir, 'LOC_With_Bank_Input_*.xlsx')
            old_files = glob.glob(loc_pattern)
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

@create_loc_with_bank_input_bp.route('/cleanup_loc_with_bank_input', methods=['POST'])
def cleanup_loc_with_bank_input_endpoint():
    """Endpoint to cleanup LOC with bank input files after download"""
    try:
        cleanup_loc_with_bank_input_files()
        return jsonify({"success": True, "message": "Cleanup completed"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

def shift_images_right(ws, start_col_idx, num_cols=1):
    """
    Shift images to the right by num_cols starting from start_col_idx
    """
    images_to_update = []
    
    for img in ws._images[:]:
        try:
            # Get image anchor position
            if hasattr(img, 'anchor'):
                anchor = img.anchor
                
                # Check if anchor has _from attribute
                if hasattr(anchor, '_from'):
                    img_col = anchor._from.col
                    img_row = anchor._from.row
                    
                    # If image is in or after the column we're inserting before, shift it right
                    if img_col >= start_col_idx:
                        images_to_update.append({
                            'image': img,
                            'old_col': img_col,
                            'new_col': img_col + num_cols,
                            'row': img_row
                        })
                elif isinstance(anchor, str):
                    # If anchor is a string like "E5", parse it
                    match = re.match(r'([A-Z]+)(\d+)', anchor)
                    if match:
                        col_letter = match.group(1)
                        row_num = int(match.group(2))
                        
                        # Convert column letter to index
                        from openpyxl.utils import column_index_from_string
                        img_col_idx = column_index_from_string(col_letter)
                        
                        if img_col_idx >= start_col_idx:
                            new_col_idx = img_col_idx + num_cols
                            new_col_letter = get_column_letter(new_col_idx)
                            new_anchor = f"{new_col_letter}{row_num}"
                            
                            # Update anchor
                            img.anchor = new_anchor
                            print(f"  📌 Moved image from {anchor} to {new_anchor}")
        except Exception as e:
            print(f"  ⚠️ Error processing image: {e}")
    
    # Update images with _from attribute and resize them
    for img_info in images_to_update:
        try:
            img = img_info['image']
            new_col = img_info['new_col']
            row = img_info['row']
            
            # Resize image by 30x (divide dimensions by 30)
            original_width = img.width
            original_height = img.height
            img.width = original_width / 30
            img.height = original_height / 30
            
            # Get the cell for the new position
            new_col_letter = get_column_letter(new_col + 1)  # +1 because openpyxl uses 0-based for _from
            new_row = row + 1  # +1 because openpyxl uses 0-based for _from
            cell = ws.cell(row=new_row, column=new_col + 1)
            
            # Update the anchor to the new cell
            img.anchor = cell.coordinate
            print(f"  📌 Moved and resized image to {cell.coordinate}")
        except Exception as e:
            print(f"  ⚠️ Error updating image position: {e}")

@create_loc_with_bank_input_bp.route('/create_loc_with_bank_input', methods=['POST'])
def create_loc_with_bank_input():
    """
    Process two LOC Excel files and merge them by inserting bank input column
    """
    try:
        print("\n" + "="*80)
        print("🚀 Creating LOC Worksheet with Bank Input")
        print("="*80)
        
        # Check if files are uploaded
        if 'loc_file1' not in request.files:
            return jsonify({"error": "No first LOC file uploaded"}), 400
        
        if 'loc_file2' not in request.files:
            return jsonify({"error": "No second LOC file uploaded"}), 400
        
        loc_file1 = request.files['loc_file1']
        loc_file2 = request.files['loc_file2']
        
        if loc_file1.filename == '' or loc_file2.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        # Create temporary directory
        temp_dir = tempfile.mkdtemp()
        print(f"📁 Created temp directory: {temp_dir}")
        
        try:
            # Save uploaded files
            file1_path = os.path.join(temp_dir, 'loc1.xlsx')
            file2_path = os.path.join(temp_dir, 'loc2.xlsx')
            
            loc_file1.save(file1_path)
            loc_file2.save(file2_path)
            print(f"💾 Saved uploaded files")
            
            # Clean up old files before generating new one
            print("\n🧹 Cleaning up old files...")
            cleanup_loc_with_bank_input_files()
            
            # Load both workbooks
            print("\n📖 Loading workbooks...")
            wb1 = load_workbook(file1_path)
            ws1 = wb1.active
            
            wb2 = load_workbook(file2_path)
            ws2 = wb2.active
            
            # Find "Yes/No given by the Bank" column in second file
            print("\n🔍 Searching for 'Yes/No given by the Bank' column in second file...")
            bank_yesno_col = None
            
            # Print all column headers for debugging
            print("  📋 Column headers in second file:")
            for col_idx in range(1, min(ws2.max_column + 1, 20)):  # Check first 20 columns
                cell_value = ws2.cell(row=1, column=col_idx).value
                print(f"    Column {get_column_letter(col_idx)}: {cell_value}")
            
            # Search for the column with multiple possible variations
            search_terms = [
                "Yes/No given by the Bank",
                "Yes/No given by Bank",
                "Yes/No Bank",
                "Bank Yes/No",
                "Bank Input"
            ]
            
            for col_idx in range(1, ws2.max_column + 1):
                cell_value = ws2.cell(row=1, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip()
                    # Check if any search term is in the cell value (case-insensitive)
                    for term in search_terms:
                        if term.lower() in cell_str.lower():
                            bank_yesno_col = col_idx
                            bank_yesno_col_letter = get_column_letter(col_idx)
                            print(f"✅ Found '{term}' in column {bank_yesno_col_letter} (index {bank_yesno_col})")
                            print(f"   Full header text: {cell_str}")
                            break
                    if bank_yesno_col:
                        break
            
            if not bank_yesno_col:
                # Collect all headers for error message
                all_headers = []
                for col_idx in range(1, ws2.max_column + 1):
                    cell_value = ws2.cell(row=1, column=col_idx).value
                    if cell_value:
                        all_headers.append(f"{get_column_letter(col_idx)}: {cell_value}")
                
                error_msg = f"The second Excel file does not have 'Yes/No given by the Bank' column. Found columns: {', '.join(all_headers) if all_headers else 'None'}"
                print(f"\n⚠️ {error_msg}")
                
                # Return a more user-friendly response instead of 400
                return jsonify({
                    "success": False,
                    "error": error_msg,
                    "message": "Column not found in second file"
                }), 200  # Changed to 200 to avoid browser error handling
            
            # Find the insertion point in first file (between B and C)
            # We want to insert the new column as column C (so it goes between current B and C)
            insert_col_idx = 3  # Column C (1-based: A=1, B=2, C=3)
            
            print(f"\n📝 Inserting new column at position {get_column_letter(insert_col_idx)} (index {insert_col_idx})")
            
            # Insert a new column at position C
            ws1.insert_cols(insert_col_idx)
            print(f"  ✅ Column inserted")
            
            # Shift images that are in column C or after to the right
            print(f"\n🖼️ Shifting images...")
            shift_images_right(ws1, insert_col_idx - 1)  # -1 because shift_images_right uses 0-based indexing
            
            # Copy the "Yes/No given by the Bank" column from file 2 to the new column C in file 1
            print(f"\n📋 Copying 'Yes/No given by the Bank' column data...")
            
            # Get column letters for conversion
            source_col_letter = get_column_letter(bank_yesno_col)
            target_col_letter = get_column_letter(insert_col_idx)
            
            for row_idx in range(1, ws2.max_row + 1):
                source_cell = ws2.cell(row=row_idx, column=bank_yesno_col)
                target_cell = ws1.cell(row=row_idx, column=insert_col_idx)
                
                # Check if source cell contains a formula
                if source_cell.data_type == 'f':  # Formula
                    formula = source_cell.value
                    
                    # Convert column references in the formula
                    # Replace source column letter with target column letter in the formula
                    converted_formula = re.sub(
                        r'\b' + source_col_letter + r'(\d+)',
                        target_col_letter + r'\1',
                        formula,
                        flags=re.IGNORECASE
                    )
                    
                    target_cell.value = converted_formula
                    print(f"  🔢 Row {row_idx}: Converted formula '{formula}' to '{converted_formula}'")
                else:
                    # Copy value as-is, but sanitize to prevent Excel injection
                    from excel_security_utils import sanitize_excel_value
                    target_cell.value = sanitize_excel_value(source_cell.value)
                
                # Copy formatting
                if source_cell.has_style:
                    target_cell.font = source_cell.font.copy()
                    target_cell.border = source_cell.border.copy()
                    target_cell.fill = source_cell.fill.copy()
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = source_cell.protection.copy()
                    target_cell.alignment = source_cell.alignment.copy()
            
            # Copy column width
            if source_col_letter in ws2.column_dimensions:
                ws1.column_dimensions[target_col_letter].width = ws2.column_dimensions[source_col_letter].width
            
            print(f"  ✅ Copied {ws2.max_row} rows")
            
            # Set background color for C1 (header) to dark blue
            print(f"\n🎨 Setting background color for column C header...")
            dark_blue_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
            white_font_bold = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            c1_cell = ws1.cell(row=1, column=3)
            c1_cell.fill = dark_blue_fill
            c1_cell.font = white_font_bold
            c1_cell.alignment = center_alignment
            print(f"  ✅ Set dark blue background for C1")
            
            # Update column widths
            print(f"\n📏 Setting column widths...")
            column_widths = {
                'A': 15,
                'B': 60,
                'C': 20,
                'D': 20,
                'E': 60,
                'F': 20,
                'G': 20,
                'H': 20,
                'I': 20,
                'J': 20,
                'K': 20
            }
            
            for col_letter, width in column_widths.items():
                ws1.column_dimensions[col_letter].width = width
                print(f"  ✅ Set column {col_letter} width to {width}")
            
            # Format all Column C data with Times New Roman, size 12
            print(f"\n📝 Formatting Column C data...")
            times_new_roman_font = Font(name='Times New Roman', size=12)
            
            for row_num in range(1, ws1.max_row + 1):
                c_cell = ws1.cell(row=row_num, column=3)  # Column C
                # Preserve existing font properties (bold, color) but ensure Times New Roman and size 12
                if c_cell.font:
                    c_cell.font = Font(
                        name='Times New Roman',
                        size=12,
                        bold=c_cell.font.bold,
                        italic=c_cell.font.italic,
                        color=c_cell.font.color
                    )
                else:
                    c_cell.font = times_new_roman_font
            
            print(f"  ✅ Formatted {ws1.max_row} rows in Column C")
            
            # Fix POC headers (F1 and G1:L1)
            print(f"\n📝 Fixing POC Attached and POC headers...")
            
            # Unmerge F1:K1 if it exists (after column insertion, the old E1:J1 became F1:K1)
            try:
                ws1.unmerge_cells('F1:K1')
                print(f"  ✅ Unmerged F1:K1")
            except:
                print(f"  ℹ️ F1:K1 was not merged or already processed")
            
            # Add "POC Attached" to F1
            f1_cell = ws1.cell(row=1, column=6)  # Column F
            f1_cell.value = "POC Attached"
            f1_cell.font = white_font_bold
            f1_cell.fill = dark_blue_fill
            f1_cell.alignment = center_alignment
            print(f"  ✅ Added 'POC Attached' to F1")
            
            # Merge G1:L1 and add "POC"
            try:
                ws1.merge_cells('G1:L1')
                print(f"  ✅ Merged G1:L1")
            except:
                print(f"  ℹ️ G1:L1 already merged")
            
            g1_cell = ws1.cell(row=1, column=7)  # Column G
            g1_cell.value = "POC"
            g1_cell.font = white_font_bold
            g1_cell.fill = dark_blue_fill
            g1_cell.alignment = center_alignment
            print(f"  ✅ Added 'POC' to G1:L1 merged cell")
            
            # Save the modified workbook
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"LOC_With_Bank_Input_{timestamp}.xlsx"
            output_path = os.path.join('static', 'uploads', output_filename)
            
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb1.save(output_path)
            wb1.close()
            wb2.close()
            
            print(f"\n✅ Modified workbook saved: {output_filename}")
            print("="*80)
            
            # Return download URL
            download_url = f"/static/uploads/{output_filename}"
            return jsonify({
                "success": True,
                "message": "LOC Worksheet with Bank Input created successfully",
                "download_url": download_url,
                "excel_file": "LOC.xlsx"
            }), 200
            
        finally:
            # Clean up temporary directory
            try:
                shutil.rmtree(temp_dir)
                print("🧹 Cleaned up temporary files")
            except Exception as e:
                print(f"⚠️ Error cleaning up temp directory: {e}")
    
    except Exception as e:
        print(f"\n❌ Error creating LOC worksheet with bank input: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

