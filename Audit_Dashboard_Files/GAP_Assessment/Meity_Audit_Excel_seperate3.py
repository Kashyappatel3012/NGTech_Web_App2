from flask import Blueprint, request, send_file
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# Create blueprint for Meity Audit Part 3
meity_audit_part3_bp = Blueprint('meity_audit_part3', __name__)

# Questions for Part 3 (51-71)
QUESTIONS_PART3 = [
    "Are all network devices and systems regularly updated with the latest patches?",
    "Is vendor due diligence conducted by the bank prior to onboarding, and is there a defined process in place for it?",
    "Are critical network devices (e.g., firewalls, routers) in HA mode for redundancy?",
    "Are critical asset consoles restricted to access only from predefined or authorized (whitelisted) systems?",
    "Are records related to logs, change management, permission for restricted access, and other security controls maintained by the bank?",
    "Are logs of critical events and assets stored and maintained for 180 days?",
    "Is an escalation matrix defined for handling cybersecurity events, issues, and policy-related actions or decisions?",
    "Are important documents, guidelines, and awareness materials shared with staff, stakeholders, and vendors, and are meetings conducted to discuss cybersecurity matters?",
    "Are all bank's email accounts protected with two-factor authentication, not shared among multiple users, and is the bank not using public domains?",
    "Is the bank's higher authority periodic reviews cybersecurity-related issues and controls?",
    "Is there a formal process implemented by the bank to revoke employee access to systems and premises upon exit or role change?",
    "Is the bank's organizational setup equipped to oversee IT security operations and handle related concerns effectively?",
    "Is there a process for securely disposing of decommissioned hardware and storage media?",
    "Is an SSL certificate implemented on the bank's website?",
    "Is there an IT Sub-Committee, IT Steering Committee, IT Strategy Committee, and IT Security Committee in the bank that meets periodically and review cyber risks and cybersecurity controls?",
    "Is backup taken and tested on a regular basis, and is there a method to verify the integrity of the backup when it is restored?",
    "Are fire extinguishers, earthing systems, and UPS units tested or refilled regularly?",
    "Are the availability statuses of important network devices and servers being monitored through a defined mechanism?",
    "Are employees aware of the need to report fraudulent transactions and cybersecurity incidents to regulators within the defined time period, and does the bank have a process in place to support this?",
    "In case the bank utilizes APIs, are appropriate security measures such as authentication, encryption, and rate-limiting implemented?",
    "Is customer data maintained at the data center in compliance with NABARD's cybersecurity guidelines and secured using adequate controls?"
]

def create_meity_audit_part3_excel(audit_data):
    """Create Excel file for Meity Audit Part 3"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Meity Audit Part 3"
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 60
    
    # Border formatting
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Header formatting
    header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Create headers
    headers = [
        "Sr. No.",
        "Requirements",
        "Complied Status (Fully Complied/Partially Complied/Not Complied/Not Applicable)",
        "Auditor's Remark"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data formatting
    data_font = Font(name='Times New Roman', size=12)
    
    # Alignment for different columns
    sr_no_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    requirement_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    status_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    remark_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Fill data rows (questions 51-71)
    for idx, question in enumerate(QUESTIONS_PART3, 51):
        row_num = idx - 50 + 1  # Row number in Excel (starts from 2)
        
        # Sr. No. (Column A)
        cell_a = ws.cell(row=row_num, column=1)
        cell_a.value = idx
        cell_a.font = data_font
        cell_a.alignment = sr_no_alignment
        cell_a.border = thin_border
        
        # Requirements (Column B)
        cell_b = ws.cell(row=row_num, column=2)
        cell_b.value = question
        cell_b.font = data_font
        cell_b.alignment = requirement_alignment
        cell_b.border = thin_border
        
        # Complied Status (Column C)
        cell_c = ws.cell(row=row_num, column=3)
        status_value = audit_data.get(f'q{idx}_status', '')
        cell_c.value = status_value
        cell_c.font = data_font
        cell_c.alignment = status_alignment
        cell_c.border = thin_border
        
        # Auditor's Remark (Column D)
        cell_d = ws.cell(row=row_num, column=4)
        remark_value = audit_data.get(f'q{idx}_remark', '')
        cell_d.value = remark_value
        cell_d.font = data_font
        cell_d.alignment = remark_alignment
        cell_d.border = thin_border
    
    # Save to file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Meity_Audit_Part3_{timestamp}.xlsx'
    filepath = os.path.join('static', 'uploads', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath, filename

@meity_audit_part3_bp.route('/process_meity_audit_part3', methods=['POST'])
def process_meity_audit_part3():
    """Process Meity Audit Part 3 form submission"""
    try:
        print("\n" + "="*80)
        print("🚀 Processing Meity Audit Part 3")
        print("="*80)
        
        # Get form data
        audit_data = {}
        for i in range(51, 72):  # Questions 51-71
            status_key = f'q{i}_status'
            remark_key = f'q{i}_remark'
            
            audit_data[status_key] = request.form.get(status_key, '')
            audit_data[remark_key] = request.form.get(remark_key, '')
            
            print(f"  Q{i}: Status={audit_data[status_key]}, Remark={audit_data[remark_key][:30]}...")
        
        # Create Excel file
        filepath, filename = create_meity_audit_part3_excel(audit_data)
        print(f"\n✅ Excel file created: {filename}")
        print("="*80)
        
        # Return file info as JSON
        from flask import jsonify
        return jsonify({
            'success': True,
            'filename': filename,
            'download_url': f'/static/uploads/{filename}'
        })
    
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        from flask import jsonify
        return jsonify({'success': False, 'error': str(e)}), 500

# Cleanup endpoint
@meity_audit_part3_bp.route('/cleanup_meity_audit_part3', methods=['POST'])
def cleanup_meity_audit_part3():
    """Clean up old Meity Audit Part 3 Excel files"""
    try:
        uploads_dir = os.path.join('static', 'uploads')
        if os.path.exists(uploads_dir):
            for filename in os.listdir(uploads_dir):
                if filename.startswith('Meity_Audit_Part3_') and filename.endswith('.xlsx'):
                    file_path = os.path.join(uploads_dir, filename)
                    try:
                        os.remove(file_path)
                        print(f"Deleted: {filename}")
                    except Exception as e:
                        print(f"Error deleting {filename}: {e}")
        
        from flask import jsonify
        return jsonify({'success': True})
    except Exception as e:
        from flask import jsonify
        return jsonify({'success': False, 'error': str(e)}), 500

