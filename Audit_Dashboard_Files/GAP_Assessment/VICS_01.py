from flask import Blueprint, request, jsonify, send_file
import json
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import tempfile
import glob

# Create blueprint for VICS Part 1
vics_part1_bp = Blueprint('vics_part1', __name__)

def cleanup_old_vics_files():
    """
    Clean up old VICS Excel files from uploads directory
    """
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            # Find all VICS Part Excel files
            vics_pattern = os.path.join(upload_dir, 'VICS_Part*.xlsx')
            old_files = glob.glob(vics_pattern)
            
            # Delete all old VICS files
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

@vics_part1_bp.route('/cleanup_vics_files', methods=['POST'])
def cleanup_vics_files_endpoint():
    """
    Endpoint to cleanup VICS files after download
    """
    try:
        cleanup_old_vics_files()
        return jsonify({"success": True, "message": "Cleanup completed"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

def create_vics_part1_excel(vics_data):
    """
    Create Excel file for VICS Part 1 with formatted data
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "VICS Part 1"
    
    # Set column widths
    column_widths = {
        'A': 20,
        'B': 60,
        'C': 20,
        'D': 20,
        'E': 20,
        'F': 60
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Merge A1 and B1
    ws.merge_cells('A1:B1')
    
    # Set header values
    headers = {
        'A1': 'A) Info Sec Processes & Controls',
        'C1': 'Max Marks',
        'D1': 'Yes/No',
        'E1': 'Marks given by the Auditor',
        'F1': "Auditor's Observation"
    }
    
    # Header styling
    header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')  # Dark Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply headers and styling
    for cell_ref, value in headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Populate data rows
    current_row = 2
    section_number = 1
    
    # Track rows that need bold formatting
    bold_rows = []
    
    # Map section keys to their display numbers
    section_mapping = {
        "it_asset_management": "1",
        "preventing_unauthorised_software": "2",
        "environment_controls": "3"
    }
    
    # Observations for each question
    observations = {
        "1.1": {
            "Implemented": "The bank maintains a comprehensive IT Assets Inventory that includes both hardware and software assets.",
            "Not Implemented": "The bank does not maintain a comprehensive IT Assets Inventory covering both hardware and software."
        },
        "1.2": {
            "Implemented": "IT assets are classified based on their criticality and sensitivity.",
            "Not Implemented": "IT assets are not classified based on criticality or sensitivity."
        },
        "1.3": {
            "Implemented": "The asset inventory includes owner and contact details for each system.",
            "Not Implemented": "Owner contact details are missing from the inventory."
        },
        "1.4": {
            "Implemented": "The IT asset inventory is reviewed and updated quarterly with version control maintained.",
            "Not Implemented": "The inventory is not updated periodically, and no version control mechanism is in place."
        },
        "1.5": {
            "Implemented": "All hardware assets are properly labeled as per the maintained inventory list.",
            "Not Implemented": "Hardware assets are not labeled or do not match the inventory records."
        },
        "2.1": {
            "Implemented": "The bank maintains a authorised software list for endpoints.",
            "Not Implemented": "The bank do not have authorised software list for endpoints."
        },
        "2.2": {
            "Implemented": "Authorized software list is communicated to users, and staff are sensitized against unauthorized installations.",
            "Not Implemented": "Authorized software list is not communicated, and staff are unaware of installation restrictions."
        },
        "2.3": {
            "Implemented": "Controls are in place to restrict or block unauthorized software installations.",
            "Not Implemented": "No mechanism exists to prevent or block installation of unauthorized software."
        },
        "2.4": {
            "Implemented": "The bank has disabled the JavaScript, Active X.",
            "Not Implemented": "The bank has not disabled the JavaScript or Active X."
        },
        "2.5": {
            "Implemented": "Internet access is limited to standalone PCs; technical controls segregate intranet and internet systems.",
            "Not Implemented": "Internet access is available on intranet-connected systems without proper segregation controls."
        },
        "3.1": {
            "Implemented": "Fire alarms are installed and tested at regular intervals to ensure functionality.",
            "Not Implemented": "Fire alarms are either not tested periodically or lack maintenance/testing records."
        },
        "3.2": {
            "Implemented": "Electrical testing including earthing is performed periodically and records are maintained.",
            "Not Implemented": "Periodic electrical or earthing testing is not conducted or lacks proper documentation."
        },
        "3.3": {
            "Implemented": "Fire extinguishers are maintained and refilled as per the prescribed lifecycle.",
            "Not Implemented": "Fire extinguishers are not refilled or maintained within their scheduled lifecycle."
        },
        "3.4": {
            "Implemented": "CCTV cameras are properly installed with sufficient lighting and recordings retained as per requirement.",
            "Not Implemented": "CCTV coverage or recording retention does not meet policy or operational requirements."
        },
        "3.5": {
            "Implemented": "Fire drills are conducted periodically and participation records are maintained.",
            "Not Implemented": "Fire drills are not conducted regularly or lack supporting documentation."
        }
    }
    
    for section_key, section_data in vics_data["sections"].items():
        section_title = section_data["title"]
        section_num = section_mapping.get(section_key, str(section_number))
        
        # Add section header row
        section_header_row = current_row
        bold_rows.append(section_header_row)  # Track for bold formatting
        
        ws.cell(row=current_row, column=1).value = section_num
        ws.cell(row=current_row, column=2).value = section_title
        ws.cell(row=current_row, column=3).value = ""
        ws.cell(row=current_row, column=4).value = ""
        ws.cell(row=current_row, column=5).value = ""
        ws.cell(row=current_row, column=6).value = ""
        current_row += 1
        
        # Track section totals
        section_total_marks = 0
        section_marks_given = 0
        
        for q_num, q_data in section_data["questions"].items():
            answer = q_data.get("answer", "")
            
            # Main question row
            ws.cell(row=current_row, column=1).value = q_num  # Column A - Question number
            ws.cell(row=current_row, column=2).value = q_data["question"]  # Column B - Question
            ws.cell(row=current_row, column=3).value = q_data["marks"]  # Column C - Max Marks
            
            # Column D - Yes/No
            yes_no = "Yes" if answer == "Implemented" else "No" if answer == "Not Implemented" else ""
            ws.cell(row=current_row, column=4).value = yes_no
            
            # Calculate marks given (full marks if Implemented, 0 if Not Implemented)
            marks_given = q_data["marks"] if answer == "Implemented" else 0
            ws.cell(row=current_row, column=5).value = marks_given  # Column E - Marks given
            
            # Column F - Observation
            observation = observations.get(q_num, {}).get(answer, "")
            ws.cell(row=current_row, column=6).value = observation
            
            # Add to section totals
            if isinstance(q_data["marks"], (int, float)):
                section_total_marks += q_data["marks"]
                section_marks_given += marks_given
            
            current_row += 1
            
            # Sub-questions
            if "sub_questions" in q_data:
                parent_answer = answer
                
                for sub_num, sub_data in q_data["sub_questions"].items():
                    ws.cell(row=current_row, column=1).value = sub_num
                    ws.cell(row=current_row, column=2).value = sub_data['question']
                    ws.cell(row=current_row, column=3).value = ""
                    ws.cell(row=current_row, column=4).value = ""
                    ws.cell(row=current_row, column=5).value = ""
                    
                    # Handle sub-question observations based on parent answer and sub-answer
                    sub_observation = ""
                    sub_answer = sub_data.get('answer', '')
                    
                    # Helper function to format date as DD/MM/YYYY
                    def format_date(date_str):
                        if not date_str:
                            return date_str
                        try:
                            # Parse date from YYYY-MM-DD format
                            from datetime import datetime
                            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                            return date_obj.strftime('%d/%m/%Y')
                        except:
                            return date_str
                    
                    if parent_answer == "Not Implemented":
                        sub_observation = "-"
                    elif parent_answer == "Implemented" and sub_answer:
                        # Special handling for each sub-question
                        if sub_num == "1.1.1":
                            sub_observation = f"The Bank has {sub_answer} hardwares."
                        elif sub_num == "1.1.2":
                            sub_observation = f"The Bank has {sub_answer} softwares."
                        elif sub_num == "1.4.1":
                            formatted_date = format_date(sub_answer)
                            sub_observation = f"Data of Last Update: {formatted_date}"
                        elif sub_num == "2.1.1":
                            sub_observation = f"No. of approved software: {sub_answer}"
                        elif sub_num == "2.2.1":
                            formatted_date = format_date(sub_answer)
                            sub_observation = f"Date of last such communication: {formatted_date}"
                        elif sub_num == "2.5.1":
                            sub_observation = f"No of system which has both access: {sub_answer}"
                        elif sub_num == "3.1.1":
                            formatted_date = format_date(sub_answer)
                            sub_observation = f"Date of last testing conducted at HO/DC/DR only: {formatted_date}"
                        elif sub_num == "3.2.1":
                            formatted_date = format_date(sub_answer)
                            sub_observation = f"Date of last testing conducted at HO/DC/DR only: {formatted_date}"
                        elif sub_num == "3.5.1":
                            formatted_date = format_date(sub_answer)
                            sub_observation = f"Date of last testing conducted at HO/DC/DR only: {formatted_date}"
                    elif parent_answer == "Implemented" and not sub_answer:
                        sub_observation = "-"
                    
                    # Special case for 2.5.1 - show for both Implemented and Not Implemented
                    if sub_num == "2.5.1" and parent_answer == "Not Implemented" and sub_answer:
                        sub_observation = f"No of system which has both access: {sub_answer}"
                    
                    ws.cell(row=current_row, column=6).value = sub_observation
                    
                    current_row += 1
        
        # Add section total row
        total_row = current_row
        bold_rows.append(total_row)  # Track for bold formatting
        
        ws.cell(row=current_row, column=1).value = ""
        ws.cell(row=current_row, column=2).value = f"Total_A.{section_num}"
        ws.cell(row=current_row, column=3).value = section_total_marks
        ws.cell(row=current_row, column=4).value = ""
        ws.cell(row=current_row, column=5).value = section_marks_given
        ws.cell(row=current_row, column=6).value = ""
        
        current_row += 1
        section_number += 1
    
    # Apply borders and alignment to all data cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for row in range(1, current_row):
        for col in range(1, 7):  # Only A-F columns
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            # Header row (row 1) should always be center-aligned
            if row == 1:
                cell.alignment = center_alignment
            # Column B (questions) and F (observations) should be left-aligned for data rows
            elif col == 2 or col == 6:  # Column B and F
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment
            
            # Set font for data rows
            if row > 1:
                # Apply bold formatting to section headers and totals
                if row in bold_rows:
                    cell.font = Font(name='Times New Roman', size=12, bold=True)
                else:
                    cell.font = Font(name='Times New Roman', size=12)
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name

@vics_part1_bp.route('/process_vics_part1', methods=['POST'])
def process_vics_part1():
    """
    Process VICS Part 1 form data and return JSON response
    """
    try:
        print("\n" + "="*80)
        print("🎯 VICS PART 1 - IT Asset Management | Preventing unauthorised software | Environment controls")
        print("="*80)
        
        # Get form data
        form_data = request.form.to_dict()
        
        # Print all form data for debugging
        print("📋 Form Data Received:")
        for key, value in form_data.items():
            print(f"  {key}: {value}")
        
        # Organize data by sections
        vics_data = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "part": "VICS Part 1: IT Asset Management | Preventing unauthorised software | Environment controls",
            "sections": {
                "it_asset_management": {
                    "title": "IT Asset Management",
                    "questions": {
                        "1.1": {
                            "question": "Whether all IT Assets (both hardware and software) have been inventoried?",
                            "marks": 1,
                            "answer": form_data.get('q1_1', ''),
                            "sub_questions": {
                                "1.1.1": {
                                    "question": "No. of hardware within bank.",
                                    "marks": "-",
                                    "answer": form_data.get('q1_1_1', ''),
                                    "type": "number"
                                },
                                "1.1.2": {
                                    "question": "No. of software within bank.",
                                    "marks": "-",
                                    "answer": form_data.get('q1_1_2', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "1.2": {
                            "question": "Whether IT assets have been classified based on their criticality (e.g. High, Medium, Low) & sensitivity based on data availability (e.g. Confidential, Private, Sensitive, Public)?",
                            "marks": 1,
                            "answer": form_data.get('q1_2', '')
                        },
                        "1.3": {
                            "question": "Whether the inventory contains asset/system/application owner contact details?",
                            "marks": 1,
                            "answer": form_data.get('q1_3', '')
                        },
                        "1.4": {
                            "question": "Whether IT asset inventory is updated periodically, at least quarterly and version controlled is maintained?",
                            "marks": 1,
                            "answer": form_data.get('q1_4', ''),
                            "sub_questions": {
                                "1.4.1": {
                                    "question": "Provide last update date.",
                                    "marks": "-",
                                    "answer": form_data.get('q1_4_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "1.5": {
                            "question": "Whether all hardware assets are labelled as per inventory?",
                            "marks": 1,
                            "answer": form_data.get('q1_5', '')
                        }
                    }
                },
                "preventing_unauthorised_software": {
                    "title": "Preventing unauthorised software",
                    "questions": {
                        "2.1": {
                            "question": "Whether the bank has centralised authorised software inventory/ register for end points?",
                            "marks": 1,
                            "answer": form_data.get('q2_1', ''),
                            "sub_questions": {
                                "2.1.1": {
                                    "question": "Provide No. of approved software.",
                                    "marks": "-",
                                    "answer": form_data.get('q2_1_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "2.2": {
                            "question": "Whether list of authorised software is communicated to the staff/ users and sensitized not to install any other software?",
                            "marks": 1,
                            "answer": form_data.get('q2_2', ''),
                            "sub_questions": {
                                "2.2.1": {
                                    "question": "Provide date of last such communication.",
                                    "marks": "-",
                                    "answer": form_data.get('q2_2_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "2.3": {
                            "question": "Whether the bank has mechanism/ control to block installation of unauthorised software?",
                            "marks": 1.5,
                            "answer": form_data.get('q2_3', '')
                        },
                        "2.4": {
                            "question": "Whether bank has mechanism to push system/ user policies (preferably centrally) or by configuring local policies including disabling of JavaScript, ActiveX, restricting macro enabled files if not used?",
                            "marks": 1.5,
                            "answer": form_data.get('q2_4', '')
                        },
                        "2.5": {
                            "question": "Whether the bank has ensured that internet usage is limited to standalone PCs or has put in place technical control if internet is provided to intranet machines (internet connected system must not be used in CBS & related operation)?",
                            "marks": 2,
                            "answer": form_data.get('q2_5', ''),
                            "sub_questions": {
                                "2.5.1": {
                                    "question": "Provide No. of intranet connected systems having internet access (with or without technical controls).",
                                    "marks": "-",
                                    "answer": form_data.get('q2_5_1', ''),
                                    "type": "number"
                                }
                            }
                        }
                    }
                },
                "environment_controls": {
                    "title": "Environment controls (Includes all premises e.g. HO, DC , DR, Branches etc.)",
                    "questions": {
                        "3.1": {
                            "question": "Whether Fire alarms installed and tested periodically?",
                            "marks": 1,
                            "answer": form_data.get('q3_1', ''),
                            "sub_questions": {
                                "3.1.1": {
                                    "question": "Provide date of last testing conducted at HO/DC/DR only.",
                                    "marks": "-",
                                    "answer": form_data.get('q3_1_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "3.2": {
                            "question": "Whether electric testing including earthing is being conducted periodically?",
                            "marks": 1,
                            "answer": form_data.get('q3_2', ''),
                            "sub_questions": {
                                "3.2.1": {
                                    "question": "Provide date of last testing conducted at HO/DC/DR only",
                                    "marks": "-",
                                    "answer": form_data.get('q3_2_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "3.3": {
                            "question": "Whether Fire extinguisher(s) is/are maintained and refilled according to its life cycle?",
                            "marks": 1,
                            "answer": form_data.get('q3_3', '')
                        },
                        "3.4": {
                            "question": "Whether CCTV cameras are installed at appropriate locations with proper lighting conditions and its recording are being kept for the identified period?",
                            "marks": 1,
                            "answer": form_data.get('q3_4', '')
                        },
                        "3.5": {
                            "question": "Whether periodic fire drill is conducted?",
                            "marks": 1,
                            "answer": form_data.get('q3_5', ''),
                            "sub_questions": {
                                "3.5.1": {
                                    "question": "Provide date of last testing conducted at HO/DC/DR only.",
                                    "marks": "-",
                                    "answer": form_data.get('q3_5_1', ''),
                                    "type": "date"
                                }
                            }
                        }
                    }
                }
            }
        }
        
        # Calculate total marks
        total_marks = 0
        implemented_count = 0
        not_implemented_count = 0
        
        for section_name, section_data in vics_data["sections"].items():
            print(f"\n📊 {section_data['title']}:")
            for q_num, q_data in section_data["questions"].items():
                if q_data["answer"]:
                    if q_data["answer"] == "Implemented":
                        implemented_count += 1
                        if isinstance(q_data["marks"], (int, float)):
                            total_marks += q_data["marks"]
                    elif q_data["answer"] == "Not Implemented":
                        not_implemented_count += 1
                    
                    print(f"  {q_num}: {q_data['answer']} (Marks: {q_data['marks']})")
                    
                    # Check sub-questions
                    if "sub_questions" in q_data:
                        for sub_q_num, sub_q_data in q_data["sub_questions"].items():
                            if sub_q_data["answer"]:
                                print(f"    {sub_q_num}: {sub_q_data['answer']} ({sub_q_data['type']})")
        
        # Add summary to data
        vics_data["summary"] = {
            "total_marks_obtained": total_marks,
            "total_questions_answered": implemented_count + not_implemented_count,
            "implemented_count": implemented_count,
            "not_implemented_count": not_implemented_count,
            "compliance_percentage": round((implemented_count / (implemented_count + not_implemented_count)) * 100, 2) if (implemented_count + not_implemented_count) > 0 else 0
        }
        
        print(f"\n📈 SUMMARY:")
        print(f"  Total Marks Obtained: {total_marks}")
        print(f"  Questions Answered: {implemented_count + not_implemented_count}")
        print(f"  Implemented: {implemented_count}")
        print(f"  Not Implemented: {not_implemented_count}")
        print(f"  Compliance Percentage: {vics_data['summary']['compliance_percentage']}%")
        print("="*80)
        
        # Clean up old VICS files before generating new one
        print("\n🧹 Cleaning up old VICS files...")
        cleanup_old_vics_files()
        
        # Generate Excel file
        excel_file_path = create_vics_part1_excel(vics_data)
        print(f"\n✅ Excel file generated: {excel_file_path}")
        
        # Save to static/uploads directory for download
        filename = f"VICS_Part_1_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_dir = "static/uploads"
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, filename)
        
        # Copy the temp file to static/uploads
        import shutil
        shutil.copy(excel_file_path, output_path)
        
        # Clean up temp file
        try:
            os.unlink(excel_file_path)
        except:
            pass
        
        return jsonify({
            "success": True,
            "message": "VICS Part 1 data processed successfully",
            "data": vics_data,
            "excel_file": filename,
            "download_url": f"/static/uploads/{filename}"
        })
        
    except Exception as e:
        print(f"❌ Error processing VICS Part 1: {e}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            "success": False,
            "message": f"Error processing VICS Part 1: {str(e)}",
            "data": None
        }), 500
