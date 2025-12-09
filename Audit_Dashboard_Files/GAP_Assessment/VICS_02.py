from flask import Blueprint, request, jsonify, send_file
import json
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import tempfile
import glob

# Create blueprint for VICS Part 2
vics_part2_bp = Blueprint('vics_part2', __name__)

def cleanup_old_vics_files():
    """
    Clean up old VICS Excel files from uploads directory
    """
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            # Find all VICS Part Excel files
            vics_pattern = os.path.join(upload_dir, 'VICS_Part*.xlsx')
            old_files = glob.glob(vics_pattern)
            
            # Delete all old VICS files
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

def create_vics_part2_excel(vics_data):
    """
    Create Excel file for VICS Part 2 with formatted data
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "VICS Part 2"
    
    # Set column widths
    column_widths = {
        'A': 20,
        'B': 60,
        'C': 20,
        'D': 20,
        'E': 20,
        'F': 60
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Merge A1 and B1
    ws.merge_cells('A1:B1')
    
    # Set header values
    headers = {
        'A1': 'A) Info Sec Processes & Controls',
        'C1': 'Max Marks',
        'D1': 'Yes/No',
        'E1': 'Marks given by the Auditor',
        'F1': "Auditor's Observation"
    }
    
    # Header styling
    header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')  # Dark Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply headers and styling
    for cell_ref, value in headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Populate data rows
    current_row = 2
    section_number = 4  # Part 2 starts from section 4
    
    # Track rows that need bold formatting
    bold_rows = []
    
    # Map section keys to their display numbers
    section_mapping = {
        "network_management": "4",
        "antivirus_and_patch_management": "5",
        "access_control": "6"
    }
    
    # Observations for each question
    observations = {
        "4.1": {
            "Implemented": "Network devices are updated promptly after the release of stable patches by OEM.",
            "Not Implemented": "Network devices are not updated regularly; patches are delayed or missing for several devices."
        },
        "4.2": {
            "Implemented": "Firewall and network device ACLs are configured only after approval from the competent authority.",
            "Not Implemented": "Access Control Lists are modified without formal approval or change management documentation."
        },
        "4.3": {
            "Implemented": "ACLs of all critical networking devices are reviewed periodically.",
            "Not Implemented": "Periodic review of ACLs is not conducted or records of such reviews are unavailable."
        },
        "4.4": {
            "Implemented": "The bank do not enable wireless (Wi-Fi) facility for the systems in bank/branches and if need to enable then take prior approval from competent authority",
            "Not Implemented": "Wireless networks are active without prior approval or adequate documentation."
        },
        "4.5": {
            "Implemented": "The bank do not enable wireless (Wi-Fi) facility for the systems usually if in special case they enable then Security controls such as encryption or authentication are properly configured.",
            "Not Implemented": "Security controls such as encryption or authentication are not properly configured on wireless networks."
        },
        "4.6": {
            "Implemented": "The bank has implemented VLANs to segregate servers and endpoints, and internet access is restricted for all servers and endpoints",
            "Not Implemented": "The bank does not implemented VLANs to segregate servers and endpoints."
        },
        "4.7": {
            "Implemented": "The bank maintains an updated IT/Network architecture diagram approved by management.",
            "Not Implemented": "No updated or management-approved IT/Network architecture document is available."
        },
        "4.8": {
            "Implemented": "All external network connections are routed through controlled and secured channels.",
            "Not Implemented": "External connections are not properly routed through secured gateways."
        },
        "5.1": {
            "Implemented": "Antivirus software is installed on all servers, PCs, and endpoints.",
            "Not Implemented": "Some systems lack antivirus installation, exposing them to potential threats."
        },
        "5.2": {
            "Implemented": "Antivirus are updated regularly across all systems.",
            "Not Implemented": "Antivirus updates are inconsistent, and some systems have outdated definitions."
        },
        "5.3": {
            "Implemented": "Security patches are applied regularly on all systems.",
            "Not Implemented": "Patch updates are irregular, and several systems have pending patches."
        },
        "5.4": {
            "Implemented": "A patch update register/log is maintained and updated after each patch cycle.",
            "Not Implemented": "No formal patch update register or recordkeeping process is in place."
        },
        "5.5": {
            "Implemented": "All systems run on supported OS versions; deprecated versions are decommissioned.",
            "Not Implemented": "Out-of-support OS versions are still in use within the bank's network."
        },
        "6.1": {
            "Implemented": "User access is granted following the least privilege principle.",
            "Not Implemented": "Users have excess privileges beyond their job role requirements."
        },
        "6.2": {
            "Implemented": "Shared or generic IDs are prohibited and monitored.",
            "Not Implemented": "Shared or generic user IDs are still in use across some systems."
        },
        "6.3": {
            "Implemented": "Shared IDs is not in used commonly, if any special case need to used then use only after formal approval.",
            "Not Implemented": "Shared user IDs are used without management approval."
        },
        "6.4": {
            "Implemented": "User access reviews are conducted periodically and documented.",
            "Not Implemented": "No periodic user access review is carried out for critical systems."
        },
        "6.5": {
            "Implemented": "Administrative privileges are limited to authorized IT personnel only.",
            "Not Implemented": "Some end users have administrative rights without justification."
        },
        "6.6": {
            "Implemented": "Centralized privilege access management system/process is in place and monitored.",
            "Not Implemented": "No centralized mechanism exists to manage privileged or super user access."
        },
        "6.7": {
            "Implemented": "RDP is disabled on endpoints and restricted to authorized servers only.",
            "Not Implemented": "RDP remains enabled on multiple endpoints without proper restriction."
        },
        "6.8": {
            "Implemented": "Internet access to intranet resources is allowed only through VPN with CISO approval.",
            "Not Implemented": "Intranet resources are accessible over the internet without VPN or formal approval."
        },
        "6.9": {
            "Implemented": "Two-factor authentication with dynamic tokens is implemented for CBS and other critical applications.",
            "Not Implemented": "2FA is not fully implemented for CBS or other critical applications."
        },
        "6.10": {
            "Implemented": "Security reviews of critical terminals are conducted by a qualified IS auditor.",
            "Not Implemented": "Security reviews of terminals have not been conducted or documented by a qualified auditor."
        }
    }
    
    # Helper function to format date as DD/MM/YYYY
    def format_date(date_str):
        if not date_str:
            return date_str
        try:
            # Parse date from YYYY-MM-DD format
            from datetime import datetime
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            return date_obj.strftime('%d/%m/%Y')
        except:
            return date_str
    
    for section_key, section_data in vics_data["sections"].items():
        section_title = section_data["title"]
        section_num = section_mapping.get(section_key, str(section_number))
        
        # Add section header row
        section_header_row = current_row
        bold_rows.append(section_header_row)  # Track for bold formatting
        
        ws.cell(row=current_row, column=1).value = section_num
        ws.cell(row=current_row, column=2).value = section_title
        ws.cell(row=current_row, column=3).value = ""
        ws.cell(row=current_row, column=4).value = ""
        ws.cell(row=current_row, column=5).value = ""
        ws.cell(row=current_row, column=6).value = ""
        current_row += 1
        
        # Track section totals
        section_total_marks = 0
        section_marks_given = 0
        
        for q_num, q_data in section_data["questions"].items():
            answer = q_data.get("answer", "")
            
            # Main question row
            ws.cell(row=current_row, column=1).value = q_num  # Column A - Question number
            ws.cell(row=current_row, column=2).value = q_data["question"]  # Column B - Question
            ws.cell(row=current_row, column=3).value = q_data["marks"]  # Column C - Max Marks
            
            # Column D - Yes/No
            yes_no = "Yes" if answer == "Implemented" else "No" if answer == "Not Implemented" else ""
            ws.cell(row=current_row, column=4).value = yes_no
            
            # Calculate marks given (full marks if Implemented, 0 if Not Implemented)
            marks_given = q_data["marks"] if answer == "Implemented" else 0
            ws.cell(row=current_row, column=5).value = marks_given  # Column E - Marks given
            
            # Column F - Observation
            observation = observations.get(q_num, {}).get(answer, "")
            ws.cell(row=current_row, column=6).value = observation
            
            # Add to section totals
            if isinstance(q_data["marks"], (int, float)):
                section_total_marks += q_data["marks"]
                section_marks_given += marks_given
            
            current_row += 1
            
            # Sub-questions
            if "sub_questions" in q_data:
                parent_answer = answer
                
                for sub_num, sub_data in q_data["sub_questions"].items():
                    ws.cell(row=current_row, column=1).value = sub_num
                    ws.cell(row=current_row, column=2).value = sub_data['question']
                    ws.cell(row=current_row, column=3).value = ""
                    ws.cell(row=current_row, column=4).value = ""
                    ws.cell(row=current_row, column=5).value = ""
                    
                    # Handle sub-question observations based on parent answer and sub-answer
                    sub_observation = ""
                    sub_answer = sub_data.get('answer', '')
                    
                    if parent_answer == "Not Implemented":
                        # Special cases for Not Implemented
                        if sub_num == "4.1.1" or sub_num == "4.3.1" or sub_num == "4.7.1" or sub_num == "6.4.1" or sub_num == "6.10.1":
                            sub_observation = "-"
                        elif sub_num == "5.1.1":
                            sub_observation = f"Total systems (in percentage) not having antivirus: {sub_answer}%" if sub_answer else "-"
                        elif sub_num == "5.2.1":
                            sub_observation = f"Total systems (in percentage) not updated within last 7 days: {sub_answer}%" if sub_answer else "-"
                        elif sub_num == "5.3.1":
                            sub_observation = f"Total systems (in percentage) not updated within last 30 days: {sub_answer}%" if sub_answer else "-"
                        elif sub_num == "5.5.1":
                            sub_observation = f"No. systems having out of support OS: {sub_answer}" if sub_answer else "-"
                        elif sub_num == "6.2.1":
                            sub_observation = f"No. of shared/ generic user ids present in the critical servers/applications: {sub_answer}" if sub_answer else "-"
                        elif sub_num == "6.7.1":
                            sub_observation = f"No. of end points (in percentage) where RDP is enabled: {sub_answer}%" if sub_answer else "-"
                        elif sub_num == "6.8.1":
                            sub_observation = f"No. of intranet systems where access is enabled over the internet without VPN: {sub_answer}" if sub_answer else "-"
                        elif sub_num == "6.9.1":
                            sub_observation = f"No. of users (in percentage) who are not enabled for 2FA for CBS: {sub_answer}%" if sub_answer else "-"
                        else:
                            sub_observation = "-"
                    elif parent_answer == "Implemented":
                        # Special handling for each sub-question when Implemented
                        if sub_num == "4.1.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of Last update: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "4.3.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last review: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "4.7.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of approval: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "5.1.1":
                            sub_observation = "Total systems (in percentage) not having antivirus: 0%"
                        elif sub_num == "5.2.1":
                            sub_observation = "Total systems (in percentage) not updated within last 7 days: 0%"
                        elif sub_num == "5.3.1":
                            sub_observation = "Total systems (in percentage) not updated within last 30 days: 0%"
                        elif sub_num == "5.5.1":
                            sub_observation = "No. systems having out of support OS: 0"
                        elif sub_num == "6.2.1":
                            sub_observation = "No. of shared/ generic user ids present in the critical servers/applications: 0"
                        elif sub_num == "6.4.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last review: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "6.7.1":
                            if sub_answer:
                                sub_observation = f"No. of end points (in percentage) where RDP is enabled: {sub_answer}%"
                            else:
                                sub_observation = "-"
                        elif sub_num == "6.8.1":
                            sub_observation = "No. of intranet systems where access is enabled over the internet without VPN: 0"
                        elif sub_num == "6.9.1":
                            sub_observation = "No. of users (in percentage) who are not enabled for 2FA for CBS: 0%"
                        elif sub_num == "6.10.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last review: {formatted_date}"
                            else:
                                sub_observation = "-"
                        else:
                            sub_observation = "-"
                    
                    ws.cell(row=current_row, column=6).value = sub_observation
                    
                    current_row += 1
        
        # Add section total row
        total_row = current_row
        bold_rows.append(total_row)  # Track for bold formatting
        
        ws.cell(row=current_row, column=1).value = ""
        ws.cell(row=current_row, column=2).value = f"Total_A.{section_num}"
        ws.cell(row=current_row, column=3).value = section_total_marks
        ws.cell(row=current_row, column=4).value = ""
        ws.cell(row=current_row, column=5).value = section_marks_given
        ws.cell(row=current_row, column=6).value = ""
        
        current_row += 1
        section_number += 1
    
    # Apply borders and alignment to all data cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for row in range(1, current_row):
        for col in range(1, 7):  # Only A-F columns
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            # Header row (row 1) should always be center-aligned
            if row == 1:
                cell.alignment = center_alignment
            # Column B (questions) and F (observations) should be left-aligned for data rows
            elif col == 2 or col == 6:  # Column B and F
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment
            
            # Set font for data rows
            if row > 1:
                # Apply bold formatting to section headers and totals
                if row in bold_rows:
                    cell.font = Font(name='Times New Roman', size=12, bold=True)
                else:
                    cell.font = Font(name='Times New Roman', size=12)
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name

@vics_part2_bp.route('/process_vics_part2', methods=['POST'])
def process_vics_part2():
    """
    Process VICS Part 2 form data and return JSON response
    """
    try:
        print("\n" + "="*80)
        print("🎯 VICS PART 2 - Network management | Antivirus and patch management | Access Control")
        print("="*80)
        
        # Get form data
        form_data = request.form.to_dict()
        
        # Print all form data for debugging
        print("📋 Form Data Received:")
        for key, value in form_data.items():
            print(f"  {key}: {value}")
        
        # Organize data by sections
        vics_data = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "part": "VICS Part 2: Network management | Antivirus and patch management | Access Control",
            "sections": {
                "network_management": {
                    "title": "Network management",
                    "questions": {
                        "4.1": {
                            "question": "Whether network devices updated regularly as soon as patch (stable) is released by Original equipment manufacturer (OEM)?",
                            "marks": 1.5,
                            "answer": form_data.get('q4_1', ''),
                            "sub_questions": {
                                "4.1.1": {
                                    "question": "Provide date of last such update.",
                                    "marks": "-",
                                    "answer": form_data.get('q4_1_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "4.2": {
                            "question": "Whether rules (Access Control List) are being configured in firewall/router/switches only after approval from competent authority?",
                            "marks": 1,
                            "answer": form_data.get('q4_2', '')
                        },
                        "4.3": {
                            "question": "Whether access control Lists of any networking devices (Firewall, Router, Wi-Fi device, Proxy etc.) are being reviewed periodically?",
                            "marks": 1.5,
                            "answer": form_data.get('q4_3', ''),
                            "sub_questions": {
                                "4.3.1": {
                                    "question": "Provide date of last review.",
                                    "marks": "-",
                                    "answer": form_data.get('q4_3_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "4.4": {
                            "question": "Whether bank ensures, that there is no wireless (wi-fi) facility for systems in bank/branches without prior approval from competent authority?",
                            "marks": 1,
                            "answer": form_data.get('q4_4', '')
                        },
                        "4.5": {
                            "question": "If wireless facility is present, whether bank has placed necessary security mechanism?",
                            "marks": 1,
                            "answer": form_data.get('q4_5', '')
                        },
                        "4.6": {
                            "question": "Whether there are separate network for internet & intranet server and user zone e.g. communication between these vlans/ network are through Access Control List?",
                            "marks": 2,
                            "answer": form_data.get('q4_6', '')
                        },
                        "4.7": {
                            "question": "Whether bank maintains IT/Network architecture diagram/document approved by the management?",
                            "marks": 1,
                            "answer": form_data.get('q4_7', ''),
                            "sub_questions": {
                                "4.7.1": {
                                    "question": "Provide date of approval.",
                                    "marks": "-",
                                    "answer": form_data.get('q4_7_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "4.8": {
                            "question": "Whether bank ensures that all the network connection from/to outside bank (e.g. internet) is routed through controlled access (e.g. through firewall and other applicable security solution)?",
                            "marks": 1,
                            "answer": form_data.get('q4_8', '')
                        }
                    }
                },
                "antivirus_and_patch_management": {
                    "title": "Antivirus and patch management",
                    "questions": {
                        "5.1": {
                            "question": "Whether Antivirus installed in all Servers/ PCs / endpoints?",
                            "marks": 1,
                            "answer": form_data.get('q5_1', ''),
                            "sub_questions": {
                                "5.1.1": {
                                    "question": "Provide total systems (in percentage) not having antivirus.",
                                    "marks": "-",
                                    "answer": form_data.get('q5_1_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "5.2": {
                            "question": "Whether Antivirus updated regularly in all Servers/ PCs / endpoints?",
                            "marks": 1,
                            "answer": form_data.get('q5_2', ''),
                            "sub_questions": {
                                "5.2.1": {
                                    "question": "Provide total systems (in percentage) not updated within last 7 days.",
                                    "marks": "-",
                                    "answer": form_data.get('q5_2_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "5.3": {
                            "question": "Whether patches are regularly updated?",
                            "marks": 1,
                            "answer": form_data.get('q5_3', ''),
                            "sub_questions": {
                                "5.3.1": {
                                    "question": "Provide total systems (in percentage) not updated within last 30 days.",
                                    "marks": "-",
                                    "answer": form_data.get('q5_3_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "5.4": {
                            "question": "Whether register for patch update is maintained?",
                            "marks": 0.5,
                            "answer": form_data.get('q5_4', '')
                        },
                        "5.5": {
                            "question": "Whether it is ensured that there is no out of support operating system/database (e.g. windows 7) version being used within the bank.",
                            "marks": 1,
                            "answer": form_data.get('q5_5', ''),
                            "sub_questions": {
                                "5.5.1": {
                                    "question": "Provide No. systems having out of support OS.",
                                    "marks": "-",
                                    "answer": form_data.get('q5_5_1', ''),
                                    "type": "number"
                                }
                            }
                        }
                    }
                },
                "access_control": {
                    "title": "Access Control",
                    "questions": {
                        "6.1": {
                            "question": "Whether user rights are provided based on minimum access required to perform their duty (least privilege)?",
                            "marks": 1,
                            "answer": form_data.get('q6_1', '')
                        },
                        "6.2": {
                            "question": "Whether bank ensures there is no shared/ generic user ids are being used?",
                            "marks": 1,
                            "answer": form_data.get('q6_2', ''),
                            "sub_questions": {
                                "6.2.1": {
                                    "question": "Provide No. of shared/ generic user ids present in the critical servers/applications.",
                                    "marks": "-",
                                    "answer": form_data.get('q6_2_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "6.3": {
                            "question": "If shared/ generic user ids are being used, has approval been obtained from competent authority?",
                            "marks": 1,
                            "answer": form_data.get('q6_3', '')
                        },
                        "6.4": {
                            "question": "Whether periodic user access review is conducted for all the critical systems?",
                            "marks": 1,
                            "answer": form_data.get('q6_4', ''),
                            "sub_questions": {
                                "6.4.1": {
                                    "question": "Provide date of last review.",
                                    "marks": "-",
                                    "answer": form_data.get('q6_4_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "6.5": {
                            "question": "Whether Admin rights restricted to end users?",
                            "marks": 2,
                            "answer": form_data.get('q6_5', '')
                        },
                        "6.6": {
                            "question": "Whether centralised system/process has been implemented to manage and monitor privileged/Super User / administrative access to critical systems?",
                            "marks": 2,
                            "answer": form_data.get('q6_6', '')
                        },
                        "6.7": {
                            "question": "Whether Remote Desktop Protocol (RDP) disabled on endpoints and limited on the servers?",
                            "marks": 2,
                            "answer": form_data.get('q6_7', ''),
                            "sub_questions": {
                                "6.7.1": {
                                    "question": "Provide No. of end points (in percentage) where RDP is enabled.",
                                    "marks": "-",
                                    "answer": form_data.get('q6_7_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "6.8": {
                            "question": "Whether access to all intranet resources are restricted over the internet, if allowed it is through virtual private network (VPN) only after approval from CISO or competent authority?",
                            "marks": 1.5,
                            "answer": form_data.get('q6_8', ''),
                            "sub_questions": {
                                "6.8.1": {
                                    "question": "Provide No. of intranet systems where access is enabled over the internet without VPN.",
                                    "marks": "-",
                                    "answer": form_data.get('q6_8_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "6.9": {
                            "question": "Does the bank have a two factor authentication (2FA) for CBS, CBS linked applications and other critical applications with second factor being dynamic?",
                            "marks": 2,
                            "answer": form_data.get('q6_9', ''),
                            "sub_questions": {
                                "6.9.1": {
                                    "question": "Provide No. of users (in percentage) who are not enabled for 2FA for CBS.",
                                    "marks": "-",
                                    "answer": form_data.get('q6_9_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "6.10": {
                            "question": "Whether security review of terminals used to access corporate internet banking applications of sponsor bank or being used for administrative access of servers is being done through a qualified IS auditor?",
                            "marks": 1,
                            "answer": form_data.get('q6_10', ''),
                            "sub_questions": {
                                "6.10.1": {
                                    "question": "Provide date of last review.",
                                    "marks": "-",
                                    "answer": form_data.get('q6_10_1', ''),
                                    "type": "date"
                                }
                            }
                        }
                    }
                }
            }
        }
        
        # Calculate total marks
        total_marks = 0
        implemented_count = 0
        not_implemented_count = 0
        
        for section_name, section_data in vics_data["sections"].items():
            print(f"\n📊 {section_data['title']}:")
            for q_num, q_data in section_data["questions"].items():
                if q_data["answer"]:
                    if q_data["answer"] == "Implemented":
                        implemented_count += 1
                        if isinstance(q_data["marks"], (int, float)):
                            total_marks += q_data["marks"]
                    elif q_data["answer"] == "Not Implemented":
                        not_implemented_count += 1
                    
                    print(f"  {q_num}: {q_data['answer']} (Marks: {q_data['marks']})")
                    
                    # Check sub-questions
                    if "sub_questions" in q_data:
                        for sub_q_num, sub_q_data in q_data["sub_questions"].items():
                            if sub_q_data["answer"]:
                                print(f"    {sub_q_num}: {sub_q_data['answer']} ({sub_q_data['type']})")
        
        # Add summary to data
        vics_data["summary"] = {
            "total_marks_obtained": total_marks,
            "total_questions_answered": implemented_count + not_implemented_count,
            "implemented_count": implemented_count,
            "not_implemented_count": not_implemented_count,
            "compliance_percentage": round((implemented_count / (implemented_count + not_implemented_count)) * 100, 2) if (implemented_count + not_implemented_count) > 0 else 0
        }
        
        print(f"\n📈 SUMMARY:")
        print(f"  Total Marks Obtained: {total_marks}")
        print(f"  Questions Answered: {implemented_count + not_implemented_count}")
        print(f"  Implemented: {implemented_count}")
        print(f"  Not Implemented: {not_implemented_count}")
        print(f"  Compliance Percentage: {vics_data['summary']['compliance_percentage']}%")
        print("="*80)
        
        # Clean up old VICS files before generating new one
        print("\n🧹 Cleaning up old VICS files...")
        cleanup_old_vics_files()
        
        # Generate Excel file
        excel_file_path = create_vics_part2_excel(vics_data)
        print(f"\n✅ Excel file generated: {excel_file_path}")
        
        # Save to static/uploads directory for download
        filename = f"VICS_Part_2_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_dir = "static/uploads"
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, filename)
        
        # Copy the temp file to static/uploads
        import shutil
        shutil.copy(excel_file_path, output_path)
        
        # Clean up temp file
        try:
            os.unlink(excel_file_path)
        except:
            pass
        
        return jsonify({
            "success": True,
            "message": "VICS Part 2 data processed successfully",
            "data": vics_data,
            "excel_file": filename,
            "download_url": f"/static/uploads/{filename}"
        })
        
    except Exception as e:
        print(f"❌ Error processing VICS Part 2: {e}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            "success": False,
            "message": f"Error processing VICS Part 2: {str(e)}",
            "data": None
        }), 500

