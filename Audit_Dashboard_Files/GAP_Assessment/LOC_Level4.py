from flask import Blueprint, request, jsonify
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import tempfile
import shutil
import glob

# Create blueprint for LOC Level 4
loc_level4_bp = Blueprint('loc_level4', __name__)

def cleanup_loc_files():
    """Clean up old LOC Excel files from uploads directory"""
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            loc_pattern = os.path.join(upload_dir, 'LOC_Level*.xlsx')
            old_files = glob.glob(loc_pattern)
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

def create_loc_level4_excel(loc_data):
    """
    Create Excel file for LOC Level 4 with formatted data
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "LOC"
    
    # Set column widths (adding column E for Section 21 details)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 30  # For Section 21 input details
    
    # Define styles
    header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Times New Roman', size=12)
    data_font_bold = Font(name='Times New Roman', size=12, bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set headers
    ws['A1'] = "Sr. No."
    ws['B1'] = "Questions"
    ws['C1'] = "Yes/No given by Auditor"
    ws['D1'] = "Auditor's Observation"
    
    # Apply header formatting
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}1']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 30
    
    current_row = 2
    
    # Define all observations (continuing in next part due to size)
    observations = {
        "20_A": {
            "Yes": "The bank has a CSOC(Cyber Security Operations Centre) in place.",
            "No": "The bank does not have a CSOC or any equivalent continuous surveillance mechanism.",
            "Not Applicable": "Not applicable as cyber security monitoring is fully outsourced to external service providers."
        },
        "22_A": {
            "Yes": "Guidelines covering eligibility, roles, responsibilities, and detailed functionalities of the CSOC have been established and formally documented.",
            "No": "Guidelines for CSOC roles, responsibilities, and functionalities are not formally defined.",
            "Not Applicable": "Not applicable as CSOC operations are fully managed by external providers."
        },
        "22_B": {
            "Yes": "CSOC operations comply with applicable laws and regulatory requirements for protection of business and customer data.",
            "No": "CSOC operations are not fully aligned with regulatory or legal requirements.",
            "Not Applicable": "Not applicable as all regulatory compliance responsibilities are handled by a third-party CSOC provider."
        },
        "22_C": {
            "Yes": "Compliance can be demonstrated through proper documentation, policies, processes, logs, and governance records.",
            "No": "Compliance cannot be adequately demonstrated due to missing or incomplete documentation or oversight.",
            "Not Applicable": "Not applicable as all CSOC compliance evidence is maintained by external service providers."
        },
        "22_D": {
            "Yes": "CSOC provides real-time or near real-time security posture updates to senior management.",
            "No": "CSOC cannot provide real-time information to senior management.",
            "Not Applicable": "Not applicable as monitoring and reporting are fully managed externally."
        },
        "22_E": {
            "Yes": "CSOC monitors network activity logs and escalates abnormal or undesired events appropriately.",
            "No": "CSOC monitoring and escalation mechanisms are insufficient or not implemented.",
            "Not Applicable": "Not applicable as log monitoring and escalation are handled by an external provider."
        },
        "22_F": {
            "Yes": "All log types and logging options are integrated with the SIEM system for centralized analysis and response.",
            "No": "Integration of all logs with SIEM is incomplete or not implemented.",
            "Not Applicable": "Not applicable as SIEM operations are fully managed by an external provider."
        },
        "23_A": {
            "Yes": "A technology framework is designed and implemented to provide proactive monitoring capabilities and ensure compliance with business and regulatory requirements.",
            "No": "The technology framework is inadequate to provide proactive monitoring or meet regulatory requirements.",
            "Not Applicable": "Not applicable as monitoring and compliance framework is fully managed by external service providers."
        },
        "23_B": {
            "Yes": "The CSOC has a security analytics engine that processes logs efficiently and provides actionable recommendations for further investigation.",
            "No": "Security analytics capabilities are insufficient or unable to provide timely insights.",
            "Not Applicable": "Not applicable as analytics and log processing are performed by a third-party CSOC provider."
        },
        "23_C": {
            "Yes": "Deep packet inspection (DPI) mechanisms are deployed to analyze network traffic for threats.",
            "No": "DPI mechanisms are not implemented within the network infrastructure.",
            "Not Applicable": "Not applicable as DPI operations are fully outsourced."
        },
        "23_D": {
            "Yes": "CSOC is equipped with malware detection and data imaging technologies to analyze and secure data effectively.",
            "No": "Malware detection and data imaging solutions are not in place.",
            "Not Applicable": "Not applicable as malware detection and imaging are fully managed externally."
        },
        "23_E": {
            "Yes": "CSOC has forensic analysis capabilities to investigate incidents and gather evidence.",
            "No": "Forensic analysis capabilities are not available within the CSOC.",
            "Not Applicable": "Not applicable as forensic investigations are outsourced to specialized vendors."
        },
        "23_F": {
            "Yes": "CSOC is designed to scale up or down based on organizational and operational requirements.",
            "No": "CSOC lacks scalability features, limiting future expansion or load handling.",
            "Not Applicable": "Not applicable as scalability considerations are managed externally."
        },
        "23_G": {
            "Yes": "CSOC operations are designed for high availability to ensure continuous monitoring.",
            "No": "CSOC does not have high availability arrangements in place.",
            "Not Applicable": "Not applicable as high availability is ensured through external service providers."
        },
        "24_A": {
            "Yes": "The bank actively participates in cyber drills conducted by CERT-IN, IDRBT, or similar agencies.",
            "No": "The bank does not participate in organized cyber drills.",
            "Not Applicable": "Not applicable if the bank's participation in such drills is managed entirely by a regulatory or parent organization."
        },
        "25_A": {
            "Yes": "Incident response capabilities are implemented across all interconnected systems and networks.",
            "No": "Incident response capabilities are limited or not consistently implemented across interconnected systems.",
            "Not Applicable": "Not applicable as all incident response activities are fully managed by external service providers."
        },
        "25_B": {
            "Yes": "Incident response capabilities are extended to cover vendor and partner networks through agreements or integrated procedures.",
            "No": "Incident response for vendor/partner networks is not established or is inconsistent.",
            "Not Applicable": "Not applicable as all vendor and partner networks are monitored and managed externally."
        },
        "25_C": {
            "Yes": "Regular tests and simulations are conducted to demonstrate readiness and recovery capabilities in a coordinated manner.",
            "No": "Readiness tests or simulations are infrequent or absent.",
            "Not Applicable": "Not applicable as recovery readiness is entirely managed by external providers or third-party vendors."
        },
        "25_D": {
            "Yes": "Policies and frameworks are implemented to align SOC, Incident Response, and Digital Forensics for business continuity.",
            "No": "Policies and frameworks for alignment are inadequate or absent.",
            "Not Applicable": "Not applicable as alignment is fully managed by external service providers."
        },
        "26_A": {
            "Yes": "KPIs and KRIs are defined and tracked to measure cybersecurity performance and risk.",
            "No": "Metrics are incomplete, inconsistent, or not formally tracked.",
            "Not Applicable": "Not applicable as performance metrics are fully managed by third-party service providers."
        },
        "26_B": {
            "Yes": "Arrangements are in place to conduct network forensics and forensic investigations when required.",
            "No": "No formal arrangements exist for network or digital forensics.",
            "Not Applicable": "Not applicable as forensic investigations are fully handled by external vendors."
        },
        "26_C": {
            "Yes": "DDoS mitigation services are in place and ready for activation as needed.",
            "No": "No standby arrangements for DDoS mitigation exist.",
            "Not Applicable": "Not applicable as DDoS mitigation is fully managed by third-party service providers."
        },
        "27_A": {
            "Yes": "The policy clearly documents the existing and proposed hardware/network architecture along with the rationale for design decisions.",
            "No": "The policy does not adequately cover current or proposed hardware/network architecture.",
            "Not Applicable": "Not applicable as hardware/network design decisions are fully managed by external vendors."
        },
        "27_B": {
            "Yes": "The policy defines standards for hardware and software aligned with the proposed architecture.",
            "No": "Hardware/software standards are missing or inadequately defined in the policy.",
            "Not Applicable": "Not applicable as hardware/software procurement and standardization are fully outsourced."
        },
        "27_C": {
            "Yes": "The policy includes a strategy for managing outsourcing, insourcing, procurement, and in-house development.",
            "No": "No clear strategy is defined in the policy for sourcing or development decisions.",
            "Not Applicable": "Not applicable as sourcing and development strategies are determined by external service providers."
        },
        "27_D": {
            "Yes": "The policy specifies the departmental organizational structure related to IT and cybersecurity functions.",
            "No": "The policy does not define departmental roles or structure clearly.",
            "Not Applicable": "Not applicable as organizational structure is determined and maintained by an external management service."
        },
        "27_E": {
            "Yes": "The policy outlines required IT competencies, gap assessment, and training/development plans.",
            "No": "Competency requirements and development plans are missing or incomplete.",
            "Not Applicable": "Not applicable as staff competency and training are managed entirely by external service providers."
        },
        "27_F": {
            "Yes": "The policy includes a strategy for monitoring technological developments and timely system updates.",
            "No": "No strategy exists to keep systems updated or monitor technology trends.",
            "Not Applicable": "Not applicable as technology monitoring and system updates are fully handled by third-party providers."
        },
        "27_G": {
            "Yes": "The policy includes provisions for independent assessment, evaluation, and monitoring of IT risks and audit findings.",
            "No": "Independent assessment and monitoring strategies are not clearly defined in the policy.",
            "Not Applicable": "Not applicable as independent assessments are conducted entirely by external audit or service providers."
        },
        "28_A": {
            "Yes": "The bank has a dedicated cyber security function staffed with skilled personnel focused on managing cybersecurity.",
            "No": "Cyber security function is inadequately staffed or lacks exclusive focus on cybersecurity management.",
            "Not Applicable": "Not applicable as cyber security responsibilities are entirely outsourced to an external provider."
        },
        "29_A": {
            "Yes": "The IT Strategy Committee is properly constituted with required board representation including technical expertise.",
            "No": "Committee does not meet composition requirements or lacks technical expertise.",
            "Not Applicable": "Not applicable as IT strategy is fully managed by external consultants or parent organization."
        },
        "29_B": {
            "Yes": "All IT policy documents, organizational structure, and investments have been reviewed and approved by the committee.",
            "No": "Approval for IT policies, structure, or investments is missing or incomplete.",
            "Not Applicable": "Not applicable as approvals are delegated entirely to external advisors or service providers."
        },
        "29_C": {
            "Yes": "Committee regularly reviews IT infrastructure performance and policy adherence.",
            "No": "Committee does not conduct regular reviews of IT performance or policy implementation.",
            "Not Applicable": "Not applicable as IT performance monitoring is performed entirely by third-party providers."
        },
        "30_A": {
            "Yes": "IT Steering Committee is constituted with representatives from IT, HR, legal, and business functions.",
            "No": "IT Steering Committee is incomplete or not formally constituted.",
            "Not Applicable": "Not applicable as IT project oversight is fully managed by external providers."
        },
        "30_B": {
            "Yes": "Progress of IT projects is reported regularly to the IT Strategy Committee by the IT Steering Committee.",
            "No": "Project reporting is inconsistent or not submitted to the IT Strategy Committee.",
            "Not Applicable": "Not applicable as project reporting is handled entirely by external consultants or vendors."
        },
        "31_A": {
            "Yes": "The CISO possesses the required technical qualifications and expertise for the role.",
            "No": "The CISO lacks necessary technical background or experience.",
            "Not Applicable": "Not applicable as CISO role is outsourced or advisory in nature."
        },
        "31_B": {
            "Yes": "Arrangements exist to scale human resources and expertise in line with business growth and technology adoption.",
            "No": "No formal arrangements exist to enhance resources or expertise.",
            "Not Applicable": "Not applicable as staffing and expertise enhancements are managed externally."
        },
        "31_C": {
            "Yes": "CISO reports independently and does not report directly to CIO/CTO, maintaining independence.",
            "No": "CISO has a direct reporting line to CIO/CTO, which may impact independence.",
            "Not Applicable": "Not applicable as CISO role is outsourced or independent reporting structure is not required."
        },
        "31_D": {
            "Yes": "CISO performance is focused on security objectives and not business targets.",
            "No": "CISO is assigned business targets, which may conflict with security priorities.",
            "Not Applicable": "Not applicable as CISO responsibilities are managed externally or advisory."
        },
        "31_E": {
            "Yes": "The Information Security Committee actively reviews security events, evaluates new threats, and reports findings to the Board.",
            "No": "The committee does not adequately review events or report to the Board.",
            "Not Applicable": "Not applicable as security event review and reporting are fully handled by external providers."
        }
    }
    
    # Define question structure
    sections = {
        "20": {
            "title": "Arrangement for continuous surveillance- Setting up of Cyber Security Operation Centre(C-SOC)",
            "questions": ["20_A"]
        },
        "21": {
            "title": "If Yes, CSOC in place then answer the following",
            "questions": ["21_A", "21_B", "21_C", "21_D", "21_E", "21_F"]
        },
        "22": {
            "title": "Expectations from C-SOC",
            "questions": ["22_A", "22_B", "22_C", "22_D", "22_E", "22_F"]
        },
        "23": {
            "title": "Steps for setting up C-SOC- Technological Aspects",
            "questions": ["23_A", "23_B", "23_C", "23_D", "23_E", "23_F", "23_G"]
        },
        "24": {
            "title": "Participation in Cyber Drills",
            "questions": ["24_A"]
        },
        "25": {
            "title": "Incident Response and Management",
            "questions": ["25_A", "25_B", "25_C", "25_D"]
        },
        "26": {
            "title": "Forensics and Metrics",
            "questions": ["26_A", "26_B", "26_C"]
        },
        "27": {
            "title": "IT Strategy and Policy",
            "questions": ["27_A", "27_B", "27_C", "27_D", "27_E", "27_F", "27_G"]
        },
        "28": {
            "title": "IT and IS governance Framework",
            "questions": ["28_A"]
        },
        "29": {
            "title": "IT stratgey Committee",
            "questions": ["29_A", "29_B", "29_C"]
        },
        "30": {
            "title": "IT Steering Committee",
            "questions": ["30_A", "30_B"]
        },
        "31": {
            "title": "CISO",
            "questions": ["31_A", "31_B", "31_C", "31_D", "31_E"]
        }
    }
    
    # Question text mapping
    question_texts = {
        "20_A": "Does the bank have continuous surveillance through Cyber Security Operations Centre (CSOC) ? (answer Yes if your Bank has CSOC of any kind - owned or shared or with ASP or MSSP, or with association or hybrid or any other)",
        "21_A": "CSOC fully owned and managed by it",
        "21_B": "Sharing CSOC with Apex Bank / Sponsor Bank/ group of DCCBs in the same State",
        "21_C": "Availing CSOC services from a shared CSOC set up by an Association / Federation",
        "21_D": "Engaged Managed Security Service Provider(MSSP)",
        "21_E": "Hybrid model - Own the infrastructure but managed by professionals from IT companies",
        "21_F": "Other",
        "22_A": "Whether guidelines for C-SOC  have been framed including eligibility criteria, roles and responsibility of the office and details of fuctionalities?",
        "22_B": "Whether CSOC complies with laws and regulations in the country for protection of business and customer data/ information?",
        "22_C": "Can the compliance for Q(93) be demonstrated? (through documentation and formulation of policies, processes, maintenance of logs, records  of activities and governance / oversight)",
        "22_D": "Ability to provide real-time/near-real time information on security posture of the RRB to senior management?",
        "22_E": "Can the C-SOC monitor logs of various network activities and escalate any abnormal / undesirable activities?",
        "22_F": "Whether all log types and logging options integrated with Security Information and Event Management (SIEM) system?",
        "23_A": "Whether, suitable technology framework designed and implemented to ensure proactive monitoring capabilities and business and regulatory requirements is available?",
        "23_B": "Does the CSOC include a security analytics engine which can process the logs within reasonable time frame and come out with possible recommendations with options for further deep dive investigations?",
        "23_C": "Whether deep packet inspection (DPI) introduced?",
        "23_D": "Does CSOC have tools and technologies for malware detection and imaging solutions for data?",
        "23_E": "Does CSOC have Forensic analysis capability?",
        "23_F": "Is the CSOC designed for scalability?",
        "23_G": "Does the CSOC offer high availability?",
        "24_A": "Whether your bank participates in cyber drills conducted under the aegis of Cert-IN, IDRBT etc.?",
        "25_A": "Have you developed incident response capabilities in all interconnected systems and networks?",
        "25_B": "Have you developed incident response capabilities for interconnected systems and networks  of vendors and partners?",
        "25_C": "Have you tested and demonstrated your readiness through collaborative and co-ordinated approach to meet the banks recovery time objectives?",
        "25_D": "Have you implemented policy and framework for aligning Security Operation Centre, Incident Response and Digital forensics to reduce the business downtime/to restore to normalcy?",
        "26_A": "Whether any comprehensive set of metrics have been developed that can provide prospective and retrospective measures, like key performance indicators, key risk indicators ?",
        "26_B": "Do you have support/arrangement for network forensics/forensic investigation?",
        "26_C": "Do you have support/arrangement for distributed denial-of-service(DDOS) mitigation services on stand-by?",
        "27_A": "Whether the Board approved Cyber Security Policy include existing and proposed hardware and network architecture for the bank and rationale?",
        "27_B": "Whether the Board approved Cyber Security Policy include Standards for hardware or software prescribed by the proposed architecture.",
        "27_C": "Whether the Board approved Cyber Security Policy include Strategy for outsourcing, in-sourcing, procuring off-the-shelf software, and in-house development.",
        "27_D": "Whether the Board approved Cyber Security Policy include Department's Organisational Structure",
        "27_E": "Whether the Board approved Cyber Security Policy include the desired number and level of IT expertise or competencies in banks human resources, plan to bridge the gap (if any) and requirements relating to training and development.",
        "27_F": "Whether the Board approved Cyber Security Policy include Strategy for keeping abreast with technology developments and to update systems as and when required.",
        "27_G": "Strategy for independent assessment, evaluation and monitoring of IT risks, findings of IT/IS/Cyber security related audits",
        "28_A": "Do you have a well equipped cyber security function (vertical / department / group) with adequate staff and expertise to focus exclusively on cyber security management?",
        "29_A": "Whether the Board Level IT strategy Committee is constituted with at least two members from the Board of which at least one member is technically qualified?",
        "29_B": "Has the IT strategy committee approved all policy documents, approved IT organizational structure, IT investment?",
        "29_C": "Does the IT Strategy Committee review performance of IT infrastructure, policy implementation?",
        "30_A": "Has the IT Steering Committee with representatives from the IT, HR, legal and business sectors been constituted?",
        "30_B": "Does the IT steering Committee report progress of IT implementation projects to the IT strategy Committee (of Board)?",
        "31_A": "Whether the CISO has the requisite technical background and expertise?",
        "31_B": "Whether arrangements have been done to enhance the human resource and expertise, if required, commensurate with increased business volume/ technology adoption and complexity?",
        "31_C": "Have you ensured that CISO has no direct relation with CIO/CTO?",
        "31_D": "Whether CISO is given any business targets?",
        "31_E": "Whether Information Security Committee is reviewing security events, asessing new developments affecting cyber security and reporting to Board?"
    }
    
    # Add "Level IV" header row
    ws[f'A{current_row}'].value = ""
    ws[f'B{current_row}'].value = "Level IV"
    ws[f'C{current_row}'].value = ""
    ws[f'D{current_row}'].value = ""
    
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}{current_row}']
        cell.font = data_font_bold
        cell.alignment = left_alignment
        cell.border = thin_border
    
    current_row += 1
    
    # Check if 20_A is No or Not Applicable
    q20_A_answer = loc_data.get('q20_A', '')
    csoc_not_available = (q20_A_answer == 'No' or q20_A_answer == 'Not Applicable')
    
    # Process each section
    for section_num, section_data in sections.items():
        # Add section header
        ws[f'A{current_row}'].value = section_num
        ws[f'B{current_row}'].value = section_data["title"]
        ws[f'C{current_row}'].value = ""
        ws[f'D{current_row}'].value = ""
        
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{current_row}']
            cell.font = data_font_bold
            cell.alignment = left_alignment if col == 'B' else center_alignment
            cell.border = thin_border
        
        current_row += 1
        
        # Add questions
        for q_id in section_data["questions"]:
            # Form data has keys like "q20_A", so add "q" prefix
            form_key = f"q{q_id}"
            answer = loc_data.get(form_key, "")
            
            # Extract letter from question ID
            letter = q_id.split('_')[1]
            
            ws[f'A{current_row}'].value = letter
            ws[f'B{current_row}'].value = question_texts.get(q_id, "")
            
            # For Section 21, handle dropdown + input field
            if section_num == "21":
                # Check if CSOC is not available (20_A is No or NA)
                if csoc_not_available:
                    # Set NA in column C and standard message in D
                    ws[f'C{current_row}'].value = "NA"
                    ws[f'D{current_row}'].value = "The bank does not have a CSOC."
                else:
                    # Get input field value
                    input_key = f"q{q_id}_input"
                    input_value = loc_data.get(input_key, "")
                    
                    # Set Yes/No/NA in column C
                    if answer == "Yes":
                        ws[f'C{current_row}'].value = "Yes"
                    elif answer == "No":
                        ws[f'C{current_row}'].value = "No"
                    elif answer == "Not Applicable":
                        ws[f'C{current_row}'].value = "NA"
                    else:
                        ws[f'C{current_row}'].value = ""
                    
                    # Set input details in column D
                    if input_value:
                        ws[f'D{current_row}'].value = input_value
                    else:
                        ws[f'D{current_row}'].value = ""
            else:
                # Normal questions - Set Yes/No/NA in column C
                if answer == "Yes":
                    ws[f'C{current_row}'].value = "Yes"
                elif answer == "No":
                    ws[f'C{current_row}'].value = "No"
                elif answer == "Not Applicable":
                    ws[f'C{current_row}'].value = "NA"
                else:
                    ws[f'C{current_row}'].value = ""
                
                # Set observation in column D
                observation = observations.get(q_id, {}).get(answer, "")
                ws[f'D{current_row}'].value = observation
            
            # Apply formatting
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].alignment = center_alignment
            ws[f'A{current_row}'].border = thin_border
            
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].alignment = left_alignment
            ws[f'B{current_row}'].border = thin_border
            
            ws[f'C{current_row}'].font = data_font
            ws[f'C{current_row}'].alignment = center_alignment
            ws[f'C{current_row}'].border = thin_border
            
            ws[f'D{current_row}'].font = data_font
            ws[f'D{current_row}'].alignment = left_alignment
            ws[f'D{current_row}'].border = thin_border
            
            current_row += 1
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name

@loc_level4_bp.route('/process_loc_level4', methods=['POST'])
def process_loc_level4():
    """
    Process LOC Level 4 form data and generate Excel
    """
    try:
        print("\n" + "="*80)
        print("🎯 LOC LEVEL 4 - List of Compliances")
        print("="*80)
        
        # Get form data
        form_data = request.form.to_dict()
        
        print("📋 Form Data Received:")
        for key, value in form_data.items():
            print(f"  {key}: {value}")
        
        # Clean up old LOC files before generating new one
        print("\n🧹 Cleaning up old LOC files...")
        cleanup_loc_files()
        
        # Generate Excel file
        print("\n📝 Generating Excel file...")
        excel_file_path = create_loc_level4_excel(form_data)
        print(f"✅ Excel file created: {excel_file_path}")
        
        # Save to static/uploads directory
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"LOC_Level4_{timestamp}.xlsx"
        static_upload_path = os.path.join('static', 'uploads', excel_filename)
        
        os.makedirs(os.path.dirname(static_upload_path), exist_ok=True)
        shutil.copy2(excel_file_path, static_upload_path)
        os.unlink(excel_file_path)
        
        print(f"✅ Excel saved to: {static_upload_path}")
        print("="*80)
        
        return jsonify({
            "success": True,
            "message": "LOC Level 4 data processed successfully",
            "download_url": f"/static/uploads/{excel_filename}",
            "excel_file": "LOC_Level4.xlsx"
        }), 200
        
    except Exception as e:
        print(f"❌ Error processing LOC Level 4: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"Error processing LOC Level 4: {str(e)}"
        }), 500

