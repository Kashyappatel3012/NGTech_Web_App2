from flask import Blueprint, request, jsonify, send_file
import json
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import tempfile
import glob

# Create blueprint for VICS Part 3
vics_part3_bp = Blueprint('vics_part3', __name__)

def cleanup_old_vics_files():
    """
    Clean up old VICS Excel files from uploads directory
    """
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            vics_pattern = os.path.join(upload_dir, 'VICS_Part*.xlsx')
            old_files = glob.glob(vics_pattern)
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

def create_vics_part3_excel(vics_data):
    """
    Create Excel file for VICS Part 3 with formatted data
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "VICS Part 3"
    
    # Set column widths
    column_widths = {
        'A': 20,
        'B': 60,
        'C': 20,
        'D': 20,
        'E': 20,
        'F': 60
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Merge A1 and B1
    ws.merge_cells('A1:B1')
    
    # Set header values
    headers = {
        'A1': 'A) Info Sec Processes & Controls',
        'C1': 'Max Marks',
        'D1': 'Yes/No',
        'E1': 'Marks given by the Auditor',
        'F1': "Auditor's Observation"
    }
    
    # Header styling
    header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')  # Dark Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply headers and styling
    for cell_ref, value in headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Populate data rows
    current_row = 2
    section_number = 7  # Part 3 starts from section 7
    
    # Track rows that need bold formatting
    bold_rows = []
    
    # Map section keys to their display numbers
    section_mapping = {
        "removable_media": "7",
        "secure_configuration_and_backup_management": "8",
        "cyber_security_framework_compliance": "9"
    }
    
    # Observations for each question
    observations = {
        "7.1": {
            "Implemented": "Use of removable media is disabled on all endpoints.",
            "Not Implemented": "Removable media ports are enabled on multiple systems, increasing the risk of data leakage."
        },
        "7.2": {
            "Implemented": "Temporary removable media access is granted only upon approval from the competent authority.",
            "Not Implemented": "Removable media access is not controlled or lacks an approval process."
        },
        "7.3": {
            "Implemented": "A register is maintained to record systems with removable media access permissions.",
            "Not Implemented": "No register is maintained for systems with removable media access."
        },
        "8.1": {
            "Implemented": "The bank maintains a secure configuration and hardening document for servers, network devices, and endpoints.",
            "Not Implemented": "No formal secure configuration or hardening document is available for IT assets."
        },
        "8.2": {
            "Implemented": "All servers and network devices are hardened as per approved configuration documents.",
            "Not Implemented": "Several systems are not hardened as per standard configuration guidelines."
        },
        "8.3": {
            "Implemented": "Configuration reviews for critical systems are conducted at least once a year.",
            "Not Implemented": "Periodic configuration reviews are not performed or not documented."
        },
        "8.4": {
            "Implemented": "Security and password management policies are enforced to all endpoints.",
            "Not Implemented": "Security policies are not enforced or are inconsistently applied across endpoints."
        },
        "8.5": {
            "Implemented": "The bank has an approved Backup and Restoration Policy defining backup frequency and retention.",
            "Not Implemented": "No formal Backup and Restoration Policy is documented or approved."
        },
        "8.6": {
            "Implemented": "Backups are taken regularly and stored securely as per policy.",
            "Not Implemented": "Backup schedules are irregular or storage practices do not meet policy requirements."
        },
        "8.7": {
            "Implemented": "Backup restoration tests are conducted periodically to validate data integrity.",
            "Not Implemented": "Backup restoration tests are not performed or lack documentation."
        },
        "8.8": {
            "Implemented": "A centralized logging system is implemented to collect and store logs from critical systems.",
            "Not Implemented": "Logs are not centrally stored or are scattered across individual systems."
        },
        "8.9": {
            "Implemented": "The Disaster Recovery site is located in a separate seismic zone to ensure operational resilience.",
            "Not Implemented": "The Disaster Recovery site is in the same or nearby seismic zone, reducing redundancy."
        },
        "9.1": {
            "Implemented": "Gap analysis of the Cyber Security Framework has been conducted to identify compliance gaps.",
            "Not Implemented": "No Cyber Security Framework gap analysis has been performed."
        },
        "9.2": {
            "Implemented": "A detailed compliance plan is prepared and reviewed periodically by top management.",
            "Not Implemented": "No formal Cyber Security Framework compliance plan is prepared or reviewed."
        }
    }
    
    # Helper function to format date as DD/MM/YYYY
    def format_date(date_str):
        if not date_str:
            return date_str
        try:
            # Parse date from YYYY-MM-DD format
            from datetime import datetime
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            return date_obj.strftime('%d/%m/%Y')
        except:
            return date_str
    
    for section_key, section_data in vics_data["sections"].items():
        section_title = section_data["title"]
        section_num = section_mapping.get(section_key, str(section_number))
        
        # Add section header row
        section_header_row = current_row
        bold_rows.append(section_header_row)  # Track for bold formatting
        
        ws.cell(row=current_row, column=1).value = section_num
        ws.cell(row=current_row, column=2).value = section_title
        ws.cell(row=current_row, column=3).value = ""
        ws.cell(row=current_row, column=4).value = ""
        ws.cell(row=current_row, column=5).value = ""
        ws.cell(row=current_row, column=6).value = ""
        current_row += 1
        
        # Track section totals
        section_total_marks = 0
        section_marks_given = 0
        
        for q_num, q_data in section_data["questions"].items():
            answer = q_data.get("answer", "")
            
            # Main question row
            ws.cell(row=current_row, column=1).value = q_num  # Column A - Question number
            ws.cell(row=current_row, column=2).value = q_data["question"]  # Column B - Question
            ws.cell(row=current_row, column=3).value = q_data["marks"]  # Column C - Max Marks
            
            # Column D - Yes/No
            yes_no = "Yes" if answer == "Implemented" else "No" if answer == "Not Implemented" else ""
            ws.cell(row=current_row, column=4).value = yes_no
            
            # Calculate marks given (full marks if Implemented, 0 if Not Implemented)
            marks_given = q_data["marks"] if answer == "Implemented" else 0
            ws.cell(row=current_row, column=5).value = marks_given  # Column E - Marks given
            
            # Column F - Observation
            observation = observations.get(q_num, {}).get(answer, "")
            ws.cell(row=current_row, column=6).value = observation
            
            # Add to section totals
            if isinstance(q_data["marks"], (int, float)):
                section_total_marks += q_data["marks"]
                section_marks_given += marks_given
            
            current_row += 1
            
            # Sub-questions
            if "sub_questions" in q_data:
                parent_answer = answer
                
                for sub_num, sub_data in q_data["sub_questions"].items():
                    ws.cell(row=current_row, column=1).value = sub_num
                    ws.cell(row=current_row, column=2).value = sub_data['question']
                    ws.cell(row=current_row, column=3).value = ""
                    ws.cell(row=current_row, column=4).value = ""
                    ws.cell(row=current_row, column=5).value = ""
                    
                    # Handle sub-question observations based on parent answer and sub-answer
                    sub_observation = ""
                    sub_answer = sub_data.get('answer', '')
                    
                    if parent_answer == "Not Implemented":
                        # Special cases for Not Implemented
                        if sub_num == "7.3.1":
                            sub_observation = f"No. of systems (in percentage) where USB and Hard drive are allowed: {sub_answer}%" if sub_answer else "-"
                        elif sub_num == "8.2.1":
                            sub_observation = f"No. of systems (in percentage) which are hardened as per document: {sub_answer}%" if sub_answer else "-"
                        elif sub_num == "8.3.1" or sub_num == "8.5.1" or sub_num == "8.7.1" or sub_num == "9.1.1" or sub_num == "9.2.1":
                            sub_observation = "-"
                        elif sub_num == "8.4.1":
                            sub_observation = f"No. of systems, where policy is not enforced: {sub_answer}" if sub_answer else "-"
                        else:
                            sub_observation = "-"
                    elif parent_answer == "Implemented":
                        # Special handling for each sub-question when Implemented
                        if sub_num == "7.3.1":
                            if sub_answer:
                                sub_observation = f"No. of systems (in percentage) where USB and Hard drive are allowed: {sub_answer}%"
                            else:
                                sub_observation = "-"
                        elif sub_num == "8.2.1":
                            sub_observation = "No. of systems (in percentage) which are hardened as per document: 100%"
                        elif sub_num == "8.3.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Last review date: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "8.4.1":
                            sub_observation = "No. of systems, where policy is not enforced: 0"
                        elif sub_num == "8.5.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of Approval: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "8.7.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of Last restoration: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "9.1.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of Gap analysis: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "9.2.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last compliance review of cyber security framework by top management: {formatted_date}"
                            else:
                                sub_observation = "-"
                        else:
                            sub_observation = "-"
                    
                    ws.cell(row=current_row, column=6).value = sub_observation
                    
                    current_row += 1
        
        # Add section total row
        total_row = current_row
        bold_rows.append(total_row)  # Track for bold formatting
        
        ws.cell(row=current_row, column=1).value = ""
        ws.cell(row=current_row, column=2).value = f"Total_A.{section_num}"
        ws.cell(row=current_row, column=3).value = section_total_marks
        ws.cell(row=current_row, column=4).value = ""
        ws.cell(row=current_row, column=5).value = section_marks_given
        ws.cell(row=current_row, column=6).value = ""
        
        current_row += 1
        section_number += 1
    
    # Apply borders and alignment to all data cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for row in range(1, current_row):
        for col in range(1, 7):  # Only A-F columns
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            # Header row (row 1) should always be center-aligned
            if row == 1:
                cell.alignment = center_alignment
            # Column B (questions) and F (observations) should be left-aligned for data rows
            elif col == 2 or col == 6:  # Column B and F
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment
            
            # Set font for data rows
            if row > 1:
                # Apply bold formatting to section headers and totals
                if row in bold_rows:
                    cell.font = Font(name='Times New Roman', size=12, bold=True)
                else:
                    cell.font = Font(name='Times New Roman', size=12)
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name

@vics_part3_bp.route('/process_vics_part3', methods=['POST'])
def process_vics_part3():
    """
    Process VICS Part 3 form data and return JSON response
    """
    try:
        print("\n" + "="*80)
        print("🎯 VICS PART 3 - Removable Media | Secure Configuration and Backup Management | Cyber Security Framework Compliance")
        print("="*80)
        
        # Get form data
        form_data = request.form.to_dict()
        
        # Print all form data for debugging
        print("📋 Form Data Received:")
        for key, value in form_data.items():
            print(f"  {key}: {value}")
        
        # Organize data by sections
        vics_data = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "part": "VICS Part 3: Removable Media | Secure Configuration and Backup Management | Cyber Security Framework Compliance",
            "sections": {
                "removable_media": {
                    "title": "Removable Media",
                    "questions": {
                        "7.1": {
                            "question": "Whether removable media disallowed in all endpoints and servers?",
                            "marks": 1,
                            "answer": form_data.get('q7_1', '')
                        },
                        "7.2": {
                            "question": "Whether access of removable media is allowed for a specific period and only after approval from authority?",
                            "marks": 1,
                            "answer": form_data.get('q7_2', '')
                        },
                        "7.3": {
                            "question": "Whether register for systems, having access to Removable media is maintained?",
                            "marks": 1,
                            "answer": form_data.get('q7_3', ''),
                            "sub_questions": {
                                "7.3.1": {
                                    "question": "Provide No. of systems (in percentage) where USB is allowed.",
                                    "marks": "-",
                                    "answer": form_data.get('q7_3_1', ''),
                                    "type": "number"
                                }
                            }
                        }
                    }
                },
                "secure_configuration_and_backup_management": {
                    "title": "Secure Configuration and Backup Management",
                    "questions": {
                        "8.1": {
                            "question": "Whether bank maintains secure configuration/ hardening document for servers/network devices/end points etc?",
                            "marks": 1,
                            "answer": form_data.get('q8_1', '')
                        },
                        "8.2": {
                            "question": "Whether all the servers and network devices are hardened as per document?",
                            "marks": 1,
                            "answer": form_data.get('q8_2', ''),
                            "sub_questions": {
                                "8.2.1": {
                                    "question": "Provide No. of systems (in percentage) which are hardened as per document.",
                                    "marks": "-",
                                    "answer": form_data.get('q8_2_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "8.3": {
                            "question": "Whether configuration review of the critical systems is conducted periodically minimum annually?",
                            "marks": 1,
                            "answer": form_data.get('q8_3', ''),
                            "sub_questions": {
                                "8.3.1": {
                                    "question": "Provide last review date.",
                                    "marks": "-",
                                    "answer": form_data.get('q8_3_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "8.4": {
                            "question": "Whether security policies (including password management) are managed and enforced on all the endpoints?",
                            "marks": 1,
                            "answer": form_data.get('q8_4', ''),
                            "sub_questions": {
                                "8.4.1": {
                                    "question": "Provide no. of systems, where policy is not enforced.",
                                    "marks": "-",
                                    "answer": form_data.get('q8_4_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "8.5": {
                            "question": "Whether Bank has Backup and Restoration Policy?",
                            "marks": 1,
                            "answer": form_data.get('q8_5', ''),
                            "sub_questions": {
                                "8.5.1": {
                                    "question": "Provide date of approval.",
                                    "marks": "-",
                                    "answer": form_data.get('q8_5_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "8.6": {
                            "question": "Whether periodic backup is being taken and safe storage of important data is being ensured as per policy?",
                            "marks": 1,
                            "answer": form_data.get('q8_6', '')
                        },
                        "8.7": {
                            "question": "Whether backup restoration of critical system is tested periodically?",
                            "marks": 1,
                            "answer": form_data.get('q8_7', ''),
                            "sub_questions": {
                                "8.7.1": {
                                    "question": "Provide date of last restoration.",
                                    "marks": "-",
                                    "answer": form_data.get('q8_7_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "8.8": {
                            "question": "Whether bank has centralized logging mechanism to store logs of all the critical systems?",
                            "marks": 1,
                            "answer": form_data.get('q8_8', '')
                        },
                        "8.9": {
                            "question": "Whether Disaster Recovery site is located in different seismic zones?",
                            "marks": 1,
                            "answer": form_data.get('q8_9', '')
                        }
                    }
                },
                "cyber_security_framework_compliance": {
                    "title": "Cyber Security Framework Compliance",
                    "questions": {
                        "9.1": {
                            "question": "Whether Gap analysis of Cyber Security Framework (Availability vs. Requirement) has been carried out?",
                            "marks": 1,
                            "answer": form_data.get('q9_1', ''),
                            "sub_questions": {
                                "9.1.1": {
                                    "question": "Provide date of gap analysis.",
                                    "marks": "-",
                                    "answer": form_data.get('q9_1_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "9.2": {
                            "question": "Whether detailed plan for compliance of cyber security framework is submitted and regularly reviewed by the top management?",
                            "marks": 2,
                            "answer": form_data.get('q9_2', ''),
                            "sub_questions": {
                                "9.2.1": {
                                    "question": "Provide date of last compliance review of cyber security framework by top management.",
                                    "marks": "-",
                                    "answer": form_data.get('q9_2_1', ''),
                                    "type": "date"
                                }
                            }
                        }
                    }
                }
            }
        }
        
        # Print organized data
        print("\n" + "="*80)
        print("📊 ORGANIZED DATA:")
        print("="*80)
        print(json.dumps(vics_data, indent=2))
        
        # Calculate total marks
        total_marks = 0
        for section_key, section_data in vics_data["sections"].items():
            print(f"\n📌 {section_data['title']}:")
            for q_num, q_data in section_data["questions"].items():
                marks = q_data.get('marks', 0)
                if marks != '-' and marks != 0:
                    total_marks += marks
                answer = q_data.get('answer', 'Not answered')
                print(f"  {q_num}. {q_data['question']}")
                print(f"      Answer: {answer} | Marks: {marks}")
                
                # Print sub-questions if any
                if 'sub_questions' in q_data:
                    for sub_num, sub_data in q_data['sub_questions'].items():
                        sub_answer = sub_data.get('answer', 'Not answered')
                        print(f"    {sub_num}. {sub_data['question']}")
                        print(f"        Answer: {sub_answer} | Type: {sub_data.get('type', 'text')}")
        
        print(f"\n💯 Total Marks for Part 3: {total_marks}")
        print("="*80 + "\n")
        
        # Clean up old VICS files before generating new one
        print("🧹 Cleaning up old VICS files...")
        cleanup_old_vics_files()
        
        # Generate Excel file
        excel_file_path = create_vics_part3_excel(vics_data)
        print(f"\n✅ Excel file generated: {excel_file_path}")
        
        # Save to static/uploads directory for download
        filename = f"VICS_Part_3_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_dir = "static/uploads"
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, filename)
        
        # Copy the temp file to static/uploads
        import shutil
        shutil.copy(excel_file_path, output_path)
        
        # Clean up temp file
        try:
            os.unlink(excel_file_path)
        except:
            pass
        
        return jsonify({
            "success": True,
            "message": "VICS Part 3 data processed successfully",
            "data": vics_data,
            "total_marks": total_marks,
            "excel_file": filename,
            "download_url": f"/static/uploads/{filename}"
        }), 200
        
    except Exception as e:
        print(f"❌ Error processing VICS Part 3: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"Error processing VICS Part 3: {str(e)}"
        }), 500

