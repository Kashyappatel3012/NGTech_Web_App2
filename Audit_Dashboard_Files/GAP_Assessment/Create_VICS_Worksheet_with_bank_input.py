from flask import Blueprint, request, jsonify
import os
import tempfile
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import glob

# Create blueprint for Create VICS Worksheet with Bank Input
create_vics_with_bank_input_bp = Blueprint('create_vics_with_bank_input', __name__)

def cleanup_vics_with_bank_input_files():
    """Clean up old VICS with bank input files from uploads directory"""
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            vics_pattern = os.path.join(upload_dir, 'VICS_With_Bank_Input_*.xlsx')
            old_files = glob.glob(vics_pattern)
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

@create_vics_with_bank_input_bp.route('/cleanup_vics_with_bank_input', methods=['POST'])
def cleanup_vics_with_bank_input_endpoint():
    """Endpoint to cleanup VICS with bank input files after download"""
    try:
        cleanup_vics_with_bank_input_files()
        return jsonify({"success": True, "message": "Cleanup completed"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

def shift_images_right(ws, start_col_idx, num_cols=1):
    """
    Shift images to the right by num_cols starting from start_col_idx
    """
    images_to_update = []
    
    for img in ws._images[:]:
        try:
            # Get image anchor position
            if hasattr(img, 'anchor'):
                anchor = img.anchor
                
                # Check if anchor has _from attribute
                if hasattr(anchor, '_from'):
                    img_col = anchor._from.col
                    img_row = anchor._from.row
                    
                    # If image is in or after the column we're inserting before, shift it right
                    if img_col >= start_col_idx:
                        images_to_update.append({
                            'image': img,
                            'old_col': img_col,
                            'new_col': img_col + num_cols,
                            'row': img_row
                        })
                elif isinstance(anchor, str):
                    # If anchor is a string like "E5", parse it
                    import re
                    match = re.match(r'([A-Z]+)(\d+)', anchor)
                    if match:
                        col_letter = match.group(1)
                        row_num = int(match.group(2))
                        
                        # Convert column letter to index
                        from openpyxl.utils import column_index_from_string
                        img_col_idx = column_index_from_string(col_letter)
                        
                        if img_col_idx >= start_col_idx:
                            new_col_idx = img_col_idx + num_cols
                            new_col_letter = get_column_letter(new_col_idx)
                            new_anchor = f"{new_col_letter}{row_num}"
                            
                            # Update anchor
                            img.anchor = new_anchor
                            print(f"  📌 Moved image from {anchor} to {new_anchor}")
        except Exception as e:
            print(f"  ⚠️ Error processing image: {e}")
    
    # Update images with _from attribute and resize them
    for img_info in images_to_update:
        try:
            img = img_info['image']
            new_col = img_info['new_col']
            row = img_info['row']
            
            # Resize image by 30x (divide dimensions by 30)
            original_width = img.width
            original_height = img.height
            img.width = original_width / 30
            img.height = original_height / 30
            
            # Get the cell for the new position
            new_col_letter = get_column_letter(new_col + 1)  # +1 because openpyxl uses 0-based for _from
            new_row = row + 1  # +1 because openpyxl uses 0-based for _from
            cell = ws.cell(row=new_row, column=new_col + 1)
            
            # Update the anchor to the new cell
            img.anchor = cell.coordinate
            print(f"  📌 Moved and resized image to {cell.coordinate}")
        except Exception as e:
            print(f"  ⚠️ Error updating image position: {e}")

@create_vics_with_bank_input_bp.route('/create_vics_with_bank_input', methods=['POST'])
def create_vics_with_bank_input():
    """
    Process two VICS Excel files and merge them by inserting bank input column
    """
    try:
        print("\n" + "="*80)
        print("🚀 Creating VICS Worksheet with Bank Input")
        print("="*80)
        
        # Check if files are uploaded
        if 'vics_file1' not in request.files:
            return jsonify({"error": "No first VICS file uploaded"}), 400
        
        if 'vics_file2' not in request.files:
            return jsonify({"error": "No second VICS file uploaded"}), 400
        
        vics_file1 = request.files['vics_file1']
        vics_file2 = request.files['vics_file2']
        
        if vics_file1.filename == '' or vics_file2.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        # Create temporary directory
        temp_dir = tempfile.mkdtemp()
        print(f"📁 Created temp directory: {temp_dir}")
        
        try:
            # Save uploaded files
            file1_path = os.path.join(temp_dir, 'vics1.xlsx')
            file2_path = os.path.join(temp_dir, 'vics2.xlsx')
            
            vics_file1.save(file1_path)
            vics_file2.save(file2_path)
            print(f"💾 Saved uploaded files")
            
            # Clean up old files before generating new one
            print("\n🧹 Cleaning up old files...")
            cleanup_vics_with_bank_input_files()
            
            # Load both workbooks
            print("\n📖 Loading workbooks...")
            wb1 = load_workbook(file1_path)
            ws1 = wb1.active
            
            wb2 = load_workbook(file2_path)
            ws2 = wb2.active
            
            # Find "Marks given by Bank" column in second file
            print("\n🔍 Searching for 'Marks given by Bank' column in second file...")
            bank_marks_col = None
            
            for col_idx in range(1, ws2.max_column + 1):
                cell_value = ws2.cell(row=1, column=col_idx).value
                if cell_value and "Marks given by Bank" in str(cell_value):
                    bank_marks_col = col_idx
                    bank_marks_col_letter = get_column_letter(col_idx)
                    print(f"✅ Found 'Marks given by Bank' in column {bank_marks_col_letter} (index {bank_marks_col})")
                    break
            
            if not bank_marks_col:
                return jsonify({"error": "Could not find 'Marks given by Bank' column in second file"}), 400
            
            # Find the insertion point in first file (between D and E)
            # We want to insert the new column as column E (so it goes between current D and E)
            insert_col_idx = 5  # Column E (1-based: A=1, B=2, C=3, D=4, E=5)
            
            print(f"\n📝 Inserting new column at position {get_column_letter(insert_col_idx)} (index {insert_col_idx})")
            
            # Insert a new column at position E
            ws1.insert_cols(insert_col_idx)
            print(f"  ✅ Column inserted")
            
            # Shift images that are in column E or after to the right
            print(f"\n🖼️ Shifting images...")
            shift_images_right(ws1, insert_col_idx - 1)  # -1 because shift_images_right uses 0-based indexing
            
            # Copy the "Marks given by Bank" column from file 2 to the new column E in file 1
            print(f"\n📋 Copying 'Marks given by Bank' column data...")
            import re
            
            # Get column letters for conversion
            source_col_letter = get_column_letter(bank_marks_col)
            target_col_letter = get_column_letter(insert_col_idx)
            
            for row_idx in range(1, ws2.max_row + 1):
                source_cell = ws2.cell(row=row_idx, column=bank_marks_col)
                target_cell = ws1.cell(row=row_idx, column=insert_col_idx)
                
                # Check if source cell contains a formula
                if source_cell.data_type == 'f':  # Formula
                    formula = source_cell.value
                    
                    # Convert column references in the formula
                    # Replace source column letter with target column letter in the formula
                    # Pattern matches column letter followed by a number (like F24, F3, F10, etc.)
                    converted_formula = re.sub(
                        r'\b' + source_col_letter + r'(\d+)',
                        target_col_letter + r'\1',
                        formula,
                        flags=re.IGNORECASE
                    )
                    
                    target_cell.value = converted_formula
                    print(f"  🔢 Row {row_idx}: Converted formula '{formula}' to '{converted_formula}'")
                else:
                    # Copy value as-is, but sanitize to prevent Excel injection
                    from excel_security_utils import sanitize_excel_value
                    target_cell.value = sanitize_excel_value(source_cell.value)
                
                # Copy formatting
                if source_cell.has_style:
                    target_cell.font = source_cell.font.copy()
                    target_cell.border = source_cell.border.copy()
                    target_cell.fill = source_cell.fill.copy()
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = source_cell.protection.copy()
                    target_cell.alignment = source_cell.alignment.copy()
            
            # Copy column width
            
            if source_col_letter in ws2.column_dimensions:
                ws1.column_dimensions[target_col_letter].width = ws2.column_dimensions[source_col_letter].width
            
            print(f"  ✅ Copied {ws2.max_row} rows")
            
            # Fix merged cells for POC column (originally H1:M1, now should be I1:N1)
            print(f"\n🔧 Fixing merged cells for POC columns...")
            
            # Find and update merged cells that involve the POC column
            # The original merge was likely G1:L1 (POC Attached) and then H1:M1 (POC)
            # After insertion, they become H1:M1 (POC Attached) and I1:N1 (POC)
            
            # Get list of merged cell ranges
            merged_cells_to_remove = []
            merged_cells_to_add = []
            
            for merged_range in list(ws1.merged_cells.ranges):
                # Check if this is in row 1, 135, 185, 224, or 253
                if merged_range.min_row in [1, 135, 185, 224, 253]:
                    # If the merge involves columns after our insertion point
                    if merged_range.min_col >= insert_col_idx:
                        merged_cells_to_remove.append(merged_range)
                        
                        # Adjust the merge range (shift right by 1)
                        new_min_col = merged_range.min_col + 1
                        new_max_col = merged_range.max_col + 1
                        new_range = f"{get_column_letter(new_min_col)}{merged_range.min_row}:{get_column_letter(new_max_col)}{merged_range.max_row}"
                        merged_cells_to_add.append(new_range)
                        print(f"  📝 Adjusted merge: {merged_range} -> {new_range}")
            
            # Remove old merges and add new ones
            for merge_range in merged_cells_to_remove:
                ws1.unmerge_cells(str(merge_range))
            
            for merge_range in merged_cells_to_add:
                ws1.merge_cells(merge_range)
            
            # Add proper headers for POC Attached (H) and POC (I)
            print(f"\n📝 Adding POC headers...")
            from openpyxl.styles import PatternFill, Font, Alignment
            
            # Define styles
            dark_blue_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
            white_font_bold = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Add "POC Attached" in H1, H135, H185, H224
            for row_num in [1, 135, 185, 224]:
                h_cell = ws1.cell(row=row_num, column=8)  # Column H
                h_cell.value = "POC Attached"
                h_cell.font = white_font_bold
                h_cell.fill = dark_blue_fill
                h_cell.alignment = center_alignment
                print(f"  ✅ Added 'POC Attached' to H{row_num}")
            
            # Merge I1:N1, I135:N135, I185:N185, I224:N224 and add "POC"
            poc_rows = [1, 135, 185, 224]
            for row_num in poc_rows:
                # Merge I to N for this row
                merge_range = f"I{row_num}:N{row_num}"
                ws1.merge_cells(merge_range)
                
                # Add "POC" text with dark blue background
                i_cell = ws1.cell(row=row_num, column=9)  # Column I
                i_cell.value = "POC"
                i_cell.font = white_font_bold
                i_cell.fill = dark_blue_fill
                i_cell.alignment = center_alignment
                print(f"  ✅ Added 'POC' and merged {merge_range}")
            
            # Set background color for E1, E135, E185, E224, E253 to dark blue
            print(f"\n🎨 Setting background color for column E headers...")
            for row_num in [1, 135, 185, 224, 253]:
                e_cell = ws1.cell(row=row_num, column=5)
                e_cell.fill = dark_blue_fill
                e_cell.font = white_font_bold
                e_cell.alignment = center_alignment
                print(f"  ✅ Set dark blue background for E{row_num}")
            
            # Update column widths
            print(f"\n📏 Setting column widths...")
            column_widths = {
                'A': 15,
                'B': 60,
                'C': 20,
                'D': 20,
                'E': 20,
                'F': 20,
                'G': 60,
                'H': 20,
                'I': 20,
                'J': 20,
                'K': 20,
                'L': 20,
                'M': 20,
                'N': 20
            }
            
            for col_letter, width in column_widths.items():
                ws1.column_dimensions[col_letter].width = width
                print(f"  ✅ Set column {col_letter} width to {width}")
            
            # Update formulas - change column E references to column F
            print(f"\n🔢 Updating formulas...")
            
            # Formula cells that need updating
            formula_updates = {
                133: "=F11+F21+F31+F44+F55+F73+F79+F95+F101+F113+F122+F132",  # F133 (was E133)
                250: "=F133+F183+F222+F249",  # F250 (was E250)
                254: "=F133",  # F254 (was =E133)
                255: "=F183",  # F255 (was =E183)
                256: "=F222",  # F256 (was =E222)
                257: "=F249",  # F257 (was =E249)
                258: "=F250"   # F258 (was =E250)
            }
            
            for row_num, formula in formula_updates.items():
                f_cell = ws1.cell(row=row_num, column=6)  # Column F
                f_cell.value = formula
                print(f"  ✅ Updated F{row_num} formula to: {formula}")
            
            # Format all Column E data with Times New Roman, size 12
            print(f"\n📝 Formatting Column E data...")
            times_new_roman_font = Font(name='Times New Roman', size=12)
            
            for row_num in range(1, ws1.max_row + 1):
                e_cell = ws1.cell(row=row_num, column=5)  # Column E
                # Preserve existing font properties (bold, color) but ensure Times New Roman and size 12
                if e_cell.font:
                    e_cell.font = Font(
                        name='Times New Roman',
                        size=12,
                        bold=e_cell.font.bold,
                        italic=e_cell.font.italic,
                        color=e_cell.font.color
                    )
                else:
                    e_cell.font = times_new_roman_font
            
            print(f"  ✅ Formatted {ws1.max_row} rows in Column E")
            
            # Save the modified workbook
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"VICS_With_Bank_Input_{timestamp}.xlsx"
            output_path = os.path.join('static', 'uploads', output_filename)
            
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb1.save(output_path)
            wb1.close()
            wb2.close()
            
            print(f"\n✅ Modified workbook saved: {output_filename}")
            print("="*80)
            
            # Return download URL
            download_url = f"/static/uploads/{output_filename}"
            return jsonify({
                "success": True,
                "message": "VICS Worksheet with Bank Input created successfully",
                "download_url": download_url,
                "excel_file": "VICS_With_Bank_Input.xlsx"
            }), 200
            
        finally:
            # Clean up temporary directory
            try:
                shutil.rmtree(temp_dir)
                print("🧹 Cleaned up temporary files")
            except Exception as e:
                print(f"⚠️ Error cleaning up temp directory: {e}")
    
    except Exception as e:
        print(f"\n❌ Error creating VICS worksheet with bank input: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

