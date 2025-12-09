from flask import Blueprint, request, jsonify
import os
import zipfile
import tempfile
import shutil
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import glob
import re

# Create blueprint for Gap Assessment Excel
gap_assessment_excel_bp = Blueprint('gap_assessment_excel', __name__)

def cleanup_gap_assessment_files():
    """Clean up old Gap Assessment Excel files from uploads directory"""
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            gap_pattern = os.path.join(upload_dir, 'Gap_Assessment_*.xlsx')
            old_files = glob.glob(gap_pattern)
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

@gap_assessment_excel_bp.route('/cleanup_gap_assessment', methods=['POST'])
def cleanup_gap_assessment_endpoint():
    """Endpoint to cleanup Gap Assessment files after download"""
    try:
        cleanup_gap_assessment_files()
        return jsonify({"success": True, "message": "Cleanup completed"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

def sort_files_by_type(excel_files):
    """
    Sort Excel files by VICS, LOC, LOE order
    """
    vics_files = []
    loc_files = []
    loe_files = []
    other_files = []
    
    for file_info in excel_files:
        filename = file_info['filename'].upper()
        
        if 'VICS' in filename:
            vics_files.append(file_info)
        elif 'LOC' in filename:
            loc_files.append(file_info)
        elif 'LOE' in filename:
            loe_files.append(file_info)
        else:
            other_files.append(file_info)
    
    # Combine in order: VICS, LOC, LOE, Others
    sorted_files = vics_files + loc_files + loe_files + other_files
    
    print(f"\n📊 File sorting:")
    print(f"  VICS files: {len(vics_files)}")
    print(f"  LOC files: {len(loc_files)}")
    print(f"  LOE files: {len(loe_files)}")
    print(f"  Other files: {len(other_files)}")
    
    return sorted_files

def copy_worksheet_with_images(source_ws, target_ws, source_wb_path):
    """
    Copy worksheet content including images and formatting
    """
    from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
    
    # Copy all cells with data and formatting
    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws.cell(row=cell.row, column=cell.column)
            
            # Copy value
            target_cell.value = cell.value
            
            # Copy formatting
            if cell.has_style:
                target_cell.font = cell.font.copy()
                target_cell.border = cell.border.copy()
                target_cell.fill = cell.fill.copy()
                target_cell.number_format = cell.number_format
                target_cell.protection = cell.protection.copy()
                target_cell.alignment = cell.alignment.copy()
    
    # Copy column dimensions
    for col_letter, col_dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dim.width
        if col_dim.hidden:
            target_ws.column_dimensions[col_letter].hidden = col_dim.hidden
    
    # Copy row dimensions
    for row_num, row_dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_num].height = row_dim.height
        if row_dim.hidden:
            target_ws.row_dimensions[row_num].hidden = row_dim.hidden
    
    # Copy merged cells
    for merged_cell in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_cell))
    
    # Copy images
    for img in source_ws._images:
        try:
            # Create a new image object
            new_img = Image(img.ref)
            
            # Copy image properties
            new_img.width = img.width
            new_img.height = img.height
            
            # Copy anchor
            if hasattr(img, 'anchor'):
                new_img.anchor = img.anchor
            
            target_ws.add_image(new_img)
        except Exception as e:
            print(f"  ⚠️ Error copying image: {e}")

@gap_assessment_excel_bp.route('/create_gap_assessment_excel', methods=['POST'])
def create_gap_assessment_excel():
    """
    Process uploaded ZIP file and create Gap Assessment Excel with VICS, LOC, LOE worksheets
    """
    try:
        print("\n" + "="*80)
        print("🚀 Creating Gap Assessment Excel")
        print("="*80)
        
        # Check if file is uploaded
        if 'gap_assessment_zip' not in request.files:
            return jsonify({"error": "No ZIP file uploaded"}), 400
        
        gap_zip = request.files['gap_assessment_zip']
        
        if gap_zip.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        # Create temporary directory
        temp_dir = tempfile.mkdtemp()
        print(f"📁 Created temp directory: {temp_dir}")
        
        try:
            # Save uploaded ZIP file
            zip_path = os.path.join(temp_dir, 'gap_assessment.zip')
            gap_zip.save(zip_path)
            print(f"💾 Saved ZIP file")
            
            # Clean up old files before generating new one
            print("\n🧹 Cleaning up old files...")
            cleanup_gap_assessment_files()
            
            # Extract ZIP file
            extract_dir = os.path.join(temp_dir, 'extracted')
            os.makedirs(extract_dir, exist_ok=True)
            
            excel_files = []
            print(f"\n📦 Extracting ZIP file...")
            
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                for file_info in zip_ref.namelist():
                    if file_info.endswith('.xlsx') and not file_info.startswith('__MACOSX'):
                        zip_ref.extract(file_info, extract_dir)
                        extracted_path = os.path.join(extract_dir, file_info)
                        filename = os.path.basename(file_info)
                        excel_files.append({
                            'filename': filename,
                            'extracted_path': extracted_path
                        })
                        print(f"  📄 Extracted: {filename}")
            
            if not excel_files:
                return jsonify({"error": "No Excel files found in ZIP"}), 400
            
            # Sort files by type (VICS, LOC, LOE)
            print(f"\n🔄 Sorting files...")
            sorted_files = sort_files_by_type(excel_files)
            
            print(f"\n📋 Processing order:")
            for idx, file_info in enumerate(sorted_files, 1):
                print(f"  {idx}. {file_info['filename']}")
            
            # Create new workbook
            print(f"\n📝 Creating combined workbook...")
            combined_wb = Workbook()
            combined_wb.remove(combined_wb.active)  # Remove default sheet
            
            sheet_counter = 0
            vics_counter = 0
            loc_counter = 0
            loe_counter = 0
            
            # Copy worksheets from each file
            for file_info in sorted_files:
                file_path = file_info['extracted_path']
                filename = file_info['filename']
                filename_upper = filename.upper()
                
                print(f"\n📖 Processing: {filename}")
                
                try:
                    source_wb = load_workbook(file_path, data_only=False)
                    
                    # Copy all sheets from this workbook
                    for source_sheet_name in source_wb.sheetnames:
                        source_ws = source_wb[source_sheet_name]
                        
                        # Determine sheet name based on file type
                        if 'VICS' in filename_upper:
                            sheet_name = 'VICS' if vics_counter == 0 else f'VICS_{vics_counter}'
                            vics_counter += 1
                        elif 'LOC' in filename_upper:
                            sheet_name = 'LOC' if loc_counter == 0 else f'LOC_{loc_counter}'
                            loc_counter += 1
                        elif 'LOE' in filename_upper:
                            sheet_name = 'LOE' if loe_counter == 0 else f'LOE_{loe_counter}'
                            loe_counter += 1
                        else:
                            sheet_name = os.path.splitext(filename)[0][:31]
                            sheet_name = sheet_name.replace('[', '').replace(']', '').replace('*', '').replace('?', '')
                        
                        # Create new sheet
                        new_ws = combined_wb.create_sheet(title=sheet_name)
                        
                        print(f"  📄 Copying sheet: {source_sheet_name} -> {sheet_name}")
                        
                        # Copy worksheet content
                        copy_worksheet_with_images(source_ws, new_ws, file_path)
                        
                        sheet_counter += 1
                        print(f"  ✅ Sheet copied successfully")
                    
                    source_wb.close()
                    
                except Exception as e:
                    print(f"  ⚠️ Error processing {filename}: {e}")
            
            # Save the combined workbook
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"Gap_Assessment_{timestamp}.xlsx"
            output_path = os.path.join('static', 'uploads', output_filename)
            
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            combined_wb.save(output_path)
            combined_wb.close()
            
            print(f"\n✅ Combined workbook created: {output_filename}")
            print(f"   Total sheets: {sheet_counter}")
            print("="*80)
            
            # Return download URL
            download_url = f"/static/uploads/{output_filename}"
            return jsonify({
                "success": True,
                "message": "Gap Assessment Excel created successfully",
                "download_url": download_url,
                "excel_file": "Gap_Assessment.xlsx"
            }), 200
            
        finally:
            # Clean up temporary directory
            try:
                shutil.rmtree(temp_dir)
                print("🧹 Cleaned up temporary files")
            except Exception as e:
                print(f"⚠️ Error cleaning up temp directory: {e}")
    
    except Exception as e:
        print(f"\n❌ Error creating Gap Assessment Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

