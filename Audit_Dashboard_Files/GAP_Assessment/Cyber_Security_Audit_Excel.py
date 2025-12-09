from flask import Blueprint, request, jsonify
import zipfile
import os
import tempfile
import shutil
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import re
import glob

# Create blueprint for Cyber Security Audit Excel
cyber_security_audit_excel_bp = Blueprint('cyber_security_audit_excel', __name__)

def cleanup_cyber_security_audit_files():
    """Clean up old Cyber Security Audit Excel files from uploads directory"""
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            pattern = os.path.join(upload_dir, 'Cyber_Security_Audit*.xlsx')
            old_files = glob.glob(pattern)
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

@cyber_security_audit_excel_bp.route('/cleanup_cyber_security_audit', methods=['POST'])
def cleanup_cyber_security_audit_endpoint():
    """Endpoint to cleanup Cyber Security Audit files after download"""
    try:
        cleanup_cyber_security_audit_files()
        return jsonify({"success": True, "message": "Cleanup completed"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

def sort_meity_files(file_list):
    """Sort Meity files in the order: Part1, Part2, Part3"""
    def get_part_number(filename):
        match = re.search(r'Part[_\s]*(\d+)', filename, re.IGNORECASE)
        if match:
            return int(match.group(1))
        return 999
    
    return sorted(file_list, key=get_part_number)

def parse_image_number(filename):
    """
    Parse image filename to extract question number
    Handles: 1.png, 1_1.jpg, 1_kjnnnvnvdfnvj.png, 1.1fduvhiufhvuih.jpg, 1ekjwnfnjf.png, etc.
    All considered as question "1"
    """
    base_name = os.path.splitext(filename)[0]
    
    # Try to extract the leading number
    match = re.match(r'^(\d+)', base_name)
    if match:
        return match.group(1)
    
    return None

def extract_images_from_zip(zip_path, extract_dir):
    """Extract images from ZIP file and organize by question number
    Returns dict: {'1': [path1, path2, ...], '5': [path1], ...}
    """
    image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif'}
    image_mapping = {}
    
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
            print(f"📂 Extracted ZIP file to: {extract_dir}")
            
            # Find all images
            for root, dirs, files in os.walk(extract_dir):
                for file in files:
                    file_ext = os.path.splitext(file)[1].lower()
                    if file_ext in image_extensions:
                        image_path = os.path.join(root, file)
                        question_num = parse_image_number(file)
                        
                        if question_num:
                            if question_num not in image_mapping:
                                image_mapping[question_num] = []
                            image_mapping[question_num].append(image_path)
                            print(f"  📷 Mapped: {file} → Question {question_num}")
        
        # Sort images within each question number group
        for q_num in image_mapping:
            image_mapping[q_num].sort()
        
        return image_mapping
    
    except Exception as e:
        print(f"❌ Error extracting images: {e}")
        return {}

def insert_images_into_worksheet(ws, image_mapping, temp_dir):
    """Insert images into worksheet based on question number in Column A"""
    try:
        print(f"\n🖼️ Inserting images into worksheet...")
        
        # Column order for image insertion
        image_columns = ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'F']
        
        annexure_number = 1
        
        # Iterate through all rows to find matching question numbers
        for row_num in range(2, ws.max_row + 1):
            cell_a = ws.cell(row=row_num, column=1)
            question_num = str(cell_a.value) if cell_a.value is not None else None
            
            if question_num and question_num in image_mapping:
                images_for_question = image_mapping[question_num]
                num_images = min(len(images_for_question), 13)  # Max 13 images
                
                if num_images > 0:
                    print(f"  📍 Row {row_num}, Question {question_num}: {num_images} image(s)")
                    
                    # Add "Annexure <no>" in Column E (red color)
                    cell_e = ws.cell(row=row_num, column=5)
                    cell_e.value = f"Annexure {annexure_number}"
                    cell_e.font = Font(name='Times New Roman', size=12, color='FF0000', bold=False)
                    cell_e.alignment = Alignment(horizontal='center', vertical='center')
                    annexure_number += 1
                    
                    # Insert images
                    for img_idx in range(num_images):
                        col_letter = image_columns[img_idx]
                        img_path = images_for_question[img_idx]
                        
                        try:
                            # Get original image dimensions using PIL
                            pil_img = PILImage.open(img_path)
                            orig_width, orig_height = pil_img.size
                            
                            # Calculate aspect ratio
                            aspect_ratio = orig_width / orig_height if orig_height > 0 else 1
                            
                            # Set height to 18 pixels and calculate width based on aspect ratio
                            target_height = 18
                            target_width = target_height * aspect_ratio
                            
                            # Create image object
                            img = ExcelImage(img_path)
                            
                            # Set image dimensions
                            img.height = target_height
                            img.width = target_width
                            
                            # Insert image at specified cell
                            cell_ref = f"{col_letter}{row_num}"
                            ws.add_image(img, cell_ref)
                            print(f"    ✅ Inserted image {img_idx + 1} at {cell_ref} (size: {target_width:.1f}x{target_height}px, original: {orig_width}x{orig_height}px)")
                            
                        except Exception as e:
                            print(f"    ⚠️ Error inserting image at {col_letter}{row_num}: {e}")
        
        print(f"  ✅ Image insertion complete")
        return True
        
    except Exception as e:
        print(f"  ❌ Error in insert_images_into_worksheet: {e}")
        import traceback
        traceback.print_exc()
        return False

def copy_worksheet_content(source_ws, target_ws, skip_first_row=False):
    """Copy content and formatting from source to target worksheet"""
    try:
        start_row = 2 if skip_first_row else 1
        
        for row in source_ws.iter_rows(min_row=start_row):
            target_row = []
            for cell in row:
                target_row.append(cell.value)
            
            # Calculate target row number
            if skip_first_row:
                target_row_num = target_ws.max_row + 1
            else:
                target_row_num = cell.row
            
            # Write values
            for col_idx, value in enumerate(target_row, 1):
                target_cell = target_ws.cell(row=target_row_num, column=col_idx)
                source_cell = source_ws.cell(row=cell.row, column=col_idx)
                
                target_cell.value = value
                
                # Copy formatting
                if source_cell.font:
                    target_cell.font = Font(
                        name=source_cell.font.name,
                        size=source_cell.font.size,
                        bold=source_cell.font.bold,
                        italic=source_cell.font.italic,
                        color=source_cell.font.color
                    )
                
                if source_cell.fill:
                    target_cell.fill = PatternFill(
                        start_color=source_cell.fill.start_color,
                        end_color=source_cell.fill.end_color,
                        fill_type=source_cell.fill.fill_type
                    )
                
                if source_cell.alignment:
                    target_cell.alignment = Alignment(
                        horizontal=source_cell.alignment.horizontal,
                        vertical=source_cell.alignment.vertical,
                        wrap_text=source_cell.alignment.wrap_text
                    )
                
                if source_cell.border:
                    target_cell.border = Border(
                        left=source_cell.border.left,
                        right=source_cell.border.right,
                        top=source_cell.border.top,
                        bottom=source_cell.border.bottom
                    )
        
        return True
    except Exception as e:
        print(f"❌ Error copying worksheet content: {e}")
        return False

@cyber_security_audit_excel_bp.route('/create_cyber_security_audit_excel', methods=['POST'])
def create_cyber_security_audit_excel():
    """Create consolidated Cyber Security Audit Excel file"""
    try:
        print("\n" + "="*80)
        print("🚀 Creating Cyber Security Audit Excel")
        print("="*80)
        
        # Get uploaded files
        meity_parts_zip = request.files.get('meity_parts_zip')
        images_zip = request.files.get('images_zip')
        
        if not meity_parts_zip or not images_zip:
            return jsonify({'success': False, 'error': 'Missing required files'}), 400
        
        # Create temporary directory
        temp_dir = tempfile.mkdtemp()
        print(f"📁 Created temp directory: {temp_dir}")
        
        try:
            # Save uploaded ZIP files
            meity_zip_path = os.path.join(temp_dir, 'meity_parts.zip')
            images_zip_path = os.path.join(temp_dir, 'images.zip')
            
            meity_parts_zip.save(meity_zip_path)
            images_zip.save(images_zip_path)
            print(f"💾 Saved uploaded ZIP files")
            
            # Extract Meity Excel files
            meity_extract_dir = os.path.join(temp_dir, 'meity_parts')
            os.makedirs(meity_extract_dir, exist_ok=True)
            
            with zipfile.ZipFile(meity_zip_path, 'r') as zip_ref:
                zip_ref.extractall(meity_extract_dir)
            
            print(f"📂 Extracted Meity parts ZIP")
            
            # Find all .xlsx files
            excel_files = []
            for root, dirs, files in os.walk(meity_extract_dir):
                for file in files:
                    if file.endswith('.xlsx') and not file.startswith('~'):
                        excel_files.append(os.path.join(root, file))
            
            print(f"📊 Found {len(excel_files)} Excel files:")
            for f in excel_files:
                print(f"  - {os.path.basename(f)}")
            
            # Sort files by part number
            sorted_files = sort_meity_files(excel_files)
            print(f"\n🔢 Sorted files:")
            for idx, f in enumerate(sorted_files, 1):
                print(f"  {idx}. {os.path.basename(f)}")
            
            # Create new workbook
            new_wb = openpyxl.Workbook()
            new_wb.remove(new_wb.active)  # Remove default sheet
            new_ws = new_wb.create_sheet(title="Meity Audit")
            
            print(f"\n📝 Creating consolidated worksheet...")
            
            # Set column widths
            new_ws.column_dimensions['A'].width = 15
            new_ws.column_dimensions['B'].width = 60
            new_ws.column_dimensions['C'].width = 25
            new_ws.column_dimensions['D'].width = 60
            new_ws.column_dimensions['E'].width = 25
            for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
                new_ws.column_dimensions[col].width = 15
            
            print(f"  ✅ Set column widths")
            
            # Copy first file completely
            if len(sorted_files) > 0:
                print(f"\n📋 Copying first file (with header)...")
                first_wb = openpyxl.load_workbook(sorted_files[0])
                first_ws = first_wb.active
                
                copy_worksheet_content(first_ws, new_ws, skip_first_row=False)
                print(f"  ✅ Copied {first_ws.max_row} rows from first file")
                first_wb.close()
            
            # Copy remaining files (skip header)
            for file_idx, file_path in enumerate(sorted_files[1:], 2):
                print(f"\n📋 Copying file {file_idx} (skipping header)...")
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                
                copy_worksheet_content(ws, new_ws, skip_first_row=True)
                print(f"  ✅ Copied {ws.max_row - 1} rows (skipped header)")
                wb.close()
            
            print(f"\n  📊 Total rows in new worksheet: {new_ws.max_row}")
            
            # Add POC columns
            print(f"\n🎨 Adding POC columns...")
            
            # E1 = "POC Attached"
            cell_e1 = new_ws.cell(row=1, column=5)
            cell_e1.value = "POC Attached"
            cell_e1.font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
            cell_e1.fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
            cell_e1.alignment = Alignment(horizontal='center', vertical='center')
            
            # Merge F1:R1 and add "POC"
            new_ws.merge_cells('F1:R1')
            cell_f1 = new_ws.cell(row=1, column=6)
            cell_f1.value = "POC"
            cell_f1.font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
            cell_f1.fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
            cell_f1.alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"  ✅ Added POC headers")
            
            # Add borders to all columns
            print(f"\n🔲 Adding borders to all columns...")
            
            # Define border styles
            full_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            left_border = Border(
                left=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            middle_border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            right_border = Border(
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            for row_num in range(1, new_ws.max_row + 1):
                # A, B, C, D columns: all side border
                for col in range(1, 5):  # A=1, B=2, C=3, D=4
                    new_ws.cell(row=row_num, column=col).border = full_border
                
                # E column: all side border
                new_ws.cell(row=row_num, column=5).border = full_border
                
                # F column: left, upper, bottom
                new_ws.cell(row=row_num, column=6).border = left_border
                
                # G-Q columns: upper, bottom only
                for col in range(7, 18):  # G=7 to Q=17 (range goes up to 17, so includes 17)
                    new_ws.cell(row=row_num, column=col).border = middle_border
                
                # R column: upper, bottom, right
                new_ws.cell(row=row_num, column=18).border = right_border
            
            print(f"  ✅ Applied borders to all columns")
            
            # Increase row heights
            print(f"\n📏 Increasing row heights...")
            for row_num in range(1, new_ws.max_row + 1):
                current_height = new_ws.row_dimensions[row_num].height
                if current_height is None:
                    current_height = 15  # Default Excel row height
                new_ws.row_dimensions[row_num].height = current_height + 15
            
            print(f"  ✅ Increased height for {new_ws.max_row} rows")
            
            # Extract and insert images
            print(f"\n🖼️ Extracting images from ZIP...")
            images_extract_dir = os.path.join(temp_dir, 'images')
            os.makedirs(images_extract_dir, exist_ok=True)
            
            image_mapping = extract_images_from_zip(images_zip_path, images_extract_dir)
            print(f"  📷 Found images for {len(image_mapping)} questions")
            
            if image_mapping:
                insert_images_into_worksheet(new_ws, image_mapping, temp_dir)
            
            # Save consolidated file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f'Cyber_Security_Audit_{timestamp}.xlsx'
            output_path = os.path.join('static', 'uploads', output_filename)
            
            # Ensure uploads directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            new_wb.save(output_path)
            print(f"\n💾 Saved: {output_filename}")
            print("="*80)
            
            return jsonify({
                'success': True,
                'excel_file': 'Cyber_Security_Audit.xlsx',
                'download_url': f'/static/uploads/{output_filename}'
            })
        
        finally:
            # Cleanup temp directory
            try:
                shutil.rmtree(temp_dir)
                print(f"🗑️ Cleaned up temp directory")
            except:
                pass
    
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

