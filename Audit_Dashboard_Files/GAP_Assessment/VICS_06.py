from flask import Blueprint, request, jsonify, send_file
import json
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import tempfile
import glob

# Create blueprint for VICS Part 6
vics_part6_bp = Blueprint('vics_part6', __name__)

def cleanup_old_vics_files():
    """Clean up old VICS Excel files from uploads directory"""
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            vics_pattern = os.path.join(upload_dir, 'VICS_Part*.xlsx')
            old_files = glob.glob(vics_pattern)
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

def create_vics_part6_excel(vics_data):
    """
    Create Excel file for VICS Part 6 with formatted data
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "VICS Part 6"
    
    # Set column widths
    column_widths = {
        'A': 20,
        'B': 60,
        'C': 20,
        'D': 20,
        'E': 20,
        'F': 60
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Merge A1 and B1
    ws.merge_cells('A1:B1')
    
    # Set header values
    headers = {
        'A1': 'C) Vendor Management',
        'C1': 'Max Marks',
        'D1': 'Yes/No',
        'E1': 'Marks given by the Auditor',
        'F1': "Auditor's Observation"
    }
    
    # Header styling
    header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')  # Dark Blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply headers and styling
    for cell_ref, value in headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Populate data rows
    current_row = 2
    section_number = 17  # Part 6 starts from section 17
    
    # Track rows that need bold formatting
    bold_rows = []
    
    # Track section total rows for grand total calculation
    section_total_rows = []
    
    # Map section keys to their display numbers
    section_mapping = {
        "vendor_onboarding": "17",
        "oversight_and_review_of_vendors": "18",
        "addressing_security_in_sla": "19",
        "compliance_adherence": "20",
        "dependency_on_vendor_staffs": "21"
    }
    
    # Observations for each question
    observations = {
        "17.1": {
            "Implemented": "The bank has a Board-approved Vendor Management Policy in place.",
            "Not Implemented": "The bank does not have an approved Vendor Management Policy."
        },
        "17.2": {
            "Implemented": "The bank maintains an updated list of all IT vendors.",
            "Not Implemented": "The bank has not maintained an updated vendor list."
        },
        "17.3": {
            "Implemented": "SLAs are reviewed and updated periodically as needed.",
            "Not Implemented": "There is no defined mechanism for reviewing or updating SLAs."
        },
        "17.4": {
            "Implemented": "The bank conducts due diligence on vendors before onboarding.",
            "Not Implemented": "Vendor due diligence is not consistently performed before onboarding."
        },
        "17.5": {
            "Implemented": "SLAs/contracts have been signed with all vendors.",
            "Not Implemented": "SLAs/contracts have not been signed with all vendors."
        },
        "17.6": {
            "Implemented": "The bank follows RBI guidelines on sharing IT resources.",
            "Not Implemented": "RBI guidelines on sharing IT resources are not implemented."
        },
        "18.1": {
            "Implemented": "Regular meetings with vendors are conducted to discuss cybersecurity matters.",
            "Not Implemented": "Regular meetings with vendors are not conducted."
        },
        "18.2": {
            "Implemented": "The bank conducts periodic reviews of vendor cybersecurity compliance.",
            "Not Implemented": "Vendor cybersecurity compliance is not reviewed periodically."
        },
        "19.1": {
            "Implemented": "SLAs include clauses to incorporate updated security requirements.",
            "Not Implemented": "SLAs do not include clauses for updating security requirements."
        },
        "19.2": {
            "Implemented": "NDAs are signed with all vendors.",
            "Not Implemented": "NDAs are not signed with vendors."
        },
        "19.3": {
            "Implemented": "SLAs define roles and responsibilities during incidents.",
            "Not Implemented": "Roles and responsibilities are not defined in SLAs."
        },
        "19.4": {
            "Implemented": "SLAs include clauses for penalties and legal action in case of breach.",
            "Not Implemented": "SLAs do not include clauses for penalties or legal action."
        },
        "19.5": {
            "Implemented": "SLAs define timelines for implementation and uptime.",
            "Not Implemented": "SLAs do not define timelines for implementation or uptime."
        },
        "19.6": {
            "Implemented": "SLAs include clauses recognizing inspection rights of RBI/NABARD.",
            "Not Implemented": "SLAs do not include clauses for inspection rights."
        },
        "19.7": {
            "Implemented": "SLAs include clauses for addressing customer grievances.",
            "Not Implemented": "SLAs do not include clauses for customer grievance handling."
        },
        "20.1": {
            "Implemented": "SLAs include clauses requiring vendors to comply with applicable regulations.",
            "Not Implemented": "SLAs do not include clauses for regulatory compliance."
        },
        "20.2": {
            "Implemented": "SLAs include the bank's right to audit vendors.",
            "Not Implemented": "SLAs do not include the right to audit."
        },
        "20.3": {
            "Implemented": "SLAs are reviewed by the legal department.",
            "Not Implemented": "SLAs are not reviewed by the legal department."
        },
        "21.1": {
            "Implemented": "Access to financial transactions is restricted to bank officials.",
            "Not Implemented": "Vendor staff have access to financial transaction systems."
        },
        "21.2": {
            "Implemented": "SLAs include clauses to ensure vendor support during transitions.",
            "Not Implemented": "SLAs do not include provisions for vendor support during transitions."
        }
    }
    
    # Helper function to format date as DD/MM/YYYY
    def format_date(date_str):
        if not date_str:
            return date_str
        try:
            # Parse date from YYYY-MM-DD format
            from datetime import datetime
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            return date_obj.strftime('%d/%m/%Y')
        except:
            return date_str
    
    for section_key, section_data in vics_data["sections"].items():
        section_title = section_data["title"]
        section_num = section_mapping.get(section_key, str(section_number))
        
        # Add section header row
        section_header_row = current_row
        bold_rows.append(section_header_row)  # Track for bold formatting
        
        ws.cell(row=current_row, column=1).value = section_num
        ws.cell(row=current_row, column=2).value = section_title
        ws.cell(row=current_row, column=3).value = ""
        ws.cell(row=current_row, column=4).value = ""
        ws.cell(row=current_row, column=5).value = ""
        ws.cell(row=current_row, column=6).value = ""
        current_row += 1
        
        # Track section totals
        section_total_marks = 0
        section_marks_given = 0
        
        for q_num, q_data in section_data["questions"].items():
            answer = q_data.get("answer", "")
            
            # Main question row
            ws.cell(row=current_row, column=1).value = q_num  # Column A - Question number
            ws.cell(row=current_row, column=2).value = q_data["question"]  # Column B - Question
            ws.cell(row=current_row, column=3).value = q_data["marks"]  # Column C - Max Marks
            
            # Column D - Yes/No
            yes_no = "Yes" if answer == "Implemented" else "No" if answer == "Not Implemented" else ""
            ws.cell(row=current_row, column=4).value = yes_no
            
            # Calculate marks given (full marks if Implemented, 0 if Not Implemented)
            marks_given = q_data["marks"] if answer == "Implemented" else 0
            ws.cell(row=current_row, column=5).value = marks_given  # Column E - Marks given
            
            # Column F - Observation
            observation = observations.get(q_num, {}).get(answer, "")
            ws.cell(row=current_row, column=6).value = observation
            
            # Add to section totals
            if isinstance(q_data["marks"], (int, float)):
                section_total_marks += q_data["marks"]
                section_marks_given += marks_given
            
            current_row += 1
            
            # Sub-questions
            if "sub_questions" in q_data:
                parent_answer = answer
                
                for sub_num, sub_data in q_data["sub_questions"].items():
                    ws.cell(row=current_row, column=1).value = sub_num
                    ws.cell(row=current_row, column=2).value = sub_data['question']
                    ws.cell(row=current_row, column=3).value = ""
                    ws.cell(row=current_row, column=4).value = ""
                    ws.cell(row=current_row, column=5).value = ""
                    
                    # Handle sub-question observations
                    sub_observation = ""
                    sub_answer = sub_data.get('answer', '')
                    
                    if parent_answer == "Not Implemented":
                        # Special cases for Not Implemented
                        if sub_num in ["17.1.1", "17.2.1", "18.1.1", "18.2.1"]:
                            sub_observation = "-"
                        elif sub_num == "17.5.1":
                            if sub_answer:
                                sub_observation = f"The bank does not maintains SLAs with {sub_answer} vendors."
                            else:
                                sub_observation = "-"
                        elif sub_num == "21.1.1":
                            if sub_answer:
                                sub_observation = f"No. of vendor staff having access to do financial transactions: {sub_answer}"
                            else:
                                sub_observation = "-"
                        else:
                            sub_observation = "-"
                    elif parent_answer == "Implemented":
                        # Special handling for Implemented
                        if sub_num == "17.1.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of Approval: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "17.2.1":
                            if sub_answer:
                                sub_observation = f"Number of Vendors: {sub_answer}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "17.5.1":
                            sub_observation = "The bank maintains SLAs with all vendors."
                        elif sub_num == "18.1.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last metting: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "18.2.1":
                            if sub_answer:
                                formatted_date = format_date(sub_answer)
                                sub_observation = f"Date of last review: {formatted_date}"
                            else:
                                sub_observation = "-"
                        elif sub_num == "21.1.1":
                            sub_observation = "No. of vendor staff having access to do financial transactions: 0"
                        else:
                            sub_observation = "-"
                    
                    ws.cell(row=current_row, column=6).value = sub_observation
                    
                    current_row += 1
        
        # Add section total row
        total_row = current_row
        bold_rows.append(total_row)  # Track for bold formatting
        section_total_rows.append(total_row)  # Track for grand total
        
        ws.cell(row=current_row, column=1).value = ""
        ws.cell(row=current_row, column=2).value = f"Total_C.{section_num}"
        ws.cell(row=current_row, column=3).value = section_total_marks
        ws.cell(row=current_row, column=4).value = ""
        ws.cell(row=current_row, column=5).value = section_marks_given
        ws.cell(row=current_row, column=6).value = ""
        
        current_row += 1
        section_number += 1
    
    # Add grand total row
    grand_total_row = current_row
    bold_rows.append(grand_total_row)
    
    # Calculate grand total from section totals
    grand_total_marks_given = sum([ws.cell(row=row, column=5).value for row in section_total_rows])
    
    ws.cell(row=current_row, column=1).value = ""
    ws.cell(row=current_row, column=2).value = "Marks in the category C"
    ws.cell(row=current_row, column=3).value = "30"
    ws.cell(row=current_row, column=4).value = ""
    ws.cell(row=current_row, column=5).value = grand_total_marks_given
    ws.cell(row=current_row, column=6).value = ""
    
    current_row += 1
    
    # Apply borders and alignment to all data cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for row in range(1, current_row):
        for col in range(1, 7):  # Only A-F columns
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            # Header row (row 1) should always be center-aligned
            if row == 1:
                cell.alignment = center_alignment
            # Column B (questions) and F (observations) should be left-aligned for data rows
            elif col == 2 or col == 6:  # Column B and F
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment
            
            # Set font for data rows
            if row > 1:
                # Apply bold formatting to section headers and totals
                if row in bold_rows:
                    cell.font = Font(name='Times New Roman', size=12, bold=True)
                else:
                    cell.font = Font(name='Times New Roman', size=12)
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name

@vics_part6_bp.route('/process_vics_part6', methods=['POST'])
def process_vics_part6():
    """
    Process VICS Part 6 form data and return JSON response
    """
    try:
        print("\n" + "="*80)
        print("🎯 VICS PART 6 - Vendor On-boarding | Oversight and Review of Vendors | Addressing security in SLA | Compliance Adherence with legal and regulatory compliances | Dependency on Vendor staffs")
        print("="*80)
        
        # Get form data
        form_data = request.form.to_dict()
        
        # Print all form data for debugging
        print("📋 Form Data Received:")
        for key, value in form_data.items():
            print(f"  {key}: {value}")
        
        # Organize data by sections
        vics_data = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "part": "VICS Part 6: Vendor On-boarding | Oversight and Review of Vendors | Addressing security in SLA | Compliance Adherence with legal and regulatory compliances | Dependency on Vendor staffs",
            "sections": {
                "vendor_onboarding": {
                    "title": "Vendor On-boarding",
                    "questions": {
                        "17.1": {
                            "question": "Whether bank has outsourcing/ vendor management policy?",
                            "marks": 1,
                            "answer": form_data.get('q17_1', ''),
                            "sub_questions": {
                                "17.1.1": {
                                    "question": "Provide date of approval.",
                                    "marks": "-",
                                    "answer": form_data.get('q17_1_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "17.2": {
                            "question": "Whether a list of third party IT vendors is maintained?",
                            "marks": 1,
                            "answer": form_data.get('q17_2', ''),
                            "sub_questions": {
                                "17.2.1": {
                                    "question": "Provide no. of vendors.",
                                    "marks": "-",
                                    "answer": form_data.get('q17_2_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "17.3": {
                            "question": "Whether there is a provision to review or update Service Level Agreements with vendor(s), whenever required?",
                            "marks": 1,
                            "answer": form_data.get('q17_3', '')
                        },
                        "17.4": {
                            "question": "Whether bank does the due diligence of vendor before on-boarding?",
                            "marks": 1,
                            "answer": form_data.get('q17_4', '')
                        },
                        "17.5": {
                            "question": "Whether contracts/ SLA have been signed between the bank and all the vendors?",
                            "marks": 2,
                            "answer": form_data.get('q17_5', ''),
                            "sub_questions": {
                                "17.5.1": {
                                    "question": "Provide no. of vendors SLA/contracts have not made with?",
                                    "marks": "-",
                                    "answer": form_data.get('q17_5_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "17.6": {
                            "question": "Whether the guidelines on sharing of Information Technology resources issued by RBI vide circular RBI/2013-14/216 dated 30 August 2013 are being followed?",
                            "marks": 2,
                            "answer": form_data.get('q17_6', '')
                        }
                    }
                },
                "oversight_and_review_of_vendors": {
                    "title": "Oversight and Review of Vendors",
                    "questions": {
                        "18.1": {
                            "question": "Whether regular meetings are taking place between vendors and bank to discuss cyber security related issues and developments?",
                            "marks": 1.5,
                            "answer": form_data.get('q18_1', ''),
                            "sub_questions": {
                                "18.1.1": {
                                    "question": "Provide date of last such meeting.",
                                    "marks": "-",
                                    "answer": form_data.get('q18_1_1', ''),
                                    "type": "date"
                                }
                            }
                        },
                        "18.2": {
                            "question": "Whether bank reviews cyber security compliance (including action on advisory and alerts) from the vendor periodically?",
                            "marks": 2,
                            "answer": form_data.get('q18_2', ''),
                            "sub_questions": {
                                "18.2.1": {
                                    "question": "Provide date of last review.",
                                    "marks": "-",
                                    "answer": form_data.get('q18_2_1', ''),
                                    "type": "date"
                                }
                            }
                        }
                    }
                },
                "addressing_security_in_sla": {
                    "title": "Addressing security in SLA",
                    "questions": {
                        "19.1": {
                            "question": "Whether SLA has provisions for updating latest security requirements?",
                            "marks": 2,
                            "answer": form_data.get('q19_1', '')
                        },
                        "19.2": {
                            "question": "Whether (Non Disclosure Agreement) NDA is signed with vendor as part of SLA or separately?",
                            "marks": 1,
                            "answer": form_data.get('q19_2', '')
                        },
                        "19.3": {
                            "question": "Whether roles and responsibilities for vendor and bank are defined in SLA in case of incident?",
                            "marks": 1.5,
                            "answer": form_data.get('q19_3', '')
                        },
                        "19.4": {
                            "question": "Whether SLA has provision for penalties and legal action on vendor in case of breach of contract?",
                            "marks": 2,
                            "answer": form_data.get('q19_4', '')
                        },
                        "19.5": {
                            "question": "Whether time-frame has been stipulated for implementing tools/applications/projects and ensuring uptime?",
                            "marks": 2,
                            "answer": form_data.get('q19_5', '')
                        },
                        "19.6": {
                            "question": "Do the outsourcing agreements include clauses to recognise the right of RBI / NABARD to inspect banks documents, records, transactions, logs processed by the service provider?",
                            "marks": 2,
                            "answer": form_data.get('q19_6', '')
                        },
                        "19.7": {
                            "question": "Whether there is a clause in SLA with vendors to address Grievances of customers including system/ technology issues?",
                            "marks": 1,
                            "answer": form_data.get('q19_7', '')
                        }
                    }
                },
                "compliance_adherence": {
                    "title": "Compliance Adherence with legal and regulatory compliances",
                    "questions": {
                        "20.1": {
                            "question": "Whether clause related to ensure regulatory compliances by the vendors are included in SLA?",
                            "marks": 1,
                            "answer": form_data.get('q20_1', '')
                        },
                        "20.2": {
                            "question": "Whether the right to audit by the bank or by its authorized auditor has been included in the SLA?",
                            "marks": 2,
                            "answer": form_data.get('q20_2', '')
                        },
                        "20.3": {
                            "question": "Whether the SLA was checked by legal department or law officer for compliance?",
                            "marks": 1,
                            "answer": form_data.get('q20_3', '')
                        }
                    }
                },
                "dependency_on_vendor_staffs": {
                    "title": "Dependency on Vendor staffs",
                    "questions": {
                        "21.1": {
                            "question": "Whether system access for doing financial transaction is restricted to bank officials only?",
                            "marks": 2,
                            "answer": form_data.get('q21_1', ''),
                            "sub_questions": {
                                "21.1.1": {
                                    "question": "Provide no. of vendor staff having access to do financial transactions.",
                                    "marks": "-",
                                    "answer": form_data.get('q21_1_1', ''),
                                    "type": "number"
                                }
                            }
                        },
                        "21.2": {
                            "question": "Whether provision for providing support for smooth transition/migration / upgradation from one system to another without hampering business continuity is included in SLA?",
                            "marks": 1,
                            "answer": form_data.get('q21_2', '')
                        }
                    }
                }
            }
        }
        
        # Print organized data
        print("\n" + "="*80)
        print("📊 ORGANIZED DATA:")
        print("="*80)
        print(json.dumps(vics_data, indent=2))
        
        # Calculate total marks
        total_marks = 0
        for section_key, section_data in vics_data["sections"].items():
            print(f"\n📌 {section_data['title']}:")
            for q_num, q_data in section_data["questions"].items():
                marks = q_data.get('marks', 0)
                if marks != '-' and marks != 0:
                    total_marks += marks
                answer = q_data.get('answer', 'Not answered')
                print(f"  {q_num}. {q_data['question']}")
                print(f"      Answer: {answer} | Marks: {marks}")
                
                # Print sub-questions if any
                if 'sub_questions' in q_data:
                    for sub_num, sub_data in q_data['sub_questions'].items():
                        sub_answer = sub_data.get('answer', 'Not answered')
                        print(f"    {sub_num}. {sub_data['question']}")
                        print(f"        Answer: {sub_answer} | Type: {sub_data.get('type', 'text')}")
        
        print(f"\n💯 Total Marks for Part 6: {total_marks}")
        print("="*80 + "\n")
        
        # Clean up old VICS files before generating new one
        print("🧹 Cleaning up old VICS files...")
        cleanup_old_vics_files()
        
        # Generate Excel file
        print("📝 Generating Excel file...")
        excel_file_path = create_vics_part6_excel(vics_data)
        print(f"✅ Excel file created: {excel_file_path}")
        
        # Save the generated file to static/uploads directory
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"VICS_Part6_{timestamp}.xlsx"
        static_upload_path = os.path.join('static', 'uploads', excel_filename)
        
        # Ensure the uploads directory exists
        os.makedirs(os.path.dirname(static_upload_path), exist_ok=True)
        
        # Copy the file
        import shutil
        shutil.copy2(excel_file_path, static_upload_path)
        
        # Clean up temp file
        os.unlink(excel_file_path)
        
        # Return JSON response with download URL
        return jsonify({
            "success": True,
            "message": "VICS Part 6 data processed successfully",
            "download_url": f"/static/uploads/{excel_filename}",
            "excel_file": excel_filename
        }), 200
        
    except Exception as e:
        print(f"❌ Error processing VICS Part 6: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"Error processing VICS Part 6: {str(e)}"
        }), 500

