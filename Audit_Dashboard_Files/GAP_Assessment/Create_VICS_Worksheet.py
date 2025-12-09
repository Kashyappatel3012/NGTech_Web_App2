from flask import Blueprint, request, jsonify
import zipfile
import os
import tempfile
import shutil
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
import re
import glob

# Create blueprint for Create VICS Worksheet
create_vics_worksheet_bp = Blueprint('create_vics_worksheet', __name__)

def cleanup_vics_consolidated_files():
    """
    Clean up old VICS Consolidated Excel files from uploads directory
    """
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            vics_pattern = os.path.join(upload_dir, 'VICS_Consolidated_Worksheet*.xlsx')
            old_files = glob.glob(vics_pattern)
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

@create_vics_worksheet_bp.route('/cleanup_vics_consolidated', methods=['POST'])
def cleanup_vics_consolidated_endpoint():
    """
    Endpoint to cleanup VICS Consolidated files after download
    """
    try:
        cleanup_vics_consolidated_files()
        return jsonify({"success": True, "message": "Cleanup completed"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

def sort_vics_files(file_list):
    """
    Sort VICS files in the order: Part_1, Part_2, Part_3, Part_4, Part_5, Part_6, Part_7
    Handles filenames with or without underscores (Part_1 or Part1)
    """
    def get_part_number(filename):
        # Extract part number from filename (handles Part_1, Part1, Part_2, Part2, etc.)
        match = re.search(r'Part[_\s]*(\d+)', filename, re.IGNORECASE)
        if match:
            return int(match.group(1))
        return 999  # Put unmatched files at the end
    
    return sorted(file_list, key=get_part_number)

def parse_image_name(filename):
    """
    Parse image filename to extract question number (e.g., '3_3' from '3_3-something.jpg')
    Returns the question number in format 'X.Y' or None if not parseable
    """
    # Remove file extension
    base_name = os.path.splitext(filename)[0]
    
    # Extract first two parts separated by underscore
    # Pattern: digits_digits (possibly followed by anything)
    match = re.match(r'(\d+)_(\d+)', base_name)
    if match:
        first_part = match.group(1)
        second_part = match.group(2)
        return f"{first_part}.{second_part}"
    
    return None

def extract_images_from_zip(zip_path, extract_dir):
    """
    Extract images from ZIP file and organize by question number
    Returns dict: {'3.3': [path1, path2, ...], '4.1': [path1], ...}
    """
    image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif'}
    image_mapping = {}
    
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
            print(f"📂 Extracted ZIP file to: {extract_dir}")
            
            # Find all images
            for root, dirs, files in os.walk(extract_dir):
                for file in files:
                    file_ext = os.path.splitext(file)[1].lower()
                    if file_ext in image_extensions:
                        image_path = os.path.join(root, file)
                        question_num = parse_image_name(file)
                        
                        if question_num:
                            if question_num not in image_mapping:
                                image_mapping[question_num] = []
                            image_mapping[question_num].append(image_path)
                            print(f"  📷 Found image: {file} → Question {question_num}")
    
    except Exception as e:
        print(f"⚠️ Error extracting images: {e}")
    
    return image_mapping

def insert_images_into_worksheet(ws, image_mapping, temp_dir):
    """
    Insert images into worksheet based on question number matching
    Images are inserted in columns H, I, J, K, L, M
    Returns list of rows that have images (for Annexure numbering)
    """
    rows_with_images = []
    
    # Define image columns in order: I, J, K, L, M, H (H is last for overflow)
    image_columns = ['I', 'J', 'K', 'L', 'M', 'H']
    
    print(f"\n📸 Inserting images into worksheet...")
    
    # Scan rows 1-250 for question numbers in column A
    for row in range(1, 251):
        cell_value = ws[f'A{row}'].value
        
        if cell_value:
            # Convert to string and check if it matches question pattern (e.g., "3.3", "14.2")
            cell_str = str(cell_value).strip()
            
            # Only match cells with format "X.Y" (not "X.Y.Z")
            if re.match(r'^\d+\.\d+$', cell_str):
                # Check if we have images for this question
                if cell_str in image_mapping:
                    images = image_mapping[cell_str]
                    print(f"  📌 Row {row}: Question {cell_str} - {len(images)} image(s)")
                    
                    # Insert images (max 6 images: I, J, K, L, M, H)
                    for idx, image_path in enumerate(images[:6]):
                        if idx < len(image_columns):
                            col = image_columns[idx]
                            
                            try:
                                # Load image using openpyxl Image (preserves quality)
                                excel_img = ExcelImage(image_path)
                                
                                # Get original dimensions
                                original_width = excel_img.width
                                original_height = excel_img.height
                                
                                # Resize by 30x (only display size, not actual image)
                                excel_img.width = max(1, original_width // 30)
                                excel_img.height = max(1, original_height // 30)
                                
                                # Add image to worksheet
                                ws.add_image(excel_img, f'{col}{row}')
                                print(f"    ✅ Inserted image in {col}{row} (size: {excel_img.width}x{excel_img.height}, original: {original_width}x{original_height})")
                                
                            except Exception as e:
                                print(f"    ⚠️ Error inserting image in {col}{row}: {e}")
                    
                    # Track this row as having images
                    if cell_str not in [ws[f'A{r}'].value for r in rows_with_images]:
                        rows_with_images.append(row)
    
    return rows_with_images

def copy_worksheet_styles(source_sheet, target_sheet, start_row=1, skip_first_row=False):
    """
    Copy cell styles, column widths, and row heights from source to target worksheet
    """
    # Copy column widths
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        if col_letter in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width
    
    # Determine starting row
    source_start_row = 2 if skip_first_row else 1
    
    # Copy data and styles
    for row_idx, row in enumerate(source_sheet.iter_rows(min_row=source_start_row), start=start_row):
        # Copy row height
        if row[0].row in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row_idx].height = source_sheet.row_dimensions[row[0].row].height
        
        for cell in row:
            target_cell = target_sheet.cell(row=row_idx, column=cell.column)
            
            # Copy value
            target_cell.value = cell.value
            
            # Copy font
            if cell.font:
                target_cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color
                )
            
            # Copy fill
            if cell.fill:
                target_cell.fill = PatternFill(
                    start_color=cell.fill.start_color,
                    end_color=cell.fill.end_color,
                    fill_type=cell.fill.fill_type
                )
            
            # Copy alignment
            if cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=cell.alignment.wrap_text
                )
            
            # Copy border
            if cell.border:
                target_cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
    
    # Copy merged cells
    for merged_range in source_sheet.merged_cells.ranges:
        if skip_first_row and merged_range.min_row == 1:
            continue
        
        offset = start_row - (2 if skip_first_row else 1)
        new_min_row = merged_range.min_row + offset if skip_first_row else merged_range.min_row + offset
        new_max_row = merged_range.max_row + offset if skip_first_row else merged_range.max_row + offset
        
        target_sheet.merge_cells(
            start_row=new_min_row,
            start_column=merged_range.min_col,
            end_row=new_max_row,
            end_column=merged_range.max_col
        )
    
    return row_idx if 'row_idx' in locals() else start_row

@create_vics_worksheet_bp.route('/create_vics_worksheet', methods=['POST'])
def create_vics_worksheet():
    """
    Process uploaded ZIP files and create consolidated VICS worksheet
    """
    try:
        print("\n" + "="*80)
        print("🎯 CREATE VICS WORKSHEET")
        print("="*80)
        
        # Clean up old VICS consolidated files
        print("\n🧹 Cleaning up old VICS consolidated files...")
        cleanup_vics_consolidated_files()
        
        # Get uploaded files
        if 'vics_parts_zip' not in request.files or 'vics_parts_zip2' not in request.files:
            return jsonify({"success": False, "message": "Both ZIP files are required"}), 400
        
        zip_file1 = request.files['vics_parts_zip']
        zip_file2 = request.files['vics_parts_zip2']
        
        # Create temporary directory
        temp_dir = tempfile.mkdtemp()
        print(f"📁 Temporary directory created: {temp_dir}")
        
        try:
            # Extract both ZIP files
            excel_files = []
            
            for zip_file in [zip_file1, zip_file2]:
                zip_path = os.path.join(temp_dir, zip_file.filename)
                zip_file.save(zip_path)
                print(f"📦 Saved ZIP file: {zip_file.filename}")
                
                # Extract ZIP
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(temp_dir)
                    print(f"📂 Extracted ZIP file: {zip_file.filename}")
            
            # Find all Excel files
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    if file.endswith('.xlsx') and not file.startswith('~$'):
                        excel_files.append(os.path.join(root, file))
            
            print(f"\n📊 Found {len(excel_files)} Excel files")
            for f in excel_files:
                print(f"  - {os.path.basename(f)}")
            
            # Sort files
            excel_files = sort_vics_files(excel_files)
            print(f"\n✅ Files sorted in order:")
            for idx, f in enumerate(excel_files, 1):
                print(f"  {idx}. {os.path.basename(f)}")
            
            # Create new workbook
            consolidated_wb = openpyxl.Workbook()
            consolidated_ws = consolidated_wb.active
            consolidated_ws.title = "VICS"
            
            # Set column widths for POC columns
            consolidated_ws.column_dimensions['G'].width = 20
            consolidated_ws.column_dimensions['H'].width = 15
            consolidated_ws.column_dimensions['I'].width = 15
            consolidated_ws.column_dimensions['J'].width = 15
            consolidated_ws.column_dimensions['K'].width = 15
            consolidated_ws.column_dimensions['L'].width = 15
            consolidated_ws.column_dimensions['M'].width = 15
            
            # Add POC header columns (will be added to row 1 after first file is copied)
            # This will be set after the first worksheet is copied
            
            current_row = 1
            
            # Process first 4 files (Parts 1-4)
            print(f"\n📝 Processing Parts 1-4...")
            for idx, excel_file in enumerate(excel_files[:4]):
                print(f"  Processing: {os.path.basename(excel_file)}")
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                skip_first_row = (idx > 0)  # Skip header for files 2, 3, 4
                current_row = copy_worksheet_styles(ws, consolidated_ws, current_row, skip_first_row)
                current_row += 1  # Move to next row after each worksheet
                
                wb.close()
            
            # Add summary row for Category A (row 133)
            print(f"\n📊 Adding Category A summary at row 133...")
            consolidated_ws['B133'] = "Marks in the category A"
            consolidated_ws['C133'] = "74"
            
            # Calculate total for E133
            total_formula = "=E11+E21+E31+E44+E55+E73+E79+E95+E101+E113+E122+E132"
            consolidated_ws['E133'] = total_formula
            
            # Apply formatting to row 133
            for col in ['B', 'C', 'E']:
                cell = consolidated_ws[f'{col}133']
                cell.font = Font(name='Times New Roman', size=12, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Merge row 134 (A134:F134) and add empty content
            consolidated_ws.merge_cells('A134:F134')
            consolidated_ws['A134'] = ""
            consolidated_ws.row_dimensions[134].height = 15
            
            current_row = 135
            
            # Process remaining files (Parts 5-7)
            print(f"\n📝 Processing Parts 5-7...")
            for idx, excel_file in enumerate(excel_files[4:7]):
                print(f"  Processing: {os.path.basename(excel_file)}")
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                # Copy worksheet (DO NOT skip first row - include header for Parts 5-7)
                current_row = copy_worksheet_styles(ws, consolidated_ws, current_row, skip_first_row=False)
                current_row += 1
                
                # Add empty merged row between worksheets (except after the last one)
                if idx < 2:  # Between 5-6 and 6-7
                    consolidated_ws.merge_cells(f'A{current_row}:F{current_row}')
                    consolidated_ws[f'A{current_row}'] = ""
                    consolidated_ws.row_dimensions[current_row].height = 15
                    current_row += 1
                
                wb.close()
            
            # Add Grand Total at row 250
            print(f"\n📊 Adding Grand Total at row 250...")
            consolidated_ws['B250'] = "Grand Total"
            consolidated_ws['C250'] = "150"
            
            # Calculate grand total
            grand_total_formula = "=E133+E183+E222+E249"
            consolidated_ws['E250'] = grand_total_formula
            
            # Apply formatting to row 250
            for col in ['B', 'C', 'E']:
                cell = consolidated_ws[f'{col}250']
                cell.font = Font(name='Times New Roman', size=12, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Add summary section (rows 253-258)
            print(f"\n📊 Adding Summary Section at rows 253-258...")
            
            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Header row (253) - Dark blue background, white text, bold, center, wrap text
            header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
            header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            consolidated_ws['B253'] = "Section"
            consolidated_ws['D253'] = "Total Marks"
            consolidated_ws['E253'] = "Total Mark given by the Auditor"
            
            for col in ['B', 'D', 'E']:
                cell = consolidated_ws[f'{col}253']
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Data rows (254-257)
            data_font = Font(name='Times New Roman', size=12)
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Row 254
            consolidated_ws['B254'] = "A) Info Sec Processes & Controls"
            consolidated_ws['D254'] = "74"
            consolidated_ws['E254'] = f"=E133"  # Reference to E133
            
            # Row 255
            consolidated_ws['B255'] = "B) Governance and Policy"
            consolidated_ws['D255'] = "30"
            consolidated_ws['E255'] = f"=E183"  # Reference to E183
            
            # Row 256
            consolidated_ws['B256'] = "C) Vendor Management"
            consolidated_ws['D256'] = "30"
            consolidated_ws['E256'] = f"=E222"  # Reference to E222
            
            # Row 257
            consolidated_ws['B257'] = "D) Cyber Crisis Management"
            consolidated_ws['D257'] = "16"
            consolidated_ws['E257'] = f"=E249"  # Reference to E249
            
            # Total row (258)
            consolidated_ws['B258'] = "Total"
            consolidated_ws['D258'] = "150"
            consolidated_ws['E258'] = f"=E250"  # Reference to E250
            
            # Apply formatting to specific cells in summary section (253-258)
            # Row 254
            for col in ['B', 'D', 'E']:
                cell = consolidated_ws[f'{col}254']
                cell.font = Font(name='Times New Roman', size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Row 255
            for col in ['B', 'D', 'E']:
                cell = consolidated_ws[f'{col}255']
                cell.font = Font(name='Times New Roman', size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Row 256
            for col in ['B', 'D', 'E']:
                cell = consolidated_ws[f'{col}256']
                cell.font = Font(name='Times New Roman', size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Row 257
            for col in ['B', 'D', 'E']:
                cell = consolidated_ws[f'{col}257']
                cell.font = Font(name='Times New Roman', size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Row 258 (Total row - center aligned and bold for B258)
            consolidated_ws['B258'].font = Font(name='Times New Roman', size=12)
            consolidated_ws['B258'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            consolidated_ws['B258'].border = thin_border
            
            for col in ['D', 'E']:
                cell = consolidated_ws[f'{col}258']
                cell.font = Font(name='Times New Roman', size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Add POC header columns
            print(f"\n📋 Adding POC header columns...")
            consolidated_ws['G1'] = "POC Attached"
            consolidated_ws['G1'].font = header_font
            consolidated_ws['G1'].fill = header_fill
            consolidated_ws['G1'].alignment = header_alignment
            consolidated_ws['G1'].border = thin_border
            
            # Merge H1:M1 for "POC"
            consolidated_ws.merge_cells('H1:M1')
            consolidated_ws['H1'] = "POC"
            consolidated_ws['H1'].font = header_font
            consolidated_ws['H1'].fill = header_fill
            consolidated_ws['H1'].alignment = header_alignment
            consolidated_ws['H1'].border = thin_border
            
            # Extract images from second ZIP file
            print(f"\n📦 Extracting images from second ZIP file...")
            image_extract_dir = os.path.join(temp_dir, 'images')
            os.makedirs(image_extract_dir, exist_ok=True)
            
            zip2_path = os.path.join(temp_dir, zip_file2.filename)
            image_mapping = extract_images_from_zip(zip2_path, image_extract_dir)
            print(f"✅ Found {len(image_mapping)} question(s) with images")
            
            # Insert images and get rows with images
            rows_with_images = insert_images_into_worksheet(consolidated_ws, image_mapping, temp_dir)
            
            # Add Annexure numbering
            print(f"\n🔢 Adding Annexure numbers...")
            annexure_counter = 1
            for row in range(1, 251):
                if row in rows_with_images:
                    consolidated_ws[f'G{row}'] = f"Annexure {annexure_counter}"
                    consolidated_ws[f'G{row}'].font = Font(name='Times New Roman', size=12, color='FF0000')  # Red color
                    consolidated_ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    print(f"  📎 Row {row}: Annexure {annexure_counter}")
                    annexure_counter += 1
            
            # Increase row height for all rows by 15px
            print(f"\n📏 Increasing row heights by 15px...")
            for row in range(1, 251):
                current_height = consolidated_ws.row_dimensions[row].height or 15
                consolidated_ws.row_dimensions[row].height = current_height + 15
            
            # Apply borders and wrap text to all cells A-F up to row 250
            print(f"\n🔲 Applying borders and wrap text to all cells (A1:F250)...")
            for row in range(1, 251):
                for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
                    cell = consolidated_ws[f'{col_letter}{row}']
                    # Apply border
                    cell.border = thin_border
                    # Apply wrap text while preserving existing alignment
                    if cell.alignment:
                        cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            wrap_text=True
                        )
                    else:
                        cell.alignment = Alignment(wrap_text=True)
            
            # Add "POC Attached" to specific cells (G135, G185, G224)
            print(f"\n📋 Adding 'POC Attached' to G135, G185, G224...")
            for row_num in [135, 185, 224]:
                consolidated_ws[f'G{row_num}'] = "POC Attached"
                consolidated_ws[f'G{row_num}'].font = header_font
                consolidated_ws[f'G{row_num}'].fill = header_fill
                consolidated_ws[f'G{row_num}'].alignment = header_alignment
                consolidated_ws[f'G{row_num}'].border = thin_border
            
            # Apply special borders to POC columns (G-M, rows 1-250)
            print(f"\n🔲 Applying special borders to POC columns (G1:M250)...")
            
            # Define border styles
            border_full = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
            border_no_right = Border(left=Side(style='thin'), right=None, 
                                     top=Side(style='thin'), bottom=Side(style='thin'))
            border_no_left_right = Border(left=None, right=None, 
                                          top=Side(style='thin'), bottom=Side(style='thin'))
            border_no_left = Border(left=None, right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            for row in range(1, 251):
                # G column - full border (skip rows 135, 185, 224 as they already have borders)
                if row not in [135, 185, 224]:
                    consolidated_ws[f'G{row}'].border = border_full
                
                # H column - left, upper, bottom (no right)
                consolidated_ws[f'H{row}'].border = border_no_right
                
                # I, J, K, L columns - upper, bottom only (no left, no right)
                for col in ['I', 'J', 'K', 'L']:
                    consolidated_ws[f'{col}{row}'].border = border_no_left_right
                
                # M column - upper, bottom, right (no left)
                consolidated_ws[f'M{row}'].border = border_no_left
            
            # Save consolidated workbook
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"VICS_Consolidated_Worksheet_{timestamp}.xlsx"
            output_path = os.path.join('static', 'uploads', output_filename)
            
            # Ensure uploads directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            consolidated_wb.save(output_path)
            consolidated_wb.close()
            
            print(f"\n✅ Consolidated worksheet created: {output_filename}")
            print("="*80)
            
            return jsonify({
                "success": True,
                "message": "VICS Consolidated Worksheet created successfully",
                "download_url": f"/static/uploads/{output_filename}",
                "excel_file": "VICS.xlsx"  # Download as VICS.xlsx
            }), 200
            
        finally:
            # Cleanup temporary directory
            shutil.rmtree(temp_dir, ignore_errors=True)
            print(f"🗑️ Temporary directory cleaned up")
    
    except Exception as e:
        print(f"❌ Error creating VICS worksheet: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"Error creating VICS worksheet: {str(e)}"
        }), 500

