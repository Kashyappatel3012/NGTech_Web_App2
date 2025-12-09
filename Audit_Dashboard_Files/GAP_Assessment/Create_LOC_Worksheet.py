from flask import Blueprint, request, jsonify, send_file
import os
import zipfile
import tempfile
import shutil
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import re
import glob

# Create blueprint for Create LOC Worksheet
create_loc_worksheet_bp = Blueprint('create_loc_worksheet', __name__)

def cleanup_loc_consolidated_files():
    """Clean up old consolidated LOC files from uploads directory"""
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            loc_pattern = os.path.join(upload_dir, 'LOC_Consolidated_Worksheet_*.xlsx')
            old_files = glob.glob(loc_pattern)
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old consolidated LOC file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

@create_loc_worksheet_bp.route('/cleanup_loc_consolidated', methods=['POST'])
def cleanup_loc_consolidated_endpoint():
    """Endpoint to cleanup consolidated LOC files after download"""
    try:
        cleanup_loc_consolidated_files()
        return jsonify({"success": True, "message": "Cleanup completed"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

def sort_loc_files(file_list):
    """
    Sort LOC files by level number (2, 3, 4)
    Handles names like: LOC_Level2, LOCLevel3, LOC_Level4 (2), etc.
    """
    def extract_level(filename):
        # Try to find pattern like "Level2", "Level_2", "Level 2"
        match = re.search(r'Level[\s_]?(\d+)', filename, re.IGNORECASE)
        if match:
            return int(match.group(1))
        return 999  # Put unmatched files at the end
    
    return sorted(file_list, key=extract_level)

def parse_image_name(filename):
    """
    Extract the X_Y pattern from image filename
    Examples: 
    - 3_D.jpg -> 3_D
    - 3_D-1.png -> 3_D
    - 3_D8327429skjkjvsksjk.jpg -> 3_D
    - 2_B.jpeg -> 2_B
    """
    # Remove extension
    name_without_ext = os.path.splitext(filename)[0]
    
    # Match pattern: number_letter (optionally followed by anything)
    match = re.match(r'^(\d+_[A-Z])', name_without_ext, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return None

def extract_images_from_zip(zip_path, extract_dir):
    """
    Extract images from ZIP and create a mapping of question numbers to image paths
    Returns: dict like {'1_A': ['path1.jpg'], '3_D': ['path1.jpg', 'path2.jpg']}
    """
    image_mapping = {}
    
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        for file_info in zip_ref.namelist():
            # Check if it's an image file
            if file_info.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                # Extract the file
                zip_ref.extract(file_info, extract_dir)
                extracted_path = os.path.join(extract_dir, file_info)
                
                # Parse the image name
                filename = os.path.basename(file_info)
                question_id = parse_image_name(filename)
                
                if question_id:
                    if question_id not in image_mapping:
                        image_mapping[question_id] = []
                    image_mapping[question_id].append(extracted_path)
                    print(f"📷 Mapped image: {filename} -> {question_id}")
    
    return image_mapping

def find_row_for_question(ws, question_id):
    """
    Find the row number for a given question ID (e.g., '2_B')
    The question ID is in column A, and must be preceded by a number row
    
    Example:
    Row 1: 1
    Row 2: A  <- If looking for 1_A, this is the row
    Row 11: 2
    Row 12: A <- If looking for 2_A, this is the row
    Row 13: B <- If looking for 2_B, this is the row
    """
    try:
        number_part, letter_part = question_id.split('_')
        number_part = number_part.strip()
        letter_part = letter_part.strip().upper()
        
        current_section = None
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            
            if cell_value is None:
                continue
                
            cell_str = str(cell_value).strip()
            
            # Check if this row contains just a number (section header)
            if cell_str.isdigit():
                current_section = cell_str
            # Check if this row contains the letter we're looking for
            elif current_section == number_part and cell_str.upper() == letter_part:
                return row
        
        return None
    except Exception as e:
        print(f"⚠️ Error finding row for {question_id}: {e}")
        return None

def insert_images_into_worksheet(ws, image_mapping, temp_dir):
    """
    Insert images into the worksheet based on the mapping
    Images are inserted into columns F-K (6-11), up to 6 images per row
    """
    for question_id, image_paths in image_mapping.items():
        row_num = find_row_for_question(ws, question_id)
        
        if row_num is None:
            print(f"⚠️ Could not find row for question: {question_id}")
            continue
        
        print(f"✅ Inserting images for {question_id} at row {row_num}")
        
        # Limit to 6 images
        images_to_insert = image_paths[:6]
        
        # Column order: G(7), H(8), I(9), J(10), K(11), F(6)
        column_order = [7, 8, 9, 10, 11, 6]
        
        for idx, img_path in enumerate(images_to_insert):
            if idx >= 6:
                break
                
            col_num = column_order[idx]
            
            try:
                # Load and resize image
                img = Image(img_path)
                
                # Reduce size by 30x (divide by 30)
                original_width = img.width
                original_height = img.height
                img.width = original_width / 30
                img.height = original_height / 30
                
                # Position the image in the cell
                cell = ws.cell(row=row_num, column=col_num)
                img.anchor = cell.coordinate
                
                ws.add_image(img)
                print(f"  📌 Inserted image {idx+1} into {get_column_letter(col_num)}{row_num}")
                
            except Exception as e:
                print(f"  ⚠️ Error inserting image {img_path}: {e}")

def copy_worksheet_content(source_ws, target_ws, skip_first_row=False):
    """
    Copy content and formatting from source worksheet to target worksheet
    """
    from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
    
    start_row = 2 if skip_first_row else 1
    
    for row_idx, row in enumerate(source_ws.iter_rows(min_row=start_row), start=1):
        target_row = row_idx if skip_first_row else row_idx
        
        for cell in row:
            target_cell = target_ws.cell(row=target_row if not skip_first_row else target_row, 
                                        column=cell.column)
            
            # Copy value
            target_cell.value = cell.value
            
            # Copy formatting
            if cell.has_style:
                target_cell.font = cell.font.copy()
                target_cell.border = cell.border.copy()
                target_cell.fill = cell.fill.copy()
                target_cell.number_format = cell.number_format
                target_cell.protection = cell.protection.copy()
                target_cell.alignment = cell.alignment.copy()
    
    # Copy column dimensions
    for col in source_ws.column_dimensions:
        if col in target_ws.column_dimensions:
            target_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width
    
    # Copy row dimensions
    for row in source_ws.row_dimensions:
        if row in target_ws.row_dimensions:
            target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height
    
    # Copy merged cells
    for merged_cell in source_ws.merged_cells.ranges:
        if not skip_first_row:
            target_ws.merge_cells(str(merged_cell))
        else:
            # Adjust merged cell range if skipping first row
            min_row = merged_cell.min_row
            if min_row > 1:
                new_range = f"{get_column_letter(merged_cell.min_col)}{min_row-1}:{get_column_letter(merged_cell.max_col)}{merged_cell.max_row-1}"
                target_ws.merge_cells(new_range)

@create_loc_worksheet_bp.route('/create_loc_worksheet', methods=['POST'])
def create_loc_worksheet():
    """
    Process uploaded LOC ZIP files and create consolidated worksheet
    """
    try:
        print("\n" + "="*80)
        print("🚀 Creating LOC Consolidated Worksheet")
        print("="*80)
        
        # Check if files are uploaded
        if 'loc_parts_zip1' not in request.files:
            return jsonify({"error": "No LOC parts ZIP file 1 uploaded"}), 400
        
        if 'loc_parts_zip2' not in request.files:
            return jsonify({"error": "No LOC parts ZIP file 2 uploaded"}), 400
        
        loc_zip1 = request.files['loc_parts_zip1']
        loc_zip2 = request.files['loc_parts_zip2']
        
        if loc_zip1.filename == '' or loc_zip2.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        # Create temporary directory
        temp_dir = tempfile.mkdtemp()
        print(f"📁 Created temp directory: {temp_dir}")
        
        try:
            # Save uploaded ZIP files
            zip1_path = os.path.join(temp_dir, 'loc_parts1.zip')
            zip2_path = os.path.join(temp_dir, 'loc_parts2.zip')
            
            loc_zip1.save(zip1_path)
            loc_zip2.save(zip2_path)
            print(f"💾 Saved ZIP files")
            
            # Extract Excel files from first ZIP
            excel_extract_dir = os.path.join(temp_dir, 'excel_files')
            os.makedirs(excel_extract_dir, exist_ok=True)
            
            excel_files = []
            with zipfile.ZipFile(zip1_path, 'r') as zip_ref:
                for file_info in zip_ref.namelist():
                    if file_info.endswith('.xlsx') and not file_info.startswith('__MACOSX'):
                        zip_ref.extract(file_info, excel_extract_dir)
                        excel_files.append(os.path.join(excel_extract_dir, file_info))
                        print(f"📄 Extracted: {file_info}")
            
            if not excel_files:
                return jsonify({"error": "No Excel files found in ZIP"}), 400
            
            # Sort Excel files
            excel_files = sort_loc_files(excel_files)
            print(f"\n📊 Sorted Excel files:")
            for idx, file in enumerate(excel_files, 1):
                print(f"  {idx}. {os.path.basename(file)}")
            
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "LOC"
            
            # Copy first file completely
            print(f"\n📋 Copying first file (with headers)...")
            first_wb = load_workbook(excel_files[0])
            first_ws = first_wb.active
            
            copy_worksheet_content(first_ws, ws, skip_first_row=False)
            first_wb.close()
            
            current_row = ws.max_row + 1
            print(f"  ✅ First file copied. Current row: {current_row}")
            
            # Copy remaining files (skip first row)
            for excel_file in excel_files[1:]:
                print(f"\n📋 Copying {os.path.basename(excel_file)} (skipping header)...")
                wb_source = load_workbook(excel_file)
                ws_source = wb_source.active
                
                # Copy content starting from row 2
                for row in ws_source.iter_rows(min_row=2):
                    for cell in row:
                        target_cell = ws.cell(row=current_row, column=cell.column)
                        target_cell.value = cell.value
                        
                        # Copy formatting
                        if cell.has_style:
                            target_cell.font = cell.font.copy()
                            target_cell.border = cell.border.copy()
                            target_cell.fill = cell.fill.copy()
                            target_cell.number_format = cell.number_format
                            target_cell.alignment = cell.alignment.copy()
                    
                    current_row += 1
                
                wb_source.close()
                print(f"  ✅ Copied. Current row: {current_row}")
            
            # Add POC columns
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            print(f"\n📝 Adding POC columns...")
            
            # E1 = "POC Attached"
            ws['E1'] = "POC Attached"
            ws['E1'].font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
            ws['E1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws['E1'].fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
            
            # Merge F1:K1 and add "POC"
            ws.merge_cells('F1:K1')
            ws['F1'] = "POC"
            ws['F1'].font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
            ws['F1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws['F1'].fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
            
            # Set column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 60
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 20
            ws.column_dimensions['K'].width = 20
            
            # Extract and insert images from second ZIP
            print(f"\n🖼️ Processing images from second ZIP...")
            images_extract_dir = os.path.join(temp_dir, 'images')
            os.makedirs(images_extract_dir, exist_ok=True)
            
            image_mapping = extract_images_from_zip(zip2_path, images_extract_dir)
            print(f"📷 Found {len(image_mapping)} question(s) with images")
            
            if image_mapping:
                insert_images_into_worksheet(ws, image_mapping, temp_dir)
            
            # Increase row heights by 15px
            print(f"\n📏 Adjusting row heights...")
            for row_num in range(1, ws.max_row + 1):
                current_height = ws.row_dimensions[row_num].height or 15
                ws.row_dimensions[row_num].height = current_height + 15
            
            # Apply borders
            print(f"\n🔲 Applying borders...")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            left_top_bottom = Border(
                left=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            top_bottom = Border(
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            right_top_bottom = Border(
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Find last row with content in column A
            last_content_row = 1
            for row_num in range(1, ws.max_row + 1):
                if ws.cell(row=row_num, column=1).value is not None:
                    last_content_row = row_num
            
            print(f"  Last content row: {last_content_row}")
            
            # Apply borders to columns A-K
            for row_num in range(1, last_content_row + 1):
                # Columns A-D: full border
                for col_num in range(1, 5):  # A-D
                    ws.cell(row=row_num, column=col_num).border = thin_border
                
                # Column E: full border
                ws.cell(row=row_num, column=5).border = thin_border
                
                # Column F: left, top, bottom only
                ws.cell(row=row_num, column=6).border = left_top_bottom
                
                # Columns G-J: top and bottom only
                for col_num in range(7, 11):  # G-J
                    ws.cell(row=row_num, column=col_num).border = top_bottom
                
                # Column K: right, top, bottom only
                ws.cell(row=row_num, column=11).border = right_top_bottom
            
            # Add Annexure numbers
            print(f"\n🔢 Adding Annexure numbers...")
            annexure_counter = 1
            
            # Create a set of rows that have images
            rows_with_images = set()
            for img in ws._images:
                try:
                    # Get the anchor information
                    if hasattr(img, 'anchor'):
                        anchor = img.anchor
                        if hasattr(anchor, '_from'):
                            # openpyxl uses 0-based indexing for anchor rows
                            img_row = anchor._from.row + 1
                            rows_with_images.add(img_row)
                        elif isinstance(anchor, str):
                            # If anchor is a string like "G5", parse it
                            import re
                            match = re.match(r'[A-Z]+(\d+)', anchor)
                            if match:
                                img_row = int(match.group(1))
                                rows_with_images.add(img_row)
                except Exception as e:
                    print(f"  ⚠️ Error checking image anchor: {e}")
            
            print(f"  📍 Found images in {len(rows_with_images)} row(s)")
            
            # Add Annexure numbers to rows with images
            for row_num in sorted(rows_with_images):
                if row_num <= last_content_row:
                    ws.cell(row=row_num, column=5).value = f"Annexure {annexure_counter}"
                    ws.cell(row=row_num, column=5).font = Font(name='Times New Roman', size=12, color='FF0000')
                    ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center', vertical='center')
                    print(f"  ✅ Added Annexure {annexure_counter} at row {row_num}")
                    annexure_counter += 1
            
            # Save the consolidated workbook
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"LOC_Consolidated_Worksheet_{timestamp}.xlsx"
            output_path = os.path.join('static', 'uploads', output_filename)
            
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            wb.close()
            
            print(f"\n✅ Consolidated worksheet created: {output_filename}")
            print("="*80)
            
            # Return download URL
            download_url = f"/static/uploads/{output_filename}"
            return jsonify({
                "success": True,
                "message": "LOC Consolidated Worksheet created successfully",
                "download_url": download_url,
                "excel_file": "LOC.xlsx"  # Custom filename for download
            }), 200
            
        finally:
            # Clean up temporary directory
            try:
                shutil.rmtree(temp_dir)
                print("🧹 Cleaned up temporary files")
            except Exception as e:
                print(f"⚠️ Error cleaning up temp directory: {e}")
    
    except Exception as e:
        print(f"\n❌ Error creating LOC worksheet: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

