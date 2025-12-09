from flask import Blueprint, request, jsonify
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import tempfile
import shutil
import glob

# Create blueprint for LOC Level 3
loc_level3_bp = Blueprint('loc_level3', __name__)

def cleanup_loc_files():
    """Clean up old LOC Excel files from uploads directory"""
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            loc_pattern = os.path.join(upload_dir, 'LOC_Level*.xlsx')
            old_files = glob.glob(loc_pattern)
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

def create_loc_level3_excel(loc_data):
    """
    Create Excel file for LOC Level 3 with formatted data
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "LOC"
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 60
    
    # Define styles
    header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Times New Roman', size=12)
    data_font_bold = Font(name='Times New Roman', size=12, bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set headers
    ws['A1'] = "Sr. No."
    ws['B1'] = "Questions"
    ws['C1'] = "Yes/No given by Auditor"
    ws['D1'] = "Auditor's Observation"
    
    # Apply header formatting
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}1']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 30
    
    current_row = 2
    
    # Define all observations
    observations = {
        "12_A": {
            "Yes": "The bank has implemented appropriate mechanisms to detect, alert, and remediate unusual activities across systems, servers, network devices, and endpoints.",
            "No": "The bank has not implemented adequate mechanisms to detect and respond to unusual or suspicious activities across its IT environment.",
            "Not Applicable": "Not applicable as system and network monitoring is managed by an external service provider."
        },
        "12_B": {
            "Yes": "Firewall rules have been appropriately defined to block unauthorized outbound connections, reverse TCP shells, and backdoor traffic.",
            "No": "Firewall rules have not been properly configured to prevent unidentified outbound connections or potential backdoor activities.",
            "Not Applicable": "Not applicable as firewall management is outsourced to a third-party entity."
        },
        "13_A": {
            "Yes": "The bank operates an ATM Switch and/or SWIFT Interface as part of its core banking operations.",
            "No": "The bank does not operate an ATM Switch or SWIFT Interface.",
            "Not Applicable": "Not applicable as payment switch operations are managed by an external entity."
        },
        "13_B": {
            "Yes": "The bank has enabled IP tables to restrict ATM Switch and server access strictly to authorized systems.",
            "No": "IP table restrictions are not implemented or inadequately configured in the ATM Switch environment.",
            "Not Applicable": "Not applicable as the bank does not have or operate an ATM Switch."
        },
        "13_C": {
            "Yes": "The bank has verified and maintained the software integrity of ATM Switch and related applications through regular validation and control checks.",
            "No": "The bank has not established adequate processes to ensure the software integrity of ATM Switch or related systems.",
            "Not Applicable": "Not applicable as the bank does not have or operate an ATM Switch."
        },
        "13_D": {
            "Yes": "PowerShell has been disabled on all desktop systems where it is not operationally required.",
            "No": "PowerShell remains enabled on desktop systems, posing potential security risks.",
            "Not Applicable": "Not applicable due to non-Windows or vendor-managed environments."
        },
        "13_E": {
            "Yes": "PowerShell has been disabled on servers where it is not essential for administrative functions.",
            "No": "PowerShell has not been disabled on non-essential servers, increasing the risk of misuse.",
            "Not Applicable": "Not applicable as PowerShell is required for all operational servers."
        },
        "13_F": {
            "Yes": "The bank has restricted default shares, including IPC$, to prevent unauthorized access and misuse.",
            "No": "Default shares such as IPC$ remain enabled and accessible, posing potential security risks.",
            "Not Applicable": "Not applicable as file sharing services are managed by a service provider."
        },
        "14_A": {
            "Yes": "The bank has undertaken software development activities during the current financial year.",
            "No": "The bank did not develop any new software applications during the current financial year.",
            "Not Applicable": "Not applicable as all development activities are outsourced."
        },
        "14_B": {
            "Yes": "Source code audits are conducted for all critical business applications by qualified personnel or service providers.",
            "No": "Source code audits for critical applications are not performed consistently or formally.",
            "Not Applicable": "Not applicable as all development activities are outsourced."
        },
        "14_C": {
            "Yes": "Security requirements for system access control are integrated and enforced across all stages of development and implementation.",
            "No": "Security requirements for access control are not consistently applied.",
            "Not Applicable": "Not applicable as all development activities are outsourced."
        },
        "14_D": {
            "Yes": "Authentication controls are implemented at every stage of system development, acquisition, and implementation.",
            "No": "Authentication security requirements are inadequately enforced.",
            "Not Applicable": "Not applicable as all development activities are outsourced."
        },
        "14_E": {
            "Yes": "Transaction authorization controls are incorporated throughout the software lifecycle.",
            "No": "Transaction authorization requirements are not consistently implemented.",
            "Not Applicable": "Not applicable as all development activities are outsourced."
        },
        "14_F": {
            "Yes": "Data integrity controls are enforced at all stages of development and deployment.",
            "No": "Data integrity requirements are inconsistently applied or absent.",
            "Not Applicable": "Not applicable as all development activities are outsourced."
        },
        "14_G": {
            "Yes": "Logging and audit trail mechanisms are integrated throughout the application lifecycle.",
            "No": "Logging and audit trail requirements are not implemented consistently.",
            "Not Applicable": "Not applicable as all development activities are outsourced."
        },
        "14_H": {
            "Yes": "Session management controls are designed and enforced across all development and implementation stages.",
            "No": "Session management requirements are insufficiently applied.",
            "Not Applicable": "Not applicable as all development activities are outsourced."
        },
        "14_I": {
            "Yes": "Security event tracking mechanisms are incorporated at all stages of development and deployment.",
            "No": "Security event tracking requirements are not consistently enforced.",
            "Not Applicable": "Not applicable as all development activities are outsourced."
        },
        "14_J": {
            "Yes": "Exception handling security controls are implemented throughout the software lifecycle.",
            "No": "Exception handling requirements are inconsistently applied or missing.",
            "Not Applicable": "Not applicable as all development activities are outsourced."
        },
        "14_K": {
            "Yes": "Development practices adopt defence-in-depth principles to ensure multiple layers of security.",
            "No": "Defence-in-depth principles are not consistently applied in development practices.",
            "Not Applicable": "Not applicable as all development activities are outsourced."
        },
        "14_L": {
            "Yes": "Application providers/OEMs provide assurance that applications are free from embedded malicious or fraudulent code.",
            "No": "No such assurance has been obtained from providers/OEMs.",
            "Not Applicable": "Not applicable as all development activities are outsourced."
        },
        "14_M": {
            "Yes": "Guidelines exist to evaluate new technologies for security threats before adoption.",
            "No": "No formal guidelines exist for evaluating new technologies for security risks.",
            "Not Applicable": "Not applicable as all development activities are outsourced."
        },
        "14_N": {
            "Yes": "Guidelines are in place for IT/security teams to ensure adequate maturity with new technologies before deploying them in critical systems.",
            "No": "There are no formal guidelines to ensure technology maturity before introducing it to critical systems.",
            "Not Applicable": "Not applicable as all critical systems are managed through pre-evaluated vendor solutions."
        },
        "15_A": {
            "Yes": "User access controls are implemented across all infrastructure modes including in-house, shared, and service provider environments.",
            "No": "User access controls are inconsistently applied across different infrastructure management modes.",
            "Not Applicable": "Not applicable as all user access management is outsourced to a third-party provider."
        },
        "15_B": {
            "Yes": "A centralised Identity and Access Management (IAM) solution is implemented for authentication and authorisation.",
            "No": "No centralised authentication/authorisation system has been implemented.",
            "Not Applicable": "Not applicable as all authentication and access management is handled by external service providers."
        },
        "15_C": {
            "Yes": "The centralised IAM system is applied for all operating system access and administration.",
            "No": "Operating system access is not controlled via the centralised IAM system.",
            "Not Applicable": "Not applicable as OS management is fully outsourced to a service provider."
        },
        "15_D": {
            "Yes": "Database access and administration are controlled through the centralised IAM system.",
            "No": "Database access is managed independently without centralised IAM controls.",
            "Not Applicable": "Not applicable as databases are maintained and managed entirely by third-party vendors."
        },
        "15_E": {
            "Yes": "Network and security device access is enforced through the centralised IAM system.",
            "No": "Access to network and security devices is not controlled centrally.",
            "Not Applicable": "Not applicable as network/security device administration is outsourced."
        },
        "15_F": {
            "Yes": "All critical applications use the centralised IAM system for access and administration.",
            "No": "Critical applications are accessed without centralised authentication and authorisation.",
            "Not Applicable": "Not applicable as all critical applications are externally managed."
        },
        "15_G": {
            "Yes": "Access to all local and remote connectivity points is controlled via a centralised authentication and authorisation system.",
            "No": "Connectivity access is not centrally controlled for all points.",
            "Not Applicable": "Not applicable as connectivity is fully managed by a third-party provider."
        },
        "15_H": {
            "Yes": "The centralised system enforces strong password policies across all users and access points.",
            "No": "Strong password policies are not consistently enforced.",
            "Not Applicable": "Not applicable as password enforcement is managed externally."
        },
        "15_I": {
            "Yes": "Two-factor/multi-factor authentication is enforced through the centralised IAM system.",
            "No": "Two-factor/multi-factor authentication is not enforced for users.",
            "Not Applicable": "Not applicable as authentication mechanisms are managed by external service providers."
        },
        "15_J": {
            "Yes": "Privileged access is governed by policies enforcing least privilege and separation of duties.",
            "No": "Policies for privileged access do not adequately enforce least privilege or segregation of duties.",
            "Not Applicable": "Not applicable as privileged access is fully managed by an external service provider."
        },
        "15_K": {
            "Yes": "Active Directory or Endpoint Management systems are implemented to enforce centralised security policies including whitelisting, blacklisting, and removable media restrictions.",
            "No": "Such centralised endpoint security policies are not enforced.",
            "Not Applicable": "Not applicable as endpoint management is entirely outsourced to a service provider."
        },
        "16_A": {
            "Yes": "Appropriate defence mechanisms are implemented across multiple points in the enterprise to prevent the installation, spread, and execution of malicious code.",
            "No": "Defences against malicious code are inadequate or not consistently applied across the enterprise.",
            "Not Applicable": "Not applicable as malware protection and endpoint security are fully managed by external service providers."
        },
        "16_B": {
            "Yes": "A mechanism to whitelist approved internet websites and systems is in place and actively enforced.",
            "No": "No formal whitelisting mechanism is implemented to control web or system access.",
            "Not Applicable": "Not applicable as web filtering and access control are handled entirely by third-party providers."
        },
        "17_A": {
            "Yes": "Procedures involve all relevant stakeholders before defining log collection scope, frequency, and storage.",
            "No": "Stakeholders are not consulted, or procedures are ad hoc for log management.",
            "Not Applicable": "Not applicable as log management is outsourced to an external service provider."
        },
        "17_B": {
            "Yes": "An established system exists to analyze and manage audit logs for detecting, responding, and recovering from attacks.",
            "No": "There is no structured system for audit log analysis or incident response.",
            "Not Applicable": "Not applicable as audit log analysis is fully managed by a third-party provider."
        },
        "17_C": {
            "Yes": "Audit logs and trails are captured consistently for all devices, system software, and applications.",
            "No": "Audit logs are not captured for all critical systems or applications.",
            "Not Applicable": "Not applicable as logging is fully performed by an external service provider."
        },
        "17_D": {
            "Yes": "Audit log settings are reviewed and validated periodically to ensure completeness and reliability.",
            "No": "Audit log settings are not regularly validated or reviewed.",
            "Not Applicable": "Not applicable as audit log validation is managed externally."
        },
        "17_E": {
            "Yes": "Logs are captured with sufficient information to uniquely identify events for auditing purposes.",
            "No": "Logs lack critical information needed for unique identification of events.",
            "Not Applicable": "Not applicable as log capture and retention are managed by a third-party provider."
        },
        "18_A": {
            "Yes": "BCP/DR capabilities are aligned with the organization's cyber resilience objectives.",
            "No": "BCP/DR capabilities are misaligned or not regularly updated to meet cyber resilience objectives.",
            "Not Applicable": "Not applicable as BCP/DR is managed externally."
        },
        "18_B": {
            "Yes": "Arrangements with third-party vendors are in place for effective incident response and management.",
            "No": "No formal arrangements exist with third-party vendors for incident response.",
            "Not Applicable": "Not applicable as incident response is handled entirely in-house or fully outsourced."
        },
        "18_C": {
            "Yes": "Mechanisms exist with third-party providers to obtain timely information on cybersecurity incidents for early mitigation.",
            "No": "No mechanisms are in place to receive timely cybersecurity incident information from vendors.",
            "Not Applicable": "Not applicable as all cybersecurity incident management is handled in-house."
        },
        "18_D": {
            "Yes": "Mechanisms exist to update and improve incident response strategies based on lessons learned.",
            "No": "Response strategies are static and do not incorporate lessons learned from past incidents.",
            "Not Applicable": "Not applicable as incident response improvement is fully managed by external providers."
        },
        "19_A": {
            "Yes": "The bank is a direct CPS member and maintains its own ATM Switch or SWIFT interface.",
            "No": "The bank is not a direct CPS member and/or does not have an ATM Switch/SWIFT interface.",
            "Not Applicable": "Not applicable as the bank does not operate any transactional switching or SWIFT interfaces."
        },
        "19_B": {
            "Yes": "Risk-based transaction monitoring is implemented across all delivery channels as part of the fraud risk management system.",
            "No": "Risk-based transaction monitoring is not implemented or is inconsistent across channels.",
            "Not Applicable": "Not applicable as the bank does not operate its own ATM/SWIFT interface."
        }
    }
    
    # Define question structure with section headers
    sections = {
        "12": {
            "title": "Network Management and Security",
            "questions": ["12_A", "12_B"]
        },
        "13": {
            "title": "Secure Configuration",
            "questions": ["13_A", "13_B", "13_C", "13_D", "13_E", "13_F"]
        },
        "14": {
            "title": "Application Security Life Cycle (ASLC)",
            "questions": ["14_A", "14_B", "14_C", "14_D", "14_E", "14_F", "14_G", "14_H", "14_I", "14_J", "14_K", "14_L", "14_M", "14_N"]
        },
        "15": {
            "title": "User Access Control",
            "questions": ["15_A", "15_B", "15_C", "15_D", "15_E", "15_F", "15_G", "15_H", "15_I", "15_J", "15_K"]
        },
        "16": {
            "title": "Advance Real-time Threat Defence and Management",
            "questions": ["16_A", "16_B"]
        },
        "17": {
            "title": "Maintenance, Monitoring and Analysis of Audit Logs",
            "questions": ["17_A", "17_B", "17_C", "17_D", "17_E"]
        },
        "18": {
            "title": "Incedent Response and Management",
            "questions": ["18_A", "18_B", "18_C", "18_D"]
        },
        "19": {
            "title": "Risk based transaction monitoring",
            "questions": ["19_A", "19_B"]
        }
    }
    
    # Question text mapping
    question_texts = {
        "12_A": "Whether appropriate mechanisms to detect and remedy any unusual activity in systems, servers, network devices and endpoints are in place?",
        "12_B": "Whether Firewall rules have been defined to block unidentified outbound connections, reverse TCP shells and other potential backdoor connections?",
        "13_A": "Do you have ATM Switch or SWIFT Interface?                                                                    If no, skip the next 2 questions.",
        "13_B": "Whether IP table is enabled to restrict access of the clients and servers in ATM Switch environment only to authorised systems?",
        "13_C": "Has the bank ensured the software integrity of the ATM Switch / SWITCH related applications?",
        "13_D": "Whether PowerShell is disabled in Desktop systems?",
        "13_E": "Whether PowerShell is disabled in Servers where it is not required?",
        "13_F": "Have the default shares including IPC$ share(inter-process communication share) been restricted?",
        "14_A": "Did the bank develop any software application during the current financial year? (If \"No\" then skip Q 54 to 63)",
        "14_B": "Whether  source code audits were conducted for critical business applications by professionally competent personnel/service providers?",
        "14_C": "Throughout the different stages of system development/ acquisition/ implementation, did you ensure security requirements relating to system access control?",
        "14_D": "Throughout the different stages of system development/ acquisition/ implementation, did you ensure security requirements relating to user authentication?",
        "14_E": "Throughout the different stages of system development/ acquisition/ implementation, did you ensure security requirements relating to transaction authorization ?",
        "14_F": "Throughout the different stages of system development/ acquisition/ implementation, did you ensure security requirements relating to data integrity?",
        "14_G": "Throughout the different stages of system development/ acquisition/ implementation, did you ensure security requirements relating to system activity logging and audit trail?",
        "14_H": "Throughout the different stages of system development/ acquisition/ implementation, did you ensure security requirements relating to   session management?",
        "14_I": "Throughout the different stages of system development/ acquisition/ implementation, did you ensure security requirements relating to security event tracking?",
        "14_J": "Throughout the different stages of system development/ acquisition/ implementation, did you ensure security requirements relating to exception handling?",
        "14_K": "Whether software/application development practices  principle of defence-in-depth has been adopted to provide layered security mechanism?",
        "14_L": "Whether your application providers/OEMs have given assurance that the applications provided by them are free from embedded malicious/fraudulent code?",
        "14_M": "Do you have proper guideline in place to ensure that adoption of new technologies are adequately evaluated for existing/evolving security threats ?",
        "14_N": "Do you have guidelines for  IT/security team of bank to first achieve reasonable level of comfort and maturity with new technologies before introducing them for critical systems of the bank?",
        "15_A": "Answer following questions (67 to 76)  irrespective of whether IT infrastructure is managed by an in-house team or whether the infrastructure is hosted at a shared location or at the service providers end. All User Acess Controls have to be ensured under all above circumstances.",
        "15_B": "Whether a centralised authentication and authorisation system through an Identity and Access Management solution has been implemented?",
        "15_C": "Is the centralised authentication and authorisation system implemented for accessing and administering operating systems?",
        "15_D": "Is the centralised authentication and authorisation system implemented for accessing and administering databases?",
        "15_E": "is the centralised authentication and authorisation system implemented for accessing and administering network and security devices/systems?",
        "15_F": "Is the centralised authentication and authorisation system implemented for accessing and administering  other critical applications (anything not mentioned from 67 to 70 above)?",
        "15_G": "Do you have centralised authentication and authorisation system for accessing all point of connectivity(local/remote)?",
        "15_H": "Does the centralised authentication and authorisation system enforce strong password policy?",
        "15_I": "Does the centralised authentication and authorisation system enforce two-factor/multi-factor authentication?",
        "15_J": "Does the bank / vendors policy for grant of privileged accesses follow the principle of least privileges and separation of duties?",
        "15_K": "Have you implemented Active Directory or Endpoint Management systems to implement centralised polciies like whitelisting /blacklisting /retricting removable media etc",
        "16_A": "Have you set in place proper defence mechanism against the installation, spread and execution of malicious code at multiple points in the enterprise?",
        "16_B": "Whether mechanism for whitelisting of internet websites/systems has been put in place?",
        "17_A": "Whether procedure laid down for consulting all stakeholders before finalising the scope, frequency and storage of log collection?",
        "17_B": "Whether there is an established system for analysis and management of audit logs which will help the bank to detect, respond, understand or recover from an attack?",
        "17_C": "Whether audit log / audit trail is captured for all devices, system software and application software?",
        "17_D": "Whether all the settings for all audit logs are validated periodically?",
        "17_E": "Whether the logs are properly captured with minimum information to uniquely identify the log?",
        "18_A": "Whether banks existing  BCP/DR capabilities are properly in tune with its cyber resilience objectives?",
        "18_B": "Do you have necessary arrangements with third party vendors/ service providers for implementation of Incident Response and Management?",
        "18_C": "Do you have necessary mechanism with third party vendor / service provider to get timely information about  any cyber security incident for early mitigation of risk and and to meet extant regulatory requirements ?",
        "18_D": "Do you have mechanism to continually improve response strategy by dynamically incorporating gained knowledge in the Incident Response Procedure of the bank / BCP?",
        "19_A": "Whether bank is a direct member of CPS and has its own ATM switch interface or SWIFT interface?",
        "19_B": "If yes whether bank has implemented Risk based transaction monitoring or surveillance as part of Fraud Risk Management Sysytems across all delivery channels?"
    }
    
    # Add "Level III" header row
    ws[f'A{current_row}'].value = ""
    ws[f'B{current_row}'].value = "Level III"
    ws[f'C{current_row}'].value = ""
    ws[f'D{current_row}'].value = ""
    
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}{current_row}']
        cell.font = data_font_bold
        cell.alignment = left_alignment
        cell.border = thin_border
    
    current_row += 1
    
    # Process each section
    for section_num, section_data in sections.items():
        # Add section header
        ws[f'A{current_row}'].value = section_num
        ws[f'B{current_row}'].value = section_data["title"]
        ws[f'C{current_row}'].value = ""
        ws[f'D{current_row}'].value = ""
        
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{current_row}']
            cell.font = data_font_bold
            cell.alignment = left_alignment if col == 'B' else center_alignment
            cell.border = thin_border
        
        current_row += 1
        
        # Add questions
        for q_id in section_data["questions"]:
            # Form data has keys like "q12_A", so add "q" prefix
            form_key = f"q{q_id}"
            answer = loc_data.get(form_key, "")
            
            # Extract letter from question ID (e.g., "12_A" -> "A")
            letter = q_id.split('_')[1]
            
            ws[f'A{current_row}'].value = letter
            ws[f'B{current_row}'].value = question_texts.get(q_id, "")
            
            # Set Yes/No/NA in column C
            if answer == "Yes":
                ws[f'C{current_row}'].value = "Yes"
            elif answer == "No":
                ws[f'C{current_row}'].value = "No"
            elif answer == "Not Applicable":
                ws[f'C{current_row}'].value = "NA"
            else:
                ws[f'C{current_row}'].value = ""
            
            # Set observation in column D
            observation = observations.get(q_id, {}).get(answer, "")
            ws[f'D{current_row}'].value = observation
            
            # Apply formatting
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].alignment = center_alignment
            ws[f'A{current_row}'].border = thin_border
            
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].alignment = left_alignment
            ws[f'B{current_row}'].border = thin_border
            
            ws[f'C{current_row}'].font = data_font
            ws[f'C{current_row}'].alignment = center_alignment
            ws[f'C{current_row}'].border = thin_border
            
            ws[f'D{current_row}'].font = data_font
            ws[f'D{current_row}'].alignment = left_alignment
            ws[f'D{current_row}'].border = thin_border
            
            current_row += 1
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name

@loc_level3_bp.route('/process_loc_level3', methods=['POST'])
def process_loc_level3():
    """
    Process LOC Level 3 form data and generate Excel
    """
    try:
        print("\n" + "="*80)
        print("🎯 LOC LEVEL 3 - List of Compliances")
        print("="*80)
        
        # Get form data
        form_data = request.form.to_dict()
        
        print("📋 Form Data Received:")
        for key, value in form_data.items():
            print(f"  {key}: {value}")
        
        # Clean up old LOC files before generating new one
        print("\n🧹 Cleaning up old LOC files...")
        cleanup_loc_files()
        
        # Generate Excel file
        print("\n📝 Generating Excel file...")
        excel_file_path = create_loc_level3_excel(form_data)
        print(f"✅ Excel file created: {excel_file_path}")
        
        # Save to static/uploads directory
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"LOC_Level3_{timestamp}.xlsx"
        static_upload_path = os.path.join('static', 'uploads', excel_filename)
        
        os.makedirs(os.path.dirname(static_upload_path), exist_ok=True)
        shutil.copy2(excel_file_path, static_upload_path)
        os.unlink(excel_file_path)
        
        print(f"✅ Excel saved to: {static_upload_path}")
        print("="*80)
        
        return jsonify({
            "success": True,
            "message": "LOC Level 3 data processed successfully",
            "download_url": f"/static/uploads/{excel_filename}",
            "excel_file": "LOC_Level3.xlsx"
        }), 200
        
    except Exception as e:
        print(f"❌ Error processing LOC Level 3: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"Error processing LOC Level 3: {str(e)}"
        }), 500
