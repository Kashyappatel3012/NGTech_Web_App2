from flask import Blueprint, request, jsonify
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import tempfile
import shutil
import glob

# Create blueprint for LOC Level 2
loc_level2_bp = Blueprint('loc_level2', __name__)

def cleanup_loc_files():
    """Clean up old LOC Excel files from uploads directory"""
    try:
        upload_dir = os.path.join('static', 'uploads')
        if os.path.exists(upload_dir):
            loc_pattern = os.path.join(upload_dir, 'LOC_Level*.xlsx')
            old_files = glob.glob(loc_pattern)
            for file_path in old_files:
                try:
                    os.remove(file_path)
                    print(f"🗑️ Deleted old file: {os.path.basename(file_path)}")
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path}: {e}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

@loc_level2_bp.route('/cleanup_loc_files', methods=['POST'])
def cleanup_loc_files_endpoint():
    """Endpoint to cleanup LOC files after download"""
    try:
        cleanup_loc_files()
        return jsonify({"success": True, "message": "Cleanup completed"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

def create_loc_level2_excel(loc_data):
    """
    Create Excel file for LOC Level 2 with formatted data
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "LOC"
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 60
    
    # Define styles
    header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00008B', end_color='00008B', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Times New Roman', size=12)
    data_font_bold = Font(name='Times New Roman', size=12, bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set headers
    ws['A1'] = "Sr. No."
    ws['B1'] = "Questions"
    ws['C1'] = "Yes/No given by Auditor"
    ws['D1'] = "Auditor's Observation"
    
    # Apply header formatting
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}1']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 30
    
    current_row = 2
    
    # Define all observations
    observations = {
        "1_A": {
            "Yes": "The bank has a centralised inventory for authorised devices and other network components connected to its network.",
            "No": "The bank does not maintain a centralised inventory for authorised devices and network components.",
            "Not Applicable": "This requirement is not applicable as the network infrastructure is managed by an external service provider."
        },
        "1_B": {
            "Yes": "The bank has implemented a multi-layered boundary defence to protect its network infrastructure.",
            "No": "The bank has not implemented a multi-layered boundary defence for its network environment.",
            "Not Applicable": "This control is not applicable as boundary defence is handled by a parent or third-party entity."
        },
        "1_C": {
            "Yes": "The bank has configured firewalls as part of its boundary defence mechanism.",
            "No": "The bank has not configured firewalls in its boundary defence setup.",
            "Not Applicable": "This requirement is not applicable as firewall management is performed by a service provider."
        },
        "1_D": {
            "Yes": "The bank has configured proxy servers for secure traffic routing and content filtering.",
            "No": "The bank has not implemented proxy servers in its boundary defence system.",
            "Not Applicable": "This control is not applicable as internet traffic is managed externally."
        },
        "1_E": {
            "Yes": "The bank has implemented a DMZ to host public-facing services securely.",
            "No": "The bank has not implemented a DMZ, increasing the exposure of external-facing systems.",
            "Not Applicable": "This requirement is not applicable as the bank does not host any public-facing applications."
        },
        "1_F": {
            "Yes": "The bank has deployed IDS/IPS solutions to monitor and protect its network perimeter.",
            "No": "The bank has not deployed IDS/IPS solutions for perimeter security monitoring.",
            "Not Applicable": "This control is not applicable as perimeter monitoring is handled by an external vendor."
        },
        "1_G": {
            "Yes": "The boundary defence system is configured to filter both inbound and outbound network traffic.",
            "No": "The boundary defence system does not adequately filter inbound and outbound network traffic.",
            "Not Applicable": "This control is not applicable as network traffic filtering is managed by a third-party provider."
        },
        "1_H": {
            "Yes": "The bank has in-house/onsite ATM facilities at certain locations.",
            "No": "The bank does not operate any in-house or onsite ATM facility.",
            "Not Applicable": "Not applicable as ATM operations are managed externally."
        },
        "1_I": {
            "Yes": "The bank has segregated LAN segments for ATM and CBS/branch networks to ensure security.",
            "No": "The bank has not segregated LAN segments for ATM and CBS/branch networks.",
            "Not Applicable": "Not applicable as the bank does not have any in-house or onsite ATM."
        },
        "2_A": {
            "Yes": "The bank has applied baseline security configurations to all categories of devices across their lifecycle.",
            "No": "The bank has not applied uniform baseline security configurations across all device categories.",
            "Not Applicable": "This requirement is not applicable due to limited device infrastructure or vendor-managed systems."
        },
        "2_B": {
            "Yes": "The bank maintains documentation for security configurations applied to each category of device.",
            "No": "The bank does not maintain documentation for device-level security configurations.",
            "Not Applicable": "Not applicable as configuration management is handled by a third-party vendor."
        },
        "2_C": {
            "Yes": "The bank periodically reviews and validates the applied security configurations.",
            "No": "The bank does not perform periodic reviews of applied configurations on devices.",
            "Not Applicable": "This requirement is not applicable as configuration review is managed externally."
        },
        "3_A": {
            "Yes": "The bank has undertaken software development activities following secure coding practices.",
            "No": "The bank does not engage in any internal software development activities.",
            "Not Applicable": "Not applicable as all application development is outsourced."
        },
        "3_B": {
            "Yes": "The bank has properly segregated development, testing, and production environments.",
            "No": "The bank has not maintained segregation between development/test and production environments.",
            "Not Applicable": "Not applicable as the bank does not conduct any development or testing in-house."
        },
        "3_C": {
            "Yes": "The bank ensures that all data used in development and testing environments is properly masked.",
            "No": "The bank does not implement adequate masking for data used during development or testing.",
            "Not Applicable": "Not applicable as the bank does not perform in-house software development or testing."
        },
        "3_D": {
            "Yes": "The bank follows secure coding standards and guidelines in all software development activities.",
            "No": "The bank does not consistently follow secure coding principles during development.",
            "Not Applicable": "This requirement is not applicable as no internal software development is carried out."
        },
        "3_E": {
            "Yes": "The bank performs security testing and secure rollouts as per global or industry best practices.",
            "No": "The bank does not perform adequate security testing before rollout of applications.",
            "Not Applicable": "Not applicable as no in-house software development or rollout activities are performed."
        },
        "4_A": {
            "Yes": "The bank has an established change management process to record and monitor production changes.",
            "No": "The bank does not have a formal change management process to track system modifications.",
            "Not Applicable": "This control is not applicable as system changes are handled by a service provider."
        },
        "4_B": {
            "Yes": "Changes in applications or systems during the year were properly recorded and approved.",
            "No": "No major system or application changes were observed during the review period.",
            "Not Applicable": "Not applicable as there were no operational changes during the audit period."
        },
        "4_C": {
            "Yes": "The bank employs configuration management processes for version and change control.",
            "No": "The bank does not have an effective configuration management process in place.",
            "Not Applicable": "This requirement is not applicable as configuration management is handled externally."
        },
        "5_A": {
            "Yes": "The bank conducts VA/PT assessments prior to implementation of internet-facing applications or systems.",
            "No": "The bank does not conduct pre-implementation VA/PT before deploying new systems.",
            "Not Applicable": "Not applicable as no new applications or systems were deployed during the review period."
        },
        "5_B": {
            "Yes": "The bank conducts periodic VA/PT for its applications and network infrastructure after deployment.",
            "No": "The bank does not conduct regular VA/PT assessments as required.",
            "Not Applicable": "Not applicable as the environment is managed by an external service provider."
        },
        "5_C": {
            "Yes": "The bank conducts vulnerability assessments for critical and DMZ-based applications at least every six months.",
            "No": "The bank has not conducted biannual vulnerability assessments for critical applications.",
            "Not Applicable": "Not applicable as no critical or DMZ-based applications are hosted by the bank."
        },
        "5_D": {
            "Yes": "The bank conducts annual penetration testing for all critical and DMZ-hosted applications.",
            "No": "The bank does not perform yearly penetration testing for critical applications.",
            "Not Applicable": "Not applicable as there are no such applications in the bank's environment."
        },
        "5_E": {
            "Yes": "The bank's Core Banking System is hosted on a shared infrastructure with appropriate access controls.",
            "No": "The bank's Core Banking System operates on a dedicated infrastructure.",
            "Not Applicable": "Not applicable as the bank does not use CBS-ASP services."
        },
        "5_F": {
            "Yes": "The bank ensures that VA/PT of CBS shared infrastructure is conducted by the service provider as per schedule.",
            "No": "VA/PT of CBS shared infrastructure has not been conducted or evidence is unavailable.",
            "Not Applicable": "Not applicable as the bank's CBS is hosted independently."
        },
        "5_G": {
            "Yes": "The bank conducts security testing before applications go live and after major updates.",
            "No": "The bank does not perform mandatory security testing before application rollout.",
            "Not Applicable": "Not applicable as no major application changes were implemented."
        },
        "5_H": {
            "Yes": "The bank remediates identified vulnerabilities in line with its risk management framework.",
            "No": "The bank has not remediated vulnerabilities as per its defined risk management timelines.",
            "Not Applicable": "Not applicable as no significant vulnerabilities were reported during the audit period."
        },
        "5_I": {
            "Yes": "VA/PT exercises are performed by certified and qualified professionals.",
            "No": "VA/PT activities are not conducted by qualified or certified personnel.",
            "Not Applicable": "Not applicable as no VA/PT assessments were performed."
        },
        "5_J": {
            "Yes": "The VA/PT findings are reviewed by the IT Sub-Committee and reported to the Board.",
            "No": "The VA/PT findings are not formally presented to the IT Sub-Committee or Board.",
            "Not Applicable": "Not applicable as no VA/PT assessments were carried out."
        },
        "5_K": {
            "Yes": "The bank has taken appropriate follow-up and remediation actions on VA/PT observations.",
            "No": "The bank has not taken adequate follow-up action on the reported VA/PT findings.",
            "Not Applicable": "Not applicable as there were no VA/PT findings to address."
        },
        "6_A": {
            "Yes": "The bank has implemented encryption mechanisms to secure data at rest for all internal and external access points.",
            "No": "The bank has not implemented adequate encryption controls for securing data at rest.",
            "Not Applicable": "Not applicable as the bank's data storage and access are managed by an external service provider."
        },
        "6_B": {
            "Yes": "The bank secures all data in transit through VPNs and other standard encryption technologies.",
            "No": "The bank does not use VPN or equivalent secure technologies for protecting data in transit.",
            "Not Applicable": "Not applicable as all external connectivity is managed and secured by a third-party provider."
        },
        "7_A": {
            "Yes": "The bank has implemented a centralised authentication mechanism ensuring all customer transactions occur through authorised applications.",
            "No": "The bank lacks a centralised authentication methodology, increasing risk of unauthorised access.",
            "Not Applicable": "Not applicable as customer authentication is managed by a service provider or parent entity."
        },
        "7_B": {
            "Yes": "The bank has implemented authentication mechanisms to securely verify and identify its applications to customers.",
            "No": "The bank has not implemented sufficient mechanisms to verify and authenticate its applications.",
            "Not Applicable": "Not applicable as the bank does not provide customer-facing applications directly."
        },
        "7_C": {
            "Yes": "The bank uses valid and updated digital certificates to verify and authenticate its applications for customers.",
            "No": "The bank does not use digital certificates to authenticate its customer-facing applications.",
            "Not Applicable": "Not applicable as the bank does not operate customer-facing applications."
        },
        "8_A": {
            "Yes": "The bank has subscribed to external anti-phishing and anti-rogue monitoring services to detect and take down fake websites or applications.",
            "No": "The bank has not subscribed to any anti-phishing or rogue application monitoring services.",
            "Not Applicable": "Not applicable as the bank does not operate internet-facing services or applications."
        },
        "9_A": {
            "Yes": "The bank encourages employees to promptly report suspicious activities and incidents to the incident response team.",
            "No": "The bank has not established adequate measures to encourage employees to report suspicious behaviour.",
            "Not Applicable": "Not applicable as incident reporting is managed externally."
        },
        "9_B": {
            "Yes": "The bank conducts mandatory cybersecurity awareness sessions for all newly recruited employees.",
            "No": "The bank does not conduct cybersecurity awareness programs for new recruits.",
            "Not Applicable": "Not applicable as recruitment and training activities are managed externally."
        },
        "9_C": {
            "Yes": "The bank conducts annual web-based quizzes and training programs for all levels of management to enhance cyber awareness.",
            "No": "The bank does not conduct annual cyber awareness training or web-based quizzes for management personnel.",
            "Not Applicable": "Not applicable as training programs are handled by a centralised or third-party entity."
        },
        "9_D": {
            "Yes": "The bank periodically sensitises Board members on new technological and cybersecurity developments.",
            "No": "The bank has not conducted awareness sessions or sensitisation programs for Board members.",
            "Not Applicable": "Not applicable as Board sensitisation is managed by a higher or parent institution."
        },
        "10_A": {
            "Yes": "The bank's systems are configured to capture and retain detailed audit logs of all user activities.",
            "No": "The bank's systems are not adequately capturing or maintaining audit logs of user actions.",
            "Not Applicable": "Not applicable as audit logging is managed by a third-party system provider."
        },
        "10_B": {
            "Yes": "An alert mechanism is in place to monitor and notify any changes made to audit log configurations.",
            "No": "The bank has not implemented an alert mechanism to track changes in audit log settings.",
            "Not Applicable": "Not applicable as audit log management is outsourced to an external vendor."
        },
        "11_A": {
            "Yes": "The bank has well-documented incident response procedures defining roles, responsibilities, and escalation processes for cyber incidents.",
            "No": "The bank does not have a formal or documented incident response plan to handle cybersecurity events.",
            "Not Applicable": "Not applicable as incident management is handled by a third-party or parent organization."
        }
    }
    
    # Define question structure with section headers
    sections = {
        "1": {
            "title": "Network Management and Security",
            "questions": ["1_A", "1_B", "1_C", "1_D", "1_E", "1_F", "1_G", "1_H", "1_I"]
        },
        "2": {
            "title": "Secure Configuration",
            "questions": ["2_A", "2_B", "2_C"]
        },
        "3": {
            "title": "Application Security Life Cycle (ASLC)",
            "questions": ["3_A", "3_B", "3_C", "3_D", "3_E"]
        },
        "4": {
            "title": "Change Management",
            "questions": ["4_A", "4_B", "4_C"]
        },
        "5": {
            "title": "Periodic Testing",
            "questions": ["5_A", "5_B", "5_C", "5_D", "5_E", "5_F", "5_G", "5_H", "5_I", "5_J", "5_K"]
        },
        "6": {
            "title": "User Access Control/Management",
            "questions": ["6_A", "6_B"]
        },
        "7": {
            "title": "Authentication Framework for Customers",
            "questions": ["7_A", "7_B", "7_C"]
        },
        "8": {
            "title": "Anti-Phishing",
            "questions": ["8_A"]
        },
        "9": {
            "title": "User/Employee/Management Awareness",
            "questions": ["9_A", "9_B", "9_C", "9_D"]
        },
        "10": {
            "title": "Audit Logs",
            "questions": ["10_A", "10_B"]
        },
        "11": {
            "title": "Incident Response and Management",
            "questions": ["11_A"]
        }
    }
    
    # Question text mapping
    question_texts = {
        "1_A": "Do you have a centralised inventory of authorised devices and other related network devices connected to banks network (within/outside banks premises)?",
        "1_B": "Whether boundary defence for banks network is multi-layered?",
        "1_C": "Whether Firewall configured in Boundary defence?",
        "1_D": "Whether proxy servers configured in boundary defence system?",
        "1_E": "Whether De-Militarized Zone (DMZ) is implemented in boundary defence system?",
        "1_F": "Whether Network-based Intrusion Prevention System, Intrusion Detection System or other perimeter networks installed?",
        "1_G": "Whether the boundary defence system put in place is able to filter both inbound and outbound traffic?",
        "1_H": "Do you have in-house or onsite ATM facility at HO or any of the branches of the bank? If no, skip the next question",
        "1_I": "Whether LAN segments for in-house/onsite ATM and CBS/branch network are different?",
        "2_A": "Whether baseline security requirements/configurations are applied to all categories of devices i.e. end-points/workstations, mobile devices, operating systems, databases, applications, network devices, security devices, security systems etc. throughout their lifecycle(from conception to deployment)?",
        "2_B": "Whether the security requirements/configurations for each and every device is documented?",
        "2_C": "Whether the security requirements/configurations applied on devices are reviewed periodically?",
        "3_A": "Is the bank developing any software application?Answer Yes even if the software development was completed during last financial year",
        "3_B": "Whether the development/test environments were / are properly segregated from production environments?",
        "3_C": "Whether the data used for development and testing of application/software was / is appropriately masked?",
        "3_D": "Whether secure coding principles were / are adhered to in case of software/application development approach?",
        "3_E": "Whether security testing and secure rollout were / are done as per global standards, in case of software/application development approach?",
        "4_A": "Whether bank has a change management process in place to record and monitor all changes that are moved/pushed into production environment by bank / software vendor?",
        "4_B": "Whether there was change in any business applications or supporting technology or service components and facilities during the last year?",
        "4_C": "Whether configuration management processes were employed by the bank / vendor?",
        "5_A": "Whether Vulnerability Assessment(VA) /Penetration Testing (VA / PT) of internet facing or web or mobile applications or servers and/or network components were conducted in pre implementation stage?",
        "5_B": "Whether periodic Vulnerability Assessment(VA) /Penetration Testing (VA / PT) of internet facing or web or mobile applications or servers and/or network components are being conducted in post implementation stage?",
        "5_C": "Whether Vulnerability Assessment(VA) of critical applications and those on DMZ are conducted at least once every six months?",
        "5_D": "Whether Penetration Testing(PT) of critical applications and those on DMZ are conducted atleast once in a year?",
        "5_E": "Whether the bank has its CBS on a shared infrastructure of an Application Service Provider(CBS-ASP) / Apex Bank / Sponsor bank? If no, skip the next question",
        "5_F": "Whether VA/PT is conducted for CBS provided by ASP (CBS -ASP) / Apex Bank / Sponsor bank on shared infrastructure?",
        "5_G": "Whether security testing of web/mobile applications is conducted before going live and after every major changes in the applications?",
        "5_H": "Whether the vulnerabilities detected through VA/PT are remedied as per banks risk management framework?",
        "5_I": "Are the VA and / or PTs conducted by professionally qualified teams?",
        "5_J": "Whether findings of the VA and / or PT reports were placed before IT sub Committee of the Board and / or Board of Directors?",
        "5_K": "Whether follow up action was taken on findings of the VA and / or PT reports by IS / IT / Audit / Top Management team?",
        "6_A": "Whether the access to Banks assets / services from within outside banks network is secured by use of encryption for data at rest?",
        "6_B": "Whether the access to Banks assets / services from within outside banks network is secured through VPN / standard technologies for in-transit data?",
        "7_A": "Whether the bank has a centralised authentication methodology to ensure transactions (including customer access credential) are only put through genuine authorised applications?",
        "7_B": "Has the bank implemented authentication framework/mechanism to securely verify and identify the banks applications to customers?",
        "7_C": "Are digital certificates being used to securely verify and identify banks application to customers?",
        "8_A": "Whether bank has subscribed to anti-phishing/anti-rogue application services from external service providers for identifying and taking down phishing websites/rogue applications?",
        "9_A": "Whether efforts are made to encourage employees to report suspicious behaviour and incidents to the incident management team?",
        "9_B": "Whether bank is conducting mandatory awareness programs on cyber security for newly recruited employees ?",
        "9_C": "Whether web-based quiz and training are conducted for all levels of management every year?",
        "9_D": "Whether arrangements are made to sensitise Board members on various technological and cyber security related developments periodically?",
        "10_A": "Whether the system captures audit logs of user actions?",
        "10_B": "Whether alert mechanism is activated to monitor changes in audit log settings?",
        "11_A": "Does the bank have written / documented incident response procedures including roles / staff / outsourced staff to handle cyber incidents?"
    }
    
    # Add "Level II" header row
    ws[f'A{current_row}'].value = ""
    ws[f'B{current_row}'].value = "Level II"
    ws[f'C{current_row}'].value = ""
    ws[f'D{current_row}'].value = ""
    
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}{current_row}']
        cell.font = data_font_bold
        cell.alignment = left_alignment
        cell.border = thin_border
    
    current_row += 1
    
    # Process each section
    for section_num, section_data in sections.items():
        # Add section header
        ws[f'A{current_row}'].value = section_num
        ws[f'B{current_row}'].value = section_data["title"]
        ws[f'C{current_row}'].value = ""
        ws[f'D{current_row}'].value = ""
        
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{current_row}']
            cell.font = data_font_bold
            cell.alignment = left_alignment if col == 'B' else center_alignment
            cell.border = thin_border
        
        current_row += 1
        
        # Add questions
        for q_id in section_data["questions"]:
            # Form data has keys like "q1_A", so add "q" prefix
            form_key = f"q{q_id}"
            answer = loc_data.get(form_key, "")
            
            # Extract letter from question ID (e.g., "1_A" -> "A")
            letter = q_id.split('_')[1]
            
            ws[f'A{current_row}'].value = letter
            ws[f'B{current_row}'].value = question_texts.get(q_id, "")
            
            # Set Yes/No/NA in column C
            if answer == "Yes":
                ws[f'C{current_row}'].value = "Yes"
            elif answer == "No":
                ws[f'C{current_row}'].value = "No"
            elif answer == "Not Applicable":
                ws[f'C{current_row}'].value = "NA"
            else:
                ws[f'C{current_row}'].value = ""
            
            # Set observation in column D
            observation = observations.get(q_id, {}).get(answer, "")
            ws[f'D{current_row}'].value = observation
            
            # Apply formatting
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].alignment = center_alignment
            ws[f'A{current_row}'].border = thin_border
            
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].alignment = left_alignment
            ws[f'B{current_row}'].border = thin_border
            
            ws[f'C{current_row}'].font = data_font
            ws[f'C{current_row}'].alignment = center_alignment
            ws[f'C{current_row}'].border = thin_border
            
            ws[f'D{current_row}'].font = data_font
            ws[f'D{current_row}'].alignment = left_alignment
            ws[f'D{current_row}'].border = thin_border
            
            current_row += 1
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name

@loc_level2_bp.route('/process_loc_level2', methods=['POST'])
def process_loc_level2():
    """
    Process LOC Level 2 form data and generate Excel
    """
    try:
        print("\n" + "="*80)
        print("🎯 LOC LEVEL 2 - List of Compliances")
        print("="*80)
        
        # Get form data
        form_data = request.form.to_dict()
        
        print("📋 Form Data Received:")
        for key, value in form_data.items():
            print(f"  {key}: {value}")
        
        # Clean up old LOC files before generating new one
        print("\n🧹 Cleaning up old LOC files...")
        cleanup_loc_files()
        
        # Generate Excel file
        print("\n📝 Generating Excel file...")
        excel_file_path = create_loc_level2_excel(form_data)
        print(f"✅ Excel file created: {excel_file_path}")
        
        # Save to static/uploads directory
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"LOC_Level2_{timestamp}.xlsx"
        static_upload_path = os.path.join('static', 'uploads', excel_filename)
        
        os.makedirs(os.path.dirname(static_upload_path), exist_ok=True)
        shutil.copy2(excel_file_path, static_upload_path)
        os.unlink(excel_file_path)
        
        print(f"✅ Excel saved to: {static_upload_path}")
        print("="*80)
        
        return jsonify({
            "success": True,
            "message": "LOC Level 2 data processed successfully",
            "download_url": f"/static/uploads/{excel_filename}",
            "excel_file": "LOC_Level2.xlsx"
        }), 200
        
    except Exception as e:
        print(f"❌ Error processing LOC Level 2: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"Error processing LOC Level 2: {str(e)}"
        }), 500

