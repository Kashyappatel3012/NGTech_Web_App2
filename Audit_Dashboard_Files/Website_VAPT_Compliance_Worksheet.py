from flask import Blueprint, request, jsonify, send_file
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from copy import copy

# Create blueprint for Website VAPT Compliance
website_vapt_compliance_bp = Blueprint('website_vapt_compliance', __name__)

def create_website_vapt_compliance_excel(form_data, excel_file_path):
    """Create Website VAPT Compliance Excel file"""
    try:
        print(f"🔍 DEBUG: Starting Website VAPT Compliance Excel creation...")
        
        # Get form data
        org_name = form_data.get('organizationName', '')
        if org_name == 'Other':
            org_name = form_data.get('organizationNameOther', '')
        
        report_id = form_data.get('reportId', '')
        report_date = form_data.get('reportDate', '')
        compliance_date = form_data.get('complianceDate', '')
        
        # Format dates
        try:
            report_date_obj = datetime.strptime(report_date, '%Y-%m-%d')
            report_date_formatted = report_date_obj.strftime('%d/%m/%Y')
        except:
            report_date_formatted = report_date
        
        try:
            compliance_date_obj = datetime.strptime(compliance_date, '%Y-%m-%d')
            compliance_date_formatted = compliance_date_obj.strftime('%d/%m/%Y')
        except:
            compliance_date_formatted = compliance_date
        
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Website VAPT"
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 23
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 45
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 55
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 40
        
        # Define styles
        header_font = Font(name='Times New Roman', size=12, bold=False)
        header_alignment_center = Alignment(horizontal='center', vertical='center')
        header_alignment_left = Alignment(horizontal='left', vertical='center')
        
        column_header_font = Font(name='Times New Roman', size=12, bold=True, color='FFFFFF')
        column_header_fill = PatternFill(start_color='FF1F497D', end_color='FF1F497D', fill_type='solid')
        column_header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Times New Roman', size=12)
        data_font_bold = Font(name='Times New Roman', size=12, bold=True)
        data_alignment_center = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin', color='FF000000'),
            right=Side(style='thin', color='FF000000'),
            top=Side(style='thin', color='FF000000'),
            bottom=Side(style='thin', color='FF000000')
        )
        
        # Row 1
        ws['A1'] = "Submitted To:"
        ws['A1'].font = header_font
        ws['A1'].alignment = header_alignment_center
        ws['A1'].border = thin_border
        
        ws.merge_cells('B1:I1')
        ws['B1'] = f"The Management, {org_name}"
        ws['B1'].font = header_font
        ws['B1'].alignment = header_alignment_left
        for col in range(2, 10):  # B to I
            ws.cell(row=1, column=col).border = thin_border
        
        # Row 2
        ws['A2'] = "Report Name:"
        ws['A2'].font = header_font
        ws['A2'].alignment = header_alignment_center
        ws['A2'].border = thin_border
        
        ws.merge_cells('B2:I2')
        ws['B2'] = f"Compliance Of Website VAPT Audit Report (Report ID no. {report_id} dated {report_date_formatted})"
        ws['B2'].font = header_font
        ws['B2'].alignment = header_alignment_left
        for col in range(2, 10):
            ws.cell(row=2, column=col).border = thin_border
        
        # Row 3
        ws['A3'] = "Confidentiality:"
        ws['A3'].font = header_font
        ws['A3'].alignment = header_alignment_center
        ws['A3'].border = thin_border
        
        ws.merge_cells('B3:I3')
        ws['B3'] = "Very High & Not for circulation"
        ws['B3'].font = header_font
        ws['B3'].alignment = header_alignment_left
        for col in range(2, 10):
            ws.cell(row=3, column=col).border = thin_border
        
        # Row 4
        ws['A4'] = "Compliance Date:"
        ws['A4'].font = header_font
        ws['A4'].alignment = header_alignment_center
        ws['A4'].border = thin_border
        
        ws.merge_cells('B4:I4')
        ws['B4'] = compliance_date_formatted
        ws['B4'].font = header_font
        ws['B4'].alignment = header_alignment_left
        for col in range(2, 10):
            ws.cell(row=4, column=col).border = thin_border
        
        # Row 5 - Column headers
        headers = [
            "Serial Number",
            "Website VAPT",
            "Sub Serial Number",
            "Reference Page No.",
            "vulnerability Name",
            "Risk Factor",
            "Compliance By Bank",
            "Status",
            "Pl attach evidence"
        ]
        
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = column_header_font
            cell.fill = column_header_fill
            cell.alignment = column_header_alignment
            cell.border = thin_border
        
        # Row 6 - Initial data
        ws['A6'] = "1"
        ws['A6'].font = data_font_bold
        ws['A6'].alignment = data_alignment_center
        ws['A6'].border = thin_border
        
        ws['B6'] = "Website VAPT"
        ws['B6'].font = data_font_bold
        ws['B6'].alignment = data_alignment_center
        ws['B6'].border = thin_border
        
        # Load user's Excel file
        print(f"📂 Loading user Excel file: {excel_file_path}")
        source_wb = load_workbook(excel_file_path)
        
        # Find the "Infra_VAPT" worksheet
        source_ws = None
        for sheet_name in source_wb.sheetnames:
            if sheet_name == "Website_VAPT":
                source_ws = source_wb[sheet_name]
                break
        
        if not source_ws:
            raise Exception("Worksheet 'Website_VAPT' not found in uploaded Excel file")
        
        print(f"✅ Found 'Website_VAPT' worksheet with {source_ws.max_row} rows")
        
        # Copy data from user's Excel
        # Starting from row 2 in source, row 6 in destination
        current_row = 6
        last_row_with_data = 5
        
        for source_row in range(2, source_ws.max_row + 1):
            # Get source cells
            source_a = source_ws.cell(row=source_row, column=1)  # A (Sub Serial)
            source_c = source_ws.cell(row=source_row, column=3)  # C (Vulnerability Name)
            source_d = source_ws.cell(row=source_row, column=4)  # D (Risk Factor)
            
            # Skip if C column is empty
            if not source_c.value:
                continue
            
            # C column (Sub Serial Number) - from user's A column
            dest_c = ws.cell(row=current_row, column=3)
            dest_c.value = source_a.value
            dest_c.font = data_font
            dest_c.alignment = data_alignment_center
            dest_c.border = thin_border
            
            # D column (Reference Page No.) - leave empty
            dest_d = ws.cell(row=current_row, column=4)
            dest_d.value = ""
            dest_d.font = data_font
            dest_d.alignment = data_alignment_center
            dest_d.border = thin_border
            
            # E column (Vulnerability Name) - from user's C column
            dest_e = ws.cell(row=current_row, column=5)
            dest_e.value = source_c.value
            dest_e.font = data_font
            dest_e.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            dest_e.border = thin_border
            
            # F column (Risk Factor) - from user's D column with formatting
            dest_f = ws.cell(row=current_row, column=6)
            dest_f.value = source_d.value
            
            # Copy formatting from source D
            if source_d.has_style:
                dest_f.font = copy(source_d.font)
                dest_f.fill = copy(source_d.fill)
                dest_f.alignment = copy(source_d.alignment)
            else:
                dest_f.font = data_font
                dest_f.alignment = data_alignment_center
            dest_f.border = thin_border
            
            # Ensure font is Times New Roman 12
            dest_f.font = Font(
                name='Times New Roman',
                size=12,
                bold=dest_f.font.bold if dest_f.font else False,
                color=dest_f.font.color if dest_f.font else None
            )
            
            # G, H, I columns - leave empty with borders
            for col in [7, 8, 9]:
                cell = ws.cell(row=current_row, column=col)
                cell.value = ""
                cell.font = data_font
                cell.border = thin_border
            
            last_row_with_data = current_row
            current_row += 1
        
        print(f"📊 Copied {current_row - 6} data rows")
        
        # Merge A6 to A(last_row) and B6 to B(last_row)
        if last_row_with_data > 6:
            print(f"🔗 Merging A6:A{last_row_with_data} and B6:B{last_row_with_data}")
            
            ws.merge_cells(f'A6:A{last_row_with_data}')
            ws.merge_cells(f'B6:B{last_row_with_data}')
            
            # Apply borders to all merged cells
            for row in range(6, last_row_with_data + 1):
                ws.cell(row=row, column=1).border = thin_border
                ws.cell(row=row, column=2).border = thin_border
        
        # Apply Times New Roman 12 to all cells
        print(f"🎨 Applying Times New Roman 12 to all cells...")
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, 10):  # A to I
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.font:
                    cell.font = Font(
                        name='Times New Roman',
                        size=12,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color,
                        underline=cell.font.underline,
                        strike=cell.font.strike
                    )
        
        print(f"✅ Website VAPT Compliance Excel created successfully")
        
        return wb
        
    except Exception as e:
        print(f"❌ Error creating Website VAPT Compliance Excel: {e}")
        import traceback
        traceback.print_exc()
        raise

@website_vapt_compliance_bp.route('/process_website_vapt_compliance', methods=['POST'])
def process_website_vapt_compliance():
    """Process Website VAPT Compliance form submission"""
    try:
        print("\n" + "="*80)
        print("🚀 Processing Website VAPT Compliance")
        print("="*80)
        
        # Get form data
        form_data = {
            'organizationName': request.form.get('organizationName'),
            'organizationNameOther': request.form.get('organizationNameOther'),
            'reportId': request.form.get('reportId'),
            'reportDate': request.form.get('reportDate'),
            'complianceDate': request.form.get('complianceDate')
        }
        
        # Handle file upload
        excel_file = request.files.get('excelFile')
        
        if not excel_file:
            return jsonify({'success': False, 'error': 'Excel file is required'}), 400
        
        # Save uploaded file temporarily
        upload_dir = os.path.join('static', 'uploads', 'temp')
        os.makedirs(upload_dir, exist_ok=True)
        
        excel_path = os.path.join(upload_dir, f'temp_website_vapt_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
        excel_file.save(excel_path)
        
        # Create the compliance Excel
        wb = create_website_vapt_compliance_excel(form_data, excel_path)
        
        # Save the final workbook
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Website_VAPT_Compliance_{timestamp}.xlsx'
        filepath = os.path.join('static', 'uploads', filename)
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        wb.save(filepath)
        
        # Clean up temporary file
        try:
            if os.path.exists(excel_path):
                os.remove(excel_path)
                print(f"   🗑️  Deleted temp file: {excel_path}")
        except Exception as e:
            print(f"   ⚠️  Warning: Could not delete temp file: {e}")
        
        print(f"\n✅ Website VAPT Compliance Excel created: {filename}")
        print("="*80)
        
        return jsonify({
            'success': True,
            'filename': filename,
            'download_url': f'/static/uploads/{filename}'
        })
    
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Clean up temporary file even if there's an error
        try:
            if 'excel_path' in locals() and os.path.exists(excel_path):
                os.remove(excel_path)
                print(f"   🗑️  Cleaned up temp file: {excel_path}")
        except Exception as cleanup_error:
            print(f"   ⚠️  Could not clean up temp file: {cleanup_error}")
        
        return jsonify({'success': False, 'error': str(e)}), 500

@website_vapt_compliance_bp.route('/cleanup_website_vapt_compliance', methods=['POST'])
def cleanup_website_vapt_compliance():
    """Clean up Website VAPT Compliance Excel file after download"""
    try:
        data = request.get_json()
        filename = data.get('filename')
        
        files_deleted = []
        
        if filename:
            # Delete the main file from static/uploads
            file_path = os.path.join('static', 'uploads', filename)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    files_deleted.append(file_path)
                    print(f"✅ Deleted: {file_path}")
                except Exception as e:
                    print(f"❌ Error deleting {file_path}: {e}")
        
        # Clean up old temp files
        temp_dir = os.path.join('static', 'uploads', 'temp')
        if os.path.exists(temp_dir):
            try:
                temp_files = [f for f in os.listdir(temp_dir) if f.startswith('temp_website_vapt_')]
                for temp_file in temp_files:
                    temp_file_path = os.path.join(temp_dir, temp_file)
                    try:
                        file_age = datetime.now().timestamp() - os.path.getmtime(temp_file_path)
                        if file_age > 300:  # 5 minutes
                            os.remove(temp_file_path)
                            files_deleted.append(temp_file_path)
                            print(f"✅ Deleted old temp file: {temp_file_path}")
                    except Exception as e:
                        print(f"❌ Error deleting temp file {temp_file_path}: {e}")
            except Exception as e:
                print(f"❌ Error cleaning temp directory: {e}")
        
        print(f"📊 Cleanup summary: {len(files_deleted)} files deleted")
        
        return jsonify({
            'success': True,
            'files_deleted': len(files_deleted)
        })
    except Exception as e:
        print(f"❌ Error cleaning up: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

