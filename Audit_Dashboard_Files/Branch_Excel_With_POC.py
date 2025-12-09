from flask import Blueprint, request, flash, redirect, url_for, send_file
from flask_login import login_required
from datetime import datetime
import os
import tempfile
import shutil
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.drawing.image import Image
import zipfile
import glob
import re

# Create blueprint
branch_excel_with_poc_bp = Blueprint('branch_excel_with_poc_bp', __name__)

# Allowed file extensions
ALLOWED_EXCEL_EXTENSIONS = {'xlsx', 'xls'}

# Mapping of audit questions in column B to image numbers
QUESTION_TEXT_TO_NUMBER = {
    "Do employees are using strong passwords?": 1,
    "Do you mandate periodical password changes?": 2,
    "Do passwords are shared among employees?": 3,
    "What happens after consecutive failed login attempts?": 4,
    "Is session timeout enforced after a pre-defined period of inactivity?": 5,
    "Whether Multiple Logins is enabled?": 6,
    "Is the CBS System isolated from the internet?": 7,
    "Is internet access restricted to only trusted and officially approved sites?": 8,
    "Is there any policy for default password while unlocking the account?": 9,
    "When a new user joins then how do you give the user ID and Password for the user by using mail or telephonic conversation?": 10,
    "SSL is using or not in CBS Application?": 11,
    "CBS is browser based or desktop based?": 12,
    "Is the CBS compatible with the latest version of the browser?": 13,
    "Whether Two Factor Authentication is implemented for CBS login?": 14,
    "CCTV Camera Present?": 15,
    "Branch Locker CCTV Camera Present?": 16,
    "CCTV camera covering complete branch area?": 17,
    "Check DVR working properly?": 18,
    "Is NTP configured in the DVR?": 19,
    "DVR is available in bank network or not?": 20,
    "Camera is working in night vision.": 21,
    "CCTV camera history recording duration.": 22,
    "Locker CCTV camera history recording duration.": 23,
    "Biometric Devices or Proximity Card is available or not?": 24,
    "ID card issues to the Employee?": 25,
    "Whether guard is available at branch or not?": 26,
    "Are secure areas controlled?": 27,
    "Smoke Detector is installed in the branch?": 28,
    "Whether Panic Switch installed?": 29,
    "If a panic switch or smoke detector is installed at the branch or head office does the notification reach higher authorities when triggered?": 30,
    "LAN ports are open or not?": 31,
    "Visitor Register is maintained or not?": 32,
    "Any record maintains for drill the environmental controls?": 33,
    "Whether external modems, data cards etc., are being used in the Branch to access internet?": 34,
    "Fire extinguisher is available or not?": 35,
    "Enough fire-extinguisher is available for area?": 36,
    "Does the Bank have made enough Battery Backup which supports Computer Systems?": 37,
    "Do the AMC is given for power backup systems?": 38,
    "Generator is available or not?": 39,
    "Is training conducted for end user with respect to the Information Security?": 40,
    "Is training conducted for end user with respect to the application usage?": 41,
    "Employees are aware about using of fire extinguisher & Panic Switch?": 42,
    "User was aware about checking CCTV recording history?": 43,
    "Is AMC is given for computer and related systems?": 44,
    "Is Complaint register maintained?": 45,
    "Antivirus available in all the desktop?": 46,
    "Antivirus updated in all the desktop?": 47,
    "Windows patch updated in all the system?": 48,
    "Are there any system on which outdated version of windows is installed?": 49,
    "Expired License antivirus installed?": 50,
    "Are controls in place to ensure users only have access to the network resources they have been specially authorized to use and are required for their duties?": 51,
    "Is there network diagram available?": 52,
    "Is structured cabling is observed?": 53,
    "Is cable tagging is observed?": 54,
    "Is provision made for backup network?": 55,
    "Networking devices are under secure location and there is any cabinet facility is provided or not?": 56,
    "Proper Cooling for Networking Devices available or not?": 57,
    "Wi-Fi is available or not?": 58,
    "Whether Firewall is available in Branch ?": 59,
    "Does User login with Administrator Rights?": 60,
    "Can Group Policy be modified?": 61,
    "Can Security Configuration Policy be modified?": 62,
    "Whether Proxy can be modified by branch users?": 63,
    "Whether internet is allowed or not?": 64,
    "Whether internet access is restricted or unrestricted?": 65,
    "Authentication Policy for internet is available or not?": 66,
    "System Password is weak or not?": 67,
    "System tagging is available or not?": 68,
    "Is NTP configured in the systems?": 69,
    "Whether USB ports/CD drives enabled in the System?": 70,
    "USB Approval process?": 71,
    "If bank have trusted USB, then check the USB open outside the network?": 72,
    "Whether Firewall is enable in systems?": 73,
    "Whether all external storage media are checked for virus?": 74,
    "Whether Remote Desktop Protocol is enabled or disabled?": 75,
    "Check the CCTV camera  LAN cable and ATM machine power cable concealed?": 76,
    "Guard is available or not?": 77,
    "CCTV Camera and check the position of camera?": 78,
    "ATM is working 24*7?": 79,
    "Whether the Do's and Don't for the user awareness is available or not?": 80,
    "Whether the ATM is network segmented or not?": 81,
    "Whether ATM Machine is properly grouted (floor and wall) ?": 82,
    "Whether access to ATM room for maintenance purpose is restricted to the authorized persons only?": 83,
    "Whether ATM power back up is supported by UPS.": 84,
    "System of ATM cash replenishment, adherence to dual control mechanism records is maintained properly?": 85,
    "Check system of ATM reconciliation status maintained?": 86,
    "Whether ATM machine preservation is done of journal print for future reference?": 87,
    "Personal Mail used or not?": 88,
    "Two factor authentications are in used or not?": 89,
    "Single mail with multiple users?": 90,
    "Whether official email ID gets open outside bank N/w?": 91,
    "Which Department is using Remote Access?": 92,
    "Remote Access approval process?": 93,
    "There are no unauthorized applications on any systems.": 94,
    "There are no Personal Data on any systems.": 95,
    "Auditor Identity was not verified.": 96,
    "Windows was not activated.": 97,
    "Password was saved in browser.": 98,
    "Password was written on wall, desk, notepad, diary.": 99,
    "Asset Movement Register was not available.": 100,
    "Dust was present on the network devices & systems.": 101,
    "Auto switchable mode was not available for connectivity.": 102,
    "Preventive Maintenance is not carried out for Systems/UPS/Network Devices.": 103
}

QUESTION_TEXT_CANONICAL = {text.lower(): text for text in QUESTION_TEXT_TO_NUMBER}

def allowed_file(filename, allowed_extensions):
    """Check if the uploaded file has an allowed extension"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in allowed_extensions

def get_question_number_from_text(question_text):
    """
    Determine the question number based on the text found in column B.
    Returns a tuple of (question_number, canonical_text) where canonical_text
    is the exact mapping key that matched (useful for logging partial matches).
    """
    if question_text is None:
        return None, None
    
    text = str(question_text).strip()
    if not text:
        return None, None
    
    # Exact match
    if text in QUESTION_TEXT_TO_NUMBER:
        return QUESTION_TEXT_TO_NUMBER[text], text
    
    text_lower = text.lower()
    
    # Case-insensitive exact match
    if text_lower in QUESTION_TEXT_CANONICAL:
        canonical_text = QUESTION_TEXT_CANONICAL[text_lower]
        return QUESTION_TEXT_TO_NUMBER[canonical_text], canonical_text
    
    # Partial match (either contains or is contained by canonical text)
    for canonical_text, number in QUESTION_TEXT_TO_NUMBER.items():
        canonical_lower = canonical_text.lower()
        if canonical_lower in text_lower or text_lower in canonical_lower:
            return number, canonical_text
    
    return None, None

def extract_and_list_images_from_zip(zip_file_path, extract_to_dir):
    """Extract ZIP file and list all image files contained within it"""
    try:
        image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp', '.svg'}
        image_files = []
        extracted_images = {}  # Dictionary to store images by base number and column
        
        with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            # Get list of all files in the ZIP
            file_list = zip_ref.namelist()
            
            print(f"Total files in ZIP: {len(file_list)}")
            print("="*50)
            print("FILES IN ZIP ARCHIVE:")
            print("="*50)
            
            for file_name in file_list:
                print(f"- {file_name}")
                
                # Check if file is an image
                file_extension = os.path.splitext(file_name)[1].lower()
                if file_extension in image_extensions:
                    image_files.append(file_name)
                    
                    # Extract image to temporary directory
                    try:
                        # Get just the filename without path
                        filename = os.path.basename(file_name)
                        # Remove extension to get the number
                        name_without_ext = os.path.splitext(filename)[0]
                        
                        # Extract base number (ignore everything after _, -, or space)
                        base_number = extract_base_number(name_without_ext)
                        
                        if base_number:
                            extracted_path = os.path.join(extract_to_dir, filename)
                            
                            # Extract the file
                            with zip_ref.open(file_name) as source, open(extracted_path, 'wb') as target:
                                target.write(source.read())
                            
                            # Store image with base number as key
                            if base_number not in extracted_images:
                                extracted_images[base_number] = []
                            extracted_images[base_number].append(extracted_path)
                            print(f"Extracted image {base_number}: {filename}")
                            
                    except Exception as e:
                        print(f"Error extracting {file_name}: {e}")
            
            print("="*50)
            print(f"IMAGE FILES FOUND: {len(image_files)}")
            print(f"EXTRACTED NUMBERED IMAGES: {sum(len(images) for images in extracted_images.values())}")
            print("="*50)
            
            if extracted_images:
                for base_num, images in sorted(extracted_images.items()):
                    print(f"Base number {base_num}: {len(images)} images")
                    for i, path in enumerate(images):
                        print(f"  - Image {i+1}: {os.path.basename(path)}")
            else:
                print("No numbered image files found in the ZIP archive.")
            
            print("="*50)
            
        return extracted_images
        
    except Exception as e:
        print(f"Error extracting ZIP file: {str(e)}")
        return {}

def extract_base_number(filename):
    """Extract base number from filename, ignoring everything after _, -, or space"""
    try:
        # Remove everything after _, -, or space
        base_part = filename.split('_')[0].split('-')[0].split(' ')[0]
        return int(base_part)
    except (ValueError, IndexError):
        return None

def attach_evidence_images_to_excel(excel_path, evidence_folder="static/uploads", extracted_images_dict=None):
    """
    Attach evidence images to Excel file based on question text in column B.
    Special handling for questions 94 and 95 based on specific text patterns in column B:
    - "Unauthorized applications have been found in the below systems" → image 94
    - "Personal Data have been found in the below systems" → image 95
    - "There are no Personal Data on any systems." → image 95
    Images are matched by number (1-95) and inserted into columns H, I, G.
    
    Args:
        excel_path: Path to the Excel file
        evidence_folder: Folder containing evidence images (default: "static/uploads")
        extracted_images_dict: Dictionary of extracted images from ZIP file (key: question number, value: list of image paths)
    """
    try:
        # Load the workbook
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Get all image files from evidence folder
        all_images = []
        if os.path.exists(evidence_folder):
            image_extensions = ['*.jpg', '*.jpeg', '*.png', '*.gif', '*.bmp', '*.tiff']
            for ext in image_extensions:
                all_images.extend(glob.glob(os.path.join(evidence_folder, ext)))
                all_images.extend(glob.glob(os.path.join(evidence_folder, ext.upper())))
            print(f"Found {len(all_images)} total images in {evidence_folder}")
        else:
            print(f"Evidence folder not found: {evidence_folder}")
        
        # Also add images from extracted_images_dict if provided
        if extracted_images_dict:
            for question_num, image_paths in extracted_images_dict.items():
                all_images.extend(image_paths)
            print(f"Added {sum(len(images) for images in extracted_images_dict.values())} images from ZIP file")
        
        print(f"Total images available: {len(all_images)}")
        
        # Helper function to find and insert images for a given question number
        def find_and_insert_images(question_number, row_idx):
            """Find images matching question number and insert them into the row"""
            nonlocal images_inserted
            matching_images = []
            number_str = str(question_number)
            
            print(f"🔍 Searching for images matching question {question_number} (pattern: '{number_str}')")
            print(f"   Total images available: {len(all_images)}")
            
            for img_path in all_images:
                img_filename = os.path.basename(img_path)
                img_name_without_ext = os.path.splitext(img_filename)[0].strip()
                
                # Match patterns (same logic as current POC image matching):
                # - Exact match: "1" matches "1.jpg", "1.png"
                # - With underscore: "1_" matches "1_bvkjvb.png", "1_anything.ong"
                # - With any non-digit character after number: "1abc.jpg" matches but "10.jpg" does NOT match for question 1
                
                if img_name_without_ext == number_str:
                    # Exact match: "1" -> matches "1.jpg"
                    matching_images.append(img_path)
                    print(f"   ✅ Matched (exact): {img_filename}")
                elif img_name_without_ext.startswith(number_str + '_'):
                    # Starts with number and underscore: "1_" -> matches "1_bvkjvb.png"
                    matching_images.append(img_path)
                    print(f"   ✅ Matched (underscore): {img_filename}")
                elif len(img_name_without_ext) > len(number_str) and img_name_without_ext.startswith(number_str):
                    # Starts with number - check next character to avoid matching longer numbers
                    next_char = img_name_without_ext[len(number_str)]
                    if not next_char.isdigit():
                        # Next character is not a digit, so it's a valid match (e.g., "1abc.jpg")
                        matching_images.append(img_path)
                        print(f"   ✅ Matched (prefix): {img_filename}")
            
            print(f"   Found {len(matching_images)} matching images for question {question_number}")
            
            # Sort images for consistent ordering
            matching_images.sort()
            
            # Insert images in columns H (8), I (9), G (7) - maximum 3 images
            if matching_images:
                image_columns = [8, 9, 7]  # H, I, G columns
                num_images_to_insert = min(len(matching_images), 3)
                
                for img_idx in range(num_images_to_insert):
                    try:
                        img_path = matching_images[img_idx]
                        col_idx = image_columns[img_idx]
                        
                        # Load and resize image
                        img = Image(img_path)
                        target_height = 25
                        aspect_ratio = img.width / img.height if img.height else 1
                        img.height = target_height
                        img.width = int(target_height * aspect_ratio)
                        
                        # Get cell reference (e.g., "H12")
                        from openpyxl.utils import get_column_letter
                        col_letter = get_column_letter(col_idx)
                        cell_ref = f"{col_letter}{row_idx}"
                        
                        # Insert image
                        ws.add_image(img, cell_ref)
                        images_inserted += 1
                        print(f"✅ Inserted evidence image {img_idx + 1} at {cell_ref} for question {question_number} (row {row_idx})")
                        
                    except Exception as e:
                        print(f"⚠️ Error inserting image at column {col_idx}, row {row_idx}: {e}")
            else:
                print(f"⚠️ No matching images found for question {question_number} at row {row_idx}")
                # Debug: print available image names
                if all_images:
                    print(f"   Available images: {[os.path.basename(img) for img in all_images[:10]]}")  # Show first 10
        
        # Iterate through rows and check column B for question matches
        images_inserted = 0
        for row_idx in range(1, ws.max_row + 1):
            # Get cell B value (handles merged cells automatically)
            cell_b = ws.cell(row=row_idx, column=2)  # Column B
            if not cell_b.value:
                continue
            
            question_text = str(cell_b.value).strip()
            
            # Special handling: Check column B for specific text patterns first
            # Check for "Unauthorized applications have been found in the below systems" in column B
            if "Unauthorized applications have been found in the below systems" in question_text:
                print(f"🔍 Found special text in column B for question 94 at row {row_idx}: '{question_text[:60]}...'")
                find_and_insert_images(94, row_idx)
                continue  # Skip normal question matching for this row
            
            # Check for "Personal Data have been found in the below systems" in column B
            if "Personal Data have been found in the below systems" in question_text:
                print(f"🔍 Found special text in column B for question 95 at row {row_idx}: '{question_text[:60]}...'")
                find_and_insert_images(95, row_idx)
                continue  # Skip normal question matching for this row
            
            # Check for "There are no Personal Data on any systems." in column B
            if "There are no Personal Data on any systems." in question_text:
                print(f"🔍 Found special text in column B for question 95 at row {row_idx}: '{question_text}'")
                find_and_insert_images(95, row_idx)
                continue  # Skip normal question matching for this row
            
            # Normal question matching from QUESTION_TEXT_TO_NUMBER mapping
            question_number, canonical_text = get_question_number_from_text(question_text)

            if not question_number:
                continue
            
            if canonical_text and question_text != canonical_text:
                print(f"⚠️ Partial match: '{question_text}' matched to question {question_number}: '{canonical_text}'")
            
            # Find and insert images for this question number
            find_and_insert_images(question_number, row_idx)
        
        # Save the workbook with images
        wb.save(excel_path)
        print(f"✅ Attached {images_inserted} evidence images to Excel file")
        
    except Exception as e:
        print(f"❌ Error attaching evidence images: {e}")
        import traceback
        traceback.print_exc()

def insert_images_to_excel(excel_path, image_mapping):
    """Insert images into Excel cells based on column B question text"""
    try:
        # Load the existing workbook
        wb = load_workbook(excel_path)
        
        # Ensure image lists are sorted for consistent processing
        for number, images in image_mapping.items():
            images.sort()
        
        total_images_inserted = 0
        
        # Process each worksheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            print(f"Processing images for sheet: {sheet_name}")
            
            max_rows = ws.max_row
            for row_idx in range(1, max_rows + 1):
                question_value = ws.cell(row=row_idx, column=2).value  # Column B
                question_number, canonical_text = get_question_number_from_text(question_value)
                
                if not question_number:
                    continue
                
                if canonical_text and str(question_value).strip() != canonical_text:
                    print(f"⚠️ Partial match: '{question_value}' matched to question {question_number}: '{canonical_text}' (row {row_idx})")
                
                if question_number not in image_mapping:
                    continue
                
                image_paths = image_mapping[question_number]
                if not image_paths:
                    continue
                
                columns = ['H', 'I', 'G']
                images_to_process = image_paths[:3]
                
                for idx, image_path in enumerate(images_to_process):
                    target_column = columns[idx]
                    target_cell = f"{target_column}{row_idx}"
                    
                    try:
                        img = Image(image_path)
                        original_width = img.width
                        original_height = img.height
                        
                        target_height = 25
                        aspect_ratio = original_width / original_height if original_height else 1
                        img.height = target_height
                        img.width = int(target_height * aspect_ratio)
                        
                        ws.add_image(img, target_cell)
                        total_images_inserted += 1
                        
                        print(f"Inserted image {question_number}_{idx + 1} into {target_cell} (size: {img.width}x{img.height}, original: {original_width}x{original_height})")
                    
                    except Exception as e:
                        print(f"Error inserting image {question_number}_{idx + 1} into {target_cell}: {e}")
                
                # Remove used images so duplicates aren't inserted elsewhere
                del image_paths[:len(images_to_process)]
        
        # Save the workbook
        wb.save(excel_path)
        print(f"Successfully inserted {total_images_inserted} images into Excel file")
        
    except Exception as e:
        print(f"Error inserting images into Excel: {e}")
        raise e

def process_excel_file(file_path):
    """Process the Excel file to add column G with borders based on column F data"""
    try:
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Process each worksheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Find the last row with data in column F
            last_row_with_data = 0
            for row in range(1, worksheet.max_row + 1):
                cell_f = worksheet[f'F{row}']
                if cell_f.value is not None and str(cell_f.value).strip() != '':
                    last_row_with_data = row
            
            print(f"Last row with data in column F: {last_row_with_data}")
            
            # If there's data in column F, add columns G, H, and I with borders
            if last_row_with_data > 0:
                # Insert new columns G, H, and I (this will shift existing columns to the right)
                worksheet.insert_cols(7, 3)  # Insert 3 columns starting at position 7 (columns G, H, I)
                
                # Define font style for header
                header_font = Font(
                    name='Times New Roman',
                    size=12,
                    bold=True
                )
                
                # Define alignment for header
                header_alignment = Alignment(
                    horizontal='center',
                    vertical='center'
                )
                
                # Define border styles
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Border for G column (no right border)
                g_border = Border(
                    left=Side(style='thin'),
                    right=Side(style=None),  # No right border
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Border for H column (no left or right border)
                h_border = Border(
                    left=Side(style=None),   # No left border
                    right=Side(style=None),  # No right border
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Border for I column (no left border)
                i_border = Border(
                    left=Side(style=None),  # No left border
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # No border style
                no_border = Border(
                    left=Side(style=None),
                    right=Side(style=None),
                    top=Side(style=None),
                    bottom=Side(style=None)
                )
                
                # Merge cells G1, H1, and I1 for the header
                worksheet.merge_cells('G1:I1')
                merged_cell = worksheet['G1']
                merged_cell.value = 'POC'  # Header
                merged_cell.font = header_font
                merged_cell.alignment = header_alignment
                merged_cell.border = thin_border
                
                # Apply borders and formatting to columns G, H, and I from row 1 to last_row_with_data
                for row in range(1, last_row_with_data + 1):
                    for col_letter in ['G', 'H', 'I']:
                        cell = worksheet[f'{col_letter}{row}']
                        
                        # Apply appropriate border style based on column
                        if col_letter == 'G':
                            cell.border = g_border  # No right border
                        elif col_letter == 'H':
                            cell.border = h_border  # No right border
                        elif col_letter == 'I':
                            cell.border = i_border  # Full border
                        
                        # Add empty content to non-header cells
                        if row > 1:
                            cell.value = ''
                
                # Set column widths for G, H, and I to 12
                for col_letter in ['G', 'H', 'I']:
                    worksheet.column_dimensions[col_letter].width = 12
                
                # Increase all row heights by 10 points
                for row in range(1, worksheet.max_row + 1):
                    current_height = worksheet.row_dimensions[row].height
                    if current_height is None:
                        # Default height is usually around 15 points
                        new_height = 30
                    else:
                        new_height = current_height + 10
                    
                    worksheet.row_dimensions[row].height = new_height
                
                print(f"Added columns G, H, and I with merged header 'POC' in G1:I1, Times New Roman 12pt bold centered, borders from G1 to I{last_row_with_data}, width set to 12")
                print(f"Increased all row heights by 10 points (total rows: {worksheet.max_row})")
        
        # Create a new temporary file for the processed workbook
        processed_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        processed_temp_file.close()
        
        # Save the processed workbook
        workbook.save(processed_temp_file.name)
        workbook.close()
        
        print(f"Processed Excel file saved to: {processed_temp_file.name}")
        return processed_temp_file.name
        
    except Exception as e:
        print(f"Error processing Excel file: {str(e)}")
        raise e

@branch_excel_with_poc_bp.route('/test_excel_with_poc_route')
def test_excel_with_poc_route():
    return "Excel with POC route is working!"

@branch_excel_with_poc_bp.route('/process_branch_excel_with_poc', methods=['POST'])
@login_required
def process_branch_excel_with_poc():
    print("="*50)
    print("BRANCH EXCEL WITH POC FORM SUBMISSION RECEIVED!")
    print("="*50)
    
    try:
        # Get form data for filename generation
        sr_no = request.form.get('srNo', '')
        branch_name = request.form.get('branchName', '')
        
        # Print form data to console
        print(f"Sr. No: {sr_no}")
        print(f"Branch Name: {branch_name}")
        
        # Check if both files are present in the request
        if 'excelFile' not in request.files or 'zipFile' not in request.files:
            flash('Both Excel file and ZIP file are required!', 'error')
            return redirect(url_for('audit_dashboard'))
        
        excel_file = request.files['excelFile']
        zip_file = request.files['zipFile']
        
        # Check if files are selected
        if excel_file.filename == '' or zip_file.filename == '':
            flash('Please select both Excel file and ZIP file!', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Validate Excel file extension
        if not allowed_file(excel_file.filename, ALLOWED_EXCEL_EXTENSIONS):
            flash('Invalid Excel file format! Please upload .xlsx or .xls files only.', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Validate ZIP file extension
        if not allowed_file(zip_file.filename, {'zip'}):
            flash('Invalid ZIP file format! Please upload .zip files only.', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Create temporary files that won't be automatically deleted
        temp_excel_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_excel_file.close()  # Close the file handle to avoid locking issues
        
        temp_zip_file = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        temp_zip_file.close()  # Close the file handle to avoid locking issues
        
        try:
            # Secure the filenames
            excel_filename = secure_filename(excel_file.filename)
            zip_filename = secure_filename(zip_file.filename)
            
            # Save uploaded files to temporary files
            excel_file.save(temp_excel_file.name)
            zip_file.save(temp_zip_file.name)
            
            print(f"Excel file saved to: {temp_excel_file.name}")
            print(f"ZIP file saved to: {temp_zip_file.name}")
            
            # Create temporary directory for extracted images
            temp_images_dir = tempfile.mkdtemp()
            
            # Extract and list images from ZIP file
            print("\n" + "="*60)
            print("PROCESSING ZIP FILE FOR IMAGES")
            print("="*60)
            extracted_images = extract_and_list_images_from_zip(temp_zip_file.name, temp_images_dir)
            
            # Process the Excel file to add column G with borders
            processed_file_path = process_excel_file(temp_excel_file.name)
            
            # Insert images into Excel if any were extracted
            if extracted_images:
                print("\n" + "="*60)
                print("INSERTING IMAGES INTO EXCEL")
                print("="*60)
                
                # Insert images into Excel by matching column B text
                insert_images_to_excel(processed_file_path, extracted_images)
            else:
                print("No numbered images found to insert into Excel.")
            
            # Attach evidence images based on column B question text
            print("\n" + "="*60)
            print("ATTACHING EVIDENCE IMAGES FROM FOLDER")
            print("="*60)
            attach_evidence_images_to_excel(processed_file_path, evidence_folder="static/uploads", extracted_images_dict=extracted_images)
            
            # Generate filename using Sr. No. and branch name (preserving spaces, no underscores)
            if branch_name and branch_name.strip():
                # Clean the branch name for filename (remove only invalid filename characters, preserve spaces)
                clean_branch_name = branch_name.strip()
                # Replace only invalid filename characters (not spaces) with nothing or remove them
                clean_branch_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '', clean_branch_name)
                
                # Add Sr. No. before branch name if provided
                if sr_no and sr_no.strip():
                    clean_sr_no = sr_no.strip()
                    # Clean Sr. No. for filename (remove only invalid characters, preserve spaces)
                    clean_sr_no = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '', clean_sr_no)
                    output_filename = f'{clean_sr_no} {clean_branch_name}.xlsx'
                else:
                    output_filename = f'{clean_branch_name}.xlsx'
            else:
                # Fallback to original Excel filename if branch name not provided
                # Remove underscores and restore spaces (if original had spaces)
                output_filename = excel_file.filename.replace('_', ' ') if excel_file.filename else 'Branch_Excel_With_POC.xlsx'
            
            print(f"Generated download filename: {output_filename}")
            
            # Verify the file was created successfully
            if os.path.exists(processed_file_path):
                print(f"\nFile processed successfully: {output_filename}")
                print(f"Found {len(extracted_images)} numbered image files in ZIP archive")
                
                # Send the file for download
                response = send_file(
                    processed_file_path,
                    as_attachment=True,
                    download_name=output_filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
                # Schedule the temporary files for deletion after response is sent
                def cleanup_temp_files():
                    try:
                        if os.path.exists(temp_excel_file.name):
                            os.unlink(temp_excel_file.name)
                        if os.path.exists(temp_zip_file.name):
                            os.unlink(temp_zip_file.name)
                        if os.path.exists(processed_file_path):
                            os.unlink(processed_file_path)
                        # Clean up temporary images directory
                        if 'temp_images_dir' in locals() and os.path.exists(temp_images_dir):
                            shutil.rmtree(temp_images_dir)
                    except Exception as e:
                        print(f"Error cleaning up temp files: {e}")
                
                # Use response callback to clean up the files after sending
                response.call_on_close(cleanup_temp_files)
                
                return response
            else:
                flash('Error processing the Excel file!', 'error')
                return redirect(url_for('audit_dashboard'))
                
        except Exception as e:
            # Clean up the temporary files if there was an error
            try:
                if os.path.exists(temp_excel_file.name):
                    os.unlink(temp_excel_file.name)
                if os.path.exists(temp_zip_file.name):
                    os.unlink(temp_zip_file.name)
                # Also clean up processed file if it exists
                if 'processed_file_path' in locals() and os.path.exists(processed_file_path):
                    os.unlink(processed_file_path)
                # Clean up temporary images directory
                if 'temp_images_dir' in locals() and os.path.exists(temp_images_dir):
                    shutil.rmtree(temp_images_dir)
            except:
                pass
            raise e
    
    except Exception as e:
        error_msg = str(e)
        print(f"Error processing file: {error_msg}")
        
        # Check if it's a 413 Request Entity Too Large error
        if '413' in error_msg or 'Request Entity Too Large' in error_msg or 'MAX_CONTENT_LENGTH' in error_msg:
            flash('File size too large! The maximum file size is 1GB. Please compress your ZIP file or split it into smaller files.', 'error')
        else:
            flash(f'Error processing file: {error_msg}', 'error')
        return redirect(url_for('audit_dashboard'))

