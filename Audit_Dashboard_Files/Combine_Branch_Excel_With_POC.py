from flask import Blueprint, request, flash, redirect, url_for, Response
from flask_login import login_required
from datetime import datetime
import os
import tempfile
import shutil
from werkzeug.utils import secure_filename
import zipfile
from openpyxl import load_workbook, Workbook
from io import BytesIO

# Create blueprint
combine_branch_excel_with_poc_bp = Blueprint('combine_branch_excel_with_poc_bp', __name__)

# Allowed file extensions
ALLOWED_ZIP_EXTENSIONS = {'zip'}

def allowed_file(filename, allowed_extensions):
    """Check if the uploaded file has an allowed extension"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in allowed_extensions

def extract_and_list_files_from_zip(zip_file_path, extract_to_dir):
    """Extract ZIP file and list all files contained within it"""
    try:
        extracted_files = []
        excel_files = []
        
        with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            # Get list of all files in the ZIP
            file_list = zip_ref.namelist()
            
            print(f"Total files in ZIP: {len(file_list)}")
            print("="*50)
            print("FILES IN ZIP ARCHIVE:")
            print("="*50)
            
            for file_name in file_list:
                print(f"- {file_name}")
                
                # Extract file to temporary directory
                try:
                    # Get the filename without path
                    filename = os.path.basename(file_name)
                    
                    # Skip directories (entries ending with /)
                    if not file_name.endswith('/'):
                        extracted_path = os.path.join(extract_to_dir, filename)
                        
                        # Create directory if it doesn't exist
                        os.makedirs(os.path.dirname(extracted_path), exist_ok=True)
                        
                        # Extract the file
                        with zip_ref.open(file_name) as source, open(extracted_path, 'wb') as target:
                            target.write(source.read())
                        
                        file_info = {
                            'original_name': file_name,
                            'extracted_path': extracted_path,
                            'filename': filename
                        }
                        
                        extracted_files.append(file_info)
                        
                        # Check if it's an Excel file
                        if filename.lower().endswith('.xlsx'):
                            excel_files.append(file_info)
                        
                        print(f"Extracted: {filename}")
                        
                except Exception as e:
                    print(f"Error extracting {file_name}: {e}")
            
            print("="*50)
            print(f"TOTAL FILES EXTRACTED: {len(extracted_files)}")
            print(f"EXCEL FILES FOUND: {len(excel_files)}")
            print("="*50)
            
            if extracted_files:
                print("EXTRACTED FILES LIST:")
                print("="*50)
                for i, file_info in enumerate(extracted_files, 1):
                    print(f"{i}. {file_info['filename']} (from: {file_info['original_name']})")
                print("="*50)
                
                if excel_files:
                    print("EXCEL FILES FOR PROCESSING:")
                    print("="*50)
                    for i, file_info in enumerate(excel_files, 1):
                        print(f"{i}. {file_info['filename']}")
                    print("="*50)
            else:
                print("No files were extracted from the ZIP archive.")
            
            print("="*50)
            
        return extracted_files, excel_files
        
    except Exception as e:
        print(f"Error extracting ZIP file: {str(e)}")
        return [], []

def combine_excel_files_to_bytes(excel_files):
    """Combine all Excel files into one workbook and return as bytes"""
    try:
        # Create a new workbook
        combined_wb = Workbook()
        
        # Remove the default sheet
        combined_wb.remove(combined_wb.active)
        
        total_sheets_added = 0
        
        print("="*60)
        print("COMBINING EXCEL FILES TO MEMORY")
        print("="*60)
        
        for excel_file in excel_files:
            file_path = excel_file['extracted_path']
            filename = excel_file['filename']
            
            # Clean filename for sheet name (Excel sheet names have restrictions)
            sheet_name = os.path.splitext(filename)[0]  # Remove .xlsx extension
            # Clean sheet name to avoid Excel issues
            sheet_name = sheet_name[:31]  # Max 31 characters
            sheet_name = sheet_name.replace('[', '').replace(']', '').replace('*', '').replace('?', '').replace('\\', '').replace('/', '')
            if not sheet_name:
                sheet_name = f"Sheet_{total_sheets_added + 1}"
            
            try:
                print(f"Processing: {filename}")
                
                # Load the source workbook with data_only=False to preserve formulas
                source_wb = load_workbook(file_path, data_only=False)
                
                # Copy each worksheet from the source workbook
                for source_sheet_name in source_wb.sheetnames:
                    source_sheet = source_wb[source_sheet_name]
                    
                    # Create a new worksheet in the combined workbook
                    new_sheet = combined_wb.create_sheet(title=sheet_name)
                    
                    # Copy all data from source sheet to new sheet
                    for row in source_sheet.iter_rows():
                        for cell in row:
                            new_cell = new_sheet.cell(row=cell.row, column=cell.column)
                            
                            # Copy cell value safely
                            try:
                                new_cell.value = cell.value
                            except Exception as value_error:
                                print(f"Warning: Could not copy value from cell {cell.coordinate}: {value_error}")
                                new_cell.value = str(cell.value) if cell.value is not None else None
                            
                            # Copy cell formatting if it exists
                            try:
                                if cell.has_style:
                                    if cell.font:
                                        new_cell.font = cell.font.copy()
                                    if cell.border:
                                        new_cell.border = cell.border.copy()
                                    if cell.fill:
                                        new_cell.fill = cell.fill.copy()
                                    if cell.number_format:
                                        new_cell.number_format = cell.number_format
                                    if cell.protection:
                                        new_cell.protection = cell.protection.copy()
                                    if cell.alignment:
                                        new_cell.alignment = cell.alignment.copy()
                            except Exception as style_error:
                                print(f"Warning: Could not copy formatting from cell {cell.coordinate}: {style_error}")
                    
                    # Copy column dimensions safely
                    try:
                        for col_letter, col_dimension in source_sheet.column_dimensions.items():
                            if col_dimension.width:
                                new_sheet.column_dimensions[col_letter].width = col_dimension.width
                            new_sheet.column_dimensions[col_letter].hidden = col_dimension.hidden
                            new_sheet.column_dimensions[col_letter].auto_size = col_dimension.auto_size
                    except Exception as col_error:
                        print(f"Warning: Could not copy column dimensions: {col_error}")
                    
                    # Copy row dimensions safely
                    try:
                        for row_num, row_dimension in source_sheet.row_dimensions.items():
                            if row_dimension.height:
                                new_sheet.row_dimensions[row_num].height = row_dimension.height
                            new_sheet.row_dimensions[row_num].hidden = row_dimension.hidden
                    except Exception as row_error:
                        print(f"Warning: Could not copy row dimensions: {row_error}")
                    
                    # Copy merged cells safely
                    try:
                        for merged_range in source_sheet.merged_cells.ranges:
                            new_sheet.merge_cells(str(merged_range))
                    except Exception as merge_error:
                        print(f"Warning: Could not copy merged cells: {merge_error}")
                    
                    # Copy images if any
                    try:
                        if hasattr(source_sheet, '_images') and source_sheet._images:
                            for image in source_sheet._images:
                                new_sheet.add_image(image)
                    except Exception as image_error:
                        print(f"Warning: Could not copy images: {image_error}")
                    
                    total_sheets_added += 1
                    print(f"  - Copied sheet '{source_sheet_name}' as '{sheet_name}'")
                
                source_wb.close()
                
            except Exception as e:
                print(f"Error processing {filename}: {e}")
                continue
        
        # Set clean workbook properties to avoid Excel recovery mode
        try:
            # Set clean workbook properties
            combined_wb.properties.title = "Combined Excel Files"
            combined_wb.properties.creator = "Audit Dashboard"
            combined_wb.properties.subject = "Branch Excel With Evidence"
            combined_wb.properties.description = "Combined Excel files from ZIP archive"
            combined_wb.properties.created = datetime.now()
            combined_wb.properties.modified = datetime.now()
            
            # Remove any problematic properties
            combined_wb.properties.keywords = None
            combined_wb.properties.category = None
            combined_wb.properties.version = None
            combined_wb.properties.lastModifiedBy = None
            combined_wb.properties.language = None
            combined_wb.properties.identifier = None
            
            # Clean up any problematic worksheet properties
            for sheet in combined_wb.worksheets:
                # Reset sheet properties to defaults
                sheet.sheet_properties.tabColor = None
                sheet.sheet_view.showGridLines = True
                sheet.sheet_view.showRowColHeaders = True
                sheet.sheet_view.showRuler = False
                sheet.sheet_view.showOutlineSymbols = False
                sheet.sheet_view.showZeros = True
                sheet.sheet_view.zoomScale = 100
                sheet.sheet_view.zoomScaleNormal = 100
                
                # Remove any problematic print settings
                if hasattr(sheet, 'page_setup'):
                    sheet.page_setup.orientation = 'portrait'
                    sheet.page_setup.paperSize = 9  # A4
                    sheet.page_setup.fitToHeight = 1
                    sheet.page_setup.fitToWidth = 1
                
                # Clear any problematic margins
                if hasattr(sheet, 'page_margins'):
                    sheet.page_margins.left = 0.7
                    sheet.page_margins.right = 0.7
                    sheet.page_margins.top = 0.75
                    sheet.page_margins.bottom = 0.75
                    sheet.page_margins.header = 0.3
                    sheet.page_margins.footer = 0.3
            
            # Save workbook to BytesIO buffer
            print("Saving combined workbook to memory...")
            output_buffer = BytesIO()
            
            # Save workbook to buffer
            combined_wb.save(output_buffer)
            combined_wb.close()
            
            # Get the bytes data
            output_buffer.seek(0)
            excel_data = output_buffer.getvalue()
            output_buffer.close()
            
            print("Workbook saved to memory successfully")
            
            # Validate the generated data
            if len(excel_data) < 1000:  # Minimum reasonable size for Excel file
                print("Warning: Generated Excel file seems too small, might be corrupted")
                return None, 0
            
            # Ensure file is completely closed and flushed
            import gc
            gc.collect()  # Force garbage collection to release file handles
            print("File handles released")
            
            print("="*60)
            print(f"COMBINATION COMPLETE")
            print(f"Total sheets added: {total_sheets_added}")
            print(f"Output size: {len(excel_data)} bytes")
            print("="*60)
            
            return excel_data, total_sheets_added
            
        except Exception as save_error:
            print(f"Error during save process: {save_error}")
            combined_wb.close()
            return None, 0
        
    except Exception as e:
        print(f"Error combining Excel files: {e}")
        return None, 0

@combine_branch_excel_with_poc_bp.route('/test_combine_with_poc_route')
def test_combine_with_poc_route():
    return "Combine Branch Excel With POC route is working!"

@combine_branch_excel_with_poc_bp.route('/combine_branch_excel_with_poc', methods=['POST'])
@login_required
def combine_branch_excel_with_poc():
    print("="*50)
    print("COMBINE BRANCH EXCEL WITH POC FORM SUBMISSION RECEIVED!")
    print("="*50)
    
    try:
        # Check if ZIP file is present in the request
        if 'zipFile' not in request.files:
            flash('ZIP file is required!', 'error')
            return redirect(url_for('audit_dashboard'))
        
        zip_file = request.files['zipFile']
        
        # Check if file is selected
        if zip_file.filename == '':
            flash('Please select a ZIP file!', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Validate ZIP file extension
        if not allowed_file(zip_file.filename, ALLOWED_ZIP_EXTENSIONS):
            flash('Invalid ZIP file format! Please upload .zip files only.', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Create temporary file that won't be automatically deleted
        temp_zip_file = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        temp_zip_file.close()  # Close the file handle to avoid locking issues
        
        try:
            # Secure the filename
            zip_filename = secure_filename(zip_file.filename)
            
            # Save uploaded file to temporary file
            zip_file.save(temp_zip_file.name)
            
            print(f"ZIP file saved to: {temp_zip_file.name}")
            
            # Create temporary directory for extracted files
            temp_extract_dir = tempfile.mkdtemp()
            
            # Extract and list files from ZIP file
            print("\n" + "="*60)
            print("PROCESSING ZIP FILE")
            print("="*60)
            extracted_files, excel_files = extract_and_list_files_from_zip(temp_zip_file.name, temp_extract_dir)
            
            if extracted_files:
                print(f"\nSuccessfully extracted {len(extracted_files)} files from ZIP archive")
                
                if excel_files:
                    print(f"Found {len(excel_files)} Excel files to combine")
                    
                    # Create output filename
                    output_filename = "Branch_Excel_With_Evidence.xlsx"
                    
                    # Combine Excel files to memory
                    excel_data, sheets_added = combine_excel_files_to_bytes(excel_files)
                    
                    if sheets_added > 0 and excel_data is not None:
                        print(f"Excel data created successfully - Size: {len(excel_data)} bytes")
                        
                        # Send the combined Excel file for download using in-memory data
                        print("Sending file for download...")
                        try:
                            response = Response(
                                excel_data,
                                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                headers={
                                    'Content-Disposition': f'attachment; filename="{output_filename}"',
                                    'Content-Length': str(len(excel_data)),
                                    'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                    'Cache-Control': 'no-cache, no-store, must-revalidate',
                                    'Pragma': 'no-cache',
                                    'Expires': '0',
                                    'X-Content-Type-Options': 'nosniff',
                                    'X-Download-Options': 'noopen',
                                    'Content-Transfer-Encoding': 'binary'
                                }
                            )
                            
                            print("File sent successfully!")
                            
                            # Schedule cleanup after response is sent
                            def cleanup_temp_files():
                                try:
                                    if os.path.exists(temp_zip_file.name):
                                        os.unlink(temp_zip_file.name)
                                    if os.path.exists(temp_extract_dir):
                                        shutil.rmtree(temp_extract_dir)
                                    print("Temporary files cleaned up")
                                except Exception as e:
                                    print(f"Error cleaning up temp files: {e}")
                            
                            response.call_on_close(cleanup_temp_files)
                            return response
                            
                        except Exception as download_error:
                            print(f"Error sending file: {download_error}")
                            flash(f'Error downloading file: {download_error}', 'error')
                            return redirect(url_for('audit_dashboard'))
                    else:
                        flash('Failed to combine Excel files!', 'error')
                        return redirect(url_for('audit_dashboard'))
                else:
                    flash('No Excel files (.xlsx) found in the ZIP archive!', 'error')
                    return redirect(url_for('audit_dashboard'))
            else:
                flash('No files found in the ZIP archive!', 'error')
                return redirect(url_for('audit_dashboard'))
                
        except Exception as e:
            # Clean up the temporary files if there was an error
            try:
                if os.path.exists(temp_zip_file.name):
                    os.unlink(temp_zip_file.name)
                # Clean up temporary extraction directory
                if 'temp_extract_dir' in locals() and os.path.exists(temp_extract_dir):
                    shutil.rmtree(temp_extract_dir)
            except:
                pass
            raise e
    
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        flash(f'Error processing file: {str(e)}', 'error')
        return redirect(url_for('audit_dashboard'))
