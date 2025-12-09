# follow_up_audit_routes.py
from flask import Blueprint, request, send_file, make_response, jsonify, session
from flask_login import current_user
import re
import io
import pandas as pd
import zipfile
import os
import math
from io import BytesIO
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# Create a Blueprint for follow up audit routes
follow_up_audit_bp = Blueprint('follow_up_audit', __name__)

# Default temp directory placeholders (overridden per request)
temp_poc_dir = None
temp_user_dir = None
temp_poc2_folder = None

def extract_vulnerability_names_from_excel(excel_file):
    """
    Extract 'Name of Vulnerability' column from Excel file.
    Returns a set of vulnerability names.
    """
    try:
        # Read the Excel file
        df = pd.read_excel(excel_file, sheet_name='Infra_VAPT')
        
        # Check if 'Name of Vulnerability' column exists
        if 'Name of Vulnerability' not in df.columns:
            print("Warning: 'Name of Vulnerability' column not found in Excel file")
            return set()
        
        # Extract vulnerability names and remove NaN values
        vulnerability_names = df['Name of Vulnerability'].dropna().astype(str).str.strip()
        
        # Return as a set for efficient comparison
        return set(vulnerability_names.tolist())
        
    except Exception as e:
        print(f"Error extracting vulnerability names from Excel: {e}")
        return set()

def extract_full_row_data_from_excel(excel_file):
    """
    Extract full row data from user's Excel file Infra_VAPT worksheet.
    Returns a dictionary with vulnerability names as keys and full row data as values.
    """
    try:
        # Read the Excel file
        df = pd.read_excel(excel_file, sheet_name='Infra_VAPT')
        
        # Check if 'Name of Vulnerability' column exists
        if 'Name of Vulnerability' not in df.columns:
            print("Warning: 'Name of Vulnerability' column not found in Excel file")
            return {}
        
        # Create a dictionary to store full row data
        row_data = {}
        
        # Iterate through each row
        for index, row in df.iterrows():
            vuln_name = row['Name of Vulnerability']
            if pd.notna(vuln_name):
                vuln_name_str = str(vuln_name).strip()
                # Store the entire row as a dictionary
                row_data[vuln_name_str] = row.to_dict()
        
        print(f"Extracted {len(row_data)} full rows from user's Excel file")
        return row_data
        
    except Exception as e:
        print(f"Error extracting full row data from Excel: {e}")
        return {}

def compare_vulnerabilities(our_vulnerabilities, user_vulnerabilities):
    """
    Compare vulnerabilities between our Excel and user's Excel.
    Returns a dictionary with status for each vulnerability.
    """
    comparison_result = {}
    
    # Find common vulnerabilities (Open)
    common_vulnerabilities = our_vulnerabilities.intersection(user_vulnerabilities)
    for vuln in common_vulnerabilities:
        comparison_result[vuln] = "Open"
    
    # Find vulnerabilities only in user's Excel (Closed)
    closed_vulnerabilities = user_vulnerabilities - our_vulnerabilities
    for vuln in closed_vulnerabilities:
        comparison_result[vuln] = "Closed"
    
    # Find vulnerabilities only in our Excel (New)
    new_vulnerabilities = our_vulnerabilities - user_vulnerabilities
    for vuln in new_vulnerabilities:
        comparison_result[vuln] = "New"
    
    return comparison_result

def add_status_column_to_excel(excel_path, comparison_result):
    """
    Populate the Status column (already created at position L) with values from comparison.
    Does NOT insert a new column - just fills existing Status column.
    """
    try:
        from openpyxl import load_workbook
        
        # Load the workbook
        wb = load_workbook(excel_path)
        
        # Get the Infra_VAPT worksheet
        if "Infra_VAPT" not in wb.sheetnames:
            print("Error: Infra_VAPT worksheet not found")
            return False
        
        ws = wb["Infra_VAPT"]
        
        # Find the Status column index (should be at position L)
        status_column_index = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and str(cell_value).strip() == 'Status':
                status_column_index = col
                break
        
        if status_column_index is None:
            print("Error: Status column not found")
            return False
        
        print(f"Found Status column at position: {status_column_index}")
        
        # Find the 'Name of Vulnerability' column
        vuln_column_index = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and 'Name of Vulnerability' in str(cell_value):
                vuln_column_index = col
                break
        
        if vuln_column_index is None:
            print("Error: 'Name of Vulnerability' column not found")
            return False
        
        # Fill status for each row based on comparison result
        status_count = {'Open': 0, 'Closed': 0, 'New': 0}
        for row in range(2, ws.max_row + 1):
            vuln_name = ws.cell(row=row, column=vuln_column_index).value
            if vuln_name:
                vuln_name_str = str(vuln_name).strip()
                # Use only first 170 characters for matching
                vuln_name_short = vuln_name_str[:170]
                
                # Check both full name and short name
                status = comparison_result.get(vuln_name_str, comparison_result.get(vuln_name_short, "Unknown"))
                ws.cell(row=row, column=status_column_index, value=status)
                
                if status in status_count:
                    status_count[status] += 1
        
        print(f"Status values populated: {status_count}")
        
        # Save the workbook
        wb.save(excel_path)
        print("Status column populated successfully")
        return True
        
    except Exception as e:
        print(f"Error populating status column: {e}")
        import traceback
        traceback.print_exc()
        return False

def extract_risk_factor_counts_from_user_excel(user_excel_file):
    """
    Extract risk factor counts from user's Excel file Infra_VAPT worksheet.
    Returns a dictionary with risk factor counts (Critical, High, Medium, Low).
    """
    try:
        from openpyxl import load_workbook
        
        # Load the user's workbook
        wb = load_workbook(user_excel_file)
        
        # Get the Infra_VAPT worksheet
        if "Infra_VAPT" not in wb.sheetnames:
            print("Warning: Infra_VAPT worksheet not found in user's Excel file")
            return {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        
        ws = wb["Infra_VAPT"]
        
        # Find Risk Factor column
        risk_factor_col = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and 'risk' in str(cell_value).lower() and 'factor' in str(cell_value).lower():
                risk_factor_col = col
                break
        
        if risk_factor_col is None:
            print("Warning: Risk Factor column not found in user's Excel file")
            return {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        
        # Count risk factors
        risk_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        
        for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
            cell_value = ws.cell(row=row, column=risk_factor_col).value
            if cell_value:
                risk_value = str(cell_value).strip().upper()
                if risk_value in risk_counts:
                    risk_counts[risk_value] += 1
                elif risk_value in ["CRITICAL"]:
                    risk_counts["Critical"] += 1
                elif risk_value in ["HIGH"]:
                    risk_counts["High"] += 1
                elif risk_value in ["MEDIUM"]:
                    risk_counts["Medium"] += 1
                elif risk_value in ["LOW"]:
                    risk_counts["Low"] += 1
        
        print(f"Risk factor counts from user's Excel: {risk_counts}")
        return risk_counts
        
    except Exception as e:
        print(f"Error extracting risk factor counts from user's Excel file: {e}")
        import traceback
        traceback.print_exc()
        return {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}

def extract_follow_up_risk_factor_counts_from_generated_excel(excel_path):
    """
    Extract risk factor counts from generated Excel file Infra_VAPT worksheet.
    Counts ALL vulnerabilities (New, Open, and Closed) for Follow Up Audit.
    Returns a dictionary with risk factor counts (Critical, High, Medium, Low).
    """
    try:
        from openpyxl import load_workbook
        
        # Load the generated workbook
        wb = load_workbook(excel_path)
        
        # Get the Infra_VAPT worksheet
        if "Infra_VAPT" not in wb.sheetnames:
            print("Warning: Infra_VAPT worksheet not found in generated Excel file")
            return {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        
        ws = wb["Infra_VAPT"]
        
        # Find Risk Factor column
        risk_factor_col = None
        
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                cell_str = str(cell_value).lower()
                if 'risk' in cell_str and 'factor' in cell_str:
                    risk_factor_col = col
                    break
        
        if risk_factor_col is None:
            print("Warning: Risk Factor column not found in generated Excel file")
            print(f"Available columns: {[ws.cell(row=1, col=c).value for c in range(1, min(ws.max_column + 1, 20))]}")
            return {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        
        # Count risk factors (count ALL rows, not just New/Open)
        risk_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        
        for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
            # Count risk factor for all rows
            risk_value = ws.cell(row=row, column=risk_factor_col).value
            if risk_value:
                risk_str = str(risk_value).strip()
                risk_upper = risk_str.upper()
                
                # Match risk factor (case-insensitive)
                if risk_upper == "CRITICAL" or risk_str == "Critical":
                    risk_counts["Critical"] += 1
                elif risk_upper == "HIGH" or risk_str == "High":
                    risk_counts["High"] += 1
                elif risk_upper == "MEDIUM" or risk_str == "Medium":
                    risk_counts["Medium"] += 1
                elif risk_upper == "LOW" or risk_str == "Low":
                    risk_counts["Low"] += 1
                else:
                    # Try to match partial strings
                    if "critical" in risk_str.lower():
                        risk_counts["Critical"] += 1
                    elif "high" in risk_str.lower() and "medium" not in risk_str.lower() and "low" not in risk_str.lower():
                        risk_counts["High"] += 1
                    elif "medium" in risk_str.lower():
                        risk_counts["Medium"] += 1
                    elif "low" in risk_str.lower():
                        risk_counts["Low"] += 1
        
        print(f"Follow-up risk factor counts (ALL vulnerabilities) from generated Excel: {risk_counts}")
        print(f"Total rows processed: {ws.max_row - 1}")
        return risk_counts
        
    except Exception as e:
        print(f"Error extracting follow-up risk factor counts from generated Excel file: {e}")
        import traceback
        traceback.print_exc()
        return {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}

def update_meta_data_with_follow_up_vulnerabilities(excel_path):
    """
    Update Meta_Data worksheet with Follow up Audit Vulnerabilities section after Infra_VAPT worksheet is ready.
    """
    try:
        from openpyxl import load_workbook
        
        # Load the workbook
        wb = load_workbook(excel_path)
        
        # Check if Meta_Data worksheet exists
        if "Meta_Data" not in wb.sheetnames:
            print("Warning: Meta_Data worksheet not found")
            return False
        
        ws_meta = wb["Meta_Data"]
        
        # Extract follow-up risk factor counts
        follow_up_counts = extract_follow_up_risk_factor_counts_from_generated_excel(excel_path)
        
        # Find the row where "FOLLOW UP AUDIT VULNERABILITIES" section should be
        follow_up_section_row = None
        for row in range(1, ws_meta.max_row + 1):
            cell_value = ws_meta.cell(row=row, column=1).value
            if cell_value and "FOLLOW UP AUDIT" in str(cell_value).upper():
                follow_up_section_row = row
                break
        
        if follow_up_section_row is None:
            print("Warning: FOLLOW UP AUDIT VULNERABILITIES section not found in Meta_Data worksheet")
            return False
        
        # Update the values in the Follow up Audit Vulnerabilities section
        # The data rows should be at follow_up_section_row + 1, +2, +3, +4
        vulnerability_levels = ['Critical', 'High', 'Medium', 'Low']
        
        for i, level in enumerate(vulnerability_levels):
            data_row = follow_up_section_row + 1 + i
            if data_row <= ws_meta.max_row:
                # Update the value in column B (index 2)
                ws_meta.cell(row=data_row, column=2, value=str(follow_up_counts[level]))
                print(f"Updated {level}: {follow_up_counts[level]}")
        
        # Save the workbook
        wb.save(excel_path)
        print(f"Successfully updated Meta_Data worksheet with follow-up vulnerability counts: {follow_up_counts}")
        return True
        
    except Exception as e:
        print(f"Error updating Meta_Data worksheet with follow-up vulnerabilities: {e}")
        import traceback
        traceback.print_exc()
        return False

def extract_poc_data_from_user_excel(user_excel_file, closed_vulnerabilities, temp_user_dir=None):
    """
    Extract POC/evidence IMAGES from user's Excel file for Closed vulnerabilities.
    Returns a dictionary with vulnerability names as keys and list of image paths as values.
    """
    try:
        from openpyxl import load_workbook
        import tempfile
        import os
        from PIL import Image as PILImage
        import io
        from datetime import datetime
        
        # Load the user's workbook
        wb = load_workbook(user_excel_file)
        
        # Get the Infra_VAPT worksheet
        if "Infra_VAPT" not in wb.sheetnames:
            print("Warning: Infra_VAPT worksheet not found in user's Excel file")
            return {}
        
        ws = wb["Infra_VAPT"]
        
        # Find POC column range (could be merged header or single column)
        poc_col_start = None
        poc_col_end = None
        
        # Check for merged cells in row 1 for POC header
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == 1 and merged_range.max_row == 1:
                first_cell = ws.cell(row=1, column=merged_range.min_col)
                cell_value = str(first_cell.value).strip() if first_cell.value else ""
                # Look for POC but not "Old POC"
                if 'POC' in cell_value and 'Old' not in cell_value:
                    poc_col_start = merged_range.min_col
                    poc_col_end = merged_range.max_col
                    break
        
        # Fallback: look for single POC column
        if poc_col_start is None:
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value and 'POC' in str(cell_value) and 'Old' not in str(cell_value):
                    poc_col_start = col
                    poc_col_end = col
                    break
        
        if poc_col_start is None:
            print("Warning: POC column(s) not found in user's Excel file")
            return {}
        
        print(f"Found POC columns from {poc_col_start} to {poc_col_end} in user's Excel")
        
        # Find Name of Vulnerability column
        vuln_col = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and 'Name of Vulnerability' in str(cell_value):
                vuln_col = col
                break
        
        if vuln_col is None:
            print("Warning: Name of Vulnerability column not found in user's Excel file")
            return {}
        
        # Extract images from the user's Excel for Closed vulnerabilities
        poc_images = {}  # Store all images for each vulnerability
        
        # Extract all images from the worksheet
        if hasattr(ws, "_images") and ws._images:
            print(f"Found {len(ws._images)} images in user's Excel")
            
            # Prepare timestamped temp folder (use temp_poc_images with timestamp)
            if not temp_user_dir:
                ts = datetime.now().strftime('%Y%m%d%H%M%S')
                temp_user_dir = f"temp_poc_images_{ts}"
            os.makedirs(temp_user_dir, exist_ok=True)
            
            for img in ws._images:
                try:
                    # Get the row and column of the image
                    row_idx = img.anchor._from.row + 1  # Excel row (1-based)
                    col_idx = img.anchor._from.col + 1  # Excel column (1-based)
                    
                    # Check if this image is in the POC column range
                    if poc_col_start <= col_idx <= poc_col_end:
                        # Get the vulnerability name for this row
                        if row_idx >= 2:  # Skip header
                            vuln_name_cell = ws.cell(row=row_idx, column=vuln_col)
                            if vuln_name_cell.value:
                                vuln_name = str(vuln_name_cell.value).strip()
                                # Use only first 170 characters for matching
                                vuln_name_short = vuln_name[:170]
                                
                                # Check if this vulnerability is in the closed list
                                is_closed = False
                                for closed_vuln in closed_vulnerabilities:
                                    if closed_vuln[:170] == vuln_name_short:
                                        is_closed = True
                                        vuln_name = closed_vuln  # Use the full name from closed list
                                        break
                                
                                if is_closed:
                                    # Get image data
                                    img_data = img._data() if callable(img._data) else img._data
                                    
                                    if img_data:
                                        # Save image to temporary file (use temp_poc_images with timestamp)
                                        if not temp_user_dir:
                                            ts = datetime.now().strftime('%Y%m%d%H%M%S')
                                            temp_dir = f"temp_poc_images_{ts}"
                                        else:
                                            temp_dir = temp_user_dir
                                        os.makedirs(temp_dir, exist_ok=True)
                                        
                                        # Create temporary file
                                        with tempfile.NamedTemporaryFile(delete=False, suffix='.png', dir=temp_dir) as tmp:
                                            tmp.write(img_data)
                                            temp_path = tmp.name
                                        
                                        # Store the image path
                                        if vuln_name not in poc_images:
                                            poc_images[vuln_name] = []
                                        poc_images[vuln_name].append(temp_path)
                                        
                                        print(f"  ✅ Extracted image for closed vulnerability '{vuln_name}' from row {row_idx}, column {col_idx}")
                
                except Exception as e:
                    print(f"  ⚠️ Error extracting image: {e}")
                    continue
        
        wb.close()
        
        print(f"Extracted images for {len(poc_images)} Closed vulnerabilities (total {sum(len(imgs) for imgs in poc_images.values())} images)")
        return poc_images
        
    except Exception as e:
        print(f"Error extracting POC images from user's Excel: {e}")
        import traceback
        traceback.print_exc()
        return {}

def sort_and_renumber_infra_vapt_worksheet(excel_path):
    """
    Sort the Infra_VAPT worksheet by risk level (Critical, High, Medium, Low) and renumber rows.
    """
    try:
        from openpyxl import load_workbook
        import pandas as pd
        
        # Load the workbook
        wb = load_workbook(excel_path)
        
        # Get the Infra_VAPT worksheet
        if "Infra_VAPT" not in wb.sheetnames:
            print("Error: Infra_VAPT worksheet not found")
            return False
        
        ws = wb["Infra_VAPT"]
        
        # Convert worksheet to DataFrame for easier manipulation
        data = []
        headers = []
        
        # Get headers from first row
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value else f"Column_{col}")
        
        # Get data from all rows (excluding header)
        for row in range(2, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(cell_value)
            data.append(row_data)
        
        if not data:
            print("No data rows found in Infra_VAPT worksheet")
            return True
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        # Find Risk Factor column
        risk_factor_col = None
        for i, col in enumerate(headers):
            if 'risk' in col.lower() and 'factor' in col.lower():
                risk_factor_col = i
                break
        
        if risk_factor_col is None:
            print("Warning: Risk Factor column not found, skipping sort")
            return True
        
        # Find POC column to exclude from sorting
        poc_col = None
        for i, col in enumerate(headers):
            if 'poc' in col.lower():
                poc_col = i
                break
        
        if poc_col is not None:
            print(f"Found POC column at index {poc_col}, will preserve during sorting")
        
        # Define risk level priority
        risk_priority = {'Critical': 1, 'High': 2, 'Medium': 3, 'Low': 4}
        
        # Create a list of rows with their risk priorities for sorting
        rows_with_priority = []
        for row_idx, row_data in enumerate(data):
            risk_value = str(row_data[risk_factor_col]).strip() if pd.notna(row_data[risk_factor_col]) else ""
            priority = risk_priority.get(risk_value, 5)
            rows_with_priority.append((priority, row_idx, row_data))
        
        # Sort by priority (risk level)
        rows_with_priority.sort(key=lambda x: x[0])
        
        # Find Sr No column for renumbering
        sr_no_col = None
        for i, col in enumerate(headers):
            if 'sr' in col.lower() and 'no' in col.lower():
                sr_no_col = i
                break
        
        # Clear the worksheet (except header)
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
        
        # Write sorted data back to worksheet, preserving POC column
        for new_row_idx, (priority, old_row_idx, row_data) in enumerate(rows_with_priority, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                # Skip POC column - keep it empty for sorted rows
                if poc_col is not None and col_idx == poc_col + 1:  # +1 because openpyxl is 1-indexed
                    ws.cell(row=new_row_idx, column=col_idx, value="")
                    continue
                
                # Set other column values
                if value is not None and pd.notna(value):
                    ws.cell(row=new_row_idx, column=col_idx, value=value)
                else:
                    ws.cell(row=new_row_idx, column=col_idx, value="")
            
            # Renumber the Sr No column
            if sr_no_col is not None:
                ws.cell(row=new_row_idx, column=sr_no_col + 1, value=new_row_idx - 1)
        
        # Apply color formatting to Risk Factor column after sorting
        if risk_factor_col is not None:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            # Create color formats for risk factors
            critical_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Dark Red
            high_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")      # Red
            medium_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")    # Orange
            low_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")       # Green
            
            # Create common formatting
            white_font = Font(color="FFFFFF", bold=True)
            center_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply formatting to each row in Risk Factor column
            for row_idx in range(2, len(rows_with_priority) + 2):  # Start from row 2 (after header)
                cell = ws.cell(row=row_idx, column=risk_factor_col + 1)  # +1 because openpyxl is 1-indexed
                risk_value = str(cell.value).strip().upper() if cell.value else ""
                
                # Apply color formatting based on risk level
                if risk_value == 'CRITICAL':
                    cell.fill = critical_fill
                    cell.font = white_font
                    cell.alignment = center_alignment
                    cell.border = border
                    cell.value = 'Critical'  # Ensure CamelCase
                elif risk_value == 'HIGH':
                    cell.fill = high_fill
                    cell.font = white_font
                    cell.alignment = center_alignment
                    cell.border = border
                    cell.value = 'High'  # Ensure CamelCase
                elif risk_value == 'MEDIUM':
                    cell.fill = medium_fill
                    cell.font = white_font
                    cell.alignment = center_alignment
                    cell.border = border
                    cell.value = 'Medium'  # Ensure CamelCase
                elif risk_value == 'LOW':
                    cell.fill = low_fill
                    cell.font = white_font
                    cell.alignment = center_alignment
                    cell.border = border
                    cell.value = 'Low'  # Ensure CamelCase
                else:
                    # Default formatting for unknown risk levels
                    cell.font = Font()
                    cell.alignment = center_alignment
                    cell.border = border
            
            print("Color formatting applied to Risk Factor column after sorting")
        
        # Apply formatting to Status column header
        status_col = None
        for col_idx, header in enumerate(headers, 1):
            if 'status' in header.lower():
                status_col = col_idx
                break
        
        if status_col is not None:
            from openpyxl.styles import PatternFill, Font, Alignment
            
            # Create Status header formatting
            status_header_fill = PatternFill(start_color="3553E8", end_color="3553E8", fill_type="solid")  # Blue
            status_header_font = Font(color="FFFFFF", bold=True)  # White bold text
            status_header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply formatting to Status header (row 1)
            status_header_cell = ws.cell(row=1, column=status_col)
            status_header_cell.fill = status_header_fill
            status_header_cell.font = status_header_font
            status_header_cell.alignment = status_header_alignment
            
            print("Status column header formatting applied (blue background, white bold text, center alignment)")
            
            # Apply formatting to all Status column cells with content
            from openpyxl.styles import Border, Side
            
            # Create cell formatting for Status column
            status_cell_alignment = Alignment(horizontal="center", vertical="center")
            status_cell_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply formatting to all Status column cells (starting from row 2)
            for row_idx in range(2, len(rows_with_priority) + 2):  # Start from row 2 (after header)
                status_cell = ws.cell(row=row_idx, column=status_col)
                
                # Only apply formatting if cell has content
                if status_cell.value and str(status_cell.value).strip():
                    status_cell.alignment = status_cell_alignment
                    status_cell.border = status_cell_border
                    print(f"DEBUG: Applied formatting to Status cell at row {row_idx}: '{status_cell.value}'")
            
            print(f"Status column cell formatting applied to {len(rows_with_priority)} rows with content")
        
        # Save the workbook
        wb.save(excel_path)
        print(f"Infra_VAPT worksheet sorted by risk level and renumbered successfully")
        print(f"Total vulnerabilities: {len(rows_with_priority)}")
        
        # Print risk level distribution
        if risk_factor_col is not None:
            risk_counts = {}
            for priority, old_row_idx, row_data in rows_with_priority:
                risk_value = str(row_data[risk_factor_col]).strip() if pd.notna(row_data[risk_factor_col]) else "Unknown"
                risk_counts[risk_value] = risk_counts.get(risk_value, 0) + 1
            
            print("Risk level distribution after sorting:")
            for risk, count in risk_counts.items():
                print(f"  {risk}: {count}")
        
        return True
        
    except Exception as e:
        print(f"Error sorting and renumbering Infra_VAPT worksheet: {e}")
        import traceback
        traceback.print_exc()
        return False

def get_existing_row_formatting(ws, sample_row=2):
    """
    Extract formatting from an existing row to maintain consistency.
    Returns a dictionary with formatting information.
    """
    try:
        from openpyxl.styles import Font, Alignment, Border, PatternFill, Side
        
        # Get formatting from a sample existing row
        sample_cell = ws.cell(row=sample_row, column=1)
        
        # Create new formatting objects based on existing cell properties
        existing_format = {
            'font': Font(
                name=sample_cell.font.name if sample_cell.font else 'Calibri',
                size=sample_cell.font.size if sample_cell.font else 11,
                bold=sample_cell.font.bold if sample_cell.font else False,
                italic=sample_cell.font.italic if sample_cell.font else False,
                color=sample_cell.font.color if sample_cell.font and sample_cell.font.color else '000000'
            ),
            'alignment': Alignment(
                horizontal=sample_cell.alignment.horizontal if sample_cell.alignment else 'center',
                vertical=sample_cell.alignment.vertical if sample_cell.alignment else 'center',
                wrap_text=sample_cell.alignment.wrap_text if sample_cell.alignment else False
            ),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'fill': PatternFill(fill_type='none')
        }
        
        return existing_format
    except Exception as e:
        print(f"Warning: Could not extract existing formatting: {e}")
        return None

def add_closed_vulnerability_rows_to_excel(excel_path, closed_vulnerabilities, user_row_data):
    """
    Add rows for Closed vulnerabilities from user's Excel to our generated Excel.
    Applies the same formatting as existing rows, including Risk Factor column formatting.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
        
        # Load the workbook
        wb = load_workbook(excel_path)
        
        # Get the Infra_VAPT worksheet
        if "Infra_VAPT" not in wb.sheetnames:
            print("Error: Infra_VAPT worksheet not found")
            return False
        
        ws = wb["Infra_VAPT"]
        
        # Get existing formatting from the worksheet
        existing_format = get_existing_row_formatting(ws)
        
        # Get the column headers from our worksheet
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value))
            else:
                headers.append(f"Column_{col}")
        
        print(f"Our worksheet headers: {headers}")
        
        # Find the Risk Factor column index
        risk_factor_col = None
        for col_idx, header in enumerate(headers, 1):
            if 'risk' in header.lower() and 'factor' in header.lower():
                risk_factor_col = col_idx
                break
        
        # Find the Status column index
        status_col = None
        for col_idx, header in enumerate(headers, 1):
            if 'status' in header.lower():
                status_col = col_idx
                break
        
        # Create formatting styles for Risk Factor column (matching existing format)
        critical_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Dark Red
        high_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")      # Red
        medium_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")    # Orange
        low_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")       # Green
        
        # Create common formatting
        white_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create default cell format (matching existing rows)
        if existing_format:
            default_format = {
                'font': existing_format['font'],
                'alignment': existing_format['alignment'],
                'border': existing_format['border']
            }
        else:
            default_format = {
                'font': Font(),
                'alignment': Alignment(horizontal="center", vertical="center"),
                'border': border
            }
        
        # Find the last row with actual content (not just max_row which may include empty rows)
        # Look for the last row that has content in the "Vulnerabilities" or "Name of Vulnerability" column
        last_row_with_content = 1  # Start with header row
        vuln_name_col = None
        for col_idx, header in enumerate(headers, 1):
            if 'name of vulnerability' in header.lower() or 'vulnerabilities' in header.lower():
                vuln_name_col = col_idx
                break
        
        # Find the actual last row with content
        if vuln_name_col:
            for row in range(ws.max_row, 0, -1):  # Check from bottom up
                cell_value = ws.cell(row=row, column=vuln_name_col).value
                if cell_value is not None and str(cell_value).strip() != "":
                    last_row_with_content = row
                    break
        else:
            # Fallback: find last row with any content in first few columns
            for row in range(ws.max_row, 0, -1):
                has_content = False
                for col in range(1, min(6, ws.max_column + 1)):  # Check first 5 columns
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None and str(cell_value).strip() != "":
                        has_content = True
                        break
                if has_content:
                    last_row_with_content = row
                    break
        
        # Find the next empty row (after the last row with content)
        next_row = last_row_with_content + 1
        print(f"DEBUG: Last row with content: {last_row_with_content}, Starting new rows at: {next_row}")
        
        # Add each Closed vulnerability row
        print(f"DEBUG: Processing {len(closed_vulnerabilities)} Closed vulnerabilities")
        rows_added = 0
        for vuln_name in closed_vulnerabilities:
            print(f"DEBUG: Processing vulnerability: {vuln_name}")
            if vuln_name in user_row_data:
                user_row = user_row_data[vuln_name]
                print(f"Adding Closed vulnerability: {vuln_name}")
                print(f"DEBUG: User row data for {vuln_name}: {dict(list(user_row.items())[:3])}")
                
                # Check if we have meaningful data for this vulnerability before adding a row
                has_meaningful_data = False
                
                # Map user's data to our worksheet columns
                # First pass: check if we have any meaningful data
                for col_idx, header in enumerate(headers, 1):
                    # Skip POC columns for this check
                    if 'POC' in header:
                        continue
                    
                    # Try to find matching column in user's data
                    cell_value = None
                    
                    # Direct column name match
                    if header in user_row:
                        cell_value = user_row[header]
                    else:
                        # Try partial matches for common variations
                        for user_col, user_value in user_row.items():
                            if user_col and header:
                                # Check for partial matches
                                if (header.lower() in user_col.lower() or 
                                    user_col.lower() in header.lower() or
                                    any(word in user_col.lower() for word in header.lower().split())):
                                    cell_value = user_value
                                    break
                    
                    # Check if we have meaningful content
                    if cell_value is not None and pd.notna(cell_value):
                        cell_str = str(cell_value).strip()
                        if cell_str != "":
                            has_meaningful_data = True
                            break
                
                # Only add row if we have meaningful data
                if not has_meaningful_data:
                    print(f"DEBUG: Skipping {vuln_name} - no meaningful data found")
                    continue
                
                # Second pass: set cell values and formatting
                for col_idx, header in enumerate(headers, 1):
                    # Get the cell (we'll set value and formatting)
                    cell = ws.cell(row=next_row, column=col_idx)
                    
                    # Skip POC columns - don't set values, but still apply basic formatting
                    if 'POC' in header:
                        # Don't set empty strings for POC columns - leave them as None
                        # But apply default border formatting so cells look consistent
                        cell.font = default_format['font']
                        cell.alignment = default_format['alignment']
                        cell.border = default_format['border']
                        continue
                    
                    # Try to find matching column in user's data
                    cell_value = None
                    
                    # Direct column name match
                    if header in user_row:
                        cell_value = user_row[header]
                    else:
                        # Try partial matches for common variations
                        for user_col, user_value in user_row.items():
                            if user_col and header:
                                # Check for partial matches
                                if (header.lower() in user_col.lower() or 
                                    user_col.lower() in header.lower() or
                                    any(word in user_col.lower() for word in header.lower().split())):
                                    cell_value = user_value
                                    break
                    
                    # Set the cell value only if we have content
                    if cell_value is not None and pd.notna(cell_value):
                        # Handle different data types
                        if isinstance(cell_value, (int, float)):
                            cell.value = cell_value
                        else:
                            cell_str = str(cell_value).strip()
                            if cell_str != "":
                                cell.value = cell_str
                    
                    # Apply formatting based on column type
                    # Apply Risk Factor column formatting
                    if col_idx == risk_factor_col and cell_value:
                        risk_value = str(cell_value).strip().upper()
                        if risk_value == 'CRITICAL':
                            cell.fill = critical_fill
                            cell.font = white_font
                            cell.alignment = center_alignment
                            cell.border = border
                            # Convert to CamelCase
                            cell.value = 'Critical'
                        elif risk_value == 'HIGH':
                            cell.fill = high_fill
                            cell.font = white_font
                            cell.alignment = center_alignment
                            cell.border = border
                            cell.value = 'High'
                        elif risk_value == 'MEDIUM':
                            cell.fill = medium_fill
                            cell.font = white_font
                            cell.alignment = center_alignment
                            cell.border = border
                            cell.value = 'Medium'
                        elif risk_value == 'LOW':
                            cell.fill = low_fill
                            cell.font = white_font
                            cell.alignment = center_alignment
                            cell.border = border
                            cell.value = 'Low'
                        else:
                            # Default formatting for unknown risk levels
                            cell.font = Font()
                            cell.alignment = center_alignment
                            cell.border = border
                    else:
                        # Apply default formatting to other columns
                        cell.font = default_format['font']
                        cell.alignment = default_format['alignment']
                        cell.border = default_format['border']
                        
                        # Special alignment for text-heavy columns
                        if any(keyword in header.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # Set status as "Closed" with proper formatting
                if status_col:
                    status_cell = ws.cell(row=next_row, column=status_col, value="Closed")
                    status_cell.font = Font()
                    status_cell.alignment = center_alignment
                    status_cell.border = border
                
                rows_added += 1
                next_row += 1
            else:
                print(f"DEBUG: Vulnerability {vuln_name} not found in user_row_data")
                print(f"DEBUG: Available keys in user_row_data: {list(user_row_data.keys())[:10]}")
        
        # Clean up any empty rows at the end (if any were created accidentally)
        # Find the actual last row with content after adding all rows
        # Re-check from the bottom to find the true last row with content
        actual_last_row = last_row_with_content
        
        # Find the last row that has actual content in key columns
        if vuln_name_col:
            # Check from max_row down to header row (row 1) to find last row with vulnerability name
            for row in range(ws.max_row, 0, -1):
                cell_value = ws.cell(row=row, column=vuln_name_col).value
                if cell_value is not None and str(cell_value).strip() != "":
                    actual_last_row = row
                    break
        else:
            # Fallback: find last row with any content in first few columns
            for row in range(ws.max_row, 0, -1):
                has_content = False
                for col in range(1, min(6, ws.max_column + 1)):  # Check first 5 columns
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None and str(cell_value).strip() != "":
                        has_content = True
                        actual_last_row = row
                        break
                if has_content:
                    break
        
        print(f"DEBUG: After adding rows - Last row with content: {actual_last_row}, Max row: {ws.max_row}")
        
        # Remove any empty rows after the last row with actual content
        # Check from max_row down to actual_last_row
        rows_deleted = 0
        if ws.max_row > actual_last_row:
            # Collect rows to delete (from bottom up to avoid index shifting issues)
            rows_to_delete = []
            
            # Check each row from max_row down to actual_last_row + 1
            for row in range(ws.max_row, actual_last_row, -1):
                is_empty = True
                
                # Check key columns first - if any key column has content, row is not empty
                key_columns = []
                if vuln_name_col:
                    key_columns.append(vuln_name_col)
                if status_col:
                    key_columns.append(status_col)
                if risk_factor_col:
                    key_columns.append(risk_factor_col)
                
                # Check key columns for content
                for col in key_columns:
                    if col:
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value is not None:
                            cell_str = str(cell_value).strip()
                            if cell_str != "":
                                is_empty = False
                                break
                
                # If key columns are empty, check all other columns
                if is_empty:
                    for col in range(1, ws.max_column + 1):
                        if col in key_columns:
                            continue  # Already checked
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value is not None:
                            cell_str = str(cell_value).strip()
                            if cell_str != "":
                                is_empty = False
                                break
                    
                    # Also check for images in this row (images might exist without cell values)
                    # If row has images, it's not empty
                    if is_empty and hasattr(ws, '_images'):
                        for image in ws._images:
                            try:
                                if hasattr(image, 'anchor'):
                                    anchor = image.anchor
                                    if hasattr(anchor, '_from'):
                                        img_row = anchor._from.row + 1  # openpyxl uses 0-based, convert to 1-based
                                        if img_row == row:
                                            is_empty = False
                                            break
                            except:
                                continue
                
                # If row is still empty, mark it for deletion
                if is_empty:
                    rows_to_delete.append(row)
            
            # Delete collected empty rows (from bottom to top to avoid index shifting)
            for row in sorted(rows_to_delete, reverse=True):
                try:
                    ws.delete_rows(row)
                    rows_deleted += 1
                    print(f"DEBUG: Deleted empty row {row}")
                except Exception as e:
                    print(f"DEBUG: Could not delete row {row}: {e}")
        
        if rows_deleted > 0:
            print(f"DEBUG: Cleaned up {rows_deleted} empty row(s) at the end")
        else:
            print(f"DEBUG: No empty rows to clean up. Last row with content: {actual_last_row}, Max row: {ws.max_row}")
        
        # Save the workbook
        wb.save(excel_path)
        print(f"Added {rows_added} Closed vulnerability rows with proper formatting")
        return True
        
    except Exception as e:
        print(f"Error adding Closed vulnerability rows: {e}")
        import traceback
        traceback.print_exc()
        return False

def extract_poc_images(evidence_files, temp_dir=None):
    """Extract POC images from evidence files zip and return mapping of vulnerability names to image paths"""
    poc_mapping = {}
    
    if not evidence_files or evidence_files.filename == '':
        return poc_mapping
    
    try:
        import zipfile
        from io import BytesIO
        from datetime import datetime
        
        # Read the zip file
        zip_data = evidence_files.read()
        
        with zipfile.ZipFile(BytesIO(zip_data), 'r') as zip_ref:
            file_list = zip_ref.namelist()
            
            # Find ALL POC folder
            poc_folder = None
            for file_path in file_list:
                if 'ALL POC' in file_path and file_path.endswith('/'):
                    poc_folder = file_path
                    break
            
            if not poc_folder:
                print("ALL POC folder not found in evidence files")
                return poc_mapping
            
            # Extract images from ALL POC folder
            image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff']
            
            # Prepare timestamped temp folder
            if not temp_dir:
                ts = datetime.now().strftime('%Y%m%d%H%M%S')
                temp_dir = f"temp_poc_images_{ts}"
            os.makedirs(temp_dir, exist_ok=True)
            
            for file_path in file_list:
                if file_path.startswith(poc_folder) and not file_path.endswith('/'):
                    # Get filename without extension
                    filename = os.path.basename(file_path)
                    name_without_ext = os.path.splitext(filename)[0]
                    
                    # Check if it's an image file
                    if any(filename.lower().endswith(ext) for ext in image_extensions):
                        # Extract the file to a temporary location
                        try:
                            with zip_ref.open(file_path) as f:
                                image_data = f.read()
                            
                            # Create temporary file
                            temp_file_path = os.path.join(temp_dir, filename)
                            
                            with open(temp_file_path, 'wb') as temp_file:
                                temp_file.write(image_data)
                            
                            # Map vulnerability name to image path(s)
                            # Support multiple images per vulnerability name
                            if name_without_ext not in poc_mapping:
                                poc_mapping[name_without_ext] = []
                            
                            # Check if this is a list or single value (for backward compatibility)
                            if isinstance(poc_mapping[name_without_ext], list):
                                poc_mapping[name_without_ext].append(temp_file_path)
                            else:
                                # Convert existing single value to list
                                poc_mapping[name_without_ext] = [poc_mapping[name_without_ext], temp_file_path]
                            
                            print(f"  📁 Extracted image: {filename}")
                            
                        except Exception as e:
                            print(f"Error extracting image {filename}: {e}")
                            continue
            
            # Calculate total number of images
            total_images = sum(len(v) if isinstance(v, list) else 1 for v in poc_mapping.values())
            print(f"Extracted {total_images} POC images for {len(poc_mapping)} unique vulnerability names")
            
    except Exception as e:
        print(f"Error extracting POC images: {e}")
    
    return poc_mapping

def normalize_vulnerability_name_for_filename(vuln_name):
    """
    Normalize vulnerability name to be compatible with file names by replacing
    invalid characters with dashes.
    
    Args:
        vuln_name (str): Original vulnerability name
        
    Returns:
        str: Normalized vulnerability name safe for file names
    """
    if not vuln_name:
        return vuln_name
    
    # Characters that are not allowed in file names
    invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
    
    normalized_name = str(vuln_name)
    
    # Replace invalid characters with dashes
    for char in invalid_chars:
        normalized_name = normalized_name.replace(char, '-')
    
    # Remove multiple consecutive dashes and trim
    normalized_name = re.sub(r'-+', '-', normalized_name)
    normalized_name = normalized_name.strip('-')
    
    return normalized_name

def load_external_images_by_name(image_directory="static/uploads"):
    """
    Load external image files and match them with vulnerability names using normalized names.
    
    Args:
        image_directory (str): Directory containing image files
        
    Returns:
        dict: Dictionary mapping vulnerability names to image data
    """
    import os
    import glob
    
    external_images = {}
    
    if not os.path.exists(image_directory):
        print(f"Image directory '{image_directory}' not found")
        return external_images
    
    # Supported image extensions
    image_extensions = ['*.jpg', '*.jpeg', '*.png', '*.bmp', '*.gif', '*.tiff']
    
    # Get all image files
    image_files = []
    for ext in image_extensions:
        image_files.extend(glob.glob(os.path.join(image_directory, ext)))
        image_files.extend(glob.glob(os.path.join(image_directory, ext.upper())))
    
    print(f"Found {len(image_files)} external image files in {image_directory}")
    
    for image_path in image_files:
        try:
            # Get filename without extension
            filename = os.path.splitext(os.path.basename(image_path))[0]
            
            # Read image data
            with open(image_path, 'rb') as f:
                image_data = f.read()
            
            # Store under the filename (which should be normalized)
            external_images[filename] = external_images.get(filename, [])
            external_images[filename].append(image_data)
            
            print(f"Loaded external image: {filename}")
            
        except Exception as e:
            print(f"Error loading external image {image_path}: {e}")
            continue
    
    return external_images

def insert_poc_images_to_excel(excel_path, poc_mapping, vulnerabilities_data):
    """Insert POC images into Excel file's POC columns (T-Z).
    Distributes images across columns: U, V, W, X, Y, Z, T (max 7 images per vulnerability).
    Returns a set of row numbers that have POC objects for border formatting"""
    rows_with_objects = set()  # Track which rows have POC objects
    
    print(f"\n🖼️ === INSERTING POC IMAGES ===")
    print(f"📁 Excel path: {excel_path}")
    # Calculate total number of images (handle both list and single value)
    total_images = sum(len(v) if isinstance(v, list) else 1 for v in poc_mapping.values())
    print(f"📊 POC mapping contains {len(poc_mapping)} unique vulnerability names with {total_images} total images")
    
    try:
        # Load external images and merge with existing poc_mapping
        external_images = load_external_images_by_name()
        
        # Convert external image data to temporary files and add to poc_mapping
        import tempfile
        import os
        
        # Track temporary files created for cleanup
        temp_files_created = []
        
        for vuln_name, image_data_list in external_images.items():
            for i, image_data in enumerate(image_data_list):
                try:
                    # Create temporary file for the image
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
                        tmp.write(image_data)
                        temp_path = tmp.name
                    
                    # Add to poc_mapping with the normalized name (support multiple images)
                    if vuln_name not in poc_mapping:
                        poc_mapping[vuln_name] = []
                    
                    # Check if this is a list or single value (for backward compatibility)
                    if isinstance(poc_mapping[vuln_name], list):
                        poc_mapping[vuln_name].append(temp_path)
                    else:
                        # Convert existing single value to list
                        poc_mapping[vuln_name] = [poc_mapping[vuln_name], temp_path]
                    
                    temp_files_created.append(temp_path)
                    print(f"Added external image to Excel processing: {vuln_name}")
                    
                except Exception as e:
                    print(f"Error processing external image {vuln_name}: {e}")
                    continue
        
        from openpyxl import load_workbook
        from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
        
        # Load the existing workbook
        wb = load_workbook(excel_path)
        
        # Get the Infra_VAPT worksheet
        if "Infra_VAPT" not in wb.sheetnames:
            print("Infra_VAPT worksheet not found")
            return rows_with_objects
        
        ws = wb["Infra_VAPT"]
        
        # Find POC columns by looking for merged "POC" header in first row
        poc_col_start = None
        poc_col_end = None
        
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == 1 and merged_range.max_row == 1:
                first_cell = ws.cell(row=1, column=merged_range.min_col)
                if first_cell.value and str(first_cell.value).strip() == "POC":
                    poc_col_start = merged_range.min_col
                    poc_col_end = merged_range.max_col
                    break
        
        if not poc_col_start or not poc_col_end:
            print("POC columns not found in worksheet")
            return rows_with_objects
        
        # Define column order for POC image insertion: U, V, W, X, Y, Z, T
        # POC columns are T-Z (7 columns): T, U, V, W, X, Y, Z
        # We want order: U, V, W, X, Y, Z, T
        image_columns = [
            poc_col_start + 1,  # U
            poc_col_start + 2,  # V
            poc_col_start + 3,  # W
            poc_col_start + 4,  # X
            poc_col_start + 5,  # Y
            poc_col_start + 6,  # Z (should be poc_col_end)
            poc_col_start       # T
        ]
        
        print(f"Found POC columns from {poc_col_start} to {poc_col_end}")
        
        # Process each row and match vulnerabilities with POC images
        for row in range(2, ws.max_row + 1):
            vulnerability_cell = ws.cell(row=row, column=2)  # Vulnerabilities column (column B)
            vulnerability_text = str(vulnerability_cell.value) if vulnerability_cell.value else ""
            
            if vulnerability_text:
                # Split vulnerabilities (they might be on separate lines)
                vulnerabilities = [v.strip() for v in vulnerability_text.split('\n') if v.strip()]
                
                # Find all matching POC images for this row
                matching_images = []
                
                # Normalize function for matching (same as Old POC)
                def normalize_for_matching(text):
                    """Normalize text for matching by replacing special chars with spaces"""
                    special_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|', '_', '-']
                    normalized = text.lower()
                    for char in special_chars:
                        normalized = normalized.replace(char, ' ')
                    # Normalize multiple spaces to single space
                    normalized = ' '.join(normalized.split())
                    return normalized
                
                def remove_trailing_numbers(text):
                    """Remove trailing numbers from text (only at the end, not in the middle)
                    Example: 'SSL Cert Expire 1' -> 'SSL Cert Expire'
                             'SSL Cert Expire 123' -> 'SSL Cert Expire'
                             'SSL 1 Cert Expire' -> 'SSL 1 Cert Expire' (unchanged)
                    """
                    if not text:
                        return text
                    text = text.strip()
                    # Use regex to remove trailing digits and any preceding whitespace
                    import re
                    # Remove trailing digits and whitespace before them
                    text = re.sub(r'\s+\d+$', '', text)
                    return text.strip()
                
                for vuln in vulnerabilities:
                    # Use only first 170 characters for matching
                    vuln_short = vuln[:170].strip()
                    
                    for image_name, image_path_or_list in poc_mapping.items():
                        # Handle both single image path and list of image paths
                        image_paths = image_path_or_list if isinstance(image_path_or_list, list) else [image_path_or_list]
                        
                        # Use only first 170 characters of image name for matching
                        image_name_short = image_name[:170].strip()
                        
                        # Normalize both names for comparison
                        vuln_normalized = normalize_for_matching(vuln_short)
                        image_normalized = normalize_for_matching(image_name_short)
                        
                        # Remove trailing numbers from image name for matching
                        # This allows "SSL Cert Expire 1" to match "SSL Cert Expire"
                        # Example: "SSL Cert Expire 1" -> "SSL Cert Expire" (matches)
                        #          "SSL Cert Expire 3" -> "SSL Cert Expire" (matches)
                        #          "SSL 1 Cert Expire" -> "SSL 1 Cert Expire" (no trailing number, unchanged)
                        image_normalized_no_trailing = remove_trailing_numbers(image_normalized)
                        
                        # Check if normalized names match (try both with and without trailing numbers)
                        matches = False
                        if vuln_normalized == image_normalized:
                            # Exact match
                            matches = True
                        elif vuln_normalized == image_normalized_no_trailing:
                            # Match after removing trailing numbers from image name
                            matches = True
                        
                        if matches:
                            # Add all images for this vulnerability (all images that match)
                            for image_path in image_paths:
                                if image_path not in [img[1] for img in matching_images]:  # Avoid duplicates
                                    matching_images.append((vuln, image_path))
                                    image_basename = os.path.basename(image_path) if os.path.exists(image_path) else image_path
                                    print(f"  🔗 Matched POC image: {image_basename} (vuln: '{vuln_short[:50]}...', image name: '{image_name_short[:50]}...')")
                
                if matching_images:
                    print(f"✅ Found {len(matching_images)} POC image(s) for row {row}")
                    try:
                        # Distribute images across columns: U, V, W, X, Y, Z, T
                        num_images_to_insert = min(len(matching_images), 7)  # Max 7 images (one per column)
                        
                        for img_idx in range(num_images_to_insert):
                            vuln, matching_image = matching_images[img_idx]
                            col_idx = image_columns[img_idx]
                            
                            if os.path.exists(matching_image):
                                try:
                                    # Load the image
                                    img = Image(matching_image)
                                    
                                    # Get original dimensions
                                    original_width = img.width
                                    original_height = img.height
                                    
                                    # Resize image to height=30px, width proportionally (maintain aspect ratio)
                                    if hasattr(img, 'width') and hasattr(img, 'height'):
                                        target_height = 30
                                        aspect_ratio = img.width / img.height if img.height else 1
                                        img.height = target_height
                                        img.width = int(target_height * aspect_ratio)
                                    
                                    # Get cell reference (e.g., "U2", "V3", etc.)
                                    from openpyxl.utils import get_column_letter
                                    col_letter = get_column_letter(col_idx)
                                    cell_ref = f"{col_letter}{row}"
                                    
                                    # Insert image at the cell
                                    ws.add_image(img, cell_ref)
                                    
                                    print(f"✅ Inserted POC image {img_idx + 1} at {cell_ref} for vulnerability: {vuln} (reduced from {original_width}x{original_height})")
                                    
                                except Exception as e:
                                    print(f"⚠️ Error inserting POC image at column {col_idx}, row {row}: {e}")
                        
                        # Track this row as having POC objects
                        if num_images_to_insert > 0:
                            rows_with_objects.add(row)
                            
                    except Exception as e:
                        print(f"Error adding images for row {row}: {e}")
                        continue
        
        # Apply custom borders to ALL POC columns (T to Z)
        from openpyxl.styles import Border, Side
        
        # Define border styles for each POC column type
        # T column: left, top, bottom (NOT right)
        t_border = Border(
            left=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # U, V, W, X, Y columns: top and bottom only (NOT left or right)
        middle_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Z column: top, bottom, right (NOT left)
        z_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            right=Side(style='thin')
        )
        
        # Find all rows that are part of the data table (have content in any column)
        table_rows = set()
        
        # Always include header row
        table_rows.add(1)
        
        # Check all rows from 2 onwards to find data rows
        for row in range(2, ws.max_row + 1):
            has_content = False
            
            # Check if any cell in this row has content (excluding POC columns)
            for col in range(1, poc_col_start):  # Check only columns before POC
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip() != "":
                    has_content = True
                    break
            
            if has_content:
                table_rows.add(row)
        
        # Apply custom borders to all table rows for each POC column
        for row_num in sorted(table_rows):
            # T column (poc_col_start): left, top, bottom
            ws.cell(row=row_num, column=poc_col_start).border = t_border
            
            # U, V, W, X, Y columns (middle columns): top, bottom only
            for col_idx in range(poc_col_start + 1, poc_col_end):
                ws.cell(row=row_num, column=col_idx).border = middle_border
            
            # Z column (poc_col_end): top, bottom, right
            ws.cell(row=row_num, column=poc_col_end).border = z_border
        
        print(f"Identified {len(table_rows)} table rows total")
        print(f"Applied custom borders to POC columns (T-Z) for {len(table_rows)} rows")
        
        # Save the workbook
        wb.save(excel_path)
        
        # Clean up temporary files created for external images
        for temp_file_path in temp_files_created:
            if os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                    print(f"Cleaned up temporary file: {temp_file_path}")
                except Exception as e:
                    print(f"Error cleaning up temporary file {temp_file_path}: {e}")
        
        print("POC images added successfully with dynamic sizing (50x reduction from original dimensions)")
        print(f"Added images to {len(rows_with_objects)} POC cells")
        print(f"Applied borders to {len(table_rows)} POC cells (all table rows including empty POC cells)")
        return rows_with_objects
        
    except Exception as e:
        print(f"Error inserting POC images: {e}")
        import traceback
        traceback.print_exc()
        return rows_with_objects

def generate_dynamic_filename(organization, end_date):
    """
    Generate dynamic filename based on end date.
    Format: Infrastructure_VAPT_Follow_Up_Report_<MONTH>_<YEAR>
    Example: Infrastructure_VAPT_Follow_Up_Report_Sep_2025
    """
    try:
        # Parse the end date (assuming format YYYY-MM-DD)
        if end_date:
            date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            month_name = date_obj.strftime('%b')
            year = date_obj.strftime('%Y')
        else:
            # Use current date if end_date is not provided
            current_date = datetime.now()
            month_name = current_date.strftime('%b')
            year = current_date.strftime('%Y')
        
        # Generate filename in the new format
        filename = f"Infrastructure_VAPT_Follow_Up_Report_{month_name}_{year}.xlsx"
        return filename
    except Exception as e:
        # Fallback to default filename if parsing fails
        current_date = datetime.now()
        month_name = current_date.strftime('%b')
        year = current_date.strftime('%Y')
        return f"Infrastructure_VAPT_Follow_Up_Report_{month_name}_{year}.xlsx"

# Import all the helper functions from the original file
# We'll need to copy all the processing functions from Infra_VAPT_First_Audit_Excel.py

def process_nmap_zip(file):
    """Process Nmap zip file and return data for Excel"""
    if file and file.filename.endswith('.zip'):
        zip_data = file.read()
        ip_ports = {}
        
        with zipfile.ZipFile(BytesIO(zip_data), 'r') as zip_ref:
            file_list = zip_ref.namelist()
            
            for file_name in file_list:
                if file_name.endswith('/'):
                    continue
                
                try:
                    with zip_ref.open(file_name) as f:
                        file_content = f.read().decode('utf-8', errors='ignore')
                    
                    ip_pattern = r"Nmap scan report for (?:[a-zA-Z0-9.-]+ )?\(?(\d+\.\d+\.\d+\.\d+)\)?"
                    port_state_pattern = r"(\d+)/(tcp|udp)\s+(open|filtered|closed|unfiltered)\s+([\w-]*)"
                    filtered_ports_pattern = r"Not shown: (\d+) filtered tcp ports"
                    
                    lines = file_content.splitlines()
                    current_ip = None
                    has_filtered_ports = False

                    for line in lines:
                        ip_match = re.search(ip_pattern, line, re.IGNORECASE)
                        if ip_match:
                            current_ip = ip_match.group(1)
                            if current_ip not in ip_ports:
                                ip_ports[current_ip] = []
                            has_filtered_ports = False
                            continue
                        
                        if not current_ip:
                            continue
                        
                        # Check for filtered ports message
                        filtered_match = re.search(filtered_ports_pattern, line)
                        if filtered_match:
                            has_filtered_ports = True
                            continue
                        
                        port_match = re.search(port_state_pattern, line)
                        if port_match:
                            port = port_match.group(1)
                            state = port_match.group(3)
                            service = port_match.group(4) or state
                            
                            if (port, service) not in ip_ports[current_ip]:
                                ip_ports[current_ip].append((port, service))
                    
                    # If no open ports found but filtered ports detected, add "Filtered" entry
                    if current_ip and has_filtered_ports and len(ip_ports[current_ip]) == 0:
                        ip_ports[current_ip].append(("Filtered", "Filtered"))

                    if file_name.endswith('.csv'):
                        try:
                            with zip_ref.open(file_name) as f:
                                csv_content = f.read()
                            
                            df = pd.read_csv(io.BytesIO(csv_content), 
                                        on_bad_lines="skip", 
                                        encoding="utf-8")
                            
                            if all(col in df.columns for col in ['host', 'port', 'service']):
                                for _, row in df.iterrows():
                                    ip = str(row['host']).strip()
                                    port = str(row['port']).strip()
                                    service = str(row['service']).strip()
                                    
                                    if ip and port and service:
                                        if ip not in ip_ports:
                                            ip_ports[ip] = []
                                        if (port, service) not in ip_ports[ip]:
                                            ip_ports[ip].append((port, service))
                        except Exception as e:
                            print(f"CSV processing error in {file_name}: {e}")
                            continue

                except Exception as e:
                    print(f"Error processing file {file_name}: {e}")
                    continue

        # Prepare data for Excel
        all_ips = sorted(ip_ports.keys())
        has_placeholder = False
        
        if len(all_ips) % 2 != 0:
            all_ips.append("-")
            ip_ports["-"] = [("-", "-")]
            has_placeholder = True

        data = []
        for i in range(0, len(all_ips), 2):
            ip1 = all_ips[i]
            ip2 = all_ips[i + 1]
            ports1 = ip_ports.get(ip1, [])
            ports2 = ip_ports.get(ip2, [])
            
            # If no ports found for an IP, add "Filtered" entry
            if len(ports1) == 0:
                ports1 = [("Filtered", "Filtered")]
            if len(ports2) == 0:
                ports2 = [("Filtered", "Filtered")]
            
            max_ports = max(len(ports1), len(ports2))

            data.append(["HOST", "PORT", "SERVICE", "HOST", "PORT", "SERVICE"])

            for j in range(max_ports):
                data.append([
                    ip1, 
                    ports1[j][0] if j < len(ports1) else "", 
                    ports1[j][1] if j < len(ports1) else "",
                    ip2, 
                    ports2[j][0] if j < len(ports2) else "", 
                    ports2[j][1] if j < len(ports2) else ""
                ])

        return data
    
    return None

def process_nessus_zip(file):
    """Process Nessus zip file and return list of DataFrames"""
    if file and file.filename.endswith('.zip'):
        zip_data = file.read()
        all_nessus_data = []
        
        with zipfile.ZipFile(BytesIO(zip_data), 'r') as zip_ref:
            file_list = zip_ref.namelist()
            nessus_files = [f for f in file_list if f.endswith('.csv') and not f.endswith('/')]
            
            for file_name in nessus_files:
                try:
                    base_name = os.path.basename(file_name)
                    first_part = base_name.split('_')[0]
                    
                    device_types = ["Server", "Switch", "Router", 'Firewall', 'Workstations', 'Access Points', 'CCTV']
                    found_device_type = None

                    for device_type in device_types:
                        if device_type in first_part or device_type.lower() in first_part.lower():
                            found_device_type = device_type
                            break

                    if found_device_type:
                        cleaned_name = first_part.replace(found_device_type, "").replace(found_device_type.lower(), "")
                        cleaned_name = cleaned_name.strip('_').strip()
                        
                        if cleaned_name:
                            branch_name = f"{cleaned_name} {found_device_type}"
                        else:
                            branch_name = found_device_type
                    else:
                        first_word = first_part.split(' ')[0]
                        branch_name = f"{first_word} Branch"

                    with zip_ref.open(file_name) as f:
                        csv_data = io.BytesIO(f.read())
                    
                    df = pd.read_csv(csv_data, 
                                   on_bad_lines='skip', 
                                   encoding='utf-8',
                                   keep_default_na=False,
                                   na_values=[])
                    
                    df['Branch Name'] = branch_name
                    all_nessus_data.append(df)
                
                except Exception as e:
                    print(f"Error processing Nessus file {file_name}: {e}")
                    continue

        return all_nessus_data
    
    return None

def create_meta_data_worksheet(workbook, form_metadata, header_format, cell_format, user_excel_file=None, generated_excel_path=None):
    """Create Meta_Data worksheet with form information from user input"""
    try:
        # Create Meta_Data worksheet
        worksheet_meta = workbook.add_worksheet("Meta_Data")
        
        # Get First Audit Report details from form (no longer from PDF)
        first_audit_report_id = form_metadata.get('firstAuditReportId', '')
        first_audit_report_date = form_metadata.get('firstAuditReportDate', '')
        
        # Format the date as DD.MM.YYYY if provided
        if first_audit_report_date:
            try:
                from datetime import datetime
                date_obj = datetime.strptime(first_audit_report_date, '%Y-%m-%d')
                first_audit_report_date = date_obj.strftime('%d.%m.%Y')
            except Exception as e:
                print(f"Error formatting first audit report date: {e}")
        
        # Extract risk factor counts from user's Excel file if provided
        risk_factor_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        if user_excel_file and user_excel_file.filename != '':
            try:
                print(f"Processing user's Excel file: {user_excel_file.filename}")
                risk_factor_counts = extract_risk_factor_counts_from_user_excel(user_excel_file)
                print(f"Extracted risk factor counts: {risk_factor_counts}")
            except Exception as e:
                print(f"Error extracting risk factor counts from user's Excel file: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("No user Excel file provided for risk factor extraction")
        
        # Extract follow-up risk factor counts from generated Excel file if provided
        follow_up_risk_factor_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        if generated_excel_path and os.path.exists(generated_excel_path):
            try:
                print(f"Processing generated Excel file: {generated_excel_path}")
                follow_up_risk_factor_counts = extract_follow_up_risk_factor_counts_from_generated_excel(generated_excel_path)
                print(f"Extracted follow-up risk factor counts: {follow_up_risk_factor_counts}")
            except Exception as e:
                print(f"Error extracting follow-up risk factor counts from generated Excel file: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("No generated Excel file provided for follow-up risk factor extraction")
        
        # Define the data structure for the metadata
        metadata_sections = [
            {
                'title': 'ORGANIZATION INFORMATION',
                'data': [
                    ('Organization Name', form_metadata.get('organization', '')),
                    ('City', form_metadata.get('city', '')),
                    ('State', form_metadata.get('state', '')),
                    ('First Audit Report ID', first_audit_report_id),
                    ('First Audit Report Date', first_audit_report_date)
                ]
            },
            {
                'title': 'AUDIT PERIOD',
                'data': [
                    ('Start Date', form_metadata.get('startDate', '')),
                    ('End Date', form_metadata.get('endDate', ''))
                ]
            },
            {
                'title': 'REPORT PREPARED BY',
                'data': [
                    ('Name', f"{form_metadata.get('preparedByTitle', '')} {form_metadata.get('preparedByName', '')}".strip()),
                ]
            },
            {
                'title': 'AUDITEE DETAILS',
                'data': [
                    ('Name', f"{form_metadata.get('auditeeTitle', '')} {form_metadata.get('auditeeName', '')}".strip()),
                    ('Designation', form_metadata.get('designation', ''))
                ]
            },
            {
                'title': 'FIRST AUDIT',
                'data': [
                    ('Critical', str(risk_factor_counts['Critical'])),
                    ('High', str(risk_factor_counts['High'])),
                    ('Medium', str(risk_factor_counts['Medium'])),
                    ('Low', str(risk_factor_counts['Low']))
                ]
            },
            {
                'title': 'FOLLOW UP AUDIT',
                'data': [
                    ('Critical', str(follow_up_risk_factor_counts['Critical'])),
                    ('High', str(follow_up_risk_factor_counts['High'])),
                    ('Medium', str(follow_up_risk_factor_counts['Medium'])),
                    ('Low', str(follow_up_risk_factor_counts['Low']))
                ]
            }
        ]
        
        # Add Bank Email Addresses section
        bank_emails = form_metadata.get('bankEmails', [])
        if bank_emails:
            bank_email_data = []
            for i, email in enumerate(bank_emails, 1):
                if email.strip():
                    bank_email_data.append((f'Email {i}', email.strip()))
            
            if bank_email_data:
                metadata_sections.append({
                    'title': 'BANK EMAIL ADDRESSES',
                    'data': bank_email_data
                })
        
        # Add Auditing Team section
        team_names = form_metadata.get('teamNames', [])
        team_designations = form_metadata.get('teamDesignations', [])
        team_emails = form_metadata.get('teamEmails', [])
        team_qualifications = form_metadata.get('teamQualifications', [])
        team_certified = form_metadata.get('teamCertified', [])
        
        if team_names:
            # Add team member details as separate entries for each member
            for i in range(len(team_names)):
                if team_names[i].strip():
                    team_member_data = [
                        (f'Team Member {i+1} - Name', team_names[i].strip()),
                        (f'Team Member {i+1} - Designation', team_designations[i] if i < len(team_designations) else ''),
                        (f'Team Member {i+1} - Email', team_emails[i] if i < len(team_emails) else ''),
                        (f'Team Member {i+1} - Qualification', team_qualifications[i] if i < len(team_qualifications) else ''),
                        (f'Team Member {i+1} - Certified', team_certified[i] if i < len(team_certified) else '')
                    ]
                    
                    metadata_sections.append({
                        'title': f'AUDITING TEAM MEMBER {i+1}',
                        'data': team_member_data
                    })
        
        # Write data to worksheet
        row = 0
        
        for section_idx, section in enumerate(metadata_sections):
            # Write section title
            worksheet_meta.write(row, 0, section['title'], header_format)
            worksheet_meta.write(row, 1, '', header_format)  # Empty cell for formatting
            row += 1
            
            # Write section data
            for field_name, field_value in section['data']:
                print(f"Writing to worksheet - Row {row+1} (Excel), Field: '{field_name}', Value: '{field_value}'")
                worksheet_meta.write(row, 0, field_name, cell_format)
                worksheet_meta.write(row, 1, field_value, cell_format)
                row += 1
            
            # Add empty row after each section
            row += 1
        
        # Set column widths
        worksheet_meta.set_column('A:A', 30)  # Field names column
        worksheet_meta.set_column('B:B', 60)  # Values column
        
        # Set row heights for better readability
        for r in range(row):
            worksheet_meta.set_row(r, 20)
        
        print(f"Created Meta_Data worksheet with {row} rows")
        
    except Exception as e:
        print(f"Error creating Meta_Data worksheet: {e}")
        import traceback
        traceback.print_exc()

@follow_up_audit_bp.route('/followup_check_vulnerabilities', methods=['POST'])
def followup_check_vulnerabilities():
    """Return both matched and unmatched vulnerabilities with full catalog details for merge management (Follow-up Audit)."""
    if 'nmapFiles' not in request.files or 'nessusFiles' not in request.files:
        return jsonify({"error": "Both Nmap and Nessus files are required"}), 400
    
    nmap_file = request.files['nmapFiles']
    nessus_file = request.files['nessusFiles']
    
    if nmap_file.filename == '' or nessus_file.filename == '':
        return jsonify({"error": "No file selected"}), 400
    
    try:
        # Process Nessus files to check vulnerabilities
        nessus_dataframes = process_nessus_zip(nessus_file)
        
        if nessus_dataframes:
            combined_nessus = pd.concat(nessus_dataframes, ignore_index=True)
            
            # Don't store the full Nessus data in session - it's too large
            
            # Calculate matched and unmatched vulnerabilities
            matched_groups = []
            unmatched_vulnerabilities = []
            
            try:
                # Filter by valid risks and normalize
                valid_risks = ['low', 'medium', 'high', 'critical']
                df_filtered = combined_nessus.copy()
                df_filtered['Risk'] = df_filtered['Risk'].astype(str).str.lower().str.strip()
                df_filtered = df_filtered[df_filtered['Risk'].isin(valid_risks)]

                # Get unique vulnerability names from Name column
                unique_vulnerabilities_list = df_filtered['Name'].dropna().drop_duplicates().astype(str).str.strip().tolist()
                unique_vulnerabilities = set(unique_vulnerabilities_list)
                
                # Load catalog to get matched vulnerabilities with full details
                catalog_path = "static/Formats_and_Catalog/Infrastructure VAPT Catalog.xlsx"
                if os.path.exists(catalog_path):
                    try:
                        catalog_df = pd.read_excel(catalog_path, sheet_name=0)
                    except Exception as e:
                        print(f"Error reading catalog file for vulnerability check: {e}")
                        catalog_df = None
                    
                    if catalog_df is not None and 'Vulnerabilities in this group' in catalog_df.columns:
                        matched_vulnerability_names = set()
                        
                        # Build matched groups with catalog details
                        for idx, row in catalog_df.iterrows():
                            vulnerabilities_in_group = str(row.get('Vulnerabilities in this group', '')).strip()
                            if pd.isna(vulnerabilities_in_group) or vulnerabilities_in_group == '':
                                continue
                            
                            # Split vulnerabilities by newlines
                            vuln_list = [v.strip() for v in vulnerabilities_in_group.split('\n') if v.strip()]
                            
                            # Find which vulnerabilities from Excel match this catalog group
                            matched_vulns_in_group = []
                            for vulnerability in unique_vulnerabilities:
                                escaped_vulnerability = re.escape(str(vulnerability))
                                pattern = rf'(?:\n|\r\n|\A){escaped_vulnerability}(?:\n|\r\n|\Z)'
                                if re.search(pattern, vulnerabilities_in_group, re.IGNORECASE):
                                    matched_vulns_in_group.append(vulnerability)
                                    matched_vulnerability_names.add(vulnerability)
                            
                            # If any vulnerabilities matched this catalog group, add it
                            # Store only essential data to keep session size small
                            if matched_vulns_in_group:
                                matched_groups.append({
                                    'catalog_id': int(idx),
                                    'group_name': str(row.get('Name of Vulnerability', ''))[:200],
                                    'risk_factor': str(row.get('Risk Factor', ''))[:20],
                                    'cvss_score': str(row.get('CVSS', ''))[:10],
                                    'matched_vulnerabilities': matched_vulns_in_group
                                })
                        
                        # Calculate unmatched vulnerabilities
                        unmatched_vulnerabilities = sorted(list(unique_vulnerabilities - matched_vulnerability_names))
                    else:
                        unmatched_vulnerabilities = sorted(list(unique_vulnerabilities))
                else:
                    unmatched_vulnerabilities = sorted(list(unique_vulnerabilities))
                        
            except Exception as e:
                print(f"Error calculating vulnerabilities: {e}")
                import traceback
                traceback.print_exc()
                return jsonify({"error": f"Error processing vulnerabilities: {str(e)}"}), 500
            
            # Initialize merge state in session with size limits (Follow-up specific)
            unmatched_limited = unmatched_vulnerabilities[:100] if len(unmatched_vulnerabilities) > 100 else unmatched_vulnerabilities
            
            session['followup_vulnerability_merge_state'] = {
                'matched_groups': matched_groups[:50],
                'unmatched_vulnerabilities': unmatched_limited,
                'merge_operations': [],
                'new_group_details': {}
            }
            
            if len(unmatched_vulnerabilities) > 100:
                print(f"⚠️ Warning: {len(unmatched_vulnerabilities)} unmatched vulnerabilities found. Limited to 100 in session.")
            
            # Return matched groups and unmatched vulnerabilities
            return jsonify({
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_vulnerabilities
            })
        else:
            return jsonify({"error": "No Nessus data found"}), 400
            
    except Exception as e:
        print(f"Error checking vulnerabilities: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error processing files: {str(e)}"}), 500

@follow_up_audit_bp.route('/submit_vulnerability_details', methods=['POST'])
def submit_vulnerability_details():
    """Handle submission of user-provided details for unmatched vulnerabilities."""
    try:
        data = request.get_json()
        
        if not data or 'vulnerability_details' not in data:
            return jsonify({"error": "No vulnerability details provided"}), 400
        
        vulnerability_details = data['vulnerability_details']
        
        # Validate required fields
        required_fields = ['vulnerabilityName', 'riskFactor', 'cveId', 'cvssScore', 
                          'auditObservation', 'impact', 'recommendation', 'referenceLink']
        
        for vuln_name, details in vulnerability_details.items():
            for field in required_fields:
                if field not in details:
                    return jsonify({"error": f"Missing required field '{field}' for vulnerability '{vuln_name}'"}), 400
        
        # Update catalog with all vulnerabilities (merged and separate)
        update_catalog_with_vulnerabilities(vulnerability_details)
        
        # Store the details in session for use in report generation
        session['unmatched_vulnerability_details'] = vulnerability_details
        
        return jsonify({"success": True, "message": "Vulnerability details saved successfully"})
        
    except Exception as e:
        print(f"Error saving vulnerability details: {e}")
        return jsonify({"error": f"Error saving vulnerability details: {str(e)}"}), 500


@follow_up_audit_bp.route('/followup_merge_with_matched', methods=['POST'])
def followup_merge_with_matched():
    """Merge an unmatched vulnerability with an existing matched group (Follow-up Audit)."""
    try:
        data = request.get_json()
        
        if not data or 'unmatched_vulnerability' not in data or 'target_group_id' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        unmatched_vuln = data['unmatched_vulnerability']
        target_group_id = data['target_group_id']
        
        # Get current merge state from session (Follow-up specific key)
        merge_state = session.get('followup_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Find the target matched group
        matched_groups = merge_state.get('matched_groups', [])
        target_group = None
        
        for group in matched_groups:
            if group['catalog_id'] == target_group_id:
                target_group = group
                break
        
        if target_group is None:
            return jsonify({"error": "Target group not found"}), 404
        
        # Add the unmatched vulnerability to the matched group
        if unmatched_vuln not in target_group['matched_vulnerabilities']:
            target_group['matched_vulnerabilities'].append(unmatched_vuln)
        
        # Remove from unmatched list
        unmatched_list = merge_state.get('unmatched_vulnerabilities', [])
        if unmatched_vuln in unmatched_list:
            unmatched_list.remove(unmatched_vuln)
        
        # Record the merge operation for undo
        merge_state['merge_operations'].append({
            'type': 'merge_with_matched',
            'unmatched_vulnerability': unmatched_vuln,
            'target_group_id': target_group_id,
            'timestamp': datetime.now().isoformat()
        })
        
        # Update session
        session['followup_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        return jsonify({
            "success": True,
            "message": "Vulnerability merged successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_list
            }
        })
        
    except Exception as e:
        print(f"Error merging with matched: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error merging: {str(e)}"}), 500


@follow_up_audit_bp.route('/followup_merge_with_unmatched', methods=['POST'])
def followup_merge_with_unmatched():
    """Merge multiple unmatched vulnerabilities into a new group (Follow-up Audit)."""
    try:
        data = request.get_json()
        
        if not data or 'vulnerabilities' not in data or 'vulnerability_details' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        vulnerabilities_to_merge = data['vulnerabilities']
        vulnerability_details = data['vulnerability_details']
        
        # Validate required fields
        required_fields = ['vulnerabilityName', 'riskFactor', 'cveId', 'cvssScore', 
                          'auditObservation', 'impact', 'recommendation', 'referenceLink']
        
        for field in required_fields:
            if field not in vulnerability_details:
                return jsonify({"error": f"Missing required field '{field}'"}), 400
        
        # Get current merge state from session
        merge_state = session.get('followup_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Create a new matched group from the merged vulnerabilities
        new_group_id = -len(merge_state.get('matched_groups', [])) - 1
        new_group = {
            'catalog_id': new_group_id,
            'group_name': vulnerability_details['vulnerabilityName'][:200],
            'risk_factor': vulnerability_details['riskFactor'][:20],
            'cvss_score': vulnerability_details['cvssScore'][:10],
            'matched_vulnerabilities': vulnerabilities_to_merge,
            'is_new_group': True
        }
        
        # Add to matched groups
        matched_groups = merge_state.get('matched_groups', [])
        matched_groups.append(new_group)
        
        # Store full details separately
        new_group_details = merge_state.get('new_group_details', {})
        new_group_details[str(new_group_id)] = vulnerability_details
        merge_state['new_group_details'] = new_group_details
        
        # Remove from unmatched list
        unmatched_list = merge_state.get('unmatched_vulnerabilities', [])
        for vuln in vulnerabilities_to_merge:
            if vuln in unmatched_list:
                unmatched_list.remove(vuln)
        
        # Record the merge operation for undo
        merge_state['merge_operations'].append({
            'type': 'merge_with_unmatched',
            'vulnerabilities': vulnerabilities_to_merge,
            'new_group_id': new_group_id,
            'timestamp': datetime.now().isoformat()
        })
        
        # Update session
        session['followup_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        # Update catalog with new vulnerability group
        # Add the merged vulnerabilities list to the details
        vulnerability_details_with_merge = vulnerability_details.copy()
        vulnerability_details_with_merge['isMerged'] = True
        vulnerability_details_with_merge['mergedVulnerabilities'] = vulnerabilities_to_merge
        
        update_catalog_with_vulnerabilities({
            vulnerability_details['vulnerabilityName']: vulnerability_details_with_merge
        })
        
        return jsonify({
            "success": True,
            "message": "Vulnerabilities merged into new group successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_list
            }
        })
        
    except Exception as e:
        print(f"Error merging unmatched vulnerabilities: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error merging: {str(e)}"}), 500


@follow_up_audit_bp.route('/followup_add_vulnerability_details', methods=['POST'])
def followup_add_vulnerability_details():
    """Add details for a single unmatched vulnerability (Follow-up Audit)."""
    try:
        data = request.get_json()
        
        if not data or 'vulnerability_name' not in data or 'vulnerability_details' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        vulnerability_name = data['vulnerability_name']
        vulnerability_details = data['vulnerability_details']
        
        # Validate required fields
        required_fields = ['vulnerabilityName', 'riskFactor', 'cveId', 'cvssScore', 
                          'auditObservation', 'impact', 'recommendation', 'referenceLink']
        
        for field in required_fields:
            if field not in vulnerability_details:
                return jsonify({"error": f"Missing required field '{field}'"}), 400
        
        # Get current merge state from session
        merge_state = session.get('followup_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Create a new matched group for this single vulnerability
        new_group_id = -len(merge_state.get('matched_groups', [])) - 1
        new_group = {
            'catalog_id': new_group_id,
            'group_name': vulnerability_details['vulnerabilityName'][:200],
            'risk_factor': vulnerability_details['riskFactor'][:20],
            'cvss_score': vulnerability_details['cvssScore'][:10],
            'matched_vulnerabilities': [vulnerability_name],
            'is_new_group': True
        }
        
        # Add to matched groups
        matched_groups = merge_state.get('matched_groups', [])
        matched_groups.append(new_group)
        
        # Store full details separately
        new_group_details = merge_state.get('new_group_details', {})
        new_group_details[str(new_group_id)] = vulnerability_details
        merge_state['new_group_details'] = new_group_details
        
        # Remove from unmatched list
        unmatched_list = merge_state.get('unmatched_vulnerabilities', [])
        if vulnerability_name in unmatched_list:
            unmatched_list.remove(vulnerability_name)
        
        # Record the operation for undo
        merge_state['merge_operations'].append({
            'type': 'add_details',
            'vulnerability': vulnerability_name,
            'new_group_id': new_group_id,
            'timestamp': datetime.now().isoformat()
        })
        
        # Update session
        session['followup_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        # Update catalog (single vulnerability - not merged)
        vulnerability_details_single = vulnerability_details.copy()
        vulnerability_details_single['isMerged'] = False
        vulnerability_details_single['actualVulnerabilityName'] = vulnerability_name  # Store the actual vulnerability name
        
        update_catalog_with_vulnerabilities({
            vulnerability_details['vulnerabilityName']: vulnerability_details_single
        })
        
        return jsonify({
            "success": True,
            "message": "Vulnerability details added successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_list
            }
        })
        
    except Exception as e:
        print(f"Error adding vulnerability details: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error adding details: {str(e)}"}), 500


@follow_up_audit_bp.route('/followup_merge_matched_groups', methods=['POST'])
def followup_merge_matched_groups():
    """Merge two matched groups together (Follow-up Audit)."""
    try:
        data = request.get_json()
        
        if not data or 'source_group_id' not in data or 'target_group_id' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        source_group_id = data['source_group_id']
        target_group_id = data['target_group_id']
        
        # Get current merge state from session
        merge_state = session.get('followup_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Find both groups
        matched_groups = merge_state.get('matched_groups', [])
        source_group = None
        target_group = None
        source_index = None
        
        for idx, group in enumerate(matched_groups):
            if group['catalog_id'] == source_group_id:
                source_group = group
                source_index = idx
            elif group['catalog_id'] == target_group_id:
                target_group = group
        
        if source_group is None or target_group is None:
            return jsonify({"error": "One or both groups not found"}), 404
        
        # Merge source group vulnerabilities into target group
        for vuln in source_group['matched_vulnerabilities']:
            if vuln not in target_group['matched_vulnerabilities']:
                target_group['matched_vulnerabilities'].append(vuln)
        
        # Remove source group
        matched_groups.pop(source_index)
        
        # Record the merge operation for undo
        merge_state['merge_operations'].append({
            'type': 'merge_matched_groups',
            'source_group_id': source_group_id,
            'target_group_id': target_group_id,
            'source_group_data': source_group,
            'timestamp': datetime.now().isoformat()
        })
        
        # Update session
        session['followup_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        return jsonify({
            "success": True,
            "message": "Groups merged successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": merge_state.get('unmatched_vulnerabilities', [])
            }
        })
        
    except Exception as e:
        print(f"Error merging matched groups: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error merging groups: {str(e)}"}), 500


@follow_up_audit_bp.route('/followup_undo_last_merge', methods=['POST'])
def followup_undo_last_merge():
    """Undo the last merge operation (Follow-up Audit)."""
    try:
        # Get current merge state from session
        merge_state = session.get('followup_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        merge_operations = merge_state.get('merge_operations', [])
        if not merge_operations:
            return jsonify({"error": "No operations to undo"}), 400
        
        # Get the last operation
        last_operation = merge_operations.pop()
        operation_type = last_operation['type']
        
        matched_groups = merge_state.get('matched_groups', [])
        unmatched_list = merge_state.get('unmatched_vulnerabilities', [])
        
        # Undo based on operation type
        if operation_type == 'merge_with_matched':
            target_group_id = last_operation['target_group_id']
            unmatched_vuln = last_operation['unmatched_vulnerability']
            
            for group in matched_groups:
                if group['catalog_id'] == target_group_id:
                    if unmatched_vuln in group['matched_vulnerabilities']:
                        group['matched_vulnerabilities'].remove(unmatched_vuln)
                    break
            
            if unmatched_vuln not in unmatched_list:
                unmatched_list.append(unmatched_vuln)
                unmatched_list.sort()
        
        elif operation_type == 'merge_with_unmatched':
            new_group_id = last_operation['new_group_id']
            vulnerabilities = last_operation['vulnerabilities']
            
            matched_groups = [g for g in matched_groups if g['catalog_id'] != new_group_id]
            
            for vuln in vulnerabilities:
                if vuln not in unmatched_list:
                    unmatched_list.append(vuln)
            unmatched_list.sort()
        
        elif operation_type == 'add_details':
            new_group_id = last_operation['new_group_id']
            vulnerability = last_operation['vulnerability']
            
            matched_groups = [g for g in matched_groups if g['catalog_id'] != new_group_id]
            
            if vulnerability not in unmatched_list:
                unmatched_list.append(vulnerability)
                unmatched_list.sort()
        
        elif operation_type == 'merge_matched_groups':
            source_group_data = last_operation['source_group_data']
            target_group_id = last_operation['target_group_id']
            
            matched_groups.append(source_group_data)
            
            for group in matched_groups:
                if group['catalog_id'] == target_group_id:
                    for vuln in source_group_data['matched_vulnerabilities']:
                        if vuln in group['matched_vulnerabilities']:
                            if vuln not in source_group_data.get('all_catalog_vulnerabilities', []):
                                continue
                            group['matched_vulnerabilities'].remove(vuln)
                    break
        
        # Update session
        merge_state['matched_groups'] = matched_groups
        merge_state['unmatched_vulnerabilities'] = unmatched_list
        merge_state['merge_operations'] = merge_operations
        session['followup_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        return jsonify({
            "success": True,
            "message": "Last operation undone successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_list
            }
        })
        
    except Exception as e:
        print(f"Error undoing merge: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error undoing: {str(e)}"}), 500


def update_catalog_with_vulnerabilities(vulnerability_details):
    """Update the Infrastructure VAPT Catalog with both merged and separate vulnerabilities."""
    try:
        catalog_path = "static/Formats_and_Catalog/Infrastructure VAPT Catalog.xlsx"
        
        if not os.path.exists(catalog_path):
            print(f"Catalog file not found at: {catalog_path}")
            return
        
        # Read the existing catalog with error handling - now reading Sheet2 (index 1)
        try:
            catalog_df = pd.read_excel(catalog_path, sheet_name=1)  # Changed from sheet_name=0 to sheet_name=1 for Sheet2
        except Exception as e:
            print(f"Error reading catalog file: {e}")
            print(f"Catalog file may be corrupted. Please check: {catalog_path}")
            return
        
        # Process all vulnerabilities (both merged and separate)
        for vuln_name, details in vulnerability_details.items():
            if details.get('isMerged', False):
                # Handle merged vulnerabilities
                merged_vulns = details.get('mergedVulnerabilities', [])
                if merged_vulns:
                    # Create a new row for the merged vulnerability group
                    new_row = {
                        'Sr No': len(catalog_df) + 1,
                        'Name of Vulnerability': details.get('vulnerabilityName', ''),
                        'Risk Factor': details.get('riskFactor', ''),
                        'CVE ID': details.get('cveId', ''),
                        'CVSS': details.get('cvssScore', ''),
                        'Audit Observation': details.get('auditObservation', ''),
                        'Impact': details.get('impact', ''),
                        'Recommendation/Countermeasure': details.get('recommendation', ''),
                        'Affected System': '',  # Empty as requested
                        'Reference Link': details.get('referenceLink', ''),
                        'Vulnerabilities in this group': '\n'.join(merged_vulns),
                        'User_name': current_user.employee_name if current_user.is_authenticated else 'Unknown',
                        'Time_stamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    # Add the new row to the catalog
                    catalog_df = pd.concat([catalog_df, pd.DataFrame([new_row])], ignore_index=True)
            else:
                # Handle separate (non-merged) vulnerabilities
                # Get the actual vulnerability name (might be different from group name)
                actual_vuln_name = details.get('actualVulnerabilityName', vuln_name)
                
                new_row = {
                    'Sr No': len(catalog_df) + 1,
                    'Name of Vulnerability': details.get('vulnerabilityName', vuln_name),  # User-provided group name
                    'Risk Factor': details.get('riskFactor', ''),
                    'CVE ID': details.get('cveId', ''),
                    'CVSS': details.get('cvssScore', ''),
                    'Audit Observation': details.get('auditObservation', ''),
                    'Impact': details.get('impact', ''),
                    'Recommendation/Countermeasure': details.get('recommendation', ''),
                    'Affected System': '',  # Empty as requested
                    'Reference Link': details.get('referenceLink', ''),
                    'Vulnerabilities in this group': actual_vuln_name,  # Actual vulnerability name from scan
                    'User_name': current_user.employee_name if current_user.is_authenticated else 'Unknown',
                    'Time_stamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                # Add the new row to the catalog
                catalog_df = pd.concat([catalog_df, pd.DataFrame([new_row])], ignore_index=True)
        
        # Save the updated catalog to Sheet2
        # Read all existing sheets first
        try:
            # Read all sheets to preserve existing data
            all_sheets = pd.read_excel(catalog_path, sheet_name=None)
            
            # Update Sheet2 with our new data
            all_sheets['Sheet2'] = catalog_df
            
            # Write all sheets back to the file
            with pd.ExcelWriter(catalog_path, engine='openpyxl') as writer:
                for sheet_name, sheet_data in all_sheets.items():
                    sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            print(f"Error saving catalog with multiple sheets: {e}")
            # Fallback: just save Sheet2
            catalog_df.to_excel(catalog_path, sheet_name='Sheet2', index=False)
        merged_count = len([v for v in vulnerability_details.values() if v.get('isMerged', False)])
        separate_count = len([v for v in vulnerability_details.values() if not v.get('isMerged', False)])
        print(f"Updated catalog with {merged_count} merged vulnerability groups and {separate_count} separate vulnerabilities")
        
    except Exception as e:
        print(f"Error updating catalog: {e}")
        import traceback
        traceback.print_exc()

@follow_up_audit_bp.route('/process_follow_up_audit_report', methods=['POST'])
def process_follow_up_audit_report():
    """Process follow-up audit report with form data for organization, city, and state."""
    try:
        if 'nmapFiles' not in request.files or 'nessusFiles' not in request.files:
            return "Both Nmap and Nessus files are required", 400
        
        nmap_file = request.files['nmapFiles']
        nessus_file = request.files['nessusFiles']
        evidence_file = request.files.get('evidenceFiles')  # Get evidence files if present
        user_excel_file = request.files.get('userExcelFile')  # Get user's Excel file for comparison
        
        if nmap_file.filename == '' or nessus_file.filename == '':
            return "No file selected", 400
        
        if not user_excel_file or user_excel_file.filename == '':
            return "User's Excel file is required for comparison", 400
        
        # Get organization, city, and state from form (with "Other" handling)
        organization = request.form.get('organization', '')
        if organization == 'other':
            organization = request.form.get('otherOrganization', '')
        
        city = request.form.get('city', '')
        if city == 'other':
            city = request.form.get('otherCity', '')
        
        state = request.form.get('state', '')
        
        # Get First Audit Report details
        first_audit_report_id = request.form.get('firstAuditReportId', '')
        first_audit_report_date = request.form.get('firstAuditReportDate', '')
        
        print(f"Form Data - Organization: {organization}, City: {city}, State: {state}")
        print(f"First Audit Report - ID: {first_audit_report_id}, Date: {first_audit_report_date}")
        
        # Capture form metadata
        form_metadata = {
            'organization': organization,  # From form
            'city': city,  # From form
            'state': state,  # From form
            'firstAuditReportId': first_audit_report_id,  # New field
            'firstAuditReportDate': first_audit_report_date,  # New field
            'startDate': request.form.get('startDate', ''),
            'endDate': request.form.get('endDate', ''),
            'preparedByTitle': request.form.get('preparedByTitle', ''),
            'preparedByName': request.form.get('preparedByName', ''),
            'auditeeTitle': request.form.get('auditeeTitle', ''),
            'auditeeName': request.form.get('auditeeName', ''),
            'designation': request.form.get('designation', ''),
            'bankEmails': request.form.getlist('bankEmail[]'),
            'teamNames': request.form.getlist('teamName[]'),
            'teamDesignations': request.form.getlist('teamDesignation[]'),
            'teamEmails': request.form.getlist('teamEmail[]'),
            'teamQualifications': request.form.getlist('teamQualification[]'),
            'teamCertified': []
        }
        
        # Handle indexed teamCertified radio buttons
        team_certified = []
        i = 0
        while True:
            certified_value = request.form.get(f'teamCertified[{i}]')
            if certified_value is None:
                break
            team_certified.append(certified_value)
            i += 1
        form_metadata['teamCertified'] = team_certified
        
        # Process both files
        nmap_data = process_nmap_zip(nmap_file)
        nessus_dataframes = process_nessus_zip(nessus_file)
        
        # Initialize unmatched_count early
        unmatched_count = 0
        
        # Create Excel file in memory with nan_inf_to_errors option
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
            workbook = writer.book
            
            # Define formats
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'vcenter',
                'align': 'center',
                'fg_color': "#3553E8",
                'font_color': 'white',
                'border': 1
            })  
            
            cell_format = workbook.add_format({
                'text_wrap': True,
                'valign': 'vcenter',
                'align': 'center',
                'border': 1
            })
            
            # Create Meta_Data worksheet as the first worksheet
            create_meta_data_worksheet(workbook, form_metadata, header_format, cell_format, user_excel_file)
            
            # Process Nmap data
            if nmap_data:
                df_nmap = pd.DataFrame(nmap_data)
                df_nmap.to_excel(writer, sheet_name='Nmap Files', index=False, header=False)
                
                worksheet_nmap = writer.sheets['Nmap Files']
                worksheet_nmap.set_column('A:F', 20)
                
                # Apply formats to Nmap worksheet
                for row_num, row_data in enumerate(nmap_data):
                    fmt = header_format if row_data == ["HOST", "PORT", "SERVICE", "HOST", "PORT", "SERVICE"] else cell_format
                    for col_num, value in enumerate(row_data):
                        worksheet_nmap.write(row_num, col_num, value, fmt)
                
                # Track IP positions for proper merging
                ip_positions = {"A": {}, "D": {}}
                
                # First pass: identify all IP positions
                for row_num in range(1, len(nmap_data)):
                    ip1 = nmap_data[row_num][0]
                    ip2 = nmap_data[row_num][3]
                    
                    if ip1 not in ip_positions["A"]:
                        ip_positions["A"][ip1] = []
                    if ip2 not in ip_positions["D"]:
                        ip_positions["D"][ip2] = []
                    
                    ip_positions["A"][ip1].append(row_num)
                    ip_positions["D"][ip2].append(row_num)
                
                # Second pass: merge cells for duplicate IPs (simplified approach to avoid overlaps)
                # Instead of complex overlap detection, we'll use a simpler approach
                # that merges only consecutive rows for each IP
                
                for ip, positions in ip_positions["A"].items():
                    if len(positions) > 1:
                        # Sort positions to ensure they're in order
                        positions.sort()
                        
                        # Group consecutive positions
                        groups = []
                        current_group = [positions[0]]
                        
                        for i in range(1, len(positions)):
                            if positions[i] == positions[i-1] + 1:
                                current_group.append(positions[i])
                        else:
                                groups.append(current_group)
                                current_group = [positions[i]]
                        groups.append(current_group)
                        
                        # Merge each group
                        for group in groups:
                            if len(group) > 1:
                                start_row = group[0]
                                end_row = group[-1]
                                try:
                                    worksheet_nmap.merge_range(start_row, 0, end_row, 0, ip, cell_format)
                                except Exception as e:
                                    print(f"Warning: Could not merge range A{start_row+1}:A{end_row+1} for IP {ip}: {e}")
                
                for ip, positions in ip_positions["D"].items():
                    if len(positions) > 1:
                        # Sort positions to ensure they're in order
                        positions.sort()
                        
                        # Group consecutive positions
                        groups = []
                        current_group = [positions[0]]
                        
                        for i in range(1, len(positions)):
                            if positions[i] == positions[i-1] + 1:
                                current_group.append(positions[i])
                        else:
                                groups.append(current_group)
                                current_group = [positions[i]]
                        groups.append(current_group)
                        
                        # Merge each group
                        for group in groups:
                            if len(group) > 1:
                                start_row = group[0]
                                end_row = group[-1]
                                try:
                                    worksheet_nmap.merge_range(start_row, 3, end_row, 3, ip, cell_format)
                                except Exception as e:
                                    print(f"Warning: Could not merge range D{start_row+1}:D{end_row+1} for IP {ip}: {e}")
                
                # Merge only empty PORT/SERVICE cells (not "Filtered" entries)
                columns_to_merge = [1, 2, 4, 5]
                for col in columns_to_merge:
                    merge_start = None
                    prev_value = None

                    for row in range(1, len(nmap_data)):
                        current_value = nmap_data[row][col]
                        
                        # Only merge truly empty cells, not "Filtered" entries
                        if current_value == "" and merge_start is None:
                            merge_start = row - 1
                            prev_value = nmap_data[merge_start][col]
                        # End merge when we hit a non-empty cell (including "Filtered")
                        elif current_value != "" and merge_start is not None:
                            if row - 1 > merge_start:
                                try:
                                    worksheet_nmap.merge_range(merge_start, col, row - 1, col, prev_value, cell_format)
                                except Exception as e:
                                    print(f"Warning: Could not merge range {chr(65+col)}{merge_start+1}:{chr(65+col)}{row} for column {col}: {e}")
                            merge_start = None
                            prev_value = None
                    
                    # Handle merge at the end of the data
                    if merge_start is not None and merge_start < len(nmap_data) - 1:
                        try:
                            worksheet_nmap.merge_range(merge_start, col, len(nmap_data) - 1, col, prev_value, cell_format)
                        except Exception as e:
                            print(f"Warning: Could not merge range {chr(65+col)}{merge_start+1}:{chr(65+col)}{len(nmap_data)} for column {col}: {e}")
            
            # Process Nessus data
            if nessus_dataframes:
                combined_nessus = pd.concat(nessus_dataframes, ignore_index=True)

                # Convert INF values to string representation
                for col in combined_nessus.columns:
                    if combined_nessus[col].dtype == 'float64':
                        combined_nessus[col] = combined_nessus[col].apply(
                            lambda x: 'INF' if math.isinf(x) else x
                        )
                
                worksheet_nessus = workbook.add_worksheet("Nessus CSV Files")
                writer.sheets["Nessus CSV Files"] = worksheet_nessus
                
                # Create cell formats (with and without border)
                wrapped_cell_with_border = workbook.add_format({
                    'text_wrap': True,
                    'valign': 'vtop',
                    'align': 'center',
                    'border': 1
                })
                wrapped_cell_no_border = workbook.add_format({
                    'text_wrap': True,
                    'valign': 'vtop',
                    'align': 'center'
                })
                
                # Create header formats (with and without border)
                wrapped_header_with_border = workbook.add_format({
                    'bold': True,
                    'text_wrap': True,
                    'valign': 'vcenter',
                    'align': 'center',
                    'fg_color': "#3553E8",
                    'font_color': 'white',
                    'border': 1
                })
                wrapped_header_no_border = workbook.add_format({
                    'bold': True,
                    'text_wrap': True,
                    'valign': 'vcenter',
                    'align': 'center',
                    'fg_color': "#3553E8",
                    'font_color': 'white'
                })
                
                # Write headers with border only for columns A-N (0..13)
                for col_num, value in enumerate(combined_nessus.columns.values):
                    header_fmt = wrapped_header_with_border if col_num <= 13 else wrapped_header_no_border
                    worksheet_nessus.write(0, col_num, value, header_fmt)
                
                # Write data rows; apply border only for columns A-N (0..13)
                for row_num in range(len(combined_nessus)):
                    for col_num in range(len(combined_nessus.columns)):
                        value = combined_nessus.iat[row_num, col_num]
                        cell_fmt = wrapped_cell_with_border if col_num <= 13 else wrapped_cell_no_border
                        
                        # Handle different data types properly
                        if pd.isna(value):
                            worksheet_nessus.write(row_num + 1, col_num, '', cell_fmt)
                        elif isinstance(value, (int, float)):
                            if isinstance(value, float) and math.isinf(value):
                                worksheet_nessus.write(row_num + 1, col_num, 'INF', cell_fmt)
                            else:
                                worksheet_nessus.write_number(row_num + 1, col_num, value, cell_fmt)
                        elif isinstance(value, str):
                            worksheet_nessus.write_string(row_num + 1, col_num, str(value), cell_fmt)
                        elif value is None:
                            worksheet_nessus.write_string(row_num + 1, col_num, 'None', cell_fmt)
                        else:
                            worksheet_nessus.write(row_num + 1, col_num, str(value), cell_fmt)
                
                # Set specific column widths for Nessus worksheet
                column_widths = [20, 20, 25, 15, 25, 20, 15, 40, 40, 80, 40, 30, 80, 40, 20]
                
                for col_num, width in enumerate(column_widths):
                    if col_num < len(combined_nessus.columns):
                        worksheet_nessus.set_column(col_num, col_num, width)
                
                if len(combined_nessus.columns) > len(column_widths):
                    for col_num in range(len(column_widths), len(combined_nessus.columns)):
                        worksheet_nessus.set_column(col_num, col_num, 25, wrapped_cell_no_border)
                
                # Set default row height only (no global format to avoid borders on all columns)
                worksheet_nessus.set_default_row(17)
                
                # Calculate unmatched vulnerabilities here while we still have access to the data
                try:
                    # Get unique vulnerability names from Name column
                    unique_vulnerabilities = set()
                    for vulnerabilities in combined_nessus['Name'].dropna():
                        unique_vulnerabilities.add(str(vulnerabilities).strip())
                    
                    # Load catalog to get matched vulnerabilities
                    catalog_path = "static/Formats_and_Catalog/Infrastructure VAPT Catalog.xlsx"
                    if os.path.exists(catalog_path):
                        try:
                            catalog_df = pd.read_excel(catalog_path, sheet_name=0)
                        except Exception as e:
                            print(f"Error reading catalog file for vulnerability calculation: {e}")
                            catalog_df = None
                        if catalog_df is not None and 'Vulnerabilities in this group' in catalog_df.columns:
                            catalog_vulnerabilities = set()
                            for vulnerabilities in catalog_df['Vulnerabilities in this group'].dropna():
                                vuln_list = [v.strip() for v in str(vulnerabilities).split(',') if v.strip()]
                                catalog_vulnerabilities.update(vuln_list)
                            
                            matched_vulnerabilities = unique_vulnerabilities.intersection(catalog_vulnerabilities)
                            unmatched_vulnerabilities = unique_vulnerabilities - matched_vulnerabilities
                            unmatched_count = len(unmatched_vulnerabilities)
                except Exception as e:
                    print(f"Error calculating unmatched vulnerabilities: {e}")
                    unmatched_count = 0
                
                # Create Scope worksheet from Nessus data
                create_scope_worksheet(workbook, combined_nessus, header_format, cell_format)
                
                # Create Summary worksheet from Nessus data
                create_summary_worksheet(workbook, combined_nessus, header_format, cell_format)
                
                # Get unmatched vulnerability details from session if available
                unmatched_details_mapping = session.get('unmatched_vulnerability_details', {})
                
                # Create Infra_VAPT worksheet from Nessus data
                create_infra_vapt_worksheet(workbook, combined_nessus, header_format, cell_format, unmatched_details_mapping)
            
        
        # Generate fixed filename for download (no timestamp)
        filename = "Infrastructure_VAPT_Follow_UP_Audit_Worksheet.xlsx"
        print(f"✅ Download filename set to: {filename}")
        
        # Create timestamped temp directories for this request
        current_ts = datetime.now().strftime('%Y%m%d%H%M%S')
        temp_poc_dir = f"temp_poc_images_{current_ts}"
        temp_poc2_folder = f"Temp_POC2_{current_ts}"
        # Use same temp_poc_dir for user POC images (no separate temp_user_dir)
        temp_user_dir = temp_poc_dir
         
        # Store evidence file for POC image insertion after sorting
        evidence_file_for_poc = evidence_file
        
        # Perform Excel comparison if user's Excel file is provided
        if user_excel_file and user_excel_file.filename != '':
            try:
                print("Starting Excel comparison process...")
                
                # Extract vulnerability names from user's Excel file
                user_vulnerabilities = extract_vulnerability_names_from_excel(user_excel_file)
                print(f"Found {len(user_vulnerabilities)} vulnerabilities in user's Excel file")
                
                # Extract full row data from user's Excel file
                user_row_data = extract_full_row_data_from_excel(user_excel_file)
                print(f"Extracted full row data for {len(user_row_data)} vulnerabilities from user's Excel file")
                
                # Save our generated Excel file temporarily for comparison
                temp_comparison_path = "temp_our_excel_for_comparison.xlsx"
                with open(temp_comparison_path, 'wb') as temp_file:
                    temp_file.write(output.getvalue())
                
                # Extract vulnerability names from our generated Excel file
                our_vulnerabilities = extract_vulnerability_names_from_excel(temp_comparison_path)
                print(f"Found {len(our_vulnerabilities)} vulnerabilities in our generated Excel file")
                
                # Compare vulnerabilities
                comparison_result = compare_vulnerabilities(our_vulnerabilities, user_vulnerabilities)
                print(f"Comparison completed. Status distribution:")
                print(f"Our vulnerabilities count: {len(our_vulnerabilities)}")
                print(f"User vulnerabilities count: {len(user_vulnerabilities)}")
                print(f"Comparison result count: {len(comparison_result)}")
                status_counts = {}
                for status in comparison_result.values():
                    status_counts[status] = status_counts.get(status, 0) + 1
                for status, count in status_counts.items():
                    print(f"  {status}: {count}")
                
                # Debug: Print some sample vulnerabilities
                if our_vulnerabilities:
                    print(f"Sample our vulnerabilities: {list(our_vulnerabilities)[:3]}")
                if user_vulnerabilities:
                    print(f"Sample user vulnerabilities: {list(user_vulnerabilities)[:3]}")
                if comparison_result:
                    print(f"Sample comparison results: {dict(list(comparison_result.items())[:3])}")
                
                # Add status column to our Excel file
                if add_status_column_to_excel(temp_comparison_path, comparison_result):
                    print("Status column added successfully")
                else:
                    print("Failed to add status column")
                
                # Handle Closed vulnerabilities - add them to our Excel
                closed_vulnerabilities = [vuln for vuln, status in comparison_result.items() if status == "Closed"]
                print(f"DEBUG: Closed vulnerabilities list: {closed_vulnerabilities}")
                print(f"DEBUG: User row data keys: {list(user_row_data.keys())[:5] if user_row_data else 'No user row data'}")
                
                if closed_vulnerabilities:
                    print(f"Found {len(closed_vulnerabilities)} Closed vulnerabilities to add to our Excel")
                    
                    # Extract POC images from user's Excel for closed vulnerabilities
                    closed_poc_images = extract_poc_data_from_user_excel(user_excel_file, closed_vulnerabilities)
                    print(f"Extracted {len(closed_poc_images)} closed vulnerabilities with images")
                    
                    if add_closed_vulnerability_rows_to_excel(temp_comparison_path, closed_vulnerabilities, user_row_data):
                        print("Closed vulnerability rows added successfully")
                        # Note: Old POC images will be inserted after sorting to ensure correct row numbers
                    else:
                        print("Failed to add Closed vulnerability rows")
                else:
                    print("DEBUG: No Closed vulnerabilities found to add")
                
                # Sort and renumber the Infra_VAPT worksheet by risk level
                print("Sorting Infra_VAPT worksheet by risk level...")
                if sort_and_renumber_infra_vapt_worksheet(temp_comparison_path):
                    print("Infra_VAPT worksheet sorted and renumbered successfully")
                else:
                    print("Warning: Failed to sort Infra_VAPT worksheet")
                
                # Add POC images after sorting (if evidence files are provided)
                if evidence_file_for_poc and evidence_file_for_poc.filename != '':
                    try:
                        print("Adding POC images after sorting...")
                        # Extract POC images
                        poc_mapping = extract_poc_images(evidence_file_for_poc, temp_dir=temp_poc_dir)
                        
                        if poc_mapping:
                            # Insert POC images using openpyxl
                            insert_poc_images_to_excel(temp_comparison_path, poc_mapping, combined_nessus)
                            print("POC images added successfully after sorting")
                        else:
                            print("No POC images found to add")
                            
                    except Exception as e:
                        print(f"Error adding POC images after sorting: {e}")
                        # Continue with normal processing even if image insertion fails
                
                # Extract images from user's uploaded Excel file
                try:
                    print("Extracting images from user's uploaded Excel file...")
                    print(f"File details - Name: {user_excel_file.filename}, Content type: {user_excel_file.content_type}")
                    
                    # Reset file pointer to beginning
                    user_excel_file.seek(0)
                    
                    # Extract images from the uploaded Excel file into a timestamped folder
                    current_ts = datetime.now().strftime('%Y%m%d%H%M%S')
                    temp_poc2_folder = f"Temp_POC2_{current_ts}"
                    result = extract_images_from_infra_vapt_worksheet(user_excel_file, output_folder=temp_poc2_folder)
                    
                    if result["success"]:
                        print(f"✅ Successfully extracted {result['images_extracted']} images from user's Excel file")
                        print(f"📁 Images saved to: {result['output_folder']}")
                        
                        # Log details of extracted images
                        if result.get("images_info"):
                            print("Extracted images details:")
                            for img in result["images_info"]:
                                print(f"  - {img['filename']} (Vulnerability: {img['vulnerability_name']})")
                    else:
                        print(f"❌ Failed to extract images: {result.get('error', 'Unknown error')}")
                        
                except Exception as e:
                    print(f"Error extracting images from user's Excel file: {e}")
                    import traceback
                    traceback.print_exc()
                    # Continue with normal processing even if image extraction fails
                
                # Add Old POC images to the generated Excel file
                try:
                    print("Adding Old POC images to generated Excel file...")
                    
                    # Match images from timestamped Temp_POC2 folder
                    old_poc_image_mapping = match_images_from_temp_poc2_folder(temp_poc2_folder)
                    
                    if old_poc_image_mapping:
                        print(f"Found {len(old_poc_image_mapping)} Old POC images to process")
                        
                        # Insert Old POC images into the Excel file
                        old_poc_rows = insert_old_poc_images_to_excel(temp_comparison_path, old_poc_image_mapping)
                        
                        if old_poc_rows:
                            print(f"✅ Successfully added Old POC images to {len(old_poc_rows)} rows")
                        else:
                            print("⚠️ No Old POC images were matched with vulnerabilities")
                    else:
                        print("ℹ️ No images found in timestamped Temp_POC2 folder for Old POC processing")
                        
                except Exception as e:
                    print(f"Error adding Old POC images: {e}")
                    import traceback
                    traceback.print_exc()
                    # Continue with normal processing even if Old POC insertion fails
                
                # Read the updated file
                with open(temp_comparison_path, 'rb') as updated_file:
                    output = BytesIO(updated_file.read())
                
                # Clean up temporary file
                if os.path.exists(temp_comparison_path):
                    os.remove(temp_comparison_path)
                    
            except Exception as e:
                print(f"Error during Excel comparison: {e}")
                import traceback
                traceback.print_exc()
        
        # Continue with normal processing even if comparison fails
        
        # If no user Excel file but evidence files are provided, add POC images after normal processing
        if not user_excel_file and evidence_file_for_poc and evidence_file_for_poc.filename != '':
            try:
                print("Adding POC images to generated Excel...")
                # Save Excel file temporarily
                temp_poc_path = "temp_poc_insertion.xlsx"
                with open(temp_poc_path, 'wb') as temp_file:
                    temp_file.write(output.getvalue())
                
                # Extract POC images
                poc_mapping = extract_poc_images(evidence_file_for_poc, temp_dir=temp_poc_dir)
                
                if poc_mapping:
                    # Insert POC images using openpyxl
                    insert_poc_images_to_excel(temp_poc_path, poc_mapping, combined_nessus)
                    print("POC images added successfully")
                    
                    # Read the updated file
                    with open(temp_poc_path, 'rb') as updated_file:
                        output = BytesIO(updated_file.read())
                
                # Clean up temporary file
                if os.path.exists(temp_poc_path):
                    os.remove(temp_poc_path)
                    
            except Exception as e:
                print(f"Error processing evidence files: {e}")
                # Continue with normal processing even if image insertion fails
        
        # Prepare response (Old POC column widths already set in create_infra_vapt_worksheet)
        print("=== Preparing final Excel response ===")
        try:
            # No need to modify widths again - they were set correctly during worksheet creation
            temp_final_path = "temp_final_excel.xlsx"
            with open(temp_final_path, 'wb') as temp_file:
                temp_file.write(output.getvalue())
            
            # Load workbook just before final save and download
            from openpyxl import load_workbook
            wb = load_workbook(temp_final_path)
            
            if 'Infra_VAPT' in wb.sheetnames:
                # Step 1: Ensure Infra_VAPT row heights are set (covers all open/closed rows)
                try:
                    ws = wb['Infra_VAPT']
                    max_row = ws.max_row
                    for r in range(2, max_row + 1):
                        val = ws.cell(row=r, column=1).value
                        if val is not None and str(val).strip() != "":
                            ws.row_dimensions[r].height = 37.5
                    print(f"✅ Applied final row heights to Infra_VAPT for {max_row-1} data rows")
                except Exception as rh_err:
                    print(f"⚠️ Could not apply final row heights: {rh_err}")
                
                # Step 2: Update Meta_Data worksheet with follow-up vulnerability counts (just before download)
                # NOTE: Closed vulnerabilities ARE included in Infra_VAPT worksheet but are NOT counted for Meta_Data
                print("="*60)
                print("UPDATING META_DATA WITH FOLLOW-UP VULNERABILITY COUNTS (JUST BEFORE DOWNLOAD)")
                print("NOTE: Counting only New/Open vulnerabilities, excluding Closed ones")
                print("="*60)
                
                # Extract follow-up risk factor counts directly from the loaded workbook
                try:
                    ws_infra = wb['Infra_VAPT']
                    
                    # Find Risk Factor column
                    risk_factor_col = None
                    for col in range(1, ws_infra.max_column + 1):
                        cell_value = ws_infra.cell(row=1, column=col).value
                        if cell_value:
                            cell_str = str(cell_value).lower()
                            if 'risk' in cell_str and 'factor' in cell_str:
                                risk_factor_col = col
                                break
                    
                    if risk_factor_col is None:
                        print("⚠️ Warning: Risk Factor column not found")
                    else:
                        # Find Status column (column L)
                        status_col = None
                        for col in range(1, ws_infra.max_column + 1):
                            cell_value = ws_infra.cell(row=1, column=col).value
                            if cell_value:
                                cell_str = str(cell_value).lower()
                                if 'status' in cell_str:
                                    status_col = col
                                    break
                        
                        if status_col is None:
                            print("⚠️ Warning: Status column not found, counting all rows")
                        
                        # Count risk factors (ONLY for New/Open status, EXCLUDE Closed)
                        # Closed vulnerabilities remain in Infra_VAPT worksheet but are not counted here
                        risk_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
                        rows_processed = 0
                        rows_skipped = 0
                        
                        for row in range(2, ws_infra.max_row + 1):
                            # Check status first - only count rows with Status = "New" or "Open"
                            if status_col:
                                status_value = ws_infra.cell(row=row, column=status_col).value
                                if status_value:
                                    status_str = str(status_value).strip().upper()
                                    # Skip Closed vulnerabilities - they are in worksheet but not counted
                                    if status_str not in ['NEW', 'OPEN']:
                                        rows_skipped += 1
                                        continue
                            
                            # Only count rows that passed status check (New or Open)
                            rows_processed += 1
                            risk_value = ws_infra.cell(row=row, column=risk_factor_col).value
                            if risk_value:
                                risk_str = str(risk_value).strip()
                                risk_upper = risk_str.upper()
                                
                                # Match risk factor (case-insensitive)
                                if risk_upper == "CRITICAL" or risk_str == "Critical":
                                    risk_counts["Critical"] += 1
                                elif risk_upper == "HIGH" or risk_str == "High":
                                    risk_counts["High"] += 1
                                elif risk_upper == "MEDIUM" or risk_str == "Medium":
                                    risk_counts["Medium"] += 1
                                elif risk_upper == "LOW" or risk_str == "Low":
                                    risk_counts["Low"] += 1
                                else:
                                    # Try to match partial strings
                                    if "critical" in risk_str.lower():
                                        risk_counts["Critical"] += 1
                                    elif "high" in risk_str.lower() and "medium" not in risk_str.lower() and "low" not in risk_str.lower():
                                        risk_counts["High"] += 1
                                    elif "medium" in risk_str.lower():
                                        risk_counts["Medium"] += 1
                                    elif "low" in risk_str.lower():
                                        risk_counts["Low"] += 1
                        
                        print(f"📊 Follow-up risk factor counts (New/Open only, excluding Closed): {risk_counts}")
                        print(f"📊 Total rows counted: {rows_processed}, Rows skipped (Closed/Other): {rows_skipped}")
                        
                        # Update Meta_Data worksheet
                        if "Meta_Data" in wb.sheetnames:
                            ws_meta = wb["Meta_Data"]
                            
                            # Find the row where "FOLLOW UP AUDIT VULNERABILITIES" section is
                            follow_up_section_row = None
                            for row in range(1, ws_meta.max_row + 1):
                                cell_value = ws_meta.cell(row=row, column=1).value
                                if cell_value and "FOLLOW UP AUDIT" in str(cell_value).upper():
                                    follow_up_section_row = row
                                    break
                            
                            if follow_up_section_row is None:
                                print("⚠️ Warning: FOLLOW UP AUDIT VULNERABILITIES section not found in Meta_Data worksheet")
                            else:
                                # Update the values in the Follow up Audit Vulnerabilities section
                                vulnerability_levels = ['Critical', 'High', 'Medium', 'Low']
                                
                                for i, level in enumerate(vulnerability_levels):
                                    data_row = follow_up_section_row + 1 + i
                                    if data_row <= ws_meta.max_row:
                                        # Update the value in column B
                                        ws_meta.cell(row=data_row, column=2, value=risk_counts[level])
                                        print(f"✅ Updated Meta_Data row {data_row} ({level}): {risk_counts[level]}")
                                print(f"✅ Successfully updated Meta_Data worksheet with follow-up vulnerability counts")
                        else:
                            print("⚠️ Warning: Meta_Data worksheet not found")
                            
                except Exception as count_err:
                    print(f"❌ Error updating Meta_Data counts: {count_err}")
                    import traceback
                    traceback.print_exc()
                
                # Step 3: Apply formatting to Infra_VAPT worksheet columns A, B, C, D, E, F, J
                # Format: Middle align (vertical center), Center (horizontal center), and Wrap Text
                try:
                    from openpyxl.styles import Alignment
                    ws_infra = wb['Infra_VAPT']
                    
                    # Define columns to format: A=1, B=2, C=3, D=4, E=5, F=6, J=10
                    columns_to_format = [1, 2, 3, 4, 5, 6, 10]
                    
                    # Create alignment format: horizontal center, vertical center, wrap text enabled
                    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # Apply formatting to all data rows (starting from row 2, row 1 is headers)
                    max_row = ws_infra.max_row
                    rows_formatted = 0
                    
                    for row in range(2, max_row + 1):
                        for col in columns_to_format:
                            cell = ws_infra.cell(row=row, column=col)
                            # Apply alignment with wrap text (preserve shrink_to_fit if it exists)
                            if cell.alignment:
                                # Set alignment with wrap text enabled, preserve shrink_to_fit
                                cell.alignment = Alignment(
                                    horizontal='center',
                                    vertical='center',
                                    wrap_text=True,
                                    shrink_to_fit=cell.alignment.shrink_to_fit if cell.alignment.shrink_to_fit else False
                                )
                            else:
                                cell.alignment = center_alignment
                        rows_formatted += 1
                    
                    print(f"✅ Applied center alignment and wrap text to columns A, B, C, D, E, F, J for {rows_formatted} data rows in Infra_VAPT worksheet")
                except Exception as format_err:
                    print(f"⚠️ Could not apply column formatting: {format_err}")
                    import traceback
                    traceback.print_exc()
                
                # Step 4: Add "N/A" to empty cells in column E (CVE ID) for rows where column A has content
                try:
                    from openpyxl.styles import Alignment
                    ws_infra = wb['Infra_VAPT']
                    max_row = ws_infra.max_row
                    column_a = 1  # Column A (Sr.No)
                    column_e = 5  # Column E (CVE ID)
                    na_count = 0
                    
                    # Iterate through all rows starting from row 2 (row 1 is headers)
                    for row in range(2, max_row + 1):
                        # Check if column A has content
                        cell_a = ws_infra.cell(row=row, column=column_a)
                        cell_a_value = cell_a.value
                        
                        # Check if column A has content (not None and not empty string)
                        if cell_a_value is not None and str(cell_a_value).strip() != "":
                            # Column A has content, check column E
                            cell_e = ws_infra.cell(row=row, column=column_e)
                            cell_e_value = cell_e.value
                            
                            # Check if column E is empty (None, empty string, or only whitespace)
                            is_empty = False
                            if cell_e_value is None:
                                is_empty = True
                            elif isinstance(cell_e_value, str):
                                if cell_e_value.strip() == "":
                                    is_empty = True
                            # Also check for numeric zero or False as empty
                            elif cell_e_value == 0 or cell_e_value is False:
                                is_empty = True
                            
                            if is_empty:
                                # Column E is empty, add "N/A"
                                cell_e.value = "N/A"
                                
                                # Apply the same formatting (center align, wrap text) to match other cells
                                cell_e.alignment = Alignment(
                                    horizontal='center',
                                    vertical='center',
                                    wrap_text=True
                                )
                                
                                na_count += 1
                    
                    if na_count > 0:
                        print(f"✅ Added 'N/A' to {na_count} empty cells in column E (CVE ID) for rows where column A has content")
                    else:
                        print(f"✅ Checked column E (CVE ID) - all cells have content for rows where column A has content")
                except Exception as na_err:
                    print(f"⚠️ Could not add N/A to empty CVE ID cells: {na_err}")
                    import traceback
                    traceback.print_exc()
                
                # Step 5: Save all changes at once
                wb.save(temp_final_path)
                print("✅ Saved workbook with all updates (row heights + Meta_Data counts + column formatting + CVE ID N/A)")
                
                # Read the modified file back into output
                with open(temp_final_path, 'rb') as temp_file:
                    output.seek(0)
                    output.truncate()
                    output.write(temp_file.read())
                
                print("✅ Excel file finalized")
            else:
                print("❌ Infra_VAPT worksheet not found")
                
        except Exception as e:
            print(f"❌ Error finalizing Excel: {e}")
            import traceback
            traceback.print_exc()
        
        print("=== Final Excel preparation completed ===")
        
        # Ensure filename is set correctly (no timestamp) - set it right before response
        filename = "Infrastructure_VAPT_Follow_UP_Audit_Worksheet.xlsx"
        print(f"✅ Final download filename: {filename}")
        
        # Prepare response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        print(f"✅ Response headers set with filename: {filename}")
        
        # Clean up temporary files after Excel generation
        try:
            # Clean up temp_final_excel.xlsx file
            if os.path.exists("temp_final_excel.xlsx"):
                os.remove("temp_final_excel.xlsx")
                print("✅ Cleaned up temp_final_excel.xlsx file")
            
            # Clean up temp_poc_images_{timestamp} folder
            if 'temp_poc_dir' in locals() and os.path.exists(temp_poc_dir):
                import shutil
                shutil.rmtree(temp_poc_dir)
                print(f"✅ Successfully deleted {temp_poc_dir} folder")
            
            # Clean up any leftover temp_poc_images_* folders (glob pattern)
            import glob
            for folder in glob.glob("temp_poc_images_*"):
                if os.path.isdir(folder):
                    try:
                        shutil.rmtree(folder)
                        print(f"✅ Successfully deleted leftover folder: {folder}")
                    except Exception as e:
                        print(f"⚠️ Error deleting leftover folder {folder}: {e}")
        except Exception as e:
            print(f"Error deleting temp_poc_images folder: {e}")
        
        # Clean up Temp_POC2_{timestamp} folder after Excel generation
        # Note: temp_user_dir is now same as temp_poc_dir, so no separate cleanup needed
        try:
            if 'temp_poc2_folder' in locals() and os.path.exists(temp_poc2_folder):
                import shutil
                shutil.rmtree(temp_poc2_folder)
                print(f"✅ Successfully deleted {temp_poc2_folder} folder")
            
            # Clean up any leftover Temp_POC2_* folders (glob pattern)
            import glob
            for folder in glob.glob("Temp_POC2_*"):
                if os.path.isdir(folder):
                    try:
                        shutil.rmtree(folder)
                        print(f"✅ Successfully deleted leftover folder: {folder}")
                    except Exception as e:
                        print(f"⚠️ Error deleting leftover folder {folder}: {e}")
        except Exception as e:
            print(f"Error deleting Temp_POC2 folder: {e}")
        
        return response
        
    except Exception as e:
        print(f"Error in process_follow_up_audit_report: {e}")
        import traceback
        traceback.print_exc()
        
        # Clean up temp_poc_images_{timestamp} folder on error as well
        try:
            if 'temp_poc_dir' in locals() and os.path.exists(temp_poc_dir):
                import shutil
                shutil.rmtree(temp_poc_dir)
                print(f"✅ Successfully deleted {temp_poc_dir} folder (error cleanup)")
            
            # Clean up any leftover temp_poc_images_* folders (glob pattern)
            import glob
            for folder in glob.glob("temp_poc_images_*"):
                if os.path.isdir(folder):
                    try:
                        shutil.rmtree(folder)
                        print(f"✅ Successfully deleted leftover folder: {folder}")
                    except Exception as e:
                        print(f"⚠️ Error deleting leftover folder {folder}: {e}")
        except Exception as e:
            print(f"Error deleting temp_poc_images folder: {e}")
        
        # Note: temp_user_dir is now same as temp_poc_dir, so cleaned up with temp_poc_dir
        # Clean up Temp_POC2_{timestamp} folder on error as well
        try:
            if 'temp_poc2_folder' in locals() and os.path.exists(temp_poc2_folder):
                import shutil
                shutil.rmtree(temp_poc2_folder)
                print(f"✅ Successfully deleted {temp_poc2_folder} folder (error cleanup)")
            
            # Clean up any leftover Temp_POC2_* folders (glob pattern)
            import glob
            for folder in glob.glob("Temp_POC2_*"):
                if os.path.isdir(folder):
                    try:
                        shutil.rmtree(folder)
                        print(f"✅ Successfully deleted leftover folder: {folder}")
                    except Exception as e:
                        print(f"⚠️ Error deleting leftover folder {folder}: {e}")
        except Exception as e:
            print(f"Error deleting Temp_POC2 folder (error cleanup): {e}")
        
        return "Error processing files", 500

@follow_up_audit_bp.route('/extract_images_from_excel', methods=['POST'])
def extract_images_from_excel_route():
    """
    Standalone route to extract images from uploaded Excel file.
    """
    try:
        # Check if file was uploaded
        if 'excelFile' not in request.files:
            return jsonify({'error': 'No Excel file uploaded'}), 400
        
        excel_file = request.files['excelFile']
        
        if excel_file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file extension
        if not excel_file.filename.lower().endswith('.xlsx'):
            return jsonify({'error': 'Please upload a valid Excel file (.xlsx)'}), 400
        
        print(f"Processing image extraction for file: {excel_file.filename}")
        
        # Extract images from the uploaded Excel file
        result = extract_images_from_infra_vapt_worksheet(excel_file)
        
        if result["success"]:
            return jsonify({
                'success': True,
                'message': f'Successfully extracted {result["images_extracted"]} images',
                'images_extracted': result["images_extracted"],
                'output_folder': result["output_folder"],
                'images_info': result.get("images_info", [])
            })
        else:
            return jsonify({
                'success': False,
                'error': result.get('error', 'Unknown error occurred')
            }), 500
            
    except Exception as e:
        print(f"Error in extract_images_from_excel_route: {e}")
        return jsonify({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }), 500
            
    except Exception as e:
        print(f"Error in process_follow_up_audit_report: {e}")
        import traceback
        traceback.print_exc()
        
        # Clean up temp_poc_images folder on error as well
        try:
            if 'temp_poc_dir' in locals() and os.path.exists(temp_poc_dir):
                import shutil
                shutil.rmtree(temp_poc_dir)
        except Exception as e:
            print(f"Error deleting temp_poc_images folder: {e}")
        
        # Note: temp_user_dir is now same as temp_poc_dir, so cleaned up with temp_poc_dir
        # Clean up Temp_POC2 folder on error as well
        try:
            if 'temp_poc2_folder' in locals() and os.path.exists(temp_poc2_folder):
                import shutil
                shutil.rmtree(temp_poc2_folder)
                print("Successfully deleted Temp_POC2 folder (error cleanup)")
        except Exception as e:
            print(f"Error deleting Temp_POC2 folder (error cleanup): {e}")
        
        return f"Error generating follow-up report: {str(e)}", 500

@follow_up_audit_bp.route('/get_images_info_from_excel', methods=['POST'])
def get_images_info_from_excel_route():
    """
    Route to get information about images in uploaded Excel file without extracting them.
    """
    try:
        # Check if file was uploaded
        if 'excelFile' not in request.files:
            return jsonify({'error': 'No Excel file uploaded'}), 400
        
        excel_file = request.files['excelFile']
        
        if excel_file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file extension
        if not excel_file.filename.lower().endswith('.xlsx'):
            return jsonify({'error': 'Please upload a valid Excel file (.xlsx)'}), 400
        
        print(f"Getting images info for file: {excel_file.filename}")
        
        # Get images information from the uploaded Excel file
        result = get_all_images_info_from_infra_vapt(excel_file)
        
        if result["success"]:
            return jsonify({
                'success': True,
                'images_found': result["images_found"],
                'images_info': result.get("images_info", []),
                'vulnerability_mapping': result.get("vulnerability_mapping", {})
            })
        else:
            return jsonify({
                'success': False,
                'error': result.get('error', 'Unknown error occurred')
            }), 500
            
    except Exception as e:
        print(f"Error in get_images_info_from_excel_route: {e}")
        return jsonify({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }), 500
        
    except Exception as e:
        print(f"Error in process_follow_up_audit_report: {e}")
        import traceback
        traceback.print_exc()
        
        # Clean up temp_poc_images_{timestamp} folder on error as well
        try:
            if 'temp_poc_dir' in locals() and os.path.exists(temp_poc_dir):
                import shutil
                shutil.rmtree(temp_poc_dir)
                print(f"✅ Successfully deleted {temp_poc_dir} folder after error")
            
            # Clean up any leftover temp_poc_images_* folders (glob pattern)
            import glob
            for folder in glob.glob("temp_poc_images_*"):
                if os.path.isdir(folder):
                    try:
                        shutil.rmtree(folder)
                        print(f"✅ Successfully deleted leftover folder: {folder}")
                    except Exception as cleanup_error:
                        print(f"⚠️ Error deleting leftover folder {folder}: {cleanup_error}")
        except Exception as cleanup_error:
            print(f"Error deleting temp_poc_images folder after error: {cleanup_error}")
        
        # Clean up Temp_POC2_{timestamp} folder on error as well
        try:
            if 'temp_poc2_folder' in locals() and os.path.exists(temp_poc2_folder):
                import shutil
                shutil.rmtree(temp_poc2_folder)
                print(f"✅ Successfully deleted {temp_poc2_folder} folder after error")
            
            # Clean up any leftover Temp_POC2_* folders (glob pattern)
            import glob
            for folder in glob.glob("Temp_POC2_*"):
                if os.path.isdir(folder):
                    try:
                        shutil.rmtree(folder)
                        print(f"✅ Successfully deleted leftover folder: {folder}")
                    except Exception as cleanup_error:
                        print(f"⚠️ Error deleting leftover folder {folder}: {cleanup_error}")
        except Exception as cleanup_error:
            print(f"Error deleting Temp_POC2 folder after error: {cleanup_error}")
        
        # Note: temp_user_dir is now same as temp_poc_dir, so cleaned up with temp_poc_dir
        return f"Error generating follow-up report: {str(e)}", 500

def convert_risk_to_camelcase(risk_value):
    """Convert risk value to CamelCase format"""
    if not risk_value:
        return ""
    
    risk_str = str(risk_value).strip().lower()
    if risk_str == 'critical':
        return 'Critical'
    elif risk_str == 'high':
        return 'High'
    elif risk_str == 'medium':
        return 'Medium'
    elif risk_str == 'low':
        return 'Low'
    else:
        # Return original value with first letter capitalized
        return risk_str.capitalize()

def create_scope_worksheet(workbook, combined_nessus, header_format, cell_format):
    """Create Scope worksheet with Branch Name and Host data from Nessus"""
    if 'Branch Name' in combined_nessus.columns and 'Host' in combined_nessus.columns:
        # Extract unique Branch Name and Host combinations
        scope_data = combined_nessus[['Branch Name', 'Host']].drop_duplicates()
        
        # Group by Branch Name and collect all unique Hosts
        branch_ip_mapping = {}
        for branch_name, group in scope_data.groupby('Branch Name'):
            ips = sorted([str(ip).strip() for ip in group['Host'].unique() if pd.notna(ip) and str(ip).strip()])
            if ips:
                branch_ip_mapping[branch_name] = ips
        
        # Create Scope worksheet
        worksheet_scope = workbook.add_worksheet("Scope")
        
        # Write headers
        worksheet_scope.write(0, 0, "Sr.No", header_format)
        worksheet_scope.write(0, 1, "BRANCH NAME", header_format)
        worksheet_scope.write(0, 2, "HOST", header_format)
        
        row = 1
        merge_ranges = {}
        serial_number = 1
        
        # Write branch names and IPs with serial numbers
        for branch_name, ips in branch_ip_mapping.items():
            if not ips:
                continue
            
            start_row = row
            worksheet_scope.write(row, 0, serial_number, cell_format)
            worksheet_scope.write(row, 1, branch_name, cell_format)
            
            for ip in ips:
                worksheet_scope.write(row, 2, ip, cell_format)
                row += 1
            
            end_row = row - 1
            
            # Only merge if there are multiple IPs for this branch
            if end_row > start_row:
                merge_ranges[branch_name] = {
                    'start_row': start_row,
                    'end_row': end_row,
                    'serial_number': serial_number
                }
            
            serial_number += 1
        
        # Apply merging for branch names and serial numbers
        for branch_name, merge_data in merge_ranges.items():
            start_row = merge_data['start_row']
            end_row = merge_data['end_row']
            serial_number = merge_data['serial_number']
            
            if start_row != end_row:
                worksheet_scope.merge_range(
                    start_row, 0, end_row, 0,
                    serial_number,
                    cell_format
                )
                worksheet_scope.merge_range(
                    start_row, 1, end_row, 1,
                    branch_name,
                    cell_format
                )
        
        # Set column widths
        worksheet_scope.set_column('A:A', 8)
        worksheet_scope.set_column('B:B', 30)
        worksheet_scope.set_column('C:C', 20)

def create_infra_vapt_worksheet(workbook, combined_nessus, header_format, cell_format, unmatched_details_mapping=None, evidence_files=None):
    """Create Infra_VAPT worksheet with data from Summary's Name column and match with catalog (using merge state for Follow-up Audit)"""
    # Check if required columns exist
    required_columns = ['Name', 'Host', 'Branch Name', 'Risk']
    missing_cols = [col for col in required_columns if col not in combined_nessus.columns]
    if missing_cols:
        print(f"Cannot create Infra_VAPT worksheet - missing columns: {missing_cols}")
        return
    
    try:
        # Get merge state from session (Follow-up specific key)
        merge_state = session.get('followup_vulnerability_merge_state', None)
        use_merge_state = merge_state is not None and 'matched_groups' in merge_state
        
        if use_merge_state:
            print("📊 [Follow-up] Using merge state from session for vulnerability grouping")
            merged_groups_from_session = merge_state.get('matched_groups', [])
            new_group_details_dict = merge_state.get('new_group_details', {})
        else:
            print("📊 [Follow-up] No merge state found - using standard catalog matching")
            merged_groups_from_session = None
            new_group_details_dict = {}
        # Create color formats for risk factors
        critical_format = workbook.add_format({
            'bg_color': '#8B0000',  # Dark Red
            'font_color': 'white',
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        high_format = workbook.add_format({
            'bg_color': '#FF0000',  # Red
            'font_color': 'white',
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        medium_format = workbook.add_format({
            'bg_color': '#FFA500',  # Orange
            'font_color': 'white',
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        low_format = workbook.add_format({
            'bg_color': '#008000',  # Green
            'font_color': 'white',
            'bold': True,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # POC column format with border (for cells with objects)
        poc_format = workbook.add_format({
            'border': 1,
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'center'
        })
        
        # POC column format without border (for empty cells)
        poc_format_no_border = workbook.add_format({
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'center'
        })
        
        # Old POC column format with specific width (for empty cells)
        old_poc_format = workbook.add_format({
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'center'
        })
        
        # Left-aligned format for specific columns
        left_align_format = workbook.add_format({
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'left',
            'border': 1
        })
        
        # Load the catalog file
        catalog_path = os.path.join('static', 'Formats_and_Catalog', 'Infrastructure VAPT Catalog.xlsx')
        
        if not os.path.exists(catalog_path):
            print(f"Catalog file not found at: {catalog_path}")
            return
        
        # Read the catalog file with error handling
        try:
            catalog_df = pd.read_excel(catalog_path)
            print(f"Catalog loaded successfully with {len(catalog_df)} rows")
        except Exception as e:
            print(f"Error reading catalog file for Infra_VAPT worksheet: {e}")
            print(f"Catalog file may be corrupted. Please check: {catalog_path}")
            return
        
        # Check if the required column exists in catalog
        if 'Vulnerabilities in this group' not in catalog_df.columns:
            print("Catalog file does not contain 'Vulnerabilities in this group' column")
            return
        
        print("Infra_VAPT worksheet creation started - catalog loaded successfully")
        
        # CREATE WORKSHEET FIRST - regardless of whether there are vulnerabilities
        # This ensures the worksheet exists even if there are no vulnerabilities
        worksheet_infra = workbook.add_worksheet("Infra_VAPT")
        print("✅ Infra_VAPT worksheet created successfully (will populate with data if available)")
        
        # Define columns to exclude from catalog
        exclude_columns = ['Sr No', 'Vulnerabilities in this group', 'Affected System']
        
        # Get catalog headers excluding the specified columns
        catalog_headers = [col for col in catalog_df.columns.tolist() if col not in exclude_columns]
        
        # Find the positions of key columns
        rec_countermeasure_col = None
        reference_link_col = None
        audit_observation_col = None
        risk_factor_col = None
        
        for i, col_name in enumerate(catalog_headers):
            if 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                rec_countermeasure_col = i
            if 'reference' in col_name.lower() and 'link' in col_name.lower():
                reference_link_col = i
            if 'audit' in col_name.lower() and 'observation' in col_name.lower():
                audit_observation_col = i
            if 'risk' in col_name.lower() and 'factor' in col_name.lower():
                risk_factor_col = i
        
        # Insert "Affected Systems" column after Recommendation/Countermeasure and before Reference Link
        if rec_countermeasure_col is not None and reference_link_col is not None:
            # Insert after Recommendation/Countermeasure
            insert_position = rec_countermeasure_col + 1
        elif rec_countermeasure_col is not None:
            # If only Recommendation/Countermeasure exists, insert after it
            insert_position = rec_countermeasure_col + 1
        elif reference_link_col is not None:
            # If only Reference Link exists, insert before it
            insert_position = reference_link_col
        else:
            # If neither exists, insert at the end
            insert_position = len(catalog_headers)
        
        # Create worksheet headers - insert "Affected Systems", "Status", "Old POC", and "POC" columns
        headers = ["Sr.No", "Vulnerabilities"]
        # Add catalog headers up to the insertion point
        headers.extend(catalog_headers[:insert_position])
        # Add "Affected Systems" column
        headers.append("Affected Systems")
        # Add remaining catalog headers
        headers.extend(catalog_headers[insert_position:])
        
        # Add Status column (position L)
        headers.append("Status")
        
        # Add Old POC columns (M to S) - 7 columns total
        headers.extend(["Old_POC_N", "Old_POC_O", "Old_POC_P", "Old_POC_Q", "Old_POC_R", "Old_POC_S", "Old_POC_M"])
        
        # Add POC columns (T to Z) - 7 columns total
        headers.extend(["POC_U", "POC_V", "POC_W", "POC_X", "POC_Y", "POC_Z", "POC_T"])
        
        # Write all headers except the ones we'll merge
        for col_num, header in enumerate(headers):
            worksheet_infra.write(0, col_num, header, header_format)
        
        # Find Status column position (should be L)
        status_col_position = None
        for i, header in enumerate(headers):
            if header == "Status":
                status_col_position = i
                break
        
        # Find Old POC column positions (M to S)
        old_poc_col_start = None
        old_poc_col_end = None
        for i, header in enumerate(headers):
            if header == "Old_POC_N":
                old_poc_col_start = i
            if header == "Old_POC_M":
                old_poc_col_end = i
                break
        
        # Find POC column positions (T to Z)
        poc_col_start = None
        poc_col_end = None
        for i, header in enumerate(headers):
            if header == "POC_U":
                poc_col_start = i
            if header == "POC_T":
                poc_col_end = i
                break
        
        # Merge M1:S1 for "Old POC" header
        if old_poc_col_start is not None and old_poc_col_end is not None:
            worksheet_infra.merge_range(0, old_poc_col_start, 0, old_poc_col_end, "Old POC", header_format)
        
        # Merge T1:Z1 for "POC" header
        if poc_col_start is not None and poc_col_end is not None:
            worksheet_infra.merge_range(0, poc_col_start, 0, poc_col_end, "POC", header_format)
        
        print(f"✅ Worksheet headers created with {len(headers)} columns")
        
        # Filter only low, medium, high, critical vulnerabilities
        valid_risks = ['low', 'medium', 'high', 'critical']
        df_filtered = combined_nessus.copy()
        df_filtered['Risk'] = df_filtered['Risk'].astype(str).str.lower().str.strip()
        df_filtered = df_filtered[df_filtered['Risk'].isin(valid_risks)]
        
        # Get unique vulnerability names from Name column
        unique_vulnerabilities = df_filtered['Name'].drop_duplicates().tolist()
        
        # If no vulnerabilities found, worksheet is already created with headers, just return
        if not unique_vulnerabilities:
            print("⚠️ No vulnerabilities found for Infra_VAPT worksheet - worksheet created with headers only")
            # Set column widths even if no data
            worksheet_infra.set_column(0, 0, 8)  # Sr.No
            worksheet_infra.set_column(1, 1, 40)  # Vulnerabilities
            # Set widths for other columns based on catalog headers
            for i in range(2, len(headers)):
                if i < len(headers) - 14:  # Before POC columns
                    worksheet_infra.set_column(i, i, 25)
                else:  # POC columns
                    worksheet_infra.set_column(i, i, 15)
            return
        
        print(f"Found {len(unique_vulnerabilities)} unique vulnerabilities to process")
        
        # Track matched catalog entries to avoid duplicates
        matched_catalog_indices = set()
        matched_vulnerabilities = set()
        vulnerability_groups = {}  # Store which vulnerabilities belong to which catalog group
        vulnerability_affected_systems = {}  # Store affected systems for each vulnerability
        vulnerability_risks = {}  # Store the highest risk for each vulnerability group
        
        # First pass: Collect affected systems for ALL vulnerabilities
        for vulnerability in unique_vulnerabilities:
            # Get all affected systems for this vulnerability (group by Branch Name)
            vuln_data = df_filtered[df_filtered['Name'] == vulnerability]
            branch_hosts = {}
            max_risk = 'low'  # Default to lowest risk
            
            for _, row in vuln_data.iterrows():
                branch = str(row['Branch Name']).strip()
                host = str(row['Host']).strip()
                risk = str(row['Risk']).lower().strip()
                
                # Track the highest risk for this vulnerability
                risk_levels = {'critical': 4, 'high': 3, 'medium': 2, 'low': 1}
                if risk_levels.get(risk, 0) > risk_levels.get(max_risk, 0):
                    max_risk = risk
                
                if branch and host:
                    if branch not in branch_hosts:
                        branch_hosts[branch] = set()
                    branch_hosts[branch].add(host)
            
            # Format the affected systems data
            formatted_systems = []
            for branch, hosts in sorted(branch_hosts.items()):
                formatted_systems.append(branch)
                for host in sorted(hosts):
                    formatted_systems.append(host)
                formatted_systems.append("")  # Add empty line between branches
            
            # Remove the last empty line if it exists
            if formatted_systems and formatted_systems[-1] == "":
                formatted_systems = formatted_systems[:-1]
            
            vulnerability_affected_systems[vulnerability] = formatted_systems
            vulnerability_risks[vulnerability] = max_risk
        
        # Use merge state if available, otherwise use standard catalog matching
        if use_merge_state and merged_groups_from_session:
            print(f"✓ [Follow-up] Using {len(merged_groups_from_session)} merged groups from session")
            
            # Build vulnerability_groups dict from merge state
            for group in merged_groups_from_session:
                catalog_idx = group['catalog_id']
                matched_catalog_indices.add(catalog_idx)
                
                # Add all vulnerabilities in this group
                vulnerability_groups[catalog_idx] = group['matched_vulnerabilities']
                for vuln in group['matched_vulnerabilities']:
                    matched_vulnerabilities.add(vuln)
                
                print(f"[Follow-up] Loaded merged group: {group['group_name']} with {len(group['matched_vulnerabilities'])} vulnerabilities")
        else:
            # Standard catalog matching (original logic)
            print("✓ [Follow-up] Using standard catalog matching")
            for vulnerability in unique_vulnerabilities:
                # Escape special characters in vulnerability name for regex
                escaped_vulnerability = re.escape(str(vulnerability))
                
                # Create pattern that matches vulnerability with newlines before and after
                pattern = rf'(?:\n|\r\n|\A){escaped_vulnerability}(?:\n|\r\n|\Z)'
                
                # Find matching rows in catalog
                matching_rows = catalog_df[
                    catalog_df['Vulnerabilities in this group'].str.contains(
                        pattern, 
                        case=False, 
                        na=False,
                        regex=True
                    )
                ]
                
                if not matching_rows.empty:
                    # Store the catalog index and the vulnerability
                    catalog_idx = matching_rows.index[0]
                    matched_catalog_indices.add(catalog_idx)
                    matched_vulnerabilities.add(vulnerability)
                    
                    # Group vulnerabilities by catalog entry
                    if catalog_idx not in vulnerability_groups:
                        vulnerability_groups[catalog_idx] = []
                    vulnerability_groups[catalog_idx].append(vulnerability)
                    
                    print(f"Matched vulnerability '{vulnerability}' with catalog index {catalog_idx}")
                else:
                    print(f"No match found for vulnerability '{vulnerability}'")
        
        print(f"Vulnerability matching completed - Matched: {len(matched_vulnerabilities)}, Total unique: {len(unique_vulnerabilities)}")
        
        # Store the risk factor from catalog for each catalog entry
        catalog_risk_values = {}
        if risk_factor_col is not None:
            for catalog_idx in matched_catalog_indices:
                # Skip negative indices (new groups from merge state)
                if catalog_idx < 0:
                    # Get risk from merge state for new groups
                    if use_merge_state and merged_groups_from_session:
                        group_data = next((g for g in merged_groups_from_session if g['catalog_id'] == catalog_idx), None)
                        if group_data:
                            catalog_risk_values[catalog_idx] = str(group_data.get('risk_factor', '')).upper().strip()
                else:
                    # Get risk from catalog for existing groups
                    catalog_risk_value = catalog_df.loc[catalog_idx, catalog_headers[risk_factor_col]]
                    if pd.isna(catalog_risk_value):
                        catalog_risk_value = ""
                    catalog_risk_values[catalog_idx] = str(catalog_risk_value).upper().strip()
        
        # Sort catalog indices by risk factor (critical, high, medium, low, then others)
        risk_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
        
        def get_risk_priority(catalog_idx):
            risk_value = catalog_risk_values.get(catalog_idx, "")
            return risk_order.get(risk_value, 4)  # Default to 4 for unknown risks
        
        sorted_catalog_indices = sorted(matched_catalog_indices, key=get_risk_priority)
        
        # Second pass: Write data to worksheet, grouping vulnerabilities by catalog entry
        row_num = 1
        
        # Only process matched vulnerabilities if there are any
        if sorted_catalog_indices:
            for catalog_idx in sorted_catalog_indices:
                # Get catalog data - either from actual catalog or from merge state (for new groups)
                if use_merge_state and catalog_idx < 0:
                    # This is a new group from merge state
                    group_data = next((g for g in merged_groups_from_session if g['catalog_id'] == catalog_idx), None)
                    if not group_data:
                        continue
                    
                    # Get full details from the separate dictionary
                    full_details = new_group_details_dict.get(str(catalog_idx), {})
                    
                    # Create a pseudo catalog row matching actual catalog column names
                    catalog_row = pd.Series({
                        'Name of Vulnerability': group_data.get('group_name', ''),
                        'Risk Factor': full_details.get('riskFactor', group_data.get('risk_factor', '')),
                        'CVE ID': full_details.get('cveId', 'N/A'),
                        'CVSS': full_details.get('cvssScore', group_data.get('cvss_score', '')),
                        'Audit Observation': full_details.get('auditObservation', ''),
                        'Impact': full_details.get('impact', ''),
                        'Recommendation/Countermeasure': full_details.get('recommendation', ''),
                        'Reference Link': full_details.get('referenceLink', ''),
                        'Affected System': '',
                        'Vulnerabilities in this group': ''
                    })
                else:
                    # Standard catalog entry
                    catalog_row = catalog_df.loc[catalog_idx]
                
                vulnerabilities_list = vulnerability_groups.get(catalog_idx, [])
                
                # Determine the highest risk for this group of vulnerabilities
                group_max_risk = 'low'
                risk_levels = {'critical': 4, 'high': 3, 'medium': 2, 'low': 1}
                for vuln in vulnerabilities_list:
                    if risk_levels.get(vulnerability_risks.get(vuln, 'low'), 0) > risk_levels.get(group_max_risk, 0):
                        group_max_risk = vulnerability_risks.get(vuln, 'low')
                
                # Collect all affected systems for all vulnerabilities in this group
                all_affected_systems = []
                branch_hosts_combined = {}
                
                for vuln in vulnerabilities_list:
                    if vuln in vulnerability_affected_systems:
                        # Parse the formatted systems to combine branches and hosts
                        current_branch = None
                        for line in vulnerability_affected_systems[vuln]:
                            if line and not line.startswith(('192.168.', '10.', '172.')) and not re.match(r'\d+\.\d+\.\d+\.\d+', line):
                                # This is a branch name
                                current_branch = line
                                if current_branch not in branch_hosts_combined:
                                    branch_hosts_combined[current_branch] = set()
                            elif line and current_branch:
                                # This is a host IP for the current branch
                                branch_hosts_combined[current_branch].add(line)
                
                # Format the combined affected systems data
                for branch, hosts in sorted(branch_hosts_combined.items()):
                    all_affected_systems.append(branch)
                    for host in sorted(hosts):
                        all_affected_systems.append(host)
                    all_affected_systems.append("")  # Add empty line between branches
                
                # Remove the last empty line if it exists
                if all_affected_systems and all_affected_systems[-1] == "":
                    all_affected_systems = all_affected_systems[:-1]
                
                # Write serial number
                worksheet_infra.write(row_num, 0, row_num, cell_format)
                
                # Write vulnerabilities (each on a new line)
                vulnerabilities_str = "\n".join(sorted(vulnerabilities_list))
                worksheet_infra.write(row_num, 1, vulnerabilities_str, cell_format)
                
                # Write catalog data up to the insertion point
                col_idx = 2
                for i in range(insert_position):
                    col_name = catalog_headers[i]
                    value = catalog_row[col_name]
                    if pd.isna(value):
                        value = ""
                    
                    # Special handling for Risk Factor column - USE CATALOG VALUE INSTEAD OF SCAN RISK
                    if risk_factor_col is not None and i == risk_factor_col:
                        # Use the risk factor from the catalog, not from the scan results
                        catalog_risk_value = str(value).strip() if value else ""
                        
                        # Apply color formatting based on the catalog risk level
                        if catalog_risk_value.upper() == 'CRITICAL':
                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                            worksheet_infra.write(row_num, col_idx, camelcase_value, critical_format)
                        elif catalog_risk_value.upper() == 'HIGH':
                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                            worksheet_infra.write(row_num, col_idx, camelcase_value, high_format)
                        elif catalog_risk_value.upper() == 'MEDIUM':
                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                            worksheet_infra.write(row_num, col_idx, camelcase_value, medium_format)
                        elif catalog_risk_value.upper() == 'LOW':
                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                            worksheet_infra.write(row_num, col_idx, camelcase_value, low_format)
                        else:
                            # If catalog doesn't have a risk factor, use the highest from scan results
                            camelcase_value = convert_risk_to_camelcase(group_max_risk)
                            if group_max_risk == 'critical':
                                worksheet_infra.write(row_num, col_idx, camelcase_value, critical_format)
                            elif group_max_risk == 'high':
                                worksheet_infra.write(row_num, col_idx, camelcase_value, high_format)
                            elif group_max_risk == 'medium':
                                worksheet_infra.write(row_num, col_idx, camelcase_value, medium_format)
                            elif group_max_risk == 'low':
                                worksheet_infra.write(row_num, col_idx, camelcase_value, low_format)
                            else:
                                worksheet_infra.write(row_num, col_idx, camelcase_value, cell_format)
                    else:
                        # Special handling for Audit Observation column
                        if audit_observation_col is not None and i == audit_observation_col:
                            # Only add the line if there are multiple vulnerabilities
                            if len(vulnerabilities_list) > 1:
                                observation_text = "It was observed that the hosts are affected by multiple vulnerabilities, which are listed below.\n\n"
                                observation_text += vulnerabilities_str
                                
                                if value:
                                    value = f"{value}\n\n{observation_text}"
                                else:
                                    value = observation_text
                            # For single vulnerability, keep original content or add basic text
                            elif not value:
                                value = "It was observed that the host is affected by a vulnerability."
                        
                        # Special handling for CVE ID column - fill with "N/A" if empty
                        if col_name == 'CVE ID' and value == "":
                            value = "N/A"
                        
                        # Apply left alignment for specific columns
                        if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                            worksheet_infra.write(row_num, col_idx, str(value), left_align_format)
                        else:
                            worksheet_infra.write(row_num, col_idx, str(value), cell_format)
                    
                    col_idx += 1
                
                # Write affected systems (formatted with branch names and IPs on separate lines)
                affected_systems_str = "\n".join(all_affected_systems)
                worksheet_infra.write(row_num, col_idx, affected_systems_str, cell_format)
                col_idx += 1
                
                # Write remaining catalog data
                for i in range(insert_position, len(catalog_headers)):
                    col_name = catalog_headers[i]
                    value = catalog_row[col_name]
                    if pd.isna(value):
                        value = ""
                    
                    # Special handling for Risk Factor column - USE CATALOG VALUE INSTEAD OF SCAN RISK
                    if risk_factor_col is not None and i == risk_factor_col:
                        # Use the risk factor from the catalog, not from the scan results
                        catalog_risk_value = str(value).strip() if value else ""
                        
                        # Apply color formatting based on the catalog risk level
                        if catalog_risk_value.upper() == 'CRITICAL':
                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                            worksheet_infra.write(row_num, col_idx, camelcase_value, critical_format)
                        elif catalog_risk_value.upper() == 'HIGH':
                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                            worksheet_infra.write(row_num, col_idx, camelcase_value, high_format)
                        elif catalog_risk_value.upper() == 'MEDIUM':
                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                            worksheet_infra.write(row_num, col_idx, camelcase_value, medium_format)
                        elif catalog_risk_value.upper() == 'LOW':
                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                            worksheet_infra.write(row_num, col_idx, camelcase_value, low_format)
                        else:
                            # If catalog doesn't have a risk factor, use the highest from scan results
                            camelcase_value = convert_risk_to_camelcase(group_max_risk)
                            if group_max_risk == 'critical':
                                worksheet_infra.write(row_num, col_idx, camelcase_value, critical_format)
                            elif group_max_risk == 'high':
                                worksheet_infra.write(row_num, col_idx, camelcase_value, high_format)
                            elif group_max_risk == 'medium':
                                worksheet_infra.write(row_num, col_idx, camelcase_value, medium_format)
                            elif group_max_risk == 'low':
                                worksheet_infra.write(row_num, col_idx, camelcase_value, low_format)
                            else:
                                worksheet_infra.write(row_num, col_idx, camelcase_value, cell_format)
                    else:
                        # Special handling for Audit Observation column
                        if audit_observation_col is not None and i == audit_observation_col:
                            # Only add the line if there are multiple vulnerabilities
                            if len(vulnerabilities_list) > 1:
                                observation_text = "It was observed that the hosts are affected by multiple vulnerabilities, which are listed below.\n\n"
                                observation_text += vulnerabilities_str
                                
                                if value:
                                    value = f"{value}\n\n{observation_text}"
                                else:
                                    value = observation_text
                            # For single vulnerability, keep original content or add basic text
                            elif not value:
                                value = "It was observed that the host is affected by a vulnerability."
                        
                        # Special handling for CVE ID column - fill with "N/A" if empty
                        if col_name == 'CVE ID' and value == "":
                            value = "N/A"
                        
                        # Apply left alignment for specific columns
                        if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                            worksheet_infra.write(row_num, col_idx, str(value), left_align_format)
                        else:
                            worksheet_infra.write(row_num, col_idx, str(value), cell_format)
                    
                    col_idx += 1
                
                # Write Status column
                worksheet_infra.write(row_num, status_col_position, "", cell_format)
                
                # Write all Old POC columns (M to S)
                for col_idx in range(old_poc_col_start, old_poc_col_end + 1):
                    worksheet_infra.write(row_num, col_idx, "", cell_format)
                
                # Write all POC columns (T to Z)
                for col_idx in range(poc_col_start, poc_col_end + 1):
                    worksheet_infra.write(row_num, col_idx, "", cell_format)
                
                row_num += 1
        
        print(f"Matched vulnerabilities processing completed - {row_num - 1} rows written")
        
        # Handle unmatched vulnerabilities - including merged ones
        unmatched_vulnerabilities = set(unique_vulnerabilities) - matched_vulnerabilities
        print(f"Total unmatched vulnerabilities to process: {len(unmatched_vulnerabilities)}")
        
        if unmatched_vulnerabilities:
            print(f"Unmatched vulnerabilities: {unmatched_vulnerabilities}")
            
            # Check if we have unmatched details mapping from user input
            if unmatched_details_mapping:
                print(f"Processing {len(unmatched_details_mapping)} user-provided vulnerability details")
                # Add vulnerabilities from unmatched_details_mapping to the processing
                for vuln_name in unmatched_details_mapping.keys():
                    if vuln_name in unmatched_vulnerabilities:
                        print(f"Processing user-provided details for vulnerability: {vuln_name}")
            else:
                print("No unmatched details mapping provided - vulnerabilities will be processed with empty details")
            
            # Process merged vulnerabilities first
            merged_vuln_processed = set()
            if unmatched_details_mapping:
                for vulnerability in unmatched_vulnerabilities:
                    if vulnerability in unmatched_details_mapping:
                        details = unmatched_details_mapping[vulnerability]
                        if details.get('isMerged', False):
                            merged_vulns = details.get('mergedVulnerabilities', [])
                            if merged_vulns:
                                # Process merged vulnerability group
                                merged_vuln_processed.update(merged_vulns)
                                
                                # Create vulnerabilities string for "Vulnerabilities in this group" column
                                vulnerabilities_in_group = "\n".join(merged_vulns)
                                
                                worksheet_infra.write(row_num, 0, row_num, cell_format)
                                
                                # Use the first vulnerability name as the main name
                                main_name = merged_vulns[0] if merged_vulns else str(vulnerability)
                                worksheet_infra.write(row_num, 1, main_name, cell_format)
                                
                                # Write catalog data with merged vulnerability details
                                col_idx = 2
                                for i in range(insert_position):
                                    col_name = catalog_headers[i]
                                    value_to_write = ""
                                    
                                    if risk_factor_col is not None and i == risk_factor_col:
                                        value_to_write = str(details.get('riskFactor', '')).upper()
                                    elif audit_observation_col is not None and i == audit_observation_col:
                                        value_to_write = details.get('auditObservation', '')
                                    elif col_name == 'CVE ID':
                                        value_to_write = details.get('cveId', '') or "N/A"
                                    elif 'cvss' in col_name.lower():
                                        value_to_write = details.get('cvss', '')
                                    elif 'impact' in col_name.lower():
                                        value_to_write = details.get('impact', '')
                                    elif 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                                        value_to_write = details.get('recommendation', '')
                                    elif 'reference' in col_name.lower() and 'link' in col_name.lower():
                                        value_to_write = details.get('referenceLink', '')
                                    
                                    # Apply risk color if applicable
                                    if risk_factor_col is not None and i == risk_factor_col and value_to_write:
                                        catalog_risk_value = str(value_to_write).upper().strip()
                                        if catalog_risk_value.upper() == 'CRITICAL':
                                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                            worksheet_infra.write(row_num, col_idx, camelcase_value, critical_format)
                                        elif catalog_risk_value.upper() == 'HIGH':
                                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                            worksheet_infra.write(row_num, col_idx, camelcase_value, high_format)
                                        elif catalog_risk_value.upper() == 'MEDIUM':
                                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                            worksheet_infra.write(row_num, col_idx, camelcase_value, medium_format)
                                        elif catalog_risk_value.upper() == 'LOW':
                                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                            worksheet_infra.write(row_num, col_idx, camelcase_value, low_format)
                                        else:
                                            worksheet_infra.write(row_num, col_idx, catalog_risk_value, cell_format)
                                    else:
                                        # Apply left alignment for specific columns
                                        if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                                            worksheet_infra.write(row_num, col_idx, value_to_write, left_align_format)
                                        else:
                                            worksheet_infra.write(row_num, col_idx, value_to_write, cell_format)
                                    col_idx += 1
                                
                                # Write empty affected systems for merged vulnerabilities
                                worksheet_infra.write(row_num, col_idx, "", cell_format)
                                col_idx += 1
                                
                                # Write remaining catalog columns
                                for i in range(insert_position, len(catalog_headers)):
                                    col_name = catalog_headers[i]
                                    value_to_write = ""
                                    
                                    if risk_factor_col is not None and i == risk_factor_col:
                                        value_to_write = str(details.get('riskFactor', '')).upper()
                                    elif audit_observation_col is not None and i == audit_observation_col:
                                        value_to_write = details.get('auditObservation', '')
                                    elif col_name == 'CVE ID':
                                        value_to_write = details.get('cveId', '') or "N/A"
                                    elif 'cvss' in col_name.lower():
                                        value_to_write = details.get('cvss', '')
                                    elif 'impact' in col_name.lower():
                                        value_to_write = details.get('impact', '')
                                    elif 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                                        value_to_write = details.get('recommendation', '')
                                    elif 'reference' in col_name.lower() and 'link' in col_name.lower():
                                        value_to_write = details.get('referenceLink', '')
                                    elif col_name == 'Vulnerabilities in this group':
                                        value_to_write = vulnerabilities_in_group
                                    
                                    # Apply risk color if applicable
                                    if risk_factor_col is not None and i == risk_factor_col and value_to_write:
                                        catalog_risk_value = str(value_to_write).upper().strip()
                                        if catalog_risk_value.upper() == 'CRITICAL':
                                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                            worksheet_infra.write(row_num, col_idx, camelcase_value, critical_format)
                                        elif catalog_risk_value.upper() == 'HIGH':
                                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                            worksheet_infra.write(row_num, col_idx, camelcase_value, high_format)
                                        elif catalog_risk_value.upper() == 'MEDIUM':
                                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                            worksheet_infra.write(row_num, col_idx, camelcase_value, medium_format)
                                        elif catalog_risk_value.upper() == 'LOW':
                                            camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                            worksheet_infra.write(row_num, col_idx, camelcase_value, low_format)
                                        else:
                                            worksheet_infra.write(row_num, col_idx, catalog_risk_value, cell_format)
                                    else:
                                        # Apply left alignment for specific columns
                                        if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                                            worksheet_infra.write(row_num, col_idx, value_to_write, left_align_format)
                                        else:
                                            worksheet_infra.write(row_num, col_idx, value_to_write, cell_format)
                                    col_idx += 1
                                
                                # Write Status column
                                worksheet_infra.write(row_num, status_col_position, "", cell_format)
                                
                                # Write all Old POC columns (M to S)
                                for col_idx in range(old_poc_col_start, old_poc_col_end + 1):
                                    worksheet_infra.write(row_num, col_idx, "", cell_format)
                                
                                # Write all POC columns (T to Z)
                                for col_idx in range(poc_col_start, poc_col_end + 1):
                                    worksheet_infra.write(row_num, col_idx, "", cell_format)
                                
                                row_num += 1
            
            # Process individual unmatched vulnerabilities (not merged)
            for vulnerability in unmatched_vulnerabilities:
                if vulnerability not in merged_vuln_processed:
                    # Get affected systems for this unmatched vulnerability
                    affected_systems = vulnerability_affected_systems.get(vulnerability, [])
                    affected_systems_str = "\n".join(affected_systems)
                    
                    worksheet_infra.write(row_num, 0, row_num, cell_format)
                    # Use edited vulnerability name if provided
                    if unmatched_details_mapping and vulnerability in unmatched_details_mapping:
                        edited_name = unmatched_details_mapping[vulnerability].get('vulnerabilityName') or str(vulnerability)
                    else:
                        edited_name = str(vulnerability)
                    worksheet_infra.write(row_num, 1, edited_name, cell_format)
                    
                    # Write empty values for catalog columns up to the insertion point
                    col_idx = 2
                    for i in range(insert_position):
                        col_name = catalog_headers[i]
                        value_to_write = ""
                        if unmatched_details_mapping and vulnerability in unmatched_details_mapping:
                            details = unmatched_details_mapping[vulnerability]
                            if risk_factor_col is not None and i == risk_factor_col:
                                value_to_write = str(details.get('riskFactor', '')).upper()
                            elif audit_observation_col is not None and i == audit_observation_col:
                                value_to_write = details.get('auditObservation', '')
                            elif col_name == 'CVE ID':
                                value_to_write = details.get('cveId', '') or "N/A"
                            elif 'cvss' in col_name.lower():
                                value_to_write = details.get('cvss', '')
                            elif 'impact' in col_name.lower():
                                value_to_write = details.get('impact', '')
                            elif 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                                value_to_write = details.get('recommendation', '')
                            elif 'reference' in col_name.lower() and 'link' in col_name.lower():
                                value_to_write = details.get('referenceLink', '')
                        # Apply risk color if applicable
                        if risk_factor_col is not None and i == risk_factor_col and value_to_write:
                            catalog_risk_value = str(value_to_write).upper().strip()
                            if catalog_risk_value.upper() == 'CRITICAL':
                                camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                worksheet_infra.write(row_num, col_idx, camelcase_value, critical_format)
                            elif catalog_risk_value.upper() == 'HIGH':
                                camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                worksheet_infra.write(row_num, col_idx, camelcase_value, high_format)
                            elif catalog_risk_value.upper() == 'MEDIUM':
                                camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                worksheet_infra.write(row_num, col_idx, camelcase_value, medium_format)
                            elif catalog_risk_value.upper() == 'LOW':
                                camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                worksheet_infra.write(row_num, col_idx, camelcase_value, low_format)
                            else:
                                worksheet_infra.write(row_num, col_idx, catalog_risk_value, cell_format)
                        else:
                            worksheet_infra.write(row_num, col_idx, value_to_write, cell_format)
                        col_idx += 1
                    
                    # Write affected systems
                    worksheet_infra.write(row_num, col_idx, affected_systems_str, cell_format)
                    col_idx += 1
                    
                    # Write remaining catalog columns, filling from details when possible
                    for i in range(insert_position, len(catalog_headers)):
                        col_name = catalog_headers[i]
                        value_to_write = ""
                        if unmatched_details_mapping and vulnerability in unmatched_details_mapping:
                            details = unmatched_details_mapping[vulnerability]
                            if risk_factor_col is not None and i == risk_factor_col:
                                value_to_write = str(details.get('riskFactor', '')).upper()
                            elif audit_observation_col is not None and i == audit_observation_col:
                                value_to_write = details.get('auditObservation', '')
                            elif col_name == 'CVE ID':
                                value_to_write = details.get('cveId', '') or "N/A"
                            elif 'cvss' in col_name.lower():
                                value_to_write = details.get('cvss', '')
                            elif 'impact' in col_name.lower():
                                value_to_write = details.get('impact', '')
                            elif 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                                value_to_write = details.get('recommendation', '')
                            elif 'reference' in col_name.lower() and 'link' in col_name.lower():
                                value_to_write = details.get('referenceLink', '')
                        # Apply risk color if applicable
                        if risk_factor_col is not None and i == risk_factor_col and value_to_write:
                            catalog_risk_value = str(value_to_write).upper().strip()
                            if catalog_risk_value.upper() == 'CRITICAL':
                                camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                worksheet_infra.write(row_num, col_idx, camelcase_value, critical_format)
                            elif catalog_risk_value.upper() == 'HIGH':
                                camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                worksheet_infra.write(row_num, col_idx, camelcase_value, high_format)
                            elif catalog_risk_value.upper() == 'MEDIUM':
                                camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                worksheet_infra.write(row_num, col_idx, camelcase_value, medium_format)
                            elif catalog_risk_value.upper() == 'LOW':
                                camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                                worksheet_infra.write(row_num, col_idx, camelcase_value, low_format)
                            else:
                                    worksheet_infra.write(row_num, col_idx, catalog_risk_value, cell_format)
                        else:
                            worksheet_infra.write(row_num, col_idx, value_to_write, cell_format)
                        col_idx += 1
                    
                    # Write Status column
                    worksheet_infra.write(row_num, status_col_position, "", cell_format)
                    
                    # Write all Old POC columns (M to S)
                    for col_idx in range(old_poc_col_start, old_poc_col_end + 1):
                        worksheet_infra.write(row_num, col_idx, "", cell_format)
                    
                    # Write all POC columns (T to Z)
                    for col_idx in range(poc_col_start, poc_col_end + 1):
                        worksheet_infra.write(row_num, col_idx, "", cell_format)
                    
                    row_num += 1
        
        print(f"Unmatched vulnerabilities processing completed - Total rows: {row_num - 1}")
        print(f"Total headers: {len(headers)}")
        print(f"Status column position: {status_col_position}")
        print(f"Old POC columns: {old_poc_col_start} to {old_poc_col_end}")
        print(f"POC columns: {poc_col_start} to {poc_col_end}")
        
        # Set specific column widths as requested (before Status column)
        column_widths = [7, 35, 30, 15, 20, 10, 60, 60, 60, 40, 50]  # First 11 columns (before Status)
        
        # Apply widths to columns before Status
        num_cols_to_set = min(len(column_widths), status_col_position if status_col_position is not None else len(headers))
        for i in range(num_cols_to_set):
            if i < len(column_widths):
                worksheet_infra.set_column(i, i, column_widths[i])
        
        # Set Status column width to 15
        if status_col_position is not None:
            worksheet_infra.set_column(status_col_position, status_col_position, 15)
        
        # Set all Old POC columns (M to S) to width 15
        if old_poc_col_start is not None and old_poc_col_end is not None:
            for col_idx in range(old_poc_col_start, old_poc_col_end + 1):
                worksheet_infra.set_column(col_idx, col_idx, 15)
        
        # Set all POC columns (T to Z) to width 15
        if poc_col_start is not None and poc_col_end is not None:
            for col_idx in range(poc_col_start, poc_col_end + 1):
                worksheet_infra.set_column(col_idx, col_idx, 15)
        
        print(f"Set column widths: Status=15, Old POC (M-S)=15 each, POC (T-Z)=15 each")
        
        print(f"Column formatting applied - {len(headers)} columns configured")
        
        # Set fixed row height of ~50px (≈37.5 points) for all content rows from second row onward
        for r in range(1, row_num):
            worksheet_infra.set_row(r, 37.5)
        
        # Ensure we have at least one data row (beyond headers)
        total_rows = max(1, row_num - 1)
        print(f"Created Infra_VAPT worksheet with {total_rows} data rows")
        print(f"Matched vulnerabilities: {len(matched_vulnerabilities)}, Unmatched: {len(unmatched_vulnerabilities)}")
        print(f"Unique catalog entries used: {len(sorted_catalog_indices)}")
        
        # If no data rows were created, add a message
        if total_rows == 0:
            print("Warning: No vulnerability data was processed for Infra_VAPT worksheet")
        
        # Old POC column width was already set immediately after headers were created
        print("Old POC column width was set to 40 immediately after headers were created")
        
        # Set row heights for all data rows (row >=2) to 50px if they have content
        for row in range(1, row_num):
            if row >= 2:
                # Check if there is any content in the row
                has_content = any(worksheet_infra.table[row][col] for col in range(len(headers))) if hasattr(worksheet_infra, 'table') else True
                if has_content:
                    worksheet_infra.set_row(row, 50)
        
    except Exception as e:
        print(f"Error creating Infra_VAPT worksheet: {e}")
        import traceback
        traceback.print_exc()

def create_summary_worksheet(workbook, combined_nessus, header_format, cell_format):
    """Create Summary worksheet with Name, Branch Name, and Host data from Nessus"""
    try:
        # Always create Summary worksheet first with headers
        worksheet_summary = workbook.add_worksheet("Summary")
        
        # Write headers with Sr.No as first column
        headers = ["Sr.No", "Name", "Branch Name", "Host"]
        for col_num, header in enumerate(headers):
            worksheet_summary.write(0, col_num, header, header_format)
        
        # Set column widths (always set, even if no data)
        worksheet_summary.set_column('A:A', 8)  # Sr.No column
        worksheet_summary.set_column('B:B', 110)  # Name column
        worksheet_summary.set_column('C:C', 50)  # Branch Name column
        worksheet_summary.set_column('D:D', 40)  # Host column
        
        # Create format for green text, bold, centered, with border (reusable)
        no_data_format = workbook.add_format({
            'font_color': '#008000',  # Green color
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'text_wrap': True
        })
        
        # Check if combined_nessus is empty or None
        if combined_nessus is None or combined_nessus.empty:
            print("No Nessus data available - creating Summary worksheet with headers and no data message")
            # Merge A3 to D3 and add message
            worksheet_summary.merge_range(2, 0, 2, 3, "Note: No Vulnerabilities are found during the Audit", no_data_format)
            print("Summary worksheet created with headers and no data message")
            return
        
        # Check if required columns exist
        required_columns = ['Name', 'Host', 'Branch Name', 'Risk']
        missing_cols = [col for col in required_columns if col not in combined_nessus.columns]
        
        if missing_cols:
            print(f"Cannot process Summary worksheet data - missing columns: {missing_cols}")
            # Still create the worksheet with headers and no data message
            # Merge A3 to D3 and add message
            worksheet_summary.merge_range(2, 0, 2, 3, "Note: No Vulnerabilities are found during the Audit", no_data_format)
            print("Summary worksheet created with headers and no data message")
            return
        
        # Filter only low, medium, high, critical vulnerabilities
        valid_risks = ['low', 'medium', 'high', 'critical']
        df_filtered = combined_nessus.copy()
        df_filtered['Risk'] = df_filtered['Risk'].astype(str).str.lower().str.strip()
        df_filtered = df_filtered[df_filtered['Risk'].isin(valid_risks)]
        
        # Select required columns
        df_summary = df_filtered[['Name', 'Branch Name', 'Host']].copy()
        
        # Clean data
        df_summary['Name'] = df_summary['Name'].astype(str).str.strip()
        df_summary['Branch Name'] = df_summary['Branch Name'].astype(str).str.strip()
        df_summary['Host'] = df_summary['Host'].astype(str).str.strip()
        
        # Remove rows with empty values
        df_summary = df_summary.dropna(subset=['Name', 'Host'])
        df_summary = df_summary[(df_summary['Name'] != '') & (df_summary['Host'] != '')]
        
        # If no data available, add message and return
        if df_summary.empty:
            print("No data available for Summary worksheet after filtering - creating worksheet with headers and no data message")
            # Merge A3 to D3 and add message (use the format already created)
            worksheet_summary.merge_range(2, 0, 2, 3, "Note: No Vulnerabilities are found during the Audit", no_data_format)
            print("Summary worksheet created with headers and no data message")
            return
        
        # Deduplicate based on Name + Host combination (vulnerability + IP)
        # Keep only unique combinations of vulnerability and IP
        df_summary = df_summary.drop_duplicates(subset=['Name', 'Host'], keep='first')
        
        # Sort data by Name, Branch Name, and Host to group similar entries
        df_summary = df_summary.sort_values(by=['Name', 'Branch Name', 'Host'])
        
        # Write data and prepare for merging
        row = 1
        current_vulnerability = None
        current_branch = None
        name_merge_start = 1
        branch_merge_start = 1
        serial_counter = 1  # Initialize serial number counter
        
        for _, row_data in df_summary.iterrows():
            name = str(row_data['Name'])
            branch = str(row_data['Branch Name'])
            host = str(row_data['Host'])
            
            # If we're starting a new vulnerability
            if name != current_vulnerability:
                # Merge previous vulnerability name cells if needed
                if current_vulnerability is not None and row > name_merge_start:
                    worksheet_summary.merge_range(
                        name_merge_start, 1, row - 1, 1,  # Name column (column B)
                        current_vulnerability,
                        cell_format
                    )
                    # Also merge serial numbers for this vulnerability
                    worksheet_summary.merge_range(
                        name_merge_start, 0, row - 1, 0,  # Sr.No column (column A)
                        serial_counter - 1,  # Use the previous serial number
                        cell_format
                    )
                current_vulnerability = name
                name_merge_start = row
                current_branch = None  # Reset branch tracking
                serial_counter += 1  # Increment serial number for new vulnerability
            
            # If we're starting a new branch for this vulnerability
            if branch != current_branch:
                # Merge previous branch name cells if needed
                if current_branch is not None and row > branch_merge_start:
                    worksheet_summary.merge_range(
                        branch_merge_start, 2, row - 1, 2,  # Branch Name column (column C)
                        current_branch,
                        cell_format
                    )
                current_branch = branch
                branch_merge_start = row
            
            # Write serial number (same for all rows of this vulnerability)
            worksheet_summary.write(row, 0, serial_counter - 1, cell_format)  # Sr.No column
            worksheet_summary.write(row, 1, name, cell_format)  # Name column
            worksheet_summary.write(row, 2, branch, cell_format)  # Branch Name column
            worksheet_summary.write(row, 3, host, cell_format)  # Host column
            
            row += 1
        
        # Merge the last group of same vulnerability names
        if current_vulnerability is not None and row > name_merge_start:
            worksheet_summary.merge_range(
                name_merge_start, 1, row - 1, 1,  # Name column (column B)
                current_vulnerability,
                cell_format
            )
            # Merge the last group of serial numbers
            worksheet_summary.merge_range(
                name_merge_start, 0, row - 1, 0,  # Sr.No column (column A)
                serial_counter - 1,
                cell_format
            )
        
        # Merge the last group of same branch names
        if current_branch is not None and row > branch_merge_start:
            worksheet_summary.merge_range(
                branch_merge_start, 2, row - 1, 2,  # Branch Name column (column C)
                current_branch,
                cell_format
            )
        
        # Column widths are already set earlier (after headers are written)
        print(f"Summary worksheet created successfully with {row - 1} data rows")
        
    except Exception as e:
        print(f"Error creating Summary worksheet: {e}")
        import traceback
        traceback.print_exc()

def match_images_from_temp_poc2_folder(temp_poc2_folder=None):
    """
    Match images from Temp_POC2 folder with vulnerability names.
    Removes numbers after underscores in image names for matching.
    
    Args:
        temp_poc2_folder (str): Path to the Temp_POC2 folder (with timestamp)
    
    Returns:
        dict: Dictionary mapping vulnerability names to image paths
    """
    # If no folder provided, create timestamped folder name
    if temp_poc2_folder is None:
        from datetime import datetime
        current_ts = datetime.now().strftime('%Y%m%d%H%M%S')
        temp_poc2_folder = f"Temp_POC2_{current_ts}"
    import os
    import re
    
    image_mapping = {}
    
    if not os.path.exists(temp_poc2_folder):
        print(f"Temp_POC2 folder not found: {temp_poc2_folder}")
        return image_mapping
    
    try:
        # Get all image files from the folder
        image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff']
        image_files = []
        
        for filename in os.listdir(temp_poc2_folder):
            if any(filename.lower().endswith(ext) for ext in image_extensions):
                image_files.append(filename)
        
        print(f"Found {len(image_files)} images in {temp_poc2_folder}")
        
        for filename in image_files:
            # Remove file extension
            name_without_ext = os.path.splitext(filename)[0]
            
            # Remove numbers only at the END of the name (e.g., "Vulnerability_2" -> "Vulnerability")
            # This handles cases like "Vulnerability_2.png" -> "Vulnerability"
            # But keeps "Vulnerability_2_SSL.png" -> "Vulnerability_2_SSL"
            clean_name = re.sub(r'_\d+$', '', name_without_ext)
            
            # Replace underscores with spaces for better matching
            clean_name = clean_name.replace('_', ' ')
            
            # Store the mapping - handle multiple images with same base name
            full_path = os.path.join(temp_poc2_folder, filename)
            if clean_name in image_mapping:
                # If key exists, append to list
                if isinstance(image_mapping[clean_name], list):
                    image_mapping[clean_name].append(full_path)
                else:
                    # Convert single item to list
                    image_mapping[clean_name] = [image_mapping[clean_name], full_path]
            else:
                # First image with this base name
                image_mapping[clean_name] = full_path
            
            print(f"Mapped image: {filename} -> {clean_name}")
        
        return image_mapping
        
    except Exception as e:
        print(f"Error matching images from Temp_POC2 folder: {e}")
        return image_mapping

def insert_old_poc_images_to_excel(excel_path, image_mapping):
    """
    Insert Old POC images into the Excel file's Old POC columns (M-S).
    Distributes images across columns: N, O, P, Q, R, S, M (max 7 images per vulnerability).
    
    Args:
        excel_path (str): Path to the Excel file
        image_mapping (dict): Dictionary mapping vulnerability names to image path(s) - can be single path or list
    
    Returns:
        set: Set of row numbers that have Old POC objects
    """
    rows_with_objects = set()
    
    print(f"\n🖼️ === INSERTING OLD POC IMAGES ===")
    print(f"📁 Excel path: {excel_path}")
    print(f"📊 Image mapping contains {len(image_mapping)} vulnerabilities")
    for vuln, paths in list(image_mapping.items())[:3]:  # Show first 3
        path_count = len(paths) if isinstance(paths, list) else 1
        print(f"  - '{vuln}': {path_count} image(s)")
    
    try:
        from openpyxl import load_workbook
        
        # Load the existing workbook
        wb = load_workbook(excel_path)
        
        # Get the Infra_VAPT worksheet
        if "Infra_VAPT" not in wb.sheetnames:
            print("Infra_VAPT worksheet not found")
            return rows_with_objects
        
        ws = wb["Infra_VAPT"]
        
        # Find Old POC columns by looking for merged "Old POC" header in first row
        old_poc_col_start = None
        old_poc_col_end = None
        
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == 1 and merged_range.max_row == 1:
                first_cell = ws.cell(row=1, column=merged_range.min_col)
                if first_cell.value and str(first_cell.value).strip() == "Old POC":
                    old_poc_col_start = merged_range.min_col
                    old_poc_col_end = merged_range.max_col
                    break
        
        if not old_poc_col_start or not old_poc_col_end:
            print("Old POC columns not found in worksheet")
            return rows_with_objects
        
        # Define column order for Old POC image insertion: N, O, P, Q, R, S, M
        # Old POC columns are M-S (7 columns): M, N, O, P, Q, R, S
        # We want order: N, O, P, Q, R, S, M
        old_image_columns = [
            old_poc_col_start + 1,  # N
            old_poc_col_start + 2,  # O
            old_poc_col_start + 3,  # P
            old_poc_col_start + 4,  # Q
            old_poc_col_start + 5,  # R
            old_poc_col_start + 6,  # S (should be old_poc_col_end)
            old_poc_col_start       # M
        ]
        
        print(f"Found Old POC columns from {old_poc_col_start} to {old_poc_col_end}")
        
        # Find Name of Vulnerability column
        vuln_name_col = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and 'name of vulnerability' in str(cell_value).lower():
                vuln_name_col = col
                break
        
        if not vuln_name_col:
            print("Name of Vulnerability column not found in worksheet")
            return rows_with_objects
        
        print(f"Found Old POC columns from {old_poc_col_start} to {old_poc_col_end}")
        print(f"Found Name of Vulnerability column at position: {vuln_name_col}")
        
        # Count how many rows we'll process
        total_rows = ws.max_row - 1  # Excluding header
        print(f"📋 Processing {total_rows} rows in Excel...")
        
        processed_count = 0
        # Normalize function for matching (defined once outside the loop)
        def normalize_for_matching(text):
            """Normalize text for matching by replacing special chars with spaces"""
            special_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|', '_', '-']
            normalized = text.lower()
            for char in special_chars:
                normalized = normalized.replace(char, ' ')
            # Normalize multiple spaces to single space
            normalized = ' '.join(normalized.split())
            return normalized
        
        # Pre-normalize all image mapping keys for efficient lookup
        normalized_image_mapping = {}
        for mapped_name, image_paths in image_mapping.items():
            mapped_name_short = mapped_name[:170].strip()
            mapped_normalized = normalize_for_matching(mapped_name_short)
            if mapped_normalized not in normalized_image_mapping:
                normalized_image_mapping[mapped_normalized] = []
            # Add all images for this normalized key
            if isinstance(image_paths, list):
                normalized_image_mapping[mapped_normalized].extend([(mapped_name, path) for path in image_paths])
            else:
                normalized_image_mapping[mapped_normalized].append((mapped_name, image_paths))
        
        print(f"📊 Pre-normalized {len(normalized_image_mapping)} unique vulnerability names from image mapping")
        for norm_key, images in list(normalized_image_mapping.items())[:5]:
            print(f"  - '{norm_key[:50]}...': {len(images)} image(s)")
        
        # Process each row and match vulnerabilities with Old POC images
        for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
            vuln_name_cell = ws.cell(row=row, column=vuln_name_col)
            if not vuln_name_cell.value:
                continue
            
            vuln_name = str(vuln_name_cell.value).strip()
            if not vuln_name or vuln_name.lower() == 'name of vulnerability':
                continue
            
            processed_count += 1
            if processed_count <= 10:  # Show first 10 for debugging
                print(f"  Row {row}: Checking vulnerability '{vuln_name[:50]}...'")
            
            # Use only first 170 characters for matching
            vuln_name_short = vuln_name[:170].strip()
            vuln_normalized = normalize_for_matching(vuln_name_short)
            
            # Find matching images using pre-normalized mapping
            matching_images = []
            
            # Check for exact normalized match first
            if vuln_normalized in normalized_image_mapping:
                matching_images = normalized_image_mapping[vuln_normalized].copy()
                if processed_count <= 10:
                    for mapped_name, image_path in matching_images:
                        print(f"  🔗 Matched image: {os.path.basename(image_path) if os.path.exists(image_path) else image_path}")
            
            # If no exact match found, try substring matching as fallback
            # But only if we haven't found any matches yet
            if not matching_images:
                for norm_key, images in normalized_image_mapping.items():
                    # Check if vulnerability name contains the mapped name or vice versa
                    if vuln_normalized in norm_key or norm_key in vuln_normalized:
                        matching_images = images.copy()
                        if processed_count <= 10:
                            for mapped_name, image_path in matching_images:
                                print(f"  🔗 Matched image (substring): {os.path.basename(image_path) if os.path.exists(image_path) else image_path}")
                        break  # Only use first substring match
            
            # Sort matching images to ensure consistent ordering
            matching_images.sort(key=lambda x: x[1])  # Sort by file path for consistent ordering
            
            if matching_images:
                print(f"✅ Found {len(matching_images)} Old POC image(s) for vulnerability '{vuln_name[:50]}...'")
                try:
                    # Distribute images across columns: N, O, P, Q, R, S, M
                    num_images_to_insert = min(len(matching_images), 7)  # Max 7 images (one per column)
                    
                    for img_idx in range(num_images_to_insert):
                        mapped_name, matching_image = matching_images[img_idx]
                        col_idx = old_image_columns[img_idx]
                        if os.path.exists(matching_image):
                            try:
                                # Get cell reference (e.g., "N2", "O3", etc.)
                                from openpyxl.utils import get_column_letter
                                col_letter = get_column_letter(col_idx)
                                cell_ref = f"{col_letter}{row}"
                                
                                # Check if an image already exists in this cell
                                image_exists = False
                                for existing_image in ws._images:
                                    try:
                                        # Check anchor position - openpyxl uses 0-based indexing
                                        if hasattr(existing_image, 'anchor'):
                                            anchor = existing_image.anchor
                                            # Handle different anchor types
                                            if hasattr(anchor, '_from'):
                                                anchor_col = anchor._from.col
                                                anchor_row = anchor._from.row
                                                # Compare with 0-based row/col (row is 1-based, col_idx is 1-based)
                                                if anchor_col == col_idx - 1 and anchor_row == row - 1:
                                                    image_exists = True
                                                    break
                                            elif hasattr(anchor, 'col') and hasattr(anchor, 'row'):
                                                # Alternative anchor format
                                                if anchor.col == col_idx - 1 and anchor.row == row - 1:
                                                    image_exists = True
                                                    break
                                    except Exception as anchor_check_err:
                                        # Skip if we can't check this image's anchor
                                        continue
                                
                                if image_exists:
                                    if processed_count <= 10:
                                        print(f"  ⏭️ Skipping Old POC image {img_idx + 1} at {cell_ref} - image already exists")
                                    continue
                                
                                # Load the image
                                from openpyxl.drawing.image import Image
                                img = Image(matching_image)
                                
                                # Get original dimensions
                                original_width = img.width
                                original_height = img.height
                                
                                # Resize image to height=30px, width proportionally (maintain aspect ratio)
                                if hasattr(img, 'width') and hasattr(img, 'height'):
                                    target_height = 30
                                    aspect_ratio = img.width / img.height if img.height else 1
                                    img.height = target_height
                                    img.width = int(target_height * aspect_ratio)
                                
                                # Insert image at the cell
                                ws.add_image(img, cell_ref)
                                
                                print(f"✅ Inserted Old POC image {img_idx + 1} at {cell_ref} for vulnerability: {vuln_name} (reduced from {original_width}x{original_height})")
                                
                            except Exception as e:
                                print(f"⚠️ Error inserting Old POC image at column {col_idx}, row {row}: {e}")
                    
                    # Mark this row as having objects
                    if num_images_to_insert > 0:
                        rows_with_objects.add(row)
                    
                except Exception as e:
                    print(f"❌ Error processing Old POC images for row {row}: {e}")
                    import traceback
                    traceback.print_exc()
                    continue
            else:
                # No images found for this vulnerability
                if processed_count <= 5:  # Log first 5 mismatches
                    print(f"  ⚠️ No images matched for '{vuln_name[:50]}...'")
        
        # Apply custom borders to ALL Old POC and POC columns (M to Z)
        from openpyxl.styles import Border, Side
        
        # Define border styles for each column type
        # M column: left, top, bottom (NOT right)
        m_border = Border(
            left=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # N, O, P, Q, R, T, U, V, W, X, Y columns: top and bottom only (NOT left or right)
        middle_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # S and Z columns: top, bottom, right (NOT left)
        right_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            right=Side(style='thin')
        )
        
        # Find all rows that are part of the data table (have content in any column)
        table_rows = set()
        
        # Always include header row
        table_rows.add(1)
        
        # Check all rows from 2 onwards to find data rows
        for row in range(2, ws.max_row + 1):
            has_content = False
            
            # Check if any cell in this row has content (excluding Old POC and POC columns)
            for col in range(1, old_poc_col_start):  # Check only columns before Old POC
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip() != "":
                    has_content = True
                    break
            
            if has_content:
                table_rows.add(row)
        
        # Apply custom borders to all table rows for Old POC and POC columns
        for row_num in sorted(table_rows):
            # M column: left, top, bottom
            ws.cell(row=row_num, column=old_poc_col_start).border = m_border
            
            # N, O, P, Q, R columns: top, bottom only
            for col_idx in range(old_poc_col_start + 1, old_poc_col_end):
                ws.cell(row=row_num, column=col_idx).border = middle_border
            
            # S column: top, bottom, right
            ws.cell(row=row_num, column=old_poc_col_end).border = right_border
        
        print(f"Identified {len(table_rows)} table rows total")
        print(f"Applied custom borders to Old POC columns (M-S) for {len(table_rows)} rows")
        
        # Save the workbook
        wb.save(excel_path)
        print(f"\n✅ OLD POC IMAGES INSERTION COMPLETE!")
        print(f"   Rows with Old POC images: {len(rows_with_objects)}")
        print(f"   Total table rows processed: {len(table_rows)}")
        print(f"   Border applied to all Old POC columns (M-S)")
        
    except Exception as e:
        print(f"Error inserting Old POC images: {e}")
        import traceback
        traceback.print_exc()
    
    return rows_with_objects

def extract_images_from_infra_vapt_worksheet(excel_file, output_folder=None):
    """
    Extract images from the Infra_VAPT worksheet, specifically from the POC column.
    Images are named based on the 'Name of Vulnerability' column.
    
    Args:
        excel_file: Either a file path (str) or Flask file upload object
        output_folder (str): Folder to save extracted images (with timestamp)
    
    Returns:
        dict: Dictionary with extraction results and statistics
    """
    # If no folder provided, create timestamped folder name
    if output_folder is None:
        from datetime import datetime
        current_ts = datetime.now().strftime('%Y%m%d%H%M%S')
        output_folder = f"Temp_POC2_{current_ts}"
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image
    import os
    from PIL import Image as PILImage
    import io
    import re
    import tempfile
    
    # Create output folder if it doesn't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        print(f"Created folder: {output_folder}")
    
    temp_file_path = None
    try:
        # Handle both file path and Flask file upload object
        if hasattr(excel_file, 'read'):  # Flask file upload object
            # Save uploaded file temporarily
            try:
                print(f"Processing Flask file upload object: {excel_file.filename}")
                
                # Ensure file pointer is at the beginning
                excel_file.seek(0)
                
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                    print(f"Creating temporary file: {temp_file.name}")
                    excel_file.save(temp_file.name)
                    temp_file_path = temp_file.name
                    print(f"Saved file to temporary path: {temp_file_path}")
                
                # Verify the temporary file exists and has content
                if not os.path.exists(temp_file_path):
                    raise Exception(f"Temporary file was not created: {temp_file_path}")
                
                file_size = os.path.getsize(temp_file_path)
                print(f"Temporary file size: {file_size} bytes")
                
                if file_size == 0:
                    raise Exception("Temporary file is empty")
                
                # Load the workbook from temporary file
                print("Loading workbook from temporary file...")
                wb = load_workbook(temp_file_path)
                print("Workbook loaded successfully")
                
            except Exception as e:
                print(f"Error handling temporary file: {e}")
                import traceback
                traceback.print_exc()
                # Clean up temporary file if it was created
                if temp_file_path and os.path.exists(temp_file_path):
                    try:
                        os.unlink(temp_file_path)
                        print(f"Cleaned up temporary file after error: {temp_file_path}")
                    except Exception as cleanup_e:
                        print(f"Warning: Could not delete temporary file {temp_file_path}: {cleanup_e}")
                raise e
        else:  # File path string
            wb = load_workbook(excel_file)
        
        # Check if Infra_VAPT worksheet exists
        if 'Infra_VAPT' not in wb.sheetnames:
            print("Error: 'Infra_VAPT' worksheet not found in the Excel file")
            return {"success": False, "error": "Infra_VAPT worksheet not found"}
        
        ws = wb['Infra_VAPT']
        
        # Find the column indices for 'Name of Vulnerability' and 'POC'
        name_col_idx = None
        poc_col_idx = None
        
        # Look for headers in the first few rows
        for row in range(1, min(10, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    cell_str = str(cell_value).strip().lower()
                    if 'name of vulnerability' in cell_str:
                        name_col_idx = col
                    elif 'poc' in cell_str:
                        poc_col_idx = col
        
        if name_col_idx is None:
            print("Error: 'Name of Vulnerability' column not found")
            return {"success": False, "error": "Name of Vulnerability column not found"}
        
        if poc_col_idx is None:
            print("Error: 'POC' column not found")
            return {"success": False, "error": "POC column not found"}
        
        print(f"Found 'Name of Vulnerability' column at index: {name_col_idx}")
        print(f"Found 'POC' column at index: {poc_col_idx}")
        
        # Get all images in the worksheet
        if not ws._images:
            print("No images found in the Infra_VAPT worksheet")
            return {"success": True, "images_extracted": 0, "message": "No images found"}
        
        print(f"Found {len(ws._images)} images in the worksheet")
        
        # Create a mapping of cell positions to vulnerability names
        vulnerability_mapping = {}
        for row in range(1, ws.max_row + 1):
            vuln_name_cell = ws.cell(row=row, column=name_col_idx)
            if vuln_name_cell.value:
                vuln_name = str(vuln_name_cell.value).strip()
                if vuln_name and vuln_name.lower() != 'name of vulnerability':
                    vulnerability_mapping[row] = vuln_name
        
        print(f"Found {len(vulnerability_mapping)} vulnerability entries")
        
        # Extract images
        total_images_extracted = 0
        images_info = []
        
        for i, image in enumerate(ws._images):
            try:
                # Get image data
                image_data = image._data()
                
                # Determine image format from signature
                if image_data.startswith(b'\xff\xd8\xff'):
                    extension = 'jpg'
                elif image_data.startswith(b'\x89PNG\r\n\x1a\n'):
                    extension = 'png'
                elif image_data.startswith(b'GIF8'):
                    extension = 'gif'
                elif image_data.startswith(b'BM'):
                    extension = 'bmp'
                else:
                    extension = 'png'  # default fallback
                
                # Get cell position from anchor
                cell_ref = get_cell_reference_from_anchor(image.anchor, ws)
                row_num = get_row_number_from_cell_ref(cell_ref)
                
                # Find the vulnerability name for this row
                vuln_name = None
                if row_num in vulnerability_mapping:
                    vuln_name = vulnerability_mapping[row_num]
                else:
                    # If exact row not found, find the closest vulnerability name
                    for v_row, v_name in vulnerability_mapping.items():
                        if abs(v_row - row_num) <= 2:  # Within 2 rows
                            vuln_name = v_name
                            break
                
                if not vuln_name:
                    vuln_name = f"Unknown_Vulnerability_{row_num}"
                    print(f"Warning: No vulnerability name found for image at row {row_num}")
                
                # Clean the vulnerability name for filename
                clean_vuln_name = re.sub(r'[<>:"/\\|?*]', '_', vuln_name)
                clean_vuln_name = clean_vuln_name.replace(' ', '_')
                clean_vuln_name = clean_vuln_name[:50]  # Limit filename length
                
                # Create filename
                image_filename = f"{clean_vuln_name}.{extension}"
                
                # Check if file already exists (multiple images for same vulnerability)
                counter = 1
                final_filename = image_filename
                full_path = os.path.join(output_folder, final_filename)
                
                while os.path.exists(full_path):
                    counter += 1
                    name_part = clean_vuln_name
                    final_filename = f"{name_part}_{counter}.{extension}"
                    full_path = os.path.join(output_folder, final_filename)
                
                # Save the image
                with open(full_path, 'wb') as img_file:
                    img_file.write(image_data)
                
                total_images_extracted += 1
                images_info.append({
                    'filename': final_filename,
                    'vulnerability_name': vuln_name,
                    'cell_reference': cell_ref,
                    'row_number': row_num,
                    'size_bytes': len(image_data),
                    'format': extension
                })
                
                print(f"Extracted image: {final_filename} for vulnerability: {vuln_name}")
                
            except Exception as e:
                print(f"Error extracting image {i+1}: {str(e)}")
        
        print(f"\nTotal images extracted: {total_images_extracted} to folder: {output_folder}")
        
        # Clean up temporary file if it was created
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                print(f"Cleaned up temporary file: {temp_file_path}")
            except Exception as e:
                print(f"Warning: Could not delete temporary file {temp_file_path}: {e}")
        
        return {
            "success": True,
            "images_extracted": total_images_extracted,
            "output_folder": output_folder,
            "images_info": images_info
        }
        
    except Exception as e:
        print(f"Error extracting images from Infra_VAPT worksheet: {str(e)}")
        
        # Clean up temporary file if it was created
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                print(f"Cleaned up temporary file after error: {temp_file_path}")
            except Exception as cleanup_e:
                print(f"Warning: Could not delete temporary file {temp_file_path}: {cleanup_e}")
        
        return {"success": False, "error": str(e)}

def get_cell_reference_from_anchor(anchor, worksheet):
    """
    Determine the cell reference from image anchor position.
    
    Args:
        anchor: The image anchor object
        worksheet: The worksheet object
    
    Returns:
        str: Cell reference (e.g., 'C5')
    """
    try:
        # For TwoCellAnchor (most common)
        if hasattr(anchor, '_from') and hasattr(anchor, 'to'):
            col_from = anchor._from.col + 1  # 0-based to 1-based
            row_from = anchor._from.row + 1  # 0-based to 1-based
            
            # Convert to cell reference
            from openpyxl.utils import get_column_letter
            cell_ref = f"{get_column_letter(col_from)}{row_from}"
            return cell_ref
        
        # For OneCellAnchor (alternative anchor type)
        elif hasattr(anchor, '_from'):
            col_from = anchor._from.col + 1
            row_from = anchor._from.row + 1
            from openpyxl.utils import get_column_letter
            cell_ref = f"{get_column_letter(col_from)}{row_from}"
            return cell_ref
            
    except Exception as e:
        print(f"Error determining cell reference: {str(e)}")
        return "Unknown"
    
    return "Unknown"

def get_row_number_from_cell_ref(cell_ref):
    """
    Extract row number from cell reference (e.g., 'C5' -> 5)
    
    Args:
        cell_ref (str): Cell reference like 'C5'
    
    Returns:
        int: Row number
    """
    try:
        import re
        match = re.match(r'[A-Z]+(\d+)', cell_ref)
        if match:
            return int(match.group(1))
        return 0
    except:
        return 0

def get_all_images_info_from_infra_vapt(excel_file):
    """
    Get information about all images in the Infra_VAPT worksheet without extracting them.
    
    Args:
        excel_file: Either a file path (str) or Flask file upload object
    
    Returns:
        dict: Dictionary containing image information and statistics
    """
    from openpyxl import load_workbook
    import tempfile
    import os
    
    temp_file_path = None
    try:
        # Handle both file path and Flask file upload object
        if hasattr(excel_file, 'read'):  # Flask file upload object
            # Save uploaded file temporarily
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                    excel_file.save(temp_file.name)
                    temp_file_path = temp_file.name
                
                # Load the workbook from temporary file
                wb = load_workbook(temp_file_path)
                
            except Exception as e:
                print(f"Error handling temporary file: {e}")
                # Clean up temporary file if it was created
                if temp_file_path and os.path.exists(temp_file_path):
                    try:
                        os.unlink(temp_file_path)
                    except:
                        pass
                raise e
        else:  # File path string
            wb = load_workbook(excel_file)
        
        if 'Infra_VAPT' not in wb.sheetnames:
            return {"success": False, "error": "Infra_VAPT worksheet not found"}
        
        ws = wb['Infra_VAPT']
        
        if not ws._images:
            return {"success": True, "images_found": 0, "message": "No images found"}
        
        images_info = []
        
        # Find the column indices for 'Name of Vulnerability'
        name_col_idx = None
        for row in range(1, min(10, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and 'name of vulnerability' in str(cell_value).strip().lower():
                    name_col_idx = col
                    break
            if name_col_idx:
                break
        
        # Create vulnerability mapping
        vulnerability_mapping = {}
        if name_col_idx:
            for row in range(1, ws.max_row + 1):
                vuln_name_cell = ws.cell(row=row, column=name_col_idx)
                if vuln_name_cell.value:
                    vuln_name = str(vuln_name_cell.value).strip()
                    if vuln_name and vuln_name.lower() != 'name of vulnerability':
                        vulnerability_mapping[row] = vuln_name
        
        for i, image in enumerate(ws._images):
            try:
                image_data = image._data()
                cell_ref = get_cell_reference_from_anchor(image.anchor, ws)
                row_num = get_row_number_from_cell_ref(cell_ref)
                
                # Find vulnerability name
                vuln_name = vulnerability_mapping.get(row_num, f"Unknown_{row_num}")
                
                # Determine image format
                if image_data.startswith(b'\xff\xd8\xff'):
                    extension = 'jpg'
                elif image_data.startswith(b'\x89PNG\r\n\x1a\n'):
                    extension = 'png'
                elif image_data.startswith(b'GIF8'):
                    extension = 'gif'
                elif image_data.startswith(b'BM'):
                    extension = 'bmp'
                else:
                    extension = 'unknown'
                
                images_info.append({
                    'index': i + 1,
                    'cell_reference': cell_ref,
                    'row_number': row_num,
                    'vulnerability_name': vuln_name,
                    'size_bytes': len(image_data),
                    'format': extension
                })
                
            except Exception as e:
                print(f"Error getting info for image {i+1}: {str(e)}")
        
        # Clean up temporary file if it was created
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                print(f"Cleaned up temporary file: {temp_file_path}")
            except Exception as e:
                print(f"Warning: Could not delete temporary file {temp_file_path}: {e}")
        
        return {
            "success": True,
            "images_found": len(images_info),
            "images_info": images_info,
            "vulnerability_mapping": vulnerability_mapping
        }
        
    except Exception as e:
        # Clean up temporary file if it was created
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                print(f"Cleaned up temporary file after error: {temp_file_path}")
            except Exception as cleanup_e:
                print(f"Warning: Could not delete temporary file {temp_file_path}: {cleanup_e}")
        
        return {"success": False, "error": str(e)}

@follow_up_audit_bp.route('/add_old_poc_images', methods=['POST'])
def add_old_poc_images_route():
    """
    Standalone route to add Old POC images to an uploaded Excel file.
    """
    try:
        # Check if file was uploaded
        if 'excelFile' not in request.files:
            return jsonify({'error': 'No Excel file uploaded'}), 400
        
        excel_file = request.files['excelFile']
        
        if excel_file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file extension
        if not excel_file.filename.lower().endswith('.xlsx'):
            return jsonify({'error': 'Please upload a valid Excel file (.xlsx)'}), 400
        
        print(f"Processing Old POC images for file: {excel_file.filename}")
        
        # Save uploaded file temporarily
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            excel_file.save(temp_file.name)
            temp_file_path = temp_file.name
        
        try:
            # Create timestamped folder name for this request
            from datetime import datetime
            current_ts = datetime.now().strftime('%Y%m%d%H%M%S')
            temp_poc2_folder_route = f"Temp_POC2_{current_ts}"
            
            # Match images from timestamped Temp_POC2 folder
            old_poc_image_mapping = match_images_from_temp_poc2_folder(temp_poc2_folder_route)
            
            if not old_poc_image_mapping:
                return jsonify({
                    'success': False,
                    'error': f'No images found in {temp_poc2_folder_route} folder'
                }), 404
            
            # Insert Old POC images into the Excel file
            old_poc_rows = insert_old_poc_images_to_excel(temp_file_path, old_poc_image_mapping)
            
            if old_poc_rows:
                # Read the updated file
                with open(temp_file_path, 'rb') as updated_file:
                    file_data = updated_file.read()
                
                # Clean up temporary file
                os.unlink(temp_file_path)
                
                # Clean up Temp_POC2 folder after processing
                try:
                    if os.path.exists(temp_poc2_folder_route):
                        import shutil
                        shutil.rmtree(temp_poc2_folder_route)
                        print(f"Successfully deleted {temp_poc2_folder_route} folder (standalone route)")
                except Exception as e:
                    print(f"Error deleting {temp_poc2_folder_route} folder (standalone route): {e}")
                
                # Return the updated Excel file
                response = make_response(file_data)
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response.headers['Content-Disposition'] = f'attachment; filename="updated_{excel_file.filename}"'
                
                return response
            else:
                return jsonify({
                    'success': False,
                    'error': 'No Old POC images were matched with vulnerabilities'
                }), 404
                
        except Exception as e:
            # Clean up temporary file on error
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
            
            # Clean up Temp_POC2 folder on error
            try:
                if 'temp_poc2_folder_route' in locals() and os.path.exists(temp_poc2_folder_route):
                    import shutil
                    shutil.rmtree(temp_poc2_folder_route)
                    print(f"Successfully deleted {temp_poc2_folder_route} folder (standalone route error cleanup)")
            except Exception as cleanup_e:
                if 'temp_poc2_folder_route' in locals():
                    print(f"Error deleting {temp_poc2_folder_route} folder (standalone route error cleanup): {cleanup_e}")
                else:
                    print(f"Error deleting Temp_POC2 folder (standalone route error cleanup): {cleanup_e}")
            
            raise e
            
    except Exception as e:
        print(f"Error in add_old_poc_images_route: {e}")
        return jsonify({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }), 500

# Example usage and testing functions
def test_image_extraction(excel_file):
    """
    Test function to demonstrate image extraction from Infra_VAPT worksheet.
    
    Args:
        excel_file (str): Path to the Excel file
    """
    print("=" * 60)
    print("TESTING IMAGE EXTRACTION FROM INFRA_VAPT WORKSHEET")
    print("=" * 60)
    
    # First, get information about all images
    print("\n1. Scanning for images in Infra_VAPT worksheet...")
    images_info = get_all_images_info_from_infra_vapt(excel_file)
    
    if images_info["success"]:
        if images_info["images_found"] > 0:
            print(f"\nFound {images_info['images_found']} images:")
            for info in images_info["images_info"]:
                print(f"  Image {info['index']}: Cell {info['cell_reference']}, "
                      f"Row {info['row_number']}, Vulnerability: {info['vulnerability_name']}, "
                      f"Size: {info['size_bytes']} bytes, Format: {info['format']}")
            
            # Extract all images to timestamped folder
            from datetime import datetime
            current_ts = datetime.now().strftime('%Y%m%d%H%M%S')
            temp_poc2_folder_test = f"Temp_POC2_{current_ts}"
            print(f"\n2. Extracting images to {temp_poc2_folder_test} folder...")
            result = extract_images_from_infra_vapt_worksheet(excel_file, output_folder=temp_poc2_folder_test)
            
            if result["success"]:
                print(f"\n✅ SUCCESS: Extracted {result['images_extracted']} images")
                print(f"📁 Output folder: {result['output_folder']}")
                
                if result["images_info"]:
                    print("\nExtracted images:")
                    for img in result["images_info"]:
                        print(f"  - {img['filename']} (Vulnerability: {img['vulnerability_name']})")
            else:
                print(f"\n❌ ERROR: {result['error']}")
        else:
            print("No images found in the Infra_VAPT worksheet.")
    else:
        print(f"\n❌ ERROR: {images_info['error']}")
    
    print("\n" + "=" * 60)

# Main execution block for testing
if __name__ == "__main__":
    # Example usage - this will be called from Flask routes with actual uploaded file
    print("Image extraction functions are ready to be used with uploaded Excel files.")
    print("Use extract_images_from_infra_vapt_worksheet(user_excel_file, output_folder=\"Temp_POC2_YYYYMMDDHHMMSS\") in your Flask routes.")

