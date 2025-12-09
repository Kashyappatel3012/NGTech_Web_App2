from flask import Blueprint, request, flash, redirect, url_for, send_file
from flask_login import login_required
from datetime import datetime
import os
import tempfile
import shutil
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side

# Create blueprint
branch_excel_bp = Blueprint('branch_excel_bp', __name__)

@branch_excel_bp.route('/test_excel_route')
def test_excel_route():
    return "Excel route is working!"

@branch_excel_bp.route('/process_branch_excel_without_poc', methods=['POST'])
@login_required
def process_branch_excel_without_poc():
    print("="*50)
    print("FORM SUBMISSION RECEIVED!")
    print("="*50)
    try:
        # Get form data
        sr_no = request.form.get('srNo')
        branch_code = request.form.get('branchCode')
        branch_name = request.form.get('branchName')
        date_of_visit = request.form.get('dateOfVisit')
        assisted_by_prefix = request.form.get('assistedByPrefix')
        assisted_by_name = request.form.get('assistedByName')
        
        # Print all form data to console
        print("\n" + "="*50)
        print("BRANCH EXCEL FORM DATA RECEIVED")
        print("="*50)
        print(f"Sr.No: {sr_no}")
        print(f"Branch Code: {branch_code}")
        print(f"Branch Name: {branch_name}")
        print(f"Date of Visit: {date_of_visit}")
        print(f"Assisted By Prefix: {assisted_by_prefix}")
        print(f"Assisted By Name: {assisted_by_name}")
        
        # Print all form data (including arrays and other fields)
        print("\nALL FORM DATA:")
        print("-" * 30)
        for key, value in request.form.items():
            print(f"{key}: {value}")
        
        # Print auditor names if present
        auditor_prefixes = request.form.getlist('auditorNamePrefix[]')
        auditor_names = request.form.getlist('auditorName[]')
        if auditor_prefixes or auditor_names:
            print(f"\nAUDITOR DETAILS:")
            print("-" * 20)
            for i, (prefix, name) in enumerate(zip(auditor_prefixes, auditor_names)):
                print(f"Auditor {i+1}: {prefix} {name}")
        
        # Print system details if present
        system_ips = request.form.getlist('systemIp[]')
        mac_addresses = request.form.getlist('macAddress[]')
        employee_names = request.form.getlist('employeeName[]')
        designations = request.form.getlist('designation[]')
        if system_ips or mac_addresses or employee_names or designations:
            print(f"\nSYSTEM DETAILS:")
            print("-" * 20)
            for i, (ip, mac, emp_name, designation) in enumerate(zip(system_ips, mac_addresses, employee_names, designations)):
                print(f"System {i+1}:")
                print(f"  IP: {ip}")
                print(f"  MAC: {mac}")
                print(f"  Employee: {emp_name}")
                print(f"  Designation: {designation}")
        
        # Print ATM details if present
        atm_ips = request.form.getlist('atmIp[]')
        has_atm = request.form.get('hasAtm')
        if atm_ips or has_atm:
            print(f"\nATM DETAILS:")
            print("-" * 20)
            print(f"Has ATM: {has_atm}")
            for i, atm_ip in enumerate(atm_ips):
                print(f"ATM IP {i+1}: {atm_ip}")
        
        # Print security assessment answers
        security_fields = [
            'strongPasswords', 'passwordChanges', 'passwordSharing', 'failedLoginAttempts',
            'sessionTimeout', 'multipleLogins', 'cbsIsolation', 'internetRestriction',
            'defaultPasswordPolicy', 'userCredentials', 'sslUsage', 'cbsType',
            'browserCompatibility', 'twoFactorAuth'
        ]
        
        print(f"\nSECURITY ASSESSMENT ANSWERS:")
        print("-" * 30)
        for field in security_fields:
            value = request.form.get(field)
            if value:
                print(f"{field}: {value}")
        
        # Print physical and environmental security answers
        physical_security_fields = [
            'cctvCameraPresent', 'cctvCameraNumber', 'branchLockerCctv', 'branchLockerCctvNumber',
            'cctvCoveringCompleteArea', 'dvrWorkingProperly', 'ntpConfiguredDvr', 'dvrInBankNetwork',
            'cameraNightVision', 'cctvHistoryDuration', 'lockerCctvHistoryDuration', 'biometricDevices',
            'idCardIssued', 'guardAvailable', 'secureAreasControlled', 'smokeDetectorInstalled',
            'panicSwitchInstalled', 'notificationReachAuthorities', 'lanPortsOpen', 'visitorRegisterMaintained',
            'environmentalControlsDrill', 'externalModemsUsed', 'fireExtinguisherAvailable', 'enoughFireExtinguisher'
        ]
        
        print(f"\nPHYSICAL AND ENVIRONMENTAL SECURITY ANSWERS:")
        print("-" * 40)
        for field in physical_security_fields:
            value = request.form.get(field)
            if value:
                print(f"{field}: {value}")
        
        # Print power backup answers
        power_backup_fields = [
            'batteryBackupSupport', 'amcPowerBackup', 'generatorAvailable'
        ]
        
        print(f"\nPOWER BACK UP ANSWERS:")
        print("-" * 25)
        for field in power_backup_fields:
            value = request.form.get(field)
            if value:
                print(f"{field}: {value}")
        
        # Print user awareness answers
        user_awareness_fields = [
            'infoSecurityTraining', 'applicationUsageTraining', 'fireExtinguisherPanicSwitchAwareness', 'cctvRecordingHistoryAwareness'
        ]
        
        print(f"\nUSER AWARENESS ANSWERS:")
        print("-" * 25)
        for field in user_awareness_fields:
            value = request.form.get(field)
            if value:
                print(f"{field}: {value}")
        
        # Print maintenance and business continuity answers
        maintenance_fields = [
            'amcComputerSystems', 'complaintRegisterMaintained'
        ]
        
        print(f"\nMAINTENANCE AND BUSINESS CONTINUITY CONTROLS ANSWERS:")
        print("-" * 50)
        for field in maintenance_fields:
            value = request.form.get(field)
            if value:
                print(f"{field}: {value}")
        
        # Print patch management answers
        patch_management_fields = [
            'antivirusAvailable', 'antivirusAvailableIps', 'antivirusUpdated', 'antivirusUpdatedIps',
            'windowsPatchUpdated', 'windowsPatchIps', 'outdatedWindowsVersion', 'outdatedWindowsIps',
            'expiredLicenseAntivirus', 'expiredLicenseIps'
        ]
        
        print(f"\nPATCH MANAGEMENT ANSWERS:")
        print("-" * 25)
        for field in patch_management_fields:
            value = request.form.get(field)
            if value:
                print(f"{field}: {value}")
        
        # Print network security answers
        network_security_fields = [
            'networkAccessControls', 'networkDiagramAvailable', 'structuredCabling', 'cableTagging',
            'backupNetwork', 'networkingDevicesSecurity', 'networkingDevicesCooling', 'wifiAvailable', 'firewallAvailable'
        ]
        
        print(f"\nNETWORK SECURITY ANSWERS:")
        print("-" * 25)
        for field in network_security_fields:
            value = request.form.get(field)
            if value:
                print(f"{field}: {value}")
        
        # Print endpoints vulnerability answers
        endpoints_vulnerability_fields = [
            'adminRightsLogin', 'adminRightsIps', 'groupPolicyModification', 'groupPolicyIps',
            'securityConfigModification', 'securityConfigIps', 'proxyModification', 'proxyIps',
            'internetAllowed', 'internetAccessRestricted', 'internetAuthPolicy', 'systemPasswordWeak',
            'systemTagging', 'systemTaggingIps', 'ntpConfigured', 'usbPortsEnabled', 'usbApprovalProcess',
            'trustedUsbCheck', 'firewallEnabled', 'firewallEnabledIps', 'externalStorageVirusCheck',
            'rdpEnabled', 'rdpIps'
        ]
        
        print(f"\nENDPOINTS VULNERABILITY ANSWERS:")
        print("-" * 30)
        for field in endpoints_vulnerability_fields:
            value = request.form.get(field)
            if value:
                print(f"{field}: {value}")
        
        # Print ATM Machine Room answers
        atm_machine_room_fields = [
            'atmCctvCableConcealed', 'atmGuardAvailable', 'atmCctvPosition', 'atmWorking247',
            'atmDosDontsAwareness', 'atmNetworkSegmented', 'atmMachineGrouted', 'atmRoomAccessRestricted',
            'atmPowerBackupUps', 'atmCashReplenishmentDualControl', 'atmReconciliationStatus', 'atmJournalPrintPreservation'
        ]
        
        print(f"\nATM MACHINE ROOM ANSWERS:")
        print("-" * 25)
        for field in atm_machine_room_fields:
            value = request.form.get(field)
            if value:
                print(f"{field}: {value}")
        
        # Print email security answers
        email_security_fields = [
            'personalMailUsed', 'emailTwoFactorAuth', 'singleMailMultipleUsers', 'officialEmailOutsideNetwork'
        ]
        
        print(f"\nEMAIL-SECURITY ANSWERS:")
        print("-" * 20)
        for field in email_security_fields:
            value = request.form.get(field)
            if value:
                print(f"{field}: {value}")
        
        # Print remote access answers
        remote_access_fields = [
            'departmentRemoteAccess', 'remoteAccessApprovalProcess'
        ]
        
        print(f"\nREMOTE ACCESS ANSWERS:")
        print("-" * 20)
        for field in remote_access_fields:
            value = request.form.get(field)
            if value:
                print(f"{field}: {value}")
        
        # Print unauthorized applications and personal data answers
        unauthorized_apps_personal_data_fields = [
            'unauthorizedApplications', 'unauthorizedAppsIps', 'unauthorizedAppsNames', 
            'personalDataPresent', 'personalDataIps'
        ]
        
        print(f"\nUNAUTHORIZED APPLICATIONS / PERSONAL DATA ANSWERS:")
        print("-" * 45)
        for field in unauthorized_apps_personal_data_fields:
            value = request.form.get(field)
            if value:
                print(f"{field}: {value}")
        
        # Print important note answers
        important_note_fields = [
            'auditorIdentityNotVerified', 'windowsNotActivated', 'windowsNotActivatedIps',
            'passwordSavedInBrowser', 'passwordBrowserIps', 'passwordWrittenOnWall',
            'assetMovementRegisterNotAvailable', 'dustPresentOnDevices', 'autoSwitchableModeNotAvailable',
            'preventiveMaintenanceNotCarriedOut'
        ]
        
        print(f"\nIMPORTANT NOTE (IF ANY) ANSWERS:")
        print("-" * 30)
        for field in important_note_fields:
            value = request.form.get(field)
            if value:
                print(f"{field}: {value}")
        
        print("="*50)
        print("END OF FORM DATA")
        print("="*50 + "\n")
        
        # Validate required fields
        if not all([sr_no, branch_code, branch_name, date_of_visit, assisted_by_prefix, assisted_by_name]):
            flash('All fields are required', 'error')
            return redirect(url_for('audit_dashboard'))
        
        # Create a new Excel workbook
        print("Creating new Excel workbook...")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Branch Audit Report"
        
        # Create the merged cell layout as specified
        print("Setting up merged cells layout...")
        
        # Merge A1 and B1 for "Sr. No" label
        worksheet.merge_cells('A1:B1')
        worksheet['A1'] = "SR. NO"
        worksheet['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Merge C1, D1, E1, F1 for the serial number value
        worksheet.merge_cells('C1:F1')
        worksheet['C1'] = sr_no
        worksheet['C1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Row 2: BRANCH CODE
        worksheet.merge_cells('A2:B2')
        worksheet['A2'] = "BRANCH CODE"
        worksheet['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        worksheet.merge_cells('C2:F2')
        worksheet['C2'] = branch_code
        worksheet['C2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Row 3: BRANCH NAME
        worksheet.merge_cells('A3:B3')
        worksheet['A3'] = "BRANCH NAME"
        worksheet['A3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        worksheet.merge_cells('C3:F3')
        worksheet['C3'] = branch_name
        worksheet['C3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Row 4: DATE OF VISITS
        worksheet.merge_cells('A4:B4')
        worksheet['A4'] = "DATE OF VISITS"
        worksheet['A4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        worksheet.merge_cells('C4:F4')
        # Format date as DD/MM/YYYY
        if date_of_visit:
            try:
                # Parse the date and format it
                date_obj = datetime.strptime(date_of_visit, '%Y-%m-%d')
                formatted_date = date_obj.strftime('%d/%m/%Y')
                worksheet['C4'] = formatted_date
            except ValueError:
                worksheet['C4'] = date_of_visit  # Use original if parsing fails
        else:
            worksheet['C4'] = date_of_visit
        worksheet['C4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Row 5: ASSISTED BY
        worksheet.merge_cells('A5:B5')
        worksheet['A5'] = "ASSISTED BY"
        worksheet['A5'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        worksheet.merge_cells('C5:F5')
        assisted_by_full = f"{assisted_by_prefix} {assisted_by_name}".strip()
        worksheet['C5'] = assisted_by_full
        worksheet['C5'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Row 6: AUDITED BY
        worksheet.merge_cells('A6:B6')
        worksheet['A6'] = "AUDITED BY"
        worksheet['A6'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        worksheet.merge_cells('C6:F6')
        worksheet['C6'] = "NG TechAssurance Private Limited"
        worksheet['C6'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Row 7: AUDITOR NAME
        worksheet.merge_cells('A7:B7')
        worksheet['A7'] = "AUDITOR NAME"
        worksheet['A7'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        worksheet.merge_cells('C7:F7')
        # Process auditor names
        auditor_names = request.form.getlist('auditorName[]')
        auditor_prefixes = request.form.getlist('auditorNamePrefix[]')
        
        auditor_list = []
        for i, (prefix, name) in enumerate(zip(auditor_prefixes, auditor_names)):
            if prefix and name:  # Only add if both prefix and name are provided
                auditor_list.append(f"{prefix} {name}".strip())
        
        # Join multiple auditors with " & " separator
        auditors_full = " & ".join(auditor_list) if auditor_list else ""
        worksheet['C7'] = auditors_full
        worksheet['C7'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Row 8: Header row for system data
        worksheet['A8'] = "SR. NO."
        worksheet['A8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        worksheet.merge_cells('B8:C8')
        worksheet['B8'] = "BRANCH CHECK POINTS"
        worksheet['B8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        worksheet['D8'] = "RISK FACTOR"
        worksheet['D8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        worksheet['E8'] = "COMPLIANCE/NON-COMPLIANCE"
        worksheet['E8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        worksheet['F8'] = "REMARKS"
        worksheet['F8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Row 9: SYSTEM IP'S header
        worksheet.merge_cells('A9:F9')
        worksheet['A9'] = "SYSTEM IP'S"
        worksheet['A9'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Row 10: Column headers for system data
        worksheet['B10'] = "System IPs"
        worksheet['B10'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        worksheet['C10'] = "MAC Address"
        worksheet['C10'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        worksheet['D10'] = "Employee Name"
        worksheet['D10'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        worksheet['E10'] = "Designation"
        worksheet['E10'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Process ATM IPs
        atm_ips = request.form.getlist('atmIp[]')
        atm_ip_text = "No ATM Facility Available"
        if atm_ips:
            # Filter out empty ATM IPs
            valid_atm_ips = [ip.strip() for ip in atm_ips if ip.strip()]
            if valid_atm_ips:
                atm_ip_text = f"ATM IP: {', '.join(valid_atm_ips)}"
        
        worksheet['F10'] = atm_ip_text
        worksheet['F10'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Apply font styling to F10
        # Create a font for the ATM IP label (bold, Times New Roman, 12)
        from openpyxl.styles import Font as FontClass
        bold_font = FontClass(name='Times New Roman', size=12, bold=True)
        normal_font = FontClass(name='Times New Roman', size=12, bold=False)
        
        # Since we can't have different fonts within the same cell in openpyxl,
        # we'll use bold for the entire cell content
        worksheet['F10'].font = bold_font
        
        # Add system data starting from row 11
        system_ips = request.form.getlist('systemIp[]')
        mac_addresses = request.form.getlist('macAddress[]')
        employee_names = request.form.getlist('employeeName[]')
        designations = request.form.getlist('designation[]')
        
        print(f"\nSYSTEM DATA POPULATION:")
        print("-" * 30)
        print(f"Number of systems: {len(system_ips)}")
        
        # Start from row 11 for the first system
        start_row = 11
        
        # Count valid systems (with all fields filled)
        valid_systems = []
        for i, (ip, mac, emp_name, designation) in enumerate(zip(system_ips, mac_addresses, employee_names, designations)):
            if ip and mac and emp_name and designation:  # Only process if all fields are filled
                valid_systems.append((ip.strip(), mac.strip(), emp_name.strip(), designation.strip()))
        
        num_valid_systems = len(valid_systems)
        print(f"Number of valid systems: {num_valid_systems}")
        
        # Merge column A from A10 to the last system row
        if num_valid_systems > 0:
            last_system_row = start_row + num_valid_systems - 1
            merge_range_a = f'A10:A{last_system_row}'
            worksheet.merge_cells(merge_range_a)
            worksheet['A10'] = ""  # Empty content for merged A column
            worksheet['A10'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            print(f"Merged A column: {merge_range_a}")
        
        # Merge column F from F11 to the last system row
        if num_valid_systems > 0:
            last_system_row = start_row + num_valid_systems - 1
            merge_range_f = f'F11:F{last_system_row}'
            worksheet.merge_cells(merge_range_f)
            worksheet['F11'] = ""  # Empty content for merged F column
            worksheet['F11'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            print(f"Merged F column: {merge_range_f}")
        
        # Populate system data
        for i, (ip, mac, emp_name, designation) in enumerate(valid_systems):
            current_row = start_row + i
            
            # Populate system data in columns B, C, D, E
            worksheet[f'B{current_row}'] = ip  # System IP
            worksheet[f'C{current_row}'] = mac  # MAC Address
            worksheet[f'D{current_row}'] = emp_name  # Employee Name
            worksheet[f'E{current_row}'] = designation  # Designation
            
            # Set alignment for the data cells
            worksheet[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            worksheet[f'C{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            worksheet[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            print(f"System {i+1} - Row {current_row}:")
            print(f"  B{current_row}: {ip}")
            print(f"  C{current_row}: {mac}")
            print(f"  D{current_row}: {emp_name}")
            print(f"  E{current_row}: {designation}")
        
        # Add "CBS ACCESS CONTROL" header below the last system data entry
        if num_valid_systems > 0:
            cbs_header_row = start_row + num_valid_systems
            worksheet.merge_cells(f'A{cbs_header_row}:F{cbs_header_row}')
            worksheet[f'A{cbs_header_row}'] = "CBS ACCESS CONTROL"
            worksheet[f'A{cbs_header_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            print(f"Added CBS ACCESS CONTROL header at row {cbs_header_row}")
            
            # Add CBS Access Control questions and risk factors
            cbs_questions = [
                "Do employees are using strong passwords?",
                "Do you mandate periodical password changes?",
                "Do passwords are shared among employees?",
                "What happens after consecutive failed login attempts?",
                "Is session timeout enforced after a pre-defined period of inactivity?",
                "Whether Multiple Logins is enabled?",
                "Is the CBS System isolated from the internet?",
                "Is internet access restricted to only trusted and officially approved sites?",
                "Is there any policy for default password while unlocking the account?",
                "When a new user joins then how do you give the user ID and Password for the user by using mail or telephonic conversation?",
                "SSL is using or not in CBS Application?",
                "CBS is browser based or desktop based?",
                "Is the CBS compatible with the latest version of the browser?",
                "Whether Two Factor Authentication is implemented for CBS login?"
            ]
            
            cbs_risk_factors = [
                "Medium",
                "Medium", 
                "Medium",
                "High",
                "High",
                "High",
                "High",
                "High",
                "Medium",
                "Low",
                "High",
                "Low",
                "Medium",
                "Medium"
            ]
            
            # Add questions and risk factors starting from the row after CBS header
            start_question_row = cbs_header_row + 1
            
            # Create font styles for CBS questions
            times_new_roman_normal = Font(name='Times New Roman', bold=False, size=12)
            
            # Get form responses for CBS questions
            form_responses = {
                'strongPasswords': request.form.get('strongPasswords'),
                'passwordChanges': request.form.get('passwordChanges'),
                'passwordSharing': request.form.get('passwordSharing'),
                'failedLoginAttempts': request.form.get('failedLoginAttempts'),
                'sessionTimeout': request.form.get('sessionTimeout'),
                'multipleLogins': request.form.get('multipleLogins'),
                'cbsIsolation': request.form.get('cbsIsolation'),
                'internetRestriction': request.form.get('internetRestriction'),
                'defaultPasswordPolicy': request.form.get('defaultPasswordPolicy'),
                'userCredentials': request.form.get('userCredentials'),
                'sslUsage': request.form.get('sslUsage'),
                'cbsType': request.form.get('cbsType'),
                'browserCompatibility': request.form.get('browserCompatibility'),
                'twoFactorAuth': request.form.get('twoFactorAuth')
            }
            
            # Define response mapping for each question
            response_mapping = {
                'strongPasswords': {
                    'Compliance': ('Compliance', 'Global password policy is followed in CBS access.'),
                    'Non-Compliance': ('Non-Compliance', 'Global password policy is not followed in CBS access.')
                },
                'passwordChanges': {
                    'Compliance': ('Compliance', 'Password expires within 30 days.'),
                    'Non-Compliance': ('Non-Compliance', 'Password does not expires after 30 days. It is mandatory to change the password within 30 days.')
                },
                'passwordSharing': {
                    'Compliance': ('Compliance', 'Employees do not share passwords with each other.'),
                    'Non-Compliance': ('Non-Compliance', 'Employees do share passwords with each other.')
                },
                'failedLoginAttempts': {
                    'Compliance': ('Compliance', 'Account gets locked after three consecutive failed login attempts.'),
                    'Non-Compliance': ('Non-Compliance', 'Account does not get locked after three consecutive failed login attempts.')
                },
                'sessionTimeout': {
                    'Compliance': ('Compliance', 'Session timeout is of 10 minutes.'),
                    'Non-Compliance': ('Non-Compliance', 'Session timeout is of 10 minutes.')
                },
                'multipleLogins': {
                    'Compliance': ('Compliance', 'Multiple logins was not enabled.'),
                    'Non-Compliance': ('Non-Compliance', 'Multiple logins was enabled.')
                },
                'cbsIsolation': {
                    'Compliance': ('Compliance', 'CBS System is isolated from the internet.'),
                    'Non-Compliance': ('Non-Compliance', 'CBS System is not isolated from the internet.')
                },
                'internetRestriction': {
                    'Compliance': ('Compliance', 'Internet access restricted to only trusted and officially approved sites.'),
                    'Non-Compliance': ('Non-Compliance', 'Unrestricted internet access is given to the user.')
                },
                'defaultPasswordPolicy': {
                    'Compliance': ('Compliance', 'There is no default password for unlocking the account, User gets a random string or OTP to unlock the account. After successful submission of OTP user gets redirected to set new password.'),
                    'Non-Compliance': ('Non-Compliance', 'Default password is in use for unlocking the account.')
                },
                'userCredentials': {
                    'Compliance': ('Compliance', 'New user password creation is done centrally from Head Office. A well established process is defined.'),
                    'Non-Compliance': ('Non-Compliance', 'New user password creation process was not defined.')
                },
                'sslUsage': {
                    'Compliance': ('Compliance', 'SSL was implemented in CBS application.'),
                    'Non-Compliance': ('Non-Compliance', 'SSL was not implemented in CBS application.')
                },
                'cbsType': {
                    'Browser Based': ('Compliance', 'CBS application was browser based.'),
                    'Desktop Based': ('Compliance', 'CBS application was Desktop based.'),
                    'Non-Compliance': ('Non-Compliance', 'Extension is used to run CBS application.')
                },
                'browserCompatibility': {
                    'Compliance': ('Compliance', 'CBS compatible with the latest version of the browser.'),
                    'Non-Compliance': ('Non-Compliance', 'CBS is not compatible with the latest version of the browser.')
                },
                'twoFactorAuth': {
                    'Biometric Based Compliance': ('Compliance', 'Biometric based two factor authentication is implemented for CBS login.'),
                    'OTP-Based Compliance': ('Compliance', 'OTP based two factor authentication is implemented for CBS login.'),
                    'Biometric and OTP Both': ('Compliance', 'Biometric based and OTP based two factor authentication is implemented for CBS login.'),
                    'Non-Compliance': ('Non-Compliance', 'Two factor authentication is not implemented for CBS login.')
                }
            }
            
            # Field names corresponding to each question
            field_names = [
                'strongPasswords', 'passwordChanges', 'passwordSharing', 'failedLoginAttempts',
                'sessionTimeout', 'multipleLogins', 'cbsIsolation', 'internetRestriction',
                'defaultPasswordPolicy', 'userCredentials', 'sslUsage', 'cbsType',
                'browserCompatibility', 'twoFactorAuth'
            ]
            
            for i, (question, risk_factor) in enumerate(zip(cbs_questions, cbs_risk_factors)):
                current_row = start_question_row + i
                field_name = field_names[i]
                response = form_responses.get(field_name, '')
                
                # Column A: Serial number
                worksheet[f'A{current_row}'] = i + 1
                worksheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'A{current_row}'].font = times_new_roman_normal
                
                # Merge columns B and C for the question
                worksheet.merge_cells(f'B{current_row}:C{current_row}')
                worksheet[f'B{current_row}'] = question
                worksheet[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'B{current_row}'].font = times_new_roman_normal
                
                # Column D: Risk factor
                worksheet[f'D{current_row}'] = risk_factor
                worksheet[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'D{current_row}'].font = times_new_roman_normal
                
                # Columns E and F: Compliance status and remarks based on user response
                compliance_status = "Not Answered"
                remarks = "No response provided for this question."
                
                if field_name in response_mapping:
                    if field_name == 'passwordChanges':
                        # Special handling for password changes - check if value is >= 30
                        try:
                            days = int(response) if response else 0
                            if days <= 30:
                                compliance_status, remarks = response_mapping[field_name]['Compliance']
                            else:
                                compliance_status, remarks = response_mapping[field_name]['Non-Compliance']
                        except (ValueError, TypeError):
                            compliance_status, remarks = "Invalid Input", "Invalid number format for password change days."
                    elif field_name == 'sessionTimeout':
                        # Special handling for session timeout
                        # If value >= 10: Compliance
                        # If value < 10: Non-Compliance
                        try:
                            minutes = int(response) if response else 0
                            if minutes <= 10:
                                # More than or equal to 10 minutes = Compliance
                                compliance_status, remarks = response_mapping[field_name]['Compliance']
                                remarks = f"Session timeout is of {minutes} minutes."
                            else:
                                # Less than 10 minutes = Non-Compliance
                                compliance_status, remarks = response_mapping[field_name]['Non-Compliance']
                                remarks = f"Session timeout is of {minutes} minutes."
                        except (ValueError, TypeError):
                            compliance_status, remarks = "Invalid Input", "Invalid number format for session timeout."
                    elif response in response_mapping[field_name]:
                        compliance_status, remarks = response_mapping[field_name][response]
                
                worksheet[f'E{current_row}'] = compliance_status
                worksheet[f'F{current_row}'] = remarks
                
                # Set alignment for E and F columns
                worksheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'F{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'E{current_row}'].font = times_new_roman_normal
                worksheet[f'F{current_row}'].font = times_new_roman_normal
                
                print(f"Added CBS question {i+1} at row {current_row} - Response: {response}, Status: {worksheet[f'E{current_row}'].value}")
            
            # Update total rows to include CBS questions
            total_rows = start_question_row + len(cbs_questions) - 1
            
            # Add "PHYSICAL AND ENVIRONMENTAL SECURITY" header after CBS questions
            physical_header_row = start_question_row + len(cbs_questions)
            worksheet.merge_cells(f'A{physical_header_row}:F{physical_header_row}')
            worksheet[f'A{physical_header_row}'] = "PHYSICAL AND ENVIRONMENTAL SECURITY"
            worksheet[f'A{physical_header_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet[f'A{physical_header_row}'].font = times_new_roman_normal  # Use normal font, will be overridden by styling loop
            print(f"Added PHYSICAL AND ENVIRONMENTAL SECURITY header at row {physical_header_row}")
            
            # Add Physical and Environmental Security questions and risk factors
            physical_questions = [
                "CCTV Camera Present?",
                "Branch Locker CCTV Camera Present?",
                "CCTV camera covering complete branch area?",
                "Check DVR working properly?",
                "Is NTP configured in the DVR?",
                "DVR is available in bank network or not?",
                "Camera is working in night vision.",
                "CCTV camera history recording duration.",
                "Locker CCTV camera history recording duration.",
                "Biometric Devices or Proximity Card is available or not?",
                "ID card issues to the Employee?",
                "Whether guard is available at branch or not?",
                "Are secure areas controlled?",
                "Smoke Detector is installed in the branch?",
                "Whether Panic Switch installed?",
                "If a panic switch or smoke detector is installed at the branch or head office does the notification reach higher authorities when triggered?",
                "LAN ports are open or not?",
                "Visitor Register is maintained or not?",
                "Any record maintains for drill the environmental controls?",
                "Whether external modems, data cards etc., are being used in the Branch to access internet?",
                "Fire extinguisher is available or not?",
                "Enough fire-extinguisher is available for area?"
            ]
            
            physical_risk_factors = [
                "High",
                "High",
                "High",
                "High",
                "Medium",
                "Medium",
                "High",
                "High",
                "High",
                "Medium",
                "Medium",
                "Medium",
                "High",
                "Medium",
                "Medium",
                "Medium",
                "Medium",
                "Medium",
                "Medium",
                "Medium",
                "Medium",
                "Medium"
            ]
            
            # Get form responses for Physical and Environmental Security questions
            physical_form_responses = {
                'cctvCameraPresent': request.form.get('cctvCameraPresent'),
                'cctvCameraNumber': request.form.get('cctvCameraNumber'),
                'branchLockerCctv': request.form.get('branchLockerCctv'),
                'branchLockerCctvNumber': request.form.get('branchLockerCctvNumber'),
                'cctvCoveringCompleteArea': request.form.get('cctvCoveringCompleteArea'),
                'dvrWorkingProperly': request.form.get('dvrWorkingProperly'),
                'ntpConfiguredDvr': request.form.get('ntpConfiguredDvr'),
                'dvrInBankNetwork': request.form.get('dvrInBankNetwork'),
                'cameraNightVision': request.form.get('cameraNightVision'),
                'cctvHistoryDuration': request.form.get('cctvHistoryDuration'),
                'lockerCctvHistoryDuration': request.form.get('lockerCctvHistoryDuration'),
                'biometricDevices': request.form.get('biometricDevices'),
                'idCardIssued': request.form.get('idCardIssued'),
                'guardAvailable': request.form.get('guardAvailable'),
                'secureAreasControlled': request.form.get('secureAreasControlled'),
                'smokeDetectorInstalled': request.form.get('smokeDetectorInstalled'),
                'panicSwitchInstalled': request.form.get('panicSwitchInstalled'),
                'notificationReachAuthorities': request.form.get('notificationReachAuthorities'),
                'lanPortsOpen': request.form.get('lanPortsOpen'),
                'visitorRegisterMaintained': request.form.get('visitorRegisterMaintained'),
                'environmentalControlsDrill': request.form.get('environmentalControlsDrill'),
                'externalModemsUsed': request.form.get('externalModemsUsed'),
                'fireExtinguisherAvailable': request.form.get('fireExtinguisherAvailable'),
                'enoughFireExtinguisher': request.form.get('enoughFireExtinguisher')
            }
            
            # Define response mapping for Physical and Environmental Security questions
            physical_response_mapping = {
                'cctvCameraPresent': {
                    'Compliance': ('Compliance', 'CCTV cameras are present and functional.'),
                    'Non-Compliance': ('Non-Compliance', 'CCTV cameras are not present or not functional.'),
                    'Unable to Check': ('Non-Compliance', 'Auditor was unable to Check.')
                },
                'branchLockerCctv': {
                    'Compliance': ('Compliance', 'Branch locker CCTV cameras are present and functional.'),
                    'Non-Compliance': ('Non-Compliance', 'Branch locker CCTV cameras are not present or not functional.'),
                    'Unable to Check': ('Non-Compliance', 'Auditor was unable to Check.')
                },
                'cctvCoveringCompleteArea': {
                    'Compliance': ('Compliance', 'CCTV cameras are covering the complete branch area.'),
                    'Non-Compliance': ('Non-Compliance', 'CCTV cameras are not covering the complete branch area.'),
                    'Unable to Check': ('Non-Compliance', 'Auditor was unable to Check.')
                },
                'dvrWorkingProperly': {
                    'Compliance': ('Compliance', 'DVR is working properly and recording footage.'),
                    'Non-Compliance': ('Non-Compliance', 'DVR is not working properly or not recording footage.')
                },
                'ntpConfiguredDvr': {
                    'Compliance': ('Compliance', 'NTP is configured in the DVR for accurate time synchronization.'),
                    'Non-Compliance': ('Non-Compliance', 'NTP is not configured in the DVR.'),
                    'Unable to Check': ('Non-Compliance', 'Auditor was unable to Check.')
                },
                'dvrInBankNetwork': {
                    'Compliance': ('Compliance', 'DVR is available in the bank network.'),
                    'Non-Compliance': ('Non-Compliance', 'DVR is not available in the bank network.')
                },
                'cameraNightVision': {
                    'Compliance': ('Compliance', 'Cameras are working in night vision mode.'),
                    'Non-Compliance': ('Non-Compliance', 'Cameras are not working in night vision mode.'),
                    'Unable to Check': ('Non-Compliance', 'Auditor was unable to Check.')
                },
                'cctvHistoryDuration': {
                    'Compliance': ('Compliance', 'CCTV camera history recording duration is adequate.'),
                    'Non-Compliance': ('Non-Compliance', 'CCTV camera history recording duration is not adequate.')
                },
                'lockerCctvHistoryDuration': {
                    'Compliance': ('Compliance', 'Locker CCTV camera history recording duration is adequate.'),
                    'Non-Compliance': ('Non-Compliance', 'Locker CCTV camera history recording duration is not adequate.')
                },
                'biometricDevices': {
                    'Compliance': ('Compliance', 'Biometric devices or proximity cards are available.'),
                    'Non-Compliance': ('Non-Compliance', 'Biometric devices or proximity cards are not available.')
                },
                'idCardIssued': {
                    'Compliance': ('Compliance', 'ID cards are issued to employees.'),
                    'Non-Compliance': ('Non-Compliance', 'ID cards are not issued to employees.')
                },
                'guardAvailable': {
                    'Compliance': ('Compliance', 'Guard is available at the branch.'),
                    'Non-Compliance': ('Non-Compliance', 'Guard is not available at the branch.')
                },
                'secureAreasControlled': {
                    'Compliance': ('Compliance', 'Secure areas are properly controlled.'),
                    'Non-Compliance': ('Non-Compliance', 'Secure areas are not properly controlled.')
                },
                'smokeDetectorInstalled': {
                    'Compliance': ('Compliance', 'Smoke detector is installed in the branch.'),
                    'Non-Compliance': ('Non-Compliance', 'Smoke detector is not installed in the branch.')
                },
                'panicSwitchInstalled': {
                    'Compliance': ('Compliance', 'Panic switch is installed.'),
                    'Non-Compliance': ('Non-Compliance', 'Panic switch is not installed.')
                },
                'notificationReachAuthorities': {
                    'Panic Switch': ('Compliance', 'Panic Switch installed and Notification reaches higher authorities when panic switch is triggered.'),
                    'Smoke Detector': ('Compliance', 'Smoke Detector installed and Notification reaches higher authorities when smoke detector is triggered.'),
                    'Panic and Smoke': ('Compliance', 'Panic Switch and Smoke Detector installed and Notification reaches higher authorities when panic switch or smoke detector is triggered.'),
                    'Non Compliance': ('Non-Compliance', 'Panic Switch and Smoke Detector not installed or not work properly.')
                },
                'lanPortsOpen': {
                    'Compliance': ('Compliance', 'LAN ports are properly secured.'),
                    'Non-Compliance': ('Non-Compliance', 'LAN ports are open and not secured.')
                },
                'visitorRegisterMaintained': {
                    'Compliance': ('Compliance', 'Visitor register is maintained properly.'),
                    'Non-Compliance': ('Non-Compliance', 'Visitor register is not maintained.')
                },
                'environmentalControlsDrill': {
                    'Compliance': ('Compliance', 'Records are maintained for environmental control drills.'),
                    'Non-Compliance': ('Non-Compliance', 'Records are not maintained for environmental control drills.')
                },
                'externalModemsUsed': {
                    'Compliance': ('Compliance', 'External modems and data cards are not being used in the branch.'),
                    'Non-Compliance': ('Non-Compliance', 'External modems and data cards are being used in the branch.')
                },
                'fireExtinguisherAvailable': {
                    'Compliance': ('Compliance', 'Fire extinguisher is available.'),
                    'Non-Compliance': ('Non-Compliance', 'Fire extinguisher is not available.')
                },
                'enoughFireExtinguisher': {
                    'Compliance': ('Compliance', 'Enough fire extinguishers are available for the area.'),
                    'Non-Compliance': ('Non-Compliance', 'Enough fire extinguishers are not available for the area.')
                }
            }
            
            # Field names corresponding to each Physical and Environmental Security question
            physical_field_names = [
                'cctvCameraPresent', 'branchLockerCctv', 'cctvCoveringCompleteArea', 'dvrWorkingProperly',
                'ntpConfiguredDvr', 'dvrInBankNetwork', 'cameraNightVision', 'cctvHistoryDuration',
                'lockerCctvHistoryDuration', 'biometricDevices', 'idCardIssued', 'guardAvailable',
                'secureAreasControlled', 'smokeDetectorInstalled', 'panicSwitchInstalled', 'notificationReachAuthorities',
                'lanPortsOpen', 'visitorRegisterMaintained', 'environmentalControlsDrill', 'externalModemsUsed',
                'fireExtinguisherAvailable', 'enoughFireExtinguisher'
            ]
            
            # Add Physical and Environmental Security questions starting from the row after header
            start_physical_question_row = physical_header_row + 1
            
            for i, (question, risk_factor) in enumerate(zip(physical_questions, physical_risk_factors)):
                current_row = start_physical_question_row + i
                field_name = physical_field_names[i]
                response = physical_form_responses.get(field_name, '')
                
                # Column A: Serial number
                worksheet[f'A{current_row}'] = i + 1
                worksheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'A{current_row}'].font = times_new_roman_normal
                
                # Merge columns B and C for the question
                worksheet.merge_cells(f'B{current_row}:C{current_row}')
                worksheet[f'B{current_row}'] = question
                worksheet[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'B{current_row}'].font = times_new_roman_normal
                
                # Column D: Risk factor
                worksheet[f'D{current_row}'] = risk_factor
                worksheet[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'D{current_row}'].font = times_new_roman_normal
                
                # Columns E and F: Compliance status and remarks based on user response
                compliance_status = "Not Answered"
                remarks = "No response provided for this question."
                
                if field_name in physical_response_mapping and response in physical_response_mapping[field_name]:
                    compliance_status, remarks = physical_response_mapping[field_name][response]
                    
                    # Special handling for CCTV camera numbers
                    if field_name == 'cctvCameraPresent' and response == 'Compliance':
                        camera_number = physical_form_responses.get('cctvCameraNumber', '')
                        if camera_number:
                            remarks = f"Total {camera_number} CCTV cameras are present and working properly."
                    elif field_name == 'branchLockerCctv' and response == 'Compliance':
                        locker_camera_number = physical_form_responses.get('branchLockerCctvNumber', '')
                        if locker_camera_number:
                            remarks = f"Total {locker_camera_number} CCTV cameras are present and working properly for Branch Locker."
                
                # Special handling for CCTV history duration fields
                if field_name == 'cctvHistoryDuration':
                    cctv_history_mode = request.form.get('cctvHistoryDurationMode', 'number')
                    if cctv_history_mode == 'unable':
                        compliance_status = 'Non-Compliance'
                        remarks = "Auditor was unable to Check."
                    else:
                        try:
                            days = int(response) if response else 0
                            if days >= 30:
                                compliance_status = 'Compliance'
                                remarks = f"Total {days} days recording was available with the bank."
                            else:
                                compliance_status = 'Non-Compliance'
                                remarks = f"Total {days} days recording was available with the bank. Minimum 30 days required."
                        except (ValueError, TypeError):
                            compliance_status = "Invalid Input"
                            remarks = "Invalid number format for CCTV history duration."
                elif field_name == 'lockerCctvHistoryDuration':
                    locker_history_mode = request.form.get('lockerCctvHistoryDurationMode', 'number')
                    if locker_history_mode == 'unable':
                        compliance_status = 'Non-Compliance'
                        remarks = "Auditor was unable to Check."
                    else:
                        try:
                            days = int(response) if response else 0
                            if days >= 180:
                                compliance_status = 'Compliance'
                                remarks = f"Total {days} days recording was available with the bank for Branch Locker."
                            else:
                                compliance_status = 'Non-Compliance'
                                remarks = f"Total {days} days recording was available with the bank for Branch Locker. Minimum 180 days required."
                        except (ValueError, TypeError):
                            compliance_status = "Invalid Input"
                            remarks = "Invalid number format for Locker CCTV history duration."
                
                worksheet[f'E{current_row}'] = compliance_status
                worksheet[f'F{current_row}'] = remarks
                
                # Set alignment for E and F columns
                worksheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'F{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'E{current_row}'].font = times_new_roman_normal
                worksheet[f'F{current_row}'].font = times_new_roman_normal
                
                print(f"Added Physical question {i+1} at row {current_row} - Response: {response}, Status: {compliance_status}")
            
            # Update total rows to include Physical and Environmental Security questions
            total_rows = start_physical_question_row + len(physical_questions) - 1
            
            # Add "POWER BACK UP" header after Physical and Environmental Security questions
            power_header_row = start_physical_question_row + len(physical_questions)
            worksheet.merge_cells(f'A{power_header_row}:F{power_header_row}')
            worksheet[f'A{power_header_row}'] = "POWER BACK UP"
            worksheet[f'A{power_header_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet[f'A{power_header_row}'].font = times_new_roman_normal  # Use normal font, will be overridden by styling loop
            print(f"Added POWER BACK UP header at row {power_header_row}")
            
            # Add Power Back Up questions and risk factors
            power_questions = [
                "Does the Bank have made enough Battery Backup which supports Computer Systems?",
                "Do the AMC is given for power backup systems?",
                "Generator is available or not?"
            ]
            
            power_risk_factors = [
                "High",
                "Medium",
                "High"
            ]
            
            # Get form responses for Power Back Up questions
            power_form_responses = {
                'batteryBackupSupport': request.form.get('batteryBackupSupport'),
                'amcPowerBackup': request.form.get('amcPowerBackup'),
                'generatorAvailable': request.form.get('generatorAvailable')
            }
            
            # Add Power Back Up questions starting from the row after header
            start_power_question_row = power_header_row + 1
            
            for i, (question, risk_factor) in enumerate(zip(power_questions, power_risk_factors)):
                current_row = start_power_question_row + i
                field_name = ['batteryBackupSupport', 'amcPowerBackup', 'generatorAvailable'][i]
                response = power_form_responses.get(field_name, '')
                
                # Column A: Serial number
                worksheet[f'A{current_row}'] = i + 1
                worksheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'A{current_row}'].font = times_new_roman_normal
                
                # Merge columns B and C for the question
                worksheet.merge_cells(f'B{current_row}:C{current_row}')
                worksheet[f'B{current_row}'] = question
                worksheet[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'B{current_row}'].font = times_new_roman_normal
                
                # Column D: Risk factor
                worksheet[f'D{current_row}'] = risk_factor
                worksheet[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'D{current_row}'].font = times_new_roman_normal
                
                # Columns E and F: Compliance status and remarks based on user response
                compliance_status = "Not Answered"
                remarks = "No response provided for this question."
                
                if field_name == 'batteryBackupSupport':
                    try:
                        hours = int(response) if response else 0
                        if hours >= 8:
                            compliance_status = 'Compliance'
                            remarks = f"UPS is present for providing power supply up to {hours} HRS."
                        else:
                            compliance_status = 'Non-Compliance'
                            remarks = f"UPS is present for providing power supply up to {hours} HRS."
                    except (ValueError, TypeError):
                        compliance_status = "Invalid Input"
                        remarks = "Invalid number format for battery backup hours."
                elif field_name == 'amcPowerBackup':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'An Annual Maintenance Contract (AMC) is maintained for the power backup systems to ensure their reliability and smooth operation.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'An Annual Maintenance Contract (AMC) is not maintained for the power backup systems.'
                elif field_name == 'generatorAvailable':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Generator is available.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Generator is not available.'
                    elif response == 'Not Applicable':
                        compliance_status = 'Not Applicable'
                        remarks = '-'
                
                # Special logic: If generator is available, override battery backup response
                if field_name == 'batteryBackupSupport' and power_form_responses.get('generatorAvailable') == 'Compliance':
                    compliance_status = 'Compliance'
                    remarks = 'Generator is available for power backup.'
                
                worksheet[f'E{current_row}'] = compliance_status
                worksheet[f'F{current_row}'] = remarks
                
                # Set alignment for E and F columns
                worksheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'F{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'E{current_row}'].font = times_new_roman_normal
                worksheet[f'F{current_row}'].font = times_new_roman_normal
                
                print(f"Added Power question {i+1} at row {current_row} - Response: {response}, Status: {compliance_status}")
            
            # Update total rows to include Power Back Up questions
            total_rows = start_power_question_row + len(power_questions) - 1
            
            # Add "USER AWARENESS" header after Power Back Up questions
            user_awareness_header_row = start_power_question_row + len(power_questions)
            worksheet.merge_cells(f'A{user_awareness_header_row}:F{user_awareness_header_row}')
            worksheet[f'A{user_awareness_header_row}'] = "USER AWARENESS"
            worksheet[f'A{user_awareness_header_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet[f'A{user_awareness_header_row}'].font = times_new_roman_normal  # Use normal font, will be overridden by styling loop
            print(f"Added USER AWARENESS header at row {user_awareness_header_row}")
            
            # Add User Awareness questions and risk factors
            user_awareness_questions = [
                "Is training conducted for end user with respect to the Information Security?",
                "Is training conducted for end user with respect to the application usage?",
                "Employees are aware about using of fire extinguisher & Panic Switch?",
                "User was aware about checking CCTV recording history?"
            ]
            
            user_awareness_risk_factors = [
                "Medium",
                "Medium",
                "Low",
                "Medium"
            ]
            
            # Get form responses for User Awareness questions
            user_awareness_form_responses = {
                'infoSecurityTraining': request.form.get('infoSecurityTraining'),
                'applicationUsageTraining': request.form.get('applicationUsageTraining'),
                'fireExtinguisherPanicSwitchAwareness': request.form.get('fireExtinguisherPanicSwitchAwareness'),
                'cctvRecordingHistoryAwareness': request.form.get('cctvRecordingHistoryAwareness')
            }
            
            # Add User Awareness questions starting from the row after header
            start_user_awareness_question_row = user_awareness_header_row + 1
            
            for i, (question, risk_factor) in enumerate(zip(user_awareness_questions, user_awareness_risk_factors)):
                current_row = start_user_awareness_question_row + i
                field_name = ['infoSecurityTraining', 'applicationUsageTraining', 'fireExtinguisherPanicSwitchAwareness', 'cctvRecordingHistoryAwareness'][i]
                response = user_awareness_form_responses.get(field_name, '')
                
                # Column A: Serial number
                worksheet[f'A{current_row}'] = i + 1
                worksheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'A{current_row}'].font = times_new_roman_normal
                
                # Merge columns B and C for the question
                worksheet.merge_cells(f'B{current_row}:C{current_row}')
                worksheet[f'B{current_row}'] = question
                worksheet[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'B{current_row}'].font = times_new_roman_normal
                
                # Column D: Risk factor
                worksheet[f'D{current_row}'] = risk_factor
                worksheet[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'D{current_row}'].font = times_new_roman_normal
                
                # Columns E and F: Compliance status and remarks based on user response
                compliance_status = "Not Answered"
                remarks = "No response provided for this question."
                
                if field_name == 'infoSecurityTraining':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'General guidelines related to cyber security such as using strong password, no sharing of password etc. are provided through email and in person training on periodical basis.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Training was not conducted for end user with respect to information security.'
                elif field_name == 'applicationUsageTraining':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Training was condcuted for end user regarding usage of application.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Training was not conducted for end user with respect to the application usage.'
                elif field_name == 'fireExtinguisherPanicSwitchAwareness':
                    if response == 'Fire Extinguishers':
                        compliance_status = 'Compliance'
                        remarks = 'Employees were aware of using fire extinguishers.'
                    elif response == 'Panic Switch':
                        compliance_status = 'Compliance'
                        remarks = 'Employees were aware of using panic switch.'
                    elif response == 'Both':
                        compliance_status = 'Compliance'
                        remarks = 'Employees were aware of using fire extinguishers and panic switch.'
                    elif response == 'Non Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Employees were not aware of using fire extinguishers and panic switch.'
                elif field_name == 'cctvRecordingHistoryAwareness':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Employees were aware about checking CCTV recording history.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Employees were not aware about checking CCTV recording history.'
                
                worksheet[f'E{current_row}'] = compliance_status
                worksheet[f'F{current_row}'] = remarks
                
                # Set alignment for E and F columns
                worksheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'F{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'E{current_row}'].font = times_new_roman_normal
                worksheet[f'F{current_row}'].font = times_new_roman_normal
                
                print(f"Added User Awareness question {i+1} at row {current_row} - Response: {response}, Status: {compliance_status}")
            
            # Update total rows to include User Awareness questions
            total_rows = start_user_awareness_question_row + len(user_awareness_questions) - 1
            
            # Add "MAINTENANCE AND BUSINESS CONTINUITY CONTROLS" header after User Awareness questions
            maintenance_header_row = start_user_awareness_question_row + len(user_awareness_questions)
            worksheet.merge_cells(f'A{maintenance_header_row}:F{maintenance_header_row}')
            worksheet[f'A{maintenance_header_row}'] = "MAINTENANCE AND BUSINESS CONTINUITY CONTROLS"
            worksheet[f'A{maintenance_header_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet[f'A{maintenance_header_row}'].font = times_new_roman_normal  # Use normal font, will be overridden by styling loop
            print(f"Added MAINTENANCE AND BUSINESS CONTINUITY CONTROLS header at row {maintenance_header_row}")
            
            # Add Maintenance and Business Continuity Controls questions and risk factors
            maintenance_questions = [
                "Is AMC is given for computer and related systems?",
                "Is Complaint register maintained?"
            ]
            
            maintenance_risk_factors = [
                "Medium",
                "Low"
            ]
            
            # Get form responses for Maintenance and Business Continuity Controls questions
            maintenance_form_responses = {
                'amcComputerSystems': request.form.get('amcComputerSystems'),
                'complaintRegisterMaintained': request.form.get('complaintRegisterMaintained')
            }
            
            # Add Maintenance and Business Continuity Controls questions starting from the row after header
            start_maintenance_question_row = maintenance_header_row + 1
            
            for i, (question, risk_factor) in enumerate(zip(maintenance_questions, maintenance_risk_factors)):
                current_row = start_maintenance_question_row + i
                field_name = ['amcComputerSystems', 'complaintRegisterMaintained'][i]
                response = maintenance_form_responses.get(field_name, '')
                
                # Column A: Serial number
                worksheet[f'A{current_row}'] = i + 1
                worksheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'A{current_row}'].font = times_new_roman_normal
                
                # Merge columns B and C for the question
                worksheet.merge_cells(f'B{current_row}:C{current_row}')
                worksheet[f'B{current_row}'] = question
                worksheet[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'B{current_row}'].font = times_new_roman_normal
                
                # Column D: Risk factor
                worksheet[f'D{current_row}'] = risk_factor
                worksheet[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'D{current_row}'].font = times_new_roman_normal
                
                # Columns E and F: Compliance status and remarks based on user response
                compliance_status = "Not Answered"
                remarks = "No response provided for this question."
                
                if field_name == 'amcComputerSystems':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'AMC is managed for computers and related systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'AMC was not managed for the Computer and related System.'
                elif field_name == 'complaintRegisterMaintained':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Complaint register is maintained in the branch.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Complaint register was not well maintained in the branch.'
                
                worksheet[f'E{current_row}'] = compliance_status
                worksheet[f'F{current_row}'] = remarks
                
                # Set alignment for E and F columns
                worksheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'F{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'E{current_row}'].font = times_new_roman_normal
                worksheet[f'F{current_row}'].font = times_new_roman_normal
                
                print(f"Added Maintenance question {i+1} at row {current_row} - Response: {response}, Status: {compliance_status}")
            
            # Update total rows to include Maintenance and Business Continuity Controls questions
            total_rows = start_maintenance_question_row + len(maintenance_questions) - 1
            
            # Add "PATCH MANAGEMENT" header after Maintenance and Business Continuity Controls questions
            patch_header_row = start_maintenance_question_row + len(maintenance_questions)
            worksheet.merge_cells(f'A{patch_header_row}:F{patch_header_row}')
            worksheet[f'A{patch_header_row}'] = "PATCH MANAGEMENT"
            worksheet[f'A{patch_header_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet[f'A{patch_header_row}'].font = times_new_roman_normal  # Use normal font, will be overridden by styling loop
            print(f"Added PATCH MANAGEMENT header at row {patch_header_row}")
            
            # Add Patch Management questions and risk factors
            patch_questions = [
                "Antivirus available in all the desktop?",
                "Antivirus updated in all the desktop?",
                "Windows patch updated in all the system?",
                "Are there any system on which outdated version of windows is installed?",
                "Expired License antivirus installed?"
            ]
            
            patch_risk_factors = [
                "Critical",
                "Critical",
                "Medium",
                "Critical",
                "High"
            ]
            
            # Get form responses for Patch Management questions
            patch_form_responses = {
                'antivirusAvailable': request.form.get('antivirusAvailable'),
                'antivirusAvailableIps': request.form.get('antivirusAvailableIps'),
                'antivirusUpdated': request.form.get('antivirusUpdated'),
                'antivirusUpdatedIps': request.form.get('antivirusUpdatedIps'),
                'windowsPatchUpdated': request.form.get('windowsPatchUpdated'),
                'windowsPatchIps': request.form.get('windowsPatchIps'),
                'outdatedWindowsVersion': request.form.get('outdatedWindowsVersion'),
                'outdatedWindowsIps': request.form.get('outdatedWindowsIps'),
                'expiredLicenseAntivirus': request.form.get('expiredLicenseAntivirus'),
                'expiredLicenseIps': request.form.get('expiredLicenseIps')
            }
            
            # Add Patch Management questions starting from the row after header
            start_patch_question_row = patch_header_row + 1
            
            for i, (question, risk_factor) in enumerate(zip(patch_questions, patch_risk_factors)):
                current_row = start_patch_question_row + i
                field_name = ['antivirusAvailable', 'antivirusUpdated', 'windowsPatchUpdated', 'outdatedWindowsVersion', 'expiredLicenseAntivirus'][i]
                response = patch_form_responses.get(field_name, '')
                
                # Column A: Serial number
                worksheet[f'A{current_row}'] = i + 1
                worksheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'A{current_row}'].font = times_new_roman_normal
                
                # Merge columns B and C for the question
                worksheet.merge_cells(f'B{current_row}:C{current_row}')
                worksheet[f'B{current_row}'] = question
                worksheet[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'B{current_row}'].font = times_new_roman_normal
                
                # Column D: Risk factor
                worksheet[f'D{current_row}'] = risk_factor
                worksheet[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'D{current_row}'].font = times_new_roman_normal
                
                # Columns E and F: Compliance status and remarks based on user response
                compliance_status = "Not Answered"
                remarks = "No response provided for this question."
                
                if field_name == 'antivirusAvailable':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Anti-virus was available in all of the systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        ips = patch_form_responses.get('antivirusAvailableIps', '')
                        if ips:
                            # Split IPs by newline and format them
                            ip_list = [ip.strip() for ip in ips.split('\n') if ip.strip()]
                            formatted_ips = '\n'.join(ip_list)
                            remarks = f'Anti-virus was not available in the below mentioned system. IPs:\n{formatted_ips}'
                        else:
                            remarks = 'Anti-virus was not available in the below mentioned system.'
                elif field_name == 'antivirusUpdated':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Anti-virus was updated in all of the systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        ips = patch_form_responses.get('antivirusUpdatedIps', '')
                        if ips:
                            # Split IPs by newline and format them
                            ip_list = [ip.strip() for ip in ips.split('\n') if ip.strip()]
                            formatted_ips = '\n'.join(ip_list)
                            remarks = f'Anti-virus was not updated in the below mentioned systems. IPs:\n{formatted_ips}'
                        else:
                            remarks = 'Anti-virus was not updated in the below mentioned systems.'
                    elif response == 'Do not Antivirus':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Anti-virus was not available in all of the systems.'
                elif field_name == 'windowsPatchUpdated':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Windows patch was updated in all of the systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        ips = patch_form_responses.get('windowsPatchIps', '')
                        if ips:
                            # Split IPs by newline and format them
                            ip_list = [ip.strip() for ip in ips.split('\n') if ip.strip()]
                            formatted_ips = '\n'.join(ip_list)
                            remarks = f'Windows patch was not updated in the below mentioned systems. IPs:\n{formatted_ips}'
                        else:
                            remarks = 'Windows patch was not updated in the below mentioned systems.'
                elif field_name == 'outdatedWindowsVersion':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Latest version of windows was installed in all of the systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        ips = patch_form_responses.get('outdatedWindowsIps', '')
                        if ips:
                            # Split IPs by newline and format them
                            ip_list = [ip.strip() for ip in ips.split('\n') if ip.strip()]
                            formatted_ips = '\n'.join(ip_list)
                            remarks = f'Outdated windows version was installed in the below mentioned system. IP:\n{formatted_ips}'
                        else:
                            remarks = 'Outdated windows version was installed in the below mentioned system.'
                elif field_name == 'expiredLicenseAntivirus':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Licensed antivirus was installed in all of the systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        ips = patch_form_responses.get('expiredLicenseIps', '')
                        if ips:
                            # Split IPs by newline and format them
                            ip_list = [ip.strip() for ip in ips.split('\n') if ip.strip()]
                            formatted_ips = '\n'.join(ip_list)
                            remarks = f'Expired license of anti-virus was installed in the below mentioned system. IP:\n{formatted_ips}'
                        else:
                            remarks = 'Expired license of anti-virus was installed in the below mentioned system.'
                
                worksheet[f'E{current_row}'] = compliance_status
                worksheet[f'F{current_row}'] = remarks
                
                # Set alignment for E and F columns
                worksheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'F{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'E{current_row}'].font = times_new_roman_normal
                worksheet[f'F{current_row}'].font = times_new_roman_normal
                
                print(f"Added Patch Management question {i+1} at row {current_row} - Response: {response}, Status: {compliance_status}")
            
            # Update total rows to include Patch Management questions
            total_rows = start_patch_question_row + len(patch_questions) - 1
            
            # Add "NETWORK SECURITY" header after Patch Management questions
            network_header_row = start_patch_question_row + len(patch_questions)
            worksheet.merge_cells(f'A{network_header_row}:F{network_header_row}')
            worksheet[f'A{network_header_row}'] = "NETWORK SECURITY"
            worksheet[f'A{network_header_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet[f'A{network_header_row}'].font = times_new_roman_normal  # Use normal font, will be overridden by styling loop
            print(f"Added NETWORK SECURITY header at row {network_header_row}")
            
            # Add Network Security questions and risk factors
            network_questions = [
                "Are controls in place to ensure users only have access to the network resources they have been specially authorized to use and are required for their duties?",
                "Is there network diagram available?",
                "Is structured cabling is observed?",
                "Is cable tagging is observed?",
                "Is provision made for backup network?",
                "Networking devices are under secure location and there is any cabinet facility is provided or not?",
                "Proper Cooling for Networking Devices available or not?",
                "Wi-Fi is available or not?",
                "Whether Firewall is available in Branch ?"
            ]
            
            network_risk_factors = [
                "High",
                "Low",
                "Low",
                "Medium",
                "Critical",
                "High",
                "High",
                "High",
                "High"
            ]
            
            # Get form responses for Network Security questions
            network_form_responses = {
                'networkAccessControls': request.form.get('networkAccessControls'),
                'networkDiagramAvailable': request.form.get('networkDiagramAvailable'),
                'structuredCabling': request.form.get('structuredCabling'),
                'cableTagging': request.form.get('cableTagging'),
                'backupNetwork': request.form.get('backupNetwork'),
                'networkingDevicesSecurity': request.form.get('networkingDevicesSecurity'),
                'networkingDevicesCooling': request.form.get('networkingDevicesCooling'),
                'wifiAvailable': request.form.get('wifiAvailable'),
                'firewallAvailable': request.form.get('firewallAvailable')
            }
            
            # Add Network Security questions starting from the row after header
            start_network_question_row = network_header_row + 1
            
            for i, (question, risk_factor) in enumerate(zip(network_questions, network_risk_factors)):
                current_row = start_network_question_row + i
                field_name = ['networkAccessControls', 'networkDiagramAvailable', 'structuredCabling', 'cableTagging', 'backupNetwork', 'networkingDevicesSecurity', 'networkingDevicesCooling', 'wifiAvailable', 'firewallAvailable'][i]
                response = network_form_responses.get(field_name, '')
                
                # Column A: Serial number
                worksheet[f'A{current_row}'] = i + 1
                worksheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'A{current_row}'].font = times_new_roman_normal
                
                # Merge columns B and C for the question
                worksheet.merge_cells(f'B{current_row}:C{current_row}')
                worksheet[f'B{current_row}'] = question
                worksheet[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'B{current_row}'].font = times_new_roman_normal
                
                # Column D: Risk factor
                worksheet[f'D{current_row}'] = risk_factor
                worksheet[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'D{current_row}'].font = times_new_roman_normal
                
                # Columns E and F: Compliance status and remarks based on user response
                compliance_status = "Not Answered"
                remarks = "No response provided for this question."
                
                if field_name == 'networkAccessControls':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Network devices were arranged in a secure location.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Network devices were not arranged in a secure location.'
                elif field_name == 'networkDiagramAvailable':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Network diagram was available at the branch.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Network diagram was not available at the branch.'
                elif field_name == 'structuredCabling':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Structured cabling was present in the branch.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Structured cabling was not present in the branch.'
                elif field_name == 'cableTagging':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Cable tagging was available in the branch.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Cable tagging was not available in the branch.'
                elif field_name == 'backupNetwork':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Secondary connectivity was available in the bank.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Branch is currently working on single connectivity.'
                elif field_name == 'networkingDevicesSecurity':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Network devices are placed under secured location, inside a network rack and it was locked properly'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Cabinet facility was there and networking devices were not placed inside the cabinet, however it was not locked.'
                elif field_name == 'networkingDevicesCooling':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Proper cooling for networking devices was available.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Proper cooling for networking devices was not available.'
                elif field_name == 'wifiAvailable':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Wi-Fi was not available in the branch.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Wi-Fi was available in the branch.'
                elif field_name == 'firewallAvailable':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Firewall is available in Branch.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Firewall is not available in Branch.'
                
                worksheet[f'E{current_row}'] = compliance_status
                worksheet[f'F{current_row}'] = remarks
                
                # Set alignment for E and F columns
                worksheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'F{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'E{current_row}'].font = times_new_roman_normal
                worksheet[f'F{current_row}'].font = times_new_roman_normal
                
                print(f"Added Network Security question {i+1} at row {current_row} - Response: {response}, Status: {compliance_status}")
            
            # Update total rows to include Network Security questions
            total_rows = start_network_question_row + len(network_questions) - 1
            
            # Add "ENDPOINTS VULNERABILITY" header after Network Security questions
            endpoints_header_row = start_network_question_row + len(network_questions)
            worksheet.merge_cells(f'A{endpoints_header_row}:F{endpoints_header_row}')
            worksheet[f'A{endpoints_header_row}'] = "ENDPOINTS VULNERABILITY"
            worksheet[f'A{endpoints_header_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet[f'A{endpoints_header_row}'].font = times_new_roman_normal  # Use normal font, will be overridden by styling loop
            print(f"Added ENDPOINTS VULNERABILITY header at row {endpoints_header_row}")
            
            # Add Endpoints Vulnerability questions and risk factors
            endpoints_questions = [
                "Does User login with Administrator Rights?",
                "Can Group Policy be modified?",
                "Can Security Configuration Policy be modified?",
                "Whether Proxy can be modified by branch users?",
                "Whether internet is allowed or not?",
                "Whether internet access is restricted or unrestricted?",
                "Authentication Policy for internet is available or not?",
                "System Password is weak or not?",
                "System tagging is available or not?",
                "Is NTP configured in the systems?",
                "Whether USB ports/CD drives enabled in the System?",
                "USB Approval process?",
                "If bank have trusted USB, then check the USB open outside the network?",
                "Whether Firewall is enable in systems?",
                "Whether all external storage media are checked for virus?",
                "Whether Remote Desktop Protocol is enabled or disabled?"
            ]
            
            endpoints_risk_factors = [
                "High",
                "High",
                "High",
                "Medium",
                "High",
                "High",
                "High",
                "High",
                "Medium",
                "Medium",
                "High",
                "High",
                "High",
                "High",
                "High",
                "High"
            ]
            
            # Get form responses for Endpoints Vulnerability questions
            endpoints_form_responses = {
                'adminRightsLogin': request.form.get('adminRightsLogin'),
                'adminRightsIps': request.form.get('adminRightsIps'),
                'groupPolicyModification': request.form.get('groupPolicyModification'),
                'groupPolicyIps': request.form.get('groupPolicyIps'),
                'securityConfigModification': request.form.get('securityConfigModification'),
                'securityConfigIps': request.form.get('securityConfigIps'),
                'proxyModification': request.form.get('proxyModification'),
                'proxyIps': request.form.get('proxyIps'),
                'internetAllowed': request.form.get('internetAllowed'),
                'internetAccessRestricted': request.form.get('internetAccessRestricted'),
                'internetAuthPolicy': request.form.get('internetAuthPolicy'),
                'systemPasswordWeak': request.form.get('systemPasswordWeak'),
                'systemTagging': request.form.get('systemTagging'),
                'systemTaggingIps': request.form.get('systemTaggingIps'),
                'ntpConfigured': request.form.get('ntpConfigured'),
                'usbPortsEnabled': request.form.get('usbPortsEnabled'),
                'usbApprovalProcess': request.form.get('usbApprovalProcess'),
                'trustedUsbCheck': request.form.get('trustedUsbCheck'),
                'firewallEnabled': request.form.get('firewallEnabled'),
                'firewallEnabledIps': request.form.get('firewallEnabledIps'),
                'externalStorageVirusCheck': request.form.get('externalStorageVirusCheck'),
                'rdpEnabled': request.form.get('rdpEnabled'),
                'rdpIps': request.form.get('rdpIps')
            }
            
            # Add Endpoints Vulnerability questions starting from the row after header
            start_endpoints_question_row = endpoints_header_row + 1
            
            for i, (question, risk_factor) in enumerate(zip(endpoints_questions, endpoints_risk_factors)):
                current_row = start_endpoints_question_row + i
                field_name = ['adminRightsLogin', 'groupPolicyModification', 'securityConfigModification', 'proxyModification', 'internetAllowed', 'internetAccessRestricted', 'internetAuthPolicy', 'systemPasswordWeak', 'systemTagging', 'ntpConfigured', 'usbPortsEnabled', 'usbApprovalProcess', 'trustedUsbCheck', 'firewallEnabled', 'externalStorageVirusCheck', 'rdpEnabled'][i]
                response = endpoints_form_responses.get(field_name, '')
                
                # Column A: Serial number
                worksheet[f'A{current_row}'] = i + 1
                worksheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'A{current_row}'].font = times_new_roman_normal
                
                # Merge columns B and C for the question
                worksheet.merge_cells(f'B{current_row}:C{current_row}')
                worksheet[f'B{current_row}'] = question
                worksheet[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'B{current_row}'].font = times_new_roman_normal
                
                # Column D: Risk factor
                worksheet[f'D{current_row}'] = risk_factor
                worksheet[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'D{current_row}'].font = times_new_roman_normal
                
                # Columns E and F: Compliance status and remarks based on user response
                compliance_status = "Not Answered"
                remarks = "No response provided for this question."
                
                if field_name == 'adminRightsLogin':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Admin rights are restricted in all of the systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        ips = endpoints_form_responses.get('adminRightsIps', '')
                        if ips:
                            ip_list = [ip.strip() for ip in ips.split('\n') if ip.strip()]
                            formatted_ips = '\n'.join(ip_list)
                            remarks = f'Admin rights are not restricted in the below listed system. IPs:\n{formatted_ips}'
                        else:
                            remarks = 'Admin rights are not restricted in the below listed system.'
                elif field_name == 'groupPolicyModification':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Group Policy cannot be modified in all of the systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        ips = endpoints_form_responses.get('groupPolicyIps', '')
                        if ips:
                            ip_list = [ip.strip() for ip in ips.split('\n') if ip.strip()]
                            formatted_ips = '\n'.join(ip_list)
                            remarks = f'Group Policy can be modified in the below listed systems. IPs:\n{formatted_ips}'
                        else:
                            remarks = 'Group Policy can be modified in the below listed systems.'
                elif field_name == 'securityConfigModification':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Security Configuration Policy cannot be modified in all of the systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        ips = endpoints_form_responses.get('securityConfigIps', '')
                        if ips:
                            ip_list = [ip.strip() for ip in ips.split('\n') if ip.strip()]
                            formatted_ips = '\n'.join(ip_list)
                            remarks = f'Security Configuration Policy can be modified in the below listed systems. IPs:\n{formatted_ips}'
                        else:
                            remarks = 'Security Configuration Policy can be modified in the below listed systems.'
                elif field_name == 'proxyModification':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Proxy cannot be modified in all of the systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Proxy can be modified in all of the systems, it is recommended to restrict user\'s rights and provide rights on need-to-know basis.'
                elif field_name == 'internetAllowed':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Internet was not allowed.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Internet was allowed.'
                elif field_name == 'internetAccessRestricted':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Restricted internet access is given to users.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Unrestricted internet access is given to users. It is recommended to implement whitelisting concept.'
                elif field_name == 'internetAuthPolicy':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Authentication policy was available for accessing the internet.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Authentication policy was not available for accessing internet.'
                    elif response == 'NA':
                        compliance_status = 'Not Applicable'
                        remarks = 'Bank does not use internet or does notgive internet access to all users.'
                elif field_name == 'systemPasswordWeak':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Strong password is being used for accessing the systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'System password is weak and common. Users need to follow standard password policy for accessing systems.'
                elif field_name == 'systemTagging':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'System tagging was present in the branch.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        ips = endpoints_form_responses.get('systemTaggingIps', '')
                        if ips:
                            ip_list = [ip.strip() for ip in ips.split('\n') if ip.strip()]
                            formatted_ips = '\n'.join(ip_list)
                            remarks = f'System tagging was not present in the below-mentioned systems. IPs:\n{formatted_ips}'
                        else:
                            remarks = 'System tagging was not present in the below-mentioned systems.'
                elif field_name == 'ntpConfigured':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'NTP was configured in systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'NTP was not configured in systems.'
                elif field_name == 'usbPortsEnabled':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'USB ports/CD drives were disabled in the systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'USB ports/CD drives were enabled in the systems.'
                elif field_name == 'usbApprovalProcess':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'USB approval process is available.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'USB approval process was not available.'
                elif field_name == 'trustedUsbCheck':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Bank has trusted USB; it does not open outside of the bank network.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Bank does not have trusted USB.'
                    elif response == 'NA':
                        compliance_status = 'Not Applicable'
                        remarks = 'Bank does not use USB.'
                elif field_name == 'firewallEnabled':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Firewall is enabled in systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        ips = endpoints_form_responses.get('firewallEnabledIps', '')
                        if ips:
                            ip_list = [ip.strip() for ip in ips.split('\n') if ip.strip()]
                            formatted_ips = '\n'.join(ip_list)
                            remarks = f'Firewall is not enabled in the below mentioned systems. IPs:\n{formatted_ips}'
                        else:
                            remarks = 'Firewall is not enabled in the below mentioned systems.'
                elif field_name == 'externalStorageVirusCheck':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'All external storage media are scanned for virus.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'External storage media are not scanned for virus.'
                elif field_name == 'rdpEnabled':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Remote Desktop Protocol (RDP) is disabled on the system.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        ips = endpoints_form_responses.get('rdpIps', '')
                        if ips:
                            ip_list = [ip.strip() for ip in ips.split('\n') if ip.strip()]
                            formatted_ips = '\n'.join(ip_list)
                            remarks = f'Remote Desktop Protocol (RDP) is not disabled on the system. IPs:\n{formatted_ips}'
                        else:
                            remarks = 'Remote Desktop Protocol (RDP) is not disabled on the system.'
                
                worksheet[f'E{current_row}'] = compliance_status
                worksheet[f'F{current_row}'] = remarks
                
                # Set alignment for E and F columns
                worksheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'F{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'E{current_row}'].font = times_new_roman_normal
                worksheet[f'F{current_row}'].font = times_new_roman_normal
                
                print(f"Added Endpoints Vulnerability question {i+1} at row {current_row} - Response: {response}, Status: {compliance_status}")
            
            # Update total rows to include Endpoints Vulnerability questions
            total_rows = start_endpoints_question_row + len(endpoints_questions) - 1
            
            # Add "ATM MACHINE ROOM" header after Endpoints Vulnerability questions
            atm_header_row = start_endpoints_question_row + len(endpoints_questions)
            worksheet.merge_cells(f'A{atm_header_row}:F{atm_header_row}')
            worksheet[f'A{atm_header_row}'] = "ATM MACHINE ROOM"
            worksheet[f'A{atm_header_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet[f'A{atm_header_row}'].font = times_new_roman_normal  # Use normal font, will be overridden by styling loop
            print(f"Added ATM MACHINE ROOM header at row {atm_header_row}")
            
            # Add ATM Machine Room questions and risk factors
            atm_questions = [
                "Check the CCTV camera  LAN cable and ATM machine power cable concealed?",
                "Guard is available or not?",
                "CCTV Camera and check the position of camera?",
                "ATM is working 24*7?",
                "Whether the Do's and Don't for the user awareness is available or not?",
                "Whether the ATM is network segmented or not?",
                "Whether ATM Machine is properly grouted (floor and wall) ?",
                "Whether access to ATM room for maintenance purpose is restricted to the authorized persons only?",
                "Whether ATM power back up is supported by UPS.",
                "System of ATM cash replenishment, adherence to dual control mechanism records is maintained properly?",
                "Check system of ATM reconciliation status maintained?",
                "Whether ATM machine preservation is done of journal print for future reference?"
            ]
            
            atm_risk_factors = [
                "High",
                "Low",
                "High",
                "Medium",
                "Medium",
                "High",
                "High",
                "High",
                "High",
                "High",
                "High",
                "High"
            ]
            
            # Get form responses for ATM Machine Room questions
            atm_form_responses = {
                'atmCctvCableConcealed': request.form.get('atmCctvCableConcealed'),
                'atmGuardAvailable': request.form.get('atmGuardAvailable'),
                'atmCctvPosition': request.form.get('atmCctvPosition'),
                'atmWorking247': request.form.get('atmWorking247'),
                'atmDosDontsAwareness': request.form.get('atmDosDontsAwareness'),
                'atmNetworkSegmented': request.form.get('atmNetworkSegmented'),
                'atmMachineGrouted': request.form.get('atmMachineGrouted'),
                'atmRoomAccessRestricted': request.form.get('atmRoomAccessRestricted'),
                'atmPowerBackupUps': request.form.get('atmPowerBackupUps'),
                'atmCashReplenishmentDualControl': request.form.get('atmCashReplenishmentDualControl'),
                'atmReconciliationStatus': request.form.get('atmReconciliationStatus'),
                'atmJournalPrintPreservation': request.form.get('atmJournalPrintPreservation')
            }
            
            # Add ATM Machine Room questions starting from the row after header
            start_atm_question_row = atm_header_row + 1
            
            for i, (question, risk_factor) in enumerate(zip(atm_questions, atm_risk_factors)):
                current_row = start_atm_question_row + i
                field_name = ['atmCctvCableConcealed', 'atmGuardAvailable', 'atmCctvPosition', 'atmWorking247', 'atmDosDontsAwareness', 'atmNetworkSegmented', 'atmMachineGrouted', 'atmRoomAccessRestricted', 'atmPowerBackupUps', 'atmCashReplenishmentDualControl', 'atmReconciliationStatus', 'atmJournalPrintPreservation'][i]
                response = atm_form_responses.get(field_name, '')
                
                # Column A: Serial number
                worksheet[f'A{current_row}'] = i + 1
                worksheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'A{current_row}'].font = times_new_roman_normal
                
                # Merge columns B and C for the question
                worksheet.merge_cells(f'B{current_row}:C{current_row}')
                worksheet[f'B{current_row}'] = question
                worksheet[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'B{current_row}'].font = times_new_roman_normal
                
                # Column D: Risk factor
                worksheet[f'D{current_row}'] = risk_factor
                worksheet[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'D{current_row}'].font = times_new_roman_normal
                
                # Columns E and F: Compliance status and remarks based on user response
                # For ATM Machine Room, treat empty responses as "Not Applicable"
                if not response or response.strip() == "":
                    compliance_status = "Not Applicable"
                    remarks = "The branch does not have an ATM facility."
                else:
                    compliance_status = "Not Answered"
                    remarks = "No response provided for this question."
                
                if field_name == 'atmCctvCableConcealed':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'CCTV camera LAN cable and ATM machine power cable concealed.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'CCTV camera LAN cable and ATM machine power cable was not concealed.'
                    elif response == 'Not Applicable':
                        compliance_status = 'Not Applicable'
                        remarks = 'The branch does not have an ATM facility.'
                elif field_name == 'atmGuardAvailable':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Guard was available.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Guard was not available.'
                    elif response == 'Not Applicable':
                        compliance_status = 'Not Applicable'
                        remarks = 'The branch does not have an ATM facility.'
                elif field_name == 'atmCctvPosition':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'CCTV Camera was installed in proper position.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'CCTV Camera was not installed in proper position.'
                    elif response == 'Not Applicable':
                        compliance_status = 'Not Applicable'
                        remarks = 'The branch does not have an ATM facility.'
                elif field_name == 'atmWorking247':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'ATM was working 24*7.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'ATM was not working 24*7.'
                    elif response == 'Not Applicable':
                        compliance_status = 'Not Applicable'
                        remarks = 'The branch does not have an ATM facility.'
                elif field_name == 'atmDosDontsAwareness':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Do\'s and Don\'t Instructions are available.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Do\'s and Don\'t Instructions are not available.'
                    elif response == 'Not Applicable':
                        compliance_status = 'Not Applicable'
                        remarks = 'The branch does not have an ATM facility.'
                elif field_name == 'atmNetworkSegmented':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'ATM is network segmented.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Network is not segmented for ATM.'
                    elif response == 'Not Applicable':
                        compliance_status = 'Not Applicable'
                        remarks = 'The branch does not have an ATM facility.'
                elif field_name == 'atmMachineGrouted':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Machine is properly grouted in floor and wall.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Machine is not properly grouted in floor and wall.'
                    elif response == 'Not Applicable':
                        compliance_status = 'Not Applicable'
                        remarks = 'The branch does not have an ATM facility.'
                elif field_name == 'atmRoomAccessRestricted':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Access to ATM room for maintenance purpose for unauthorized person was restricted.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Access to ATM room for maintenance purpose was not restricted.'
                    elif response == 'Not Applicable':
                        compliance_status = 'Not Applicable'
                        remarks = 'The branch does not have an ATM facility.'
                elif field_name == 'atmPowerBackupUps':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'ATM power back up was supported by UPS.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'ATM power back up was not available.'
                    elif response == 'Not Applicable':
                        compliance_status = 'Not Applicable'
                        remarks = 'The branch does not have an ATM facility.'
                elif field_name == 'atmCashReplenishmentDualControl':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'System of ATM cash replenishment, adherence to dual control mechanism records is maintained properly.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'System of ATM cash replenishment, adherence to dual control mechanism records is not maintained properly.'
                    elif response == 'Not Applicable':
                        compliance_status = 'Not Applicable'
                        remarks = 'The branch does not have an ATM facility.'
                elif field_name == 'atmReconciliationStatus':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'System of ATM reconciliation status maintained.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'System of ATM reconciliation status is not maintained.'
                    elif response == 'Not Applicable':
                        compliance_status = 'Not Applicable'
                        remarks = 'The branch does not have an ATM facility.'
                elif field_name == 'atmJournalPrintPreservation':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'ATM machine preserves journal print for future reference.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'ATM machine does not preserve journal print for future reference.'
                    elif response == 'Not Applicable':
                        compliance_status = 'Not Applicable'
                        remarks = 'The branch does not have an ATM facility.'
                
                worksheet[f'E{current_row}'] = compliance_status
                worksheet[f'F{current_row}'] = remarks
                
                # Set alignment for E and F columns
                worksheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'F{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'E{current_row}'].font = times_new_roman_normal
                worksheet[f'F{current_row}'].font = times_new_roman_normal
                
                print(f"Added ATM Machine Room question {i+1} at row {current_row} - Response: {response}, Status: {compliance_status}")
            
            # Update total rows to include ATM Machine Room questions
            total_rows = start_atm_question_row + len(atm_questions) - 1
            
            # Add "EMAIL-SECURITY" header after ATM Machine Room questions
            email_header_row = start_atm_question_row + len(atm_questions)
            worksheet.merge_cells(f'A{email_header_row}:F{email_header_row}')
            worksheet[f'A{email_header_row}'] = "EMAIL-SECURITY"
            worksheet[f'A{email_header_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet[f'A{email_header_row}'].font = times_new_roman_normal  # Use normal font, will be overridden by styling loop
            print(f"Added EMAIL-SECURITY header at row {email_header_row}")
            
            # Add Email Security questions and risk factors
            email_questions = [
                "Personal Mail used or not?",
                "Two factor authentications are in used or not?",
                "Single mail with multiple users?",
                "Whether official email ID gets open outside bank N/w?"
            ]
            
            email_risk_factors = [
                "High",
                "High",
                "Medium",
                "High"
            ]
            
            # Get form responses for Email Security questions
            email_form_responses = {
                'personalMailUsed': request.form.get('personalMailUsed'),
                'emailTwoFactorAuth': request.form.get('emailTwoFactorAuth'),
                'singleMailMultipleUsers': request.form.get('singleMailMultipleUsers'),
                'officialEmailOutsideNetwork': request.form.get('officialEmailOutsideNetwork')
            }
            
            # Add Email Security questions starting from the row after header
            start_email_question_row = email_header_row + 1
            
            for i, (question, risk_factor) in enumerate(zip(email_questions, email_risk_factors)):
                current_row = start_email_question_row + i
                field_name = ['personalMailUsed', 'emailTwoFactorAuth', 'singleMailMultipleUsers', 'officialEmailOutsideNetwork'][i]
                response = email_form_responses.get(field_name, '')
                
                # Column A: Serial number
                worksheet[f'A{current_row}'] = i + 1
                worksheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'A{current_row}'].font = times_new_roman_normal
                
                # Merge columns B and C for the question
                worksheet.merge_cells(f'B{current_row}:C{current_row}')
                worksheet[f'B{current_row}'] = question
                worksheet[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'B{current_row}'].font = times_new_roman_normal
                
                # Column D: Risk factor
                worksheet[f'D{current_row}'] = risk_factor
                worksheet[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'D{current_row}'].font = times_new_roman_normal
                
                # Columns E and F: Compliance status and remarks based on user response
                compliance_status = "Not Answered"
                remarks = "No response provided for this question."
                
                if field_name == 'personalMailUsed':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Bank official email id is in use.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Personal email id is in use.'
                elif field_name == 'emailTwoFactorAuth':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Two factor authentication was enabled for mail.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Two factor authentication was not enabled for mail. It is recommended to enable Two factor authentication.'
                elif field_name == 'singleMailMultipleUsers':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Dedicated mail id is provided to each user.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Single mail is used by multiple users. Accountability cannot be established in case of any data theft.'
                elif field_name == 'officialEmailOutsideNetwork':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Email id opens only in the branch network.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Bank official email id can be accessed outside of the bank network.'
                
                worksheet[f'E{current_row}'] = compliance_status
                worksheet[f'F{current_row}'] = remarks
                
                # Set alignment for E and F columns
                worksheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'F{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'E{current_row}'].font = times_new_roman_normal
                worksheet[f'F{current_row}'].font = times_new_roman_normal
                
                print(f"Added Email Security question {i+1} at row {current_row} - Response: {response}, Status: {compliance_status}")
            
            # Update total rows to include Email Security questions
            total_rows = start_email_question_row + len(email_questions) - 1
            
            # Add "REMOTE ACCESS" header after Email Security questions
            remote_header_row = start_email_question_row + len(email_questions)
            worksheet.merge_cells(f'A{remote_header_row}:F{remote_header_row}')
            worksheet[f'A{remote_header_row}'] = "REMOTE ACCESS"
            worksheet[f'A{remote_header_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet[f'A{remote_header_row}'].font = times_new_roman_normal  # Use normal font, will be overridden by styling loop
            print(f"Added REMOTE ACCESS header at row {remote_header_row}")
            
            # Add Remote Access questions and risk factors
            remote_questions = [
                "Which Department is using Remote Access?",
                "Remote Access approval process?"
            ]
            
            remote_risk_factors = [
                "Medium",
                "Medium"
            ]
            
            # Get form responses for Remote Access questions
            remote_form_responses = {
                'departmentRemoteAccess': request.form.get('departmentRemoteAccess'),
                'remoteAccessApprovalProcess': request.form.get('remoteAccessApprovalProcess')
            }
            
            # Add Remote Access questions starting from the row after header
            start_remote_question_row = remote_header_row + 1
            
            for i, (question, risk_factor) in enumerate(zip(remote_questions, remote_risk_factors)):
                current_row = start_remote_question_row + i
                field_name = ['departmentRemoteAccess', 'remoteAccessApprovalProcess'][i]
                response = remote_form_responses.get(field_name, '')
                
                # Column A: Serial number
                worksheet[f'A{current_row}'] = i + 1
                worksheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'A{current_row}'].font = times_new_roman_normal
                
                # Merge columns B and C for the question
                worksheet.merge_cells(f'B{current_row}:C{current_row}')
                worksheet[f'B{current_row}'] = question
                worksheet[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'B{current_row}'].font = times_new_roman_normal
                
                # Column D: Risk factor
                worksheet[f'D{current_row}'] = risk_factor
                worksheet[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'D{current_row}'].font = times_new_roman_normal
                
                # Columns E and F: Compliance status and remarks based on user response
                compliance_status = "Not Answered"
                remarks = "No response provided for this question."
                
                if field_name == 'departmentRemoteAccess':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'IT department is using remote access.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Other Department or Person is using remote access.'
                    elif response == 'NA':
                        compliance_status = 'Not Applicable'
                        remarks = 'Bank does not use remote access.'
                elif field_name == 'remoteAccessApprovalProcess':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'Remote access approval process was available.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Remote access approval process was not available.'
                    elif response == 'NA':
                        compliance_status = 'Not Applicable'
                        remarks = 'Bank does not use remote access.'
                
                worksheet[f'E{current_row}'] = compliance_status
                worksheet[f'F{current_row}'] = remarks
                
                # Set alignment for E and F columns
                worksheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'F{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'E{current_row}'].font = times_new_roman_normal
                worksheet[f'F{current_row}'].font = times_new_roman_normal
                
                print(f"Added Remote Access question {i+1} at row {current_row} - Response: {response}, Status: {compliance_status}")
            
            # Update total rows to include Remote Access questions
            total_rows = start_remote_question_row + len(remote_questions) - 1
            
            # Add "UNAUTHORIZED APPLICATIONS / PERSONAL DATA" header after Remote Access questions
            unauthorized_header_row = start_remote_question_row + len(remote_questions)
            worksheet.merge_cells(f'A{unauthorized_header_row}:F{unauthorized_header_row}')
            worksheet[f'A{unauthorized_header_row}'] = "UNAUTHORIZED APPLICATIONS / PERSONAL DATA"
            worksheet[f'A{unauthorized_header_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet[f'A{unauthorized_header_row}'].font = times_new_roman_normal  # Use normal font, will be overridden by styling loop
            print(f"Added UNAUTHORIZED APPLICATIONS / PERSONAL DATA header at row {unauthorized_header_row}")
            
            # Add Unauthorized Applications / Personal Data questions and risk factors
            unauthorized_questions = [
                "unauthorizedApplications",
                "personalDataPresent"
            ]
            
            unauthorized_risk_factors = [
                "Medium",
                "Low"
            ]
            
            # Get form responses for Unauthorized Applications / Personal Data questions
            unauthorized_form_responses = {
                'unauthorizedApplications': request.form.get('unauthorizedApplications'),
                'unauthorizedAppsIps': request.form.get('unauthorizedAppsIps'),
                'unauthorizedAppsNames': request.form.get('unauthorizedAppsNames'),
                'personalDataPresent': request.form.get('personalDataPresent'),
                'personalDataIps': request.form.get('personalDataIps')
            }
            
            # Add Unauthorized Applications / Personal Data questions starting from the row after header
            start_unauthorized_question_row = unauthorized_header_row + 1
            
            for i, (question, risk_factor) in enumerate(zip(unauthorized_questions, unauthorized_risk_factors)):
                current_row = start_unauthorized_question_row + i
                field_name = ['unauthorizedApplications', 'personalDataPresent'][i]
                response = unauthorized_form_responses.get(field_name, '')
                
                # Column A: Serial number
                worksheet[f'A{current_row}'] = i + 1
                worksheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'A{current_row}'].font = times_new_roman_normal
                
                # Merge columns B and C for the question - show IP addresses instead of field names
                worksheet.merge_cells(f'B{current_row}:C{current_row}')
                
                # Display descriptive text with IP addresses or compliance message in B-C merged cell
                if field_name == 'unauthorizedApplications':
                    if response == 'Compliance':
                        worksheet[f'B{current_row}'] = "There are no unauthorized applications on any systems."
                    else:
                        ips = unauthorized_form_responses.get('unauthorizedAppsIps', '')
                        if ips:
                            ip_list = [ip.strip() for ip in ips.split('\n') if ip.strip()]
                            formatted_ips = '\n'.join(ip_list)
                            worksheet[f'B{current_row}'] = f"Unauthorized applications have been found in the below systems IP's\n{formatted_ips}"
                        else:
                            worksheet[f'B{current_row}'] = question
                elif field_name == 'personalDataPresent':
                    if response == 'Compliance':
                        worksheet[f'B{current_row}'] = "There are no Personal Data on any systems."
                    else:
                        ips = unauthorized_form_responses.get('personalDataIps', '')
                        if ips:
                            ip_list = [ip.strip() for ip in ips.split('\n') if ip.strip()]
                            formatted_ips = '\n'.join(ip_list)
                            worksheet[f'B{current_row}'] = f"Personal Data have been found in the below systems IP's\n{formatted_ips}"
                        else:
                            worksheet[f'B{current_row}'] = question
                else:
                    worksheet[f'B{current_row}'] = question
                
                worksheet[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'B{current_row}'].font = times_new_roman_normal
                
                # Column D: Risk factor
                worksheet[f'D{current_row}'] = risk_factor
                worksheet[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'D{current_row}'].font = times_new_roman_normal
                
                # Columns E and F: Compliance status and remarks based on user response
                compliance_status = "Not Answered"
                remarks = "No response provided for this question."
                
                if field_name == 'unauthorizedApplications':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'There are no unauthorized applications on any systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        app_names = unauthorized_form_responses.get('unauthorizedAppsNames', '')
                        if app_names:
                            app_list = [app.strip() for app in app_names.split('\n') if app.strip()]
                            formatted_apps = '\n'.join(app_list)
                            remarks = f'Below listed Unauthorised Applications have been found in systems.\n{formatted_apps}'
                        else:
                            remarks = 'Below listed Unauthorised Applications have been found in systems.'
                elif field_name == 'personalDataPresent':
                    if response == 'Compliance':
                        compliance_status = 'Compliance'
                        remarks = 'There are no Personal Data on any systems.'
                    elif response == 'Non-Compliance':
                        compliance_status = 'Non-Compliance'
                        remarks = 'Personal files were available in the system.'
                
                worksheet[f'E{current_row}'] = compliance_status
                worksheet[f'F{current_row}'] = remarks
                
                # Set alignment for E and F columns
                worksheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'F{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet[f'E{current_row}'].font = times_new_roman_normal
                worksheet[f'F{current_row}'].font = times_new_roman_normal
                
                print(f"Added Unauthorized Applications / Personal Data question {i+1} at row {current_row} - Response: {response}, Status: {compliance_status}")
            
            # Update total rows to include Unauthorized Applications / Personal Data questions
            total_rows = start_unauthorized_question_row + len(unauthorized_questions) - 1
            
            # Get form responses for Important Note questions first to check if any are Non-Compliance
            important_note_form_responses = {
                'auditorIdentityNotVerified': request.form.get('auditorIdentityNotVerified'),
                'windowsNotActivated': request.form.get('windowsNotActivated'),
                'windowsNotActivatedIps': request.form.get('windowsNotActivatedIps'),
                'passwordSavedInBrowser': request.form.get('passwordSavedInBrowser'),
                'passwordBrowserIps': request.form.get('passwordBrowserIps'),
                'passwordWrittenOnWall': request.form.get('passwordWrittenOnWall'),
                'assetMovementRegisterNotAvailable': request.form.get('assetMovementRegisterNotAvailable'),
                'dustPresentOnDevices': request.form.get('dustPresentOnDevices'),
                'autoSwitchableModeNotAvailable': request.form.get('autoSwitchableModeNotAvailable'),
                'preventiveMaintenanceNotCarriedOut': request.form.get('preventiveMaintenanceNotCarriedOut')
            }
            
            # Check if any Important Note items are Non-Compliance
            important_note_questions = [
                "Auditor Identity was not verified.",
                "Windows was not activated.",
                "Password was saved in browser.",
                "Password was written on wall, desk, notepad, diary.",
                "Asset Movement Register was not available.",
                "Dust was present on the network devices & systems.",
                "Auto switchable mode was not available for connectivity.",
                "Preventive Maintenance is not carried out for Systems/UPS/Network Devices."
            ]
            
            important_note_risk_factors = [
                "Medium",
                "High",
                "Medium",
                "High",
                "Medium",
                "Medium",
                "High",
                "High"
            ]
            
            field_names = ['auditorIdentityNotVerified', 'windowsNotActivated', 'passwordSavedInBrowser', 'passwordWrittenOnWall', 'assetMovementRegisterNotAvailable', 'dustPresentOnDevices', 'autoSwitchableModeNotAvailable', 'preventiveMaintenanceNotCarriedOut']
            
            # Check if any responses are Non-Compliance
            has_non_compliance = any(important_note_form_responses.get(field_name, '') == 'Non-Compliance' for field_name in field_names)
            
            if has_non_compliance:
                # Add "Important Note" header only if there are Non-Compliance items
                important_note_header_row = start_unauthorized_question_row + len(unauthorized_questions)
                worksheet.merge_cells(f'A{important_note_header_row}:F{important_note_header_row}')
                worksheet[f'A{important_note_header_row}'] = "Important Note"
                worksheet[f'A{important_note_header_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                worksheet[f'A{important_note_header_row}'].font = times_new_roman_normal  # Use normal font, will be overridden by styling loop
                print(f"Added Important Note header at row {important_note_header_row}")
                
                # Add Important Note questions starting from the row after header
                start_important_note_question_row = important_note_header_row + 1
                current_important_note_row = start_important_note_question_row
                important_note_serial_number = 1
                
                for i, (question, risk_factor) in enumerate(zip(important_note_questions, important_note_risk_factors)):
                    field_name = field_names[i]
                    response = important_note_form_responses.get(field_name, '')
                    
                    # Only add row if response is Non-Compliance
                    if response == 'Non-Compliance':
                        # Column A: Serial number
                        worksheet[f'A{current_important_note_row}'] = important_note_serial_number
                        worksheet[f'A{current_important_note_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        worksheet[f'A{current_important_note_row}'].font = times_new_roman_normal
                        
                        # Merge columns B and C for the question
                        worksheet.merge_cells(f'B{current_important_note_row}:C{current_important_note_row}')
                        worksheet[f'B{current_important_note_row}'] = question
                        worksheet[f'B{current_important_note_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        worksheet[f'B{current_important_note_row}'].font = times_new_roman_normal
                        
                        # Column D: Risk factor
                        worksheet[f'D{current_important_note_row}'] = risk_factor
                        worksheet[f'D{current_important_note_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        worksheet[f'D{current_important_note_row}'].font = times_new_roman_normal
                        
                        # Columns E and F: Compliance status and remarks based on user response
                        compliance_status = "Non-Compliance"
                        remarks = "No response provided for this question."
                        
                        if field_name == 'auditorIdentityNotVerified':
                            remarks = 'Auditor Identity was not verified.'
                        elif field_name == 'windowsNotActivated':
                            ips = important_note_form_responses.get('windowsNotActivatedIps', '')
                            if ips:
                                ip_list = [ip.strip() for ip in ips.split('\n') if ip.strip()]
                                formatted_ips = '\n'.join(ip_list)
                                remarks = f'Windows is not activated on the following systems:\n{formatted_ips}'
                            else:
                                remarks = 'Windows is not activated on the following systems.'
                        elif field_name == 'passwordSavedInBrowser':
                            ips = important_note_form_responses.get('passwordBrowserIps', '')
                            if ips:
                                ip_list = [ip.strip() for ip in ips.split('\n') if ip.strip()]
                                formatted_ips = '\n'.join(ip_list)
                                remarks = f'Passwords were saved in the browser on the following systems:\n{formatted_ips}'
                            else:
                                remarks = 'Passwords were saved in the browser on the following systems.'
                        elif field_name == 'passwordWrittenOnWall':
                            remarks = 'Password was written on wall, desk, notepad, diary.'
                        elif field_name == 'assetMovementRegisterNotAvailable':
                            remarks = 'Asset Movement Register was not available.'
                        elif field_name == 'dustPresentOnDevices':
                            remarks = 'Dust was present on the network devices & systems.'
                        elif field_name == 'autoSwitchableModeNotAvailable':
                            remarks = 'Auto switchable mode was not available for connectivity.'
                        elif field_name == 'preventiveMaintenanceNotCarriedOut':
                            remarks = 'Preventive Maintenance is not carried out for Systems/UPS/Network Devices.'
                        
                        worksheet[f'E{current_important_note_row}'] = compliance_status
                        worksheet[f'F{current_important_note_row}'] = remarks
                        
                        # Set alignment for E and F columns
                        worksheet[f'E{current_important_note_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        worksheet[f'F{current_important_note_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        worksheet[f'E{current_important_note_row}'].font = times_new_roman_normal
                        worksheet[f'F{current_important_note_row}'].font = times_new_roman_normal
                        
                        print(f"Added Important Note question {important_note_serial_number} at row {current_important_note_row} - Response: {response}, Status: {compliance_status}")
                        
                        # Increment row and serial number for next item
                        current_important_note_row += 1
                        important_note_serial_number += 1
                
                # Update total rows to include Important Note questions (only non-compliance items)
                total_rows = current_important_note_row - 1
            else:
                # No Non-Compliance items, so no Important Note section needed
                print("All Important Note items are Compliance - skipping Important Note section")
                total_rows = start_unauthorized_question_row + len(unauthorized_questions) - 1
        
        # Add some basic styling to make it look better
        
        # Create Times New Roman font styles with size 12
        times_new_roman_bold = Font(name='Times New Roman', bold=True, size=12)
        times_new_roman_normal = Font(name='Times New Roman', bold=False, size=12)
        
        # Create border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Calculate the number of system data rows needed
        if num_valid_systems > 0:
            # total_rows is already calculated in the CBS section above
            pass  # Use the total_rows from CBS section
        else:
            total_rows = 10  # Just the header rows if no systems
        
        # Apply styling to all rows (1 to total_rows)
        for row_num in range(1, total_rows + 1):  # Rows 1 to total_rows
            # Apply borders to all cells in each row
            for col in range(1, 7):  # Columns A-F (1-6)
                cell = worksheet.cell(row=row_num, column=col)
                cell.border = thin_border
                
                # Apply text wrapping to all cells
                cell.alignment = cell.alignment.copy(wrap_text=True) if cell.alignment else Alignment(wrap_text=True)
                
                # Apply specific font styling based on row
                if row_num in [8, 9, 10]:  # Rows 8, 9, 10 - all bold
                    cell.font = times_new_roman_bold
                elif row_num <= 7:  # Rows 1-7 - mixed styling
                    if col in [1, 2]:  # Columns A-B (labels)
                        cell.font = times_new_roman_bold
                    else:  # Columns C-F (values)
                        cell.font = times_new_roman_normal
                elif row_num >= 11:  # System data rows, CBS header, CBS questions, Physical header, Physical questions, Power header, Power questions, User Awareness header, User Awareness questions, Maintenance header, Maintenance questions, Patch header, Patch questions, Network header, Network questions, Endpoints header, Endpoints questions, ATM header, ATM questions, Email header, Email questions, Remote header, Remote questions, Unauthorized header, Unauthorized questions, Important Note header, and Important Note questions
                    if num_valid_systems > 0 and row_num == (11 + num_valid_systems):  # CBS ACCESS CONTROL header
                        cell.font = times_new_roman_bold
                    elif num_valid_systems > 0 and row_num == (11 + num_valid_systems + 1 + len(cbs_questions)):  # PHYSICAL AND ENVIRONMENTAL SECURITY header
                        cell.font = times_new_roman_bold
                    elif num_valid_systems > 0 and row_num == (11 + num_valid_systems + 1 + len(cbs_questions) + 1 + len(physical_questions)):  # POWER BACK UP header
                        cell.font = times_new_roman_bold
                    elif num_valid_systems > 0 and row_num == (11 + num_valid_systems + 1 + len(cbs_questions) + 1 + len(physical_questions) + 1 + len(power_questions)):  # USER AWARENESS header
                        cell.font = times_new_roman_bold
                    elif num_valid_systems > 0 and row_num == (11 + num_valid_systems + 1 + len(cbs_questions) + 1 + len(physical_questions) + 1 + len(power_questions) + 1 + len(user_awareness_questions)):  # MAINTENANCE AND BUSINESS CONTINUITY CONTROLS header
                        cell.font = times_new_roman_bold
                    elif num_valid_systems > 0 and row_num == (11 + num_valid_systems + 1 + len(cbs_questions) + 1 + len(physical_questions) + 1 + len(power_questions) + 1 + len(user_awareness_questions) + 1 + len(maintenance_questions)):  # PATCH MANAGEMENT header
                        cell.font = times_new_roman_bold
                    elif num_valid_systems > 0 and row_num == (11 + num_valid_systems + 1 + len(cbs_questions) + 1 + len(physical_questions) + 1 + len(power_questions) + 1 + len(user_awareness_questions) + 1 + len(maintenance_questions) + 1 + len(patch_questions)):  # NETWORK SECURITY header
                        cell.font = times_new_roman_bold
                    elif num_valid_systems > 0 and row_num == (11 + num_valid_systems + 1 + len(cbs_questions) + 1 + len(physical_questions) + 1 + len(power_questions) + 1 + len(user_awareness_questions) + 1 + len(maintenance_questions) + 1 + len(patch_questions) + 1 + len(network_questions)):  # ENDPOINTS VULNERABILITY header
                        cell.font = times_new_roman_bold
                    elif num_valid_systems > 0 and row_num == (11 + num_valid_systems + 1 + len(cbs_questions) + 1 + len(physical_questions) + 1 + len(power_questions) + 1 + len(user_awareness_questions) + 1 + len(maintenance_questions) + 1 + len(patch_questions) + 1 + len(network_questions) + 1 + len(endpoints_questions)):  # ATM MACHINE ROOM header
                        cell.font = times_new_roman_bold
                    elif num_valid_systems > 0 and row_num == (11 + num_valid_systems + 1 + len(cbs_questions) + 1 + len(physical_questions) + 1 + len(power_questions) + 1 + len(user_awareness_questions) + 1 + len(maintenance_questions) + 1 + len(patch_questions) + 1 + len(network_questions) + 1 + len(endpoints_questions) + 1 + len(atm_questions)):  # EMAIL-SECURITY header
                        cell.font = times_new_roman_bold
                    elif num_valid_systems > 0 and row_num == (11 + num_valid_systems + 1 + len(cbs_questions) + 1 + len(physical_questions) + 1 + len(power_questions) + 1 + len(user_awareness_questions) + 1 + len(maintenance_questions) + 1 + len(patch_questions) + 1 + len(network_questions) + 1 + len(endpoints_questions) + 1 + len(atm_questions) + 1 + len(email_questions)):  # REMOTE ACCESS header
                        cell.font = times_new_roman_bold
                    elif num_valid_systems > 0 and row_num == (11 + num_valid_systems + 1 + len(cbs_questions) + 1 + len(physical_questions) + 1 + len(power_questions) + 1 + len(user_awareness_questions) + 1 + len(maintenance_questions) + 1 + len(patch_questions) + 1 + len(network_questions) + 1 + len(endpoints_questions) + 1 + len(atm_questions) + 1 + len(email_questions) + 1 + len(remote_questions)):  # UNAUTHORIZED APPLICATIONS / PERSONAL DATA header
                        cell.font = times_new_roman_bold
                    elif num_valid_systems > 0 and row_num == (11 + num_valid_systems + 1 + len(cbs_questions) + 1 + len(physical_questions) + 1 + len(power_questions) + 1 + len(user_awareness_questions) + 1 + len(maintenance_questions) + 1 + len(patch_questions) + 1 + len(network_questions) + 1 + len(endpoints_questions) + 1 + len(atm_questions) + 1 + len(email_questions) + 1 + len(remote_questions) + 1 + len(unauthorized_questions)):  # Important Note header
                        cell.font = times_new_roman_bold
                    else:  # System data rows, CBS questions, Physical questions, Power questions, User Awareness questions, Maintenance questions, Patch questions, Network questions, Endpoints questions, ATM questions, Email questions, Remote questions, Unauthorized questions, and Important Note questions - normal font
                        cell.font = times_new_roman_normal
            
            # Auto-adjust row height for text wrapping
            worksheet.row_dimensions[row_num].height = None  # Let Excel auto-adjust
        
        # Set column widths as specified
        worksheet.column_dimensions['A'].width = 10
        worksheet.column_dimensions['B'].width = 28
        worksheet.column_dimensions['C'].width = 31
        worksheet.column_dimensions['D'].width = 21
        worksheet.column_dimensions['E'].width = 23
        worksheet.column_dimensions['F'].width = 48
        
        print(f"Excel layout created successfully!")
        print(f"Sr. No label in merged cells A1:B1")
        print(f"Sr. No value '{sr_no}' in merged cells C1:F1")
        
        # Create a temporary file to save the Excel
        temp_dir = tempfile.mkdtemp()
        temp_excel = os.path.join(temp_dir, "branch_audit_report.xlsx")
        
        # Save the workbook
        print("Saving Excel file...")
        workbook.save(temp_excel)
        print(f"Excel file saved to: {temp_excel}")
        
        # Generate filename using Sr. No. and branch name
        if branch_name and branch_name.strip():
            # Clean the branch name for filename (remove invalid characters)
            clean_branch_name = branch_name.strip()
            # Replace invalid filename characters with underscores
            import re
            clean_branch_name = re.sub(r'[<>:"/\\|?*]', '_', clean_branch_name)
            # Remove multiple consecutive underscores
            clean_branch_name = re.sub(r'_+', '_', clean_branch_name)
            # Remove leading/trailing underscores
            clean_branch_name = clean_branch_name.strip('_')
            
            # Add Sr. No. before branch name if provided
            if sr_no and sr_no.strip():
                clean_sr_no = sr_no.strip()
                # Clean Sr. No. for filename (remove invalid characters)
                clean_sr_no = re.sub(r'[<>:"/\\|?*]', '_', clean_sr_no)
                clean_sr_no = re.sub(r'_+', '_', clean_sr_no)
                clean_sr_no = clean_sr_no.strip('_')
                filename = f'{clean_sr_no} {clean_branch_name}.xlsx'
            else:
                filename = f'{clean_branch_name}.xlsx'
        else:
            # Fallback to branch code if branch name is not provided
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'Branch_Audit_Report_{branch_code}_{timestamp}.xlsx'
        print(f"Generated filename: {filename}")
        
        # Send file for download
        print("Sending file for download...")
        
        try:
            # Return HTML with immediate download and redirect
            from flask import Response
            import base64
            
            # Read the file content
            with open(temp_excel, 'rb') as f:
                file_content = f.read()
            
            # Encode file content to base64 for JavaScript
            file_base64 = base64.b64encode(file_content).decode('utf-8')
            
            # Create minimal HTML response with immediate download and redirect
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Downloading...</title>
                <meta charset="UTF-8">
            </head>
            <body>
                <script>
                    // Convert base64 to blob and download immediately
                    const byteCharacters = atob('{file_base64}');
                    const byteNumbers = new Array(byteCharacters.length);
                    for (let i = 0; i < byteCharacters.length; i++) {{
                        byteNumbers[i] = byteCharacters.charCodeAt(i);
                    }}
                    const byteArray = new Uint8Array(byteNumbers);
                    const blob = new Blob([byteArray], {{ type: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' }});
                    
                    // Create download link and trigger download
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = '{filename}';
                    a.style.display = 'none';
                    document.body.appendChild(a);
                    a.click();
                    
                    // Clean up
                    window.URL.revokeObjectURL(url);
                    document.body.removeChild(a);
                    
                    // Immediately redirect to audit dashboard
                    window.location.href = '/audit_dashboard';
                </script>
            </body>
            </html>
            """
            
            print(f"File sent successfully: {filename}")
            return Response(html_content, mimetype='text/html')
                
        except Exception as send_error:
            print(f"Error sending file: {send_error}")
                # Clean up on error
            shutil.rmtree(temp_dir, ignore_errors=True)
            print(f"Cleaned up temporary directory: {temp_dir}")
            raise send_error
        
    except Exception as e:
        # Clean up temporary files in case of error
        if 'temp_dir' in locals():
            shutil.rmtree(temp_dir, ignore_errors=True)
            print(f"Cleaned up temporary directory: {temp_dir}")
        
        print(f"ERROR PROCESSING REQUEST: {str(e)}")
        print(f"Exception type: {type(e)}")
        import traceback
        print(f"Full traceback: {traceback.format_exc()}")
        
        flash(f'Error processing request: {str(e)}', 'error')
        return redirect(url_for('audit_dashboard'))
