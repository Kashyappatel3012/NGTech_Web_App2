"""
User Activity Logger Module
Handles logging all user activities to Excel files with automatic file/worksheet management
"""
import os
import json
import time
import shutil
import threading
from datetime import datetime
from contextlib import nullcontext
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# File locks dictionary to prevent concurrent access to the same file
_file_locks = {}
_locks_lock = threading.Lock()

def get_file_lock(filename):
    """Get or create a reentrant lock for a specific file"""
    with _locks_lock:
        if filename not in _file_locks:
            _file_locks[filename] = threading.RLock()  # Use RLock for reentrant locking
        return _file_locks[filename]

def get_column_letter_wrapper(col_num):
    """Wrapper for get_column_letter to handle compatibility"""
    try:
        return get_column_letter(col_num)
    except:
        # Fallback for older openpyxl versions
        letters = ''
        while col_num > 0:
            col_num -= 1
            letters = chr(65 + (col_num % 26)) + letters
            col_num //= 26
        return letters

# Constants
MAX_ROWS_PER_WORKSHEET = 1040000
LOGS_DIR = os.path.join('static', 'Logs')

def ensure_logs_directory():
    """Ensure the logs directory exists"""
    os.makedirs(LOGS_DIR, exist_ok=True)

def get_current_date_string():
    """Get current date in YYYY-MM-DD format"""
    return datetime.now().strftime('%Y-%m-%d')

def get_current_datetime_string():
    """Get current date and time in YYYY-MM-DD HH:MM:SS format"""
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def get_current_month_string():
    """Get current month in MMM_YYYY format (e.g., Dec_2025)"""
    return datetime.now().strftime('%b_%Y')

def get_excel_filename(month_str=None):
    """Get Excel filename for a given month (e.g., Dec_2025.xlsx)"""
    if month_str is None:
        month_str = get_current_month_string()
    return os.path.join(LOGS_DIR, f'{month_str}.xlsx')

def get_worksheet_name(date_str=None):
    """Get worksheet name for a given date (e.g., 2025-12-02)"""
    if date_str is None:
        date_str = get_current_date_string()
    return date_str

def get_or_create_workbook(month_str=None, max_retries=3, file_lock=None):
    """Get existing workbook or create new one for a given month with error handling and file locking"""
    ensure_logs_directory()
    filename = get_excel_filename(month_str)
    
    # If lock is not provided, get one (for cases where function is called directly)
    if file_lock is None:
        file_lock = get_file_lock(filename)
        lock_context = file_lock
    else:
        # Lock already acquired by caller, use null context (no-op)
        lock_context = nullcontext()
    
    # Acquire lock before accessing the file (or use existing lock)
    with lock_context:
        if os.path.exists(filename):
            # Try to load the workbook with retry logic and error handling
            for attempt in range(max_retries):
                try:
                    # Check if file is accessible and not corrupted
                    file_size = os.path.getsize(filename)
                    if file_size == 0:
                        print(f"Warning: Log file {filename} is empty, creating new workbook")
                        wb = Workbook()
                        if 'Sheet' in wb.sheetnames:
                            wb.remove(wb['Sheet'])
                        return wb, filename
                    
                    # Wait a bit if file is very small (might be in the process of being written)
                    if file_size < 1000:  # Less than 1KB is suspicious
                        time.sleep(0.2)
                    
                    # Try to load the workbook
                    wb = load_workbook(filename, read_only=False, keep_links=False)
                    return wb, filename
                    
                except (EOFError, IOError, OSError, PermissionError, Exception) as e:
                    error_msg = str(e).lower()
                    print(f"Error loading workbook (attempt {attempt + 1}/{max_retries}): {e}")
                    
                    # Check if file is locked by another process
                    if 'being used by another process' in error_msg or 'permission denied' in error_msg:
                        if attempt < max_retries - 1:
                            # Wait longer for locked files
                            wait_time = 0.5 * (2 ** attempt)
                            print(f"File is locked, waiting {wait_time:.2f} seconds before retry...")
                            time.sleep(wait_time)
                            continue
                    
                    if attempt < max_retries - 1:
                        # Wait a bit before retrying (exponential backoff)
                        time.sleep(0.2 * (2 ** attempt))
                        continue
                    else:
                        # Last attempt failed - file might be corrupted
                        print(f"Failed to load workbook after {max_retries} attempts. File may be corrupted.")
                        
                        # Create backup of corrupted file
                        try:
                            backup_filename = f"{filename}.corrupted_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                            # Try multiple times to create backup
                            for backup_attempt in range(3):
                                try:
                                    shutil.copy2(filename, backup_filename)
                                    print(f"Created backup of corrupted file: {backup_filename}")
                                    break
                                except Exception as backup_error:
                                    if backup_attempt < 2:
                                        time.sleep(0.3)
                                        continue
                                    else:
                                        print(f"Could not create backup after 3 attempts: {backup_error}")
                        except Exception as backup_error:
                            print(f"Could not create backup: {backup_error}")
                        
                        # Create a new workbook
                        print(f"Creating new workbook to replace corrupted file: {filename}")
                        try:
                            # Try multiple times to remove the corrupted file
                            for remove_attempt in range(3):
                                try:
                                    os.remove(filename)
                                    break
                                except Exception as remove_error:
                                    if remove_attempt < 2:
                                        time.sleep(0.3)
                                        continue
                                    else:
                                        print(f"Could not remove corrupted file after 3 attempts: {remove_error}")
                                        # Continue anyway - we'll overwrite it
                        except Exception as remove_error:
                            print(f"Could not remove corrupted file: {remove_error}")
                        
                        # Create new workbook and save it immediately to replace corrupted file
                        wb = Workbook()
                        if 'Sheet' in wb.sheetnames:
                            wb.remove(wb['Sheet'])
                        # Save the new workbook immediately
                        try:
                            wb.save(filename)
                            time.sleep(0.1)  # Wait a bit after saving
                        except Exception as save_error:
                            print(f"Warning: Could not save new workbook immediately: {save_error}")
                        return wb, filename
        else:
            # File doesn't exist, create new workbook
            wb = Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            return wb, filename

def get_or_create_worksheet(wb, date_str=None):
    """Get existing worksheet or create new one for a given date (one worksheet per day)"""
    if date_str is None:
        date_str = get_current_date_string()
    
    worksheet_name = date_str
    
    # Check if worksheet exists
    if worksheet_name in wb.sheetnames:
        ws = wb[worksheet_name]
        # Check if worksheet has reached max rows (very unlikely for one day, but handle it)
        if ws.max_row >= MAX_ROWS_PER_WORKSHEET:
            # Create new worksheet with timestamp for same day
            worksheet_name = f"{date_str}_{datetime.now().strftime('%H%M%S')}"
            ws = wb.create_sheet(worksheet_name)
            initialize_worksheet(ws)
            return ws, False
        return ws, False
    else:
        # Create new worksheet for this day
        ws = wb.create_sheet(worksheet_name)
        initialize_worksheet(ws)
        return ws, False

def check_if_suspicious_activity(ws, user_id, activity_type, request_url, ip_address, response_status, request_data):
    """
    Check if the current activity is suspicious by analyzing activities in the last 1 minute
    Returns True if suspicious (repeated more than 3 times in 1 minute), False otherwise
    """
    try:
        from collections import defaultdict
        from datetime import datetime, timedelta
        
        # Main activities that are not suspicious
        main_activities = [
            'login', 'logout', 'report_generation', 'generate', 'download', 
            'export', 'create_report', 'audit', 'certificate', 'excel'
        ]
        
        activity_type_lower = (activity_type or '').lower()
        request_url_lower = (request_url or '').lower()
        
        # Check if it's a main activity (not suspicious)
        is_main_activity = any(main_act in activity_type_lower or main_act in request_url_lower 
                              for main_act in main_activities)
        if is_main_activity:
            return False
        
        # Get current time
        current_time = datetime.now()
        one_minute_ago = current_time - timedelta(minutes=1)
        
        # Track activities within the last 1 minute only
        activity_counts = defaultdict(int)
        
        # Start from row 2 (skip header) and check all rows
        start_row = 2
        
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, values_only=False):
            if row[0].value is None:  # Skip empty rows
                continue
            
            try:
                # Get timestamp from row (column 0)
                row_timestamp_str = str(row[0].value) if row[0].value else ''
                if not row_timestamp_str:
                    continue
                
                # Parse timestamp
                row_timestamp = datetime.strptime(row_timestamp_str, '%Y-%m-%d %H:%M:%S')
                
                # Only consider activities within the last 1 minute
                if row_timestamp < one_minute_ago:
                    continue  # Skip activities older than 1 minute
                
                # Get activity data from row (columns: Timestamp, User ID, Username, Employee Name, Department, IP, Activity Type, Description, Method, URL, ...)
                row_user_id = str(row[1].value) if row[1].value else ''
                row_activity_type = (str(row[6].value) or '').lower() if row[6].value else ''
                row_request_url = (str(row[9].value) or '') if row[9].value else ''
                
                # Create key for tracking (user_id, activity_type, request_url)
                activity_key = (row_user_id, row_activity_type, row_request_url)
                
                # Count activities within last 1 minute
                activity_counts[activity_key] += 1
            except Exception as e:
                # Skip rows with invalid timestamps or data
                continue
        
        # Check for repeated activities in the last 1 minute
        # Same user, same activity, same URL - if 7+ times in 1 minute, mark as suspicious
        user_id_str = str(user_id) if user_id else ''
        activity_key = (user_id_str, activity_type_lower, request_url or '')
        
        # If we see 6+ occurrences in the last 1 minute, this current one would be the 7th, making it suspicious
        if activity_counts[activity_key] >= 6:
            return True
        
        # Check for failed login attempts - if 2+ back-to-back failed attempts (any time), mark as suspicious
        # Reset count after successful login
        if 'login' in activity_type_lower or 'login' in (request_url or '').lower() or 'failed_attempt' in activity_type_lower:
            # Check for failed login indicators
            failed_login_indicators = ['failed', 'fail', 'error', 'invalid', 'unauthorized', '401', '403', 'failed_attempt']
            activity_desc_lower = (activity_type or '').lower()
            request_data_str = str(request_data or '').lower()
            
            # Check if this is a failed login attempt
            is_failed_login = (
                'failed_attempt' in activity_type_lower or
                any(indicator in activity_desc_lower or indicator in request_data_str 
                   for indicator in failed_login_indicators) or
                (response_status and str(response_status) in ['401', '403'])
            )
            
            if is_failed_login:
                # Count consecutive failed login attempts (back-to-back, not time-based)
                # Look backwards through rows to find the last successful login or count consecutive failures
                # Only count for the same user
                failed_login_count = 0
                user_id_str = str(user_id) if user_id else ''
                
                # Get all rows in reverse order (most recent first)
                all_rows = list(ws.iter_rows(min_row=start_row, max_row=ws.max_row, values_only=False))
                all_rows.reverse()  # Start from most recent
                
                # Find current row index (will be added after max_row, so we check existing rows)
                for row_idx, row in enumerate(all_rows):
                    if row[0].value is None:
                        continue
                    try:
                        # Check if this row is for the same user
                        row_user_id = str(row[1].value) if row[1].value else ''
                        if row_user_id != user_id_str:
                            continue  # Skip rows from different users
                        
                        row_activity_type = (str(row[6].value) or '').lower() if row[6].value else ''
                        row_activity_desc = (str(row[7].value) or '').lower() if row[7].value else ''
                        row_request_data = (str(row[10].value) or '').lower() if row[10].value else ''
                        row_response_status = str(row[11].value) if row[11].value else ''
                        row_request_url = (str(row[9].value) or '').lower() if row[9].value else ''
                        
                        # Check if this is a login-related activity
                        if ('login' in row_activity_type or 'login' in row_request_url or 
                            'failed_attempt' in row_activity_type):
                            
                            # Check if it's a successful login (reset counter)
                            row_is_success = (
                                'login' in row_activity_type and 
                                'failed' not in row_activity_type and
                                'failed_attempt' not in row_activity_type and
                                not any(indicator in row_activity_desc for indicator in ['failed', 'fail', 'error', 'invalid']) and
                                row_response_status not in ['401', '403']
                            )
                            
                            if row_is_success:
                                # Found successful login, stop counting (counter resets)
                                break
                            
                            # Check if it's a failed login
                            row_is_failed = (
                                'failed_attempt' in row_activity_type or
                                any(indicator in row_activity_desc or indicator in row_request_data 
                                   for indicator in failed_login_indicators) or
                                row_response_status in ['401', '403']
                            )
                            
                            if row_is_failed:
                                failed_login_count += 1
                            else:
                                # Not a login activity or not failed, stop counting
                                break
                        else:
                            # Not a login-related activity, stop counting
                            break
                    except:
                        continue
                
                # If we see 1+ consecutive failed login attempts, this current one would be the 2nd, making it suspicious
                if failed_login_count >= 1:
                    return True
        
        # Check for error responses (4xx, 5xx) - these are always suspicious
        if response_status:
            try:
                status_code = int(response_status)
                if status_code >= 400:
                    return True
            except:
                pass
        
        # Check for suspicious request patterns - these are always suspicious
        suspicious_patterns = [
            'sql', 'script', 'union', 'select', 'drop', 'delete', 'insert', 'update',
            'exec', 'eval', 'cmd', 'shell', 'system', 'passwd', 'password', 'admin',
            'test', 'debug', 'trace', 'backup', 'config', 'env', '.env'
        ]
        
        if request_data:
            request_data_str = str(request_data).lower()
            for pattern in suspicious_patterns:
                if pattern in request_data_str:
                    return True
        
        return False
        
    except Exception as e:
        print(f"Error checking suspicious activity: {e}")
        import traceback
        traceback.print_exc()
        return False

def initialize_worksheet(ws):
    """Initialize worksheet with headers"""
    headers = [
        'Timestamp',
        'User ID',
        'Username',
        'Employee Name',
        'Department',
        'IP Address',
        'Activity Type',
        'Activity Description',
        'Request Method',
        'Request URL',
        'Request Data',
        'Response Status',
        'Session ID',
        'User Agent',
        'Additional Details'
    ]
    
    # Set headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Set column widths
    column_widths = {
        'A': 20,  # Timestamp
        'B': 10,  # User ID
        'C': 15,  # Username
        'D': 20,  # Employee Name
        'E': 15,  # Department
        'F': 18,  # IP Address
        'G': 20,  # Activity Type
        'H': 40,  # Activity Description
        'I': 12,  # Request Method
        'J': 50,  # Request URL
        'K': 50,  # Request Data
        'L': 12,  # Response Status
        'M': 30,  # Session ID
        'N': 50,  # User Agent
        'O': 50   # Additional Details
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'

def log_user_activity(
    user_id=None,
    username=None,
    employee_name=None,
    department=None,
    ip_address=None,
    activity_type=None,
    activity_description=None,
    request_method=None,
    request_url=None,
    request_data=None,
    response_status=None,
    session_id=None,
    user_agent=None,
    additional_details=None
):
    """
    Log user activity to Excel file
    
    Args:
        user_id: User ID
        username: Username
        employee_name: Employee name
        department: Department
        ip_address: IP address
        activity_type: Type of activity (login, logout, data_access, etc.)
        activity_description: Description of the activity
        request_method: HTTP method (GET, POST, etc.)
        request_url: Request URL
        request_data: Request data (JSON string or dict)
        response_status: Response status code
        session_id: Session ID
        user_agent: User agent string
        additional_details: Additional details (JSON string or dict)
    """
    try:
        ensure_logs_directory()
        
        # Get current date and month
        current_date = get_current_date_string()
        current_month = get_current_month_string()
        current_datetime = get_current_datetime_string()
        
        # Prepare data
        log_data = [
            current_datetime,  # Timestamp
            user_id or '',
            username or '',
            employee_name or '',
            department or '',
            ip_address or '',
            activity_type or '',
            activity_description or '',
            request_method or '',
            request_url or '',
            json.dumps(request_data) if isinstance(request_data, dict) else (request_data or ''),
            response_status or '',
            session_id or '',
            user_agent or '',
            json.dumps(additional_details) if isinstance(additional_details, dict) else (additional_details or '')
        ]
        
        # Get file lock for the log file
        filename = get_excel_filename(current_month)
        file_lock = get_file_lock(filename)
        
        # Acquire lock before accessing the file
        with file_lock:
            # Get or create workbook for current month with retry logic
            max_save_retries = 3
            for save_attempt in range(max_save_retries):
                wb = None
                try:
                    # Pass the lock to avoid double-locking
                    wb, filename = get_or_create_workbook(current_month, file_lock=file_lock)
                    
                    # Get or create worksheet for current date (one worksheet per day)
                    ws, _ = get_or_create_worksheet(wb, current_date)
                    
                    # Find next row
                    next_row = ws.max_row + 1
                    
                    # Check if this activity is suspicious
                    is_suspicious = check_if_suspicious_activity(
                        ws, user_id, activity_type, request_url, ip_address, response_status, request_data
                    )
                    
                    # Define cell fill colors
                    if is_suspicious:
                        cell_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # Light red background
                    else:
                        cell_fill = None  # Normal (no fill)
                    
                    # Write data
                    for col_num, value in enumerate(log_data, 1):
                        cell = ws.cell(row=next_row, column=col_num, value=value)
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        # Apply red background if suspicious
                        if is_suspicious:
                            cell.fill = cell_fill
                    
                    # Save workbook with error handling
                    try:
                        # Save the workbook
                        wb.save(filename)
                        
                        # Small delay to ensure file is fully written to disk
                        time.sleep(0.05)
                        
                        # Ensure workbook is properly closed
                        try:
                            wb.close()
                            wb = None  # Mark as closed
                        except:
                            pass
                        
                        # Additional small delay after closing to ensure file handle is released
                        time.sleep(0.05)
                        
                        return True
                        
                    except (IOError, OSError, PermissionError) as save_error:
                        error_msg = str(save_error).lower()
                        print(f"Error saving workbook (attempt {save_attempt + 1}/{max_save_retries}): {save_error}")
                        
                        # Close workbook before retry
                        if wb:
                            try:
                                wb.close()
                                wb = None
                            except:
                                pass
                        
                        if save_attempt < max_save_retries - 1:
                            # Wait before retrying (longer for permission errors)
                            if 'permission' in error_msg or 'being used' in error_msg:
                                wait_time = 0.5 * (2 ** save_attempt)
                            else:
                                wait_time = 0.2 * (2 ** save_attempt)
                            time.sleep(wait_time)
                            continue
                        else:
                            # Last attempt failed
                            print(f"Failed to save workbook after {max_save_retries} attempts")
                            raise
                            
                except Exception as wb_error:
                    print(f"Error in workbook operation (attempt {save_attempt + 1}/{max_save_retries}): {wb_error}")
                    
                    # Ensure workbook is closed
                    if wb:
                        try:
                            wb.close()
                            wb = None
                        except:
                            pass
                    
                    if save_attempt < max_save_retries - 1:
                        # Wait before retrying
                        time.sleep(0.2 * (2 ** save_attempt))
                        continue
                    else:
                        # Last attempt failed
                        print(f"Failed to log activity after {max_save_retries} attempts")
                        raise
        
    except Exception as e:
        print(f"Error logging user activity: {e}")
        import traceback
        traceback.print_exc()
        return False

def get_logs_from_excel(start_date=None, end_date=None, limit=None):
    """
    Read logs from Excel files
    
    Args:
        start_date: Start date (YYYY-MM-DD) or datetime
        end_date: End date (YYYY-MM-DD) or datetime
        limit: Maximum number of logs to return
    
    Returns:
        List of log dictionaries
    """
    try:
        ensure_logs_directory()
        logs = []
        
        # Convert dates to datetime if strings
        if isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
        if isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
        
        # Get all Excel files in logs directory (monthly files like Dec_2025.xlsx)
        excel_files = [f for f in os.listdir(LOGS_DIR) if f.endswith('.xlsx') and not f.startswith('~$')]
        # Sort by month/year (most recent first)
        excel_files.sort(reverse=True)
        
        for excel_file in excel_files:
            file_path = os.path.join(LOGS_DIR, excel_file)
            
            # Get file lock for reading (prevents conflicts with concurrent writes)
            file_lock = get_file_lock(file_path)
            
            # Acquire lock before reading the file
            with file_lock:
                try:
                    # Check if file is accessible and not empty
                    if os.path.getsize(file_path) == 0:
                        print(f"Skipping empty file: {excel_file}")
                        continue
                    
                    # Try to load workbook with error handling and retry logic
                    max_read_retries = 3
                    wb = None
                    for read_attempt in range(max_read_retries):
                        try:
                            wb = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
                            break  # Success, exit retry loop
                        except (EOFError, IOError, OSError, PermissionError) as load_error:
                            error_msg = str(load_error).lower()
                            print(f"Error loading Excel file {excel_file} (attempt {read_attempt + 1}/{max_read_retries}): {load_error}")
                            
                            # Check if file is locked by another process
                            if 'being used by another process' in error_msg or 'permission denied' in error_msg:
                                if read_attempt < max_read_retries - 1:
                                    # Wait longer for locked files
                                    wait_time = 0.5 * (2 ** read_attempt)
                                    print(f"File is locked, waiting {wait_time:.2f} seconds before retry...")
                                    time.sleep(wait_time)
                                    continue
                            
                            if read_attempt < max_read_retries - 1:
                                # Wait a bit before retrying (exponential backoff)
                                time.sleep(0.2 * (2 ** read_attempt))
                                continue
                            else:
                                # Last attempt failed - file may be corrupted or locked
                                print(f"Failed to load workbook after {max_read_retries} attempts. Skipping file.")
                                # Try to create backup and skip this file
                                try:
                                    backup_filename = f"{file_path}.corrupted_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                                    shutil.copy2(file_path, backup_filename)
                                    print(f"Created backup of corrupted file: {backup_filename}")
                                except:
                                    pass
                                break  # Skip to next file
                    
                    if wb is None:
                        continue  # Skip this file if we couldn't load it
                    
                    # Iterate through all worksheets
                    for sheet_name in wb.sheetnames:
                        try:
                            ws = wb[sheet_name]
                            
                            # Skip header row
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                if len(row) < 1 or not row[0]:
                                    continue
                                
                                try:
                                    log_timestamp = datetime.strptime(str(row[0]), '%Y-%m-%d %H:%M:%S')
                                    
                                    # Filter by date range
                                    if start_date and log_timestamp < start_date:
                                        continue
                                    if end_date and log_timestamp > end_date:
                                        continue
                                    
                                    log_entry = {
                                        'timestamp': row[0],
                                        'user_id': row[1] if len(row) > 1 else '',
                                        'username': row[2] if len(row) > 2 else '',
                                        'employee_name': row[3] if len(row) > 3 else '',
                                        'department': row[4] if len(row) > 4 else '',
                                        'ip_address': row[5] if len(row) > 5 else '',
                                        'activity_type': row[6] if len(row) > 6 else '',
                                        'activity_description': row[7] if len(row) > 7 else '',
                                        'request_method': row[8] if len(row) > 8 else '',
                                        'request_url': row[9] if len(row) > 9 else '',
                                        'request_data': row[10] if len(row) > 10 else '',
                                        'response_status': row[11] if len(row) > 11 else '',
                                        'session_id': row[12] if len(row) > 12 else '',
                                        'user_agent': row[13] if len(row) > 13 else '',
                                        'additional_details': row[14] if len(row) > 14 else ''
                                    }
                                    
                                    logs.append(log_entry)
                                    
                                    # Check limit
                                    if limit and len(logs) >= limit:
                                        wb.close()
                                        return logs
                                        
                                except Exception as e:
                                    print(f"Error parsing log row: {e}")
                                    continue
                        except Exception as sheet_error:
                            print(f"Error reading worksheet {sheet_name} from {excel_file}: {sheet_error}")
                            continue
                    
                    wb.close()
                    
                except Exception as e:
                    print(f"Error reading Excel file {excel_file}: {e}")
                    continue
        
        return logs
        
    except Exception as e:
        print(f"Error getting logs: {e}")
        import traceback
        traceback.print_exc()
        return []

