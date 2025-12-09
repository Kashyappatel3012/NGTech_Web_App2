import os
import re
import zipfile
import tempfile
import calendar
from datetime import datetime, date
from collections import defaultdict

from flask import Blueprint, current_app, flash, redirect, request, send_from_directory, url_for, abort, jsonify, send_file
from flask_login import login_required, current_user
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

admin_dashboard_bp = Blueprint('admin_dashboard_bp', __name__)

MIN_TRACKER_DATE = date(2025, 10, 1)


@admin_dashboard_bp.route('/admin/download_daily_activity_tracker', methods=['POST'])
@login_required
def download_daily_activity_tracker():
    if current_user.department != "Admin":
        abort(403)

    tracker_date_str = request.form.get('tracker_date')
    if not tracker_date_str:
        flash('Please select a date for the activity tracker.', 'error')
        return redirect(url_for('admin_dashboard'))

    try:
        tracker_date = datetime.strptime(tracker_date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date format. Please pick a valid date.', 'error')
        return redirect(url_for('admin_dashboard'))

    today = datetime.now().date()
    if tracker_date < MIN_TRACKER_DATE or tracker_date > today:
        flash('Please choose a date between 01 Oct 2025 and today.', 'error')
        return redirect(url_for('admin_dashboard'))

    filename = f"{tracker_date.day}_{tracker_date.strftime('%b')}_{tracker_date.year}.xlsx"
    tracker_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Everyday_Updated_Work')
    file_path = os.path.join(tracker_dir, filename)

    if not os.path.exists(file_path):
        formatted_date = tracker_date.strftime('%d %b %Y')
        flash(f'No activity tracker found for {formatted_date}.', 'error')
        return redirect(url_for('admin_dashboard'))

    return send_from_directory(tracker_dir, filename, as_attachment=True)


@admin_dashboard_bp.route('/admin/get_available_tracker_dates', methods=['GET'])
@login_required
def get_available_tracker_dates():
    """Get list of dates for which activity tracker files exist"""
    if current_user.department != "Admin":
        abort(403)
    
    try:
        tracker_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Everyday_Updated_Work')
        
        if not os.path.exists(tracker_dir):
            return jsonify({'success': True, 'dates': []})
        
        available_dates = []
        # Pattern to match: DD_Mon_YYYY.xlsx (e.g., 28_Oct_2025.xlsx)
        date_pattern = re.compile(r'^(\d{1,2})_([A-Za-z]{3})_(\d{4})\.xlsx$')
        
        month_map = {
            'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
            'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
        }
        
        for filename in os.listdir(tracker_dir):
            if filename.endswith('.xlsx'):
                match = date_pattern.match(filename)
                if match:
                    try:
                        day = int(match.group(1))
                        month_str = match.group(2)
                        year = int(match.group(3))
                        
                        if month_str in month_map:
                            month = month_map[month_str]
                            file_date = date(year, month, day)
                            
                            # Only include dates from Oct 1, 2025 onwards and not in the future
                            today = datetime.now().date()
                            if file_date >= MIN_TRACKER_DATE and file_date <= today:
                                available_dates.append({
                                    'date': file_date.isoformat(),
                                    'display': file_date.strftime('%d %b %Y'),
                                    'filename': filename
                                })
                    except (ValueError, KeyError):
                        # Skip invalid date formats
                        continue
        
        # Sort dates in descending order (newest first)
        available_dates.sort(key=lambda x: x['date'], reverse=True)
        
        return jsonify({'success': True, 'dates': available_dates})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_dashboard_bp.route('/admin/download_time_range_tracker', methods=['POST'])
@login_required
def download_time_range_tracker():
    """Download all activity tracker files between start_date and end_date as a zip file"""
    if current_user.department != "Admin":
        abort(403)

    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')
    
    if not start_date_str or not end_date_str:
        flash('Please select both start date and end date.', 'error')
        return redirect(url_for('admin_dashboard'))

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date format. Please select valid dates.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Validate date range
    if start_date > end_date:
        flash('Start date must be before or equal to end date.', 'error')
        return redirect(url_for('admin_dashboard'))

    today = datetime.now().date()
    if start_date < MIN_TRACKER_DATE or end_date > today:
        flash(f'Please choose dates between {MIN_TRACKER_DATE.strftime("%d %b %Y")} and today.', 'error')
        return redirect(url_for('admin_dashboard'))

    tracker_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Everyday_Updated_Work')
    
    if not os.path.exists(tracker_dir):
        flash('Activity tracker directory not found.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Pattern to match: DD_Mon_YYYY.xlsx (e.g., 28_Oct_2025.xlsx)
    date_pattern = re.compile(r'^(\d{1,2})_([A-Za-z]{3})_(\d{4})\.xlsx$')
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
    }

    # Find all files within the date range
    files_to_zip = []
    for filename in os.listdir(tracker_dir):
        if filename.endswith('.xlsx'):
            match = date_pattern.match(filename)
            if match:
                try:
                    day = int(match.group(1))
                    month_str = match.group(2)
                    year = int(match.group(3))
                    
                    if month_str in month_map:
                        month = month_map[month_str]
                        file_date = date(year, month, day)
                        
                        # Check if file date is within the range
                        if start_date <= file_date <= end_date:
                            file_path = os.path.join(tracker_dir, filename)
                            if os.path.exists(file_path):
                                files_to_zip.append((file_path, filename))
                except (ValueError, KeyError):
                    # Skip invalid date formats
                    continue

    if not files_to_zip:
        flash(f'No activity tracker files found between {start_date.strftime("%d %b %Y")} and {end_date.strftime("%d %b %Y")}.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Create a temporary zip file
    try:
        temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        temp_zip_path = temp_zip.name
        temp_zip.close()

        # Create zip file with all matching files
        with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path, filename in files_to_zip:
                zipf.write(file_path, filename)

        # Generate zip filename
        zip_filename = f"Activity_Tracker_{start_date.strftime('%d_%b_%Y')}_to_{end_date.strftime('%d_%b_%Y')}.zip"

        # Send the zip file
        return send_file(
            temp_zip_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name=zip_filename
        )
    except Exception as e:
        flash(f'Error creating zip file: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


@admin_dashboard_bp.route('/admin/generate_attendance_record', methods=['POST'])
@login_required
def generate_attendance_record():
    """Generate attendance record Excel file for selected month and year"""
    if current_user.department != "Admin":
        abort(403)

    month_str = request.form.get('month')
    year_str = request.form.get('year')

    if not month_str or not year_str:
        flash('Please select both month and year.', 'error')
        return redirect(url_for('admin_dashboard'))

    try:
        month = int(month_str)
        year = int(year_str)
        
        if month < 1 or month > 12:
            flash('Invalid month selected.', 'error')
            return redirect(url_for('admin_dashboard'))
            
        if year < 2020 or year > 2030:
            flash('Invalid year selected.', 'error')
            return redirect(url_for('admin_dashboard'))
    except ValueError:
        flash('Invalid month or year format.', 'error')
        return redirect(url_for('admin_dashboard'))

    tracker_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Everyday_Updated_Work')
    
    if not os.path.exists(tracker_dir):
        flash('Activity tracker directory not found.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Pattern to match: DD_Mon_YYYY.xlsx (e.g., 28_Oct_2025.xlsx)
    date_pattern = re.compile(r'^(\d{1,2})_([A-Za-z]{3})_(\d{4})\.xlsx$')
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
    }

    # Find all files matching the selected month and year
    matching_files = []
    for filename in os.listdir(tracker_dir):
        if filename.endswith('.xlsx'):
            match = date_pattern.match(filename)
            if match:
                try:
                    day = int(match.group(1))
                    month_str_file = match.group(2)
                    year_file = int(match.group(3))
                    
                    if month_str_file in month_map:
                        month_file = month_map[month_str_file]
                        
                        # Check if file matches selected month and year
                        if month_file == month and year_file == year:
                            file_path = os.path.join(tracker_dir, filename)
                            if os.path.exists(file_path):
                                matching_files.append(file_path)
                except (ValueError, KeyError):
                    continue

    if not matching_files:
        month_name = calendar.month_name[month]
        flash(f'No activity tracker files found for {month_name} {year}.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Process all matching files and calculate attendance
    employee_attendance = defaultdict(float)  # {employee_name: total_working_days}
    
    for file_path in matching_files:
        try:
            workbook = load_workbook(file_path, data_only=True)
            
            # Check if "Updated Work" worksheet exists
            if 'Updated Work' not in workbook.sheetnames:
                continue
            
            worksheet = workbook['Updated Work']
            
            # Find header row (assuming headers are in row 1)
            header_row = 1
            employee_name_col = None
            total_time_col = None
            
            # Find column indices for Employee Name (Column A) and Total Time (Column D)
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip().lower()
                    if 'employee name' in cell_str or col_idx == 1:
                        employee_name_col = col_idx
                    if 'total time' in cell_str or col_idx == 4:
                        total_time_col = col_idx
            
            # Default to Column A (1) and Column D (4) if not found
            if employee_name_col is None:
                employee_name_col = 1
            if total_time_col is None:
                total_time_col = 4
            
            # Process data rows (starting from row 2)
            for row_idx in range(header_row + 1, worksheet.max_row + 1):
                employee_name_cell = worksheet.cell(row=row_idx, column=employee_name_col)
                total_time_cell = worksheet.cell(row=row_idx, column=total_time_col)
                
                employee_name = employee_name_cell.value
                total_time_value = total_time_cell.value
                
                # Skip empty rows
                if not employee_name:
                    continue
                
                employee_name = str(employee_name).strip()
                
                # Convert total_time to float
                try:
                    if total_time_value is None:
                        total_time = 0.0
                    else:
                        total_time = float(total_time_value)
                except (ValueError, TypeError):
                    total_time = 0.0
                
                # Calculate working day based on Total Time
                # > 6 hours = 1 working day
                # 3-6 hours = 0.5 working day
                # < 3 hours = 0 working day
                if total_time > 6:
                    working_day = 1.0
                elif total_time >= 3:
                    working_day = 0.5
                else:
                    working_day = 0.0
                
                employee_attendance[employee_name] += working_day
                
        except Exception as e:
            print(f"Error processing file {file_path}: {str(e)}")
            continue

    if not employee_attendance:
        flash('No attendance data found in the files.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Get total days in the selected month
    total_days_in_month = calendar.monthrange(year, month)[1]

    # Create Excel workbook with attendance summary
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Record"
        
        # Header row
        headers = ['Sr. No', 'Employee Name', 'Working Day', 'Non-Working Day']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        row_num = 2
        for sr_no, (employee_name, working_days) in enumerate(sorted(employee_attendance.items()), start=1):
            non_working_days = total_days_in_month - working_days
            
            ws.cell(row=row_num, column=1, value=sr_no).border = border
            ws.cell(row=row_num, column=2, value=employee_name).border = border
            ws.cell(row=row_num, column=3, value=working_days).border = border
            ws.cell(row=row_num, column=4, value=non_working_days).border = border
            
            # Center align Sr. No and numeric columns
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')
            
            row_num += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        
        # Save to temporary file
        month_name = calendar.month_name[month]
        temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_excel_path = temp_excel.name
        temp_excel.close()
        
        wb.save(temp_excel_path)
        
        # Generate filename
        excel_filename = f"Attendance_Record_{month_name}_{year}.xlsx"
        
        # Send the Excel file
        return send_file(
            temp_excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=excel_filename
        )
        
    except Exception as e:
        flash(f'Error generating attendance record: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


@admin_dashboard_bp.route('/admin/generate_personal_activity_tracker', methods=['POST'])
@login_required
def generate_personal_activity_tracker():
    """Generate personal activity tracker Excel file for selected employee, month and year"""
    if current_user.department != "Admin":
        abort(403)

    employee_name = request.form.get('employee_name', '').strip()
    month_str = request.form.get('month')
    year_str = request.form.get('year')

    if not employee_name:
        flash('Please enter employee name.', 'error')
        return redirect(url_for('admin_dashboard'))

    if not month_str or not year_str:
        flash('Please select both month and year.', 'error')
        return redirect(url_for('admin_dashboard'))

    try:
        month = int(month_str)
        year = int(year_str)
        
        if month < 1 or month > 12:
            flash('Invalid month selected.', 'error')
            return redirect(url_for('admin_dashboard'))
            
        if year < 2020 or year > 2030:
            flash('Invalid year selected.', 'error')
            return redirect(url_for('admin_dashboard'))
    except ValueError:
        flash('Invalid month or year format.', 'error')
        return redirect(url_for('admin_dashboard'))

    tracker_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Everyday_Updated_Work')
    
    if not os.path.exists(tracker_dir):
        flash('Activity tracker directory not found.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Pattern to match: DD_Mon_YYYY.xlsx (e.g., 28_Oct_2025.xlsx)
    date_pattern = re.compile(r'^(\d{1,2})_([A-Za-z]{3})_(\d{4})\.xlsx$')
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
    }

    # Find all files matching the selected month and year
    matching_files = []
    for filename in os.listdir(tracker_dir):
        if filename.endswith('.xlsx'):
            match = date_pattern.match(filename)
            if match:
                try:
                    day = int(match.group(1))
                    month_str_file = match.group(2)
                    year_file = int(match.group(3))
                    
                    if month_str_file in month_map:
                        month_file = month_map[month_str_file]
                        
                        # Check if file matches selected month and year
                        if month_file == month and year_file == year:
                            file_path = os.path.join(tracker_dir, filename)
                            if os.path.exists(file_path):
                                matching_files.append((file_path, filename))
                except (ValueError, KeyError):
                    continue

    if not matching_files:
        month_name = calendar.month_name[month]
        flash(f'No activity tracker files found for {month_name} {year}.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Process all matching files and extract data for the selected employee
    all_rows = []  # List of dictionaries with employee data
    
    for file_path, filename in matching_files:
        try:
            workbook = load_workbook(file_path, data_only=True)
            
            # Check if "Updated Work" worksheet exists
            if 'Updated Work' not in workbook.sheetnames:
                continue
            
            worksheet = workbook['Updated Work']
            
            # Find header row (assuming headers are in row 1)
            header_row = 1
            
            # Find column indices
            employee_name_col = None
            date_col = None
            total_time_col = None
            task_columns = {}  # {task_num: {'task': col, 'client': col, 'time': col}}
            
            # Scan header row to find all columns
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip()
                    cell_lower = cell_str.lower()
                    
                    # Employee Name (Column A or header contains "employee name")
                    if col_idx == 1 or 'employee name' in cell_lower:
                        employee_name_col = col_idx
                    
                    # Date (Column B or header contains "date")
                    if col_idx == 2 or ('date' in cell_lower and 'client' not in cell_lower):
                        date_col = col_idx
                    
                    # Total Time (Column D or header contains "total time")
                    if col_idx == 4 or 'total time' in cell_lower:
                        total_time_col = col_idx
                    
                    # Task columns - pattern: Task1, Task1 Client Name, Time For Task1, etc.
                    task_match = re.match(r'^task(\d+)$', cell_lower)
                    if task_match:
                        task_num = int(task_match.group(1))
                        if task_num not in task_columns:
                            task_columns[task_num] = {}
                        task_columns[task_num]['task'] = col_idx
                    
                    client_match = re.match(r'^task(\d+)\s+client\s+name$', cell_lower)
                    if client_match:
                        task_num = int(client_match.group(1))
                        if task_num not in task_columns:
                            task_columns[task_num] = {}
                        task_columns[task_num]['client'] = col_idx
                    
                    time_match = re.match(r'^time\s+for\s+task(\d+)$', cell_lower)
                    if time_match:
                        task_num = int(time_match.group(1))
                        if task_num not in task_columns:
                            task_columns[task_num] = {}
                        task_columns[task_num]['time'] = col_idx
            
            # Default to Column A, B, D if not found
            if employee_name_col is None:
                employee_name_col = 1
            if date_col is None:
                date_col = 2
            if total_time_col is None:
                total_time_col = 4
            
            # Process data rows (starting from row 2)
            for row_idx in range(header_row + 1, worksheet.max_row + 1):
                employee_name_cell = worksheet.cell(row=row_idx, column=employee_name_col)
                employee_name_value = employee_name_cell.value
                
                # Skip empty rows
                if not employee_name_value:
                    continue
                
                employee_name_value = str(employee_name_value).strip()
                
                # Filter by selected employee name (case-insensitive)
                if employee_name_value.lower() != employee_name.lower():
                    continue
                
                # Extract Date
                date_cell = worksheet.cell(row=row_idx, column=date_col)
                date_value = date_cell.value
                if date_value:
                    if isinstance(date_value, datetime):
                        date_str = date_value.strftime('%d-%m-%Y')
                    else:
                        date_str = str(date_value)
                else:
                    date_str = ''
                
                # Extract Total Time
                total_time_cell = worksheet.cell(row=row_idx, column=total_time_col)
                total_time_value = total_time_cell.value
                try:
                    if total_time_value is None:
                        total_time = ''
                    else:
                        total_time = str(total_time_value)
                except:
                    total_time = ''
                
                # Extract all tasks
                task_list = []
                for task_num in sorted(task_columns.keys()):
                    task_info = task_columns[task_num]
                    task_text = ''
                    client_text = ''
                    time_text = ''
                    
                    if 'task' in task_info:
                        task_cell = worksheet.cell(row=row_idx, column=task_info['task'])
                        if task_cell.value:
                            task_text = str(task_cell.value).strip()
                    
                    if 'client' in task_info:
                        client_cell = worksheet.cell(row=row_idx, column=task_info['client'])
                        if client_cell.value:
                            client_text = str(client_cell.value).strip()
                    
                    if 'time' in task_info:
                        time_cell = worksheet.cell(row=row_idx, column=task_info['time'])
                        if time_cell.value:
                            time_text = str(time_cell.value).strip()
                    
                    # Combine task info: "Task - Task Client Name - Time for Task"
                    if task_text or client_text or time_text:
                        task_combined = ' - '.join(filter(None, [task_text, client_text, time_text]))
                        if task_combined:
                            task_list.append(task_combined)
                
                # Combine all tasks with newlines
                task_combined_str = '\n'.join(task_list) if task_list else ''
                
                # Add row data
                all_rows.append({
                    'employee_name': employee_name_value,
                    'task': task_combined_str,
                    'date': date_str,
                    'total_time': total_time
                })
                
        except Exception as e:
            print(f"Error processing file {file_path}: {str(e)}")
            continue

    if not all_rows:
        flash(f'No data found for employee "{employee_name}" in {calendar.month_name[month]} {year}.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Create Excel workbook with personal activity tracker
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Personal Activity Tracker"
        
        # Header row
        headers = ['Sr. No', 'Employee Name', 'Task', 'Date', 'Total Time']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Center alignment with text wrap for all cells
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Data rows
        row_num = 2
        for sr_no, row_data in enumerate(all_rows, start=1):
            # Create all cells with border
            cell1 = ws.cell(row=row_num, column=1, value=sr_no)
            cell2 = ws.cell(row=row_num, column=2, value=row_data['employee_name'])
            cell3 = ws.cell(row=row_num, column=3, value=row_data['task'])
            cell4 = ws.cell(row=row_num, column=4, value=row_data['date'])
            cell5 = ws.cell(row=row_num, column=5, value=row_data['total_time'])
            
            # Apply border and alignment to all cells
            for cell in [cell1, cell2, cell3, cell4, cell5]:
                cell.border = border
                cell.alignment = center_alignment
            
            row_num += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # Set row heights for better visibility of wrapped text
        for row_idx in range(2, row_num):
            ws.row_dimensions[row_idx].height = 60
        
        # Save to temporary file
        month_name = calendar.month_name[month]
        safe_employee_name = re.sub(r'[^\w\s-]', '', employee_name).strip().replace(' ', '_')
        temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_excel_path = temp_excel.name
        temp_excel.close()
        
        wb.save(temp_excel_path)
        
        # Generate filename
        excel_filename = f"Personal_Activity_Tracker_{safe_employee_name}_{month_name}_{year}.xlsx"
        
        # Send the Excel file
        return send_file(
            temp_excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=excel_filename
        )
        
    except Exception as e:
        flash(f'Error generating personal activity tracker: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


@admin_dashboard_bp.route('/admin/generate_client_wise_activity_tracker', methods=['POST'])
@login_required
def generate_client_wise_activity_tracker():
    """Generate client-wise activity tracker Excel file for selected month and year"""
    if current_user.department != "Admin":
        abort(403)

    month_str = request.form.get('month')
    year_str = request.form.get('year')

    if not month_str or not year_str:
        flash('Please select both month and year.', 'error')
        return redirect(url_for('admin_dashboard'))

    try:
        month = int(month_str)
        year = int(year_str)
        
        if month < 1 or month > 12:
            flash('Invalid month selected.', 'error')
            return redirect(url_for('admin_dashboard'))
            
        if year < 2020 or year > 2030:
            flash('Invalid year selected.', 'error')
            return redirect(url_for('admin_dashboard'))
    except ValueError:
        flash('Invalid month or year format.', 'error')
        return redirect(url_for('admin_dashboard'))

    tracker_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Everyday_Updated_Work')
    
    if not os.path.exists(tracker_dir):
        flash('Activity tracker directory not found.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Pattern to match: DD_Mon_YYYY.xlsx (e.g., 28_Oct_2025.xlsx)
    date_pattern = re.compile(r'^(\d{1,2})_([A-Za-z]{3})_(\d{4})\.xlsx$')
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
    }

    # Find all files matching the selected month and year
    matching_files = []
    for filename in os.listdir(tracker_dir):
        if filename.endswith('.xlsx'):
            match = date_pattern.match(filename)
            if match:
                try:
                    day = int(match.group(1))
                    month_str_file = match.group(2)
                    year_file = int(match.group(3))
                    
                    if month_str_file in month_map:
                        month_file = month_map[month_str_file]
                        
                        # Check if file matches selected month and year
                        if month_file == month and year_file == year:
                            file_path = os.path.join(tracker_dir, filename)
                            if os.path.exists(file_path):
                                matching_files.append((file_path, filename, day, month_str_file, year_file))
                except (ValueError, KeyError):
                    continue

    if not matching_files:
        month_name = calendar.month_name[month]
        flash(f'No activity tracker files found for {month_name} {year}.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Process all matching files and aggregate data by Client Name
    client_data = defaultdict(lambda: {'total_time': 0.0, 'details': [], 'dates': set()})
    
    for file_path, filename, day, month_str_file, year_file in matching_files:
        try:
            workbook = load_workbook(file_path, data_only=True)
            
            # Check if "Client_Time" worksheet exists
            if 'Client_Time' not in workbook.sheetnames:
                continue
            
            worksheet = workbook['Client_Time']
            
            # Format date from filename: 21_Nov_2025 -> 21 Nov 2025
            date_str = f"{day} {month_str_file} {year_file}"
            
            # Find header row (assuming headers are in row 1)
            header_row = 1
            
            # Find column indices
            client_name_col = None
            time_col = None
            details_col = None
            
            # Scan header row to find columns
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip()
                    cell_lower = cell_str.lower()
                    
                    # Client Name (Column B or header contains "client name")
                    if col_idx == 2 or 'client name' in cell_lower:
                        client_name_col = col_idx
                    
                    # Time (Column C or header contains "time")
                    if col_idx == 3 or ('time' in cell_lower and 'total' not in cell_lower):
                        time_col = col_idx
                    
                    # Details/Employees and Task (Column D or header contains "employees" or "task" or "details")
                    if col_idx == 4 or 'employees' in cell_lower or ('task' in cell_lower and 'client' not in cell_lower) or 'details' in cell_lower:
                        details_col = col_idx
            
            # Default to Column B, C, D if not found
            if client_name_col is None:
                client_name_col = 2
            if time_col is None:
                time_col = 3
            if details_col is None:
                details_col = 4
            
            # Process data rows (starting from row 2)
            for row_idx in range(header_row + 1, worksheet.max_row + 1):
                client_name_cell = worksheet.cell(row=row_idx, column=client_name_col)
                client_name_value = client_name_cell.value
                
                # Skip empty rows
                if not client_name_value:
                    continue
                
                client_name_value = str(client_name_value).strip()
                
                # Extract Time
                time_cell = worksheet.cell(row=row_idx, column=time_col)
                time_value = time_cell.value
                try:
                    if time_value is None:
                        time_val = 0.0
                    else:
                        time_val = float(time_value)
                except (ValueError, TypeError):
                    time_val = 0.0
                
                # Extract Details
                details_cell = worksheet.cell(row=row_idx, column=details_col)
                details_value = details_cell.value
                details_str = str(details_value).strip() if details_value else ''
                
                # Append date to details if details exist
                if details_str:
                    details_with_date = f"{details_str} - {date_str}"
                    client_data[client_name_value]['details'].append(details_with_date)
                
                # Aggregate data by client name
                client_data[client_name_value]['total_time'] += time_val
                client_data[client_name_value]['dates'].add(date_str)
                
        except Exception as e:
            print(f"Error processing file {file_path}: {str(e)}")
            continue

    if not client_data:
        flash('No client data found in the files for the selected month and year.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Create Excel workbook with client-wise activity tracker
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Client Wise Activity Tracker"
        
        # Header row
        headers = ['Sr. No', 'Client Name', 'Total Time', 'Date', 'Details']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Center alignment with text wrap for all cells
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Data rows - sort by client name
        row_num = 2
        for sr_no, (client_name, data) in enumerate(sorted(client_data.items()), start=1):
            # Combine all dates
            dates_str = ', '.join(sorted(data['dates']))
            
            # Combine all details with newlines
            details_str = '\n'.join(data['details']) if data['details'] else ''
            
            # Create all cells
            cell1 = ws.cell(row=row_num, column=1, value=sr_no)
            cell2 = ws.cell(row=row_num, column=2, value=client_name)
            cell3 = ws.cell(row=row_num, column=3, value=data['total_time'])
            cell4 = ws.cell(row=row_num, column=4, value=dates_str)
            cell5 = ws.cell(row=row_num, column=5, value=details_str)
            
            # Apply border and alignment to all cells
            for cell in [cell1, cell2, cell3, cell4, cell5]:
                cell.border = border
                cell.alignment = center_alignment
            
            row_num += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 50
        
        # Set row heights for better visibility of wrapped text
        for row_idx in range(2, row_num):
            ws.row_dimensions[row_idx].height = 60
        
        # Save to temporary file
        month_name = calendar.month_name[month]
        temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_excel_path = temp_excel.name
        temp_excel.close()
        
        wb.save(temp_excel_path)
        
        # Generate filename
        excel_filename = f"Client_Wise_Activity_Tracker_{month_name}_{year}.xlsx"
        
        # Send the Excel file
        return send_file(
            temp_excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=excel_filename
        )
        
    except Exception as e:
        flash(f'Error generating client-wise activity tracker: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


@admin_dashboard_bp.route('/admin/download_approved_extra_work', methods=['POST'])
@login_required
def download_approved_extra_work():
    """Download approved extra work file for selected month and year"""
    if current_user.department != "Admin":
        abort(403)

    month_str = request.form.get('month')
    year_str = request.form.get('year')

    if not month_str or not year_str:
        flash('Please select both month and year.', 'error')
        return redirect(url_for('admin_dashboard'))

    try:
        month = int(month_str)
        year = int(year_str)
        
        if month < 1 or month > 12:
            flash('Invalid month selected.', 'error')
            return redirect(url_for('admin_dashboard'))
            
        if year < 2020 or year > 2030:
            flash('Invalid year selected.', 'error')
            return redirect(url_for('admin_dashboard'))
    except ValueError:
        flash('Invalid month or year format.', 'error')
        return redirect(url_for('admin_dashboard'))

    approved_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Approved_Extra_work')
    
    if not os.path.exists(approved_dir):
        flash('Approved extra work directory not found.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Map month number to abbreviation
    month_map = {
        1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
        7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
    }
    
    month_abbr = month_map.get(month)
    if not month_abbr:
        flash('Invalid month selected.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    # Construct filename: Nov_2025.xlsx
    filename = f"{month_abbr}_{year}.xlsx"
    file_path = os.path.join(approved_dir, filename)
    
    if not os.path.exists(file_path):
        month_name = calendar.month_name[month]
        flash(f'No approved extra work file found for {month_name} {year}.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    return send_from_directory(approved_dir, filename, as_attachment=True)


@admin_dashboard_bp.route('/admin/generate_attendance_with_extra_work', methods=['POST'])
@login_required
def generate_attendance_with_extra_work():
    """Generate attendance record with extra work Excel file for selected month and year"""
    if current_user.department != "Admin":
        abort(403)

    month_str = request.form.get('month')
    year_str = request.form.get('year')

    if not month_str or not year_str:
        flash('Please select both month and year.', 'error')
        return redirect(url_for('admin_dashboard'))

    try:
        month = int(month_str)
        year = int(year_str)
        
        if month < 1 or month > 12:
            flash('Invalid month selected.', 'error')
            return redirect(url_for('admin_dashboard'))
            
        if year < 2020 or year > 2030:
            flash('Invalid year selected.', 'error')
            return redirect(url_for('admin_dashboard'))
    except ValueError:
        flash('Invalid month or year format.', 'error')
        return redirect(url_for('admin_dashboard'))

    tracker_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Everyday_Updated_Work')
    approved_extra_work_dir = os.path.join(current_app.root_path, 'static', 'Activity_Tracker', 'Approved_Extra_work')
    
    if not os.path.exists(tracker_dir):
        flash('Activity tracker directory not found.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Pattern to match: DD_Mon_YYYY.xlsx (e.g., 28_Oct_2025.xlsx)
    date_pattern = re.compile(r'^(\d{1,2})_([A-Za-z]{3})_(\d{4})\.xlsx$')
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
    }
    
    month_names = {v: k for k, v in month_map.items()}

    # Find all files matching the selected month and year
    matching_files = []
    for filename in os.listdir(tracker_dir):
        if filename.endswith('.xlsx'):
            match = date_pattern.match(filename)
            if match:
                try:
                    day = int(match.group(1))
                    month_str_file = match.group(2)
                    year_file = int(match.group(3))
                    
                    if month_str_file in month_map:
                        month_file = month_map[month_str_file]
                        
                        # Check if file matches selected month and year
                        if month_file == month and year_file == year:
                            file_path = os.path.join(tracker_dir, filename)
                            if os.path.exists(file_path):
                                matching_files.append(file_path)
                except (ValueError, KeyError):
                    continue

    if not matching_files:
        month_name = calendar.month_name[month]
        flash(f'No activity tracker files found for {month_name} {year}.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Process all matching files and calculate attendance (same logic as attendance record)
    employee_attendance = defaultdict(float)  # {employee_name: total_working_days}
    
    for file_path in matching_files:
        try:
            workbook = load_workbook(file_path, data_only=True)
            
            # Check if "Updated Work" worksheet exists
            if 'Updated Work' not in workbook.sheetnames:
                continue
            
            worksheet = workbook['Updated Work']
            
            # Find header row (assuming headers are in row 1)
            header_row = 1
            employee_name_col = None
            total_time_col = None
            
            # Find column indices for Employee Name (Column A) and Total Time (Column D)
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip().lower()
                    if 'employee name' in cell_str or col_idx == 1:
                        employee_name_col = col_idx
                    if 'total time' in cell_str or col_idx == 4:
                        total_time_col = col_idx
            
            # Default to Column A (1) and Column D (4) if not found
            if employee_name_col is None:
                employee_name_col = 1
            if total_time_col is None:
                total_time_col = 4
            
            # Process data rows (starting from row 2)
            for row_idx in range(header_row + 1, worksheet.max_row + 1):
                employee_name_cell = worksheet.cell(row=row_idx, column=employee_name_col)
                total_time_cell = worksheet.cell(row=row_idx, column=total_time_col)
                
                employee_name = employee_name_cell.value
                total_time_value = total_time_cell.value
                
                # Skip empty rows
                if not employee_name:
                    continue
                
                employee_name = str(employee_name).strip()
                
                # Convert total_time to float
                try:
                    if total_time_value is None:
                        total_time = 0.0
                    else:
                        total_time = float(total_time_value)
                except (ValueError, TypeError):
                    total_time = 0.0
                
                # Calculate working day based on Total Time
                # > 6 hours = 1 working day
                # 3-6 hours = 0.5 working day
                # < 3 hours = 0 working day
                if total_time > 6:
                    working_day = 1.0
                elif total_time >= 3:
                    working_day = 0.5
                else:
                    working_day = 0.0
                
                employee_attendance[employee_name] += working_day
                
        except Exception as e:
            print(f"Error processing file {file_path}: {str(e)}")
            continue

    # Process approved extra work files
    extra_work_attendance = defaultdict(float)  # {employee_name: extra_working_days}
    
    if os.path.exists(approved_extra_work_dir):
        # Pattern to match: Mon_YYYY.xlsx (e.g., Nov_2025.xlsx)
        file_pattern = re.compile(r'^([A-Za-z]{3})_(\d{4})\.xlsx$')
        
        for filename in os.listdir(approved_extra_work_dir):
            if filename.endswith('.xlsx'):
                match = file_pattern.match(filename)
                if match:
                    try:
                        month_str_file = match.group(1)
                        year_file = int(match.group(2))
                        
                        if month_str_file in month_map:
                            month_file = month_map[month_str_file]
                            
                            # Check if file matches selected month and year
                            if month_file == month and year_file == year:
                                file_path = os.path.join(approved_extra_work_dir, filename)
                                
                                try:
                                    workbook = load_workbook(file_path, data_only=True)
                                    
                                    # Check if "Approved Extra Work" worksheet exists
                                    if 'Approved Extra Work' not in workbook.sheetnames:
                                        continue
                                    
                                    worksheet = workbook['Approved Extra Work']
                                    
                                    # Find header row (assuming headers are in row 1)
                                    header_row = 1
                                    employee_name_col = None
                                    time_col = None
                                    
                                    # Find column indices
                                    for col_idx in range(1, worksheet.max_column + 1):
                                        cell_value = worksheet.cell(row=header_row, column=col_idx).value
                                        if cell_value:
                                            cell_str = str(cell_value).strip().lower()
                                            if 'employee name' in cell_str or col_idx == 1:
                                                employee_name_col = col_idx
                                            if 'time' in cell_str or col_idx == 3:
                                                time_col = col_idx
                                    
                                    # Default to Column A (1) and Column C (3) if not found
                                    if employee_name_col is None:
                                        employee_name_col = 1
                                    if time_col is None:
                                        time_col = 3
                                    
                                    # Process data rows (starting from row 2)
                                    for row_idx in range(header_row + 1, worksheet.max_row + 1):
                                        employee_name_cell = worksheet.cell(row=row_idx, column=employee_name_col)
                                        time_cell = worksheet.cell(row=row_idx, column=time_col)
                                        
                                        employee_name = employee_name_cell.value
                                        time_value = time_cell.value
                                        
                                        # Skip empty rows
                                        if not employee_name:
                                            continue
                                        
                                        employee_name = str(employee_name).strip()
                                        
                                        # Convert time to float
                                        try:
                                            if time_value is None:
                                                time_val = 0.0
                                            else:
                                                time_val = float(time_value)
                                        except (ValueError, TypeError):
                                            time_val = 0.0
                                        
                                        # Calculate extra working day based on Time
                                        # > 6 hours = 1 working day
                                        # 3-6 hours = 0.5 working day
                                        # < 3 hours = 0 working day
                                        if time_val > 6:
                                            extra_working_day = 1.0
                                        elif time_val >= 3:
                                            extra_working_day = 0.5
                                        else:
                                            extra_working_day = 0.0
                                        
                                        extra_work_attendance[employee_name] += extra_working_day
                                        
                                except Exception as e:
                                    print(f"Error processing extra work file {file_path}: {str(e)}")
                                    continue
                    except (ValueError, KeyError):
                        continue

    if not employee_attendance and not extra_work_attendance:
        flash('No attendance data found in the files.', 'error')
        return redirect(url_for('admin_dashboard'))

    # Get total days in the selected month
    total_days_in_month = calendar.monthrange(year, month)[1]

    # Combine regular attendance and extra work attendance
    all_employees = set(employee_attendance.keys()) | set(extra_work_attendance.keys())
    
    # Create Excel workbook with attendance summary including extra work
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Record with Extra Work"
        
        # Header row
        headers = ['Sr. No', 'Employee Name', 'Working Day', 'Extra Work Day', 'Total Working Day', 'Non-Working Day']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        row_num = 2
        for sr_no, employee_name in enumerate(sorted(all_employees), start=1):
            regular_working_days = employee_attendance.get(employee_name, 0.0)
            extra_working_days = extra_work_attendance.get(employee_name, 0.0)
            total_working_days = regular_working_days + extra_working_days
            non_working_days = total_days_in_month - total_working_days
            
            ws.cell(row=row_num, column=1, value=sr_no).border = border
            ws.cell(row=row_num, column=2, value=employee_name).border = border
            ws.cell(row=row_num, column=3, value=regular_working_days).border = border
            ws.cell(row=row_num, column=4, value=extra_working_days).border = border
            ws.cell(row=row_num, column=5, value=total_working_days).border = border
            ws.cell(row=row_num, column=6, value=non_working_days).border = border
            
            # Center align Sr. No and numeric columns
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='center')
            
            row_num += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        
        # Save to temporary file
        month_name = calendar.month_name[month]
        temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_excel_path = temp_excel.name
        temp_excel.close()
        
        wb.save(temp_excel_path)
        
        # Generate filename
        excel_filename = f"Attendance_Record_with_Extra_Work_{month_name}_{year}.xlsx"
        
        # Send the Excel file
        return send_file(
            temp_excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=excel_filename
        )
        
    except Exception as e:
        flash(f'Error generating attendance record with extra work: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))


@admin_dashboard_bp.route('/admin/get_all_employees', methods=['GET'])
@login_required
def get_all_employees():
    """Get all employees for dropdown"""
    if current_user.department != "Admin":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Get db instance
        try:
            db = current_app.extensions['sqlalchemy']
        except (KeyError, AttributeError):
            from app import db
        
        from app import User
        users = db.session.query(User).all()
        employees = [{'id': user.id, 'employee_name': user.employee_name} for user in users]
        return jsonify({'success': True, 'employees': employees})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_dashboard_bp.route('/admin/get_active_employees', methods=['GET'])
@login_required
def get_active_employees():
    """Get all active employees with their details"""
    if current_user.department != "Admin":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Get db instance
        try:
            db = current_app.extensions['sqlalchemy']
        except (KeyError, AttributeError):
            from app import db
        
        from app import User, UserStatus
        
        # Query all users and filter by active status
        all_users = db.session.query(User).all()
        
        employees = []
        for user in all_users:
            # Check if user has status and if it's active
            # If no status record, consider user as active (default behavior)
            is_user_active = True
            if user.status:
                is_user_active = user.status.is_active
            
            # Only include if active
            if is_user_active:
                # Get department/team from user.department
                # Employee ID is the username
                employees.append({
                    'employee_name': user.employee_name,
                    'employee_id': user.username,
                    'department_team': user.department
                })
        
        # Sort by employee name
        employees.sort(key=lambda x: x['employee_name'])
        
        return jsonify({'success': True, 'employees': employees})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_dashboard_bp.route('/admin/get_all_employees_past_record', methods=['GET'])
@login_required
def get_all_employees_past_record():
    """Get all employees (active and deleted) with their creation and deletion times"""
    if current_user.department != "Admin":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Get db instance
        try:
            db = current_app.extensions['sqlalchemy']
        except (KeyError, AttributeError):
            from app import db
        
        from app import User, UserStatus
        
        # Query all users (including deleted ones - they're soft deleted now)
        all_users = db.session.query(User).all()
        
        employees = []
        for user in all_users:
            # Check if user is active or deleted
            is_user_active = True
            if user.status:
                is_user_active = user.status.is_active
            
            # Get creation time - use created_at if available, otherwise use EmployeeData.created_at as fallback
            creation_time = None
            if hasattr(user, 'created_at') and user.created_at:
                creation_time = user.created_at.strftime('%d-%m-%Y %H:%M:%S')
            elif user.employee_data and hasattr(user.employee_data, 'created_at') and user.employee_data.created_at:
                creation_time = user.employee_data.created_at.strftime('%d-%m-%Y %H:%M:%S')
            else:
                creation_time = "N/A"
            
            # Get deletion time
            deletion_time = "-"
            if hasattr(user, 'deleted_at') and user.deleted_at:
                deletion_time = user.deleted_at.strftime('%d-%m-%Y %H:%M:%S')
            
            # Determine status
            status = "Activate"
            if not is_user_active or (hasattr(user, 'deleted_at') and user.deleted_at):
                status = "Deactivate"
            
            employees.append({
                'employee_name': user.employee_name,
                'employee_id': user.username,
                'department_team': user.department,
                'creation_time': creation_time,
                'deletion_time': deletion_time,
                'status': status
            })
        
        # Sort by creation time (newest first), then by employee name
        employees.sort(key=lambda x: (
            x['creation_time'] != "N/A",
            x['creation_time'] if x['creation_time'] != "N/A" else "",
            x['employee_name']
        ), reverse=True)
        
        return jsonify({'success': True, 'employees': employees})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_dashboard_bp.route('/admin/get_all_employees_for_performance', methods=['GET'])
@login_required
def get_all_employees_for_performance():
    """Get all active employees for performance rating modals"""
    if current_user.department != "Admin":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Get db instance
        try:
            db = current_app.extensions['sqlalchemy']
        except (KeyError, AttributeError):
            from app import db
        
        from app import User, UserStatus
        
        # Query all active users (not deleted and active)
        all_users = db.session.query(User).all()
        
        employees = []
        for user in all_users:
            # Check if user is deleted
            is_deleted = False
            if hasattr(user, 'deleted_at') and user.deleted_at:
                is_deleted = True
            
            # Check if user is inactive
            is_inactive = False
            if user.status and not user.status.is_active:
                is_inactive = True
            
            # Only include users that are not deleted and are active
            if not is_deleted and not is_inactive:
                employees.append({
                    'id': user.id,
                    'employee_name': user.employee_name
                })
        
        # Sort by employee name
        employees.sort(key=lambda x: x['employee_name'])
        
        return jsonify({'success': True, 'employees': employees})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_dashboard_bp.route('/admin/save_performance_data', methods=['POST'])
@login_required
def save_performance_data():
    """Save performance data for all employees (Technical Skills or Client Satisfaction)"""
    if current_user.department != "Admin":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from app import User, Performance
        from datetime import datetime
        
        # Get db instance using current_app context
        try:
            db = current_app.extensions['sqlalchemy']
        except (KeyError, AttributeError):
            from app import db
        
        metric_type = request.form.get('metric_type', '').strip()
        
        if not metric_type:
            return jsonify({'success': False, 'message': 'Metric type is required'}), 400
        
        # Get current month and year
        current_date = datetime.now()
        current_month = current_date.month
        current_year = current_date.year
        
        # Get all employee ratings from form
        employee_ratings = {}
        for key, value in request.form.items():
            if key.startswith('employee_'):
                try:
                    employee_id_str = key.replace('employee_', '')
                    # Validate employee_id with type safety
                    from app import validate_type_safe
                    id_valid, employee_id, id_error = validate_type_safe(employee_id_str, int, min_value=1)
                    if not id_valid:
                        return jsonify({'success': False, 'message': f'Invalid employee ID: {id_error}'}), 400
                    
                    # Validate rating with type safety
                    rating_valid, rating, rating_error = validate_type_safe(value, float, min_value=1.0, max_value=10.0)
                    if not rating_valid:
                        return jsonify({'success': False, 'message': f'Rating for employee {employee_id} must be between 1 and 10'}), 400
                    
                    employee_ratings[employee_id] = rating
                except Exception as e:
                    return jsonify({'success': False, 'message': f'Error processing rating: {str(e)}'}), 400
        
        if not employee_ratings:
            return jsonify({'success': False, 'message': 'No ratings provided'}), 400
        
        # Save performance data for each employee
        saved_count = 0
        for employee_id, rating in employee_ratings.items():
            # Get or create performance record
            performance = db.session.query(Performance).filter_by(
                user_id=employee_id,
                month=current_month,
                year=current_year
            ).first()
            
            if not performance:
                performance = Performance(
                    user_id=employee_id,
                    month=current_month,
                    year=current_year
                )
                db.session.add(performance)
            
            # Update the specific metric
            if metric_type == 'technical_skills':
                performance.technical_skills = rating
            elif metric_type == 'client_satisfaction':
                performance.client_satisfaction = rating
            else:
                return jsonify({'success': False, 'message': 'Invalid metric type'}), 400
            
            saved_count += 1
        
        db.session.commit()
        
        # Update Excel file
        try:
            update_performance_excel_admin(current_month, current_year, metric_type, current_app)
        except Exception as e:
            print(f"Warning: Could not update Excel file: {e}")
            # Don't fail the request if Excel update fails
        
        metric_name = metric_type.replace('_', ' ').title()
        return jsonify({
            'success': True,
            'message': f'{metric_name} ratings saved successfully for {saved_count} employee(s)!'
        })
        
    except Exception as e:
        try:
            db.session.rollback()
        except:
            pass
        print(f"Error saving performance data: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error saving performance data: {str(e)}'}), 500


def update_performance_excel_admin(month, year, metric_type, app_context=None):
    """Update Performance_data.xlsx with new performance data for Admin metrics"""
    try:
        import os
        from openpyxl import load_workbook, Workbook
        from app import User, Performance
        
        # Get db instance using current_app context
        if app_context:
            db = app_context.extensions['sqlalchemy']
        else:
            from flask import current_app
            try:
                db = current_app.extensions['sqlalchemy']
            except (KeyError, AttributeError):
                from app import db
        
        excel_path = os.path.join('static', 'Performance', 'Performance_data.xlsx')
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        
        # Worksheet name format: "MM_YYYY" (e.g., "11_2025")
        worksheet_name = f"{month:02d}_{year}"
        
        # Load or create workbook
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # Get or create worksheet
        if worksheet_name in wb.sheetnames:
            ws = wb[worksheet_name]
        else:
            ws = wb.create_sheet(worksheet_name)
            # Add headers
            ws['A1'] = 'Employee Name'
            ws['B1'] = 'Punctuality'
            ws['C1'] = 'Behaviour'
            ws['D1'] = 'Team Coordination'
            ws['E1'] = 'Communication Skills'
            ws['F1'] = 'Technical Skills'
            ws['G1'] = 'Client Satisfaction'
            
            # Style headers
            from openpyxl.styles import Font, PatternFill
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
        
        # Ensure all headers exist (in case worksheet was created by HR dashboard)
        headers = {
            'A1': 'Employee Name',
            'B1': 'Punctuality',
            'C1': 'Behaviour',
            'D1': 'Team Coordination',
            'E1': 'Communication Skills',
            'F1': 'Technical Skills',
            'G1': 'Client Satisfaction'
        }
        
        for cell_ref, header_value in headers.items():
            if ws[cell_ref].value != header_value:
                ws[cell_ref] = header_value
        
        # Style headers if not already styled
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Get all users (active only)
        all_users = db.session.query(User).all()
        active_users = []
        for user in all_users:
            is_deleted = False
            if hasattr(user, 'deleted_at') and user.deleted_at:
                is_deleted = True
            is_inactive = False
            if user.status and not user.status.is_active:
                is_inactive = True
            if not is_deleted and not is_inactive:
                active_users.append(user)
        
        # Get performance data for current month
        performances = db.session.query(Performance).filter_by(month=month, year=year).all()
        perf_dict = {p.user_id: p for p in performances}
        
        # Clear existing data (keep headers)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
        
        # Add employee data
        row = 2
        for user in active_users:
            ws.cell(row=row, column=1, value=user.employee_name or user.username)
            
            perf = perf_dict.get(user.id)
            if perf:
                ws.cell(row=row, column=2, value=perf.punctuality or 0)
                ws.cell(row=row, column=3, value=perf.behaviour or 0)
                ws.cell(row=row, column=4, value=perf.team_coordination or 0)
                ws.cell(row=row, column=5, value=perf.communication_skills or 0)
                ws.cell(row=row, column=6, value=perf.technical_skills or 0)
                ws.cell(row=row, column=7, value=perf.client_satisfaction or 0)
            else:
                ws.cell(row=row, column=2, value=0)
                ws.cell(row=row, column=3, value=0)
                ws.cell(row=row, column=4, value=0)
                ws.cell(row=row, column=5, value=0)
                ws.cell(row=row, column=6, value=0)
                ws.cell(row=row, column=7, value=0)
            
            row += 1
        
        # Save workbook
        wb.save(excel_path)
        print(f"✅ Excel file updated: {excel_path}, Worksheet: {worksheet_name}")
        
    except Exception as e:
        print(f"❌ Error updating Excel file: {e}")
        import traceback
        traceback.print_exc()
        raise e


@admin_dashboard_bp.route('/admin/get_employee_last_month_performance/<int:employee_id>', methods=['GET'])
@login_required
def get_employee_last_month_performance(employee_id):
    """Get last month's performance data for a specific employee"""
    if current_user.department != "Admin":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from app import User, Performance
        from datetime import datetime, timedelta
        
        # Get db instance
        try:
            db = current_app.extensions['sqlalchemy']
        except (KeyError, AttributeError):
            from app import db
        
        # Calculate last month
        current_date = datetime.now()
        last_month = current_date.month - 1 if current_date.month > 1 else 12
        last_month_year = current_date.year if current_date.month > 1 else current_date.year - 1
        
        # Get performance data
        performance = db.session.query(Performance).filter_by(
            user_id=employee_id,
            month=last_month,
            year=last_month_year
        ).first()
        
        if performance:
            return jsonify({
                'success': True,
                'performance': {
                    'punctuality': performance.punctuality or 0,
                    'client_satisfaction': performance.client_satisfaction or 0,
                    'behaviour': performance.behaviour or 0,
                    'communication_skills': performance.communication_skills or 0,
                    'technical_skills': performance.technical_skills or 0,
                    'team_coordination': performance.team_coordination or 0
                }
            })
        else:
            # Return zeros if no data exists
            return jsonify({
                'success': True,
                'performance': {
                    'punctuality': 0,
                    'client_satisfaction': 0,
                    'behaviour': 0,
                    'communication_skills': 0,
                    'technical_skills': 0,
                    'team_coordination': 0
                }
            })
            
    except Exception as e:
        current_app.logger.error(f"Error fetching last month performance: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_dashboard_bp.route('/admin/get_employee_last_year_performance/<int:employee_id>', methods=['GET'])
@login_required
def get_employee_last_year_performance(employee_id):
    """Get last 12 months performance for a specific employee with all 6 metrics separately (for bar charts)"""
    if current_user.department != "Admin":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from app import Performance
        from datetime import datetime, timedelta
        
        # Get db instance
        try:
            db = current_app.extensions['sqlalchemy']
        except (KeyError, AttributeError):
            from app import db
        
        current_date = datetime.now()
        performance_history = []
        
        for i in range(12):
            # Calculate month and year for each of the previous 12 months
            month_offset = i + 1
            target_date = current_date - timedelta(days=30*month_offset)
            month = target_date.month
            year = target_date.year
            
            performance = db.session.query(Performance).filter_by(
                user_id=employee_id,
                month=month,
                year=year
            ).first()
            
            if performance:
                # Return all 6 metrics separately
                performance_history.append({
                    'month': month,
                    'year': year,
                    'punctuality': performance.punctuality or 0,
                    'client_satisfaction': performance.client_satisfaction or 0,
                    'behaviour': performance.behaviour or 0,
                    'communication_skills': performance.communication_skills or 0,
                    'technical_skills': performance.technical_skills or 0,
                    'team_coordination': performance.team_coordination or 0,
                    'month_name': target_date.strftime('%b'),
                    'year_short': target_date.strftime('%y')
                })
            else:
                # If no data exists for this month, add a placeholder with 0 for all metrics
                performance_history.append({
                    'month': month,
                    'year': year,
                    'punctuality': 0,
                    'client_satisfaction': 0,
                    'behaviour': 0,
                    'communication_skills': 0,
                    'technical_skills': 0,
                    'team_coordination': 0,
                    'month_name': target_date.strftime('%b'),
                    'year_short': target_date.strftime('%y')
                })
        
        # Reverse to get chronological order (oldest to newest)
        performance_history.reverse()
        
        return jsonify({
            'success': True,
            'performance_history': performance_history
        })
        
    except Exception as e:
        current_app.logger.error(f"Error fetching last year performance: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_dashboard_bp.route('/admin/get_employee_performance_growth/<int:employee_id>', methods=['GET'])
@login_required
def get_employee_performance_growth(employee_id):
    """Get performance history for growth line chart"""
    if current_user.department != "Admin":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from app import Performance
        from datetime import datetime, timedelta
        
        # Get db instance
        try:
            db = current_app.extensions['sqlalchemy']
        except (KeyError, AttributeError):
            from app import db
        
        current_date = datetime.now()
        performance_history = []
        
        # Get last 12 months of data
        for i in range(12):
            # Calculate month and year for each of the previous 12 months
            month_offset = i + 1
            target_date = current_date - timedelta(days=30*month_offset)
            month = target_date.month
            year = target_date.year
            
            performance = db.session.query(Performance).filter_by(
                user_id=employee_id,
                month=month,
                year=year
            ).first()
            
            if performance:
                # Calculate average of all performance metrics
                avg_score = (
                    (performance.punctuality or 0) + 
                    (performance.client_satisfaction or 0) + 
                    (performance.behaviour or 0) + 
                    (performance.communication_skills or 0) + 
                    (performance.technical_skills or 0) + 
                    (performance.team_coordination or 0)
                ) / 6
                
                performance_history.append({
                    'month': month,
                    'year': year,
                    'average': round(avg_score, 2),
                    'month_name': target_date.strftime('%b'),
                    'year_short': target_date.strftime('%y')
                })
            else:
                # If no data exists for this month, add a placeholder with 0
                performance_history.append({
                    'month': month,
                    'year': year,
                    'average': 0,
                    'month_name': target_date.strftime('%b'),
                    'year_short': target_date.strftime('%y')
                })
        
        # Reverse to get chronological order (oldest to newest)
        performance_history.reverse()
        
        return jsonify({
            'success': True,
            'performance_history': performance_history
        })
        
    except Exception as e:
        current_app.logger.error(f"Error fetching performance growth: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_dashboard_bp.route('/admin/download_performance_excel', methods=['GET'])
@login_required
def download_performance_excel():
    """Download the Performance_data.xlsx file"""
    if current_user.department != "Admin":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        import os
        excel_path = os.path.join('static', 'Performance', 'Performance_data.xlsx')
        
        # Check if file exists
        if not os.path.exists(excel_path):
            return jsonify({'success': False, 'message': 'Performance data file not found'}), 404
        
        # Send the file for download
        return send_file(
            excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Performance_data.xlsx'
        )
        
    except Exception as e:
        current_app.logger.error(f"Error downloading performance Excel file: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_dashboard_bp.route('/admin/get_logs', methods=['GET'])
@login_required
def get_logs():
    """Get activity logs based on time period"""
    if current_user.department != "Admin":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        from Admin_Dashboard_Files.user_activity_logger import get_logs_from_excel
        from datetime import datetime, timedelta
        from collections import defaultdict
        
        period = request.args.get('period', 'all')
        
        # Calculate date range based on period
        end_date = datetime.now()
        
        if period == '1day':
            # Exactly 24 hours ago
            start_date = end_date - timedelta(hours=24)
        elif period == '1week':
            # Exactly 7 days (168 hours) ago
            start_date = end_date - timedelta(days=7)
        elif period == '1month':
            # Exactly 30 days ago
            start_date = end_date - timedelta(days=30)
        else:  # 'all' - show last 1 year
            # Exactly 365 days (1 year) ago
            start_date = end_date - timedelta(days=365)
        
        # Get logs
        logs = get_logs_from_excel(start_date=start_date, end_date=end_date, limit=50000)
        
        # Show all logs (no filtering) - but still mark suspicious activities for highlighting
        # Detect suspicious activities to add flags for UI highlighting
        if logs:
            logs_with_flags = detect_suspicious_activities(logs)
            # Create a map of suspicious/main activity flags by log key
            flags_map = {}
            for flagged_log in logs_with_flags:
                log_key = (
                    flagged_log.get('timestamp', ''),
                    flagged_log.get('user_id', ''),
                    flagged_log.get('activity_type', ''),
                    flagged_log.get('request_url', '')
                )
                flags_map[log_key] = {
                    'is_suspicious': flagged_log.get('is_suspicious', False),
                    'is_main_activity': flagged_log.get('is_main_activity', False),
                    'suspicion_reasons': flagged_log.get('suspicion_reasons', [])
                }
            
            # Add flags to all logs
            for log in logs:
                log_key = (
                    log.get('timestamp', ''),
                    log.get('user_id', ''),
                    log.get('activity_type', ''),
                    log.get('request_url', '')
                )
                if log_key in flags_map:
                    log['is_suspicious'] = flags_map[log_key]['is_suspicious']
                    log['is_main_activity'] = flags_map[log_key]['is_main_activity']
                    log['suspicion_reasons'] = flags_map[log_key]['suspicion_reasons']
                else:
                    log['is_suspicious'] = False
                    log['is_main_activity'] = False
                    log['suspicion_reasons'] = []
        
        # Sort by timestamp (most recent first)
        logs.sort(key=lambda x: x.get('timestamp', ''), reverse=True)
        
        return jsonify({
            'success': True,
            'logs': logs,
            'count': len(logs)
        })
        
    except Exception as e:
        current_app.logger.error(f"Error getting logs: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


def detect_suspicious_activities(logs):
    """
    Detect suspicious activities from logs
    Returns list of suspicious logs with 'is_suspicious' flag
    """
    suspicious_logs = []
    
    if not logs:
        return suspicious_logs
    
    # Track activity patterns within 1-minute windows
    # Group activities by minute windows and count occurrences
    minute_window_activities = defaultdict(lambda: defaultdict(int))  # (minute_window, activity_key) -> count
    
    # Main activities that should always be shown (only login, report generation, and download)
    main_activities_to_show = [
        'login', 'report_generation', 'generate', 'download', 
        'export', 'create_report'
    ]
    
    # Process logs to identify patterns within 1-minute windows
    for log in logs:
        user_id = log.get('user_id', '')
        username = log.get('username', '')
        activity_type = log.get('activity_type', '').lower()
        request_url = log.get('request_url', '')
        ip_address = log.get('ip_address', '')
        timestamp = log.get('timestamp', '')
        
        # Check if it's a main activity to show (only login, report generation, download)
        is_main_activity = any(main_act in activity_type or main_act in request_url.lower() 
                              for main_act in main_activities_to_show)
        
        # Create key for tracking
        activity_key = (str(user_id), activity_type, request_url)
        
        # Track time-based patterns (activities per minute)
        try:
            if timestamp:
                from datetime import datetime
                log_time = datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S')
                minute_window = log_time.strftime('%Y-%m-%d %H:%M')
                minute_window_activities[minute_window][activity_key] += 1
        except:
            pass
    
    # Identify suspicious logs
    for log in logs:
        user_id = log.get('user_id', '')
        username = log.get('username', '')
        activity_type = log.get('activity_type', '').lower()
        request_url = log.get('request_url', '')
        ip_address = log.get('ip_address', '')
        timestamp = log.get('timestamp', '')
        
        is_suspicious = False
        suspicion_reasons = []
        
        # Check if it's a main activity to show (only login, report generation, download)
        # For login: only show successful logins (not failed attempts)
        is_main_activity = any(main_act in activity_type or main_act in request_url.lower() 
                              for main_act in main_activities_to_show)
        
        # For login activities, only show successful logins (not failed attempts)
        if 'login' in activity_type or 'login' in (request_url or '').lower():
            # Check if it's a failed login attempt
            failed_login_indicators = ['failed', 'fail', 'error', 'invalid', 'unauthorized', '401', '403', 'failed_attempt']
            activity_desc = log.get('activity_description', '').lower()
            request_data = log.get('request_data', '')
            request_data_str = str(request_data).lower() if request_data else ''
            response_status = log.get('response_status', '')
            
            is_failed_login = (
                'failed_attempt' in activity_type or
                any(indicator in activity_desc or indicator in request_data_str 
                   for indicator in failed_login_indicators) or
                (response_status and str(response_status) in ['401', '403'])
            )
            
            # Only show successful logins as main activity
            if not is_failed_login:
                log['is_suspicious'] = False
                log['is_main_activity'] = True
                suspicious_logs.append(log)
                continue  # Skip suspicious check for successful logins
            # Failed logins will continue to suspicious check below (don't continue here)
        elif is_main_activity:
            # For report generation and download activities, always show
            log['is_suspicious'] = False
            log['is_main_activity'] = True
            suspicious_logs.append(log)
            continue
        
        # Check for repeated activities within 1 minute (same user, same activity, same URL - 3+ times in 1 minute)
        activity_key = (str(user_id), activity_type, request_url)
        try:
            if timestamp:
                from datetime import datetime
                log_time = datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S')
                minute_window = log_time.strftime('%Y-%m-%d %H:%M')
                
                # Count occurrences in the same minute window
                count_in_minute = minute_window_activities.get(minute_window, {}).get(activity_key, 0)
                
                # If 7+ times in the same minute, mark as suspicious
                if count_in_minute >= 7:
                    is_suspicious = True
                    suspicion_reasons.append(f"Repeated activity ({count_in_minute} times in 1 minute)")
        except:
            pass
        
        # Check for failed login attempts - if 2+ back-to-back failed attempts (any time), mark as suspicious
        # Reset count after successful login
        if 'login' in activity_type or 'login' in (request_url or '').lower() or 'failed_attempt' in activity_type:
            # Check for failed login indicators
            failed_login_indicators = ['failed', 'fail', 'error', 'invalid', 'unauthorized', '401', '403', 'failed_attempt']
            activity_desc = log.get('activity_description', '').lower()
            request_data = log.get('request_data', '')
            request_data_str = str(request_data).lower() if request_data else ''
            response_status = log.get('response_status', '')
            
            # Check if this is a failed login attempt
            is_failed_login = (
                'failed_attempt' in activity_type or
                any(indicator in activity_desc or indicator in request_data_str 
                   for indicator in failed_login_indicators) or
                (response_status and str(response_status) in ['401', '403'])
            )
            
            # Check if this is a successful login (resets the failed attempt counter)
            is_successful_login = (
                'login' in activity_type and 
                'failed' not in activity_type and
                'failed_attempt' not in activity_type and
                not any(indicator in activity_desc for indicator in ['failed', 'fail', 'error', 'invalid']) and
                (not response_status or str(response_status) not in ['401', '403'])
            )
            
            if is_failed_login:
                # Count consecutive failed login attempts (back-to-back, not time-based)
                # Look backwards through logs to find the last successful login or count consecutive failures
                failed_login_count = 0
                try:
                    if timestamp:
                        from datetime import datetime
                        log_time = datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S')
                        
                        # Sort logs by timestamp to check in chronological order
                        sorted_logs = sorted([l for l in logs if l.get('timestamp')], 
                                             key=lambda x: datetime.strptime(str(x.get('timestamp')), '%Y-%m-%d %H:%M:%S'))
                        
                        # Find this log's position and check backwards for consecutive failures
                        current_log_index = None
                        for idx, other_log in enumerate(sorted_logs):
                            if (other_log.get('timestamp') == timestamp and 
                                other_log.get('user_id') == user_id):
                                current_log_index = idx
                                break
                        
                        if current_log_index is not None:
                            # Check backwards from current log for consecutive failed attempts
                            for i in range(current_log_index - 1, -1, -1):
                                other_log = sorted_logs[i]
                                
                                # Only check logs from the same user
                                if other_log.get('user_id') != user_id:
                                    continue
                                
                                other_activity_type = other_log.get('activity_type', '').lower()
                                other_activity_desc = other_log.get('activity_description', '').lower()
                                other_request_data = str(other_log.get('request_data', '')).lower()
                                other_response = other_log.get('response_status', '')
                                
                                # Check if this is a login-related activity
                                if ('login' in other_activity_type or 'login' in (other_log.get('request_url', '') or '').lower() or
                                    'failed_attempt' in other_activity_type):
                                    
                                    # Check if it's a successful login (reset counter)
                                    other_is_success = (
                                        'login' in other_activity_type and 
                                        'failed' not in other_activity_type and
                                        'failed_attempt' not in other_activity_type and
                                        not any(indicator in other_activity_desc for indicator in ['failed', 'fail', 'error', 'invalid']) and
                                        (not other_response or str(other_response) not in ['401', '403'])
                                    )
                                    
                                    if other_is_success:
                                        # Found successful login, stop counting (counter resets)
                                        break
                                    
                                    # Check if it's a failed login
                                    other_is_failed = (
                                        'failed_attempt' in other_activity_type or
                                        any(indicator in other_activity_desc or indicator in other_request_data 
                                          for indicator in failed_login_indicators) or
                                        (other_response and str(other_response) in ['401', '403'])
                                    )
                                    
                                    if other_is_failed:
                                        failed_login_count += 1
                                    else:
                                        # Not a login activity or not failed, stop counting
                                        break
                                else:
                                    # Not a login-related activity, stop counting
                                    break
                        
                        # If we see 1+ consecutive failed login attempts, this current one would be the 2nd, making it suspicious
                        if failed_login_count >= 1:
                            is_suspicious = True
                            suspicion_reasons.append(f"Consecutive failed login attempts ({failed_login_count + 1} failed attempts)")
                except Exception as e:
                    print(f"Error checking failed login attempts: {e}")
                    pass
        
        # Check for error responses (4xx, 5xx)
        response_status = log.get('response_status', '')
        if response_status:
            try:
                status_code = int(response_status)
                if status_code >= 400:
                    is_suspicious = True
                    suspicion_reasons.append(f"Error response ({status_code})")
            except:
                pass
        
        # Check for suspicious request patterns
        suspicious_patterns = [
            'sql', 'script', 'union', 'select', 'drop', 'delete', 'insert', 'update',
            'exec', 'eval', 'cmd', 'shell', 'system', 'passwd', 'password', 'admin',
            'test', 'debug', 'trace', 'backup', 'config', 'env', '.env'
        ]
        
        request_data = log.get('request_data', '')
        if request_data:
            request_data_lower = str(request_data).lower()
            for pattern in suspicious_patterns:
                if pattern in request_data_lower:
                    is_suspicious = True
                    suspicion_reasons.append(f"Suspicious pattern detected")
                    break
        
        # If suspicious, add to list
        if is_suspicious:
            log['is_suspicious'] = True
            log['suspicion_reasons'] = suspicion_reasons
            suspicious_logs.append(log)
    
    return suspicious_logs


@admin_dashboard_bp.route('/admin/get_available_log_files', methods=['GET'])
@login_required
def get_available_log_files():
    """Get list of available log files (monthly files)"""
    if current_user.department != "Admin":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        import os
        from Admin_Dashboard_Files.user_activity_logger import LOGS_DIR, ensure_logs_directory
        from datetime import datetime
        
        ensure_logs_directory()
        
        # Get all Excel files in logs directory
        log_files = []
        if os.path.exists(LOGS_DIR):
            excel_files = [f for f in os.listdir(LOGS_DIR) if f.endswith('.xlsx') and not f.startswith('~$')]
            
            for excel_file in excel_files:
                # Extract month and year from filename (e.g., Dec_2025.xlsx -> Dec 2025)
                try:
                    # Remove .xlsx extension
                    file_base = excel_file.replace('.xlsx', '')
                    # Split by underscore
                    parts = file_base.split('_')
                    if len(parts) == 2:
                        month_abbr = parts[0]  # Dec
                        year = parts[1]  # 2025
                        # Format as "Dec 2025"
                        display_name = f"{month_abbr} {year}"
                        log_files.append({
                            'filename': excel_file,
                            'display_name': display_name,
                            'month': month_abbr,
                            'year': year
                        })
                except:
                    # If parsing fails, use filename as display name
                    log_files.append({
                        'filename': excel_file,
                        'display_name': excel_file.replace('.xlsx', ''),
                        'month': '',
                        'year': ''
                    })
        
        # Sort by year and month (most recent first)
        def sort_key(file_info):
            try:
                month_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                month_idx = month_order.index(file_info['month']) if file_info['month'] in month_order else 0
                return (int(file_info['year']), month_idx)
            except:
                return (0, 0)
        
        log_files.sort(key=sort_key, reverse=True)
        
        return jsonify({
            'success': True,
            'log_files': log_files
        })
        
    except Exception as e:
        current_app.logger.error(f"Error getting log files: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_dashboard_bp.route('/admin/download_log_file/<filename>', methods=['GET'])
@login_required
def download_log_file(filename):
    """Download a specific log file"""
    if current_user.department != "Admin":
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        import os
        import shutil
        import tempfile
        from Admin_Dashboard_Files.user_activity_logger import LOGS_DIR
        
        # Security: ensure filename doesn't contain path traversal
        filename = os.path.basename(filename)
        
        file_path = os.path.join(LOGS_DIR, filename)
        
        # Check if file exists
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'message': 'File not found'}), 404
        
        # Copy file to temporary location to avoid file locking issues
        # This ensures the original file can remain open/writable while we serve the copy
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        
        try:
            # Read file content into memory first to ensure we get a complete, uncorrupted copy
            # This avoids issues if the file is being written to during download
            import time
            max_retries = 3
            retry_delay = 0.1  # 100ms
            
            file_content = None
            for attempt in range(max_retries):
                try:
                    # Read the entire file into memory
                    with open(file_path, 'rb') as f:
                        file_content = f.read()
                    break  # Success, exit retry loop
                except (PermissionError, IOError) as e:
                    if attempt < max_retries - 1:
                        time.sleep(retry_delay)
                        retry_delay *= 2  # Exponential backoff
                        continue
                    else:
                        raise e
            
            if file_content is None:
                raise Exception("Failed to read file after retries")
            
            # Write the content to temporary file
            with open(temp_file.name, 'wb') as f:
                f.write(file_content)
            
            # Send the temporary file for download
            response = send_file(
                temp_file.name,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
            # Schedule cleanup of temporary file after response is sent
            def cleanup_temp_file():
                try:
                    if os.path.exists(temp_file.name):
                        os.unlink(temp_file.name)
                except Exception as e:
                    current_app.logger.error(f"Error cleaning up temp file: {e}")
            
            response.call_on_close(cleanup_temp_file)
            
            return response
            
        except Exception as e:
            # Clean up temp file if error occurs
            try:
                if os.path.exists(temp_file.name):
                    os.unlink(temp_file.name)
            except:
                pass
            raise e
        
    except Exception as e:
        current_app.logger.error(f"Error downloading log file: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500



