# API VAPT Follow-Up Audit Excel Report Generator
from flask import Blueprint, request, send_file, make_response, jsonify
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import zipfile
import os
import re
import shutil
import json
import tempfile
from io import BytesIO

# Create a Blueprint for API VAPT Follow-Up routes
api_vapt_followup_bp = Blueprint('api_vapt_followup', __name__)

def generate_api_followup_filename(organization, end_date):
    """
    Generate static filename for API VAPT Follow-Up Audit.
    Format: API_VAPT_Follow_Up_Audit
    """
    return "API_VAPT_Follow_Up_Audit_Report.xlsx"

def normalize_name_for_matching(name):
    """
    Normalize name for matching by:
    1. Replacing special characters with '-'
    2. Trimming to first 170 characters
    3. Removing trailing numbers (e.g., 'abc 1' -> 'abc', 'abc_1' -> 'abc', but 'ab1c' stays 'ab1c')
    """
    if not name:
        return ""
    
    # Replace special characters with '-'
    normalized = re.sub(r'[<>\:"/\\|?*]', '-', str(name))
    
    # Remove trailing whitespace
    normalized = normalized.strip()
    
    # Remove trailing numbers (only at the end of the name)
    normalized = re.sub(r'[\s_\-]+\d+$', '', normalized)
    
    # Take only first 170 characters
    normalized = normalized[:170]
    
    return normalized.lower()

def copy_worksheet_with_images(source_wb, source_ws_name, target_wb, target_ws_name):
    """Copy worksheet from source workbook to target workbook including images and formatting
    Returns tuple: (success: bool, temp_image_files: list)"""
    temp_image_files = []
    
    try:
        source_ws = source_wb[source_ws_name]
        target_ws = target_wb.create_sheet(target_ws_name)
        
        # Copy all cells with data and formatting
        for row in source_ws.iter_rows():
            for cell in row:
                target_cell = target_ws.cell(row=cell.row, column=cell.column)
                
                # Copy value
                target_cell.value = cell.value
                
                # Copy formatting
                if cell.has_style:
                    try:
                        target_cell.font = cell.font.copy() if cell.font else Font()
                        target_cell.border = cell.border.copy() if cell.border else Border()
                        target_cell.fill = cell.fill.copy() if cell.fill else PatternFill()
                        target_cell.number_format = cell.number_format
                        target_cell.protection = cell.protection.copy() if cell.protection else None
                        target_cell.alignment = cell.alignment.copy() if cell.alignment else Alignment()
                    except:
                        pass
        
        # Copy column dimensions
        for col_letter, col_dim in source_ws.column_dimensions.items():
            if col_dim.width:
                target_ws.column_dimensions[col_letter].width = col_dim.width
            if col_dim.hidden:
                target_ws.column_dimensions[col_letter].hidden = col_dim.hidden
        
        # Copy row dimensions
        for row_num, row_dim in source_ws.row_dimensions.items():
            if row_dim.height:
                target_ws.row_dimensions[row_num].height = row_dim.height
            if row_dim.hidden:
                target_ws.row_dimensions[row_num].hidden = row_dim.hidden
        
        # Copy merged cells
        for merged_cell in source_ws.merged_cells.ranges:
            try:
                target_ws.merge_cells(str(merged_cell))
            except:
                pass
        
        # Copy images
        if hasattr(source_ws, '_images'):
            for img in source_ws._images:
                try:
                    # Get image data
                    if hasattr(img, '_data'):
                        img_data = img._data() if callable(img._data) else img._data
                    else:
                        continue
                    
                    # Create temporary file for image (don't delete it yet!)
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
                        tmp.write(img_data)
                        temp_img_path = tmp.name
                    
                    # Keep track of temp file for cleanup later
                    temp_image_files.append(temp_img_path)
                    
                    # Create new image object
                    new_img = Image(temp_img_path)
                    
                    # Copy image properties
                    if hasattr(img, 'width'):
                        new_img.width = img.width
                    if hasattr(img, 'height'):
                        new_img.height = img.height
                    
                    # Copy anchor if available
                    if hasattr(img, 'anchor'):
                        new_img.anchor = img.anchor
                    
                    target_ws.add_image(new_img)
                    
                except Exception as e:
                    print(f"  ⚠️ Error copying image: {e}")
                    # If image copy failed, still try to clean up the temp file
                    if 'temp_img_path' in locals() and temp_img_path in temp_image_files:
                        temp_image_files.remove(temp_img_path)
                    if 'temp_img_path' in locals():
                        try:
                            os.unlink(temp_img_path)
                        except:
                            pass
                    continue
            
        print(f"✅ Copied worksheet '{source_ws_name}' to '{target_ws_name}' with {len(temp_image_files)} images")
        return True, temp_image_files
            
    except Exception as e:
        print(f"❌ Error copying worksheet: {e}")
        import traceback
        traceback.print_exc()
        # Clean up any temp files created before error
        for temp_file in temp_image_files:
            try:
                os.unlink(temp_file)
            except:
                pass
        return False, []

def find_first_empty_cell_in_row(ws, row_num=1):
    """Find the first empty cell in a row, checking A1, B1, C1, etc."""
    col = 1
    while True:
        cell = ws.cell(row=row_num, column=col)
        # Check if cell is part of a merged range
        if isinstance(cell, type(ws.cell(row=1, column=1))):
            # Check if this cell is the top-left of a merged range
            is_merged = False
            for merged_range in ws.merged_cells.ranges:
                if (merged_range.min_row <= row_num <= merged_range.max_row and 
                    merged_range.min_col <= col <= merged_range.max_col):
                    # This cell is in a merged range
                    # Check if it's the top-left cell
                    if merged_range.min_row == row_num and merged_range.min_col == col:
                        # It's the top-left, check its value
                        if cell.value is None or str(cell.value).strip() == '':
                            return col
                    is_merged = True
                    break
            
            if not is_merged:
                # Not merged, check value normally
                if cell.value is None or str(cell.value).strip() == '':
                    return col
        else:
            # It's a MergedCell, skip it
            pass
        
        col += 1
        if col > 1000:  # Safety limit
            break
    return col

def get_top_left_cell_of_merged_range(ws, row, col):
    """Get the top-left cell of a merged range if the given cell is merged, otherwise return the cell itself"""
    cell = ws.cell(row=row, column=col)
    
    # Check if this cell is part of a merged range
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and 
            merged_range.min_col <= col <= merged_range.max_col):
            # Return the top-left cell of the merged range
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    
    # Not merged, return the cell itself
    return cell

def apply_borders_to_content_range(ws):
    """Apply borders to all cells in the content range"""
    try:
        from openpyxl.styles import Border, Side
        
        # Create thin border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Find last column in row 1 with content
        # Check 10 cells ahead if empty, and if so, add 6 more columns for borders
        last_col = 0
        col = 1
        
        while col <= ws.max_column:
            cell_value = ws.cell(row=1, column=col).value
            has_content = cell_value is not None and str(cell_value).strip() != ""
            
            if has_content:
                # Check if next 10 cells are empty
                next_10_empty = True
                for check_col in range(col + 1, min(col + 11, ws.max_column + 1)):
                    check_value = ws.cell(row=1, column=check_col).value
                    if check_value is not None and str(check_value).strip() != "":
                        next_10_empty = False
                        break
                
                if next_10_empty:
                    # Next 10 cells are empty, add 6 more columns for borders
                    # So border goes from col to col+6 (7 columns total)
                    last_col = min(col + 6, ws.max_column)
                    print(f"📍 Found content at {get_column_letter(col)}, next 10 empty, applying borders to {get_column_letter(col)}-{get_column_letter(last_col)}")
                    break
                else:
                    # Next 10 cells have content, continue
                    last_col = col
                    col += 1
            else:
                # Check next 10 cells (col+1 to col+10)
                found_content = False
                for check_col in range(col + 1, min(col + 11, ws.max_column + 1)):
                    check_value = ws.cell(row=1, column=check_col).value
                    if check_value is not None and str(check_value).strip() != "":
                        found_content = True
                        last_col = check_col
                        col = check_col + 1
                        break
                
                if not found_content:
                    # No content in next 10 cells, stop here
                    # last_col is the last cell with content before this empty sequence
                    break
        
        # Ensure we have at least column A
        if last_col == 0:
            last_col = 1
        
        print(f"📍 Last column for borders in row 1: {get_column_letter(last_col)} (col {last_col})")
        
        # Find last row in column A with content (only check 1 empty cell)
        last_row = 1
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value is not None and str(cell_value).strip() != "":
                last_row = row
            elif row > 1:
                # If we find an empty cell after row 1, stop checking (only need to check 1 empty)
                break
        
        print(f"📍 Last row with content in column A: {last_row}")
        
        # Apply borders to all cells in range A1 to last_col, last_row
        # Handle merged cells properly
        processed_merged_ranges = set()
        
        for row in range(1, last_row + 1):
            for col in range(1, last_col + 1):
                # Check if this cell is part of a merged range
                is_in_merged = False
                merged_range = None
                
                for mr in ws.merged_cells.ranges:
                    if (mr.min_row <= row <= mr.max_row and 
                        mr.min_col <= col <= mr.max_col):
                        is_in_merged = True
                        merged_range = mr
                        break
                
                if is_in_merged:
                    # Only process the top-left cell of merged range
                    if (merged_range.min_row, merged_range.min_col) not in processed_merged_ranges:
                        # Apply border to all cells in merged range
                        for mr_row in range(merged_range.min_row, merged_range.max_row + 1):
                            for mr_col in range(merged_range.min_col, merged_range.max_col + 1):
                                if (mr_row <= last_row and mr_col <= last_col):
                                    cell = ws.cell(row=mr_row, column=mr_col)
                                    cell.border = thin_border
                        
                        processed_merged_ranges.add((merged_range.min_row, merged_range.min_col))
                else:
                    # Regular cell, apply border
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
        
        print(f"✅ Applied borders to range A1:{get_column_letter(last_col)}{last_row}")
        
    except Exception as e:
        print(f"⚠️ Error applying borders: {e}")
        import traceback
        traceback.print_exc()

def find_latest_status_column_in_user_excel(user_excel_path, ws_name="API VAPT"):
    """Find the latest Status column in user's Excel file by date"""
    try:
        user_wb = load_workbook(user_excel_path)
        
        # Find worksheet
        ws = None
        for sheet_name in user_wb.sheetnames:
            if ws_name.lower() in sheet_name.lower() or "api vapt" in sheet_name.lower() or "api_vapt" in sheet_name.lower():
                ws = user_wb[sheet_name]
                break
        
        if not ws:
            user_wb.close()
            return None, None
        
        # Find Status columns in row 1
        status_columns = []
        
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                cell_str = str(cell_value).strip()
                if cell_str.lower().startswith('status'):
                    # Try to extract date from "Status - DD.MM.YYYY" format
                    import re
                    date_match = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', cell_str)
                    if date_match:
                        day = int(date_match.group(1))
                        month = int(date_match.group(2))
                        year = int(date_match.group(3))
                        date_obj = datetime(year, month, day)
                        status_columns.append({
                            'col': col,
                            'header': cell_str,
                            'date': date_obj
                        })
        
        user_wb.close()
        
        if not status_columns:
            print("⚠️ No Status columns found in user Excel")
            return None, None
        
        # Sort by date descending and get the latest
        status_columns.sort(key=lambda x: x['date'], reverse=True)
        latest_status = status_columns[0]
        
        print(f"📍 Found {len(status_columns)} Status columns, using latest: {latest_status['header']} (col {get_column_letter(latest_status['col'])})")
        
        return latest_status['col'], latest_status['date']
                                    
    except Exception as e:
        print(f"❌ Error finding Status column in user Excel: {e}")
        import traceback
        traceback.print_exc()
        return None, None

def get_vulnerability_status_from_user_excel(user_excel_path, ws_name, status_col, name_col, vuln_name):
    """Get the status of a vulnerability from user's Excel file"""
    try:
        user_wb = load_workbook(user_excel_path)
        
        # Find worksheet
        ws = None
        for sheet_name in user_wb.sheetnames:
            if ws_name.lower() in sheet_name.lower() or "api vapt" in sheet_name.lower() or "api_vapt" in sheet_name.lower():
                ws = user_wb[sheet_name]
                break
        
        if not ws or not status_col:
            user_wb.close()
            return None
        
        # Normalize vulnerability name for matching
        normalized_vuln = normalize_name_for_matching(vuln_name)
        
        # Search for the vulnerability in the worksheet
        for row_idx in range(2, ws.max_row + 1):
            name_cell = ws.cell(row=row_idx, column=name_col)
            if not name_cell or not name_cell.value:
                continue
            
            row_vuln_name = str(name_cell.value).strip()
            if not row_vuln_name:
                continue
            
            # Normalize and compare
            normalized_row = normalize_name_for_matching(row_vuln_name)
            
            # Check if names match
            if (normalized_vuln == normalized_row or 
                normalized_vuln in normalized_row or 
                normalized_row in normalized_vuln):
                
                # Get status from the Status column
                status_cell = ws.cell(row=row_idx, column=status_col)
                status_value = str(status_cell.value).strip().upper() if status_cell.value else ""
                
                user_wb.close()
                return status_value
        
        user_wb.close()
        return None
        
    except Exception as e:
        print(f"❌ Error getting vulnerability status from user Excel: {e}")
        import traceback
        traceback.print_exc()
        return None

def extract_poc_images_from_user_excel(user_excel_path, ws_name="API VAPT"):
    """Extract POC images from user's Excel file and return mapping"""
    poc_mapping = {}
    temp_dir = None
    
    try:
        wb = load_workbook(user_excel_path)
        
        # Find worksheet
        ws = None
        for sheet_name in wb.sheetnames:
            if ws_name.lower() in sheet_name.lower() or "api vapt" in sheet_name.lower() or "api_vapt" in sheet_name.lower():
                ws = wb[sheet_name]
                break
        
        if not ws:
            print(f"⚠️ Worksheet '{ws_name}' not found in user Excel")
            wb.close()
            return poc_mapping, temp_dir
        
        # Find POC column (merged cells in row 1 containing "POC")
        poc_col_start = None
        poc_col_end = None
        
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == 1 and merged_range.max_row == 1:
                first_cell = ws.cell(row=1, column=merged_range.min_col)
                if first_cell.value and 'poc' in str(first_cell.value).lower():
                    poc_col_start = merged_range.min_col
                    poc_col_end = merged_range.max_col
                    break
                
        if not poc_col_start:
            print("⚠️ POC column not found in user Excel")
            wb.close()
            return poc_mapping, temp_dir
        
        print(f"📍 Found POC columns from {get_column_letter(poc_col_start)} to {get_column_letter(poc_col_end)}")
        
        # Find "Name of Vulnerability" column
        name_col = None
        for col in range(1, min(10, ws.max_column + 1)):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and ('name of vulnerability' in str(cell_value).lower() or 'vulnerability name' in str(cell_value).lower()):
                name_col = col
                break
                
        if not name_col:
            name_col = 3  # Default to column C
        
        # Create temp directory for images
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        temp_dir = f"Temp_POC_API_{timestamp}"
        os.makedirs(temp_dir, exist_ok=True)
        
        # Extract images from worksheet and map to vulnerabilities
        if hasattr(ws, '_images'):
            for idx, img in enumerate(ws._images):
                try:
                    # Get image data
                    if hasattr(img, '_data'):
                        img_data = img._data() if callable(img._data) else img._data
                    else:
                        continue
                    
                    # Get image position
                    if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                        img_row = img.anchor._from.row + 1
                        img_col = img.anchor._from.col + 1
                    else:
                        continue
                    
                    # Check if image is in POC column range
                    if img_col < poc_col_start or img_col > poc_col_end:
                        continue
                    
                    # Get vulnerability name from the row (column C or name_col)
                    name_cell = ws.cell(row=img_row, column=name_col)
                    if not name_cell or not name_cell.value:
                        continue
                    
                    vuln_name = str(name_cell.value).strip()
                    if not vuln_name:
                        continue
                    
                    # Save image to temp file
                    img_filename = f"image_{img_row}_{idx}.png"
                    img_path = os.path.join(temp_dir, img_filename)
                    
                    with open(img_path, 'wb') as f:
                        f.write(img_data)
                    
                    # Normalize name for matching
                    normalized = normalize_name_for_matching(vuln_name)
                    if normalized not in poc_mapping:
                        poc_mapping[normalized] = []
                    
                    poc_mapping[normalized].append({
                        'path': img_path,
                        'original_name': vuln_name,
                        'row': img_row,
                        'col': img_col
                    })
                    
                    print(f"  ✅ Extracted image from row {img_row}, col {get_column_letter(img_col)}: {vuln_name[:50]}")
                    
                except Exception as e:
                    print(f"  ⚠️ Error extracting image {idx}: {e}")
                    continue
        
        wb.close()
        print(f"✅ Extracted {sum(len(imgs) for imgs in poc_mapping.values())} POC images from {len(poc_mapping)} vulnerabilities")
        
    except Exception as e:
        print(f"❌ Error extracting POC images from user Excel: {e}")
        import traceback
        traceback.print_exc()
    
    return poc_mapping, temp_dir

def copy_metadata_sections_from_user_excel(user_wb, target_ws, sections_to_copy):
    """Copy specific sections from user's Meta_Data worksheet to target worksheet"""
    try:
        # Find Meta_Data worksheet in user workbook
        user_meta_ws = None
        for sheet_name in user_wb.sheetnames:
            if 'meta' in sheet_name.lower() or 'metadata' in sheet_name.lower():
                user_meta_ws = user_wb[sheet_name]
                break
        
        if not user_meta_ws:
            print("⚠️ Meta_Data worksheet not found in user Excel")
            return 0
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Find current row in target worksheet
        current_row = target_ws.max_row + 1
        if current_row == 1:
            current_row = 1
        
        sections_copied = 0
        
        # Find and copy each section
        for section_title in sections_to_copy:
            # Find section in user Meta_Data worksheet
            section_start_row = None
            
            for row_idx in range(1, user_meta_ws.max_row + 1):
                cell_value = user_meta_ws.cell(row=row_idx, column=1).value
                if cell_value and str(cell_value).strip().upper() == section_title.upper():
                    section_start_row = row_idx
                    break
            
            if section_start_row:
                # Determine how many rows to copy based on section type
                rows_to_copy = 0
                section_upper = section_title.upper()
                
                if section_upper == "ORGANIZATION INFORMATION":
                    # Copy 4 rows: the header row + 3 rows after it
                    rows_to_copy = 4
                elif section_upper == "APPLICATION TYPE":
                    # Copy 2 rows: the header row + 1 row after it
                    rows_to_copy = 2
                elif section_upper.startswith("ASSET "):
                    # Copy 6 rows: the header row + 5 rows after it
                    rows_to_copy = 7
                else:
                    # Default: copy until next section or end of worksheet
                    # Find end of section (next non-empty cell in column A that looks like a section header)
                    section_end_row = None
                    for end_idx in range(section_start_row + 1, user_meta_ws.max_row + 1):
                        next_cell = user_meta_ws.cell(row=end_idx, column=1).value
                        if next_cell and str(next_cell).strip():
                            # Check if it's a section header (all caps or mostly uppercase)
                            cell_str = str(next_cell).strip()
                            if cell_str.isupper() or (len(cell_str) > 3 and sum(1 for c in cell_str if c.isupper()) > len(cell_str) * 0.5):
                                section_end_row = end_idx - 1
                                break
                    if not section_end_row:
                        section_end_row = user_meta_ws.max_row
                    rows_to_copy = section_end_row - section_start_row + 1
                
                # Calculate end row
                section_end_row = min(section_start_row + rows_to_copy - 1, user_meta_ws.max_row)
                
                # Copy section
                for row_idx in range(section_start_row, section_end_row + 1):
                    source_row = user_meta_ws[row_idx]
                    target_row = current_row
                    
                    # Copy cells with formatting
                    for col_idx, source_cell in enumerate(source_row, 1):
                        target_cell = target_ws.cell(row=target_row, column=col_idx)
                        target_cell.value = source_cell.value
                        
                        # Copy formatting
                        if source_cell.has_style:
                            try:
                                target_cell.font = source_cell.font.copy() if source_cell.font else data_font
                                target_cell.border = source_cell.border.copy() if source_cell.border else thin_border
                                target_cell.fill = source_cell.fill.copy() if source_cell.fill else None
                                target_cell.alignment = source_cell.alignment.copy() if source_cell.alignment else data_alignment
                            except:
                                target_cell.font = data_font
                                target_cell.border = thin_border
                                target_cell.alignment = data_alignment
                        
                        # Copy column width
                        if col_idx <= 2:
                            source_col_letter = get_column_letter(col_idx)
                            if source_col_letter in user_meta_ws.column_dimensions:
                                source_width = user_meta_ws.column_dimensions[source_col_letter].width
                                if source_width:
                                    target_ws.column_dimensions[source_col_letter].width = source_width
                    
                    # Copy row height
                    if row_idx in user_meta_ws.row_dimensions:
                        source_height = user_meta_ws.row_dimensions[row_idx].height
                        if source_height:
                            target_ws.row_dimensions[target_row].height = source_height
                    
                    current_row += 1
                
                # Add spacing row
                current_row += 1
                sections_copied += 1
                print(f"✅ Copied section '{section_title}'")
        
        return sections_copied
        
    except Exception as e:
        print(f"❌ Error copying metadata sections: {e}")
        import traceback
        traceback.print_exc()
        return 0

def copy_revalidation_cycle_sections_from_user_excel(user_wb, target_ws):
    """Copy all existing Revalidation Cycle sections from user's Excel to target worksheet"""
    try:
        # Find Meta_Data worksheet in user workbook
        user_meta_ws = None
        for sheet_name in user_wb.sheetnames:
            if 'meta' in sheet_name.lower() or 'metadata' in sheet_name.lower():
                user_meta_ws = user_wb[sheet_name]
                break
        
        if not user_meta_ws:
            print("⚠️ Meta_Data worksheet not found in user Excel for Revalidation Cycles")
            return 0
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Find current row in target worksheet
        current_row = target_ws.max_row + 1
        if current_row == 1:
            current_row = 1
        
        cycles_copied = 0
        
        # Find all Revalidation Cycle sections
        cycle_sections = []
        i = 1
        while i <= user_meta_ws.max_row:
            cell_value = user_meta_ws.cell(row=i, column=1).value
            if cell_value:
                value_str = str(cell_value).strip()
                if 'revalidation cycle' in value_str.lower():
                    # Found a Revalidation Cycle section
                    section_start = i
                    section_end = i
                    
                    # A Revalidation Cycle section typically has:
                    # - Header row (REVALIDATION CYCLE - X)
                    # - Start Date row
                    # - End Date row
                    # So we need to find where this section ends
                    
                    # Look ahead to find the end of this section
                    found_start_date = False
                    found_end_date = False
                    
                    for j in range(i + 1, user_meta_ws.max_row + 1):
                        next_cell = user_meta_ws.cell(row=j, column=1).value
                        if next_cell:
                            next_str = str(next_cell).strip().lower()
                            
                            # Check if it's the next Revalidation Cycle section
                            if 'revalidation cycle' in next_str:
                                # Found next cycle, end current section at previous row
                                section_end = j - 1
                                break
                            
                            # Track if we found Start Date and End Date
                            if 'start date' in next_str:
                                found_start_date = True
                            elif 'end date' in next_str:
                                found_end_date = True
                            
                            # If we found both dates and next row is empty or another section
                            if found_start_date and found_end_date:
                                # Check if next row is empty or another section
                                if j < user_meta_ws.max_row:
                                    next_next_cell = user_meta_ws.cell(row=j + 1, column=1).value
                                    if not next_next_cell or (next_next_cell and 'revalidation cycle' not in str(next_next_cell).strip().lower()):
                                        # Next row is empty or not a cycle, so this section ends here
                                        section_end = j
                                        break
                    
                    # If no end found, use worksheet end
                    if section_end == section_start:
                        # At minimum, include Start Date and End Date rows
                        # Look for them
                        for j in range(i + 1, min(i + 10, user_meta_ws.max_row + 1)):
                            next_cell = user_meta_ws.cell(row=j, column=1).value
                            if next_cell:
                                next_str = str(next_cell).strip().lower()
                                if 'end date' in next_str:
                                    section_end = j
                                    break
                        
                        # If still not found, use worksheet end
                        if section_end == section_start:
                            section_end = user_meta_ws.max_row
                    
                    cycle_sections.append((section_start, section_end))
                    i = section_end + 1
                else:
                    i += 1
            else:
                i += 1
        
        # Copy each Revalidation Cycle section
        for section_start, section_end in cycle_sections:
            # Add one empty row before each section (except the first one)
            if cycles_copied > 0:
                current_row += 1
            
            # Copy the section
            for row_idx in range(section_start, section_end + 1):
                source_row = user_meta_ws[row_idx]
                target_row = current_row
                
                # Copy cells with formatting
                for col_idx, source_cell in enumerate(source_row, 1):
                    target_cell = target_ws.cell(row=target_row, column=col_idx)
                    target_cell.value = source_cell.value
                    
                    # Copy formatting
                    if source_cell.has_style:
                        try:
                            target_cell.font = source_cell.font.copy() if source_cell.font else data_font
                            target_cell.border = source_cell.border.copy() if source_cell.border else thin_border
                            target_cell.fill = source_cell.fill.copy() if source_cell.fill else None
                            target_cell.alignment = source_cell.alignment.copy() if source_cell.alignment else data_alignment
                        except:
                            target_cell.font = data_font
                            target_cell.border = thin_border
                            target_cell.alignment = data_alignment
                
                # Copy row height
                if row_idx in user_meta_ws.row_dimensions:
                    source_height = user_meta_ws.row_dimensions[row_idx].height
                    if source_height:
                        target_ws.row_dimensions[target_row].height = source_height
                
                current_row += 1
            
            cycles_copied += 1
            print(f"✅ Copied Revalidation Cycle section from rows {section_start}-{section_end}")
        
        return cycles_copied
        
    except Exception as e:
        print(f"❌ Error copying Revalidation Cycle sections: {e}")
        import traceback
        traceback.print_exc()
        return 0

def find_next_revalidation_cycle_number(user_excel_path=None, target_ws=None):
    """Find the next available Revalidation Cycle number by checking user Excel first, then target worksheet"""
    cycle_numbers = []
    
    # First, check user Excel file if provided
    if user_excel_path and os.path.exists(user_excel_path):
        try:
            user_wb = load_workbook(user_excel_path)
            user_meta_ws = None
            
            # Find Meta_Data worksheet in user workbook
            for sheet_name in user_wb.sheetnames:
                if 'meta' in sheet_name.lower() or 'metadata' in sheet_name.lower():
                    user_meta_ws = user_wb[sheet_name]
                    break
            
            if user_meta_ws:
                for row_idx in range(1, user_meta_ws.max_row + 1):
                    cell_value = user_meta_ws.cell(row=row_idx, column=1).value
                    if cell_value:
                        value_str = str(cell_value).strip()
                        if 'revalidation cycle' in value_str.lower():
                            # Extract number
                            match = re.search(r'revalidation cycle\s*-\s*(\d+)', value_str, re.IGNORECASE)
                            if match:
                                cycle_numbers.append(int(match.group(1)))
            
            user_wb.close()
        except Exception as e:
            print(f"⚠️ Error checking user Excel for Revalidation Cycles: {e}")
    
    # Also check target worksheet if provided
    if target_ws:
        for row_idx in range(1, target_ws.max_row + 1):
            cell_value = target_ws.cell(row=row_idx, column=1).value
            if cell_value:
                value_str = str(cell_value).strip()
                if 'revalidation cycle' in value_str.lower():
                    # Extract number
                    match = re.search(r'revalidation cycle\s*-\s*(\d+)', value_str, re.IGNORECASE)
                    if match:
                        cycle_numbers.append(int(match.group(1)))
    
    if not cycle_numbers:
        return 1
    
    return max(cycle_numbers) + 1

def create_api_followup_metadata_worksheet(wb, form_data, user_excel_path=None):
    """Create Meta_Data worksheet for follow-up audit"""
    try:
        print("📊 Creating Meta_Data worksheet...")
        ws = wb.create_sheet("Meta_Data")
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 60
        
        # Copy sections from user Excel if provided
        user_wb = None
        if user_excel_path and os.path.exists(user_excel_path):
            try:
                user_wb = load_workbook(user_excel_path)
                
                # Sections to copy
                sections_to_copy = ['ORGANIZATION INFORMATION', 'APPLICATION TYPE']
                
                # Add ASSET sections (1-10)
                for i in range(1, 11):
                    sections_to_copy.append(f'ASSET {i}')
                
                # Copy sections
                copy_metadata_sections_from_user_excel(user_wb, ws, sections_to_copy)
                
            except Exception as e:
                print(f"⚠️ Error copying sections from user Excel: {e}")
        
        # Copy all existing Revalidation Cycle sections from user Excel
        if user_wb:
            try:
                cycles_copied = copy_revalidation_cycle_sections_from_user_excel(user_wb, ws)
                print(f"✅ Copied {cycles_copied} existing Revalidation Cycle section(s) from user Excel")
            except Exception as e:
                print(f"⚠️ Error copying Revalidation Cycle sections: {e}")
        
        # Close user workbook if it was opened
        if user_wb:
            user_wb.close()
        
        # Add new Revalidation Cycle section
        # Check user Excel for existing cycles, then check current worksheet
        cycle_number = find_next_revalidation_cycle_number(user_excel_path, ws)
        cycle_title = f'REVALIDATION CYCLE - {cycle_number}'
        
        # Find current row
        current_row = ws.max_row + 1
        if current_row == 1:
            current_row = 1
        
        # Add one empty row before new Revalidation Cycle section
        current_row += 1
        
        # Write Revalidation Cycle section
        cell = ws.cell(row=current_row, column=1, value=cycle_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        cell = ws.cell(row=current_row, column=2, value='')
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
            
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Format dates
        start_date = form_data.get('startDate', '')
        end_date = form_data.get('endDate', '')
        
        if start_date:
            try:
                date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                start_date = date_obj.strftime('%d.%m.%Y')
            except:
                pass
        
        if end_date:
            try:
                date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                end_date = date_obj.strftime('%d.%m.%Y')
            except:
                pass
        
        # Write Start Date and End Date
        for field_name, field_value in [('Start Date', start_date), ('End Date', end_date)]:
            cell = ws.cell(row=current_row, column=1, value=field_name)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=current_row, column=2, value=field_value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
                
            ws.row_dimensions[current_row].height = 20
            current_row += 1
        
        print(f"✅ Meta_Data worksheet created")
        
    except Exception as e:
        print(f"❌ Error creating Meta_Data worksheet: {e}")
        import traceback
        traceback.print_exc()

@api_vapt_followup_bp.route('/process_api_vapt_follow_up_audit_report', methods=['POST'])
def process_api_follow_up_audit_report():
    """Main endpoint to process API VAPT Follow-Up Audit Report."""
    temp_files = []
    temp_dirs = []
    
    try:
        print("\n" + "="*80)
        print("🚀 Processing API VAPT Follow-Up Audit Report")
        print("="*80)
        
        # Extract form data
        form_data = {
            'startDate': request.form.get('startDate'),
            'endDate': request.form.get('endDate'),
        }
        
        # Get vulnerability data from JSON
        vulnerability_data_json = request.form.get('vulnerabilityData')
        if not vulnerability_data_json:
            return jsonify({"error": "Vulnerability data not provided"}), 400
        
        vulnerability_data = json.loads(vulnerability_data_json)
        print(f"📋 Processing {len(vulnerability_data)} vulnerabilities")
        
        # Get filtered vulnerabilities (NA/Closed) from JSON
        filtered_vulnerabilities = []
        filtered_vuln_json = request.form.get('filteredVulnerabilities')
        if filtered_vuln_json:
            filtered_vulnerabilities = json.loads(filtered_vuln_json)
            print(f"📋 Found {len(filtered_vulnerabilities)} filtered vulnerabilities (NA/Closed)")
        
        # Get user Excel file
        user_excel_file = request.files.get('userExcelFile')
        if not user_excel_file or user_excel_file.filename == '':
            return jsonify({"error": "User Excel file not provided"}), 400
        
        # Save user Excel to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            user_excel_file.save(tmp.name)
            user_excel_path = tmp.name
            temp_files.append(user_excel_path)
        
        print(f"📁 User Excel file: {user_excel_file.filename}")
        
        # Load user Excel workbook
        user_wb = load_workbook(user_excel_path)
        
        # Find "API VAPT" worksheet
        user_ws_name = None
        for sheet_name in user_wb.sheetnames:
            if 'api vapt' in sheet_name.lower() or 'api_vapt' in sheet_name.lower():
                user_ws_name = sheet_name
                break
        
        if not user_ws_name:
            # Try first worksheet
            user_ws_name = user_wb.sheetnames[0]
        
        print(f"📄 Using worksheet: {user_ws_name}")
        
        # Create new workbook
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Copy "API VAPT" worksheet from user Excel
        copy_success, temp_image_files = copy_worksheet_with_images(user_wb, user_ws_name, wb, "API VAPT")
        if not copy_success:
            user_wb.close()
            return jsonify({"error": "Failed to copy worksheet from user Excel"}), 500
        
        # Add temp image files to cleanup list (will be cleaned up after saving)
        temp_files.extend(temp_image_files)
        user_wb.close()
        
        # Get the copied worksheet
        ws = wb["API VAPT"]
        
        # Find first empty cell in row 1
        first_empty_col = find_first_empty_cell_in_row(ws, 1)
        print(f"📍 First empty column in row 1: {get_column_letter(first_empty_col)}")
        
        # Get header style from A1
        header_cell = ws.cell(row=1, column=1)
        header_font = header_cell.font.copy() if header_cell.font else Font(name='Arial', size=11, bold=True)
        header_fill = header_cell.fill.copy() if header_cell.fill else PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = header_cell.alignment.copy() if header_cell.alignment else Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_border = header_cell.border.copy() if header_cell.border else Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Format end date for header
        end_date = form_data.get('endDate', '')
        if end_date:
            try:
                date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                end_date_formatted = date_obj.strftime('%d.%m.%Y')
            except:
                end_date_formatted = end_date
        else:
            end_date_formatted = ''
        
        # Add Status column
        status_col = first_empty_col
        status_header = f"Status - {end_date_formatted}"
        # Get the actual cell (handle merged cells)
        status_cell = get_top_left_cell_of_merged_range(ws, 1, status_col)
        status_cell.value = status_header
        status_cell.font = header_font
        status_cell.fill = header_fill
        status_cell.alignment = header_alignment
        status_cell.border = header_border
        
        # Add Remark column
        remark_col = status_col + 1
        remark_header = f"Remark - {end_date_formatted}"
        # Get the actual cell (handle merged cells)
        remark_cell = get_top_left_cell_of_merged_range(ws, 1, remark_col)
        remark_cell.value = remark_header
        remark_cell.font = header_font
        remark_cell.fill = header_fill
        remark_cell.alignment = header_alignment
        remark_cell.border = header_border
        
        # Add POC column (merged 7 cells)
        poc_col_start = remark_col + 1
        poc_col_end = poc_col_start + 6
        poc_header = f"POC - {end_date_formatted}"
        # Get the actual cell (handle merged cells)
        poc_cell = get_top_left_cell_of_merged_range(ws, 1, poc_col_start)
        poc_cell.value = poc_header
        poc_cell.font = header_font
        poc_cell.fill = header_fill
        poc_cell.alignment = header_alignment
        poc_cell.border = header_border
        
        # Merge POC cells
        ws.merge_cells(start_row=1, start_column=poc_col_start, end_row=1, end_column=poc_col_end)
        
        # Set column widths: Status=20, Remark=30, POC columns=20 each
        ws.column_dimensions[get_column_letter(status_col)].width = 20
        ws.column_dimensions[get_column_letter(remark_col)].width = 30
        for poc_col in range(poc_col_start, poc_col_end + 1):
            ws.column_dimensions[get_column_letter(poc_col)].width = 20
        
        # Create alignment style (center vertical and horizontal, wrap text)
        center_wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Create font styles for status colors
        red_font = Font(color='FF0000', bold=True)  # Red for "Open"
        green_font = Font(color='008000', bold=True)  # Green for "Closed"
        
        # Find "Name of Vulnerability" column (usually column C)
        name_col = None
        for col in range(1, min(10, ws.max_column + 1)):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and ('name of vulnerability' in str(cell_value).lower() or 'vulnerability name' in str(cell_value).lower()):
                name_col = col
                break
        
        if not name_col:
            name_col = 3  # Default to column C
        
        print(f"📍 Name of Vulnerability column: {get_column_letter(name_col)}")
        
        # Create mapping of vulnerability names to data
        vuln_data_map = {}
        for vuln in vulnerability_data:
            name = vuln.get('name', '').strip()
            if name:
                normalized = normalize_name_for_matching(name)
                vuln_data_map[normalized] = vuln
        
        # Create mapping of filtered vulnerability names (NA/Closed)
        filtered_vuln_map = {}
        for vuln in filtered_vulnerabilities:
            name = vuln.get('name', '').strip()
            if name:
                normalized = normalize_name_for_matching(name)
                filtered_vuln_map[normalized] = vuln
        
        # Extract POC images from user Excel
        poc_mapping, temp_dir = extract_poc_images_from_user_excel(user_excel_path, user_ws_name)
        if temp_dir:
            temp_dirs.append(temp_dir)
        
        # Find latest Status column in user's Excel and create status mapping
        user_status_col, user_status_date = find_latest_status_column_in_user_excel(user_excel_path, user_ws_name)
        print(f"📍 Latest Status column in user Excel: {get_column_letter(user_status_col) if user_status_col else 'None'}")
        
        # Create mapping of vulnerability names to their statuses from user Excel
        user_status_map = {}
        if user_status_col:
            try:
                user_wb = load_workbook(user_excel_path)
                user_ws = None
                for sheet_name in user_wb.sheetnames:
                    if user_ws_name.lower() in sheet_name.lower() or "api vapt" in sheet_name.lower() or "api_vapt" in sheet_name.lower():
                        user_ws = user_wb[sheet_name]
                        break
                
                if user_ws:
                    # Find name column in user Excel
                    user_name_col = None
                    for col in range(1, min(10, user_ws.max_column + 1)):
                        cell_value = user_ws.cell(row=1, column=col).value
                        if cell_value and ('name of vulnerability' in str(cell_value).lower() or 'vulnerability name' in str(cell_value).lower()):
                            user_name_col = col
                            break
                    
                    if not user_name_col:
                        user_name_col = 3  # Default to column C
                    
                    # Build status mapping
                    for row_idx in range(2, user_ws.max_row + 1):
                        name_cell = user_ws.cell(row=row_idx, column=user_name_col)
                        if not name_cell or not name_cell.value:
                            continue
                        
                        vuln_name = str(name_cell.value).strip()
                        if not vuln_name:
                            continue
                        
                        # Get status
                        status_cell = user_ws.cell(row=row_idx, column=user_status_col)
                        status_value = str(status_cell.value).strip().upper() if status_cell.value else ""
                        
                        # Normalize and store
                        normalized = normalize_name_for_matching(vuln_name)
                        user_status_map[normalized] = status_value
                    
                    print(f"📍 Created status mapping for {len(user_status_map)} vulnerabilities from user Excel")
                
                user_wb.close()
            except Exception as e:
                print(f"⚠️ Error creating status mapping: {e}")
        
        # Process each row and add status/remark
        for row_idx in range(2, ws.max_row + 1):
            # Get vulnerability name from the row
            name_cell = ws.cell(row=row_idx, column=name_col)
            if not name_cell or not name_cell.value:
                continue

            vuln_name = str(name_cell.value).strip()
            if not vuln_name:
                continue
            
            # Normalize name for matching
            normalized = normalize_name_for_matching(vuln_name)
            
            # Check status from user's Excel mapping
            user_status = user_status_map.get(normalized)
            if not user_status:
                # Try partial matching
                for norm_name, status_val in user_status_map.items():
                    if normalized == norm_name or normalized in norm_name or norm_name in normalized:
                        user_status = status_val
                        break
            
            if user_status:
                print(f"  📍 Vulnerability '{vuln_name[:50]}' has status '{user_status}' in user Excel")
            
            # Find matching vulnerability data
            matched_vuln = None
            for norm_name, vuln_data in vuln_data_map.items():
                # Check exact match or partial match
                if normalized == norm_name or normalized in norm_name or norm_name in normalized:
                    matched_vuln = vuln_data
                    break
            
            # Check if status is Closed, Closed With Exception, or NA (from user Excel or filtered list)
            is_closed_or_na = False
            if user_status and (user_status == 'CLOSED' or user_status == 'CLOSED WITH EXCEPTION' or user_status == 'NA'):
                is_closed_or_na = True
            elif not user_status:
                # Check filtered vulnerabilities list
                matched_filtered = None
                for norm_name, filtered_vuln in filtered_vuln_map.items():
                    if normalized == norm_name or normalized in norm_name or norm_name in normalized:
                        matched_filtered = filtered_vuln
                        break
                if matched_filtered:
                    is_closed_or_na = True
            
            if is_closed_or_na:
                # Add "NA" for Status - handle merged cells
                status_cell = get_top_left_cell_of_merged_range(ws, row_idx, status_col)
                status_cell.value = "NA"
                status_cell.alignment = center_wrap_alignment
                
                # Add "NA" for Remark - handle merged cells
                remark_cell = get_top_left_cell_of_merged_range(ws, row_idx, remark_col)
                remark_cell.value = "NA"
                remark_cell.alignment = center_wrap_alignment
                
                # Merge POC cells and add "NA"
                poc_cell = get_top_left_cell_of_merged_range(ws, row_idx, poc_col_start)
                poc_cell.value = "NA"
                poc_cell.alignment = center_wrap_alignment
                
                # Merge POC cells for this row
                try:
                    ws.merge_cells(start_row=row_idx, start_column=poc_col_start, end_row=row_idx, end_column=poc_col_end)
                except:
                    pass
                
                # Apply alignment to all POC cells in merged range
                for poc_col in range(poc_col_start, poc_col_end + 1):
                    poc_cell_temp = ws.cell(row=row_idx, column=poc_col)
                    poc_cell_temp.alignment = center_wrap_alignment
                
                print(f"  ✅ Added NA for Closed/NA vulnerability: {vuln_name[:50]}")
            elif matched_vuln:
                # Add status - handle merged cells
                status_value = matched_vuln.get('status', '')
                status_cell = get_top_left_cell_of_merged_range(ws, row_idx, status_col)
                status_cell.value = status_value
                status_cell.alignment = center_wrap_alignment
                
                # Apply color based on status
                status_upper = str(status_value).strip().upper()
                if status_upper == 'OPEN':
                    status_cell.font = red_font
                elif status_upper == 'CLOSED' or status_upper == 'CLOSED WITH EXCEPTION':
                    status_cell.font = green_font
                
                # Add remark - handle merged cells
                remark_cell = get_top_left_cell_of_merged_range(ws, row_idx, remark_col)
                remark_cell.value = matched_vuln.get('remark', '')
                remark_cell.alignment = center_wrap_alignment
                
                # Add POC images
                # Image column order: V, W, X, Y, Z, AA, U
                # If poc_col_start is U (21), then: V=22, W=23, X=24, Y=25, Z=26, AA=27, U=21
                image_columns = [
                    poc_col_start + 1,  # V (first after start)
                    poc_col_start + 2,  # W
                    poc_col_start + 3,  # X
                    poc_col_start + 4,  # Y
                    poc_col_start + 5,  # Z
                    poc_col_start + 6,  # AA
                    poc_col_start       # U (start column last)
                ]
                
                # Find matching images using improved matching logic
                matching_images = []
                vuln_short = vuln_name[:170]
                vuln_normalized = normalized
                vuln_short_lower = vuln_short.lower()
                vuln_normalized_lower = vuln_normalized.lower()
                
                for norm_img_name, img_list in poc_mapping.items():
                    # Check if image name matches vulnerability name
                    # Match using both original and normalized names (ignoring last trailing number)
                    matches = (
                        vuln_short_lower in norm_img_name or 
                        norm_img_name in vuln_short_lower or
                        vuln_normalized_lower == norm_img_name or
                        vuln_normalized_lower in norm_img_name or
                        norm_img_name in vuln_normalized_lower
                    )
                    
                    if matches:
                        matching_images.extend(img_list)
                
                # Sort images by original name (to handle numbered images)
                matching_images.sort(key=lambda x: x.get('original_name', ''))
                
                # Insert images
                for img_idx, img_info in enumerate(matching_images[:7]):  # Max 7 images
                    if img_idx < len(image_columns):
                        col = image_columns[img_idx]
                        img_path = img_info['path']
                        
                        try:
                            img = Image(img_path)
                            
                            # Get original dimensions
                            original_width = img.width
                            original_height = img.height
                            
                            # Resize image to height=30px, width proportionally (maintain aspect ratio)
                            target_height = 30
                            aspect_ratio = original_width / original_height if original_height > 0 else 1
                            img.height = target_height
                            img.width = int(target_height * aspect_ratio)
                            
                            # Calculate cell position
                            cell_ref = f"{get_column_letter(col)}{row_idx}"
                            img.anchor = cell_ref
                            
                            ws.add_image(img)
                            
                            # Apply alignment to POC cell (even if it has an image)
                            poc_cell_img = ws.cell(row=row_idx, column=col)
                            poc_cell_img.alignment = center_wrap_alignment
                            
                            print(f"  ✅ Added image to {cell_ref}")
                        except Exception as e:
                            print(f"  ⚠️ Error adding image to row {row_idx}, col {col}: {e}")
                
                # Apply alignment to all POC cells (including empty ones)
                for poc_col in range(poc_col_start, poc_col_end + 1):
                    poc_cell_temp = ws.cell(row=row_idx, column=poc_col)
                    poc_cell_temp.alignment = center_wrap_alignment
            else:
                # Check if this vulnerability was filtered out (NA/Closed)
                matched_filtered = None
                for norm_name, filtered_vuln in filtered_vuln_map.items():
                    # Check exact match or partial match
                    if normalized == norm_name or normalized in norm_name or norm_name in normalized:
                        matched_filtered = filtered_vuln
                        break
                
                if matched_filtered:
                    # Add "NA" for Status - handle merged cells
                    status_cell = get_top_left_cell_of_merged_range(ws, row_idx, status_col)
                    status_cell.value = "NA"
                    
                    # Add "NA" for Remark - handle merged cells
                    remark_cell = get_top_left_cell_of_merged_range(ws, row_idx, remark_col)
                    remark_cell.value = "NA"
                    
                    # Merge POC cells and add "NA"
                    poc_cell = get_top_left_cell_of_merged_range(ws, row_idx, poc_col_start)
                    poc_cell.value = "NA"
                    
                    # Merge POC cells for this row
                    try:
                        ws.merge_cells(start_row=row_idx, start_column=poc_col_start, end_row=row_idx, end_column=poc_col_end)
                    except:
                        pass
                    
                    print(f"  ✅ Added NA for filtered vulnerability: {vuln_name[:50]}")
        
        # Create Meta_Data worksheet
        create_api_followup_metadata_worksheet(wb, form_data, user_excel_path)
        
        # Apply borders to all cells in the content range
        apply_borders_to_content_range(ws)
        
        # Save to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            wb.save(temp_file.name)
            temp_file_path = temp_file.name
            temp_files.append(temp_file_path)
        
        # Read the final file
        output = BytesIO()
        with open(temp_file_path, 'rb') as f:
            output.write(f.read())
        output.seek(0)
        
        filename = generate_api_followup_filename('', form_data.get('endDate', ''))
        
        print(f"✅ Excel file created: {filename}")
        
        response = make_response(send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        ))
        
        return response
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

    finally:
        # Clean up temporary files
        for temp_file in temp_files:
            try:
                os.unlink(temp_file)
            except:
                pass
        
        # Clean up temporary directories
        for temp_dir in temp_dirs:
            try:
                shutil.rmtree(temp_dir)
            except:
                pass
