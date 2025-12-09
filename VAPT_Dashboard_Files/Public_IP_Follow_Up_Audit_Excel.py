# Public IP VAPT Follow-Up Audit Excel Report Generator
from flask import Blueprint, request, send_file, make_response, jsonify, session
import re
import io
import pandas as pd
import zipfile
import os
import math
from io import BytesIO
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import copy

# Create a Blueprint for Public IP VAPT Follow-Up routes
public_ip_vapt_followup_bp = Blueprint('public_ip_vapt_followup', __name__)

# Global variables to store user's old Excel data
PUBLIC_IP_VAPT_USER_ROW_DATA = {}
PUBLIC_IP_VAPT_USER_VULNERABILITIES = set()

def extract_vulnerability_names_from_excel(excel_file):
    """
    Extract 'Name of Vulnerability' column (Column C) from Excel file.
    Returns a set of vulnerability names.
    """
    try:
        # Read the Excel file
        df = pd.read_excel(excel_file, sheet_name='Public_IP_VAPT')
        
        print(f"\n🔍 DEBUG: extract_vulnerability_names_from_excel")
        print(f"   Excel file shape: {df.shape}")
        print(f"   Available columns: {list(df.columns)[:10]}")
        
        # Check if 'Name of Vulnerability' column exists
        if 'Name of Vulnerability' not in df.columns:
            print("⚠️ Warning: 'Name of Vulnerability' column not found in Excel file")
            print(f"   Looking for columns containing 'name' or 'vulnerability':")
            matching_cols = [col for col in df.columns if 'name' in str(col).lower() or 'vulnerability' in str(col).lower()]
            print(f"      Found: {matching_cols}")
            return set()
        
        # Extract vulnerability names from 'Name of Vulnerability' column and remove NaN values
        vulnerability_names = df['Name of Vulnerability'].dropna().astype(str).str.strip()
        
        # Return as a set for efficient comparison
        result = set(vulnerability_names.tolist())
        print(f"   Extracted {len(result)} vulnerability names from old Excel (Name of Vulnerability column)")
        return result
        
    except Exception as e:
        print(f"❌ Error extracting vulnerability names from Excel: {e}")
        import traceback
        traceback.print_exc()
        return set()

def extract_full_row_data_from_excel(excel_file):
    """
    Extract full row data from user's Excel file Public_IP_VAPT worksheet.
    Returns a dictionary with vulnerability names as keys and full row data as values.
    """
    try:
        # Read the Excel file
        df = pd.read_excel(excel_file, sheet_name='Public_IP_VAPT')
        
        # Check if 'Name of Vulnerability' column exists
        if 'Name of Vulnerability' not in df.columns:
            print("Warning: 'Name of Vulnerability' column not found in Excel file")
            return {}
        
        # Create a dictionary to store full row data
        row_data = {}
        
        # Iterate through each row
        for index, row in df.iterrows():
            vuln_name = row['Name of Vulnerability']
            if pd.notna(vuln_name):
                vuln_name_str = str(vuln_name).strip()
                # Store the entire row as a dictionary
                row_data[vuln_name_str] = row.to_dict()
        
        print(f"Extracted {len(row_data)} full rows from user's Excel file")
        return row_data
        
    except Exception as e:
        print(f"Error extracting full row data from Excel: {e}")
        return {}

def convert_risk_to_camelcase(risk_value):
    """Convert risk value to CamelCase format"""
    if not risk_value:
        return ""
    
    risk_str = str(risk_value).strip().lower()
    if risk_str == 'critical':
        return 'Critical'
    elif risk_str == 'high':
        return 'High'
    elif risk_str == 'medium':
        return 'Medium'
    elif risk_str == 'low':
        return 'Low'
    else:
        # Return original value with first letter capitalized
        return risk_str.capitalize()

def compare_vulnerabilities(our_vulnerabilities, user_vulnerabilities):
    """
    Compare vulnerabilities between our Excel and user's Excel.
    Returns a dictionary with status for each vulnerability.
    """
    comparison_result = {}
    
    # Find common vulnerabilities (Open)
    common_vulnerabilities = our_vulnerabilities.intersection(user_vulnerabilities)
    for vuln in common_vulnerabilities:
        comparison_result[vuln] = "Open"
    
    # Find vulnerabilities only in user's Excel (Closed)
    closed_vulnerabilities = user_vulnerabilities - our_vulnerabilities
    for vuln in closed_vulnerabilities:
        comparison_result[vuln] = "Closed"
    
    # Find vulnerabilities only in our Excel (New)
    new_vulnerabilities = our_vulnerabilities - user_vulnerabilities
    for vuln in new_vulnerabilities:
        comparison_result[vuln] = "New"
    
    return comparison_result

def generate_public_ip_followup_filename(organization, end_date):
    """
    Generate static filename for Public IP VAPT Follow-Up Audit.
    Format: Public_IP_VAPT_Follow_Up_Audit
    """
    return "Public_IP_VAPT_Follow_Up_Audit_Report.xlsx"

def extract_poc_images(evidence_files, temp_dir=None):
    """Extract POC images from evidence files zip and return mapping of vulnerability names to image paths"""
    poc_mapping = {}
    
    if not evidence_files or evidence_files.filename == '':
        return poc_mapping
    
    try:
        import zipfile
        from io import BytesIO
        from datetime import datetime
        
        # Generate timestamp for temporary folder if not provided (format: DDMMYYYYHHMMSS)
        if temp_dir is None:
            current_ts = datetime.now().strftime('%d%m%Y%H%M%S')
            temp_dir = f"temp_poc_images_{current_ts}"
        
        # Read the zip file
        zip_data = evidence_files.read()
        
        with zipfile.ZipFile(BytesIO(zip_data), 'r') as zip_ref:
            file_list = zip_ref.namelist()
            
            # Find ALL POC folder
            poc_folder = None
            for file_path in file_list:
                if 'ALL POC' in file_path and file_path.endswith('/'):
                    poc_folder = file_path
                    break
            
            if not poc_folder:
                print("ALL POC folder not found in evidence files")
                return poc_mapping
            
            # Extract images from ALL POC folder
            image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff']
            
            for file_path in file_list:
                if file_path.startswith(poc_folder) and not file_path.endswith('/'):
                    # Get filename without extension
                    filename = os.path.basename(file_path)
                    name_without_ext = os.path.splitext(filename)[0]
                    
                    # Check if it's an image file
                    if any(filename.lower().endswith(ext) for ext in image_extensions):
                        # Extract the file to a temporary location
                        try:
                            with zip_ref.open(file_path) as f:
                                image_data = f.read()
                            
                            # Create temporary file
                            os.makedirs(temp_dir, exist_ok=True)
                            temp_file_path = os.path.join(temp_dir, filename)
                            
                            with open(temp_file_path, 'wb') as temp_file:
                                temp_file.write(image_data)
                            
                            # Map vulnerability name to image path
                            poc_mapping[name_without_ext] = temp_file_path
                            
                        except Exception as e:
                            print(f"Error extracting image {filename}: {e}")
                            continue
            
            print(f"Extracted {len(poc_mapping)} POC images")
            
    except Exception as e:
        print(f"Error extracting POC images: {e}")
    
    return poc_mapping


def insert_poc_images_to_excel(excel_path, poc_mapping, vulnerabilities_data):
    """Insert POC images directly into Excel using openpyxl with reduced size"""
    rows_with_objects = set()  # Track which rows have POC objects
    
    try:
        from openpyxl import load_workbook
        
        # Load the existing workbook
        wb = load_workbook(excel_path)
        
        # Get the Public_IP_VAPT worksheet
        if "Public_IP_VAPT" not in wb.sheetnames:
            print("Public_IP_VAPT worksheet not found")
            return rows_with_objects
        
        ws = wb["Public_IP_VAPT"]
        
        # Find POC columns by looking at the merged header in first row
        poc_col_start = None
        poc_col_end = None
        
        # Check for merged cells in row 1
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == 1 and merged_range.max_row == 1:
                # Check if this merged cell contains "POC"
                first_cell = ws.cell(row=1, column=merged_range.min_col)
                if first_cell.value and str(first_cell.value).strip() == "POC":
                    poc_col_start = merged_range.min_col
                    poc_col_end = merged_range.max_col
                    break
        
        if not poc_col_start or not poc_col_end:
            print("POC columns not found in worksheet")
            return rows_with_objects
        
        # Define the column order for image insertion: M, N, O, P, Q, R, L
        image_columns = [
            poc_col_start + 1,  # M
            poc_col_start + 2,  # N
            poc_col_start + 3,  # O
            poc_col_start + 4,  # P
            poc_col_start + 5,  # Q
            poc_col_start + 6,  # R
            poc_col_start       # L
        ]
        
        print(f"Found POC columns from column {poc_col_start} to {poc_col_end}")
        
        # Process each row and match vulnerabilities with POC images
        for row in range(2, ws.max_row + 1):
            vulnerability_cell = ws.cell(row=row, column=2)  # Vulnerabilities column (column B)
            vulnerability_text = str(vulnerability_cell.value) if vulnerability_cell.value else ""
            
            if vulnerability_text:
                # Split vulnerabilities (they might be on separate lines)
                vulnerabilities = [v.strip() for v in vulnerability_text.split('\n') if v.strip()]
                
                # Find all matching POC images for this row
                matching_images = []
                for vuln in vulnerabilities:
                    # Use only first 170 characters for matching
                    vuln_short = vuln[:170].lower()
                    
                    for image_name, image_path in poc_mapping.items():
                        # Use only first 170 characters of image name for matching
                        image_name_short = image_name[:170].lower()
                        
                        # Check if vulnerability name matches image name (case-insensitive, first 170 chars only)
                        if (vuln_short in image_name_short or image_name_short in vuln_short):
                            if image_path not in [img[1] for img in matching_images]:  # Avoid duplicates
                                matching_images.append((vuln, image_path))
                
                if matching_images:
                    try:
                        # Distribute images across columns: M, N, O, P, Q, R, L
                        num_images_to_insert = min(len(matching_images), 7)  # Max 7 images (one per column)
                        
                        for img_idx in range(num_images_to_insert):
                            vuln, matching_image = matching_images[img_idx]
                            col_idx = image_columns[img_idx]
                            
                            if os.path.exists(matching_image):
                                try:
                                    # Load the image
                                    from openpyxl.drawing.image import Image as XLImage
                                    img = XLImage(matching_image)
                                    
                                    # Get original dimensions
                                    original_width = img.width
                                    original_height = img.height
                                    
                                    # Resize image by reducing to 1/30th (30x reduction)
                                    if hasattr(img, 'width') and hasattr(img, 'height'):
                                        img.width = img.width / 30
                                        img.height = img.height / 30
                                    
                                    # Get cell reference (e.g., "M2", "N3", etc.)
                                    from openpyxl.utils import get_column_letter
                                    col_letter = get_column_letter(col_idx)
                                    cell_ref = f"{col_letter}{row}"
                                    
                                    # Insert image at the cell
                                    ws.add_image(img, cell_ref)
                                    
                                    print(f"✅ Inserted image {img_idx + 1} at {cell_ref} for vulnerability: {vuln} (reduced from {original_width}x{original_height})")
                                    
                                except Exception as e:
                                    print(f"⚠️ Error inserting image at column {col_idx}, row {row}: {e}")
                        
                        # Track this row as having POC objects
                        if num_images_to_insert > 0:
                            rows_with_objects.add(row)
                            
                    except Exception as e:
                        print(f"Error adding images for row {row}: {e}")
                        continue
        
        # Apply custom borders to ALL POC columns (L, M, N, O, P, Q, R)
        from openpyxl.styles import Border, Side
        
        # Define border styles for each POC column type
        # L column: left, above, below (NOT right)
        left_border = Border(
            left=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # M, N, O, P, Q columns: above and below only (NOT left or right)
        middle_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # R column: above, below, right (NOT left)
        right_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            right=Side(style='thin')
        )
        
        # Find all rows that are part of the data table (have content in any column)
        table_rows = set()
        
        # Always include header row
        table_rows.add(1)
        
        # Check all rows from 2 onwards to find data rows
        for row in range(2, ws.max_row + 1):
            has_content = False
            
            # Check if any cell in this row has content (excluding POC columns)
            for col in range(1, poc_col_start):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip() != "":
                    has_content = True
                    break
            
            if has_content:
                table_rows.add(row)
        
        # Apply custom borders to all table rows for each POC column
        for row_num in sorted(table_rows):
            # L column (poc_col_start): left, top, bottom
            ws.cell(row=row_num, column=poc_col_start).border = left_border
            
            # M, N, O, P, Q columns (middle columns): top, bottom only
            for col_idx in range(poc_col_start + 1, poc_col_end):
                ws.cell(row=row_num, column=col_idx).border = middle_border
            
            # R column (poc_col_end): top, bottom, right
            ws.cell(row=row_num, column=poc_col_end).border = right_border
        
        # Set row height to 60px (approximately 45 units in Excel) for rows with images
        for row_num in rows_with_objects:
            # Row height in Excel: 1 point ≈ 1.33 pixels, so 60px ≈ 45 points
            ws.row_dimensions[row_num].height = 45
        
        # Save the workbook
        wb.save(excel_path)
        print(f"\n✅ POC IMAGES INSERTION COMPLETE!")
        print(f"   Rows with POC images: {len(rows_with_objects)}")
        print(f"   Applied custom borders to {len(table_rows)} table rows")
        
    except Exception as e:
        print(f"Error inserting POC images: {e}")
        import traceback
        traceback.print_exc()
    
    return rows_with_objects

def process_nmap_zip_public_ip(nmap_file):
    """
    Process Nmap ZIP file for Public IP VAPT - Same logic as Infrastructure VAPT
    Returns: List of 6-element rows [HOST, PORT, SERVICE, HOST, PORT, SERVICE]
    Two IPs are placed side by side (columns A-C for first IP, columns D-F for second IP)
    """
    try:
        print("📁 Processing Nmap ZIP file for Public IP VAPT...")
        ip_ports = {}
        zip_data = nmap_file.read()
        
        with zipfile.ZipFile(BytesIO(zip_data), 'r') as zip_ref:
            file_list = zip_ref.namelist()
            
            for file_name in file_list:
                if file_name.endswith('/'):
                    continue
                
                try:
                    with zip_ref.open(file_name) as f:
                        file_content = f.read().decode('utf-8', errors='ignore')
                    
                    ip_pattern = r"Nmap scan report for (?:[a-zA-Z0-9.-]+ )?\(?(\d+\.\d+\.\d+\.\d+)\)?"
                    port_state_pattern = r"(\d+)/(tcp|udp)\s+(open|filtered|closed|unfiltered)\s+([\w-]*)"
                    filtered_ports_pattern = r"Not shown: (\d+) filtered tcp ports"
                    
                    lines = file_content.splitlines()
                    current_ip = None
                    has_filtered_ports = False

                    for line in lines:
                        ip_match = re.search(ip_pattern, line, re.IGNORECASE)
                        if ip_match:
                            current_ip = ip_match.group(1)
                            if current_ip not in ip_ports:
                                ip_ports[current_ip] = []
                            has_filtered_ports = False
                            continue
                        
                        if not current_ip:
                            continue
                        
                        filtered_match = re.search(filtered_ports_pattern, line)
                        if filtered_match:
                            has_filtered_ports = True
                            continue
                        
                        port_match = re.search(port_state_pattern, line)
                        if port_match:
                            port = port_match.group(1)
                            state = port_match.group(3)
                            service = port_match.group(4) or state
                            
                            if (port, service) not in ip_ports[current_ip]:
                                ip_ports[current_ip].append((port, service))
                    
                    if current_ip and has_filtered_ports and len(ip_ports[current_ip]) == 0:
                        ip_ports[current_ip].append(("Filtered", "Filtered"))

                    if file_name.endswith('.csv'):
                        try:
                            with zip_ref.open(file_name) as f:
                                csv_content = f.read()
                            
                            df = pd.read_csv(BytesIO(csv_content), 
                                        on_bad_lines="skip", 
                                        encoding="utf-8")
                            
                            df.columns = df.columns.str.lower().str.strip()
                            
                            if all(col in df.columns for col in ['host', 'port', 'service']):
                                for _, row in df.iterrows():
                                    ip = str(row['host']).strip()
                                    port = str(row['port']).strip()
                                    service = str(row['service']).strip()
                                    
                                    if ip and port and service:
                                        if ip not in ip_ports:
                                            ip_ports[ip] = []
                                        if (port, service) not in ip_ports[ip]:
                                            ip_ports[ip].append((port, service))
                        except Exception as e:
                            print(f"CSV processing error in {file_name}: {e}")
                            continue

                except Exception as e:
                    print(f"Error processing file {file_name}: {e}")
                    continue
        
        # Prepare data for Excel - same as Infrastructure VAPT
        all_ips = sorted(ip_ports.keys())
        has_placeholder = False
        
        if len(all_ips) % 2 != 0:
            all_ips.append("-")
            ip_ports["-"] = [("-", "-")]
            has_placeholder = True

        data = []
        for i in range(0, len(all_ips), 2):
            ip1 = all_ips[i]
            ip2 = all_ips[i + 1]
            ports1 = ip_ports.get(ip1, [])
            ports2 = ip_ports.get(ip2, [])
            
            if len(ports1) == 0:
                ports1 = [("Filtered", "Filtered")]
            if len(ports2) == 0:
                ports2 = [("Filtered", "Filtered")]
            
            max_ports = max(len(ports1), len(ports2))

            data.append(["HOST", "PORT", "SERVICE", "HOST", "PORT", "SERVICE"])

            for j in range(max_ports):
                data.append([
                    ip1, 
                    ports1[j][0] if j < len(ports1) else "", 
                    ports1[j][1] if j < len(ports1) else "",
                    ip2, 
                    ports2[j][0] if j < len(ports2) else "", 
                    ports2[j][1] if j < len(ports2) else ""
                ])
        
        print(f"✅ Processed {len(ip_ports)} IP addresses with ports")
        return data
        
    except Exception as e:
        print(f"❌ Error processing Nmap ZIP: {e}")
        import traceback
        traceback.print_exc()
        return []


def process_nessus_zip_public_ip(nessus_file):
    """
    Process Nessus ZIP file for Public IP VAPT - direct copy of data
    Returns: List of DataFrames (one per CSV file)
    """
    try:
        print("📁 Processing Nessus ZIP file for Public IP VAPT...")
        nessus_dataframes = []
        
        with zipfile.ZipFile(nessus_file, 'r') as zip_ref:
            csv_files = [f for f in zip_ref.namelist() if f.endswith('.csv')]
            
            if not csv_files:
                print("⚠️ No CSV files found in Nessus ZIP")
                return nessus_dataframes
            
            print(f"📄 Found {len(csv_files)} Nessus CSV files")
            
            for csv_file in csv_files:
                try:
                    with zip_ref.open(csv_file) as file:
                        df = pd.read_csv(file)
                        
                        # Just store the dataframe as-is for direct copy
                        nessus_dataframes.append(df)
                        print(f"✅ Loaded {csv_file}: {len(df)} rows")
                        
                except Exception as e:
                    print(f"⚠️ Error processing {csv_file}: {e}")
                    continue
        
        print(f"✅ Processed {len(nessus_dataframes)} Nessus files")
        return nessus_dataframes
        
    except Exception as e:
        print(f"❌ Error processing Nessus ZIP: {e}")
        import traceback
        traceback.print_exc()
        return []


@public_ip_vapt_followup_bp.route('/public_ip_vapt_followup_check_vulnerabilities', methods=['POST'])
def public_ip_vapt_followup_check_vulnerabilities():
    """Return both matched and unmatched vulnerabilities using Public IP VAPT Catalog."""
    if 'nmapFiles' not in request.files or 'nessusFiles' not in request.files:
        return jsonify({"error": "Both Nmap and Nessus files are required"}), 400
    
    nmap_file = request.files['nmapFiles']
    nessus_file = request.files['nessusFiles']
    
    if nmap_file.filename == '' or nessus_file.filename == '':
        return jsonify({"error": "No file selected"}), 400
    
    try:
        # Process Nessus files to check vulnerabilities
        nessus_dataframes = process_nessus_zip_public_ip(nessus_file)
        
        if nessus_dataframes:
            combined_nessus = pd.concat(nessus_dataframes, ignore_index=True)
            
            # Calculate matched and unmatched vulnerabilities
            matched_groups = []
            unmatched_vulnerabilities = []
            
            try:
                # Filter by valid risks and normalize
                valid_risks = ['low', 'medium', 'high', 'critical']
                df_filtered = combined_nessus.copy()
                df_filtered['Risk'] = df_filtered['Risk'].astype(str).str.lower().str.strip()
                df_filtered = df_filtered[df_filtered['Risk'].isin(valid_risks)]

                # Get unique vulnerability names from Name column
                unique_vulnerabilities_list = df_filtered['Name'].dropna().drop_duplicates().astype(str).str.strip().tolist()
                unique_vulnerabilities = set(unique_vulnerabilities_list)
                
                # Load Public IP VAPT Catalog
                catalog_path = "static/Formats_and_Catalog/Public IP VAPT Catalog.xlsx"
                if os.path.exists(catalog_path):
                    try:
                        catalog_df = pd.read_excel(catalog_path, sheet_name=0)
                    except Exception as e:
                        print(f"Error reading Public IP VAPT catalog file: {e}")
                        catalog_df = None
                    
                    if catalog_df is not None and 'Vulnerabilities in this group' in catalog_df.columns:
                        matched_vulnerability_names = set()
                        
                        # Build matched groups with catalog details
                        for idx, row in catalog_df.iterrows():
                            vulnerabilities_in_group = str(row.get('Vulnerabilities in this group', '')).strip()
                            if pd.isna(vulnerabilities_in_group) or vulnerabilities_in_group == '':
                                continue
                            
                            # Split vulnerabilities by newlines
                            vuln_list = [v.strip() for v in vulnerabilities_in_group.split('\n') if v.strip()]
                            
                            # Find which vulnerabilities from Excel match this catalog group
                            matched_vulns_in_group = []
                            for vulnerability in unique_vulnerabilities:
                                # Use only first 170 characters for matching
                                vuln_short = str(vulnerability)[:170]
                                escaped_vulnerability = re.escape(vuln_short)
                                pattern = rf'(?:\n|\r\n|\A){escaped_vulnerability}'
                                if re.search(pattern, vulnerabilities_in_group, re.IGNORECASE):
                                    matched_vulns_in_group.append(vulnerability)
                                    matched_vulnerability_names.add(vulnerability)
                            
                            # If any vulnerabilities matched this catalog group, add it
                            if matched_vulns_in_group:
                                matched_groups.append({
                                    'catalog_id': int(idx),
                                    'group_name': str(row.get('Name of Vulnerability', ''))[:200],
                                    'risk_factor': str(row.get('Risk Factor', ''))[:20],
                                    'cvss_score': str(row.get('CVSS Score', ''))[:10],
                                    'matched_vulnerabilities': matched_vulns_in_group
                                })
                        
                        # Calculate unmatched vulnerabilities
                        unmatched_vulnerabilities = sorted(list(unique_vulnerabilities - matched_vulnerability_names))
                        
            except Exception as e:
                print(f"Error calculating vulnerabilities: {e}")
                import traceback
                traceback.print_exc()
                return jsonify({"error": f"Error processing vulnerabilities: {str(e)}"}), 500
            
            # Initialize merge state in session - keep it minimal to avoid cookie size issues
            unmatched_limited = unmatched_vulnerabilities[:100] if len(unmatched_vulnerabilities) > 100 else unmatched_vulnerabilities
            
            # Limit matched_groups to 50 and truncate strings
            limited_matched_groups = matched_groups[:50]
            for group in limited_matched_groups:
                if 'group_name' in group and len(group['group_name']) > 100:
                    group['group_name'] = group['group_name'][:100]
            
            session['public_ip_vapt_followup_vulnerability_merge_state'] = {
                'matched_groups': limited_matched_groups,
                'unmatched_vulnerabilities': unmatched_limited,
                'merge_operations': []  # Initialize merge_operations for undo functionality
            }
            
            if len(unmatched_vulnerabilities) > 100:
                print(f"⚠️ Warning: {len(unmatched_vulnerabilities)} unmatched vulnerabilities found. Limited to 100 in session.")
            
            # Return matched groups and unmatched vulnerabilities
            return jsonify({
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_vulnerabilities
            })
        else:
            return jsonify({"error": "No Nessus data found"}), 400
            
    except Exception as e:
        print(f"Error checking vulnerabilities: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error processing files: {str(e)}"}), 500


# =============================================================================
# MERGE OPERATIONS
# =============================================================================

@public_ip_vapt_followup_bp.route('/public_ip_vapt_followup_merge_with_matched', methods=['POST'])
def public_ip_vapt_followup_merge_with_matched():
    """Merge an unmatched vulnerability with an existing matched group."""
    try:
        data = request.get_json()
        
        if not data or 'unmatched_vulnerability' not in data or 'target_group_id' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        unmatched_vuln = data['unmatched_vulnerability']
        target_group_id = data['target_group_id']
        
        # Get current merge state from session
        merge_state = session.get('public_ip_vapt_followup_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Find the target matched group
        matched_groups = merge_state.get('matched_groups', [])
        target_group = None
        target_index = None
        
        for idx, group in enumerate(matched_groups):
            if group['catalog_id'] == target_group_id:
                target_group = group
                target_index = idx
                break
        
        if target_group is None:
            return jsonify({"error": "Target group not found"}), 404
        
        # Add the unmatched vulnerability to the matched group
        if unmatched_vuln not in target_group['matched_vulnerabilities']:
            target_group['matched_vulnerabilities'].append(unmatched_vuln)
            print(f"✓ Added '{unmatched_vuln}' to group '{target_group['group_name']}'")
        
        # Remove from unmatched list
        unmatched_list = merge_state.get('unmatched_vulnerabilities', [])
        print(f"DEBUG: Unmatched list before removal: {len(unmatched_list)} items")
        
        if unmatched_vuln in unmatched_list:
            unmatched_list.remove(unmatched_vuln)
            print(f"✓ Removed '{unmatched_vuln}' from unmatched list")
        
        # Explicitly update the merge_state
        merge_state['unmatched_vulnerabilities'] = unmatched_list
        
        # Store minimal merge operation info (only last operation to minimize session size)
        from datetime import datetime
        merge_state['merge_operations'] = [{
            'type': 'merge_with_matched',
            'target_group_id': target_group_id,
            'unmatched_vulnerability': unmatched_vuln,
            'timestamp': datetime.now().isoformat()
        }]
        
        # Update session
        session['public_ip_vapt_followup_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        print(f"DEBUG: Final state - Matched: {len(matched_groups)}, Unmatched: {len(unmatched_list)}")
        
        return jsonify({
            "success": True,
            "message": "Vulnerability merged successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_list
            }
        })
        
    except Exception as e:
        print(f"Error merging with matched: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error merging: {str(e)}"}), 500


@public_ip_vapt_followup_bp.route('/public_ip_vapt_followup_merge_with_unmatched', methods=['POST'])
def public_ip_vapt_followup_merge_with_unmatched():
    """Merge multiple unmatched vulnerabilities into a new group."""
    try:
        data = request.get_json()
        
        if not data or 'vulnerabilities' not in data or 'vulnerability_details' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        vulnerabilities_to_merge = data['vulnerabilities']
        vulnerability_details = data['vulnerability_details']
        
        # Validate required fields
        required_fields = ['vulnerabilityName', 'riskFactor', 'cveId', 'cvssScore', 
                          'auditObservation', 'impact', 'recommendation', 'referenceLink']
        
        for field in required_fields:
            if field not in vulnerability_details:
                return jsonify({"error": f"Missing required field '{field}'"}), 400
        
        # Get current merge state from session
        merge_state = session.get('public_ip_vapt_followup_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Create a new matched group
        new_group_id = -len(merge_state.get('matched_groups', [])) - 1
        new_group = {
            'catalog_id': new_group_id,
            'group_name': vulnerability_details['vulnerabilityName'][:200],
            'risk_factor': vulnerability_details['riskFactor'][:20],
            'cvss_score': vulnerability_details['cvssScore'][:10],
            'matched_vulnerabilities': vulnerabilities_to_merge,
            'is_new_group': True
        }
        
        # Add to matched groups
        matched_groups = merge_state.get('matched_groups', [])
        matched_groups.append(new_group)
        
        # Store full details separately
        # Add merged vulnerability names to details
        vulnerability_details['isMerged'] = True
        vulnerability_details['mergedVulnerabilities'] = vulnerabilities_to_merge
        
        # Add actual vulnerability names to details for worksheet display
        # Store the first vulnerability name as actualVulnerabilityName for lookup
        if vulnerabilities_to_merge:
            vulnerability_details['actualVulnerabilityName'] = vulnerabilities_to_merge[0]
        
        new_group_details = merge_state.get('new_group_details', {})
        new_group_details[str(new_group_id)] = vulnerability_details
        merge_state['new_group_details'] = new_group_details
        
        # Remove from unmatched list
        unmatched_list = merge_state.get('unmatched_vulnerabilities', [])
        for vuln in vulnerabilities_to_merge:
            if vuln in unmatched_list:
                unmatched_list.remove(vuln)
        
        # Explicitly update the merge_state
        merge_state['unmatched_vulnerabilities'] = unmatched_list
        merge_state['matched_groups'] = matched_groups
        
        # Store minimal merge operation info (only last operation to minimize session size)
        from datetime import datetime
        merge_state['merge_operations'] = [{
            'type': 'merge_with_unmatched',
            'new_group_id': new_group_id,
            'vulnerabilities': vulnerabilities_to_merge,
            'timestamp': datetime.now().isoformat()
        }]
        
        # Update session
        session['public_ip_vapt_followup_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        # Update catalog
        vulnerability_details_with_merge = vulnerability_details.copy()
        vulnerability_details_with_merge['isMerged'] = True
        vulnerability_details_with_merge['mergedVulnerabilities'] = vulnerabilities_to_merge
        
        update_public_ip_catalog_with_vulnerabilities({
            vulnerability_details['vulnerabilityName']: vulnerability_details_with_merge
        })
        
        return jsonify({
            "success": True,
            "message": "Vulnerabilities merged into new group successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_list
            }
        })
        
    except Exception as e:
        print(f"Error merging with unmatched: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error merging: {str(e)}"}), 500


@public_ip_vapt_followup_bp.route('/public_ip_vapt_followup_add_vulnerability_details', methods=['POST'])
def public_ip_vapt_followup_add_vulnerability_details():
    """Add detailed information for a single unmatched vulnerability."""
    try:
        data = request.get_json()
        
        if not data or 'vulnerability_name' not in data or 'vulnerability_details' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        vuln_name = data['vulnerability_name']
        vuln_details = data['vulnerability_details']
        
        # Get current merge state
        merge_state = session.get('public_ip_vapt_followup_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Create a new matched group for this single vulnerability
        new_group_id = -len(merge_state.get('matched_groups', [])) - 1
        new_group = {
            'catalog_id': new_group_id,
            'group_name': vuln_details['vulnerabilityName'][:200],
            'risk_factor': vuln_details['riskFactor'][:20],
            'cvss_score': vuln_details['cvssScore'][:10],
            'matched_vulnerabilities': [vuln_name],
            'is_new_group': True
        }
        
        # Add to matched groups
        matched_groups = merge_state.get('matched_groups', [])
        matched_groups.append(new_group)
        
        # Store full details separately
        # Add actual vulnerability name to details for worksheet display
        vuln_details['actualVulnerabilityName'] = vuln_name  # Store the original vulnerability name
        
        new_group_details = merge_state.get('new_group_details', {})
        new_group_details[str(new_group_id)] = vuln_details
        merge_state['new_group_details'] = new_group_details
        
        # Remove from unmatched list
        unmatched_list = merge_state.get('unmatched_vulnerabilities', [])
        if vuln_name in unmatched_list:
            unmatched_list.remove(vuln_name)
        
        # Explicitly update the merge_state
        merge_state['unmatched_vulnerabilities'] = unmatched_list
        merge_state['matched_groups'] = matched_groups
        
        # Store minimal merge operation info (only last operation to minimize session size)
        from datetime import datetime
        merge_state['merge_operations'] = [{
            'type': 'add_details',
            'new_group_id': new_group_id,
            'vulnerability': vuln_name,
            'timestamp': datetime.now().isoformat()
        }]
        
        # Update session
        session['public_ip_vapt_followup_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        # Update catalog
        # Add actual vulnerability name to details for catalog storage
        vuln_details_with_actual = vuln_details.copy()
        vuln_details_with_actual['actualVulnerabilityName'] = vuln_name  # Store the original vulnerability name
        
        update_public_ip_catalog_with_vulnerabilities({
            vuln_details['vulnerabilityName']: vuln_details_with_actual
        })
        
        return jsonify({
            "success": True,
            "message": "Vulnerability details added successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_list
            }
        })
        
    except Exception as e:
        print(f"Error adding vulnerability details: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error adding details: {str(e)}"}), 500


@public_ip_vapt_followup_bp.route('/public_ip_vapt_followup_merge_matched_groups', methods=['POST'])
def public_ip_vapt_followup_merge_matched_groups():
    """Merge two existing matched groups together."""
    try:
        data = request.get_json()
        
        if not data or 'source_group_id' not in data or 'target_group_id' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        source_group_id = data['source_group_id']
        target_group_id = data['target_group_id']
        
        # Get current merge state
        merge_state = session.get('public_ip_vapt_followup_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        matched_groups = merge_state.get('matched_groups', [])
        
        # Find source and target groups
        source_group = None
        target_group = None
        source_index = None
        
        for idx, group in enumerate(matched_groups):
            if group['catalog_id'] == source_group_id:
                source_group = group
                source_index = idx
            if group['catalog_id'] == target_group_id:
                target_group = group
        
        if source_group is None or target_group is None:
            return jsonify({"error": "Source or target group not found"}), 404
        
        # Merge vulnerabilities from source to target
        for vuln in source_group['matched_vulnerabilities']:
            if vuln not in target_group['matched_vulnerabilities']:
                target_group['matched_vulnerabilities'].append(vuln)
        
        # Store source group data for undo (before removing it)
        source_group_data = source_group.copy()
        
        # Remove source group
        matched_groups.pop(source_index)
        
        # Explicitly update the merge_state
        merge_state['matched_groups'] = matched_groups
        
        # Store minimal merge operation info (only last operation to minimize session size)
        from datetime import datetime
        if 'merge_operations' not in merge_state:
            merge_state['merge_operations'] = []
        # Keep only the last operation to minimize session size
        merge_state['merge_operations'] = [{
            'type': 'merge_matched_groups',
            'source_group_id': source_group_id,
            'target_group_id': target_group_id,
            'source_group_data': source_group_data,  # Store source group to restore on undo
            'timestamp': datetime.now().isoformat()
        }]
        
        # Update session
        session['public_ip_vapt_followup_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        return jsonify({
            "success": True,
            "message": "Groups merged successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": merge_state.get('unmatched_vulnerabilities', [])
            }
        })
        
    except Exception as e:
        print(f"Error merging matched groups: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error merging groups: {str(e)}"}), 500


@public_ip_vapt_followup_bp.route('/public_ip_vapt_followup_undo_last_merge', methods=['POST'])
def public_ip_vapt_followup_undo_last_merge():
    """Undo the last merge operation."""
    try:
        merge_state = session.get('public_ip_vapt_followup_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        merge_operations = merge_state.get('merge_operations', [])
        if not merge_operations:
            # Return success with message instead of error to avoid frontend error
            return jsonify({
                "success": False,
                "message": "No operations to undo",
                "updated_state": {
                    "matched_groups": merge_state.get('matched_groups', []),
                    "unmatched_vulnerabilities": merge_state.get('unmatched_vulnerabilities', [])
                }
            })
        
        # Get the last operation
        last_operation = merge_operations.pop()
        operation_type = last_operation.get('type')
        
        matched_groups = merge_state.get('matched_groups', [])
        unmatched_list = merge_state.get('unmatched_vulnerabilities', [])
        
        # Undo based on operation type
        if operation_type == 'merge_with_matched':
            target_group_id = last_operation['target_group_id']
            unmatched_vuln = last_operation['unmatched_vulnerability']
            
            # Remove from target group
            for group in matched_groups:
                if group['catalog_id'] == target_group_id:
                    if unmatched_vuln in group['matched_vulnerabilities']:
                        group['matched_vulnerabilities'].remove(unmatched_vuln)
                    break
            
            # Add back to unmatched list
            if unmatched_vuln not in unmatched_list:
                unmatched_list.append(unmatched_vuln)
                unmatched_list.sort()
        
        elif operation_type == 'merge_with_unmatched':
            new_group_id = last_operation['new_group_id']
            vulnerabilities = last_operation['vulnerabilities']
            
            # Remove the new group
            matched_groups = [g for g in matched_groups if g['catalog_id'] != new_group_id]
            
            # Remove from new_group_details
            new_group_details = merge_state.get('new_group_details', {})
            if str(new_group_id) in new_group_details:
                del new_group_details[str(new_group_id)]
            merge_state['new_group_details'] = new_group_details
            
            # Add back to unmatched list
            for vuln in vulnerabilities:
                if vuln not in unmatched_list:
                    unmatched_list.append(vuln)
            unmatched_list.sort()
        
        elif operation_type == 'add_details':
            new_group_id = last_operation['new_group_id']
            vulnerability = last_operation['vulnerability']
            
            # Remove the new group
            matched_groups = [g for g in matched_groups if g['catalog_id'] != new_group_id]
            
            # Remove from new_group_details
            new_group_details = merge_state.get('new_group_details', {})
            if str(new_group_id) in new_group_details:
                del new_group_details[str(new_group_id)]
            merge_state['new_group_details'] = new_group_details
            
            # Add back to unmatched list
            if vulnerability not in unmatched_list:
                unmatched_list.append(vulnerability)
                unmatched_list.sort()
        
        elif operation_type == 'merge_matched_groups':
            source_group_data = last_operation.get('source_group_data')
            target_group_id = last_operation['target_group_id']
            
            if source_group_data:
                # Restore the source group
                matched_groups.append(source_group_data)
                
                # Remove merged vulnerabilities from target group
                for group in matched_groups:
                    if group['catalog_id'] == target_group_id:
                        for vuln in source_group_data['matched_vulnerabilities']:
                            if vuln in group['matched_vulnerabilities']:
                                group['matched_vulnerabilities'].remove(vuln)
                        break
        
        # Update session
        merge_state['matched_groups'] = matched_groups
        merge_state['unmatched_vulnerabilities'] = unmatched_list
        merge_state['merge_operations'] = merge_operations
        session['public_ip_vapt_followup_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        return jsonify({
            "success": True,
            "message": "Last operation undone successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_list
            }
        })
        
    except Exception as e:
        print(f"Error undoing operation: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error undoing: {str(e)}"}), 500


@public_ip_vapt_followup_bp.route('/public_ip_vapt_followup_add_manual_vulnerability', methods=['POST'])
def public_ip_vapt_followup_add_manual_vulnerability():
    """Add a manually entered vulnerability by user."""
    try:
        data = request.get_json()
        
        if not data or 'vulnerability_details' not in data:
            return jsonify({"error": "No vulnerability details provided"}), 400
        
        vuln_details = data['vulnerability_details']
        
        # Validate required fields
        required_fields = ['vulnerabilityName', 'riskFactor', 'cveId', 'cvssScore', 
                          'auditObservation', 'impact', 'recommendation', 'referenceLink']
        
        for field in required_fields:
            if field not in vuln_details:
                return jsonify({"error": f"Missing required field '{field}'"}), 400
        
        # Get current merge state
        merge_state = session.get('public_ip_vapt_followup_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Create a new matched group for the manual vulnerability
        new_group_id = -len(merge_state.get('matched_groups', [])) - 1000  # Use -1000+ for manual entries
        new_group = {
            'catalog_id': new_group_id,
            'group_name': vuln_details['vulnerabilityName'][:200],
            'risk_factor': vuln_details['riskFactor'][:20],
            'cvss_score': vuln_details['cvssScore'][:10],
            'matched_vulnerabilities': [vuln_details['vulnerabilityName']],  # Manual vulnerability
            'is_new_group': True,
            'is_manual': True  # Flag to indicate this was manually added
        }
        
        # Add to matched groups
        matched_groups = merge_state.get('matched_groups', [])
        matched_groups.append(new_group)
        
        # Store full details
        new_group_details = merge_state.get('new_group_details', {})
        new_group_details[str(new_group_id)] = vuln_details
        merge_state['new_group_details'] = new_group_details
        
        # Track manually added vulnerabilities
        manual_vulns = merge_state.get('manually_added_vulnerabilities', [])
        manual_vulns.append(vuln_details['vulnerabilityName'])
        merge_state['manually_added_vulnerabilities'] = manual_vulns
        
        # Update the merge_state
        merge_state['matched_groups'] = matched_groups
        
        # Update session
        session['public_ip_vapt_followup_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        # Update catalog
        vuln_details['isManual'] = True
        # Add actual vulnerability name to details for catalog storage (manual = same as group name)
        vuln_details['actualVulnerabilityName'] = vuln_details['vulnerabilityName']
        
        update_public_ip_catalog_with_vulnerabilities({
            vuln_details['vulnerabilityName']: vuln_details
        })
        
        return jsonify({
            "success": True,
            "message": "Manual vulnerability added successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": merge_state.get('unmatched_vulnerabilities', [])
            },
            "new_vulnerability": {
                "group_id": new_group_id,
                "group_name": vuln_details['vulnerabilityName'],
                "risk_factor": vuln_details['riskFactor']
            }
        })
        
    except Exception as e:
        print(f"Error adding manual vulnerability: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error adding manual vulnerability: {str(e)}"}), 500


def update_public_ip_catalog_with_vulnerabilities(vulnerability_details_dict):
    """Update the Public IP VAPT Catalog with both merged and separate vulnerabilities."""
    try:
        from flask_login import current_user
        
        catalog_path = "static/Formats_and_Catalog/Public IP VAPT Catalog.xlsx"
        
        if not os.path.exists(catalog_path):
            print(f"Catalog file not found at: {catalog_path}")
            return
        
        # Read the existing catalog with error handling - now reading Sheet2 (index 1)
        try:
            catalog_df = pd.read_excel(catalog_path, sheet_name=1)  # Changed from sheet_name=0 to sheet_name=1 for Sheet2
        except Exception as e:
            print(f"Error reading catalog file: {e}")
            print(f"Catalog file may be corrupted. Please check: {catalog_path}")
            return
        
        # Process all vulnerabilities (both merged and separate)
        for vuln_name, details in vulnerability_details_dict.items():
            if details.get('isMerged', False):
                # Handle merged vulnerabilities
                merged_vulns = details.get('mergedVulnerabilities', [])
                if merged_vulns:
                    # Truncate vulnerability names to first 170 characters for catalog storage
                    truncated_merged_vulns = [str(v)[:170] for v in merged_vulns]
                    
                    # Create a new row for the merged vulnerability group
                    new_row = {
                        'Sr No': len(catalog_df) + 1,
                        'Name of Vulnerability': details.get('vulnerabilityName', ''),
                        'Risk Factor': details.get('riskFactor', ''),
                        'CVE/CWE ID': details.get('cveId', ''),
                        'CVSS': details.get('cvssScore', ''),
                        'Audit Observation': details.get('auditObservation', ''),
                        'Impact': details.get('impact', ''),
                        'Recommendation/Countermeasure': details.get('recommendation', ''),
                        'Affected System': '',  # Empty as requested
                        'Reference Link': details.get('referenceLink', ''),
                        'Vulnerabilities in this group': '\n'.join(truncated_merged_vulns),
                        'User_name': current_user.employee_name if current_user.is_authenticated else 'Unknown',
                        'Time_stamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    # Add the new row to the catalog
                    catalog_df = pd.concat([catalog_df, pd.DataFrame([new_row])], ignore_index=True)
            else:
                # Handle separate (non-merged) vulnerabilities
                # Get the actual vulnerability name (might be different from group name)
                actual_vuln_name = details.get('actualVulnerabilityName', vuln_name)
                # Truncate to first 170 characters for catalog storage
                actual_vuln_name_short = str(actual_vuln_name)[:170]
                
                new_row = {
                    'Sr No': len(catalog_df) + 1,
                    'Name of Vulnerability': details.get('vulnerabilityName', vuln_name),  # User-provided group name
                    'Risk Factor': details.get('riskFactor', ''),
                    'CVE/CWE ID': details.get('cveId', ''),
                    'CVSS': details.get('cvssScore', ''),
                    'Audit Observation': details.get('auditObservation', ''),
                    'Impact': details.get('impact', ''),
                    'Recommendation/Countermeasure': details.get('recommendation', ''),
                    'Affected System': '',  # Empty as requested
                    'Reference Link': details.get('referenceLink', ''),
                    'Vulnerabilities in this group': actual_vuln_name_short,  # Truncated to 170 chars
                    'User_name': current_user.employee_name if current_user.is_authenticated else 'Unknown',
                    'Time_stamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                # Add the new row to the catalog
                catalog_df = pd.concat([catalog_df, pd.DataFrame([new_row])], ignore_index=True)
        
        # Save the updated catalog to Sheet2
        # Read all existing sheets first
        try:
            # Read all sheets to preserve existing data
            all_sheets = pd.read_excel(catalog_path, sheet_name=None)
            
            # Update Sheet2 with our new data
            all_sheets['Sheet2'] = catalog_df
            
            # Write all sheets back to the file
            with pd.ExcelWriter(catalog_path, engine='openpyxl') as writer:
                for sheet_name, sheet_df in all_sheets.items():
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            print(f"✅ Successfully updated Public IP VAPT Catalog (Sheet2)")
        
        except Exception as e:
            print(f"Error updating catalog file: {e}")
            import traceback
            traceback.print_exc()
        
    except Exception as e:
        print(f"Error updating Public IP VAPT Catalog: {e}")
        import traceback
        traceback.print_exc()


# =============================================================================
# REPORT GENERATION
# =============================================================================

def create_public_ip_nmap_worksheet(wb, nmap_data_list):
    """
    Create Nmap Scan worksheet for Public IP VAPT.
    Format: 6 columns (HOST, PORT, SERVICE, HOST, PORT, SERVICE) with merge logic
    Same as Infrastructure VAPT - columns A-F
    """
    try:
        print("📊 Creating Nmap Scan worksheet...")
        ws = wb.create_sheet("Nmap Files", 0)
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=10)
        data_alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Set column widths A-F
        ws.column_dimensions['A'].width = 20  # HOST
        ws.column_dimensions['B'].width = 20  # PORT
        ws.column_dimensions['C'].width = 20  # SERVICE
        ws.column_dimensions['D'].width = 20  # HOST
        ws.column_dimensions['E'].width = 20  # PORT
        ws.column_dimensions['F'].width = 20  # SERVICE
        
        # Write all data rows with formatting
        current_row = 1
        total_ports = 0
        
        # Track IP positions for merging (columns A and D)
        ip_positions = {"A": {}, "D": {}}
        
        # First pass: Write data and identify IP positions
        for row_num, row_data in enumerate(nmap_data_list):
            for col_num, value in enumerate(row_data):
                cell = ws.cell(row=current_row, column=col_num + 1, value=value)
                
                # Apply header format if this is a header row
                if row_data == ["HOST", "PORT", "SERVICE", "HOST", "PORT", "SERVICE"]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                else:
                    cell.font = data_font
                    cell.alignment = data_alignment_center
                    
                    # Track IP positions for columns A and D (indices 0 and 3)
                    if col_num == 0 and value and value != "HOST" and value != "-":
                        if value not in ip_positions["A"]:
                            ip_positions["A"][value] = []
                        ip_positions["A"][value].append(current_row)
                    
                    if col_num == 3 and value and value != "HOST" and value != "-":
                        if value not in ip_positions["D"]:
                            ip_positions["D"][value] = []
                        ip_positions["D"][value].append(current_row)
                    
                    if value == "":
                        total_ports += 1
                
                # Apply border to EACH cell
            cell.border = thin_border
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1
            
        # Second pass: Merge contiguous IP cells in columns A and D
        for col_letter, ip_dict in ip_positions.items():
            col_index = ord(col_letter) - ord('A')
            
            for ip, row_nums in ip_dict.items():
                row_nums.sort()
                
                # Group contiguous rows
                groups = []
                current_group = [row_nums[0]]
                
                for i in range(1, len(row_nums)):
                    if row_nums[i] == row_nums[i-1] + 1:
                        current_group.append(row_nums[i])
                    else:
                        groups.append(current_group)
                        current_group = [row_nums[i]]
                
                groups.append(current_group)
                
                # Merge each group
                for group in groups:
                    if len(group) > 1:
                        start_row = group[0]
                        end_row = group[-1]
                        ws.merge_cells(start_row=start_row, start_column=col_index + 1, end_row=end_row, end_column=col_index + 1)
                        # Reapply border to the merged cell
                        merged_cell = ws.cell(row=start_row, column=col_index + 1)
                        merged_cell.border = thin_border
        
        # Handle placeholder "-" IP merging in column D
        has_placeholder = any("-" in row for row in nmap_data_list)
        if has_placeholder:
            for row_num in range(1, len(nmap_data_list)):
                if nmap_data_list[row_num][3] == "-":
                    placeholder_rows = []
                    current_row_idx = row_num
                    while current_row_idx < len(nmap_data_list) and nmap_data_list[current_row_idx][3] == "-":
                        placeholder_rows.append(current_row_idx + 1)  # +1 because row_num starts from 1
                        current_row_idx += 1
                    
                    if len(placeholder_rows) > 1:
                        start_row = placeholder_rows[0]
                        end_row = placeholder_rows[-1]
                        ws.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)
                        # Reapply border to the merged cell
                        merged_cell = ws.cell(row=start_row, column=4)
                        merged_cell.border = thin_border
                    break
        
        # Merge empty PORT/SERVICE cells with previous entry (columns 1,2,4,5 which are B, C, E, F)
        columns_to_merge = [1, 2, 4, 5]  # B, C, E, F
        for col in columns_to_merge:
            # First, find sequences of empty cells and merge them with the previous non-empty cell
            merge_start = None
            prev_filled_row = None
            
            for row_idx in range(1, len(nmap_data_list) + 1):
                current_value = ws.cell(row=row_idx, column=col + 1).value
                
                if current_value == "":
                    # Empty cell - start tracking merge if not started
                    if merge_start is None:
                        merge_start = row_idx
                else:
                    # Non-empty cell - if we have a range of empty cells, merge them
                    if merge_start is not None:
                        # Merge from merge_start-1 to merge_start range
                        # The previous filled row should be merge_start - 1
                        if merge_start > 1:
                            prev_filled_value = ws.cell(row=merge_start - 1, column=col + 1).value
                            if prev_filled_value != "":  # Only merge if previous row has a value
                                ws.merge_cells(start_row=merge_start - 1, start_column=col + 1, end_row=row_idx - 1, end_column=col + 1)
                                ws.cell(row=merge_start - 1, column=col + 1).value = prev_filled_value
                                # Reapply border to the merged cell
                                merged_cell = ws.cell(row=merge_start - 1, column=col + 1)
                                merged_cell.border = thin_border
                    merge_start = None
                    prev_filled_row = row_idx
            
            # Handle merging at the end
            if merge_start is not None and merge_start <= len(nmap_data_list):
                if merge_start > 1:
                    prev_filled_value = ws.cell(row=merge_start - 1, column=col + 1).value
                    if prev_filled_value != "":
                        ws.merge_cells(start_row=merge_start - 1, start_column=col + 1, end_row=len(nmap_data_list), end_column=col + 1)
                        ws.cell(row=merge_start - 1, column=col + 1).value = prev_filled_value
                        # Reapply border to the merged cell
                        merged_cell = ws.cell(row=merge_start - 1, column=col + 1)
                        merged_cell.border = thin_border
        
        # Final safety pass: Reapply borders to all cells
        for row_idx in range(1, len(nmap_data_list) + 1):
            for col_idx in range(1, 7):  # Columns A-F
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                except:
                    pass  # Skip merged cells that can't be accessed directly
        
        print(f"✅ Nmap Files worksheet created with {len(nmap_data_list)} rows")
        
    except Exception as e:
        print(f"❌ Error creating Nmap worksheet: {e}")
        import traceback
        traceback.print_exc()


def create_public_ip_nessus_csv_worksheet(wb, nessus_dataframes):
    """
    Create Nessus CSV Files worksheet - direct copy from Nessus files.
    """
    try:
        print("📊 Creating Nessus CSV Files worksheet...")
        
        if not nessus_dataframes:
            print("⚠️ No Nessus data to create worksheet")
            return
        
        # Combine all Nessus dataframes
        combined_df = pd.concat(nessus_dataframes, ignore_index=True)
        
        ws = wb.create_sheet("Nessus CSV Files")
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Write headers (column names from DataFrame)
        for col_idx, column_name in enumerate(combined_df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        ws.row_dimensions[1].height = 30
        
        # Write data rows
        for row_idx, data_row in enumerate(combined_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            ws.row_dimensions[row_idx].height = 20
        
        print(f"✅ Nessus CSV Files worksheet created with {len(combined_df)} rows")
        
    except Exception as e:
        print(f"❌ Error creating Nessus CSV Files worksheet: {e}")
        import traceback
        traceback.print_exc()


def create_public_ip_scope_worksheet(wb, nessus_dataframes):
    """
    Create Scope worksheet with Sr.No and Host columns from Nessus CSV Files.
    Same format as Infrastructure VAPT reference
    """
    try:
        print("📊 Creating Scope worksheet...")
        
        if not nessus_dataframes:
            print("⚠️ No Nessus data for Scope worksheet")
            return
        
        # Combine all Nessus dataframes
        combined_df = pd.concat(nessus_dataframes, ignore_index=True)
        
        # Extract unique Host values
        if 'Host' not in combined_df.columns:
            print("⚠️ 'Host' column not found in Nessus data")
            return
        
        unique_hosts = combined_df['Host'].dropna().unique().tolist()
        unique_hosts = [str(host).strip() for host in unique_hosts if str(host).strip()]
        unique_hosts = sorted(unique_hosts)  # Sort for consistent output
        
        ws = wb.create_sheet("Scope")
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        data_font = Font(name='Arial', size=10)
        data_alignment_center = Alignment(horizontal='center', vertical='center')
        data_alignment_left = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 10  # Sr.No
        ws.column_dimensions['B'].width = 30  # Host
        
        # Create headers
        headers = ['Sr.No', 'Host']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 30
        
        # Add unique hosts with serial numbers
        for idx, host in enumerate(unique_hosts, start=2):
            # Sr.No
            cell = ws.cell(row=idx, column=1, value=idx - 1)
            cell.font = data_font
            cell.alignment = data_alignment_center
            cell.border = thin_border
            
            # Host
            cell = ws.cell(row=idx, column=2, value=host)
            cell.font = data_font
            cell.alignment = data_alignment_center
            cell.border = thin_border
            
            ws.row_dimensions[idx].height = 20
        
        print(f"✅ Scope worksheet created with {len(unique_hosts)} hosts")
        
    except Exception as e:
        print(f"❌ Error creating Scope worksheet: {e}")
        import traceback
        traceback.print_exc()


def create_public_ip_summary_worksheet(wb, nessus_dataframes):
    """
    Create Summary worksheet for Public IP VAPT.
    Format: 3 columns (Sr.No, Name, Host) with Name column merged for same vulnerabilities
    Same format as Infrastructure VAPT reference
    """
    try:
        print("📊 Creating Summary worksheet...")
        
        # Always create the worksheet with headers
        ws = wb.create_sheet("Summary")
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=10)
        data_alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 10  # Sr.No
        ws.column_dimensions['B'].width = 60  # Name
        ws.column_dimensions['C'].width = 20  # Host
        
        # Create headers
        headers = ['Sr.No', 'Name', 'Host']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 30
        
        if not nessus_dataframes:
            print("⚠️ No Nessus data for Summary worksheet - created empty worksheet with headers")
            return
        
        # Combine all Nessus dataframes
        combined_df = pd.concat(nessus_dataframes, ignore_index=True)
        
        # Filter by valid risks
        valid_risks = ['low', 'medium', 'high', 'critical']
        df_filtered = combined_df.copy()
        df_filtered['Risk'] = df_filtered['Risk'].astype(str).str.lower().str.strip()
        df_filtered = df_filtered[df_filtered['Risk'].isin(valid_risks)]
        
        # Select only Host, Name columns (no Risk column in output)
        if not all(col in df_filtered.columns for col in ['Host', 'Name']):
            print("⚠️ Required columns not found in Nessus data - created empty worksheet with headers")
            return
        
        summary_data = df_filtered[['Host', 'Name']].copy()
        
        # Clean data
        summary_data['Host'] = summary_data['Host'].astype(str).str.strip()
        summary_data['Name'] = summary_data['Name'].astype(str).str.strip()
        
        # Remove empty rows
        summary_data = summary_data.dropna()
        summary_data = summary_data[(summary_data['Host'] != '') & (summary_data['Name'] != '')]
        
        if summary_data.empty:
            print("⚠️ No valid data found for Summary worksheet after filtering - created empty worksheet with headers")
            return
        
        # Remove duplicates based on Name + Host combination
        summary_data = summary_data.drop_duplicates(subset=['Name', 'Host'], keep='first')
        
        # Sort by Name, then Host to group similar vulnerabilities
        summary_data = summary_data.sort_values(['Name', 'Host'])
        
        # Write data with merging logic
        row = 2
        current_vulnerability = None
        name_merge_start = 2
        serial_counter = 1
        
        for _, row_data in summary_data.iterrows():
            name = str(row_data['Name'])
            host = str(row_data['Host'])
            
            # If we're starting a new vulnerability
            if name != current_vulnerability:
                # Merge previous vulnerability name cells if needed
                if current_vulnerability is not None and row > name_merge_start:
                    # Merge Name column
                    ws.merge_cells(start_row=name_merge_start, start_column=2, end_row=row-1, end_column=2)
                    # Merge Sr.No column
                    ws.merge_cells(start_row=name_merge_start, start_column=1, end_row=row-1, end_column=1)
                
                current_vulnerability = name
                name_merge_start = row
                serial_counter += 1
            
            # Write Sr.No (only for first occurrence of each vulnerability)
            if name != current_vulnerability or row == name_merge_start:
                cell = ws.cell(row=row, column=1, value=serial_counter - 1)
                cell.font = data_font
                cell.alignment = data_alignment_center
                cell.border = thin_border
            
            # Write Name (only for first occurrence of each vulnerability)
            if name != current_vulnerability or row == name_merge_start:
                cell = ws.cell(row=row, column=2, value=name)
                cell.font = data_font
                cell.alignment = data_alignment_center
                cell.border = thin_border
            
            # Write Host (always)
            cell = ws.cell(row=row, column=3, value=host)
            cell.font = data_font
            cell.alignment = data_alignment_center
            cell.border = thin_border
            
            ws.row_dimensions[row].height = 20
            row += 1
        
        # Merge the last vulnerability if needed
        if current_vulnerability is not None and row > name_merge_start:
            # Merge Name column
            ws.merge_cells(start_row=name_merge_start, start_column=2, end_row=row-1, end_column=2)
            # Merge Sr.No column
            ws.merge_cells(start_row=name_merge_start, start_column=1, end_row=row-1, end_column=1)
        
        print(f"✅ Summary worksheet created with {len(summary_data)} entries")
        
    except Exception as e:
        print(f"❌ Error creating Summary worksheet: {e}")
        import traceback
        traceback.print_exc()


def create_public_ip_Public_IP_VAPT_worksheet(wb, nessus_dataframes):
    """
    Create Public_IP_VAPT worksheet for Public IP VAPT Follow-Up Audit.
    Includes Status and Old POC columns for follow-up audit tracking.
    """
    try:
        print("📊 Creating Public_IP_VAPT worksheet...")
        
        # Get merge state from session (Follow-up specific key) - check before processing data
        merge_state = session.get('public_ip_vapt_followup_vulnerability_merge_state', None)
        use_merge_state = merge_state is not None and 'matched_groups' in merge_state
        
        if use_merge_state:
            print("📊 [Follow-up] Using merge state from session for vulnerability grouping")
            merged_groups_from_session = merge_state.get('matched_groups', [])
            new_group_details_dict = merge_state.get('new_group_details', {})
            manually_added_vulnerabilities = merge_state.get('manually_added_vulnerabilities', [])
        else:
            print("📊 [Follow-up] No merge state found - using standard catalog matching")
            merged_groups_from_session = None
            new_group_details_dict = {}
            manually_added_vulnerabilities = []
        
        # Load the Public IP VAPT catalog file first (needed for headers)
        catalog_path = os.path.join('static', 'Formats_and_Catalog', 'Public IP VAPT Catalog.xlsx')
        
        if not os.path.exists(catalog_path):
            print(f"Catalog file not found at: {catalog_path}")
            return
        
        # Read the catalog file with error handling - read Sheet1 (index 0)
        try:
            catalog_df = pd.read_excel(catalog_path, sheet_name=0)
        except Exception as e:
            print(f"Error reading catalog file for Public_IP_VAPT worksheet: {e}")
            print(f"Catalog file may be corrupted. Please check: {catalog_path}")
            return
        
        # Check if the required column exists in catalog
        if 'Vulnerabilities in this group' not in catalog_df.columns:
            print("Catalog file does not contain 'Vulnerabilities in this group' column")
            return
        
        # Always create Public_IP_VAPT worksheet with headers first
        ws = wb.create_sheet("Public_IP_VAPT")
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=10)
        data_alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Risk color formats
        critical_fill = PatternFill(start_color='FF8B0000', end_color='FF8B0000', fill_type='solid')
        high_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        medium_fill = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')
        low_fill = PatternFill(start_color='FF008000', end_color='FF008000', fill_type='solid')
        
        critical_font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
        high_font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
        medium_font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
        low_font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Define columns to exclude from catalog
        exclude_columns = ['Sr No', 'Vulnerabilities in this group', 'Affected System']
        
        # Get catalog headers excluding the specified columns
        catalog_headers = [col for col in catalog_df.columns.tolist() if col not in exclude_columns]
        # Normalize column names: replace "CVE ID" with "CVE/CWE ID" for consistency
        catalog_headers = ['CVE/CWE ID' if col == 'CVE ID' else col for col in catalog_headers]
        # Also normalize the catalog_df column names to match
        if 'CVE ID' in catalog_df.columns:
            catalog_df = catalog_df.rename(columns={'CVE ID': 'CVE/CWE ID'})
        
        # Find the positions of key columns
        rec_countermeasure_col = None
        reference_link_col = None
        audit_observation_col = None
        risk_factor_col = None
        
        for i, col_name in enumerate(catalog_headers):
            if 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                rec_countermeasure_col = i
            if 'reference' in col_name.lower() and 'link' in col_name.lower():
                reference_link_col = i
            if 'audit' in col_name.lower() and 'observation' in col_name.lower():
                audit_observation_col = i
            if 'risk' in col_name.lower() and 'factor' in col_name.lower():
                risk_factor_col = i
        
        # Insert "Affected Systems" column after Recommendation/Countermeasure and before Reference Link
        if rec_countermeasure_col is not None and reference_link_col is not None:
            insert_position = rec_countermeasure_col + 1
        elif rec_countermeasure_col is not None:
            insert_position = rec_countermeasure_col + 1
        elif reference_link_col is not None:
            insert_position = reference_link_col
        else:
            insert_position = len(catalog_headers)
        
        # Create worksheet headers
        headers = ["Sr.No", "Vulnerabilities"]
        headers.extend(catalog_headers[:insert_position])
        headers.append("Affected Systems")
        headers.extend(catalog_headers[insert_position:])
        
        # Add Status column
        headers.append("Status")
        
        # Add Old POC columns (7 columns)
        headers.extend(["Old_POC_N", "Old_POC_O", "Old_POC_P", "Old_POC_Q", "Old_POC_R", "Old_POC_S", "Old_POC_M"])
        
        # Add POC columns (7 columns)
        headers.extend(["POC_U", "POC_V", "POC_W", "POC_X", "POC_Y", "POC_Z", "POC_T"])
        
        # Set column widths
        ws.column_dimensions['A'].width = 8   # Sr.No
        ws.column_dimensions['B'].width = 50  # Vulnerabilities
        for i in range(2, len(headers) - 1):  # Catalog columns
            col_letter = get_column_letter(i + 1)  # i is 0-based, columns are 1-based
            ws.column_dimensions[col_letter].width = 20
        col_letter = get_column_letter(len(headers))  # Affected Systems
        ws.column_dimensions[col_letter].width = 30
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 30
        
        # Find Status column position
        status_col_position = None
        for i, header in enumerate(headers, start=1):
            if header == "Status":
                status_col_position = i
                break
        
        # Find Old POC column positions
        old_poc_col_start = None
        old_poc_col_end = None
        for i, header in enumerate(headers, start=1):
            if header == "Old_POC_N":
                old_poc_col_start = i
            if header == "Old_POC_M":
                old_poc_col_end = i
                break
        
        # Find POC column positions
        poc_col_start = None
        poc_col_end = None
        for i, header in enumerate(headers, start=1):
            if header == "POC_U":
                poc_col_start = i
            if header == "POC_T":
                poc_col_end = i
                break
        
        # Set POC and Old POC column widths
        if old_poc_col_start is not None and old_poc_col_end is not None:
            for col in range(old_poc_col_start, old_poc_col_end + 1):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 20
        
        if poc_col_start is not None and poc_col_end is not None:
            for col in range(poc_col_start, poc_col_end + 1):
                col_letter = get_column_letter(col)  # Convert to column letter
                ws.column_dimensions[col_letter].width = 20  # Set POC column width to 20
        
        # Merge Old POC header cells
        if old_poc_col_start is not None and old_poc_col_end is not None:
            ws.merge_cells(start_row=1, start_column=old_poc_col_start, end_row=1, end_column=old_poc_col_end)
            old_poc_header_cell = ws.cell(row=1, column=old_poc_col_start, value="Old POC")
            old_poc_header_cell.font = header_font
            old_poc_header_cell.fill = header_fill
            old_poc_header_cell.alignment = header_alignment
            old_poc_header_cell.border = thin_border
        
        # Merge POC header cells
        if poc_col_start is not None and poc_col_end is not None:
            ws.merge_cells(start_row=1, start_column=poc_col_start, end_row=1, end_column=poc_col_end)
            poc_header_cell = ws.cell(row=1, column=poc_col_start, value="POC")
            poc_header_cell.font = header_font
            poc_header_cell.fill = header_fill
            poc_header_cell.alignment = header_alignment
            poc_header_cell.border = thin_border
        
        # Check if we have Nessus data
        if not nessus_dataframes:
            print("⚠️ No Nessus data for Public_IP_VAPT worksheet - created empty worksheet with headers")
            return
        
        # Get global comparison data
        comparison_result = {}
        try:
            combined_nessus = pd.concat(nessus_dataframes, ignore_index=True) if nessus_dataframes else pd.DataFrame()
            our_vulnerabilities = set()
            
            if not combined_nessus.empty and 'Name' in combined_nessus.columns:
                unique_names = combined_nessus['Name'].drop_duplicates().tolist()
                our_vulnerabilities = set(unique_names)
            
            comparison_result = compare_vulnerabilities(our_vulnerabilities, PUBLIC_IP_VAPT_USER_VULNERABILITIES)
            print(f"📊 Comparison result: {len(comparison_result)} vulnerabilities with status assigned")
        except Exception as e:
            print(f"⚠️ Could not get comparison data: {e}")
            comparison_result = {}
        
        # Combine all Nessus dataframes
        combined_df = pd.concat(nessus_dataframes, ignore_index=True)
        
        # Check if required columns exist
        required_columns = ['Name', 'Host', 'Risk']
        missing_cols = [col for col in required_columns if col not in combined_df.columns]
        if missing_cols:
            print(f"Cannot create Public_IP_VAPT worksheet - missing columns: {missing_cols} - created empty worksheet with headers")
            return
        
        # Filter by valid risks
        valid_risks = ['low', 'medium', 'high', 'critical']
        df_filtered = combined_df.copy()
        df_filtered['Risk'] = df_filtered['Risk'].astype(str).str.lower().str.strip()
        df_filtered = df_filtered[df_filtered['Risk'].isin(valid_risks)]
        
        # Get unique vulnerability names
        unique_vulnerabilities = df_filtered['Name'].drop_duplicates().tolist()
        
        # Add manual vulnerabilities to unique list if they're not already there
        for manual_vuln in manually_added_vulnerabilities:
            if manual_vuln not in unique_vulnerabilities:
                unique_vulnerabilities.append(manual_vuln)
        
        if not unique_vulnerabilities:
            print("⚠️ No vulnerabilities found for Public_IP_VAPT worksheet - worksheet created with headers only")
            # Set column widths even if no data (similar to Infra_VAPT)
            ws.column_dimensions['A'].width = 8   # Sr.No
            ws.column_dimensions['B'].width = 50  # Vulnerabilities
            # Set widths for other columns based on catalog headers
            for col_idx in range(3, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                if col_idx < len(headers) - 13:  # Before POC columns
                    ws.column_dimensions[col_letter].width = 20
                else:  # POC columns
                    ws.column_dimensions[col_letter].width = 20
            # Don't return - allow closed vulnerabilities to be added later
            print("⚠️ Worksheet created with headers - closed vulnerabilities from user Excel will be added later")
            return
        
        print(f"Found {len(unique_vulnerabilities)} unique vulnerabilities to process")
        
        # Track matched catalog entries
        matched_catalog_indices = set()
        matched_vulnerabilities = set()
        vulnerability_groups = {}
        vulnerability_affected_systems = {}
        vulnerability_risks = {}
        catalog_risk_values = {}
        
        # First pass: Collect affected systems for ALL vulnerabilities
        for vulnerability in unique_vulnerabilities:
            # Check if this is a manual vulnerability (not in scan data)
            is_manual = vulnerability in manually_added_vulnerabilities
            
            if is_manual:
                # For manual vulnerabilities, get affected systems from merge state
                affected_systems_list = []
                
                # Find the group that contains this manual vulnerability
                if use_merge_state and merged_groups_from_session:
                    for group in merged_groups_from_session:
                        if vulnerability in group.get('matched_vulnerabilities', []):
                            group_id = group['catalog_id']
                            details = new_group_details_dict.get(str(group_id), {})
                            
                            # Get affected systems from details
                            if 'affectedSystems' in details and details['affectedSystems']:
                                affected_systems_str = str(details['affectedSystems']).strip()
                                if affected_systems_str:
                                    # Split by newlines or commas and clean up
                                    if '\n' in affected_systems_str:
                                        affected_systems_list = [s.strip() for s in affected_systems_str.split('\n') if s.strip()]
                                    elif ',' in affected_systems_str:
                                        affected_systems_list = [s.strip() for s in affected_systems_str.split(',') if s.strip()]
                                    else:
                                        affected_systems_list = [affected_systems_str]
                                
                                print(f"✓ Found affected systems for manual vulnerability '{vulnerability}': {affected_systems_list}")
                                break
                    
                    # If not found in merged groups, check all new_group_details (in case group was merged)
                    if not affected_systems_list:
                        for group_id_str, details in new_group_details_dict.items():
                            # Check if this vulnerability name matches
                            if details.get('vulnerabilityName') == vulnerability or details.get('actualVulnerabilityName') == vulnerability:
                                if 'affectedSystems' in details and details['affectedSystems']:
                                    affected_systems_str = str(details['affectedSystems']).strip()
                                    if affected_systems_str:
                                        if '\n' in affected_systems_str:
                                            affected_systems_list = [s.strip() for s in affected_systems_str.split('\n') if s.strip()]
                                        elif ',' in affected_systems_str:
                                            affected_systems_list = [s.strip() for s in affected_systems_str.split(',') if s.strip()]
                                        else:
                                            affected_systems_list = [affected_systems_str]
                                        
                                        print(f"✓ Found affected systems for manual vulnerability '{vulnerability}' from details: {affected_systems_list}")
                                        break
                
                vulnerability_affected_systems[vulnerability] = affected_systems_list
                vulnerability_risks[vulnerability] = 'low'  # Default risk
            else:
                # Get all affected systems for this vulnerability (just Host, no Branch Name)
                vuln_data = df_filtered[df_filtered['Name'] == vulnerability]
                hosts = set()
                max_risk = 'low'
                
                for _, row in vuln_data.iterrows():
                    host = str(row['Host']).strip()
                    risk = str(row['Risk']).lower().strip()
                    
                    # Track the highest risk
                    risk_levels = {'critical': 4, 'high': 3, 'medium': 2, 'low': 1}
                    if risk_levels.get(risk, 0) > risk_levels.get(max_risk, 0):
                        max_risk = risk
                    
                    if host:
                        hosts.add(host)
                
                # Format affected systems (just hosts, no branches)
                formatted_systems = sorted(list(hosts))
                vulnerability_affected_systems[vulnerability] = formatted_systems
                vulnerability_risks[vulnerability] = max_risk
        
        # Use merge state if available, otherwise use standard catalog matching
        if use_merge_state and merged_groups_from_session:
            print(f"✓ [Follow-up] Using {len(merged_groups_from_session)} merged groups from session")
            
            # Build vulnerability_groups dict from merge state
            for group in merged_groups_from_session:
                catalog_idx = group['catalog_id']
                matched_catalog_indices.add(catalog_idx)
                
                # Add all vulnerabilities in this group
                vulnerability_groups[catalog_idx] = group['matched_vulnerabilities']
                for vuln in group['matched_vulnerabilities']:
                    matched_vulnerabilities.add(vuln)
                
                # Store risk factor
                catalog_risk_values[catalog_idx] = str(group.get('risk_factor', '')).upper().strip()
        else:
            # Standard catalog matching
            print("✓ [Follow-up] Using standard catalog matching")
            for vulnerability in unique_vulnerabilities:
                # Skip manual vulnerabilities in standard matching
                if vulnerability in manually_added_vulnerabilities:
                    continue
                
                # Use only first 170 characters for matching
                vuln_short = str(vulnerability)[:170]
                escaped_vulnerability = re.escape(vuln_short)
                pattern = rf'(?:\n|\r\n|\A){escaped_vulnerability}'
                
                # Find matching rows in catalog
                matching_rows = catalog_df[
                    catalog_df['Vulnerabilities in this group'].str.contains(
                        pattern, 
                        case=False, 
                        na=False,
                        regex=True
                    )
                ]
                
                if not matching_rows.empty:
                    catalog_idx = matching_rows.index[0]
                    matched_catalog_indices.add(catalog_idx)
                    matched_vulnerabilities.add(vulnerability)
                    
                    if catalog_idx not in vulnerability_groups:
                        vulnerability_groups[catalog_idx] = []
                    vulnerability_groups[catalog_idx].append(vulnerability)
                    
                    # Store the risk factor from catalog
                    if risk_factor_col is not None:
                        catalog_risk_value = catalog_df.loc[catalog_idx, catalog_headers[risk_factor_col]]
                        if pd.isna(catalog_risk_value):
                            catalog_risk_value = ""
                        catalog_risk_values[catalog_idx] = str(catalog_risk_value).upper().strip()
        
        # Sort catalog indices by risk factor
        risk_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
        
        def get_risk_priority(catalog_idx):
            risk_value = catalog_risk_values.get(catalog_idx, "")
            return risk_order.get(risk_value, 4)
        
        sorted_catalog_indices = sorted(matched_catalog_indices, key=get_risk_priority)
        
        # Second pass: Write data to worksheet
        row_num = 2
        
        for catalog_idx in sorted_catalog_indices:
            # Get catalog data - either from actual catalog or from merge state
            if use_merge_state and catalog_idx < 0:
                # New group from merge state (including manual vulnerabilities)
                group_data = next((g for g in merged_groups_from_session if g['catalog_id'] == catalog_idx), None)
                if not group_data:
                    continue
                
                # Get full details
                full_details = new_group_details_dict.get(str(catalog_idx), {})
                
                if full_details:
                    print(f"✓ Found details for catalog_idx {catalog_idx}: {list(full_details.keys())}")
                else:
                    print(f"⚠️ No details found for catalog_idx {catalog_idx} in new_group_details_dict")
                    print(f"   Available keys in new_group_details_dict: {list(new_group_details_dict.keys())}")
                
                # Create pseudo catalog row dynamically using catalog_headers
                # Map full_details keys to catalog column names
                catalog_row_dict = {}
                for col_name in catalog_headers:
                    if 'risk' in col_name.lower() and 'factor' in col_name.lower():
                        catalog_row_dict[col_name] = full_details.get('riskFactor', group_data.get('risk_factor', ''))
                    elif col_name == 'CVE/CWE ID':
                        catalog_row_dict[col_name] = full_details.get('cveId', 'N/A')
                    elif 'cvss' in col_name.lower():
                        catalog_row_dict[col_name] = full_details.get('cvssScore', group_data.get('cvss_score', ''))
                    elif 'audit' in col_name.lower() and 'observation' in col_name.lower():
                        catalog_row_dict[col_name] = full_details.get('auditObservation', '')
                    elif 'impact' in col_name.lower():
                        catalog_row_dict[col_name] = full_details.get('impact', '')
                    elif 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                        catalog_row_dict[col_name] = full_details.get('recommendation', '')
                    elif 'reference' in col_name.lower() and 'link' in col_name.lower():
                        catalog_row_dict[col_name] = full_details.get('referenceLink', '')
                    elif 'name' in col_name.lower() and 'vulnerability' in col_name.lower():
                        catalog_row_dict[col_name] = group_data.get('group_name', '')
                    else:
                        catalog_row_dict[col_name] = ''
                
                # Add excluded columns for completeness (won't be used)
                catalog_row_dict['Name of Vulnerability'] = group_data.get('group_name', '')
                catalog_row_dict['Affected System'] = ''
                catalog_row_dict['Vulnerabilities in this group'] = ''
                
                catalog_row = pd.Series(catalog_row_dict)
            else:
                # Standard catalog entry
                catalog_row = catalog_df.loc[catalog_idx]
            
            vulnerabilities_list = vulnerability_groups.get(catalog_idx, [])
            
            # Determine the highest risk for this group
            group_max_risk = 'low'
            risk_levels = {'critical': 4, 'high': 3, 'medium': 2, 'low': 1}
            for vuln in vulnerabilities_list:
                if risk_levels.get(vulnerability_risks.get(vuln, 'low'), 0) > risk_levels.get(group_max_risk, 0):
                    group_max_risk = vulnerability_risks.get(vuln, 'low')
            
            # Collect all affected systems for all vulnerabilities in this group
            all_affected_systems = []
            hosts_combined = set()
            
            for vuln in vulnerabilities_list:
                if vuln in vulnerability_affected_systems:
                    hosts_combined.update(vulnerability_affected_systems[vuln])
            
            # Format affected systems (just sorted hosts)
            all_affected_systems = sorted(list(hosts_combined))
            affected_systems_str = "\n".join(all_affected_systems) if all_affected_systems else ""
            
            # Write serial number
            cell = ws.cell(row=row_num, column=1, value=row_num - 1)
            cell.font = data_font
            cell.alignment = data_alignment_center
            cell.border = thin_border
            
            # Write vulnerabilities
            vulnerabilities_str = "\n".join(sorted(vulnerabilities_list))
            cell = ws.cell(row=row_num, column=2, value=vulnerabilities_str)
            cell.font = data_font
            cell.alignment = data_alignment_center
            cell.border = thin_border
            
            # Write catalog data up to insertion point
            col_idx = 3
            for i in range(insert_position):
                col_name = catalog_headers[i]
                value = catalog_row.get(col_name, "")
                if pd.isna(value):
                    value = ""
                
                # Special handling for Risk Factor column
                if risk_factor_col is not None and i == risk_factor_col:
                    catalog_risk_value = str(value).strip() if value else ""
                    
                    # Apply color formatting
                    if catalog_risk_value.upper() == 'CRITICAL':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = critical_font
                        cell.fill = critical_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    elif catalog_risk_value.upper() == 'HIGH':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = high_font
                        cell.fill = high_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    elif catalog_risk_value.upper() == 'MEDIUM':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = medium_font
                        cell.fill = medium_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    elif catalog_risk_value.upper() == 'LOW':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = low_font
                        cell.fill = low_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    else:
                        # Use highest from scan results
                        camelcase_value = convert_risk_to_camelcase(group_max_risk)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        if group_max_risk == 'critical':
                            cell.font = critical_font
                            cell.fill = critical_fill
                        elif group_max_risk == 'high':
                            cell.font = high_font
                            cell.fill = high_fill
                        elif group_max_risk == 'medium':
                            cell.font = medium_font
                            cell.fill = medium_fill
                        elif group_max_risk == 'low':
                            cell.font = low_font
                            cell.fill = low_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                else:
                    # Special handling for Audit Observation
                    if audit_observation_col is not None and i == audit_observation_col:
                        if len(vulnerabilities_list) > 1:
                            observation_text = "It was observed that the hosts are affected by multiple vulnerabilities, which are listed below.\n\n"
                            observation_text += vulnerabilities_str
                            if value:
                                value = f"{value}\n\n{observation_text}"
                            else:
                                value = observation_text
                        elif not value:
                            value = "It was observed that the host is affected by a vulnerability."
                    
                    # Special handling for CVE/CWE ID
                    if col_name == 'CVE/CWE ID' and value == "":
                        value = "N/A"
                    
                    # Apply left alignment for specific columns
                    cell = ws.cell(row=row_num, column=col_idx, value=str(value))
                    cell.font = data_font
                    if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                        cell.alignment = data_alignment_left
                    else:
                        cell.alignment = data_alignment_center
                    cell.border = thin_border
                
                col_idx += 1
            
            # Write affected systems
            cell = ws.cell(row=row_num, column=col_idx, value=affected_systems_str)
            cell.font = data_font
            cell.alignment = data_alignment_center
            cell.border = thin_border
            col_idx += 1
            
            # Write remaining catalog data
            for i in range(insert_position, len(catalog_headers)):
                col_name = catalog_headers[i]
                value = catalog_row.get(col_name, "")
                if pd.isna(value):
                    value = ""
                
                # Special handling for Risk Factor column
                if risk_factor_col is not None and i == risk_factor_col:
                    catalog_risk_value = str(value).strip() if value else ""
                    
                    # Apply color formatting (same logic as above)
                    if catalog_risk_value.upper() == 'CRITICAL':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = critical_font
                        cell.fill = critical_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    elif catalog_risk_value.upper() == 'HIGH':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = high_font
                        cell.fill = high_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    elif catalog_risk_value.upper() == 'MEDIUM':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = medium_font
                        cell.fill = medium_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    elif catalog_risk_value.upper() == 'LOW':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = low_font
                        cell.fill = low_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    else:
                        camelcase_value = convert_risk_to_camelcase(group_max_risk)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        if group_max_risk == 'critical':
                            cell.font = critical_font
                            cell.fill = critical_fill
                        elif group_max_risk == 'high':
                            cell.font = high_font
                            cell.fill = high_fill
                        elif group_max_risk == 'medium':
                            cell.font = medium_font
                            cell.fill = medium_fill
                        elif group_max_risk == 'low':
                            cell.font = low_font
                            cell.fill = low_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                else:
                    # Special handling for Audit Observation
                    if audit_observation_col is not None and i == audit_observation_col:
                        if len(vulnerabilities_list) > 1:
                            observation_text = "It was observed that the hosts are affected by multiple vulnerabilities, which are listed below.\n\n"
                            observation_text += vulnerabilities_str
                            if value:
                                value = f"{value}\n\n{observation_text}"
                            else:
                                value = observation_text
                        elif not value:
                            value = "It was observed that the host is affected by a vulnerability."
                    
                    # Special handling for CVE/CWE ID
                    if col_name == 'CVE/CWE ID' and value == "":
                        value = "N/A"
                    
                    # Apply left alignment for specific columns
                    cell = ws.cell(row=row_num, column=col_idx, value=str(value))
                    cell.font = data_font
                    if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                        cell.alignment = data_alignment_left
                    else:
                        cell.alignment = data_alignment_center
                    cell.border = thin_border
                
                col_idx += 1
            
            # Write Status column
            if status_col_position is not None:
                # Determine status based on first vulnerability in the list
                status_value = "New"  # Default to New
                if vulnerabilities_list:
                    first_vuln = vulnerabilities_list[0]
                    if first_vuln in comparison_result:
                        status_value = comparison_result[first_vuln]
                
                cell = ws.cell(row=row_num, column=status_col_position, value=status_value)
                cell.font = data_font
                cell.alignment = data_alignment_center
                cell.border = thin_border
            
            # Write all Old POC columns
            if old_poc_col_start is not None and old_poc_col_end is not None:
                for poc_col in range(old_poc_col_start, old_poc_col_end + 1):
                    cell = ws.cell(row=row_num, column=poc_col, value="")
                    cell.font = data_font
                    cell.alignment = data_alignment_center
                    cell.border = thin_border
            
            # Write all POC columns
            if poc_col_start is not None and poc_col_end is not None:
                for poc_col in range(poc_col_start, poc_col_end + 1):
                    cell = ws.cell(row=row_num, column=poc_col, value="")
                    cell.font = data_font
                    cell.alignment = data_alignment_center
                    cell.border = thin_border
            
            ws.row_dimensions[row_num].height = 37.5
            row_num += 1
        
        # Handle unmatched vulnerabilities - including merged ones
        unmatched_vulnerabilities = set(unique_vulnerabilities) - matched_vulnerabilities
        
        # Also include manual vulnerabilities that weren't in matched groups
        for manual_vuln in manually_added_vulnerabilities:
            found_in_groups = False
            for group in (merged_groups_from_session or []):
                if manual_vuln in group.get('matched_vulnerabilities', []):
                    found_in_groups = True
                    break
            if not found_in_groups:
                unmatched_vulnerabilities.add(manual_vuln)
        
        if unmatched_vulnerabilities:
            print(f"Unmatched vulnerabilities: {len(unmatched_vulnerabilities)}")
            
            for vulnerability in unmatched_vulnerabilities:
                # Get affected systems - first try from vulnerability_affected_systems dict
                affected_systems = vulnerability_affected_systems.get(vulnerability, [])
                
                # Get details from merge state if available
                details = None
                if use_merge_state:
                    # Find in new_group_details - check both matched and unmatched groups
                    for group in merged_groups_from_session:
                        if vulnerability in group.get('matched_vulnerabilities', []):
                            group_id = group['catalog_id']
                            details = new_group_details_dict.get(str(group_id), {})
                            break
                    
                    # If not found in matched groups, try to find in unmatched details
                    # For manual vulnerabilities or unmatched vulnerabilities that had details added
                    if not details:
                        # Search through all new_group_details to find this vulnerability
                        # Check both vulnerabilityName and actualVulnerabilityName
                        for group_id_str, group_details in new_group_details_dict.items():
                            if (group_details.get('vulnerabilityName') == vulnerability or 
                                group_details.get('actualVulnerabilityName') == vulnerability):
                                details = group_details
                                # Also get affected systems from details if not already set
                                if not affected_systems and 'affectedSystems' in details:
                                    affected_systems_str = str(details['affectedSystems']).strip()
                                    if affected_systems_str:
                                        if '\n' in affected_systems_str:
                                            affected_systems = [s.strip() for s in affected_systems_str.split('\n') if s.strip()]
                                        elif ',' in affected_systems_str:
                                            affected_systems = [s.strip() for s in affected_systems_str.split(',') if s.strip()]
                                        else:
                                            affected_systems = [affected_systems_str]
                                break
                
                affected_systems_str = "\n".join(affected_systems) if affected_systems else ""
                
                # Write serial number
                cell = ws.cell(row=row_num, column=1, value=row_num - 1)
                cell.font = data_font
                cell.alignment = data_alignment_center
                cell.border = thin_border
                
                # Write vulnerability name
                vuln_name = vulnerability
                if details and details.get('vulnerabilityName'):
                    vuln_name = details.get('vulnerabilityName')
                
                cell = ws.cell(row=row_num, column=2, value=vuln_name)
                cell.font = data_font
                cell.alignment = data_alignment_center
                cell.border = thin_border
                
                # Write catalog data columns
                col_idx = 3
                for i in range(insert_position):
                    col_name = catalog_headers[i]
                    value_to_write = ""
                    
                    if details:
                        if risk_factor_col is not None and i == risk_factor_col:
                            value_to_write = str(details.get('riskFactor', '')).upper()
                        elif audit_observation_col is not None and i == audit_observation_col:
                            value_to_write = details.get('auditObservation', '')
                        elif col_name == 'CVE/CWE ID':
                            value_to_write = details.get('cveId', '') or "N/A"
                        elif 'cvss' in col_name.lower():
                            value_to_write = details.get('cvssScore', '')
                        elif 'impact' in col_name.lower():
                            value_to_write = details.get('impact', '')
                        elif 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                            value_to_write = details.get('recommendation', '')
                        elif 'reference' in col_name.lower() and 'link' in col_name.lower():
                            value_to_write = details.get('referenceLink', '')
                    
                    # Apply risk color if applicable
                    if risk_factor_col is not None and i == risk_factor_col and value_to_write:
                        catalog_risk_value = str(value_to_write).upper().strip()
                        cell = ws.cell(row=row_num, column=col_idx, value=catalog_risk_value)
                        if catalog_risk_value == 'CRITICAL':
                            cell.font = critical_font
                            cell.fill = critical_fill
                        elif catalog_risk_value == 'HIGH':
                            cell.font = high_font
                            cell.fill = high_fill
                        elif catalog_risk_value == 'MEDIUM':
                            cell.font = medium_font
                            cell.fill = medium_fill
                        elif catalog_risk_value == 'LOW':
                            cell.font = low_font
                            cell.fill = low_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    else:
                        cell = ws.cell(row=row_num, column=col_idx, value=value_to_write)
                        cell.font = data_font
                        if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                            cell.alignment = data_alignment_left
                        else:
                            cell.alignment = data_alignment_center
                        cell.border = thin_border
                    
                    col_idx += 1
                
                # Write affected systems
                cell = ws.cell(row=row_num, column=col_idx, value=affected_systems_str)
                cell.font = data_font
                cell.alignment = data_alignment_center
                cell.border = thin_border
                col_idx += 1
                
                # Write remaining catalog columns
                for i in range(insert_position, len(catalog_headers)):
                    col_name = catalog_headers[i]
                    value_to_write = ""
                    
                    if details:
                        if risk_factor_col is not None and i == risk_factor_col:
                            value_to_write = str(details.get('riskFactor', '')).upper()
                        elif audit_observation_col is not None and i == audit_observation_col:
                            value_to_write = details.get('auditObservation', '')
                        elif col_name == 'CVE/CWE ID':
                            value_to_write = details.get('cveId', '') or "N/A"
                        elif 'cvss' in col_name.lower():
                            value_to_write = details.get('cvssScore', '')
                        elif 'impact' in col_name.lower():
                            value_to_write = details.get('impact', '')
                        elif 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                            value_to_write = details.get('recommendation', '')
                        elif 'reference' in col_name.lower() and 'link' in col_name.lower():
                            value_to_write = details.get('referenceLink', '')
                    
                    # Apply risk color if applicable
                    if risk_factor_col is not None and i == risk_factor_col and value_to_write:
                        catalog_risk_value = str(value_to_write).upper().strip()
                        cell = ws.cell(row=row_num, column=col_idx, value=catalog_risk_value)
                        if catalog_risk_value == 'CRITICAL':
                            cell.font = critical_font
                            cell.fill = critical_fill
                        elif catalog_risk_value == 'HIGH':
                            cell.font = high_font
                            cell.fill = high_fill
                        elif catalog_risk_value == 'MEDIUM':
                            cell.font = medium_font
                            cell.fill = medium_fill
                        elif catalog_risk_value == 'LOW':
                            cell.font = low_font
                            cell.fill = low_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    else:
                        cell = ws.cell(row=row_num, column=col_idx, value=value_to_write)
                        cell.font = data_font
                        if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                            cell.alignment = data_alignment_left
                        else:
                            cell.alignment = data_alignment_center
                        cell.border = thin_border
                    
                    col_idx += 1
                
                # Write Status column
                if status_col_position is not None:
                    cell = ws.cell(row=row_num, column=status_col_position, value="")
                    cell.font = data_font
                    cell.alignment = data_alignment_center
                    cell.border = thin_border
                
                # Write all Old POC columns
                if old_poc_col_start is not None and old_poc_col_end is not None:
                    for poc_col in range(old_poc_col_start, old_poc_col_end + 1):
                        cell = ws.cell(row=row_num, column=poc_col, value="")
                        cell.font = data_font
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                
                # Write all POC columns
                if poc_col_start is not None and poc_col_end is not None:
                    for poc_col in range(poc_col_start, poc_col_end + 1):
                        cell = ws.cell(row=row_num, column=poc_col, value="")
                        cell.font = data_font
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                
                ws.row_dimensions[row_num].height = 37.5
                row_num += 1
        
        # Set column widths according to specified array
        column_widths = [7, 35, 30, 15, 20, 10, 60, 60, 60, 40, 50, 30, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25]
        
        # Apply column widths to columns A through Z (26 columns)
        for col_idx, width in enumerate(column_widths, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
        
        print(f"✅ Public_IP_VAPT worksheet created with {row_num - 2} rows")
        
    except Exception as e:
        print(f"❌ Error creating Public_IP_VAPT worksheet: {e}")
        import traceback
        traceback.print_exc()


def extract_risk_factor_counts_from_user_excel_public_ip(user_excel_file):
    """
    Extract risk factor counts from user's Excel file Public_IP_VAPT worksheet.
    Returns a dictionary with risk factor counts (Critical, High, Medium, Low).
    """
    try:
        from openpyxl import load_workbook
        
        # Load the user's workbook
        wb = load_workbook(user_excel_file)
        
        # Get the Public_IP_VAPT worksheet
        if "Public_IP_VAPT" not in wb.sheetnames:
            print("Warning: Public_IP_VAPT worksheet not found in user's Excel file")
            return {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        
        ws = wb["Public_IP_VAPT"]
        
        # Find Risk Factor column
        risk_factor_col = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and 'risk' in str(cell_value).lower() and 'factor' in str(cell_value).lower():
                risk_factor_col = col
                break
        
        if risk_factor_col is None:
            print("Warning: Risk Factor column not found in user's Excel file")
            return {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        
        # Count risk factors
        risk_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        
        for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
            cell_value = ws.cell(row=row, column=risk_factor_col).value
            if cell_value:
                risk_value = str(cell_value).strip().upper()
                if risk_value in risk_counts:
                    risk_counts[risk_value] += 1
                elif risk_value in ["CRITICAL"]:
                    risk_counts["Critical"] += 1
                elif risk_value in ["HIGH"]:
                    risk_counts["High"] += 1
                elif risk_value in ["MEDIUM"]:
                    risk_counts["Medium"] += 1
                elif risk_value in ["LOW"]:
                    risk_counts["Low"] += 1
        
        print(f"Risk factor counts from user's Excel: {risk_counts}")
        return risk_counts
        
    except Exception as e:
        print(f"Error extracting risk factor counts from user's Excel file: {e}")
        import traceback
        traceback.print_exc()
        return {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}


def extract_follow_up_risk_factor_counts_from_generated_excel_public_ip(excel_path):
    """
    Extract risk factor counts from generated Excel file Public_IP_VAPT worksheet.
    Only counts vulnerabilities with Status 'New' or 'Open' (excludes 'Closed').
    Returns a dictionary with risk factor counts (Critical, High, Medium, Low).
    """
    try:
        from openpyxl import load_workbook
        
        # Load the generated workbook
        wb = load_workbook(excel_path)
        
        # Get the Public_IP_VAPT worksheet
        if "Public_IP_VAPT" not in wb.sheetnames:
            print("Warning: Public_IP_VAPT worksheet not found in generated Excel file")
            return {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        
        ws = wb["Public_IP_VAPT"]
        
        # Find Risk Factor and Status columns
        risk_factor_col = None
        status_col = None
        
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                cell_str = str(cell_value).lower()
                if 'risk' in cell_str and 'factor' in cell_str:
                    risk_factor_col = col
                elif 'status' in cell_str:
                    status_col = col
        
        if risk_factor_col is None:
            print("Warning: Risk Factor column not found in generated Excel file")
            return {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        
        if status_col is None:
            print("Warning: Status column not found in generated Excel file")
            return {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        
        # Count risk factors (only for New/Open status)
        risk_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        
        for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
            # Check status first
            status_value = ws.cell(row=row, column=status_col).value
            if not status_value:
                continue
                
            status_str = str(status_value).strip().upper()
            # Only count if status is 'New' or 'Open'
            if status_str not in ['NEW', 'OPEN']:
                continue
            
            # Count risk factor
            risk_value = ws.cell(row=row, column=risk_factor_col).value
            if risk_value:
                risk_str = str(risk_value).strip().upper()
                if risk_str in risk_counts:
                    risk_counts[risk_str] += 1
                elif risk_str in ["CRITICAL"]:
                    risk_counts["Critical"] += 1
                elif risk_str in ["HIGH"]:
                    risk_counts["High"] += 1
                elif risk_str in ["MEDIUM"]:
                    risk_counts["Medium"] += 1
                elif risk_str in ["LOW"]:
                    risk_counts["Low"] += 1
        
        print(f"Follow-up risk factor counts (New/Open only) from generated Excel: {risk_counts}")
        return risk_counts
        
    except Exception as e:
        print(f"Error extracting follow-up risk factor counts from generated Excel file: {e}")
        import traceback
        traceback.print_exc()
        return {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}


def create_public_ip_metadata_worksheet(wb, form_data, user_excel_file=None, generated_excel_path=None):
    """
    Create Meta_Data worksheet - structured format same as Infrastructure VAPT reference.
    """
    try:
        print("📊 Creating Meta_Data worksheet...")
        ws = wb.create_sheet("Meta_Data")
        
        # Get First Audit Report details from form
        first_audit_report_id = form_data.get('firstAuditReportId', '')
        first_audit_report_date = form_data.get('firstAuditReportDate', '')
        
        # Format the date as DD.MM.YYYY if provided
        if first_audit_report_date:
            try:
                from datetime import datetime
                date_obj = datetime.strptime(first_audit_report_date, '%Y-%m-%d')
                first_audit_report_date = date_obj.strftime('%d.%m.%Y')
            except Exception as e:
                print(f"Error formatting first audit report date: {e}")
        
        # Extract risk factor counts from user's Excel file if provided
        risk_factor_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        if user_excel_file and user_excel_file.filename != '':
            try:
                print(f"Processing user's Excel file: {user_excel_file.filename}")
                risk_factor_counts = extract_risk_factor_counts_from_user_excel_public_ip(user_excel_file)
                print(f"Extracted risk factor counts: {risk_factor_counts}")
            except Exception as e:
                print(f"Error extracting risk factor counts from user's Excel file: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("No user Excel file provided for risk factor extraction")
        
        # Extract follow-up risk factor counts from generated Excel file if provided
        follow_up_risk_factor_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
        if generated_excel_path and os.path.exists(generated_excel_path):
            try:
                print(f"Processing generated Excel file: {generated_excel_path}")
                follow_up_risk_factor_counts = extract_follow_up_risk_factor_counts_from_generated_excel_public_ip(generated_excel_path)
                print(f"Extracted follow-up risk factor counts: {follow_up_risk_factor_counts}")
            except Exception as e:
                print(f"Error extracting follow-up risk factor counts from generated Excel file: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("No generated Excel file provided for follow-up risk factor extraction")
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 35  # Field names
        ws.column_dimensions['B'].width = 60  # Values
        
        # Define the data structure for the metadata
        metadata_sections = [
            {
                'title': 'ORGANIZATION INFORMATION',
                'data': [
                    ('Organization Name', form_data.get('organization', '')),
                    ('City', form_data.get('city', '')),
                    ('State', form_data.get('state', '')),
                    ('First Audit Report ID', first_audit_report_id),
                    ('First Audit Report Date', first_audit_report_date)
                ]
            },
            {
                'title': 'AUDIT PERIOD',
                'data': [
                    ('Start Date', form_data.get('startDate', '')),
                    ('End Date', form_data.get('endDate', ''))
                ]
            },
            {
                'title': 'REPORT PREPARED BY',
                'data': [
                    ('Name', f"{form_data.get('preparedByTitle', '')} {form_data.get('preparedByName', '')}".strip()),
                ]
            },
            {
                'title': 'AUDITEE DETAILS',
                'data': [
                    ('Name', f"{form_data.get('auditeeTitle', '')} {form_data.get('auditeeName', '')}".strip()),
                    ('Designation', form_data.get('designation', ''))
                ]
            },
            {
                'title': 'FIRST AUDIT',
                'data': [
                    ('Critical', str(risk_factor_counts['Critical'])),
                    ('High', str(risk_factor_counts['High'])),
                    ('Medium', str(risk_factor_counts['Medium'])),
                    ('Low', str(risk_factor_counts['Low']))
                ]
            },
            {
                'title': 'FOLLOW UP AUDIT',
                'data': [
                    ('Critical', str(follow_up_risk_factor_counts['Critical'])),
                    ('High', str(follow_up_risk_factor_counts['High'])),
                    ('Medium', str(follow_up_risk_factor_counts['Medium'])),
                    ('Low', str(follow_up_risk_factor_counts['Low']))
                ]
            }
        ]
        
        # Add Email Addresses section
        bank_emails = form_data.get('bankEmails', [])
        if bank_emails:
            email_data = []
            for i, email in enumerate(bank_emails, 1):
                if email.strip():
                    email_data.append((f'Email {i}', email.strip()))
            
            if email_data:
                metadata_sections.append({
                    'title': 'ORGANIZATION EMAIL ADDRESSES',
                    'data': email_data
                })
        
        # Add Auditing Team sections
        team_names = form_data.get('teamNames', [])
        team_designations = form_data.get('teamDesignations', [])
        team_emails = form_data.get('teamEmails', [])
        team_qualifications = form_data.get('teamQualifications', [])
        team_certified = form_data.get('teamCertified', [])
        
        if team_names:
            for i in range(len(team_names)):
                if team_names[i].strip():
                    team_member_data = [
                        (f'Team Member {i+1} - Name', team_names[i].strip()),
                        (f'Team Member {i+1} - Designation', team_designations[i] if i < len(team_designations) else ''),
                        (f'Team Member {i+1} - Email', team_emails[i] if i < len(team_emails) else ''),
                        (f'Team Member {i+1} - Qualification', team_qualifications[i] if i < len(team_qualifications) else ''),
                        (f'Team Member {i+1} - Certified', 
                         'Yes' if team_certified[i].lower() == 'yes' else 'No' if team_certified[i].lower() == 'no' else team_certified[i] 
                         if i < len(team_certified) else '')
                    ]
                    
                    metadata_sections.append({
                        'title': f'AUDITING TEAM MEMBER {i+1}',
                        'data': team_member_data
                    })
        
        # Write data to worksheet
        row = 1
        
        for section in metadata_sections:
            # Write section title (merged across both columns)
            cell = ws.cell(row=row, column=1, value=section['title'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=row, column=2, value='')
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            ws.row_dimensions[row].height = 25
            row += 1
            
            # Write section data
            for field_name, field_value in section['data']:
                cell = ws.cell(row=row, column=1, value=field_name)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                cell = ws.cell(row=row, column=2, value=field_value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                ws.row_dimensions[row].height = 20
                row += 1
            
            # Add empty row after each section for spacing
            row += 1
        
        print(f"✅ Meta_Data worksheet created with {row} rows")
        
    except Exception as e:
        print(f"❌ Error creating Meta_Data worksheet: {e}")
        import traceback
        traceback.print_exc()


def extract_old_poc_images_from_user_excel(user_excel_file, target_vulnerabilities, output_folder="Temp_POC2_Website"):
    """
    Extract POC/evidence IMAGES from user's Excel file for specified vulnerabilities.
    Saves images to a local folder and returns a dictionary mapping vulnerability names to image paths.
    """
    try:
        from openpyxl import load_workbook
        import tempfile
        import os
        import shutil
        
        # Create output folder if it doesn't exist
        if os.path.exists(output_folder):
            shutil.rmtree(output_folder)
        os.makedirs(output_folder)
        print(f"Created folder: {output_folder}")
        
        # Handle Flask file upload object - save to temp file first
        temp_file_path = None
        try:
            if hasattr(user_excel_file, 'read'):
                user_excel_file.seek(0)
                # Save upload file to temporary file
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                    temp_file.write(user_excel_file.read())
                    temp_file_path = temp_file.name
                print(f"Saved user Excel to: {temp_file_path}")
                wb = load_workbook(temp_file_path)
            else:
                wb = load_workbook(user_excel_file)
        except Exception as e:
            print(f"Error loading workbook: {e}")
            import traceback
            traceback.print_exc()
            return {}
        
        # Get the Public_IP_VAPT worksheet
        if "Public_IP_VAPT" not in wb.sheetnames:
            print("Warning: Public_IP_VAPT worksheet not found in user's Excel file")
            wb.close()
            if temp_file_path and os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
            return {}
        
        ws = wb["Public_IP_VAPT"]
        
        # Find POC column range in user's Excel (first audit has "POC", not "Old POC")
        poc_col_start = None
        poc_col_end = None
        
        # Check for merged cells in row 1 for POC header
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == 1 and merged_range.max_row == 1:
                first_cell = ws.cell(row=1, column=merged_range.min_col)
                cell_value = str(first_cell.value).strip().lower() if first_cell.value else ""
                # Look for "POC" (not "Old POC" since user's Excel is from first audit)
                if 'poc' in cell_value and 'old' not in cell_value:
                    poc_col_start = merged_range.min_col
                    poc_col_end = merged_range.max_col
                    print(f"Found merged POC columns from {poc_col_start} to {poc_col_end}")
                    break
        
        # If not found in merged cells, look for individual POC columns
        if poc_col_start is None:
            # Look for individual columns that contain "POC" but not "Old"
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    cell_str = str(cell_value).strip().lower()
                    if 'poc' in cell_str and 'old' not in cell_str:
                        if poc_col_start is None:
                            poc_col_start = col
                        poc_col_end = col
        
        if poc_col_start is None:
            print("Warning: POC column(s) not found in user's Excel file")
            print(f"DEBUG: Looking through headers...")
            for col in range(1, min(ws.max_column + 1, 30)):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    print(f"  Column {col}: {cell_value}")
            wb.close()
            if temp_file_path and os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
            return {}
        
        print(f"Found POC columns from {poc_col_start} to {poc_col_end} in user's Excel (will be placed in Old POC columns)")
        
        # Find Name of Vulnerability column
        vuln_col = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and 'Name of Vulnerability' in str(cell_value):
                vuln_col = col
                break
        
        if vuln_col is None:
            print("Warning: Name of Vulnerability column not found in user's Excel file")
            wb.close()
            if temp_file_path and os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
            return {}
        
        # Extract images and save to folder
        image_counter = {}
        images_by_vulnerability = {}
        
        # Check if worksheet has images
        print(f"DEBUG: ws._images exists: {hasattr(ws, '_images')}")
        if hasattr(ws, "_images"):
            print(f"DEBUG: len(ws._images): {len(ws._images) if ws._images else 0}")
        
        # Extract all images from the worksheet
        if hasattr(ws, "_images") and ws._images:
            print(f"Found {len(ws._images)} images in user's Excel")
            
            for img in ws._images:
                try:
                    row_idx = img.anchor._from.row + 1
                    col_idx = img.anchor._from.col + 1
                    
                    # Check if this image is in the POC column range (from first audit)
                    if poc_col_start <= col_idx <= poc_col_end:
                        if row_idx >= 2:
                            vuln_name_cell = ws.cell(row=row_idx, column=vuln_col)
                            if vuln_name_cell.value:
                                vuln_name = str(vuln_name_cell.value).strip()
                                vuln_name_short = vuln_name[:170]
                                
                                # Check if this vulnerability is in the target list
                                for target_vuln in target_vulnerabilities:
                                    if target_vuln[:170] == vuln_name_short:
                                        # Normalize vulnerability name for filename
                                        safe_name = vuln_name.replace('\\', '_').replace('/', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
                                        
                                        # Counter for multiple images
                                        if safe_name not in image_counter:
                                            image_counter[safe_name] = 0
                                        image_counter[safe_name] += 1
                                        
                                        # Get image data
                                        img_data = img._data() if callable(img._data) else img._data
                                        
                                        if img_data:
                                            # Save image to folder
                                            image_number = image_counter[safe_name]
                                            filename = f"{safe_name}_{image_number}.png"
                                            file_path = os.path.join(output_folder, filename)
                                            
                                            with open(file_path, 'wb') as img_file:
                                                img_file.write(img_data)
                                            
                                            # Store the image path
                                            if target_vuln not in images_by_vulnerability:
                                                images_by_vulnerability[target_vuln] = []
                                            images_by_vulnerability[target_vuln].append(file_path)
                                            
                                            print(f"  ✅ Extracted Old POC image for vulnerability '{target_vuln}': {filename}")
                                        break
                
                except Exception as e:
                    print(f"  ⚠️ Error extracting image: {e}")
                    continue
        
        wb.close()
        if temp_file_path and os.path.exists(temp_file_path):
            os.unlink(temp_file_path)
        
        print(f"Extracted Old POC images for {len(images_by_vulnerability)} vulnerabilities")
        return images_by_vulnerability
        
    except Exception as e:
        print(f"Error extracting Old POC images from user's Excel: {e}")
        import traceback
        traceback.print_exc()
        return {}

def insert_old_poc_images_for_open_vulnerabilities(excel_path, old_poc_images, open_vulnerabilities):
    """
    Insert Old POC images for Open vulnerabilities into the existing worksheet.
    Uses the same approach as Infrastructure VAPT Follow-up.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image
        from openpyxl.utils import get_column_letter
        import os
        import re
        
        # Load the workbook
        wb = load_workbook(excel_path)
        
        # Get the Public_IP_VAPT worksheet
        if "Public_IP_VAPT" not in wb.sheetnames:
            print("Public_IP_VAPT worksheet not found")
            wb.close()
            return False
        
        ws = wb["Public_IP_VAPT"]
        
        # Find Old POC columns by looking for merged header
        old_poc_col_start = None
        old_poc_col_end = None
        
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == 1 and merged_range.max_row == 1:
                first_cell = ws.cell(row=1, column=merged_range.min_col)
                if first_cell.value and str(first_cell.value).strip() == "Old POC":
                    old_poc_col_start = merged_range.min_col
                    old_poc_col_end = merged_range.max_col
                    break
        
        if not old_poc_col_start or not old_poc_col_end:
            print("Old POC columns not found in worksheet")
            wb.close()
            return False
        
        # Define column order for Old POC image insertion: N, O, P, Q, R, S, M
        old_image_columns = [
            old_poc_col_start + 1,  # N
            old_poc_col_start + 2,  # O
            old_poc_col_start + 3,  # P
            old_poc_col_start + 4,  # Q
            old_poc_col_start + 5,  # R
            old_poc_col_end,        # S
            old_poc_col_start       # M
        ]
        
        print(f"Found Old POC columns from {old_poc_col_start} to {old_poc_col_end}")
        
        # Find Name of Vulnerability column
        vuln_name_col = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and 'name of vulnerability' in str(cell_value).lower():
                vuln_name_col = col
                break
        
        if not vuln_name_col:
            print("Name of Vulnerability column not found in worksheet")
            wb.close()
            return False
        
        print(f"Found Name of Vulnerability column at position: {vuln_name_col}")
        
        # Normalize vulnerability names for matching
        def normalize_for_matching(text):
            """Normalize text for matching by replacing special chars with spaces"""
            special_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|', '_', '-']
            normalized = text.lower()
            for char in special_chars:
                normalized = normalized.replace(char, ' ')
            # Normalize multiple spaces to single space
            normalized = ' '.join(normalized.split())
            return normalized
        
        rows_with_objects = set()
        
        # Process each row
        for row in range(2, ws.max_row + 1):
            vuln_name_cell = ws.cell(row=row, column=vuln_name_col)
            if not vuln_name_cell.value:
                continue
            
            vuln_name = str(vuln_name_cell.value).strip()
            if not vuln_name:
                continue
            
            # Use only first 170 characters for matching
            vuln_name_short = vuln_name[:170].strip()
            vuln_normalized = normalize_for_matching(vuln_name_short)
            
            # Find matching images
            matching_images = []
            for mapped_name, image_paths in old_poc_images.items():
                mapped_name_short = mapped_name[:170].strip()
                mapped_normalized = normalize_for_matching(mapped_name_short)
                
                if vuln_normalized == mapped_normalized:
                    # Handle both single path and list of paths
                    if isinstance(image_paths, list):
                        for image_path in image_paths:
                            matching_images.append((mapped_name, image_path))
                    else:
                        matching_images.append((mapped_name, image_paths))
            
            # Sort matching images
            matching_images.sort(key=lambda x: x[1])
            
            if matching_images:
                print(f"✅ Found {len(matching_images)} Old POC image(s) for '{vuln_name[:50]}...'")
                
                num_images_to_insert = min(len(matching_images), 7)
                
                for img_idx in range(num_images_to_insert):
                    mapped_name, matching_image = matching_images[img_idx]
                    col_idx = old_image_columns[img_idx]
                    
                    if os.path.exists(matching_image):
                        try:
                            img = Image(matching_image)
                            
                            # Resize image
                            img.width = img.width / 30
                            img.height = img.height / 30
                            
                            col_letter = get_column_letter(col_idx)
                            cell_ref = f"{col_letter}{row}"
                            
                            ws.add_image(img, cell_ref)
                            print(f"  ✅ Inserted Old POC image {img_idx + 1} at {cell_ref}")
                            
                        except Exception as e:
                            print(f"  ⚠️ Error inserting image at column {col_idx}, row {row}: {e}")
                
                if num_images_to_insert > 0:
                    rows_with_objects.add(row)
        
        # Apply custom borders to ALL Old POC and POC columns
        from openpyxl.styles import Border, Side
        
        # Define border styles for each column type
        # M column: left, top, bottom (NOT right)
        m_border = Border(
            left=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # N, O, P, Q, R columns: top and bottom only (NOT left or right)
        middle_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # S column: top, bottom, right (NOT left)
        right_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            right=Side(style='thin')
        )
        
        # Find all table rows (rows with content)
        table_rows = set()
        table_rows.add(1)  # Header row
        
        for row in range(2, ws.max_row + 1):
            has_content = False
            # Check if any cell in this row has content (excluding Old POC and POC columns)
            for col in range(1, old_poc_col_start):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip() != "":
                    has_content = True
                    break
            
            if has_content:
                table_rows.add(row)
        
        # Apply custom borders to all table rows for Old POC columns
        for row_num in sorted(table_rows):
            # M column: left, top, bottom
            ws.cell(row=row_num, column=old_poc_col_start).border = m_border
            
            # N, O, P, Q, R columns: top, bottom only
            for col_idx in range(old_poc_col_start + 1, old_poc_col_end):
                ws.cell(row=row_num, column=col_idx).border = middle_border
            
            # S column: top, bottom, right
            ws.cell(row=row_num, column=old_poc_col_end).border = right_border
        
        print(f"Applied custom borders to Old POC columns for {len(table_rows)} rows")
        
        # Save the workbook
        wb.save(excel_path)
        wb.close()
        
        print(f"✅ Inserted Old POC images for {len(rows_with_objects)} Open vulnerabilities")
        return True
        
    except Exception as e:
        print(f"Error inserting Old POC images for Open vulnerabilities: {e}")
        import traceback
        traceback.print_exc()
        return False

def sort_and_renumber_public_ip_vapt_worksheet(excel_path):
    """
    Sort the Public_IP_VAPT worksheet by risk level (Critical, High, Medium, Low) and renumber rows.
    Also handles image repositioning when rows are sorted.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image
        from openpyxl.utils import get_column_letter
        import pandas as pd
        
        # Load the workbook
        wb = load_workbook(excel_path)
        
        # Get the Public_IP_VAPT worksheet
        if "Public_IP_VAPT" not in wb.sheetnames:
            print("Error: Public_IP_VAPT worksheet not found")
            wb.close()
            return False
        
        ws = wb["Public_IP_VAPT"]
        
        # Store images info for tracking
        num_images = len(ws._images) if hasattr(ws, '_images') and ws._images else 0
        print(f"Found {num_images} images before sorting")
        
        # Convert worksheet to DataFrame for easier manipulation
        data = []
        headers = []
        
        # Get headers from first row
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value else f"Column_{col}")
        
        # Get data from all rows (excluding header)
        for row in range(2, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(cell_value)
            data.append(row_data)
        
        if not data:
            print("No data rows found in Public_IP_VAPT worksheet")
            wb.close()
            return True
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        # Find Risk Factor column
        risk_factor_col = None
        for i, col in enumerate(headers):
            if 'risk' in col.lower() and 'factor' in col.lower():
                risk_factor_col = i
                break
        
        if risk_factor_col is None:
            print("Warning: Risk Factor column not found, skipping sort")
            wb.close()
            return True
        
        # Define risk level priority (case-insensitive matching)
        risk_priority = {'critical': 1, 'high': 2, 'medium': 3, 'low': 4}
        
        # Create a list of rows with their risk priorities for sorting
        rows_with_priority = []
        for row_idx, row_data in enumerate(data):
            risk_value = str(row_data[risk_factor_col]).strip() if pd.notna(row_data[risk_factor_col]) else ""
            # Normalize to lowercase for case-insensitive matching
            risk_value_lower = risk_value.lower()
            priority = risk_priority.get(risk_value_lower, 5)
            rows_with_priority.append((priority, row_idx, row_data))
        
        # Sort by priority (risk level)
        rows_with_priority.sort(key=lambda x: x[0])
        
        # Find Sr No column for renumbering
        sr_no_col = None
        for i, col in enumerate(headers):
            if 'sr' in col.lower() and 'no' in col.lower():
                sr_no_col = i
                break
        
        # Clear data cells (keep header) - but preserve row heights
        row_heights = {}  # Store original row heights
        for row in range(2, ws.max_row + 1):
            row_heights[row] = ws.row_dimensions[row].height
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
        
        # Note: We don't clear images here - we'll update their anchor positions
        # after sorting the data
        
        # Write sorted data back to worksheet
        new_row_to_old_row = {}  # Map: new row -> old row (1-based)
        for new_row_idx, (priority, old_row_idx, row_data) in enumerate(rows_with_priority, start=2):
            actual_old_row = old_row_idx + 2  # Convert back to 1-based row number
            new_row_to_old_row[new_row_idx] = actual_old_row
            
            for col_idx, value in enumerate(row_data, start=1):
                if value is not None and pd.notna(value):
                    ws.cell(row=new_row_idx, column=col_idx, value=value)
                else:
                    ws.cell(row=new_row_idx, column=col_idx, value="")
            
            # Renumber the Sr No column (sequentially 1, 2, 3, 4...)
            if sr_no_col is not None:
                ws.cell(row=new_row_idx, column=sr_no_col + 1, value=new_row_idx - 1)
            
            # Preserve row height if it was set
            if actual_old_row in row_heights and row_heights[actual_old_row] is not None:
                ws.row_dimensions[new_row_idx].height = row_heights[actual_old_row]
        
        # Update image anchor positions to match sorted data
        print("📸 Updating image anchor positions after sorting...")
        images_moved = 0
        
        # Update the anchor row for each image
        if hasattr(ws, '_images') and ws._images:
            for img in ws._images:
                try:
                    if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                        old_row = img.anchor._from.row + 1  # Convert to 1-based
                        old_col = img.anchor._from.col + 1
                        
                        # Find which new row this old row corresponds to
                        # Get list of old row numbers that exist in data
                        valid_old_rows = set()
                        for _, old_idx, _ in rows_with_priority:
                            actual_old_row = old_idx + 2  # Convert to 1-based
                            valid_old_rows.add(actual_old_row)
                        
                        if old_row in valid_old_rows:
                            # Find the new row position for this old row
                            # We have new_row_to_old_row mapping, so we need the reverse
                            old_row_to_new_row = {v: k for k, v in new_row_to_old_row.items()}
                            
                            if old_row in old_row_to_new_row:
                                new_row = old_row_to_new_row[old_row]
                                
                                # Update the anchor to point to new row
                                from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
                                img.anchor._from.row = new_row - 1  # Convert back to 0-based
                                
                                # If it's a TwoCellAnchor, also update the 'to' anchor
                                if hasattr(img.anchor, '_to') and hasattr(img.anchor._to, 'row'):
                                    row_diff = img.anchor._to.row - img.anchor._from.row
                                    img.anchor._to.row = new_row - 1 + row_diff
                                
                                images_moved += 1
                                
                except Exception as e:
                    print(f"⚠️ Error updating image anchor: {e}")
                    import traceback
                    traceback.print_exc()
                    continue
        
        print(f"✅ Updated {images_moved} image anchor positions")
        
        # Apply color formatting to Risk Factor column
        if risk_factor_col is not None:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            # Create color formats
            critical_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
            high_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            medium_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            low_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
            
            white_font = Font(color="FFFFFF", bold=True)
            center_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply formatting to each row in Risk Factor column
            for row_idx in range(2, len(rows_with_priority) + 2):
                cell = ws.cell(row=row_idx, column=risk_factor_col + 1)
                risk_value = str(cell.value).strip().upper() if cell.value else ""
                
                if risk_value == 'CRITICAL':
                    cell.fill = critical_fill
                    cell.font = white_font
                    cell.alignment = center_alignment
                    cell.border = border
                    cell.value = 'Critical'
                elif risk_value == 'HIGH':
                    cell.fill = high_fill
                    cell.font = white_font
                    cell.alignment = center_alignment
                    cell.border = border
                    cell.value = 'High'
                elif risk_value == 'MEDIUM':
                    cell.fill = medium_fill
                    cell.font = white_font
                    cell.alignment = center_alignment
                    cell.border = border
                    cell.value = 'Medium'
                elif risk_value == 'LOW':
                    cell.fill = low_fill
                    cell.font = white_font
                    cell.alignment = center_alignment
                    cell.border = border
                    cell.value = 'Low'
                else:
                    cell.font = Font()
                    cell.alignment = center_alignment
                    cell.border = border
        
        # Apply borders to Status column (no color, just border)
        status_col = None
        for col_idx, header in enumerate(headers, 1):
            if 'status' in header.lower():
                status_col = col_idx
                break
        
        if status_col is not None:
            from openpyxl.styles import Border, Side, Alignment
            
            status_cell_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            status_cell_alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply to all Status column cells
            for row_idx in range(2, len(rows_with_priority) + 2):
                status_cell = ws.cell(row=row_idx, column=status_col)
                if status_cell.value:
                    status_cell.border = status_cell_border
                    status_cell.alignment = status_cell_alignment
        
        # Save the workbook
        wb.save(excel_path)
        wb.close()
        
        print(f"Public_IP_VAPT worksheet sorted by risk level and renumbered successfully")
        print(f"Total vulnerabilities: {len(rows_with_priority)}")
        
        # Print risk level distribution
        if risk_factor_col is not None:
            risk_counts = {}
            for priority, old_row_idx, row_data in rows_with_priority:
                risk_value = str(row_data[risk_factor_col]).strip() if pd.notna(row_data[risk_factor_col]) else "Unknown"
                risk_counts[risk_value] = risk_counts.get(risk_value, 0) + 1
            
            print("Risk level distribution after sorting:")
            for risk, count in risk_counts.items():
                print(f"  {risk}: {count}")
        
        return True
        
    except Exception as e:
        print(f"Error sorting and renumbering Public_IP_VAPT worksheet: {e}")
        import traceback
        traceback.print_exc()
        return False

def apply_old_poc_column_borders(ws, old_poc_col_start, old_poc_col_end):
    """
    Apply custom borders to Old POC columns (M-S) similar to Infrastructure VAPT.
    - M column: left, top, bottom (NOT right)
    - N, O, P, Q, R columns: top and bottom only
    - S column: top, bottom, right (NOT left)
    """
    try:
        from openpyxl.styles import Border, Side
        
        # Define border styles for each column type
        m_border = Border(
            left=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        middle_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        right_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            right=Side(style='thin')
        )
        
        # Find all table rows
        table_rows = set()
        table_rows.add(1)  # Header
        
        for row in range(2, ws.max_row + 1):
            has_content = False
            for col in range(1, old_poc_col_start):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip() != "":
                    has_content = True
                    break
            
            if has_content:
                table_rows.add(row)
        
        # Apply borders
        for row_num in sorted(table_rows):
            ws.cell(row=row_num, column=old_poc_col_start).border = m_border
            
            for col_idx in range(old_poc_col_start + 1, old_poc_col_end):
                ws.cell(row=row_num, column=col_idx).border = middle_border
            
            ws.cell(row=row_num, column=old_poc_col_end).border = right_border
        
        print(f"Applied custom borders to Old POC columns for {len(table_rows)} rows")
        return True
        
    except Exception as e:
        print(f"Error applying Old POC column borders: {e}")
        return False

def add_closed_vulnerabilities_to_excel(excel_path, closed_vulnerabilities, user_row_data, wb, old_poc_images=None):
    """
    Add rows for Closed vulnerabilities from user's Excel to our generated Excel.
    Applies the same formatting as existing rows and inserts Old POC images.
    """
    if old_poc_images is None:
        old_poc_images = {}
    
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
        
        # Load the workbook from file
        wb = load_workbook(excel_path)
        
        # Get the Public_IP_VAPT worksheet
        if "Public_IP_VAPT" not in wb.sheetnames:
            print("Error: Public_IP_VAPT worksheet not found")
            wb.close()
            return False
        
        ws = wb["Public_IP_VAPT"]
        
        # Get headers from first row
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value else f"Column_{col}")
        
        # Find the Status column index
        status_col = None
        for col_idx, header in enumerate(headers, 1):
            if header and 'status' in str(header).lower():
                status_col = col_idx
                break
        
        # Create formatting styles
        critical_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
        high_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        medium_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        low_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
        
        white_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Add each Closed vulnerability row
        print(f"\n📊 DEBUG: add_closed_vulnerabilities_to_excel called")
        print(f"   Number of closed vulnerabilities: {len(closed_vulnerabilities)}")
        print(f"   Number of user_row_data entries: {len(user_row_data)}")
        print(f"   Excel path: {excel_path}")
        
        print(f"Adding {len(closed_vulnerabilities)} Closed vulnerabilities to Excel")
        for vuln_name in closed_vulnerabilities:
            print(f"   Processing: {vuln_name[:100]}...")
            if vuln_name in user_row_data:
                user_row = user_row_data[vuln_name]
                print(f"   ✅ Found in user_row_data. Adding Closed vulnerability: {vuln_name[:50]}")
                
                # Find Old POC column start and end
                old_poc_start = None
                old_poc_end = None
                for col_idx, header in enumerate(headers, 1):
                    if 'Old_POC_1' in str(header):
                        old_poc_start = col_idx
                    elif 'Old_POC_7' in str(header):
                        old_poc_end = col_idx
                        break
                
                # Map user's data to our worksheet columns
                for col_idx, header in enumerate(headers, 1):
                    # Skip Old POC and POC columns - will be handled separately
                    if 'Old POC' in str(header) or 'POC_' in str(header):
                        ws.cell(row=next_row, column=col_idx, value="")
                        continue
                    
                    # Try to find matching column in user's data
                    cell_value = None
                    
                    # Direct column name match
                    if header in user_row:
                        cell_value = user_row[header]
                    else:
                        # Try partial matches
                        for user_col, user_value in user_row.items():
                            if user_col and header:
                                if (str(header).lower() in str(user_col).lower() or 
                                    str(user_col).lower() in str(header).lower()):
                                    cell_value = user_value
                                    break
                    
                    # Set the cell value
                    cell = ws.cell(row=next_row, column=col_idx)
                    if cell_value is not None and pd.notna(cell_value):
                        cell.value = str(cell_value)
                    else:
                        cell.value = ""
                    
                    # Apply formatting
                    cell.font = Font()
                    cell.alignment = center_alignment
                    cell.border = border
                
                # Set status as "Closed" with proper formatting (no background color)
                if status_col:
                    status_cell = ws.cell(row=next_row, column=status_col, value="Closed")
                    status_cell.font = Font()
                    status_cell.alignment = center_alignment
                    status_cell.border = border
                    # No background color for Status column
                
                # Insert Old POC images if available - use Infrastructure VAPT approach
                if vuln_name in old_poc_images:
                    # Find Old POC columns by looking for merged "Old POC" header
                    old_poc_col_start = None
                    old_poc_col_end = None
                    for merged_range in ws.merged_cells.ranges:
                        if merged_range.min_row == 1 and merged_range.max_row == 1:
                            first_cell = ws.cell(row=1, column=merged_range.min_col)
                            if first_cell.value and str(first_cell.value).strip() == "Old POC":
                                old_poc_col_start = merged_range.min_col
                                old_poc_col_end = merged_range.max_col
                                break
                    
                    if old_poc_col_start:
                        # Define column order for Old POC image insertion: N, O, P, Q, R, S, M
                        old_image_columns = [
                            old_poc_col_start + 1,  # N
                            old_poc_col_start + 2,  # O
                            old_poc_col_start + 3,  # P
                            old_poc_col_start + 4,  # Q
                            old_poc_col_start + 5,  # R
                            old_poc_col_end,        # S
                            old_poc_col_start       # M
                        ]
                        
                        try:
                            from openpyxl.drawing.image import Image
                            from openpyxl.utils import get_column_letter
                            
                            image_paths = old_poc_images[vuln_name]
                            num_images = min(len(image_paths), 7)
                            
                            for img_idx, img_path in enumerate(image_paths[:num_images]):
                                if os.path.exists(img_path):
                                    img = Image(img_path)
                                    # Resize image (30x reduction like Infrastructure)
                                    img.width = img.width / 30
                                    img.height = img.height / 30
                                    
                                    col_idx = old_image_columns[img_idx]
                                    col_letter = get_column_letter(col_idx)
                                    cell_ref = f'{col_letter}{next_row}'
                                    
                                    ws.add_image(img, cell_ref)
                                    print(f"  ✅ Inserted Old POC image {img_idx+1} for '{vuln_name}' at {cell_ref}")
                        except Exception as e:
                            print(f"  ⚠️ Error inserting Old POC image: {e}")
                
                next_row += 1
                print(f"   ✅ Row added at {next_row-1}")
            else:
                print(f"   ❌ DEBUG: Vulnerability not found in user_row_data")
                print(f"      Looking for: {vuln_name[:100]}...")
                print(f"      Available keys (first 3):")
                for i, key in enumerate(list(user_row_data.keys())[:3]):
                    print(f"         {i+1}. {key[:100]}...")
        
        # Save the workbook
        print(f"\n📊 Saving workbook to: {excel_path}")
        wb.save(excel_path)
        wb.close()
        
        print(f"✅ Added {len(closed_vulnerabilities)} Closed vulnerability rows")
        return True
        
    except Exception as e:
        print(f"Error adding Closed vulnerability rows: {e}")
        import traceback
        traceback.print_exc()
        return False

@public_ip_vapt_followup_bp.route('/process_public_ip_vapt_follow_up_audit_report', methods=['POST'])
def process_public_ip_follow_up_audit_report():
    """Main endpoint to process Public IP VAPT Follow-Up Audit Report."""
    try:
        print("\n" + "="*80)
        print("🚀 Processing Public IP VAPT Follow-Up Audit Report")
        print("="*80)
        
        # Validate files
        if 'nmapFiles' not in request.files or 'nessusFiles' not in request.files:
            return jsonify({"error": "Both Nmap and Nessus files are required"}), 400
        
        nmap_file = request.files['nmapFiles']
        nessus_file = request.files['nessusFiles']
        evidence_file = request.files.get('evidenceFiles')
        user_excel_file = request.files.get('userExcelFile')  # Get user's old Excel file for comparison
        
        if nmap_file.filename == '' or nessus_file.filename == '':
            return jsonify({"error": "Please select all required files"}), 400
        
        if not user_excel_file or user_excel_file.filename == '':
            return jsonify({"error": "User's old Excel file is required for comparison"}), 400
        
        # Extract form data
        form_data = {
            'organization': request.form.get('organization'),
            'otherOrganization': request.form.get('otherOrganization'),
            'city': request.form.get('city'),
            'otherCity': request.form.get('otherCity'),
            'state': request.form.get('state'),
            'startDate': request.form.get('startDate'),
            'endDate': request.form.get('endDate'),
            'preparedByTitle': request.form.get('preparedByTitle'),
            'preparedByName': request.form.get('preparedByName'),
            'auditeeTitle': request.form.get('auditeeTitle'),
            'auditeeName': request.form.get('auditeeName'),
            'designation': request.form.get('designation'),
            'firstAuditReportId': request.form.get('firstAuditReportId'),
            'firstAuditReportDate': request.form.get('firstAuditReportDate')
        }
        
        # Handle "Other" selections
        if form_data['organization'] == 'other':
            form_data['organization'] = form_data['otherOrganization']
        if form_data['city'] == 'other':
            form_data['city'] = form_data['otherCity']
        
        # Extract email addresses
        bank_emails = request.form.getlist('bankEmail[]')
        form_data['bankEmails'] = [email for email in bank_emails if email.strip()]
        
        # Extract team member details
        team_names = request.form.getlist('teamName[]')
        team_designations = request.form.getlist('teamDesignation[]')
        team_emails = request.form.getlist('teamEmail[]')
        team_qualifications = request.form.getlist('teamQualification[]')
        
        # Extract team certified values
        team_certified = []
        for i in range(len(team_names)):
            certified_value = request.form.get(f'teamCertified[{i}]', 'no')
            team_certified.append(certified_value)
        
        form_data['teamNames'] = team_names
        form_data['teamDesignations'] = team_designations
        form_data['teamEmails'] = team_emails
        form_data['teamQualifications'] = team_qualifications
        form_data['teamCertified'] = team_certified
        
        # Process files
        print("📁 Processing Nmap files...")
        nmap_data = process_nmap_zip_public_ip(nmap_file)  # Returns dict: {ip: [(port, service), ...]}
        
        print("📁 Processing Nessus files...")
        nessus_dataframes = process_nessus_zip_public_ip(nessus_file)
        
        if not nessus_dataframes:
            return jsonify({"error": "No valid Nessus data found"}), 400
        
        # Extract vulnerabilities from user's old Excel file for comparison
        print("📁 Extracting vulnerabilities from user's old Excel file...")
        user_vulnerabilities = extract_vulnerability_names_from_excel(user_excel_file)
        user_row_data = extract_full_row_data_from_excel(user_excel_file)
        
        print(f"📊 Found {len(user_vulnerabilities)} vulnerabilities in user's old Excel file")
        print(f"📊 Extracted {len(user_row_data)} full rows from user's old Excel file")
        
        # DEBUG: Show sample vulnerabilities from user's Excel
        if user_vulnerabilities:
            print(f"   Sample vulnerabilities from user's Excel (first 3):")
            for i, vuln in enumerate(list(user_vulnerabilities)[:3]):
                print(f"      {i+1}. {vuln[:100]}...")
        
        # Store user row data globally for later use in worksheet creation
        global PUBLIC_IP_VAPT_USER_ROW_DATA, PUBLIC_IP_VAPT_USER_VULNERABILITIES
        PUBLIC_IP_VAPT_USER_ROW_DATA = user_row_data
        PUBLIC_IP_VAPT_USER_VULNERABILITIES = user_vulnerabilities
        
        # Create Excel workbook
        print("📊 Creating Excel workbook...")
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Compare vulnerabilities
        print("📊 Comparing vulnerabilities with previous audit...")
        combined_nessus = pd.concat(nessus_dataframes, ignore_index=True) if nessus_dataframes else pd.DataFrame()
        our_vulnerabilities = set()
        
        if not combined_nessus.empty and 'Name' in combined_nessus.columns:
            unique_names = combined_nessus['Name'].drop_duplicates().tolist()
            our_vulnerabilities = set(unique_names)
        
        comparison_result = compare_vulnerabilities(our_vulnerabilities, user_vulnerabilities)
        print(f"📊 Comparison result: {len(comparison_result)} vulnerabilities with status assigned")
        
        # Create Excel workbook
        print("📊 Creating Excel workbook...")
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create all worksheets
        if nmap_data:
            create_public_ip_nmap_worksheet(wb, nmap_data)  # Pass list of 6-element rows
        else:
            print("⚠️ No Nmap data found, skipping Nmap worksheet")
        create_public_ip_nessus_csv_worksheet(wb, nessus_dataframes)
        create_public_ip_scope_worksheet(wb, nessus_dataframes)
        create_public_ip_summary_worksheet(wb, nessus_dataframes)
        create_public_ip_Public_IP_VAPT_worksheet(wb, nessus_dataframes)  # Add Public_IP_VAPT worksheet
        create_public_ip_metadata_worksheet(wb, form_data, user_excel_file)
        
        # Save to BytesIO initially
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Save to temporary file for processing
        temp_excel_path = "temp_public_ip_vapt_followup.xlsx"
        with open(temp_excel_path, 'wb') as temp_file:
            temp_file.write(output.getvalue())
        
        # Load the workbook from file
        from openpyxl import load_workbook
        wb = load_workbook(temp_excel_path)
        
        # Get closed vulnerabilities
        # IMPORTANT: Extract vulnerabilities from the GENERATED Excel file (Public_IP_VAPT worksheet)
        # NOT from raw Nessus data, because we need to compare catalog-matched names
        print(f"\n🔍 DEBUG: Extracting vulnerabilities from generated Excel (Public_IP_VAPT worksheet)...")
        
        # Read the generated Excel file that was just created
        try:
            generated_df = pd.read_excel(temp_excel_path, sheet_name='Public_IP_VAPT')
            
            print(f"   Generated Excel shape: {generated_df.shape}")
            print(f"   Columns: {list(generated_df.columns)[:10]}")
            
            # IMPORTANT: Compare using 'Name of Vulnerability' column (Column C)
            # This matches what we extract from user's old Excel
            if 'Name of Vulnerability' in generated_df.columns:
                # Extract vulnerability names from the generated Excel
                our_vulnerabilities_list = generated_df['Name of Vulnerability'].dropna().astype(str).str.strip()
                our_vulnerabilities = set(our_vulnerabilities_list.tolist())
                
                print(f"   Found {len(our_vulnerabilities)} unique vulnerabilities from generated Excel (Name of Vulnerability column)")
                
                # DEBUG: Show sample
                if our_vulnerabilities:
                    print(f"   Sample vulnerabilities from generated Excel (first 3):")
                    for i, vuln in enumerate(list(our_vulnerabilities)[:3]):
                        print(f"      {i+1}. {vuln[:100]}...")
            else:
                print(f"   ⚠️ 'Name of Vulnerability' column not found in generated Excel")
                print(f"      Available columns: {list(generated_df.columns)[:10]}")
                our_vulnerabilities = set()
                
        except Exception as e:
            print(f"   ⚠️ Error reading generated Excel: {e}")
            import traceback
            traceback.print_exc()
            our_vulnerabilities = set()
        
        comparison_result = compare_vulnerabilities(our_vulnerabilities, user_vulnerabilities)
        
        # Find closed vulnerabilities
        closed_vulnerabilities = user_vulnerabilities - our_vulnerabilities
        
        # DEBUG: Print vulnerability counts
        print(f"\n🔍 DEBUG: Vulnerability comparison:")
        print(f"   Our vulnerabilities (from Nessus): {len(our_vulnerabilities)}")
        print(f"   User vulnerabilities (from old Excel): {len(user_vulnerabilities)}")
        print(f"   Closed vulnerabilities: {len(closed_vulnerabilities)}")
        
        # DEBUG: Print sample vulnerabilities from each source
        if our_vulnerabilities:
            print(f"   Sample 'Our vulnerabilities' (first 2):")
            for i, vuln in enumerate(list(our_vulnerabilities)[:2]):
                print(f"      {i+1}. {vuln[:100]}...")
        
        if user_vulnerabilities:
            print(f"   Sample 'User vulnerabilities' (first 2):")
            for i, vuln in enumerate(list(user_vulnerabilities)[:2]):
                print(f"      {i+1}. {vuln[:100]}...")
        
        if closed_vulnerabilities:
            print(f"   Closed vulnerability names (first 3):")
            for i, vuln in enumerate(list(closed_vulnerabilities)[:3]):
                print(f"      {i+1}. {vuln[:100]}...")
        
        # Generate timestamp for temporary folders (format: DDMMYYYYHHMMSS)
        from datetime import datetime
        current_ts = datetime.now().strftime('%d%m%Y%H%M%S')
        temp_poc2_website_folder = f"Temp_POC2_Website_{current_ts}"
        print(f"📁 Using timestamped folder: {temp_poc2_website_folder}")
        
        # Initialize temp_poc_images_folder variable for cleanup
        temp_poc_images_folder = None
        
        # Find all vulnerabilities that need Old POC images (Open and Closed)
        open_vulnerabilities = user_vulnerabilities.intersection(our_vulnerabilities)
        vulnerabilities_needing_old_poc = open_vulnerabilities.union(closed_vulnerabilities)
        
        # Extract Old POC images from user's Excel for Open and Closed vulnerabilities
        old_poc_images = {}
        if vulnerabilities_needing_old_poc:
            print(f"📊 Extracting Old POC images for {len(vulnerabilities_needing_old_poc)} vulnerabilities (Open: {len(open_vulnerabilities)}, Closed: {len(closed_vulnerabilities)})...")
            old_poc_images = extract_old_poc_images_from_user_excel(user_excel_file, vulnerabilities_needing_old_poc, temp_poc2_website_folder)
            print(f"📊 Extracted Old POC images for {len(old_poc_images)} vulnerabilities")
        
        # Insert Old POC images for Open vulnerabilities
        if old_poc_images and open_vulnerabilities:
            print(f"📊 Inserting Old POC images for {len([v for v in open_vulnerabilities if v in old_poc_images])} Open vulnerabilities...")
            insert_old_poc_images_for_open_vulnerabilities(temp_excel_path, old_poc_images, open_vulnerabilities)
            # Reload wb after saving
            wb = load_workbook(temp_excel_path)
        
        # Add closed vulnerabilities if any exist
        print(f"\n🔍 DEBUG: About to check closed vulnerabilities")
        print(f"   closed_vulnerabilities set: {closed_vulnerabilities}")
        print(f"   len(closed_vulnerabilities): {len(closed_vulnerabilities)}")
        
        if closed_vulnerabilities:
            print(f"📊 Adding {len(closed_vulnerabilities)} Closed vulnerabilities from user's Excel...")
            result = add_closed_vulnerabilities_to_excel(temp_excel_path, closed_vulnerabilities, user_row_data, None, old_poc_images)
            print(f"📊 add_closed_vulnerabilities_to_excel returned: {result}")
            # Function already saves the workbook
        else:
            print("⚠️ DEBUG: No closed vulnerabilities to add")
        
        # Apply Old POC column borders to all rows
        wb = load_workbook(temp_excel_path)
        ws = wb["Public_IP_VAPT"]
        
        # Find Old POC columns
        old_poc_col_start = None
        old_poc_col_end = None
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == 1 and merged_range.max_row == 1:
                first_cell = ws.cell(row=1, column=merged_range.min_col)
                if first_cell.value and str(first_cell.value).strip() == "Old POC":
                    old_poc_col_start = merged_range.min_col
                    old_poc_col_end = merged_range.max_col
                    break
        
        if old_poc_col_start and old_poc_col_end:
            apply_old_poc_column_borders(ws, old_poc_col_start, old_poc_col_end)
            wb.save(temp_excel_path)
        
        # Clean up Old POC folder after inserting images
        try:
            import shutil
            import glob
            
            # Remove the specific timestamped folder if it was created
            if temp_poc2_website_folder and os.path.exists(temp_poc2_website_folder):
                shutil.rmtree(temp_poc2_website_folder)
                print(f"✅ Successfully deleted {temp_poc2_website_folder} folder")
            
            # Also clean up any old Temp_POC2_Website* folders that might exist (excluding the one we just deleted)
            old_folders = glob.glob("Temp_POC2_Website*")
            for old_folder in old_folders:
                if os.path.isdir(old_folder) and (not temp_poc2_website_folder or old_folder != temp_poc2_website_folder):
                    try:
                        shutil.rmtree(old_folder)
                        print(f"✅ Successfully deleted old folder: {old_folder}")
                    except Exception as e:
                        print(f"⚠️ Could not delete old folder {old_folder}: {e}")
        except Exception as e:
            print(f"⚠️ Error cleaning up Temp_POC2_Website folders: {e}")
        
        # Read back the updated Excel
        with open(temp_excel_path, 'rb') as updated_file:
            output = BytesIO(updated_file.read())
        
        # Clean up temp file
        if os.path.exists(temp_excel_path):
            os.remove(temp_excel_path)
        
        # Insert POC images if evidence file is provided
        if evidence_file and evidence_file.filename != '':
            try:
                print("Processing POC images from evidence files...")
                
                # Save Excel file temporarily
                temp_excel_path = "temp_website_poc.xlsx"
                with open(temp_excel_path, 'wb') as temp_file:
                    temp_file.write(output.getvalue())
                
                # Generate timestamp for temporary POC images folder (format: DDMMYYYYHHMMSS)
                from datetime import datetime
                current_ts_poc = datetime.now().strftime('%d%m%Y%H%M%S')
                temp_poc_images_folder = f"temp_poc_images_{current_ts_poc}"
                print(f"📁 Using timestamped POC images folder: {temp_poc_images_folder}")
                
                # Extract POC images
                poc_mapping = extract_poc_images(evidence_file, temp_poc_images_folder)
                
                if poc_mapping:
                    print(f"Found {len(poc_mapping)} POC images to insert")
                    # Combine Nessus dataframes for vulnerability matching
                    combined_nessus = pd.concat(nessus_dataframes, ignore_index=True) if nessus_dataframes else pd.DataFrame()
                    
                    # Insert POC images using openpyxl
                    insert_poc_images_to_excel(temp_excel_path, poc_mapping, combined_nessus)
                    
                    # Read the updated file
                    with open(temp_excel_path, 'rb') as updated_file:
                        output = BytesIO(updated_file.read())
                
                # Clean up temporary file
                if os.path.exists(temp_excel_path):
                    os.remove(temp_excel_path)
                    
            except Exception as e:
                print(f"Error processing POC images: {e}")
                import traceback
                traceback.print_exc()
                # Continue with original file if POC processing fails
        
        # Sort vulnerabilities by Risk Factor and renumber (AT THE END - after all vulnerabilities added)
        try:
            print("📊 Sorting Public_IP_VAPT worksheet by Risk Factor...")
            
            # Save current state to file before sorting
            temp_excel_path_final = "temp_public_ip_vapt_final_sort.xlsx"
            with open(temp_excel_path_final, 'wb') as temp_file:
                temp_file.write(output.getvalue())
            
            # Sort the worksheet (after ALL vulnerabilities including unmatched are added)
            sort_and_renumber_public_ip_vapt_worksheet(temp_excel_path_final)
            
            # Update Meta_Data worksheet with follow-up vulnerability counts
            print("📊 Updating Meta_Data worksheet with follow-up vulnerability counts...")
            if 'Public_IP_VAPT' in load_workbook(temp_excel_path_final).sheetnames:
                # Re-create Meta_Data worksheet with updated counts
                wb_temp = load_workbook(temp_excel_path_final)
                if 'Meta_Data' in wb_temp.sheetnames:
                    wb_temp.remove(wb_temp['Meta_Data'])
                create_public_ip_metadata_worksheet(wb_temp, form_data, user_excel_file, temp_excel_path_final)
                
                # Reorder worksheets to the desired sequence: Meta_Data, Nmap Files, Nessus CSV Files, Scope, Summary, Public_IP_VAPT
                desired_order = ["Meta_Data", "Nmap Files", "Nessus CSV Files", "Scope", "Summary", "Public_IP_VAPT"]
                current_sheets = wb_temp.sheetnames.copy()

                # Create a list of sheets in the desired order (only include sheets that exist)
                ordered_sheets = []
                for sheet_name in desired_order:
                    if sheet_name in current_sheets:
                        ordered_sheets.append(sheet_name)

                # Add any remaining sheets that weren't in desired order (shouldn't happen, but just in case)
                for sheet_name in current_sheets:
                    if sheet_name not in ordered_sheets:
                        ordered_sheets.append(sheet_name)

                # Reorder by manipulating _sheets list directly
                # Build a new list of sheets in the desired order
                reordered_sheets = []
                for sheet_name in ordered_sheets:
                    if sheet_name in wb_temp.sheetnames:
                        reordered_sheets.append(wb_temp[sheet_name])

                # Add any remaining sheets that weren't in desired order (shouldn't happen, but just in case)
                for sheet_name in wb_temp.sheetnames:
                    if sheet_name not in ordered_sheets:
                        reordered_sheets.append(wb_temp[sheet_name])

                # Replace _sheets with reordered list
                wb_temp._sheets = reordered_sheets

                print(f"✅ Worksheets reordered: {wb_temp.sheetnames}")
                
                wb_temp.save(temp_excel_path_final)
                wb_temp.close()
                print("✅ Successfully updated Meta_Data worksheet with follow-up vulnerabilities")
            else:
                print("❌ Public_IP_VAPT worksheet not found")
            
            # Read the sorted file
            with open(temp_excel_path_final, 'rb') as sorted_file:
                output = BytesIO(sorted_file.read())
            
            # Clean up temp file
            if os.path.exists(temp_excel_path_final):
                os.remove(temp_excel_path_final)
                
        except Exception as e:
            print(f"⚠️ Error sorting worksheet: {e}")
            import traceback
            traceback.print_exc()
        
        # Generate filename
        filename = generate_public_ip_followup_filename(form_data['organization'], form_data['endDate'])
        
        print(f"✅ Report generated: {filename}")
        print("="*80 + "\n")
        
        response = send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
        # Clean up temp_poc_images folder after Excel generation
        try:
            import shutil
            import glob
            
            # Remove the specific timestamped folder if it was created
            if temp_poc_images_folder and os.path.exists(temp_poc_images_folder):
                shutil.rmtree(temp_poc_images_folder)
                print(f"✅ Successfully deleted {temp_poc_images_folder} folder")
            
            # Also clean up any old temp_poc_images* folders that might exist (excluding the one we just deleted)
            old_folders = glob.glob("temp_poc_images*")
            for old_folder in old_folders:
                if os.path.isdir(old_folder) and (not temp_poc_images_folder or old_folder != temp_poc_images_folder):
                    try:
                        shutil.rmtree(old_folder)
                        print(f"✅ Successfully deleted old folder: {old_folder}")
                    except Exception as e:
                        print(f"⚠️ Could not delete old folder {old_folder}: {e}")
        except Exception as e:
            print(f"Error deleting temp_poc_images folders: {e}")
        
        return response
        
    except Exception as e:
        print(f"❌ Error processing report: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error generating report: {str(e)}"}), 500
