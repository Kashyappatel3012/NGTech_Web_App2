"""
Extra Work Module - Handles extra work submissions
"""
from flask import Blueprint, request, jsonify
from flask_login import current_user
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

extra_work_bp = Blueprint('extra_work', __name__)

def get_month_name(month):
    """Convert month number to month name"""
    month_names = {
        1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr',
        5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Aug',
        9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
    }
    return month_names.get(month, 'Jan')

def generate_filename():
    """Generate filename based on current month"""
    now = datetime.now()
    month = get_month_name(now.month)
    year = str(now.year)
    return f"{month}_{year}.xlsx"

def get_file_path():
    """Get the full path for the extra work file"""
    filename = generate_filename()
    base_dir = os.path.join('static', 'Activity_Tracker', 'Extra_Work')
    os.makedirs(base_dir, exist_ok=True)
    return os.path.join(base_dir, filename)

def format_date_for_display(date_str):
    """Convert date string to readable format"""
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        return date_obj.strftime('%d-%m-%Y')
    except:
        return date_str

def create_extra_work_excel(employee_name, employee_team, date_str, time, task, task_description, client_name, concerned_person):
    """Create or update the extra work Excel file"""
    file_path = get_file_path()
    
    # Check if file exists
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Extra Work"
        
        # Set headers
        headers = ['Employee Name', 'Employee Team', 'Date', 'Time', 'Task', 'Task Description', 'Client', 'Concerned Person', 'Timestamp']
        
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col).value = header
        
        # Format headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col in range(1, 10):  # A to I
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 80
        ws.column_dimensions['G'].width = 50
        ws.column_dimensions['H'].width = 40
        ws.column_dimensions['I'].width = 35
    
    # Get current timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Format date
    formatted_date = format_date_for_display(date_str)
    
    # Prepare row data
    next_row = ws.max_row + 1
    
    ws.cell(row=next_row, column=1).value = employee_name
    ws.cell(row=next_row, column=2).value = employee_team
    ws.cell(row=next_row, column=3).value = formatted_date
    ws.cell(row=next_row, column=4).value = time
    ws.cell(row=next_row, column=5).value = task
    ws.cell(row=next_row, column=6).value = task_description
    ws.cell(row=next_row, column=7).value = client_name
    ws.cell(row=next_row, column=8).value = concerned_person
    ws.cell(row=next_row, column=9).value = timestamp
    
    # Apply border and alignment
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, 10):
        cell = ws.cell(row=next_row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
    
    # Set row height
    ws.row_dimensions[next_row].height = 25
    
    # Save workbook and ensure it's properly closed
    try:
        wb.save(file_path)
        wb.close()
        print(f"✅ Extra work saved to: {file_path}")
        return True
    except Exception as e:
        # Ensure workbook is closed even if save fails
        try:
            wb.close()
        except:
            pass
        print(f"❌ Error saving extra work: {e}")
        raise

@extra_work_bp.route('/submit_extra_work', methods=['POST'])
def submit_extra_work():
    """Handle extra work submission"""
    try:
        print("\n" + "="*80)
        print("📋 Processing Extra Work Submission")
        print("="*80)
        
        # Get form data
        date_str = request.form.get('date', '')
        employee_name = request.form.get('employeeName', '').strip()
        employee_team = request.form.get('employeeTeam', '').strip()
        time = request.form.get('time', '').strip()
        task = request.form.get('task', '').strip()
        task_description = request.form.get('taskDescription', '').strip()
        client_name = request.form.get('clientName', '').strip()
        concerned_person = request.form.get('concernedPerson', '').strip()
        
        # Validate data
        if not employee_name or not employee_team:
            return jsonify({'success': False, 'error': 'Employee information is missing'}), 400
        
        if not time or not task or not task_description or not client_name or not concerned_person:
            return jsonify({'success': False, 'error': 'All fields are required'}), 400
        
        print(f"📅 Date: {date_str}")
        print(f"👤 Employee: {employee_name}")
        print(f"👥 Team: {employee_team}")
        print(f"⏰ Time: {time}")
        print(f"📝 Task: {task}")
        
        # Create or update Excel file
        create_extra_work_excel(employee_name, employee_team, date_str, time, task, task_description, client_name, concerned_person)
        
        print("✅ Extra work submitted successfully!")
        
        return jsonify({
            'success': True,
            'message': 'Extra work submitted successfully!',
            'filename': generate_filename()
        })
        
    except Exception as e:
        print(f"❌ Error submitting extra work: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }), 500

