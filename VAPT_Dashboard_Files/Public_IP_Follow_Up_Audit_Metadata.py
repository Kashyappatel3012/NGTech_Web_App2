from flask import Blueprint, request, jsonify, send_file
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils.units import pixels_to_EMU
from datetime import datetime
import shutil

# Create blueprint for Public IP Follow Up Audit MetaData
public_ip_follow_up_audit_metadata_bp = Blueprint('public_ip_follow_up_audit_metadata', __name__)

def create_public_ip_follow_up_audit_metadata_excel(form_data):
    """Create Public IP Follow Up Audit MetaData Excel file"""
    try:
        print(f"🔍 DEBUG: Starting Excel creation...")
        
        # Create a fresh workbook instead of loading template
        # This avoids the hang issue with load_workbook()
        print(f"🔍 DEBUG: Creating new workbook from scratch...")
        wb = Workbook()
        print(f"🔍 DEBUG: Workbook created")
        
        # Remove the default sheet if it exists
        if wb.active and wb.active.title == 'Sheet':
            print(f"🔍 DEBUG: Removing default sheet...")
            wb.remove(wb.active)
            print(f"🔍 DEBUG: Default sheet removed")
        
        # Create new worksheet "1. Auditor Details"
        print(f"🔍 DEBUG: Creating '1. Auditor Details' worksheet...")
        ws = wb.create_sheet("1. Auditor Details", 0)
        print(f"🔍 DEBUG: Worksheet created at index 0")
        
        print(f"📝 Creating '1. Auditor Details' worksheet...")
        
        # Set column widths
        print(f"🔍 DEBUG: Setting column widths...")
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        print(f"🔍 DEBUG: Column widths set")
        
        # Define colors
        light_sky_blue = 'FFBDD7EE'
        creamy_peach = 'FFF4B084'
        olive = 'FFA9D08E'  # Parrot/Light green
        
        # Define border
        thin_border = Border(
            left=Side(style='thin', color='FF000000'),  
            right=Side(style='thin', color='FF000000'),
            top=Side(style='thin', color='FF000000'),
            bottom=Side(style='thin', color='FF000000')
        )
        
        # Row 1: Merge A1:E1 - "Details of Auditing Organisation"
        print(f"🔍 DEBUG: Creating Row 1...")
        ws.merge_cells('A1:C1')
        cell_a1 = ws.cell(row=1, column=1)
        cell_a1.value = "Details of Auditing Organisation"
        cell_a1.font = Font(name='Arial', size=15, bold=True, color='FF000000')
        cell_a1.fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        cell_a1.alignment = Alignment(horizontal='center', vertical='center')
        cell_a1.border = thin_border
        ws.row_dimensions[1].height = 35
        print(f"🔍 DEBUG: Row 1 created")
        
        # Row 2: Headers
        print(f"🔍 DEBUG: Creating Row 2 headers...")
        headers = {
            'A2': "S.no.",
            'B2': "Question",
            'C2': "Response"
        }
        
        for cell_ref, value in headers.items():
            cell = ws[cell_ref]
            cell.value = value
            cell.font = Font(name='Arial', size=15, bold=True, color='FF000000')
            cell.fill = PatternFill(start_color=creamy_peach, end_color=creamy_peach, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        ws.row_dimensions[2].height = 28
        print(f"🔍 DEBUG: Row 2 created")
        
        # Format dates from user input
        print(f"🔍 DEBUG: Formatting dates...")
        start_date = form_data.get('startAuditDate', '')
        end_date = form_data.get('endAuditDate', '')
        print(f"🔍 DEBUG: Start date: {start_date}, End date: {end_date}")
        
        # Convert dates to DD/MM/YYYY format
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            start_date_formatted = start_date_obj.strftime('%d/%m/%Y')
        except Exception as e:
            print(f"🔍 DEBUG: Error formatting start date: {e}")
            start_date_formatted = start_date
        
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            end_date_formatted = end_date_obj.strftime('%d/%m/%Y')
        except Exception as e:
            print(f"🔍 DEBUG: Error formatting end date: {e}")
            end_date_formatted = end_date
        
        print(f"🔍 DEBUG: Dates formatted: {start_date_formatted}, {end_date_formatted}")
        
        # Data rows - Row 3 (A3, B3, C3)
        print(f"🔍 DEBUG: Creating Row 3...")
        ws['A3'].value = "1"
        ws['A3'].font = Font(name='Arial', size=15, color='FF000000')
        ws['A3'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A3'].border = thin_border
        
        ws['B3'].value = "Name of Organisation"
        ws['B3'].font = Font(name='Arial', size=15, color='FF000000')
        ws['B3'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws['B3'].alignment = Alignment(horizontal='left', vertical='center')
        ws['B3'].border = thin_border
        
        ws['C3'].value = "NG TechAssurance Private Limited"
        ws['C3'].font = Font(name='Arial', size=15, color='FF000000')
        ws['C3'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C3'].border = thin_border
        
        ws.row_dimensions[3].height = 32
        print(f"🔍 DEBUG: Row 3 created")
        
        # Row 4 (A4, B4, C4)
        print(f"🔍 DEBUG: Creating Row 4...")
        ws['A4'].value = "2"
        ws['A4'].font = Font(name='Arial', size=15, color='FF000000')
        ws['A4'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A4'].border = thin_border
        
        ws['B4'].value = "Data Validated by"
        ws['B4'].font = Font(name='Arial', size=15, color='FF000000')
        ws['B4'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws['B4'].alignment = Alignment(horizontal='left', vertical='center')
        ws['B4'].border = thin_border
        
        ws['C4'].value = "Response"
        ws['C4'].font = Font(name='Arial', size=15, color='FF000000')
        ws['C4'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws['C4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C4'].border = thin_border
        
        ws.row_dimensions[4].height = 32
        print(f"🔍 DEBUG: Row 4 created")
        
        # Row 5 (A5, B5, C5)
        print(f"🔍 DEBUG: Creating Row 5...")
        ws['A5'].value = "3"
        ws['A5'].font = Font(name='Arial', size=15, color='FF000000')
        ws['A5'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A5'].border = thin_border
        
        ws['B5'].value = "Designation"
        ws['B5'].font = Font(name='Arial', size=15, color='FF000000')
        ws['B5'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws['B5'].alignment = Alignment(horizontal='left', vertical='center')
        ws['B5'].border = thin_border
        
        ws['C5'].value = "CEO/Director"
        ws['C5'].font = Font(name='Arial', size=15, color='FF000000')
        ws['C5'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws['C5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C5'].border = thin_border
        
        ws.row_dimensions[5].height = 32
        print(f"🔍 DEBUG: Row 5 created")
        
        # Row 6 (A6, B6, C6)
        print(f"🔍 DEBUG: Creating Row 6...")
        ws['A6'].value = "4"
        ws['A6'].font = Font(name='Arial', size=15, color='FF000000')
        ws['A6'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A6'].border = thin_border
        
        ws['B6'].value = "Contact Details (Email)"
        ws['B6'].font = Font(name='Arial', size=15, color='FF000000')
        ws['B6'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws['B6'].alignment = Alignment(horizontal='left', vertical='center')
        ws['B6'].border = thin_border
        
        ws['C6'].value = "admin@ngtech.co.in"
        ws['C6'].font = Font(name='Arial', size=15, color='FF000000')
        ws['C6'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws['C6'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C6'].border = thin_border
        
        ws.row_dimensions[6].height = 32
        print(f"🔍 DEBUG: Row 6 created")
        
        # Row 7 (A7, B7, C7)
        print(f"🔍 DEBUG: Creating Row 7...")
        ws['A7'].value = "5"
        ws['A7'].font = Font(name='Arial', size=15, color='FF000000')
        ws['A7'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A7'].border = thin_border
        
        ws['B7'].value = "Contact Details (Mobile)"
        ws['B7'].font = Font(name='Arial', size=15, color='FF000000')
        ws['B7'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws['B7'].alignment = Alignment(horizontal='left', vertical='center')
        ws['B7'].border = thin_border
        
        ws['C7'].value = "9979580410"
        ws['C7'].font = Font(name='Arial', size=15, color='FF000000')
        ws['C7'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws['C7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C7'].border = thin_border
        
        ws.row_dimensions[7].height = 32
        print(f"🔍 DEBUG: Row 7 created")
        
        # Row 8 (A8, B8, C8)
        print(f"🔍 DEBUG: Creating Row 8...")
        ws['A8'].value = "6"
        ws['A8'].font = Font(name='Arial', size=15, color='FF000000')
        ws['A8'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A8'].border = thin_border
        
        ws['B8'].value = "Audit Start Date"
        ws['B8'].font = Font(name='Arial', size=15, color='FF000000')
        ws['B8'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws['B8'].alignment = Alignment(horizontal='left', vertical='center')
        ws['B8'].border = thin_border
        
        ws['C8'].value = start_date_formatted
        ws['C8'].font = Font(name='Arial', size=15, color='FF000000')
        ws['C8'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws['C8'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C8'].border = thin_border
        
        ws.row_dimensions[8].height = 32
        print(f"🔍 DEBUG: Row 8 created")
        
        # Row 9 (A9, B9, C9)
        print(f"🔍 DEBUG: Creating Row 9...")
        ws['A9'].value = "7"
        ws['A9'].font = Font(name='Arial', size=15, color='FF000000')
        ws['A9'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws['A9'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A9'].border = thin_border
        
        ws['B9'].value = "Audit End Date"
        ws['B9'].font = Font(name='Arial', size=15, color='FF000000')
        ws['B9'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws['B9'].alignment = Alignment(horizontal='left', vertical='center')
        ws['B9'].border = thin_border
        
        ws['C9'].value = end_date_formatted
        ws['C9'].font = Font(name='Arial', size=15, color='FF000000')
        ws['C9'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws['C9'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C9'].border = thin_border
        
        ws.row_dimensions[9].height = 32
        print(f"🔍 DEBUG: Row 9 created")
        
        print(f"  ✅ Created '1. Auditor Details' worksheet")
        
        # Create "0. Instructions" worksheet
        print(f"🔍 DEBUG: Creating '0. Instructions' worksheet...")
        ws_instructions = wb.create_sheet("0. Instructions", 0)
        
        # Set column widths (A to M = 13)
        for col in range(1, 14):  # A to M (1 to 13)
            ws_instructions.column_dimensions[get_column_letter(col)].width = 13
        
        # Set row heights for rows 4-13 to 15
        for row in range(4, 14):  # Rows 4 to 13
            ws_instructions.row_dimensions[row].height = 35
        
        # Define colors
        creamy_peach = 'FFF4B084'
        light_sky_blue = 'FFBDD7EE'
        olive = 'FFA9D08E'
        dark_blue = 'FF00008B'
        red_color = 'FFFF0000'
        
        # Define border for A1-M13 range
        thin_border = Border(
            left=Side(style='thin', color='FF000000'),
            right=Side(style='thin', color='FF000000'),
            top=Side(style='thin', color='FF000000'),
            bottom=Side(style='thin', color='FF000000')
        )
        
        # Merge A1:M3 for header with logo
        ws_instructions.merge_cells('A1:M3')
        cell_a1_inst = ws_instructions['A1']
        
        # Create rich text with mixed formatting (dark blue for main text, red for TLP: AMBER)
        # InlineFont parameters: rFont (font name), sz (size), b (bold), color
        # Note: Since we can't have different alignments in same cell, we'll center "TLP: AMBER" using spaces
        rich_text = CellRichText(
            TextBlock(InlineFont(rFont='Arial', sz=15, b=True, color=dark_blue), 
                     "\nFramework for Assessing Vulnerabilities and Audit Landscape - Instructions for Auditing Organisations\n\n                                                                                                                   "),
            TextBlock(InlineFont(rFont='Arial', sz=15, b=True, color=red_color), 
                     "TLP: AMBER")
        )
        cell_a1_inst.value = rich_text
        cell_a1_inst.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        # Add light sky blue background to A1
        cell_a1_inst.fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws_instructions.row_dimensions[1].height = 60  # Increased height for 3 merged rows
        
        # Add CERT-In logo to A1:M3 (top-right position with padding)
        try:
            logo_path = os.path.join('static', 'Images', 'Certin_Logo.jpg')
            if os.path.exists(logo_path):
                logo = Image(logo_path)
                # Make logo bigger to fit the cell height
                logo.height = 74  # Slightly smaller than cell height (60px) for padding
                logo.width = 217  # Maintain aspect ratio (approximately 3:1 ratio)
                
                # Add image with custom anchor for padding (5px from right and 5px from top)
                # Create anchor marker for top-left position (column K, row 1) with offsets
                from_marker = AnchorMarker(col=10, colOff=pixels_to_EMU(14), row=0, rowOff=pixels_to_EMU(14))  # K1 with 5px top and left offset
                # Calculate the to_marker based on image size
                to_marker = AnchorMarker(col=13, colOff=pixels_to_EMU(-10), row=2, rowOff=0)  # M3 with 5px right padding
                
                # Create two-cell anchor
                logo.anchor = TwoCellAnchor(editAs='oneCell', _from=from_marker, to=to_marker)
                
                ws_instructions.add_image(logo)
                print(f"🔍 DEBUG: Added CERT-In logo at top-right corner with 5px padding")
            else:
                print(f"⚠️ Warning: Logo not found at {logo_path}")
        except Exception as e:
            print(f"⚠️ Warning: Could not add logo: {e}")
        
        # Apply border to A1
        ws_instructions['A1'].border = thin_border
        
        # Row 4: Headers
        ws_instructions['A4'].value = "S.No."
        ws_instructions['A4'].font = Font(name='Arial', size=12, bold=True, color='FF000000')
        ws_instructions['A4'].fill = PatternFill(start_color=creamy_peach, end_color=creamy_peach, fill_type='solid')
        ws_instructions['A4'].alignment = Alignment(horizontal='center', vertical='center')
        ws_instructions['A4'].border = thin_border
        
        ws_instructions.merge_cells('B4:M4')
        ws_instructions['B4'].value = "Instructions"
        ws_instructions['B4'].font = Font(name='Arial', size=12, bold=True, color='FF000000')
        ws_instructions['B4'].fill = PatternFill(start_color=creamy_peach, end_color=creamy_peach, fill_type='solid')
        ws_instructions['B4'].alignment = Alignment(horizontal='center', vertical='center')
        # Apply border to all cells in merged range B4:M4
        for col in range(2, 14):  # B to M (2 to 13)
            ws_instructions.cell(row=4, column=col).border = thin_border
        
        # Row 5
        ws_instructions['A5'].value = "1"
        ws_instructions['A5'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['A5'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws_instructions['A5'].alignment = Alignment(horizontal='center', vertical='center')
        ws_instructions['A5'].border = thin_border
        
        ws_instructions.merge_cells('B5:M5')
        ws_instructions['B5'].value = "This document is intended for CERT-In empanelled organisations to submit audit data after completion of audit."
        ws_instructions['B5'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['B5'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws_instructions['B5'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        # Apply border to all cells in merged range B5:M5
        for col in range(2, 14):
            ws_instructions.cell(row=5, column=col).border = thin_border
        
        # Row 6
        ws_instructions['A6'].value = "2"
        ws_instructions['A6'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['A6'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws_instructions['A6'].alignment = Alignment(horizontal='center', vertical='center')
        ws_instructions['A6'].border = thin_border
        
        ws_instructions.merge_cells('B6:M6')
        ws_instructions['B6'].value = "Organisations to provide data for both first audit and follow up audit separately in the prescribed format within 5 days of audit completion."
        ws_instructions['B6'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['B6'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws_instructions['B6'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in range(2, 14):
            ws_instructions.cell(row=6, column=col).border = thin_border
        
        # Row 7
        ws_instructions['A7'].value = "3"
        ws_instructions['A7'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['A7'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws_instructions['A7'].alignment = Alignment(horizontal='center', vertical='center')
        ws_instructions['A7'].border = thin_border
        
        ws_instructions.merge_cells('B7:M7')
        ws_instructions['B7'].value = "Data needs to be provided ONLY in the format provided in sheet no. 1 to Sheet no. 4."
        ws_instructions['B7'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['B7'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws_instructions['B7'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in range(2, 14):
            ws_instructions.cell(row=7, column=col).border = thin_border
        
        # Row 8
        ws_instructions['A8'].value = "4"
        ws_instructions['A8'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['A8'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws_instructions['A8'].alignment = Alignment(horizontal='center', vertical='center')
        ws_instructions['A8'].border = thin_border
        
        ws_instructions.merge_cells('B8:M8')
        ws_instructions['B8'].value = "Don't change or modify any of the field as data needs to be processed by automated systems in a structured way."
        ws_instructions['B8'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['B8'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws_instructions['B8'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in range(2, 14):
            ws_instructions.cell(row=8, column=col).border = thin_border
        
        # Row 9
        ws_instructions['A9'].value = "5"
        ws_instructions['A9'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['A9'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws_instructions['A9'].alignment = Alignment(horizontal='center', vertical='center')
        ws_instructions['A9'].border = thin_border
        
        ws_instructions.merge_cells('B9:M9')
        ws_instructions['B9'].value = "Filled excel sheet needs to be submitted to auditdata@cert-in.org.in"
        ws_instructions['B9'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['B9'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws_instructions['B9'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in range(2, 14):
            ws_instructions.cell(row=9, column=col).border = thin_border
        
        # Row 10
        ws_instructions['A10'].value = "6"
        ws_instructions['A10'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['A10'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws_instructions['A10'].alignment = Alignment(horizontal='center', vertical='center')
        ws_instructions['A10'].border = thin_border
        
        ws_instructions.merge_cells('B10:M10')
        ws_instructions['B10'].value = "Executive summaries and audit reports of the completed audits are to be provided along with the data."
        ws_instructions['B10'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['B10'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws_instructions['B10'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in range(2, 14):
            ws_instructions.cell(row=10, column=col).border = thin_border
        
        # Row 11
        ws_instructions['A11'].value = "7"
        ws_instructions['A11'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['A11'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws_instructions['A11'].alignment = Alignment(horizontal='center', vertical='center')
        ws_instructions['A11'].border = thin_border
        
        ws_instructions.merge_cells('B11:M11')
        ws_instructions['B11'].value = "Senior Management of organisations must ensure accuracy of data."
        ws_instructions['B11'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['B11'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws_instructions['B11'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in range(2, 14):
            ws_instructions.cell(row=11, column=col).border = thin_border
        
        # Row 12
        ws_instructions['A12'].value = "8"
        ws_instructions['A12'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['A12'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws_instructions['A12'].alignment = Alignment(horizontal='center', vertical='center')
        ws_instructions['A12'].border = thin_border
        
        ws_instructions.merge_cells('B12:M12')
        ws_instructions['B12'].value = "As per empanelment terms and condition, it is mandatory for the auditor to share the consolidated report related to information Security audits with CERT-In as per the format provided by CERT-In. Non-compliance of the same may lead to termination of empanelment"
        ws_instructions['B12'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['B12'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws_instructions['B12'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in range(2, 14):
            ws_instructions.cell(row=12, column=col).border = thin_border
        
        # Row 13
        ws_instructions['A13'].value = "9"
        ws_instructions['A13'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['A13'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws_instructions['A13'].alignment = Alignment(horizontal='center', vertical='center')
        ws_instructions['A13'].border = thin_border
        
        ws_instructions.merge_cells('B13:M13')
        ws_instructions['B13'].value = "This document is marked as TLP:AMBER. Document should not be shared outside the recipient organisation and also it should not be shared or exchanged over public channels."
        ws_instructions['B13'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['B13'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws_instructions['B13'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in range(2, 14):
            ws_instructions.cell(row=13, column=col).border = thin_border
        
        # Row 15: Copyright notice
        ws_instructions.merge_cells('A15:I15')
        ws_instructions['A15'].value = "© Indian Computer Emergency Response Team (CERT-In) Version 2.2 March 2024"
        ws_instructions['A15'].font = Font(name='Arial', size=12, color='FF000000')
        ws_instructions['A15'].alignment = Alignment(horizontal='left', vertical='center')
        
        print(f"  ✅ Created '0. Instructions' worksheet")
        
        # Create "2. Audits Completed" worksheet
        print(f"🔍 DEBUG: Creating '2. Audits Completed' worksheet...")
        ws_audits = wb.create_sheet("2. Audits Completed")
        
        # Set all column widths to 40
        for col in range(1, 25):  # A to X (1 to 24)
            ws_audits.column_dimensions[get_column_letter(col)].width = 40
        
        # Set row heights
        ws_audits.row_dimensions[1].height = 40  # Row 1
        ws_audits.row_dimensions[2].height = 35  # Row 2
        
        # Define additional colors for this worksheet
        light_mustard_yellow = 'FFFFFFDB58'
        very_creamy_peach = 'FFF8CBAD'  # New color for audits completed
        sky_blue = 'FFBDD7EE'  # Match light_sky_blue
        olive = 'FFA9D08E'  # For consistency
        light_sky_blue = 'FFBDD7EE'
        creamy_peach = 'FFF8CBAD'
        
        # Row 1: Merge A1:X1
        ws_audits.merge_cells('A1:X1')
        ws_audits['A1'].value = "Details of completed audits"
        ws_audits['A1'].font = Font(name='Arial', size=14, bold=True, color='FF000000')
        ws_audits['A1'].fill = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type='solid')
        ws_audits['A1'].alignment = Alignment(horizontal='center', vertical='center')
        # Apply border to all cells in merged range A1:X1
        for col in range(1, 25):
            ws_audits.cell(row=1, column=col).border = thin_border
        
        # Row 2: Multiple merged sections
        # A2:K2 - Empty with sky blue
        ws_audits.merge_cells('A2:K2')
        ws_audits['A2'].value = ""
        ws_audits['A2'].font = Font(name='Arial', size=14, bold=True, color='FF000000')
        ws_audits['A2'].fill = PatternFill(start_color=sky_blue, end_color=sky_blue, fill_type='solid')
        ws_audits['A2'].alignment = Alignment(horizontal='center', vertical='center')
        for col in range(1, 12):  # A to K
            ws_audits.cell(row=2, column=col).border = thin_border
        
        # L2:P2 - Questionnaire text with light mustard yellow
        ws_audits.merge_cells('L2:P2')
        ws_audits['L2'].value = "Questionnaire related to audit criteria and checklist(s) of auditing organisation w.r.t. regulatory, statutory and domestic requirements"
        ws_audits['L2'].font = Font(name='Arial', size=14, bold=True, color='FF000000')
        ws_audits['L2'].fill = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type='solid')
        ws_audits['L2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col in range(12, 17):  # L to P
            ws_audits.cell(row=2, column=col).border = thin_border
        
        # Q2:U2 - Patching Status with light green
        ws_audits.merge_cells('Q2:U2')
        ws_audits['Q2'].value = "Patching Status"
        ws_audits['Q2'].font = Font(name='Arial', size=14, bold=True, color='FF000000')
        ws_audits['Q2'].fill = PatternFill(start_color=olive, end_color=olive, fill_type='solid')
        ws_audits['Q2'].alignment = Alignment(horizontal='center', vertical='center')
        for col in range(17, 22):  # Q to U
            ws_audits.cell(row=2, column=col).border = thin_border
        
        # V2:X2 - Empty with sky blue
        ws_audits.merge_cells('V2:X2')
        ws_audits['V2'].value = ""
        ws_audits['V2'].font = Font(name='Arial', size=14, bold=True, color='FF000000')
        ws_audits['V2'].fill = PatternFill(start_color=sky_blue, end_color=sky_blue, fill_type='solid')
        ws_audits['V2'].alignment = Alignment(horizontal='center', vertical='center')
        for col in range(22, 25):  # V to X
            ws_audits.cell(row=2, column=col).border = thin_border
        
        # Row 3: Headers
        headers_row3 = [
            "SNO.",
            "Name of Auditee organization",
            "Category of Organization",
            "Sector of Organization",
            "Sub Sector",
            "Type of audit",
            "Please Specify type of Audit(if selected Any other)",
            "Details of web site/name of application/no. of servers, perimeter devices, other infrastructure",
            "Reason for conducting Audit",
            "Standards Used for conducting Audit",
            "Challenges in conducting Audits",
            "Compliance to financial sector regulatory bodies guidelines/circulars/directions like from RBI, IRDAI, SEBI, NPCI.(Yes/No/Not Applicable)",
            "Compliance to regulatory bodies guidelines/circulars/directions like from UIDAI, CCA, DoT, TRAI.",
            "Compliance to CERT-In directions under sub-section (6) of section 70B of the IT Act, 2000 dated 28 April, 2022.",
            "Compliance to CEA (Cyber Security in Power Sector) guidelines 2021.",
            "Any other domestic regulation, directions, guidelines (Please Specify)",
            "Type of Audit Report(First Audit Report/Follow up Report)",
            "No. of vulnerabilities/ security issues reported during FIRST Audit",
            "No. of days taken by Client to patch the vulnerabilities after submission of Audit report ( In case follow-up audit was conducted)",
            "Durations (in days) between end of FIRST audit and start of FOLLOW-UP/SECOND audit",
            "No. of open issues after follow-up audits",
            "Geographical detail of Client (State/UT)",
            "Date of Audit Completion",
            "Date of last audit of same infrastructure"
        ]
        
        for idx, header in enumerate(headers_row3, start=1):
            cell = ws_audits.cell(row=3, column=idx)
            cell.value = header
            cell.font = Font(name='Arial', size=12, bold=True, color='FF000000')
            cell.fill = PatternFill(start_color=very_creamy_peach, end_color=very_creamy_peach, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Row 4: Data from form
        # A4 = "1"
        ws_audits['A4'].value = "1"
        ws_audits['A4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['A4'].alignment = Alignment(horizontal='left', vertical='center')
        ws_audits['A4'].border = thin_border
        
        # B4 = Name of Auditee Organization (check for dropdown and "Other" field)
        auditee_org = form_data.get('nameOfAuditeeOrg', '')
        if auditee_org == 'Other':
            auditee_org = form_data.get('nameOfAuditeeOrgOther', '')
        ws_audits['B4'].value = auditee_org
        ws_audits['B4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['B4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['B4'].border = thin_border
        
        # C4 = Category of Organization
        ws_audits['C4'].value = form_data.get('categoryOfOrg', '')
        ws_audits['C4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['C4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['C4'].border = thin_border
        
        # D4 = Sector of Organization
        ws_audits['D4'].value = form_data.get('sectorOfOrg', '')
        ws_audits['D4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['D4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['D4'].border = thin_border
        
        # E4 = Sub Sector
        ws_audits['E4'].value = form_data.get('subSector', '')
        ws_audits['E4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['E4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['E4'].border = thin_border
        
        # F4 = Type of Audit
        type_of_audit = form_data.get('typeOfAudit', '')
        ws_audits['F4'].value = type_of_audit
        ws_audits['F4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['F4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['F4'].border = thin_border
        
        # G4 = Type of Audit Other Specify (NA if not "Other")
        if type_of_audit and 'other' in type_of_audit.lower():
            ws_audits['G4'].value = form_data.get('typeOfAuditOther', '')
        else:
            ws_audits['G4'].value = "NA"
        ws_audits['G4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['G4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['G4'].border = thin_border
        
        # H4 = Details of website/application/servers
        ws_audits['H4'].value = form_data.get('detailsOfWebsite', '')
        ws_audits['H4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['H4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['H4'].border = thin_border
        
        # I4 = Reason for conducting Audit
        ws_audits['I4'].value = form_data.get('reasonForAudit', '')
        ws_audits['I4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['I4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['I4'].border = thin_border
        
        # J4 = Standards Used
        ws_audits['J4'].value = form_data.get('standardsUsed', '')
        ws_audits['J4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['J4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['J4'].border = thin_border
        
        # K4 = Challenges
        ws_audits['K4'].value = form_data.get('challenges', '')
        ws_audits['K4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['K4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['K4'].border = thin_border
        
        # L4 = Compliance to financial sector
        ws_audits['L4'].value = form_data.get('complianceFinancial', '')
        ws_audits['L4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['L4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['L4'].border = thin_border
        
        # M4 = Compliance to regulatory bodies
        ws_audits['M4'].value = form_data.get('complianceRegulatory', '')
        ws_audits['M4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['M4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['M4'].border = thin_border
        
        # N4 = Compliance to CERT-In
        ws_audits['N4'].value = form_data.get('complianceCertIn', '')
        ws_audits['N4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['N4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['N4'].border = thin_border
        
        # O4 = Compliance to CEA
        ws_audits['O4'].value = form_data.get('complianceCEA', '')
        ws_audits['O4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['O4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['O4'].border = thin_border
        
        # P4 = Any other domestic regulation
        ws_audits['P4'].value = form_data.get('complianceOther', '')
        ws_audits['P4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['P4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['P4'].border = thin_border
        
        # Q4 = Type of Audit Report
        ws_audits['Q4'].value = form_data.get('typeOfAuditReport', '')
        ws_audits['Q4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['Q4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['Q4'].border = thin_border
        
        # R4 = No. of vulnerabilities
        ws_audits['R4'].value = form_data.get('numVulnerabilities', '')
        ws_audits['R4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['R4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['R4'].border = thin_border
        
        # S4 = No. of days taken to patch
        ws_audits['S4'].value = form_data.get('numDaysToPatch', '')
        ws_audits['S4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['S4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['S4'].border = thin_border
        
        # T4 = Duration between audits
        ws_audits['T4'].value = form_data.get('durationBetweenAudits', '')
        ws_audits['T4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['T4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['T4'].border = thin_border
        
        # U4 = No. of open issues
        ws_audits['U4'].value = form_data.get('numOpenIssues', '')
        ws_audits['U4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['U4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['U4'].border = thin_border
        
        # V4 = Geographical detail
        ws_audits['V4'].value = form_data.get('geographicalState', '')
        ws_audits['V4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['V4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['V4'].border = thin_border
        
        # W4 = Date of Audit Completion
        audit_completion_date = form_data.get('dateAuditCompletion', '')
        try:
            if audit_completion_date:
                date_obj = datetime.strptime(audit_completion_date, '%Y-%m-%d')
                audit_completion_date = date_obj.strftime('%d/%m/%Y')
        except:
            pass
        ws_audits['W4'].value = audit_completion_date
        ws_audits['W4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['W4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['W4'].border = thin_border
        
        # X4 = Date of last audit
        last_audit_date = form_data.get('dateLastAudit', '')
        try:
            if last_audit_date:
                date_obj = datetime.strptime(last_audit_date, '%Y-%m-%d')
                last_audit_date = date_obj.strftime('%d/%m/%Y')
        except:
            pass
        ws_audits['X4'].value = last_audit_date
        ws_audits['X4'].font = Font(name='Arial', size=12, color='FF000000')
        ws_audits['X4'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_audits['X4'].border = thin_border
        
        print(f"  ✅ Created '2. Audits Completed' worksheet")
        
        # Create "3. Security Issues" worksheet
        print(f"🔍 DEBUG: Creating '3. Security Issues' worksheet...")
        ws_security = wb.create_sheet("3. Security Issues")
        
        # Set column widths
        ws_security.column_dimensions['A'].width = 15
        ws_security.column_dimensions['B'].width = 60
        ws_security.column_dimensions['C'].width = 25
        ws_security.column_dimensions['D'].width = 40
        ws_security.column_dimensions['E'].width = 30
        ws_security.column_dimensions['F'].width = 30
        ws_security.column_dimensions['G'].width = 80
        ws_security.column_dimensions['H'].width = 35
        ws_security.column_dimensions['I'].width = 25
        ws_security.column_dimensions['J'].width = 30
        ws_security.column_dimensions['K'].width = 35
        
        # Row 1: Merge A1:K1 with mixed color text
        ws_security.merge_cells('A1:K1')
        ws_security.row_dimensions[1].height = 25
        
        # Create rich text with black "Details of Vulnerability / Security Issues" and red "(Organization Wise)"
        rich_text_security = CellRichText(
            TextBlock(InlineFont(rFont='Arial', sz=14, b=True, color='FF000000'), 
                     "Details of Vulnerability / Security Issues "),
            TextBlock(InlineFont(rFont='Arial', sz=14, b=True, color=red_color), 
                     "(Organization Wise)")
        )
        ws_security['A1'].value = rich_text_security
        ws_security['A1'].fill = PatternFill(start_color=light_sky_blue, end_color=light_sky_blue, fill_type='solid')
        ws_security['A1'].alignment = Alignment(horizontal='center', vertical='center')
        # Apply border to all cells in merged range A1:K1
        for col in range(1, 12):  # A to K
            ws_security.cell(row=1, column=col).border = thin_border
        
        # Row 2: Headers
        headers_security = [
            "SNO.",
            "Name of Auditee organization",
            "Audited infrastructure details",
            "Type of Audit",
            "Please Specify type of Audit (if, selected any other)",
            "Type of Audit Report(First Audit Report/Follow up Report)",
            "Name of Vulnerability / Security Issues as per CVE/CWE/Mitre DB Only",
            "CVE/CWE Reference Number",
            "Severity of Issue",
            "Cumulative count of no. of occurrence of vulnerability in audit.",
            "Attributing Factor "
        ]
        
        for idx, header in enumerate(headers_security, start=1):
            cell = ws_security.cell(row=2, column=idx)
            cell.value = header
            cell.font = Font(name='Arial', size=12, bold=True, color='FF000000')
            cell.fill = PatternFill(start_color=very_creamy_peach, end_color=very_creamy_peach, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Get vulnerability status from form_data
        vulnerability_status = form_data.get('vulnerabilityStatus', 'has')
        print(f"🔍 DEBUG: Vulnerability status: {vulnerability_status}")
        
        # Check if "0 Vulnerability" is selected
        if vulnerability_status == '0':
            # Merge cells A4-K4 and add the note message
            ws_security.merge_cells('A4:K4')
            
            # Define green color (RGB: 0, 128, 0 or 00FF0080 in hex with alpha)
            green_color = 'FF008000'  # Dark green
            
            # Set the merged cell value
            note_text = "Note: All previously reported vulnerabilities were verified as closed during the follow-up audit."
            ws_security['A4'].value = note_text
            ws_security['A4'].font = Font(name='Arial', size=12, bold=True, color=green_color)
            ws_security['A4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws_security['A4'].border = thin_border
            
            # Apply border to all cells in merged range A4-K4
            for col in range(1, 12):  # A to K (1 to 11)
                ws_security.cell(row=4, column=col).border = thin_border
            
            ws_security.row_dimensions[4].height = 40  # Set appropriate row height
            
            print(f"  ✅ Created '3. Security Issues' worksheet with no vulnerability note")
        else:
            # Get vulnerability details from form_data
            vuln_details = form_data.get('vulnerabilities', [])
            print(f"🔍 DEBUG: Vulnerability details count: {len(vuln_details)}")
            print(f"🔍 DEBUG: Vulnerability details: {vuln_details}")
            
            # Get common data
            auditee_org_security = form_data.get('nameOfAuditeeOrg', '')
            if auditee_org_security == 'Other':
                auditee_org_security = form_data.get('nameOfAuditeeOrgOther', '')
            
            type_of_audit_security = form_data.get('typeOfAudit', '')
            type_of_audit_other_security = form_data.get('typeOfAuditOther', '') if type_of_audit_security and 'other' in type_of_audit_security.lower() else "NA"
            type_of_audit_report_security = form_data.get('typeOfAuditReport', '')
            
            # Add data rows (starting from row 3)
            for idx, vuln in enumerate(vuln_details, start=1):
                row_num = idx + 2  # Row 3, 4, 5, etc.
                
                # A column: Serial number
                ws_security.cell(row=row_num, column=1).value = str(idx)
                ws_security.cell(row=row_num, column=1).font = Font(name='Arial', size=12, color='FF000000')
                ws_security.cell(row=row_num, column=1).alignment = Alignment(horizontal='left', vertical='center')
                ws_security.cell(row=row_num, column=1).border = thin_border
                
                # B column: Name of Auditee organization (same for all entries)
                ws_security.cell(row=row_num, column=2).value = auditee_org_security
                ws_security.cell(row=row_num, column=2).font = Font(name='Arial', size=12, color='FF000000')
                ws_security.cell(row=row_num, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws_security.cell(row=row_num, column=2).border = thin_border
                
                # C column: Audited infrastructure details
                ws_security.cell(row=row_num, column=3).value = vuln.get('infrastructure', '')
                ws_security.cell(row=row_num, column=3).font = Font(name='Arial', size=12, color='FF000000')
                ws_security.cell(row=row_num, column=3).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws_security.cell(row=row_num, column=3).border = thin_border
                
                # D column: Type of Audit (same for all entries)
                ws_security.cell(row=row_num, column=4).value = type_of_audit_security
                ws_security.cell(row=row_num, column=4).font = Font(name='Arial', size=12, color='FF000000')
                ws_security.cell(row=row_num, column=4).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws_security.cell(row=row_num, column=4).border = thin_border
                
                # E column: Type of Audit Other (same for all entries)
                ws_security.cell(row=row_num, column=5).value = type_of_audit_other_security
                ws_security.cell(row=row_num, column=5).font = Font(name='Arial', size=12, color='FF000000')
                ws_security.cell(row=row_num, column=5).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws_security.cell(row=row_num, column=5).border = thin_border
                
                # F column: Type of Audit Report (same for all entries)
                ws_security.cell(row=row_num, column=6).value = type_of_audit_report_security
                ws_security.cell(row=row_num, column=6).font = Font(name='Arial', size=12, color='FF000000')
                ws_security.cell(row=row_num, column=6).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws_security.cell(row=row_num, column=6).border = thin_border
                
                # G column: Name of Vulnerability
                ws_security.cell(row=row_num, column=7).value = vuln.get('vulnerability_name', '')
                ws_security.cell(row=row_num, column=7).font = Font(name='Arial', size=12, color='FF000000')
                ws_security.cell(row=row_num, column=7).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws_security.cell(row=row_num, column=7).border = thin_border
                
                # H column: CVE/CWE Reference Number
                ws_security.cell(row=row_num, column=8).value = vuln.get('cve_reference', '')
                ws_security.cell(row=row_num, column=8).font = Font(name='Arial', size=12, color='FF000000')
                ws_security.cell(row=row_num, column=8).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws_security.cell(row=row_num, column=8).border = thin_border
                
                # I column: Severity
                ws_security.cell(row=row_num, column=9).value = vuln.get('severity', '')
                ws_security.cell(row=row_num, column=9).font = Font(name='Arial', size=12, color='FF000000')
                ws_security.cell(row=row_num, column=9).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws_security.cell(row=row_num, column=9).border = thin_border
                
                # J column: Cumulative count
                ws_security.cell(row=row_num, column=10).value = vuln.get('count', '')
                ws_security.cell(row=row_num, column=10).font = Font(name='Arial', size=12, color='FF000000')
                ws_security.cell(row=row_num, column=10).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws_security.cell(row=row_num, column=10).border = thin_border
                
                # K column: Attributing Factor
                ws_security.cell(row=row_num, column=11).value = vuln.get('attributing_factor', '')
                ws_security.cell(row=row_num, column=11).font = Font(name='Arial', size=12, color='FF000000')
                ws_security.cell(row=row_num, column=11).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws_security.cell(row=row_num, column=11).border = thin_border
            
            print(f"  ✅ Created '3. Security Issues' worksheet with {len(vuln_details)} entries")
        
        # Create "4. Manpower involved in audits" worksheet
        print(f"🔍 DEBUG: Creating '4. Manpower involved in audits' worksheet...")
        ws_manpower = wb.create_sheet("4. Manpower involved in audits")
        
        # Set column widths
        ws_manpower.column_dimensions['A'].width = 10
        ws_manpower.column_dimensions['B'].width = 45
        ws_manpower.column_dimensions['C'].width = 45
        ws_manpower.column_dimensions['D'].width = 18
        ws_manpower.column_dimensions['E'].width = 18
        ws_manpower.column_dimensions['F'].width = 18
        ws_manpower.column_dimensions['G'].width = 18
        ws_manpower.column_dimensions['H'].width = 18
        ws_manpower.column_dimensions['I'].width = 18
        ws_manpower.column_dimensions['J'].width = 18
        ws_manpower.column_dimensions['K'].width = 45
        ws_manpower.column_dimensions['L'].width = 40
        
        # Row 1 and 2: Headers with merged cells
        # A1:A2 - S.No
        ws_manpower.merge_cells('A1:A2')
        ws_manpower['A1'].value = "S.No"
        ws_manpower['A1'].font = Font(name='Arial', size=12, bold=True, color='FF000000')
        ws_manpower['A1'].fill = PatternFill(start_color=very_creamy_peach, end_color=very_creamy_peach, fill_type='solid')
        ws_manpower['A1'].alignment = Alignment(horizontal='center', vertical='center')
        for row in [1, 2]:
            ws_manpower.cell(row=row, column=1).border = thin_border
        
        # B1:B2 - Name of employee
        ws_manpower.merge_cells('B1:B2')
        ws_manpower['B1'].value = "Name of employee"
        ws_manpower['B1'].font = Font(name='Arial', size=12, bold=True, color='FF000000')
        ws_manpower['B1'].fill = PatternFill(start_color=very_creamy_peach, end_color=very_creamy_peach, fill_type='solid')
        ws_manpower['B1'].alignment = Alignment(horizontal='center', vertical='center')
        for row in [1, 2]:
            ws_manpower.cell(row=row, column=2).border = thin_border
        
        # C1:C2 - E-mail ID
        ws_manpower.merge_cells('C1:C2')
        ws_manpower['C1'].value = "E-mail ID"
        ws_manpower['C1'].font = Font(name='Arial', size=12, bold=True, color='FF000000')
        ws_manpower['C1'].fill = PatternFill(start_color=very_creamy_peach, end_color=very_creamy_peach, fill_type='solid')
        ws_manpower['C1'].alignment = Alignment(horizontal='center', vertical='center')
        for row in [1, 2]:
            ws_manpower.cell(row=row, column=3).border = thin_border
        
        # D1:K1 - Certifications(Yes/No)
        ws_manpower.merge_cells('D1:K1')
        ws_manpower['D1'].value = "Certifications(Yes/No)"
        ws_manpower['D1'].font = Font(name='Arial', size=12, bold=True, color='FF000000')
        ws_manpower['D1'].fill = PatternFill(start_color=very_creamy_peach, end_color=very_creamy_peach, fill_type='solid')
        ws_manpower['D1'].alignment = Alignment(horizontal='center', vertical='center')
        for col in range(4, 12):  # D to K
            ws_manpower.cell(row=1, column=col).border = thin_border
        
        # Row 2: Individual certification headers
        cert_headers = ["CISSP", "CISA", "CISM", "ISO", "DISA", "OSCP", "CEH", "Others"]
        for idx, cert in enumerate(cert_headers, start=4):  # Start from column D (4)
            ws_manpower.cell(row=2, column=idx).value = cert
            ws_manpower.cell(row=2, column=idx).font = Font(name='Arial', size=12, bold=True, color='FF000000')
            ws_manpower.cell(row=2, column=idx).fill = PatternFill(start_color=very_creamy_peach, end_color=very_creamy_peach, fill_type='solid')
            ws_manpower.cell(row=2, column=idx).alignment = Alignment(horizontal='center', vertical='center')
            ws_manpower.cell(row=2, column=idx).border = thin_border
        
        # L1:L2 - Experience
        ws_manpower.merge_cells('L1:L2')
        ws_manpower['L1'].value = "Experience in Cyber Security Audit (No. of Years)"
        ws_manpower['L1'].font = Font(name='Arial', size=12, bold=True, color='FF000000')
        ws_manpower['L1'].fill = PatternFill(start_color=very_creamy_peach, end_color=very_creamy_peach, fill_type='solid')
        ws_manpower['L1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in [1, 2]:
            ws_manpower.cell(row=row, column=12).border = thin_border
        
        # Get employee details from form_data
        emp_details = form_data.get('employees', [])
        print(f"🔍 DEBUG: Employee details count: {len(emp_details)}")
        print(f"🔍 DEBUG: Employee details: {emp_details}")
        
        # Add first fixed entry (Niraj Goyal)
        print(f"🔍 DEBUG: Adding first fixed entry for Niraj Goyal...")
        row_num = 3
        
        # A column: Serial number (1)
        ws_manpower.cell(row=row_num, column=1).value = "1"
        ws_manpower.cell(row=row_num, column=1).font = Font(name='Arial', size=12, color='FF000000')
        ws_manpower.cell(row=row_num, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws_manpower.cell(row=row_num, column=1).border = thin_border
        
        # B column: Name of employee
        ws_manpower.cell(row=row_num, column=2).value = "Niraj Goyal"
        ws_manpower.cell(row=row_num, column=2).font = Font(name='Arial', size=12, color='FF000000')
        ws_manpower.cell(row=row_num, column=2).alignment = Alignment(horizontal='left', vertical='center')
        ws_manpower.cell(row=row_num, column=2).border = thin_border
        
        # C column: Email ID
        ws_manpower.cell(row=row_num, column=3).value = "admin@ngtech.co.in"
        ws_manpower.cell(row=row_num, column=3).font = Font(name='Arial', size=12, color='FF000000')
        ws_manpower.cell(row=row_num, column=3).alignment = Alignment(horizontal='left', vertical='center')
        ws_manpower.cell(row=row_num, column=3).border = thin_border
        
        # D column: CISSP
        ws_manpower.cell(row=row_num, column=4).value = ""
        ws_manpower.cell(row=row_num, column=4).font = Font(name='Arial', size=12, color='FF000000')
        ws_manpower.cell(row=row_num, column=4).alignment = Alignment(horizontal='left', vertical='center')
        ws_manpower.cell(row=row_num, column=4).border = thin_border
        
        # E column: CISA
        ws_manpower.cell(row=row_num, column=5).value = ""
        ws_manpower.cell(row=row_num, column=5).font = Font(name='Arial', size=12, color='FF000000')
        ws_manpower.cell(row=row_num, column=5).alignment = Alignment(horizontal='left', vertical='center')
        ws_manpower.cell(row=row_num, column=5).border = thin_border
        
        # F column: CISM
        ws_manpower.cell(row=row_num, column=6).value = ""
        ws_manpower.cell(row=row_num, column=6).font = Font(name='Arial', size=12, color='FF000000')
        ws_manpower.cell(row=row_num, column=6).alignment = Alignment(horizontal='left', vertical='center')
        ws_manpower.cell(row=row_num, column=6).border = thin_border
        
        # G column: ISO
        ws_manpower.cell(row=row_num, column=7).value = ""
        ws_manpower.cell(row=row_num, column=7).font = Font(name='Arial', size=12, color='FF000000')
        ws_manpower.cell(row=row_num, column=7).alignment = Alignment(horizontal='left', vertical='center')
        ws_manpower.cell(row=row_num, column=7).border = thin_border
        
        # H column: DISA
        ws_manpower.cell(row=row_num, column=8).value = "Yes"
        ws_manpower.cell(row=row_num, column=8).font = Font(name='Arial', size=12, color='FF000000')
        ws_manpower.cell(row=row_num, column=8).alignment = Alignment(horizontal='left', vertical='center')
        ws_manpower.cell(row=row_num, column=8).border = thin_border
        
        # I column: OSCP
        ws_manpower.cell(row=row_num, column=9).value = ""
        ws_manpower.cell(row=row_num, column=9).font = Font(name='Arial', size=12, color='FF000000')
        ws_manpower.cell(row=row_num, column=9).alignment = Alignment(horizontal='left', vertical='center')
        ws_manpower.cell(row=row_num, column=9).border = thin_border
        
        # J column: CEH
        ws_manpower.cell(row=row_num, column=10).value = "Yes"
        ws_manpower.cell(row=row_num, column=10).font = Font(name='Arial', size=12, color='FF000000')
        ws_manpower.cell(row=row_num, column=10).alignment = Alignment(horizontal='left', vertical='center')
        ws_manpower.cell(row=row_num, column=10).border = thin_border
        
        # K column: Others
        ws_manpower.cell(row=row_num, column=11).value = "-"
        ws_manpower.cell(row=row_num, column=11).font = Font(name='Arial', size=12, color='FF000000')
        ws_manpower.cell(row=row_num, column=11).alignment = Alignment(horizontal='left', vertical='center')
        ws_manpower.cell(row=row_num, column=11).border = thin_border
        
        # L column: Experience
        ws_manpower.cell(row=row_num, column=12).value = "9 Years"
        ws_manpower.cell(row=row_num, column=12).font = Font(name='Arial', size=12, color='FF000000')
        ws_manpower.cell(row=row_num, column=12).alignment = Alignment(horizontal='left', vertical='center')
        ws_manpower.cell(row=row_num, column=12).border = thin_border
        
        print(f"  ✅ Added first fixed entry for Niraj Goyal")
        
        # Add data rows from user input (starting from row 4)
        for idx, emp in enumerate(emp_details, start=2):
            row_num = idx + 2  # Row 3, 4, 5, etc.
            
            # A column: Serial number
            ws_manpower.cell(row=row_num, column=1).value = str(idx)
            ws_manpower.cell(row=row_num, column=1).font = Font(name='Arial', size=12, color='FF000000')
            ws_manpower.cell(row=row_num, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws_manpower.cell(row=row_num, column=1).border = thin_border
            
            # B column: Name of employee
            ws_manpower.cell(row=row_num, column=2).value = emp.get('name', '')
            ws_manpower.cell(row=row_num, column=2).font = Font(name='Arial', size=12, color='FF000000')
            ws_manpower.cell(row=row_num, column=2).alignment = Alignment(horizontal='left', vertical='center')
            ws_manpower.cell(row=row_num, column=2).border = thin_border
            
            # C column: Email ID
            ws_manpower.cell(row=row_num, column=3).value = emp.get('email', '')
            ws_manpower.cell(row=row_num, column=3).font = Font(name='Arial', size=12, color='FF000000')
            ws_manpower.cell(row=row_num, column=3).alignment = Alignment(horizontal='left', vertical='center')
            ws_manpower.cell(row=row_num, column=3).border = thin_border
            
            # D column: CISSP
            cissp_value = emp.get('cissp', '')
            ws_manpower.cell(row=row_num, column=4).value = '' if cissp_value == 'No' else cissp_value
            ws_manpower.cell(row=row_num, column=4).font = Font(name='Arial', size=12, color='FF000000')
            ws_manpower.cell(row=row_num, column=4).alignment = Alignment(horizontal='left', vertical='center')
            ws_manpower.cell(row=row_num, column=4).border = thin_border
            
            # E column: CISA
            cisa_value = emp.get('cisa', '')
            ws_manpower.cell(row=row_num, column=5).value = '' if cisa_value == 'No' else cisa_value
            ws_manpower.cell(row=row_num, column=5).font = Font(name='Arial', size=12, color='FF000000')
            ws_manpower.cell(row=row_num, column=5).alignment = Alignment(horizontal='left', vertical='center')
            ws_manpower.cell(row=row_num, column=5).border = thin_border
            
            # F column: CISM
            cism_value = emp.get('cism', '')
            ws_manpower.cell(row=row_num, column=6).value = '' if cism_value == 'No' else cism_value
            ws_manpower.cell(row=row_num, column=6).font = Font(name='Arial', size=12, color='FF000000')
            ws_manpower.cell(row=row_num, column=6).alignment = Alignment(horizontal='left', vertical='center')
            ws_manpower.cell(row=row_num, column=6).border = thin_border
            
            # G column: ISO
            iso_value = emp.get('iso', '')
            ws_manpower.cell(row=row_num, column=7).value = '' if iso_value == 'No' else iso_value
            ws_manpower.cell(row=row_num, column=7).font = Font(name='Arial', size=12, color='FF000000')
            ws_manpower.cell(row=row_num, column=7).alignment = Alignment(horizontal='left', vertical='center')
            ws_manpower.cell(row=row_num, column=7).border = thin_border
            
            # H column: DISA
            disa_value = emp.get('disa', '')
            ws_manpower.cell(row=row_num, column=8).value = '' if disa_value == 'No' else disa_value
            ws_manpower.cell(row=row_num, column=8).font = Font(name='Arial', size=12, color='FF000000')
            ws_manpower.cell(row=row_num, column=8).alignment = Alignment(horizontal='left', vertical='center')
            ws_manpower.cell(row=row_num, column=8).border = thin_border
            
            # I column: OSCP
            oscp_value = emp.get('oscp', '')
            ws_manpower.cell(row=row_num, column=9).value = '' if oscp_value == 'No' else oscp_value
            ws_manpower.cell(row=row_num, column=9).font = Font(name='Arial', size=12, color='FF000000')
            ws_manpower.cell(row=row_num, column=9).alignment = Alignment(horizontal='left', vertical='center')
            ws_manpower.cell(row=row_num, column=9).border = thin_border
            
            # J column: CEH
            ceh_value = emp.get('ceh', '')
            ws_manpower.cell(row=row_num, column=10).value = '' if ceh_value == 'No' else ceh_value
            ws_manpower.cell(row=row_num, column=10).font = Font(name='Arial', size=12, color='FF000000')
            ws_manpower.cell(row=row_num, column=10).alignment = Alignment(horizontal='left', vertical='center')
            ws_manpower.cell(row=row_num, column=10).border = thin_border
            
            # K column: Others
            others_value = emp.get('others', '')
            ws_manpower.cell(row=row_num, column=11).value = '' if others_value == 'No' else others_value
            ws_manpower.cell(row=row_num, column=11).font = Font(name='Arial', size=12, color='FF000000')
            ws_manpower.cell(row=row_num, column=11).alignment = Alignment(horizontal='left', vertical='center')
            ws_manpower.cell(row=row_num, column=11).border = thin_border
            
            # L column: Experience (add "Years" if it's a number)
            experience_value = emp.get('experience', '')
            if experience_value and str(experience_value).strip():
                # Check if it's already formatted with "Years"
                if 'Years' not in str(experience_value):
                    ws_manpower.cell(row=row_num, column=12).value = f"{experience_value} Years"
                else:
                    ws_manpower.cell(row=row_num, column=12).value = experience_value
            else:
                ws_manpower.cell(row=row_num, column=12).value = ""
            ws_manpower.cell(row=row_num, column=12).font = Font(name='Arial', size=12, color='FF000000')
            ws_manpower.cell(row=row_num, column=12).alignment = Alignment(horizontal='left', vertical='center')
            ws_manpower.cell(row=row_num, column=12).border = thin_border
        
        total_entries = 1 + len(emp_details)  # 1 fixed entry + user entries
        print(f"  ✅ Created '4. Manpower involved in audits' worksheet with {total_entries} entries (1 fixed + {len(emp_details)} user entries)")
        
        # Save to file
        print(f"🔍 DEBUG: Preparing to save file...")
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Public_IP_Follow_Up_Audit_MetaData_{timestamp}.xlsx'
        filepath = os.path.join('static', 'uploads', filename)
        print(f"🔍 DEBUG: File will be saved to: {filepath}")
        
        # Ensure directory exists
        print(f"🔍 DEBUG: Creating uploads directory if needed...")
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        print(f"🔍 DEBUG: Directory ready")
        
        print(f"🔍 DEBUG: Saving workbook (this may take a moment)...")
        wb.save(filepath)
        print(f"🔍 DEBUG: Workbook saved successfully!")
        print(f"💾 Saved: {filename}")
        
        print(f"🔍 DEBUG: Returning filepath and filename...")
        return filepath, filename
    
    except Exception as e:
        print(f"❌ Error creating metadata Excel: {e}")
        import traceback
        traceback.print_exc()
        raise

def get_public_ip_follow_up_vulnerability_details():
    """Safely collect vulnerability details from form with safety limit"""
    vuln_details = []
    max_entries = 50  # Safety limit
    
    for i in range(max_entries):
        infrastructure = request.form.get(f'vuln_{i}_infrastructure')
        if not infrastructure:  # Check if empty or None
            continue
            
        vuln_details.append({
            'infrastructure': infrastructure,
            'vulnerability_name': request.form.get(f'vuln_{i}_name', ''),
            'cve_reference': request.form.get(f'vuln_{i}_cve', ''),
            'severity': request.form.get(f'vuln_{i}_severity', ''),
            'count': request.form.get(f'vuln_{i}_count', ''),
            'attributing_factor': request.form.get(f'vuln_{i}_factor', '')
        })
    
    return vuln_details

def get_public_ip_follow_up_employee_details():
    """Safely collect employee details from form with safety limit"""
    employee_details = []
    max_entries = 20  # Safety limit
    
    for i in range(max_entries):
        emp_name = request.form.get(f'emp_{i}_name')
        if not emp_name:  # Check if empty or None
            continue
            
        employee_details.append({
            'name': emp_name,
            'email': request.form.get(f'emp_{i}_email', ''),
            'cissp': request.form.get(f'emp_{i}_cissp', ''),
            'cisa': request.form.get(f'emp_{i}_cisa', ''),
            'cism': request.form.get(f'emp_{i}_cism', ''),
            'iso': request.form.get(f'emp_{i}_iso', ''),
            'disa': request.form.get(f'emp_{i}_disa', ''),
            'oscp': request.form.get(f'emp_{i}_oscp', ''),
            'ceh': request.form.get(f'emp_{i}_ceh', ''),
            'others': request.form.get(f'emp_{i}_others', ''),
            'experience': request.form.get(f'emp_{i}_experience', '')
        })
    
    return employee_details

@public_ip_follow_up_audit_metadata_bp.route('/process_public_ip_follow_up_audit_metadata', methods=['POST'])
def process_public_ip_follow_up_audit_metadata():
    """Process Public IP Follow Up Audit MetaData form submission"""
    try:
        print("\n" + "="*80)
        print("🚀 Processing Public IP Follow Up Audit MetaData")
        print("="*80)
        print(f"🔍 DEBUG: Request method: {request.method}")
        print(f"🔍 DEBUG: Form keys count: {len(list(request.form.keys()))}")
        
        # Get form data
        form_data = {
            'startAuditDate': request.form.get('startAuditDate'),
            'endAuditDate': request.form.get('endAuditDate'),
            'nameOfAuditeeOrg': request.form.get('nameOfAuditeeOrg'),
            'nameOfAuditeeOrgOther': request.form.get('nameOfAuditeeOrgOther'),
            'categoryOfOrg': request.form.get('categoryOfOrg'),
            'sectorOfOrg': request.form.get('sectorOfOrg'),
            'subSector': request.form.get('subSector'),
            'typeOfAudit': request.form.get('typeOfAudit'),
            'typeOfAuditOther': request.form.get('typeOfAuditOther'),
            'detailsOfWebsite': request.form.get('detailsOfWebsite'),
            'reasonForAudit': request.form.get('reasonForAudit'),
            'standardsUsed': request.form.get('standardsUsed'),
            'challenges': request.form.get('challenges'),
            'complianceFinancial': request.form.get('complianceFinancial'),
            'complianceRegulatory': request.form.get('complianceRegulatory'),
            'complianceCertIn': request.form.get('complianceCertIn'),
            'complianceCEA': request.form.get('complianceCEA'),
            'complianceOther': request.form.get('complianceOther'),
            'typeOfAuditReport': request.form.get('typeOfAuditReport'),
            'numVulnerabilities': request.form.get('numVulnerabilities'),
            'numDaysToPatch': request.form.get('numDaysToPatch'),
            'durationBetweenAudits': request.form.get('durationBetweenAudits'),
            'numOpenIssues': request.form.get('numOpenIssues'),
            'geographicalState': request.form.get('geographicalState'),
            'dateAuditCompletion': request.form.get('dateAuditCompletion'),
            'dateLastAudit': request.form.get('dateLastAudit'),
            'vulnerabilityStatus': request.form.get('vulnerabilityStatus', 'has')
        }
        
        print(f"🔍 DEBUG: Collecting vulnerability details...")
        # Get vulnerability and employee details using safe methods
        form_data['vulnerabilities'] = get_public_ip_follow_up_vulnerability_details()
        print(f"🔍 DEBUG: Vulnerabilities collected: {len(form_data['vulnerabilities'])}")
        
        print(f"🔍 DEBUG: Collecting employee details...")
        form_data['employees'] = get_public_ip_follow_up_employee_details()
        print(f"🔍 DEBUG: Employees collected: {len(form_data['employees'])}")
        
        print(f"  📋 Form data collected")
        print(f"  👥 Vulnerabilities: {len(form_data['vulnerabilities'])}")
        print(f"  👥 Employees: {len(form_data['employees'])}")
        
        print(f"🔍 DEBUG: Creating Excel file...")
        # Create Excel file
        filepath, filename = create_public_ip_follow_up_audit_metadata_excel(form_data)
        print(f"\n✅ Excel file created: {filename}")
        print(f"🔍 DEBUG: File path: {filepath}")
        print("="*80)
        
        # Return file info as JSON
        return jsonify({
            'success': True,
            'filename': filename,
            'download_url': f'/static/uploads/{filename}'
        })
    
    except Exception as e:
        # Log error securely (server-side only)
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error processing public IP follow up audit metadata: {type(e).__name__}: {str(e)}", exc_info=True)
        # Return safe error message to client
        return jsonify({'success': False, 'error': 'An error occurred processing the request. Please try again later.'}), 500

# Cleanup endpoint
@public_ip_follow_up_audit_metadata_bp.route('/cleanup_public_ip_follow_up_audit_metadata', methods=['POST'])
def cleanup_public_ip_follow_up_audit_metadata():
    """Clean up old VAPT Public IP Follow Up Audit MetaData Excel files"""
    try:
        uploads_dir = os.path.join('static', 'uploads')
        if os.path.exists(uploads_dir):
            for filename in os.listdir(uploads_dir):
                if filename.startswith('Public_IP_Follow_Up_Audit_MetaData_') and filename.endswith('.xlsx'):
                    file_path = os.path.join(uploads_dir, filename)
                    try:
                        os.remove(file_path)
                        print(f"Deleted: {filename}")
                    except Exception as e:
                        print(f"Error deleting {filename}: {e}")
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

