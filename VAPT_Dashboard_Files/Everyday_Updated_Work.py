"""
Everyday Updated Work Module - Handles daily work updates
"""
from flask import Blueprint, request, jsonify
from flask_login import current_user
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

everyday_updated_work_bp = Blueprint('everyday_updated_work', __name__)

def get_month_name(month):
    """Convert month number to month name"""
    month_names = {
        1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr',
        5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Aug',
        9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
    }
    return month_names.get(month, 'Jan')

def generate_filename():
    """Generate filename based on current date"""
    now = datetime.now()
    day = str(now.day)
    month = get_month_name(now.month)
    year = str(now.year)
    return f"{day}_{month}_{year}.xlsx"

def get_file_path():
    """Get the full path for the update work file"""
    filename = generate_filename()
    base_dir = os.path.join('static', 'Activity_Tracker', 'Everyday_Updated_Work')
    os.makedirs(base_dir, exist_ok=True)
    return os.path.join(base_dir, filename)

def format_date_for_display(date_str):
    """Convert date string to readable format"""
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        return date_obj.strftime('%d-%m-%Y')
    except:
        return date_str

def create_client_time_worksheet(wb, all_data):
    """Create Client_Time worksheet with aggregated client data"""
    if 'Client_Time' in wb.sheetnames:
        ws_client = wb['Client_Time']
    else:
        ws_client = wb.create_sheet('Client_Time')
    
    # Set headers
    ws_client['A1'] = 'Sr.No'
    ws_client['B1'] = 'Client Name'
    ws_client['C1'] = 'Time'
    ws_client['D1'] = 'Employees and Task'
    
    # Format headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col in ['A1', 'B1', 'C1', 'D1']:
        cell = ws_client[col]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Set column widths
    ws_client.column_dimensions['A'].width = 10
    ws_client.column_dimensions['B'].width = 40
    ws_client.column_dimensions['C'].width = 20
    ws_client.column_dimensions['D'].width = 100
    
    # Aggregate data by client
    client_data = {}
    for row_data in all_data:
        employee_name = row_data.get('employee_name', '')
        tasks = []
        for i in range(10):
            task_key = f'task_{i}'
            client_key = f'client_{i}'
            time_key = f'time_{i}'
            
            task = row_data.get(task_key, '').strip()
            client = row_data.get(client_key, '').strip()
            time = row_data.get(time_key, '')
            
            if task and client and time:
                if client not in client_data:
                    client_data[client] = {'total_time': 0, 'entries': []}
                
                try:
                    time_val = float(time)
                    client_data[client]['total_time'] += time_val
                    client_data[client]['entries'].append({
                        'employee': employee_name,
                        'task': task,
                        'time': time_val
                    })
                except:
                    pass
    
    # Write aggregated data
    row = 2
    sr_no = 1
    for client_name, data in client_data.items():
        ws_client.cell(row=row, column=1).value = sr_no
        ws_client.cell(row=row, column=2).value = client_name
        ws_client.cell(row=row, column=3).value = data['total_time']
        
        # Create employee-task details
        details = []
        for entry in data['entries']:
            details.append(f"{entry['employee']} - {entry['task']} - {entry['time']}")
        
        ws_client.cell(row=row, column=4).value = '\n'.join(details)
        ws_client.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Apply border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, 5):
            cell = ws_client.cell(row=row, column=col)
            cell.border = thin_border
        
        sr_no += 1
        row += 1

def create_update_work_excel(employee_name, employee_team, tasks_data):
    """Create or update the Excel file"""
    file_path = get_file_path()
    all_rows_data = []
    
    # Load existing data if file exists
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Read existing rows
        for row in range(2, ws.max_row + 1):
            row_data = {
                'employee_name': ws.cell(row=row, column=1).value or '',
                'date': ws.cell(row=row, column=2).value or '',
                'employee_team': ws.cell(row=row, column=3).value or '',
                'total_time': ws.cell(row=row, column=4).value or '',
                'timestamp': ws.cell(row=row, column=5).value or ''
            }
            
            # Read task data
            col = 6
            for i in range(10):
                row_data[f'task_{i}'] = ws.cell(row=row, column=col).value or ''
                row_data[f'client_{i}'] = ws.cell(row=row, column=col+1).value or ''
                row_data[f'time_{i}'] = ws.cell(row=row, column=col+2).value or ''
                row_data[f'status_{i}'] = ws.cell(row=row, column=col+3).value or ''
                row_data[f'description_{i}'] = ws.cell(row=row, column=col+4).value or ''
                col += 5
            
            all_rows_data.append(row_data)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Updated Work"
        
        # Set headers
        headers = []
        header_values = ['Employee Name', 'Date', 'Employee Team', 'Total Time', 'TimeStamp']
        
        for i in range(1, 11):
            header_values.extend([
                f'Task{i}',
                f'Task{i} Client Name',
                f'Time For Task{i}',
                f'Status For Task{i}',
                f'Description For Task{i}'
            ])
        
        for i, value in enumerate(header_values, start=1):
            cell = ws.cell(row=1, column=i)
            cell.value = value
        
        # Format headers A1 to BC1
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col in range(1, 56):  # Columns A to BC (55 columns)
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Set column widths for first 5 columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        
        # Set widths for task columns (every 5 columns: Task, Client Name, Time, Status, Description)
        for i in range(1, 11):
            # Task column
            task_col = 5 + (i - 1) * 5 + 1
            ws.column_dimensions[ws.cell(row=1, column=task_col).column_letter].width = 40
            
            # Client Name column
            client_col = task_col + 1
            ws.column_dimensions[ws.cell(row=1, column=client_col).column_letter].width = 40
            
            # Time column
            time_col = task_col + 2
            ws.column_dimensions[ws.cell(row=1, column=time_col).column_letter].width = 15
            
            # Status column
            status_col = task_col + 3
            ws.column_dimensions[ws.cell(row=1, column=status_col).column_letter].width = 30
            
            # Description column
            desc_col = task_col + 4
            ws.column_dimensions[ws.cell(row=1, column=desc_col).column_letter].width = 60
    
    # Calculate total time
    total_time = sum(float(task.get('time', 0)) for task in tasks_data if task.get('time'))
    
    # Get current timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Format date
    date_str = request.form.get('date', '')
    formatted_date = format_date_for_display(date_str)
    
    # Prepare row data
    row_data = {
        'employee_name': employee_name,
        'date': formatted_date,
        'employee_team': employee_team,
        'total_time': total_time,
        'timestamp': timestamp
    }
    
    # Add task data (up to 10 tasks)
    col = 6
    for i in range(10):
        task_data = tasks_data[i] if i < len(tasks_data) else {}
        row_data[f'task_{i}'] = task_data.get('task', '')
        row_data[f'client_{i}'] = task_data.get('clientName', '')
        row_data[f'time_{i}'] = task_data.get('time', '')
        row_data[f'status_{i}'] = task_data.get('status', '')
        row_data[f'description_{i}'] = task_data.get('description', '')
    
    # Add new row data
    all_rows_data.append(row_data)
    
    # Rebuild worksheet
    ws = wb.active
    ws.delete_rows(2, ws.max_row)  # Clear existing data rows
    
    # Write all data
    for row_idx, data in enumerate(all_rows_data, start=2):
        ws.cell(row=row_idx, column=1).value = data['employee_name']
        ws.cell(row=row_idx, column=2).value = data['date']
        ws.cell(row=row_idx, column=3).value = data['employee_team']
        ws.cell(row=row_idx, column=4).value = data['total_time']
        ws.cell(row=row_idx, column=5).value = data['timestamp']
        
        col = 6
        for i in range(10):
            ws.cell(row=row_idx, column=col).value = data.get(f'task_{i}', '')
            ws.cell(row=row_idx, column=col+1).value = data.get(f'client_{i}', '')
            ws.cell(row=row_idx, column=col+2).value = data.get(f'time_{i}', '')
            ws.cell(row=row_idx, column=col+3).value = data.get(f'status_{i}', '')
            ws.cell(row=row_idx, column=col+4).value = data.get(f'description_{i}', '')
            col += 5
        
        # Apply border and alignment
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, 56):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
        
        ws.row_dimensions[row_idx].height = 25
    
    # Create/update Client_Time worksheet
    create_client_time_worksheet(wb, all_rows_data)
    
    # Save workbook and ensure it's properly closed
    try:
        wb.save(file_path)
        wb.close()
        print(f"✅ Updated work saved to: {file_path}")
        return True
    except Exception as e:
        # Ensure workbook is closed even if save fails
        try:
            wb.close()
        except:
            pass
        print(f"❌ Error saving updated work: {e}")
        raise

def check_existing_entry(employee_name, date_str):
    """Check if user already submitted update work for today"""
    file_path = get_file_path()
    
    if not os.path.exists(file_path):
        return False
    
    try:
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        
        # Check rows 2 onwards
        for row in range(2, ws.max_row + 1):
            row_employee_name = ws.cell(row=row, column=1).value
            row_date = ws.cell(row=row, column=2).value
            
            formatted_date = format_date_for_display(date_str)
            if str(row_employee_name).strip().lower() == str(employee_name).strip().lower() and str(row_date) == formatted_date:
                wb.close()
                return True
        
        wb.close()
        return False
    except Exception as e:
        print(f"Error checking existing entry: {e}")
        try:
            wb.close()
        except:
            pass
        return False

@everyday_updated_work_bp.route('/check_update_work_status', methods=['POST'])
def check_update_work_status():
    """Check if user has already submitted update work for today"""
    try:
        date_str = request.form.get('date', '')
        employee_name = request.form.get('employeeName', '').strip()
        
        if not employee_name:
            return jsonify({'submitted': False, 'error': 'Employee name missing'}), 400
        
        submitted = check_existing_entry(employee_name, date_str)
        return jsonify({'submitted': submitted})
        
    except Exception as e:
        print(f"Error checking update work status: {e}")
        return jsonify({'submitted': False, 'error': str(e)}), 500

@everyday_updated_work_bp.route('/submit_update_work', methods=['POST'])
def submit_update_work():
    """Handle update work submission"""
    try:
        print("\n" + "="*80)
        print("📋 Processing Update Work Submission")
        print("="*80)
        
        # Get form data
        date_str = request.form.get('date', '')
        employee_name = request.form.get('employeeName', '').strip()
        employee_team = request.form.get('employeeTeam', '').strip()
        
        # Get tasks data
        tasks = request.form.getlist('task[]')
        client_names = request.form.getlist('clientName[]')
        times = request.form.getlist('time[]')
        statuses = request.form.getlist('status[]')
        descriptions = request.form.getlist('description[]')
        
        # Validate data
        if not employee_name or not employee_team:
            return jsonify({'success': False, 'error': 'Employee information is missing'}), 400
        
        if not tasks or len(tasks) == 0:
            return jsonify({'success': False, 'error': 'At least one task is required'}), 400
        
        # Check if user already submitted today
        if check_existing_entry(employee_name, date_str):
            return jsonify({'success': False, 'error': 'You have already updated your work for today'}), 400
        
        # Prepare tasks data
        tasks_data = []
        for i in range(len(tasks)):
            if tasks[i].strip():
                tasks_data.append({
                    'task': tasks[i].strip(),
                    'clientName': client_names[i].strip() if i < len(client_names) else '',
                    'time': times[i] if i < len(times) else '0',
                    'status': statuses[i] if i < len(statuses) else '',
                    'description': descriptions[i].strip() if i < len(descriptions) else ''
                })
        
        print(f"📅 Date: {date_str}")
        print(f"👤 Employee: {employee_name}")
        print(f"👥 Team: {employee_team}")
        print(f"📝 Tasks: {len(tasks_data)}")
        
        # Create or update Excel file
        create_update_work_excel(employee_name, employee_team, tasks_data)
        
        print("✅ Update work submitted successfully!")
        
        return jsonify({
            'success': True,
            'message': 'Work updated successfully!',
            'filename': generate_filename()
        })
        
    except Exception as e:
        print(f"❌ Error submitting update work: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }), 500

