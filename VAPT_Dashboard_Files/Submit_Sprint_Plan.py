"""
Submit Sprint Plan Module - Handles sprint plan submissions for upcoming days
"""
from flask import Blueprint, request, jsonify
from flask_login import current_user
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

submit_sprint_plan_bp = Blueprint('submit_sprint_plan', __name__)

def get_month_name(month):
    """Convert month number to month name"""
    month_names = {
        1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr',
        5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Aug',
        9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
    }
    return month_names.get(month, 'Jan')

def generate_filename():
    """Generate filename based on current date"""
    now = datetime.now()
    day = str(now.day)
    month = get_month_name(now.month)
    year = str(now.year)
    return f"{day}_{month}_{year}.xlsx"

def get_file_path():
    """Get the full path for the sprint plan file"""
    filename = generate_filename()
    base_dir = os.path.join('static', 'Activity_Tracker', 'Submit_Sprint_Plan')
    os.makedirs(base_dir, exist_ok=True)
    return os.path.join(base_dir, filename)

def get_monday_of_current_week():
    """Get the Monday of the current week"""
    today = datetime.now()
    days_since_monday = today.weekday()
    monday = today - timedelta(days=days_since_monday)
    return monday.date()

def format_sprint_plan(dates, tasks_array, client_names_array, times_array):
    """Format sprint plan data into a single string"""
    sprint_plan_parts = []
    
    for i, date in enumerate(dates):
        if not date or i >= len(tasks_array) or i >= len(client_names_array) or i >= len(times_array):
            continue
        
        date_str = date if isinstance(date, str) else date.strftime('%d-%m-%Y')
        tasks = tasks_array[i] if i < len(tasks_array) else []
        client_names = client_names_array[i] if i < len(client_names_array) else []
        times = times_array[i] if i < len(times_array) else []
        
        for j in range(len(tasks)):
            task = tasks[j].strip() if j < len(tasks) and tasks[j] else ''
            client = client_names[j].strip() if j < len(client_names) and client_names[j] else ''
            time = times[j] if j < len(times) and times[j] else ''
            
            if task and client and time:
                sprint_plan_parts.append(f"{date_str} - {client} - {task} - {time}")
    
    return '\n'.join(sprint_plan_parts)

def calculate_total_time(times_array):
    """Calculate total time from all days and tasks"""
    total = 0.0
    for times_list in times_array:
        for time_val in times_list:
            try:
                total += float(time_val)
            except:
                pass
    return total

def create_sprint_plan_excel(employee_name, employee_team, dates, tasks_array, client_names_array, times_array):
    """Create or update the sprint plan Excel file"""
    file_path = get_file_path()
    
    # Check if file exists
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sprint Plan"
        
        # Set headers
        headers = ['Employee Name', 'Employee Team', 'Total Time', 'TimeStamp', 'Sprint Plan']
        
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col).value = header
        
        # Format headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col in range(1, 6):  # A to E
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 100
    
    # Calculate total time
    total_time = calculate_total_time(times_array)
    
    # Get current timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Format sprint plan
    sprint_plan = format_sprint_plan(dates, tasks_array, client_names_array, times_array)
    
    # Prepare row data
    next_row = ws.max_row + 1
    
    ws.cell(row=next_row, column=1).value = employee_name
    ws.cell(row=next_row, column=2).value = employee_team
    ws.cell(row=next_row, column=3).value = total_time
    ws.cell(row=next_row, column=4).value = timestamp
    ws.cell(row=next_row, column=5).value = sprint_plan
    ws.cell(row=next_row, column=5).alignment = Alignment(wrap_text=True, vertical='top')
    
    # Apply border and alignment
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, 6):
        cell = ws.cell(row=next_row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
    
    # Set row height
    ws.row_dimensions[next_row].height = 25
    
    # Save workbook and ensure it's properly closed
    try:
        wb.save(file_path)
        wb.close()
        print(f"✅ Sprint plan saved to: {file_path}")
        return True
    except Exception as e:
        # Ensure workbook is closed even if save fails
        try:
            wb.close()
        except:
            pass
        print(f"❌ Error saving sprint plan: {e}")
        raise

def check_existing_entry(employee_name):
    """Check if user already submitted sprint plan for this week"""
    file_path = get_file_path()
    
    if not os.path.exists(file_path):
        return False
    
    try:
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        
        # Get Monday of current week
        monday = get_monday_of_current_week()
        monday_str = monday.strftime('%d-%m-%Y')
        
        # Check rows 2 onwards
        for row in range(2, ws.max_row + 1):
            row_employee_name = ws.cell(row=row, column=1).value
            
            if str(row_employee_name).strip().lower() == str(employee_name).strip().lower():
                # Check if submitted this week by checking timestamp
                timestamp = ws.cell(row=row, column=4).value
                if timestamp:
                    try:
                        timestamp_date = datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S').date()
                        # If timestamp is after Monday of current week, it's this week
                        if timestamp_date >= monday:
                            wb.close()
                            return True
                    except:
                        pass
        
        wb.close()
        return False
    except Exception as e:
        print(f"Error checking existing entry: {e}")
        try:
            wb.close()
        except:
            pass
        return False

@submit_sprint_plan_bp.route('/check_sprint_plan_status', methods=['POST'])
def check_sprint_plan_status():
    """Check if user has already submitted sprint plan for this week"""
    try:
        employee_name = request.form.get('employeeName', '').strip()
        
        if not employee_name:
            return jsonify({'submitted': False, 'error': 'Employee name missing'}), 400
        
        submitted = check_existing_entry(employee_name)
        return jsonify({'submitted': submitted})
        
    except Exception as e:
        print(f"Error checking sprint plan status: {e}")
        return jsonify({'submitted': False, 'error': str(e)}), 500

@submit_sprint_plan_bp.route('/submit_sprint_plan', methods=['POST'])
def submit_sprint_plan():
    """Handle sprint plan submission"""
    try:
        print("\n" + "="*80)
        print("📋 Processing Sprint Plan Submission")
        print("="*80)
        
        # Get form data
        employee_name = request.form.get('employeeName', '').strip()
        employee_team = request.form.get('employeeTeam', '').strip()
        
        # Get dates array
        dates = request.form.getlist('dates[]')
        
        # Get tasks, client names, and times arrays
        # Since we have nested arrays, we need to process them differently
        form_data = request.form.to_dict(flat=False)
        
        tasks_dict = {}
        client_names_dict = {}
        times_dict = {}
        
        for key, value in form_data.items():
            if key.startswith('tasks['):
                # Extract day index from key like "tasks[0][]"
                day_index = int(key.split('[')[1].split(']')[0])
                if day_index not in tasks_dict:
                    tasks_dict[day_index] = []
                tasks_dict[day_index].extend(value)
            elif key.startswith('clientNames['):
                day_index = int(key.split('[')[1].split(']')[0])
                if day_index not in client_names_dict:
                    client_names_dict[day_index] = []
                client_names_dict[day_index].extend(value)
            elif key.startswith('times['):
                day_index = int(key.split('[')[1].split(']')[0])
                if day_index not in times_dict:
                    times_dict[day_index] = []
                times_dict[day_index].extend(value)
        
        # Convert to lists in order
        tasks_array = []
        client_names_array = []
        times_array = []
        
        max_index = max(tasks_dict.keys()) if tasks_dict else 0
        for i in range(max_index + 1):
            tasks_array.append(tasks_dict.get(i, []))
            client_names_array.append(client_names_dict.get(i, []))
            times_array.append(times_dict.get(i, []))
        
        # Validate data
        if not employee_name or not employee_team:
            return jsonify({'success': False, 'error': 'Employee information is missing'}), 400
        
        if not dates or len(dates) == 0:
            return jsonify({'success': False, 'error': 'At least one day is required'}), 400
        
        # Check if user already submitted this week
        if check_existing_entry(employee_name):
            return jsonify({'success': False, 'error': 'You have already submitted your sprint plan for this week'}), 400
        
        print(f"👤 Employee: {employee_name}")
        print(f"👥 Team: {employee_team}")
        print(f"📅 Days: {len(dates)}")
        
        # Create or update Excel file
        create_sprint_plan_excel(employee_name, employee_team, dates, tasks_array, client_names_array, times_array)
        
        print("✅ Sprint plan submitted successfully!")
        
        return jsonify({
            'success': True,
            'message': 'Sprint plan submitted successfully!',
            'filename': generate_filename()
        })
        
    except Exception as e:
        print(f"❌ Error submitting sprint plan: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }), 500

