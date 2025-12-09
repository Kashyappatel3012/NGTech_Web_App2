"""
Everyday Workplan Module - Handles daily work plan submissions
"""
from flask import Blueprint, request, jsonify
from flask_login import current_user
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import re

everyday_workplan_bp = Blueprint('everyday_workplan', __name__)

def get_month_name(month):
    """Convert month number to month name"""
    month_names = {
        1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr',
        5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Aug',
        9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
    }
    return month_names.get(month, 'Jan')

def generate_filename():
    """Generate filename based on current date"""
    now = datetime.now()
    day = str(now.day)
    month = get_month_name(now.month)
    year = str(now.year)
    return f"{day}_{month}_{year}.xlsx"

def get_file_path():
    """Get the full path for the workplan file"""
    filename = generate_filename()
    base_dir = os.path.join('static', 'Activity_Tracker', 'Everyday_Workplan')
    os.makedirs(base_dir, exist_ok=True)
    return os.path.join(base_dir, filename)

def format_date_for_display(date_str):
    """Convert date string to readable format"""
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        return date_obj.strftime('%d-%m-%Y')
    except:
        return date_str

def create_workplan_excel(employee_name, employee_team, tasks_data):
    """Create or update the workplan Excel file"""
    file_path = get_file_path()
    
    # Check if file exists
    if os.path.exists(file_path):
        # Load existing workbook
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Work Plan"
        
        # Set headers
        headers = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 
                   'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'Q1', 'R1', 'S1', 'T1', 
                   'U1', 'V1', 'W1', 'X1', 'Y1', 'Z1', 'AA1', 'AB1', 'AC1', 'AD1',
                   'AE1', 'AF1', 'AG1', 'AH1', 'AI1', 'AJ1', 'AK1', 'AL1', 'AM1', 'AN1',
                   'AO1', 'AP1', 'AQ1', 'AR1', 'AS1']
        
        header_values = [
            'Employee Name', 'Date', 'Employee Team', 'Total Time', 'TimeStamp',
            'Task1', 'Task1 Description', 'Task1 Client Name', 'Task1 Time',
            'Task2', 'Task2 Description', 'Task2 Client Name', 'Task2 Time',
            'Task3', 'Task3 Description', 'Task3 Client Name', 'Task3 Time',
            'Task4', 'Task4 Description', 'Task4 Client Name', 'Task4 Time',
            'Task5', 'Task5 Description', 'Task5 Client Name', 'Task5 Time',
            'Task6', 'Task6 Description', 'Task6 Client Name', 'Task6 Time',
            'Task7', 'Task7 Description', 'Task7 Client Name', 'Task7 Time',
            'Task8', 'Task8 Description', 'Task8 Client Name', 'Task8 Time',
            'Task9', 'Task9 Description', 'Task9 Client Name', 'Task9 Time',
            'Task10', 'Task10 Description', 'Task10 Client Name', 'Task10 Time'
        ]
        
        # Apply headers
        for cell, value in zip(headers, header_values):
            ws[cell] = value
        
        # Format headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Format A1 to AS1 (all columns from A to AS)
        for col in range(1, 46):  # Columns A to AS (1 to 45)
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Set column widths for first 5 columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        
        # Set widths for task columns (F to AS)
        # Pattern: Task (40), Description (60), Client Name (40), Time (15)
        # Columns: F, G, H, I (Task1), J, K, L, M (Task2), etc.
        for i in range(1, 11):  # 10 tasks
            # Task column (e.g., F, J, N, R, V, Z, AA, AD, AH, AL, AP)
            task_col = 5 + (i - 1) * 4 + 1  # Column F=6, J=10, N=14, etc.
            ws.column_dimensions[ws.cell(row=1, column=task_col).column_letter].width = 40
            
            # Description column (G, K, O, S, W, AA, AE, AI, AM, AQ)
            desc_col = task_col + 1
            ws.column_dimensions[ws.cell(row=1, column=desc_col).column_letter].width = 60
            
            # Client Name column (H, L, P, T, X, AB, AF, AJ, AN, AR)
            client_col = task_col + 2
            ws.column_dimensions[ws.cell(row=1, column=client_col).column_letter].width = 40
            
            # Time column (I, M, Q, U, Y, AC, AG, AK, AO, AS)
            time_col = task_col + 3
            ws.column_dimensions[ws.cell(row=1, column=time_col).column_letter].width = 15
    
    # Calculate total time
    total_time = sum(float(task.get('time', 0)) for task in tasks_data if task.get('time'))
    
    # Get current timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Format date
    date_str = request.form.get('date', '')
    formatted_date = format_date_for_display(date_str)
    
    # Prepare row data
    row_data = [
        employee_name,                    # A - Employee Name
        formatted_date,                   # B - Date
        employee_team,                    # C - Employee Team
        total_time,                       # D - Total Time
        timestamp                          # E - TimeStamp
    ]
    
    # Add task data (up to 10 tasks)
    for i in range(10):
        task_data = tasks_data[i] if i < len(tasks_data) else {}
        row_data.extend([
            task_data.get('task', ''),                          # Task
            task_data.get('taskDescription', ''),               # Task Description
            task_data.get('clientName', ''),                    # Client Name
            task_data.get('time', '')                           # Time
        ])
    
    # Find next empty row
    next_row = ws.max_row + 1
    
    # Write data to row
    for col, value in enumerate(row_data, start=1):
        cell = ws.cell(row=next_row, column=col)
        cell.value = value
        cell.alignment = Alignment(vertical='center')
        
        # Apply border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    # Set row height
    ws.row_dimensions[next_row].height = 25
    
    # Save workbook and ensure it's properly closed
    try:
        wb.save(file_path)
        wb.close()
        print(f"✅ Work plan saved to: {file_path}")
        return True
    except Exception as e:
        # Ensure workbook is closed even if save fails
        try:
            wb.close()
        except:
            pass
        print(f"❌ Error saving work plan: {e}")
        raise

def check_existing_entry(employee_name, date_str):
    """Check if user already submitted a work plan for today"""
    file_path = get_file_path()
    
    if not os.path.exists(file_path):
        return False
    
    try:
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        
        # Check rows 2 onwards (row 1 is header)
        for row in range(2, ws.max_row + 1):
            row_employee_name = ws.cell(row=row, column=1).value
            row_date = ws.cell(row=row, column=2).value
            
            # Format the date for comparison
            try:
                formatted_date = format_date_for_display(date_str)
                # Check if same employee and same date
                if str(row_employee_name).strip().lower() == str(employee_name).strip().lower() and str(row_date) == formatted_date:
                    wb.close()
                    return True
            except:
                continue
        
        wb.close()
        return False
    except Exception as e:
        print(f"Error checking existing entry: {e}")
        try:
            wb.close()
        except:
            pass
        return False

@everyday_workplan_bp.route('/check_work_plan_status', methods=['POST'])
def check_work_plan_status():
    """Check if user has already submitted work plan for today"""
    try:
        date_str = request.form.get('date', '')
        employee_name = request.form.get('employeeName', '').strip()
        
        if not employee_name:
            return jsonify({'submitted': False, 'error': 'Employee name missing'}), 400
        
        submitted = check_existing_entry(employee_name, date_str)
        return jsonify({'submitted': submitted})
        
    except Exception as e:
        print(f"Error checking work plan status: {e}")
        return jsonify({'submitted': False, 'error': str(e)}), 500

@everyday_workplan_bp.route('/submit_work_plan', methods=['POST'])
def submit_work_plan():
    """Handle work plan submission"""
    try:
        print("\n" + "="*80)
        print("📋 Processing Work Plan Submission")
        print("="*80)
        
        # Get form data
        date_str = request.form.get('date', '')
        employee_name = request.form.get('employeeName', '').strip()
        employee_team = request.form.get('employeeTeam', '').strip()
        
        # Get tasks data
        tasks = request.form.getlist('task[]')
        task_descriptions = request.form.getlist('taskDescription[]')
        client_names = request.form.getlist('clientName[]')
        times = request.form.getlist('time[]')
        
        # Validate data
        if not employee_name or not employee_team:
            return jsonify({'success': False, 'error': 'Employee information is missing'}), 400
        
        if not tasks or len(tasks) == 0:
            return jsonify({'success': False, 'error': 'At least one task is required'}), 400
        
        # Check if user already submitted today
        if check_existing_entry(employee_name, date_str):
            return jsonify({'success': False, 'error': 'You have already submitted your work plan for today'}), 400
        
        # Prepare tasks data
        tasks_data = []
        for i in range(len(tasks)):
            if tasks[i].strip():  # Only add non-empty tasks
                tasks_data.append({
                    'task': tasks[i].strip(),
                    'taskDescription': task_descriptions[i].strip() if i < len(task_descriptions) else '',
                    'clientName': client_names[i].strip() if i < len(client_names) else '',
                    'time': times[i] if i < len(times) else '0'
                })
        
        print(f"📅 Date: {date_str}")
        print(f"👤 Employee: {employee_name}")
        print(f"👥 Team: {employee_team}")
        print(f"📝 Tasks: {len(tasks_data)}")
        
        # Create or update Excel file
        create_workplan_excel(employee_name, employee_team, tasks_data)
        
        print("✅ Work plan submitted successfully!")
        
        return jsonify({
            'success': True,
            'message': 'Work plan submitted successfully!',
            'filename': generate_filename()
        })
        
    except Exception as e:
        print(f"❌ Error submitting work plan: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }), 500

