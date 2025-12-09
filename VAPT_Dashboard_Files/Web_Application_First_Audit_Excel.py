# Web Application VAPT First Audit Excel Report Generator
from flask import Blueprint, request, send_file, make_response, jsonify
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
import zipfile
import os
import re
import shutil

# Create a Blueprint for Web Application VAPT routes
web_app_vapt_bp = Blueprint('web_app_vapt', __name__)

def generate_web_app_filename(organization, end_date):
    """
    Generate static filename for Web Application VAPT First Audit.
    Format: Web_Application_VAPT_First_Audit
    """
    return "Web_Application_VAPT_First_Audit_Report.xlsx"

def normalize_name_for_matching(name):
    """
    Normalize name for matching by:
    1. Replacing special characters with '-'
    2. Trimming to first 170 characters
    3. Removing trailing numbers (e.g., 'abc 1' -> 'abc', 'abc_1' -> 'abc', but 'ab1c' stays 'ab1c')
    """
    if not name:
        return ""
    
    # Replace special characters with '-'
    normalized = re.sub(r'[<>\:"/\\|?*]', '-', str(name))
    
    # Remove trailing whitespace
    normalized = normalized.strip()
    
    # Remove trailing numbers (only at the end of the name)
    # This regex removes patterns like " 1", "_1", "_123" etc. at the end
    normalized = re.sub(r'[\s_\-]+\d+$', '', normalized)
    normalized = re.sub(r'^\d+[\s_\-]', '', normalized)  # Also handle at start if needed
    
    # Take only first 170 characters
    normalized = normalized[:170]
    
    return normalized.lower()

def extract_poc_images(evidence_files):
    """Extract POC images from evidence files zip and return mapping of vulnerability names to image paths"""
    poc_mapping = {}
    temp_dir = None
    
    if not evidence_files or evidence_files.filename == '':
        return poc_mapping, temp_dir
    
    try:
        from io import BytesIO
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        temp_dir = f"temp_poc_images_{timestamp}"
        
        # Read the zip file
        zip_data = evidence_files.read()
        
        with zipfile.ZipFile(BytesIO(zip_data), 'r') as zip_ref:
            file_list = zip_ref.namelist()
            
            # Extract ALL images from anywhere in the zip file
            image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff']
            
            print(f"Processing {len(file_list)} files in ZIP")
            
            for file_path in file_list:
                # Skip directories
                if file_path.endswith('/'):
                    continue
                
                # Get filename without extension
                filename = os.path.basename(file_path)
                name_without_ext = os.path.splitext(filename)[0]
                
                # Check if it's an image file
                is_image = any(filename.lower().endswith(ext) for ext in image_extensions)
                
                if is_image:
                    print(f"  Found image: {file_path}")
                    
                    # Extract the file to a temporary location
                    try:
                        if temp_dir is None:
                            temp_dir = f"temp_poc_images_{timestamp}"
                        os.makedirs(temp_dir, exist_ok=True)
                        with zip_ref.open(file_path) as f:
                            image_data = f.read()
                        
                        temp_file_path = os.path.join(temp_dir, filename)
                        
                        with open(temp_file_path, 'wb') as temp_file:
                            temp_file.write(image_data)
                        
                        # Map vulnerability name to image path
                        poc_mapping[name_without_ext] = temp_file_path
                        print(f"    ✅ Mapped: {name_without_ext} -> {filename}")
                        
                    except Exception as e:
                        print(f"    ❌ Error extracting image {filename}: {e}")
                        continue
            
            print(f"Extracted {len(poc_mapping)} POC images from ZIP file")
            
    except Exception as e:
        print(f"Error extracting POC images: {e}")
    
    return poc_mapping, temp_dir

def insert_poc_images_to_excel(excel_path, poc_mapping, vulnerabilities_data):
    """Insert POC images directly into Excel using openpyxl with 30x30 pixel size
    Returns a set of row numbers that have POC objects for border formatting"""
    rows_with_objects = set()  # Track which rows have POC objects
    
    try:
        # Load the existing workbook
        wb = load_workbook(excel_path)
        
        # Get the Web Application VAPT worksheet
        if "Web Application VAPT" not in wb.sheetnames:
            print("Web Application VAPT worksheet not found")
            return rows_with_objects
        
        ws = wb["Web Application VAPT"]
        
        # Find POC columns (L to R, which is columns 12 to 18)
        poc_col_start = 12  # Column L
        poc_col_end = 18   # Column R
        
        # Define the column order for image insertion: M, N, O, P, Q, R, L
        image_columns = [13, 14, 15, 16, 17, 18, 12]  # M, N, O, P, Q, R, L
        
        print(f"POC columns from column {poc_col_start} to {poc_col_end}")
        
        # Process each row and match vulnerabilities with POC images
        for row in range(2, ws.max_row + 1):
            vulnerabilities_cell = ws.cell(row=row, column=2)  # Vulnerabilities column (column B)
            vulnerabilities_text = str(vulnerabilities_cell.value) if vulnerabilities_cell.value else ""
            
            if vulnerabilities_text:
                # Split vulnerabilities text into individual entries (supports comma or newline separated)
                vuln_entries = re.split(r'[,\n]+', vulnerabilities_text)
                
                # Find matching POC images for each vulnerability entry
                matching_images = []
                for raw_entry in vuln_entries:
                    vuln_name = raw_entry.strip()
                    if not vuln_name:
                        continue
                    
                    normalized_vuln = normalize_name_for_matching(vuln_name.rstrip())
                    normalized_vuln_base = re.sub(r'[\s_\-]*\d+$', '', normalized_vuln)
                    
                    for image_name, image_path in poc_mapping.items():
                        normalized_image = normalize_name_for_matching(image_name.rstrip())
                        normalized_image_base = re.sub(r'[\s_\-]*\d+$', '', normalized_image)
                        
                        if normalized_vuln_base and normalized_image_base == normalized_vuln_base:
                            if image_path not in [img[1] for img in matching_images]:
                                matching_images.append((vuln_name, image_path))
                
                if matching_images:
                    try:
                        # Distribute images across columns: M, N, O, P, Q, R, L
                        num_images_to_insert = min(len(matching_images), 7)  # Max 7 images (one per column)
                        
                        for img_idx in range(num_images_to_insert):
                            vuln, matching_image = matching_images[img_idx]
                            col_idx = image_columns[img_idx]
                            
                            if os.path.exists(matching_image):
                                try:
                                    # Load the image
                                    img = Image(matching_image)
                                    
                                    # Get original dimensions
                                    original_width = img.width
                                    original_height = img.height
                                    
                                    # Resize image by reducing to 1/30th (30x reduction)
                                    if hasattr(img, 'width') and hasattr(img, 'height'):
                                        img.width = img.width / 30
                                        img.height = img.height / 30
                                    
                                    # Get cell reference (e.g., "M2", "N3", etc.)
                                    from openpyxl.utils import get_column_letter
                                    col_letter = get_column_letter(col_idx)
                                    cell_ref = f"{col_letter}{row}"
                                    
                                    # Insert image at the cell
                                    ws.add_image(img, cell_ref)
                                    
                                    print(f"✅ Inserted image {img_idx + 1} at {cell_ref} for vulnerability: {vuln} (reduced from {original_width}x{original_height})")
                                    
                                except Exception as e:
                                    print(f"⚠️ Error inserting image at column {col_idx}, row {row}: {e}")
                        
                        # Track this row as having POC objects
                        if num_images_to_insert > 0:
                            rows_with_objects.add(row)
                            
                    except Exception as e:
                        print(f"Error adding images for row {row}: {e}")
        
        # Save the workbook
        wb.save(excel_path)
        
    except Exception as e:
        print(f"Error inserting POC images: {e}")
        import traceback
        traceback.print_exc()
    
    return rows_with_objects

def create_web_app_metadata_worksheet(wb, form_data):
    """
    Create Meta_Data worksheet - structured format same as other VAPT files.
    """
    try:
        print("📊 Creating Meta_Data worksheet...")
        ws = wb.create_sheet("Meta_Data")
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 35  # Field names
        ws.column_dimensions['B'].width = 60  # Values
        
        # Define the data structure for the metadata
        metadata_sections = [
            {
                'title': 'ORGANIZATION INFORMATION',
                'data': [
                    ('Organization Name', form_data.get('organization', '')),
                    ('City', form_data.get('city', '')),
                    ('State', form_data.get('state', ''))
                ]
            },
            {
                'title': 'AUDIT PERIOD',
                'data': [
                    ('Start Date', form_data.get('startDate', '')),
                    ('End Date', form_data.get('endDate', ''))
                ]
            },
            {
                'title': 'APPLICATION TYPE',
                'data': [
                    ('Type', form_data.get('applicationType', 'Internal'))
                ]
            },
            {
                'title': 'REPORT PREPARED BY',
                'data': [
                    ('Name', f"{form_data.get('preparedByTitle', '')} {form_data.get('preparedByName', '')}".strip()),
                ]
            },
            {
                'title': 'AUDITEE DETAILS',
                'data': [
                    ('Name', f"{form_data.get('auditeeTitle', '')} {form_data.get('auditeeName', '')}".strip()),
                    ('Designation', form_data.get('designation', ''))
                ]
            }
        ]
        
        # Add Email Addresses section
        bank_emails = form_data.get('bankEmails', [])
        if bank_emails:
            email_data = []
            for i, email in enumerate(bank_emails, 1):
                if email.strip():
                    email_data.append((f'Email {i}', email.strip()))
            
            if email_data:
                metadata_sections.append({
                    'title': 'ORGANIZATION EMAIL ADDRESSES',
                    'data': email_data
                })
        
        # Add Auditing Team sections
        team_names = form_data.get('teamNames', [])
        team_designations = form_data.get('teamDesignations', [])
        team_emails = form_data.get('teamEmails', [])
        team_qualifications = form_data.get('teamQualifications', [])
        team_certified = form_data.get('teamCertified', [])
        
        if team_names:
            for i in range(len(team_names)):
                if team_names[i].strip():
                    team_member_data = [
                        (f'Team Member {i+1} - Name', team_names[i].strip()),
                        (f'Team Member {i+1} - Designation', team_designations[i] if i < len(team_designations) else ''),
                        (f'Team Member {i+1} - Email', team_emails[i] if i < len(team_emails) else ''),
                        (f'Team Member {i+1} - Qualification', team_qualifications[i] if i < len(team_qualifications) else ''),
                        (f'Team Member {i+1} - Certified', 
                         'Yes' if team_certified[i].lower() == 'yes' else 'No' if team_certified[i].lower() == 'no' else team_certified[i] 
                         if i < len(team_certified) else '')
                    ]
                    
                    metadata_sections.append({
                        'title': f'AUDITING TEAM MEMBER {i+1}',
                        'data': team_member_data
                    })
        
        # Add Asset details sections
        assets = form_data.get('assets', [])
        application_type = form_data.get('applicationType', 'Internal')
        ip_field_label = 'External IP Address' if str(application_type).strip().lower() == 'external' else 'Internal IP Addresses'
        if assets:
            for i, asset in enumerate(assets):
                asset_data = [
                    (f'Asset {i+1} - Description', asset.get('description', '')),
                    (f'Asset {i+1} - Criticality', asset.get('criticality', '')),
                    (f'Asset {i+1} - {ip_field_label}', asset.get('ip', '')),
                    (f'Asset {i+1} - URL', asset.get('url', '')),
                    (f'Asset {i+1} - Version', asset.get('version', ''))
                ]
                
                metadata_sections.append({
                    'title': f'ASSET {i+1}',
                    'data': asset_data
                })
        
        # Write data to worksheet
        row = 1
        
        for section in metadata_sections:
            # Write section title (merged across both columns)
            cell = ws.cell(row=row, column=1, value=section['title'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=row, column=2, value='')
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            ws.row_dimensions[row].height = 25
            row += 1
            
            # Write section data
            for field_name, field_value in section['data']:
                cell = ws.cell(row=row, column=1, value=field_name)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                cell = ws.cell(row=row, column=2, value=field_value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                ws.row_dimensions[row].height = 20
                row += 1
            
            # Add empty row after each section for spacing
            row += 1
        
        print(f"✅ Meta_Data worksheet created with {row} rows")
        
    except Exception as e:
        print(f"❌ Error creating Meta_Data worksheet: {e}")
        import traceback
        traceback.print_exc()

def create_web_app_vapt_worksheet(wb, vulnerabilities_data):
    """
    Create Web Application VAPT worksheet with headers and vulnerability data.
    
    Args:
        wb: openpyxl Workbook object
        vulnerabilities_data: List of dictionaries containing vulnerability information
    """
    try:
        print("📊 Creating Web Application VAPT worksheet...")
        
        # Create worksheet
        ws = wb.create_sheet("Web Application VAPT")
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=10)
        data_alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        thin_side = Side(style='thin')
        poc_left_border = Border(left=thin_side, top=thin_side, bottom=thin_side)
        poc_middle_border = Border(top=thin_side, bottom=thin_side)
        poc_right_border = Border(right=thin_side, top=thin_side, bottom=thin_side)

        # Define headers
        headers = [
            "Sr No",                          # A
            "Vulnerabilities",                 # B
            "Name of Vulnerability",           # C
            "Risk Factor",                     # D
            "CVE/CWE ID",                      # E
            "CVSS",                            # F
            "Audit Observation",               # G
            "Impact",                          # H
            "Recommendation / Countermeasure", # I
            "Reference Link",                  # J
            "Affected URL",                    # K
            "POC"                             # L-R (merged)
        ]
        
        # Set column widths
        ws.column_dimensions['A'].width = 8   # Sr No
        ws.column_dimensions['B'].width = 50  # Vulnerabilities
        ws.column_dimensions['C'].width = 35  # Name of Vulnerability
        ws.column_dimensions['D'].width = 15  # Risk Factor
        ws.column_dimensions['E'].width = 15  # CVE/CWE ID
        ws.column_dimensions['F'].width = 10  # CVSS
        ws.column_dimensions['G'].width = 40  # Audit Observation
        ws.column_dimensions['H'].width = 40  # Impact
        ws.column_dimensions['I'].width = 40  # Recommendation / Countermeasure
        ws.column_dimensions['J'].width = 30  # Reference Link
        ws.column_dimensions['K'].width = 30  # Affected URL
        # POC columns (L to R) - width set to 20 each
        for col in range(12, 19):  # Columns L to R
            col_letter = chr(64 + col)
            ws.column_dimensions[col_letter].width = 20
        
        # Write headers for columns A to K
        for col_idx, header in enumerate(headers[:11], start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Merge POC header cells (L to R, which is columns 12 to 18)
        ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=18)
        poc_cell = ws.cell(row=1, column=12, value="POC")
        poc_cell.font = header_font
        poc_cell.fill = header_fill
        poc_cell.alignment = header_alignment
        poc_cell.border = thin_border
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        # Check if there are vulnerabilities or if it's 0 Vulnerabilities case
        if not vulnerabilities_data:
            # Merge A3-R3 and add note
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=18)
            note_cell = ws.cell(row=3, column=1, value="Note: No vulnerabilities were identified by the auditor during the audit.")
            note_cell.font = Font(name='Arial', size=12, bold=True, color='FF008000')  # Green color
            note_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            note_cell.border = thin_border
            # Apply border to all cells in the merged range to ensure complete border
            for col in range(1, 19):  # Columns A to R (1 to 18)
                cell = ws.cell(row=3, column=col)
                cell.border = thin_border
            ws.row_dimensions[3].height = 50
            print("✅ Added 'No vulnerabilities' note to worksheet")
        else:
            # Sort vulnerabilities by risk factor: Critical, High, Medium, Low, then others
            risk_priority = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
            sorted_vulnerabilities = sorted(
                vulnerabilities_data,
                key=lambda item: risk_priority.get(str(item.get('riskFactor', '')).strip().lower(), 4)
            )

            # Write vulnerability data (set row height to 25 for all data rows)
            row_num = 2
            for idx, vuln_data in enumerate(sorted_vulnerabilities, start=1):
                # Sr No
                ws.cell(row=row_num, column=1, value=idx).border = thin_border
                ws.cell(row=row_num, column=1).font = data_font
                ws.cell(row=row_num, column=1).alignment = data_alignment_center
                
                # Vulnerabilities (Column B)
                ws.cell(row=row_num, column=2, value=vuln_data.get('vulnerabilities', '')).border = thin_border
                ws.cell(row=row_num, column=2).font = data_font
                ws.cell(row=row_num, column=2).alignment = data_alignment_left
                
                # Name of Vulnerability (Column C)
                ws.cell(row=row_num, column=3, value=vuln_data.get('name', '')).border = thin_border
                ws.cell(row=row_num, column=3).font = data_font
                ws.cell(row=row_num, column=3).alignment = data_alignment_left
                
                # Risk Factor
                ws.cell(row=row_num, column=4, value=vuln_data.get('riskFactor', '')).border = thin_border
                ws.cell(row=row_num, column=4).font = data_font
                ws.cell(row=row_num, column=4).alignment = data_alignment_center
                
                # CVE/CWE ID
                ws.cell(row=row_num, column=5, value=vuln_data.get('cve', '')).border = thin_border
                ws.cell(row=row_num, column=5).font = data_font
                ws.cell(row=row_num, column=5).alignment = data_alignment_center
                
                # CVSS
                ws.cell(row=row_num, column=6, value=vuln_data.get('cvss', '')).border = thin_border
                ws.cell(row=row_num, column=6).font = data_font
                ws.cell(row=row_num, column=6).alignment = data_alignment_center
                
                # Audit Observation
                ws.cell(row=row_num, column=7, value=vuln_data.get('observation', '')).border = thin_border
                ws.cell(row=row_num, column=7).font = data_font
                ws.cell(row=row_num, column=7).alignment = data_alignment_left
                
                # Impact
                ws.cell(row=row_num, column=8, value=vuln_data.get('impact', '')).border = thin_border
                ws.cell(row=row_num, column=8).font = data_font
                ws.cell(row=row_num, column=8).alignment = data_alignment_left
                
                # Recommendation / Countermeasure
                ws.cell(row=row_num, column=9, value=vuln_data.get('recommendation', '')).border = thin_border
                ws.cell(row=row_num, column=9).font = data_font
                ws.cell(row=row_num, column=9).alignment = data_alignment_left
                
                # Reference Link
                ws.cell(row=row_num, column=10, value=vuln_data.get('reference', '')).border = thin_border
                ws.cell(row=row_num, column=10).font = data_font
                ws.cell(row=row_num, column=10).alignment = data_alignment_left
                
                # Affected URL
                ws.cell(row=row_num, column=11, value=vuln_data.get('affectedSystem', '')).border = thin_border
                ws.cell(row=row_num, column=11).font = data_font
                ws.cell(row=row_num, column=11).alignment = data_alignment_left
                
                # POC columns (L to R) - leave empty for now
                for col in range(12, 19):
                    cell = ws.cell(row=row_num, column=col)
                    if col == 12:
                        cell.border = poc_left_border
                    elif col == 18:
                        cell.border = poc_right_border
                    else:
                        cell.border = poc_middle_border
                
                # Set row height to 25px
                ws.row_dimensions[row_num].height = 25
                
                row_num += 1

            # Increase row heights for populated rows in column A by 70
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value not in (None, ''):
                    current_height = ws.row_dimensions[row].height
                    if current_height is None:
                        current_height = 15
                    ws.row_dimensions[row].height = current_height + 70

            # Apply risk factor coloring (column D)
            risk_color_map = {
                'high': 'FFFF0000',      # Red
                'critical': 'FF8B0000',  # Dark Red
                'medium': 'FFFFA500',    # Orange
                'low': 'FF008000'        # Green
            }
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=4)  # Column D
                if cell.value:
                    risk_value = str(cell.value).strip().lower()
                    if risk_value in risk_color_map:
                        fill_color = risk_color_map[risk_value]
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                        cell.font = Font(
                            name=data_font.name,
                            size=data_font.size,
                            bold=True,
                            color='FFFFFFFF'
                        )
                        cell.alignment = data_alignment_center
        
        print(f"✅ Created Web Application VAPT worksheet with {len(vulnerabilities_data)} vulnerabilities")
        
    except Exception as e:
        print(f"❌ Error creating Web Application VAPT worksheet: {e}")
        import traceback
        traceback.print_exc()

@web_app_vapt_bp.route('/process_web_app_vapt_first_audit_report', methods=['POST'])
def process_web_app_first_audit_report():
    """Main endpoint to process Web Application VAPT First Audit Report."""
    try:
        print("\n" + "="*80)
        print("🚀 Processing Web Application VAPT First Audit Report")
        print("="*80)
        
        # Extract form data
        form_data = {
            'organization': request.form.get('organization'),
            'otherOrganization': request.form.get('otherOrganization'),
            'city': request.form.get('city'),
            'otherCity': request.form.get('otherCity'),
            'state': request.form.get('state'),
            'startDate': request.form.get('startDate'),
            'endDate': request.form.get('endDate'),
            'preparedByTitle': request.form.get('preparedByTitle'),
            'preparedByName': request.form.get('preparedByName'),
            'auditeeTitle': request.form.get('auditeeTitle'),
            'auditeeName': request.form.get('auditeeName'),
            'designation': request.form.get('designation'),
            'applicationType': request.form.get('applicationType', 'Internal')
        }
        
        # Handle "Other" selections
        if form_data['organization'] == 'other':
            form_data['organization'] = form_data['otherOrganization']
        if form_data['city'] == 'other':
            form_data['city'] = form_data['otherCity']
        
        # Extract email addresses
        bank_emails = request.form.getlist('bankEmail[]')
        form_data['bankEmails'] = [email for email in bank_emails if email.strip()]
        
        # Extract team member details
        team_names = request.form.getlist('teamName[]')
        team_designations = request.form.getlist('teamDesignation[]')
        team_emails = request.form.getlist('teamEmail[]')
        team_qualifications = request.form.getlist('teamQualification[]')
        
        # Extract team certified values
        team_certified = []
        for i in range(len(team_names)):
            certified_value = request.form.get(f'teamCertified[{i}]', 'no')
            team_certified.append(certified_value)
        
        form_data['teamNames'] = team_names
        form_data['teamDesignations'] = team_designations
        form_data['teamEmails'] = team_emails
        form_data['teamQualifications'] = team_qualifications
        form_data['teamCertified'] = team_certified
        
        # Extract asset details
        asset_descriptions = request.form.getlist('assetDescription[]')
        asset_criticality = request.form.getlist('assetCriticality[]')
        asset_internal_ips = request.form.getlist('assetInternalIp[]')
        asset_urls = request.form.getlist('assetUrl[]')
        asset_versions = request.form.getlist('assetVersion[]')
        
        assets = []
        for i in range(len(asset_descriptions)):
            description = asset_descriptions[i].strip() if i < len(asset_descriptions) and asset_descriptions[i] else ''
            if not description:
                continue
            criticality = asset_criticality[i].strip() if i < len(asset_criticality) and asset_criticality[i] else ''
            ip_value = asset_internal_ips[i].strip() if i < len(asset_internal_ips) and asset_internal_ips[i] else ''
            url = asset_urls[i].strip() if i < len(asset_urls) and asset_urls[i] else ''
            version = asset_versions[i].strip() if i < len(asset_versions) and asset_versions[i] else ''
            
            assets.append({
                'description': description,
                'criticality': criticality,
                'ip': ip_value,
                'url': url,
                'version': version
            })
        
        if not assets:
            return jsonify({"error": "At least one asset detail is required"}), 400
        
        form_data['assets'] = assets
        
        # Check vulnerability status
        vulnerability_status = request.form.get('webAppVulnerabilityStatus', 'has')
        
        # Extract vulnerability data
        vulnerabilities_data = []
        evidence_file = None
        
        if vulnerability_status == 'has':
            vuln_vulnerabilities = request.form.getlist('vulnVulnerabilities[]')
            vuln_names = request.form.getlist('vulnName[]')
            vuln_risk_factors = request.form.getlist('vulnRiskFactor[]')
            vuln_cves = request.form.getlist('vulnCVE[]')
            vuln_cvss = request.form.getlist('vulnCVSS[]')
            vuln_observations = request.form.getlist('vulnObservation[]')
            vuln_impacts = request.form.getlist('vulnImpact[]')
            vuln_recommendations = request.form.getlist('vulnRecommendation[]')
            vuln_references = request.form.getlist('vulnReference[]')
            vuln_affected_systems = request.form.getlist('vulnAffectedSystem[]')
            
            # Build vulnerabilities data list
            for i in range(len(vuln_names)):
                vulnerability = {
                    'vulnerabilities': vuln_vulnerabilities[i] if i < len(vuln_vulnerabilities) else '',
                    'name': vuln_names[i],
                    'riskFactor': vuln_risk_factors[i],
                    'cve': vuln_cves[i] if i < len(vuln_cves) else '',
                    'cvss': vuln_cvss[i] if i < len(vuln_cvss) else '',
                    'observation': vuln_observations[i],
                    'impact': vuln_impacts[i],
                    'recommendation': vuln_recommendations[i],
                    'reference': vuln_references[i] if i < len(vuln_references) else '',
                    'affectedSystem': vuln_affected_systems[i]
                }
                vulnerabilities_data.append(vulnerability)
            
            if not vulnerabilities_data:
                return jsonify({"error": "At least one vulnerability is required"}), 400
            
            print(f"📝 Processing {len(vulnerabilities_data)} vulnerabilities...")
            
            # Check if evidence file is provided
            evidence_file = request.files.get('evidenceFiles')
            if not evidence_file or evidence_file.filename == '':
                return jsonify({"error": "Evidence ZIP file is required"}), 400
        else:
            print("📝 No vulnerabilities identified (0 Vulnerabilities selected)")
        
        if evidence_file:
            print("📁 Evidence file received: ", evidence_file.filename)
        
        # Create Excel workbook
        print("📊 Creating Excel workbook...")
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create Web Application VAPT worksheet
        create_web_app_vapt_worksheet(wb, vulnerabilities_data)
        
        # Create Meta_Data worksheet
        create_web_app_metadata_worksheet(wb, form_data)
        
        # Save to temporary file first
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            wb.save(temp_file.name)
            temp_file_path = temp_file.name
        
        # Extract POC images from evidence file (only if vulnerabilities exist)
        temp_images_dir = None
        if evidence_file and vulnerabilities_data:
            poc_mapping, temp_images_dir = extract_poc_images(evidence_file)
            
            if poc_mapping:
                print(f"📁 Found {len(poc_mapping)} POC images to insert")
                # Insert POC images into the Excel file
                rows_with_objects = insert_poc_images_to_excel(temp_file_path, poc_mapping, vulnerabilities_data)
                print(f"✅ Inserted POC images into {len(rows_with_objects)} rows")
        
        # Read the final Excel file
        from io import BytesIO
        output = BytesIO()
        with open(temp_file_path, 'rb') as f:
            output.write(f.read())
        output.seek(0)
        
        # Clean up temporary file
        try:
            os.unlink(temp_file_path)
        except:
            pass
        
        # Clean up temporary images directory
        if temp_images_dir and os.path.exists(temp_images_dir):
            try:
                shutil.rmtree(temp_images_dir)
            except Exception as cleanup_error:
                print(f"⚠️ Unable to remove temp images directory {temp_images_dir}: {cleanup_error}")
        
        
        # Generate filename
        filename = generate_web_app_filename(
            form_data['organization'],
            form_data['endDate']
        )
        
        print(f"✅ Excel file created successfully: {filename}")
        
        # Create response
        response = make_response(send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        ))
        
        return response
        
    except Exception as e:
        print(f"❌ Error processing Web Application VAPT First Audit Report: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

