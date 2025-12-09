# Website VAPT First Audit Excel Report Generator
from flask import Blueprint, request, send_file, make_response, jsonify, session
import re
import io
import pandas as pd
import zipfile
import os
import math
from io import BytesIO
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import copy

# Create a Blueprint for Website VAPT routes
website_vapt_bp = Blueprint('website_vapt', __name__)

def convert_risk_to_camelcase(risk_value):
    """Convert risk value to CamelCase format"""
    if not risk_value:
        return ""
    
    risk_str = str(risk_value).strip().lower()
    if risk_str == 'critical':
        return 'Critical'
    elif risk_str == 'high':
        return 'High'
    elif risk_str == 'medium':
        return 'Medium'
    elif risk_str == 'low':
        return 'Low'
    else:
        # Return original value with first letter capitalized
        return risk_str.capitalize()

def generate_website_filename(organization, end_date):
    """
    Generate static filename for Website VAPT First Audit.
    Format: Website_VAPT_First_Audit
    """
    return "Website_VAPT_First_Audit_Report.xlsx"

def extract_poc_images(evidence_files, temp_dir=None):
    """Extract POC images from evidence files zip and return mapping of vulnerability names to image paths"""
    poc_mapping = {}
    
    if not evidence_files or evidence_files.filename == '':
        return poc_mapping
    
    try:
        import zipfile
        from io import BytesIO
        from datetime import datetime
        
        # Generate timestamp for temporary folder if not provided (format: DDMMYYYYHHMMSS)
        if temp_dir is None:
            current_ts = datetime.now().strftime('%d%m%Y%H%M%S')
            temp_dir = f"temp_poc_images_{current_ts}"
        
        # Read the zip file
        zip_data = evidence_files.read()
        
        with zipfile.ZipFile(BytesIO(zip_data), 'r') as zip_ref:
            file_list = zip_ref.namelist()
            
            # Find ALL POC folder
            poc_folder = None
            for file_path in file_list:
                if 'ALL POC' in file_path and file_path.endswith('/'):
                    poc_folder = file_path
                    break
            
            if not poc_folder:
                print("ALL POC folder not found in evidence files")
                return poc_mapping
            
            # Extract images from ALL POC folder
            image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff']
            
            for file_path in file_list:
                if file_path.startswith(poc_folder) and not file_path.endswith('/'):
                    # Get filename without extension
                    filename = os.path.basename(file_path)
                    name_without_ext = os.path.splitext(filename)[0]
                    
                    # Check if it's an image file
                    if any(filename.lower().endswith(ext) for ext in image_extensions):
                        # Extract the file to a temporary location
                        try:
                            with zip_ref.open(file_path) as f:
                                image_data = f.read()
                            
                            # Create temporary file
                            os.makedirs(temp_dir, exist_ok=True)
                            temp_file_path = os.path.join(temp_dir, filename)
                            
                            with open(temp_file_path, 'wb') as temp_file:
                                temp_file.write(image_data)
                            
                            # Map vulnerability name to image path
                            poc_mapping[name_without_ext] = temp_file_path
                            
                        except Exception as e:
                            print(f"Error extracting image {filename}: {e}")
                            continue
            
            print(f"Extracted {len(poc_mapping)} POC images")
            
    except Exception as e:
        print(f"Error extracting POC images: {e}")
    
    return poc_mapping


def normalize_vulnerability_name_for_filename(vuln_name):
    """
    Normalize vulnerability name to be compatible with file names by replacing
    invalid characters with dashes (same logic as Infra_VAPT_First_Audit_Excel.py).
    
    Args:
        vuln_name (str): Original vulnerability name
        
    Returns:
        str: Normalized vulnerability name safe for file names
    """
    if not vuln_name:
        return vuln_name
    
    import re
    
    # Characters that are not allowed in file names
    invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
    
    normalized_name = str(vuln_name)
    
    # Replace invalid characters with dashes
    for char in invalid_chars:
        normalized_name = normalized_name.replace(char, '-')
    
    # Remove multiple consecutive dashes and trim
    normalized_name = re.sub(r'-+', '-', normalized_name)
    normalized_name = normalized_name.strip('-')
    
    return normalized_name


def insert_poc_images_to_excel(excel_path, poc_mapping, vulnerabilities_data):
    """Insert POC images directly into Excel using openpyxl with reduced size
    Returns a set of row numbers that have POC objects for border formatting"""
    rows_with_objects = set()  # Track which rows have POC objects
    
    print(f"\n🖼️ === INSERTING POC IMAGES FOR WEBSITE_VAPT ===")
    print(f"📁 Excel path: {excel_path}")
    print(f"📊 POC mapping contains {len(poc_mapping)} image entries")
    
    if not poc_mapping:
        print("⚠️ No POC images to insert")
        return rows_with_objects
    
    try:
        # Load the workbook
        wb = load_workbook(excel_path)
        
        # Check if Website_VAPT worksheet exists
        if 'Website_VAPT' not in wb.sheetnames:
            print("⚠️ Website_VAPT worksheet not found - skipping POC image insertion")
            return rows_with_objects
        
        ws = wb['Website_VAPT']
        
        # Find POC columns by looking at the merged header in first row
        # We're looking for the merged "POC" header which spans L, M, N, O, P, Q, R columns
        poc_col_start = None
        poc_col_end = None
        
        # Check for merged cells in row 1
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == 1 and merged_range.max_row == 1:
                # Check if this merged cell contains "POC"
                first_cell = ws.cell(row=1, column=merged_range.min_col)
                if first_cell.value and str(first_cell.value).strip() == "POC":
                    poc_col_start = merged_range.min_col
                    poc_col_end = merged_range.max_col
                    break
        
        if not poc_col_start or not poc_col_end:
            print("⚠️ POC columns not found in Website_VAPT worksheet")
            wb.save(excel_path)
            return rows_with_objects
        
        print(f"📍 POC columns found: {get_column_letter(poc_col_start)} to {get_column_letter(poc_col_end)}")
        
        # Define the column order for image insertion: M, N, O, P, Q, R, L
        # POC columns are: L, M, N, O, P, Q, R (7 columns total)
        # Insertion order: M, N, O, P, Q, R, L
        image_columns = [
            poc_col_start + 1,  # M
            poc_col_start + 2,  # N
            poc_col_start + 3,  # O
            poc_col_start + 4,  # P
            poc_col_start + 5,  # Q
            poc_col_start + 6,  # R (should be poc_col_end)
            poc_col_start       # L
        ]
        
        # Process each row and match vulnerabilities with POC images
        for row in range(2, ws.max_row + 1):
            vulnerability_cell = ws.cell(row=row, column=2)  # Vulnerabilities column (column B)
            vulnerability_text = str(vulnerability_cell.value) if vulnerability_cell.value else ""
            
            if vulnerability_text:
                # Split vulnerabilities (they might be on separate lines)
                vulnerabilities = [v.strip() for v in vulnerability_text.split('\n') if v.strip()]
                
                # Find all matching POC images for this row
                matching_images = []
                for vuln in vulnerabilities:
                    # Normalize the vulnerability name for filename matching
                    normalized_vuln = normalize_vulnerability_name_for_filename(vuln)
                    
                    # Use only first 170 characters for matching (to handle long vulnerability names)
                    vuln_short = vuln[:170].lower()
                    normalized_vuln_short = normalized_vuln[:170].lower()
                    
                    for image_name, image_path in poc_mapping.items():
                        # Use only first 170 characters of image name for matching
                        image_name_short = image_name[:170].lower()
                        
                        # Check if vulnerability name matches image name (case-insensitive, first 170 chars only)
                        if (vuln_short in image_name_short or image_name_short in vuln_short or
                            normalized_vuln_short == image_name_short):
                            if image_path not in [img[1] for img in matching_images]:  # Avoid duplicates
                                matching_images.append((vuln, image_path))
                
                if matching_images:
                    try:
                        # Distribute images across columns: M, N, O, P, Q, R, L
                        num_images_to_insert = min(len(matching_images), 7)  # Max 7 images (one per column)
                        
                        for img_idx in range(num_images_to_insert):
                            vuln, matching_image = matching_images[img_idx]
                            col_idx = image_columns[img_idx]
                            
                            if os.path.exists(matching_image):
                                try:
                                    # Load the image
                                    img = Image(matching_image)
                                    
                                    # Get original dimensions
                                    original_width = img.width
                                    original_height = img.height
                                    
                                    # Resize image to height=30px, width proportionally (maintain aspect ratio)
                                    if hasattr(img, 'width') and hasattr(img, 'height'):
                                        target_height = 30
                                        aspect_ratio = img.width / img.height if img.height > 0 else 1
                                        img.height = target_height
                                        img.width = int(target_height * aspect_ratio)
                                    
                                    # Get cell reference (e.g., "M2", "N3", etc.)
                                    col_letter = get_column_letter(col_idx)
                                    cell_ref = f"{col_letter}{row}"
                                    
                                    # Insert image at the cell
                                    ws.add_image(img, cell_ref)
                                    
                                    print(f"✅ Inserted image {img_idx + 1} at {cell_ref} for vulnerability: {vuln} (original {original_width}x{original_height} → resized to {img.width}x{img.height})")
                                    
                                except Exception as e:
                                    print(f"⚠️ Error inserting image at column {col_idx}, row {row}: {e}")
                        
                        # Track this row as having POC objects
                        if num_images_to_insert > 0:
                            rows_with_objects.add(row)
                            
                    except Exception as e:
                        print(f"Error adding images for row {row}: {e}")
                        continue
        
        # Apply custom borders to ALL POC columns (L, M, N, O, P, Q, R)
        from openpyxl.styles import Border, Side
        
        # Define border styles for each POC column type
        # L column: left, above, below (NOT right)
        left_border = Border(
            left=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # M, N, O, P, Q columns: above and below only (NOT left or right)
        middle_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # R column: above, below, right (NOT left)
        right_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            right=Side(style='thin')
        )
        
        # Find all rows that are part of the data table (have content in any column)
        table_rows = set()
        
        # Always include header row
        table_rows.add(1)
        
        # Check all rows from 2 onwards to find data rows
        for row in range(2, ws.max_row + 1):
            has_content = False
            
            # Check if any cell in this row has content (excluding POC columns)
            for col in range(1, poc_col_start):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip() != "":
                    has_content = True
                    break
            
            if has_content:
                table_rows.add(row)
        
        # Apply custom borders to all table rows for each POC column
        for row_num in sorted(table_rows):
            # L column (poc_col_start): left, top, bottom
            ws.cell(row=row_num, column=poc_col_start).border = left_border
            
            # M, N, O, P, Q columns (middle columns): top, bottom only
            for col_idx in range(poc_col_start + 1, poc_col_end):
                ws.cell(row=row_num, column=col_idx).border = middle_border
            
            # R column (poc_col_end): top, bottom, right
            ws.cell(row=row_num, column=poc_col_end).border = right_border
        
        # Set row height to ~50px (≈37.5 points) for rows with images to fit 30px images comfortably
        for row_num in rows_with_objects:
            # Row height in Excel: 1 point ≈ 1.33 pixels, so 50px ≈ 37.5 points
            ws.row_dimensions[row_num].height = 37.5
        
        # Save the workbook
        wb.save(excel_path)
        print(f"✅ Successfully inserted POC images into Website_VAPT worksheet")
        print(f"Added images to {len(rows_with_objects)} rows, distributed across columns M, N, O, P, Q, R, L")
        print(f"Applied custom borders to POC columns for {len(table_rows)} rows")
        
    except Exception as e:
        print(f"❌ Error inserting POC images: {e}")
        import traceback
        traceback.print_exc()
    
    return rows_with_objects

def process_nmap_zip_website(nmap_file):
    """
    Process Nmap ZIP file for Website VAPT - Same logic as Infrastructure VAPT
    Returns: Dictionary with IP addresses as keys and list of (port, service) tuples as values
    Output format: Grouped by IP for merging HOST column
    """
    try:
        print("📁 Processing Nmap ZIP file for Website VAPT...")
        ip_ports = {}
        
        zip_data = nmap_file.read()
        
        with zipfile.ZipFile(BytesIO(zip_data), 'r') as zip_ref:
            file_list = zip_ref.namelist()
            
            for file_name in file_list:
                if file_name.endswith('/'):
                    continue
                
                try:
                    with zip_ref.open(file_name) as f:
                        file_content = f.read().decode('utf-8', errors='ignore')
                    
                    # Regex patterns (same as Infrastructure VAPT)
                    ip_pattern = r"Nmap scan report for (?:[a-zA-Z0-9.-]+ )?\(?(\d+\.\d+\.\d+\.\d+)\)?"
                    port_state_pattern = r"(\d+)/(tcp|udp)\s+(open|filtered|closed|unfiltered)\s+([\w-]*)"
                    filtered_ports_pattern = r"Not shown: (\d+) filtered tcp ports"
                    
                    lines = file_content.splitlines()
                    current_ip = None
                    has_filtered_ports = False

                    for line in lines:
                        ip_match = re.search(ip_pattern, line, re.IGNORECASE)
                        if ip_match:
                            current_ip = ip_match.group(1)
                            if current_ip not in ip_ports:
                                ip_ports[current_ip] = []
                            has_filtered_ports = False
                            continue
                        
                        if not current_ip:
                            continue
                        
                        # Check for filtered ports message
                        filtered_match = re.search(filtered_ports_pattern, line)
                        if filtered_match:
                            has_filtered_ports = True
                            continue
                        
                        port_match = re.search(port_state_pattern, line)
                        if port_match:
                            port = port_match.group(1)
                            state = port_match.group(3)
                            service = port_match.group(4) or state
                            
                            if (port, service) not in ip_ports[current_ip]:
                                ip_ports[current_ip].append((port, service))
                    
                    # If no open ports found but filtered ports detected, add "Filtered" entry
                    if current_ip and has_filtered_ports and len(ip_ports[current_ip]) == 0:
                        ip_ports[current_ip].append(("Filtered", "Filtered"))

                    # Also process CSV files if present
                    if file_name.endswith('.csv'):
                        try:
                            with zip_ref.open(file_name) as f:
                                csv_content = f.read()
                            
                            df = pd.read_csv(BytesIO(csv_content), 
                                        on_bad_lines="skip", 
                                        encoding="utf-8")
                            
                            # Check for columns (case-insensitive)
                            df.columns = df.columns.str.lower().str.strip()
                            
                            if all(col in df.columns for col in ['host', 'port', 'service']):
                                for _, row in df.iterrows():
                                    ip = str(row['host']).strip()
                                    port = str(row['port']).strip()
                                    service = str(row['service']).strip()
                                    
                                    if ip and port and service:
                                        if ip not in ip_ports:
                                            ip_ports[ip] = []
                                        if (port, service) not in ip_ports[ip]:
                                            ip_ports[ip].append((port, service))
                        except Exception as e:
                            print(f"CSV processing error in {file_name}: {e}")
                            continue

                except Exception as e:
                    print(f"Error processing file {file_name}: {e}")
                    continue
        
        print(f"✅ Processed {len(ip_ports)} IP addresses with ports")
        return ip_ports
        
    except Exception as e:
        print(f"❌ Error processing Nmap ZIP: {e}")
        import traceback
        traceback.print_exc()
        return {}


def process_nessus_zip_website(nessus_file):
    """
    Process Nessus ZIP file for Website VAPT - direct copy of data
    Returns: List of DataFrames (one per CSV file)
    """
    try:
        print("📁 Processing Nessus ZIP file for Website VAPT...")
        nessus_dataframes = []
        
        with zipfile.ZipFile(nessus_file, 'r') as zip_ref:
            csv_files = [f for f in zip_ref.namelist() if f.endswith('.csv')]
            
            if not csv_files:
                print("⚠️ No CSV files found in Nessus ZIP")
                return nessus_dataframes
            
            print(f"📄 Found {len(csv_files)} Nessus CSV files")
            
            for csv_file in csv_files:
                try:
                    with zip_ref.open(csv_file) as file:
                        df = pd.read_csv(file)
                        
                        # Just store the dataframe as-is for direct copy
                        nessus_dataframes.append(df)
                        print(f"✅ Loaded {csv_file}: {len(df)} rows")
                        
                except Exception as e:
                    print(f"⚠️ Error processing {csv_file}: {e}")
                    continue
        
        print(f"✅ Processed {len(nessus_dataframes)} Nessus files")
        return nessus_dataframes
        
    except Exception as e:
        print(f"❌ Error processing Nessus ZIP: {e}")
        import traceback
        traceback.print_exc()
        return []


@website_vapt_bp.route('/check_website_vapt_vulnerabilities', methods=['POST'])
def check_website_vapt_vulnerabilities():
    """Return both matched and unmatched vulnerabilities using Website VAPT Catalog."""
    if 'nmapFiles' not in request.files or 'nessusFiles' not in request.files:
        return jsonify({"error": "Both Nmap and Nessus files are required"}), 400
    
    nmap_file = request.files['nmapFiles']
    nessus_file = request.files['nessusFiles']
    
    if nmap_file.filename == '' or nessus_file.filename == '':
        return jsonify({"error": "No file selected"}), 400
    
    try:
        # Process Nessus files to check vulnerabilities
        nessus_dataframes = process_nessus_zip_website(nessus_file)
        
        if nessus_dataframes:
            combined_nessus = pd.concat(nessus_dataframes, ignore_index=True)
            
            # Calculate matched and unmatched vulnerabilities
            matched_groups = []
            unmatched_vulnerabilities = []
            
            try:
                # Filter by valid risks and normalize
                valid_risks = ['low', 'medium', 'high', 'critical']
                df_filtered = combined_nessus.copy()
                df_filtered['Risk'] = df_filtered['Risk'].astype(str).str.lower().str.strip()
                df_filtered = df_filtered[df_filtered['Risk'].isin(valid_risks)]

                # Get unique vulnerability names from Name column
                unique_vulnerabilities_list = df_filtered['Name'].dropna().drop_duplicates().astype(str).str.strip().tolist()
                unique_vulnerabilities = set(unique_vulnerabilities_list)
                
                # Load Website VAPT Catalog
                catalog_path = "static/Formats_and_Catalog/Website VAPT Catalog.xlsx"
                if os.path.exists(catalog_path):
                    try:
                        catalog_df = pd.read_excel(catalog_path, sheet_name=0)
                    except Exception as e:
                        print(f"Error reading Website VAPT catalog file: {e}")
                        catalog_df = None
                    
                    if catalog_df is not None and 'Vulnerabilities in this group' in catalog_df.columns:
                        matched_vulnerability_names = set()
                        
                        # Build matched groups with catalog details
                        for idx, row in catalog_df.iterrows():
                            vulnerabilities_in_group = str(row.get('Vulnerabilities in this group', '')).strip()
                            if pd.isna(vulnerabilities_in_group) or vulnerabilities_in_group == '':
                                continue
                            
                            # Split vulnerabilities by newlines
                            vuln_list = [v.strip() for v in vulnerabilities_in_group.split('\n') if v.strip()]
                            
                            # Find which vulnerabilities from Excel match this catalog group
                            matched_vulns_in_group = []
                            for vulnerability in unique_vulnerabilities:
                                # Use only first 170 characters for matching
                                vuln_short = str(vulnerability)[:170]
                                escaped_vulnerability = re.escape(vuln_short)
                                pattern = rf'(?:\n|\r\n|\A){escaped_vulnerability}'
                                if re.search(pattern, vulnerabilities_in_group, re.IGNORECASE):
                                    matched_vulns_in_group.append(vulnerability)
                                    matched_vulnerability_names.add(vulnerability)
                            
                            # If any vulnerabilities matched this catalog group, add it
                            if matched_vulns_in_group:
                                matched_groups.append({
                                    'catalog_id': int(idx),
                                    'group_name': str(row.get('Name of Vulnerability', ''))[:200],
                                    'risk_factor': str(row.get('Risk Factor', ''))[:20],
                                    'cvss_score': str(row.get('CVSS Score', ''))[:10],
                                    'matched_vulnerabilities': matched_vulns_in_group
                                })
                        
                        # Calculate unmatched vulnerabilities
                        unmatched_vulnerabilities = sorted(list(unique_vulnerabilities - matched_vulnerability_names))
                        
            except Exception as e:
                print(f"Error calculating vulnerabilities: {e}")
                import traceback
                traceback.print_exc()
                return jsonify({"error": f"Error processing vulnerabilities: {str(e)}"}), 500
            
            # Initialize merge state in session
            unmatched_limited = unmatched_vulnerabilities[:100] if len(unmatched_vulnerabilities) > 100 else unmatched_vulnerabilities
            
            session['website_vapt_vulnerability_merge_state'] = {
                'matched_groups': matched_groups[:50],
                'unmatched_vulnerabilities': unmatched_limited,
                'merge_operations': [],
                'new_group_details': {},
                'manually_added_vulnerabilities': []  # New: Track manually added vulnerabilities
            }
            
            if len(unmatched_vulnerabilities) > 100:
                print(f"⚠️ Warning: {len(unmatched_vulnerabilities)} unmatched vulnerabilities found. Limited to 100 in session.")
            
            # Return matched groups and unmatched vulnerabilities
            return jsonify({
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_vulnerabilities
            })
        else:
            return jsonify({"error": "No Nessus data found"}), 400
            
    except Exception as e:
        print(f"Error checking vulnerabilities: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error processing files: {str(e)}"}), 500


# =============================================================================
# MERGE OPERATIONS
# =============================================================================

@website_vapt_bp.route('/merge_website_vapt_with_matched', methods=['POST'])
def merge_website_with_matched():
    """Merge an unmatched vulnerability with an existing matched group."""
    try:
        data = request.get_json()
        
        if not data or 'unmatched_vulnerability' not in data or 'target_group_id' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        unmatched_vuln = data['unmatched_vulnerability']
        target_group_id = data['target_group_id']
        
        # Get current merge state from session
        merge_state = session.get('website_vapt_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Find the target matched group
        matched_groups = merge_state.get('matched_groups', [])
        target_group = None
        target_index = None
        
        for idx, group in enumerate(matched_groups):
            if group['catalog_id'] == target_group_id:
                target_group = group
                target_index = idx
                break
        
        if target_group is None:
            return jsonify({"error": "Target group not found"}), 404
        
        # Add the unmatched vulnerability to the matched group
        if unmatched_vuln not in target_group['matched_vulnerabilities']:
            target_group['matched_vulnerabilities'].append(unmatched_vuln)
            print(f"✓ Added '{unmatched_vuln}' to group '{target_group['group_name']}'")
        
        # Remove from unmatched list
        unmatched_list = merge_state.get('unmatched_vulnerabilities', [])
        print(f"DEBUG: Unmatched list before removal: {len(unmatched_list)} items")
        
        if unmatched_vuln in unmatched_list:
            unmatched_list.remove(unmatched_vuln)
            print(f"✓ Removed '{unmatched_vuln}' from unmatched list")
        
        # Explicitly update the merge_state
        merge_state['unmatched_vulnerabilities'] = unmatched_list
        
        # Record the merge operation for undo
        merge_state['merge_operations'].append({
            'type': 'merge_with_matched',
            'unmatched_vulnerability': unmatched_vuln,
            'target_group_id': target_group_id,
            'timestamp': datetime.now().isoformat()
        })
        
        # Update session
        session['website_vapt_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        print(f"DEBUG: Final state - Matched: {len(matched_groups)}, Unmatched: {len(unmatched_list)}")
        
        return jsonify({
            "success": True,
            "message": "Vulnerability merged successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_list
            }
        })
        
    except Exception as e:
        print(f"Error merging with matched: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error merging: {str(e)}"}), 500


@website_vapt_bp.route('/merge_website_vapt_with_unmatched', methods=['POST'])
def merge_website_with_unmatched():
    """Merge multiple unmatched vulnerabilities into a new group."""
    try:
        data = request.get_json()
        
        if not data or 'vulnerabilities' not in data or 'vulnerability_details' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        vulnerabilities_to_merge = data['vulnerabilities']
        vulnerability_details = data['vulnerability_details']
        
        # Validate required fields
        required_fields = ['vulnerabilityName', 'riskFactor', 'cveId', 'cvssScore', 
                          'auditObservation', 'impact', 'recommendation', 'referenceLink']
        
        for field in required_fields:
            if field not in vulnerability_details:
                return jsonify({"error": f"Missing required field '{field}'"}), 400
        
        # Get current merge state from session
        merge_state = session.get('website_vapt_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Create a new matched group
        new_group_id = -len(merge_state.get('matched_groups', [])) - 1
        new_group = {
            'catalog_id': new_group_id,
            'group_name': vulnerability_details['vulnerabilityName'][:200],
            'risk_factor': vulnerability_details['riskFactor'][:20],
            'cvss_score': vulnerability_details['cvssScore'][:10],
            'matched_vulnerabilities': vulnerabilities_to_merge,
            'is_new_group': True
        }
        
        # Add to matched groups
        matched_groups = merge_state.get('matched_groups', [])
        matched_groups.append(new_group)
        
        # Store full details separately
        # Add merged vulnerability names to details
        vulnerability_details['isMerged'] = True
        vulnerability_details['mergedVulnerabilities'] = vulnerabilities_to_merge
        
        new_group_details = merge_state.get('new_group_details', {})
        new_group_details[str(new_group_id)] = vulnerability_details
        merge_state['new_group_details'] = new_group_details
        
        # Remove from unmatched list
        unmatched_list = merge_state.get('unmatched_vulnerabilities', [])
        for vuln in vulnerabilities_to_merge:
            if vuln in unmatched_list:
                unmatched_list.remove(vuln)
        
        # Explicitly update the merge_state
        merge_state['unmatched_vulnerabilities'] = unmatched_list
        merge_state['matched_groups'] = matched_groups
        
        # Record the merge operation
        merge_state['merge_operations'].append({
            'type': 'merge_with_unmatched',
            'vulnerabilities': vulnerabilities_to_merge,
            'new_group_id': new_group['catalog_id'],
            'timestamp': datetime.now().isoformat()
        })
        
        # Update session
        session['website_vapt_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        # Update catalog
        vulnerability_details_with_merge = vulnerability_details.copy()
        vulnerability_details_with_merge['isMerged'] = True
        vulnerability_details_with_merge['mergedVulnerabilities'] = vulnerabilities_to_merge
        
        update_website_catalog_with_vulnerabilities({
            vulnerability_details['vulnerabilityName']: vulnerability_details_with_merge
        })
        
        return jsonify({
            "success": True,
            "message": "Vulnerabilities merged into new group successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_list
            }
        })
        
    except Exception as e:
        print(f"Error merging with unmatched: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error merging: {str(e)}"}), 500


@website_vapt_bp.route('/add_website_vapt_vulnerability_details', methods=['POST'])
def add_website_vulnerability_details():
    """Add detailed information for a single unmatched vulnerability."""
    try:
        data = request.get_json()
        
        if not data or 'vulnerability_name' not in data or 'vulnerability_details' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        vuln_name = data['vulnerability_name']
        vuln_details = data['vulnerability_details']
        
        # Get current merge state
        merge_state = session.get('website_vapt_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Create a new matched group for this single vulnerability
        new_group_id = -len(merge_state.get('matched_groups', [])) - 1
        new_group = {
            'catalog_id': new_group_id,
            'group_name': vuln_details['vulnerabilityName'][:200],
            'risk_factor': vuln_details['riskFactor'][:20],
            'cvss_score': vuln_details['cvssScore'][:10],
            'matched_vulnerabilities': [vuln_name],
            'is_new_group': True
        }
        
        # Add to matched groups
        matched_groups = merge_state.get('matched_groups', [])
        matched_groups.append(new_group)
        
        # Store full details
        # Add actual vulnerability name to details for worksheet display
        vuln_details['actualVulnerabilityName'] = vuln_name  # Store the original vulnerability name
        
        new_group_details = merge_state.get('new_group_details', {})
        new_group_details[str(new_group_id)] = vuln_details
        merge_state['new_group_details'] = new_group_details
        
        # Remove from unmatched list
        unmatched_list = merge_state.get('unmatched_vulnerabilities', [])
        if vuln_name in unmatched_list:
            unmatched_list.remove(vuln_name)
        
        # Explicitly update the merge_state
        merge_state['unmatched_vulnerabilities'] = unmatched_list
        merge_state['matched_groups'] = matched_groups
        
        # Update session
        session['website_vapt_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        # Update catalog
        # Add actual vulnerability name to details for catalog storage
        vuln_details_with_actual = vuln_details.copy()
        vuln_details_with_actual['actualVulnerabilityName'] = vuln_name  # Store the original vulnerability name
        
        update_website_catalog_with_vulnerabilities({
            vuln_details['vulnerabilityName']: vuln_details_with_actual
        })
        
        return jsonify({
            "success": True,
            "message": "Vulnerability details added successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": unmatched_list
            }
        })
        
    except Exception as e:
        print(f"Error adding vulnerability details: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error adding details: {str(e)}"}), 500


@website_vapt_bp.route('/merge_website_vapt_matched_groups', methods=['POST'])
def merge_website_matched_groups():
    """Merge two existing matched groups together."""
    try:
        data = request.get_json()
        
        if not data or 'source_group_id' not in data or 'target_group_id' not in data:
            return jsonify({"error": "Missing required parameters"}), 400
        
        source_group_id = data['source_group_id']
        target_group_id = data['target_group_id']
        
        # Get current merge state
        merge_state = session.get('website_vapt_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        matched_groups = merge_state.get('matched_groups', [])
        
        # Find source and target groups
        source_group = None
        target_group = None
        source_index = None
        
        for idx, group in enumerate(matched_groups):
            if group['catalog_id'] == source_group_id:
                source_group = group
                source_index = idx
            if group['catalog_id'] == target_group_id:
                target_group = group
        
        if source_group is None or target_group is None:
            return jsonify({"error": "Source or target group not found"}), 404
        
        # Merge vulnerabilities from source to target
        for vuln in source_group['matched_vulnerabilities']:
            if vuln not in target_group['matched_vulnerabilities']:
                target_group['matched_vulnerabilities'].append(vuln)
        
        # Remove source group
        matched_groups.pop(source_index)
        
        # Explicitly update the merge_state
        merge_state['matched_groups'] = matched_groups
        
        # Record the merge operation
        merge_state['merge_operations'].append({
            'type': 'merge_matched_groups',
            'source_group_id': source_group_id,
            'target_group_id': target_group_id,
            'timestamp': datetime.now().isoformat()
        })
        
        # Update session
        session['website_vapt_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        return jsonify({
            "success": True,
            "message": "Groups merged successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": merge_state.get('unmatched_vulnerabilities', [])
            }
        })
        
    except Exception as e:
        print(f"Error merging matched groups: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error merging groups: {str(e)}"}), 500


@website_vapt_bp.route('/website_vapt_undo_last_merge', methods=['POST'])
def website_undo_last_merge():
    """Undo the last merge operation."""
    try:
        merge_state = session.get('website_vapt_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        merge_operations = merge_state.get('merge_operations', [])
        if not merge_operations:
            return jsonify({"error": "No operations to undo"}), 400
        
        # Get the last operation
        last_operation = merge_operations.pop()
        
        # Note: Full undo implementation would require more complex state management
        # For now, we'll just provide feedback
        print(f"Undo operation: {last_operation['type']}")
        
        # Update session
        session['website_vapt_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        return jsonify({
            "success": True,
            "message": "Last operation undone (simplified)",
            "updated_state": {
                "matched_groups": merge_state.get('matched_groups', []),
                "unmatched_vulnerabilities": merge_state.get('unmatched_vulnerabilities', [])
            }
        })
        
    except Exception as e:
        print(f"Error undoing operation: {e}")
        return jsonify({"error": f"Error undoing: {str(e)}"}), 500


@website_vapt_bp.route('/add_website_manual_vulnerability', methods=['POST'])
def add_website_manual_vulnerability():
    """Add a manually entered vulnerability by user."""
    try:
        data = request.get_json()
        
        if not data or 'vulnerability_details' not in data:
            return jsonify({"error": "No vulnerability details provided"}), 400
        
        vuln_details = data['vulnerability_details']
        
        # Validate required fields
        required_fields = ['vulnerabilityName', 'riskFactor', 'cveId', 'cvssScore', 
                          'auditObservation', 'impact', 'recommendation', 'referenceLink']
        
        for field in required_fields:
            if field not in vuln_details:
                return jsonify({"error": f"Missing required field '{field}'"}), 400
        
        # Get current merge state
        merge_state = session.get('website_vapt_vulnerability_merge_state', {})
        if not merge_state:
            return jsonify({"error": "No vulnerability data found in session"}), 400
        
        # Create a new matched group for the manual vulnerability
        new_group_id = -len(merge_state.get('matched_groups', [])) - 1000  # Use -1000+ for manual entries
        new_group = {
            'catalog_id': new_group_id,
            'group_name': vuln_details['vulnerabilityName'][:200],
            'risk_factor': vuln_details['riskFactor'][:20],
            'cvss_score': vuln_details['cvssScore'][:10],
            'matched_vulnerabilities': [vuln_details['vulnerabilityName']],  # Manual vulnerability
            'is_new_group': True,
            'is_manual': True  # Flag to indicate this was manually added
        }
        
        # Add to matched groups
        matched_groups = merge_state.get('matched_groups', [])
        matched_groups.append(new_group)
        
        # Store full details
        new_group_details = merge_state.get('new_group_details', {})
        new_group_details[str(new_group_id)] = vuln_details
        merge_state['new_group_details'] = new_group_details
        
        # Debug: Log affected systems if provided
        if 'affectedSystems' in vuln_details:
            print(f"✓ Manual vulnerability '{vuln_details['vulnerabilityName']}' - Affected Systems received: {vuln_details['affectedSystems'][:100]}...")
        else:
            print(f"⚠️ Manual vulnerability '{vuln_details['vulnerabilityName']}' - No affectedSystems field in details. Keys: {list(vuln_details.keys())}")
        
        # Track manually added vulnerabilities
        manual_vulns = merge_state.get('manually_added_vulnerabilities', [])
        manual_vulns.append(vuln_details['vulnerabilityName'])
        merge_state['manually_added_vulnerabilities'] = manual_vulns
        
        # Update the merge_state
        merge_state['matched_groups'] = matched_groups
        
        # Update session
        session['website_vapt_vulnerability_merge_state'] = merge_state
        session.modified = True
        
        # Update catalog
        vuln_details['isManual'] = True
        # Add actual vulnerability name to details for catalog storage (manual = same as group name)
        vuln_details['actualVulnerabilityName'] = vuln_details['vulnerabilityName']
        
        update_website_catalog_with_vulnerabilities({
            vuln_details['vulnerabilityName']: vuln_details
        })
        
        return jsonify({
            "success": True,
            "message": "Manual vulnerability added successfully",
            "updated_state": {
                "matched_groups": matched_groups,
                "unmatched_vulnerabilities": merge_state.get('unmatched_vulnerabilities', [])
            },
            "new_vulnerability": {
                "group_id": new_group_id,
                "group_name": vuln_details['vulnerabilityName'],
                "risk_factor": vuln_details['riskFactor']
            }
        })
        
    except Exception as e:
        print(f"Error adding manual vulnerability: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error adding manual vulnerability: {str(e)}"}), 500


def update_website_catalog_with_vulnerabilities(vulnerability_details_dict):
    """Update the Website VAPT Catalog with both merged and separate vulnerabilities."""
    try:
        from flask_login import current_user
        
        catalog_path = "static/Formats_and_Catalog/Website VAPT Catalog.xlsx"
        
        if not os.path.exists(catalog_path):
            print(f"Catalog file not found at: {catalog_path}")
            return
        
        # Read the existing catalog with error handling - now reading Sheet2 (index 1)
        try:
            catalog_df = pd.read_excel(catalog_path, sheet_name=1)  # Changed from sheet_name=0 to sheet_name=1 for Sheet2
        except Exception as e:
            print(f"Error reading catalog file: {e}")
            print(f"Catalog file may be corrupted. Please check: {catalog_path}")
            return
        
        # Process all vulnerabilities (both merged and separate)
        for vuln_name, details in vulnerability_details_dict.items():
            if details.get('isMerged', False):
                # Handle merged vulnerabilities
                merged_vulns = details.get('mergedVulnerabilities', [])
                if merged_vulns:
                    # Truncate vulnerability names to first 170 characters for catalog storage
                    truncated_merged_vulns = [str(v)[:170] for v in merged_vulns]
                    
                    # Create a new row for the merged vulnerability group
                    new_row = {
                        'Sr No': len(catalog_df) + 1,
                        'Name of Vulnerability': details.get('vulnerabilityName', ''),
                        'Risk Factor': details.get('riskFactor', ''),
                        'CVE/CWE ID': details.get('cveId', ''),
                        'CVSS': details.get('cvssScore', ''),
                        'Audit Observation': details.get('auditObservation', ''),
                        'Impact': details.get('impact', ''),
                        'Recommendation/Countermeasure': details.get('recommendation', ''),
                        'Affected System': '',  # Empty as requested
                        'Reference Link': details.get('referenceLink', ''),
                        'Vulnerabilities in this group': '\n'.join(truncated_merged_vulns),
                        'User_name': current_user.employee_name if current_user.is_authenticated else 'Unknown',
                        'Time_stamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    # Add the new row to the catalog
                    catalog_df = pd.concat([catalog_df, pd.DataFrame([new_row])], ignore_index=True)
            else:
                # Handle separate (non-merged) vulnerabilities
                # Get the actual vulnerability name (might be different from group name)
                actual_vuln_name = details.get('actualVulnerabilityName', vuln_name)
                # Truncate to first 170 characters for catalog storage
                actual_vuln_name_short = str(actual_vuln_name)[:170]
                
                new_row = {
                    'Sr No': len(catalog_df) + 1,
                    'Name of Vulnerability': details.get('vulnerabilityName', vuln_name),  # User-provided group name
                    'Risk Factor': details.get('riskFactor', ''),
                    'CVE/CWE ID': details.get('cveId', ''),
                    'CVSS': details.get('cvssScore', ''),
                    'Audit Observation': details.get('auditObservation', ''),
                    'Impact': details.get('impact', ''),
                    'Recommendation/Countermeasure': details.get('recommendation', ''),
                    'Affected System': '',  # Empty as requested
                    'Reference Link': details.get('referenceLink', ''),
                    'Vulnerabilities in this group': actual_vuln_name_short,  # Truncated to 170 chars
                    'User_name': current_user.employee_name if current_user.is_authenticated else 'Unknown',
                    'Time_stamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                # Add the new row to the catalog
                catalog_df = pd.concat([catalog_df, pd.DataFrame([new_row])], ignore_index=True)
        
        # Save the updated catalog to Sheet2
        # Read all existing sheets first
        try:
            # Read all sheets to preserve existing data
            all_sheets = pd.read_excel(catalog_path, sheet_name=None)
            
            # Update Sheet2 with our new data
            all_sheets['Sheet2'] = catalog_df
            
            # Write all sheets back to the file
            with pd.ExcelWriter(catalog_path, engine='openpyxl') as writer:
                for sheet_name, sheet_df in all_sheets.items():
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            print(f"✅ Successfully updated Website VAPT Catalog (Sheet2)")
        
        except Exception as e:
            print(f"Error updating catalog file: {e}")
            import traceback
            traceback.print_exc()
        
    except Exception as e:
        print(f"Error updating Website VAPT Catalog: {e}")
        import traceback
        traceback.print_exc()


# =============================================================================
# REPORT GENERATION
# =============================================================================

def create_website_nmap_worksheet(wb, ip_ports_dict):
    """
    Create Nmap Scan worksheet for Website VAPT.
    Format: 3 columns (HOST, PORT, SERVICE) with HOST column merged for each IP
    Same logic as Infrastructure VAPT but with 3 columns instead of 6
    """
    try:
        print("📊 Creating Nmap Scan worksheet...")
        ws = wb.create_sheet("Nmap Files", 0)
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=10)
        data_alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 25  # HOST
        ws.column_dimensions['B'].width = 15  # PORT
        ws.column_dimensions['C'].width = 30  # SERVICE
        
        # Sort IPs for consistent output
        sorted_ips = sorted(ip_ports_dict.keys())
        
        current_row = 1
        total_ports = 0
        
        # Process each IP
        for ip in sorted_ips:
            ports = ip_ports_dict.get(ip, [])
            
            # If no ports found, add "Filtered" entry
            if len(ports) == 0:
                ports = [("Filtered", "Filtered")]
            
            # Write HOST PORT SERVICE headers before each new IP (only write headers in HOST column, PORT and SERVICE will be empty)
            cell = ws.cell(row=current_row, column=1, value="HOST")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=current_row, column=2, value="PORT")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=current_row, column=3, value="SERVICE")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            ws.row_dimensions[current_row].height = 30
            current_row += 1
            
            start_row = current_row
            
            # Write all ports for this IP
            for port, service in ports:
                # PORT column
                cell = ws.cell(row=current_row, column=2, value=port)
                cell.font = data_font
                cell.alignment = data_alignment_center
                cell.border = thin_border
                
                # SERVICE column
                cell = ws.cell(row=current_row, column=3, value=service)
                cell.font = data_font
                cell.alignment = data_alignment_center
                cell.border = thin_border
                
                ws.row_dimensions[current_row].height = 20
                current_row += 1
                total_ports += 1
            
            end_row = current_row - 1
            
            # Merge HOST column for this IP if it has multiple ports (start from after header row)
            if end_row > start_row:
                # Merge cells from start_row to end_row in column A
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            
            # Write HOST value (will appear in merged cell)
            cell = ws.cell(row=start_row, column=1, value=ip)
            cell.font = data_font
            cell.alignment = data_alignment_center
            cell.border = thin_border
            
            # Apply border to all cells in the merged range (data rows only, header already has border)
            for row_idx in range(start_row, end_row + 1):
                ws.cell(row=row_idx, column=1).border = thin_border
        
        print(f"✅ Nmap Files worksheet created with {len(sorted_ips)} IPs and {total_ports} port entries")
        
    except Exception as e:
        print(f"❌ Error creating Nmap worksheet: {e}")
        import traceback
        traceback.print_exc()


def create_website_nessus_csv_worksheet(wb, nessus_dataframes):
    """
    Create Nessus CSV Files worksheet - direct copy from Nessus files.
    """
    try:
        print("📊 Creating Nessus CSV Files worksheet...")
        
        if not nessus_dataframes:
            print("⚠️ No Nessus data to create worksheet")
            return
        
        # Combine all Nessus dataframes
        combined_df = pd.concat(nessus_dataframes, ignore_index=True)
        
        ws = wb.create_sheet("Nessus CSV Files")
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Write headers (column names from DataFrame)
        for col_idx, column_name in enumerate(combined_df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        ws.row_dimensions[1].height = 30
        
        # Write data rows
        for row_idx, data_row in enumerate(combined_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            ws.row_dimensions[row_idx].height = 20
        
        print(f"✅ Nessus CSV Files worksheet created with {len(combined_df)} rows")
        
    except Exception as e:
        print(f"❌ Error creating Nessus CSV Files worksheet: {e}")
        import traceback
        traceback.print_exc()


def create_website_scope_worksheet(wb, nessus_dataframes):
    """
    Create Scope worksheet with Sr.No and Host columns from Nessus CSV Files.
    Same format as Infrastructure VAPT reference
    """
    try:
        print("📊 Creating Scope worksheet...")
        
        if not nessus_dataframes:
            print("⚠️ No Nessus data for Scope worksheet")
            return
        
        # Combine all Nessus dataframes
        combined_df = pd.concat(nessus_dataframes, ignore_index=True)
        
        # Extract unique Host values
        if 'Host' not in combined_df.columns:
            print("⚠️ 'Host' column not found in Nessus data")
            return
        
        unique_hosts = combined_df['Host'].dropna().unique().tolist()
        unique_hosts = [str(host).strip() for host in unique_hosts if str(host).strip()]
        unique_hosts = sorted(unique_hosts)  # Sort for consistent output
        
        ws = wb.create_sheet("Scope")
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        data_font = Font(name='Arial', size=10)
        data_alignment_center = Alignment(horizontal='center', vertical='center')
        data_alignment_left = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 10  # Sr.No
        ws.column_dimensions['B'].width = 30  # Host
        
        # Create headers
        headers = ['Sr.No', 'Host']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 30
        
        # Add unique hosts with serial numbers
        for idx, host in enumerate(unique_hosts, start=2):
            # Sr.No
            cell = ws.cell(row=idx, column=1, value=idx - 1)
            cell.font = data_font
            cell.alignment = data_alignment_center
            cell.border = thin_border
            
            # Host
            cell = ws.cell(row=idx, column=2, value=host)
            cell.font = data_font
            cell.alignment = data_alignment_center
            cell.border = thin_border
            
            ws.row_dimensions[idx].height = 20
        
        print(f"✅ Scope worksheet created with {len(unique_hosts)} hosts")
        
    except Exception as e:
        print(f"❌ Error creating Scope worksheet: {e}")
        import traceback
        traceback.print_exc()


def create_website_summary_worksheet(wb, nessus_dataframes):
    """
    Create Summary worksheet for Website VAPT.
    Format: 3 columns (Sr.No, Name, Host) with Name column merged for same vulnerabilities
    Same format as Infrastructure VAPT reference
    """
    try:
        print("📊 Creating Summary worksheet...")
        
        # Always create Summary worksheet first with headers
        ws = wb.create_sheet("Summary")
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        note_font = Font(name='Arial', size=12, bold=True, color='FF008000')  # Green color
        note_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Set column widths (always set, even if no data)
        ws.column_dimensions['A'].width = 10  # Sr.No
        ws.column_dimensions['B'].width = 60  # Name
        ws.column_dimensions['C'].width = 20  # Host
        
        # Create headers
        headers = ['Sr.No', 'Name', 'Host']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 30
        
        # Check if nessus_dataframes is empty or None
        if not nessus_dataframes:
            print("⚠️ No Nessus data for Summary worksheet - creating empty worksheet with note")
            # Merge cells A3-C3 and add note
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
            note_cell = ws.cell(row=3, column=1, value="Note: No vulnerabilities were identified during the Website VAPT Audit.")
            note_cell.font = note_font
            note_cell.alignment = note_alignment
            
            # Apply border to all cells in the merged range
            for col in range(1, 4):  # Columns A, B, C
                cell = ws.cell(row=3, column=col)
                cell.border = thin_border
            
            ws.row_dimensions[3].height = 30
            print("✅ Summary worksheet created with headers and no data message")
            return
        
        # Combine all Nessus dataframes
        combined_df = pd.concat(nessus_dataframes, ignore_index=True)
        
        # Filter by valid risks
        valid_risks = ['low', 'medium', 'high', 'critical']
        df_filtered = combined_df.copy()
        df_filtered['Risk'] = df_filtered['Risk'].astype(str).str.lower().str.strip()
        df_filtered = df_filtered[df_filtered['Risk'].isin(valid_risks)]
        
        # Select only Host, Name columns (no Risk column in output)
        if not all(col in df_filtered.columns for col in ['Host', 'Name']):
            print("⚠️ Required columns not found in Nessus data - worksheet created with headers only")
            # Merge cells A3-C3 and add note
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
            note_cell = ws.cell(row=3, column=1, value="Note: No vulnerabilities were identified during the Website VAPT Audit.")
            note_cell.font = note_font
            note_cell.alignment = note_alignment
            
            # Apply border to all cells in the merged range
            for col in range(1, 4):  # Columns A, B, C
                cell = ws.cell(row=3, column=col)
                cell.border = thin_border
            
            ws.row_dimensions[3].height = 30
            print("✅ Summary worksheet created with headers and no data message")
            return
        
        summary_data = df_filtered[['Host', 'Name']].copy()
        
        # Clean data
        summary_data['Host'] = summary_data['Host'].astype(str).str.strip()
        summary_data['Name'] = summary_data['Name'].astype(str).str.strip()
        
        # Remove empty rows
        summary_data = summary_data.dropna()
        summary_data = summary_data[(summary_data['Host'] != '') & (summary_data['Name'] != '')]
        
        if summary_data.empty:
            print("⚠️ No valid data found for Summary worksheet after filtering - worksheet created with headers and no data message")
            # Merge cells A3-C3 and add note (worksheet already created with headers)
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
            note_cell = ws.cell(row=3, column=1, value="Note: No vulnerabilities were identified during the Website VAPT Audit.")
            note_cell.font = note_font
            note_cell.alignment = note_alignment
            
            # Apply border to all cells in the merged range
            for col in range(1, 4):  # Columns A, B, C
                cell = ws.cell(row=3, column=col)
                cell.border = thin_border
            
            ws.row_dimensions[3].height = 30
            print("✅ Summary worksheet created with headers and no data message")
            return
        
        # Remove duplicates based on Name + Host combination
        summary_data = summary_data.drop_duplicates(subset=['Name', 'Host'], keep='first')
        
        # Sort by Name, then Host to group similar vulnerabilities
        summary_data = summary_data.sort_values(['Name', 'Host'])
        
        # Define data styles (header styles already defined above)
        data_font = Font(name='Arial', size=10)
        data_alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Write data with merging logic
        row = 2
        current_vulnerability = None
        name_merge_start = 2
        serial_counter = 1
        
        for _, row_data in summary_data.iterrows():
            name = str(row_data['Name'])
            host = str(row_data['Host'])
            
            # If we're starting a new vulnerability
            if name != current_vulnerability:
                # Merge previous vulnerability name cells if needed
                if current_vulnerability is not None and row > name_merge_start:
                    # Merge Name column
                    ws.merge_cells(start_row=name_merge_start, start_column=2, end_row=row-1, end_column=2)
                    # Merge Sr.No column
                    ws.merge_cells(start_row=name_merge_start, start_column=1, end_row=row-1, end_column=1)
                
                current_vulnerability = name
                name_merge_start = row
                serial_counter += 1
            
            # Write Sr.No (only for first occurrence of each vulnerability)
            if name != current_vulnerability or row == name_merge_start:
                cell = ws.cell(row=row, column=1, value=serial_counter - 1)
                cell.font = data_font
                cell.alignment = data_alignment_center
                cell.border = thin_border
            
            # Write Name (only for first occurrence of each vulnerability)
            if name != current_vulnerability or row == name_merge_start:
                cell = ws.cell(row=row, column=2, value=name)
                cell.font = data_font
                cell.alignment = data_alignment_center
                cell.border = thin_border
            
            # Write Host (always)
            cell = ws.cell(row=row, column=3, value=host)
            cell.font = data_font
            cell.alignment = data_alignment_center
            cell.border = thin_border
            
            ws.row_dimensions[row].height = 20
            row += 1
        
        # Merge the last vulnerability if needed
        if current_vulnerability is not None and row > name_merge_start:
            # Merge Name column
            ws.merge_cells(start_row=name_merge_start, start_column=2, end_row=row-1, end_column=2)
            # Merge Sr.No column
            ws.merge_cells(start_row=name_merge_start, start_column=1, end_row=row-1, end_column=1)
        
        print(f"✅ Summary worksheet created with {len(summary_data)} entries")
        
    except Exception as e:
        print(f"❌ Error creating Summary worksheet: {e}")
        import traceback
        traceback.print_exc()

def create_website_vapt_worksheet(wb, nessus_dataframes):
    """
    Create Website_VAPT worksheet with data from Summary's Name column and match with catalog (using merge state).
    Similar to Infra_VAPT worksheet but adapted for Website VAPT data (no Branch Name, just Host).
    Includes manual vulnerabilities from merge state.
    """
    try:
        print("📊 Creating Website_VAPT worksheet...")
        
        # Load the catalog file first (needed for headers)
        catalog_path = os.path.join('static', 'Formats_and_Catalog', 'Website VAPT Catalog.xlsx')
        catalog_df = None
        
        if os.path.exists(catalog_path):
            # Read the catalog file with error handling - read Sheet1 (index 0)
            try:
                catalog_df = pd.read_excel(catalog_path, sheet_name=0)
                print(f"Catalog loaded successfully with {len(catalog_df)} rows")
            except Exception as e:
                print(f"Error reading catalog file for Website_VAPT worksheet: {e}")
                print(f"Catalog file may be corrupted. Will create worksheet with basic headers only.")
        else:
            print(f"Catalog file not found at: {catalog_path} - will create worksheet with basic headers only")
        
        # CREATE WORKSHEET FIRST - regardless of whether there are vulnerabilities
        # This ensures the worksheet exists even if there are no vulnerabilities
        ws = wb.create_sheet("Website_VAPT")
        print("✅ Website_VAPT worksheet created successfully (will populate with data if available)")
        
        # Define columns to exclude from catalog
        exclude_columns = ['Sr No', 'Vulnerabilities in this group', 'Affected System']
        
        # Get catalog headers excluding the specified columns
        if catalog_df is not None and 'Vulnerabilities in this group' in catalog_df.columns:
            catalog_headers = [col for col in catalog_df.columns.tolist() if col not in exclude_columns]
            # Normalize column names: replace "CVE ID" with "CVE/CWE ID" for consistency
            catalog_headers = ['CVE/CWE ID' if col == 'CVE ID' else col for col in catalog_headers]
            # Also normalize the catalog_df column names to match
            if 'CVE ID' in catalog_df.columns:
                catalog_df = catalog_df.rename(columns={'CVE ID': 'CVE/CWE ID'})
        else:
            # Use default headers if catalog is not available
            catalog_headers = []
            print("Using default headers as catalog is not available")
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        note_font = Font(name='Arial', size=12, bold=True, color='FF008000')  # Green color
        note_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Find the positions of key columns
        rec_countermeasure_col = None
        reference_link_col = None
        
        for i, col_name in enumerate(catalog_headers):
            if 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                rec_countermeasure_col = i
            if 'reference' in col_name.lower() and 'link' in col_name.lower():
                reference_link_col = i
        
        # Insert "Affected Systems" column after Recommendation/Countermeasure and before Reference Link
        if rec_countermeasure_col is not None and reference_link_col is not None:
            insert_position = rec_countermeasure_col + 1
        elif rec_countermeasure_col is not None:
            insert_position = rec_countermeasure_col + 1
        elif reference_link_col is not None:
            insert_position = reference_link_col
        else:
            insert_position = len(catalog_headers)
        
        # Create worksheet headers
        headers = ["Sr.No", "Vulnerabilities"]
        headers.extend(catalog_headers[:insert_position])
        headers.append("Affected Systems")
        headers.extend(catalog_headers[insert_position:])
        
        # Add POC columns at the end (7 columns: L, M, N, O, P, Q, R)
        poc_headers = ["POC_L", "POC_M", "POC_N", "POC_O", "POC_P", "POC_Q", "POC_R"]
        headers.extend(poc_headers)
        
        # Set column widths
        column_widths = [7, 35, 30, 15, 20, 10, 60, 60, 60, 40, 50, 30, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25]
        for i, width in enumerate(column_widths):
            if i < len(headers):
                col_letter = get_column_letter(i + 1)
                ws.column_dimensions[col_letter].width = width
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Find POC column positions and merge them
        poc_col_start = None
        poc_col_end = None
        for i, header in enumerate(headers, start=1):
            if header == "POC_L":
                poc_col_start = i
            if header == "POC_R":
                poc_col_end = i
                break
        
        # Merge POC header cells and set value to "POC"
        if poc_col_start is not None and poc_col_end is not None:
            ws.merge_cells(start_row=1, start_column=poc_col_start, end_row=1, end_column=poc_col_end)
            poc_header_cell = ws.cell(row=1, column=poc_col_start, value="POC")
            poc_header_cell.font = header_font
            poc_header_cell.fill = header_fill
            poc_header_cell.alignment = header_alignment
            poc_header_cell.border = thin_border
        
        ws.row_dimensions[1].height = 30
        
        print(f"✅ Worksheet headers created with {len(headers)} columns")
        
        # Get merge state from session - check before processing data
        merge_state = session.get('website_vapt_vulnerability_merge_state', None)
        use_merge_state = merge_state is not None and 'matched_groups' in merge_state
        
        if use_merge_state:
            print("📊 Using merge state from session for vulnerability grouping")
            merged_groups_from_session = merge_state.get('matched_groups', [])
            new_group_details_dict = merge_state.get('new_group_details', {})
            manually_added_vulnerabilities = merge_state.get('manually_added_vulnerabilities', [])
        else:
            print("📊 No merge state found - using standard catalog matching")
            merged_groups_from_session = None
            new_group_details_dict = {}
            manually_added_vulnerabilities = []
        
        # Check if we have Nessus data or manual vulnerabilities
        if not nessus_dataframes and not manually_added_vulnerabilities:
            print("⚠️ No Nessus data or manual vulnerabilities for Website_VAPT worksheet - worksheet created with headers only")
            # Merge cells A3-R3 (columns 1-18) and add note
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=18)
            note_cell = ws.cell(row=3, column=1, value="Note: No vulnerabilities were identified during the Website VAPT Audit.")
            note_cell.font = note_font
            note_cell.alignment = note_alignment
            
            # Apply border to all cells in the merged range
            for col in range(1, 19):  # Columns A through R (1-18)
                cell = ws.cell(row=3, column=col)
                cell.border = thin_border
            
            ws.row_dimensions[3].height = 30
            print("✅ Website_VAPT worksheet created with headers and no data message")
            return
        
        # Combine all Nessus dataframes (if available)
        if nessus_dataframes:
            combined_nessus = pd.concat(nessus_dataframes, ignore_index=True)
            
            # Check if required columns exist
            required_columns = ['Name', 'Host', 'Risk']
            missing_cols = [col for col in required_columns if col not in combined_nessus.columns]
            if missing_cols:
                print(f"⚠️ Missing required columns: {missing_cols} - worksheet created with headers only")
                # Merge cells A3-R3 (columns 1-18) and add note (worksheet already created with headers)
                ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=18)
                note_cell = ws.cell(row=3, column=1, value="Note: No vulnerabilities were identified during the Website VAPT Audit.")
                note_cell.font = note_font
                note_cell.alignment = note_alignment
                
                # Apply border to all cells in the merged range
                for col in range(1, 19):  # Columns A through R (1-18)
                    cell = ws.cell(row=3, column=col)
                    cell.border = thin_border
                
                ws.row_dimensions[3].height = 30
                print("✅ Website_VAPT worksheet created with headers only")
                return
            
            # Filter only low, medium, high, critical vulnerabilities
            valid_risks = ['low', 'medium', 'high', 'critical']
            df_filtered = combined_nessus.copy()
            df_filtered['Risk'] = df_filtered['Risk'].astype(str).str.lower().str.strip()
            df_filtered = df_filtered[df_filtered['Risk'].isin(valid_risks)]
            
            # Get unique vulnerability names from Name column
            unique_vulnerabilities = df_filtered['Name'].drop_duplicates().tolist()
        else:
            unique_vulnerabilities = []
        
        # Add manual vulnerabilities to unique list if they're not already there
        for manual_vuln in manually_added_vulnerabilities:
            if manual_vuln not in unique_vulnerabilities:
                unique_vulnerabilities.append(manual_vuln)
        
        if not unique_vulnerabilities:
            print("⚠️ No vulnerabilities found for Website_VAPT worksheet - worksheet created with headers only")
            # Merge cells A3-R3 (columns 1-18) and add note (worksheet already created with headers)
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=18)
            note_cell = ws.cell(row=3, column=1, value="Note: No vulnerabilities were identified during the Website VAPT Audit.")
            note_cell.font = note_font
            note_cell.alignment = note_alignment
            
            # Apply border to all cells in the merged range
            for col in range(1, 19):  # Columns A through R (1-18)
                cell = ws.cell(row=3, column=col)
                cell.border = thin_border
            
            ws.row_dimensions[3].height = 30
            print("✅ Website_VAPT worksheet created with headers only")
            return
        
        # Check if catalog is available for matching
        if catalog_df is None or 'Vulnerabilities in this group' not in catalog_df.columns:
            print("⚠️ Catalog file not available - worksheet created with headers only")
            # Merge cells A3-R3 (columns 1-18) and add note (worksheet already created with headers)
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=18)
            note_cell = ws.cell(row=3, column=1, value="Note: No vulnerabilities were identified during the Website VAPT Audit.")
            note_cell.font = note_font
            note_cell.alignment = note_alignment
            
            # Apply border to all cells in the merged range
            for col in range(1, 19):  # Columns A through R (1-18)
                cell = ws.cell(row=3, column=col)
                cell.border = thin_border
            
            ws.row_dimensions[3].height = 30
            print("✅ Website_VAPT worksheet created with headers only")
            return
        
        print(f"Found {len(unique_vulnerabilities)} unique vulnerabilities to process")
        
        # Define data styles (header styles already defined above)
        data_font = Font(name='Arial', size=10)
        data_alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Risk color formats
        critical_fill = PatternFill(start_color='FF8B0000', end_color='FF8B0000', fill_type='solid')
        high_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        medium_fill = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')
        low_fill = PatternFill(start_color='FF008000', end_color='FF008000', fill_type='solid')
        
        critical_font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
        high_font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
        medium_font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
        low_font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
        
        # Find additional column positions (catalog_headers already defined above)
        audit_observation_col = None
        risk_factor_col = None
        
        for i, col_name in enumerate(catalog_headers):
            if 'audit' in col_name.lower() and 'observation' in col_name.lower():
                audit_observation_col = i
            if 'risk' in col_name.lower() and 'factor' in col_name.lower():
                risk_factor_col = i
        
        # Track matched catalog entries
        matched_catalog_indices = set()
        matched_vulnerabilities = set()
        vulnerability_groups = {}
        vulnerability_affected_systems = {}
        vulnerability_risks = {}
        catalog_risk_values = {}
        
        # First pass: Collect affected systems for ALL vulnerabilities
        for vulnerability in unique_vulnerabilities:
            # Check if this is a manual vulnerability (not in scan data)
            is_manual = vulnerability in manually_added_vulnerabilities
            
            if is_manual:
                # For manual vulnerabilities, get affected systems from merge state
                affected_systems_list = []
                
                # Find the group that contains this manual vulnerability
                if use_merge_state and merged_groups_from_session:
                    for group in merged_groups_from_session:
                        if vulnerability in group.get('matched_vulnerabilities', []):
                            group_id = group['catalog_id']
                            details = new_group_details_dict.get(str(group_id), {})
                            
                            # Get affected systems from details
                            if 'affectedSystems' in details and details['affectedSystems']:
                                affected_systems_str = str(details['affectedSystems']).strip()
                                if affected_systems_str:
                                    # Split by newlines or commas and clean up
                                    if '\n' in affected_systems_str:
                                        affected_systems_list = [s.strip() for s in affected_systems_str.split('\n') if s.strip()]
                                    elif ',' in affected_systems_str:
                                        affected_systems_list = [s.strip() for s in affected_systems_str.split(',') if s.strip()]
                                    else:
                                        affected_systems_list = [affected_systems_str]
                                
                                print(f"✓ Found affected systems for manual vulnerability '{vulnerability}': {affected_systems_list}")
                                break
                    
                    # If not found in merged groups, check all new_group_details (in case group was merged)
                    if not affected_systems_list:
                        for group_id_str, details in new_group_details_dict.items():
                            # Check if this vulnerability name matches
                            if details.get('vulnerabilityName') == vulnerability or details.get('actualVulnerabilityName') == vulnerability:
                                if 'affectedSystems' in details and details['affectedSystems']:
                                    affected_systems_str = str(details['affectedSystems']).strip()
                                    if affected_systems_str:
                                        if '\n' in affected_systems_str:
                                            affected_systems_list = [s.strip() for s in affected_systems_str.split('\n') if s.strip()]
                                        elif ',' in affected_systems_str:
                                            affected_systems_list = [s.strip() for s in affected_systems_str.split(',') if s.strip()]
                                        else:
                                            affected_systems_list = [affected_systems_str]
                                        
                                        print(f"✓ Found affected systems for manual vulnerability '{vulnerability}' from details: {affected_systems_list}")
                                        break
                
                vulnerability_affected_systems[vulnerability] = affected_systems_list
                vulnerability_risks[vulnerability] = 'low'  # Default risk
            else:
                # Get all affected systems for this vulnerability (just Host, no Branch Name)
                vuln_data = df_filtered[df_filtered['Name'] == vulnerability]
                hosts = set()
                max_risk = 'low'
                
                for _, row in vuln_data.iterrows():
                    host = str(row['Host']).strip()
                    risk = str(row['Risk']).lower().strip()
                    
                    # Track the highest risk
                    risk_levels = {'critical': 4, 'high': 3, 'medium': 2, 'low': 1}
                    if risk_levels.get(risk, 0) > risk_levels.get(max_risk, 0):
                        max_risk = risk
                    
                    if host:
                        hosts.add(host)
                
                # Format affected systems (just hosts, no branches)
                formatted_systems = sorted(list(hosts))
                vulnerability_affected_systems[vulnerability] = formatted_systems
                vulnerability_risks[vulnerability] = max_risk
        
        # Use merge state if available, otherwise use standard catalog matching
        if use_merge_state and merged_groups_from_session:
            print(f"✓ Using {len(merged_groups_from_session)} merged groups from session")
            
            # Build vulnerability_groups dict from merge state
            for group in merged_groups_from_session:
                catalog_idx = group['catalog_id']
                matched_catalog_indices.add(catalog_idx)
                
                # Add all vulnerabilities in this group
                vulnerability_groups[catalog_idx] = group['matched_vulnerabilities']
                for vuln in group['matched_vulnerabilities']:
                    matched_vulnerabilities.add(vuln)
                
                # Store risk factor
                catalog_risk_values[catalog_idx] = str(group.get('risk_factor', '')).upper().strip()
        else:
            # Standard catalog matching
            print("✓ Using standard catalog matching")
            for vulnerability in unique_vulnerabilities:
                # Skip manual vulnerabilities in standard matching
                if vulnerability in manually_added_vulnerabilities:
                    continue
                
                # Use only first 170 characters for matching
                vuln_short = str(vulnerability)[:170]
                escaped_vulnerability = re.escape(vuln_short)
                pattern = rf'(?:\n|\r\n|\A){escaped_vulnerability}'
                
                # Find matching rows in catalog
                matching_rows = catalog_df[
                    catalog_df['Vulnerabilities in this group'].str.contains(
                        pattern, 
                        case=False, 
                        na=False,
                        regex=True
                    )
                ]
                
                if not matching_rows.empty:
                    catalog_idx = matching_rows.index[0]
                    matched_catalog_indices.add(catalog_idx)
                    matched_vulnerabilities.add(vulnerability)
                    
                    if catalog_idx not in vulnerability_groups:
                        vulnerability_groups[catalog_idx] = []
                    vulnerability_groups[catalog_idx].append(vulnerability)
                    
                    # Store the risk factor from catalog
                    if risk_factor_col is not None:
                        catalog_risk_value = catalog_df.loc[catalog_idx, catalog_headers[risk_factor_col]]
                        if pd.isna(catalog_risk_value):
                            catalog_risk_value = ""
                        catalog_risk_values[catalog_idx] = str(catalog_risk_value).upper().strip()
        
        # Sort catalog indices by risk factor
        risk_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
        
        def get_risk_priority(catalog_idx):
            risk_value = catalog_risk_values.get(catalog_idx, "")
            return risk_order.get(risk_value, 4)
        
        sorted_catalog_indices = sorted(matched_catalog_indices, key=get_risk_priority)
        
        # Second pass: Write data to worksheet
        row_num = 2
        
        for catalog_idx in sorted_catalog_indices:
            # Get catalog data - either from actual catalog or from merge state
            if use_merge_state and catalog_idx < 0:
                # New group from merge state (including manual vulnerabilities)
                group_data = next((g for g in merged_groups_from_session if g['catalog_id'] == catalog_idx), None)
                if not group_data:
                    continue
                
                # Get full details
                full_details = new_group_details_dict.get(str(catalog_idx), {})
                
                # Create pseudo catalog row
                catalog_row = pd.Series({
                    'Name of Vulnerability': group_data.get('group_name', ''),
                    'Risk Factor': full_details.get('riskFactor', group_data.get('risk_factor', '')),
                    'CVE/CWE ID': full_details.get('cveId', 'N/A'),
                    'CVSS': full_details.get('cvssScore', group_data.get('cvss_score', '')),
                    'Audit Observation': full_details.get('auditObservation', ''),
                    'Impact': full_details.get('impact', ''),
                    'Recommendation/Countermeasure': full_details.get('recommendation', ''),
                    'Reference Link': full_details.get('referenceLink', ''),
                    'Affected System': '',
                    'Vulnerabilities in this group': ''
                })
            else:
                # Standard catalog entry
                catalog_row = catalog_df.loc[catalog_idx]
            
            vulnerabilities_list = vulnerability_groups.get(catalog_idx, [])
            
            # Determine the highest risk for this group
            group_max_risk = 'low'
            risk_levels = {'critical': 4, 'high': 3, 'medium': 2, 'low': 1}
            for vuln in vulnerabilities_list:
                if risk_levels.get(vulnerability_risks.get(vuln, 'low'), 0) > risk_levels.get(group_max_risk, 0):
                    group_max_risk = vulnerability_risks.get(vuln, 'low')
            
            # Collect all affected systems for all vulnerabilities in this group
            all_affected_systems = []
            hosts_combined = set()
            
            for vuln in vulnerabilities_list:
                if vuln in vulnerability_affected_systems:
                    hosts_combined.update(vulnerability_affected_systems[vuln])
            
            # Format affected systems (just sorted hosts)
            all_affected_systems = sorted(list(hosts_combined))
            affected_systems_str = "\n".join(all_affected_systems) if all_affected_systems else ""
            
            # Write serial number
            cell = ws.cell(row=row_num, column=1, value=row_num - 1)
            cell.font = data_font
            cell.alignment = data_alignment_center
            cell.border = thin_border
            
            # Write vulnerabilities
            vulnerabilities_str = "\n".join(sorted(vulnerabilities_list))
            cell = ws.cell(row=row_num, column=2, value=vulnerabilities_str)
            cell.font = data_font
            cell.alignment = data_alignment_center
            cell.border = thin_border
            
            # Write catalog data up to insertion point
            col_idx = 3
            for i in range(insert_position):
                col_name = catalog_headers[i]
                value = catalog_row.get(col_name, "")
                if pd.isna(value):
                    value = ""
                
                # Special handling for Risk Factor column
                if risk_factor_col is not None and i == risk_factor_col:
                    catalog_risk_value = str(value).strip() if value else ""
                    
                    # Apply color formatting
                    if catalog_risk_value.upper() == 'CRITICAL':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = critical_font
                        cell.fill = critical_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    elif catalog_risk_value.upper() == 'HIGH':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = high_font
                        cell.fill = high_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    elif catalog_risk_value.upper() == 'MEDIUM':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = medium_font
                        cell.fill = medium_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    elif catalog_risk_value.upper() == 'LOW':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = low_font
                        cell.fill = low_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    else:
                        # Use highest from scan results
                        camelcase_value = convert_risk_to_camelcase(group_max_risk)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        if group_max_risk == 'critical':
                            cell.font = critical_font
                            cell.fill = critical_fill
                        elif group_max_risk == 'high':
                            cell.font = high_font
                            cell.fill = high_fill
                        elif group_max_risk == 'medium':
                            cell.font = medium_font
                            cell.fill = medium_fill
                        elif group_max_risk == 'low':
                            cell.font = low_font
                            cell.fill = low_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                else:
                    # Special handling for Audit Observation
                    if audit_observation_col is not None and i == audit_observation_col:
                        if len(vulnerabilities_list) > 1:
                            observation_text = "It was observed that the hosts are affected by multiple vulnerabilities, which are listed below.\n\n"
                            observation_text += vulnerabilities_str
                            if value:
                                value = f"{value}\n\n{observation_text}"
                            else:
                                value = observation_text
                        elif not value:
                            value = "It was observed that the host is affected by a vulnerability."
                    
                    # Special handling for CVE/CWE ID
                    if col_name == 'CVE/CWE ID' and value == "":
                        value = "N/A"
                    
                    # Apply left alignment for specific columns
                    cell = ws.cell(row=row_num, column=col_idx, value=str(value))
                    cell.font = data_font
                    if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                        cell.alignment = data_alignment_left
                    else:
                        cell.alignment = data_alignment_center
                    cell.border = thin_border
                
                col_idx += 1
            
            # Write affected systems
            cell = ws.cell(row=row_num, column=col_idx, value=affected_systems_str)
            cell.font = data_font
            cell.alignment = data_alignment_center
            cell.border = thin_border
            col_idx += 1
            
            # Write remaining catalog data
            for i in range(insert_position, len(catalog_headers)):
                col_name = catalog_headers[i]
                value = catalog_row.get(col_name, "")
                if pd.isna(value):
                    value = ""
                
                # Special handling for Risk Factor column
                if risk_factor_col is not None and i == risk_factor_col:
                    catalog_risk_value = str(value).strip() if value else ""
                    
                    # Apply color formatting (same logic as above)
                    if catalog_risk_value.upper() == 'CRITICAL':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = critical_font
                        cell.fill = critical_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    elif catalog_risk_value.upper() == 'HIGH':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = high_font
                        cell.fill = high_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    elif catalog_risk_value.upper() == 'MEDIUM':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = medium_font
                        cell.fill = medium_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    elif catalog_risk_value.upper() == 'LOW':
                        camelcase_value = convert_risk_to_camelcase(catalog_risk_value)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        cell.font = low_font
                        cell.fill = low_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    else:
                        camelcase_value = convert_risk_to_camelcase(group_max_risk)
                        cell = ws.cell(row=row_num, column=col_idx, value=camelcase_value)
                        if group_max_risk == 'critical':
                            cell.font = critical_font
                            cell.fill = critical_fill
                        elif group_max_risk == 'high':
                            cell.font = high_font
                            cell.fill = high_fill
                        elif group_max_risk == 'medium':
                            cell.font = medium_font
                            cell.fill = medium_fill
                        elif group_max_risk == 'low':
                            cell.font = low_font
                            cell.fill = low_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                else:
                    # Special handling for Audit Observation
                    if audit_observation_col is not None and i == audit_observation_col:
                        if len(vulnerabilities_list) > 1:
                            observation_text = "It was observed that the hosts are affected by multiple vulnerabilities, which are listed below.\n\n"
                            observation_text += vulnerabilities_str
                            if value:
                                value = f"{value}\n\n{observation_text}"
                            else:
                                value = observation_text
                        elif not value:
                            value = "It was observed that the host is affected by a vulnerability."
                    
                    # Special handling for CVE/CWE ID
                    if col_name == 'CVE/CWE ID' and value == "":
                        value = "N/A"
                    
                    # Apply left alignment for specific columns
                    cell = ws.cell(row=row_num, column=col_idx, value=str(value))
                    cell.font = data_font
                    if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                        cell.alignment = data_alignment_left
                    else:
                        cell.alignment = data_alignment_center
                    cell.border = thin_border
                
                col_idx += 1
            
            # Write POC columns (empty initially)
            if poc_col_start is not None and poc_col_end is not None:
                for poc_col in range(poc_col_start, poc_col_end + 1):
                    cell = ws.cell(row=row_num, column=poc_col, value="")
                    cell.border = thin_border
            
            ws.row_dimensions[row_num].height = 37.5
            row_num += 1
        
        # Handle unmatched vulnerabilities (including manual ones that weren't matched)
        unmatched_vulnerabilities = set(unique_vulnerabilities) - matched_vulnerabilities
        
        # Also include manual vulnerabilities that weren't in matched groups
        for manual_vuln in manually_added_vulnerabilities:
            found_in_groups = False
            for group in (merged_groups_from_session or []):
                if manual_vuln in group.get('matched_vulnerabilities', []):
                    found_in_groups = True
                    break
            if not found_in_groups:
                unmatched_vulnerabilities.add(manual_vuln)
        
        if unmatched_vulnerabilities:
            print(f"Unmatched vulnerabilities: {len(unmatched_vulnerabilities)}")
            
            for vulnerability in unmatched_vulnerabilities:
                # Get affected systems - first try from vulnerability_affected_systems dict
                affected_systems = vulnerability_affected_systems.get(vulnerability, [])
                
                # Get details from merge state if available
                details = None
                if use_merge_state:
                    # Find in new_group_details - check both matched and unmatched groups
                    for group in merged_groups_from_session:
                        if vulnerability in group.get('matched_vulnerabilities', []):
                            group_id = group['catalog_id']
                            details = new_group_details_dict.get(str(group_id), {})
                            break
                    
                    # If not found in matched groups, try to find in unmatched details
                    # For manual vulnerabilities that might not be in any group yet
                    if not details and vulnerability in manually_added_vulnerabilities:
                        # Search through all new_group_details to find this vulnerability
                        for group_id_str, group_details in new_group_details_dict.items():
                            if group_details.get('vulnerabilityName') == vulnerability:
                                details = group_details
                                # Also get affected systems from details if not already set
                                if not affected_systems and 'affectedSystems' in details:
                                    affected_systems_str = str(details['affectedSystems']).strip()
                                    if affected_systems_str:
                                        if '\n' in affected_systems_str:
                                            affected_systems = [s.strip() for s in affected_systems_str.split('\n') if s.strip()]
                                        elif ',' in affected_systems_str:
                                            affected_systems = [s.strip() for s in affected_systems_str.split(',') if s.strip()]
                                        else:
                                            affected_systems = [affected_systems_str]
                                break
                
                affected_systems_str = "\n".join(affected_systems) if affected_systems else ""
                
                # Write serial number
                cell = ws.cell(row=row_num, column=1, value=row_num - 1)
                cell.font = data_font
                cell.alignment = data_alignment_center
                cell.border = thin_border
                
                # Write vulnerability name
                vuln_name = vulnerability
                if details and details.get('vulnerabilityName'):
                    vuln_name = details.get('vulnerabilityName')
                
                cell = ws.cell(row=row_num, column=2, value=vuln_name)
                cell.font = data_font
                cell.alignment = data_alignment_center
                cell.border = thin_border
                
                # Write catalog data columns
                col_idx = 3
                for i in range(insert_position):
                    col_name = catalog_headers[i]
                    value_to_write = ""
                    
                    if details:
                        if risk_factor_col is not None and i == risk_factor_col:
                            value_to_write = str(details.get('riskFactor', '')).upper()
                        elif audit_observation_col is not None and i == audit_observation_col:
                            value_to_write = details.get('auditObservation', '')
                        elif col_name == 'CVE/CWE ID':
                            value_to_write = details.get('cveId', '') or "N/A"
                        elif 'cvss' in col_name.lower():
                            value_to_write = details.get('cvssScore', '')
                        elif 'impact' in col_name.lower():
                            value_to_write = details.get('impact', '')
                        elif 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                            value_to_write = details.get('recommendation', '')
                        elif 'reference' in col_name.lower() and 'link' in col_name.lower():
                            value_to_write = details.get('referenceLink', '')
                    
                    # Apply risk color if applicable
                    if risk_factor_col is not None and i == risk_factor_col and value_to_write:
                        catalog_risk_value = str(value_to_write).upper().strip()
                        cell = ws.cell(row=row_num, column=col_idx, value=catalog_risk_value)
                        if catalog_risk_value == 'CRITICAL':
                            cell.font = critical_font
                            cell.fill = critical_fill
                        elif catalog_risk_value == 'HIGH':
                            cell.font = high_font
                            cell.fill = high_fill
                        elif catalog_risk_value == 'MEDIUM':
                            cell.font = medium_font
                            cell.fill = medium_fill
                        elif catalog_risk_value == 'LOW':
                            cell.font = low_font
                            cell.fill = low_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    else:
                        cell = ws.cell(row=row_num, column=col_idx, value=value_to_write)
                        cell.font = data_font
                        if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                            cell.alignment = data_alignment_left
                        else:
                            cell.alignment = data_alignment_center
                        cell.border = thin_border
                    
                    col_idx += 1
                
                # Write affected systems
                cell = ws.cell(row=row_num, column=col_idx, value=affected_systems_str)
                cell.font = data_font
                cell.alignment = data_alignment_center
                cell.border = thin_border
                col_idx += 1
                
                # Write remaining catalog columns
                for i in range(insert_position, len(catalog_headers)):
                    col_name = catalog_headers[i]
                    value_to_write = ""
                    
                    if details:
                        if risk_factor_col is not None and i == risk_factor_col:
                            value_to_write = str(details.get('riskFactor', '')).upper()
                        elif audit_observation_col is not None and i == audit_observation_col:
                            value_to_write = details.get('auditObservation', '')
                        elif col_name == 'CVE/CWE ID':
                            value_to_write = details.get('cveId', '') or "N/A"
                        elif 'cvss' in col_name.lower():
                            value_to_write = details.get('cvssScore', '')
                        elif 'impact' in col_name.lower():
                            value_to_write = details.get('impact', '')
                        elif 'recommendation' in col_name.lower() or 'countermeasure' in col_name.lower():
                            value_to_write = details.get('recommendation', '')
                        elif 'reference' in col_name.lower() and 'link' in col_name.lower():
                            value_to_write = details.get('referenceLink', '')
                    
                    # Apply risk color if applicable
                    if risk_factor_col is not None and i == risk_factor_col and value_to_write:
                        catalog_risk_value = str(value_to_write).upper().strip()
                        cell = ws.cell(row=row_num, column=col_idx, value=catalog_risk_value)
                        if catalog_risk_value == 'CRITICAL':
                            cell.font = critical_font
                            cell.fill = critical_fill
                        elif catalog_risk_value == 'HIGH':
                            cell.font = high_font
                            cell.fill = high_fill
                        elif catalog_risk_value == 'MEDIUM':
                            cell.font = medium_font
                            cell.fill = medium_fill
                        elif catalog_risk_value == 'LOW':
                            cell.font = low_font
                            cell.fill = low_fill
                        cell.alignment = data_alignment_center
                        cell.border = thin_border
                    else:
                        cell = ws.cell(row=row_num, column=col_idx, value=value_to_write)
                        cell.font = data_font
                        if any(keyword in col_name.lower() for keyword in ['audit observation', 'impact', 'recommendation', 'countermeasure', 'reference link']):
                            cell.alignment = data_alignment_left
                        else:
                            cell.alignment = data_alignment_center
                        cell.border = thin_border
                    
                    col_idx += 1
                
                # Write POC columns
                if poc_col_start is not None and poc_col_end is not None:
                    for poc_col in range(poc_col_start, poc_col_end + 1):
                        cell = ws.cell(row=row_num, column=poc_col, value="")
                        cell.border = thin_border
                
                ws.row_dimensions[row_num].height = 37.5
                row_num += 1
        
        # Set column widths
        column_widths = [7, 35, 30, 15, 20, 10, 60, 60, 60, 40, 50, 30]
        num_cols_to_set = min(len(column_widths), poc_col_start - 1 if poc_col_start else len(headers))
        
        for i in range(num_cols_to_set):
            if i < len(column_widths):
                col_letter = get_column_letter(i + 1)
                ws.column_dimensions[col_letter].width = column_widths[i]
        
        # Set POC columns width
        if poc_col_start is not None and poc_col_end is not None:
            for col_idx in range(poc_col_start, poc_col_end + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 16
        
        print(f"✅ Website_VAPT worksheet created with {row_num - 2} rows")
        print(f"Matched vulnerabilities: {len(matched_vulnerabilities)}, Unmatched: {len(unmatched_vulnerabilities)}")
        print(f"Unique catalog entries used: {len(sorted_catalog_indices)}")
        
    except Exception as e:
        print(f"❌ Error creating Website_VAPT worksheet: {e}")
        import traceback
        traceback.print_exc()


def create_website_metadata_worksheet(wb, form_data):
    """
    Create Meta_Data worksheet - structured format same as Infrastructure VAPT reference.
    """
    try:
        print("📊 Creating Meta_Data worksheet...")
        ws = wb.create_sheet("Meta_Data")
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 35  # Field names
        ws.column_dimensions['B'].width = 60  # Values
        
        # Define the data structure for the metadata
        metadata_sections = [
            {
                'title': 'ORGANIZATION INFORMATION',
                'data': [
                    ('Organization Name', form_data.get('organization', '')),
                    ('City', form_data.get('city', '')),
                    ('State', form_data.get('state', ''))
                ]
            },
            {
                'title': 'AUDIT PERIOD',
                'data': [
                    ('Start Date', form_data.get('startDate', '')),
                    ('End Date', form_data.get('endDate', ''))
                ]
            },
            {
                'title': 'REPORT PREPARED BY',
                'data': [
                    ('Name', f"{form_data.get('preparedByTitle', '')} {form_data.get('preparedByName', '')}".strip()),
                ]
            },
            {
                'title': 'AUDITEE DETAILS',
                'data': [
                    ('Name', f"{form_data.get('auditeeTitle', '')} {form_data.get('auditeeName', '')}".strip()),
                    ('Designation', form_data.get('designation', ''))
                ]
            }
        ]
        
        # Add Email Addresses section
        bank_emails = form_data.get('bankEmails', [])
        if bank_emails:
            email_data = []
            for i, email in enumerate(bank_emails, 1):
                if email.strip():
                    email_data.append((f'Email {i}', email.strip()))
            
            if email_data:
                metadata_sections.append({
                    'title': 'ORGANIZATION EMAIL ADDRESSES',
                    'data': email_data
                })
        
        # Add Auditing Team sections
        team_names = form_data.get('teamNames', [])
        team_designations = form_data.get('teamDesignations', [])
        team_emails = form_data.get('teamEmails', [])
        team_qualifications = form_data.get('teamQualifications', [])
        team_certified = form_data.get('teamCertified', [])
        
        if team_names:
            for i in range(len(team_names)):
                if team_names[i].strip():
                    team_member_data = [
                        (f'Team Member {i+1} - Name', team_names[i].strip()),
                        (f'Team Member {i+1} - Designation', team_designations[i] if i < len(team_designations) else ''),
                        (f'Team Member {i+1} - Email', team_emails[i] if i < len(team_emails) else ''),
                        (f'Team Member {i+1} - Qualification', team_qualifications[i] if i < len(team_qualifications) else ''),
                        (f'Team Member {i+1} - Certified', 
                         'Yes' if team_certified[i].lower() == 'yes' else 'No' if team_certified[i].lower() == 'no' else team_certified[i] 
                         if i < len(team_certified) else '')
                    ]
                    
                    metadata_sections.append({
                        'title': f'AUDITING TEAM MEMBER {i+1}',
                        'data': team_member_data
                    })
        
        # Write data to worksheet
        row = 1
        
        for section in metadata_sections:
            # Write section title (merged across both columns)
            cell = ws.cell(row=row, column=1, value=section['title'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=row, column=2, value='')
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            ws.row_dimensions[row].height = 25
            row += 1
            
            # Write section data
            for field_name, field_value in section['data']:
                cell = ws.cell(row=row, column=1, value=field_name)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                cell = ws.cell(row=row, column=2, value=field_value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                ws.row_dimensions[row].height = 20
                row += 1
            
            # Add empty row after each section for spacing
            row += 1
        
        print(f"✅ Meta_Data worksheet created with {row} rows")
        
    except Exception as e:
        print(f"❌ Error creating Meta_Data worksheet: {e}")
        import traceback
        traceback.print_exc()


@website_vapt_bp.route('/process_website_vapt_first_audit_report', methods=['POST'])
def process_website_first_audit_report():
    """Main endpoint to process Website VAPT First Audit Report."""
    try:
        print("\n" + "="*80)
        print("🚀 Processing Website VAPT First Audit Report")
        print("="*80)
        
        # Validate files
        if 'nmapFiles' not in request.files or 'nessusFiles' not in request.files:
            return jsonify({"error": "Both Nmap and Nessus files are required"}), 400
        
        nmap_file = request.files['nmapFiles']
        nessus_file = request.files['nessusFiles']
        evidence_file = request.files.get('evidenceFiles')
        
        if nmap_file.filename == '' or nessus_file.filename == '':
            return jsonify({"error": "Please select all required files"}), 400
        
        # Extract form data
        form_data = {
            'organization': request.form.get('organization'),
            'otherOrganization': request.form.get('otherOrganization'),
            'city': request.form.get('city'),
            'otherCity': request.form.get('otherCity'),
            'state': request.form.get('state'),
            'startDate': request.form.get('startDate'),
            'endDate': request.form.get('endDate'),
            'preparedByTitle': request.form.get('preparedByTitle'),
            'preparedByName': request.form.get('preparedByName'),
            'auditeeTitle': request.form.get('auditeeTitle'),
            'auditeeName': request.form.get('auditeeName'),
            'designation': request.form.get('designation')
        }
        
        # Handle "Other" selections
        if form_data['organization'] == 'other':
            form_data['organization'] = form_data['otherOrganization']
        if form_data['city'] == 'other':
            form_data['city'] = form_data['otherCity']
        
        # Extract email addresses
        bank_emails = request.form.getlist('bankEmail[]')
        form_data['bankEmails'] = [email for email in bank_emails if email.strip()]
        
        # Extract team member details
        team_names = request.form.getlist('teamName[]')
        team_designations = request.form.getlist('teamDesignation[]')
        team_emails = request.form.getlist('teamEmail[]')
        team_qualifications = request.form.getlist('teamQualification[]')
        
        # Extract team certified values
        team_certified = []
        for i in range(len(team_names)):
            certified_value = request.form.get(f'teamCertified[{i}]', 'no')
            team_certified.append(certified_value)
        
        form_data['teamNames'] = team_names
        form_data['teamDesignations'] = team_designations
        form_data['teamEmails'] = team_emails
        form_data['teamQualifications'] = team_qualifications
        form_data['teamCertified'] = team_certified
        
        # Process files
        print("📁 Processing Nmap files...")
        nmap_data = process_nmap_zip_website(nmap_file)  # Returns dict: {ip: [(port, service), ...]}
        
        print("📁 Processing Nessus files...")
        nessus_dataframes = process_nessus_zip_website(nessus_file)
        
        if not nessus_dataframes:
            return jsonify({"error": "No valid Nessus data found"}), 400
        
        # Create Excel workbook
        print("📊 Creating Excel workbook...")
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create all worksheets (order: Meta_Data, Nmap Files, Nessus CSV Files, Scope, Summary, Website_VAPT)
        # Create Meta_Data first (will be at index 0)
        create_website_metadata_worksheet(wb, form_data)
        
        # Create Nmap Files
        if nmap_data:
            create_website_nmap_worksheet(wb, nmap_data)  # Pass dictionary
        else:
            print("⚠️ No Nmap data found, skipping Nmap worksheet")
        
        # Create Nessus CSV Files
        create_website_nessus_csv_worksheet(wb, nessus_dataframes)
        
        # Create Scope
        create_website_scope_worksheet(wb, nessus_dataframes)
        
        # Create Summary
        create_website_summary_worksheet(wb, nessus_dataframes)
        
        # Create Website_VAPT worksheet
        create_website_vapt_worksheet(wb, nessus_dataframes)
        
        # Reorder worksheets to the desired sequence: Meta_Data, Nmap Files, Nessus CSV Files, Scope, Summary, Website_VAPT
        desired_order = ["Meta_Data", "Nmap Files", "Nessus CSV Files", "Scope", "Summary", "Website_VAPT"]
        
        # Get current sheet names (make a copy to avoid iteration issues)
        current_sheets = list(wb.sheetnames)
        
        # Create a list of sheet names in the desired order (only include sheets that exist)
        ordered_sheet_names = []
        for sheet_name in desired_order:
            if sheet_name in current_sheets:
                ordered_sheet_names.append(sheet_name)
        
        # Add any remaining sheets that weren't in desired order (shouldn't happen, but just in case)
        for sheet_name in current_sheets:
            if sheet_name not in ordered_sheet_names:
                ordered_sheet_names.append(sheet_name)
        
        # Reorder sheets by manipulating the workbook's _sheets list directly
        ordered_sheets = []
        for sheet_name in ordered_sheet_names:
            if sheet_name in wb.sheetnames:
                ordered_sheets.append(wb[sheet_name])
        
        # Replace the workbook's sheet list with the ordered list
        wb._sheets = ordered_sheets
        
        print(f"📋 Worksheets reordered: {ordered_sheet_names}")
        
        # Save to BytesIO initially
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Initialize temp_poc_images_folder variable for cleanup
        temp_poc_images_folder = None
        
        # Insert POC images if evidence file is provided
        if evidence_file and evidence_file.filename != '':
            try:
                print("Processing POC images from evidence files...")
                
                # Save Excel file temporarily
                temp_excel_path = "temp_website_poc.xlsx"
                with open(temp_excel_path, 'wb') as temp_file:
                    temp_file.write(output.getvalue())
                
                # Generate timestamp for temporary POC images folder (format: DDMMYYYYHHMMSS)
                from datetime import datetime
                current_ts = datetime.now().strftime('%d%m%Y%H%M%S')
                temp_poc_images_folder = f"temp_poc_images_{current_ts}"
                print(f"📁 Using timestamped POC images folder: {temp_poc_images_folder}")
                
                # Extract POC images
                poc_mapping = extract_poc_images(evidence_file, temp_poc_images_folder)
                
                if poc_mapping:
                    print(f"Found {len(poc_mapping)} POC images to insert")
                    # Combine Nessus dataframes for vulnerability matching
                    combined_nessus = pd.concat(nessus_dataframes, ignore_index=True) if nessus_dataframes else pd.DataFrame()
                    
                    # Insert POC images into Website_VAPT worksheet
                    if 'Website_VAPT' in wb.sheetnames:
                        rows_with_objects = insert_poc_images_to_excel(temp_excel_path, poc_mapping, combined_nessus)
                        print(f"✅ Inserted POC images into Website_VAPT worksheet")
                    else:
                        print("⚠️ Website_VAPT worksheet not found, skipping POC image insertion")
                    
                    # Read the updated file
                    with open(temp_excel_path, 'rb') as updated_file:
                        output = BytesIO(updated_file.read())
                
                # Clean up temporary file
                if os.path.exists(temp_excel_path):
                    os.remove(temp_excel_path)
                    
            except Exception as e:
                print(f"Error processing POC images: {e}")
                import traceback
                traceback.print_exc()
                # Continue with original file if POC processing fails
        
        
        # Generate filename
        filename = generate_website_filename(form_data['organization'], form_data['endDate'])
        
        print(f"✅ Report generated: {filename}")
        print("="*80 + "\n")
        
        response = send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
        # Clean up temp_poc_images folder after Excel generation
        try:
            import shutil
            # Clean up timestamped temp_poc_images folder
            if temp_poc_images_folder and os.path.exists(temp_poc_images_folder):
                shutil.rmtree(temp_poc_images_folder)
                print(f"✅ Successfully deleted {temp_poc_images_folder} folder")
            # Also clean up any old temp_poc_images folders (backward compatibility)
            import glob
            old_folders = glob.glob("temp_poc_images*")
            for folder in old_folders:
                if os.path.isdir(folder):
                    try:
                        shutil.rmtree(folder)
                        print(f"✅ Cleaned up old folder: {folder}")
                    except:
                        pass
        except Exception as e:
            print(f"⚠️ Error deleting temp_poc_images folder: {e}")
        
        return response
        
    except Exception as e:
        print(f"❌ Error processing report: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error generating report: {str(e)}"}), 500
